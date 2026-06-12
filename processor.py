"""
Balance Sheet Year-Shift Processor  (v9 — single-pass, Render-safe)

Key improvements over v8:
  • ONE workbook open total (was 4). Collects values + formulas in one pass.
  • all_formula_rows protection removed — wrong concept and memory-heavy.
    Instead: only clear a CY cell if THAT SPECIFIC CY column cell is a
    plain constant (not a formula). This is correct and sufficient.
  • HFPL / "Current year" / "Previous year" header detection added.
  • XML edits via regex on raw bytes — never ET.tostring → no namespace corruption.
  • Falls back gracefully when lxml absent (sharedStrings via ET is safe there).
  • BIG_ROWS raised; big sheets skip CY/PY shift but still get date text updates.
"""

import re
import html
import zipfile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter

# ── thresholds ─────────────────────────────────────────────────────────────────
BIG_ROWS        = 3000
BIG_COLS        = 150
HEADER_SCAN_ROWS = 20

# ── XML namespaces ─────────────────────────────────────────────────────────────
_NS   = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

# ── Fallback column map  (lower-cased sheet name → [(cy_col, py_col)]) ────────
#    Used when auto-detect finds nothing.
SHEET_COL_MAP = {
    # DP Thapar format
    "bs":            [("E", "F")],
    "p&l":           [("E", "F")],
    "notes to bs":   [("D", "E")],
    "notes to p&l":  [("D", "E")],
    "details":       [("D", "E")],
    "gross profit":  [("B", "C"), ("F", "G")],
    # HFPL / HUG FOODS format
    "notes":         [("H", "J")],
    "othr notes":    [("H", "J")],
    "share cap":     [("H", "J")],
    "provision":     [("H", "J")],
    "provisions":    [("H", "J")],
    "cash flow":     [("D", "F")],
    "dep co":        [("D", "F")],
    "consump":       [("F", "H")],
    "dep":           [("D", "F")],
}

# Sheets that only contain text — skip CY/PY shift entirely
TEXT_ONLY_SHEETS = {s.strip().lower() for s in [
    "notes to accounts", "Fixed Assets C. Yr.", "Fixed Assets P. Yr.",
    "FA2022", "Tax audit ", "Tax Audit report", "PPE",
    "acc policies",
]}

# Capital sheets need special year-shift logic (Bugs 4 & 5)
CAPITAL_SHEET_NAMES = {"capital"}

def detect_fixed_asset_sheet_names(sheetnames):
    """Return (cy_sheet_name, py_sheet_name) for fixed-asset sheets."""
    cy_sn = py_sn = None
    for sn in sheetnames or []:
        sl = (sn or "").strip().lower()
        if "p. yr" in sl or "p.yr" in sl or ("fixed" in sl and (" p." in sl or "p. " in sl)):
            py_sn = sn
        elif "fixed asset" in sl or (sl.startswith("fa") and "2022" not in sl) or "ppe" in sl:
            if cy_sn is None:
                cy_sn = sn
    return cy_sn, py_sn

# Sheets whose data must never be touched (raw transaction dumps etc.)
RAW_DATA_SHEETS = {s.strip().lower() for s in [
    "new trial", "summary trial", "purchase report", "sale report",
    "stk", "other details", "debtors", "creditors",
    "purchase report pivot", "sales report pivot", "control",
    "pending", "legal case", "provisions",
]}


# ═══════════════════════════════════════════════════════════════════════════════
#  Date-replacement pairs
# ═══════════════════════════════════════════════════════════════════════════════

def _date_replacements(cy: str, ny: str) -> list:
    po = str(int(cy) - 1)
    PH, PH_R = "__NEWCY__", "__FYRNG__"

    patterns = [
        "31.03.{y}", "31 March, {y}", "31 March {y}",
        "31st March, {y}", "31st March {y}",
        "31ST MARCH ,{y}", "31ST MARCH, {y}", "31ST MARCH {y}",
        "31 MARCH, {y}", "31 MARCH {y}",
        "31st MARCH, {y}", "31st MARCH {y}",
        "year ended 31 March, {y}", "year ended, 31st March, {y}",
        "year end 31 March, {y}",
        "year ending 31.03.{y}", "YEAR ENDING 31.03.{y}",
        "YEAR ENDING 31ST MARCH ,{y}", "YEAR ENDING 31ST MARCH, {y}",
        "YEAR ENDING 31ST MARCH {y}",
        "as at 31 March, {y}", "AS AT 31ST MARCH {y}",
        "AS AT 31ST MARCH, {y}", "AS AT 31 MARCH, {y}",
        "AS AT 31st MARCH, {y}", "AS AT 31st MARCH {y}",
        "for the year ended, 31st March, {y}",
        "for the year ended 31 March, {y}",
        "FOR THE YEAR ENDED 31ST MARCH, {y}",
        "FOR THE YEAR ENDED 31ST MARCH {y}",
        "FOR THE YEAR ENDED 31st MARCH, {y}",
        "FOR THE YEAR ENDED 31st MARCH {y}",
    ]

    # Step A: CY dates → placeholder
    a = [(p.format(y=cy), p.format(y=PH)) for p in patterns]

    # Step B: old PY dates → CY dates  (so they become the new PY column header)
    b = []
    for old_tpl, new_tpl in [
        ("1st April {po}",       "1st April {cy}"),
        ("1 April {po}",         "1 April {cy}"),
        ("01.04.{po}",           "01.04.{cy}"),
        ("31.03.{po}",           "31.03.{cy}"),
        ("31st March {po}",      "31st March {cy}"),
        ("31st March, {po}",     "31st March, {cy}"),
        ("31 March, {po}",       "31 March, {cy}"),
        ("31 March {po}",        "31 March {cy}"),
        ("31ST MARCH {po}",      "31ST MARCH {cy}"),
        ("31ST MARCH, {po}",     "31ST MARCH, {cy}"),
        ("31st MARCH {po}",      "31st MARCH {cy}"),
        ("31st MARCH, {po}",     "31st MARCH, {cy}"),
        ("AS AT 31ST MARCH {po}","AS AT 31ST MARCH {cy}"),
        ("AS AT 31ST MARCH, {po}","AS AT 31ST MARCH, {cy}"),
        ("AS AT 31st MARCH {po}","AS AT 31st MARCH {cy}"),
        ("AS AT 31st MARCH, {po}","AS AT 31st MARCH, {cy}"),
    ]:
        b.append((old_tpl.format(po=po, cy=cy), new_tpl.format(po=po, cy=cy)))

    # Step C: placeholder → NY dates
    c = [(p.format(y=PH), p.format(y=ny)) for p in patterns]

    # Step D: fiscal year range strings  e.g. "2024-25" → "2025-26"
    po_s = po[2:]; cy_s = cy[2:]; ny_s = ny[2:]
    pp = str(int(po) - 1); pp_s = pp[2:]
    r = [
        (f"{po}-{cy_s}", PH_R),          # "2024-25" → placeholder
        (f"{po}-{cy}",   f"{PH_R}L"),     # "2024-2025" → placeholder+L
        (f"{pp}-{po_s}", f"{po}-{cy_s}"), # "2023-24" → "2024-25"
        (f"{pp}-{po}",   f"{po}-{cy}"),   # "2023-2024" → "2024-2025"
        (f"{PH_R}L",     f"{cy}-{ny}"),   # placeholder+L → "2025-2026"
        (PH_R,           f"{cy}-{ny_s}"), # placeholder → "2025-26"
    ]
    return a + b + c + r


def _apply_pairs(text: str, pairs: list) -> str:
    for old, new in pairs:
        text = text.replace(old, new)
    return text


# ═══════════════════════════════════════════════════════════════════════════════
#  Column-detection helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _has_year_date(val: str, yr: str) -> bool:
    """Return True if val is a date-bearing string for year yr."""
    if not isinstance(val, str):
        return False
    flat = val.replace("\n", " ").replace("\r", " ")
    if yr not in flat:
        return False
    return ("31.03." in flat or any(m in flat for m in (
        "March", "MARCH", "march",
        "year end", "Year end", "YEAR END",
        "as at", "As at", "AS AT",
    )))


def _has_cy_py_label(val: str) -> tuple:
    """
    Detect HFPL-style headers: 'Current year' / 'Previous year'.
    Returns (is_cy, is_py).
    """
    if not isinstance(val, str):
        return False, False
    lower = val.strip().lower()
    is_cy = lower in ("current year", "current year ", "cy", "current")
    is_py = lower in ("previous year", "previous year ", "py", "previous")
    return is_cy, is_py


_COL_RE = re.compile(r'^([A-Z]+)(\d+)$')

# Excel built-in date/time number-format IDs (ECMA-376 §18.8.30)
_BUILTIN_DATE_FMT_IDS = set(range(14, 23)) | set(range(27, 37)) | set(range(45, 48)) | {56}


def _is_date_format_code(code: str) -> bool:
    """True if a numFmt formatCode string represents a date/time format."""
    # Strip locale/currency prefixes like [$-F800], [$-409]
    cleaned = re.sub(r'\[\$[^\]]*\]', '', code)
    # Strip quoted literals and escaped characters (e.g. \, \ )
    cleaned = re.sub(r'"[^"]*"', '', cleaned)
    cleaned = re.sub(r'\\.', '', cleaned)
    cleaned_lower = cleaned.lower()
    has_date_letters = any(ch in cleaned_lower for ch in 'ymdh')
    has_number_placeholder = '#' in cleaned or '0' in cleaned
    return has_date_letters and not has_number_placeholder


def _build_date_style_set(filepath: str) -> set:
    """
    Parse xl/styles.xml and return the set of cellXfs style indices (the
    values used in <c s="N"> attributes) whose number format is a date/time
    format. Used to detect when writing a numeric value into a cell would
    cause it to render as a date (e.g. 7080 -> "1919-05-20").
    """
    try:
        with zipfile.ZipFile(filepath) as z:
            styles_xml = z.read("xl/styles.xml").decode("utf-8", errors="replace")
    except Exception:
        return set()

    # Collect custom numFmt definitions: {numFmtId: formatCode}
    custom_fmts = {}
    m = re.search(r'<numFmts\b[^>]*>(.*?)</numFmts>', styles_xml, re.DOTALL)
    if m:
        for fid, code in re.findall(
            r'<numFmt\s+numFmtId="(\d+)"\s+formatCode="([^"]*)"\s*/>', m.group(1)
        ):
            custom_fmts[int(fid)] = html.unescape(code)

    # Walk cellXfs in order — each <xf> corresponds to style index = its
    # position in the list (0-based), matching the <c s="N"> attribute.
    date_style_indices = set()
    m2 = re.search(r'<cellXfs\b[^>]*>(.*?)</cellXfs>', styles_xml, re.DOTALL)
    if m2:
        xf_records = re.findall(r'<xf\b[^>]*?(?:/>|>.*?</xf>)', m2.group(1), re.DOTALL)
        for idx, xf in enumerate(xf_records):
            fm = re.search(r'numFmtId="(\d+)"', xf)
            if not fm:
                continue
            numfmt_id = int(fm.group(1))
            if numfmt_id in _BUILTIN_DATE_FMT_IDS:
                date_style_indices.add(idx)
            elif numfmt_id in custom_fmts and _is_date_format_code(custom_fmts[numfmt_id]):
                date_style_indices.add(idx)

    return date_style_indices




def _col_idx(letter: str) -> int:
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - 64)
    return n


def _detect_columns(ws, closing_year: int) -> list:
    """
    Scan header rows for CY/PY column markers.
    Supports both:
      • date-bearing strings  ("As at 31st March, 2025")
      • label strings         ("Current year" / "Previous year")
    Returns [(cy_letter, py_letter), ...].
    """
    cy_s, py_s = str(closing_year), str(closing_year - 1)
    cy_cols, py_cols = set(), set()

    for row in ws.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS, max_col=60):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            if not isinstance(v, str):
                continue
            if v.startswith("="):
                continue

            # Date-bearing detection
            if _has_year_date(v, cy_s):
                cy_cols.add(cell.column)
            if _has_year_date(v, py_s):
                py_cols.add(cell.column)

            # Label detection ("Current year" / "Previous year")
            is_cy_lbl, is_py_lbl = _has_cy_py_label(v)
            if is_cy_lbl:
                cy_cols.add(cell.column)
            if is_py_lbl:
                py_cols.add(cell.column)

    # Also detect fiscal year range labels e.g. "2024-25", "2023-24"
    cy_s2 = cy_s[2:]   # "25"
    py_s2 = py_s[2:]   # "24"
    cy_range = f"{py_s}-{cy_s2}"   # "2024-25"
    py_range = f"{str(int(py_s)-1)[2:]}-{py_s2}" if False else None  # not needed

    for row in ws.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS, max_col=60):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            if not isinstance(v, str) or v.startswith("="):
                continue
            stripped = v.strip()
            if stripped == cy_range:
                cy_cols.add(cell.column)

    if not cy_cols:
        return []

    # BUG 1 FIX: Remove col 1 (A) from cy_cols — it's almost always a 
    # label/title column containing text like "Current Year (CY)" but no data.
    # Real CY data columns are always col 2 (B) or higher.
    cy_cols.discard(1)
    py_cols.discard(1)
    py_cols.discard(2)  # Col B is rarely a PY data col (usually labels)

    if not cy_cols:
        return []

    pairs, used = [], set()
    for cc in sorted(cy_cols):
        # Look for a PY col to the right (offset 1, 2, or 3)
        for off in [1, 2, 3]:
            cand = cc + off
            if cand in py_cols and cand not in used:
                pairs.append((get_column_letter(cc), get_column_letter(cand)))
                used.add(cand)
                break
        # If no PY col found but CY found, don't force it
    return pairs


# ═══════════════════════════════════════════════════════════════════════════════
#  ZIP helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _parse_dim(dim_str: str) -> tuple:
    if ":" not in dim_str:
        return 1, 1
    _, end = dim_str.split(":")
    m = _COL_RE.match(end)
    if not m:
        return 1, 1
    return int(m.group(2)), _col_idx(m.group(1))


def _get_sizes_from_zip(filepath: str) -> dict:
    """Read sheet dimensions from XML headers — no cell parsing, very fast."""
    sizes = {}
    dim_re = re.compile(rb'<dimension ref="([^"]+)"')
    with zipfile.ZipFile(filepath) as z:
        for item in z.infolist():
            fn = item.filename
            if fn.startswith("xl/worksheets/sheet") and fn.endswith(".xml"):
                header = z.read(fn)[:2000]
                m = dim_re.search(header)
                sizes[fn] = _parse_dim(m.group(1).decode()) if m else (0, 0)
    return sizes


def _sheet_file_map(z) -> dict:
    """Return {sheet_name: 'xl/worksheets/sheetN.xml'}."""
    root = ET.fromstring(z.read("xl/workbook.xml"))
    srid = {}
    for el in root.iter(f"{{{_NS}}}sheet"):
        srid[el.get("name")] = el.get(f"{{{_NS_R}}}id")
    rr = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
    rf = {r.get("Id"): r.get("Target") for r in rr}
    out = {}
    for name, rid in srid.items():
        t = rf.get(rid, "")
        t = t.lstrip("/")
        if t and not t.startswith("xl/"):
            t = f"xl/{t}"
        if t:
            out[name] = t
    return out


# ═══════════════════════════════════════════════════════════════════════════════
#  SINGLE-PASS workbook scan  (ONE open, collect everything)
# ═══════════════════════════════════════════════════════════════════════════════

def _scan_workbook(filepath: str, closing_year: int) -> tuple:
    """
    Open the workbook ONCE (read_only=True, data_only=True).

    Returns:
        sizes        {sheet_name: (rows, cols)}
        shift_map    {sheet_name: [(cy_letter, py_letter), ...]}
        cy_values    {sheet_name: {cy_col_letter: {row_int: float}}}
        cy_formulas  {sheet_name: {cy_col_letter: set_of_row_ints}}

    Strategy:
    - data_only=True means formula cells return their last-cached value.
    - We read each CY column cell: if the cached value is numeric → record it
      in cy_values.  If the raw XML shows it was a formula, we mark it in
      cy_formulas so we know NOT to clear it later.

    Detecting formula cells with data_only=True:
    - openpyxl sets cell.data_type == 'n' for numbers, but for formula cells
      that evaluated to a number, it also returns cell.value as a number.
    - The only way to distinguish formula vs constant is to check the raw XML.
    - We do that efficiently via a single regex pass on each sheet's XML bytes,
      extracting all formula-cell refs before opening with openpyxl.
    """

    # ── Step 1: sizes from ZIP (no cell parsing) ───────────────────────────
    with zipfile.ZipFile(filepath) as z:
        smap = _sheet_file_map(z)
        file_sizes = {}
        dim_re = re.compile(rb'<dimension ref="([^"]+)"')
        for fn in smap.values():
            try:
                header = z.read(fn)[:2000]
                m = dim_re.search(header)
                file_sizes[fn] = _parse_dim(m.group(1).decode()) if m else (0, 0)
            except Exception:
                file_sizes[fn] = (0, 0)

    sizes = {sn: file_sizes.get(sf, (0, 0)) for sn, sf in smap.items()}
    big_names = {n for n, (r, c) in sizes.items() if r > BIG_ROWS or c > BIG_COLS}

    # ── Step 2: extract formula-cell refs directly from ZIP XML (regex, no ET)
    #    formula_refs[sheet_file] = set of "ColRow" refs like {"E5","E12"}
    formula_refs = {}  # {sheet_file: set_of_refs}
    # A formula cell in OOXML looks like: <c r="E5" ...><f ...>...</f><v>...</v></c>
    # We just need to find all refs that contain a <f> tag.
    _f_cell_re = re.compile(rb'<c\b[^>]*\br="([A-Z]+\d+)"[^>]*>[^<]*<f[ />]')
    with zipfile.ZipFile(filepath) as z:
        for sn, sf in smap.items():
            if sn in big_names:
                continue
            sl = sn.strip().lower()
            if sl in TEXT_ONLY_SHEETS or sl in RAW_DATA_SHEETS:
                continue
            try:
                xml_data = z.read(sf)
                refs = set(m.group(1).decode() for m in _f_cell_re.finditer(xml_data))
                formula_refs[sf] = refs
            except Exception:
                formula_refs[sf] = set()

    # ── Step 3: single openpyxl open (data_only → gets cached values) ─────
    fb = {k.strip().lower(): v for k, v in SHEET_COL_MAP.items()}
    shift_map  = {}
    cy_values  = {}
    cy_formulas = {}

    wb = load_workbook(filepath, read_only=True, data_only=True)
    try:
        for sn in wb.sheetnames:
            sl = sn.strip().lower()
            if sn in big_names:
                continue
            if sl in TEXT_ONLY_SHEETS or sl in RAW_DATA_SHEETS:
                continue

            ws = wb[sn]

            # Detect CY/PY columns
            det = _detect_columns(ws, closing_year)
            if not det:
                det = fb.get(sl, [])
            if not det:
                continue

            shift_map[sn] = det

            # Build formula-row sets from pre-extracted refs
            sf = smap.get(sn, "")
            sheet_frefs = formula_refs.get(sf, set())

            cy_vals_sheet   = {}
            cy_frows_sheet  = {}

            for cy_l, _ in det:
                ci = _col_idx(cy_l)
                vals  = {}
                frows = set()

                for row in ws.iter_rows(min_col=ci, max_col=ci):
                    for cell in row:
                        if isinstance(cell, MergedCell):
                            continue
                        if not hasattr(cell, 'row') or cell.row is None:
                            continue
                        rn = cell.row
                        ref = f"{cy_l}{rn}"

                        if ref in sheet_frefs:
                            # It's a formula cell — record row, grab cached value too
                            frows.add(rn)
                            if isinstance(cell.value, (int, float)) and cell.value is not None:
                                vals[rn] = float(cell.value)
                            elif cell.value is None:
                                # Formula evaluated to empty/zero — record as 0
                                vals[rn] = 0.0
                        elif isinstance(cell.value, (int, float)) and cell.value is not None:
                            vals[rn] = float(cell.value)
                        elif isinstance(cell.value, str) and cell.value.strip() in ("-", "—", "–", "-"):
                            # BUG 2 FIX: dash string = zero. Copy 0 to PY column.
                            vals[rn] = 0.0

                cy_vals_sheet[cy_l]  = vals
                cy_frows_sheet[cy_l] = frows

            cy_values[sn]   = cy_vals_sheet
            cy_formulas[sn] = cy_frows_sheet

    finally:
        wb.close()

    # ── Step 4: Scan capital sheet for CY/PY row data (Bugs 4 & 5) ──────
    #
    # Templates vary in TWO ways that must both be detected dynamically:
    #
    # (a) COLUMN LAYOUT — the number and order of data columns (C onward)
    #     differs between firms. A 2-partner firm may have 7 columns
    #     (opening, introduced, interest, salary, withdrawals, profit,
    #     closing); a single-proprietor firm may have only 5 (opening,
    #     introduced, withdrawals, profit, closing). We detect each
    #     column's ROLE from its header text in the header row (the row
    #     containing "Sr. No." / "Name of...") rather than hardcoding
    #     column positions.
    #
    # (b) PY ROW STRUCTURE — two patterns exist:
    #     Pattern A (multi-partner, e.g. 2 partners): a separate
    #       "Previous Year (PY)" SECTION HEADER followed by per-partner
    #       data rows (Sr.No + Name + values), mirroring the CY block.
    #     Pattern B (single proprietor): the "Previous Year (PY)" label
    #       and the PY data values are in the SAME ROW (no separate
    #       Sr.No/Name columns for PY).
    #
    # Strategy:
    #   1. Find the header row (contains "Sr. No.") and derive col_roles
    #      = {col_num: role} from its text.
    #   2. Find all "Curret/Current Year (CY)" and "Previous Year (PY)"
    #      section header rows.
    #   3. Find CY data rows: col A = Sr. No. (int), col B = name (str),
    #      numeric data in the detected columns. Take rows belonging to
    #      the FIRST CY block.
    #   4. For PY: first look for Pattern-A data rows in the LAST PY
    #      block (Sr.No + Name rows after a PY section header). If none
    #      found, fall back to Pattern B: an inline row whose col A text
    #      contains "previous year" AND which itself has numeric data in
    #      the detected columns.
    #   5. Match CY rows to PY rows by name (Pattern A) or 1:1 positional
    #      (Pattern B, since there's only one row of each).
    cap_data = {}
    wb2 = load_workbook(filepath, read_only=True, data_only=True)
    try:
        for sn in wb2.sheetnames:
            if sn.strip().lower() not in CAPITAL_SHEET_NAMES:
                continue
            ws_cap = wb2[sn]

            # --- (a) Detect header row and column roles ---
            header_row = None
            for r in range(1, 20):
                a = ws_cap.cell(r, 1).value
                if isinstance(a, str) and a.strip().lower() in ("sr. no.", "sr no.", "sr no"):
                    header_row = r
                    break

            col_roles = {}
            if header_row:
                for c in range(3, 12):
                    v = ws_cap.cell(header_row, c).value
                    if not isinstance(v, str):
                        continue
                    vl = v.strip().lower()
                    if "march" in vl:
                        col_roles[c] = "closing"
                    elif "april" in vl or ("as at" in vl and "1st" in vl):
                        col_roles[c] = "opening"
                    elif "introduced" in vl:
                        col_roles[c] = "introduced"
                    elif "interest" in vl:
                        col_roles[c] = "interest"
                    elif "salary" in vl:
                        col_roles[c] = "salary"
                    elif "withdraw" in vl:
                        col_roles[c] = "withdrawals"
                    elif "profit" in vl or "loss" in vl:
                        col_roles[c] = "profit"

            if not col_roles:
                # Fall back to the legacy fixed 7-column layout if header
                # text couldn't be parsed (shouldn't normally happen).
                col_roles = {3: "opening", 4: "introduced", 5: "interest",
                              6: "salary", 7: "withdrawals", 8: "profit", 9: "closing"}

            data_cols = sorted(col_roles.keys())

            # --- (b) Section headers ---
            cy_hdrs, py_hdrs = [], []
            for row in ws_cap.iter_rows(min_row=1, max_row=60):
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    v = cell.value
                    if isinstance(v, str):
                        vl = v.strip().lower()
                        if "curret year" in vl or "current year" in vl:
                            cy_hdrs.append(cell.row)
                        elif "previous year" in vl:
                            py_hdrs.append(cell.row)

            if not cy_hdrs:
                continue

            # --- CY/PY data rows: Sr.No (int) + name (str) + numeric data ---
            data_rows = []
            for r in range(1, 60):
                a = ws_cap.cell(r, 1).value
                b = ws_cap.cell(r, 2).value
                if isinstance(a, int) and isinstance(b, str):
                    bl = b.strip().lower()
                    if len(b.strip()) > 2 and bl not in (
                        "name of partners", "name of proprietor", "sr. no.", "particulars"
                    ):
                        if any(isinstance(ws_cap.cell(r, c).value, (int, float))
                               for c in data_cols):
                            data_rows.append((r, b.strip()))

            if not data_rows:
                continue

            all_hdrs = sorted(set(cy_hdrs + py_hdrs))

            def _rows_in_block(start_hdr):
                idx = all_hdrs.index(start_hdr)
                end = all_hdrs[idx + 1] if idx + 1 < len(all_hdrs) else 10_000
                return [(r, name) for r, name in data_rows if start_hdr < r < end]

            cy_block = _rows_in_block(cy_hdrs[0])
            if not cy_block:
                continue

            # --- PY rows: Pattern A (separate data rows) ---
            py_block = _rows_in_block(py_hdrs[-1]) if py_hdrs else []

            # --- PY rows: Pattern B (inline "Previous Year (PY)" row with
            #     its own data in the SAME row, col A = the label text) ---
            py_inline_rows = []
            if not py_block:
                for r in range(1, 60):
                    a = ws_cap.cell(r, 1).value
                    if isinstance(a, str) and "previous year" in a.strip().lower():
                        if any(isinstance(ws_cap.cell(r, c).value, (int, float))
                               for c in data_cols):
                            py_inline_rows.append(r)

            # --- Build (cy_row, py_row, name) pairs ---
            fixed_pairs = []
            if py_block:
                # Pattern A: match by partner name, fallback positional
                py_by_name = {name.strip().lower(): r for r, name in py_block}
                used_py_rows = set()
                pairs = []
                for cy_r, cy_name in cy_block:
                    py_r = py_by_name.get(cy_name.strip().lower())
                    pairs.append((cy_r, py_r, cy_name))
                    if py_r:
                        used_py_rows.add(py_r)
                unmatched_py = [r for r, _ in py_block if r not in used_py_rows]
                for cy_r, py_r, cy_name in pairs:
                    if py_r is None and unmatched_py:
                        py_r = unmatched_py.pop(0)
                    fixed_pairs.append((cy_r, py_r, cy_name))
            elif py_inline_rows:
                # Pattern B: single proprietor — one CY row, one inline PY row
                for i, (cy_r, cy_name) in enumerate(cy_block):
                    py_r = py_inline_rows[i] if i < len(py_inline_rows) else None
                    fixed_pairs.append((cy_r, py_r, cy_name))
            else:
                for cy_r, cy_name in cy_block:
                    fixed_pairs.append((cy_r, None, cy_name))

            # --- Read CY row values using detected column roles ---
            row_pairs = []
            for cy_r, py_r, cy_name in fixed_pairs:
                cy_vals = {}
                for col_num, role in col_roles.items():
                    v = ws_cap.cell(cy_r, col_num).value
                    if isinstance(v, (int, float)):
                        cy_vals[role] = float(v)
                row_pairs.append({
                    "cy_row": cy_r,
                    "py_row": py_r,
                    "name": cy_name,
                    "col_roles": col_roles,
                    "cy_opening":    cy_vals.get("opening"),
                    "cy_introduced": cy_vals.get("introduced"),
                    "cy_interest":   cy_vals.get("interest"),
                    "cy_salary":     cy_vals.get("salary"),
                    "cy_withdrawals":cy_vals.get("withdrawals"),
                    "cy_profit":     cy_vals.get("profit"),
                    "cy_closing":    cy_vals.get("closing"),
                })

            if row_pairs:
                cap_data[sn] = {"row_pairs": row_pairs}
    except Exception:
        pass

    return sizes, shift_map, cy_values, cy_formulas, cap_data


# ═══════════════════════════════════════════════════════════════════════════════
#  XML sheet manipulation  (regex on raw bytes — never ET.tostring)
# ═══════════════════════════════════════════════════════════════════════════════

def _process_sheet_xml(xml_bytes: bytes, col_pairs: list,
                       cy_vals: dict, cy_formulas: dict,
                       date_style_indices: set = None,
                       shift_map: dict = None) -> bytes:
    """
    Surgical byte-level edits on a worksheet XML:

    For every row:
      • PY cell (non-formula): replace its <v> with the CY value for that row.
      • CY cell (constant, i.e. NOT in cy_formulas): remove its <v>.

    Never calls ET.tostring — preserves all namespace prefixes exactly.
    """
    # Build lookup maps
    col_info  = {}   # cy_letter → (py_letter, {row: val}, formula_row_set)
    py_to_cy  = {}   # py_letter → cy_letter  (reverse map)
    for cy_l, py_l in col_pairs:
        col_info[cy_l] = (py_l,
                          cy_vals.get(cy_l, {}),
                          cy_formulas.get(cy_l, set()))
        py_to_cy[py_l] = cy_l

    cy_letters = set(col_info.keys())
    py_letters = set(py_to_cy.keys())
    date_style_indices = date_style_indices or set()
    shift_map = shift_map or {}

    # Build a quick lookup of "is this sheet!cell a PY-column cell that will
    # be correctly mirrored by its own shift?" — used below to decide whether
    # a cross-sheet formula in a PY cell is safe to keep.
    # shift_map: {sheet_name: [(cy_letter, py_letter), ...]}
    py_cols_by_sheet = {
        sn: {py_l for _, py_l in pairs} for sn, pairs in shift_map.items()
    }

    # Build the change dict: {cell_ref_str: ("set_v", new_val_str) | ("clear_v", None)}
    # We do this via a lightweight ET parse (read-only, no tostring ever called)
    changes = {}
    # style_changes: {cell_ref_str: new_style_index_str} — used to swap a PY
    # cell's date-formatted style for the CY cell's (number-formatted) style
    # when writing a numeric value into it (see DATE-FORMAT FIX above).
    style_changes = {}
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return xml_bytes

    sd = root.find(f"{{{_NS}}}sheetData")
    if sd is None:
        return xml_bytes

    for row_el in sd:
        for cell_el in row_el:
            ref = cell_el.get("r", "")
            m = _COL_RE.match(ref)
            if not m:
                continue
            cl, rn = m.group(1), int(m.group(2))

            if cl in py_letters:
                cy_l = py_to_cy[cl]
                py_l, vals, frows = col_info[cy_l]
                if rn in vals:
                    # DATE-FORMAT FIX: if this PY cell's current style is a
                    # date/time number format (often a leftover empty styled
                    # cell, e.g. <c r="E29" s="418"/> with numFmtId=168
                    # "[$-F800]dddd, mmmm dd, yyyy"), writing a plain number
                    # into it (e.g. 7080) would make Excel render it as a
                    # date ("1919-05-20"). Detect this and record the CY
                    # cell's style index so we can swap the PY cell's style
                    # to match — the CY cell's style is known-good since it
                    # was already displaying this number correctly.
                    py_s_val = cell_el.get("s")
                    style_fix = None
                    if (py_s_val is not None and date_style_indices
                            and int(py_s_val) in date_style_indices):
                        cy_ref_for_style = f"{cy_l}{rn}"
                        for _row_el in sd:
                            for _cell_el in _row_el:
                                if _cell_el.get("r") == cy_ref_for_style:
                                    cy_s_val = _cell_el.get("s")
                                    if cy_s_val is not None and int(cy_s_val) not in date_style_indices:
                                        style_fix = cy_s_val
                                    break
                            if style_fix is not None:
                                break

                    # If PY cell already has a formula (<f> tag), keep the formula
                    # intact and only update the cached <v> so it shows correctly
                    # until Excel recalculates. This preserves cross-sheet references
                    # like ='notes to bs'!E20 and =SUM(F12:F20) in BS/P&L PY columns.
                    py_f_el = cell_el.find(f"{{{_NS}}}f")
                    py_has_formula = (py_f_el is not None)

                    if py_has_formula and rn in frows:
                        # FORMULA-RANGE MISMATCH FIX:
                        # Both PY and CY cells are formulas (e.g. PY=SUM(E14:E14),
                        # CY=SUM(D14:D15)). After the shift, CY's range becomes the
                        # correct range for PY (just translated CY-col → PY-col).
                        # If we keep the old PY formula's range as-is, rows that were
                        # only in CY's range (e.g. "Round Off" row 15, or row 41/45
                        # in employee benefits) get dropped from the new PY total.
                        #
                        # Find the corresponding CY cell's <f> text in this same sheet
                        # and, if it's a simple range formula (SUM/range refs using
                        # only the cy_l column), translate cy_l -> py_l and use that
                        # as the new PY formula. Otherwise fall back to old behaviour.
                        cy_ref = f"{cy_l}{rn}"
                        cy_f_text = None
                        for _row_el in sd:
                            for _cell_el in _row_el:
                                if _cell_el.get("r") == cy_ref:
                                    _f = _cell_el.find(f"{{{_NS}}}f")
                                    if _f is not None and _f.text:
                                        cy_f_text = _f.text
                                    break
                            if cy_f_text is not None:
                                break

                        new_py_formula = None
                        if cy_f_text and "!" not in cy_f_text:
                            # Only translate if the CY formula references ONLY the
                            # cy_l column (e.g. SUM(D14:D15), D41+D42, etc.) — safe
                            # to do a column-letter substitution without breaking
                            # cross-sheet refs or multi-column formulas. The "!"
                            # check excludes cross-sheet refs like =Details!D19,
                            # where "D19" would otherwise be misread as a same-
                            # sheet column-D reference.
                            col_refs = set(re.findall(r'\b([A-Z]+)\d+\b', cy_f_text))
                            if col_refs == {cy_l}:
                                translated = re.sub(
                                    rf'\b{cy_l}(\d+)\b', rf'{py_l}\1', cy_f_text
                                )
                                new_py_formula = translated

                        if new_py_formula:
                            changes[ref] = ("set_f", (new_py_formula, _fmt_num(vals[rn])))
                        else:
                            changes[ref] = ("set_v_keep_f", _fmt_num(vals[rn]))
                    elif py_has_formula:
                        # CY cell is a plain constant (not a formula), but PY cell
                        # has its own formula — e.g. PY = '=44317.6-82.3' (a stale
                        # self-contained arithmetic formula from a prior year's
                        # "Rebate & Discount" entry), while CY = 8002 (this year's
                        # actual figure as a plain number).
                        #
                        # If we kept PY's old formula (set_v_keep_f), Excel would
                        # recalculate it on open, silently overwriting the correct
                        # value and unbalancing the sheet.
                        #
                        # Only keep the PY formula if it's a cross-sheet reference
                        # (e.g. ='notes to bs'!E20) AND the referenced cell is
                        # itself a PY column in a sheet that's part of this same
                        # shift (shift_map) — in that case, after both sheets are
                        # shifted, the referenced cell will correctly mirror the
                        # old CY data, just like the original formula intended
                        # (e.g. bs!E8 -> 'notes to bs'!E20, where 'notes to bs'
                        # also shifts D->E).
                        #
                        # Any other formula — self-contained arithmetic, SUM of
                        # its own column, or a cross-sheet ref to a column that
                        # ISN'T part of a coordinated PY shift (e.g. a stale
                        # ='GROSS PROFIT'!E10 reference in an "Other Expenses"
                        # line item) — gets overwritten with the CY constant,
                        # exactly like a plain value would be.
                        py_f_text = py_f_el.text if py_f_el is not None else None
                        keep_formula = False
                        if py_f_text and "!" in py_f_text:
                            m_ref = re.match(
                                r"^=?\s*(?:'([^']+)'|([A-Za-z0-9_ ]+))!\$?([A-Z]+)\$?\d+\s*$",
                                py_f_text.strip()
                            )
                            if m_ref:
                                ref_sheet = (m_ref.group(1) or m_ref.group(2)).strip()
                                ref_col = m_ref.group(3)
                                ref_py_cols = py_cols_by_sheet.get(ref_sheet, set())
                                if ref_col in ref_py_cols:
                                    keep_formula = True
                        if keep_formula:
                            changes[ref] = ("set_v_keep_f", _fmt_num(vals[rn]))
                        else:
                            changes[ref] = ("set_v_overwrite", _fmt_num(vals[rn]))
                    else:
                        changes[ref] = ("set_v_overwrite", _fmt_num(vals[rn]))

                    if style_fix is not None:
                        style_changes[ref] = style_fix
                else:
                    # BUG 6 FIX: Only clear PY cell when CY cell is TRULY empty
                    # (has no <v> tag at all). Skip if CY has a string value
                    # (shared string ref like date headers) — those get updated
                    # by _update_inline_strings separately.
                    cy_ref = f"{cy_l}{rn}"
                    cy_cell_el = None
                    for _row_el in sd:
                        for _cell_el in _row_el:
                            if _cell_el.get("r") == cy_ref:
                                cy_cell_el = _cell_el
                                break
                        if cy_cell_el is not None: break
                    # Only clear PY if CY cell has NO value at all
                    cy_has_any_value = (
                        cy_cell_el is not None and
                        cy_cell_el.find(f"{{{_NS}}}v") is not None
                    )
                    if not cy_has_any_value:
                        changes[ref] = ("clear_v", None)

            if cl in cy_letters:
                _, vals, frows = col_info[cl]
                if rn in vals:
                    if rn in frows:
                        # Formula cell: clear cached <v> but KEEP <f> formula intact
                        # Excel will recalculate when user opens and enters new data
                        changes[ref] = ("clear_v_keep_f", None)
                    else:
                        # Constant cell: remove <v> entirely
                        changes[ref] = ("clear_v", None)

    # BUG 1 FIX: Build insertions dict for PY cells that don't exist in XML
    # These are CY values where there's no existing PY <c> element to overwrite
    insertions = {}  # {row_num: {py_letter: val_str}}
    existing_py_refs = set()
    for row_el in sd:
        for cell_el in row_el:
            ref = cell_el.get("r", "")
            m2 = _COL_RE.match(ref)
            if m2 and m2.group(1) in py_letters:
                existing_py_refs.add(ref)

    for cy_l, (py_l, vals, frows) in col_info.items():
        for rn, val in vals.items():
            py_ref = f"{py_l}{rn}"
            if py_ref not in existing_py_refs:
                # PY cell doesn't exist — need to insert it
                insertions.setdefault(rn, {})[py_l] = _fmt_num(val)

    if not changes and not insertions:
        return xml_bytes

    # Apply changes via regex on the raw text — row by row for safety
    text = xml_bytes.decode("utf-8", errors="replace")

    def _fix_cell(cm):
        full = cm.group(0)
        ref_m = re.search(r'\br="([A-Z]+\d+)"', full)
        if not ref_m:
            return full
        ref = ref_m.group(1)
        if ref not in changes:
            return full
        action, new_val = changes[ref]
        if action == "clear_v":
            # Remove both formula <f> and cached value <v> so cell is truly blank
            # (keeping <f> would cause formula to recalculate, showing stale data)
            full = re.sub(r'<f[^>]*>.*?</f>', '', full, flags=re.DOTALL)
            full = re.sub(r'<f[^>]*/>', '', full)
            full = re.sub(r'<v>[^<]*</v>', '', full)
            full = re.sub(r'<v\s*/>', '', full)
        elif action == "clear_v_keep_f":
            # Keep <f> formula tag, just clear the cached <v> value
            # Excel recalculates formula when file is opened
            full = re.sub(r'<v>[^<]*</v>', '', full)
            full = re.sub(r'<v\s*/>', '', full)
        elif action == "set_f":
            # Replace the cell's formula entirely with a new formula (translated
            # from the CY cell's range formula), and set the cached <v> to the
            # CY-computed total so it displays correctly until recalculation.
            new_formula, new_val = new_val
            # Escape XML special chars in the formula text
            esc_formula = (new_formula.replace("&", "&amp;")
                                       .replace("<", "&lt;")
                                       .replace(">", "&gt;"))
            if full.rstrip().endswith('/>'):
                full = re.sub(r'/>\s*$', '></c>', full.rstrip())
            # Replace existing <f>...</f> or <f .../> with the new formula
            if re.search(r'<f\b[^>]*>.*?</f>', full, flags=re.DOTALL):
                full = re.sub(r'<f\b[^>]*>.*?</f>', f'<f>{esc_formula}</f>',
                               full, flags=re.DOTALL)
            elif re.search(r'<f\b[^>]*/>', full):
                full = re.sub(r'<f\b[^>]*/>', f'<f>{esc_formula}</f>', full)
            else:
                full = full.replace('</c>', f'<f>{esc_formula}</f></c>')
            # Update or insert the cached <v>
            if '<v>' in full:
                full = re.sub(r'<v>[^<]*</v>', f'<v>{new_val}</v>', full)
            else:
                full = full.replace('</c>', f'<v>{new_val}</v></c>')
            full = re.sub(r'\s*t="s"', '', full)
        elif action in ("set_v", "set_v_overwrite", "set_v_keep_f"):
            if action == "set_v_overwrite":
                # Remove any <f>...</f> formula tag first (convert formula → value)
                full = re.sub(r'<f\b[^>]*>.*?</f>', '', full, flags=re.DOTALL)
                full = re.sub(r'<f\b[^>]*/>', '', full)
            # set_v_keep_f: keep formula tag, just update cached <v> value
            # (for PY cells that have cross-sheet formulas like ='notes to bs'!E20)
            # BUG 1 FIX: self-closing empty cells (<c r="E16" s="814"/>) end with />
            # not </c>, so replace('</c>', ...) silently fails. Convert to paired tag.
            if full.rstrip().endswith('/>'):
                full = re.sub(r'/>\s*$', '></c>', full.rstrip())
            if '<v>' in full:
                full = re.sub(r'<v>[^<]*</v>', f'<v>{new_val}</v>', full)
            else:
                full = full.replace('</c>', f'<v>{new_val}</v></c>')
            # Remove string-type attribute — it's a number now (only for overwrite)
            if action != "set_v_keep_f":
                full = re.sub(r'\s*t="s"', '', full)

        # DATE-FORMAT FIX: if this cell's old style is a date/time format and
        # we recorded a replacement style (from the corresponding CY cell),
        # swap the s="..." attribute so the new numeric value doesn't render
        # as a date (e.g. 7080 -> "1919-05-20").
        if ref in style_changes:
            new_style = style_changes[ref]
            if re.search(r'\bs="\d+"', full):
                full = re.sub(r'\bs="\d+"', f's="{new_style}"', full, count=1)
            else:
                # No existing s= attribute — add one right after the r="..." ref
                full = re.sub(r'(\br="[A-Z]+\d+")', rf'\1 s="{new_style}"', full, count=1)

        return full

    def _fix_row(rm):
        row_xml = rm.group(0)
        # Match self-closing cells first, then paired cells
        row_xml = re.sub(
            r'<c\b[^>]*/>\s*|<c\b[^>]*>.*?</c>\s*',
            _fix_cell, row_xml, flags=re.DOTALL
        )
        # BUG 1 FIX: Insert new PY cells for rows where PY cell doesn't exist in XML
        # Get current row number from the row element
        row_num_m = re.search(r'<row\b[^>]*\br="(\d+)"', row_xml)
        if row_num_m:
            rn = int(row_num_m.group(1))
            if rn in insertions:
                # For each PY letter that needs a new cell in this row
                existing_refs = set(re.findall(r'r="([A-Z]+\d+)"', row_xml))
                new_cells = ""
                for py_l, val_str in sorted(insertions[rn].items(),
                                            key=lambda x: _col_idx(x[0])):
                    ref = f"{py_l}{rn}"
                    if ref not in existing_refs:
                        new_cells += f'<c r="{ref}"><v>{val_str}</v></c>'
                if new_cells:
                    row_xml = row_xml.replace("</row>", new_cells + "</row>")
        return row_xml

    text = re.sub(r'<row\b[^>]*>.*?</row>', _fix_row, text, flags=re.DOTALL)
    return text.encode("utf-8")


def _fmt_num(v: float) -> str:
    """Format a float for XML: drop .0 suffix for whole numbers."""
    if v == int(v):
        return str(int(v))
    return repr(v)  # repr gives enough precision without scientific notation

def _fmt_num_for_py(v: float) -> str:
    """Format CY value for writing to PY cell. Zero stays as zero (not dash) 
    because PY column may have its own dash formatting via cell format."""
    return _fmt_num(v)


def _process_capital_sheet(xml_bytes: bytes, cap_data: dict) -> bytes:
    """
    Special handler for capital sheet (Bugs 4 & 5 — multi-partner aware):

    cap_data["row_pairs"] is a list of dicts, one per partner/proprietor:
      cy_row, py_row        — row numbers (py_row may be None)
      cy_opening..cy_closing — this partner's CY-row values (cols C-I) as read
                               from the ORIGINAL file, before any shift

    For each pair, after the year shift:
      1. PY row (py_row): becomes a mirror of the OLD CY row's data —
         overwrite cols C-I (opening, introduced, interest, salary,
         withdrawals, profit, closing) with the recorded cy_* constants.
         This is "last year's actuals" for the new PY column.
      2. CY row (cy_row): becomes the new year's starting point —
         opening (col C) = old CY closing (cy_closing), as a constant.
         Introduced/interest/salary/withdrawals (cols D-G) are cleared
         (new year, no entries yet). Profit (col H) and closing (col I)
         keep their formulas so they auto-recalculate from the shifted
         p&l sheet and the new opening balance.
    """
    if not cap_data:
        return xml_bytes

    row_pairs = cap_data.get("row_pairs") or []
    if not row_pairs:
        return xml_bytes

    # Build changes dict: {cell_ref: ("set_v_overwrite", val) | ("clear_fv", None)}
    changes = {}

    for pair in row_pairs:
        cy_row = pair.get("cy_row")
        py_row = pair.get("py_row")
        col_roles = pair.get("col_roles") or {}
        # role -> column letter, for this template's actual layout
        role_to_col = {role: get_column_letter(c) for c, role in col_roles.items()}

        if py_row:
            # PY row becomes a full mirror of the old CY row's data, using
            # this template's own column layout (role_to_col).
            for role, col_letter in role_to_col.items():
                ref = f"{col_letter}{py_row}"
                val = pair.get(f"cy_{role}")
                if val is not None:
                    changes[ref] = ("set_v_overwrite", _fmt_num(float(val)))
                else:
                    changes[ref] = ("clear_v", None)

        if cy_row:
            # New CY row: opening = old CY closing (constant), using
            # whichever column this template uses for "opening".
            cy_closing = pair.get("cy_closing")
            opening_col = role_to_col.get("opening")
            if cy_closing is not None and opening_col:
                changes[f"{opening_col}{cy_row}"] = ("set_v_overwrite", _fmt_num(float(cy_closing)))
            # Clear introduced/interest/salary/withdrawals — new year has no
            # entries yet. Use clear_fv to remove both formula AND cached
            # value (some templates use formulas like =75242.8+884000).
            for role in ("introduced", "interest", "salary", "withdrawals"):
                col_letter = role_to_col.get(role)
                if col_letter:
                    ref = f"{col_letter}{cy_row}"
                    changes[ref] = ("clear_fv", None)
            # Profit and closing columns keep their formulas — they
            # auto-recalculate from the shifted p&l sheet and the new
            # opening balance, so we leave them untouched.

    if not changes:
        return xml_bytes



    text = xml_bytes.decode("utf-8", errors="replace")

    def _fix_cap_cell(cm):
        full = cm.group(0)
        ref_m = re.search(r'\br="([A-Z]+\d+)"', full)
        if not ref_m:
            return full
        ref = ref_m.group(1)
        if ref not in changes:
            return full
        action, new_val = changes[ref]
        if action == "clear_v":
            full = re.sub(r'<v>[^<]*</v>', '', full)
            full = re.sub(r'<v\s*/>', '', full)
        elif action == "clear_fv":
            # Remove BOTH formula and cached value (for CY intro/withdrawals)
            full = re.sub(r'<f\b[^>]*>.*?</f>', '', full, flags=re.DOTALL)
            full = re.sub(r'<f\b[^>]*/>', '', full)
            full = re.sub(r'<v>[^<]*</v>', '', full)
            full = re.sub(r'<v\s*/>', '', full)
        elif action == "set_v_overwrite":
            full = re.sub(r'<f\b[^>]*>.*?</f>', '', full, flags=re.DOTALL)
            full = re.sub(r'<f\b[^>]*/>', '', full)
            if '<v>' in full:
                full = re.sub(r'<v>[^<]*</v>', f'<v>{new_val}</v>', full)
            else:
                full = full.replace('</c>', f'<v>{new_val}</v></c>')
            full = re.sub(r'\s*t="s"', '', full)
        return full

    def _fix_cap_row(rm):
        row_xml = rm.group(0)
        row_xml = re.sub(
            r'<c\b[^>]*/> *|<c\b[^>]*>.*?</c> *',
            _fix_cap_cell, row_xml, flags=re.DOTALL
        )
        # Insert new cells for refs in changes that didn't exist
        row_num_m = re.search(r'<row\b[^>]*\br="(\d+)"', row_xml)
        if row_num_m:
            rn = int(row_num_m.group(1))
            existing = set(re.findall(r'r="([A-Z]+\d+)"', row_xml))
            new_cells = ""
            for ref, (action, val) in changes.items():
                r_m = re.match(r'([A-Z]+)(\d+)', ref)
                if r_m and int(r_m.group(2)) == rn and ref not in existing:
                    if action == "set_v_overwrite" and val:
                        new_cells += f'<c r="{ref}"><v>{val}</v></c>'
            if new_cells:
                row_xml = row_xml.replace("</row>", new_cells + "</row>")
        return row_xml

    text = re.sub(r'<row\b[^>]*>.*?</row>', _fix_cap_row, text, flags=re.DOTALL)
    return text.encode("utf-8")


# ═══════════════════════════════════════════════════════════════════════════════
#  Shared-strings date replacement
# ═══════════════════════════════════════════════════════════════════════════════

def _update_shared_strings(xml_bytes: bytes, pairs: list) -> bytes:
    """
    Replace date strings in sharedStrings.xml.
    Handles plain <t> cells and rich-text <r><t> runs (superscript ordinals etc.).
    Uses lxml when available (namespace-safe); falls back to stdlib ET.
    sharedStrings.xml itself doesn't use xl: / r: prefixes in its content,
    so stdlib ET is safe here.
    """
    try:
        from lxml import etree as letree
        root = letree.fromstring(xml_bytes)
        NS_L = "{" + _NS + "}"
        changed = False
        for si in root:
            t_els = si.findall(f".//{NS_L}t")
            if not t_els:
                continue
            if len(t_els) == 1:
                old = t_els[0].text or ""
                new = _apply_pairs(old, pairs)
                if new != old:
                    t_els[0].text = new
                    changed = True
            else:
                originals = [t.text or "" for t in t_els]
                full_new = _apply_pairs("".join(originals), pairs)
                if full_new != "".join(originals):
                    changed = True
                    _redistribute_rich_text(t_els, originals, full_new)
        if not changed:
            return xml_bytes
        return letree.tostring(root, xml_declaration=True,
                               encoding="UTF-8", standalone=True)
    except ImportError:
        pass

    # stdlib ET fallback
    root = ET.fromstring(xml_bytes)
    changed = False
    for si in root:
        t_els = list(si.iter(f"{{{_NS}}}t"))
        if not t_els:
            continue
        if len(t_els) == 1:
            old = t_els[0].text or ""
            new = _apply_pairs(old, pairs)
            if new != old:
                t_els[0].text = new
                changed = True
        else:
            originals = [t.text or "" for t in t_els]
            full_new = _apply_pairs("".join(originals), pairs)
            if full_new != "".join(originals):
                changed = True
                _redistribute_rich_text(t_els, originals, full_new)
    if not changed:
        return xml_bytes
    ET.register_namespace("", _NS)
    return ET.tostring(root, xml_declaration=True, encoding="UTF-8")


def _redistribute_rich_text(t_els, originals: list, full_new: str):
    """Redistribute replaced text back into rich-text <t> runs."""
    full_old = "".join(originals)
    if len(full_new) == len(full_old):
        pos = 0
        for i, t_el in enumerate(t_els):
            run_len = len(originals[i])
            t_el.text = full_new[pos:pos + run_len]
            pos += run_len
        return

    # Find common prefix / suffix, assign changed middle to overlapping runs
    pfx = 0
    while pfx < len(full_old) and pfx < len(full_new) and full_old[pfx] == full_new[pfx]:
        pfx += 1
    sfx = 0
    while (sfx < len(full_old) - pfx and sfx < len(full_new) - pfx
           and full_old[-(sfx + 1)] == full_new[-(sfx + 1)]):
        sfx += 1

    change_start   = pfx
    change_end_old = len(full_old) - sfx
    change_end_new = len(full_new) - sfx
    new_middle     = full_new[change_start:change_end_new]

    pos = 0
    new_texts = []
    middle_assigned = False
    for orig in originals:
        rstart, rend = pos, pos + len(orig)
        if rend <= change_start:
            new_texts.append(orig)
        elif rstart >= change_end_old:
            new_texts.append(orig)
        else:
            before = orig[:max(0, change_start - rstart)]
            after  = orig[max(0, change_end_old - rstart):]
            if not middle_assigned:
                new_texts.append(before + new_middle + after)
                middle_assigned = True
            else:
                new_texts.append(after)
        pos = rend

    for i, t_el in enumerate(t_els):
        t_el.text = new_texts[i] if i < len(new_texts) else ""


def _update_inline_strings(xml_bytes: bytes, pairs: list) -> bytes:
    """Replace date strings in worksheet inline strings (raw text substitution)."""
    text = xml_bytes.decode("utf-8", errors="replace")
    new_text = _apply_pairs(text, pairs)
    if new_text == text:
        return xml_bytes
    return new_text.encode("utf-8")


# ═══════════════════════════════════════════════════════════════════════════════
#  External reference cleaner
# ═══════════════════════════════════════════════════════════════════════════════

# Patterns that indicate a formula references an external workbook or DDE link.
# These will break (#REF!) when the file is opened on a different computer.
_EXT_REF_RE = re.compile(
    rb'\|'              # DDE link separator (=Excel.Sheet.8|'\\server\...')
    rb'|\[\d'           # External link ref like [1]Sheet!A1
    rb'|\\\\',          # UNC network path \\server\share
)

def _strip_external_formulas(xml_bytes: bytes) -> bytes:
    """Convert external-reference formulas to their cached values.

    For every <c> cell that contains a <f> formula matching an external
    reference pattern, remove the <f>...</f> tag but keep the <v> cached
    value.  This prevents #REF! errors when the file is opened on a
    different computer where the referenced file isn't accessible.
    """
    # Quick check: if no external-looking formulas exist, skip
    if not _EXT_REF_RE.search(xml_bytes):
        return xml_bytes

    text = xml_bytes.decode("utf-8", errors="replace")
    count = 0

    def _clean_cell(cm):
        nonlocal count
        cell_xml = cm.group(0)
        # Check if cell has a <f> tag with external ref
        f_match = re.search(r'<f[^>]*>(.*?)</f>', cell_xml, re.DOTALL)
        if not f_match:
            # Also check self-closing <f ... />
            f_match = re.search(r'<f[^>]*/>', cell_xml)
        if not f_match:
            return cell_xml

        formula_content = f_match.group(0).encode("utf-8", errors="replace")
        if not _EXT_REF_RE.search(formula_content):
            return cell_xml

        # This formula has an external reference — remove <f> but keep <v>
        cell_xml = re.sub(r'<f[^>]*>.*?</f>', '', cell_xml, flags=re.DOTALL)
        cell_xml = re.sub(r'<f[^>]*/>', '', cell_xml)
        count += 1
        return cell_xml

    text = re.sub(
        r'<c\b[^>]*/>\s*|<c\b[^>]*>.*?</c>\s*',
        _clean_cell, text, flags=re.DOTALL
    )

    return text.encode("utf-8") if count else xml_bytes


# ═══════════════════════════════════════════════════════════════════════════════
#  Main entry point
# ═══════════════════════════════════════════════════════════════════════════════

def process(input_path: str, output_path: str,
            closing_year: int, new_year: int) -> dict:
    pairs = _date_replacements(str(closing_year), str(new_year))
    log   = []

    # ── Single-pass scan ──────────────────────────────────────────────────────
    sizes, shift_map, cy_values, cy_formulas, cap_data = _scan_workbook(input_path, closing_year)
    big_names = {n for n, (r, c) in sizes.items() if r > BIG_ROWS or c > BIG_COLS}

    # Pre-compute which cellXfs style indices use a date/time number format,
    # so PY cells with leftover date-formatted (but empty) styles don't render
    # newly-written numeric values as dates (see DATE-FORMAT FIX).
    date_style_indices = _build_date_style_set(input_path)

    ext_count = 0   # track external refs cleaned

    # ── ZIP-level edit loop ───────────────────────────────────────────────────
    with zipfile.ZipFile(input_path, "r") as zi:
        smap = _sheet_file_map(zi)

        with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zo:
            for item in zi.infolist():
                fn   = item.filename
                data = zi.read(fn)

                # Skip externalLinks — they reference files on other machines
                if "externalLinks" in fn:
                    continue

                # shared strings — date text replacement only
                if fn == "xl/sharedStrings.xml":
                    data = _update_shared_strings(data, pairs)
                    zo.writestr(item, data)
                    continue

                # worksheet XMLs
                if fn.startswith("xl/worksheets/") and fn.endswith(".xml"):
                    sheet_name = next(
                        (sn for sn, sf in smap.items() if sf == fn), None
                    )
                    sl = (sheet_name or "").strip().lower()

                    # Always strip external formulas from all sheets
                    before_len = len(data)
                    data = _strip_external_formulas(data)
                    if len(data) != before_len:
                        ext_count += 1

                    if sheet_name and sheet_name in shift_map:
                        # Full CY→PY shift + date update
                        data = _process_sheet_xml(
                            data,
                            shift_map[sheet_name],
                            cy_values.get(sheet_name, {}),
                            cy_formulas.get(sheet_name, {}),
                            date_style_indices,
                            shift_map,
                        )
                        data = _update_inline_strings(data, pairs)
                        desc = ", ".join(f"{c}→{p}" for c, p in shift_map[sheet_name])
                        log.append(
                            f"✓ {sheet_name}: CY→PY copied ({desc}), "
                            f"CY constants cleared, dates updated"
                        )
                    elif sheet_name and sl in RAW_DATA_SHEETS:
                        # Don't touch raw data sheets at all
                        log.append(f"— {sheet_name}: skipped (raw data sheet)")
                    elif sheet_name and sheet_name.strip().lower() in CAPITAL_SHEET_NAMES:
                        # Capital sheet: special CY/PY row shift (Bugs 4 & 5)
                        cap_info = cap_data.get(sheet_name, {})
                        if cap_info:
                            data = _process_capital_sheet(data, cap_info)
                            log.append(
                                f"✓ {sheet_name}: capital CY→PY shifted, "
                                f"CY opening updated, additions/withdrawals cleared"
                            )
                        data = _update_inline_strings(data, pairs)
                    elif sheet_name and sheet_name not in big_names:
                        # Small sheet with no CY/PY columns — just update dates
                        data = _update_inline_strings(data, pairs)
                        log.append(f"· {sheet_name}: dates updated")
                    elif sheet_name in big_names:
                        log.append(
                            f"* {sheet_name}: preserved unchanged (large sheet "
                            f"{sizes[sheet_name][0]} rows)"
                        )

                    zo.writestr(item, data)
                    continue

                # workbook.xml.rels — remove references to external links
                if fn == "xl/workbook.xml.rels" or fn.endswith(".rels"):
                    text_rels = data.decode("utf-8", errors="replace")
                    if "externalLinks" in text_rels:
                        text_rels = re.sub(
                            r'<Relationship[^>]*Target="externalLinks[^"]*"[^>]*/>\s*',
                            '', text_rels
                        )
                        data = text_rels.encode("utf-8")

                # workbook.xml — remove <externalReferences> section
                if fn == "xl/workbook.xml":
                    text_wb = data.decode("utf-8", errors="replace")
                    if "externalReference" in text_wb:
                        text_wb = re.sub(
                            r'<externalReferences>.*?</externalReferences>\s*',
                            '', text_wb, flags=re.DOTALL
                        )
                        data = text_wb.encode("utf-8")

                # [Content_Types].xml — remove external link content types
                if fn == "[Content_Types].xml":
                    text_ct = data.decode("utf-8", errors="replace")
                    if "externalLink" in text_ct:
                        text_ct = re.sub(
                            r'<Override[^>]*externalLink[^>]*/>\s*',
                            '', text_ct
                        )
                        data = text_ct.encode("utf-8")

                # everything else — pass through byte-perfect
                zo.writestr(item, data)

    if ext_count:
        log.append(f"🔗 External references converted to values in {ext_count} sheet(s)")

    return {"status": "success", "log": log, "output": output_path}


if __name__ == "__main__":
    import sys
    if len(sys.argv) != 5:
        print("Usage: python processor.py input.xlsx output.xlsx closing_year new_year")
        sys.exit(1)
    result = process(sys.argv[1], sys.argv[2], int(sys.argv[3]), int(sys.argv[4]))
    for line in result["log"]:
        print(line)
    print(f"\nSaved → {result['output']}")
