"""
Trial Balance → Balance Sheet Processor
Reads a trial balance, classifies accounts under BS/P&L heads,
and injects aggregated values into a BS template.
Zero formatting change in the output BS file.
Memory-efficient: single workbook open, read_only where possible.

v2026-07-16 comprehensive fixes — Details sheet injection:

  CLASSIFICATION FIXES
  ────────────────────
  Fix A — Group name aliases: added "CUSTOMERS", "SUPPLIERS", "PAYABLES",
    "OTHERS ASSETS" and Tally "TRADING/DIRECT EXPENSES/INCOMES",
    "PROFIT & LOSS EXPENSES/INCOMES" etc. to GROUP_HEAD_MAP and the
    inline group-level classifier.  JEANS WORLD TB uses these names;
    without them accounts fell through to low-confidence / unclassified.

  CREDITOR SECTION FIXES
  ──────────────────────
  Fix B — "Advance from Customers" sub-section INSIDE creditor SUM:
    JEANS WORLD / SHREE CRAFT templates put the advance-from-cust sub-rows
    INSIDE the =SUM(D21:D55) creditor block. The old code stopped the
    cred_end_row scan at the "Advance from Customers" header (treating it
    as a section boundary), so those rows were excluded from the search
    range. New: scan continues through mixed-case sub-headers; only stops
    at a new TOP-LEVEL (all-caps, no amount) header or the SUM/TOTAL row.
    cred_total_row is now tracked for the insert_rows fallback.

  Fix C — New creditor overflow uses insert_rows before TOTAL:
    Old code searched for a random blank row anywhere on the sheet when
    the creditor section was full, silently placing amounts outside the
    SUM formula. New: inserts a row immediately before the creditor TOTAL
    row, rewrites the SUM formula to include the new row, and updates all
    cross-sheet formulas that reference Details rows at/below the insert.

  ADVANCE FROM CUSTOMERS FIXES
  ─────────────────────────────
  Fix D — Detect layout variant (inside vs outside creditor SUM):
    Added _afc_inside_creditor flag. When the section is inside the
    creditor block, _afc_end = cred_end_row−1 so name-matching searches
    all adv-from-cust pre-existing rows. Insert fallback added (insert
    before afc total row + update SUM).

  ADVANCE TO SUPPLIERS FIXES (also includes yesterday's Fashion Adda fix)
  ────────────────────────────────────────────────────────────────────────
  Fix E — Scan starts from row 1 (not _recv_total_row): header can sit
    before the recv section's TOTAL row.
  Fix F — Broken regex-as-string: "adv.*supp" tested with `in` is a
    literal 8-char substring check, never a regex match. Replaced.
  Fix G — Insert fallback: inserts before TOTAL + updates SUM formula.
  Fix H — _adv_inside_recv flag: when header is inside the recv SUM
    range, amounts already count toward the recv total (no separate insert).
"""

import re
import os
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from copy import copy


# ═══════════════════════════════════════════════════════════════════════
# ACCOUNT CLASSIFICATION RULES
# ═══════════════════════════════════════════════════════════════════════

BS_HEADS = {
    # --- LIABILITIES ---
    "capital": {
        "label": "Owner's Capital / Partners Capital",
        "side": "liability",
        "keywords": [
            "capital", "partner", "proprietor", "owner", "equity",
            "share capital", "reserves", "surplus", "retained earning",
            "general reserve", "capital reserve", "securities premium",
            "profit & loss", "profit and loss", "p&l appropriation",
            "current account", "drawing", "partner current",
            "capital account", "share application",
        ],
        "negative_keywords": ["capital gain", "capital goods", "working capital loan"],
    },
    "lt_borrowings": {
        "label": "Long Term Borrowings",
        "side": "liability",
        "keywords": [
            "long term loan", "term loan", "secured loan", "unsecured loan",
            "mortgage", "debenture", "long term borrowing",
            "loan from bank", "loan from director", "loan from partner",
            "vehicle loan", "car loan", "home loan", "housing loan",
            "hypothecation", "loan payable",
        ],
        "negative_keywords": ["short term", "od", "overdraft", "cc limit"],
    },
    "st_borrowings": {
        "label": "Short Term Borrowings",
        "side": "liability",
        "keywords": [
            "short term loan", "overdraft", "od account", "cash credit",
            "cc limit", "cc account", "working capital", "packing credit",
            "bank od", "bank overdraft", "short term borrowing",
        ],
        "negative_keywords": [],
    },
    "trade_payables": {
        "label": "Trade Payables",
        "side": "liability",
        "keywords": [
            "trade payable", "creditor", "sundry creditor",
            "accounts payable", "supplier", "purchase payable",
            "bills payable", "trade creditor",
        ],
        "negative_keywords": [],
    },
    "advance_from_customer": {
        # Credit-balance debtor → advance received from customer (liability)
        # Shown under Sundry Creditors group as "Advance from Customer".
        "label": "Advance from Customer",
        "side": "liability",
        "keywords": [
            "advance from customer", "customer advance",
            "advance received", "advance from buyer",
            "advance from m/s", "advance from mr", "advance from ms",
            "advance from ",  # generic fallback — catches "Advance from XYZ Ltd"
        ],
        "negative_keywords": [
            "advance from supplier", "advance from vendor",
        ],
    },
    "other_cl": {
        "label": "Other Current Liabilities",
        "side": "liability",
        "keywords": [
            "other current liabilit", "statutory", "tds payable",
            "gst payable", "gst output", "output cgst", "output sgst",
            "output igst", "tax payable", "duty payable", "cess payable",
            "salary payable", "wages payable", "rent payable",
            "interest payable", "expense payable", "outstanding",
            "advance from customer", "customer advance", "security deposit received",
            "audit fee payable", "professional fee payable",
            "electricity payable", "telephone payable",
            "provision for expense", "payable",
            "income received in advance", "unearned",
        ],
        "negative_keywords": ["provision for tax", "provision for depreciation", "provision for bad debt"],
    },
    "st_provisions": {
        "label": "Short Term Provisions",
        "side": "liability",
        "keywords": [
            "provision for tax", "provision for income tax",
            "provision for depreciation", "provision for bad debt",
            "provision for doubtful", "short term provision",
            "provision for gratuity", "provision for bonus",
            "provision for leave", "provision for warranty",
        ],
        "negative_keywords": [],
    },

    # --- ASSETS ---
    "fixed_assets": {
        "label": "Fixed Assets / PPE",
        "side": "asset",
        "keywords": [
            "fixed asset", "property", "plant", "equipment", "ppe",
            "land", "building", "furniture", "fixture", "vehicle",
            "computer", "machinery", "office equipment", "electrical",
            "air condition", "ac ", "motor car", "scooter", "bike",
            "mobile", "telephone instrument", "printer", "laptop",
            "intangible", "goodwill", "patent", "trademark", "copyright",
            "software", "leasehold", "capital wip", "cwip",
        ],
        "negative_keywords": [
            "depreciation", "accumulated", "provision for",
            "repair", "maintenance", "rent",
        ],
    },
    "non_current_investments": {
        "label": "Non-Current Investments",
        "side": "asset",
        "keywords": [
            "investment", "shares", "debenture held", "mutual fund",
            "fixed deposit", "fdr", "fd ", "nsc ", "kvp",
            "government securities", "bonds",
            "investment in subsidiary", "investment in associate",
        ],
        "negative_keywords": ["provision for investment"],
    },
    "inventories": {
        "label": "Inventories / Stock",
        "side": "asset",
        "keywords": [
            "inventor", "stock", "closing stock", "opening stock",
            "raw material", "finished good", "work in progress",
            "wip", "stores", "spare", "packing material",
            "stock in trade", "goods in transit",
        ],
        "negative_keywords": ["stock broker"],
    },
    "trade_rec": {
        "label": "Trade Receivables",
        "side": "asset",
        "keywords": [
            "trade receivable", "debtor", "sundry debtor",
            "accounts receivable", "bills receivable",
            "trade debtor", "receivable from customer",
        ],
        "negative_keywords": [],
    },
    "cash_bank": {
        "label": "Cash and Bank Balances",
        "side": "asset",
        "keywords": [
            "cash", "bank", "cash in hand", "cash at bank",
            "petty cash", "savings account", "current account bank",
            "bank account", "bank balance", "cheque in hand",
            "imprest",
        ],
        "negative_keywords": ["cash credit", "cc account", "overdraft", "od account", "bank od",
                              "interest", "loan interest", "loan a/c", "loan account",
                              "bank charge", "processing fee"],
    },
    "stla": {
        "label": "Short Term Loans & Advances",
        "side": "asset",
        "keywords": [
            "loan given", "advance to", "loan to",
            "advance to staff", "staff advance",
            "prepaid", "deposit", "security deposit paid",
            "tds receivable", "tcs receivable", "tcs collection", "tcs ",
            "input tax",
            "input cgst", "input sgst", "input igst", "gst input",
            "advance tax", "self assessment tax", "mat credit",
            "cenvat", "vat input", "excise input",
            "income tax refund", "refund receivable",
            "advance recoverable",
        ],
        "negative_keywords": ["advance from customer", "customer advance",
                              "advance to supplier", "advance to customer",
                              "supplier advance"],
    },
    "advance_to_supplier": {
        # Debit-balance creditor → advance paid to supplier (asset)
        # Listed under Sundry Debtors group as "Advance to Supplier / Customer".
        "label": "Advance to Supplier",
        "side": "asset",
        "keywords": [
            "advance to supplier", "advance to customer", "supplier advance",
            "advance paid to supplier", "advance to vendor", "vendor advance",
        ],
        "negative_keywords": ["advance from customer", "customer advance"],
    },
    "other_current_assets": {
        "label": "Other Current Assets",
        "side": "asset",
        "keywords": [
            "other current asset", "accrued income", "interest accrued",
            "accrued interest", "interest receivable",
            "other receivable", "other asset",
        ],
        "negative_keywords": [],
    },

    # --- P&L HEADS ---
    "revenue": {
        "label": "Revenue from Operations / Sales",
        "side": "pl",
        "keywords": [
            "sale", "revenue", "income from operation",
            "turnover", "gross receipt", "service income",
            "service revenue", "consulting income", "fee received",
            "commission received", "commission income", "commission recd",
            "export sale", "domestic sale", "local sale",
            # Professional-services income (CA/consulting firms etc.) —
            # "Professional Income (...)", "Professional Charges" are the
            # primary revenue line items in such templates' "Revenue from
            # operations" note.
            "professional income", "professional charges", "professional fee",
            "professional fees", "consultancy charges", "consultancy fee",
        ],
        "negative_keywords": ["sale return", "sales return", "sale of asset", "sale of investment"],
    },
    "other_income": {
        "label": "Other Income",
        "side": "pl",
        "keywords": [
            "other income", "interest received", "interest income",
            "dividend received", "dividend income", "rental income",
            "rent received", "miscellaneous income", "misc income",
            "profit on sale", "gain on sale", "exchange gain",
            "discount received", "discount earned", "commission earned",
            "bad debt recovered", "insurance claim received",
            "interest on fd", "interest on deposit", "bank interest received",
            "bank interest recd", "interest recd",
            # Pension/annuity income received by the proprietor
            "pension received", "pension recd",
        ],
        "negative_keywords": [],
    },
    "purchases": {
        "label": "Purchases / Cost of Material",
        "side": "pl",
        "keywords": [
            "purchase", "cost of material", "cost of goods",
            "import purchase", "local purchase", "domestic purchase",
            "raw material consumed", "material consumed",
            "sub contract", "job work", "labour charge",
            "freight inward", "carriage inward", "octroi",
            "custom duty", "clearing charge",
        ],
        "negative_keywords": ["purchase return"],
    },
    "employee_expenses": {
        "label": "Employee / Salary Expenses",
        "side": "pl",
        "keywords": [
            # Salary / remuneration (very high priority — matched first)
            "salary", "salaries", "salary to partner", "partner salary",
            "partner remuneration", "director remuneration",
            "director salary", "stipend",
            # Wages
            "wage", "labour charge", "labor charge",
            # Bonus, gratuity, leave
            "bonus", "gratuity", "leave with wages", "leave salary",
            # Welfare & statutory
            "staff welfare", "labour welfare fund", "labor welfare fund",
            "epf", "esi ", "e.s.i", "pf contribution",
            "employee benefit", "employee provident",
            # Other
            "incentive", "overtime",
            "labour refreshment", "labor refreshment",
        ],
        "negative_keywords": [
            "salary payable", "wages payable", "bonus payable",
            "leave with wages payable", "e.s.i payable", "esi payable",
            "salary & bonus payable", "wages & salary payable",
            "labour welfare fund payable", "labor welfare fund payable",
            "lwf payable",
        ],
    },
    "finance_cost": {
        "label": "Finance Cost",
        "side": "pl",
        "keywords": [
            "interest paid", "interest on loan", "bank charge",
            "bank interest", "bank cc intt", "cc interest",
            "bank od interest", "overdraft interest",
            "loan interest", "loan 1 interest", "loan 2 interest",
            "loan 3 interest", "loan 4 interest",
            "car loan interest", "machine loan interest", "machinery loan interest",
            "top up loan interest", "top up car loan interest",
            "interest to partner", "interest on unsecured",
            "intt on tempu", "interest on tempu", "interest on hdfc",
            "interest on term loan", "term loan interest",
            "interest on term loan", "interest on secured",
            "interest on lap", "bank interest on lap", "lap interest",
            "interest on lap loan",
            "processing fee", "cersai charge",
            "life insurance machinery loan",
            # FIX (Bug 4): Additional patterns that appear in Tally TBs
            "bank interest on term", "bank interest on", "interest charges",
            "machine loan", "machinery loan",
            "interest on car loan", "interest on vehicle",
            "interest on mortgage", "mortgage interest",
            "interest on cc", "interest on od",
        ],
        "negative_keywords": ["interest received", "interest income",
                              "intt paid on late payment", "bank charges and interest",
                              "bank charges", "processing fee"],
    },
    "other_expenses": {
        "label": "Other Expenses / Indirect Expenses",
        "side": "pl",
        "keywords": [
            "expense", "rent", "electricity", "telephone",
            "internet", "travelling", "conveyance", "vehicle running",
            "petrol", "diesel", "fuel", "repair", "maintenance",
            "insurance", "audit fee", "professional fee", "legal fee",
            "printing", "stationery", "postage", "courier",
            "advertisement", "marketing", "donation",
            "discount allowed", "bad debt",
            "miscellaneous", "office expense", "general expense",
            "entertainment", "subscription", "membership",
            "rate", "tax", "municipal", "water charge",
            "loading", "unloading", "packing",
            "commission paid", "brokerage", "agency",
            "foreign exchange loss", "exchange diff",
            "penalty", "fine", "late fee", "round off",
            "festival", "gift", "welfare",
        ],
        "negative_keywords": [
            "salary", "wage", "depreciation", "purchase",
            "payable", "outstanding", "provision",
        ],
    },
    "depreciation": {
        "label": "Depreciation",
        "side": "pl",
        "keywords": [
            "depreciation", "amortisation", "amortization",
            "dep on", "accumulated depreciation",
        ],
        "negative_keywords": ["provision for depreciation"],
    },
    # NEW — separate bucket for Direct Expenses (Electricity, Wages, Power & Fuel)
    # so they post into the Trading A/c "Direct Expenses" sub-rows instead of
    # being collapsed into Purchases.
    "direct_expenses": {
        "label": "Direct Expenses",
        "side": "pl",
        "keywords": [
            "wages a/c", "wages account", "factory wages", "labour wages",
            "electricity exp", "power and fuel", "power & fuel",
            "oil & lubricant", "oil and lubricant",
            "factory expense", "production expense",
        ],
        "negative_keywords": ["wages payable", "electricity payable"],
    },
}

# Priority order for classification (most specific first)
CLASSIFICATION_PRIORITY = [
    "depreciation",
    "advance_from_customer", "advance_to_supplier",
    "trade_payables", "trade_rec",
    "finance_cost",   # Must be before cash_bank so "LOAN INTEREST" doesn't match "bank"
    # FIX: direct_expenses must be checked BEFORE employee_expenses.
    # employee_expenses has a broad "wage" keyword (meant to catch combined
    # "Salary & Wages" accounts), but direct_expenses has the more specific
    # "wages a/c" / "factory wages" / "labour wages" phrasings for
    # manufacturing labour cost — which belongs in the Trading A/c's Direct
    # Expenses section, not the P&L's Employee Benefits note. With the old
    # order, a plain "WAGES A/C" account (TB group: Expenses Direct/Mfg.)
    # matched employee_expenses' generic "wage" keyword first and never
    # got a chance to match direct_expenses' more specific keyword — so it
    # was classified as employee_expenses entirely, then ALSO got written
    # into the wrong notes-to-p&l sub-row (its large amount landing under
    # "Salaries" and bumping the real Salary account down into "Bonus"),
    # while GROSS PROFIT's actual "Wages" row stayed blank — corrupting
    # both the P&L employee-expense breakdown AND the Trading Account's
    # Cost of Material Consumed figure (which depends on Direct Expenses
    # being present).
    "direct_expenses",
    "employee_expenses",
    "cash_bank", "inventories",
    "st_provisions",
    "st_borrowings", "lt_borrowings",
    "purchases", "revenue", "other_income",
    "fixed_assets", "non_current_investments",
    "stla",
    "capital",
    "other_cl", "other_current_assets",
    "other_expenses",
]


# ═══════════════════════════════════════════════════════════════════════
# PDF TRIAL BALANCE PARSER
# ═══════════════════════════════════════════════════════════════════════

def parse_tb_pdf(pdf_path):
    """Parse a Trial Balance PDF (Tally/Busy/Excel-exported) into account rows.

    HYBRID approach (fixes Dr/Cr column-attribution bug):
      • `extract_tables()` is authoritative for the Debit vs Credit column —
        each table row arrives as ``[name, debit_str, credit_str]`` so a
        credit-balance row on the debtor side (e.g. MEERA FORGING with only
        a credit amount) is preserved exactly as it appears in the PDF.
      • `extract_text()` lines are walked in parallel ONLY to capture the
        group/section headers ("SUNDRY DEBTORS", "SECURED LOANS", …) which
        appear BETWEEN tables and are NOT inside the extracted tables.

    Also handles Tally "Group-Wise" PDFs where pdfplumber merges multiple
    account names into a single cell with \\n separators (common when the TB
    has no ruling lines between rows on a page). In that case we split the
    merged cell and zip names with amounts.

    Returns the same shape as detect_tb_structure for seamless integration.
    """
    import pdfplumber, re

    num_re = re.compile(r'([\d,]+\.\d{2})')

    # Section/group keywords found in Indian TB PDFs. Used to reset
    # `current_group` when walking the text lines.
    GROUP_KEYWORDS = {
        "bank accounts", "bank account",
        "capital account", "capital accounts",
        "cash-in-hand", "cash in hand",
        "fixed assets",
        "direct expenses", "indirect expenses",
        "indirect incomes", "indirect income",
        "direct incomes", "direct income",
        "purchase account", "purchase accounts", "purchases", "purchase",
        "sales account", "sales accounts", "sales", "sale",
        "sundry creditors", "sundry debtors",
        "sundry payables", "sundry payable",
        "sundry receivables", "sundry receivable",
        "loans & advances (asset)", "loans & advances",
        "loans and advances", "loans (liability)", "loans liability",
        "current liabilities", "current assets",
        "deposits (asset)", "deposits",
        "duties & taxes", "duties and taxes",
        "secured loans", "unsecured loans", "unsecure loans",
        "stock-in-hand", "stock in hand",
        "investments", "investment",
        "profit & loss account", "profit and loss account",
        "profit & loss a/c",
        "provisions", "reserves & surplus", "reserves and surplus",
        "misc. expenses (asset)", "miscellaneous expenses",
        "branch/divisions", "suspense a/c",
        "expenses (indirect/admn.)", "expenses (indirect/admn)",
        "expenses (direct/mfg.)", "expenses (direct/mfg)",
        "income (indirect)", "income (direct/opr.)",
        "revenue accounts", "profit & loss",
        "provisions/expenses payable",
        "securities & deposits (asset)",
        "loans (liability)",
    }

    def _norm_header(line):
        return re.sub(r"[\s\-:]+$", "", line.strip().lower())

    def _is_group_header(line):
        return _norm_header(line) in GROUP_KEYWORDS

    def _to_float(v):
        if v is None: return 0.0
        s = str(v).strip()
        if not s: return 0.0
        s = (s.replace(",", "").replace("₹", "")
               .replace("Rs.", "").replace("Rs", "")
               .replace("(", "-").replace(")", ""))
        try:
            return float(s)
        except ValueError:
            return 0.0

    SKIP_NAMES = {"particulars", "trial balance",
                  "debit amount", "credit amount",
                  "total", "grand total", "opening balance",
                  "closing balance", "balance c/d", "balance b/d",
                  "sub total", "net total",
                  "totals c/o", "totals b/d",
                  "account/group name", "debit bal.", "credit bal."}

    accounts = []
    running_group = ""

    # ── Detect Tally "merged-cell" Group-Wise PDF format ──────────────────
    # This format (Penguin Packages style) has pdfplumber merging many accounts
    # into single cells because the PDF has no ruling lines between rows.
    # Cell [0] = names joined by \n, Cell [1] = debit amounts joined by \n,
    # Cell [2] = credit amounts joined by \n (some cells are empty strings)
    # Each PAGE has rows: [header, "Totals b/d", [merged data], ..., "Totals c/o"]
    # We detect it by checking if any table cell in col 0 contains multiple \n-separated names
    _is_merged_cell_format = False
    try:
        with pdfplumber.open(pdf_path) as _pdf_mc:
            for _page_mc in _pdf_mc.pages[:2]:
                _tables_mc = _page_mc.extract_tables() or []
                for _tbl_mc in _tables_mc:
                    for _row_mc in _tbl_mc:
                        if _row_mc and _row_mc[0]:
                            _cell0 = str(_row_mc[0] or "")
                            if _cell0.count('\n') > 2:  # multiple accounts merged in one cell
                                _is_merged_cell_format = True
                                break
                    if _is_merged_cell_format:
                        break
                if _is_merged_cell_format:
                    break
    except Exception:
        pass

    if _is_merged_cell_format:
        # ── Merged-cell Tally PDF: combined text+table parser ──────────────
        # Tally Group-Wise PDFs have no ruling lines between rows, so pdfplumber
        # merges all account names on a page into one cell with \n separators.
        # Strategy:
        #   1. Walk TEXT LINES per page to track current_group (groups appear as
        #      plain text lines with their name followed by a group subtotal amount)
        #   2. Use the same text lines to assign each account to its group
        #   3. Use TABLE CELLS to get the authoritative Dr/Cr split per account:
        #      check whether the account's amount appears in col B (Dr) or col C (Cr)
        import re as _re_mc

        _NUM_RE_MC = _re_mc.compile(r'([\d,]+\.\d{2})')
        _KNOWN_GROUPS_MC = {
            "capital account", "reserves & surplus",
            "current assets", "bank accounts", "cash-in-hand",
            "loans & advances (asset)", "securities & deposits (asset)",
            "stock-in-hand", "sundry debtors",
            "current liabilities", "duties & taxes", "provisions/expenses payable",
            "sundry creditors",
            "fixed assets", "investments", "loans (liability)",
            "bank o/d account", "secured loans", "unsecured loans",
            "pre-operative expenses", "profit & loss", "revenue accounts",
            "expenses (direct/mfg.)", "expenses (indirect/admn.)",
            "income (direct/opr.)", "income (indirect)",
            "purchase", "sale", "suspense account",
        }
        _SKIP_MC = {"totals c/o", "totals b/d", "grand total", "total",
                    "account/group name", "debit bal.", "credit bal."}
        _TITLE_RE_MC = _re_mc.compile(
            r"^(page \d|penguin|ludhiana|gstin|e-1|trial balance|as on|all accounts)",
            _re_mc.I
        )

        def _parse_merged_pdf(pdf_path_mc):
            mc_accounts = []
            seen_mc = set()
            cur_grp = ""

            try:
                with pdfplumber.open(pdf_path_mc) as _pdf_m:
                    for _page in _pdf_m.pages:
                        _text = _page.extract_text() or ""
                        _lines = [l.strip() for l in _text.split('\n') if l.strip()]
                        _tables = _page.extract_tables() or []

                        # Build name → group mapping by walking text lines in order
                        _page_grp = cur_grp
                        _line_grp = {}  # {name_lower: group}
                        for _ln in _lines:
                            if _TITLE_RE_MC.match(_ln.lower()): continue
                            if any(_ln.lower().startswith(s) for s in ("totals", "grand total", "contd.")): continue
                            _nm = _NUM_RE_MC.sub('', _ln).strip()
                            if not _nm: continue
                            _nml = _re_mc.sub(r"[\s\-:]+$", "", _nm.lower())
                            if _nml in _KNOWN_GROUPS_MC:
                                _page_grp = _nm.strip()
                                continue
                            _line_grp[_nm.lower().strip()] = _page_grp
                        cur_grp = _page_grp  # carry across pages

                        # Build flat set of Dr cell strings and Cr cell strings for quick lookup
                        _all_dr_strs = set()
                        _all_cr_strs = set()
                        for _tbl in _tables:
                            for _row in _tbl:
                                if not _row or len(_row) < 3: continue
                                _dr_raw = str(_row[1] or "")
                                _cr_raw = str(_row[2] or "")
                                for _v in _NUM_RE_MC.findall(_dr_raw):
                                    _all_dr_strs.add(_v.replace(',',''))
                                for _v in _NUM_RE_MC.findall(_cr_raw):
                                    _all_cr_strs.add(_v.replace(',',''))

                        # Walk text lines again to emit accounts with correct Dr/Cr
                        _page_grp2 = list(_line_grp.values())[0] if _line_grp else cur_grp
                        _cur_g2 = list(_line_grp.values())[0] if _line_grp else ""
                        for _ln in _lines:
                            if _TITLE_RE_MC.match(_ln.lower()): continue
                            if any(_ln.lower().startswith(s) for s in ("totals", "grand total", "contd.")): continue
                            _nm = _NUM_RE_MC.sub('', _ln).strip()
                            if not _nm: continue
                            _nml = _re_mc.sub(r"[\s\-:]+$", "", _nm.lower())
                            if _nml in _KNOWN_GROUPS_MC:
                                _cur_g2 = _nm.strip()
                                continue
                            _nums = _NUM_RE_MC.findall(_ln)
                            if not _nums: continue
                            _amt_str = _nums[0].replace(',', '')
                            try:
                                _amt = float(_amt_str)
                            except ValueError:
                                continue
                            # Determine Dr or Cr by checking which table column this amount appears in
                            if _amt_str in _all_dr_strs and _amt_str not in _all_cr_strs:
                                _dr_v, _cr_v = _amt, 0.0
                            elif _amt_str in _all_cr_strs and _amt_str not in _all_dr_strs:
                                _dr_v, _cr_v = 0.0, _amt
                            elif _amt_str in _all_dr_strs and _amt_str in _all_cr_strs:
                                # Ambiguous: use group heuristic
                                _cgl = _cur_g2.lower()
                                _is_cr = any(kw in _cgl for kw in (
                                    "capital","loan","liability","creditor","income","sale",
                                    "payable","provision","surplus","unsecured","secured",
                                ))
                                _dr_v, _cr_v = (0.0, _amt) if _is_cr else (_amt, 0.0)
                            else:
                                # Amount not in any table cell directly — use group heuristic
                                _cgl = _cur_g2.lower()
                                _is_cr = any(kw in _cgl for kw in (
                                    "capital","loan","liability","creditor","income","sale",
                                    "payable","provision","surplus","unsecured","secured",
                                ))
                                _dr_v, _cr_v = (0.0, _amt) if _is_cr else (_amt, 0.0)

                            _grp = _line_grp.get(_nm.lower().strip(), _cur_g2)
                            _key = f"{_nm}_{_dr_v:.2f}_{_cr_v:.2f}"
                            if _key not in seen_mc:
                                seen_mc.add(_key)
                                mc_accounts.append({
                                    "row": len(mc_accounts),
                                    "key": f"{_nm}_{len(mc_accounts)}",
                                    "name": _nm,
                                    "group": _grp,
                                    "debit": _dr_v,
                                    "credit": _cr_v,
                                    "net": _dr_v - _cr_v,
                                })
            except Exception:
                pass
            return mc_accounts

        try:
            _mc_accts = _parse_merged_pdf(pdf_path)
            if len(_mc_accts) > 10:
                return {
                    "format_type": "PDF",
                    "sheet_name": "PDF",
                    "header_row": 0,
                    "data_start_row": 0,
                    "account_col": 0,
                    "debit_col": 1,
                    "credit_col": 2,
                    "net_col": None,
                    "accounts": _mc_accts,
                }
        except Exception:
            pass
        # Fall through to standard parsers if merged-cell parse failed

    # ── Detect Group-Wise Trial Balance format (plain text, no tables) ──
    _is_group_wise = False
    try:
        with pdfplumber.open(pdf_path) as _pdf_detect:
            _ft = _pdf_detect.pages[0].extract_text() or ""
            if "Group Wise Trial Balance" in _ft or "Group wise Trial Balance" in _ft:
                _is_group_wise = True
    except Exception:
        pass

    if _is_group_wise:
        import re as _re_gw
        _GROUP_SIDE_GW = {
            'CURRENT LIABILITIES': 'credit', 'DUTIES AND TAXES': 'credit',
            'SUNDRY CREDITORS': 'credit', 'SUNDRY PAYABLE': 'credit',
            'CAPITAL ACCOUNT': 'credit', 'SECURED LOANS': 'credit',
            'UNSECURED LOANS': 'credit', 'ADVANCE FROM CUSTOMERS': 'credit',
            'SALES ACCOUNTS': 'credit', 'INDIRECT INCOME': 'credit',
            'BANK ACCOUNTS': 'both',      # ← BEFORE 'CURRENT ASSETS' so it matches first
            'CURRENT ASSETS': 'debit', 'CASH-IN-HAND': 'debit',
            'STOCK-IN-HAND': 'debit', 'SUNDRY DEBTORS': 'debit',
            'FIXED ASSETS': 'debit', 'PURCHASE ACCOUNTS': 'debit',
            'DIRECT EXPENSES': 'debit', 'INDIRECT EXPENSES': 'debit',
        }
        _SKIP_GW  = ['Group Wise', 'Closing Balance', 'P a r t i c u l',
                     'Debit  Credit', 'Debit\tCredit', 'Total :', 'Grand Total',
                     'Continue to next', 'Station', 'Page:',
                     'A.S.TRADERS', 'A.S. TRADERS', '321-I', 'BRS NAGAR',
                     'FOCAL POINT', 'FASHION ADDA', 'SHREE CRAFT']
        _KNOWN_GRP_WORDS = ['LIABILIT','ASSET','LOAN','EXPENSE','INCOME','ACCOUNT',
                            'CREDITOR','DEBTOR','CAPITAL','SALES','PURCHASE',
                            'STOCK','CASH','BANK','FIXED','ADVANCE','PAYABLE']
        _amt_re_gw = _re_gw.compile(r'^(.+?)\s+([\d,]+\.\d{2})\s*(?:([\d,]+\.\d{2}))?\s*$')
        _grp_re_gw = _re_gw.compile(r'^([A-Z][A-Z\s&\./\(\)\-0-9,\']+?)(?:\s*-{3,}\s*\(\s*(.+?)\s*\))?\s*$')
        _cg = "UNKNOWN"; _cs = 'debit'

        def _pi_gw(s):
            return float(s.replace(',', '').strip())

        def _side_gw(grp):
            gu = grp.upper()
            for k, s in _GROUP_SIDE_GW.items():
                if k in gu: return s
            return 'debit'

        with pdfplumber.open(pdf_path) as _pdf_gw:
            for _page_gw in _pdf_gw.pages:
                for _ln in (_page_gw.extract_text() or "").split('\n'):
                    _ln = _ln.strip()
                    if not _ln: continue
                    if any(sk in _ln for sk in _SKIP_GW): continue
                    _am_gw = _amt_re_gw.match(_ln)
                    if _am_gw:
                        _name_gw = _am_gw.group(1).strip()
                        # Same fix as the table-based parser below: strip a
                        # leading merged "S.No." prefix if present.
                        _name_gw = _re_gw.sub(r"^\d{1,3}\.\s*", "", _name_gw).strip()
                        _f = _pi_gw(_am_gw.group(2))
                        _s = _pi_gw(_am_gw.group(3)) if _am_gw.group(3) else None
                        if _s is not None:
                            _d, _c = _f, _s
                        elif _cs == 'credit':
                            _d, _c = 0, _f
                        elif _cs == 'both':
                            # CC A/C with large balance = overdraft (liability) → credit side
                            # Small CC balance (< 50,000) = positive deposit → debit
                            _nl = _name_gw.lower()
                            if ('cc a/c' in _nl or 'cc a/c no' in _nl) and _f > 50000:
                                _d, _c = 0, _f   # large CC = overdraft liability
                            else:
                                _d, _c = _f, 0   # small balance or regular account = asset
                        else:
                            _d, _c = _f, 0
                        accounts.append({
                            "name": _name_gw, "group": _cg,
                            "debit": _d, "credit": _c,
                            "net": _d - _c,
                            "key": f"{_name_gw}_{len(accounts)}",
                        })
                    else:
                        _gm_gw = _grp_re_gw.match(_ln)
                        if _gm_gw and not _ln.startswith('('):
                            _mn = _gm_gw.group(1).strip()
                            _sb = (_gm_gw.group(2) or "").strip()
                            _proposed = f"{_mn} ({_sb})" if _sb else _mn
                            # Only accept as group if it has a financial keyword
                            if any(w in _proposed.upper() for w in _KNOWN_GRP_WORDS):
                                _cg = _proposed
                                _cs = _side_gw(_cg)
                            # else: keep previous group (company header line)
        # Skip the table-based parser
        import sys as _sys_gw
        _gw_parsed = True
    else:
        _gw_parsed = False

    if not _gw_parsed:
      try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                # Text lines on this page (used only for group-header tracking)
                text_lines = [ln.strip() for ln in
                              (page.extract_text() or "").splitlines()
                              if ln.strip()]

                # Tables on this page (authoritative for Dr / Cr columns)
                tables = page.extract_tables() or []
                table_rows = []
                for tbl in tables:
                    for row in tbl:
                        if not row:
                            continue
                        if all((c is None or str(c).strip() == "")
                               for c in row):
                            continue
                        table_rows.append(row)

                # Set of names appearing as first-cell of a data row — used
                # to know which text lines are data (vs group headers).
                tr_first_cells = {str(row[0]).strip()
                                  for row in table_rows
                                  if row and row[0]}

                # Walk text lines IN ORDER to attach the most recently seen
                # group header to each data-row name on this page.
                page_group = running_group
                line_to_group = {}
                for ln in text_lines:
                    if _is_group_header(ln):
                        page_group = _norm_header(ln)
                        continue
                    name_part = re.sub(
                        r"\s+-?[\d,]+\.\d{1,2}(\s+-?[\d,]+\.\d{1,2})?\s*$",
                        "", ln).strip()
                    if name_part and name_part in tr_first_cells:
                        line_to_group.setdefault(name_part, page_group)

                # Emit account rows from the tables
                for row in table_rows:
                    cells = [str(c).strip() if c is not None else ""
                             for c in row]
                    while len(cells) < 3:
                        cells.append("")
                    name, dr_str, cr_str = cells[0], cells[1], cells[2]

                    # FIX: some Trial Balance PDF layouts (e.g. K.D. Knitwear's)
                    # have no ruling line between the "S.No." and "Account"
                    # columns, so pdfplumber's extract_tables() merges them
                    # into ONE cell — e.g. "1. Sales" instead of separate
                    # "1." / "Sales" cells. That stray leading serial number
                    # survives into the account name and silently breaks every
                    # downstream exact/fuzzy name match: "1. Sales" doesn't
                    # match the "Sales" classification keyword cleanly, and
                    # _norm_gst("1. Sales") normalises to "1." instead of ""
                    # (which would at least be recognised as a generic/empty
                    # name) — so it neither matches a known head by keyword
                    # nor gets handled by the single-candidate-row shortcut in
                    # the GROSS PROFIT sales-row matcher, and ends up
                    # spuriously appended as a brand new row instead of being
                    # written into the correct existing template row.
                    name = re.sub(r"^\d{1,3}\.\s*", "", name).strip()

                    if not name:
                        # Subtotal row — ignore.
                        continue
                    nl = name.lower().strip()
                    if nl in SKIP_NAMES:
                        continue
                    if re.search(r"continued on page", name, re.I):
                        continue
                    if re.match(
                            r"^(total|grand total|sub total|opening|closing|balance)\b",
                            name, re.I):
                        continue

                    dr = _to_float(dr_str)
                    cr = _to_float(cr_str)
                    if dr == 0 and cr == 0:
                        continue

                    grp = line_to_group.get(name, page_group)
                    accounts.append({
                        "row":    len(accounts),
                        "key":    f"{name}_{len(accounts)}",
                        "name":   name,
                        "group":  grp,
                        "debit":  dr,
                        "credit": cr,
                        "net":    dr - cr,
                    })

                running_group = page_group
      except Exception:
        # Fall back to text-only parser if pdfplumber table extraction fails
        try:
            return _parse_tb_pdf_text_fallback(pdf_path)
        except Exception:
            return None

    # If table extraction found no rows (PDFs without ruled lines), fall back
    if not accounts and not _gw_parsed:
        try:
            return _parse_tb_pdf_text_fallback(pdf_path)
        except Exception:
            return None

    return {
        "format_type":   "PDF",
        "sheet_name":    "PDF",
        "header_row":    0,
        "data_start_row":0,
        "account_col":   0,
        "debit_col":     1,
        "credit_col":    2,
        "net_col":       None,
        "accounts":      accounts,
    }


def _parse_tb_pdf_text_fallback(pdf_path):
    """Legacy text-only PDF parser kept as a fallback for PDFs whose tables
    are not extractable by pdfplumber (e.g. no ruled lines).

    NOTE: this path infers Dr/Cr from `current_group` alone, so a debtor with
    a CREDIT balance will appear in the debit column. The primary
    `parse_tb_pdf` should be used wherever possible.
    """
    import pdfplumber, re

    all_lines = []
    # FIX: a line with only ONE amount on it gives no way to tell, from
    # text alone, whether that number was under the "Debit Bal." or
    # "Credit Bal." column — the previous logic guessed purely from the
    # account's GROUP (e.g. "Sundry Creditors" -> assume credit). That
    # heuristic is right for the vast majority of accounts in a group,
    # but wrong for genuine exceptions a TB can legitimately contain —
    # e.g. a creditor with a DEBIT balance (supplier overpaid in advance)
    # sitting inside an otherwise credit-side "Sundry Creditors" group, or
    # a debtor with a CREDIT balance inside "Sundry Debtors". Those
    # specific accounts got silently flipped to the wrong side with no
    # way to detect it from group alone. Using each number's actual X
    # coordinate against the PDF's own "Debit Bal." / "Credit Bal."
    # column positions resolves these exactly, instead of guessing.
    _pos_amount_side = {}   # {amount_text: 'debit'|'credit'}
    with pdfplumber.open(pdf_path) as pdf:
        for page_idx, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            for line in text.split("\n"):
                l = line.strip()
                if l:
                    all_lines.append(l)
            try:
                words = page.extract_words()
            except Exception:
                words = []
            debit_x = credit_x = None
            for w in words:
                if w["text"].strip().rstrip(".") == "Debit" and debit_x is None:
                    debit_x = (w["x0"], w["x1"])
                elif w["text"].strip().rstrip(".") == "Credit" and credit_x is None:
                    credit_x = (w["x0"], w["x1"])
            if debit_x is None or credit_x is None:
                continue
            # Column boundary = midpoint between the two header labels'
            # nearest edges; anything left of it is the debit column,
            # right of it is the credit column. Keyed by the amount text
            # itself (not page/line position) — threading exact line
            # coordinates through the existing flat extract_text()-based
            # pipeline below isn't practical, and exact decimal amounts
            # (e.g. "6,00,000.00") are distinctive enough in practice that
            # this is a large accuracy improvement over guessing purely
            # from group name, even though it isn't a perfect 1:1 mapping
            # in the rare case the exact same amount appears more than
            # once across both columns in the same document.
            amt_re_w = re.compile(r'^[\d,]+\.\d{2}$')
            for w in words:
                if amt_re_w.match(w["text"]):
                    side = "debit" if w["x0"] < (debit_x[1] + credit_x[0]) / 2.0 else "credit"
                    _pos_amount_side[w["text"]] = side

    if not all_lines:
        return None

    num_re = re.compile(r'([\d,]+\.\d{2})')
    skip_patterns = {"trial balance", "as on ", "page no", "continued",
                     "focal point", "punjab", "phase-", "e-254"}
    # FIX: the previous EXACT-match whitelist only covered a handful of
    # literal group names ("capital account", "secured loans", ...),
    # so any TB using slightly different (but extremely common) phrasing —
    # "Advance From Customers", "Provisions/Expenses Payable", "Sundry
    # Payable" (singular), "PURCHASE A/C" mixed in with liability-side
    # groups, etc. — silently fell through to the debit-side default,
    # flipping the Dr/Cr sign for every account in that group. Switched to
    # a substring-based keyword check covering the actual recurring
    # liability/income vocabulary instead of an exhaustive literal list.
    _credit_group_keywords = (
        "capital", "secured loan", "unsecured loan", "creditor",
        "payable", "sale", "income", "profit & loss account",
        "profit and loss account", "current liabilit", "duties & tax",
        "duties and tax", "advance from customer", "reserve", "provision",
        "outstanding expense", "bills payable",
    )

    def _is_credit_group(group_name):
        gl = group_name.lower()
        return any(kw in gl for kw in _credit_group_keywords)

    current_group = ""
    accounts = []
    company_name = ""

    for line in all_lines:
        ll = line.lower().strip()
        if any(s in ll for s in skip_patterns):
            continue
        if ll == "particulars debit amount credit amount":
            continue
        # FIX: this PDF layout (and likely others like it) repeats a
        # column-header line — e.g. "S.No. Account Debit Bal. (`) Credit
        # Bal. (`)" — directly after EVERY "Group : X" section header,
        # right before that group's actual account rows. Since this line
        # has no decimal-formatted amount in it, it fell through to the
        # "no numbers found -> treat as a new group name" branch below,
        # immediately overwriting the just-set, correct group name with
        # this header text on every single section — so every account in
        # the whole document ended up tagged with the same wrong "group"
        # (the literal column-header string) instead of its real group.
        # Any classification logic that relies on the TB group (rather
        # than the account name alone) then silently broke for every
        # account in the file.
        if ("debit" in ll and "credit" in ll) or \
           (ll.startswith("s.no") and "account" in ll):
            continue
        # "Group : X" lines should set current_group to just "X" (the
        # literal "Group :" prefix isn't part of any GROUP_HEAD_MAP /
        # GROUP_KEYWORDS entry, so leaving it in would prevent the group
        # from ever being recognised downstream).
        if ll.startswith("group"):
            _gm = re.match(r"^group\s*:?\s*(.+)$", line.strip(), re.I)
            if _gm:
                current_group = _gm.group(1).strip()
                continue
        if not company_name and line.isupper() and len(line) > 3 and not num_re.search(line):
            company_name = line
            continue
        if company_name and line.strip() == company_name:
            continue
        nums = num_re.findall(line)
        name = num_re.sub('', line).strip()
        # Same fix as the primary table-based parser: strip a leading
        # merged "S.No." prefix if present (e.g. "1. Sales" -> "Sales").
        name = re.sub(r"^\d{1,3}\.\s*", "", name).strip()
        if not nums:
            if name and len(name) > 1 and name not in ("0.01", ""):
                nl = name.lower()
                if not any(s in nl for s in ["phase-", "focal", "punjab",
                           "ludhiana-", "delhi-", "mumbai-", "address"]):
                    current_group = name
            continue
        if not name:
            continue
        # FIX: "Total" (per-group subtotal) and "Grand Total" (document
        # total) lines DO contain two decimal-formatted numbers, so they
        # don't fall into the "no numbers -> group name" branch above —
        # they were instead falling all the way through to here and being
        # appended to `accounts` as if they were genuine individual
        # accounts (e.g. an account literally named "Total" with the
        # group's debit/credit subtotal as its amount, and a final "Grand
        # Total" entry with the whole TB's grand total). Besides being
        # nonsensical entries that show up in the Manual Review / mapping
        # UI, every such line silently double-counts that group's total
        # into the aggregated figures alongside its individual accounts.
        # The primary table-based parser already filters these out; this
        # fallback parser was missing the same guard.
        if re.match(r"^(total|grand total|sub\s*total|opening|closing|"
                    r"balance\s*(c/d|b/d))\b", name, re.I):
            continue
        dr_amt = 0.0
        cr_amt = 0.0
        if len(nums) == 1:
            val = float(nums[0].replace(',', ''))
            # Prefer the actual column position from the PDF (exact) over
            # the group-name heuristic (a reasonable guess, but wrong for
            # any account whose balance is on the opposite side from its
            # group's usual convention).
            _pos_side = _pos_amount_side.get(nums[0])
            if _pos_side == "credit":
                cr_amt = val
            elif _pos_side == "debit":
                dr_amt = val
            elif _is_credit_group(current_group):
                cr_amt = val
            else:
                dr_amt = val
        elif len(nums) >= 2:
            dr_amt = float(nums[0].replace(',', ''))
            cr_amt = float(nums[1].replace(',', ''))
        accounts.append({
            "row":    len(accounts),
            "key":    f"{name}_{len(accounts)}",
            "name":   name,
            "group":  current_group,
            "debit":  dr_amt,
            "credit": cr_amt,
            "net":    dr_amt - cr_amt,
        })
    if not accounts:
        return None
    return {
        "format_type":   "PDF",
        "sheet_name":    "PDF",
        "header_row":    0,
        "data_start_row":0,
        "account_col":   0,
        "debit_col":     1,
        "credit_col":    2,
        "net_col":       None,
        "accounts":      accounts,
    }

def convert_pdf_tb_to_xlsx(pdf_path, xlsx_path):
    """Convert a PDF trial balance to an xlsx file for processing."""
    result = parse_tb_pdf(pdf_path)
    if not result or not result["accounts"]:
        return False

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Trial Balance"

    # Header row
    ws.cell(1, 1, "Particulars")
    ws.cell(1, 2, "Debit")
    ws.cell(1, 3, "Credit")

    current_group = ""
    r = 2
    for acct in result["accounts"]:
        if acct["group"] != current_group:
            current_group = acct["group"]
            if current_group:
                ws.cell(r, 1, current_group)
                r += 1

        ws.cell(r, 1, acct["name"])
        if acct["debit"]:
            ws.cell(r, 2, acct["debit"])
        if acct["credit"]:
            ws.cell(r, 3, acct["credit"])
        r += 1

    wb.save(xlsx_path)
    return True


# ═══════════════════════════════════════════════════════════════════════
# TB AUTO-DETECTION ENGINE
# ═══════════════════════════════════════════════════════════════════════

def _detect_hierarchical_indented_tb(rows, sheet_name):
    """
    Parse a Tally-style Group-Wise Trial Balance that uses LEADING SPACES to denote
    hierarchy (e.g. Penguin Packages format):
      indent=0  → Top-level group header (Capital Account, Current Assets, Fixed Assets…)
      indent=4  → Sub-group or direct account under top-level group
      indent=6  → Individual account under sub-group

    Key insight: Group headers at every indent level have amounts equal to the SUM of their
    children. We must emit ONLY leaf accounts (accounts with no children) to avoid
    double-counting. A row is a leaf if the NEXT non-empty row has indent <= current.

    This format is identified by:
    1. Col A contains "Account/Group Name" (or similar) as the header
    2. Leading whitespace is used for hierarchy (not a separate code column)
    3. Col B = Debit, Col C = Credit
    """
    import re as _re
    SKIP_NAMES = {
        "total", "grand total", "totals c/o", "totals b/d",
        "account/group name", "particulars",
        "debit bal.", "credit bal.",
    }
    SKIP_STARTS = ("totals c/o", "totals b/d", "total")

    # Detect if this is the hierarchical format:
    # Must have "Account/Group Name" in col 0 header AND some rows with leading spaces
    header_row_idx = None
    has_indented = False
    for ri, row in enumerate(rows[:10]):
        if not row: continue
        nm = str(row[0] or "").strip()
        if "account/group name" in nm.lower() or \
           ("account" in nm.lower() and "name" in nm.lower()):
            # Check that col 1 and 2 have debit/credit markers
            col1 = str(row[1] or "").strip().lower()
            col2 = str(row[2] or "").strip().lower()
            if ("debit" in col1 or "dr" in col1) and ("credit" in col2 or "cr" in col2):
                header_row_idx = ri
                break

    if header_row_idx is None:
        return None  # Not this format

    # Check for indented rows after header
    for row in rows[header_row_idx + 1: header_row_idx + 20]:
        if not row: continue
        nm = str(row[0] or "")
        if nm and nm != nm.lstrip():  # has leading spaces
            has_indented = True
            break

    if not has_indented:
        return None  # Not indented — different format, let standard parser handle it

    # Now parse the hierarchical data
    # Strategy: collect all data rows, then identify leaf nodes.
    # A leaf node = a row where the next row at same or LOWER indent has NO child rows
    # between them (i.e., the very next non-empty row is at indent <= current row's indent).
    # BUT: some accounts ARE direct children of top-level groups (indent=4 with no indent=6 below)
    # and those are also leaves.

    data_rows = []  # [(row_idx, indent, name, dr, cr)]
    for ri in range(header_row_idx + 1, len(rows)):
        row = rows[ri]
        if not row: continue
        nm = str(row[0] or "")
        if not nm.strip(): continue
        indent = len(nm) - len(nm.lstrip())
        name = nm.strip()
        if not name: continue
        # Skip totals
        if name.lower() in SKIP_NAMES: continue
        if any(name.lower().startswith(s) for s in SKIP_STARTS): continue
        # Skip "Grand Total" or pure number rows
        if name.lower() in ("grand total", "sub total"): break
        # Get Dr/Cr
        dr = row[1] if len(row) > 1 else None
        cr = row[2] if len(row) > 2 else None
        if not isinstance(dr, (int, float)): dr = 0.0
        if not isinstance(cr, (int, float)): cr = 0.0
        dr, cr = float(dr), float(cr)
        data_rows.append((ri, indent, name, dr, cr))

    if not data_rows:
        return None

    # Identify leaf nodes: row i is a leaf if the next row has indent <= row i's indent,
    # meaning row i has no children below it.
    # Exception: if a row has indent=0 AND has amounts AND no children → it IS a leaf
    # (e.g. a standalone group with only one item that IS the total itself).
    accounts = []
    _KNOWN_GROUP_HEADERS = {
        "capital account", "current assets", "current liabilities",
        "fixed assets", "investments", "loans (liability)", "loans (asset)",
        "revenue accounts", "suspense account", "pre-operative expenses",
        "profit & loss", "bank accounts", "cash-in-hand",
    }

    # Track current group at each indent level for classification hints
    group_stack = {}  # {indent: name}

    for i, (ri, indent, name, dr, cr) in enumerate(data_rows):
        # Update group stack
        group_stack[indent] = name
        # Clear deeper levels
        for k in list(group_stack.keys()):
            if k > indent:
                del group_stack[k]

        # Check if this is a leaf node
        # Look at next row's indent
        if i + 1 < len(data_rows):
            next_indent = data_rows[i + 1][1]
            has_children = next_indent > indent
        else:
            has_children = False

        # Skip if this is a group header (has children below it)
        if has_children:
            continue

        # Skip rows with ZERO amounts — they're empty group headers or placeholders
        if dr == 0 and cr == 0:
            continue

        # Skip known top-level group names even if they appear as leaves
        if name.lower().strip() in _KNOWN_GROUP_HEADERS:
            continue

        # Determine group from stack (parent group = closest ancestor group)
        parent_group = ""
        for lvl in sorted(group_stack.keys(), reverse=True):
            if lvl < indent:
                parent_group = group_stack[lvl]
                break

        accounts.append({
            "row": ri,
            "key": f"{name}_{ri}",
            "name": name,
            "group": parent_group,
            "debit": dr,
            "credit": cr,
            "net": dr - cr,
        })

    if len(accounts) < 5:
        return None  # Too few accounts — probably wrong format

    return {
        "format_type": 1,  # Dr/Cr separate columns
        "sheet_name": sheet_name,
        "header_row": header_row_idx,
        "data_start_row": header_row_idx + 1,
        "account_col": 0,
        "debit_col": 1,
        "credit_col": 2,
        "net_col": None,
        "accounts": accounts,
    }


def detect_tb_structure(file_path):
    """
    Auto-detect the structure of a trial balance file.
    Returns dict with: format_type, header_row, data_start_row,
    account_col, debit_col, credit_col, net_col, sheet_name, accounts[]
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    results = []

    for sname in wb.sheetnames:
        ws = wb[sname]
        rows_data = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows_data.append(list(row))
            if i >= 2000:   # handle large TBs up to 2000 rows
                break

        if len(rows_data) < 2:
            continue

        # FIX: Try hierarchical indented format FIRST (Tally Group-Wise TB with spaces)
        # This format fails with standard _detect_columns because "All Accounts" triggers
        # a false acct_col before the real header row.
        hier = _detect_hierarchical_indented_tb(rows_data, sname)
        if hier and len(hier.get("accounts", [])) > 5:
            results.append(hier)
            continue

        det = _detect_columns(rows_data, sname)
        if det:
            results.append(det)

    wb.close()

    if not results:
        return {"error": "Could not detect trial balance structure. Please check the file format."}

    # Pick the best detection (most accounts found)
    best = max(results, key=lambda d: len(d.get("accounts", [])))
    return best


def _detect_columns(rows, sheet_name):
    """Detect which column is what in the TB."""
    header_row = None
    acct_col = None
    dr_col = None
    cr_col = None
    net_col = None
    format_type = None

    # Scan first 15 rows for headers
    for ri, row in enumerate(rows[:15]):
        if not row:
            continue
        row_lower = [str(c).strip().lower() if c else "" for c in row]

        # FIX: Skip document title rows that contain the word "name" (or "account")
        # in a narrative context — e.g. "TRIAL BALANCE NAME WISE" or "ACCOUNT STATEMENT".
        # A genuine column-header row always has MULTIPLE non-empty cells (one per column);
        # a title row usually has text only in col 0 and None elsewhere. Skip single-cell
        # rows whose text spans common title patterns so they don't steal acct_col.
        _non_empty = [v for v in row_lower if v]
        if len(_non_empty) == 1:
            _title = _non_empty[0]
            if any(kw in _title for kw in [
                "trial balance", "balance sheet", "profit & loss",
                "profit and loss", "account statement", "ledger report",
                "from ", "as on ", "as at ",
            ]):
                continue

        # FIX: Skip rows that have "account"/"name" keyword but NO debit/credit keyword
        # in ANY cell of the same row — these are title rows like "As On: 31-3-2026  All Accounts"
        # where "all accounts" triggers acct_col but the row has no Dr/Cr header.
        # A real column-header row always has the debit/credit column labels in the same row.
        _row_has_dr_cr = any(
            ("debit" in v or "credit" in v or v in ("dr", "cr", "dr.", "cr.", "dr bal", "cr bal",
             "debit bal.", "credit bal.", "debit balance", "credit balance"))
            for v in row_lower if v
        )
        _row_has_acct = any(
            any(k in v for k in ["particular", "account", "ledger", "name", "head", "description"])
            for v in row_lower if v
        )
        # If this row has "account" keyword but NO dr/cr keyword, it might be a title row
        # unless it's a single-column name-only format (which we handle after)
        if _row_has_acct and not _row_has_dr_cr and len(_non_empty) <= 2:
            # Very likely a title/date row — skip (e.g. "As On : 31-3-2026  All Accounts")
            # But don't skip if we haven't set acct_col yet and this could be a name-only TB
            # We'll let it fall through if the row has only 1-2 cells; real headers have ≥3
            continue

        # Look for account name column
        for ci, val in enumerate(row_lower):
            if not val:
                continue

            # Account name indicators
            if any(k in val for k in [
                "particular", "account", "ledger", "name", "head",
                "description", "party name",
            ]):
                if acct_col is None:
                    acct_col = ci
                    header_row = ri

            # Debit column — handle "Debit (₹)", "Dr.", "Debit Amount" etc
            if (val in ("dr", "debit", "dr balance", "debit balance",
                       "dr amount", "debit amount", "dr bal", "debit bal") or
                    val.startswith("debit") or val.startswith("dr ") or
                    val == "dr." or "debit" in val.split("(")[0].strip()):
                dr_col = ci

            # Credit column — handle "Credit (₹)", "Cr.", "Credit Amount" etc
            if (val in ("cr", "credit", "cr balance", "credit balance",
                       "cr amount", "credit amount", "cr bal", "credit bal") or
                    val.startswith("credit") or val.startswith("cr ") or
                    val == "cr." or "credit" in val.split("(")[0].strip()):
                cr_col = ci

            # Net/Amount column
            if val in ("amount", "net balance", "net amount", "balance",
                       "closing balance", "closing", "net"):
                net_col = ci

            # Opening column (Type 3)
            if val in ("opening", "opening balance"):
                pass  # We mostly care about closing

            # Closing column (Type 3)
            if val in ("closing", "closing balance"):
                if net_col is None:
                    net_col = ci

    # If no header found, try heuristic: first column with text, next columns with numbers
    if acct_col is None:
        for ri, row in enumerate(rows[:15]):
            if not row:
                continue
            text_cols = []
            num_cols = []
            for ci, val in enumerate(row):
                if val is None:
                    continue
                if isinstance(val, str) and len(val.strip()) > 2 and not _is_number_str(val):
                    text_cols.append(ci)
                elif isinstance(val, (int, float)) or _is_number_str(str(val)):
                    num_cols.append(ci)
            if len(text_cols) >= 1 and len(num_cols) >= 1:
                acct_col = text_cols[0]
                header_row = ri
                break

    if acct_col is None:
        return None

    # Determine format type
    # Re-check: if both dr and cr columns found but assigned same index, fix
    if dr_col is not None and cr_col is not None and dr_col != cr_col:
        format_type = 1  # Dr/Cr separate columns
    elif dr_col is not None and cr_col is not None and dr_col == cr_col:
        # Conflict - reset and try harder
        dr_col = None; cr_col = None
        format_type = None
    elif net_col is not None:
        format_type = 4  # Single amount (negative = credit)
    else:
        # Try to auto-detect number columns after the account column
        for ri, row in enumerate(rows[header_row + 1: header_row + 10], header_row + 1):
            if not row:
                continue
            num_cols_found = []
            for ci, val in enumerate(row):
                if ci == acct_col:
                    continue
                if isinstance(val, (int, float)) and val != 0:
                    num_cols_found.append(ci)
                elif isinstance(val, str) and _is_number_str(val):
                    num_cols_found.append(ci)
            if len(num_cols_found) >= 2:
                dr_col = num_cols_found[0]
                cr_col = num_cols_found[1]
                format_type = 1
                break
            elif len(num_cols_found) == 1:
                net_col = num_cols_found[0]
                format_type = 4
                break

    if format_type is None:
        return None

    # Fix: if format detected as 4 (single net column) but there are actually
    # two number columns (like Debit col A and Credit col B in this TB),
    # detect that and switch to format 1
    if format_type == 4 and net_col is not None and dr_col is None:
        # Check if there are two numeric columns around the net_col
        for ri2 in range(header_row + 1, min(header_row + 10, len(rows))):
            row2 = rows[ri2]
            if not row2: continue
            num_cols_found = []
            for ci2, v2 in enumerate(row2):
                if ci2 == acct_col: continue
                if isinstance(v2, (int, float)) and v2 != 0:
                    num_cols_found.append(ci2)
            if len(num_cols_found) >= 2:
                # Two numeric columns found — treat as Dr/Cr
                format_type = 1
                dr_col = num_cols_found[0]
                cr_col = num_cols_found[1]
                net_col = None
                break

    # Data starts after header
    data_start = header_row + 1

    # ── Accode/Acname TB format detection ────────────────────────────────────
    # Some TB exports (e.g. Tally name-wise format) use:
    #   Col A = Accode (numeric account code like "00009") OR group-header text
    #   Col B = Acname (the actual account name) OR None (for group-header rows)
    #   Col C = Debit, Col D = Credit
    #
    # Group-header rows look like: [' CAPITAL ', None, None, None, ...]
    # Data rows look like:         ['00009', 'CAPITAL A/C PROP...', 0, 6736804.66, ...]
    #
    # In this format acct_col is detected as 1 (col B, "Acname" contains "name"),
    # but group headers only have text in col A — so row[acct_col]=None for them,
    # making them invisible to the group-tracking logic. We detect this format
    # by checking if col 0 (A) regularly contains short numeric-code-like strings
    # when col 1 (B) has the real account name.
    _is_accode_format = False
    if acct_col == 1:
        # Check a sample of data rows: if col 0 looks like a numeric account code
        # (4-6 digit string) while col 1 has the name, it's accode format.
        _accode_like = 0
        _checked = 0
        for _ri2 in range(data_start, min(data_start + 30, len(rows))):
            _r2 = rows[_ri2]
            if not _r2 or len(_r2) < 2:
                continue
            _a, _b = _r2[0], _r2[1]
            if _b is not None and isinstance(_b, str) and len(_b.strip()) > 2:
                _checked += 1
                if isinstance(_a, str) and re.match(r'^[A-Z0-9]{2,8}$', _a.strip()):
                    _accode_like += 1
        if _checked > 0 and _accode_like / _checked > 0.5:
            _is_accode_format = True

    # Extract accounts — handles both flat and hierarchical TB formats
    accounts = []
    total_keywords = {"total", "grand total", "difference", "net total",
                      "closing balance", "opening balance total",
                      "balance c/d", "balance b/d", "group total"}

    # Extra skip patterns for reconciliation/difference entries that aren't real accounts
    _skip_name_patterns = re.compile(
        r'^(difference in opening|opening balance difference|'
        r'balance difference|rounding|round off difference)\b',
        re.I
    )

    current_group = None  # Track current group header for hierarchical TBs

    for ri in range(data_start, len(rows)):
        row = rows[ri]
        if not row or ri >= len(rows):
            continue

        # ── Accode/Acname format: group headers are in col A, names in col B ──
        if _is_accode_format:
            col_a = row[0] if len(row) > 0 else None
            col_b = row[1] if len(row) > 1 else None
            # Group header: col A has text, col B is None/empty
            if isinstance(col_a, str) and col_a.strip() and (col_b is None or str(col_b).strip() == ''):
                g = col_a.strip()
                # Skip rows that are actually totals/headers (Grand Total, etc.)
                if not re.match(r'^(grand total|total)\b', g, re.I):
                    current_group = g
                continue
            # Skip accode-only rows (col A = accode, col B = None, no amounts)
            if col_b is None or str(col_b).strip() == '':
                continue

        acct_name = row[acct_col] if acct_col < len(row) else None
        if not acct_name or not isinstance(acct_name, str):
            continue
        acct_name = acct_name.strip()
        if not acct_name or len(acct_name) < 2:
            continue

        # Skip totals/subtotals
        if acct_name.lower().strip() in total_keywords:
            continue
        if re.match(r'^(total|grand total|group total|sub total|net total)\b', acct_name, re.I):
            continue

        # Skip reconciliation/difference entries
        if _skip_name_patterns.match(acct_name):
            continue

        # Get amounts
        dr_amt = 0
        cr_amt = 0
        net_amt = 0

        if format_type == 1 and dr_col is not None and cr_col is not None:
            dr_val = row[dr_col] if dr_col < len(row) else None
            cr_val = row[cr_col] if cr_col < len(row) else None
            dr_amt = _to_float(dr_val)
            cr_amt = _to_float(cr_val)

            # FIX: Some Tally exports (e.g. Chadha Sons Format 1) have the
            # credit column header in col E but individual account credits
            # are placed in col D (the column immediately after the debit
            # column col C). The header-based auto-detection gives cr_col=E,
            # but D is empty in the header so it's missed. As a result, all
            # accounts whose credits sit in col D (creditors, unsecured loans,
            # capital etc.) get net=0 and are silently dropped.
            # Fix: if cr_col != dr_col+1 AND the row has a non-zero value in
            # dr_col+1, use that as an additional credit amount.
            alt_cr_col = dr_col + 1
            if cr_col != alt_cr_col and alt_cr_col < len(row):
                alt_cr_val = row[alt_cr_col]
                alt_cr_amt = _to_float(alt_cr_val)
                if alt_cr_amt != 0 and cr_amt == 0:
                    cr_amt = alt_cr_amt   # use adjacent column as credit

            net_amt = dr_amt - cr_amt
        elif format_type == 4 and net_col is not None:
            # For hierarchical format: check BOTH debit and credit columns
            # even if format_type was detected as 4
            if dr_col is None and cr_col is None:
                # Try columns 1 and 2 as dr/cr if they have data
                dr_try = row[1] if len(row) > 1 else None
                cr_try = row[2] if len(row) > 2 else None
                dr_f = _to_float(dr_try)
                cr_f = _to_float(cr_try)
                if dr_f != 0 or cr_f != 0:
                    dr_amt = dr_f
                    cr_amt = cr_f
                    net_amt = dr_amt - cr_amt
                else:
                    net_val = row[net_col] if net_col < len(row) else None
                    net_amt = _to_float(net_val)
                    if net_amt > 0:
                        dr_amt = net_amt
                    else:
                        cr_amt = abs(net_amt)
            else:
                net_val = row[net_col] if net_col < len(row) else None
                net_amt = _to_float(net_val)
                if net_amt > 0:
                    dr_amt = net_amt
                else:
                    cr_amt = abs(net_amt)

        # Detect group headers: rows with name but zero amounts
        # These help classify sub-items in hierarchical TBs
        if dr_amt == 0 and cr_amt == 0 and net_amt == 0:
            # This is likely a group/category header
            current_group = acct_name
            continue

        accounts.append({
            "row": ri,
            "key": f"{acct_name}_{ri}",   # unique key for JS mapping
            "name": acct_name,
            "group": current_group,  # parent group for classification hint
            "debit": dr_amt,
            "credit": cr_amt,
            "net": net_amt,
        })

    return {
        "format_type": format_type,
        "sheet_name": sheet_name,
        "header_row": header_row,
        "data_start_row": data_start,
        "account_col": acct_col,
        "debit_col": dr_col,
        "credit_col": cr_col,
        "net_col": net_col,
        "accounts": accounts,
    }


def _is_number_str(s):
    """Check if a string looks like a number."""
    if not s:
        return False
    s = s.strip().replace(",", "").replace("(", "-").replace(")", "")
    try:
        float(s)
        return True
    except (ValueError, TypeError):
        return False


def _to_float(val):
    """Convert a value to float, handling strings and None."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        s = val.strip().replace(",", "").replace("₹", "").replace("Rs.", "").replace("Rs", "")
        # Handle parentheses = negative
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        s = s.strip()
        try:
            return float(s)
        except (ValueError, TypeError):
            return 0.0
    return 0.0


# ═══════════════════════════════════════════════════════════════════════
# ACCOUNT CLASSIFIER
# ═══════════════════════════════════════════════════════════════════════

# Group header → BS head mapping for hierarchical TBs
GROUP_HEAD_MAP = {
    "capital account":           "capital",
    "capital accounts":          "capital",
    "capital":                   "capital",
    "reserves & surplus":        "capital",
    "reserves and surplus":      "capital",
    "bank accounts":             "cash_bank",
    "bank account":              "cash_bank",
    "bank":                      "cash_bank",
    "cash-in-hand":              "cash_bank",
    "cash in hand":              "cash_bank",
    "cash":                      "cash_bank",
    "fixed assets":              "fixed_assets",
    "investments":               "non_current_investments",
    "sundry creditors":          "trade_payables",
    "suppliers":                 "trade_payables",   # Tally name-wise format
    "sundry debtors":            "trade_rec",
    "sundry debtor":             "trade_rec",
    "sundry receivables":        "trade_rec",
    "sundry receivable":         "trade_rec",
    "customers":                 "trade_rec",        # Tally name-wise format
    "purchase account":          "purchases",
    "purchase accounts":         "purchases",
    "purchases":                 "purchases",
    "sales account":             "revenue",
    "sales accounts":            "revenue",
    "sales":                     "revenue",
    "stock-in-hand":             "inventories",
    "stock in hand":             "inventories",
    # "indirect expenses" → other_expenses (high confidence via group header).
    # Previously this was commented out, causing FREIGHT, ADVERTISEMENT, ACCOUNTANCY
    # CHARGES etc. (all under "Indirect Expenses" in Tally) to fall through to
    # low-confidence classification and land in wrong buckets (purchases or direct_expenses).
    "indirect expenses":         "other_expenses",
    "indirect expense":          "other_expenses",
    # FIX: Tally uses "Expenses (Indirect/Admn.)" and "Expenses (Direct/Mfg.)"
    # as group names — these weren't in GROUP_HEAD_MAP so AMC CHARGES, GENERATOR EXP
    # etc. fell through to low-confidence other_current_assets instead of other_expenses.
    "expenses (indirect/admn.)": "other_expenses",
    "expenses (indirect/admn)":  "other_expenses",
    "expenses (indirect":        "other_expenses",   # catches any variation
    "profit & loss expenses":    "other_expenses",   # also catches this format
    # Tally "Expenses (Direct/Mfg.)" = wages, electricity, freight (direct expenses)
    "expenses (direct/mfg.)":    "direct_expenses",
    "expenses (direct/mfg)":     "direct_expenses",
    "expenses (direct":          "direct_expenses",
    # FIX (Issue 2/4): Direct expenses (Electricity, Wages, Power & Fuel)
    # need their OWN injection target on the Trading A/c. They were
    # previously collapsed into `purchases`, which made them overwrite the
    # purchase header row instead of landing in the Direct Expenses sub-rows.
    "direct expenses":           "direct_expenses",
    "direct expense":            "direct_expenses",
    "indirect income":           "other_income",
    "indirect incomes":          "other_income",
    # FIX: Tally "Income (Indirect)" group name
    "income (indirect)":         "other_income",
    "income (direct/opr.)":      "revenue",
    "income (direct/opr)":       "revenue",
    "income (direct)":           "revenue",
    "direct income":             "revenue",
    "direct incomes":            "revenue",
    # NOTE: "trading/direct expenses" is intentionally NOT in GROUP_HEAD_MAP.
    # It contains a mix of purchases, opening stock, and direct expenses —
    # each of which is correctly routed via name-level rules in _classify_single
    # (the direct_expenses group rule with opening-stock/purchase exemptions).
    # Putting it in GROUP_HEAD_MAP would cause the group-override in Step 1
    # to fire before those name-level rules, silently misclassifying all
    # opening stock and purchase accounts as purchases (or direct_expenses).
    "trading/direct incomes":    "revenue",          # sales
    "profit & loss expenses":    "other_expenses",   # indirect expenses
    "profit & loss incomes":     "other_income",     # indirect incomes
    "profit & loss expense":     "other_expenses",
    "profit & loss income":      "other_income",
    "others assets":             "stla",             # misc assets (advance tax, GST rcvbl)
    "other assets":              "stla",
    "others assets (asset)":     "stla",
    "payables":                  "other_cl",         # salary/ESI/rent payable (JEANS WORLD)
    "sundry payables":           "other_cl",
    "sundry payable":            "other_cl",
    "current liabilities":       "other_cl",
    # FIX (2026-07-16): additional Tally / Busy group name variants seen in
    # real client TBs (JEANS WORLD, Penguin Packages, Shree Craft etc.)
    "provisions/expenses payable": "other_cl",      # already present below, kept for safety
    "provisions & expenses payable": "other_cl",
    "salary & bonus payable":    "other_cl",
    "salary and bonus payable":  "other_cl",
    "expenses payable":          "other_cl",
    # Tally name-wise format group names for trading items
    "trading/direct expenses":   "direct_expenses",
    "trading/direct incomes":    "revenue",
    "profit & loss expenses":    "other_expenses",
    "profit & loss incomes":     "other_income",
    "profit and loss expenses":  "other_expenses",
    "profit and loss incomes":   "other_income",
    # FIX: "Provisions/Expenses Payable" group
    "provisions/expenses payable": "other_cl",
    "current assets":            "other_current_assets",
    "provisions":                "st_provisions",
    "unsecure loans":            "lt_borrowings",
    "unsecured loans":           "lt_borrowings",
    "secured loans":             "lt_borrowings",
    "loans (liability)":         "lt_borrowings",
    "loans liability":           "lt_borrowings",
    # FIX (Bug 2): the LOANS & ADVANCES (ASSET) group in Tally/Busy TBs
    # represents short-term advances (GST credit, TDS/TCS receivable, loans
    # given out, deposits). Map to `stla` so those items are classified with
    # HIGH confidence under "Short-term loans & advances" rather than falling
    # through to the generic low-confidence other_current_assets bucket.
    "loans & advances (asset)":  "stla",
    "loans & advances":          "stla",
    "loans and advances":        "stla",
    "deposits (asset)":          "stla",
    "deposits":                  "stla",
    # FIX: Tally "Securities & Deposits (Asset)" = FDR, Security Deposits, TDS Receivable
    "securities & deposits (asset)": "stla",
    "securities and deposits (asset)": "stla",
    "securities & deposits":     "stla",
    # FIX: Tally "Purchase" group contains individual purchase accounts + wages + electricity
    # Map to purchases (direct expenses like Wages/Electricity handled by name rules below)
    "purchase":                  "purchases",
    "purchases":                 "purchases",
    # FIX: Tally "Sale" group
    "sale":                      "revenue",
    "sales accounts":            "revenue",
    "sales account":             "revenue",
    "duties & taxes":            "other_cl",   # GST payable = current liability
    "duties and taxes":          "other_cl",
    "duties & taxes (gst)":      "other_cl",
    "duties and taxes (gst)":    "other_cl",
    "advance from customers":    "other_cl",   # advance received = current liability
    "misc. expenses (asset)":    "misc_expenditure",
    "miscellaneous expenses":    "misc_expenditure",
    "profit & loss account":     "capital",
    "profit and loss account":   "capital",
    "profit & loss a/c":         "capital",
}


def classify_accounts(accounts):
    """
    Classify each TB account under a BS/P&L head.
    Uses group headers as classification hints for hierarchical TBs.
    Returns list of dicts with added 'bs_head' and 'confidence' keys.
    """
    classified = []
    for i, acct in enumerate(accounts):
        name = acct["name"]
        group = acct.get("group", "")
        bs_head, confidence = _classify_single(name, acct["net"], group)
        acct_copy = dict(acct)
        acct_copy["bs_head"] = bs_head
        acct_copy["confidence"] = confidence
        # Ensure every account has a unique key for the JS mapping UI.
        # Without this, accounts from the group-wise PDF parser (which don't
        # add a key field) all share `undefined` as their JS key, causing
        # userMappings[undefined] to be overwritten repeatedly and all
        # accounts to appear under the last account's bs_head in the UI.
        if not acct_copy.get("key"):
            acct_copy["key"] = f"{name}_{i}"
        classified.append(acct_copy)
    return classified


def _classify_single(name, net_amount, group=None):
    """Classify a single account name. Returns (head_key, confidence)."""
    name_lower = name.lower().strip()
    group_lower = (group or "").lower().strip()
    # FIX: Tally name-wise TB format wraps group names in spaces: ' CAPITAL ',
    # ' CUSTOMERS ', etc. Strip them before lookup so GROUP_HEAD_MAP entries
    # (which have no surrounding spaces) match correctly.
    group_lower = group_lower.strip()
    # FIX: Tally/Busy TBs very commonly name groups like "PURCHASE A/C",
    # "SALE A/C", "CURRENT LIABILITIES A/C" etc. — using the "A/C"
    # abbreviation instead of the full word "ACCOUNT". GROUP_HEAD_MAP only
    # has entries for the spelled-out forms ("purchase account", "purchase
    # accounts", "purchases"), so "purchase a/c" matched NEITHER the exact
    # lookup nor the partial-substring fallback, meaning the group-based
    # override never fired for these — letting generic name-keyword
    # matching decide instead. That's how accounts like "RAW MATERIAL" and
    # "PACKING MATERIAL" (literally listed under Group: PURCHASE A/C, i.e.
    # purchases made during the year) got silently misclassified as
    # `inventories` instead of `purchases`: those exact phrases are ALSO
    # legitimate inventory keywords (for TBs where they really do represent
    # closing stock), and with no group override to disambiguate, the
    # name-keyword match won by default — hugely inflating the "closing
    # stock" figure and corrupting Net Profit / Total Assets on the
    # website's summary page (a multi-crore RAW MATERIAL purchase is not a
    # multi-crore stock balance). Normalising "a/c" -> "account" here lets
    # the group lookup work the same way it already does for the
    # spelled-out form, for every classification call site.
    group_lower = re.sub(r"\ba/c\b", "account", group_lower)

    # Rule: TB group explicitly indicates direct/manufacturing expenses
    # (e.g. "Expenses (Direct/Mfg.)", "Direct Expenses", "Direct Exp -
    # Manufacturing") — route to direct_expenses regardless of word order,
    # since none of these phrasings match any GROUP_HEAD_MAP entry (exact
    # or partial-substring) due to the differing word order, which
    # previously let accounts like "WAGES A/C" fall through to generic
    # name-keyword matching and get misclassified as employee_expenses.
    if group_lower and re.search(r"\bdirect\b", group_lower) and (
            "expense" in group_lower or "mfg" in group_lower
            or "manufactur" in group_lower):
        # EXCEPTION: opening/closing stock always goes to inventories
        if "opening stock" in name_lower or "closing stock" in name_lower:
            return "inventories", "high"
        # EXCEPTION: purchase accounts always go to purchases
        # (Tally name-wise TBs put purchases inside "TRADING/DIRECT EXPENSES")
        if "purchase" in name_lower:
            return "purchases", "high"
        return "direct_expenses", "high"

    # FIX: Tally "Purchase" group contains BOTH pure purchase accounts AND
    # factory direct expenses (WAGES, ELECTRICITY EXPENSES, Freight).
    # Since GROUP_HEAD_MAP maps "purchase" → "purchases" (Step 1), these
    # direct expenses would be mislabelled as purchases without this guard.
    # Check name keywords BEFORE the GROUP_HEAD_MAP lookup for this group.
    if group_lower in ("purchase", "purchases"):
        # Opening/closing stock → inventories
        if "opening stock" in name_lower or "closing stock" in name_lower:
            return "inventories", "high"
        # Direct expense name keywords → direct_expenses
        _de_kws = BS_HEADS["direct_expenses"]["keywords"]
        if any(kw in name_lower for kw in _de_kws):
            return "direct_expenses", "high"
        # Plain "WAGES" / "ELECTRICITY" in Purchase group = direct factory cost
        # (not a purchase account). These are Tally accounts mis-grouped under
        # "Purchase" but are really manufacturing costs.
        _direct_in_purchase = [
            "wages", "electricity exp", "electricity charges",
            "power & fuel", "power and fuel",
            "labour", "labor", "freight",
        ]
        if any(kw in name_lower for kw in _direct_in_purchase):
            return "direct_expenses", "high"
        # Otherwise → purchases (actual purchase account)

    # ── Smart rules based on name + balance sign ──────────────────────
    # Rule: "loan" in name + CREDIT balance = borrowing (not fixed asset)
    if "loan" in name_lower and net_amount < 0:
        if any(kw in name_lower for kw in ["secured", "hypothec", "mortgage"]):
            return "lt_borrowings", "high"
        if any(kw in name_lower for kw in ["unsecure", "unsecured"]):
            return "lt_borrowings", "high"
        # Generic loan with credit balance = borrowing
        return "lt_borrowings", "high"

    # Rule: cheque issued / not cleared with credit balance = outstanding liability (OCL)
    # Must check BEFORE the "bank in name" rule since these don't always contain "bank"
    if net_amount < 0 and any(kw in name_lower for kw in [
            "cheque issued", "chq issued", "not cleared", "not presented", "issued not"]):
        return "other_cl", "high"

    # Rule: "bank" in name + CREDIT balance = secured loan/OD/CC
    if ("bank" in name_lower or "a/c" in name_lower) and net_amount < 0:
        # Cash Credit / Overdraft facilities are WORKING CAPITAL (short-term),
        # not long-term loans — even though they're "secured" against assets.
        # Must check this BEFORE the generic "loan"/"term loan" check below,
        # since CC/OD account names often also contain "loan"-adjacent words.
        if any(kw in name_lower for kw in ["od", "overdraft", "cash credit"]) or \
           re.search(r'(?<![a-z])cc(?![a-z])', name_lower):
            return "st_borrowings", "high"
        if any(kw in name_lower for kw in ["loan", "machinery", "vehicle", "term loan"]):
            return "lt_borrowings", "high"
        # Cheque issued/not cleared with credit balance = outstanding liability
        if any(kw in name_lower for kw in ["cheque issued", "chq issued",
               "not cleared", "not presented", "issued not"]):
            return "other_cl", "high"
        # Bank account with negative balance = bank overdraft = short term borrowing
        if any(kw in group_lower for kw in ["bank", "cash"]):
            return "st_borrowings", "high"

    # Rule: "round off" / "roundoff" = other_expenses (even if credit)
    if "round off" in name_lower or "roundoff" in name_lower:
        return "other_expenses", "high"

    # ── Name-priority overrides: run BEFORE group check ──────────────
    # These specific P&L sub-categories must override a broad group like
    # "indirect expenses" → other_expenses which fires too early.
    # Order matters: check most specific first.

    # Opening stock: always inventories regardless of group.
    # Tally name-wise format places OPENING STOCK inside the
    # "TRADING/DIRECT EXPENSES" group alongside purchases — without this
    # guard it would be classified as `purchases` via the group mapping,
    # making it inflate the purchases total and never reach the opening
    # stock row in notes to p&l.
    if "opening stock" in name_lower:
        return "inventories", "high"

    # Depreciation: must always go to depreciation head
    for kw in BS_HEADS["depreciation"]["keywords"]:
        if kw in name_lower:
            return "depreciation", "high"

    # FIX Bug 2/3/4/5: PAYABLES group accounts are LIABILITIES (other_cl),
    # even when their name contains salary/ESI/wage keywords.
    # E.g. "SALARY ARTI DEVI" under group "PAYABLES" is salary PAYABLE
    # (a current liability), NOT salary expense in P&L.
    # Without this guard the employee_expenses keyword "salary"/"wage" below
    # fires first — landing SALARY PAYABLE amounts in notes to p&l Employee
    # Benefits rows instead of notes to bs OCL, producing wrong figures in
    # BOTH the P&L (ESI shows 12,331 = SALARY JASBIR instead of 41,454) AND
    # the balance sheet (salary payable not shown as liability).
    _is_payables_group = bool(group_lower) and any(
        kw in group_lower for kw in (
            "payable", "payables", "outstanding", "accrued",
        )
    )
    if _is_payables_group:
        # Any credit-balance account in a "payables"-named group is a
        # current liability.  Debit-balance would be an advance/receivable
        # but TB payables groups virtually never carry debit balances.
        return "other_cl", "high"

    # Employee expenses: salary, wages etc. (but not "salary payable")
    # Use word-boundary matching for short/ambiguous keywords to avoid
    # false matches (e.g. "esi" matching "designer", "esi " is safer)
    _emp_neg = BS_HEADS["employee_expenses"].get("negative_keywords", [])
    if not any(nk in name_lower for nk in _emp_neg):
        _emp_match = False
        for kw in BS_HEADS["employee_expenses"]["keywords"]:
            # For very short keywords (≤4 chars), require word boundary (space or start/end)
            if len(kw) <= 4:
                import re as _re
                if _re.search(r'(?<![a-z])' + _re.escape(kw) + r'(?![a-z])', name_lower):
                    _emp_match = True
                    break
            elif kw in name_lower:
                _emp_match = True
                break
        if _emp_match:
            return "employee_expenses", "high"

    # Finance cost: bank interest, cc intt etc. (but not late payment interest)
    # IMPORTANT (Issue 3): "Bank Interest Recd." / "Bank Interest Received" is
    # INCOME (a credit balance, net_amount < 0), not an expense — even though
    # its name contains "bank interest" which matches a finance_cost keyword.
    # Only classify as finance_cost (an EXPENSE head) when the account has a
    # DEBIT balance (net_amount > 0, money paid out). A credit-balance
    # "interest" account is interest INCOME and must fall through to be
    # classified as other_income/revenue instead.
    _fin_neg = BS_HEADS["finance_cost"].get("negative_keywords", [])
    if net_amount > 0 and not any(nk in name_lower for nk in _fin_neg):
        for kw in BS_HEADS["finance_cost"]["keywords"]:
            if kw in name_lower:
                return "finance_cost", "high"

    # Step 1a: Group-based override for compound group names (group-wise TBs)
    if group_lower:
        # CURRENT LIABILITIES (DUTIES AND TAXES) → other_cl
        if "current liabilit" in group_lower and "sundry creditor" not in group_lower:
            return "other_cl", "high"
        # ADVANCE FROM CUSTOMERS → other_cl (advance received = current liability)
        if "advance from customer" in group_lower:
            return "other_cl", "high"
        # SUNDRY PAYABLE / PAYABLES → other_cl (JEANS WORLD "PAYABLES" group)
        if "sundry payable" in group_lower or group_lower in ("payables", "payable"):
            return "other_cl", "high"
        # SUNDRY CREDITORS / SUPPLIERS → trade_payables
        if "sundry creditor" in group_lower:
            return "trade_payables", "high"
        # FIX (2026-07-16): "SUPPLIERS" group in JEANS WORLD / Tally name-wise
        # format is already in GROUP_HEAD_MAP but hits the _GENERIC_BARE_GROUPS
        # guard before the map lookup. Explicitly catch it here first.
        if group_lower in ("suppliers", "supplier"):
            return "trade_payables", "high"
        # SUNDRY DEBTORS / CUSTOMERS → trade_rec
        if "sundry debtor" in group_lower:
            return "trade_rec", "high"
        # FIX (2026-07-16): "CUSTOMERS" group in JEANS WORLD / Tally name-wise
        if group_lower in ("customers", "customer"):
            return "trade_rec", "high"
        # FIX (2026-07-16): OTHERS ASSETS in Tally name-wise = stla
        if group_lower in ("others assets", "other assets", "others assets (asset)"):
            return "stla", "high"
        # FIX: Tally "Duties & Taxes" group contains BOTH:
        #   - GST Input (CGST Input, SGST Input) → DEBIT balance → asset (stla)
        #   - GST Payable → CREDIT balance → liability (other_cl)
        # The GROUP_HEAD_MAP maps "duties & taxes" → other_cl, which is WRONG for
        # debit-balance rows. Override here based on sign before hitting GROUP_HEAD_MAP.
        if ("duties" in group_lower and "tax" in group_lower) or \
           ("duties & taxes" in group_lower) or ("duties and taxes" in group_lower):
            if net_amount > 0:
                # Debit balance = GST Input Credit = asset
                return "stla", "high"
            else:
                # Credit balance = GST Payable = liability
                return "other_cl", "high"

    # Step 1: Check if group header directly maps to a head.
    # FIX: only apply this for SPECIFIC group labels (e.g. "current assets
    # (bank accounts)", "current assets (sundry debtors)") — NOT for a
    # bare generic group like plain "current assets" or "current
    # liabilities" with no sub-bracket. A bare generic group should defer
    # to Step 2's name-keyword matching first, since many TBs dump
    # unrelated accounts (Prepaid Insurance, TCS Collection, Round Off,
    # etc.) directly under the bare top-level group with no sub-heading —
    # and those specific keywords (e.g. "prepaid", "tcs") are far more
    # reliable classifiers than the generic group bucket. Without this,
    # "PREPAID INSURANCE" / "TCS COLLECTION" under bare "CURRENT ASSETS"
    # always landed in the generic Other Current Assets bucket instead of
    # Short Term Loans & Advances, where the template actually expects
    # them and has dedicated injection logic.
    _GENERIC_BARE_GROUPS = {
        "current assets", "current liabilities", "current liability",
        # NOTE: "customers", "suppliers", "bank", "cash" are NOT generic bare
        # groups — they have specific GROUP_HEAD_MAP entries that should fire
        # directly, not defer to name-keyword matching.
    }
    if (group_lower and group_lower in GROUP_HEAD_MAP
            and group_lower not in _GENERIC_BARE_GROUPS):
        group_head = GROUP_HEAD_MAP[group_lower]
        # FIX: guard against the TB parser mis-assigning a group header
        # as the last sub-account of the previous group.  Common case:
        # "PURCHASE ACCOUNTS" (row immediately after the SALES ACCOUNTS
        # Total row) gets group='SALES ACCOUNTS' → revenue, even though
        # its own name and sign both scream purchases.
        # If the derived group_head disagrees with what the account name
        # directly maps to in GROUP_HEAD_MAP, and the name's own
        # GROUP_HEAD_MAP entry exists, trust the name.
        name_head = GROUP_HEAD_MAP.get(name_lower)
        if name_head and name_head != group_head:
            return name_head, "high"
        # Also apply a sign sanity check: revenue accounts should have a
        # credit balance (net < 0).  A debit-balance account classified
        # as revenue via its group is almost certainly a mis-assigned row.
        if group_head == "revenue" and net_amount > 0:
            # Re-classify using name keywords instead
            pass  # fall through to name-based classification below
        else:
            return group_head, "high"

    # Step 2: Try each head in priority order by name
    for head_key in CLASSIFICATION_PRIORITY:
        # finance_cost is an EXPENSE head — already handled above with the
        # correct net_amount > 0 guard (Issue 3 fix). Skip it here for
        # credit-balance accounts so e.g. "Bank Interest Recd." (net < 0)
        # doesn't get re-matched as an expense via this generic loop.
        if head_key == "finance_cost" and net_amount <= 0:
            continue
        head = BS_HEADS[head_key]
        # Check negative keywords first (exclusions)
        excluded = False
        for nk in head.get("negative_keywords", []):
            if nk in name_lower:
                excluded = True
                break
        if excluded:
            continue

        # Check positive keywords
        for kw in head["keywords"]:
            if kw in name_lower:
                return head_key, "high"

    # Step 3: Partial group match (also handles the bare generic groups
    # deferred above, since "current assets" / "current liabilities" are
    # present in GROUP_HEAD_MAP and will match here as a substring of
    # themselves if Step 2's keyword check didn't already return).
    if group_lower:
        for grp_key, head_key in GROUP_HEAD_MAP.items():
            if grp_key in group_lower or group_lower in grp_key:
                return head_key, "low"

    # Step 4: Use net amount sign as hint
    # Debit balance (positive net) → likely asset or expense
    # Credit balance (negative net) → likely liability or income
    if net_amount > 0:
        return "other_current_assets", "low"
    elif net_amount < 0:
        return "other_cl", "low"
    else:
        return "unclassified", "none"


def get_aggregated_values(classified_accounts):
    """
    Aggregate classified accounts into BS head totals.
    Returns dict: {head_key: total_amount}

    Sign-aware reclassification rules (Indian accounting standards):
    - Creditor with DEBIT balance (net > 0) → advance to supplier → stla
    - Debtor with CREDIT balance (net < 0)  → advance from customer → other_cl
    These accounts are reclassified before aggregation.
    """
    totals = {}
    for acct in classified_accounts:
        head = acct["bs_head"]
        net  = acct["net"]

        if head == "unclassified":
            continue

        # ── Sign-aware auto-reclassification ──────────────────────────
        # Creditor with debit balance = advance paid to supplier (asset).
        # Aggregate into stla, but flag with explicit UI head
        # "advance_to_supplier" so the Review-Mapping page can show it.
        if head == "trade_payables" and net > 0:
            head = "stla"
            acct["bs_head"]    = "advance_to_supplier"
            acct["reclassified_from"] = "trade_payables"

        # Debtor with credit balance = advance received from customer (liability)
        # Aggregate into other_cl, flag as "advance_from_customer".
        elif head == "trade_rec" and net < 0:
            head = "other_cl"
            acct["bs_head"]    = "advance_from_customer"
            acct["reclassified_from"] = "trade_rec"

        # Provision with DEBIT balance = receivable/advance (asset not liability)
        # e.g. TCS GST A/C (debit) = TCS paid to govt = advance to revenue authority
        elif head == "st_provisions" and net > 0:
            head = "stla"
            acct["bs_head"]      = "stla"
            acct["reclassified_from"] = "st_provisions"
            acct["stla_subtype"]  = "revenue_authority"  # flag for D129 row

        # stla accounts that are TCS/TDS by name also get revenue_authority subtype
        # (covers accounts classified via group "duties & taxes" → stla directly,
        #  bypassing the st_provisions reclassification path above)
        elif head == "stla" and net > 0:
            name_l = acct.get("name", "").lower()
            if "tcs" in name_l and not acct.get("stla_subtype"):
                acct["stla_subtype"] = "revenue_authority"
            elif "excess tds" in name_l and not acct.get("stla_subtype"):
                acct["stla_subtype"] = "revenue_authority"

        # Provision with CREDIT balance but name doesn't match genuine provisions
        # → reclassify to other_cl to prevent spurious amounts in ST Provisions.
        # Genuine provisions: tax, gratuity, bonus, leave, warranty.
        elif head == "st_provisions" and net < 0:
            VALID_PROVISION_KEYWORDS = {
                "provision for tax", "provision for income",
                "provision for gratuity", "provision for bonus",
                "provision for leave", "provision for warranty",
                "provision for bad", "provision for doubtful",
                "short term provision",
            }
            name_l = acct.get("name", "").lower()
            if not any(kw in name_l for kw in VALID_PROVISION_KEYWORDS):
                head = "other_cl"
                acct["bs_head"] = "other_cl"
                acct["reclassified_from"] = "st_provisions"

        # ── Explicit user-mapped advance heads ─────────────────────────
        # When the user picks "Advance from Customer" / "Advance to Supplier"
        # from the dropdown, the amount must feed the parent BS bucket while
        # the injector still knows the origin via `reclassified_from`.
        elif head == "advance_from_customer":
            head = "other_cl"
            acct["reclassified_from"] = acct.get("reclassified_from") or "trade_rec"

        elif head == "advance_to_supplier":
            head = "stla"
            acct["reclassified_from"] = acct.get("reclassified_from") or "trade_payables"

        amt = abs(net)
        if head not in totals:
            totals[head] = 0
        totals[head] += amt
    return totals


# ═══════════════════════════════════════════════════════════════════════
# BS TEMPLATE INJECTION — Notes-Aware Engine
# ═══════════════════════════════════════════════════════════════════════
# This BS template is 100% formula-driven on the bs sheet.
# Every cell like E7, E11, E17 etc. on the bs sheet pulls from:
#   capital!G10       → owners capital closing balance
#   'notes to bs'!D20 → long-term borrowings total
#   'notes to bs'!D48 → trade payables (creditors total)
#   'notes to bs'!D69 → other current liabilities total
#   'Fixed Assets C. Yr.'!I31 → PPE net block
#   'notes to p&l'!D24 → inventories (closing stock)
#   'notes to bs'!D104 → trade receivables total
#   'notes to bs'!D116 → cash & bank total
#   'notes to bs'!D133 → short-term loans & advances
#
# Strategy: inject into the SOURCE sheets (the plain-value cells),
# never into the formula cells on the bs sheet.
# The formulas then auto-propagate values to bs and other sheets.

def _is_formula(val):
    return isinstance(val, str) and val.strip().startswith("=")


def _get_writable_cell(ws, row, col):
    """
    Return the writable cell at (row, col).
    If the target is a MergedCell, return the top-left anchor of the merged range
    (which is the only writable cell in that range). If no merged range is found,
    returns None to indicate the cell cannot be written.
    """
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return cell
    # Cell is part of a merged range — find the top-left anchor
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= col <= merged_range.max_col):
            anchor = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            if not isinstance(anchor, MergedCell):
                return anchor
            return None
    return None


def _safe_set(ws, row, col, value):
    """
    Write a plain numeric value only if the target cell is not a formula.
    MergedCell-aware: if the target is a MergedCell, write to the top-left anchor
    of its merged range instead (provided the anchor isn't a formula).
    Returns True on success, False otherwise.
    """
    try:
        cell = _get_writable_cell(ws, row, col)
        if cell is None:
            return False
        if _is_formula(cell.value):
            return False
        cell.value = round(float(value), 2)
        return True
    except (AttributeError, TypeError, ValueError):
        # Defensive: never let a single write blow up the entire pipeline
        return False


def _safe_write(ws, row, col, value):
    """
    Write any value (number, string, etc.) to the target cell safely.
    MergedCell-aware (writes to the merged range's anchor cell).
    Skips formula cells. Returns True on success, False otherwise.
    Use this in place of direct `ws.cell(r, c).value = ...` assignments.
    """
    try:
        cell = _get_writable_cell(ws, row, col)
        if cell is None:
            return False
        if _is_formula(cell.value):
            return False
        if isinstance(value, (int, float)):
            cell.value = round(float(value), 2)
        else:
            cell.value = value
        return True
    except (AttributeError, TypeError, ValueError):
        return False


def _detect_notes_structure(wb):
    """
    Scan the Notes sheets and build an injection map:
    { head_key: [(sheet_name, row, col, label), ...] }
    This maps each BS head to the plain-value target cell(s) where
    we should write the aggregated amount.

    Returns (injection_map, details) where details contains per-head
    info about what was found.
    """
    # ── Capital sheet ────────────────────────────────────────────────
    # The capital closing balance is computed by the formula:
    #   G10 = C10+D10-E10+F10   (opening + intro - drawings + profit)
    # We can only inject into C8 (opening), D8 (intro), E8 (drawings).
    # Profit (F8) comes from p&l sheet — we cannot inject that.
    # For a fresh TB injection the cleanest approach is:
    #   Write the closing balance from TB directly into a single
    #   "lump sum" row that doesn't break the formula structure.
    # We inject into capital!D8 (Capital Introduced) as a net plug
    # if the opening balance is already in C8 via =G11 formula.
    # In practice: set capital!D8 = (TB_closing - capital!C8_resolved)
    # But since C8=G11=prev year closing which is read-only, safest is:
    # Write total capital to capital!G10 if it's not a formula,
    # else find a writable row.

    capital_map = []
    if "capital" in wb.sheetnames:
        ws_cap = wb["capital"]
        # Scan rows 8-15 for plain-value cells in column G (closing bal)
        for r in range(8, 16):
            v = ws_cap.cell(r, 7).value
            if v is not None and not _is_formula(v):
                capital_map.append(("capital", r, 7, "Closing Balance"))
                break
        # If G10 is a formula (=C10+D10-E10+F10), inject into D8 (Capital Introduced)
        # after zeroing withdrawals so closing = opening + introduced
        if not capital_map:
            # We'll use a special "capital_plug" strategy
            capital_map.append(("capital", "plug", None, "Capital plug via D8"))

    # ── notes to bs ─────────────────────────────────────────────────
    notes_map = {}
    if "notes to bs" in wb.sheetnames:
        ws_n = wb["notes to bs"]
        for row in ws_n.iter_rows(min_row=1, max_row=200, min_col=2, max_col=4, values_only=True):
            b_val, _, d_val = row
            if b_val is not None:
                r = ws_n.cell(1,1).row  # placeholder
                break
        # Fast scan using iter_rows
        for r_idx, row in enumerate(ws_n.iter_rows(min_row=1, max_row=200,
                                                     min_col=2, max_col=4), 1):
            b_val = row[0].value; d_val = row[2].value
            label = str(b_val).strip().lower() if b_val else ""
            if d_val is not None and not _is_formula(d_val):
                notes_map[r_idx] = (label, d_val)

    # ── Details sheet ────────────────────────────────────────────────
    details_map = {}
    if "Details" in wb.sheetnames:
        ws_det = wb["Details"]
        for r_idx, row in enumerate(ws_det.iter_rows(min_row=1, max_row=200,
                                                       min_col=2, max_col=4), 1):
            b_val = row[0].value; d_val = row[2].value
            if b_val is not None and not _is_formula(d_val if d_val is not None else "x"):
                details_map[r_idx] = (str(b_val).strip().lower(), d_val)

    # ── GROSS PROFIT / notes to p&l for inventory ────────────────────
    gp_map = {}
    if "GROSS PROFIT" in wb.sheetnames:
        ws_gp = wb["GROSS PROFIT"]
        for r in range(1, 30):
            b_val = ws_gp.cell(r, 4).value  # E-side labels
            e_val = ws_gp.cell(r, 5).value
            if b_val and "closing stock" in str(b_val).lower() and e_val is not None and not _is_formula(e_val):
                gp_map["closing_stock"] = ("GROSS PROFIT", r, 5)
                break

    return {
        "capital_map": capital_map,
        "notes_map": notes_map,
        "details_map": details_map,
        "gp_map": gp_map,
    }


def _fuzzy_match_name(tb_name, template_name):
    """Check if two party names roughly match (for creditor/debtor matching)."""
    import re

    # Finance-cost abbreviation aliases: Tally short names → normalised forms
    # so "BANK CC INTT" matches a template row labelled "Bank Interest".
    FINANCE_ALIASES = {
        "bank cc intt":       "bank interest",
        "bank cc interest":   "bank interest",
        "bank od intt":       "bank interest",
        "bank od interest":   "bank interest",
        "cc interest":        "bank interest",
        "od interest":        "bank interest",
        # "bank charges and interest" is NOT a loan interest — it's a bank service charge
        # and belongs in other_expenses. Do NOT alias it to "bank interest".
    }

    def normalize(s):
        s = s.lower().strip()
        # Apply finance alias before generic normalisation
        if s in FINANCE_ALIASES:
            s = FINANCE_ALIASES[s]
        s = re.sub(r'\bm/s\.?\s*', '', s)
        # FIX (Bug 5): strip city/location suffixes that appear in TB names
        # (e.g. "A.G. ENTERPRISES, LUDHIANA" → "A.G. ENTERPRISES") BEFORE
        # removing punctuation, so the city removal regex works on the word
        # boundary correctly.
        for city in ['ludhiana', 'delhi', 'jalandhar', 'surat', 'ahmedabad',
                     'ahemadabad', 'varanasi', 'mumbai', 'ambala', 'citi',
                     'chandigarh', 'patiala', 'amritsar', 'bathinda', 'moga',
                     'gurugram', 'gurgaon', 'noida', 'faridabad', 'kolkata',
                     'chennai', 'hyderabad', 'pune', 'bangalore', 'bengaluru']:
            s = re.sub(r',?\s*\b' + city + r'\b', '', s).strip()
        # FIX (Bug 5): Remove dots from initials before stripping punctuation,
        # so "A.G." → "AG", "A.K." → "AK", "Pvt." → "Pvt" etc.
        # This prevents "a g" ≠ "ag" mismatches that broke creditor/debtor
        # matching when the TB used dot-separated initials but the template
        # had them run together.
        s = re.sub(r'(?<=[a-z])\.(?=[a-z])', '', s)  # e.g. "a.g." → "ag"
        s = re.sub(r'(?<=\w)\.(?=\s)', ' ', s)        # trailing dot on words
        s = re.sub(r'[^a-z0-9\s]', ' ', s)
        s = re.sub(r'\s+', ' ', s).strip()
        return s

    a = normalize(tb_name)
    b = normalize(template_name)
    if not a or not b:
        return False
    # Exact after normalization
    if a == b:
        return True
    # Simple singular/plural stemming: "salary" should match "salaries",
    # "expense" should match "expenses", etc. Without this, a TB account
    # like "SALARY" never matches a template label "Salaries" and falls
    # through to "find next empty row" — which can silently duplicate an
    # amount that's ALREADY pre-filled in the template under a slightly
    # different spelling, double-counting it in the section's Total.
    def _stem(w):
        if w.endswith("ies") and len(w) > 4:
            return w[:-3] + "y"
        if w.endswith("es") and len(w) > 3:
            return w[:-2]
        if w.endswith("s") and len(w) > 3:
            return w[:-1]
        return w
    a_words = a.split()
    b_words = b.split()
    if len(a_words) == 1 and len(b_words) == 1:
        if _stem(a_words[0]) == _stem(b_words[0]):
            return True
    # Only allow substring if the shorter is at least 6 chars
    # AND the match is not just a common word like 'textiles'
    COMMON_WORDS = {'textiles', 'trading', 'enterprises', 'creation', 'fashion',
                    'fabrics', 'industries', 'pvt', 'ltd', 'co', 'and', 'sons',
                    # FIX: "products", "store", "traders", "corporation",
                    # "international" and similar generic business-type
                    # words appear across many UNRELATED company names —
                    # treating any one of them as "distinctive" caused
                    # false-positive matches, e.g. "PARAS MIRACLE POLY
                    # PRODUCTS PVT. LTD." incorrectly matching "G.M.
                    # Products" purely because both contain "Products",
                    # silently swapping the two companies' amounts with
                    # each other.
                    'products', 'store', 'stores', 'traders', 'corporation',
                    'international', 'company', 'industry', 'general',
                    'national', 'private', 'limited', 'agency', 'agencies',
                    # Textile/garment industry generics — every second account
                    # name ends with one of these, so they are NOT distinctive.
                    'knitwears', 'knitwear', 'knit', 'garments', 'garment',
                    'readymade', 'hosiery', 'collection', 'collections',
                    'wool', 'woollen', 'fab', 'fabs', 'impex', 'trendz',
                    'house', 'super', 'bazar', 'bazaar', 'shop', 'centre',
                    'center', 'emporium', 'imporium',
                    # FIX (Bug 3 — Details debtor/creditor false positives):
                    # Generic business-type suffixes that appear in almost
                    # every company name in printing/packaging/pharma sectors.
                    # Previously "ABHINANDAN PRINTERS" matched "Sahil Printers."
                    # because "printers" (8 chars ≥ 7) passed the single-long-
                    # word test — the only shared word between two completely
                    # different companies. Adding these prevents such matches.
                    'printers', 'printer', 'printing', 'press', 'presses',
                    'packaging', 'packers', 'packing', 'packwell',
                    'forgings', 'forging', 'forge', 'stamping',
                    'pharmaceuticals', 'pharma', 'meditech', 'biotech',
                    'herbal', 'healthcare', 'remedies',
                    'engineering', 'engineers', 'electrical', 'electronics',
                    'mechanical', 'automation',
                    'associates', 'partnership', 'syndicate', 'syndication',
                    'brothers', 'distributors', 'supplier', 'suppliers',
                    'solutions', 'services', 'service',
                    'paper', 'papers', 'board', 'boards', 'stationery',
                    'label', 'labels', 'offset', 'graphics', 'graphic',
                    'creations', 'creators', 'designs', 'designers',
                    'exporters', 'importers', 'exports', 'imports',
                    'hospital', 'hospitals', 'clinic', 'clinics',
                    'foods', 'food', 'agro', 'organic', 'organics',
                    'steel', 'steels', 'alloys', 'metals', 'metal',
                    'tubes', 'tube', 'pipes', 'pipe', 'wires', 'wire',
                    }
    if len(a) >= 6 and len(b) >= 6:
        shorter, longer = (a, b) if len(a) <= len(b) else (b, a)
        if shorter in longer:
            # Verify the match isn't just on common words
            unique_words = [w for w in shorter.split() if w not in COMMON_WORDS and len(w) > 3]
            if unique_words:
                return True
    # Check first 2+ significant words match exactly.
    # Require BOTH names to have at least 2 significant words for this
    # path — otherwise a single short shared word (e.g. "Goel" appearing
    # in both "Anjali Goel" and "Goel Trading Co.") would wrongly count
    # as a full match via min(2, 1, 2) = 1.
    words_a = [w for w in a.split() if len(w) > 3 and w not in COMMON_WORDS]
    words_b = [w for w in b.split() if len(w) > 3 and w not in COMMON_WORDS]
    if len(words_a) >= 2 and len(words_b) >= 2:
        common = sum(1 for w in words_a if w in words_b)
        if common >= 2:
            return True
    # Single distinctive word match (>=7 chars, not common)
    if words_a and words_b:
        long_a = [w for w in words_a if len(w) >= 7]
        long_b = [w for w in words_b if len(w) >= 7]
        if any(w in long_b for w in long_a):
            return True
    return False


def _col_letter(col_idx):
    """Convert 0-indexed column to letter."""
    from openpyxl.utils import get_column_letter
    return get_column_letter(col_idx + 1)



def _load_sheet_cache(wb, sheet_name, max_row=200, max_col=10):
    """Pre-load all cell values into dict for fast O(1) lookup."""
    cache = {}
    if sheet_name not in wb.sheetnames:
        return cache
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                cache[(cell.row, cell.column)] = cell.value
    return cache


def inject_into_bs(bs_template_path, output_path, aggregated_values,
                   mapping_overrides=None, individual_accounts=None):
    """
    Inject aggregated TB values into BS template by writing to SOURCE sheets.

    The BS sheet itself is all formulas — writing there does nothing.
    Instead we write to:
      capital      → owner's capital closing balance
      notes to bs  → long-term loans, trade payables, OCL, cash, receivables,
                     short-term loans
      Details      → individual creditor/debtor amounts (matched by name)
      GROSS PROFIT → closing stock (inventories)
      Fixed Assets C. Yr. → asset additions (col C = >180 days)

    individual_accounts: list of {name, bs_head, net, debit, credit}
    from the classified TB accounts — used for name-matching creditors/debtors.
    """
    import shutil
    log = []

    # Copy template to output first
    shutil.copy2(bs_template_path, output_path)

    wb = load_workbook(output_path)

    # FIX: some uploaded BS templates contain leftover ArrayFormula cells
    # referencing an EXTERNAL workbook that no longer exists (e.g.
    # "=[1]!'!A D Garments!R21C18:R24C18'" — the "[1]!" prefix is an
    # external-link index pointing at a file that isn't embedded or
    # available). Excel can't resolve these and shows #VALUE! every time
    # the file is reopened, which then cascades through any downstream
    # formula that depends on that cell (Cost of Materials Consumed →
    # Total Expenses → Profit → Capital → the Previous Year balance
    # sheet column) — almost certainly the cause of "previous year
    # balance sheet doesn't match" reports for templates carrying this
    # kind of dead reference. openpyxl still preserves the LAST cached
    # value Excel computed before the link broke (visible via
    # data_only=True), so the safest repair is to replace the broken
    # array formula with that last-known plain number, removing the
    # dead link while keeping the figure the template already expected.
    try:
        from openpyxl.worksheet.formula import ArrayFormula as _ArrayFormula
        wb_cached = load_workbook(output_path, data_only=True)
        _broken_links_fixed = 0
        _broken_refs_fixed = 0
        for _sn in wb.sheetnames:
            _ws = wb[_sn]
            _ws_cached = wb_cached[_sn] if _sn in wb_cached.sheetnames else None
            for _row in _ws.iter_rows():
                for _cell in _row:
                    if isinstance(_cell.value, _ArrayFormula):
                        _formula_text = getattr(_cell.value, "text", "") or ""
                        if "[1]!" in _formula_text or "[2]!" in _formula_text or "[3]!" in _formula_text:
                            _cached_val = None
                            if _ws_cached is not None:
                                _cached_val = _ws_cached.cell(_cell.row, _cell.column).value
                            if isinstance(_cached_val, (int, float)):
                                _cell.value = _cached_val
                                _broken_links_fixed += 1
                    elif isinstance(_cell.value, str) and _cell.value.startswith("=") \
                            and "#REF!" in _cell.value:
                        # FIX: plain string formulas containing a #REF!
                        # fragment (e.g. "='notes to p&l'!#REF!") happen
                        # when a row/column the formula pointed to was
                        # deleted from the source sheet at some point in
                        # the template's history. These always evaluate
                        # to #VALUE!/#REF! and never contribute anything
                        # meaningful. Only touch cells where this matters
                        # for arithmetic (not label/text columns A and D,
                        # which sometimes legitimately hold a cross-sheet
                        # text-label formula) — replace with the last
                        # cached value if numeric, otherwise 0, so totals
                        # depending on the column don't show #VALUE!.
                        if _cell.column != 1:
                            _cached_val = None
                            if _ws_cached is not None:
                                _cached_val = _ws_cached.cell(_cell.row, _cell.column).value
                            _cell.value = _cached_val if isinstance(_cached_val, (int, float)) else 0
                            _broken_refs_fixed += 1
        if _broken_links_fixed:
            log.append(
                f"⚠ Repaired {_broken_links_fixed} broken external-link "
                f"formula(s) in the uploaded template (dead reference to "
                f"a missing source workbook) — replaced with last-known "
                f"cached value(s) so they no longer show #VALUE! on open."
            )
        if _broken_refs_fixed:
            log.append(
                f"⚠ Repaired {_broken_refs_fixed} broken #REF! formula(s) "
                f"in value columns of the uploaded template (a row/column "
                f"the formula pointed to no longer exists) — replaced "
                f"with last-known cached value(s) or 0."
            )
    except Exception as _link_fix_exc:
        pass

    structure = _detect_notes_structure(wb)
    injected = []
    skipped = []

    # Queue for any row-insertion requests that must be deferred until
    # every other (cache-dependent) injection step has finished reading
    # and writing — see the detailed explanation at the sale-section
    # overflow handling below for why this matters.
    _deferred_section_expansions = []

    # Pre-load sheet caches for fast lookup (avoids slow ws.cell(r,c) in loops)
    _cache = {}
    for _sn in ["notes to bs", "notes to p&l", "Details", "GROSS PROFIT",
                "Fixed Assets C. Yr.", "capital"]:
        _cache[_sn] = _load_sheet_cache(wb, _sn, max_row=250, max_col=12)

    # Resolved (data_only) cache for "notes to p&l" — used where we need a
    # formula cell's CACHED VALUE (e.g. D5 = "=10898146+793859" -> 11692005)
    # to compare against a TB total, rather than the formula text itself.
    _npl_cache_do = {}
    try:
        _wb_do_tmp = load_workbook(bs_template_path, data_only=True)
        _npl_cache_do = _load_sheet_cache(_wb_do_tmp, "notes to p&l", max_row=250, max_col=12)
        _wb_do_tmp.close()
    except Exception:
        pass

    # ────────────────────────────────────────────────────────────────
    # 1. CAPITAL — DO NOT inject from TB.
    #    TB only has closing balance. Additions/withdrawals come from
    #    ledger. User's BS template capital sheet is preserved as-is.
    # ────────────────────────────────────────────────────────────────
    capital_amt = aggregated_values.get("capital", 0)
    if capital_amt:
        log.append(f"· Capital from TB: {capital_amt:,.2f} — skipped (fill from ledger)")

    # ────────────────────────────────────────────────────────────────
    # 2. NOTES TO BS — inject by scanning plain-value cells
    # ────────────────────────────────────────────────────────────────
    if "notes to bs" in wb.sheetnames:
        ws_n = wb["notes to bs"]

        def inject_notes_row(row, col, amount, label):
            if _safe_set(ws_n, row, col, amount):
                injected.append(f"notes to bs!{_col_letter(col-1)}{row} ({label}) = {amount:,.2f}")
                return True
            else:
                skipped.append(f"notes to bs R{row}C{col} ({label}) is formula")
                return False

        # Long-term borrowings → match individual loan accounts to template rows
        ltb_amt = aggregated_values.get("lt_borrowings", 0)
        if ltb_amt and individual_accounts:
            ltb_accounts = [a for a in individual_accounts
                            if a.get("bs_head") == "lt_borrowings"
                            and abs(a.get("net", 0)) > 0]

            # FIX: Defend against re-processing a previously-generated
            # (and possibly buggy) output file as if it were a pristine
            # template. Some uploaded "templates" already contain stray
            # numeric values in UNLABELED rows of the LT-borrowings section
            # (col B empty, col D has a leftover number from an earlier
            # run) which get silently summed by the section's Total
            # formula alongside our freshly-injected figures, double-
            # counting that amount. Clear any such label-less numeric cell
            # in rows 7-24 of the LT borrowings section before injecting —
            # a legitimate template row always has a label in col B.
            for _r in range(7, 25):
                _b = ws_n.cell(_r, 2).value
                _d = ws_n.cell(_r, 4).value
                if (_b is None or str(_b).strip() == "") and isinstance(_d, (int, float)) and _d != 0:
                    if not _is_formula(_d):
                        ws_n.cell(_r, 4).value = None
                        log.append(
                            f"· Cleared stray unlabeled value at notes to bs!D{_r} "
                            f"({_d:,.2f}) — likely leftover from a previous run"
                        )

            # Separate secured vs unsecured based on Tally group
            def _is_unsecured(acct):
                g = (acct.get("group") or "").lower()
                n = (acct.get("name") or "").lower()
                return "unsecure" in g or "unsecure" in n

            secured_accounts   = [a for a in ltb_accounts if not _is_unsecured(a)]
            unsecured_accounts = [a for a in ltb_accounts if _is_unsecured(a)]

            # Build template label → row map for LT borrowing section
            ltb_template = {}
            for r in range(7, 24):
                b_val = ws_n.cell(r, 2).value
                if b_val and isinstance(b_val, str) and len(b_val.strip()) > 2:
                    lbl = b_val.strip().lower()
                    # Exclude section headers but NOT loan names that happen to contain 'from'
                    HEADER_LABELS = {'from banks', 'from related parties', 'from other parties',
                                     'secured', 'unsecured', 'term loans'}
                    if 'total' in lbl or lbl.strip() in HEADER_LABELS:
                        continue
                    ltb_template[r] = lbl

            # Find the unsecured section start row (row after "Unsecured" header)
            unsecured_start = None
            secured_end = None
            for r in range(7, 24):
                b_val = ws_n.cell(r, 2).value
                if b_val and isinstance(b_val, str):
                    bl = b_val.strip().lower()
                    if 'unsecured' in bl and 'total' not in bl:
                        unsecured_start = r + 1
                    if 'total' in bl and 'secured' in bl and 'unsecured' not in bl and secured_end is None:
                        secured_end = r
            if unsecured_start is None:
                unsecured_start = 15  # fallback
            if secured_end is None:
                secured_end = 11

            written_rows = set()

            def _inject_ltb_list(acct_list, fallback_start, fallback_end):
                # Sort by name length DESC so more specific accounts (longer names) match first
                # This prevents "Axis ML3" from stealing a row meant for a more specific name
                for acct in sorted(acct_list, key=lambda a: -len(a["name"])):
                    amt  = abs(acct["net"])
                    name = acct["name"]
                    matched_row = None
                    # Try exact/fuzzy match — but avoid matching generic labels to wrong loans
                    for r, lbl in sorted(ltb_template.items()):
                        if r in written_rows:
                            continue
                        if _fuzzy_match_name(name, lbl):
                            matched_row = r
                            break
                    if matched_row:
                        if _safe_set(ws_n, matched_row, 4, amt):
                            written_rows.add(matched_row)
                            injected.append(f"notes to bs!D{matched_row} ({name}) = {amt:,.2f}")
                    else:
                        # No match — find empty row in the correct section
                        for r in range(fallback_start, fallback_end + 5):
                            if r in written_rows:
                                continue
                            d_val = ws_n.cell(r, 4).value
                            b_val = ws_n.cell(r, 2).value
                            if (d_val is None or d_val == 0) and (b_val is None or str(b_val).strip() == ""):
                                _safe_set(ws_n, r, 2, name)
                                if _safe_set(ws_n, r, 4, amt):
                                    written_rows.add(r)
                                    injected.append(f"notes to bs!D{r} (new: {name}) = {amt:,.2f}")
                                break

            # Secured → rows 7 to secured_end-1 as fallback
            _inject_ltb_list(secured_accounts, 7, secured_end - 1)

            # ── FIX: Unsecured loans often live in a dedicated "Details"
            # sheet section (e.g. "UNSECURED LOAN" / "FROM RELATED PARTIES")
            # with one row PER LENDER NAME, summed via a Total formula that
            # 'notes to bs' references (e.g. notes to bs!D18='=Details!D16').
            # The generic _inject_ltb_list() above only targets 'notes to bs'
            # directly — since that target cell is itself a formula pointing
            # to Details, the write gets silently skipped and unsecured loans
            # (e.g. 9 individually-named lenders) never reach the BS at all.
            #
            # Try matching unsecured accounts by NAME to a "Details" sheet
            # section first; only fall back to the generic notes-to-bs
            # injection for any names that don't match.
            unsecured_remaining = list(unsecured_accounts)
            if "Details" in wb.sheetnames and unsecured_accounts:
                ws_det_ltb = wb["Details"]
                # Find a section whose header mentions "unsecured loan"
                # and locate its name-labeled rows + Total row.
                section_start, section_end = None, None
                for r in range(1, min(ws_det_ltb.max_row, 60) + 1):
                    a_val = ws_det_ltb.cell(r, 1).value
                    if a_val and "unsecured loan" in str(a_val).strip().lower():
                        section_start = r
                        break
                if section_start:
                    for r in range(section_start, section_start + 30):
                        b_val = ws_det_ltb.cell(r, 2).value
                        if b_val and str(b_val).strip().lower() == "total":
                            section_end = r
                            break
                if section_start and section_end:
                    still_unmatched_ltb = []
                    for acct in unsecured_remaining:
                        name = acct["name"]
                        amt = abs(acct["net"])
                        matched_r = None
                        for r in range(section_start, section_end):
                            b_val = ws_det_ltb.cell(r, 2).value
                            if b_val and _fuzzy_match_name(name, str(b_val)):
                                matched_r = r
                                break
                        if matched_r:
                            if _safe_write(ws_det_ltb, matched_r, 4, amt):
                                injected.append(
                                    f"Details!D{matched_r} (unsecured loan: {name}) = {amt:,.2f}"
                                )
                                written_rows.add(("details_ltb", matched_r))
                                continue
                        still_unmatched_ltb.append(acct)
                    unsecured_remaining = still_unmatched_ltb

            # Unsecured → rows unsecured_start onward as fallback (any
            # accounts that didn't match a Details-sheet name row above)
            _inject_ltb_list(unsecured_remaining, unsecured_start, unsecured_start + 5)

            if not written_rows and ltb_amt:
                # Fallback: write total to "from banks" row
                for r in range(7, 11):
                    d = ws_n.cell(r, 4).value
                    if d is None or not _is_formula(str(d)):
                        inject_notes_row(r, 4, abs(ltb_amt), "Long-term borrowings total")
                        break

        # Short-term borrowings → find the labeled CC/OD row dynamically.
        # FIX: previously hardcoded to "first writable row from 26", which
        # often lands on a SECTION HEADER row (e.g. R26="Secured") that has
        # no value cell wired into the Total formula chain feeding bs!E15.
        # The actual writable target is the row labeled with the bank/CC/OD
        # account name (e.g. "HDFC Bank CC A/c"), found by scanning for a
        # label containing cc/od/overdraft/cash credit/loan repayable.
        stb_amt = aggregated_values.get("st_borrowings", 0)
        if stb_amt:
            placed = False
            target_row = None
            # First pass: prefer a row with a SPECIFIC bank/CC account name
            # (e.g. "HDFC Bank CC A/c") over a generic section label like
            # "Loans repayable on demand", since the generic label row is
            # often just a sub-header with no value of its own — the real
            # writable cell is the specific account name row beneath it.
            for r in range(7, 50):
                b_val = ws_n.cell(r, 2).value
                if not b_val:
                    continue
                bl = str(b_val).strip().lower()
                if "total" in bl:
                    continue
                if any(kw in bl for kw in ["cc a/c", "cc account", "overdraft", "cash credit", "od a/c"]):
                    target_row = r
                    break
            if target_row is None:
                # Second pass: fall back to the generic "loans repayable" label
                for r in range(7, 50):
                    b_val = ws_n.cell(r, 2).value
                    if not b_val:
                        continue
                    bl = str(b_val).strip().lower()
                    if "loans repayable on demand" in bl and "total" not in bl:
                        target_row = r
                        break
            if target_row:
                d_existing = ws_n.cell(target_row, 4).value
                if d_existing is None or not _is_formula(d_existing):
                    placed = inject_notes_row(target_row, 4, stb_amt, "Short-term borrowings")
            if not placed:
                # Fallback to old behaviour (scan from row 26) if no
                # specific CC/OD label was found in this template.
                for r in range(26, 32):
                    d = ws_n.cell(r, 4).value
                    b = ws_n.cell(r, 2).value
                    # Skip section-header-only rows (no specific account label)
                    if b and any(kw in str(b).lower() for kw in ["secured", "unsecured", "total"]):
                        continue
                    if d is None or (not _is_formula(str(d))):
                        placed = inject_notes_row(r, 4, stb_amt, "Short-term borrowings")
                        break
            if not placed:
                skipped.append(f"short_term_borrowings {stb_amt:,.2f}: no writable row in notes to bs")

        # ── Cheques Issued But Not Cleared → "Cheque Issued" labelled row ──
        # Only CREDIT-balance cheque accounts are liabilities (issued cheques).
        # Find the row by label (not hardcoded D68) so it works for any template.
        if individual_accounts:
            cheque_accounts = [a for a in individual_accounts
                               if ("cheque" in a.get("name","").lower()
                               or "chq" in a.get("name","").lower())
                               and a.get("net", 0) < 0]   # credit balance = liability
            total_cheques = sum(abs(a["net"]) for a in cheque_accounts)
            if total_cheques > 0:
                cheque_row = None
                for r in range(50, 100):
                    lbl = (ws_n.cell(r, 2).value or "").strip().lower()
                    if "cheque issued" in lbl or "chq issued" in lbl or "not presented" in lbl:
                        cheque_row = r
                        break
                if cheque_row:
                    ws_n.cell(cheque_row, 4).value = total_cheques
                    injected.append(f"notes to bs!D{cheque_row} (Cheques issued) = {total_cheques:,.2f}")
                else:
                    # Fallback to OCL section last named row
                    for r in range(55, 90):
                        lbl = (ws_n.cell(r, 2).value or "").strip().lower()
                        if "total" in lbl and "current" in lbl:
                            ws_n.cell(r-1, 4).value = total_cheques
                            injected.append(f"notes to bs!D{r-1} (Cheques issued fallback) = {total_cheques:,.2f}")
                            break

        # Other current liabilities → match by name to template rows
        ocl_amt = aggregated_values.get("other_cl", 0)
        if ocl_amt:
            # Detect OCL section boundaries
            ocl_start = None
            ocl_end = None
            for r in range(50, 100):
                b = ws_n.cell(r, 2).value
                if b and 'other current liabilit' in str(b).lower() and 'total' not in str(b).lower():
                    ocl_start = r + 1
                if ocl_start and b and 'total' in str(b).lower() and 'other' in str(b).lower():
                    ocl_end = r
                    break
            if not ocl_start:
                ocl_start = 59
            if not ocl_end:
                ocl_end = 67

            # Build template label → row map for OCL section
            ocl_template = {}
            for r in range(ocl_start, ocl_end):
                b = ws_n.cell(r, 2).value
                if b and isinstance(b, str) and len(b.strip()) > 2:
                    lbl = b.strip().lower()
                    if 'total' not in lbl:
                        ocl_template[r] = lbl

            if individual_accounts:
                ocl_accounts = [a for a in individual_accounts
                                if a.get("bs_head") == "other_cl"
                                and abs(a.get("net", 0)) > 0
                                # Exclude cheque accounts — already injected above
                                and "cheque" not in a.get("name","").lower()
                                and "chq" not in a.get("name","").lower()]
                written_ocl = set()
                
                # Phase 1: fuzzy match to template
                for acct in ocl_accounts:
                    for r, lbl in ocl_template.items():
                        if r in written_ocl:
                            continue
                        if _fuzzy_match_name(acct["name"], lbl):
                            if _safe_set(ws_n, r, 4, abs(acct["net"])):
                                written_ocl.add(r)
                                injected.append(f"notes to bs!D{r} (OCL: {acct['name']}) = {abs(acct['net']):,.2f}")
                            break
                
                # Phase 2: unmatched → aggregate into category rows (NEVER insert_rows)
                # Group unmatched accounts by category keyword and add to the best
                # matching named row. This avoids row-shifting entirely.
                CATEGORY_KEYWORDS_OCL = [
                    ("esi",          ["esi"]),
                    ("wages",        ["wages", "wage"]),
                    ("tds payable",  ["tds"]),
                    ("leave",        ["leave with wages", "leave"]),
                    ("bonus",        ["bonus"]),
                    ("audit",        ["audit"]),
                    ("professional", ["professional"]),
                    ("salary",       ["salary payable", "salary"]),
                ]
                # Accumulate unmatched amounts by category
                cat_pending: dict = {}
                for acct in ocl_accounts:
                    if any(_fuzzy_match_name(acct["name"],
                           ocl_template.get(r, "")) for r in written_ocl):
                        continue   # already matched in Phase 1
                    name_l = acct["name"].lower()
                    amt    = abs(acct["net"])
                    matched_cat = None
                    for cat, kws in CATEGORY_KEYWORDS_OCL:
                        if any(k in name_l for k in kws):
                            matched_cat = cat
                            break
                    if not matched_cat:
                        matched_cat = "other"
                    cat_pending[matched_cat] = cat_pending.get(matched_cat, 0) + amt

                # Write accumulated category totals to matching template rows
                for cat, amt in cat_pending.items():
                    if amt == 0:
                        continue
                    for r, lbl in ocl_template.items():
                        if r in written_ocl:
                            continue
                        cat_kws = next(
                            (kws for c, kws in CATEGORY_KEYWORDS_OCL if c == cat), []
                        )
                        if any(k in lbl for k in cat_kws):
                            existing = ws_n.cell(r, 4).value or 0
                            if not _is_formula(str(existing)):
                                new_val = (existing if isinstance(existing,(int,float)) else 0) + amt
                                if _safe_set(ws_n, r, 4, new_val):
                                    written_ocl.add(r)
                                    injected.append(
                                        f"notes to bs!D{r} (OCL cat '{cat}') = {new_val:,.2f}"
                                    )
                            break
                    else:
                        # No template row for this category — find first truly empty row
                        for r in range(ocl_start, ocl_end):
                            if r in written_ocl:
                                continue
                            d = ws_n.cell(r, 4).value
                            if d is None or d == 0:
                                if _safe_set(ws_n, r, 4, amt):
                                    written_ocl.add(r)
                                    injected.append(
                                        f"notes to bs!D{r} (OCL overflow '{cat}') = {amt:,.2f}"
                                    )
                                break
                        else:
                            skipped.append(
                                f"OCL category '{cat}' = {amt:,.2f}: no row — template full"
                            )

        # ── Cash & Bank — label-driven, no hardcoded row numbers ─────────────
        cash_bank_amt = aggregated_values.get("cash_bank", 0)
        if individual_accounts:
            cash_accounts = [a for a in individual_accounts
                             if a.get("bs_head") in ("cash_bank",)
                             and a.get("net", 0) > 0]
            cash_only = [a for a in cash_accounts if "cash" in a["name"].lower()]
            bank_only = [a for a in cash_accounts
                         if "cash" not in a["name"].lower()
                         and "cheque" not in a["name"].lower()]

            # Locate Cash in Hand row by scanning for the label (not hardcoded row 109)
            cash_in_hand_row = None
            for r in range(90, 170):
                lbl = (ws_n.cell(r, 2).value or "").strip().lower()
                if "cash in hand" in lbl or "cash on hand" in lbl or lbl == "cash":
                    cash_in_hand_row = r
                    break

            if cash_in_hand_row and cash_only:
                total_cash = sum(abs(a["net"]) for a in cash_only)
                # Always write cash (overwrite even formula PY values in CY col)
                ws_n.cell(cash_in_hand_row, 4).value = total_cash
                injected.append(
                    f"notes to bs!D{cash_in_hand_row} (Cash in hand) = {total_cash:,.2f}"
                )

            # Locate bank rows by scanning for bank name labels.
            # FIX: match SPECIFIC distinguishing words (account numbers,
            # "cc"/"od" suffixes) BEFORE generic words like "hdfc" — with
            # multiple accounts at the same bank (e.g. a savings A/c and a
            # separate CC A/c), matching on "hdfc" alone for both accounts
            # caused the second one to silently overwrite the first in the
            # same template row.
            written_bank_rows = set()
            for acct in bank_only:
                a_name_l = acct["name"].lower()
                # Specific words: digits-only tokens (account numbers) or
                # cc/od (even though short, these distinguish CC accounts)
                specific_words = [w for w in a_name_l.split()
                                  if w.isdigit() or w in ("cc", "od")]
                generic_words  = [w for w in a_name_l.split()
                                  if len(w) > 3 and w not in
                                  {"bank","a/c","ltd","pvt","the","current",
                                   "assets","accounts","balance","balances",
                                   "hand","cash","account","other","total"}
                                  and w not in specific_words]
                matched_row = None
                # Pass 1: specific words only
                if specific_words:
                    for r in range(90, 170):
                        if r in written_bank_rows:
                            continue
                        lbl = (ws_n.cell(r, 2).value or "").strip().lower()
                        if not lbl or "total" in lbl or "cash in hand" in lbl or "cash on hand" in lbl:
                            continue
                        if any(w in lbl for w in specific_words):
                            matched_row = r
                            break
                # Pass 2: generic words, only on rows not already claimed
                if matched_row is None and generic_words:
                    for r in range(90, 170):
                        if r in written_bank_rows:
                            continue
                        lbl = (ws_n.cell(r, 2).value or "").strip().lower()
                        if not lbl or "total" in lbl or "cash in hand" in lbl or "cash on hand" in lbl:
                            continue
                        if any(w in lbl for w in generic_words):
                            matched_row = r
                            break
                if matched_row is not None:
                    ws_n.cell(matched_row, 4).value = abs(acct["net"])
                    written_bank_rows.add(matched_row)
                    injected.append(
                        f"notes to bs!D{matched_row} ({acct['name']}) = {abs(acct['net']):,.2f}"
                    )
                    continue
                # Fallback: find empty bank row near section
                for r in range(90, 170):
                    if r in written_bank_rows:
                        continue
                    lbl = (ws_n.cell(r, 2).value or "").strip().lower()
                    col_a_val = str(ws_n.cell(r, 1).value or "").strip()
                    # Skip: section headers (A/B/C letters in col A), totals,
                    # "other bank balances" header rows, cash-in-hand rows
                    if col_a_val in ("A", "B", "C", "D"):
                        continue
                    if not lbl or "total" in lbl or "cash in hand" in lbl \
                            or "cash on hand" in lbl or lbl.startswith("other ") \
                            or "other bank" in lbl or "bank balance" in lbl:
                        continue
                    if any(k in lbl for k in ["bank", "a/c"]):
                        d = ws_n.cell(r, 4).value
                        if d is None or d == 0:
                            ws_n.cell(r, 4).value = abs(acct["net"])
                            written_bank_rows.add(r)
                            injected.append(
                                f"notes to bs!D{r} ({acct['name']}) = {abs(acct['net']):,.2f}"
                            )
                            break

        elif cash_bank_amt:
            # Fallback: find Cash in hand row by label
            for r in range(90, 170):
                lbl = (ws_n.cell(r, 2).value or "").strip().lower()
                if "cash in hand" in lbl:
                    ws_n.cell(r, 4).value = cash_bank_amt
                    injected.append(f"notes to bs!D{r} (Cash and bank lump) = {cash_bank_amt:,.2f}")
                    break

        # Short-term loans & advances → D128 (GST), D129 (TCS GST), D130 (Excess TDS)
        # Special case: "TDS claimable from Facebook" → D137 (Other current assets)
        stla_amt = aggregated_values.get("stla", 0)
        if stla_amt and individual_accounts:
            stla_accounts = [a for a in individual_accounts
                             if a.get("bs_head") == "stla"
                             and abs(a.get("net", 0)) > 0
                             and "cheque" not in a.get("name","").lower()
                             and "chq" not in a.get("name","").lower()
                             and a.get("reclassified_from") != "trade_payables"]
            for acct in stla_accounts:
                name_l = acct["name"].lower()
                placed = False

                # Find the STLA Revenue Authorities section dynamically
                # by scanning for label keywords (works for any template layout)
                def _find_rev_auth_row(keyword):
                    """Scan notes to bs for a row matching keyword under Revenue Authorities."""
                    in_rev_section = False
                    for r in range(110, 180):
                        lbl = (ws_n.cell(r, 2).value or "").strip().lower()
                        if "advance to revenue" in lbl or "revenue authorit" in lbl:
                            in_rev_section = True
                        if in_rev_section:
                            if keyword in lbl:
                                return r
                            if "total" in lbl and "a+b" in lbl:
                                break  # past section
                    return None

                # ── Facebook TDS → Other current assets ─────────────────────
                if "facebook" in name_l:
                    fb_row = None
                    for r in range(130, 180):
                        lbl = (ws_n.cell(r, 2).value or "").strip().lower()
                        if "facebook" in lbl or "recoverable" in lbl:
                            fb_row = r; break
                    if fb_row and _safe_set(ws_n, fb_row, 4, abs(acct["net"])):
                        injected.append(f"notes to bs!D{fb_row} (TDS from Facebook) = {abs(acct['net']):,.2f}")
                    placed = True

                # ── TCS → TCS row under Revenue Authorities ─────────────────
                elif "tcs" in name_l:
                    row = _find_rev_auth_row("tcs")
                    if row:
                        existing = ws_n.cell(row, 4).value
                        base = float(existing) if isinstance(existing,(int,float)) else 0
                        ws_n.cell(row, 4).value = base + abs(acct["net"])
                        injected.append(
                            f"notes to bs!D{row} (TCS) = {base+abs(acct['net']):,.2f}"
                        )
                        placed = True

                # ── GST (CGST/SGST/IGST/refund) → one row per account ───────
                # Check GST BEFORE TDS so "GST TDS 2024-25" goes to GST section
                elif any(k in name_l for k in ["gst","igst","cgst","sgst"]):
                    # Find next AVAILABLE (empty) GST row; each account gets its own row
                    placed_gst = False
                    for r in range(110, 180):
                        lbl = (ws_n.cell(r, 2).value or "").strip().lower()
                        if "advance to revenue" in lbl or "revenue authorit" in lbl:
                            for rr in range(r+1, r+15):
                                lbl2 = (ws_n.cell(rr, 2).value or "").strip().lower()
                                if "total" in lbl2: break
                                if "gst" in lbl2 or "igst" in lbl2 or "cgst" in lbl2 or "sgst" in lbl2:
                                    d = ws_n.cell(rr, 4).value
                                    # Only write to empty rows — gives each account its own slot
                                    if d is None or d == 0:
                                        ws_n.cell(rr, 4).value = abs(acct["net"])
                                        injected.append(f"notes to bs!D{rr} (GST: {acct['name']}) = {abs(acct['net']):,.2f}")
                                        placed_gst = True
                                        break
                            break
                    if placed_gst:
                        placed = True

                # ── TDS / Excess TDS → separate rows per account ─────────────
                elif "tds" in name_l:
                    # Use _find_rev_auth_row to get first matching TDS label row,
                    # but find an EMPTY one (each TDS account → its own row).
                    placed_tds = False
                    for r in range(110, 180):
                        lbl = (ws_n.cell(r, 2).value or "").strip().lower()
                        if "advance to revenue" in lbl or "revenue authorit" in lbl:
                            for rr in range(r+1, r+15):
                                lbl2 = (ws_n.cell(rr, 2).value or "").strip().lower()
                                if "total" in lbl2: break
                                # Match TDS-labelled rows (not GST rows)
                                if ("tds" in lbl2 or "excess tds" in lbl2) and \
                                        "gst" not in lbl2:
                                    d = ws_n.cell(rr, 4).value
                                    if d is None or d == 0:
                                        ws_n.cell(rr, 4).value = abs(acct["net"])
                                        injected.append(f"notes to bs!D{rr} (TDS: {acct['name']}) = {abs(acct['net']):,.2f}")
                                        placed_tds = True
                                        break
                            break
                    if placed_tds:
                        placed = True

                if not placed:
                    # Generic: find first empty row in STLA Revenue Auth area
                    for r in range(110, 180):
                        lbl = (ws_n.cell(r, 2).value or "").strip().lower()
                        if "advance to revenue" in lbl or "revenue authorit" in lbl:
                            for rr in range(r+1, r+15):
                                lbl2 = (ws_n.cell(rr, 2).value or "").strip().lower()
                                if "total" in lbl2:
                                    break
                                d = ws_n.cell(rr, 4).value
                                if d is None or d == 0:
                                    ws_n.cell(rr, 4).value = abs(acct["net"])
                                    injected.append(
                                        f"notes to bs!D{rr} ({acct['name']}) = {abs(acct['net']):,.2f}"
                                    )
                                    placed = True
                                    break
                            break
                    if not placed:
                        skipped.append(f"STLA '{acct['name']}' = {abs(acct['net']):,.2f}: no row found")
        elif stla_amt:
            # Lump sum fallback
            for r in range(110, 180):
                lbl = (ws_n.cell(r, 2).value or "").strip().lower()
                if "advance to revenue" in lbl or "gst" in lbl:
                    ws_n.cell(r, 4).value = stla_amt
                    injected.append(f"notes to bs!D{r} (STLA lump) = {stla_amt:,.2f}")
                    break

    # ────────────────────────────────────────────────────────────────
    # 3. DETAILS — individual creditor/debtor amounts (CACHE-BASED)
    # ────────────────────────────────────────────────────────────────
    if "Details" in wb.sheetnames and individual_accounts:
        ws_det  = wb["Details"]
        det_cache = _cache.get("Details", {})

        # FIX: this previously hardcoded range(21, 63) as "the creditor
        # section", assuming a fixed template layout. In templates where
        # the Unsecured Loan "FROM OTHER PARTIES" sub-section sits above
        # Sundry Creditors (e.g. AS Traders: rows 18-24 are unsecured
        # loans, "SUNDRY CREDITORS" header is at row 26), this caused
        # unmatched creditor names to be written into the BLANK rows of
        # the unsecured-loan sub-section instead of the real creditor
        # section — corrupting that section and shifting/hiding the
        # actual "SUNDRY CREDITORS" header and its rows beneath stray
        # data. Now locates the section dynamically by scanning for the
        # "SUNDRY CREDITORS" header text and using the row directly
        # below it as the start, falling back to the old hardcoded
        # range only if the header text can't be found at all.
        # ── CREDITOR SECTION DETECTION ──────────────────────────────────────
        #
        # FIX (2026-07-16) — Three layout variants handled:
        #
        # Variant A (Fashion Adda): SUM covers only named creditors.
        #   "Advance from Customers" is a SEPARATE section below debtors.
        #
        # Variant B (JEANS WORLD / SHREE CRAFT): "Advance from Customers"
        #   sub-section sits INSIDE the Sundry Creditors SUM range.
        #   e.g. JEANS WORLD: SUM(D21:D55) covers creditors (rows 21-27)
        #   AND advance-from-cust sub-rows (rows 28-55). The old code stopped
        #   at the "Advance from Customers" header (row 28) as an ALL-CAPS
        #   boundary → cred_end_row=28, excluding the advance-from-cust rows.
        #   New accounts that are advance-from-cust then had nowhere to go.
        #   FIX: detect the SUM formula row explicitly and set cred_end_row
        #   to ONE BEFORE the SUM row, covering all rows inside the formula.
        #
        # We track cred_total_row (the row holding the =SUM formula) for use
        # in the insert_rows overflow path below.
        cred_start_row, cred_end_row = 21, 63
        cred_total_row = None   # row that holds =SUM(...) covering the section

        for r in range(1, 300):
            lbl = det_cache.get((r, 1)) or det_cache.get((r, 2))
            if not (lbl and "sundry creditor" in str(lbl).strip().lower()):
                continue
            cred_start_row = r + 2  # skip header + PARTICULARS sub-header
            cred_end_row   = cred_start_row + 80  # safe fallback

            for r2 in range(cred_start_row, cred_start_row + 300):
                lbl2 = det_cache.get((r2, 1)) or det_cache.get((r2, 2))
                lbl2_s = str(lbl2).strip() if lbl2 else ""
                d2  = det_cache.get((r2, 4))
                e2  = det_cache.get((r2, 5))

                # Explicit TOTAL/SUM row — stop BEFORE it; record it for later.
                # The SUM row itself is the canonical end of the creditor block.
                _is_total = lbl2_s.lower() in ("total", "totals")
                _is_sum   = (isinstance(d2, str) and "sum" in d2.lower()) or \
                            (isinstance(e2, str) and "sum" in e2.lower())
                if _is_total or _is_sum:
                    cred_end_row  = r2          # SUM row itself (excluded from data scan)
                    cred_total_row = r2
                    break

                # Mixed-case sub-section headers inside the creditor block
                # (e.g. "Advance from Customers", "Due to MSME Creditors").
                # OLD behaviour: stop here → excluded those rows from the section.
                # NEW behaviour: keep scanning — they're inside the SUM range.
                # We ONLY stop for a fully-capitalised section header that starts
                # an entirely new top-level section AND has no amount in D or E
                # (i.e. it's clearly not just a creditor whose name happens to be
                # in caps).
                if lbl2 and lbl2_s.isupper() and len(lbl2_s) > 5 \
                        and "m/s" not in lbl2_s.lower() \
                        and "advance" not in lbl2_s.lower() \
                        and "msme" not in lbl2_s.lower() \
                        and "due to" not in lbl2_s.lower():
                    has_amount = (d2 is not None or e2 is not None)
                    if not has_amount:
                        # New top-level section — genuine end of creditor block
                        cred_end_row = r2
                        break
                    # ALL-CAPS data row (creditor name in caps) — include it
                    cred_end_row = r2 + 1
            break

        # Build name→row map from det_cache (no cell access needed)
        cred_row_map = {}
        for r in range(cred_start_row, cred_end_row):
            b = det_cache.get((r, 2))
            if b and str(b).strip():
                cred_row_map[str(b).strip()] = r

        written_rows = set()
        creditor_accounts = [a for a in individual_accounts
                             if a.get("bs_head") == "trade_payables"
                             and abs(a.get("net", 0)) > 0]
        # Debit-balance creditors are reclassified to stla (advance to supplier)
        # — they must NOT appear in creditor rows, handled below in stla section
        adv_to_supplier = [a for a in individual_accounts
                           if a.get("reclassified_from") == "trade_payables"]
        unmatched_creditors = []
        for acct in creditor_accounts:
            best_row = None
            for tmpl_name, row in cred_row_map.items():
                if row not in written_rows and _fuzzy_match_name(acct["name"], tmpl_name):
                    best_row = row; break
            if best_row:
                # FIX: use direct write (not _safe_write) for name-matched
                # creditor rows. The template may have old-year values at
                # these rows (contaminated template). _safe_write would block
                # because the cell is non-zero, leaving the wrong value.
                # Since we explicitly matched this account to this row by
                # name, we should always overwrite with the new TB value.
                cell = ws_det.cell(best_row, 4)
                from openpyxl.cell.cell import MergedCell
                if isinstance(cell, MergedCell):
                    skipped.append(f"Details!D{best_row} ({acct['name']}): cell is merged")
                else:
                    cell.value = abs(acct["net"])
                    # Ensure numeric format (prevent date-formatted cells)
                    if cell.number_format and 'y' in cell.number_format.lower():
                        cell.number_format = '#,##0.00'
                    injected.append(f"Details!D{best_row} ({acct['name']}) = {abs(acct['net']):,.2f}")
                    written_rows.add(best_row)
            else:
                unmatched_creditors.append(acct)

        # Pass 2a: fully blank rows (no name, no value) within section
        blank_rows = [r for r in range(cred_start_row, cred_end_row)
                      if r not in written_rows
                      and det_cache.get((r, 4)) is None
                      and (det_cache.get((r, 2)) is None
                           or str(det_cache.get((r, 2)) or "").strip() in ("", " ", "Nil", "-"))]

        # Pass 2b: template rows whose prior-year value (col E) is 0/None AND whose
        # name didn't match any TB account this year — these are former parties that
        # no longer owe us money, safe to replace with a new creditor name.
        zero_py_rows = [r for r in range(cred_start_row, cred_end_row)
                        if r not in written_rows and r not in blank_rows
                        and (det_cache.get((r, 5)) in (None, 0, "")
                             or det_cache.get((r, 4)) in (None, 0, ""))
                        and det_cache.get((r, 2)) is not None
                        and str(det_cache.get((r, 2)) or "").strip() not in ("", " ", "Nil", "-")]

        reuse_rows = blank_rows + zero_py_rows  # prefer fully blank first

        for i, acct in enumerate(unmatched_creditors):
            if i < len(reuse_rows):
                r = reuse_rows[i]
                ws_det.cell(r, 1).value = "M/s."
                ws_det.cell(r, 2).value = acct["name"]
                cell_d = ws_det.cell(r, 4)
                if hasattr(cell_d, 'number_format') and cell_d.number_format:
                    fmt = cell_d.number_format.lower()
                    if 'y' in fmt or 'd' in fmt or 'm' in fmt:
                        cell_d.number_format = '#,##0.00'
                cell_d.value = abs(acct["net"])
                written_rows.add(r)
                injected.append(f"Details!D{r} (creditor reuse row: {acct['name']}) = {abs(acct['net']):,.2f}")
            else:
                # FIX (2026-07-16): section has no spare blank rows left.
                # Insert a new row before the TOTAL/SUM row and update the
                # SUM formula so the new party is included in the total.
                # This is the same pattern used for the trade-receivables
                # overflow path and is safer than searching for a random
                # blank row elsewhere on the sheet (which puts the value
                # outside the formula range, causing the silent drop).
                _cred_insert_at = cred_total_row if cred_total_row else (cred_end_row + 1)
                try:
                    ws_det.insert_rows(_cred_insert_at)
                    ws_det.cell(_cred_insert_at, 1).value = "M/s."
                    ws_det.cell(_cred_insert_at, 2).value = acct["name"]
                    ws_det.cell(_cred_insert_at, 4).value = abs(acct["net"])
                    written_rows.add(_cred_insert_at)
                    # The insert shifts everything below by 1; update the SUM formula
                    _new_total_r = _cred_insert_at + 1  # TOTAL row moved down
                    _total_fv = ws_det.cell(_new_total_r, 4).value
                    if isinstance(_total_fv, str) and "SUM" in _total_fv.upper():
                        ws_det.cell(_new_total_r, 4).value = (
                            f"=SUM(D{cred_start_row}:D{_cred_insert_at})"
                        )
                        injected.append(
                            f"Details!D{_new_total_r} creditor TOTAL updated "
                            f"=SUM(D{cred_start_row}:D{_cred_insert_at})"
                        )
                    injected.append(
                        f"Details!D{_cred_insert_at} (creditor INSERT ROW: "
                        f"{acct['name']}) = {abs(acct['net']):,.2f}"
                    )
                    # Keep cred_total_row in sync for the next overflow party
                    cred_total_row = _new_total_r
                    cred_end_row   = _cred_insert_at  # new last data row
                    # Update cross-sheet formulas that reference Details rows
                    # at or below the insertion point
                    _cred_cross = re.compile(
                        r"'?Details'?!(\$?)([A-Z]+)(\$?)(\d+)"
                        r"(:(\$?)([A-Z]+)(\$?)(\d+))?"
                    )
                    def _cred_shift(mm, _ins=_cred_insert_at):
                        rn = int(mm.group(4))
                        nr = rn + 1 if rn >= _ins else rn
                        if mm.group(5):
                            rn2 = int(mm.group(9))
                            nr2 = rn2 + 1 if rn2 >= _ins else rn2
                            return (f"'Details'!{mm.group(1)}{mm.group(2)}"
                                    f"{mm.group(3)}{nr}:"
                                    f"{mm.group(6)}{mm.group(7)}{nr2}")
                        return f"'Details'!{mm.group(1)}{mm.group(2)}{mm.group(3)}{nr}"
                    for _osn in wb.sheetnames:
                        if _osn == "Details": continue
                        for _or in wb[_osn].iter_rows():
                            for _oc in _or:
                                _ov = _oc.value
                                if not (isinstance(_ov, str) and _ov.startswith("=")
                                        and "Details" in _ov):
                                    continue
                                _nv = _cred_cross.sub(_cred_shift, _ov)
                                if _nv != _ov:
                                    _oc.value = _nv
                                    log.append(
                                        f"✓ {_osn}!{_oc.coordinate} repaired: "
                                        f"{_ov} → {_nv} (creditor row insert at "
                                        f"Details!{_cred_insert_at})"
                                    )
                except Exception as _ce:
                    skipped.append(
                        f"Trade payable '{acct['name']}' ({abs(acct['net']):,.2f}): "
                        f"insert failed ({_ce}) — not placed"
                    )

        # Trade Receivables — auto-detect section bounds (not hardcoded rows)
        receivable_accounts = [a for a in individual_accounts
                               if a.get("bs_head") == "trade_rec"
                               and abs(a.get("net", 0)) > 0]
        recv_written = set()

        # FIX (Bug 2 — recv section detection):
        # The previous code used range(50, 200) for the scan, but in this template
        # the TRADE RECEIVABLE header sits at row 49 — ONE ROW BEFORE the scan starts.
        # So _recv_section_start never got updated from its initial fallback value of 74,
        # causing rows 51-73 to be skipped entirely (ABHINANDAN PRINTERS, ACE designers,
        # Bene Hygeine, AMP Pharmaceuticals, Bing Hospitality, etc. never matched).
        #
        # Fix: start scan from row 1 (or at least row 40) to guarantee we catch headers
        # at any position. Also separate the "find first <6months header" pass from the
        # "find total row" pass so the two-header template (row 49 = <6months,
        # row 141 = >6months) doesn't cause the start to jump to 143 before total is found.
        _recv_section_start = None   # will be set when first trade-receivable header found
        _recv_total_row = None
        _recv_gt6_start = None       # >6months section start (ignored for injection)
        for _sr in range(1, 250):
            _a = ws_det.cell(_sr, 1).value
            _b = ws_det.cell(_sr, 2).value
            _lbl = str(_a or _b or "").lower()
            _d_val = ws_det.cell(_sr, 4).value

            if "trade receivable" in _lbl or "trade rec" in _lbl:
                if _recv_section_start is None:
                    # First (< 6 months) header — this is where our debtors go
                    _recv_section_start = _sr + 2  # skip header + PARTICULARS row
                else:
                    # Second (> 6 months) header — record but don't use for injection
                    _recv_gt6_start = _sr + 2

            # TOTAL / SUM row ends the <6months debtor section
            if _recv_section_start and _recv_total_row is None:
                _is_total = ("total" in _lbl and _sr > _recv_section_start)
                _is_sum   = (isinstance(_d_val, str) and "sum" in _d_val.lower()
                             and _sr > _recv_section_start)
                if _is_total or _is_sum:
                    # Make sure we haven't accidentally crossed into the >6months section
                    if _recv_gt6_start is None or _sr < _recv_gt6_start:
                        _recv_total_row = _sr
                        break   # found both start and end — done

        if _recv_section_start is None:
            _recv_section_start = 74   # hard fallback for unexpected templates
        recv_end_row = (_recv_total_row - 1) if _recv_total_row else (_recv_section_start + 80)

        # Floor guard: debtor scan must stay BELOW the creditor/AFC section.
        # cred_end_row is the SUM row of the creditor block; the recv section
        # should start well after it. Using max() protects against stale
        # _recv_section_start values that predate AFC insert_rows operations.
        _recv_scan_start = max(_recv_section_start, cred_end_row + 5)

        for acct in receivable_accounts:
            placed = False
            # First try to match by name within the trade-receivable section.
            # FIX (2026-07-16): read LIVE from ws_det (not det_cache) because
            # det_cache was built before AFC insert_rows operations shifted all
            # debtor template rows. det_cache[(66,2)] still says "Bawa Garments"
            # but after AFC inserts "Bawa Garments" is now at sheet row 83.
            # Using det_cache causes debtor amounts to land at the (now AFC) row 66.
            for r in range(_recv_scan_start, recv_end_row + 1):
                b = ws_det.cell(r, 2).value
                if b and _fuzzy_match_name(acct["name"], str(b)) and r not in recv_written:
                    if _safe_write(ws_det, r, 4, abs(acct["net"])):
                        injected.append(f"Details!D{r} ({acct['name']}) = {abs(acct['net']):,.2f}")
                        recv_written.add(r); placed = True
                    else:
                        existing = ws_det.cell(r, 4).value
                        try:
                            _ex_f = float(existing) if existing not in (None, "") and not str(existing).startswith("=") else None
                        except (TypeError, ValueError):
                            _ex_f = None
                        if _ex_f is not None and abs(_ex_f - abs(acct["net"])) < 1:
                            recv_written.add(r); placed = True
                            injected.append(f"Details!D{r} ({acct['name']}) already = {existing} ✓")
                    break
            if not placed:
                # Pass 2a: fully blank row (no name, no D value)
                # Pass 2b: named row with zero/None PY value that no TB account matched —
                # prior-year debtor no longer owes us, safe to replace with new debtor
                _recv_blank   = []
                _recv_zero_py = []
                # FIX (2026-07-16 v2): for debtors, only use TRULY BLANK rows
                # (col B has no name). Reusing zero-PY named rows (old approach)
                # caused debtor writes to land in AFC rows that were inserted into
                # the same range by a prior AFC insert_rows pass, polluting the
                # creditor block with debtor amounts and producing wrong SUMs.
                # Use insert_rows for any debtor that has no blank slot.
                # Use live ws_det reads (not stale det_cache) for blank-row scan.
                _recv_floor = max(_recv_section_start, cred_end_row + 5)
                for r in range(_recv_floor, recv_end_row + 1):
                    if r in recv_written: continue
                    b_r = ws_det.cell(r, 2).value   # live read
                    d_r = ws_det.cell(r, 4).value   # live read
                    b_empty = b_r is None or str(b_r).strip() in ("", " ", "Nil", "-")
                    if b_empty and (d_r is None or d_r == 0):
                        _recv_blank.append(r)
                _recv_reuse = _recv_blank  # no zero-PY reuse for debtors
                if _recv_reuse:
                    r = _recv_reuse[0]
                    ws_det.cell(r, 1).value = "M/s."
                    ws_det.cell(r, 2).value = acct["name"]
                    cell_rd = ws_det.cell(r, 4)
                    if hasattr(cell_rd, 'number_format') and cell_rd.number_format:
                        if any(c in cell_rd.number_format.lower() for c in ('y', 'mmm')):
                            cell_rd.number_format = '#,##0.00'
                    cell_rd.value = abs(acct["net"])
                    recv_written.add(r); placed = True
                    injected.append(f"Details!D{r} (recv reuse: {acct['name']}) = {abs(acct['net']):,.2f}")
            if not placed:
                # Section full — insert new row before total and update formula
                try:
                    insert_at = recv_end_row + 1  # insert at position of total row
                    ws_det.insert_rows(insert_at)
                    ws_det.cell(insert_at, 1).value = "M/s."   # FIX: add M/s. prefix
                    ws_det.cell(insert_at, 2).value = acct["name"]
                    ws_det.cell(insert_at, 4).value = abs(acct["net"])
                    recv_written.add(insert_at)
                    recv_end_row = insert_at  # expand section end
                    injected.append(f"Details!D{insert_at} (recv NEW ROW: {acct['name']}) = {abs(acct['net']):,.2f}")
                    det_cache[(insert_at, 2)] = acct["name"]
                    det_cache[(insert_at, 4)] = abs(acct["net"])

                    # Update total formula to cover full section including new row
                    _pending_total_row   = recv_end_row + 1
                    _pending_total_start = _recv_section_start
                    # Re-write the total formula to cover start→new end
                    _total_r = (_recv_total_row + 1) if _recv_total_row else (recv_end_row + 1)
                    _total_f = ws_det.cell(_total_r, 4).value
                    if isinstance(_total_f, str) and "SUM" in _total_f.upper():
                        ws_det.cell(_total_r, 4).value = (
                            f"=SUM(D{_pending_total_start}:D{recv_end_row})"
                        )
                        injected.append(f"Details!D{_total_r} TOTAL formula updated to cover D{_pending_total_start}:D{recv_end_row}")
                    elif _total_f is None or _total_f == 0:
                        ws_det.cell(_total_r, 4).value = (
                            f"=SUM(D{_pending_total_start}:D{recv_end_row})"
                        )

                    _recv_cross_pattern = re.compile(
                        r"'?Details'?!(\$?)([A-Z]+)(\$?)(\d+)"
                        r"(:(\$?)([A-Z]+)(\$?)(\d+))?"
                    )

                    def _recv_shift_match(_mm, _ins_row=recv_end_row):
                        _row_num = int(_mm.group(4))
                        _new_row = _row_num + 1 if _row_num >= _ins_row else _row_num
                        if _mm.group(5):
                            _row_num2 = int(_mm.group(9))
                            _new_row2 = _row_num2 + 1 if _row_num2 >= _ins_row else _row_num2
                            return (
                                f"'Details'!{_mm.group(1)}{_mm.group(2)}{_mm.group(3)}{_new_row}:"
                                f"{_mm.group(6)}{_mm.group(7)}{_new_row2}"
                            )
                        return f"'Details'!{_mm.group(1)}{_mm.group(2)}{_mm.group(3)}{_new_row}"

                    for _other_sn in wb.sheetnames:
                        if _other_sn == "Details":
                            continue
                        _ws_other = wb[_other_sn]
                        for _orow in _ws_other.iter_rows():
                            for _ocell in _orow:
                                _ov = _ocell.value
                                if not (isinstance(_ov, str) and _ov.startswith("=") and "Details" in _ov):
                                    continue
                                _new_ov = _recv_cross_pattern.sub(_recv_shift_match, _ov)
                                if _new_ov != _ov:
                                    _ocell.value = _new_ov
                                    log.append(
                                        f"✓ {_other_sn}!{_ocell.coordinate} formula repaired: "
                                        f"{_ov} → {_new_ov} (reference into Details shifted after "
                                        f"row inserted at Details!{recv_end_row} for overflow debtor "
                                        f"'{acct['name']}')"
                                    )

                    # FIX: the Total/sub-total formulas LIVING ON the Details
                    # sheet itself (e.g. "=D95+SUM(D97:D100)") have the exact
                    # same staleness problem — insert_rows() shifted their
                    # CELL POSITION down by one row along with everything
                    # else below the insertion point, but never touched the
                    # formula TEXT inside them. Any such formula cell that
                    # itself sits at or below the insertion row almost
                    # certainly moved together with the data it's summing,
                    # so its bare (no sheet-prefix) row references need the
                    # same +1 shift applied for every row number >= the
                    # insertion point.
                    _recv_bare_pattern = re.compile(r"(?<!['!])(\$?)([A-Z]{1,2})(\$?)(\d+)")
                    for _srow in ws_det.iter_rows(min_row=recv_end_row):
                        for _scell in _srow:
                            if _scell.row == recv_end_row:
                                continue  # the row we just wrote into — not a formula
                            _sv = _scell.value
                            if not (isinstance(_sv, str) and _sv.startswith("=")):
                                continue
                            if "!" in _sv:
                                continue  # cross-sheet ref, not handled here

                            def _recv_shift_bare(_mm, _ins_row=recv_end_row):
                                _row_num = int(_mm.group(4))
                                if _row_num >= _ins_row:
                                    return f"{_mm.group(1)}{_mm.group(2)}{_mm.group(3)}{_row_num + 1}"
                                return _mm.group(0)

                            _new_sv = _recv_bare_pattern.sub(_recv_shift_bare, _sv)
                            if _new_sv != _sv:
                                _scell.value = _new_sv
                                log.append(
                                    f"✓ Details!{_scell.coordinate} formula repaired: "
                                    f"{_sv} → {_new_sv} (same-sheet reference shifted after "
                                    f"row inserted at Details!{recv_end_row} for overflow debtor "
                                    f"'{acct['name']}')"
                                )
                except Exception as e:
                    skipped.append(f"recv '{acct['name']}': insert failed: {e}")

        # NOTE: LT borrowings are written to notes to bs!D8 directly (Section 2 above).
        # Do NOT also write to Details D7-D12 — that would double-count because
        # notes to bs!D16 = =Details!D9 (formula), creating a second entry.

        # ── Advance to Suppliers (debit-balance creditors) ───────────────────
        # FIX (2026-07-16) — full rewrite of this section:
        #
        # Bug 1 (scan start too late): scan started from _recv_total_row,
        #   missing headers that sit BEFORE the recv total (Fashion Adda row 77
        #   vs recv total row 80).  Now scans the entire sheet from row 1.
        #
        # Bug 2 (broken regex-as-string): "adv.*supp" tested with `in` can
        #   never match — it's a literal 8-char string.  Replaced with correct
        #   plain substring checks.
        #
        # Bug 3 (no insert fallback): when section had no blank rows the amount
        #   was silently dropped.  Now inserts a row before TOTAL + updates SUM.
        #
        # Bug 4 (JEANS WORLD / SHREE CRAFT layout): advance-to-supplier can be
        #   a sub-section INSIDE the trade-receivables SUM range (JEANS WORLD
        #   row 105 "ADVANCE TO SUPPLIERS" header sits inside =SUM(D66:D106)).
        #   The old code only searched BELOW recv_total_row so it found nothing
        #   and fell back to placing accounts AFTER the SUM row (outside formula).
        #   New: search the full Details sheet; if the header falls inside the
        #   recv section, treat those rows as part of recv (already inside SUM).
        _adv_supp_start  = None
        _adv_supp_end    = None
        _adv_supp_total  = None   # row number of the TOTAL / SUM row
        _adv_inside_recv = False  # True when the section is inside the recv SUM

        for _sr2 in range(1, 300):   # full sheet scan
            _a2   = ws_det.cell(_sr2, 1).value
            _lbl2 = str(_a2 or "").strip().lower()
            if ("advance to supplier" in _lbl2
                    or ("advance" in _lbl2 and "supplier" in _lbl2)
                    or ("advance" in _lbl2 and "supp" in _lbl2
                        and "from" not in _lbl2)):
                _adv_supp_start = _sr2 + 1  # first data row = row after header
                # Check whether this header lives inside the recv SUM range
                if (_recv_section_start is not None
                        and _recv_section_start < _sr2 <= recv_end_row):
                    _adv_inside_recv = True
            # Once we have a start, find the TOTAL / SUM boundary
            if _adv_supp_start and _sr2 > _adv_supp_start:
                _d2 = ws_det.cell(_sr2, 4).value
                _is_total2 = "total" in _lbl2
                _is_sum2   = isinstance(_d2, str) and "sum" in _d2.lower()
                if _is_total2 or _is_sum2:
                    _adv_supp_total = _sr2
                    _adv_supp_end   = _sr2 - 1
                    break
                # Stop at a new top-level section header (ALL-CAPS, no amount in D)
                if (_a2 and str(_a2).strip().isupper() and len(str(_a2).strip()) > 3
                        and "m/s" not in str(_a2).strip().lower()
                        and _d2 is None):
                    _adv_supp_end = _sr2 - 1
                    break

        if _adv_supp_start is None:
            _adv_supp_start = recv_end_row + 3
        if _adv_supp_end is None:
            _adv_supp_end = _adv_supp_start + 10

        for acct in adv_to_supplier:
            placed = False
            # Pass 1: name-match within the advance-to-suppliers section
            for r in range(_adv_supp_start, _adv_supp_end + 1):
                b = ws_det.cell(r, 2).value
                if b and _fuzzy_match_name(acct["name"], str(b)) and r not in recv_written:
                    cell = ws_det.cell(r, 4)
                    from openpyxl.cell.cell import MergedCell
                    if not isinstance(cell, MergedCell):
                        cell.value = abs(acct["net"])
                        if cell.number_format and 'y' in cell.number_format.lower():
                            cell.number_format = '#,##0.00'
                        injected.append(
                            f"Details!D{r} (adv-to-supplier: {acct['name']}) "
                            f"= {abs(acct['net']):,.2f}"
                        )
                        recv_written.add(r); placed = True
                    break
            if not placed:
                # Pass 2: first genuinely blank row in the section
                for r in range(_adv_supp_start, _adv_supp_end + 1):
                    _b2 = ws_det.cell(r, 2).value
                    _d2 = ws_det.cell(r, 4).value
                    if (_d2 is None and r not in recv_written
                            and (_b2 is None or str(_b2).strip() in ("", "Nil", "-"))):
                        ws_det.cell(r, 1).value = "M/s."
                        ws_det.cell(r, 2).value = acct["name"]
                        ws_det.cell(r, 4).value = abs(acct["net"])
                        injected.append(
                            f"Details!D{r} (adv-to-supplier new slot: {acct['name']}) "
                            f"= {abs(acct['net']):,.2f}"
                        )
                        recv_written.add(r); placed = True; break
            if not placed:
                # Pass 3: insert row before TOTAL and update SUM
                _insert_at = _adv_supp_total if _adv_supp_total else (_adv_supp_end + 1)
                try:
                    ws_det.insert_rows(_insert_at)
                    ws_det.cell(_insert_at, 1).value = "M/s."
                    ws_det.cell(_insert_at, 2).value = acct["name"]
                    ws_det.cell(_insert_at, 4).value = abs(acct["net"])
                    recv_written.add(_insert_at)
                    _new_total_r = _insert_at + 1
                    _new_end     = _insert_at
                    _total_f = ws_det.cell(_new_total_r, 4).value
                    if isinstance(_total_f, str) and "SUM" in _total_f.upper():
                        ws_det.cell(_new_total_r, 4).value = (
                            f"=SUM(D{_adv_supp_start}:D{_new_end})"
                        )
                        injected.append(
                            f"Details!D{_new_total_r} adv-to-supp TOTAL updated "
                            f"=SUM(D{_adv_supp_start}:D{_new_end})"
                        )
                    injected.append(
                        f"Details!D{_insert_at} (adv-to-supplier INSERT ROW: "
                        f"{acct['name']}) = {abs(acct['net']):,.2f}"
                    )
                    _adv_supp_total = _new_total_r
                    _adv_supp_end   = _new_end
                    placed = True
                except Exception as _adv_ins_exc:
                    skipped.append(
                        f"Advance-to-supplier '{acct['name']}' "
                        f"({abs(acct['net']):,.2f}): insert failed "
                        f"({_adv_ins_exc}) — not placed"
                    )

        # ── Advance from Customers (credit-balance debtors / creditors) ─────
        # FIX (2026-07-16):
        #
        # Layout A (Fashion Adda): separate section below debtors, with its
        #   own rows outside the creditor SUM.
        #
        # Layout B (JEANS WORLD / SHREE CRAFT): "Advance from Customers"
        #   sub-header sits INSIDE the Sundry Creditors SUM range.
        #   e.g. JEANS WORLD Details rows 28-55 are the adv-from-cust sub-rows,
        #   but they're summed together with the creditors (=SUM(D21:D55)).
        #   Detection: scan for the header; if it falls inside [cred_start_row,
        #   cred_end_row], the section IS the creditor rows from header+1
        #   to cred_end_row. We still write to col D and name-match correctly;
        #   the SUM already covers these rows so no formula update needed.
        #
        # Add insert_rows fallback so new adv-from-cust accounts that don't
        # have a pre-existing row in the template get inserted before the
        # creditor TOTAL and counted in the formula.
        adv_from_customer = [a for a in individual_accounts
                             if a.get("reclassified_from") == "trade_rec"
                             or a.get("bs_head") == "advance_from_customer"]
        if adv_from_customer and "Details" in wb.sheetnames:
            ws_det_afc = wb["Details"]

            # Detect "Advance from Customers" sub-section (full sheet scan)
            _afc_start, _afc_end = None, None
            _afc_inside_creditor  = False
            _afc_total_row        = None
            for _sr3 in range(1, 300):
                _lbl3a = str(ws_det_afc.cell(_sr3, 1).value or "").strip().lower()
                _lbl3b = str(ws_det_afc.cell(_sr3, 2).value or "").strip().lower()
                _lbl3  = _lbl3a or _lbl3b
                if "advance from customer" in _lbl3 or "advance from buyer" in _lbl3:
                    _afc_start = _sr3 + 1
                    # Is this inside the creditor SUM block?
                    if cred_start_row < _sr3 < cred_end_row:
                        _afc_inside_creditor = True
                        _afc_end = cred_end_row - 1   # last data row before cred SUM
                        _afc_total_row = cred_total_row
                if _afc_start and _sr3 > _afc_start and not _afc_inside_creditor:
                    _lbl3x = str(ws_det_afc.cell(_sr3, 1).value or "").strip()
                    _d3    = ws_det_afc.cell(_sr3, 4).value
                    _is_t  = _lbl3x.lower() in ("total", "totals")
                    _is_s  = isinstance(_d3, str) and "sum" in _d3.lower()
                    if _is_t or _is_s:
                        _afc_total_row = _sr3
                        _afc_end       = _sr3 - 1
                        break
                    if (_lbl3x and _lbl3x.isupper() and len(_lbl3x) > 3
                            and "m/s" not in _lbl3x.lower()
                            and "advance" not in _lbl3x.lower()
                            and _d3 is None):
                        _afc_end = _sr3 - 1
                        break
            if _afc_start is None:
                _afc_start = cred_end_row + 3   # safe fallback below creditors
            if _afc_end is None:
                _afc_end = _afc_start + 20

            for acct in adv_from_customer:
                placed = False
                # Pass 1: name-match in the advance-from-customers section
                for r in range(_afc_start, _afc_end + 1):
                    b = ws_det_afc.cell(r, 2).value
                    if b and _fuzzy_match_name(acct["name"], str(b)) and r not in written_rows:
                        cell_d = ws_det_afc.cell(r, 4)
                        from openpyxl.cell.cell import MergedCell
                        if not isinstance(cell_d, MergedCell):
                            cell_d.value = abs(acct["net"])
                            if cell_d.number_format and 'y' in cell_d.number_format.lower():
                                cell_d.number_format = '#,##0.00'
                            injected.append(
                                f"Details!D{r} (adv-from-customer: {acct['name']}) "
                                f"= {abs(acct['net']):,.2f}"
                            )
                            written_rows.add(r); placed = True
                        break
                if not placed:
                    # Pass 2a: fully blank row (no name)
                    # Pass 2b: named row with zero/None PY value that didn't match any TB account
                    # (prior-year party no longer present → safe to replace with new account)
                    _afc_blank = []
                    _afc_zero_py = []
                    for r in range(_afc_start, _afc_end + 1):
                        if r in written_rows: continue
                        b2 = ws_det_afc.cell(r, 2).value
                        d2 = ws_det_afc.cell(r, 4).value
                        e2 = ws_det_afc.cell(r, 5).value
                        b2_empty = b2 is None or not str(b2).strip() or str(b2).strip() in (" ", "Nil", "-")
                        if b2_empty and (d2 is None or d2 == 0):
                            _afc_blank.append(r)
                        elif not b2_empty and (e2 in (None, 0) and (d2 is None or d2 == 0)):
                            _afc_zero_py.append(r)
                    _afc_reuse = _afc_blank + _afc_zero_py
                    if _afc_reuse:
                        r = _afc_reuse[0]
                        ws_det_afc.cell(r, 1).value = "M/s."
                        ws_det_afc.cell(r, 2).value = acct["name"]
                        ws_det_afc.cell(r, 4).value = abs(acct["net"])
                        injected.append(
                            f"Details!D{r} (adv-from-cust reuse: {acct['name']}) "
                            f"= {abs(acct['net']):,.2f}"
                        )
                        written_rows.add(r); placed = True
                if not placed:
                    # Pass 3: insert before TOTAL row and update SUM
                    _afc_insert = (_afc_total_row if _afc_total_row
                                   else (_afc_end + 1))
                    try:
                        ws_det_afc.insert_rows(_afc_insert)
                        ws_det_afc.cell(_afc_insert, 1).value = "M/s."
                        ws_det_afc.cell(_afc_insert, 2).value = acct["name"]
                        ws_det_afc.cell(_afc_insert, 4).value = abs(acct["net"])
                        written_rows.add(_afc_insert)
                        _afc_new_total = _afc_insert + 1
                        _afc_total_f   = ws_det_afc.cell(_afc_new_total, 4).value
                        # When AFC is inside the creditor SUM block, the formula
                        # must start from cred_start_row (covers all creditors too)
                        _sum_start = cred_start_row if _afc_inside_creditor else _afc_start
                        if isinstance(_afc_total_f, str) and "SUM" in _afc_total_f.upper():
                            ws_det_afc.cell(_afc_new_total, 4).value = (
                                f"=SUM(D{_sum_start}:D{_afc_insert})"
                            )
                            injected.append(
                                f"Details!D{_afc_new_total} afc TOTAL updated "
                                f"=SUM(D{_sum_start}:D{_afc_insert})"
                            )
                        injected.append(
                            f"Details!D{_afc_insert} (adv-from-cust INSERT ROW: "
                            f"{acct['name']}) = {abs(acct['net']):,.2f}"
                        )
                        _afc_total_row = _afc_new_total
                        _afc_end       = _afc_insert
                        placed = True
                    except Exception as _afc_ins_exc:
                        skipped.append(
                            f"Adv-from-customer '{acct['name']}': "
                            f"insert failed ({_afc_ins_exc}) — not placed"
                        )

        # ── Re-detect trade-receivable section after AFC inserts ──────────
        # AFC insert_rows operations shift the TRADE RECEIVABLE header and all
        # debtor rows downward.  _recv_section_start was detected from the
        # template BEFORE those inserts, so it now points at the wrong rows.
        # Re-scan ws_det (current state) to find the updated section boundaries.
        if adv_from_customer and "Details" in wb.sheetnames:
            _new_recv_start = None
            _new_recv_total = None
            for _rs in range(1, 400):
                _rv = ws_det.cell(_rs, 1).value
                _rl = str(_rv or "").strip().lower()
                if "trade receivable" in _rl and "less than" in _rl or \
                   ("trade receivable" in _rl and "<6" in _rl.replace(" ", "")):
                    _new_recv_start = _rs + 2   # skip header + PARTICULARS
                if _new_recv_start and _rs > _new_recv_start:
                    _rd = ws_det.cell(_rs, 4).value
                    _is_t = _rl in ("total", "totals")
                    _is_s = isinstance(_rd, str) and "sum" in _rd.lower()
                    if _is_t or _is_s:
                        _new_recv_total = _rs
                        break
            if _new_recv_start and _new_recv_start != _recv_section_start:
                log.append(
                    f"✓ Trade-receivable section re-detected after AFC inserts: "
                    f"start {_recv_section_start} → {_new_recv_start}, "
                    f"total row {_recv_total_row} → {_new_recv_total}"
                )
                _recv_section_start = _new_recv_start
                _recv_total_row     = _new_recv_total
                recv_end_row = (_new_recv_total - 1) if _new_recv_total else (_new_recv_start + 80)


        # ── DEFERRED: TOTAL / NBS / POST-PASS repaired AFTER AFC inserts ──────
        # (moved here so the scan sees the FULL post-AFC debtor range)
        # NOTE: The deferred TOTAL formula write (_pending_total_row) is intentionally
        # removed. _pending_total_row was set to the pre-AFC row number of the TOTAL row,
        # but after AFC insert_rows that row becomes a debtor data row. Writing a SUM
        # formula there corrupts the data. The final repair scan below correctly
        # detects and sets the TOTAL formula on the real TOTAL row (post-AFC).

        # ── FINAL REPAIR: re-detect actual debtor range from live sheet ─────
        # recv_written holds PRE-AFC-insert row numbers. After AFC inserts (+N rows),
        # those rows shifted. The notes-to-bs SUM and Details TOTAL must use
        # POST-INSERT positions. Re-scan the live Details sheet to find the actual
        # TOTAL row and the true min/max of written debtor data rows.
        _actual_min = min(recv_written) if recv_written else None
        _actual_max = max(recv_written) if recv_written else None
        if recv_written and "Details" in wb.sheetnames:
            _ws_det_live = wb["Details"]
            _lr_header = None
            for _lr in range(1, 500):
                _la = _ws_det_live.cell(_lr, 1).value
                _lb = _ws_det_live.cell(_lr, 2).value
                _ll = str(_la or _lb or "").strip().lower()
                if ("trade receivable" in _ll
                        and ("less" in _ll or "<6" in _ll.replace(" ", ""))):
                    _lr_header = _lr
                    break
            _lr_total = None
            if _lr_header:
                for _lr in range(_lr_header + 1, _lr_header + 300):
                    _la2 = _ws_det_live.cell(_lr, 1).value
                    _lb2 = _ws_det_live.cell(_lr, 2).value
                    _ld2 = _ws_det_live.cell(_lr, 4).value
                    _is_t2 = str(_la2 or "").strip().upper() in ("TOTAL", "TOTALS")
                    _b2_empty = _lb2 is None or str(_lb2).strip() == ""
                    _has_sum_d2 = (isinstance(_ld2, str) and "SUM" in _ld2.upper()
                                   and _lr > _lr_header + 5)
                    # If a DATA row (col B has a name) has a SUM formula in D,
                    # it's a stale formula from the insert_rows overflow path.
                    # Clear it so the row just holds None (correct for a debtor
                    # with no current-year balance).
                    if _has_sum_d2 and not _b2_empty:
                        _ws_det_live.cell(_lr, 4).value = None
                        log.append(
                            f"✓ Details!D{_lr}: cleared stale SUM from debtor row "
                            f"(B={repr(str(_lb2)[:30])})"
                        )
                        continue  # don't treat this as the TOTAL row
                    # A real TOTAL row: A='TOTAL' or D has SUM AND B is empty
                    _is_s2 = _has_sum_d2 and _b2_empty
                    if _is_t2 or _is_s2:
                        _lr_total = _lr
                        break
            # Collect all rows with data in the debtor area (post-inserts)
            _actual_recv_rows = []
            if _lr_header:
                _scan_end2 = (_lr_total - 1) if _lr_total else (_lr_header + 200)
                for _lr in range(_lr_header + 2, _scan_end2 + 1):
                    _ld3 = _ws_det_live.cell(_lr, 4).value
                    if isinstance(_ld3, (int, float)) and _ld3 != 0:
                        _actual_recv_rows.append(_lr)
            if _actual_recv_rows:
                _actual_min = min(_actual_recv_rows)
                _actual_max = max(_actual_recv_rows)
                if _lr_total:
                    _correct_sum = f"=SUM(D{_actual_min}:D{_actual_max})"
                    _old_total_f = _ws_det_live.cell(_lr_total, 4).value
                    if _old_total_f != _correct_sum:
                        _ws_det_live.cell(_lr_total, 4).value = _correct_sum
                        log.append(
                            f"✓ Details!D{_lr_total} recv TOTAL repaired: "
                            f"{_old_total_f} → {_correct_sum}"
                        )

        # ── FIX: repair notes to bs Trade Receivables formula ───────────
        # Dynamically locate the "Unsecured Considered good" row for the
        # <6months debtor section in notes to bs and rewrite it to SUM the
        # full actual range written into Details.
        #
        # FIX (2026-07-16): the old condition checked whether the existing
        # formula's row reference was WITHIN recv_written range (min_r:max_r).
        # If it was (e.g. =Details!D108 and 108 is between 82 and 123), it
        # skipped the repair even though D108 is a SINGLE CELL pointing at one
        # debtor, not a SUM of all debtors.  Now always replace any single-cell
        # Details ref with the full SUM formula.
        if recv_written and "notes to bs" in wb.sheetnames:
            ws_nbs_fix = wb["notes to bs"]
            # Use live-detected range (post-inserts) if available; fallback to recv_written
            min_r = _actual_min if _actual_min else min(recv_written)
            max_r = _actual_max if _actual_max else max(recv_written)

            _lt6_row, _gt6_row = None, None
            _found_lt6_header = False
            for _nr in range(1, 300):
                _nb   = ws_nbs_fix.cell(_nr, 2).value   # col B = always the text label
                _nlbl = str(_nb or "").strip().lower()
                if "trade receivable" in _nlbl:
                    _found_lt6_header = True
                if _found_lt6_header and "(b)" in _nlbl and "unsecured" in _nlbl:
                    if _lt6_row is None:
                        _lt6_row = _nr
                    else:
                        _gt6_row = _nr
                        break

            if _lt6_row is None: _lt6_row = 80
            if _gt6_row is None: _gt6_row = 86

            _lt6_val = ws_nbs_fix.cell(_lt6_row, 4).value
            # Always replace if it's a single-cell ref OR a stale SUM
            _is_single_ref = (isinstance(_lt6_val, str) and
                              "Details" in _lt6_val and
                              "SUM" not in _lt6_val.upper())
            _is_stale_sum  = (isinstance(_lt6_val, str) and
                              "SUM" in _lt6_val.upper() and
                              f"D{min_r}" not in _lt6_val)
            if _is_single_ref or _is_stale_sum or _lt6_val is None:
                _nbs_min = _actual_min if _actual_min else min_r
                _nbs_max = _actual_max if _actual_max else max_r
                ws_nbs_fix.cell(_lt6_row, 4).value = (
                    f"=SUM(Details!D{_nbs_min}:D{_nbs_max})"
                )
                ws_nbs_fix.cell(_gt6_row, 4).value = 0
                log.append(
                    f"✓ Repaired 'notes to bs'!D{_lt6_row} → "
                    f"SUM(Details!D{_nbs_min}:D{_nbs_max}) "
                    f"(Trade Receivables; was: {repr(_lt6_val)})"
                )

        # ── POST-PASS: fix stale SUM formulas in Details sheet itself ────
        # After insert_rows operations (AFC and debtor overflow), the Details
        # sheet may have SUM formulas whose range no longer covers the actual
        # data.  e.g. =SUM(D66:D106) at row 124 when debtors are in D82:D123.
        # Scan all SUM formulas in Details whose range falls entirely OUTSIDE
        # the current recv_written range and rewrite them.
        if recv_written and "Details" in wb.sheetnames:
            import re as _dre
            _sum_re = _dre.compile(r'=SUM\(D(\d+):D(\d+)\)', _dre.IGNORECASE)
            # FIX: use _actual_min/_actual_max from the final repair scan (post-AFC
            # live-sheet detection), not pre-AFC recv_written which has shifted row numbers.
            _min_rw = _actual_min if "_actual_min" in dir() and _actual_min else min(recv_written)
            _max_rw = _actual_max if "_actual_max" in dir() and _actual_max else max(recv_written)
            _ws_det_pp = wb["Details"]
            for _pp_r in range(1, 300):
                _pp_v = _ws_det_pp.cell(_pp_r, 4).value
                if not (isinstance(_pp_v, str) and "SUM" in _pp_v.upper()):
                    continue
                _mm = _sum_re.match(_pp_v)
                if not _mm:
                    continue
                _sr, _er = int(_mm.group(1)), int(_mm.group(2))
                # Only fix SUM formulas that belong to the trade-receivable
                # section (row >= _recv_section_start). The same condition
                # must NOT fire on unsecured-loan or creditor sub-totals that
                # happen to use the D column.
                _in_recv_area = (_pp_r >= _recv_section_start) if _recv_section_start else (_pp_r > 60)
                # Stale: row is in the recv area AND range doesn't fully cover data
                if _in_recv_area and not (_sr <= _min_rw and _er >= _max_rw):
                    _new_f = f"=SUM(D{_min_rw}:D{_max_rw})"
                    _ws_det_pp.cell(_pp_r, 4).value = _new_f
                    log.append(
                        f"✓ Details!D{_pp_r}: repaired stale SUM "
                        f"{_pp_v} → {_new_f}"
                    )

    # ────────────────────────────────────────────────────────────────
    # 4. CLOSING STOCK / INVENTORIES
    # ────────────────────────────────────────────────────────────────
    # GROSS PROFIT!E17 has formula =B24-SUM(E11:E14) — closing stock
    # is the BALANCING FIGURE on the sales side.
    # DO NOT write to E17 — the formula calculates it automatically
    # once sales (E11:E14) and total (B24/E24) are filled.
    #
    # notes to p&l!D24 = ='GROSS PROFIT'!E17 (formula) — also auto.
    #
    # Opening stock (D17) IS written (it's a plain constant from TB).
    # Closing stock flows automatically — no injection needed here.
    inventory_amt = aggregated_values.get("inventories", 0)
    if inventory_amt:
        log.append(f"· Closing stock {inventory_amt:,.2f} — NOT written to E17 (auto-calculated by formula =B24-SUM(E11:E14))")

    # ────────────────────────────────────────────────────────────────
    # 5. FIXED ASSETS — DO NOT inject from TB.
    #    TB only has WDV closing. Additions/sales come from ledger.
    #    User's BS template FA chart is preserved as-is.
    # ────────────────────────────────────────────────────────────────
    fixed_assets_amt = aggregated_values.get("fixed_assets", 0)
    if fixed_assets_amt:
        log.append(f"· Fixed Assets from TB: {fixed_assets_amt:,.2f} — skipped (fill from ledger)")

    # ────────────────────────────────────────────────────────────────
    # 6. SHORT TERM PROVISIONS
    # ────────────────────────────────────────────────────────────────
    stp_amt = aggregated_values.get("st_provisions", 0)
    if stp_amt and individual_accounts:
        prov_accounts = [a for a in individual_accounts
                         if a.get("bs_head") == "st_provisions"
                         and abs(a.get("net", 0)) > 0]
        if "notes to bs" in wb.sheetnames:
            ws_n = wb["notes to bs"]
            # TCS GST A/C → row 73 area  
            for acct in prov_accounts:
                name_l = acct["name"].lower()
                if "tcs" in name_l and _safe_set(ws_n, 73, 4, abs(acct["net"])):
                    injected.append(f"notes to bs!D73 (TCS provision) = {abs(acct['net']):,.2f}")

    # ────────────────────────────────────────────────────────────────
    # 7. P&L NOTES INJECTION (notes to p&l + GROSS PROFIT)
    # ────────────────────────────────────────────────────────────────
    # P&L sheet cells are ALL formulas pulling from notes to p&l and
    # GROSS PROFIT. We must write to the source sheets, never p&l directly.
    #
    # Chain:
    #  p&l!E7  (Revenue)      ← notes to p&l!D7  ← notes to p&l!D6
    #                                                ← SUM(GROSS PROFIT!E11:E14)
    #  p&l!E12 (Cost Mat.)    ← notes to p&l!D26 ← D17+D20+D22-D24
    #                              D17=E24 (formula←prev yr closing)
    #                              D20=SUM(GROSS PROFIT!B14:B18)
    #                              D24='GROSS PROFIT'!E17 (formula=auto)
    #  p&l!E13 (Employee)     ← notes to p&l!D34 ← SUM(D31:D33)  → write D31
    #  p&l!E14 (Finance cost) ← notes to p&l!D40 ← SUM(D38:D39)  → write D38
    #  p&l!E15 (Depreciation) ← notes to p&l!D53 ← D52
    #                              D52='Fixed Assets C. Yr.'!H31 (formula=auto)
    #  p&l!E16 (Other exp)    ← notes to p&l!D79 ← SUM(D57:D78)  → write D57:D78

    if individual_accounts and "notes to p&l" in wb.sheetnames:
        ws_npl = wb["notes to p&l"]

        # ── A. Sales → GROSS PROFIT!E11:E14 ─────────────────────────
        # GROSS PROFIT!E11 = Sale GST 12% Interstate
        # GROSS PROFIT!E12 = Sale GST 12% Within State
        # GROSS PROFIT!E13 = Sale GST 5% Interstate
        # GROSS PROFIT!E14 = Sale GST 5% Within State
        # Shared normaliser used by BOTH the sales-side and purchases-side
        # GROSS PROFIT label matching below. Previously this was defined
        # only inside the sales block (gated by `revenue_prefilled`), so
        # when that block was skipped, the purchases block below (which is
        # NOT gated by `revenue_prefilled`) would crash with
        # UnboundLocalError if it ever reached a non-empty purchase_accounts
        # list.
        def _norm_gst(s):
            """Normalise sale/purchase account names for row matching.
            Maps all synonym pairs to canonical forms so that e.g.
            'SALE GST 12% INTERSTATE' and 'Sales 12% Intrastate' match
            the same template row regardless of exact wording.
            """
            s = str(s).lower()
            s = re.sub(r"\b(sale|sales|purchase|purchases|gst|a/c)\b", " ", s)
            # FIX: "withinstate" (no space, no word boundary) was never
            # being normalised to match "within state" (with space) —
            # both forms appear as separate template row labels in some
            # workbooks (e.g. Fashion Adda's GROSS PROFIT sheet has both
            # "Purchase GST 5% withinstate" at one row AND a second,
            # differently-spelled "PURCHASE GST 5% WITHIN STATE" row),
            # so a single TB account matching the "within state" spelling
            # only matched ONE of the two rows, while the OTHER row's
            # pre-existing value was treated as if blank and the same
            # amount got duplicated into it via the unmatched-account
            # fallback — silently double-counting that purchase category.
            # Collapsing both spellings to one canonical form before the
            # rest of the synonym normalisation ensures they always
            # resolve to the same row.
            s = s.replace("withinstate", "within state")
            s = s.replace("interstate", "inter state")
            # Synonym normalisation
            s = s.replace("intrastate", "within state")
            s = s.replace("within state", "intrastate")   # canonical = intrastate
            s = s.replace("within state", "intrastate")
            s = s.replace("intra state", "intrastate")
            s = s.replace("inter state", "interstate")
            s = re.sub(r"\s+", " ", s).strip()
            return s

        sale_accounts = [a for a in individual_accounts
                         if a.get("bs_head") == "revenue"
                         and abs(a.get("net", 0)) > 0]

        # FIX (Issue 1 — revised): the previous "pre-filled" detector checked
        # whether TB revenue amounts matched ANY numeric literal anywhere in
        # notes to p&l!D5:D7 (including individual addends of a formula like
        # "=10898146+793859"). That produced FALSE POSITIVES: if even one
        # addend happened to equal a TB group total, the WHOLE row's
        # resolved value (which may be wrong/stale, e.g. 11,692,005 instead
        # of the correct 10,898,146) was treated as "already correct" and
        # never fixed — leaving D5 blank/stale while the real CY revenue
        # never got written anywhere.
        #
        # New approach: reconcile each of rows 5-7 INDIVIDUALLY.
        #   1. Match TB revenue accounts to a row via its label in column B
        #      (e.g. accounts containing "professional" -> the row labelled
        #      "Professional Income"; "commission" -> "Commission Recd...").
        #   2. Compare that TB group's total to the row's RESOLVED D-value.
        #   3. If they already match (within tolerance) -> leave the row
        #      alone (genuinely pre-filled correctly).
        #   4. If they DON'T match -> overwrite D{row} with the correct TB
        #      total as a plain number (replacing any stale formula/value).
        # TB accounts that don't match any row 5-7 label fall through to the
        # existing GROSS PROFIT injection below, exactly as before.
        npl_cache_rev = _cache.get("notes to p&l", {})

        rev_row_labels = {}   # {row: normalised label}
        # FIX: previously hardcoded range(5, 8) — only 3 rows — missing row 8
        # (Sale SGST/CGST 5%) and any template with more than 2 sale sub-rows.
        # Now dynamically scan from row 5 until the "Total revenue" row.
        _rev_total_row_found = None
        for r in range(5, 25):
            b_val = npl_cache_rev.get((r, 2))
            if b_val:
                _bl = str(b_val).strip().lower()
                if "total" in _bl and ("revenue" in _bl or "operation" in _bl or "sale" in _bl):
                    _rev_total_row_found = r
                    break
                if _bl and "total" not in _bl and "revenue from" not in _bl:
                    rev_row_labels[r] = _bl

        def _rev_row_for(acct_name):
            nl = acct_name.lower()
            # FIX: matching used to require only ANY ONE shared word
            # between the row label and the account name (e.g. label
            # "Sale Local" → words ["sale","local"], and since virtually
            # every revenue account name contains some form of
            # "sale"/"sales", every single revenue account matched this
            # row via the word "sale" alone — including "SALES CENTRAL",
            # which has its own separate, correctly-matched row, causing
            # it (and every other revenue account) to be folded into
            # "Sale Local" too and double-counted. Now excludes the
            # generic "sale"/"sales" word from consideration and
            # requires a genuinely distinguishing word (e.g. "local",
            # "central", "export") to actually match.
            best_r, best_score = None, 0
            for r, lbl in rev_row_labels.items():
                if not lbl:
                    continue
                lbl_words = [w for w in lbl.replace("/", " ").split()
                             if len(w) > 3 and w not in ("sale", "sales")]
                if not lbl_words:
                    continue
                score = sum(1 for w in lbl_words if w in nl)
                if score > best_score:
                    best_score = score
                    best_r = r
            return best_r

        rev_row_totals = {}   # {row: tb_total}
        consumed_accounts = set()  # id() of accounts matched to a row
        for a in sale_accounts:
            r = _rev_row_for(a["name"])
            if r is not None:
                rev_row_totals[r] = rev_row_totals.get(r, 0.0) + abs(a["net"])
                consumed_accounts.add(id(a))

        for r, tb_total in rev_row_totals.items():
            existing_resolved = _npl_cache_do.get((r, 4))
            existing_val = existing_resolved if isinstance(existing_resolved, (int, float)) else None
            if existing_val is not None and abs(existing_val - tb_total) < max(1.0, tb_total * 0.001):
                log.append(
                    f"· notes to p&l!D{r} already correct ({existing_val:,.2f}) — not overwritten"
                )
                continue
            if _safe_write(ws_npl, r, 4, tb_total):
                injected.append(f"notes to p&l!D{r} ({rev_row_labels.get(r,'')}) = {tb_total:,.2f}")
            else:
                # The cell holds a stale formula (e.g. "=10898146+793859")
                # whose RESOLVED value doesn't match this year's TB total —
                # _safe_write/_safe_set both refuse to touch formula cells,
                # so overwrite directly with a plain number here.
                cell = _get_writable_cell(ws_npl, r, 4)
                if cell is not None:
                    cell.value = round(float(tb_total), 2)
                    injected.append(
                        f"notes to p&l!D{r} ({rev_row_labels.get(r,'')}) = {tb_total:,.2f} "
                        f"(overwrote stale formula)"
                    )

        # Remaining TB revenue accounts (not matched to rows 5-7) continue
        # through the existing GROSS PROFIT injection pipeline, unchanged.
        sale_accounts = [a for a in sale_accounts if id(a) not in consumed_accounts]
        tb_revenue_total = sum(abs(a["net"]) for a in sale_accounts)
        revenue_prefilled = (tb_revenue_total == 0)
        if sale_accounts:
            log.append(
                f"· {len(sale_accounts)} revenue account(s) totalling "
                f"{tb_revenue_total:,.2f} did not match a notes to p&l!D5:D7 "
                f"label — passing through to GROSS PROFIT injection"
            )

        # Defined here (shared scope) so the purchases block below — which
        # is NOT gated by `revenue_prefilled` — can also use it without an
        # UnboundLocalError if the sales block above is skipped.
        ws_gp = wb["GROSS PROFIT"] if "GROSS PROFIT" in wb.sheetnames else None

        if not revenue_prefilled and "GROSS PROFIT" in wb.sheetnames:
            # FIX (Issues 2 & 4): Trading-account formatting.
            #
            # Previously this block used HARD-CODED row numbers
            # (12% Within → row 12, 18% Within → row 14 …) which assumed
            # one specific template layout. If the user's BS template has a
            # different layout (e.g. extra rows, different section order,
            # different GST rates listed), the sale amounts landed on the
            # WRONG rows — sometimes on the TOTAL row, overwriting its formula
            # and showing the same number on every row.
            #
            # New strategy: SCAN column D of the GROSS PROFIT sheet for sale-side
            # row labels (B11:D24 area), build a label→row map dynamically, and
            # write each TB sale account into the row whose label best matches.
            # Unknown rate codes get appended below the last labelled row.
            gp_cache = _cache.get("GROSS PROFIT", {})

            # FIX: same pattern as the purchases-side fix above. Some
            # templates (e.g. AD Garments) have GROSS PROFIT's "Sales
            # GST" row driven entirely by a formula pulling from notes
            # to p&l (col E = "='notes to p&l'!D8"), meaning sales are
            # meant to be entered there, not directly in GROSS PROFIT.
            # That same template ALSO has several pre-existing broken
            # "='notes to p&l'!#REF!" formulas scattered in column D
            # (left over from a deleted row/column reference) which
            # evaluate as effectively blank in the cache, making the
            # code think there's room to append "unmatched" sale
            # accounts there — landing them well past the TOTAL row
            # (outside its SUM range, so excluded from any total) and,
            # worse, overwriting a genuine pre-existing formula row's
            # label with an unrelated sale account name. Detect the
            # formula-driven pattern up front and skip direct injection
            # entirely in that case — sales already get written via the
            # notes to p&l revenue-row matching above, so nothing is
            # lost by skipping this block.
            _gp_sale_is_formula_driven = False
            # FIX: the sales-side "label, value" column pair isn't
            # consistently D/E across templates — AD Garments uses E/F
            # instead (label in column E, formula in column F), one
            # column over from what this check originally assumed.
            # Check both layouts so the formula-driven pattern is
            # correctly detected regardless of which the template uses.
            _found_sales_gst_row = False
            for r in range(1, 40):
                for _lbl_col, _val_col in ((4, 5), (5, 6)):
                    lbl = gp_cache.get((r, _lbl_col))
                    if lbl and "sales" in str(lbl).strip().lower() and "gst" in str(lbl).strip().lower():
                        _found_sales_gst_row = True
                        val = gp_cache.get((r, _val_col))
                        if isinstance(val, str) and val.startswith("="):
                            _gp_sale_is_formula_driven = True
                        break
                if _found_sales_gst_row:
                    break

            # Accounts whose value gets written into GROSS PROFIT (either
            # matched to an existing row or appended as a new one) end up
            # here — used below to prevent the separate "inject into notes
            # to p&l" step from writing the SAME figure a second time as a
            # redundant duplicate row.
            _gp_handled_sale_ids = set()

            if _gp_sale_is_formula_driven:
                injected.append(
                    "  [GP] Sales: GROSS PROFIT's Sales GST row is a "
                    "formula pulling from 'notes to p&l' — skipping direct "
                    "injection here; sales already handled via the notes "
                    "to p&l revenue matching above."
                )
                sale_label_rows = {}
                sale_total_rows = set()
                unmatched_sales = list(sale_accounts)
            else:
                # Build label→row map by scanning the Sales side of GROSS PROFIT.
                # Sale rows live in col D (label) with amounts in col E —
                # EXCEPT some templates (this one included) shift the whole
                # sales-side table one column right, using col E (label) /
                # col F (amount) instead. The formula-driven detection above
                # already checks both (4,5) and (5,6) layouts; this scan was
                # hardcoded to (4,5) only, so on an E/F-layout template it
                # picked up unrelated column-D header text instead — in this
                # case the literal date string '31.03.2022' (a 3rd "previous
                # year" column header), which then got treated as if it were
                # a genuine sale sub-row label.
                # FIX (Bug 2): Also detect E/F layout by checking whether col E
                # contains sale-related labels (Sale, Job Work, etc.) even when
                # those labels don't contain "GST" — some templates (Penguin
                # Packages style) have plain "Sale" and "Job Work" sub-rows on
                # the sales side without GST-rate differentiation.
                _sale_lbl_col, _sale_val_col = (4, 5)
                # First: check for "sales gst" in col E (existing detection)
                _col_e_has_sales_gst = False
                for r in range(1, 40):
                    _probe = gp_cache.get((r, 5))
                    if _probe and "sales" in str(_probe).strip().lower() \
                            and "gst" in str(_probe).strip().lower():
                        _col_e_has_sales_gst = True
                        break
                if _col_e_has_sales_gst:
                    _sale_lbl_col, _sale_val_col = (5, 6)
                else:
                    # FIX (Bug 2): check whether the TOTAL row in col E/F layout
                    # sums a column F range. If TOTAL is in col E and value in F,
                    # it's E/F layout. Also check if col E has "SALES" header.
                    # Additionally check col 5 (E) for any non-date, non-empty
                    # non-formula text that looks like a sale sub-row label
                    # (e.g. "Sale", "Job Work", "Sale GST 12%").
                    _col_e_sale_labels = 0
                    for r in range(1, 40):
                        _probe_e = gp_cache.get((r, 5))
                        if not _probe_e:
                            continue
                        _pe_s = str(_probe_e).strip().lower()
                        # Skip formulas, dates, headers
                        if _pe_s.startswith("=") or re.match(r'^\d', _pe_s):
                            continue
                        if _pe_s in ("sales", "closing stock", "total", "particulars",
                                     "current year", "previous year", "rs.", "p"):
                            continue
                        if any(w in _pe_s for w in ("sale", "job work", "service", "export",
                                                      "domestic", "local sale", "interstate")):
                            # Check col F has a value (numeric or formula) → confirms E/F layout
                            _probe_f = gp_cache.get((r, 6))
                            if _probe_f is not None:
                                _col_e_sale_labels += 1
                    if _col_e_sale_labels >= 1:
                        _sale_lbl_col, _sale_val_col = (5, 6)

                _date_like = re.compile(r'^\d{1,2}[./]\d{1,2}[./]\d{2,4}$')

                sale_label_rows = {}
                sale_total_rows = set()
                for r in range(1, 40):
                    lbl = gp_cache.get((r, _sale_lbl_col))
                    if not lbl:
                        continue
                    lbl_s = str(lbl).strip()
                    if not lbl_s:
                        continue
                    # Skip formula-text labels (broken #REF! references or
                    # cross-sheet pulls) — these aren't real sale sub-row
                    # labels and shouldn't influence row placement.
                    if lbl_s.startswith("="):
                        continue
                    # Skip date-shaped strings (e.g. '31.03.2022' year-header
                    # text) — never a genuine sale sub-row label.
                    if _date_like.match(lbl_s):
                        continue
                    lbl_low = lbl_s.lower()
                    # Skip section headers (SALES, CLOSING STOCK), generic
                    # column headers (PARTICULARS), and totals
                    if lbl_low in ("sales", "closing stock", "particulars", "") \
                            or "total" in lbl_low:
                        if "total" in lbl_low:
                            sale_total_rows.add(r)
                        continue
                    if "as certified" in lbl_low or "proprietor" in lbl_low:
                        continue
                    # This is a sub-item row — record it
                    _norm_lbl = _norm_gst(lbl_s)
                    if not _norm_lbl:
                        # Label normalised to nothing (e.g. it was ONLY the
                        # word "Sales"/"GST"/etc, stripped entirely) — an
                        # empty key would later match EVERY account via the
                        # "name_norm in lbl_norm" substring fallback below
                        # (an empty string is a substring of anything in
                        # Python), causing spurious matches. Skip it.
                        continue
                    sale_label_rows[_norm_lbl] = (r, lbl_s)

                # Place each TB sale into its matching row
                sale_row_totals = {}     # {row: amount}
                unmatched_sales = []
                # Track which accounts get a value written into GROSS PROFIT
                # (matched to an existing row OR appended as a new row there)
                # so the separate "also inject into notes to p&l" step below
                # never duplicates the SAME figure as a second, redundant
                # line — notes to p&l!D6 already pulls from GROSS PROFIT via
                # formula, so writing it again elsewhere just shows the same
                # sale twice and adds a row that wasn't in the original
                # template, which is exactly what should never happen:
                # existing template formulas stay untouched, only their
                # SOURCE cell gets the value, and row insertion (handled by
                # the cross-sheet shift fix elsewhere) is the only thing
                # that should ever shift anything.
                _gp_handled_sale_ids = set()
                for acct in sale_accounts:
                    name_norm = _norm_gst(acct["name"])
                    best_row = None
                    best_label = None
                    if not name_norm:
                        # Account name normalised to nothing (e.g. the
                        # account is literally just "Sales") — there's no
                        # specific GST-rate signal to match against. If the
                        # template has exactly one real sale sub-row, that's
                        # almost certainly where this generic figure belongs.
                        # FIX: even when there are multiple GST-rate sub-rows
                        # (e.g. "Sales GST Local 12%", "Local 18%", "Local 5%"),
                        # a TB with a single undifferentiated "Sales" account
                        # has no rate signal to distribute across sub-rows —
                        # the entire amount belongs in the first available row.
                        # Appending as a new row (the old behaviour when count
                        # != 1) overrode the PURCHASES / CLOSING STOCK section
                        # with a stray "Sales" label row and the amount in the
                        # wrong column, corrupting the GROSS PROFIT layout.
                        if len(sale_label_rows) >= 1:
                            best_row, best_label = next(iter(sale_label_rows.values()))
                            sale_row_totals[best_row] = sale_row_totals.get(best_row, 0) + abs(acct["net"])
                            _gp_handled_sale_ids.add(id(acct))
                        else:
                            unmatched_sales.append(acct)
                        continue
                    # Prefer exact normalised match, else substring either-way
                    if name_norm in sale_label_rows:
                        best_row, best_label = sale_label_rows[name_norm]
                    else:
                        for lbl_norm, (r, lbl_orig) in sale_label_rows.items():
                            if lbl_norm and (lbl_norm in name_norm or name_norm in lbl_norm):
                                best_row, best_label = r, lbl_orig
                                break
                    if best_row is None:
                        unmatched_sales.append(acct)
                        continue
                    sale_row_totals[best_row] = sale_row_totals.get(best_row, 0) + abs(acct["net"])
                    _gp_handled_sale_ids.add(id(acct))

                # Write each label-matched row — NEVER write to a Total row
                for row, amt in sale_row_totals.items():
                    if row in sale_total_rows:
                        continue
                    if _safe_write(ws_gp, row, _sale_val_col, amt):
                        injected.append(
                            f"GROSS PROFIT!{chr(64+_sale_val_col)}{row} (Sale {amt:,.2f})"
                        )

                # Unmatched sales — append below the last sale sub-row but ABOVE total
                # Also hard-stop before any "CLOSING STOCK", "PURCHASES", or
                # formula-bearing row to avoid overwriting those sections.
                if unmatched_sales and sale_label_rows:
                    last_sub_row = max(r for r, _ in sale_label_rows.values())
                    # Find the next section header row below last_sub_row
                    _sale_stop_row = None
                    for _ssr in range(last_sub_row + 1, 40):
                        _sv = gp_cache.get((_ssr, _sale_lbl_col))
                        if _sv and str(_sv).strip().lower() in (
                                'closing stock', 'closing stock ', 'purchases',
                                'total', 'gross profit', 'gross profit '):
                            _sale_stop_row = _ssr
                            break
                    next_row = last_sub_row + 1
                    for acct in unmatched_sales:
                        if next_row in sale_total_rows:
                            break
                        if _sale_stop_row and next_row >= _sale_stop_row:
                            break
                        if _safe_write(ws_gp, next_row, _sale_lbl_col, acct["name"]):
                            _safe_write(ws_gp, next_row, _sale_val_col, abs(acct["net"]))
                            _gp_handled_sale_ids.add(id(acct))
                            injected.append(
                                f"GROSS PROFIT!{chr(64+_sale_val_col)}{next_row} "
                                f"(Sale unmatched: {acct['name']}) = {abs(acct['net']):,.2f}"
                            )
                            next_row += 1

            # Also inject individual sales into notes to p&l (rows 7-14)
            if ws_npl:
                import re as _re2
                def _extract_rate(name):
                    m = _re2.search(r'(\d+)\s*%', name)
                    return m.group(1) if m else None

                still_unmatched_sale2 = []
                for acct in sale_accounts:
                    if id(acct) in _gp_handled_sale_ids:
                        # Already written into GROSS PROFIT above, and
                        # notes to p&l!D6 pulls from there via formula
                        # (e.g. ='GROSS PROFIT'!F10) — writing the same
                        # figure again here would duplicate it as a second,
                        # redundant "Sales" line that was never part of the
                        # original template.
                        continue
                    amt = abs(acct["net"])
                    if amt <= 0: continue
                    acct_rate = _extract_rate(acct["name"])
                    wrote = False
                    for r in range(7, 15):
                        cell_name = ws_npl.cell(r, 2).value
                        if not cell_name: continue
                        tmpl_rate = _extract_rate(str(cell_name))
                        if acct_rate and tmpl_rate and acct_rate == tmpl_rate:
                            acct_l = acct["name"].lower()
                            tmpl_l = str(cell_name).lower()
                            if (("within" in acct_l) == ("within" in tmpl_l)) or \
                               ("inter" not in acct_l and "within" not in acct_l):
                                if _safe_write(ws_npl, r, 4, amt):
                                    injected.append(f"notes to p&l!D{r} (Sale: {acct['name']}) = {amt:,.2f}")
                                    wrote = True
                                break
                    if not wrote:
                        still_unmatched_sale2.append(acct)

                # FIX: the rows-7-to-9 fallback above has only 3 slots —
                # in templates where the revenue section is genuinely
                # tight (e.g. AD Garments has exactly "Sale Local" /
                # "Sale Central" with NO spare row before the TOTAL),
                # additional individual customer/freight accounts had
                # nowhere to go and were silently dropped with no log
                # entry at all. Inserting rows HERE, mid-pipeline, was
                # tried and found to actively corrupt OTHER injection
                # steps further down this function: many of them (e.g.
                # the Other Income block) read from `_cache`/`npl_cache`
                # dictionaries that were pre-loaded ONCE at the top of
                # this function and assume row positions stay stable for
                # the rest of its execution. Inserting rows mid-way
                # shifts physical row positions but does NOT update
                # those caches, so later steps end up writing into
                # whatever unrelated content now occupies the OLD
                # (stale) row numbers they were still relying on.
                # The safe fix is to defer ALL row insertions to a single
                # pass at the very end of the function, after every
                # cache-dependent step has already finished reading and
                # writing. Queue this request instead of acting on it now.
                if still_unmatched_sale2:
                    _deferred_section_expansions.append({
                        "sheet": "notes to p&l",
                        "section_header_substr": "revenue from operation",
                        "total_label": "total revenue from operation",
                        "accounts": still_unmatched_sale2,
                        "label_col": 2,
                        "value_col": 4,
                    })
                    injected.append(
                        f"· {len(still_unmatched_sale2)} sale account(s) totalling "
                        f"{sum(abs(a['net']) for a in still_unmatched_sale2):,.2f} queued "
                        f"for safe section expansion at end of processing (revenue section "
                        f"had no spare rows)"
                    )

        # ── B. Purchases → GROSS PROFIT!B14:B18 ─────────────────────
        # Row 14=Purchase GST 12% Interstate, 15=12% WS, 16=18% WS
        # Row 17=5% Interstate, 18=5% WS
        if "GROSS PROFIT" in wb.sheetnames:
            purchase_accounts = [a for a in individual_accounts
                                  if a.get("bs_head") == "purchases"
                                  and abs(a.get("net", 0)) > 0]

            # FIX: some templates (e.g. AD Garments) have a single
            # "Purchase GST" row in GROSS PROFIT whose VALUE CELL is a
            # formula like "=SUM('notes to p&l'!D21:D26)" — meaning
            # purchases are meant to be entered in the notes to p&l
            # sheet, and GROSS PROFIT just pulls the total automatically.
            # Other templates (e.g. Fashion Adda) instead have 5+
            # separate GST-rate sub-rows directly in GROSS PROFIT with
            # plain numeric cells expecting direct injection. Running
            # the direct-injection logic below on a formula-driven
            # template caused two serious corruptions: (1) the lone
            # "Purchase GST" row's normalised label was empty/unhelpful,
          # so account-matching never found a target and every account
            # fell into the "unmatched" fallback; (2) that fallback then
            # treated unrelated DIRECT EXPENSES rows (whose labels are
            # formula text like "='notes to p&l'!B29", which still
            # produces a non-empty normalised string) as valid "purchase
            # label rows" purely by accident, computing the wrong
            # insertion point and overwriting their labels with
            # unrelated purchase account names. Detect this pattern up
            # front and skip the whole block — defer entirely to the
            # notes to p&l injection further below.
            _gp_purchase_is_formula_driven = False
            gp_cache_probe = _cache.get("GROSS PROFIT", {})
            for r in range(1, 40):
                lbl = gp_cache_probe.get((r, 1))
                if lbl and "purchase" in str(lbl).strip().lower() and "gst" in str(lbl).strip().lower():
                    val = gp_cache_probe.get((r, 2))
                    if isinstance(val, str) and val.startswith("="):
                        _gp_purchase_is_formula_driven = True
                    break

            # Accounts whose value gets written into GROSS PROFIT end up
            # here — used below to prevent the "inject into notes to p&l"
            # step from duplicating the same figure as a redundant row.
            _gp_handled_purch_ids = set()

            if _gp_purchase_is_formula_driven:
                injected.append(
                    "  [GP] Purchases: GROSS PROFIT's Purchase row is a "
                    "formula pulling from 'notes to p&l' — skipping direct "
                    "injection here, handled in the notes to p&l section below."
                )
                purch_label_rows = {}
                unmatched_purch = purchase_accounts
                purch_total_rows = set()
                purch_row_totals = {}
            else:
                # FIX (Issues 2 & 4): Same label-driven approach as Sales.
                # Purchases live on the DEBIT side of GROSS PROFIT —
                # column A holds labels, column B holds amounts.
                gp_cache = _cache.get("GROSS PROFIT", {})
                purch_label_rows = {}
                purch_total_rows = set()
                # FIX: only scan rows between "PURCHASES" and the next
                # major header (DIRECT EXPENSES / GROSS PROFIT).
                # Previously scanning from row 1 picked up title rows
                # ("TRADING ACCOUNT...", "PARTICULARS") and Direct
                # Expenses rows ("Cartage Inward") as fake purchase
                # targets, causing all real purchase accounts to be
                # "unmatched" and appended at wrong positions that
                # overwrote GROSS PROFIT formula cells.
                _purch_start_row = None
                _purch_end_row = None
                for _pr in range(1, 40):
                    _plbl = gp_cache.get((_pr, 1))
                    if not _plbl: continue
                    _plbl_l = str(_plbl).strip().lower().rstrip()
                    if _plbl_l == "purchases" and _purch_start_row is None:
                        _purch_start_row = _pr + 1
                    elif _purch_start_row and _plbl_l in (
                            "direct expenses", "gross profit", "gross profit ",
                            "closing stock", "closing stock ", "total", "sales"):
                        _purch_end_row = _pr
                        break
                if _purch_start_row is None:
                    _purch_start_row = 1
                if _purch_end_row is None:
                    _purch_end_row = 40

                for r in range(_purch_start_row, _purch_end_row):
                    lbl = gp_cache.get((r, 1))  # col A
                    if not lbl:
                        continue
                    lbl_s = str(lbl).strip()
                    if not lbl_s:
                        continue
                    if lbl_s.startswith("="):
                        continue
                    lbl_low = lbl_s.lower()
                    if lbl_low in ("purchases", "direct expenses", "opening stock",
                                    "gross profit"):
                        continue
                    if "total" in lbl_low:
                        purch_total_rows.add(r)
                        continue
                    norm = _norm_gst(lbl_s)
                    if not norm:
                        continue
                    purch_label_rows[norm] = (r, lbl_s)

                purch_row_totals = {}
                unmatched_purch = []
                _gp_handled_purch_ids = set()
                for acct in purchase_accounts:
                    name_norm = _norm_gst(acct["name"])
                    best_row = None
                    if name_norm in purch_label_rows:
                        best_row = purch_label_rows[name_norm][0]
                    else:
                        for lbl_norm, (r, _) in purch_label_rows.items():
                            if lbl_norm and (lbl_norm in name_norm or name_norm in lbl_norm):
                                best_row = r
                                break
                    if best_row is None:
                        unmatched_purch.append(acct)
                    else:
                        purch_row_totals[best_row] = purch_row_totals.get(best_row, 0) + abs(acct["net"])
                        _gp_handled_purch_ids.add(id(acct))

                for row, amt in purch_row_totals.items():
                    if row in purch_total_rows:
                        continue   # never overwrite TOTAL row
                    if _safe_write(ws_gp, row, 2, amt):
                        injected.append(f"GROSS PROFIT!B{row} (Purchase {amt:,.2f})")

                # Append unmatched purchase accounts below last sub-row, above TOTAL
                # Hard cap: never write at or beyond the Direct Expenses /
                # Gross Profit header row (_purch_end_row) — those rows
                # contain important formulas that must not be overwritten.
                if unmatched_purch and purch_label_rows:
                    last_sub_row = max(r for r, _ in purch_label_rows.values())
                    next_row = last_sub_row + 1

                    # FIX: when all TB purchases are unmatched because the
                    # template uses GST-rate sub-rows (e.g. "Purchases GST
                    # 12% Local") while the TB has generic accounts (e.g.
                    # "GOODS PURCHASE A/C"), appending after last_sub_row
                    # MIGHT land outside the formula range that notes to
                    # p&l's "Purchases" cell pulls from (e.g.
                    # =SUM('GROSS PROFIT'!B13:B17)). Even if next_row is
                    # still within the GROSS PROFIT section bounds, a
                    # value written there is invisible to that formula.
                    # Detect the actual formula range end from notes to
                    # p&l (if available) and compare against next_row.
                    _formula_range_end = _purch_end_row  # default: section end
                    if ws_npl:
                        import re as _re_fr
                        for _fr_r in range(1, 50):
                            _fr_v = ws_npl.cell(_fr_r, 4).value
                            if isinstance(_fr_v, str) and 'GROSS PROFIT' in _fr_v:
                                _m = _re_fr.search(
                                    r"'GROSS PROFIT'!\$?[A-Z]+\$?(\d+):\$?[A-Z]+\$?(\d+)",
                                    _fr_v)
                                if _m:
                                    _formula_range_end = int(_m.group(2)) + 1
                                    break

                    _all_unmatched = len(unmatched_purch) == len(purchase_accounts)
                    _outside_formula = next_row >= _formula_range_end
                    if _all_unmatched and _outside_formula and purch_label_rows:
                        # Find first sub-row with an empty B cell (writable)
                        _first_avail = None
                        for _pr_r, _pr_lbl in sorted(purch_label_rows.values()):
                            if ws_gp.cell(_pr_r, 2).value is None:
                                _first_avail = _pr_r
                                break
                        if _first_avail:
                            _total = sum(abs(a["net"]) for a in unmatched_purch)
                            if _safe_write(ws_gp, _first_avail, 2, _total):
                                for a in unmatched_purch:
                                    _gp_handled_purch_ids.add(id(a))
                                injected.append(
                                    f"GROSS PROFIT!B{_first_avail} (Purchases aggregated "
                                    f"into first sub-row, {len(unmatched_purch)} unmatched "
                                    f"accounts totalling {_total:,.2f})"
                                )
                                unmatched_purch = []  # all handled

                    # FIX: skip appending if the template's purchase sub-rows
                    # already contain the same total amount (template contaminated
                    # with previous-year data). Common case: "PURCHASE ACCOUNTS"
                    # (group-level total account, normalized empty name → unmatched)
                    # when the template already has "Purchase Local GST 5%" = same
                    # amount from the prior year — appending would write it twice.
                    if unmatched_purch and purch_label_rows:
                        _existing_total = sum(
                            abs(ws_gp.cell(r, 2).value or 0)
                            for r, _ in purch_label_rows.values()
                            if isinstance(ws_gp.cell(r, 2).value, (int, float))
                        )
                        _unmatched_total = sum(abs(a["net"]) for a in unmatched_purch)
                        if _existing_total > 0 and abs(_existing_total - _unmatched_total) < 1:
                            # Template already has this total — skip to avoid doubling
                            for a in unmatched_purch:
                                _gp_handled_purch_ids.add(id(a))
                                skipped.append(
                                    f"purch '{a['name']}': template already has {_existing_total:,.2f} "
                                    f"in existing sub-rows — skipping to prevent doubling"
                                )
                            unmatched_purch = []

                    for acct in unmatched_purch:
                        if next_row in purch_total_rows:
                            break
                        if next_row >= _purch_end_row:
                            skipped.append(
                                f"purch '{acct['name']}': no room in GROSS PROFIT "
                                f"purchase section (next_row {next_row} >= end {_purch_end_row})"
                            )
                            continue
                        if _safe_write(ws_gp, next_row, 1, acct["name"]):
                            _safe_write(ws_gp, next_row, 2, abs(acct["net"]))
                            _gp_handled_purch_ids.add(id(acct))
                            injected.append(
                                f"GROSS PROFIT!B{next_row} (Purchase unmatched: {acct['name']}) = {abs(acct['net']):,.2f}"
                            )
                            next_row += 1

            # ── B-bis. Purchases → notes to p&l (rate-keyword sub-rows) ──
            # FIX: GROSS PROFIT!B14 ("Purchase GST") is itself a FORMULA
            # (=SUM('notes to p&l'!D22:D29)) in many templates — _safe_write
            # silently skips formula cells, so the purchases total appeared
            # to vanish even though purch_row_totals had the right number.
            # The TRUE writable target is the detail rows in 'notes to p&l'
            # (e.g. R22 "Purchase GST Central 0%", R26 "Purchases GST Local
            # 18%", etc.) — same pattern as the Sales injection above.
            # Without this, purchase figures never reached the template,
            # Closing Stock (the Trading A/c balancing figure) absorbed the
            # entire missing purchase amount and went hugely negative, and
            # any TRULY unmatched purchase account fell through to GROSS
            # PROFIT's "append below last row" logic — which collided with
            # and overwrote the DIRECT EXPENSES section directly below.
            if ws_npl:
                import re as _re3
                def _extract_rate3(name):
                    m = _re3.search(r'(\d+)\s*%', name)
                    return m.group(1) if m else None

                # Find the Purchases sub-section bounds in notes to p&l:
                # starts after a row labelled "Purchases" (col B), ends at
                # the next section label (col A non-empty, e.g. "(c)").
                # FIX (Bug 3): also detect the sub-section by finding the first
                # row INSIDE the Purchases section that the GROSS PROFIT Purchase
                # formula references via ='notes to p&l'!D{row}. That target row
                # is the TRUE destination for the purchase total. Some templates
                # have a first sub-row labelled '-Purchase GST (Net of Discounts)'
                # (which B14 in GROSS PROFIT references as D24) while the second
                # sub-row 'Purchase' (D25) has the generic purchase label —
                # rate/name matching always landed in D25, leaving D24 empty
                # and causing GROSS PROFIT!B14 to show nothing.
                purch_start, purch_end = None, None
                _npl_purch_formula_row = None   # row that GP formula references
                for r in range(1, 60):
                    lbl_b = ws_npl.cell(r, 2).value
                    if lbl_b and str(lbl_b).strip().lower() == "purchases":
                        purch_start = r + 1
                        break
                if purch_start:
                    for r in range(purch_start, purch_start + 20):
                        lbl_a = ws_npl.cell(r, 1).value
                        if lbl_a and str(lbl_a).strip():
                            purch_end = r
                            break
                    if purch_end is None:
                        purch_end = purch_start + 12

                # Detect which notes to p&l row the GROSS PROFIT Purchase row
                # references (if any), so we write to that exact row.
                if ws_gp:
                    import re as _re_gp_ref
                    for _gpr in range(1, 40):
                        _gp_a = ws_gp.cell(_gpr, 1).value
                        _gp_b = ws_gp.cell(_gpr, 2).value
                        if _gp_a and "purchase" in str(_gp_a).lower():
                            if isinstance(_gp_b, str) and "notes to p&l" in _gp_b:
                                _m_ref = _re_gp_ref.search(r"!D(\d+)", _gp_b)
                                if _m_ref:
                                    _npl_purch_formula_row = int(_m_ref.group(1))
                            break

                still_unmatched_purch = []
                if purch_start and purch_end:
                    for acct in purchase_accounts:
                        if id(acct) in _gp_handled_purch_ids:
                            # Already written into GROSS PROFIT above —
                            # writing it again here would duplicate the
                            # same figure as a second, redundant row.
                            continue
                        amt = abs(acct["net"])
                        if amt <= 0:
                            continue
                        acct_l = acct["name"].lower()
                        acct_rate = _extract_rate3(acct["name"])
                        wrote = False

                        # FIX (Bug 3): If the GROSS PROFIT formula references a
                        # specific notes to p&l row and that row is empty,
                        # write to it directly (regardless of label matching)
                        # when the account has no GST rate (generic purchase).
                        if (not acct_rate and _npl_purch_formula_row and
                                purch_start <= _npl_purch_formula_row < purch_end):
                            _target_cell = ws_npl.cell(_npl_purch_formula_row, 4)
                            if _target_cell.value is None or _target_cell.value == 0:
                                if _safe_write(ws_npl, _npl_purch_formula_row, 4, amt):
                                    injected.append(
                                        f"notes to p&l!D{_npl_purch_formula_row} "
                                        f"(Purchase GP-formula target: {acct['name']}) = {amt:,.2f}"
                                    )
                                    wrote = True
                            elif not _is_formula(_target_cell.value):
                                # Already has a value — accumulate
                                existing_v = float(_target_cell.value or 0)
                                if _safe_write(ws_npl, _npl_purch_formula_row, 4, existing_v + amt):
                                    injected.append(
                                        f"notes to p&l!D{_npl_purch_formula_row} "
                                        f"(Purchase GP-formula target accum: {acct['name']}) = {existing_v+amt:,.2f}"
                                    )
                                    wrote = True

                        if not wrote:
                            for r in range(purch_start, purch_end):
                                cell_name = ws_npl.cell(r, 2).value
                                if not cell_name:
                                    continue
                                tmpl_l = str(cell_name).lower()
                                tmpl_rate = _extract_rate3(str(cell_name))
                                rate_match = acct_rate and tmpl_rate and acct_rate == tmpl_rate
                                region_match = (
                                    ("central" in acct_l) == ("central" in tmpl_l)
                                    and ("local" in acct_l) == ("local" in tmpl_l)
                                )
                                if rate_match and region_match:
                                    if _safe_write(ws_npl, r, 4, amt):
                                        injected.append(
                                            f"notes to p&l!D{r} (Purchase: {acct['name']}) = {amt:,.2f}"
                                        )
                                        wrote = True
                                    break
                        # FIX: the rate/region match above requires BOTH
                        # the TB account name AND the template label to
                        # contain a GST percentage (e.g. "5%", "18%") —
                        # but some clients' purchase accounts have no
                        # rate suffix at all (e.g. "PURCHASE GARMENTS GST
                        # LOCAL" vs template label "Purchase Garments Gst
                        # Local", which match near-exactly by NAME but
                        # neither has a numeric rate to compare). Without
                        # this fallback, such accounts always fell
                        # through to "unmatched" and got folded into a
                        # single row together with unrelated accounts,
                        # silently combining distinct line items into
                        # one figure.
                        if not wrote and not acct_rate:
                            acct_words = {w for w in re.findall(r"[a-z]+", acct_l)
                                          if w not in ("purchase", "purchases", "gst", "a", "c")}
                            best_r, best_overlap = None, 0
                            for r in range(purch_start, purch_end):
                                cell_name = ws_npl.cell(r, 2).value
                                if not cell_name:
                                    continue
                                tmpl_l2 = str(cell_name).lower()
                                tmpl_words = {w for w in re.findall(r"[a-z]+", tmpl_l2)
                                              if w not in ("purchase", "purchases", "gst", "a", "c")}
                                # FIX: a single shared generic word (e.g.
                                # "garments", present in BOTH "Purchase
                                # Garments Gst Local" and "Purchase
                                # Garments Igst") isn't enough to
                                # distinguish between two similarly-named
                                # categories — pick the template row with
                                # the MOST matching words, not just the
                                # first one with any overlap at all, so
                                # the more specific distinguishing word
                                # ("local" vs "igst") decides the match.
                                overlap = len(acct_words & tmpl_words)
                                if overlap > best_overlap:
                                    best_overlap = overlap
                                    best_r = r
                            if best_r is not None and best_overlap > 0:
                                if _safe_write(ws_npl, best_r, 4, amt):
                                    injected.append(
                                        f"notes to p&l!D{best_r} (Purchase name-match: {acct['name']}) = {amt:,.2f}"
                                    )
                                    wrote = True
                        if not wrote:
                            still_unmatched_purch.append(acct)

                    # Any genuinely unmatched purchase accounts get appended
                    # as NEW rows strictly inside the Purchases sub-section
                    # (between purch_start and purch_end), never spilling
                    # into Direct Expenses below.
                    next_row = purch_start
                    # Find first truly empty row within the sub-section
                    while next_row < purch_end and ws_npl.cell(next_row, 2).value:
                        next_row += 1
                    for acct in still_unmatched_purch:
                        amt = abs(acct.get("net", 0))
                        if amt <= 0:
                            continue
                        if next_row >= purch_end:
                            # No room left in the template's sub-section —
                            # fold into the last purchase row rather than
                            # overwrite Direct Expenses.
                            last_r = purch_end - 1
                            existing = ws_npl.cell(last_r, 4).value
                            existing_amt = existing if isinstance(existing, (int, float)) else 0
                            _safe_write(ws_npl, last_r, 4, existing_amt + amt)
                            injected.append(
                                f"notes to p&l!D{last_r} (Purchase folded, no room: {acct['name']}) = {amt:,.2f}"
                            )
                            continue
                        if _safe_write(ws_npl, next_row, 2, acct["name"]):
                            _safe_write(ws_npl, next_row, 4, amt)
                            injected.append(
                                f"notes to p&l!D{next_row} (Purchase new: {acct['name']}) = {amt:,.2f}"
                            )
                            next_row += 1

        # ── B2. DIRECT EXPENSES → GROSS PROFIT Direct-Expenses sub-rows ──
        # FIX (Issues 2/4): Direct expenses (Electricity Exp, Wages A/c,
        # Oil & Lubricant) must NOT be collapsed onto the Purchases header.
        # They live in their own labelled sub-rows below "DIRECT EXPENSES"
        # on the debit side of the Trading A/c.
        if "GROSS PROFIT" in wb.sheetnames:
            de_accounts = [a for a in individual_accounts
                           if a.get("bs_head") == "direct_expenses"
                           and abs(a.get("net", 0)) > 0]
            if de_accounts:
                gp_cache_de = _cache.get("GROSS PROFIT", {})
                # Find the DIRECT EXPENSES header row
                de_header_row = None
                for r in range(1, 40):
                    lbl = gp_cache_de.get((r, 1))
                    if lbl and "direct expense" in str(lbl).lower():
                        de_header_row = r
                        break
                # Find the first "total" / "gross profit" row after it
                de_end_row = None
                if de_header_row:
                    for r in range(de_header_row + 1, de_header_row + 15):
                        lbl = gp_cache_de.get((r, 1))
                        if lbl and ("total" in str(lbl).lower()
                                    or "gross profit" in str(lbl).lower()):
                            de_end_row = r
                            break

                if de_header_row and de_end_row and de_end_row > de_header_row + 1:
                    # Build label→row map for the direct-expense sub-rows
                    de_rows = {}
                    for r in range(de_header_row + 1, de_end_row):
                        lbl = gp_cache_de.get((r, 1))
                        if lbl:
                            de_rows[str(lbl).strip().lower()] = r

                    used_rows = set()
                    for acct in de_accounts:
                        name_l = acct["name"].lower()
                        amt = abs(acct["net"])
                        placed = False
                        # Match against existing labelled rows
                        for lbl, r in de_rows.items():
                            if r in used_rows:
                                continue
                            if lbl in name_l or name_l in lbl or any(
                                w in lbl and w in name_l
                                for w in ("electricity", "wages", "oil",
                                          "lubricant", "power", "fuel")
                            ):
                                if _safe_write(ws_gp, r, 2, amt):
                                    injected.append(
                                        f"GROSS PROFIT!B{r} (Direct Exp: {acct['name']}) = {amt:,.2f}"
                                    )
                                    used_rows.add(r)
                                    placed = True
                                    break
                        if placed:
                            continue
                        # Append as new labelled sub-row before TOTAL.
                        # Use the LIVE worksheet (not the stale cache) to
                        # check whether a row is still free — earlier
                        # writes in the same run (e.g. purchase appends)
                        # won't be reflected in the cache, so we'd
                        # accidentally overwrite them.
                        for r in range(de_header_row + 1, de_end_row):
                            if r in used_rows:
                                continue
                            cur_lbl = ws_gp.cell(r, 1).value
                            cur_amt = ws_gp.cell(r, 2).value
                            if (not cur_lbl or str(cur_lbl).strip() == "") and (
                                cur_amt is None or cur_amt == 0
                            ):
                                if _safe_write(ws_gp, r, 1, acct["name"]):
                                    _safe_write(ws_gp, r, 2, amt)
                                    used_rows.add(r)
                                    injected.append(
                                        f"GROSS PROFIT!B{r} (Direct Exp new: {acct['name']}) = {amt:,.2f}"
                                    )
                                    placed = True
                                    break
                        if not placed:
                            skipped.append(
                                f"Direct expense '{acct['name']}' = {amt:,.2f}: no free row in Trading A/c"
                            )

        # ── C. Opening Stock → notes to p&l Opening Stock row ────────────
        # Dynamically find the "Inventories at the beginning of the year" /
        # "Opening stock" row in notes to p&l (not hardcoded row 17,
        # which breaks for templates with 2+ revenue sub-rows that push
        # all subsequent rows down — e.g. after a deferred section expansion).
        opening_stock_accounts = [a for a in individual_accounts
                                   if "opening stock" in a.get("name","").lower()
                                   and abs(a.get("net", 0)) > 0]
        if opening_stock_accounts:
            total_opening = sum(abs(a["net"]) for a in opening_stock_accounts)
            # Scan notes to p&l for the opening stock row (col B contains
            # "beginning" or "opening stock" or "inventories at the beginning")
            _os_row = None
            for _r in range(1, 60):
                _b = ws_npl.cell(_r, 2).value
                if _b:
                    _bl = str(_b).strip().lower()
                    if ("beginning" in _bl or "opening stock" in _bl or
                            ("inventori" in _bl and "beginning" in _bl)):
                        _os_row = _r
                        break
            if _os_row is None:
                _os_row = 20  # fallback: template default
            if _safe_write(ws_npl, _os_row, 4, total_opening):
                injected.append(f"notes to p&l!D{_os_row} (Opening stock) = {total_opening:,.2f}")

        # ── C2. OTHER INCOME (Rebate & Discount, Interest Received, etc.) ──
        # FIX (Issue 3): "REBATE & DISCOUNT" / other indirect-income accounts
        # were being classified correctly as `other_income` but never written
        # anywhere — so the amount silently disappeared from the BS / P&L.
        #
        # The standard "Notes to P&L" template carries an "Other Income"
        # section. We locate it by scanning column B of `notes to p&l` for a
        # row whose label contains "other income" / "other incomes", then
        # write each TB other-income account into a sub-row beneath it.
        oi_accounts = [a for a in individual_accounts
                       if a.get("bs_head") == "other_income"
                       and abs(a.get("net", 0)) > 0]
        if oi_accounts:
            npl_cache_for_oi = _cache.get("notes to p&l", {})
            oi_header_row = None
            oi_total_row  = None
            for r in range(1, 200):
                b = npl_cache_for_oi.get((r, 2))
                if not b:
                    continue
                bs = str(b).strip().lower()
                if "other income" in bs and "total" not in bs:
                    oi_header_row = r
                    break
            if oi_header_row:
                for r in range(oi_header_row + 1, oi_header_row + 20):
                    b = npl_cache_for_oi.get((r, 2))
                    if b and "total" in str(b).lower() and "other" in str(b).lower():
                        oi_total_row = r
                        break

            # FIX (Issues 1/2/3 continued): _fuzzy_match_name's generic
            # "single distinctive word >=7 chars" rule matches on the word
            # "received"/"RECEIVED" alone — which both "Bank Interest Recd."
            # and "PENSION RECEIVED FROM LIC" contain. This caused "Pension
            # received" to fuzzy-match the "Bank interest received" row
            # (and vice versa), so both accounts landed on the same row,
            # silently dropping one of them.
            #
            # For Other Income sub-row matching specifically, require that
            # at least one MEANINGFUL word (excluding generic terms like
            # "received"/"recd"/"income") also matches between the TB
            # account name and the template row label.
            _OI_GENERIC_WORDS = {"received", "recd", "income", "interest", "other"}

            def _oi_row_match(tb_name, row_label):
                import re as _re_oi
                def _words(s):
                    s = _re_oi.sub(r'[^a-z0-9\s]', ' ', s.lower())
                    return set(w for w in s.split() if len(w) > 3)
                a_words = _words(tb_name) - _OI_GENERIC_WORDS
                b_words = _words(row_label) - _OI_GENERIC_WORDS
                return bool(a_words & b_words)

            if oi_header_row and oi_total_row and oi_total_row > oi_header_row + 1:
                for acct in oi_accounts:
                    amt = abs(acct["net"])
                    placed = False
                    # First try matching an existing labelled sub-row, requiring
                    # at least one meaningful (non-generic) shared word.
                    for r in range(oi_header_row + 1, oi_total_row):
                        b = npl_cache_for_oi.get((r, 2))
                        if b and _oi_row_match(acct["name"], str(b)):
                            if _safe_write(ws_npl, r, 4, amt):
                                injected.append(
                                    f"notes to p&l!D{r} (Other Income: {acct['name']}) = {amt:,.2f}"
                                )
                                placed = True
                                break
                    if placed:
                        continue
                    # Otherwise find first empty sub-row
                    for r in range(oi_header_row + 1, oi_total_row):
                        d_val = npl_cache_for_oi.get((r, 4))
                        b_val = npl_cache_for_oi.get((r, 2))
                        if (d_val is None or d_val == 0) and (
                            b_val is None or str(b_val).strip() == ""
                        ):
                            if _safe_write(ws_npl, r, 2, acct["name"]):
                                _safe_write(ws_npl, r, 4, amt)
                                injected.append(
                                    f"notes to p&l!D{r} (Other Income new: {acct['name']}) = {amt:,.2f}"
                                )
                                placed = True
                                break
                    if not placed:
                        skipped.append(
                            f"Other Income '{acct['name']}' = {amt:,.2f}: no row in Other Income section"
                        )
            else:
                skipped.append(
                    "Other Income section not found in 'notes to p&l'. "
                    "Add a row labelled 'Other Income' to capture: "
                    + ", ".join(a["name"] for a in oi_accounts)
                )

        # ── D. Employee Expenses → match to template rows dynamically ──
        # Detect Employee benefits section in notes to p&l
        emp_start = None
        emp_end = None
        for r in range(30, 60):
            b = ws_npl.cell(r, 2).value
            if not b: continue
            bl = str(b).lower()
            if not emp_start and 'employee benefit' in bl and 'total' not in bl:
                emp_start = r + 1
            if emp_start and 'total' in bl and 'employee' in bl:
                emp_end = r
                break
        if not emp_start: emp_start = 41
        if not emp_end: emp_end = 46

        # All employee-type accounts (including ESI, bonus, leave with wages)
        emp_accounts = [a for a in individual_accounts
                        if a.get("bs_head") == "employee_expenses"
                        and "payable" not in a.get("name","").lower()
                        and abs(a.get("net", 0)) > 0]
        # Also include salary, ESI, bonus, leave from other_expenses ONLY.
        # FIX: this used to also sweep in "direct_expenses"-classified
        # accounts whose name happened to contain "wage"/"labor" — but
        # those accounts (e.g. manufacturing "WAGES A/C") already get
        # written into GROSS PROFIT's own Direct Expenses sub-rows via the
        # dedicated B2 injection block above. Sweeping them in here too
        # wrote the SAME figure a second time into the P&L's Employee
        # Benefits note (and, since "Salaries" was usually the first open
        # slot, often landed there and bumped the real Salary account down
        # into "Bonus") — duplicating the amount and corrupting both
        # sections' labelling.
        for a in individual_accounts:
            name_l = a.get("name","").lower()
            if a.get("bs_head") == "other_expenses":
                if any(kw in name_l for kw in ["salary", "wage", "e.s.i", "esi ",
                       "bonus", "leave with", "labour refreshment", "labor",
                       "labour welfare", "labor welfare", "lwf"]):
                    if "payable" not in name_l and abs(a.get("net", 0)) > 0:
                        if a not in emp_accounts:
                            emp_accounts.append(a)

        if emp_accounts:
            # FIX (Bug 1): Sort employee accounts so Salary comes before Bonus
            # in the injection order.  Previously the list was in TB row order,
            # which could put BONUS (row 153) BEFORE Salary (row 178) and cause
            # BONUS to claim the "Salaries" template row while Salary landed
            # in the wrong slot (e.g. "Staff welfare expenses").
            # Priority: salary/salaries first, then wages, then bonus, then rest.
            def _emp_sort_key(acct):
                n = acct.get("name","").lower()
                if "salary" in n or "salari" in n or "salaries" in n:
                    return 0
                if "wage" in n:
                    return 1
                if "bonus" in n:
                    return 3
                if "esi" in n or "e.s.i" in n:
                    return 4
                return 2
            emp_accounts = sorted(emp_accounts, key=_emp_sort_key)

            # FIX (Bug 1): Expand template label synonyms for salary matching.
            # "Salaries" in the template should match TB accounts named
            # "Salary", "SALARY", "BONUS" should match "Bonus", etc.
            # The existing _fuzzy_match_name handles substring/edit distance,
            # but we add explicit canonical mappings to guarantee matches even
            # when template uses plural ("Salaries") vs TB singular ("Salary").
            _EMP_CANONICAL = {
                "salary": ["salary", "salaries", "salari", "remuneration"],
                "salaries": ["salary", "salaries", "salari"],
                "wage": ["wages", "wage", "wages a/c", "wages account"],
                "wages": ["wages", "wage", "wages a/c"],
                "bonus": ["bonus"],
                "gratuity": ["gratuity"],
                "esi": ["esi", "e.s.i", "e.s.i.", "esi charges", "esic"],
                "e.s.i": ["esi", "e.s.i", "esic", "e.s.i charges", "e.s.i. charges"],
                "leave": ["leave", "leave with wages", "leave salary", "lwf",
                          "labour welfare", "labor welfare", "leave with"],
                "labour welfare": ["labour welfare fund", "labor welfare fund", "lwf"],
                "staff welfare": ["staff welfare", "staff welfare expenses"],
                "epf": ["epf", "pf contribution", "provident fund"],
                "pf": ["epf", "pf contribution", "provident fund"],
            }

            emp_template = {}
            for r in range(emp_start, emp_end):
                b = ws_npl.cell(r, 2).value
                if b and isinstance(b, str) and len(b.strip()) > 2:
                    emp_template[r] = b.strip().lower()
            
            written_emp = set()
            for acct in emp_accounts:
                matched = False
                acct_name_l = acct["name"].lower()
                for r, lbl in emp_template.items():
                    if r in written_emp: continue
                    # FIX (Bug 1): Use canonical synonym matching BEFORE
                    # generic fuzzy match.  This ensures "Salary" account
                    # always matches a template row labelled "Salaries"
                    # (or vice-versa), "Bonus" matches "Bonus", etc.,
                    # regardless of plural/case variations.
                    canonical_hit = False
                    for canon_tmpl, acct_synonyms in _EMP_CANONICAL.items():
                        if canon_tmpl in lbl:
                            if any(syn in acct_name_l for syn in acct_synonyms):
                                canonical_hit = True
                                break
                    if canonical_hit or _fuzzy_match_name(acct["name"], lbl):
                        if _safe_set(ws_npl, r, 4, abs(acct["net"])):
                            written_emp.add(r)
                            injected.append(f"notes to p&l!D{r} (Employee: {acct['name']}) = {abs(acct['net']):,.2f}")
                            matched = True
                        break
                if not matched:
                    # FIX: before creating a new row, check whether this
                    # EXACT amount already sits in another row of this
                    # section (e.g. a contaminated/previously-generated
                    # template that already has "Salaries"=240000 filled
                    # in, while our account is named "SALARY" — a near-
                    # miss the fuzzy matcher doesn't catch). Writing the
                    # same amount again would double it in the section's
                    # SUM-based Total. Skip in that case instead.
                    amt = abs(acct["net"])
                    already_present = False
                    for r in range(emp_start, emp_end):
                        if r in written_emp:
                            continue
                        existing_d = ws_npl.cell(r, 4).value
                        if isinstance(existing_d, (int, float)) and abs(existing_d - amt) < 0.5:
                            already_present = True
                            written_emp.add(r)
                            log.append(
                                f"· Employee expense '{acct['name']}' ({amt:,.2f}) "
                                f"matches existing notes to p&l!D{r} — not duplicated"
                            )
                            break
                    if already_present:
                        continue
                    # Find empty row in section
                    for r in range(emp_start, emp_end):
                        if r not in written_emp:
                            d = ws_npl.cell(r, 4).value
                            if d is None or d == 0:
                                b = ws_npl.cell(r, 2).value
                                if b is None or str(b).strip() == "":
                                    _safe_set(ws_npl, r, 2, acct["name"])
                                _safe_set(ws_npl, r, 4, abs(acct["net"]))
                                written_emp.add(r)
                                injected.append(f"notes to p&l!D{r} (Employee new: {acct['name']}) = {abs(acct['net']):,.2f}")
                                break

        # ── E. Finance Cost → match to template rows dynamically ──
        # Detect Finance cost section
        fin_start = None
        fin_end = None
        for r in range(30, 80):
            b = ws_npl.cell(r, 2).value
            if not b: continue
            bl = str(b).lower()
            if not fin_start and 'finance cost' in bl and 'total' not in bl:
                fin_start = r + 1
            if fin_start and 'total' in bl and 'finance' in bl:
                fin_end = r
                break
        if not fin_start: fin_start = 50
        if not fin_end: fin_end = 56

        # All finance cost accounts (interest on loans, bank charges related to loans)
        FINANCE_KEYWORDS = ["interest", "loan interest", "car loan interest",
                           "machine loan", "top up", "bank cc intt", "bank cc",
                           "bank interest", "cc interest", "bank od interest",
                           "overdraft interest", "interest on unsecured",
            "intt on tempu", "interest on tempu", "interest on hdfc",
            "interest on term loan", "term loan interest",
                           "interest on loan", "interest on term",
                           "interest paid to", "interest to partner"]
        
        finance_accounts = [a for a in individual_accounts
                           if abs(a.get("net", 0)) > 0
                           and any(kw in a.get("name","").lower() for kw in FINANCE_KEYWORDS)
                           and a.get("bs_head") not in ("other_expenses", "other_income", "revenue")]
                           # ^ FIX (Issue 3): "Bank Interest Recd." matches the
                           # "bank interest" keyword in FINANCE_KEYWORDS, but if
                           # it was classified as other_income/revenue (i.e. it's
                           # a CREDIT/income account, not an expense), it must
                           # NOT be treated as a finance cost — otherwise it gets
                           # written into notes to p&l!D38 "Bank interest &
                           # charges" (an EXPENSE row) as a positive expense.
        # Also include accounts explicitly classified as finance_cost
        for a in individual_accounts:
            if a.get("bs_head") == "finance_cost" and abs(a.get("net", 0)) > 0:
                if a not in finance_accounts:
                    finance_accounts.append(a)

        if finance_accounts:
            fin_template = {}
            for r in range(fin_start, fin_end):
                b = ws_npl.cell(r, 2).value
                if b and isinstance(b, str) and len(b.strip()) > 2:
                    fin_template[r] = b.strip().lower()
            
            # FIX (Bug 4): Finance synonym map so "Bank Interest on LAP"
            # matches "Bank Interest on Term Loan", "Interest Charges" matches
            # "Interest Charges", "Interest on Car Loan" matches
            # "Interest on Car loan", etc.
            _FIN_SYNONYMS = {
                "bank interest": ["bank interest", "bank interest on",
                                  "interest on bank", "cc interest", "overdraft"],
                "bank interest on term loan": ["bank interest on term",
                                               "bank interest on lap",
                                               "interest on term loan",
                                               "term loan interest"],
                "bank interest on lap": ["bank interest on lap",
                                         "bank interest on term",
                                         "lap interest", "interest on lap"],
                "interest charges": ["interest charges", "interest on loan",
                                     "bank interest", "bank cc"],
                "interest on car loan": ["interest on car", "car loan interest",
                                         "car loan", "vehicle loan interest"],
                "interest to partner": ["interest to partner", "partner interest",
                                        "interest on capital"],
            }

            written_fin = set()
            for acct in finance_accounts:
                matched = False
                acct_name_l = acct["name"].lower()
                for r, lbl in fin_template.items():
                    if r in written_fin: continue
                    # Try canonical synonym match first
                    _fin_syn_hit = False
                    for canon, syns in _FIN_SYNONYMS.items():
                        if canon in lbl:
                            if any(s in acct_name_l for s in syns):
                                _fin_syn_hit = True
                                break
                    if _fin_syn_hit or _fuzzy_match_name(acct["name"], lbl):
                        if _safe_set(ws_npl, r, 4, abs(acct["net"])):
                            written_fin.add(r)
                            injected.append(f"notes to p&l!D{r} (Finance: {acct['name']}) = {abs(acct['net']):,.2f}")
                            matched = True
                        break
                if not matched:
                    for r in range(fin_start, fin_end):
                        if r not in written_fin:
                            d = ws_npl.cell(r, 4).value
                            if d is None or d == 0:
                                b = ws_npl.cell(r, 2).value
                                if b is None or str(b).strip() == "":
                                    _safe_set(ws_npl, r, 2, acct["name"])
                                _safe_set(ws_npl, r, 4, abs(acct["net"]))
                                written_fin.add(r)
                                injected.append(f"notes to p&l!D{r} (Finance new: {acct['name']}) = {abs(acct['net']):,.2f}")
                                break
                    else:
                        # FIX (Bug 4): no empty row found — accumulate into
                        # last writable finance row rather than silently drop.
                        last_r = fin_end - 1
                        existing_v = ws_npl.cell(last_r, 4).value
                        if existing_v is not None and not _is_formula(existing_v):
                            new_v = float(existing_v or 0) + abs(acct["net"])
                            _safe_set(ws_npl, last_r, 4, new_v)
                            injected.append(
                                f"notes to p&l!D{last_r} (Finance overflow: {acct['name']}) = {new_v:,.2f}"
                            )

        # Combined finance+employee accounts for exclusion from Other Expenses
        finance_acct_names = {a["name"].lower() for a in finance_accounts}
        emp_acct_names = {a["name"].lower() for a in emp_accounts} if emp_accounts else set()

        # ── F. Other Expenses → notes to p&l!D57:D78 ─────────────────
        # Finance-cost keywords to exclude from other_expenses
        # Only EXACT bank/loan interest excluded — NOT "intt paid on late payment of tds"
        BANK_INT_KEYWORDS = ["bank interest", "bank cc intt", "cc interest",
                             "bank od interest", "overdraft interest", "bank charges"]
        LOAN_INT_KEYWORDS = ["loan interest", "car loan interest", "machine loan",
                             "top up", "interest on loan", "interest on term",
                             "interest on unsecured",
            "intt on tempu", "interest on tempu", "interest on hdfc",
            "interest on term loan", "term loan interest", "interest paid to",
                             "interest to partner"]
        FINANCE_KEYWORDS = set(BANK_INT_KEYWORDS + LOAN_INT_KEYWORDS)
        finance_acct_names = {a["name"].lower() for a in finance_accounts}

        # Also exclude salary, depreciation, and employee expenses — handled separately
        EXCLUDE_FROM_OTHER = finance_acct_names | emp_acct_names | {"salary","depreciation","dep on","amort"}

        other_exp_accounts = [a for a in individual_accounts
                              if a.get("bs_head") == "other_expenses"
                              and abs(a.get("net", 0)) > 0
                              and a.get("name","").lower() not in EXCLUDE_FROM_OTHER
                              and a.get("name","").lower() not in finance_acct_names
                              and a.get("name","").lower() not in emp_acct_names
                              and not any(kw in a.get("name","").lower()
                                          for kw in {"salary","depreciation","dep on","amort",
                                                     "salary to partner", "interest to partner"})]

        # FIX (Issue 2): "INTT PAID ON LATE PAYMENT OF TDS" must stay in Other Expenses.
        # Previously these were rerouted to D62 (Bank Charges), which was wrong
        # because interest on late payment of TDS is a statutory penal expense,
        # not a bank charge. We now leave them in `other_exp_accounts` so they
        # get placed into a row of the Other Expenses section (D57:D78), and we
        # also write the account name into column B for that row so the label is visible.
        # (No-op block kept for clarity — see unmatched-placement logic below where
        #  the account name is now written to column B as well.)

        # Build template label → row map for Other Expenses section
        # Different templates place Other Expenses at different rows.
        # Detect the actual section start by searching for "Other Expenses" header
        exp_section_start = 72  # default
        exp_section_end = 98    # default
        npl_cache = _cache.get("notes to p&l", {})
        
        for r in range(50, 100):
            b_val = npl_cache.get((r, 2))
            if b_val and "other expense" in str(b_val).lower():
                exp_section_start = r + 1
                break
        
        # Find section end (next "Total" or blank section)
        for r in range(exp_section_start, min(exp_section_start + 30, 120)):
            b_val = npl_cache.get((r, 2))
            if b_val and "total" in str(b_val).lower() and "other" in str(b_val).lower():
                exp_section_end = r
                break
        
        exp_row_map = {}
        for r in range(exp_section_start, exp_section_end):
            b_val = npl_cache.get((r, 2))
            if b_val:
                exp_row_map[r] = str(b_val).strip().lower()

        # For each TB account, find the best matching template row
        # Strategy: for each template row label, find the TB account whose
        # name best matches — not the other way around.
        # This prevents mismatches from greedy matching.

        # Build reverse: for each TB account → best matching row
        written_exp_rows = set()
        account_row_assignments = {}  # acct_key → row

        for acct in other_exp_accounts:
            name_l = acct["name"].lower().strip()
            best_row = None
            best_score = 0

            for r, lbl in exp_row_map.items():
                if r in written_exp_rows:
                    continue
                score = 0
                # Score based on common significant words
                name_words = set(w for w in name_l.replace("."," ").split() if len(w) > 2)
                lbl_words  = set(w for w in lbl.replace("."," ").split() if len(w) > 2)
                common = name_words & lbl_words
                if common:
                    score = sum(len(w) for w in common)
                # Bonus for exact substring
                if lbl in name_l or name_l in lbl:
                    score += 20
                # Key abbreviation matches
                abbrev_map = {
                    "exp": "expenses",
                    "exp.": "expenses",
                    "adda": "",
                    "advertisment": "advertisement",
                    "advertisement": "advertisement",
                    "stationery": "stationary",
                    "stationary": "stationary",
                }
                for short, full in abbrev_map.items():
                    if short in name_l and (short in lbl or full in lbl):
                        score += 5

                if score > best_score:
                    best_score = score
                    best_row = r

            if best_row and best_score > 2:
                account_row_assignments[acct["name"]] = best_row
                written_exp_rows.add(best_row)

        # Write matched accounts
        for acct in other_exp_accounts:
            amt = abs(acct["net"])
            row = account_row_assignments.get(acct["name"])
            if row:
                d_val = ws_npl.cell(row, 4).value
                if not _is_formula(str(d_val or "")):
                    if _safe_write(ws_npl, row, 4, amt):
                        injected.append(f"notes to p&l!D{row} ({acct['name']}) = {amt:,.2f}")
            else:
                # No template match — find first empty D row in 57:78
                # FIX (Issue 2): also write the account name to column B so the
                # label is visible alongside the amount (otherwise interest on
                # late payment of TDS would appear as an unlabelled amount).
                placed = False
                for r in range(exp_section_start, exp_section_end):
                    if r not in written_exp_rows:
                        d_val = _cache.get("notes to p&l", {}).get((r, 4))
                        b_val = _cache.get("notes to p&l", {}).get((r, 2))
                        if d_val is None or d_val == 0:
                            if _safe_write(ws_npl, r, 4, amt):
                                # Write the account name to col B if the row is unlabelled
                                if b_val is None or str(b_val).strip() == "":
                                    _safe_write(ws_npl, r, 2, acct["name"])
                                written_exp_rows.add(r)
                                injected.append(f"notes to p&l!D{r} (unmatched: {acct['name']}) = {amt:,.2f}")
                                placed = True
                            break
                if not placed:
                    skipped.append(f"Other expense '{acct['name']}' = {amt:,.2f}: no row in notes to p&l")

    # ── Depreciation → Fixed Assets C. Yr. sheet H31 ─────────────────
    # notes to p&l!D52 = 'Fixed Assets C. Yr.'!H31 (formula, auto)
    # We write the TB depreciation total directly to H31 so it flows through.
    dep_total = aggregated_values.get("depreciation", 0)
    if dep_total and "Fixed Assets C. Yr." in wb.sheetnames:
        ws_fa = wb["Fixed Assets C. Yr."]
        fa_cell = ws_fa.cell(31, 8)  # H31
        if fa_cell.value is None or (isinstance(fa_cell.value, (int, float)) and fa_cell.value == 0):
            if _safe_set(ws_fa, 31, 8, dep_total):
                injected.append(f"Fixed Assets C. Yr.!H31 (Depreciation) = {dep_total:,.2f}")
            else:
                skipped.append(f"Depreciation {dep_total:,.2f}: Fixed Assets H31 is formula — not overwritten")
        else:
            log.append(f"· Depreciation H31 already has value {fa_cell.value} — not overwritten")

    # ── Deferred section expansions (row insertions) ─────────────────
    # Process any queued "this section had no spare rows" requests now,
    # at the very end — after every other injection step above has
    # finished reading from `_cache`/`npl_cache_*` dictionaries. Doing
    # this earlier would shift physical row positions while those
    # caches (built once near the top of the function) still hold the
    # OLD positions, causing later steps to write into whatever
    # unrelated content ends up at those now-stale row numbers.
    for _exp in _deferred_section_expansions:
        _sheet_name = _exp["sheet"]
        if _sheet_name not in wb.sheetnames:
            skipped.append(
                f"Deferred section expansion skipped: sheet '{_sheet_name}' not found"
            )
            continue
        _ws_exp = wb[_sheet_name]
        _label_col = _exp["label_col"]
        _value_col = _exp["value_col"]
        _header_row, _total_row = None, None
        for _r in range(1, 60):
            _b = _ws_exp.cell(_r, _label_col).value
            if not _b:
                continue
            _bl = str(_b).strip().lower()
            if _exp["section_header_substr"] in _bl and _exp["total_label"] not in _bl:
                _header_row = _r
            if _bl == _exp["total_label"]:
                _total_row = _r
                break
        if not _total_row:
            skipped.append(
                f"Deferred section expansion skipped: couldn't locate "
                f"'{_exp['total_label']}' row in {_sheet_name} — accounts not placed: "
                + ", ".join(a["name"] for a in _exp["accounts"])
            )
            continue
        # First check if there's already a genuinely blank row between
        # the section header and its TOTAL — if so, reuse it instead of
        # inserting, since insertion is only truly needed when the
        # section is completely full.
        _blank_rows_avail = []
        if _header_row:
            for _r in range(_header_row + 1, _total_row):
                if _ws_exp.cell(_r, _label_col).value is None \
                        and _ws_exp.cell(_r, _value_col).value is None:
                    _blank_rows_avail.append(_r)
        _accounts_remaining = list(_exp["accounts"])
        for _r in _blank_rows_avail:
            if not _accounts_remaining:
                break
            _acct = _accounts_remaining.pop(0)
            _amt = abs(_acct["net"])
            _ws_exp.cell(_r, _label_col).value = _acct["name"]
            _ws_exp.cell(_r, _value_col).value = round(float(_amt), 2)
            log.append(
                f"✓ {_sheet_name}!{chr(64+_value_col)}{_r} ({_acct['name']}) = {_amt:,.2f}"
            )
        if not _accounts_remaining:
            continue
        _n_new = len(_accounts_remaining)
        try:
            _ws_exp.insert_rows(_total_row, amount=_n_new)
            _new_total_row = _total_row + _n_new
            for _i, _acct in enumerate(_accounts_remaining):
                _r = _total_row + _i
                _amt = abs(_acct["net"])
                _ws_exp.cell(_r, _label_col).value = _acct["name"]
                _ws_exp.cell(_r, _value_col).value = round(float(_amt), 2)
                log.append(
                    f"✓ {_sheet_name}!{chr(64+_value_col)}{_r} (section expanded: "
                    f"{_acct['name']}) = {_amt:,.2f}"
                )
            # Repair the (now-shifted) TOTAL formula's own range to
            # include the newly-inserted rows, in both the value column
            # and, if present, the column immediately after it (commonly
            # the previous-year figure in a matching SUM pattern).
            for _col in (_value_col, _value_col + 1):
                _old_f = _ws_exp.cell(_new_total_row, _col).value
                _coll = chr(64 + _col)
                _m = re.match(rf"=SUM\({_coll}(\d+):{_coll}(\d+)\)", str(_old_f or ""))
                if _m:
                    _rng_start = int(_m.group(1))
                    _new_f = f"=SUM({_coll}{_rng_start}:{_coll}{_new_total_row - 1})"
                    _ws_exp.cell(_new_total_row, _col).value = _new_f
                    log.append(
                        f"✓ {_sheet_name}!{_coll}{_new_total_row} formula extended: "
                        f"{_old_f} → {_new_f} (to include the new rows)"
                    )
            # FIX: formulas in OTHER sheets entirely (not just the sheet
            # being expanded) can reference ANY row in this section by
            # absolute position — not just its TOTAL row. insert_rows()
            # only updates cell positions WITHIN the sheet it's called
            # on; it has no way to know about or fix cross-sheet
            # references elsewhere in the workbook, so EVERY formula
            # pointing at row >= the insertion point in this sheet keeps
            # using its OLD row number — silently grabbing whatever
            # unrelated content (a different section's header, a
            # different total, an individual account row) now sits
            # there instead. This was caught happening to AD Garments'
            # "Total Employee benefits expense" reference, not just the
            # revenue total, confirming a single targeted fix isn't
            # enough — any reference into this sheet at or below the
            # insertion row needs shifting by the same amount that
            # insert_rows() shifted the sheet's own internal rows.
            _cross_sheet_fixes = 0
            # FIX: Excel range syntax like "'notes to p&l'!D21:D26" only
            # has the sheet-name prefix on the FIRST cell of the range —
            # the original simpler pattern only matched that first cell,
            # leaving the second (e.g. "D26") completely unshifted and
            # producing a nonsensical inverted range like "D29:D26"
            # (start past end) once the first cell got shifted alone.
            # This pattern explicitly also matches an optional trailing
            # ":Cell2" range continuation and shifts both halves
            # together.
            _cross_ref_pattern = re.compile(
                rf"'?{re.escape(_sheet_name)}'?!(\$?)([A-Z]+)(\$?)(\d+)"
                rf"(:(\$?)([A-Z]+)(\$?)(\d+))?"
            )
            for _other_sn in wb.sheetnames:
                if _other_sn == _sheet_name:
                    continue
                _ws_other = wb[_other_sn]
                for _orow in _ws_other.iter_rows():
                    for _ocell in _orow:
                        _ov = _ocell.value
                        if not (isinstance(_ov, str) and _ov.startswith("=") and _sheet_name in _ov):
                            continue

                        def _shift_match(_mm):
                            _row_num = int(_mm.group(4))
                            _new_row = _row_num + _n_new if _row_num >= _total_row else _row_num
                            if _mm.group(5):  # has a :Cell2 range continuation
                                _row_num2 = int(_mm.group(9))
                                _new_row2 = _row_num2 + _n_new if _row_num2 >= _total_row else _row_num2
                                return (
                                    f"'{_sheet_name}'!{_mm.group(1)}{_mm.group(2)}"
                                    f"{_mm.group(3)}{_new_row}:"
                                    f"{_mm.group(6)}{_mm.group(7)}{_new_row2}"
                                )
                            return (
                                f"'{_sheet_name}'!{_mm.group(1)}{_mm.group(2)}"
                                f"{_mm.group(3)}{_new_row}"
                            )

                        _new_ov = _cross_ref_pattern.sub(_shift_match, _ov)
                        if _new_ov != _ov:
                            _ocell.value = _new_ov
                            _cross_sheet_fixes += 1
                            log.append(
                                f"✓ {_other_sn}!{_ocell.coordinate} formula repaired: "
                                f"{_ov} → {_new_ov} (reference into {_sheet_name} shifted "
                                f"after {_n_new} row(s) inserted there)"
                            )
            if _cross_sheet_fixes == 0:
                log.append(
                    f"· No cross-sheet references found needing adjustment "
                    f"after expanding {_sheet_name} (or none needed fixing)"
                )

            # FIX: formulas WITHIN the same sheet that reference an
            # absolute row number via a SUM(...) range or arithmetic
            # (e.g. "=SUM(D18:D33)-D35", no sheet-name prefix needed
            # since it's an implicit same-sheet reference) have the
            # exact same staleness problem — insert_rows() shifts cell
            # POSITIONS but not formula TEXT, so a formula like this
            # that sat above the insertion point but referenced rows
            # below it keeps the OLD row numbers. This was caught
            # happening to "Total (a+b+c-d)" in notes to p&l, which
            # still said "D18:D33" after the section below it shifted
            # by 8 rows. Bare references (no sheet prefix) are riskier
            # to blanket-shift than cross-sheet ones — a cell could
            # contain "=D5+D10" where D5 is a genuinely unrelated
            # constant reference that was never meant to move — so this
            # only touches cells that are THEMSELVES below the
            # insertion point (meaning they moved along with the data
            # they reference, so their relative-position intent is
            # almost certainly to track the section they're part of)
            # and only matches bare column+row patterns not already
            # preceded by a sheet name/quote (negative lookbehind).
            _same_sheet_fixes = 0
            _bare_ref_pattern = re.compile(r"(?<!['!])(\$?)([A-Z]{1,2})(\$?)(\d+)")
            # FIX: scanning from _new_total_row missed formulas that sit
            # ABOVE the new total but still reference rows that shifted
            # — e.g. "Total (a+b+c-d)" itself lives just above its own
            # section's end, referencing rows further up that also
            # moved. Any cell from the original insertion point (_total_row)
            # downward could hold a stale reference, so the scan must
            # start there, not at the new total row.
            for _srow in _ws_exp.iter_rows(min_row=_total_row):
                for _scell in _srow:
                    if _scell.row == _new_total_row:
                        continue  # already correctly extended above — don't re-shift it
                    _sv = _scell.value
                    if not (isinstance(_sv, str) and _sv.startswith("=")):
                        continue
                    if _sheet_name in _sv or "!" in _sv:
                        continue  # already handled by cross-sheet logic, or has its own sheet ref

                    def _shift_bare(_mm):
                        _row_num = int(_mm.group(4))
                        if _row_num >= _total_row:
                            return f"{_mm.group(1)}{_mm.group(2)}{_mm.group(3)}{_row_num + _n_new}"
                        return _mm.group(0)

                    _new_sv = _bare_ref_pattern.sub(_shift_bare, _sv)
                    if _new_sv != _sv:
                        _scell.value = _new_sv
                        _same_sheet_fixes += 1
                        log.append(
                            f"✓ {_sheet_name}!{_scell.coordinate} formula repaired: "
                            f"{_sv} → {_new_sv} (same-sheet reference shifted after "
                            f"{_n_new} row(s) inserted above it)"
                        )
        except Exception as _exp_exc:
            skipped.append(
                f"Deferred section expansion failed for {_sheet_name} "
                f"({_exp_exc}): " + ", ".join(a["name"] for a in _exp["accounts"])
            )

    for s in skipped:
        log.append(f"⚠ SKIPPED: {s}")
    for inj_msg in injected:
        log.append(f"✓ {inj_msg}")

    wb.save(output_path)
    wb.close()

    # ── FORMULA CORRECTION PASS ──────────────────────────────────────────────
    # For KNOWN templates (Shree Craft, Fashion Adda): call inject_with_config
    # which has the exact config-driven injection. This overrides the heuristic output.
    try:
        import importlib.util as _ilu, os as _os
        _inj_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "bs_inject_config.py")
        if _os.path.exists(_inj_path):
            _spec = _ilu.spec_from_file_location("_bsic_corr", _inj_path)
            _bsic = _ilu.module_from_spec(_spec)
            _spec.loader.exec_module(_bsic)
            # Check if this is a known template
            from openpyxl import load_workbook as _lwb0
            _wb0 = _lwb0(bs_template_path)
            _tk, _cfg = _bsic.get_config(_wb0) if hasattr(_bsic, 'get_config') else (None, None)
            _wb0.close()
            if _tk:
                # Known template — run full config-driven injection (replaces heuristic output)
                _bsic.inject_with_config(
                    tb_path  = tb_path,
                    template = bs_template_path,
                    output   = output_path,
                )
                # Skip the manual fixes below — inject_with_config handled everything
                raise StopIteration("config injection done")
    except StopIteration:
        pass  # clean exit after config injection
    except Exception as _ce:
        pass  # config injection failed, fall through to manual fixes

    # Manual formula fixes for unknown templates (Fashion Adda fallback)
    try:
        from openpyxl import load_workbook as _lwb
        _wb2 = _lwb(output_path)
        _sheets = _wb2.sheetnames

        # notes to p&l fixes
        # FIX: these row-number-specific patches (row 6, row 8, row 20)
        # were written for Fashion Adda's exact layout, but ran
        # unconditionally for ANY template reaching this fallback path —
        # including AD Garments, whose row 20 ALSO happens to hold a
        # legitimate "=SUM(...)" formula (its own "Total other incomes"
        # total) purely by coincidence of row numbering, which got
        # blindly overwritten with Fashion-Adda-specific content
        # ("=SUM('GROSS PROFIT'!B14:B19)") that has nothing to do with
        # that template's actual structure. Now verifies the row LABELS
        # actually match what these specific patches expect before
        # touching anything, so templates with different layouts that
        # happen to share a row number are left alone.
        if "notes to p&l" in _sheets:
            _npl = _wb2["notes to p&l"]
            _r6_label = str(_npl.cell(6, 2).value or "").strip().lower()
            _r20_label = str(_npl.cell(20, 2).value or "").strip().lower()
            # Fashion Adda layout: GROSS PROFIT sales in col E (E11:E16).
            # KD Knitwear / others: sales in col F (F10). Distinguish so
            # this patch never fires on a template whose row-20 label
            # also contains "purchase" but uses a different GP structure.
            _fa_gp_check = False
            if "GROSS PROFIT" in _wb2.sheetnames:
                _gp_chk = _wb2["GROSS PROFIT"]
                _fa_gp_check = any(
                    isinstance(_gp_chk.cell(r, 5).value, (int, float))
                    and not isinstance(_gp_chk.cell(r, 6).value, (int, float))
                    for r in range(9, 20)
                )
            _is_fashion_adda_layout = (
                "revenue" in str(_npl.cell(5, 2).value or "").lower()
                and "sale" in _r6_label
                and ("purchase" in _r20_label or _r20_label == "")
                and _fa_gp_check
            )
            if _is_fashion_adda_layout:
                # Fix 1: Revenue formula — must cover all 6 sale rows (E11:E16)
                if str(_npl.cell(6,4).value).startswith("=SUM"):
                    _npl.cell(6,4).value = "=SUM('GROSS PROFIT'!E11:E16)"
                # Fix 2: Clear stray SALE GST 12% INTERSTATE value written outside revenue box
                if isinstance(_npl.cell(8,4).value, (int, float)):
                    _npl.cell(8,4).value = None
                # Fix 3: Purchase formula — must cover all 5 purchase rows (B14:B19)
                if str(_npl.cell(20,4).value).startswith("=SUM"):
                    _npl.cell(20,4).value = "=SUM('GROSS PROFIT'!B14:B19)"
            # Fix 4: Merge INTT PAID ON LATE TDS (296) into Bank Charges (D62)
            _d62 = _npl.cell(62,4).value
            _d70 = _npl.cell(70,4).value
            if isinstance(_d70, (int, float)) and _d70 == 296:
                if isinstance(_d62, (int, float)):
                    _npl.cell(62,4).value = _d62 + 296
                _npl.cell(70,4).value = None

        # GROSS PROFIT fixes
        if "GROSS PROFIT" in _sheets:
            _gp = _wb2["GROSS PROFIT"]
            # Fix 5: Closing stock back-calc must use E11:E16
            if "E11:E14" in str(_gp.cell(17,5).value):
                _gp.cell(17,5).value = "=B24-SUM(E11:E16)"
            # Fix 6: Gross Profit % must use E11:E16
            if "E11:E14" in str(_gp.cell(22,2).value):
                _gp.cell(22,2).value = "=SUM(E11:E16)*8.5%"
            # Fix 7: Kill any #REF! in old data columns
            # FIX: previously scanned ALL columns 1-12 including label
            # columns (A, D) where a "#REF!" formula might be a
            # cross-sheet LABEL pull (e.g. "='notes to p&l'!#REF!" used
            # as a row's display text) rather than a numeric value —
            # zeroing those destroys the row's identity/label, not just
            # a broken number. Restrict to the columns that are
            # genuininely meant to hold amounts (B, C, E, F) so label
            # cells are left alone even if they contain a broken
            # reference; a broken label is a template problem to flag,
            # not silently paper over with a 0 that could be mistaken
            # for a meaningful figure.
            for _r in range(1, 30):
                for _c in (2, 3, 5, 6):
                    if "#REF!" in str(_wb2["GROSS PROFIT"].cell(_r,_c).value or ""):
                        _wb2["GROSS PROFIT"].cell(_r,_c).value = 0

        _wb2.save(output_path)
        _wb2.close()
    except Exception as _fe:
        pass  # formula fix failed silently — do not break existing flow

    # ── GENERIC BS FORMULA REPAIR PASS ──────────────────────────────
    # Some BS templates have a Note Number in column D but are MISSING
    # the actual cross-reference formula in columns E/F (e.g. Trade
    # Payables row pointing nowhere) — likely a template authoring gap,
    # not something our injection caused. Detect any bs row whose Note
    # No. matches a "notes to bs" or "notes to p&l" section, but whose
    # E/F cells are empty, and auto-link them to that section's Total row.
    try:
        from openpyxl import load_workbook as _lwb3
        _wb3 = _lwb3(output_path)
        if "bs" in _wb3.sheetnames and "notes to bs" in _wb3.sheetnames:
            _bs = _wb3["bs"]
            _nbs = _wb3["notes to bs"]

            # Build a map: note_number -> row in 'notes to bs' where that
            # note's "Total ..." row lives (the row whose label starts
            # with "total" and is the LAST such row before the next note
            # number marker in column A, e.g. "=bs!D16").
            note_total_rows = {}
            current_note = None
            note_start_row = None
            note_subject = None

            def _find_best_total_row(start_r, end_r, subject):
                """Within [start_r, end_r), find the row whose label both
                contains 'total' AND best matches the note's subject text
                (e.g. 'Trade payables' → prefer 'Total Trade payables' over
                a generic 'Total' inside an unrelated MSME disclosure)."""
                candidates = []
                for rr in range(start_r, end_r):
                    b_val = _nbs.cell(rr, 2).value
                    if b_val and "total" in str(b_val).strip().lower():
                        candidates.append((rr, str(b_val).strip().lower()))
                if not candidates:
                    return None
                if subject:
                    subj_words = set(subject.lower().split())
                    scored = []
                    for rr, lbl in candidates:
                        lbl_words = set(lbl.replace("total", "").split())
                        score = len(subj_words & lbl_words)
                        scored.append((score, rr))
                    scored.sort(key=lambda x: (-x[0], x[1]))
                    if scored[0][0] > 0:
                        return scored[0][1]
                # No subject match — use the FIRST total row (most likely
                # to be the section's own total, before any nested
                # disclosure sub-totals further down)
                return candidates[0][0]

            for r in range(1, _nbs.max_row + 1):
                a_val = _nbs.cell(r, 1).value
                if isinstance(a_val, str) and a_val.strip().startswith("=bs!D"):
                    # New note section begins — finalize the previous one
                    if current_note is not None and note_start_row is not None:
                        best = _find_best_total_row(note_start_row, r, note_subject)
                        if best:
                            note_total_rows[current_note] = best
                    # Extract the note number this section corresponds to
                    import re as _re_note
                    m = _re_note.search(r'D(\d+)', a_val)
                    if m:
                        current_note = int(_bs.cell(int(m.group(1)), 4).value) \
                            if isinstance(_bs.cell(int(m.group(1)), 4).value, (int, float)) \
                            else None
                    note_subject = _nbs.cell(r, 2).value  # e.g. "Trade payables"
                    note_start_row = r + 1
            # Finalize the last section
            if current_note is not None and note_start_row is not None:
                best = _find_best_total_row(note_start_row, _nbs.max_row + 1, note_subject)
                if best:
                    note_total_rows[current_note] = best

            # Now scan the bs sheet for rows with a Note No. (col D) but
            # missing E/F formulas, and link them up.
            repaired = []
            for r in range(1, _bs.max_row + 1):
                note_no = _bs.cell(r, 4).value
                if not isinstance(note_no, (int, float)):
                    continue
                e_val = _bs.cell(r, 5).value
                f_val = _bs.cell(r, 6).value
                if e_val is not None or f_val is not None:
                    continue  # already has something — don't touch
                # FIX: skip if the immediately following rows (within 5)
                # already carry E/F formulas that cover the same section.
                # Example: bs row 16 "Trade Payables" has Note 4 and E16=None,
                # but rows 17 (MSME) and 18 (non-MSME) already have E17/E18
                # formulas that appear in the =SUM(E12:E20) total. Writing
                # E16 here would count Trade Payables twice in that sum.
                has_sub_formulas = False
                for sub_r in range(r + 1, min(r + 6, _bs.max_row + 1)):
                    sub_note = _bs.cell(sub_r, 4).value
                    sub_e = _bs.cell(sub_r, 5).value
                    # If a sub-row has a Note number (different section starts)
                    # or a label in col B (next header), stop scanning
                    if isinstance(sub_note, (int, float)) and sub_note != note_no:
                        break
                    if sub_e is not None and isinstance(sub_e, str) and sub_e.startswith('='):
                        has_sub_formulas = True
                        break
                if has_sub_formulas:
                    continue  # sub-rows already handle this — don't write header row
                total_row = note_total_rows.get(int(note_no))
                if total_row:
                    _bs.cell(r, 5).value = f"='notes to bs'!D{total_row}"
                    _bs.cell(r, 6).value = f"='notes to bs'!E{total_row}"
                    repaired.append(
                        f"bs!E{r}/F{r} (Note {int(note_no)}) → "
                        f"'notes to bs'!D{total_row}/E{total_row}"
                    )

            if repaired:
                _wb3.save(output_path)
                log.append(f"✓ Auto-repaired {len(repaired)} missing bs formula(s):")
                for rep in repaired:
                    log.append(f"    {rep}")
        _wb3.close()
    except Exception as _fe2:
        pass  # repair pass failed silently — do not break existing flow

    # ─────────────────────────────────────────────────────────────
    # FIX (Issue 3): On-screen totals must mirror what the downloaded BS shows.
    #
    # The old logic simply summed asset-side vs. liability-side TB aggregates,
    # which always tallies (debits = credits in any TB) — so the UI permanently
    # displayed equal Assets/Liabilities and zero profit, even though the
    # downloaded Excel file (after recomputing formulas in Excel/LibreOffice)
    # shows real numbers with a P&L balance.
    #
    # We now compute the actual BS figures the spreadsheet will resolve to:
    #   1. Net Profit = Revenue − (Opening Stock + Purchases + Employee
    #                  + Finance Cost + Depreciation + Other Expenses) + Closing Stock
    #   2. Closing Stock is the balancing figure in the Trading A/c:
    #      Closing Stock = (Opening Stock + Purchases) − (Sales − Gross Profit)
    #      For display purposes we use the inventory aggregate from TB if present,
    #      otherwise treat closing stock as the Trading A/c balancing figure.
    # ─────────────────────────────────────────────────────────────
    revenue_total       = aggregated_values.get("revenue", 0)
    other_income_total  = aggregated_values.get("other_income", 0)   # FIX Issue 3
    purchases_total     = aggregated_values.get("purchases", 0)
    direct_exp_total    = aggregated_values.get("direct_expenses", 0)
    employee_total      = aggregated_values.get("employee_expenses", 0)
    other_exp_total     = aggregated_values.get("other_expenses", 0)
    depreciation_total  = aggregated_values.get("depreciation", 0)
    inventories_total   = aggregated_values.get("inventories", 0)

    # Opening stock is included inside the `inventories` aggregate in some TBs
    # and as a separate line in others. Use individual_accounts to separate.
    opening_stock = 0.0
    closing_stock_from_tb = 0.0
    if individual_accounts:
        for a in individual_accounts:
            n = a.get("name", "").lower()
            if "opening stock" in n:
                opening_stock += abs(a.get("net", 0))
            elif "closing stock" in n:
                closing_stock_from_tb += abs(a.get("net", 0))

    # If no explicit opening/closing rows, the TB's undifferentiated stock
    # balance must be treated as CLOSING stock, never opening. A trial
    # balance is, by definition, a snapshot as of the balance sheet date
    # (period END) — there is no such thing as "opening stock" sitting in
    # a TB, since opening stock is necessarily LAST year's closing figure
    # and isn't available from this year's TB at all. Treating it as
    # opening stock (the previous behaviour) silently flipped the sign of
    # this figure in the Net Profit calculation — subtracting it as a cost
    # instead of adding it as a closing asset — which alone was enough to
    # swing a genuinely profitable, tallying business into a large fake
    # loss and a large fake Assets/Liabilities mismatch on the website's
    # summary page, even though the actual generated spreadsheet (which
    # computes everything via its own template formulas, not this
    # redundant aggregate) was correct.
    if opening_stock == 0 and closing_stock_from_tb == 0:
        closing_stock_from_tb = inventories_total

    # Closing stock used in the BS = explicit value if TB has one,
    # otherwise the Trading A/c balancing figure.
    # Trading A/c balancing figure rule:
    #   Closing Stock = max(0, (Opening Stock + Purchases) − Sales) when Sales < Cost
    # We instead compute Net Profit using a single equation that does NOT rely
    # on a separately-supplied closing stock.
    if closing_stock_from_tb > 0:
        closing_stock_bs = closing_stock_from_tb
    else:
        # Closing stock acts as the credit-side balancing figure on the
        # Trading A/c. Without an explicit value, we cannot derive it
        # independently — leave at 0 and let Net Profit absorb the difference.
        closing_stock_bs = 0.0

    # Net Profit = Sales + Other Income + Closing Stock
    #              − Opening Stock − Purchases − Employee
    #              − Other Expenses − Depreciation
    # FIX Issue 3: Other Income (e.g. Rebate & Discount) was being silently
    # ignored from the P&L computation, causing the on-screen "Profit" figure
    # to under-report by that amount.
    net_profit = (
        revenue_total + other_income_total + closing_stock_bs
        - opening_stock - purchases_total - direct_exp_total
        - employee_total - other_exp_total - depreciation_total
    )

    # ── Asset side ──
    total_assets = (
        aggregated_values.get("fixed_assets", 0)
        + aggregated_values.get("non_current_investments", 0)
        + closing_stock_bs
        + aggregated_values.get("trade_rec", 0)
        + aggregated_values.get("cash_bank", 0)
        + aggregated_values.get("stla", 0)
        + aggregated_values.get("other_current_assets", 0)
    )

    # ── Liability side (Capital + Profit + Borrowings + Payables + OCL + Provisions) ──
    capital_total = aggregated_values.get("capital", 0)
    total_liabilities = (
        capital_total
        + net_profit
        + aggregated_values.get("lt_borrowings", 0)
        + aggregated_values.get("st_borrowings", 0)
        + aggregated_values.get("trade_payables", 0)
        + aggregated_values.get("other_cl", 0)
        + aggregated_values.get("st_provisions", 0)
    )

    diff = abs(total_assets - total_liabilities)

    log.append(f"\n📊 P&L Summary:")
    log.append(f"  Revenue:           {revenue_total:>15,.2f}")
    log.append(f"  Other Income:      {other_income_total:>15,.2f}")
    log.append(f"  Opening Stock:     {opening_stock:>15,.2f}")
    log.append(f"  Purchases:         {purchases_total:>15,.2f}")
    log.append(f"  Direct Expenses:   {direct_exp_total:>15,.2f}")
    log.append(f"  Employee Exp:      {employee_total:>15,.2f}")
    log.append(f"  Other Expenses:    {other_exp_total:>15,.2f}")
    log.append(f"  Depreciation:      {depreciation_total:>15,.2f}")
    log.append(f"  Closing Stock:     {closing_stock_bs:>15,.2f}")
    log.append(f"  → Net Profit:      {net_profit:>15,.2f}")

    log.append(f"\n📊 Balance Sheet Totals (computed):")
    log.append(f"  Total Assets:      {total_assets:>15,.2f}")
    log.append(f"  Total Liabilities: {total_liabilities:>15,.2f}")
    if diff < 1:
        log.append(f"  ✅ Balance Sheet TALLIES!")
    else:
        log.append(f"  ❌ Difference: {diff:,.2f}")
        if total_assets > total_liabilities:
            log.append(f"  Assets exceed Liabilities by {diff:,.2f}")
        else:
            log.append(f"  Liabilities exceed Assets by {diff:,.2f}")

    return {
        "status": "success",
        "output": output_path,
        "log": log,
        "injected_count": len(injected),
        "skipped_count": len(skipped),
        "tally_ok": diff < 1,
        "total_assets": total_assets,
        "total_liabilities": total_liabilities,
        "net_profit": net_profit,
        "revenue": revenue_total,
        "other_income": other_income_total,           # NEW Issue 3
        "opening_stock": opening_stock,
        "closing_stock": closing_stock_bs,
        "capital": capital_total,
    }


def detect_bs_template(file_path):
    """
    Detect BS template structure — returns info for the UI.
    Primary sheet is always the bs/balance sheet sheet.
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    result = {"sheets": []}

    for sname in wb.sheetnames:
        ws = wb[sname]
        rows_data = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows_data.append(list(row))
            if i >= 100:
                break
        sheet_info = _detect_bs_columns(rows_data, sname)
        if sheet_info:
            result["sheets"].append(sheet_info)

    wb.close()

    if not result["sheets"]:
        return {"error": "Could not detect Balance Sheet template structure."}

    best = max(result["sheets"], key=lambda s: len(s.get("head_rows", {})))
    result["primary"] = best
    return result



# ═══════════════════════════════════════════════════════════════════════
# MAIN PIPELINE (called from app.py routes)
# ═══════════════════════════════════════════════════════════════════════

def analyze_trial_balance(tb_path):
    """
    Step 1: Read and analyze the trial balance.
    Returns structure info + classified accounts.
    Supports both .xlsx and .pdf input files.
    """
    # Handle PDF input — parse and use directly
    if tb_path.lower().endswith('.pdf'):
        detection = parse_tb_pdf(tb_path)
        if not detection or not detection.get("accounts"):
            return {"error": "Could not parse trial balance PDF. Ensure it's a text-based PDF (not scanned image)."}
    else:
        detection = detect_tb_structure(tb_path)
        if "error" in detection:
            return detection

    classified = classify_accounts(detection["accounts"])

    # Pre-flag debit-creditor / credit-debtor as the explicit advance heads
    # so the Review-Mapping UI groups them correctly.
    for acct in classified:
        head = acct.get("bs_head")
        net  = acct.get("net", 0)
        if head == "trade_payables" and net > 0:
            acct["bs_head"] = "advance_to_supplier"
            acct["reclassified_from"] = "trade_payables"
        elif head == "trade_rec" and net < 0:
            acct["bs_head"] = "advance_from_customer"
            acct["reclassified_from"] = "trade_rec"

    # Separate by confidence
    high_conf = [a for a in classified if a["confidence"] == "high"]
    low_conf = [a for a in classified if a["confidence"] == "low"]
    unclassified = [a for a in classified if a["confidence"] == "none"]

    # ── FIX (Issue 1): Build an explicit "Manual Review" list so the UI can
    # show the user WHICH accounts need attention. Previously the front-end
    # only knew the COUNT ("6 Manual") but had no way to drill down. We now
    # return the full account objects, plus a flat list of dropdown options
    # so the UI can render a `<select>` with every valid bs_head label.
    bs_head_options = [
        {"key": k, "label": v["label"], "side": v["side"]}
        for k, v in BS_HEADS.items()
    ]
    # Add the two sign-aware advance heads (they aren't in BS_HEADS as separate
    # buckets — they're virtual heads that aggregate into stla / other_cl)
    bs_head_options.append({"key": "advance_to_supplier",
                            "label": "Advance to Supplier (Cr-side debit)",
                            "side": "asset"})
    bs_head_options.append({"key": "advance_from_customer",
                            "label": "Advance from Customer (Dr-side credit)",
                            "side": "liability"})

    # Manual-review payload — what the UI shows under the "Manual" tab.
    # `dr_cr` is filled so the UI can colour debit vs credit balances and
    # show them at the end of the BS/P&L group they ALMOST belong to.
    manual_review = []
    for a in unclassified:
        manual_review.append({
            "row":      a.get("row"),
            "name":     a.get("name"),
            "group":    a.get("group"),
            "debit":    a.get("debit", 0),
            "credit":   a.get("credit", 0),
            "net":      a.get("net", 0),
            "dr_cr":    "Dr" if a.get("net", 0) >= 0 else "Cr",
            "bs_head":  a.get("bs_head", "unclassified"),
            # Suggest the most likely side so the UI can pre-position the row
            "suggested_side": "asset" if a.get("net", 0) >= 0 else "liability",
        })

    return {
        "status": "success",
        "detection": {
            "format_type": detection["format_type"],
            "sheet_name": detection["sheet_name"],
            "header_row": detection["header_row"],
            "data_start_row": detection["data_start_row"],
            "account_col": detection["account_col"],
            "debit_col": detection["debit_col"],
            "credit_col": detection["credit_col"],
            "net_col": detection["net_col"],
        },
        "accounts": classified,
        "manual_review": manual_review,            # NEW — list of "Manual X" items
        "bs_head_options": bs_head_options,        # NEW — dropdown options for UI
        "summary": {
            "total_accounts":    len(classified),
            "high_confidence":   len(high_conf),
            "low_confidence":    len(low_conf),
            "unclassified":      len(unclassified),
            "manual_count":      len(unclassified),  # alias for UI clarity
        },
    }


def process_tb_to_bs(tb_path, bs_template_path, output_path, user_mapping=None):
    """
    Full pipeline: analyze TB → classify → inject into BS.

    user_mapping: dict of overrides. Supports TWO formats:
      1. {"ACCOUNT NAME": "bs_head"}   — matched by name (case-insensitive)
      2. {"ACCOUNT NAME_rownum": "bs_head"} — matched by unique key

    This is the SINGLE SOURCE OF TRUTH for user overrides.
    The Flask /tb-process route must pass user_mapping here.
    """
    # Step 1: Analyze TB (handles both xlsx and pdf)
    # If PDF, convert to temporary xlsx for the injection step
    import os, tempfile
    actual_tb_path = tb_path
    tmp_xlsx = None
    if tb_path.lower().endswith('.pdf'):
        tmp_xlsx = tempfile.mktemp(suffix='.xlsx')
        if not convert_pdf_tb_to_xlsx(tb_path, tmp_xlsx):
            return {"error": "Could not convert PDF trial balance to Excel."}
        actual_tb_path = tmp_xlsx

    analysis = analyze_trial_balance(tb_path)  # Use original path for detection
    if "error" in analysis:
        if tmp_xlsx and os.path.exists(tmp_xlsx):
            os.remove(tmp_xlsx)
        return analysis

    accounts = analysis["accounts"]

    # Step 2: Apply user overrides — match by unique key first, then by name
    if user_mapping:
        name_map = {}
        key_map  = {}
        for raw_key, head in user_mapping.items():
            if not head or head in ("auto", "", None):
                continue
            rk = str(raw_key).strip()
            # Key format: "ACCOUNT NAME_123"  (ends with underscore + digits)
            parts = rk.rsplit("_", 1)
            if len(parts) == 2 and parts[1].isdigit():
                key_map[rk] = head
            else:
                name_map[rk.upper()] = head

        for acct in accounts:
            acct_key  = str(acct.get("key", "")).strip()
            acct_name = str(acct.get("name", "")).strip().upper()
            if acct_key in key_map:
                acct["bs_head"]    = key_map[acct_key]
                acct["confidence"] = "user"
            elif acct_name in name_map:
                acct["bs_head"]    = name_map[acct_name]
                acct["confidence"] = "user"

    # Step 3: Aggregate by bs_head (uses overridden heads)
    aggregated = get_aggregated_values(accounts)

    # Step 4 + 5: Config-driven injection via bs_inject_config.py
    # Uses template_configs.py exact row maps — correct formulas, no heuristics.
    # Falls back to old heuristic injector (which now has formula correction pass built-in).
    try:
        import importlib.util, os
        _inj_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "bs_inject_config.py")
        if os.path.exists(_inj_path):
            spec = importlib.util.spec_from_file_location("bs_inject_config", _inj_path)
            inj  = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(inj)
            result = inj.inject_with_config(
                tb_path  = tb_path,
                template = bs_template_path,
                output   = output_path,
            )
            result["analysis"]            = analysis
            result["aggregated"]          = aggregated
            result["classified_accounts"] = accounts
        else:
            # bs_inject_config.py not found — use heuristic injector
            # (formula correction pass is now built into inject_into_bs above)
            result = inject_into_bs(
                bs_template_path, output_path, aggregated,
                mapping_overrides=None,
                individual_accounts=accounts,
            )
            result["analysis"]            = analysis
            result["aggregated"]          = aggregated
            result["classified_accounts"] = accounts
    except Exception as _inj_exc:
        # Hard fallback
        result = inject_into_bs(
            bs_template_path, output_path, aggregated,
            mapping_overrides=None,
            individual_accounts=accounts,
        )
        result["post_process_log"]    = [f"Config injector error: {_inj_exc}"]
        result["analysis"]            = analysis
        result["aggregated"]          = aggregated
        result["classified_accounts"] = accounts

    return result
