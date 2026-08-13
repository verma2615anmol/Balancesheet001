"""
BS Annual Updater — Multi-Tool CA Dashboard
Auth + Upload-Based Plans + Admin Panel
"""

import re
import os
import uuid
import json
import hashlib
import secrets
from datetime import datetime, timedelta
from functools import wraps
from flask import (Flask, request, send_file, jsonify,
                   render_template_string, session, redirect, url_for, g)
from processor import detect_fixed_asset_sheet_names
from lumid_compat import process

# ── Database driver selection ────────────────────────────────────────────────
# If DATABASE_URL is set (Supabase/PostgreSQL), use psycopg2.
# Otherwise, fall back to SQLite for local development.
DATABASE_URL = os.environ.get("DATABASE_URL")
USE_POSTGRES = bool(DATABASE_URL)

if USE_POSTGRES:
    import psycopg2
    import psycopg2.extras
else:
    import sqlite3

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024

# ── Global error handlers — ensure /process NEVER returns HTML ────────────
@app.errorhandler(500)
def handle_500(e):
    if request.path == "/process":
        return jsonify({"status": "error", "message": f"Server error: {e}"}), 500
    return "Internal Server Error", 500

@app.errorhandler(Exception)
def handle_exception(e):
    if request.path == "/process":
        return jsonify({"status": "error", "message": f"Server error: {e}"}), 500
    raise e

UPLOAD_DIR = "/tmp/bs_uploads"
OUTPUT_DIR = "/tmp/bs_outputs"
DB_PATH    = os.environ.get("DB_PATH", "users.db")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

FREE_UPLOADS         = 2
UPLOAD_VALIDITY_DAYS = 90

ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "sumit_admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "Admin@Secure123")
CONTACT_EMAIL  = "sumitverma2880@gmail.com"
CONTACT_UPI    = "sumit2615verma@okhdfcbank"

PLANS = {
    "free":     {"label": "Free",         "uploads": 2,   "price": 0},
    "starter":  {"label": "Starter",      "uploads": 10,  "price": 60},
    "standard": {"label": "Standard",     "uploads": 25,  "price": 130},
    "pro":      {"label": "Professional", "uploads": 60,  "price": 270},
    "firm":     {"label": "Firm",         "uploads": 150, "price": 600},
    "ca":       {"label": "CA Firm",      "uploads": 500, "price": 1000},
}

# ══════════════════════════════════════════════════════════════════════════════
#  DATABASE — PostgreSQL (Supabase) with SQLite fallback for local dev
# ══════════════════════════════════════════════════════════════════════════════

def get_db():
    if "db" not in g:
        if USE_POSTGRES:
            g.db = psycopg2.connect(DATABASE_URL)
            g.db.autocommit = False
        else:
            g.db = sqlite3.connect(DB_PATH, detect_types=sqlite3.PARSE_DECLTYPES)
            g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop("db", None)
    if db:
        try: db.close()
        except: pass

def _db_fetchone(sql, params=()):
    db = get_db()
    if USE_POSTGRES:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(sql, params)
        row = cur.fetchone()
        cur.close()
        return row
    else:
        return db.execute(sql, params).fetchone()

def _db_fetchall(sql, params=()):
    db = get_db()
    if USE_POSTGRES:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(sql, params)
        rows = cur.fetchall()
        cur.close()
        return rows
    else:
        return db.execute(sql, params).fetchall()

def _db_execute(sql, params=()):
    db = get_db()
    if USE_POSTGRES:
        cur = db.cursor()
        cur.execute(sql, params)
        cur.close()
        db.commit()
    else:
        db.execute(sql, params)
        db.commit()

def _placeholder(n=1):
    """Return the correct placeholder for the DB driver: %s for PG, ? for SQLite."""
    return "%s" if USE_POSTGRES else "?"

def _ph(sql_with_qmarks):
    """Convert ? placeholders to %s for PostgreSQL."""
    if USE_POSTGRES:
        return sql_with_qmarks.replace("?", "%s")
    return sql_with_qmarks

def init_db():
    if USE_POSTGRES:
        conn = psycopg2.connect(DATABASE_URL)
        cur = conn.cursor()
        cur.execute("""CREATE TABLE IF NOT EXISTS users (
            id            SERIAL PRIMARY KEY,
            username      TEXT    UNIQUE NOT NULL,
            password      TEXT    NOT NULL,
            plan          TEXT    NOT NULL DEFAULT 'free',
            is_admin      INTEGER NOT NULL DEFAULT 0,
            uploads_total INTEGER NOT NULL DEFAULT 2,
            uploads_used  INTEGER NOT NULL DEFAULT 0,
            validity_end  TEXT,
            created_at    TEXT    NOT NULL)""")
        cur.execute("""CREATE TABLE IF NOT EXISTS usage_log (
            id           SERIAL PRIMARY KEY,
            user_id      INTEGER NOT NULL REFERENCES users(id),
            filename     TEXT,
            processed_at TEXT    NOT NULL)""")
        # Insert admin if not exists
        cur.execute("SELECT id FROM users WHERE username=%s", (ADMIN_USERNAME,))
        if not cur.fetchone():
            cur.execute("""INSERT INTO users
                (username,password,plan,is_admin,uploads_total,uploads_used,created_at)
                VALUES (%s,%s,'firm',1,999999,0,%s)""",
                (ADMIN_USERNAME, _hash(ADMIN_PASSWORD), datetime.utcnow().isoformat()))
        conn.commit()
        cur.close()
        conn.close()
    else:
        db = sqlite3.connect(DB_PATH)
        db.execute("""CREATE TABLE IF NOT EXISTS users (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            username      TEXT    UNIQUE NOT NULL,
            password      TEXT    NOT NULL,
            plan          TEXT    NOT NULL DEFAULT 'free',
            is_admin      INTEGER NOT NULL DEFAULT 0,
            uploads_total INTEGER NOT NULL DEFAULT 2,
            uploads_used  INTEGER NOT NULL DEFAULT 0,
            validity_end  TEXT,
            created_at    TEXT    NOT NULL)""")
        db.execute("""CREATE TABLE IF NOT EXISTS usage_log (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id      INTEGER NOT NULL,
            filename     TEXT,
            processed_at TEXT    NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id))""")
        db.execute("""INSERT OR IGNORE INTO users
            (username,password,plan,is_admin,uploads_total,uploads_used,created_at)
            VALUES (?,?,'firm',1,999999,0,?)""",
            (ADMIN_USERNAME, _hash(ADMIN_PASSWORD), datetime.utcnow().isoformat()))
        db.commit()
        db.close()

def _hash(p): return hashlib.sha256(p.encode("utf-8")).hexdigest()
def get_user_by_name(u): return _db_fetchone(_ph("SELECT * FROM users WHERE username=?"), (u,))
def get_user_by_id(i):   return _db_fetchone(_ph("SELECT * FROM users WHERE id=?"), (i,))
def uploads_remaining(user): return max(0, user["uploads_total"] - user["uploads_used"])

def log_usage(user_id, filename):
    _db_execute(_ph("UPDATE users SET uploads_used=uploads_used+1 WHERE id=?"), (user_id,))
    _db_execute(_ph("INSERT INTO usage_log (user_id,filename,processed_at) VALUES (?,?,?)"),
               (user_id, filename, datetime.utcnow().isoformat()))

def add_uploads(user_id, plan_key):
    user    = get_user_by_id(user_id)
    extra   = PLANS[plan_key]["uploads"]
    rem     = uploads_remaining(user)
    new_tot = user["uploads_used"] + rem + extra
    validity = (datetime.utcnow() + timedelta(days=UPLOAD_VALIDITY_DAYS)).isoformat()
    _db_execute(_ph("UPDATE users SET plan=?,uploads_total=?,validity_end=? WHERE id=?"),
               (plan_key, new_tot, validity, user_id))

def create_user(username, password, plan_key):
    uploads  = PLANS[plan_key]["uploads"]
    validity = None if plan_key == "free" else (datetime.utcnow() + timedelta(days=UPLOAD_VALIDITY_DAYS)).isoformat()
    _db_execute(_ph("""INSERT INTO users
        (username,password,plan,is_admin,uploads_total,uploads_used,validity_end,created_at)
        VALUES (?,?,?,0,?,0,?,?)"""),
        (username, _hash(password), plan_key, uploads, validity, datetime.utcnow().isoformat()))

def del_user(uid):
    _db_execute(_ph("DELETE FROM usage_log WHERE user_id=?"), (uid,))
    _db_execute(_ph("DELETE FROM users WHERE id=?"), (uid,))

def all_users(): return _db_fetchall("SELECT * FROM users ORDER BY id")

# ══════════════════════════════════════════════════════════════════════════════
#  AUTH DECORATORS
# ══════════════════════════════════════════════════════════════════════════════

def login_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if "uid" not in session: return redirect(url_for("login_page"))
        return f(*a, **kw)
    return dec

def admin_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if "uid" not in session: return redirect(url_for("login_page"))
        u = get_user_by_id(session["uid"])
        if not u or not u["is_admin"]: return "Access denied.", 403
        return f(*a, **kw)
    return dec

# ══════════════════════════════════════════════════════════════════════════════
#  SHARED CSS
# ══════════════════════════════════════════════════════════════════════════════

BASE_CSS = """
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:ital,wght@0,400;0,500;0,600;0,700;0,800;1,700&family=Inter:wght@400;500;600&display=swap');
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{
  /* ── Option 1: Professional Premium — Teal + Amber ── */
  --brand:#14B8A6;--brand-d:#0F766E;--brand-dk:#0D5C55;--brand-l:#F0FDFB;--brand-m:#99F6E4;
  --accent:#F59E0B;--accent-d:#EFA600;--accent-l:#FFFBEB;--accent-vl:#FEF3C7;
  --purple:#6D28D9;--purple-l:#EDE9FE;
  --green:#059669;--green-l:#ECFDF5;
  --red:#DC2626;--red-l:#FEF2F2;
  --ink:#0F172A;--ink2:#1E3A4A;--muted:#4B6A72;
  /* Subtle green-tinted borders instead of grey */
  --border:#E0F2EE;--border2:#99D6CB;
  /* Very slightly tinted card backgrounds */
  --bg:#F8FAFC;--bg2:#EEF9F7;--card:#FAFCFB;--white:#fff;
  --radius:18px;--radius-sm:11px;
  --shadow:0 1px 4px rgba(15,118,110,.06),0 4px 18px rgba(15,118,110,.08);
  --shadow-md:0 4px 10px rgba(15,118,110,.08),0 14px 44px rgba(15,118,110,.14);
  --shadow-lg:0 8px 24px rgba(15,118,110,.10),0 24px 64px rgba(15,118,110,.18);
  --font-head:'Plus Jakarta Sans',sans-serif;
  --font-body:'Inter',sans-serif;
}
body{font-family:var(--font-body);background:var(--bg);color:var(--ink);min-height:100vh;-webkit-font-smoothing:antialiased}

/* ── NAV ─────────────────────────────────────────────────────────── */
nav{background:rgba(255,255,255,.97);backdrop-filter:blur(14px);
    border-bottom:1px solid var(--border);padding:0 16px;
    display:flex;align-items:center;justify-content:space-between;height:54px;
    position:sticky;top:0;z-index:200;box-shadow:0 1px 0 var(--border)}
@media(min-width:769px){nav{padding:0 28px;height:62px}}
.logo{font-family:var(--font-head);font-size:19px;font-weight:800;
      color:var(--brand-d);letter-spacing:-.6px;text-decoration:none;display:flex;align-items:center;gap:2px}
@media(min-width:769px){.logo{font-size:21px}}
.logo-dot{width:7px;height:7px;background:var(--accent);border-radius:50%;margin-bottom:2px;flex-shrink:0}
.logo span{color:var(--accent)}
.nav-right{display:flex;align-items:center;gap:8px}
@media(min-width:769px){.nav-right{gap:10px}}
.nav-user{font-size:12.5px;color:var(--muted);display:flex;align-items:center;gap:6px}
.nav-user strong{color:var(--ink);font-weight:600}
/* Hide username text + plan badge on mobile — keep only avatar */
@media(max-width:768px){
  .nav-user strong,.nav-user .badge,.nav-sep,.nav-link{display:none!important}
  nav .nav-btn.ghost{display:none!important}
}
.nav-avatar{width:30px;height:30px;background:linear-gradient(135deg,var(--brand),var(--brand-d));border-radius:50%;
            display:inline-flex;align-items:center;justify-content:center;
            font-size:12px;font-weight:700;color:#fff;box-shadow:0 2px 8px rgba(15,118,110,.35)}
.nav-btn{background:linear-gradient(135deg,var(--brand),var(--brand-d));color:#fff;padding:7px 14px;border-radius:var(--radius-sm);
         font-size:12px;font-weight:700;text-decoration:none;letter-spacing:.01em;
         transition:opacity .18s,transform .15s,box-shadow .18s;white-space:nowrap;
         font-family:var(--font-body);display:inline-flex;align-items:center;gap:5px;
         box-shadow:0 2px 12px rgba(15,118,110,.28)}
@media(min-width:769px){.nav-btn{padding:8px 18px;font-size:12.5px}}
.nav-btn:hover{opacity:.93;transform:translateY(-1px);box-shadow:0 5px 20px rgba(15,118,110,.38)}
.nav-btn:active{transform:scale(0.97)!important}
.nav-btn.ghost{background:transparent;color:var(--ink2);border:1.5px solid var(--border2);box-shadow:none}
.nav-btn.ghost:hover{background:var(--bg2);border-color:var(--brand-d);color:var(--brand-d)}
.nav-btn.dash{background:var(--bg2);color:var(--ink2);border:1.5px solid var(--border);font-weight:600;box-shadow:none}
.nav-btn.dash:hover{background:var(--brand-l);color:var(--brand-d);border-color:var(--brand);transform:translateY(-1px)}
.nav-link{font-size:12.5px;color:var(--muted);text-decoration:none;font-weight:500;padding:4px 2px;
          transition:color .18s;position:relative}
.nav-link::after{content:'';position:absolute;bottom:-2px;left:0;width:0;height:1.5px;
  background:var(--brand-d);border-radius:99px;transition:width .22s ease}
.nav-link:hover{color:var(--brand-d)}
.nav-link:hover::after{width:100%}
.nav-sep{width:1px;height:20px;background:var(--border);margin:0 2px}

/* ── BADGES ────────────────────────────────────────────────────────── */
.badge{display:inline-flex;align-items:center;font-size:10px;font-weight:700;
       padding:2px 9px;border-radius:99px;text-transform:uppercase;letter-spacing:.05em}
.b-free{background:#F1F5F9;color:var(--muted)}
.b-starter{background:var(--green-l);color:#065F46}
.b-standard{background:var(--brand-l);color:var(--brand-d)}
.b-pro{background:var(--accent-vl);color:#92400E}
.b-firm{background:var(--purple-l);color:#5B21B6}
.b-ca{background:#FDF2F8;color:#9D174D}

/* ── FOOTER ─────────────────────────────────────────────────────────── */
footer{background:#071812;color:#94A3B8;font-size:12px;padding:0}
.ft-main{display:grid;grid-template-columns:2fr 1fr 1.5fr;gap:44px;
         padding:48px 56px;max-width:1280px;margin:0 auto}
.ft-brand-name{color:#fff;font-family:var(--font-head);font-size:20px;font-weight:800;
               margin-bottom:12px;letter-spacing:-.3px}
.ft-brand-name span{color:var(--accent)}
.ft-brand-desc{font-size:12.5px;line-height:1.85;color:#94A3B8;max-width:340px;margin-bottom:14px}
.ft-col-title{color:#fff;font-size:13px;font-weight:700;margin-bottom:15px;letter-spacing:.03em;
              text-transform:uppercase;font-size:11px}
.ft-links{list-style:none;padding:0;margin:0}
.ft-links li{margin-bottom:13px}
.ft-links a{color:#94A3B8;text-decoration:none;font-size:12.5px;
            transition:color .18s,padding-left .18s}
.ft-links a:hover{color:var(--accent);padding-left:4px}
.ft-contact-name{color:#fff;font-weight:700;font-size:13.5px;margin-bottom:8px}
.ft-contact-addr{color:#94A3B8;font-size:12px;line-height:1.9;margin-bottom:10px}
.ft-contact-line{color:#94A3B8;font-size:12px;margin-bottom:6px}
.ft-contact-line a{color:#6EE7B7;text-decoration:none;transition:color .18s}
.ft-contact-line a:hover{color:var(--accent)}
.ft-socials{display:flex;gap:10px;margin-top:16px}
.ft-socials a{width:36px;height:36px;background:#0D2E24;border-radius:10px;
              display:flex;align-items:center;justify-content:center;
              color:#94A3B8;transition:background .2s,color .2s,transform .2s}
.ft-socials a:hover{background:var(--brand-d);color:#fff;transform:translateY(-2px)}
.ft-socials svg{width:16px;height:16px;fill:currentColor}
.ft-bottom{background:#030D09;border-top:1px solid #0D2E24;
           padding:14px 56px;display:flex;justify-content:space-between;
           align-items:center;flex-wrap:wrap;gap:8px}
.ft-bottom-left{font-size:11px;color:#475569}
.ft-bottom-right{font-size:11px;color:#475569}
@media(max-width:768px){
  .ft-main{grid-template-columns:1fr;padding:32px 24px;gap:28px}
  .ft-bottom{padding:12px 24px;flex-direction:column;text-align:center}
}

/* ── WA FLOAT ─────────────────────────────────────────────────────────── */
.wa-float{position:fixed;bottom:24px;left:24px;width:50px;height:50px;
          background:#25D366;border-radius:50%;display:flex;align-items:center;
          justify-content:center;box-shadow:0 4px 16px rgba(37,211,102,.4);
          z-index:999;text-decoration:none;transition:transform .2s,box-shadow .2s}
.wa-float:hover{transform:scale(1.1);box-shadow:0 6px 24px rgba(37,211,102,.5)}
.wa-float svg{width:22px;height:22px;fill:#fff}

/* ── HELP MODAL ─────────────────────────────────────────────────────────── */
.help-btn{position:fixed;bottom:86px;right:20px;width:42px;height:42px;
          background:linear-gradient(135deg,var(--brand),var(--brand-d));color:#fff;border-radius:50%;
          display:flex;align-items:center;justify-content:center;
          font-size:18px;font-weight:800;cursor:pointer;
          box-shadow:0 4px 14px rgba(15,118,110,.4);z-index:998;border:none;
          transition:transform .2s,box-shadow .2s;text-decoration:none}
.help-btn:hover{transform:scale(1.1);box-shadow:0 6px 20px rgba(15,118,110,.5)}
.help-overlay{display:none;position:fixed;inset:0;background:rgba(7,24,18,.6);
              z-index:1001;align-items:center;justify-content:center;padding:16px}
.help-overlay.open{display:flex}
.help-modal{background:#fff;border-radius:20px;max-width:540px;width:100%;
            max-height:82vh;overflow-y:auto;box-shadow:0 24px 72px rgba(0,0,0,.22)}
.help-modal-head{padding:22px 24px 16px;border-bottom:1px solid var(--border);
                 display:flex;justify-content:space-between;align-items:center}
.help-modal-head h3{font-size:16px;font-weight:800;color:var(--ink);font-family:var(--font-head)}
.help-close{background:none;border:none;font-size:22px;cursor:pointer;color:var(--muted);line-height:1}
.help-modal-body{padding:20px 24px}
.help-step{display:flex;gap:14px;margin-bottom:18px;align-items:flex-start}
.help-step-num{min-width:28px;height:28px;background:linear-gradient(135deg,var(--brand),var(--brand-d));color:#fff;
               border-radius:50%;display:flex;align-items:center;justify-content:center;
               font-size:12px;font-weight:800;flex-shrink:0;margin-top:1px}
.help-step-body h4{font-size:13px;font-weight:700;margin-bottom:3px;color:var(--ink)}
.help-step-body p{font-size:12px;color:var(--muted);line-height:1.6;margin:0}
.help-tip{background:var(--brand-l);border:1px solid var(--brand-m);border-radius:8px;
          padding:10px 14px;font-size:12px;color:var(--brand-d);margin-top:4px;line-height:1.6}

/* ── ANIMATIONS (CSS-only, Render-safe) ─────────────────────────────── */
@keyframes fadeUp{from{opacity:0;transform:translateY(18px)}to{opacity:1;transform:translateY(0)}}
@keyframes fadeIn{from{opacity:0}to{opacity:1}}
@keyframes slideIn{from{opacity:0;transform:translateX(-12px)}to{opacity:1;transform:translateX(0)}}
@keyframes meshShift{0%,100%{background-position:0% 50%}50%{background-position:100% 50%}}
@keyframes pulseRing{0%,100%{opacity:.6;transform:scale(1)}50%{opacity:.25;transform:scale(1.08)}}
@keyframes shimmer{0%{background-position:-200% 0}100%{background-position:200% 0}}
.anim-up{animation:fadeUp .45s cubic-bezier(.22,.68,0,1.2) both}
.anim-in{animation:fadeIn .35s ease-out both}
.anim-d1{animation-delay:.08s}.anim-d2{animation-delay:.16s}.anim-d3{animation-delay:.24s}
.anim-d4{animation-delay:.32s}.anim-d5{animation-delay:.40s}.anim-d6{animation-delay:.48s}
@media(prefers-reduced-motion:reduce){*{animation-duration:.01ms!important;transition-duration:.01ms!important}}

/* ── SHARED PAGE WRAPPER (privacy / story / how-to-use) ─────────────── */
.page-wrap{max-width:820px;margin:48px auto;padding:0 28px 80px}
.page-hero{margin-bottom:36px}
.page-eyebrow{display:inline-flex;align-items:center;gap:6px;
  background:var(--brand-l);color:var(--brand-d);border:1px solid var(--brand-m);
  padding:5px 14px;border-radius:99px;font-size:11px;font-weight:700;
  letter-spacing:.06em;text-transform:uppercase;margin-bottom:12px}
.page-title{font-family:var(--font-head);font-size:clamp(24px,4vw,38px);font-weight:800;
  color:var(--ink);letter-spacing:-.6px;line-height:1.15;margin-bottom:10px}
.page-title em{font-style:italic;color:var(--brand-d)}
.page-sub{font-size:14.5px;color:var(--muted);line-height:1.8;max-width:580px}
.page-date{font-size:11.5px;color:var(--muted);margin-top:6px}
.page-divider{height:1px;background:linear-gradient(90deg,var(--border),transparent);
  margin:32px 0}
.page-section{margin-bottom:36px}
.page-section h2{font-family:var(--font-head);font-size:16px;font-weight:800;
  color:var(--brand-d);margin-bottom:10px;letter-spacing:-.2px}
.page-section p,.page-section li{font-size:13.5px;line-height:1.85;color:#374151}
.page-section ul{padding-left:20px;margin-bottom:12px}
.page-section ul li{margin-bottom:6px}
.page-section .warn{color:#B91C1C;font-weight:600;background:#FEF2F2;
  padding:10px 14px;border-radius:8px;border-left:4px solid #EF4444;
  font-size:13px;line-height:1.7}
.page-section a{color:var(--brand-d);text-decoration:underline}
"""

# ══════════════════════════════════════════════════════════════════════════════
#  LOGIN PAGE
# ══════════════════════════════════════════════════════════════════════════════

PRIVACY_TEMPLATE = """<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Privacy Policy – CA Toolkit</title>
<style>""" + BASE_CSS + """</style></head><body>
<nav>
  <a href="/" class="logo">CA<span class="logo-dot"></span><span>Toolkit</span></a>
  <div class="nav-right"><a href="/" class="nav-btn dash">← Back to Dashboard</a></div>
</nav>
<div class="page-wrap anim-up">
  <div class="page-hero">
    <div class="page-eyebrow">🔒 Legal</div>
    <h1 class="page-title">Privacy <em>Policy</em></h1>
    <p class="page-sub">We built CA Toolkit to save your time — not to sell your data. Here's exactly what we collect and why.</p>
    <p class="page-date">Last updated: June 2026</p>
  </div>
  <div class="page-divider"></div>
  <div class="page-section">
    <h2>1. Data We Collect</h2>
    <p>We collect only the minimum information needed to run the platform: your username for account creation, and uploaded Excel or PDF files solely to process your request. We do not collect your email unless you contact us directly.</p>
  </div>
  <div class="page-section">
    <h2>2. File Handling</h2>
    <ul>
      <li>Uploaded files are processed entirely in memory on our servers — they are never written to permanent storage.</li>
      <li>Files are automatically deleted within minutes of processing.</li>
      <li>We never read, analyse, or share the contents of your financial files with any third party.</li>
    </ul>
  </div>
  <div class="page-section">
    <h2>3. No Ads · No Tracking</h2>
    <p>CA Toolkit does not serve advertisements and does not use third-party tracking or analytics cookies. We do not sell your data — full stop.</p>
  </div>
  <div class="page-section">
    <h2>4. Account Data</h2>
    <p>Your username and plan information are stored securely in our database (Supabase/PostgreSQL). We do not store any payment card details — all payments are handled via UPI or direct bank transfer.</p>
  </div>
  <div class="page-section">
    <h2>5. Account Creation</h2>
    <p>Accounts on CA Toolkit are <strong>created only by the administrator</strong>. There is no self-registration. If you need access, contact us on WhatsApp or email and we will set up your account manually.</p>
  </div>
  <div class="page-section">
    <h2>6. Refund Policy</h2>
    <div class="warn">No refund is issued once the first upload of a paid plan has been used. Unused credits on free plans are non-transferable. Please verify the tool works for your use case using your 2 free uploads before purchasing a plan.</div>
  </div>
  <div class="page-section">
    <h2>7. Contact</h2>
    <p>For any privacy concerns or access requests, reach us at:</p>
    <ul>
      <li>Email: <a href="mailto:sumitverma2880@gmail.com">sumitverma2880@gmail.com</a></li>
      <li>WhatsApp: <a href="https://wa.me/918427651580">+91 84276 51580</a></li>
    </ul>
  </div>
</div>
<footer>
  <div class="ft-bottom" style="justify-content:center;background:#071812;border-top:1px solid #0D2E24;padding:14px 24px">
    <span class="ft-bottom-left">©2026 CA Toolkit · All Rights Reserved · <a href="/privacy" style="color:#6B7280;text-decoration:none">Privacy Policy</a></span>
  </div>
</footer>
<a href="https://wa.me/918427651580" target="_blank" class="wa-float" title="WhatsApp Support"><svg viewBox="0 0 24 24"><path d="M17.472 14.382c-.297-.149-1.758-.867-2.03-.967-.273-.099-.471-.148-.67.15-.197.297-.767.966-.94 1.164-.173.199-.347.223-.644.075-.297-.15-1.255-.463-2.39-1.475-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.298-.347.446-.52.149-.174.198-.298.298-.497.099-.198.05-.371-.025-.52-.075-.149-.669-1.612-.916-2.207-.242-.579-.487-.5-.669-.51-.173-.008-.371-.01-.57-.01-.198 0-.52.074-.792.372-.272.297-1.04 1.016-1.04 2.479 0 1.462 1.065 2.875 1.213 3.074.149.198 2.096 3.2 5.077 4.487.709.306 1.262.489 1.694.625.712.227 1.36.195 1.871.118.571-.085 1.758-.719 2.006-1.413.248-.694.248-1.289.173-1.413-.074-.124-.272-.198-.57-.347m-5.421 7.403h-.004a9.87 9.87 0 01-5.031-1.378l-.361-.214-3.741.982.998-3.648-.235-.374a9.86 9.86 0 01-1.51-5.26c.001-5.45 4.436-9.884 9.888-9.884 2.64 0 5.122 1.03 6.988 2.898a9.825 9.825 0 012.893 6.994c-.003 5.45-4.437 9.884-9.885 9.884m8.413-18.297A11.815 11.815 0 0012.05 0C5.495 0 .16 5.335.157 11.892c0 2.096.547 4.142 1.588 5.945L.057 24l6.305-1.654a11.882 11.882 0 005.683 1.448h.005c6.554 0 11.89-5.335 11.893-11.893a11.821 11.821 0 00-3.48-8.413z"/></svg></a>
</body></html>"""

WA_SVG = """<svg viewBox="0 0 24 24"><path d="M17.472 14.382c-.297-.149-1.758-.867-2.03-.967-.273-.099-.471-.148-.67.15-.197.297-.767.966-.94 1.164-.173.199-.347.223-.644.075-.297-.15-1.255-.463-2.39-1.475-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.298-.347.446-.52.149-.174.198-.298.298-.497.099-.198.05-.371-.025-.52-.075-.149-.669-1.612-.916-2.207-.242-.579-.487-.5-.669-.51-.173-.008-.371-.01-.57-.01-.198 0-.52.074-.792.372-.272.297-1.04 1.016-1.04 2.479 0 1.462 1.065 2.875 1.213 3.074.149.198 2.096 3.2 5.077 4.487.709.306 1.262.489 1.694.625.712.227 1.36.195 1.871.118.571-.085 1.758-.719 2.006-1.413.248-.694.248-1.289.173-1.413-.074-.124-.272-.198-.57-.347m-5.421 7.403h-.004a9.87 9.87 0 01-5.031-1.378l-.361-.214-3.741.982.998-3.648-.235-.374a9.86 9.86 0 01-1.51-5.26c.001-5.45 4.436-9.884 9.888-9.884 2.64 0 5.122 1.03 6.988 2.898a9.825 9.825 0 012.893 6.994c-.003 5.45-4.437 9.884-9.885 9.884m8.413-18.297A11.815 11.815 0 0012.05 0C5.495 0 .16 5.335.157 11.892c0 2.096.547 4.142 1.588 5.945L.057 24l6.305-1.654a11.882 11.882 0 005.683 1.448h.005c6.554 0 11.89-5.335 11.893-11.893a11.821 11.821 0 00-3.48-8.413z"/></svg>"""

_PAGE_NAV = """<nav>
  <a href="/" class="logo">CA<span class="logo-dot"></span><span>Toolkit</span></a>
  <div class="nav-right">
    <a href="/story" class="nav-link">Our Story</a>
    <a href="/how-to-use" class="nav-link">How to Use</a>
    <a href="/ca-tools-hub" class="nav-link">CA Tools Hub</a>
    <a href="/pricing" class="nav-link">Pricing</a>
    <div class="nav-sep"></div>
    <a href="/" class="nav-btn dash">← Dashboard</a>
  </div>
</nav>"""

_PAGE_FOOTER = """<footer>
  <div class="ft-bottom" style="justify-content:center;background:#071812;border-top:1px solid #0D2E24;padding:14px 24px">
    <span class="ft-bottom-left" style="text-align:center;line-height:2">©2026 CA Toolkit · All Rights Reserved ·
      <a href="/privacy" style="color:#6B7280;text-decoration:none">Privacy</a> ·
      <a href="/story" style="color:#6B7280;text-decoration:none">Our Story</a> ·
      <a href="/how-to-use" style="color:#6B7280;text-decoration:none">How to Use</a> ·
      <a href="/ca-tools-hub" style="color:#6B7280;text-decoration:none">CA Tools Hub</a> ·
      <a href="/pricing" style="color:#6B7280;text-decoration:none">Pricing</a>
    </span>
  </div>
</footer>
<a href="https://wa.me/918427651580" target="_blank" class="wa-float" title="WhatsApp Support">""" + WA_SVG + """</a>
<button id="pageScrollTop" onclick="window.scrollTo({top:0,behavior:'scroll'})" title="Back to top" aria-label="Scroll to top" style="position:fixed;bottom:82px;right:24px;width:40px;height:40px;background:var(--white);border:1.5px solid var(--border2);border-radius:50%;display:flex;align-items:center;justify-content:center;cursor:pointer;font-size:17px;color:var(--brand-d);box-shadow:var(--shadow-md);opacity:0;transform:translateY(10px);transition:opacity .26s,transform .26s;z-index:997;pointer-events:none">↑</button>
<script>(function(){var b=document.getElementById('pageScrollTop');if(!b)return;window.addEventListener('scroll',function(){var s=window.scrollY>320;b.style.opacity=s?'1':'0';b.style.transform=s?'translateY(0)':' translateY(10px)';b.style.pointerEvents=s?'auto':'none';},{passive:true});})();</script>"""

# ══════════════════════════════════════════════════════════════════════════════
#  STORY PAGE  — /story
# ══════════════════════════════════════════════════════════════════════════════

STORY_TEMPLATE = """<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Our Story – CA Toolkit</title>
<style>""" + BASE_CSS + """
/* Story-specific extras */
.story-grid{display:grid;grid-template-columns:1fr 1fr;gap:48px;align-items:start;margin-bottom:48px}
@media(max-width:768px){.story-grid{grid-template-columns:1fr;gap:28px}}
.story-quote-block{background:linear-gradient(135deg,var(--brand-l),#E0F7F4);
  border-left:4px solid var(--brand-d);border-radius:0 14px 14px 0;
  padding:20px 24px;font-size:14.5px;font-style:italic;color:var(--ink2);line-height:1.8;margin:20px 0}
.story-quote-block cite{display:block;margin-top:10px;font-size:12px;font-style:normal;
  color:var(--muted);font-weight:700}
.stat-grid{display:grid;grid-template-columns:1fr 1fr;gap:14px}
.stat-tile{background:var(--card);border:1.5px solid var(--border);border-radius:14px;
  padding:20px 16px;text-align:center;transition:border-color .2s,box-shadow .2s,transform .2s}
.stat-tile:hover{border-color:var(--brand);box-shadow:var(--shadow-md);transform:translateY(-3px)}
.stat-tile .si{font-size:28px;margin-bottom:8px}
.stat-tile .sn{font-family:var(--font-head);font-size:24px;font-weight:800;color:var(--brand-d);margin-bottom:3px}
.stat-tile .sl{font-size:11.5px;color:var(--muted);line-height:1.4}
.timeline{list-style:none;padding:0;border-left:2px solid var(--border);margin-left:12px}
.tl-item{position:relative;padding:0 0 28px 28px}
.tl-item:last-child{padding-bottom:0}
.tl-dot{position:absolute;left:-9px;top:4px;width:16px;height:16px;
  background:linear-gradient(135deg,var(--brand),var(--brand-d));border-radius:50%;
  border:2px solid white;box-shadow:0 0 0 2px var(--brand-m)}
.tl-label{font-size:10.5px;font-weight:700;color:var(--accent-d);text-transform:uppercase;
  letter-spacing:.06em;margin-bottom:4px}
.tl-text{font-size:13px;color:var(--ink2);line-height:1.7}
</style></head><body>
""" + _PAGE_NAV + """
<div class="page-wrap anim-up">
  <div class="page-hero">
    <div class="page-eyebrow">✦ Our Story</div>
    <h1 class="page-title">Built out of frustration — <em>and necessity</em></h1>
    <p class="page-sub">Why a CA Article from Ludhiana spent weekends building automation tools that every CA firm in India needed.</p>
  </div>
  <div class="page-divider"></div>

  <div class="story-grid">
    <div>
      <div class="page-section">
        <h2>The Problem</h2>
        <p>Every financial year-end, CA firms across India face the same grind. A comparative balance sheet needs to be "rolled over" — current year figures become previous year, CY cells are cleared, and every single date reference in the file gets updated. For a typical CA firm handling 50–100 clients, this means days of repetitive, error-prone manual work. One wrong cell, one missed date, and the entire BS is wrong.</p>
        <p>Then there's the Trial Balance. After closing books in Tally, someone has to manually pick every figure and paste it into the CA-format balance sheet. Account by account. For every client. Every year.</p>
      </div>
      <div class="story-quote-block">
        "I spent 3 hours rolling over one client's balance sheet manually — updating dates, copying PY figures, clearing CY cells. Then I looked at my list and saw 47 more clients waiting. That's when I decided to build the tool."
        <cite>— Sumit Verma, Article of G.D. Singla &amp; Co. &amp; Founder of CA Toolkit</cite>
      </div>
      <div class="page-section">
        <h2>The Solution</h2>
        <p>CA Toolkit automates the repetitive, mechanical parts of CA work — the parts that don't require professional judgment, just time and patience. The year-shift tool does in 8 seconds what used to take 2–3 hours. The TB→BS tool does in minutes what used to take half a day.</p>
        <p>Every tool is built from real experience in a CA office, not from theory. If you've sat through a year-end close in a CA firm, you know exactly why these tools exist.</p>
      </div>
    </div>
    <div>
      <div class="stat-grid" style="margin-bottom:28px">
        <div class="stat-tile anim-up anim-d1"><div class="si">⏱️</div><div class="sn">3 hrs</div><div class="sl">saved per client on year-shift alone</div></div>
        <div class="stat-tile anim-up anim-d2"><div class="si">📂</div><div class="sn">∞</div><div class="sl">CA templates supported — any format</div></div>
        <div class="stat-tile anim-up anim-d3"><div class="si">✅</div><div class="sn">100%</div><div class="sl">formatting &amp; formulas preserved</div></div>
        <div class="stat-tile anim-up anim-d4"><div class="si">⚡</div><div class="sn">&lt;10s</div><div class="sl">processing time per file</div></div>
      </div>
      <div class="page-section">
        <h2>How It Grew</h2>
        <ul class="timeline">
          <li class="tl-item"><div class="tl-dot"></div><div class="tl-label">March 2024</div><div class="tl-text">First version of the year-shift tool — built over a weekend, used internally for 50 clients.</div></li>
          <li class="tl-item"><div class="tl-dot"></div><div class="tl-label">July 2024</div><div class="tl-text">TB→Balance Sheet tool started — solving the manual figure-picking problem after every Tally close.</div></li>
          <li class="tl-item"><div class="tl-dot"></div><div class="tl-label">Jan 2025</div><div class="tl-text">GST Reconciliation tool added — Books vs GSTR 3B, month-wise and state-wise, in seconds.</div></li>
          <li class="tl-item"><div class="tl-dot"></div><div class="tl-label">2025–26</div><div class="tl-text">Free tools added — Tax Calculator, TDS/TCS, Depreciation, MSME, Capital Gains. CA Toolkit goes public.</div></li>
          <li class="tl-item"><div class="tl-dot"></div><div class="tl-label">Now</div><div class="tl-text">9+ tools live, more coming. Every tool built from real CA office experience.</div></li>
        </ul>
      </div>
    </div>
  </div>

  <div class="page-divider"></div>
  <div class="page-section">
    <h2>Who Built This</h2>
    <p>CA Toolkit is built and maintained by <strong>Sumit Verma</strong>, a CA Article from Ludhiana, Punjab. Every tool on this platform comes from a real problem encountered while working in a CA firm — not from a product roadmap. If you have a repetitive CA task that you think could be automated, reach out.</p>
    <ul>
      <li>Email: <a href="mailto:sumitverma2880@gmail.com">sumitverma2880@gmail.com</a></li>
      <li>WhatsApp: <a href="https://wa.me/918427651580">+91 84276 51580</a></li>
    </ul>
  </div>
</div>
""" + _PAGE_FOOTER + """
</body></html>"""

# ══════════════════════════════════════════════════════════════════════════════
#  HOW TO USE PAGE  — /how-to-use
# ══════════════════════════════════════════════════════════════════════════════

HOW_TO_USE_TEMPLATE = """<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>How to Use – CA Toolkit</title>
<style>""" + BASE_CSS + """
/* How-to-use page extras */
.tool-guide{background:var(--card);border:1.5px solid var(--border);border-radius:var(--radius);
  padding:28px 28px 24px;margin-bottom:28px;position:relative;overflow:hidden;
  transition:border-color .22s,box-shadow .22s,transform .2s}
.tool-guide:hover{border-color:var(--brand);box-shadow:var(--shadow-md);transform:translateY(-3px)}
.tool-guide::before{content:'';position:absolute;top:0;left:0;right:0;height:3.5px;
  background:var(--tg-grad,linear-gradient(90deg,var(--brand),var(--brand-d)))}
.tg-header{display:flex;align-items:center;gap:14px;margin-bottom:16px}
.tg-icon{font-size:28px;width:52px;height:52px;border-radius:13px;
  display:flex;align-items:center;justify-content:center;flex-shrink:0;
  box-shadow:0 2px 8px rgba(0,0,0,.08)}
.tg-tag{display:inline-flex;align-items:center;font-size:10px;font-weight:700;
  padding:3px 10px;border-radius:99px;margin-bottom:6px;letter-spacing:.04em}
.tg-title{font-family:var(--font-head);font-size:17px;font-weight:800;color:var(--ink);margin-bottom:4px}
.tg-sub{font-size:12.5px;color:var(--muted);line-height:1.6}
.tg-body{font-size:13px;color:var(--ink2);line-height:1.8;margin-bottom:16px}
.tg-body strong{color:var(--ink);font-weight:600}
.tg-steps{list-style:none;padding:0;counter-reset:step;margin-bottom:18px}
.tg-steps li{display:flex;gap:12px;align-items:flex-start;padding:10px 0;
  border-bottom:1px solid var(--border);font-size:13px;color:var(--ink2);line-height:1.6}
.tg-steps li:last-child{border:none;padding-bottom:0}
.tg-steps li::before{counter-increment:step;content:counter(step);
  min-width:24px;height:24px;background:linear-gradient(135deg,var(--brand),var(--brand-d));
  color:#fff;border-radius:50%;display:flex;align-items:center;justify-content:center;
  font-size:11px;font-weight:800;flex-shrink:0;margin-top:1px}
.tg-tip{background:var(--accent-vl);border:1px solid #FDE68A;border-radius:10px;
  padding:12px 16px;font-size:12.5px;color:#78350F;line-height:1.7;margin-top:6px}
.tg-link{display:inline-flex;align-items:center;gap:6px;margin-top:14px;
  font-size:13px;font-weight:700;color:var(--brand-d);text-decoration:none;
  padding:9px 20px;border-radius:99px;border:1.5px solid var(--brand-m);
  background:var(--brand-l);transition:background .18s,border-color .18s,transform .18s}
.tg-link:hover{background:var(--brand-m);border-color:var(--brand);transform:translateX(3px)}
.universal-steps{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:36px}
@media(max-width:768px){.universal-steps{grid-template-columns:repeat(2,1fr)}}
@media(max-width:420px){.universal-steps{grid-template-columns:1fr}}
/* Light clean workflow cards — not heavy dark green */
.us-card{background:var(--white);border:1.5px solid var(--border);border-radius:16px;
  padding:22px 18px;position:relative;overflow:hidden;
  transition:border-color .22s,box-shadow .22s,transform .2s}
.us-card:hover{border-color:var(--brand);box-shadow:var(--shadow-md);transform:translateY(-3px)}
.us-card::before{content:'';position:absolute;top:0;left:0;right:0;height:3px;
  background:linear-gradient(90deg,var(--brand),var(--accent))}
.us-num{width:34px;height:34px;background:linear-gradient(135deg,var(--accent),var(--accent-d));
  border-radius:11px;display:flex;align-items:center;justify-content:center;
  font-size:14px;font-weight:800;color:#fff;margin-bottom:14px;
  box-shadow:0 3px 10px rgba(245,158,11,.35)}
.us-title{font-size:13.5px;font-weight:800;color:var(--ink);margin-bottom:6px;
  font-family:var(--font-head)}
.us-desc{font-size:12px;color:var(--muted);line-height:1.65}
</style></head><body>
""" + _PAGE_NAV + """
<div class="page-wrap anim-up">
  <div class="page-hero">
    <div class="page-eyebrow">📖 Guide</div>
    <h1 class="page-title">How to use <em>CA Toolkit</em></h1>
    <p class="page-sub">Step-by-step instructions for every tool. No installation, no learning curve — just upload and download.</p>
  </div>
  <div class="page-divider"></div>

  <!-- Universal flow -->
  <div class="page-section">
    <h2>General Workflow — same for every tool</h2>
    <p style="margin-bottom:20px">Every tool on CA Toolkit follows this 4-step pattern:</p>
    <div class="universal-steps">
      <div class="us-card anim-up anim-d1"><div class="us-num">1</div><div class="us-title">Get access</div><div class="us-desc">Accounts are created by the administrator only. Contact us on WhatsApp or email to get your account set up.</div></div>
      <div class="us-card anim-up anim-d2"><div class="us-num">2</div><div class="us-title">Upload your file</div><div class="us-desc">Drag and drop your Excel or PDF file. Processed entirely in memory — never stored permanently.</div></div>
      <div class="us-card anim-up anim-d3"><div class="us-num">3</div><div class="us-title">Review &amp; confirm</div><div class="us-desc">Some tools (like TB→BS) show you an intermediate step to verify auto-mappings before generating.</div></div>
      <div class="us-card anim-up anim-d4"><div class="us-num">4</div><div class="us-title">Download instantly</div><div class="us-desc">Your processed file downloads in seconds. All formatting, formulas, and print layout remain intact.</div></div>
    </div>
  </div>

  <div class="page-divider"></div>
  <div class="page-section"><h2>Premium Tool Guides</h2></div>

  <!-- Tool 1: Year Shift -->
  <div class="tool-guide anim-up anim-d1" style="--tg-grad:linear-gradient(90deg,#14B8A6,#0F766E,#F59E0B)">
    <div class="tg-header">
      <div class="tg-icon" style="background:linear-gradient(135deg,#E0F7F4,#CCFBF1)">📊</div>
      <div>
        <div class="tg-tag" style="background:#F0FDFB;color:#0F766E;border:1px solid #99F6E4">⭐ Premium · Balance Sheet Year-Shift</div>
        <div class="tg-title">Roll over a comparative Balance Sheet</div>
        <div class="tg-sub">Shifts CY→PY, clears CY, updates all dates — in under 10 seconds</div>
      </div>
    </div>
    <div class="tg-body">
      Every comparative balance sheet needs a yearly "roll-over" — the current year (CY) column becomes the previous year (PY) column, CY cells are cleared for new data, and every date reference in the file is updated to the new financial year. This tool handles all of that automatically, without changing a single formula or cell format.
      <br/><br/><strong>What it handles:</strong> CY→PY column shift, CY cell clearing, all date text formats ("31st March 2024", "31.03.2024", "Year ended March 31, 2024"), formula preservation, and external link removal.
    </div>
    <ul class="tg-steps">
      <li>Sign in to your account and go to <strong>Balance Sheet Year-Shift</strong></li>
      <li>Upload your existing comparative balance sheet (<strong>.xlsx format</strong>)</li>
      <li>Enter the <strong>closing financial year</strong> (e.g. 2025 — the year you are rolling over from)</li>
      <li>Enter the <strong>new financial year</strong> (e.g. 2026 — the year you are rolling into)</li>
      <li>Click <strong>Process</strong> — wait 5–10 seconds</li>
      <li>Download the output file — open in Excel and verify: CY column is blank, PY has old CY values, all dates are updated ✓</li>
    </ul>
    <div class="tg-tip">💡 <strong>Tip:</strong> The tool preserves all formulas and does not remove any formatting. If a cell had a formula referencing another sheet, it will still have that formula — just with the old CY value cleared.</div>
    <a href="/tool/converter" class="tg-link">Open Year-Shift Tool →</a>
  </div>

  <!-- Tool 2: TB to BS -->
  <div class="tool-guide anim-up anim-d2" style="--tg-grad:linear-gradient(90deg,#F59E0B,#D97706)">
    <div class="tg-header">
      <div class="tg-icon" style="background:linear-gradient(135deg,#FFFBEB,#FEF3C7)">🗂️</div>
      <div>
        <div class="tg-tag" style="background:#FFFBEB;color:#92400E;border:1px solid #FDE68A">⭐ Premium · Trial Balance → Balance Sheet</div>
        <div class="tg-title">Auto-fill a BS template from a Trial Balance</div>
        <div class="tg-sub">Reads your TB, classifies every account, injects figures into your template</div>
      </div>
    </div>
    <div class="tg-body">
      After closing books in Tally or any accounting software, you have a Trial Balance. Manually picking each figure and filling it into a CA-format balance sheet is slow and error-prone. This tool reads your TB, auto-classifies every account into the correct BS/P&amp;L head (capital, borrowings, trade payables, fixed assets, etc.), and injects the aggregated values into your template — <strong>without changing a single cell's formatting or formula</strong>.
    </div>
    <ul class="tg-steps">
      <li>Sign in and go to <strong>Balance Sheet from Trial Balance</strong></li>
      <li>Upload your <strong>Trial Balance</strong> (.xlsx — export from Tally or any software). Must have account name + Dr/Cr balance columns</li>
      <li>Upload your <strong>BS Template</strong> (.xlsx — your existing CA-format balance sheet with CY column ready to fill)</li>
      <li>Enter <strong>client name</strong> and <strong>financial year</strong></li>
      <li>Click <strong>Analyse</strong> — the tool auto-maps every account to a BS head. Review the mappings and <strong>override any that are wrong</strong></li>
      <li>Enter <strong>fixed asset additions, sales, and depreciation</strong> for the year if prompted</li>
      <li>Enter <strong>capital account movements</strong> (drawings, fresh capital) if applicable</li>
      <li>Click <strong>Generate Balance Sheet</strong> — download the filled template ✓</li>
    </ul>
    <div class="tg-tip">💡 <strong>Tip:</strong> The tool never changes your template's formatting. Only the CY figures are filled in. If a cell had a formula (like a sum), it will be overwritten with the calculated figure — so review total cells if your template uses live formulas.</div>
    <a href="/tool/tb-to-bs" class="tg-link">Open TB → BS Tool →</a>
  </div>

  <!-- Tool 3: GST Recon -->
  <div class="tool-guide anim-up anim-d3" style="--tg-grad:linear-gradient(90deg,#7C3AED,#6D28D9)">
    <div class="tg-header">
      <div class="tg-icon" style="background:linear-gradient(135deg,#EDE9FE,#DDD6FE)">📋</div>
      <div>
        <div class="tg-tag" style="background:#EDE9FE;color:#5B21B6;border:1px solid #C4B5FD">⭐ Premium · GST Reconciliation</div>
        <div class="tg-title">Books vs GSTR 3B — find the differences</div>
        <div class="tg-sub">Month-wise, state-wise reconciliation report in seconds</div>
      </div>
    </div>
    <div class="tg-body">
      GST reconciliation — comparing sales as per books with GSTR 3B returns — is mandatory for every registered business during audit and ITR filing. Doing it manually across 12 months and multiple states takes hours. This tool parses your GSTR 3B PDFs, compares them with your sales summary, and gives you a complete difference report instantly.
    </div>
    <ul class="tg-steps">
      <li>Sign in and go to <strong>GST Reconciliation</strong></li>
      <li>Download the <strong>Sales Summary Template</strong> from the tool page and fill in your month-wise, state-wise sales figures from books</li>
      <li>Collect all 12 months of <strong>GSTR 3B PDFs</strong> (downloaded from the GST portal) and zip them into a single .zip file</li>
      <li>Upload the <strong>Sales Summary Excel</strong> and the <strong>GSTR 3B ZIP</strong></li>
      <li>Click <strong>Reconcile</strong> — the tool auto-parses each PDF and matches figures month by month</li>
      <li>Download the <strong>Reconciliation Report</strong> — differences are highlighted, state-wise breakdown included ✓</li>
    </ul>
    <div class="tg-tip">💡 <strong>Tip:</strong> Make sure your GSTR 3B PDFs are the official portal downloads (not scanned copies). The tool reads structured PDF text — scanned images won't work.</div>
    <a href="/tool/gst-reconciliation" class="tg-link">Open GST Recon Tool →</a>
  </div>

  <!-- Tool 4: T-Shape BS Converter -->
  <div class="tool-guide anim-up anim-d4" style="--tg-grad:linear-gradient(90deg,#EF4444,#DC2626,#F97316)">
    <div class="tg-header">
      <div class="tg-icon" style="background:linear-gradient(135deg,#FEE2E2,#FECACA)">&#128209;</div>
      <div>
        <div class="tg-tag" style="background:#FEF2F2;color:#991B1B;border:1px solid #FECACA">&#11088; Premium &middot; T-Shape BS Converter</div>
        <div class="tg-title">Convert a T-shaped XLS into a comparative Balance Sheet</div>
        <div class="tg-sub">Extracts every annexure &mdash; fills PY column automatically across all notes</div>
      </div>
    </div>
    <div class="tg-body">
      Some CA firms produce balance sheets in a T-shaped format (.xls) where liabilities and assets sit side by side in the same rows. This tool reads that format forensically &mdash; extracting every annexure (capital account, creditors, debtors, unsecured loans, fixed assets, cash &amp; bank, short-term loans, and more) &mdash; and fills the PY column of the standard comparative output template. The CY column is left blank (yellow-highlighted) for the CA to fill.
    </div>
    <ul class="tg-steps">
      <li>Sign in and go to <strong>T-Shape BS Converter</strong></li>
      <li>Upload the <strong>T-shaped .xls file</strong></li>
      <li>Enter <strong>Closing Year</strong> (e.g. 2024 &mdash; the year of the T-shaped BS) and <strong>New Year</strong> (e.g. 2025)</li>
      <li>Optionally enter the <strong>client name</strong> in the Output Name field</li>
      <li>Click <strong>Process</strong> &mdash; the tool auto-detects the T-shaped format and converts it</li>
      <li>Download the output .xlsx &mdash; open in Excel, verify PY values, then fill the CY column &#10003;</li>
    </ul>
    <div class="tg-tip">&#128161; <strong>Tip:</strong> The tool auto-detects T-shaped .xls files &mdash; no separate mode needed. Just upload the file and it will route automatically.</div>
    <a href="/tool/tshape" class="tg-link">Open T-Shape Converter &#8594;</a>
  </div>

  <div class="page-divider"></div>
  <div class="page-section">
    <h2>Need help or access?</h2>
    <p>Accounts are created by the administrator only — there is no self-registration. To get access or for any help using the tools, reach out directly:</p>
    <ul>
      <li>Email: <a href="mailto:sumitverma2880@gmail.com">sumitverma2880@gmail.com</a></li>
      <li>WhatsApp: <a href="https://wa.me/918427651580">+91 84276 51580</a></li>
    </ul>
  </div>
</div>
""" + _PAGE_FOOTER + """
</body></html>"""

# ══════════════════════════════════════════════════════════════════════════════

LOGIN_T = """<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Login – CA Toolkit</title>
<style>
""" + BASE_CSS + """
body{display:flex;flex-direction:column;align-items:center;justify-content:center;padding:24px}
.auth-card{background:var(--white);border:1px solid var(--border);border-radius:var(--radius);
           box-shadow:var(--shadow);width:100%;max-width:420px;padding:40px}
.auth-logo{font-size:22px;font-weight:800;color:var(--brand);margin-bottom:4px}
.auth-logo span{color:var(--accent)}
.auth-sub{font-size:13px;color:var(--muted);margin-bottom:28px}
h2{font-size:20px;font-weight:700;margin-bottom:4px}
label{display:block;font-size:11px;font-weight:600;text-transform:uppercase;
      letter-spacing:.04em;color:var(--muted);margin-bottom:5px}
.field{margin-bottom:18px}
input{width:100%;border:1.5px solid var(--border);border-radius:8px;padding:10px 14px;
      font-family:inherit;font-size:14px;color:var(--ink);background:var(--white);
      outline:none;transition:border-color .2s}
input:focus{border-color:var(--brand)}
.btn{width:100%;background:var(--brand);color:#fff;border:none;border-radius:8px;
     padding:12px;font-family:inherit;font-size:14px;font-weight:600;cursor:pointer;transition:background .2s}
.btn:hover{background:var(--brand-d)}
.alert{padding:10px 14px;border-radius:8px;font-size:13px;margin-bottom:18px}
.ae{background:#FEF2F2;border:1px solid #FECACA;color:#991B1B}
.lr{text-align:center;margin-top:16px;font-size:13px;color:var(--muted)}
.lr a{color:var(--brand);text-decoration:none;font-weight:500}
</style></head><body>
<div class="auth-card">
  <div class="auth-logo">CA<span>Toolkit</span></div>
  <div class="auth-sub">Professional tools for Indian CAs &amp; Accountants</div>
  <h2>Sign in</h2>
  <p style="font-size:13px;color:var(--muted);margin-bottom:24px">Enter your credentials to continue</p>
  {% if error %}<div class="alert ae">{{ error }}</div>{% endif %}
  <form method="POST" action="/login">
    <div class="field"><label>Username</label>
      <input type="text" name="username" placeholder="Enter username" required autocomplete="username"/></div>
    <div class="field"><label>Password</label>
      <input type="password" name="password" placeholder="Enter password" required autocomplete="current-password"/></div>
    <button class="btn" type="submit">Sign In →</button>
  </form>
  <div class="lr">Need access? <a href="mailto:{{ email }}">Contact admin</a></div>
</div><a href="https://wa.me/918427651580" target="_blank" class="wa-float" title="WhatsApp Support"><svg viewBox="0 0 24 24"><path d="M17.472 14.382c-.297-.149-1.758-.867-2.03-.967-.273-.099-.471-.148-.67.15-.197.297-.767.966-.94 1.164-.173.199-.347.223-.644.075-.297-.15-1.255-.463-2.39-1.475-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.298-.347.446-.52.149-.174.198-.298.298-.497.099-.198.05-.371-.025-.52-.075-.149-.669-1.612-.916-2.207-.242-.579-.487-.5-.669-.51-.173-.008-.371-.01-.57-.01-.198 0-.52.074-.792.372-.272.297-1.04 1.016-1.04 2.479 0 1.462 1.065 2.875 1.213 3.074.149.198 2.096 3.2 5.077 4.487.709.306 1.262.489 1.694.625.712.227 1.36.195 1.871.118.571-.085 1.758-.719 2.006-1.413.248-.694.248-1.289.173-1.413-.074-.124-.272-.198-.57-.347m-5.421 7.403h-.004a9.87 9.87 0 01-5.031-1.378l-.361-.214-3.741.982.998-3.648-.235-.374a9.86 9.86 0 01-1.51-5.26c.001-5.45 4.436-9.884 9.888-9.884 2.64 0 5.122 1.03 6.988 2.898a9.825 9.825 0 012.893 6.994c-.003 5.45-4.437 9.884-9.885 9.884m8.413-18.297A11.815 11.815 0 0012.05 0C5.495 0 .16 5.335.157 11.892c0 2.096.547 4.142 1.588 5.945L.057 24l6.305-1.654a11.882 11.882 0 005.683 1.448h.005c6.554 0 11.89-5.335 11.893-11.893a11.821 11.821 0 00-3.48-8.413z"/></svg></a>
</body></html>"""

# ══════════════════════════════════════════════════════════════════════════════
#  DASHBOARD — tool selection homepage
# ══════════════════════════════════════════════════════════════════════════════

DASHBOARD_T = """<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Dashboard – CA Toolkit</title>
<style>
""" + BASE_CSS + """
/* ── DASHBOARD NAV EXTRAS ───────────────────────────────────────── */
.nav-links-center{display:flex;align-items:center;gap:4px;position:absolute;
  left:50%;transform:translateX(-50%)}
.nav-link-item{font-size:13px;color:var(--muted);text-decoration:none;font-weight:500;
  padding:6px 12px;border-radius:8px;transition:background .18s,color .18s;white-space:nowrap}
.nav-link-item:hover{background:var(--bg2);color:var(--brand-d)}
.nav-link-item.active{color:var(--brand-d);font-weight:600}
@media(max-width:768px){.nav-links-center{display:none}}
/* Mobile hamburger */
.hamburger{display:none;background:none;border:none;cursor:pointer;
  padding:6px;border-radius:8px;color:var(--ink2);transition:background .18s;
  flex-direction:column;justify-content:center;gap:4px}
.hamburger:hover{background:var(--bg2)}
.hamburger span{display:block;width:20px;height:2px;background:currentColor;border-radius:99px;transition:transform .22s,opacity .22s}
@media(max-width:768px){.hamburger{display:flex}}
#mobileMenu{display:none;position:fixed;top:54px;left:0;right:0;
  background:rgba(255,255,255,.98);backdrop-filter:blur(14px);
  border-bottom:1px solid var(--border);z-index:199;
  padding:12px 16px 16px;box-shadow:0 8px 24px rgba(15,118,110,.12)}
#mobileMenu.open{display:block}
#mobileMenu a{display:block;padding:11px 14px;font-size:14px;font-weight:500;
  color:var(--ink2);text-decoration:none;border-radius:10px;
  transition:background .18s,color .18s;margin-bottom:2px}
#mobileMenu a:hover,#mobileMenu a.active{background:var(--brand-l);color:var(--brand-d);font-weight:600}
#mobileMenu .mob-divider{height:1px;background:var(--border);margin:8px 0}
/* Bigger logo */
.logo-lg{font-size:24px!important;letter-spacing:-.8px!important}
/* Amber Sign In button */
.nav-btn-amber{background:linear-gradient(135deg,#F59E0B,#EFA600)!important;
  box-shadow:0 2px 12px rgba(245,158,11,.35)!important}
.nav-btn-amber:hover{box-shadow:0 4px 20px rgba(245,158,11,.5)!important}

/* ── TRUST STRIP ────────────────────────────────────────────────── */
.trust-strip{background:var(--white);border-bottom:1px solid var(--border);
  padding:10px 24px}
.trust-inner{max-width:1320px;margin:0 auto;display:flex;align-items:center;
  justify-content:center;gap:28px;flex-wrap:wrap}
.trust-item{display:inline-flex;align-items:center;gap:6px;
  font-size:12px;color:var(--muted);font-weight:500;white-space:nowrap}
.trust-item .ti-check{width:16px;height:16px;background:linear-gradient(135deg,var(--brand),var(--brand-d));
  border-radius:50%;display:inline-flex;align-items:center;justify-content:center;
  font-size:9px;color:#fff;flex-shrink:0}

/* ── DASHBOARD HERO ─────────────────────────────────────────────── */
/* ── DASHBOARD HERO ─────────────────────────────────────────────── */
@keyframes heroGlow{0%,100%{opacity:.7}50%{opacity:1}}
.dash-hero{
  background:linear-gradient(135deg,#0A5244 0%,#0E7E6F 40%,#12A896 70%,#0A5244 100%);
  padding:28px 16px 30px;text-align:center;position:relative;overflow:hidden}
@media(min-width:769px){.dash-hero{padding:24px 24px 26px}}
.dash-hero::before{content:'';position:absolute;inset:0;
  background:radial-gradient(ellipse 70% 60% at 50% 40%,rgba(245,158,11,.13) 0%,transparent 70%);
  animation:heroGlow 8s ease-in-out infinite;
  pointer-events:none}
.dash-hero::after{content:'';position:absolute;inset:0;
  background:url("data:image/svg+xml,%3Csvg width='32' height='32' viewBox='0 0 32 32' xmlns='http://www.w3.org/2000/svg'%3E%3Crect x='15' y='15' width='2' height='2' rx='1' fill='%23ffffff' fill-opacity='0.04'/%3E%3C/svg%3E");
  pointer-events:none}
.dash-hero-inner{position:relative;max-width:740px;margin:0 auto}
.dash-hero-badge{display:inline-flex;align-items:center;gap:6px;
  background:rgba(245,158,11,.18);backdrop-filter:blur(6px);
  color:#FDE68A;border:1px solid rgba(245,158,11,.35);border-radius:99px;
  padding:5px 14px;font-size:11px;font-weight:700;margin-bottom:10px;letter-spacing:.04em}
@media(min-width:769px){.dash-hero-badge{padding:5px 16px;font-size:11.5px}}
.dash-hero h1{font-family:var(--font-head);font-size:clamp(24px,6vw,42px);
  font-weight:800;line-height:1.12;letter-spacing:-1px;color:#fff;margin-bottom:8px}
.dash-hero h1 em{font-style:italic;color:#A7F3D0}
.dash-hero p{font-size:15px;color:rgba(255,255,255,.82);line-height:1.7;
  max-width:480px;margin:0 auto 18px;letter-spacing:.01em}
@media(min-width:769px){.dash-hero p{font-size:17px;line-height:1.75}}
.hero-cta-row{display:flex;gap:10px;justify-content:center;flex-wrap:wrap;margin-bottom:10px}
.hero-link-row{display:flex;gap:12px;justify-content:center;flex-wrap:wrap}
@media(min-width:769px){.hero-link-row{gap:18px}}
.hero-link{font-size:11.5px;color:rgba(255,255,255,.5);text-decoration:none;
  transition:color .2s;letter-spacing:.01em}
@media(min-width:769px){.hero-link{font-size:12px}}
.hero-link:hover{color:rgba(255,255,255,.88)}
.hero-cta{display:inline-flex;align-items:center;gap:6px;padding:11px 22px;border-radius:99px;
  font-size:13px;font-weight:700;text-decoration:none;transition:transform .18s,box-shadow .18s}
@media(min-width:769px){.hero-cta{padding:10px 24px}}
.hero-cta.primary{background:linear-gradient(135deg,#F59E0B,#EFA600);color:#fff;
  box-shadow:0 4px 22px rgba(245,158,11,.45)}
.hero-cta.primary:hover{transform:translateY(-2px);box-shadow:0 7px 30px rgba(245,158,11,.55)}
.hero-cta.secondary{background:rgba(255,255,255,.13);color:#fff;border:1.5px solid rgba(255,255,255,.32)}
.hero-cta.secondary:hover{background:rgba(255,255,255,.22);transform:translateY(-2px)}

/* ── USAGE STRIP ─────────────────────────────────────────────────── */
.usage-strip{max-width:1080px;margin:-18px auto 0;padding:0 16px;position:relative;z-index:10}
@media(min-width:769px){.usage-strip{padding:0 24px}}
.usage-box{background:var(--white);border:1px solid var(--border);
  border-radius:var(--radius);padding:12px 16px;
  display:flex;align-items:center;justify-content:space-between;
  flex-wrap:nowrap;gap:10px;box-shadow:var(--shadow-md);overflow:hidden}
@media(min-width:769px){.usage-box{padding:16px 22px;gap:14px}}
.usage-left{display:flex;align-items:center;gap:10px;flex:1;min-width:0;overflow:hidden}
@media(min-width:769px){.usage-left{gap:16px}}
.usage-icon{width:36px;height:36px;background:linear-gradient(135deg,var(--brand),var(--brand-d));border-radius:10px;
  display:flex;align-items:center;justify-content:center;font-size:16px;flex-shrink:0;
  box-shadow:0 2px 10px rgba(13,148,136,.3)}
@media(min-width:769px){.usage-icon{width:40px;height:40px;font-size:18px;border-radius:11px}}
.usage-text{min-width:0;overflow:hidden}
.usage-text strong{display:block;font-size:13px;font-weight:700;color:var(--ink);
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
@media(min-width:769px){.usage-text strong{font-size:14px}}
.usage-text span{font-size:11px;color:var(--muted);white-space:nowrap;overflow:hidden;text-overflow:ellipsis;display:block}
@media(min-width:769px){.usage-text span{font-size:12px}}
/* Hide progress bar on mobile — saves space */
.usage-bar-wrap{flex:1;max-width:180px}
@media(max-width:768px){.usage-bar-wrap{display:none}}
.usage-bar-bg{background:var(--bg2);border-radius:99px;height:7px;overflow:hidden}
.usage-bar-fill{height:100%;border-radius:99px}
.usage-validity{font-size:11px;color:var(--muted);margin-top:3px}
.upgrade-btn{display:inline-flex;align-items:center;gap:5px;
  background:linear-gradient(135deg,var(--accent),var(--accent-d));
  color:#fff;padding:8px 14px;border-radius:99px;
  font-size:12px;font-weight:700;text-decoration:none;white-space:nowrap;flex-shrink:0;
  transition:opacity .18s,transform .15s,box-shadow .18s;
  box-shadow:0 3px 14px rgba(245,158,11,.38)}
@media(min-width:769px){.upgrade-btn{padding:9px 22px;font-size:12.5px}}
.upgrade-btn:hover{opacity:.93;transform:translateY(-2px);box-shadow:0 6px 22px rgba(245,158,11,.48)}

/* ── SECTION HEADER ─────────────────────────────────────────────── */
.section-wrap{max-width:1320px;margin:0 auto;padding:40px 24px 0}
.section-hd{display:flex;align-items:center;gap:12px;margin-bottom:22px}
.section-hd h2{font-family:var(--font-head);font-size:18px;font-weight:800;
  color:var(--ink);letter-spacing:-.3px}
.section-hd-badge{display:inline-flex;align-items:center;gap:5px;
  background:var(--brand-l);color:var(--brand-d);
  padding:4px 13px;border-radius:99px;font-size:11px;font-weight:700;
  border:1px solid var(--brand-m)}

/* ── TOOL CARDS ─────────────────────────────────────────────────── */
.tools-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:20px;margin-bottom:48px}
@media(max-width:1100px){.tools-grid{grid-template-columns:repeat(3,1fr)}}
@media(max-width:768px){.tools-grid{grid-template-columns:repeat(2,1fr)}}
@media(max-width:480px){.tools-grid{grid-template-columns:1fr}}
.tools-grid-4{grid-template-columns:repeat(4,1fr)}
@media(max-width:1100px){.tools-grid-4{grid-template-columns:repeat(3,1fr)}}
@media(max-width:768px){.tools-grid-4{grid-template-columns:repeat(2,1fr)}}
@media(max-width:480px){.tools-grid-4{grid-template-columns:1fr}}

.tool-card{background:var(--card);border:1.5px solid var(--border);
  border-radius:var(--radius);padding:24px 22px 20px;
  text-decoration:none;color:var(--ink);
  transition:border-color .22s,box-shadow .25s,transform .22s;
  position:relative;overflow:hidden;display:flex;flex-direction:column}
.tool-card:hover{border-color:var(--brand);
  box-shadow:0 0 0 3px rgba(20,184,166,.13),0 18px 52px rgba(15,118,110,.24);
  transform:translateY(-6px) scale(1.012)}
.tool-card.disabled{cursor:default;opacity:.6}
.tool-card.disabled:hover{border-color:var(--border);box-shadow:none;transform:none}
/* Accent top stripe — brightens on hover */
.tool-card.premium-card::before{content:'';position:absolute;top:0;left:0;right:0;
  height:3.5px;background:linear-gradient(90deg,var(--brand),#0FADA0,var(--accent));
  transition:height .22s,opacity .22s}
.tool-card.premium-card:hover::before{height:4.5px}
.tool-card.free-card::before{content:'';position:absolute;top:0;left:0;right:0;
  height:3.5px;background:linear-gradient(90deg,var(--green),#34D399);
  transition:height .22s}
.tool-card.free-card:hover::before{height:4.5px}
/* Shimmer on hover */
.tool-card:not(.disabled)::after{content:'';position:absolute;inset:0;
  background:linear-gradient(135deg,rgba(15,118,110,.04) 0%,transparent 60%);
  opacity:0;transition:opacity .25s;pointer-events:none}
.tool-card:not(.disabled):hover::after{opacity:1}

.tool-icon{width:48px;height:48px;border-radius:13px;display:flex;
  align-items:center;justify-content:center;font-size:23px;margin-bottom:16px;flex-shrink:0;
  box-shadow:0 2px 8px rgba(0,0,0,.07);transition:transform .22s,box-shadow .22s}
.tool-card:hover .tool-icon{transform:scale(1.08);box-shadow:0 4px 14px rgba(0,0,0,.12)}
.tool-card h2{font-family:var(--font-head);font-size:14.5px;font-weight:700;
  margin-bottom:7px;color:var(--ink);line-height:1.3}
.tool-card p{font-size:12px;color:var(--muted);line-height:1.7;
  margin-bottom:18px;flex:1}
.tool-footer{display:flex;align-items:center;justify-content:space-between;margin-top:auto}
.tool-tag{display:inline-flex;align-items:center;gap:4px;
  font-size:10.5px;font-weight:700;padding:4px 10px;border-radius:99px;letter-spacing:.02em}
.tag-live-prem{background:var(--brand-l);color:var(--brand-d);border:1px solid var(--brand-m)}
.tag-live-free{background:var(--green-l);color:#065F46;border:1px solid #A7F3D0}
.tag-login{background:var(--accent-vl);color:#92400E;border:1px solid #FDE68A}
.tag-soon{background:var(--bg2);color:var(--muted)}
.tool-arrow{width:30px;height:30px;background:var(--bg2);border-radius:9px;
  display:flex;align-items:center;justify-content:center;
  font-size:14px;color:var(--muted);
  transition:background .2s,color .2s,transform .22s,box-shadow .22s}
.tool-card:not(.disabled):hover .tool-arrow{background:var(--brand-d);color:#fff;
  transform:translateX(5px);box-shadow:0 2px 8px rgba(15,118,110,.35)}

/* Corner badge — uniform for all 3 premium cards */
.corner-badge{position:absolute;top:15px;right:15px;
  font-size:9.5px;font-weight:700;padding:3px 9px;border-radius:7px;letter-spacing:.04em}
.cb-prem{background:linear-gradient(135deg,#F0FDFA,#CCFBF1);color:var(--brand-d);border:1px solid var(--brand-m)}
.cb-free{background:var(--green-l);color:#065F46;border:1px solid #A7F3D0}
.cb-lock{background:var(--accent-vl);color:#92400E;border:1px solid #FDE68A}

/* ── FREE TOOLS SECTION ─────────────────────────────────────────── */
.free-section-wrap{background:linear-gradient(180deg,#F0FDF9 0%,#F8FAFC 100%);
  padding:32px 24px 48px;margin-top:8px;border-top:1px solid #C6F0E8;
  border-bottom:1px solid #C6F0E8}
.free-section-inner{max-width:1320px;margin:0 auto}
.section-hd-free{display:flex;align-items:center;gap:12px;margin-bottom:10px}
.section-hd-free h2{font-family:var(--font-head);font-size:20px;font-weight:800;
  color:#065F46;letter-spacing:-.3px}
.section-sub-free{font-size:13px;color:var(--muted);margin-bottom:22px;line-height:1.6}

/* ── COMING SOON CARD ───────────────────────────────────────────── */
.tool-card.coming{border-style:dashed;border-color:#C6F0E8}

/* ── STATS ROW ──────────────────────────────────────────────────── */
.stats-row{background:linear-gradient(135deg,#042F2E 0%,#0D9488 60%,#134E4A 100%);
  padding:28px 24px;margin:0}
.stats-inner{max-width:960px;margin:0 auto;
  display:grid;grid-template-columns:repeat(4,1fr);gap:0;text-align:center}
.stat-item{padding:10px 0;border-right:1px solid rgba(255,255,255,.12)}
.stat-item:last-child{border:none}
.stat-icon{font-size:20px;margin-bottom:6px;display:block}
.stat-n{font-family:var(--font-head);font-size:26px;font-weight:800;color:#fff;margin-bottom:3px;
  animation:countUp .5s ease-out both}
.stat-n em{font-style:normal;color:#6EE7B7}
.stat-l{font-size:11.5px;color:rgba(255,255,255,.6);letter-spacing:.02em}
@media(max-width:640px){.stats-inner{grid-template-columns:repeat(2,1fr)}.stat-item:nth-child(2){border-right:none}.stat-item{border-bottom:1px solid rgba(255,255,255,.1);padding:14px 0}}

/* ── ABOUT / HOW IT WORKS SECTION ───────────────────────────────── */
.about-section{background:var(--white);padding:60px 24px;border-top:1px solid var(--border)}
.about-inner{max-width:1100px;margin:0 auto}
.about-eyebrow{display:inline-flex;align-items:center;gap:6px;
  background:var(--brand-l);color:var(--brand-d);border:1px solid var(--brand-m);
  padding:5px 14px;border-radius:99px;font-size:11px;font-weight:700;
  letter-spacing:.06em;text-transform:uppercase;margin-bottom:14px}
.about-title{font-family:var(--font-head);font-size:clamp(22px,3.5vw,34px);font-weight:800;
  color:var(--ink);letter-spacing:-.5px;line-height:1.2;margin-bottom:10px}
.about-title em{font-style:italic;color:var(--brand)}
.about-sub{font-size:15px;color:var(--muted);line-height:1.75;max-width:620px;margin-bottom:52px}

/* Story strip */
.story-strip{display:grid;grid-template-columns:1fr 1fr;gap:48px;align-items:center;margin-bottom:60px}
@media(max-width:768px){.story-strip{grid-template-columns:1fr;gap:28px}}
.story-text h3{font-family:var(--font-head);font-size:20px;font-weight:800;color:var(--ink);
  margin-bottom:12px;letter-spacing:-.3px}
.story-text p{font-size:13.5px;color:var(--muted);line-height:1.85;margin-bottom:12px}
.story-text p strong{color:var(--ink);font-weight:600}
.story-quote{background:linear-gradient(135deg,var(--brand-l),#E0F2FE);
  border-left:4px solid var(--brand);border-radius:0 12px 12px 0;
  padding:18px 22px;font-size:14px;font-style:italic;color:var(--ink2);line-height:1.7;
  margin-top:16px}
.story-quote cite{display:block;margin-top:8px;font-size:12px;font-style:normal;
  color:var(--muted);font-weight:600}
.story-visual{display:grid;grid-template-columns:1fr 1fr;gap:14px}
.sv-card{background:var(--bg);border:1.5px solid var(--border);border-radius:14px;
  padding:18px 16px;text-align:center}
.sv-card .sv-icon{font-size:28px;margin-bottom:8px}
.sv-card .sv-num{font-family:var(--font-head);font-size:22px;font-weight:800;color:var(--brand);margin-bottom:3px}
.sv-card .sv-lbl{font-size:11px;color:var(--muted);line-height:1.4}

/* Tool deep-dives */
.tool-deep-wrap{display:grid;grid-template-columns:1fr 1fr 1fr;gap:24px;margin-bottom:60px}
@media(max-width:900px){.tool-deep-wrap{grid-template-columns:1fr 1fr}}
@media(max-width:560px){.tool-deep-wrap{grid-template-columns:1fr}}
.tool-deep{background:var(--bg);border:1.5px solid var(--border);border-radius:16px;
  padding:24px 22px;position:relative;overflow:hidden}
.tool-deep::before{content:'';position:absolute;top:0;left:0;right:0;height:3px;
  background:var(--td-color,var(--brand))}
.tool-deep-icon{font-size:30px;margin-bottom:14px}
.tool-deep h3{font-family:var(--font-head);font-size:16px;font-weight:800;color:var(--ink);
  margin-bottom:8px;letter-spacing:-.2px}
.tool-deep-tag{display:inline-flex;align-items:center;font-size:10px;font-weight:700;
  padding:2px 9px;border-radius:99px;margin-bottom:12px;letter-spacing:.04em;
  background:var(--td-bg,var(--brand-l));color:var(--td-c,var(--brand-d))}
.tool-deep p{font-size:12.5px;color:var(--muted);line-height:1.75;margin-bottom:14px}
.tool-deep-steps{list-style:none;padding:0;counter-reset:step}
.tool-deep-steps li{display:flex;gap:9px;align-items:flex-start;
  font-size:12px;color:var(--ink2);margin-bottom:8px;line-height:1.5}
.tool-deep-steps li::before{counter-increment:step;content:counter(step);
  min-width:20px;height:20px;background:var(--td-color,var(--brand));color:#fff;
  border-radius:50%;display:flex;align-items:center;justify-content:center;
  font-size:10px;font-weight:700;flex-shrink:0;margin-top:1px}
.tool-deep-link{display:inline-flex;align-items:center;gap:6px;
  font-size:12.5px;font-weight:700;color:var(--brand-d);text-decoration:none;
  padding:8px 16px;border-radius:99px;border:1.5px solid var(--brand-m);
  background:var(--brand-l);transition:background .18s,border-color .18s}
.tool-deep-link:hover{background:var(--brand-m);border-color:var(--brand)}

/* How to use steps */
.howto-section{background:linear-gradient(135deg,#042F2E,#0D4F47);
  padding:56px 24px;border-radius:20px;margin-bottom:48px;position:relative;overflow:hidden}
.howto-section::before{content:'';position:absolute;inset:0;
  background:radial-gradient(ellipse 60% 50% at 80% 50%,rgba(245,158,11,.12),transparent);
  pointer-events:none}
.howto-inner{max-width:900px;margin:0 auto;position:relative}
.howto-title{font-family:var(--font-head);font-size:clamp(20px,3vw,28px);font-weight:800;
  color:#fff;letter-spacing:-.4px;margin-bottom:8px}
.howto-sub{font-size:13.5px;color:rgba(255,255,255,.65);margin-bottom:36px;line-height:1.7}
.howto-steps{display:grid;grid-template-columns:repeat(4,1fr);gap:20px}
@media(max-width:768px){.howto-steps{grid-template-columns:repeat(2,1fr)}}
@media(max-width:420px){.howto-steps{grid-template-columns:1fr}}
.hs-item{background:rgba(255,255,255,.07);border:1px solid rgba(255,255,255,.12);
  border-radius:14px;padding:20px 16px;position:relative}
.hs-num{width:32px;height:32px;background:linear-gradient(135deg,var(--accent),var(--accent-d));
  border-radius:10px;display:flex;align-items:center;justify-content:center;
  font-size:13px;font-weight:800;color:#fff;margin-bottom:12px;
  box-shadow:0 3px 10px rgba(245,158,11,.4)}
.hs-title{font-size:13px;font-weight:700;color:#fff;margin-bottom:5px}
.hs-desc{font-size:11.5px;color:rgba(255,255,255,.6);line-height:1.6}
</style></head><body>

<!-- ── NAV ─────────────────────────────────────────────────────── -->
<nav style="position:relative">
  <a href="/" class="logo logo-lg">CA<span class="logo-dot"></span><span>Toolkit</span></a>
  <!-- Centre nav links (desktop only) -->
  <div class="nav-links-center">
    <a href="/#premium" class="nav-link-item active">Tools</a>
    <a href="/how-to-use" class="nav-link-item">How to Use</a>
    <a href="/story" class="nav-link-item">About</a>
    <a href="/ca-tools-hub" class="nav-link-item">CA Tools Hub</a>
    <a href="/pricing" class="nav-link-item">Pricing</a>
  </div>
  <div class="nav-right">
    {% if username %}
    <div class="nav-user">
      <span class="nav-avatar">{{ username[0].upper() }}</span>
      <strong>{{ username }}</strong>
      <span class="badge b-{{ plan }}">{{ plan_label }}</span>
      {% if is_admin %}<span class="badge" style="background:#EDE9FE;color:#5B21B6">Admin</span>{% endif %}
    </div>
    <div class="nav-sep"></div>
    {% if is_admin %}<a href="/admin" class="nav-btn ghost">⚙ Admin</a>{% endif %}
    <a href="/logout" class="nav-link">Sign out</a>
    {% else %}
    <a href="/login" class="nav-btn nav-btn-amber">Sign In →</a>
    {% endif %}
    <!-- Hamburger for mobile -->
    <button class="hamburger" id="hamburgerBtn" onclick="toggleMobileMenu()" aria-label="Open menu">
      <span></span><span></span><span></span>
    </button>
  </div>
</nav>
<!-- Mobile nav drawer -->
<div id="mobileMenu">
  <a href="/#premium" class="active">⚡ Tools</a>
  <a href="/how-to-use">📖 How to Use</a>
  <a href="/story">✦ About</a>
  <a href="/ca-tools-hub">🧰 CA Tools Hub</a>
  <a href="/pricing">💳 Pricing</a>
  <div class="mob-divider"></div>
  {% if username %}
  <a href="/logout" style="color:var(--muted)">Sign out ({{ username }})</a>
  {% if is_admin %}<a href="/admin">⚙ Admin Panel</a>{% endif %}
  {% else %}
  <a href="/login" style="color:var(--brand-d);font-weight:700">Sign In →</a>
  {% endif %}
</div>

<!-- ── TRUST STRIP ────────────────────────────────────────────── -->
<div class="trust-strip">
  <div class="trust-inner">
    <span class="trust-item"><span class="ti-check">✓</span> Built for Indian CAs</span>
    <span class="trust-item"><span class="ti-check">✓</span> Excel Compatible</span>
    <span class="trust-item"><span class="ti-check">✓</span> Formatting Always Preserved</span>
    <span class="trust-item"><span class="ti-check">✓</span> 🔒 Files Auto-Deleted After Processing</span>
    <span class="trust-item"><span class="ti-check">✓</span> 🔐 No Data Stored Permanently</span>
    <span class="trust-item"><span class="ti-check">✓</span> No Software Installation</span>
  </div>
</div>

<!-- ── HERO ─────────────────────────────────────────────────────── -->
<div class="dash-hero">
  <div class="dash-hero-inner anim-up">
    <div class="dash-hero-badge">🇮🇳 Made for Indian CAs &amp; Accountants</div>
    <h1>Your Complete <em>CA Toolkit</em></h1>
    <p>Professional automation tools built by a CA Article — saving hours of manual work every year.</p>
    <div class="hero-cta-row">
      {% if not username %}
      <a href="/login" class="hero-cta primary">Sign In &amp; Get Started →</a>
      <a href="#premium" class="hero-cta secondary">Explore All Tools ↓</a>
      {% else %}
      <a href="#premium" class="hero-cta primary">Explore All Tools ↓</a>
      <a href="/how-to-use" class="hero-cta secondary">How to Use</a>
      {% endif %}
    </div>
    <div class="hero-link-row">
      <a href="/story" class="hero-link">✦ Our Story</a>
      <a href="/how-to-use" class="hero-link">📖 Tool Guide</a>
      <a href="/privacy" class="hero-link">🔒 Privacy Policy</a>
    </div>
  </div>
</div>

<!-- ── USAGE STRIP ────────────────────────────────────────────── -->
{% if username %}
<div class="usage-strip">
  <div class="usage-box">
    <div class="usage-left">
      <div class="usage-icon">⚡</div>
      <div class="usage-text">
        <strong>{{ uploads_remaining }} uploads remaining</strong>
        <span>{{ uploads_used }} / {{ uploads_total }} used
          {% if validity_end %}&nbsp;· Valid till {{ validity_end[:10] }}{% endif %}
        </span>
      </div>
    </div>
    <div class="usage-bar-wrap">
      <div class="usage-bar-bg">
        <div class="usage-bar-fill"
             style="width:{{ bar_pct }}%;background:{{ '#DC2626' if uploads_remaining==0 else '#D97706' if uploads_remaining<=3 else '#059669' }}">
        </div>
      </div>
    </div>
    <a href="/pricing" class="upgrade-btn">⬆ Upgrade Plan</a>
  </div>
</div>
{% endif %}

<!-- ── PREMIUM TOOLS ───────────────────────────────────────────── -->
<div class="section-wrap" id="premium">
  <div class="section-hd">
    <h2>⚡ Premium Tools</h2>
    <span class="section-hd-badge">{% if username %}✓ {{ uploads_remaining }} uploads left{% else %}Login required{% endif %}</span>
  </div>
  <div class="tools-grid">

    {% if username %}
    <a href="/tool/converter" class="tool-card premium-card anim-up anim-d1">
      <span class="corner-badge cb-prem">⭐ Premium</span>
    {% else %}
    <a href="/login" class="tool-card premium-card anim-up anim-d1">
      <span class="corner-badge cb-lock">🔒 Sign In</span>
    {% endif %}
      <div class="tool-icon" style="background:linear-gradient(135deg,#E0F2FE,#BAE6FD)">📊</div>
      <h2>Balance Sheet Year-Shift</h2>
      <p>Roll over your comparative Excel balance sheet to any financial year. Shifts CY→PY, clears CY, restores formulas and updates every date — in seconds.</p>
      <div class="tool-footer">
        {% if username %}<span class="tool-tag tag-live-prem">✓ Live · Premium</span>
        {% else %}<span class="tool-tag tag-login">🔒 Login to Use</span>{% endif %}
        <span class="tool-arrow">→</span>
      </div>
    </a>

    {% if username %}
    <a href="/tool/gst-reconciliation" class="tool-card premium-card anim-up anim-d2">
      <span class="corner-badge cb-prem">⭐ Premium</span>
    {% else %}
    <a href="/login" class="tool-card premium-card anim-up anim-d2">
      <span class="corner-badge cb-lock">🔒 Sign In</span>
    {% endif %}
      <div class="tool-icon" style="background:linear-gradient(135deg,#FEF9C3,#FDE68A)">📋</div>
      <h2>GST Reconciliation</h2>
      <p>Compare Sales as per Books vs GSTR 3B returns. Upload your sales summary and GSTR 3B PDFs (ZIP) — month-wise, state-wise difference report instantly.</p>
      <div class="tool-footer">
        {% if username %}<span class="tool-tag tag-live-prem">✓ Live · Premium</span>
        {% else %}<span class="tool-tag tag-login">🔒 Login to Use</span>{% endif %}
        <span class="tool-arrow">→</span>
      </div>
    </a>

    {% if username %}
    <a href="/tool/tb-to-bs" class="tool-card premium-card anim-up anim-d3">
      <span class="corner-badge cb-prem">⭐ Premium</span>
    {% else %}
    <a href="/login" class="tool-card premium-card anim-up anim-d3">
      <span class="corner-badge cb-lock">🔒 Sign In</span>
    {% endif %}
      <div class="tool-icon" style="background:linear-gradient(135deg,#DCFCE7,#A7F3D0)">🗂️</div>
      <h2>Balance Sheet from Trial Balance</h2>
      <p>Upload your trial balance and BS template — auto-maps accounts and fills CY figures. Zero formatting change in your template.</p>
      <div class="tool-footer">
        {% if username %}<span class="tool-tag tag-live-prem">✓ Live · Premium</span>
        {% else %}<span class="tool-tag tag-login">🔒 Login to Use</span>{% endif %}
        <span class="tool-arrow">→</span>
      </div>
    </a>

    {% if username %}
    <a href="/tool/tshape" class="tool-card premium-card anim-up anim-d4">
      <span class="corner-badge cb-prem">&#11088; Premium</span>
    {% else %}
    <a href="/login" class="tool-card premium-card anim-up anim-d4">
      <span class="corner-badge cb-lock">&#128274; Sign In</span>
    {% endif %}
      <div class="tool-icon" style="background:linear-gradient(135deg,#F0FDFB,#CCFBF1)">&#128209;</div>
      <h2>T-Shape BS Converter</h2>
      <p>Convert a T-shaped balance sheet (.xls) into the standard comparative format. Auto-fills PY column — creditors, debtors, capital, fixed assets, loans and more.</p>
      <div class="tool-footer">
        {% if username %}<span class="tool-tag tag-live-prem">&#10003; Live &middot; Premium</span>
        {% else %}<span class="tool-tag tag-login">&#128274; Login to Use</span>{% endif %}
        <span class="tool-arrow">&#8594;</span>
      </div>
    </a>

  </div>

  <!-- ── FREE TOOLS ──────────────────────────────────────────────── -->
</div><!-- close section-wrap -->

<div class="free-section-wrap">
  <div class="free-section-inner">
    <div class="section-hd-free">
      <h2>🆓 Free Tools</h2>
      <span class="section-hd-badge" style="background:#DCFCE7;color:#065F46;border-color:#A7F3D0">No login needed</span>
    </div>
    <p class="section-sub-free">Six calculators built for everyday CA work — no account required, no uploads counted.</p>
  <div class="tools-grid tools-grid-4">

    <a href="/tool/tax-calculator" class="tool-card free-card anim-up anim-d1">
      <span class="corner-badge cb-free">Free</span>
      <div class="tool-icon" style="background:linear-gradient(135deg,#ECFDF5,#A7F3D0)">🧮</div>
      <h2>Income Tax Calculator</h2>
      <p>Old &amp; new regime for PY 2025-26. Income under 5 heads, TDS/TCS, surcharge &amp; cess — all built in.</p>
      <div class="tool-footer">
        <span class="tool-tag tag-live-free">✓ Live · Free</span>
        <span class="tool-arrow">→</span>
      </div>
    </a>

    <a href="/tool/tds-calculator" class="tool-card free-card anim-up anim-d2">
      <span class="corner-badge cb-free">Free</span>
      <div class="tool-icon" style="background:linear-gradient(135deg,#F0FDFA,#CCFBF1)">📑</div>
      <h2>TDS / TCS Calculator</h2>
      <p>TDS or TCS as per IT Act 2025 (Sec 393/394). New payment codes, rates, late deposit interest.</p>
      <div class="tool-footer">
        <span class="tool-tag tag-live-free">✓ Live · Free</span>
        <span class="tool-arrow">→</span>
      </div>
    </a>

    <a href="/tool/depreciation-calculator" class="tool-card free-card anim-up anim-d3">
      <span class="corner-badge cb-free">Free</span>
      <div class="tool-icon" style="background:linear-gradient(135deg,#ECFDF5,#BBF7D0)">🏭</div>
      <h2>Depreciation Calculator</h2>
      <p>Companies Act 2013 (WDV/SLM) and Income Tax Act. Full schedule with opening/closing WDV.</p>
      <div class="tool-footer">
        <span class="tool-tag tag-live-free">✓ Live · Free</span>
        <span class="tool-arrow">→</span>
      </div>
    </a>

    <a href="/tool/msme-calculator" class="tool-card free-card anim-up anim-d4">
      <span class="corner-badge cb-free">Free</span>
      <div class="tool-icon" style="background:linear-gradient(135deg,#F0FDFA,#99F6E4)">📄</div>
      <h2>MSME Disallowance</h2>
      <p>Check MSME payment compliance under Sec 43B(h). Overdue payments highlighted with total disallowance.</p>
      <div class="tool-footer">
        <span class="tool-tag tag-live-free">✓ Live · Free</span>
        <span class="tool-arrow">→</span>
      </div>
    </a>

    <a href="/tool/capital-gains-calculator" class="tool-card free-card anim-up anim-d5">
      <span class="corner-badge cb-free">Free</span>
      <div class="tool-icon" style="background:linear-gradient(135deg,#DCFCE7,#A7F3D0)">💰</div>
      <h2>Capital Gains Calculator</h2>
      <p>LTCG/STCG on property, shares, MF. Old vs new regime, indexation, zero-tax sale price.</p>
      <div class="tool-footer">
        <span class="tool-tag tag-live-free">✓ Live · Free</span>
        <span class="tool-arrow">→</span>
      </div>
    </a>

    <div class="tool-card coming disabled anim-up anim-d6">
      <div class="tool-icon" style="background:linear-gradient(135deg,#F0FDFA,#CCFBF1)">🚀</div>
      <h2>More Tools Coming</h2>
      <p>New utilities added regularly based on your feedback. Stay tuned!</p>
      <div class="tool-footer">
        <span class="tool-tag tag-soon">Coming Soon</span>
        <span></span>
      </div>
    </div>

  </div>
  </div>
</div>

<!-- ── STATS BAR ──────────────────────────────────────────────── -->
<div class="stats-row">
  <div class="stats-inner">
    <div class="stat-item anim-up anim-d1"><span class="stat-icon">🧰</span><div class="stat-n"><em>9</em>+</div><div class="stat-l">CA Tools Live</div></div>
    <div class="stat-item anim-up anim-d2"><span class="stat-icon">✔️</span><div class="stat-n">100%</div><div class="stat-l">Formatting Preserved</div></div>
    <div class="stat-item anim-up anim-d3"><span class="stat-icon">⚡</span><div class="stat-n">&lt;<em>10s</em></div><div class="stat-l">Processing Time</div></div>
    <div class="stat-item anim-up anim-d4"><span class="stat-icon">📄</span><div class="stat-n"><em>∞</em></div><div class="stat-l">Templates Supported</div></div>
  </div>
</div>

<!-- ── LEARN MORE STRIP ──────────────────────────────────────────── -->
<div style="background:linear-gradient(135deg,#0B5D4A,#0E8A7B);padding:40px 24px;margin:0">
  <div style="max-width:900px;margin:0 auto;display:grid;grid-template-columns:1fr 1fr 1fr;gap:20px;text-align:center"
       class="lm-grid">
    <a href="/story" style="text-decoration:none;background:rgba(255,255,255,.08);border:1px solid rgba(255,255,255,.15);
       border-radius:14px;padding:22px 18px;display:block;transition:background .2s,transform .2s"
       onmouseover="this.style.background='rgba(255,255,255,.15)';this.style.transform='translateY(-3px)'"
       onmouseout="this.style.background='rgba(255,255,255,.08)';this.style.transform='translateY(0)'">
      <div style="font-size:26px;margin-bottom:10px">✦</div>
      <div style="font-family:'Plus Jakarta Sans',sans-serif;font-size:15px;font-weight:800;color:#fff;margin-bottom:6px">Our Story</div>
      <div style="font-size:12px;color:rgba(255,255,255,.6);line-height:1.6">Why CA Toolkit was built — and who built it</div>
    </a>
    <a href="/how-to-use" style="text-decoration:none;background:rgba(255,255,255,.08);border:1px solid rgba(255,255,255,.15);
       border-radius:14px;padding:22px 18px;display:block;transition:background .2s,transform .2s"
       onmouseover="this.style.background='rgba(255,255,255,.15)';this.style.transform='translateY(-3px)'"
       onmouseout="this.style.background='rgba(255,255,255,.08)';this.style.transform='translateY(0)'">
      <div style="font-size:26px;margin-bottom:10px">📖</div>
      <div style="font-family:'Plus Jakarta Sans',sans-serif;font-size:15px;font-weight:800;color:#fff;margin-bottom:6px">How to Use</div>
      <div style="font-size:12px;color:rgba(255,255,255,.6);line-height:1.6">Step-by-step guide for every tool</div>
    </a>
    <a href="/privacy" style="text-decoration:none;background:rgba(255,255,255,.08);border:1px solid rgba(255,255,255,.15);
       border-radius:14px;padding:22px 18px;display:block;transition:background .2s,transform .2s"
       onmouseover="this.style.background='rgba(255,255,255,.15)';this.style.transform='translateY(-3px)'"
       onmouseout="this.style.background='rgba(255,255,255,.08)';this.style.transform='translateY(0)'">
      <div style="font-size:26px;margin-bottom:10px">🔒</div>
      <div style="font-family:'Plus Jakarta Sans',sans-serif;font-size:15px;font-weight:800;color:#fff;margin-bottom:6px">Privacy Policy</div>
      <div style="font-size:12px;color:rgba(255,255,255,.6);line-height:1.6">How we handle your data and files</div>
    </a>
  </div>
</div>
<style>@media(max-width:640px){.lm-grid{grid-template-columns:1fr!important}}</style>

<footer>
  <div class="ft-main">
    <div>
      <div class="ft-brand-name">CA<span>Toolkit</span></div>
      <p class="ft-brand-desc">Automation tools for Indian Chartered Accountants — built by a CA Article from Ludhiana, Punjab. Saving hours of manual work every year-end.</p>
      <div style="margin-top:12px;display:flex;flex-direction:column;gap:5px">
        <span style="font-size:11.5px;color:#6EE7B7">🔒 Files deleted automatically after processing</span>
        <span style="font-size:11.5px;color:#6EE7B7">🔐 No financial data stored permanently</span>
        <span style="font-size:11.5px;color:#6EE7B7">🇮🇳 Built &amp; operated in India</span>
      </div>
      <div class="ft-socials">
        <a href="https://wa.me/918427651580" target="_blank" title="WhatsApp"><svg viewBox="0 0 24 24"><path d="M17.472 14.382c-.297-.149-1.758-.867-2.03-.967-.273-.099-.471-.148-.67.15-.197.297-.767.966-.94 1.164-.173.199-.347.223-.644.075-.297-.15-1.255-.463-2.39-1.475-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.298-.347.446-.52.149-.174.198-.298.298-.497.099-.198.05-.371-.025-.52-.075-.149-.669-1.612-.916-2.207-.242-.579-.487-.5-.669-.51-.173-.008-.371-.01-.57-.01-.198 0-.52.074-.792.372-.272.297-1.04 1.016-1.04 2.479 0 1.462 1.065 2.875 1.213 3.074.149.198 2.096 3.2 5.077 4.487.709.306 1.262.489 1.694.625.712.227 1.36.195 1.871.118.571-.085 1.758-.719 2.006-1.413.248-.694.248-1.289.173-1.413-.074-.124-.272-.198-.57-.347m-5.421 7.403h-.004a9.87 9.87 0 01-5.031-1.378l-.361-.214-3.741.982.998-3.648-.235-.374a9.86 9.86 0 01-1.51-5.26c.001-5.45 4.436-9.884 9.888-9.884 2.64 0 5.122 1.03 6.988 2.898a9.825 9.825 0 012.893 6.994c-.003 5.45-4.437 9.884-9.885 9.884m8.413-18.297A11.815 11.815 0 0012.05 0C5.495 0 .16 5.335.157 11.892c0 2.096.547 4.142 1.588 5.945L.057 24l6.305-1.654a11.882 11.882 0 005.683 1.448h.005c6.554 0 11.89-5.335 11.893-11.893a11.821 11.821 0 00-3.48-8.413z"/></svg></a>
      </div>
    </div>
    <div>
      <div class="ft-col-title">Tools & Pages</div>
      <ul class="ft-links">
        <li><a href="/tool/converter">BS Year-Shift</a></li>
        <li><a href="/tool/tb-to-bs">TB → Balance Sheet</a></li>
        <li><a href="/tool/gst-reconciliation">GST Reconciliation</a></li>
        <li><a href="/tool/tshape">T-Shape BS Converter</a></li>
        <li><a href="/tool/tax-calculator">Tax Calculator</a></li>
        <li><a href="/tool/tds-calculator">TDS / TCS Calculator</a></li>
        <li><a href="/story">Our Story</a></li>
        <li><a href="/how-to-use">How to Use</a></li>
        <li><a href="/privacy">Privacy Policy</a></li>
      </ul>
    </div>
    <div>
      <div class="ft-col-title">Contact &amp; Support</div>
      <div class="ft-contact-name">Sumit Verma</div>
      <div class="ft-contact-addr">CA Article · Ludhiana, Punjab<br/>Creator of CA Toolkit</div>
      <div class="ft-contact-line">📧 <a href="mailto:sumitverma2880@gmail.com">sumitverma2880@gmail.com</a></div>
      <div class="ft-contact-line">💬 <a href="https://wa.me/918427651580">WhatsApp · +91 84276 51580</a></div>
      <div class="ft-contact-line" style="margin-top:10px;font-size:11px;color:#64748B">Accounts created by admin only.<br/>Contact to get access.</div>
    </div>
  </div>
  <div class="ft-bottom">
    <span class="ft-bottom-left">©2026 CA Toolkit · All Rights Reserved ·
      <a href="/privacy" style="color:#475569;text-decoration:none">Privacy</a> ·
      <a href="/story" style="color:#475569;text-decoration:none">Our Story</a> ·
      <a href="/how-to-use" style="color:#475569;text-decoration:none">How to Use</a> ·
      <span style="color:#F87171">No refund after first upload is used</span></span>
    <span class="ft-bottom-right">Built for Indian CAs · Ludhiana, Punjab</span>
  </div>
</footer>
<a href="https://wa.me/918427651580" target="_blank" class="wa-float" title="WhatsApp Support">
  <svg viewBox="0 0 24 24"><path d="M17.472 14.382c-.297-.149-1.758-.867-2.03-.967-.273-.099-.471-.148-.67.15-.197.297-.767.966-.94 1.164-.173.199-.347.223-.644.075-.297-.15-1.255-.463-2.39-1.475-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.298-.347.446-.52.149-.174.198-.298.298-.497.099-.198.05-.371-.025-.52-.075-.149-.669-1.612-.916-2.207-.242-.579-.487-.5-.669-.51-.173-.008-.371-.01-.57-.01-.198 0-.52.074-.792.372-.272.297-1.04 1.016-1.04 2.479 0 1.462 1.065 2.875 1.213 3.074.149.198 2.096 3.2 5.077 4.487.709.306 1.262.489 1.694.625.712.227 1.36.195 1.871.118.571-.085 1.758-.719 2.006-1.413.248-.694.248-1.289.173-1.413-.074-.124-.272-.198-.57-.347m-5.421 7.403h-.004a9.87 9.87 0 01-5.031-1.378l-.361-.214-3.741.982.998-3.648-.235-.374a9.86 9.86 0 01-1.51-5.26c.001-5.45 4.436-9.884 9.888-9.884 2.64 0 5.122 1.03 6.988 2.898a9.825 9.825 0 012.893 6.994c-.003 5.45-4.437 9.884-9.885 9.884m8.413-18.297A11.815 11.815 0 0012.05 0C5.495 0 .16 5.335.157 11.892c0 2.096.547 4.142 1.588 5.945L.057 24l6.305-1.654a11.882 11.882 0 005.683 1.448h.005c6.554 0 11.89-5.335 11.893-11.893a11.821 11.821 0 00-3.48-8.413z"/></svg>
</a>
</a>

<!-- ── SCROLL TO TOP ──────────────────────────────────────────────── -->
<button id="scrollTopBtn" onclick="window.scrollTo({top:0,behavior:'smooth'})" title="Back to top" aria-label="Scroll to top" style="position:fixed;bottom:82px;right:24px;width:42px;height:42px;background:var(--white);border:1.5px solid var(--border2);border-radius:50%;display:flex;align-items:center;justify-content:center;cursor:pointer;font-size:18px;color:var(--brand-d);box-shadow:var(--shadow-md);opacity:0;transform:translateY(12px);transition:opacity .28s,transform .28s;z-index:997;pointer-events:none">↑</button>

<script>
// Scroll-to-top
(function(){var b=document.getElementById('scrollTopBtn');window.addEventListener('scroll',function(){b.style.opacity=window.scrollY>400?'1':'0';b.style.transform=window.scrollY>400?'translateY(0)':' translateY(12px)';b.style.pointerEvents=window.scrollY>400?'auto':'none';},{passive:true});})();
// IntersectionObserver scroll-reveal
(function(){var els=document.querySelectorAll('.reveal');if(!els.length)return;var obs=new IntersectionObserver(function(entries){entries.forEach(function(e,i){if(e.isIntersecting){setTimeout(function(){e.target.style.opacity='1';e.target.style.transform='translateY(0)';},i*60);obs.unobserve(e.target);}});},{threshold:0.1});els.forEach(function(el){el.style.opacity='0';el.style.transform='translateY(22px)';el.style.transition='opacity .48s cubic-bezier(.22,.68,0,1.2),transform .48s cubic-bezier(.22,.68,0,1.2)';obs.observe(el);});})();
// Mobile hamburger
function toggleMobileMenu(){var m=document.getElementById('mobileMenu');var b=document.getElementById('hamburgerBtn');m.classList.toggle('open');b.classList.toggle('open');}
document.addEventListener('click',function(e){var m=document.getElementById('mobileMenu');if(m&&m.classList.contains('open')&&!m.contains(e.target)&&!document.getElementById('hamburgerBtn').contains(e.target)){m.classList.remove('open');}});
</script>
</body></html>"""

# ══════════════════════════════════════════════════════════════════════════════
#  BALANCE SHEET CONVERTER TOOL PAGE════════════════════════════════════════════════════════════════════════════

CONVERTER_T = """<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Balance Sheet Year-Shift – CA Toolkit</title>

<style>
""" + BASE_CSS + """
.nav-links{display:flex;gap:20px;list-style:none}
.nav-links a{text-decoration:none;color:var(--muted);font-size:13px;font-weight:500;transition:color .2s}
.nav-links a:hover{color:var(--brand)}
.hero{text-align:center;padding:56px 24px 40px;max-width:700px;margin:0 auto}
.hero-badge{display:inline-flex;align-items:center;gap:6px;background:var(--brand-l);
            color:var(--brand);border:1px solid #C7D2FE;border-radius:99px;
            padding:5px 14px;font-size:12px;font-weight:600;margin-bottom:18px}
h1{font-size:clamp(24px,4vw,40px);font-weight:800;line-height:1.15;
   letter-spacing:-.5px;margin-bottom:14px}
h1 em{font-style:normal;color:var(--accent)}
.hero p{font-size:15px;color:var(--muted);line-height:1.7;max-width:520px;margin:0 auto 28px}
.stats{display:flex;justify-content:center;gap:36px;flex-wrap:wrap;
       padding:16px 24px;background:var(--white);
       border-top:1px solid var(--border);border-bottom:1px solid var(--border)}
.stat-n{font-size:20px;font-weight:800;color:var(--brand)}
.stat-l{font-size:11px;color:var(--muted);margin-top:2px}
.main{max-width:1080px;margin:0 auto;padding:40px 24px;
      display:grid;grid-template-columns:1fr 1fr;gap:24px;align-items:start}
@media(max-width:768px){.main{grid-template-columns:1fr}}
.card{background:var(--white);border-radius:var(--radius);border:1px solid var(--border);
      box-shadow:var(--shadow);overflow:hidden;animation:fadeUp .4s ease-out both}
.card-head{padding:16px 20px;border-bottom:1px solid var(--border);
           display:flex;align-items:center;gap:10px}
.card-head .icon{width:32px;height:32px;border-radius:8px;display:flex;
                 align-items:center;justify-content:center;font-size:16px}
.card-head h2{font-size:14px;font-weight:700}
.card-head p{font-size:12px;color:var(--muted);margin-top:1px}
.card-body{padding:20px}
.usage-row{display:flex;justify-content:space-between;align-items:center;
           font-size:12px;font-weight:600;margin-bottom:5px}
.usage-bar-bg{background:#F3F4F6;border-radius:99px;height:6px;overflow:hidden;margin-bottom:14px}
.usage-bar-fill{height:100%;border-radius:99px;transition:width .4s}
.field{margin-bottom:16px}
label{display:block;font-size:11px;font-weight:600;text-transform:uppercase;
      letter-spacing:.04em;color:var(--muted);margin-bottom:5px}
.hint{font-size:11px;color:var(--muted);margin-top:4px}
.dropzone{border:2px dashed var(--border);border-radius:10px;padding:24px 14px;
          text-align:center;cursor:pointer;transition:all .2s;position:relative;background:var(--bg)}
.dropzone:hover,.dropzone.drag{border-color:var(--brand);background:#EFF6FF}
.dropzone input{position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%;pointer-events:none}
.dz-icon{font-size:26px;margin-bottom:6px}
.dz-text{font-size:12px;color:var(--muted)}
.dz-text strong{color:var(--brand)}
.dz-file{font-size:12px;font-weight:600;color:var(--green);margin-top:5px;display:none}
.row2{display:grid;grid-template-columns:1fr 1fr;gap:12px}
input[type=number],input[type=text]{width:100%;border:1.5px solid var(--border);
  border-radius:8px;padding:9px 12px;font-family:inherit;font-size:13px;
  color:var(--ink);background:var(--white);transition:border-color .2s;outline:none}
input:focus{border-color:var(--brand)}
.btn{width:100%;background:var(--brand);color:#fff;border:none;border-radius:10px;
     padding:12px;font-family:inherit;font-size:14px;font-weight:700;cursor:pointer;
     transition:background .2s;display:flex;align-items:center;justify-content:center;gap:8px}
.btn:hover{background:var(--brand-d)}
.btn:disabled{background:#93C5FD;cursor:not-allowed}
.spinner{width:16px;height:16px;border:2px solid rgba(255,255,255,.3);
         border-top-color:#fff;border-radius:50%;animation:spin .7s linear infinite;display:none}
@keyframes spin{to{transform:rotate(360deg)}}
#status{margin-top:12px;border-radius:8px;padding:12px 14px;font-size:13px;display:none;line-height:1.6}
#status.success{background:#ECFDF5;border:1px solid #A7F3D0;color:#065F46}
#status.error{background:#FEF2F2;border:1px solid #FECACA;color:#991B1B}
.log-list{margin-top:6px;padding-left:14px;font-size:11px;color:#374151;line-height:2}
.dl-btn{display:none;margin-top:10px;width:100%;background:var(--green);color:#fff;
        border:none;border-radius:10px;padding:11px;font-family:inherit;font-size:13px;
        font-weight:600;cursor:pointer;text-decoration:none;text-align:center;transition:background .2s}
.dl-btn:hover{background:#059669}
.steps{padding:0;list-style:none;counter-reset:step}
.steps li{display:flex;gap:10px;align-items:flex-start;padding:12px 0;border-bottom:1px solid var(--border)}
.steps li:last-child{border:none}
.steps li::before{counter-increment:step;content:counter(step);min-width:24px;height:24px;
                  background:var(--brand);color:#fff;border-radius:50%;display:flex;
                  align-items:center;justify-content:center;font-size:11px;font-weight:700;margin-top:1px}
.steps li strong{display:block;font-size:12px;font-weight:600;margin-bottom:2px}
.steps li span{font-size:11px;color:var(--muted)}
.features{max-width:1080px;margin:0 auto;padding:0 24px 40px;
          display:grid;grid-template-columns:repeat(3,1fr);gap:16px}
@media(max-width:640px){.features{grid-template-columns:1fr}}
.feat{background:var(--white);border:1px solid var(--border);border-radius:var(--radius);
      padding:20px;text-align:center}
.feat .fi{font-size:26px;margin-bottom:8px}
.feat h3{font-size:13px;font-weight:700;margin-bottom:4px}
.feat p{font-size:12px;color:var(--muted);line-height:1.6}
.pricing-section{background:var(--white);border-top:1px solid var(--border);
                 border-bottom:1px solid var(--border);padding:48px 24px}
.pricing-section h2{text-align:center;font-size:24px;font-weight:800;margin-bottom:6px}
.psub{text-align:center;color:var(--muted);font-size:13px;margin-bottom:32px}
.plans{max-width:1080px;margin:0 auto;
       display:grid;grid-template-columns:repeat(6,1fr);gap:14px}
@media(max-width:900px){.plans{grid-template-columns:repeat(3,1fr)}}
@media(max-width:480px){.plans{grid-template-columns:1fr}}
.plan{border:1.5px solid var(--border);border-radius:var(--radius);padding:20px 16px;position:relative}
.plan.pop{border-color:var(--brand)}
.plan-badge{position:absolute;top:-10px;left:50%;transform:translateX(-50%);
            background:var(--brand);color:#fff;font-size:10px;font-weight:700;
            padding:2px 10px;border-radius:99px;white-space:nowrap}
.plan-name{font-size:10px;font-weight:700;color:var(--muted);text-transform:uppercase;
           letter-spacing:.06em;margin-bottom:6px}
.plan-price{font-size:24px;font-weight:800;color:var(--ink);margin-bottom:2px}
.plan-uploads{font-size:12px;font-weight:700;color:var(--brand);margin-bottom:2px}
.plan-validity{font-size:10px;color:var(--muted);margin-bottom:14px}
.plan ul{list-style:none;margin-bottom:16px}
.plan ul li{font-size:11px;padding:3px 0;display:flex;gap:5px}
.plan ul li::before{content:"✓";color:var(--green);font-weight:700}
.plan-btn{display:block;text-align:center;padding:8px;border-radius:7px;
          font-size:11px;font-weight:600;text-decoration:none;transition:all .2s;
          border:1.5px solid var(--brand);color:var(--brand)}
.plan-btn:hover{background:var(--brand);color:#fff}
.plan.pop .plan-btn{background:var(--brand);color:#fff}
.no-refund-note{text-align:center;font-size:11px;color:var(--muted);margin-top:14px;font-weight:500}
.faq-section{max-width:720px;margin:0 auto;padding:40px 24px}
.faq-section h2{font-size:20px;font-weight:800;text-align:center;margin-bottom:24px}
details{border:1px solid var(--border);border-radius:10px;margin-bottom:8px}
summary{padding:12px 16px;font-size:13px;font-weight:600;cursor:pointer;
        list-style:none;display:flex;justify-content:space-between;align-items:center}
summary::after{content:"＋";color:var(--muted)}
details[open] summary::after{content:"－"}
details p{padding:0 16px 12px;font-size:12px;color:var(--muted);line-height:1.7}
.contact-section{background:#EFF6FF;border-top:1px solid #BFDBFE;padding:36px 24px;text-align:center}
.contact-section h2{font-size:18px;font-weight:800;margin-bottom:6px}
.contact-section p{font-size:13px;color:var(--muted);margin-bottom:14px}
.contact-grid{display:flex;justify-content:center;gap:20px;flex-wrap:wrap}
.contact-item{background:var(--white);border:1px solid var(--border);border-radius:10px;
              padding:14px 20px;font-size:13px}
.contact-item strong{display:block;font-size:10px;text-transform:uppercase;
                     letter-spacing:.05em;color:var(--muted);margin-bottom:3px}
.contact-item a{color:var(--brand);text-decoration:none;font-weight:600}
.limit-banner{max-width:640px;margin:0 auto 0;padding:0 24px}
.limit-box{background:#FEF2F2;border:1px solid #FECACA;border-radius:var(--radius);
           padding:20px 24px;text-align:center;margin-top:16px}
.limit-box h3{font-size:15px;font-weight:700;color:#991B1B;margin-bottom:8px}
.limit-box p{font-size:13px;color:#7F1D1D;line-height:1.7;margin-bottom:10px}
.limit-box a{color:var(--brand);font-weight:600;text-decoration:none}
.toast{position:fixed;bottom:24px;right:24px;background:var(--ink);color:#fff;
       padding:11px 18px;border-radius:10px;font-size:13px;font-weight:500;
       transform:translateY(80px);transition:transform .3s;z-index:999}
.toast.show{transform:translateY(0)}
@media(max-width:480px){.row2{grid-template-columns:1fr}}
</style></head><body>

<nav>
  <a href="/" class="logo">CA<span class="logo-dot"></span><span>Toolkit</span></a>
  <ul class="nav-links">
    <li><a href="#tool">Tool</a></li>
    <li><a href="#pricing">Pricing</a></li>
    <li><a href="#faq">FAQ</a></li>
    <li><a href="#contact">Contact</a></li>
  </ul>
  <div class="nav-right">
    {% if username %}
    <div class="nav-user">
      <span class="nav-avatar">{{ username[0].upper() }}</span>
      <strong>{{ username }}</strong>
      <span class="badge b-{{ plan }}">{{ plan_label }}</span>
      {% if is_admin %}<span class="badge" style="background:#EDE9FE;color:#5B21B6">Admin</span>{% endif %}
    </div>
    <div class="nav-sep"></div>
    {% if is_admin %}<a href="/admin" class="nav-btn ghost">⚙ Admin</a>{% endif %}
    {% endif %}
    <a href="/" class="nav-btn dash">⬅ Dashboard</a>
    {% if username %}<a href="/logout" class="nav-link">Sign out</a>
    {% else %}<a href="/login" class="nav-btn">Sign In →</a>{% endif %}
  </div>
</nav>

<section class="hero">
  <div class="hero-badge">🇮🇳 CA Tool · Balance Sheet Year-Shift</div>
  <h1>Roll Over to <em>Any Financial Year</em><br/>in Seconds</h1>
  <p>Upload your comparative Excel balance sheet — shifts CY→PY, clears CY column, restores all formulas, and updates every date automatically.</p>
</section>

{% if uploads_left == 0 %}
<div class="limit-banner">
  <div class="limit-box">
    <h3>🔒 No uploads remaining</h3>
    <p>You've used all your uploads. Contact us to recharge your account.<br/>
       Pay via UPI and email your screenshot — upgraded within a few hours.</p>
    <p>📧 <a href="mailto:{{ contact_email }}">{{ contact_email }}</a> &nbsp;|&nbsp;
       💳 UPI: <strong>{{ contact_upi }}</strong></p>
    <p style="font-size:11px;color:#9CA3AF;margin-top:8px">No refund after first upload is used.</p>
  </div>
</div>
{% endif %}

<div class="stats">
  <div class="stat"><div class="stat-n">100%</div><div class="stat-l">Formatting preserved</div></div>
  <div class="stat"><div class="stat-n">All sheets</div><div class="stat-l">Processed at once</div></div>
  <div class="stat"><div class="stat-n">&lt;10 sec</div><div class="stat-l">Processing time</div></div>
  <div class="stat"><div class="stat-n">Any format</div><div class="stat-l">Works with all CA templates</div></div>
</div>

<div class="main" id="tool">
  <div class="card">
    <div class="card-head">
      <div class="icon" style="background:#EFF6FF">📊</div>
      <div>
        <h2>Process Your Balance Sheet</h2>
        <p>{{ plan_label }} · {{ uploads_left }} upload{{ 's' if uploads_left != 1 else '' }} remaining</p>
      </div>
    </div>
    <div class="card-body">
      <div class="usage-row">
        <span style="color:var(--muted)">Uploads used</span>
        <span><strong>{{ uploads_used }}</strong> / {{ uploads_total }}
          {% if validity_end %}<span style="color:#9CA3AF;font-weight:400"> · expires {{ validity_end[:10] }}</span>{% endif %}
        </span>
      </div>
      <div class="usage-bar-bg">
        <div class="usage-bar-fill"
             style="width:{{ bar_pct }}%;background:{{ '#EF4444' if uploads_left==0 else '#F59E0B' if uploads_left<=3 else '#10B981' }}">
        </div>
      </div>
      <div class="field">
        <label>Upload Excel File (.xlsx / .xls)</label>
        <div class="dropzone" id="dropzone">
          <input type="file" id="xlFile" accept=".xlsx,.xls,.xlsb" {{ 'disabled' if uploads_left==0 else '' }}/>
          <div class="dz-icon">📁</div>
          <div class="dz-text"><strong>Click to browse</strong> or drag &amp; drop</div>
          <div class="dz-text" style="margin-top:3px">.xlsx or .xls · Max 20 MB</div>
          <div class="dz-text" style="margin-top:2px;font-size:11px;color:var(--muted)">(.xlsb also accepted — formatting not preserved)</div>
          <div class="dz-file" id="dzFile"></div>
        </div>
      </div>
      <div class="row2">
        <div class="field">
          <label>Closing Year (CY)</label>
          <input type="number" id="closingYear" placeholder="e.g. 2025" min="2000" max="2100"
                 {{ 'disabled' if uploads_left==0 else '' }}/>
          <p class="hint">Year ending 31.03.YYYY</p>
        </div>
        <div class="field">
          <label>New Year</label>
          <input type="number" id="newYear" placeholder="Auto-filled" readonly/>
          <p class="hint">Auto-filled</p>
        </div>
      </div>
      <div class="field">
        <label>Output Filename <span style="font-weight:400;text-transform:none;color:var(--muted)">(optional)</span></label>
        <input type="text" id="outputName" placeholder="e.g. ClientName_BS"
               {{ 'disabled' if uploads_left==0 else '' }}/>
        <p class="hint">Leave blank to auto-generate</p>
      </div>
      <button class="btn" id="processBtn" onclick="processFile()"
              {{ 'disabled' if uploads_left==0 else '' }}>
        <span id="btnText">⚡ Process &amp; Download</span>
        <div class="spinner" id="spinner"></div>
      </button>
      <div id="status"></div>
      <a id="dlBtn" class="dl-btn" href="#">⬇&nbsp; Download Processed File</a>
    </div>
  </div>

  <div>
    <div class="card" style="margin-bottom:18px">
      <div class="card-head">
        <div class="icon" style="background:#F0FDF4">✅</div>
        <div><h2>How It Works</h2><p>4 steps, fully automatic</p></div>
      </div>
      <div class="card-body">
        <ol class="steps">
          <li><strong>Upload your Excel file</strong>
              <span>Your FY comparative balance sheet with CY and PY columns</span></li>
          <li><strong>Auto-detects all CY/PY columns</strong>
              <span>Scans every sheet and finds correct data columns automatically</span></li>
          <li><strong>Shifts CY → PY, clears CY</strong>
              <span>Values become PY. Formulas and cross-sheet links restored. CY cleared for fresh entry</span></li>
          <li><strong>Updates every date</strong>
              <span>All date strings across every sheet updated in one shot</span></li>
        </ol>
      </div>
    </div>
    <div class="card">
      <div class="card-head">
        <div class="icon" style="background:#FFFBEB">🔒</div>
        <div><h2>Your Data is Safe</h2><p>Privacy first</p></div>
      </div>
      <div class="card-body">
        <ul class="steps">
          <li><strong>Deleted immediately after download</strong>
              <span>File removed from server the moment processing completes</span></li>
          <li><strong>HTTPS encrypted</strong><span>All transfers encrypted end-to-end</span></li>
          <li><strong>No data stored</strong><span>We never read, store, or share your financial data</span></li>
        </ul>
      </div>
    </div>
  </div>
</div>

<div class="features">
  <div class="feat"><div class="fi">🧮</div><h3>Formulas Preserved</h3><p>Every SUM and cross-sheet reference in the PY column is snapshotted and restored automatically.</p></div>
  <div class="feat"><div class="fi">🎨</div><h3>Formatting Intact</h3><p>Fonts, borders, colors, merged cells, column widths — everything preserved exactly.</p></div>
  <div class="feat"><div class="fi">📅</div><h3>All Dates Updated</h3><p>Every date string across every sheet updated in one shot.</p></div>
  <div class="feat"><div class="fi">🗂️</div><h3>All Sheets at Once</h3><p>BS, P&L, Notes, Capital, Fixed Assets — every sheet processed together.</p></div>
  <div class="feat"><div class="fi">🔄</div><h3>Any CA Template</h3><p>Auto-detects column positions. Works with any firm's template.</p></div>
  <div class="feat"><div class="fi">⚡</div><h3>Instant Results</h3><p>What took 30–45 minutes of manual work now takes under 10 seconds.</p></div>
</div>

<section class="pricing-section" id="pricing">
  <h2>Simple Pricing</h2>
  <p class="psub">Upload-based · 3-month validity · Uploads stack when you recharge</p>
  <div class="plans">
    <div class="plan">
      <div class="plan-name">Free</div>
      <div class="plan-price">₹0</div>
      <div class="plan-uploads">2 uploads</div>
      <div class="plan-validity">Try it out</div>
      <ul><li>All features</li><li>All sheet types</li><li>Up to 20 MB</li></ul>
      <a href="#tool" class="plan-btn">Get Started</a>
    </div>
    <div class="plan">
      <div class="plan-name">Starter</div>
      <div class="plan-price">₹60</div>
      <div class="plan-uploads">10 uploads</div>
      <div class="plan-validity">3 month validity</div>
      <ul><li>All features</li><li>All sheet types</li><li>Up to 20 MB</li></ul>
      <a href="#contact" class="plan-btn">Contact to Buy</a>
    </div>
    <div class="plan pop">
      <div class="plan-badge">Most Popular</div>
      <div class="plan-name">Standard</div>
      <div class="plan-price">₹130</div>
      <div class="plan-uploads">25 uploads</div>
      <div class="plan-validity">3 month validity</div>
      <ul><li>All features</li><li>Priority support</li><li>Up to 20 MB</li></ul>
      <a href="#contact" class="plan-btn">Contact to Buy</a>
    </div>
    <div class="plan">
      <div class="plan-name">Professional</div>
      <div class="plan-price">₹270</div>
      <div class="plan-uploads">60 uploads</div>
      <div class="plan-validity">3 month validity</div>
      <ul><li>All features</li><li>Priority support</li><li>Up to 20 MB</li></ul>
      <a href="#contact" class="plan-btn">Contact to Buy</a>
    </div>
    <div class="plan">
      <div class="plan-name">Firm</div>
      <div class="plan-price">₹600</div>
      <div class="plan-uploads">150 uploads</div>
      <div class="plan-validity">3 month validity</div>
      <ul><li>All features</li><li>WhatsApp support</li><li>Up to 20 MB</li></ul>
      <a href="#contact" class="plan-btn">Contact to Buy</a>
    </div>
    <div class="plan">
      <div class="plan-name">CA Firm</div>
      <div class="plan-price">₹1,000</div>
      <div class="plan-uploads">500 uploads</div>
      <div class="plan-validity">3 month validity</div>
      <ul><li>All features + GST Recon</li><li>WhatsApp support</li><li>Best for CA firms</li></ul>
      <a href="#contact" class="plan-btn">Contact to Buy</a>
    </div>
  </div>
  <p class="no-refund-note">⚠ No refund after first upload is used &nbsp;·&nbsp; Unused uploads stack when you recharge before expiry</p>
</section>

<section class="faq-section" id="faq">
  <h2>Frequently Asked Questions</h2>
  <details><summary>Which Excel formats are supported?</summary>
    <p>.xlsx (Excel 2007+), .xls (legacy Excel), and .xlsb (Excel Binary) are all supported — .xls and .xlsb files are automatically converted before processing.</p></details>
  <details><summary>Will it work with my firm's custom template?</summary>
    <p>Yes. Auto-detects CY/PY columns by scanning date headers like "31.03.2025". Works with any Indian CA template.</p></details>
  <details><summary>Are my formulas and formatting safe?</summary>
    <p>Yes. Formulas in PY column are snapshotted before and restored after. Formatting is never touched.</p></details>
  <details><summary>What happens to my uploaded file?</summary>
    <p>Deleted from our server immediately after you download the result. We never store or share your data.</p></details>
  <details><summary>How do I purchase a plan?</summary>
    <p>Pay via UPI to <strong>{{ contact_upi }}</strong> and email your screenshot to <strong>{{ contact_email }}</strong>. Account upgraded within a few hours.</p></details>
  <details><summary>Do unused uploads carry over when I recharge?</summary>
    <p>Yes. Remaining uploads stack on top of the new plan if you recharge before expiry.</p></details>
</section>

<section class="contact-section" id="contact">
  <h2>Purchase a Plan</h2>
  <p>Pay via UPI and send your payment screenshot to our email. We'll upgrade your account within a few hours.</p>
  <div class="contact-grid">
    <div class="contact-item">
      <strong>Email</strong>
      <a href="mailto:{{ contact_email }}">{{ contact_email }}</a>
    </div>
    <div class="contact-item">
      <strong>UPI Payment</strong>
      <span style="font-weight:600;color:var(--ink)">{{ contact_upi }}</span>
    </div>
  </div>
</section>

<footer>
  <div class="ft-main">
    <div>
      <div class="ft-brand-name">CA<span>Toolkit</span></div>
      <p class="ft-brand-desc">CA Toolkit is a comprehensive utility platform built by a CA Article from Ludhiana, Punjab, providing automation tools for Indian Chartered Accountants. The platform saves hours of manual work every year — from Balance Sheet year-shift to GST reconciliation, tax calculations, and more.</p>
    </div>
    <div>
      <div class="ft-col-title">Know More</div>
      <ul class="ft-links">
        <li><a href="/">Home</a></li>
        <li><a href="/">BS Year Shift</a></li>
        <li><a href="/tool/tb-to-bs">TB → Balance Sheet</a></li>
        <li><a href="/tool/tax-calculator">Tax Calculator</a></li>
        <li><a href="/privacy">Privacy Policy</a></li>
      </ul>
    </div>
    <div>
      <div class="ft-col-title">Contact Us</div>
      <div class="ft-contact-name">CA Toolkit</div>
      <div class="ft-contact-addr">Built for Indian Chartered Accountants<br/>Created by CA Article · Ludhiana, Punjab</div>
      <div class="ft-contact-line">Support · <a href="https://wa.me/918427651580" style="color:#9CA3AF">WhatsApp Chat</a></div>
      <div class="ft-socials">
        <a href="https://wa.me/918427651580" target="_blank" title="WhatsApp"><svg viewBox="0 0 24 24"><path d="M17.472 14.382c-.297-.149-1.758-.867-2.03-.967-.273-.099-.471-.148-.67.15-.197.297-.767.966-.94 1.164-.173.199-.347.223-.644.075-.297-.15-1.255-.463-2.39-1.475-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.298-.347.446-.52.149-.174.198-.298.298-.497.099-.198.05-.371-.025-.52-.075-.149-.669-1.612-.916-2.207-.242-.579-.487-.5-.669-.51-.173-.008-.371-.01-.57-.01-.198 0-.52.074-.792.372-.272.297-1.04 1.016-1.04 2.479 0 1.462 1.065 2.875 1.213 3.074.149.198 2.096 3.2 5.077 4.487.709.306 1.262.489 1.694.625.712.227 1.36.195 1.871.118.571-.085 1.758-.719 2.006-1.413.248-.694.248-1.289.173-1.413-.074-.124-.272-.198-.57-.347m-5.421 7.403h-.004a9.87 9.87 0 01-5.031-1.378l-.361-.214-3.741.982.998-3.648-.235-.374a9.86 9.86 0 01-1.51-5.26c.001-5.45 4.436-9.884 9.888-9.884 2.64 0 5.122 1.03 6.988 2.898a9.825 9.825 0 012.893 6.994c-.003 5.45-4.437 9.884-9.885 9.884m8.413-18.297A11.815 11.815 0 0012.05 0C5.495 0 .16 5.335.157 11.892c0 2.096.547 4.142 1.588 5.945L.057 24l6.305-1.654a11.882 11.882 0 005.683 1.448h.005c6.554 0 11.89-5.335 11.893-11.893a11.821 11.821 0 00-3.48-8.413z"/></svg></a>
      </div>
    </div>
  </div>
  <div class="ft-bottom">
    <span class="ft-bottom-left">©2026 CA Toolkit · All Rights Reserved · <a href="/privacy" style="color:#6B7280;text-decoration:none">Privacy Policy</a> · <span style="color:#EF4444">No refund after first upload is used</span></span>
    <span class="ft-bottom-right">Built for Indian CAs · Ludhiana, Punjab</span>
  </div>
</footer>
<div class="toast" id="toast"></div>

<script>
const dz=document.getElementById('dropzone'),fi=document.getElementById('xlFile'),dzFile=document.getElementById('dzFile');
if(dz&&fi){
  dz.addEventListener('dragover',e=>{e.preventDefault();dz.classList.add('drag');});
  dz.addEventListener('dragleave',()=>dz.classList.remove('drag'));
  dz.addEventListener('drop',e=>{e.preventDefault();dz.classList.remove('drag');
    if(e.dataTransfer.files.length){fi.files=e.dataTransfer.files;showFile(fi.files[0]);}});
  fi.addEventListener('change',()=>{if(fi.files.length)showFile(fi.files[0]);});
}
function showFile(f){dzFile.textContent='✓ '+f.name;dzFile.style.display='block';}
document.getElementById('closingYear').addEventListener('input',function(){
  const v=parseInt(this.value);if(!isNaN(v))document.getElementById('newYear').value=v+1;});
async function processFile(){
  const f=fi?fi.files[0]:null,cYr=parseInt(document.getElementById('closingYear').value),
        nYr=parseInt(document.getElementById('newYear').value),
        oNm=document.getElementById('outputName').value.trim(),
        btn=document.getElementById('processBtn'),sp=document.getElementById('spinner'),
        bt=document.getElementById('btnText'),dl=document.getElementById('dlBtn');
  if(!f){showStatus('error','✗ Please select an Excel file first.');return;}
  if(isNaN(cYr)){showStatus('error','✗ Enter a valid closing year.');return;}
  btn.disabled=true;sp.style.display='block';bt.textContent='Processing…';
  dl.style.display='none';showStatus('','');
  const fd=new FormData();
  fd.append('file',f);fd.append('closing_year',cYr);fd.append('new_year',nYr);fd.append('output_name',oNm);
  try{
    const res=await fetch('/process',{method:'POST',body:fd});
    const ct=res.headers.get('content-type')||'';
    if(!ct.includes('application/json')){
      showStatus('error','✗ Server error (non-JSON response). Please try again or contact support.');return;
    }
    const data=await res.json();
    if(data.status==='success'){
      const logHtml='<ul class="log-list">'+data.log.map(l=>`<li>${l}</li>`).join('')+'</ul>';
      showStatus('success','✓ Done! Your file is ready.'+logHtml);
      dl.href='/download/'+data.file_id+'?fn='+encodeURIComponent(data.filename);dl.download=data.filename;
      dl.textContent='⬇  Download — '+data.filename;dl.style.display='block';
      toast('Processed successfully!');
    }else{showStatus('error','✗ '+data.message);}
  }catch(e){showStatus('error','✗ Network error: '+e.message);}
  finally{btn.disabled=false;sp.style.display='none';bt.textContent='⚡ Process & Download';}
}
function showStatus(t,m){const e=document.getElementById('status');e.className=t;e.innerHTML=m;e.style.display=m?'block':'none';}
function toast(msg){const t=document.getElementById('toast');t.textContent=msg;t.classList.add('show');setTimeout(()=>t.classList.remove('show'),3000);}
</script>
<a href="https://wa.me/918427651580" target="_blank" class="wa-float" title="WhatsApp Support"><svg viewBox="0 0 24 24"><path d="M17.472 14.382c-.297-.149-1.758-.867-2.03-.967-.273-.099-.471-.148-.67.15-.197.297-.767.966-.94 1.164-.173.199-.347.223-.644.075-.297-.15-1.255-.463-2.39-1.475-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.298-.347.446-.52.149-.174.198-.298.298-.497.099-.198.05-.371-.025-.52-.075-.149-.669-1.612-.916-2.207-.242-.579-.487-.5-.669-.51-.173-.008-.371-.01-.57-.01-.198 0-.52.074-.792.372-.272.297-1.04 1.016-1.04 2.479 0 1.462 1.065 2.875 1.213 3.074.149.198 2.096 3.2 5.077 4.487.709.306 1.262.489 1.694.625.712.227 1.36.195 1.871.118.571-.085 1.758-.719 2.006-1.413.248-.694.248-1.289.173-1.413-.074-.124-.272-.198-.57-.347m-5.421 7.403h-.004a9.87 9.87 0 01-5.031-1.378l-.361-.214-3.741.982.998-3.648-.235-.374a9.86 9.86 0 01-1.51-5.26c.001-5.45 4.436-9.884 9.888-9.884 2.64 0 5.122 1.03 6.988 2.898a9.825 9.825 0 012.893 6.994c-.003 5.45-4.437 9.884-9.885 9.884m8.413-18.297A11.815 11.815 0 0012.05 0C5.495 0 .16 5.335.157 11.892c0 2.096.547 4.142 1.588 5.945L.057 24l6.305-1.654a11.882 11.882 0 005.683 1.448h.005c6.554 0 11.89-5.335 11.893-11.893a11.821 11.821 0 00-3.48-8.413z"/></svg></a>

<button class="help-btn" onclick="openHelp()" title="How to use this tool">?</button>
<div class="help-overlay" id="helpOverlay">
  <div class="help-modal">
    <div class="help-modal-head"><h3>How to Use — Balance Sheet Year Shift</h3><button class="help-close" onclick="closeHelp()">&#10005;</button></div>
    <div class="help-modal-body"><div class="help-step"><div class="help-step-num">1</div><div class="help-step-body"><h4>Upload BS File</h4><p>Click or drag-drop your comparative Excel balance sheet (.xlsx). It needs CY and PY columns with date headers like '31.03.2025'.</p></div></div><div class="help-step"><div class="help-step-num">2</div><div class="help-step-body"><h4>Enter Years</h4><p>Set Closing Year (e.g. 2025) and New Year (2026). New Year = Closing Year + 1.</p></div></div><div class="help-step"><div class="help-step-num">3</div><div class="help-step-body"><h4>Optional: Custom Filename</h4><p>Enter a custom output name, or leave blank for auto-naming.</p></div></div><div class="help-step"><div class="help-step-num">4</div><div class="help-step-body"><h4>Click Process</h4><p>The tool shifts CY→PY, clears CY columns, updates all dates, and rolls over Fixed Assets automatically.</p></div></div><div class="help-step"><div class="help-step-num">5</div><div class="help-step-body"><h4>Download</h4><p>Download the result. Your file is auto-deleted from our server within minutes.</p></div></div><div class="help-tip">✅ Works with DP Thapar, HFPL, Atultex, and most Indian CA firm templates. Formulas in PY are preserved.</div></div>
  </div>
</div>
<script>function openHelp(){document.getElementById('helpOverlay').classList.add('open')}function closeHelp(){document.getElementById('helpOverlay').classList.remove('open')}document.getElementById('helpOverlay').addEventListener('click',function(e){if(e.target===this)closeHelp()})</script>
</body></html>"""

# ══════════════════════════════════════════════════════════════════════════════
#  INCOME TAX CALCULATOR — PY 2025-26 / AY 2026-27
# ══════════════════════════════════════════════════════════════════════════════


# ══════════════════════════════════════════════════════════════════════════════
#  INCOME TAX CALCULATOR — Multi-Year: PY 2023-24 to PY 2026-27
# ══════════════════════════════════════════════════════════════════════════════

TAX_CALC_T = r"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Income Tax Calculator – CA Toolkit</title>

<style>
""" + BASE_CSS + r"""
.nav-links{display:flex;gap:20px;list-style:none}
.nav-links a{text-decoration:none;color:var(--muted);font-size:13px;font-weight:500;transition:color .2s}
.nav-links a:hover{color:var(--brand)}

.hero{text-align:center;padding:44px 24px 32px;max-width:700px;margin:0 auto}
.hero-badge{display:inline-flex;align-items:center;gap:6px;background:#FFFBEB;
            color:#92400E;border:1px solid #FDE68A;border-radius:99px;
            padding:5px 14px;font-size:12px;font-weight:600;margin-bottom:18px}
h1{font-size:clamp(22px,3.5vw,34px);font-weight:800;line-height:1.15;letter-spacing:-.5px;margin-bottom:12px}
h1 em{font-style:normal;color:var(--accent)}
.hero p{font-size:14px;color:var(--muted);line-height:1.7;max-width:520px;margin:0 auto}

.main-wrap{max-width:1200px;margin:0 auto;padding:0 24px 48px}

.regime-toggle{display:flex;justify-content:center;gap:12px;margin-bottom:28px;flex-wrap:wrap}
.regime-btn{padding:10px 28px;border-radius:10px;border:2px solid var(--border);
            background:var(--white);font-family:inherit;font-size:13px;font-weight:700;
            cursor:pointer;transition:all .2s;color:var(--muted)}
.regime-btn.active{border-color:var(--brand);background:#EFF6FF;color:var(--brand);box-shadow:0 2px 12px rgba(29,78,216,.12)}
.regime-btn:hover{border-color:var(--brand)}

.calc-grid{display:grid;grid-template-columns:1fr 1fr;gap:20px;align-items:start}
@media(max-width:860px){.calc-grid{grid-template-columns:1fr}}
.card{background:var(--white);border-radius:var(--radius);border:1px solid var(--border);
      box-shadow:var(--shadow);overflow:hidden;margin-bottom:20px}
.card-head{padding:14px 20px;border-bottom:1px solid var(--border);
           display:flex;align-items:center;gap:10px}
.card-head .icon{width:32px;height:32px;border-radius:8px;display:flex;
                 align-items:center;justify-content:center;font-size:16px}
.card-head h2{font-size:14px;font-weight:700}
.card-head p{font-size:12px;color:var(--muted);margin-top:1px}
.card-body{padding:20px}

.section-title{font-size:12px;font-weight:700;color:var(--brand);text-transform:uppercase;
               letter-spacing:.06em;margin:16px 0 10px;padding-bottom:6px;
               border-bottom:1px solid var(--border);display:flex;align-items:center;gap:6px}
.section-title:first-child{margin-top:0}
.field{margin-bottom:12px}
label{display:block;font-size:11px;font-weight:600;text-transform:uppercase;
      letter-spacing:.04em;color:var(--muted);margin-bottom:4px}
.hint{font-size:10px;color:var(--muted);margin-top:3px;font-style:italic}
input[type=number]{width:100%;border:1.5px solid var(--border);border-radius:8px;
    padding:9px 12px;font-family:inherit;font-size:13px;color:var(--ink);
    background:var(--white);outline:none;transition:border-color .2s}
input[type=number]:focus{border-color:var(--brand)}
input[type=number]::-webkit-outer-spin-button,
input[type=number]::-webkit-inner-spin-button{-webkit-appearance:none;margin:0}
input[type=number]{-moz-appearance:textfield}
select{width:100%;border:1.5px solid var(--border);border-radius:8px;padding:9px 12px;
       font-family:inherit;font-size:13px;color:var(--ink);background:var(--white);outline:none}
.row2{display:grid;grid-template-columns:1fr 1fr;gap:10px}
.row3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px}
@media(max-width:480px){.row2,.row3{grid-template-columns:1fr}}

.btn-calc{width:100%;background:var(--brand);color:#fff;border:none;border-radius:10px;
          padding:13px;font-family:inherit;font-size:14px;font-weight:700;cursor:pointer;
          transition:all .2s;display:flex;align-items:center;justify-content:center;gap:8px;
          margin-top:8px}
.btn-calc:hover{background:var(--brand-d);transform:translateY(-1px);box-shadow:0 4px 16px rgba(29,78,216,.2)}
.btn-reset{width:100%;background:#F3F4F6;color:var(--ink);border:none;border-radius:10px;
           padding:10px;font-family:inherit;font-size:13px;font-weight:600;cursor:pointer;
           transition:background .2s;margin-top:8px}
.btn-reset:hover{background:#E5E7EB}

.result-panel{display:none}
.result-panel.show{display:block}
.result-row{display:flex;justify-content:space-between;align-items:center;
            padding:10px 0;border-bottom:1px solid var(--border);font-size:13px}
.result-row:last-child{border-bottom:none}
.result-row .lbl{color:var(--muted);font-weight:500}
.result-row .val{font-weight:700;color:var(--ink);text-align:right}
.result-row.total{padding:14px 0;font-size:15px}
.result-row.total .lbl{color:var(--ink);font-weight:800}
.result-row.total .val{color:var(--brand);font-size:17px}
.result-row.refund .val{color:var(--green)}
.result-row.payable .val{color:var(--red)}
.result-row.sub{font-size:12px;padding:6px 0}
.result-row.sub .lbl{padding-left:16px;font-size:11px}
.result-row.sub .val{font-size:12px}

.compare-box{background:linear-gradient(135deg,#EFF6FF,#FFFBEB);border:2px solid var(--brand);
             border-radius:var(--radius);padding:20px;text-align:center;margin-top:16px}
.compare-box h3{font-size:14px;font-weight:800;margin-bottom:6px}
.compare-box .savings{font-size:28px;font-weight:800;color:var(--green);margin:8px 0}
.compare-box .regime-winner{font-size:13px;color:var(--muted)}
.compare-box .regime-winner strong{color:var(--ink)}
.compare-table{width:100%;margin-top:14px;font-size:12px;border-collapse:collapse}
.compare-table th{text-align:center;font-size:10px;text-transform:uppercase;letter-spacing:.06em;
                  color:var(--muted);padding:6px 8px;border-bottom:1.5px solid var(--border)}
.compare-table td{text-align:center;padding:8px;border-bottom:1px solid var(--border);font-weight:600}
.compare-table .winner{background:#ECFDF5;color:#065F46;border-radius:6px}

.slab-table{width:100%;font-size:12px;border-collapse:collapse;margin-top:8px}
.slab-table th{text-align:left;font-size:10px;text-transform:uppercase;letter-spacing:.06em;
               color:var(--muted);padding:6px 8px;border-bottom:1.5px solid var(--border)}
.slab-table td{padding:7px 8px;border-bottom:1px solid var(--border);font-size:12px}
.slab-table tr:last-child td{border-bottom:none}
.slab-table .amt{text-align:right;font-weight:700;font-family:'Inter',monospace}

.disclaimer{font-size:11px;color:var(--muted);line-height:1.6;margin-top:16px;
            padding:12px;background:#F9FAFB;border-radius:8px;border:1px solid var(--border)}
.disclaimer.future{background:#FFFBEB;border-color:#FDE68A}
.toast{position:fixed;bottom:24px;right:24px;background:var(--ink);color:#fff;
       padding:11px 18px;border-radius:10px;font-size:13px;font-weight:500;
       transform:translateY(80px);transition:transform .3s;z-index:999}
.toast.show{transform:translateY(0)}

.print-btn{display:inline-flex;align-items:center;gap:5px;background:#F3F4F6;color:var(--ink);
           border:1px solid var(--border);border-radius:8px;padding:7px 14px;font-size:12px;
           font-weight:600;cursor:pointer;font-family:inherit;transition:all .2s;margin-top:8px}
.print-btn:hover{background:#E5E7EB}

.year-pills{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:6px}
.year-pill{padding:6px 14px;border-radius:8px;border:1.5px solid var(--border);background:var(--white);
           font-size:12px;font-weight:600;cursor:pointer;transition:all .2s;color:var(--muted);font-family:inherit}
.year-pill.active{border-color:var(--brand);background:#EFF6FF;color:var(--brand)}
.year-pill:hover{border-color:var(--brand)}
.year-pill .future-tag{font-size:9px;background:#FDE68A;color:#92400E;padding:1px 5px;border-radius:4px;margin-left:4px;font-weight:700}

.at-group-label{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:var(--muted);margin:12px 0 6px;padding:3px 8px;background:var(--surface);border-radius:5px;display:inline-block}
.at-btn-row{display:flex;flex-wrap:wrap;gap:6px;margin-bottom:4px}
.at-btn{padding:7px 13px;border:1.5px solid var(--border);border-radius:20px;background:#fff;font-size:12px;font-weight:500;color:var(--ink);cursor:pointer;transition:all .15s;white-space:nowrap;font-family:inherit}
.at-btn:hover{border-color:var(--brand);color:var(--brand);background:#EFF6FF}
.at-btn.active{border-color:var(--brand);background:var(--brand);color:#fff;font-weight:700;box-shadow:0 2px 8px rgba(37,99,235,.25)}

.mat-row:last-child{border-bottom:none}
.mat-row .ml{color:#78350F;font-weight:500}
.mat-row .mv{font-weight:700;color:#92400E}
.mat-row.mt{font-size:13px;padding:10px 0;border-top:1.5px solid #FDE68A;margin-top:4px}
.mat-row.mt .ml{color:#451A03;font-weight:800}
.mat-row.mt .mv{color:#B45309;font-size:14px}
.mat-badge{display:inline-flex;align-items:center;gap:5px;padding:4px 12px;border-radius:20px;font-size:11px;font-weight:700;margin-top:10px}
.mat-badge.normal{background:#ECFDF5;color:#065F46}
.mat-badge.mat{background:#FEF3C7;color:#92400E}

.at-table{width:100%;border-collapse:collapse;font-size:12px;margin-bottom:6px}
.at-table th{text-align:left;font-size:10px;text-transform:uppercase;letter-spacing:.06em;color:var(--muted);padding:7px 10px;border-bottom:2px solid var(--border);background:#F9FAFB}
.at-table td{padding:10px 10px;border-bottom:1px solid var(--border);vertical-align:top}
.at-table tr:last-child td{border-bottom:none}
.at-table .due{font-weight:700;color:var(--brand)}
.at-table .pct{font-weight:800;font-size:14px;color:#1E40AF;text-align:center}
.at-table .cumul{font-size:11px;color:var(--muted)}
.at-table .amt-cell{font-weight:700;color:var(--ink);text-align:right}
.at-table tr.overdue td{background:#FEF2F2}
.at-table tr.upcoming td{background:#FFFBEB}
.at-table tr.done td{background:#F0FDF4}

.assessee-badge{display:inline-flex;align-items:center;gap:5px;padding:4px 12px;border-radius:20px;font-size:11px;font-weight:700;background:#EFF6FF;color:#1E40AF;margin-bottom:10px}

/* ═══════════════════════════════════════════
   ANIMATIONS & MICRO-INTERACTIONS
   ═══════════════════════════════════════════ */

/* ── Scroll-reveal cards ── */
.reveal{opacity:0;transform:translateY(28px);transition:opacity .55s cubic-bezier(.22,1,.36,1),transform .55s cubic-bezier(.22,1,.36,1)}
.reveal.visible{opacity:1;transform:translateY(0)}
.reveal-delay-1{transition-delay:.08s}
.reveal-delay-2{transition-delay:.16s}
.reveal-delay-3{transition-delay:.24s}
.reveal-delay-4{transition-delay:.32s}

/* ── Result panel slide-in ── */
@keyframes slideUp{from{opacity:0;transform:translateY(32px)}to{opacity:1;transform:translateY(0)}}
@keyframes fadeIn{from{opacity:0}to{opacity:1}}
.result-panel.show{animation:slideUp .5s cubic-bezier(.22,1,.36,1) forwards}
.result-row{animation:fadeIn .3s ease both}

/* ── Calculate button states ── */
.btn-calc{position:relative;overflow:hidden}
.btn-calc .btn-text{transition:opacity .2s}
.btn-calc .btn-spinner{position:absolute;display:none;width:18px;height:18px;border:2px solid rgba(255,255,255,.3);border-top-color:#fff;border-radius:50%;animation:spin .7s linear infinite}
@keyframes spin{to{transform:rotate(360deg)}}
.btn-calc.loading .btn-text{opacity:0}
.btn-calc.loading .btn-spinner{display:block}
.btn-calc .ripple{position:absolute;border-radius:50%;background:rgba(255,255,255,.35);transform:scale(0);animation:rippleAnim .55s linear;pointer-events:none}
@keyframes rippleAnim{to{transform:scale(4);opacity:0}}

/* ── Progress bar ── */
#calcProgress{position:fixed;top:0;left:0;width:0;height:3px;background:linear-gradient(90deg,var(--brand),#60A5FA,var(--green));z-index:9999;transition:width .35s ease;border-radius:0 3px 3px 0;box-shadow:0 0 8px rgba(37,99,235,.5)}

/* ── Number counter ── */
.count-anim{display:inline-block;transition:transform .1s}

/* ── Tax Donut Chart ── */
#taxChartWrap{margin-top:20px;padding:16px;background:#F9FAFB;border-radius:12px;border:1px solid var(--border)}
#taxChartWrap h3{font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.05em;color:var(--muted);margin-bottom:14px;text-align:center}
.donut-container{display:flex;align-items:center;gap:20px;flex-wrap:wrap;justify-content:center}
.donut-svg{flex-shrink:0;filter:drop-shadow(0 4px 12px rgba(0,0,0,.08))}
.donut-legend{display:flex;flex-direction:column;gap:8px;min-width:140px}
.donut-legend-item{display:flex;align-items:center;gap:8px;font-size:12px}
.donut-dot{width:10px;height:10px;border-radius:50%;flex-shrink:0}
.donut-label{color:var(--muted);font-weight:500}
.donut-val{font-weight:700;color:var(--ink);margin-left:auto}
.donut-segment{transition:stroke-dasharray .8s cubic-bezier(.22,1,.36,1),stroke-dashoffset .8s cubic-bezier(.22,1,.36,1)}

/* ── Regime bar chart ── */
#regimeChartWrap{margin-top:16px;padding:16px;background:#F9FAFB;border-radius:12px;border:1px solid var(--border)}
#regimeChartWrap h3{font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.05em;color:var(--muted);margin-bottom:14px;text-align:center}
.bar-chart{display:flex;flex-direction:column;gap:10px}
.bar-row{display:flex;align-items:center;gap:10px;font-size:12px}
.bar-label{width:90px;font-weight:600;color:var(--muted);font-size:11px;text-align:right;flex-shrink:0}
.bar-track{flex:1;height:22px;background:#E5E7EB;border-radius:6px;overflow:hidden;position:relative}
.bar-fill{height:100%;border-radius:6px;width:0;transition:width 1s cubic-bezier(.22,1,.36,1);display:flex;align-items:center;justify-content:flex-end;padding-right:8px}
.bar-fill span{font-size:10px;font-weight:700;color:#fff;white-space:nowrap}
.bar-val{width:90px;font-weight:700;font-size:11px;color:var(--ink);flex-shrink:0}

/* ── Confetti ── */
.confetti-piece{position:fixed;width:8px;height:8px;top:-10px;border-radius:2px;pointer-events:none;z-index:9998;animation:confettiFall linear forwards}
@keyframes confettiFall{0%{transform:translateY(0) rotate(0deg);opacity:1}100%{transform:translateY(110vh) rotate(720deg);opacity:0}}

/* ── Assessee button pop ── */
.at-btn{transition:all .18s cubic-bezier(.34,1.56,.64,1)}
.at-btn.active{transform:scale(1.05)}
.at-btn:active{transform:scale(.95)}

/* ── Input focus ring glow ── */
input[type=number]:focus{border-color:var(--brand);box-shadow:0 0 0 3px rgba(37,99,235,.12)}
select:focus{border-color:var(--brand);box-shadow:0 0 0 3px rgba(37,99,235,.12);outline:none}

/* ── Card hover lift ── */
.card{transition:box-shadow .25s,transform .25s}
.card:hover{box-shadow:0 8px 32px rgba(0,0,0,.10);transform:translateY(-2px)}

/* ── Year pill bounce ── */
.year-pill{transition:all .2s cubic-bezier(.34,1.56,.64,1)}
.year-pill.active{transform:scale(1.06)}

/* ── Regime btn pop ── */
.regime-btn{transition:all .2s cubic-bezier(.34,1.56,.64,1)}
.regime-btn.active{transform:scale(1.04)}

/* ── Result row stagger ── */
@keyframes rowIn{from{opacity:0;transform:translateX(-10px)}to{opacity:1;transform:translateX(0)}}
.result-row{animation:rowIn .3s ease both}

/* ── Total row pulse ── */
@keyframes totalPulse{0%{transform:scale(1)}50%{transform:scale(1.02)}100%{transform:scale(1)}}
.result-row.total{animation:rowIn .4s ease both, totalPulse .4s ease .5s}

/* ── Toast slide & bounce ── */
@keyframes toastIn{0%{transform:translateY(80px) scale(.9)}70%{transform:translateY(-4px) scale(1.02)}100%{transform:translateY(0) scale(1)}}
.toast.show{animation:toastIn .4s cubic-bezier(.34,1.56,.64,1) forwards}

/* ── Hero text shimmer on load ── */
@keyframes shimmer{0%{background-position:200% center}100%{background-position:-200% center}}
.hero-shimmer{background:linear-gradient(90deg,var(--brand) 0%,#60A5FA 40%,var(--brand) 80%);background-size:200% auto;-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;animation:shimmer 3s linear infinite}

/* ── Nav scroll shadow ── */
nav{transition:box-shadow .3s}
nav.scrolled{box-shadow:0 4px 24px rgba(0,0,0,.10)}

@media print{nav,footer,.hero,.regime-toggle,.card:first-child,.btn-calc,.btn-reset,.print-btn,.toast,.year-pills,#calcProgress{display:none!important}
             .result-panel{display:block!important}.calc-grid{display:block!important}
             .card{box-shadow:none!important;border:1px solid #ccc!important}}
</style></head><body>

<div id="calcProgress"></div>
<nav>
  <a href="/" class="logo">CA<span class="logo-dot"></span><span>Toolkit</span></a>
  <ul class="nav-links">
    <li><a href="#input">Calculator</a></li>
    <li><a href="#result-section">Results</a></li>
  </ul>
  <div class="nav-right">
    {% if username %}<div class="nav-user"><span class="nav-avatar">{{ username[0].upper() }}</span><strong>{{ username }}</strong><span class="badge b-{{ plan }}">{{ plan_label }}</span>{% if is_admin %}<span class="badge" style="background:#EDE9FE;color:#5B21B6">Admin</span>{% endif %}</div><div class="nav-sep"></div>{% if is_admin %}<a href="/admin" class="nav-btn ghost">⚙ Admin</a>{% endif %}{% endif %}
    <a href="/" class="nav-btn dash">⬅ Dashboard</a>
    {% if username %}<a href="/logout" class="nav-link">Sign out</a>{% else %}<a href="/login" class="nav-btn">Sign In →</a>{% endif %}
  </div>
</nav>

<section class="hero">
  <div class="hero-badge">🧮 Multi-Year · PY 2023-24 to PY 2026-27</div>
  <h1>Income Tax <em class="hero-shimmer">Calculator</em></h1>
  <p>Calculate tax under Old &amp; New Regime for any year from PY 2023-24 to PY 2026-27. Income under 5 heads, deductions, TDS/TCS — instant comparison with slab-wise breakup.</p>
</section>

<div class="main-wrap">

<!-- Regime Toggle -->
<div class="regime-toggle">
  <button class="regime-btn active" onclick="setRegime('new')" id="btn-new">🆕 New Regime (Default)</button>
  <button class="regime-btn" onclick="setRegime('old')" id="btn-old">📜 Old Regime</button>
  <button class="regime-btn" onclick="setRegime('both')" id="btn-both">⚖️ Compare Both</button>
</div>

<div class="calc-grid" id="input">
  <!-- LEFT: Input Section -->
  <div>
    <!-- ──── BASIC INFO ──── -->
    <div class="card reveal reveal-delay-1">
      <div class="card-head">
        <div class="icon" style="background:#EFF6FF">👤</div>
        <div><h2>Basic Information</h2><p>Assessee details &amp; Assessment Year</p></div>
      </div>
      <div class="card-body">
        <div class="field">
          <label>Assessment Year</label>
          <div class="year-pills" id="yearPills">
            <button class="year-pill" onclick="setYear('2023-24')">PY 2023-24<br/><span style="font-size:10px;font-weight:400;color:var(--muted)">AY 2024-25</span></button>
            <button class="year-pill" onclick="setYear('2024-25')">PY 2024-25<br/><span style="font-size:10px;font-weight:400;color:var(--muted)">AY 2025-26</span></button>
            <button class="year-pill active" onclick="setYear('2025-26')">PY 2025-26<br/><span style="font-size:10px;font-weight:400;color:var(--muted)">AY 2026-27</span></button>
            <button class="year-pill" onclick="setYear('2026-27')">PY 2026-27<br/><span style="font-size:10px;font-weight:400;color:var(--muted)">AY 2027-28</span><span class="future-tag">Upcoming</span></button>
          </div>
          <div id="futureYearNote" style="display:none;margin-top:6px;padding:8px 12px;background:#FFFBEB;border:1px solid #FDE68A;border-radius:8px;font-size:11px;color:#92400E;font-weight:500">
            ⚠️ PY 2026-27 rates are based on Union Budget 2026 (no changes from PY 2025-26). Final rates subject to any future amendments.
          </div>
        </div>
        <div class="field">
          <label>Assessee Name <span style="font-weight:400;text-transform:none">(optional)</span></label>
          <input type="text" id="assesseeName" placeholder="e.g. Rajesh Kumar / ABC Pvt Ltd" style="border:1.5px solid var(--border);border-radius:8px;padding:9px 12px;font-family:inherit;font-size:13px;width:100%"/>
        </div>

        <!-- ── TYPE OF ASSESSEE — visible button grid ── -->
        <div class="field">
          <label>Type of Assessee</label>
          <!-- hidden select keeps existing JS logic intact -->
          <select id="assesseeType" onchange="onAssesseeTypeChange()" style="display:none">
            <option value="individual_below60" selected>Individual – Below 60 yrs</option>
            <option value="individual_senior">Individual – Senior Citizen (60–80)</option>
            <option value="individual_supersenior">Individual – Super Senior Citizen (80+)</option>
            <option value="individual_nri">Individual – Non-Resident (NRI)</option>
            <option value="huf">HUF (Hindu Undivided Family)</option>
            <option value="firm">Partnership Firm / LLP</option>
            <option value="company_domestic">Domestic Company</option>
            <option value="company_foreign">Foreign Company</option>
            <option value="company_mfg_new">Domestic Company – New Mfg (Sec 115BAB)</option>
            <option value="company_small">Domestic Company – Small (≤ ₹400 Cr, Sec 115BA)</option>
            <option value="aop_boi">AOP / BOI</option>
            <option value="cooperative">Co-operative Society</option>
            <option value="trust_aop">Trust / AOP (Registered)</option>
            <option value="local_authority">Local Authority</option>
            <option value="artificial_person">Artificial Juridical Person</option>
          </select>

          <!-- Group: Individuals -->
          <div class="at-group-label">👤 Individual</div>
          <div class="at-btn-row">
            <button class="at-btn active" data-val="individual_below60" onclick="selectAssessee(this,'individual_below60')">Below 60 yrs</button>
            <button class="at-btn" data-val="individual_senior" onclick="selectAssessee(this,'individual_senior')">Senior (60–80)</button>
            <button class="at-btn" data-val="individual_supersenior" onclick="selectAssessee(this,'individual_supersenior')">Super Senior (80+)</button>
            <button class="at-btn" data-val="individual_nri" onclick="selectAssessee(this,'individual_nri')">NRI</button>
          </div>

          <!-- Group: HUF -->
          <div class="at-group-label">🏠 HUF</div>
          <div class="at-btn-row">
            <button class="at-btn" data-val="huf" onclick="selectAssessee(this,'huf')">HUF (Hindu Undivided Family)</button>
          </div>

          <!-- Group: Firm / LLP -->
          <div class="at-group-label">🤝 Firm / LLP</div>
          <div class="at-btn-row">
            <button class="at-btn" data-val="firm" onclick="selectAssessee(this,'firm')">Partnership Firm / LLP</button>
          </div>

          <!-- Group: Companies -->
          <div class="at-group-label">🏢 Company</div>
          <div class="at-btn-row">
            <button class="at-btn" data-val="company_domestic" onclick="selectAssessee(this,'company_domestic')">Domestic Co.</button>
            <button class="at-btn" data-val="company_foreign" onclick="selectAssessee(this,'company_foreign')">Foreign Co.</button>
            <button class="at-btn" data-val="company_mfg_new" onclick="selectAssessee(this,'company_mfg_new')">New Mfg Co. (115BAB)</button>
            <button class="at-btn" data-val="company_small" onclick="selectAssessee(this,'company_small')">Small Co. (115BA)</button>
          </div>

          <!-- Group: Others -->
          <div class="at-group-label">🏛️ Others</div>
          <div class="at-btn-row">
            <button class="at-btn" data-val="aop_boi" onclick="selectAssessee(this,'aop_boi')">AOP / BOI</button>
            <button class="at-btn" data-val="cooperative" onclick="selectAssessee(this,'cooperative')">Co-operative Society</button>
            <button class="at-btn" data-val="trust_aop" onclick="selectAssessee(this,'trust_aop')">Trust</button>
            <button class="at-btn" data-val="local_authority" onclick="selectAssessee(this,'local_authority')">Local Authority</button>
            <button class="at-btn" data-val="artificial_person" onclick="selectAssessee(this,'artificial_person')">Artificial Juridical Person</button>
          </div>
        </div>

        <!-- Individual-only fields -->
        <div id="individualFields">
          <div class="row2">
            <div class="field">
              <label>Age Category</label>
              <div id="ageBadge" style="padding:8px 14px;background:#EFF6FF;border-radius:8px;font-size:12px;font-weight:600;color:#1E40AF;display:inline-block">Below 60 years</div>
            </div>
            <div class="field">
              <label>Residential Status</label>
              <select id="residentialStatus">
                <option value="resident">Resident</option>
                <option value="nri">Non-Resident</option>
              </select>
            </div>
          </div>
        </div>

        <!-- Company-specific info -->
        <div id="companyFields" style="display:none">
          <div id="matInfo" style="padding:10px 14px;background:#EFF6FF;border:1px solid #BFDBFE;border-radius:8px;font-size:11px;color:#1E40AF;margin-top:4px">
            <strong>ℹ️ MAT u/s 115JB:</strong> Tax payable is higher of normal tax or 15% of Book Profit (+ surcharge + cess). Enter Book Profit in the MAT section below.
          </div>
        </div>
        <!-- Firm info -->
        <div id="firmFields" style="display:none">
          <div style="padding:10px 14px;background:#F0FDF4;border:1px solid #BBF7D0;border-radius:8px;font-size:11px;color:#065F46;margin-top:4px">
            <strong>ℹ️ Firm / LLP:</strong> Flat 30% on total income + surcharge (12% if income &gt; ₹1 Cr) + cess 4%. AMT @ 18.5% of Adjusted Total Income applies u/s 115JC.
          </div>
        </div>
        <!-- AOP / Co-op info -->
        <div id="aopFields" style="display:none">
          <div style="padding:10px 14px;background:#FFF7ED;border:1px solid #FED7AA;border-radius:8px;font-size:11px;color:#92400E;margin-top:4px">
            <strong>ℹ️ AOP / BOI / Co-op / Trust:</strong> Taxed at applicable slab rates as per specific provisions. AMT @ 18.5% may apply u/s 115JC.
          </div>
        </div>
      </div>
    </div>

    <!-- ──── INCOME HEADS (Dynamic per assessee) ──── -->
    <div class="card">
      <div class="card-head">
        <div class="icon" style="background:#F0FDF4">💰</div>
        <div><h2>Income Details</h2><p id="incomeCardSub">Gross Total Income computation</p></div>
      </div>
      <div class="card-body">

        <!-- ═══ INDIVIDUAL / HUF / AOP / TRUST (all 5 heads) ═══ -->
        <div id="inc_individual">
          <div class="section-title">📋 1. Income from Salary</div>
          <div class="row2">
            <div class="field">
              <label>Gross Salary</label>
              <input type="number" id="grossSalary" placeholder="0" min="0"/>
            </div>
            <div class="field">
              <label>Exempt Allowances (HRA, LTA etc.)</label>
              <input type="number" id="exemptAllow" placeholder="0" min="0"/>
              <p class="hint">Old regime: HRA, LTA etc. New regime: mostly nil</p>
            </div>
          </div>
          <div class="field">
            <label>Standard Deduction</label>
            <input type="number" id="stdDeduction" value="75000" placeholder="75000" min="0"/>
            <p class="hint" id="stdDedHint">₹75,000 for PY 2024-25 onwards. ₹50,000 for PY 2023-24.</p>
          </div>

          <div class="section-title">🏠 2. Income from House Property</div>
          <div class="row2">
            <div class="field">
              <label>Net Annual Value / Rental Income</label>
              <input type="number" id="houseIncome" placeholder="0"/>
              <p class="hint">Can be negative for self-occupied (loss)</p>
            </div>
            <div class="field">
              <label>Interest on Home Loan (Sec 24b)</label>
              <input type="number" id="homeLoanInterest" placeholder="0" min="0"/>
              <p class="hint">Max ₹2L self-occupied</p>
            </div>
          </div>

          <div class="section-title">💼 3. Profits from Business / Profession</div>
          <div class="field">
            <label>Net Profit from Business / Profession</label>
            <input type="number" id="businessIncome" placeholder="0"/>
            <p class="hint">After all business deductions. Can be loss (negative)</p>
          </div>

          <div class="section-title">📈 4. Capital Gains</div>
          <div id="transitionalNotice" style="display:none;margin-bottom:12px;padding:10px 14px;background:#FFFBEB;border:1px solid #FDE68A;border-radius:8px;font-size:11px;color:#92400E;line-height:1.6">
            <strong>⚠️ PY 2024-25 Transitional Year:</strong> Budget 2024 changed capital gains rates from 23 July 2024. Enter gains sold <strong>before 23 July</strong> separately from gains sold <strong>on/after 23 July</strong>.
          </div>
          <div id="preJulyFields" style="display:none">
            <div style="font-size:11px;font-weight:700;color:#92400E;text-transform:uppercase;letter-spacing:.05em;margin:8px 0 8px;padding:4px 10px;background:#FFFBEB;border-radius:6px;display:inline-block">📅 Pre-23 July 2024 (Old Rates)</div>
            <div class="row2">
              <div class="field">
                <label>STCG 111A — Pre-July (equity) @ 15%</label>
                <input type="number" id="stcg111aPreJuly" placeholder="0" min="0"/>
              </div>
              <div class="field">
                <label>LTCG 112A — Pre-July (equity) @ 10%</label>
                <input type="number" id="ltcg112aPreJuly" placeholder="0" min="0"/>
                <p class="hint">Exempt up to ₹1 lakh (old limit)</p>
              </div>
            </div>
            <div class="row2">
              <div class="field">
                <label>LTCG Other — Pre-July @ 20%</label>
                <input type="number" id="ltcgOtherPreJuly" placeholder="0" min="0"/>
                <p class="hint">With indexation benefit (pre-July rule)</p>
              </div>
              <div class="field"></div>
            </div>
            <div style="font-size:11px;font-weight:700;color:var(--brand);text-transform:uppercase;letter-spacing:.05em;margin:12px 0 8px;padding:4px 10px;background:#EFF6FF;border-radius:6px;display:inline-block">📅 Post-23 July 2024 (New Rates)</div>
          </div>
          <div class="row2">
            <div class="field">
              <label id="stcg111aLabel">STCG u/s 111A (equity, STT paid)</label>
              <input type="number" id="stcg111a" placeholder="0" min="0"/>
              <p class="hint" id="stcgRateHint">Rate varies by year</p>
            </div>
            <div class="field">
              <label>STCG — Other (non-equity)</label>
              <input type="number" id="stcgOther" placeholder="0" min="0"/>
              <p class="hint">Taxed at slab rates</p>
            </div>
          </div>
          <div class="row2">
            <div class="field">
              <label id="ltcg112aLabel">LTCG u/s 112A (equity, STT paid)</label>
              <input type="number" id="ltcg112a" placeholder="0" min="0"/>
              <p class="hint" id="ltcgRateHint">Rate &amp; exemption varies by year</p>
            </div>
            <div class="field">
              <label id="ltcgOtherLabel">LTCG — Other (property, debt etc.)</label>
              <input type="number" id="ltcgOther" placeholder="0" min="0"/>
              <p class="hint" id="ltcgOtherHint">Rate varies by year</p>
            </div>
          </div>

          <div id="bfLossSection" style="margin-top:4px;padding:12px 14px;background:#FFF7ED;border:1px solid #FED7AA;border-radius:10px">
            <div style="font-size:11.5px;font-weight:700;color:#92400E;margin-bottom:10px;display:flex;align-items:center;gap:6px">
              ↩️ Brought Forward Capital Losses (B/F C/F Losses)
              <span style="font-size:10px;font-weight:500;color:#B45309;background:#FEF3C7;padding:2px 7px;border-radius:99px">Set-off u/s 74</span>
            </div>
            <div class="row2">
              <div class="field">
                <label>B/F Short-Term Capital Loss (STCL)</label>
                <input type="number" id="bf_stcl" placeholder="0" min="0"/>
                <p class="hint">Enter as positive. Can be set off against any STCG or LTCG.</p>
              </div>
              <div class="field">
                <label>B/F Long-Term Capital Loss (LTCL)</label>
                <input type="number" id="bf_ltcl" placeholder="0" min="0"/>
                <p class="hint">Enter as positive. Can only be set off against LTCG.</p>
              </div>
            </div>
          </div>

          <div class="section-title">📦 5. Income from Other Sources</div>
          <div class="row2">
            <div class="field">
              <label>Interest / Dividends / Other Income</label>
              <input type="number" id="otherIncome" placeholder="0" min="0"/>
            </div>
            <div class="field">
              <label>Winnings (lottery, games etc.)</label>
              <input type="number" id="winningsIncome" placeholder="0" min="0"/>
              <p class="hint">Taxed at 30% flat</p>
            </div>
          </div>
        </div>

        <!-- ═══ COMPANY ═══ -->
        <div id="inc_company" style="display:none">
          <div style="margin-bottom:14px;padding:10px 14px;background:#EFF6FF;border:1px solid #BFDBFE;border-radius:8px;font-size:11px;color:#1E40AF">
            <strong>ℹ️ Company Income:</strong> Enter net taxable income computed as per IT Act provisions (after all allowable business deductions, depreciation, etc.).
          </div>

          <div class="section-title">💼 Business / Profession Income</div>
          <div class="field">
            <label>Net Taxable Business Income (₹)</label>
            <input type="number" id="co_businessIncome" placeholder="0"/>
            <p class="hint">Net profit after all IT Act allowable deductions &amp; depreciation</p>
          </div>

          <div class="section-title">📈 Capital Gains</div>
          <div class="row2">
            <div class="field">
              <label>STCG u/s 111A (equity, STT paid)</label>
              <input type="number" id="co_stcg111a" placeholder="0" min="0"/>
            </div>
            <div class="field">
              <label>STCG — Other assets</label>
              <input type="number" id="co_stcgOther" placeholder="0" min="0"/>
            </div>
          </div>
          <div class="row2">
            <div class="field">
              <label>LTCG u/s 112A (equity)</label>
              <input type="number" id="co_ltcg112a" placeholder="0" min="0"/>
            </div>
            <div class="field">
              <label>LTCG — Other assets</label>
              <input type="number" id="co_ltcgOther" placeholder="0" min="0"/>
            </div>
          </div>

          <div class="section-title">📦 Other Sources</div>
          <div class="field">
            <label>Interest / Dividend / Other Income (₹)</label>
            <input type="number" id="co_otherIncome" placeholder="0" min="0"/>
          </div>

          <div class="section-title">⚖️ MAT — Book Profit (Sec 115JB)</div>
          <div class="row2">
            <div class="field">
              <label>Book Profit u/s 115JB (₹)</label>
              <input type="number" id="co_bookProfit" placeholder="0" min="0" oninput="syncMatBookProfit()"/>
              <p class="hint">Net profit per P&amp;L + mandatory add-backs as per Sch VII</p>
            </div>
            <div class="field">
              <label>Turnover / Gross Receipts (₹)</label>
              <input type="number" id="co_turnover" placeholder="0" min="0"/>
              <p class="hint">For reference / threshold checks</p>
            </div>
          </div>
        </div>

        <!-- ═══ FIRM / LLP ═══ -->
        <div id="inc_firm" style="display:none">
          <div style="margin-bottom:14px;padding:10px 14px;background:#F0FDF4;border:1px solid #BBF7D0;border-radius:8px;font-size:11px;color:#065F46">
            <strong>ℹ️ Firm / LLP Income:</strong> Firms are not allowed salary to partners for tax purposes beyond limits. Enter total firm income after all deductions. Remuneration &amp; interest to partners allowed u/s 40(b) already deducted.
          </div>

          <div class="section-title">💼 Business / Profession Income</div>
          <div class="field">
            <label>Net Taxable Business Income (₹)</label>
            <input type="number" id="firm_businessIncome" placeholder="0"/>
            <p class="hint">Firm profit after partner remuneration / interest u/s 40(b)</p>
          </div>

          <div class="section-title">📈 Capital Gains</div>
          <div class="row2">
            <div class="field">
              <label>STCG (equity u/s 111A)</label>
              <input type="number" id="firm_stcg111a" placeholder="0" min="0"/>
            </div>
            <div class="field">
              <label>LTCG (equity u/s 112A)</label>
              <input type="number" id="firm_ltcg112a" placeholder="0" min="0"/>
            </div>
          </div>
          <div class="row2">
            <div class="field">
              <label>STCG — Other</label>
              <input type="number" id="firm_stcgOther" placeholder="0" min="0"/>
            </div>
            <div class="field">
              <label>LTCG — Other</label>
              <input type="number" id="firm_ltcgOther" placeholder="0" min="0"/>
            </div>
          </div>

          <div class="section-title">📦 Other Sources</div>
          <div class="field">
            <label>Interest / Other Income (₹)</label>
            <input type="number" id="firm_otherIncome" placeholder="0" min="0"/>
          </div>

          <div class="section-title">⚖️ AMT — Adjusted Total Income (Sec 115JC)</div>
          <div class="field">
            <label>Adjusted Total Income for AMT (₹)</label>
            <input type="number" id="firm_amtAti" placeholder="0" min="0" oninput="syncAmtAti()"/>
            <p class="hint">GTI + deductions claimed u/s 10AA / 35AD / 80H–80RRB added back</p>
          </div>
        </div>

        <!-- ═══ CO-OPERATIVE SOCIETY ═══ -->
        <div id="inc_coop" style="display:none">
          <div style="margin-bottom:14px;padding:10px 14px;background:#FFF7ED;border:1px solid #FED7AA;border-radius:8px;font-size:11px;color:#92400E">
            <strong>ℹ️ Co-operative Society:</strong> Taxed on slab — 10% up to ₹10K / 20% up to ₹20K / 30% above. Surcharge 12% if income &gt; ₹1 Cr. Cess 4%.
          </div>
          <div class="section-title">💼 Business Income</div>
          <div class="field">
            <label>Net Taxable Income (₹)</label>
            <input type="number" id="coop_businessIncome" placeholder="0"/>
          </div>
          <div class="section-title">📦 Other Sources</div>
          <div class="field">
            <label>Interest / Dividend / Other (₹)</label>
            <input type="number" id="coop_otherIncome" placeholder="0" min="0"/>
          </div>
        </div>

        <!-- ═══ LOCAL AUTHORITY ═══ -->
        <div id="inc_local" style="display:none">
          <div style="margin-bottom:14px;padding:10px 14px;background:#F5F3FF;border:1px solid #DDD6FE;border-radius:8px;font-size:11px;color:#4C1D95">
            <strong>ℹ️ Local Authority:</strong> Flat 30% on total income + 4% cess. No surcharge.
          </div>
          <div class="section-title">💼 Business / Property / Other Income</div>
          <div class="field">
            <label>Net Taxable Income (₹)</label>
            <input type="number" id="local_income" placeholder="0"/>
          </div>
        </div>

      </div>
    </div>

    <!-- ──── DEDUCTIONS (OLD REGIME) ──── -->
    <div class="card" id="deductions-card">
      <div class="card-head">
        <div class="icon" style="background:#FFFBEB">🧾</div>
        <div><h2>Deductions (Chapter VI-A)</h2><p>Applicable in Old Regime only (except 80CCD(2))</p></div>
      </div>
      <div class="card-body">
        <div class="row2">
          <div class="field">
            <label>80C (PPF, LIC, ELSS, etc.)</label>
            <input type="number" id="ded80c" placeholder="0" min="0" max="150000"/>
            <p class="hint">Max ₹1,50,000</p>
          </div>
          <div class="field">
            <label>80CCD(1B) — NPS Extra</label>
            <input type="number" id="ded80ccd1b" placeholder="0" min="0" max="50000"/>
            <p class="hint">Max ₹50,000</p>
          </div>
        </div>
        <div class="row2">
          <div class="field">
            <label>80CCD(2) — Employer NPS</label>
            <input type="number" id="ded80ccd2" placeholder="0" min="0"/>
            <p class="hint">Available in both regimes</p>
          </div>
          <div class="field">
            <label>80D — Medical Insurance</label>
            <input type="number" id="ded80d" placeholder="0" min="0"/>
            <p class="hint">₹25K self + ₹25K/₹50K parents</p>
          </div>
        </div>
        <div class="row2">
          <div class="field">
            <label>80E — Education Loan Interest</label>
            <input type="number" id="ded80e" placeholder="0" min="0"/>
          </div>
          <div class="field">
            <label>80G — Donations</label>
            <input type="number" id="ded80g" placeholder="0" min="0"/>
          </div>
        </div>
        <div class="row2">
          <div class="field">
            <label>80TTA/80TTB — Savings Interest</label>
            <input type="number" id="ded80tta" placeholder="0" min="0"/>
            <p class="hint">₹10K (80TTA) / ₹50K seniors (80TTB)</p>
          </div>
          <div class="field">
            <label>Other Deductions (80DD, 80DDB, etc.)</label>
            <input type="number" id="dedOther" placeholder="0" min="0"/>
          </div>
        </div>
      </div>
    </div>

    <!-- ──── MAT / AMT ──── -->
    <div class="card" id="matAmtCard" style="display:none">
      <div class="card-head">
        <div class="icon" style="background:#FEF3C7">⚖️</div>
        <div><h2 id="matAmtCardTitle">MAT / AMT Computation</h2><p id="matAmtCardSub">Minimum Alternate Tax u/s 115JB / 115JC</p></div>
      </div>
      <div class="card-body">
        <!-- MAT (Companies) -->
        <div id="matSection">
          <div class="section-title">📋 MAT u/s 115JB — Companies</div>
          <div class="row2">
            <div class="field">
              <label>Book Profit u/s 115JB (₹)</label>
              <input type="number" id="matBookProfit" placeholder="0" min="0" oninput="computeMatAmt()"/>
              <p class="hint">Net profit per P&amp;L + mandatory add-backs (Sch VII items)</p>
            </div>
            <div class="field">
              <label>MAT Rate</label>
              <select id="matRate" onchange="computeMatAmt()">
                <option value="0.15" selected>15% – Domestic Company (general)</option>
                <option value="0.09">9% – New Mfg Co. u/s 115BAB</option>
                <option value="0.075">7.5% – Co. in IFSC u/s 115A(4)</option>
              </select>
            </div>
          </div>
          <div id="matResult" style="display:none;margin-top:12px;padding:14px;background:#FFFBEB;border:1.5px solid #FDE68A;border-radius:10px">
            <div style="font-size:12px;font-weight:700;color:#92400E;margin-bottom:8px">MAT Computation</div>
            <div id="matResultRows"></div>
          </div>
        </div>
        <!-- AMT (Non-companies) -->
        <div id="amtSection" style="display:none">
          <div class="section-title">📋 AMT u/s 115JC — Firms / LLPs / Individuals / HUF / AOP</div>
          <div class="field">
            <label>Adjusted Total Income (ATI) for AMT (₹)</label>
            <input type="number" id="amtAti" placeholder="0" min="0" oninput="computeMatAmt()"/>
            <p class="hint">GTI + deductions claimed u/s 10AA / 35AD / 80H–80RRB added back (u/s 115JC)</p>
          </div>
          <div id="amtResult" style="display:none;margin-top:12px;padding:14px;background:#FFFBEB;border:1.5px solid #FDE68A;border-radius:10px">
            <div style="font-size:12px;font-weight:700;color:#92400E;margin-bottom:8px">AMT Computation</div>
            <div id="amtResultRows"></div>
          </div>
        </div>
        <div style="margin-top:12px;padding:10px 14px;background:#F0FDF4;border:1px solid #BBF7D0;border-radius:8px;font-size:11px;color:#065F46">
          <strong>📌 Credit:</strong> If normal tax &gt; MAT/AMT, MAT/AMT credit u/s 115JAA/115JD is carried forward for up to 15 years and can be set off in future years when normal tax exceeds MAT/AMT.
        </div>
      </div>
    </div>

    <!-- ──── TDS / TCS / ADVANCE TAX ──── -->
    <div class="card">
      <div class="card-head">
        <div class="icon" style="background:#F5F3FF">🏦</div>
        <div><h2>Tax Already Paid</h2><p>TDS, TCS &amp; Advance Tax</p></div>
      </div>
      <div class="card-body">
        <div class="row3">
          <div class="field">
            <label>TDS (Estimated)</label>
            <input type="number" id="tds" placeholder="0" min="0"/>
          </div>
          <div class="field">
            <label>TCS (Estimated)</label>
            <input type="number" id="tcs" placeholder="0" min="0"/>
          </div>
          <div class="field">
            <label>Advance Tax Paid</label>
            <input type="number" id="advanceTax" placeholder="0" min="0"/>
          </div>
        </div>
      </div>
    </div>

    <button class="btn-calc" id="calcBtn" onclick="calculateTax(event)">
      <span class="btn-text">🧮 Calculate Tax</span>
      <div class="btn-spinner"></div>
    </button>
    <button class="btn-reset" onclick="resetForm()">↺ Reset All Fields</button>
  </div>

  <!-- RIGHT: Results Section -->
  <div id="result-section">
    <div class="result-panel" id="resultPanel">

      <div class="card" id="singleResult" style="display:none">
        <div class="card-head">
          <div class="icon" style="background:#ECFDF5">📊</div>
          <div><h2 id="resultTitle">Tax Computation</h2><p id="resultSubtitle"></p></div>
        </div>
        <div class="card-body">
          <div id="resultBody"></div>
          <!-- Donut chart -->
          <div id="taxChartWrap" style="display:none">
            <h3>Tax Breakdown</h3>
            <div class="donut-container">
              <svg class="donut-svg" width="140" height="140" viewBox="0 0 140 140">
                <circle cx="70" cy="70" r="54" fill="none" stroke="#F3F4F6" stroke-width="22"/>
                <circle id="donut-base" class="donut-segment" cx="70" cy="70" r="54" fill="none" stroke="#2563EB" stroke-width="22" stroke-dasharray="0 339.3" stroke-dashoffset="84.8" stroke-linecap="round"/>
                <circle id="donut-surcharge" class="donut-segment" cx="70" cy="70" r="54" fill="none" stroke="#F59E0B" stroke-width="22" stroke-dasharray="0 339.3" stroke-dashoffset="84.8" stroke-linecap="round"/>
                <circle id="donut-cess" class="donut-segment" cx="70" cy="70" r="54" fill="none" stroke="#10B981" stroke-width="22" stroke-dasharray="0 339.3" stroke-dashoffset="84.8" stroke-linecap="round"/>
                <text x="70" y="65" text-anchor="middle" font-size="10" fill="#6B7280" font-weight="600" font-family="Inter,sans-serif">Total Tax</text>
                <text id="donut-center-val" x="70" y="82" text-anchor="middle" font-size="13" fill="#111827" font-weight="800" font-family="Inter,sans-serif">₹0</text>
              </svg>
              <div class="donut-legend" id="donutLegend"></div>
            </div>
          </div>
          <button class="print-btn" onclick="window.print()">🖨️ Print / Save PDF</button>
        </div>
      </div>

      <div id="compareResult" style="display:none">
        <div class="compare-box">
          <h3>⚖️ Regime Comparison — <span id="compareYearLabel"></span></h3>
          <div class="regime-winner" id="regimeWinner"></div>
          <div class="savings" id="savingsAmt"></div>
          <table class="compare-table">
            <thead><tr><th></th><th>🆕 New Regime</th><th>📜 Old Regime</th></tr></thead>
            <tbody id="compareBody"></tbody>
          </table>
        </div>

        <!-- Regime bar chart -->
        <div id="regimeChartWrap">
          <h3>📊 Visual Comparison</h3>
          <div class="bar-chart" id="regimeBarChart"></div>
        </div>

        <div class="card" style="margin-top:16px">
          <div class="card-head">
            <div class="icon" style="background:#EFF6FF">📊</div>
            <div><h2>New Regime — Detailed</h2></div>
          </div>
          <div class="card-body"><div id="newRegimeDetail"></div></div>
        </div>

        <div class="card" style="margin-top:16px">
          <div class="card-head">
            <div class="icon" style="background:#FFFBEB">📊</div>
            <div><h2>Old Regime — Detailed</h2></div>
          </div>
          <div class="card-body"><div id="oldRegimeDetail"></div></div>
        </div>

        <button class="print-btn" onclick="window.print()">🖨️ Print / Save PDF</button>
      </div>

      <div class="card" style="margin-top:16px" id="slabCard">
        <div class="card-head">
          <div class="icon" style="background:#F0FDF4">📋</div>
          <div><h2>Slab-wise Tax Breakup</h2><p id="slabRegimeLabel"></p></div>
        </div>
        <div class="card-body" id="slabBody"></div>
      </div>

      <div class="disclaimer" id="disclaimerBox">
        <strong>⚠ Disclaimer:</strong> This calculator is for estimation purposes only. Actual tax liability may differ
        based on specific exemptions, deductions, and interpretations. Always consult a qualified Chartered Accountant
        for final tax computation. Surcharge marginal relief is indicative. Special rate incomes (capital gains, winnings)
        are not eligible for Section 87A rebate.
      </div>
      <div class="disclaimer future" id="futureDisclaimer" style="display:none;margin-top:8px">
        <strong>📅 Future Year Note:</strong> PY 2026-27 (AY 2027-28) rates are based on the Union Budget 2026 which
        retained PY 2025-26 slab rates without changes. These rates are subject to any future amendments or notifications
        by the Government. Always verify with the latest Finance Act before finalizing.
      </div>

      <!-- ──── MAT/AMT RESULT IN RESULTS PANEL ──── -->
      <div id="matAmtResultCard" style="display:none;margin-top:16px">
        <div class="card">
          <div class="card-head">
            <div class="icon" style="background:#FEF3C7">⚖️</div>
            <div><h2 id="matAmtResTitle">MAT / AMT Summary</h2><p>Minimum Alternate Tax computation result</p></div>
          </div>
          <div class="card-body" id="matAmtResBody"></div>
        </div>
      </div>

      <!-- ──── ADVANCE TAX SCHEDULE (PY 2026-27 only) ──── -->
      <div id="advanceTaxCard" style="display:none;margin-top:16px">
        <div class="card">
          <div class="card-head">
            <div class="icon" style="background:#EFF6FF">📅</div>
            <div><h2>Advance Tax Schedule — PY 2026-27</h2><p>Instalment-wise liability u/s 208 &amp; 211</p></div>
          </div>
          <div class="card-body">
            <div style="padding:10px 14px;background:#EFF6FF;border:1px solid #BFDBFE;border-radius:8px;font-size:11px;color:#1E40AF;margin-bottom:14px">
              <strong>ℹ️ Who must pay?</strong> Every assessee whose estimated tax liability for the year is ₹10,000 or more (after TDS/TCS) must pay advance tax. Senior citizens (60+) with no business income are exempt u/s 207.
            </div>
            <div id="advanceTaxTable"></div>
            <div style="margin-top:14px;padding:10px 14px;background:#FFF7ED;border:1px solid #FED7AA;border-radius:8px;font-size:11px;color:#92400E">
              <strong>⚠️ Interest for default:</strong> u/s 234B — 1% per month on 90% of assessed tax not paid as advance tax. u/s 234C — 1% per month for 3 months on shortfall per instalment (single month for last instalment). u/s 234A — 1% per month on self-assessment tax if return filed late.
            </div>
          </div>
        </div>
      </div>
    </div>

    <!-- Before calculation: show slab reference -->
    <div id="preCalcInfo">
      <div class="card" id="refSlabCard">
        <div class="card-head">
          <div class="icon" style="background:#EFF6FF">📋</div>
          <div><h2 id="refSlabTitle">New Regime Slab Rates</h2><p id="refSlabSub"></p></div>
        </div>
        <div class="card-body" id="refSlabBody"></div>
      </div>
      <div class="card">
        <div class="card-head">
          <div class="icon" style="background:#FFFBEB">📋</div>
          <div><h2>Old Regime Slab Rates</h2><p>Unchanged across all years</p></div>
        </div>
        <div class="card-body">
          <table class="slab-table">
            <thead><tr><th>Income Slab</th><th style="text-align:right">Rate</th></tr></thead>
            <tbody>
              <tr><td>Up to ₹2,50,000</td><td class="amt">Nil</td></tr>
              <tr><td>₹2,50,001 – ₹5,00,000</td><td class="amt">5%</td></tr>
              <tr><td>₹5,00,001 – ₹10,00,000</td><td class="amt">20%</td></tr>
              <tr><td>Above ₹10,00,000</td><td class="amt">30%</td></tr>
            </tbody>
          </table>
          <p class="hint" style="margin-top:10px">Senior citizens (60-80): exempt up to ₹3L. Super seniors (80+): exempt up to ₹5L. Rebate u/s 87A: up to ₹5L → max ₹12,500.</p>
        </div>
      </div>
      <div class="card">
        <div class="card-head">
          <div class="icon" style="background:#F5F3FF">📋</div>
          <div><h2>Special Rate Incomes</h2><p id="refSpecialSub"></p></div>
        </div>
        <div class="card-body" id="refSpecialBody"></div>
      </div>
    </div>
  </div>
</div>
</div>

<footer>
  <div class="ft-main">
    <div>
      <div class="ft-brand-name">CA<span>Toolkit</span></div>
      <p class="ft-brand-desc">CA Toolkit is a comprehensive utility platform built by a CA Article from Ludhiana, Punjab, providing automation tools for Indian Chartered Accountants. The platform saves hours of manual work every year — from Balance Sheet year-shift to GST reconciliation, tax calculations, and more.</p>
    </div>
    <div>
      <div class="ft-col-title">Know More</div>
      <ul class="ft-links">
        <li><a href="/">Home</a></li>
        <li><a href="/">BS Year Shift</a></li>
        <li><a href="/tool/tb-to-bs">TB → Balance Sheet</a></li>
        <li><a href="/tool/tax-calculator">Tax Calculator</a></li>
        <li><a href="/privacy">Privacy Policy</a></li>
      </ul>
    </div>
    <div>
      <div class="ft-col-title">Contact Us</div>
      <div class="ft-contact-name">CA Toolkit</div>
      <div class="ft-contact-addr">Built for Indian Chartered Accountants<br/>Created by CA Article · Ludhiana, Punjab</div>
    </div>
  </div>
  <div class="ft-bottom">
    <span class="ft-bottom-left">©2026 CA Toolkit · All Rights Reserved</span>
    <span class="ft-bottom-right">Built for Indian CAs · Ludhiana, Punjab</span>
  </div>
</footer>
<div class="toast" id="toast"></div>

<script>
/* ═══════════════════════════════════════════════════════════════════════
   INCOME TAX CALCULATOR — Multi-Year Engine (Enhanced)
   PY 2023-24 | PY 2024-25 | PY 2025-26 | PY 2026-27 (upcoming)
   Features: All Assessee Types · MAT/AMT · Advance Tax Schedule
   ═══════════════════════════════════════════════════════════════════════ */

let currentRegime = 'new';
let currentYear = '2025-26';

/* ── ASSESSEE TYPE CONFIGURATION ──────────────────────────────────── */
const ASSESSEE_CFG = {
  // Individuals
  individual_below60:    { label:'Individual – Below 60', group:'individual', age:'below60', canAmt:true },
  individual_senior:     { label:'Individual – Senior Citizen (60–80)', group:'individual', age:'senior', canAmt:true },
  individual_supersenior:{ label:'Individual – Super Senior (80+)', group:'individual', age:'supersenior', canAmt:true },
  individual_nri:        { label:'Individual – NRI', group:'individual', age:'below60', isNRI:true, canAmt:true },
  // HUF
  huf:                   { label:'HUF', group:'individual', age:'below60', canAmt:true },
  // Firms / LLPs
  firm:                  { label:'Firm / LLP', group:'firm', flatRate:0.30, surchargeThreshold:1e7, surchargeRate:0.12, canAmt:true },
  // Companies
  company_domestic:      { label:'Domestic Company', group:'company', flatRate:0.22, surchargeThreshold:1e7, surchargeRateLow:0.07, surchargeRateHigh:0.12, canMat:true },
  company_foreign:       { label:'Foreign Company', group:'company', flatRate:0.40, surchargeRateLow:0.02, surchargeRateHigh:0.05, surchargeThreshold:1e7, canMat:true },
  company_mfg_new:       { label:'New Mfg Co. u/s 115BAB', group:'company', flatRate:0.15, surchargeRate:0.10, matRate:0.09, canMat:true },
  company_small:         { label:'Domestic Co. ≤₹400Cr (Sec 115BA)', group:'company', flatRate:0.25, surchargeRateLow:0.07, surchargeRateHigh:0.12, surchargeThreshold:1e7, canMat:true },
  // Others
  aop_boi:               { label:'AOP / BOI', group:'individual', age:'below60', canAmt:true },
  cooperative:           { label:'Co-operative Society', group:'coop', canAmt:true },
  trust_aop:             { label:'Trust / AOP (Registered)', group:'individual', age:'below60', canAmt:true },
  local_authority:       { label:'Local Authority', group:'local', flatRate:0.30, canAmt:false },
  artificial_person:     { label:'Artificial Juridical Person', group:'individual', age:'below60', canAmt:true },
};

function getAssesseeCfg() {
  const t = document.getElementById('assesseeType').value;
  return ASSESSEE_CFG[t] || ASSESSEE_CFG['individual_below60'];
}

/* ── Assessee button selector ─────────────────────────────────────── */
function selectAssessee(btn, val) {
  document.querySelectorAll('.at-btn').forEach(b => b.classList.remove('active'));
  btn.classList.add('active');
  document.getElementById('assesseeType').value = val;
  onAssesseeTypeChange();
}

function onAssesseeTypeChange() {
  const t = document.getElementById('assesseeType').value;
  const ac = ASSESSEE_CFG[t];

  const isInd  = ac.group === 'individual';
  const isCo   = ac.group === 'company';
  const isFirm = ac.group === 'firm';
  const isCoop = ac.group === 'coop';
  const isLocal= ac.group === 'local';

  // ── Income panel visibility ──────────────────────────────────────
  document.getElementById('inc_individual').style.display = (isInd) ? 'block' : 'none';
  document.getElementById('inc_company').style.display    = isCo   ? 'block' : 'none';
  document.getElementById('inc_firm').style.display       = isFirm ? 'block' : 'none';
  document.getElementById('inc_coop').style.display       = isCoop ? 'block' : 'none';
  document.getElementById('inc_local').style.display      = isLocal? 'block' : 'none';

  // Card subtitle
  const subMap = {
    individual:'All 5 Heads of Income', huf:'All 5 Heads of Income',
    firm:'Business, Capital Gains & Other Sources',
    company_domestic:'Business, Capital Gains & MAT',
    company_foreign:'Business, Capital Gains & MAT',
    company_mfg_new:'Business, Capital Gains & MAT',
    company_small:'Business, Capital Gains & MAT',
    aop_boi:'All applicable Heads of Income',
    cooperative:'Business & Other Sources',
    trust_aop:'All applicable Heads of Income',
    local_authority:'Net Taxable Income',
    artificial_person:'All 5 Heads of Income',
  };
  document.getElementById('incomeCardSub').textContent = subMap[t] || 'Gross Total Income computation';

  // ── Age badge for individuals ────────────────────────────────────
  const ageBadge = document.getElementById('ageBadge');
  if (ageBadge) {
    const ageMap = {
      individual_below60:'Below 60 years', individual_senior:'Senior Citizen (60–80)',
      individual_supersenior:'Super Senior Citizen (80+)', individual_nri:'Non-Resident (NRI)', huf:'HUF',
      aop_boi:'AOP / BOI', trust_aop:'Trust', artificial_person:'Juridical Person',
    };
    ageBadge.textContent = ageMap[t] || '';
  }

  // ── Show/hide Basic Info sub-sections ───────────────────────────
  document.getElementById('individualFields').style.display  = isInd  ? 'block' : 'none';
  document.getElementById('companyFields').style.display     = isCo   ? 'block' : 'none';
  document.getElementById('firmFields').style.display        = isFirm ? 'block' : 'none';
  document.getElementById('aopFields').style.display         = (isCoop || isLocal) ? 'block' : 'none';

  // ── Regime toggle visibility ─────────────────────────────────────
  const regimeToggle = document.querySelector('.regime-toggle');
  if (isCo || isFirm || isLocal) {
    regimeToggle.style.display = 'none';
    currentRegime = 'new';
  } else {
    regimeToggle.style.display = 'flex';
  }

  // ── Deductions card ──────────────────────────────────────────────
  const dc = document.getElementById('deductions-card');
  if (isCo || isFirm || isCoop || isLocal) {
    dc.style.display = 'none';
  } else {
    dc.style.display = 'block';
    dc.style.opacity = currentRegime === 'new' ? '0.4' : '1';
    dc.style.pointerEvents = currentRegime === 'new' ? 'none' : 'auto';
  }

  // ── B/F Capital Loss fields — only for individuals / HUF / AOP (not co/firm/coop/local) ──
  const bfLossEl = document.getElementById('bfLossSection');
  if (bfLossEl) bfLossEl.style.display = (isInd) ? 'block' : 'none';

  // ── MAT / AMT card ───────────────────────────────────────────────
  const matCard = document.getElementById('matAmtCard');
  if (ac.canMat) {
    matCard.style.display = 'block';
    document.getElementById('matSection').style.display = 'block';
    document.getElementById('amtSection').style.display = 'none';
    document.getElementById('matAmtCardTitle').textContent = 'MAT Computation';
    document.getElementById('matAmtCardSub').textContent = 'Minimum Alternate Tax u/s 115JB';
    const mr = document.getElementById('matRate');
    if (t === 'company_mfg_new') mr.value = '0.09';
    else mr.value = '0.15';
  } else if (ac.canAmt && (isFirm || isCoop)) {
    matCard.style.display = 'block';
    document.getElementById('matSection').style.display = 'none';
    document.getElementById('amtSection').style.display = 'block';
    document.getElementById('matAmtCardTitle').textContent = 'AMT Computation';
    document.getElementById('matAmtCardSub').textContent = 'Alternate Minimum Tax u/s 115JC';
  } else {
    matCard.style.display = 'none';
  }

  // ── NRI auto-set ─────────────────────────────────────────────────
  if (ac.isNRI) {
    const rs = document.getElementById('residentialStatus');
    if (rs) rs.value = 'nri';
  }

  computeMatAmt();
}

/* ── YEAR-SPECIFIC TAX CONFIGURATIONS ─────────────────────────────── */
const YEAR_CONFIG = {
  '2023-24': {
    label: 'Tax Year 2024-25',
    ayLabel: 'AY 2024-25',
    isFuture: false,
    stdDeduction: 50000,
    newSlabs: [
      { upto: 300000,  rate: 0 },
      { upto: 600000,  rate: 0.05 },
      { upto: 900000,  rate: 0.10 },
      { upto: 1200000, rate: 0.15 },
      { upto: 1500000, rate: 0.20 },
      { upto: Infinity, rate: 0.30 },
    ],
    rebateNew: { limit: 700000, max: 25000 },
    rebateOld: { limit: 500000, max: 12500 },
    stcg111aRate: 0.15,
    ltcg112aRate: 0.10,
    ltcg112aExempt: 100000,
    ltcgOtherRate: 0.20,
    ltcgOtherLabel: '20% (with indexation)',
    maxSurchargeNew: 0.25,
  },
  '2024-25': {
    label: 'Tax Year 2025-26',
    ayLabel: 'AY 2025-26',
    isFuture: false,
    stdDeduction: 75000,
    hasTransitional: true,
    newSlabs: [
      { upto: 300000,  rate: 0 },
      { upto: 700000,  rate: 0.05 },
      { upto: 1000000, rate: 0.10 },
      { upto: 1200000, rate: 0.15 },
      { upto: 1500000, rate: 0.20 },
      { upto: Infinity, rate: 0.30 },
    ],
    rebateNew: { limit: 700000, max: 25000 },
    rebateOld: { limit: 500000, max: 12500 },
    stcg111aRate: 0.20,
    ltcg112aRate: 0.125,
    ltcg112aExempt: 125000,
    ltcgOtherRate: 0.125,
    ltcgOtherLabel: '12.5% (post July 2024)',
    stcg111aRateOld: 0.15,
    ltcg112aRateOld: 0.10,
    ltcg112aExemptOld: 100000,
    ltcgOtherRateOld: 0.20,
    ltcgOtherLabelOld: '20% with indexation (pre July 2024)',
    maxSurchargeNew: 0.25,
  },
  '2025-26': {
    label: 'Tax Year 2026-27',
    ayLabel: 'AY 2026-27',
    isFuture: false,
    stdDeduction: 75000,
    newSlabs: [
      { upto: 400000,  rate: 0 },
      { upto: 800000,  rate: 0.05 },
      { upto: 1200000, rate: 0.10 },
      { upto: 1600000, rate: 0.15 },
      { upto: 2000000, rate: 0.20 },
      { upto: 2400000, rate: 0.25 },
      { upto: Infinity, rate: 0.30 },
    ],
    rebateNew: { limit: 1200000, max: 60000 },
    rebateOld: { limit: 500000, max: 12500 },
    stcg111aRate: 0.20,
    ltcg112aRate: 0.125,
    ltcg112aExempt: 125000,
    ltcgOtherRate: 0.125,
    ltcgOtherLabel: '12.5%',
    maxSurchargeNew: 0.25,
  },
  '2026-27': {
    label: 'Tax Year 2027-28',
    ayLabel: 'AY 2027-28',
    isFuture: true,
    stdDeduction: 75000,
    newSlabs: [
      { upto: 400000,  rate: 0 },
      { upto: 800000,  rate: 0.05 },
      { upto: 1200000, rate: 0.10 },
      { upto: 1600000, rate: 0.15 },
      { upto: 2000000, rate: 0.20 },
      { upto: 2400000, rate: 0.25 },
      { upto: Infinity, rate: 0.30 },
    ],
    rebateNew: { limit: 1200000, max: 60000 },
    rebateOld: { limit: 500000, max: 12500 },
    stcg111aRate: 0.20,
    ltcg112aRate: 0.125,
    ltcg112aExempt: 125000,
    ltcgOtherRate: 0.125,
    ltcgOtherLabel: '12.5%',
    maxSurchargeNew: 0.25,
  },
};

/* ── OLD REGIME SLABS ─────────────────────────────────────────────── */
const OLD_SLABS_BELOW60 = [
  { upto: 250000,  rate: 0 },
  { upto: 500000,  rate: 0.05 },
  { upto: 1000000, rate: 0.20 },
  { upto: Infinity, rate: 0.30 },
];
const OLD_SLABS_SENIOR = [
  { upto: 300000,  rate: 0 },
  { upto: 500000,  rate: 0.05 },
  { upto: 1000000, rate: 0.20 },
  { upto: Infinity, rate: 0.30 },
];
const OLD_SLABS_SUPERSENIOR = [
  { upto: 500000,  rate: 0 },
  { upto: 1000000, rate: 0.20 },
  { upto: Infinity, rate: 0.30 },
];
/* Co-operative Society slabs */
const COOP_SLABS = [
  { upto: 10000,  rate: 0.10 },
  { upto: 20000,  rate: 0.20 },
  { upto: Infinity, rate: 0.30 },
];

function getOldSlabs() {
  const ac = getAssesseeCfg();
  if (ac.group === 'coop') return COOP_SLABS;
  const age = ac.age || 'below60';
  if (age === 'supersenior') return OLD_SLABS_SUPERSENIOR;
  if (age === 'senior') return OLD_SLABS_SENIOR;
  return OLD_SLABS_BELOW60;
}

function cfg() { return YEAR_CONFIG[currentYear]; }

/* ── UI: Year selection ───────────────────────────────────────────── */
function setYear(yr) {
  currentYear = yr;
  document.querySelectorAll('.year-pill').forEach(b => b.classList.remove('active'));
  event.currentTarget.classList.add('active');

  const c = cfg();
  document.getElementById('stdDeduction').value = c.stdDeduction;
  document.getElementById('stdDedHint').textContent =
    c.stdDeduction === 50000 ? '₹50,000 for PY 2023-24' : '₹75,000 for PY 2024-25 onwards';

  const isTrans = c.hasTransitional || false;
  document.getElementById('transitionalNotice').style.display = isTrans ? 'block' : 'none';
  document.getElementById('preJulyFields').style.display = isTrans ? 'block' : 'none';

  if (isTrans) {
    document.getElementById('stcg111aLabel').textContent = 'STCG 111A — Post-July (equity) @ 20%';
    document.getElementById('ltcg112aLabel').textContent = 'LTCG 112A — Post-July (equity) @ 12.5%';
    document.getElementById('ltcgOtherLabel').textContent = 'LTCG Other — Post-July @ 12.5%';
  } else {
    document.getElementById('stcg111aLabel').textContent = 'STCG u/s 111A (equity, STT paid)';
    document.getElementById('ltcg112aLabel').textContent = 'LTCG u/s 112A (equity, STT paid)';
    document.getElementById('ltcgOtherLabel').textContent = 'LTCG — Other (property, debt etc.)';
  }

  document.getElementById('stcgRateHint').textContent =
    'Taxed at ' + (c.stcg111aRate * 100) + '%' + (isTrans ? ' (post 23 July 2024)' : '');
  document.getElementById('ltcgRateHint').textContent =
    (c.ltcg112aRate * 100) + '% above ₹' + (c.ltcg112aExempt / 100000).toFixed(2).replace('.00','') + ' lakh exemption';
  document.getElementById('ltcgOtherHint').textContent = c.ltcgOtherLabel;

  document.getElementById('futureYearNote').style.display = c.isFuture ? 'block' : 'none';

  if (!isTrans) {
    ['stcg111aPreJuly','ltcg112aPreJuly','ltcgOtherPreJuly'].forEach(id => {
      const el = document.getElementById(id);
      if (el) el.value = '';
    });
  }

  updateRefSlabs();
  document.getElementById('resultPanel').classList.remove('show');
  document.getElementById('preCalcInfo').style.display = 'block';
  // Hide advance tax card until recalculated
  document.getElementById('advanceTaxCard').style.display = 'none';
  document.getElementById('matAmtResultCard').style.display = 'none';
}

function setRegime(r) {
  currentRegime = r;
  document.querySelectorAll('.regime-btn').forEach(b => b.classList.remove('active'));
  document.getElementById('btn-' + r).classList.add('active');
  const dc = document.getElementById('deductions-card');
  if (r === 'new') { dc.style.opacity = '0.4'; dc.style.pointerEvents = 'none'; }
  else { dc.style.opacity = '1'; dc.style.pointerEvents = 'auto'; }
}

function updateRefSlabs() {
  const c = cfg();
  const slabs = c.newSlabs;
  document.getElementById('refSlabTitle').textContent = 'New Regime Slab Rates';
  document.getElementById('refSlabSub').textContent = c.label + (c.isFuture ? ' (Estimated)' : '');
  let h = '<table class="slab-table"><thead><tr><th>Income Slab</th><th style="text-align:right">Rate</th></tr></thead><tbody>';
  let prev = 0;
  for (const s of slabs) {
    const from = '₹' + prev.toLocaleString('en-IN');
    const to = s.upto === Infinity ? '& above' : '₹' + s.upto.toLocaleString('en-IN');
    const label = s.upto === Infinity ? 'Above ' + from : from + ' – ' + to;
    h += '<tr><td>' + (prev === 0 ? 'Up to ' + to : label) + '</td><td class="amt">' + (s.rate === 0 ? 'Nil' : (s.rate*100)+'%') + '</td></tr>';
    prev = s.upto;
  }
  h += '</tbody></table>';
  const rebateInfo = c.rebateNew.limit >= 1200000
    ? 'Rebate u/s 87A: Income up to ₹' + (c.rebateNew.limit/100000) + ' lakh → zero tax (max rebate ₹' + c.rebateNew.max.toLocaleString('en-IN') + ').'
    : 'Rebate u/s 87A: Income up to ₹' + (c.rebateNew.limit/100000) + ' lakh → zero tax (max rebate ₹' + c.rebateNew.max.toLocaleString('en-IN') + ').';
  h += '<p class="hint" style="margin-top:10px">' + rebateInfo + ' Standard deduction ₹' + c.stdDeduction.toLocaleString('en-IN') + ' for salaried.</p>';
  document.getElementById('refSlabBody').innerHTML = h;

  document.getElementById('refSpecialSub').textContent = c.label;
  let sp = '<table class="slab-table"><thead><tr><th>Type</th><th style="text-align:right">Rate</th></tr></thead><tbody>';
  sp += '<tr><td>STCG u/s 111A (equity, STT paid)</td><td class="amt">' + (c.stcg111aRate*100) + '%</td></tr>';
  sp += '<tr><td>LTCG u/s 112A (equity) above ₹' + (c.ltcg112aExempt/100000).toFixed(2).replace('.00','') + 'L</td><td class="amt">' + (c.ltcg112aRate*100) + '%</td></tr>';
  sp += '<tr><td>LTCG — Other assets</td><td class="amt">' + c.ltcgOtherLabel + '</td></tr>';
  sp += '<tr><td>Winnings (lottery, games, etc.)</td><td class="amt">30%</td></tr>';
  sp += '</tbody></table>';
  sp += '<p class="hint" style="margin-top:10px">Health &amp; Education Cess @ 4% on tax + surcharge. Max surcharge under new regime: 25%.</p>';
  document.getElementById('refSpecialBody').innerHTML = sp;
}

function syncMatBookProfit() {
  const v = document.getElementById('co_bookProfit').value;
  const el = document.getElementById('matBookProfit');
  if (el) el.value = v;
  computeMatAmt();
}
function syncAmtAti() {
  const v = document.getElementById('firm_amtAti').value;
  const el = document.getElementById('amtAti');
  if (el) el.value = v;
  computeMatAmt();
}

/* helper: read value by id safely */
function v(id) { return parseFloat(document.getElementById(id).value) || 0; }
function fmt(n) {
  if (n < 0) return '-₹' + Math.abs(Math.round(n)).toLocaleString('en-IN');
  return '₹' + Math.round(n).toLocaleString('en-IN');
}

/* ── Slab-based tax ───────────────────────────────────────────────── */
function calcSlabTax(taxableIncome, slabs) {
  let tax = 0, prev = 0;
  const breakup = [];
  for (const slab of slabs) {
    if (taxableIncome <= prev) break;
    const chunk = Math.min(taxableIncome, slab.upto) - prev;
    const t = chunk * slab.rate;
    breakup.push({ from: prev, to: Math.min(taxableIncome, slab.upto), rate: slab.rate, amount: chunk, tax: t });
    tax += t;
    prev = slab.upto;
  }
  return { tax, breakup };
}

/* ── Surcharge ────────────────────────────────────────────────────── */
function calcSurcharge(tax, totalIncome, isNewRegime) {
  if (totalIncome <= 5000000) return 0;
  let rate = 0;
  const maxNew = cfg().maxSurchargeNew;
  if (isNewRegime) {
    if (totalIncome <= 10000000) rate = 0.10;
    else if (totalIncome <= 20000000) rate = 0.15;
    else rate = maxNew;
  } else {
    if (totalIncome <= 10000000) rate = 0.10;
    else if (totalIncome <= 20000000) rate = 0.15;
    else if (totalIncome <= 50000000) rate = 0.25;
    else rate = 0.37;
  }
  return tax * rate;
}

// Compute marginal relief separately so it can be shown in detail
function calcMarginalRelief(normalTax, specialTax, totalIncome, surcharge, slabs, isNew, normalTaxable) {
  // Marginal relief: ensures that crossing a surcharge threshold doesn't leave
  // the taxpayer worse off than if they had earned just at the threshold.
  //
  // Standard ICAI formula for each threshold boundary T:
  //   maxAllowed = (taxAtT + lowerSurchargeAtT) + (actualIncome - T)
  //   where taxAtT = slab_tax_on_(T - specialIncomeGross) + specialTax
  //         lowerSurchargeAtT = taxAtT × rateBelow (rate that applied just below T)
  //   relief = max(0, actualTaxPlusSurcharge - maxAllowed)
  //
  // specialIncomeGross = gross LTCG + STCG income (kept constant at threshold)
  const specialIncomeGross = totalIncome - (normalTaxable !== undefined ? normalTaxable : 0);

  // Surcharge rates: {threshold: rate_BELOW_this_threshold}
  const lowerRates = {5000000: 0, 10000000: 0.10, 20000000: 0.15, 50000000: 0.25};
  const thresholds  = [5000000, 10000000, 20000000, 50000000];

  const totalBaseTax        = normalTax + specialTax;
  const totalWithSurcharge  = totalBaseTax + surcharge;

  let relief      = 0;
  let reliefDetail = { applies: false, threshold: 0, taxAtTh: 0, surchargeAtTh: 0, excess: 0, maxAllowed: 0, actualWithSurcharge: 0 };

  for (const th of thresholds) {
    if (totalIncome <= th) break;

    const excess       = totalIncome - th;
    const lowerRate    = lowerRates[th] || 0;

    // Normal income if total GTI were exactly at threshold (keep special income same)
    const normalAtTh   = Math.max(0, th - specialIncomeGross);

    // Tax at threshold
    const taxAtThNormal  = calcSlabTax(normalAtTh, slabs).tax;
    const taxAtTh        = taxAtThNormal + specialTax;          // special rate tax stays same

    // Surcharge at the LOWER rate (rate applicable just below this threshold)
    const surchargeAtTh  = taxAtTh * lowerRate;

    // Maximum tax+surcharge the taxpayer should pay
    const maxAllowed     = taxAtTh + surchargeAtTh + excess;

    if (totalWithSurcharge > maxAllowed) {
      const r = totalWithSurcharge - maxAllowed;
      if (r > relief) {
        relief       = r;
        reliefDetail = { applies: true, threshold: th, taxAtTh, surchargeAtTh, excess, maxAllowed, actualWithSurcharge: totalWithSurcharge };
      }
    }
  }
  return { relief: Math.max(0, relief), detail: reliefDetail };
}

function calcSurchargeCapped(tax, totalIncome) {
  if (totalIncome <= 5000000) return 0;
  let rate = Math.min(0.15, totalIncome > 10000000 ? 0.15 : 0.10);
  return tax * rate;
}

/* ── Company / Firm tax computation ───────────────────────────────── */
function computeForCompanyFirm(ac, totalIncome) {
  const flatRate = ac.flatRate;
  let baseTax = totalIncome * flatRate;

  // Surcharge for companies
  let surcharge = 0;
  if (ac.group === 'company') {
    const low = ac.surchargeRateLow || 0;
    const high = ac.surchargeRateHigh || 0;
    const th = ac.surchargeThreshold || 1e7;
    if (totalIncome > th) surcharge = baseTax * high;
    else if (totalIncome > 1e7) surcharge = baseTax * low;
    else surcharge = baseTax * low;
  } else if (ac.group === 'firm') {
    if (totalIncome > 1e7) surcharge = baseTax * (ac.surchargeRate || 0.12);
  }
  const cess = (baseTax + surcharge) * 0.04;
  const totalTax = baseTax + surcharge + cess;
  return { baseTax, surcharge, cess, totalTax, flatRate };
}

/* ── MAT / AMT Computation ────────────────────────────────────────── */
function computeMatAmt() {
  const t = document.getElementById('assesseeType').value;
  const ac = ASSESSEE_CFG[t];
  if (!ac) return;

  if (ac.canMat) {
    const bookProfit = parseFloat(document.getElementById('matBookProfit').value) || 0;
    if (!bookProfit) { document.getElementById('matResult').style.display='none'; return; }
    const matRate = parseFloat(document.getElementById('matRate').value) || 0.15;
    const matBase = bookProfit * matRate;
    let surcharge = 0;
    if (ac.surchargeRateLow) surcharge = matBase * (bookProfit > 1e7 ? (ac.surchargeRateHigh||0) : (ac.surchargeRateLow||0));
    const matCess = (matBase + surcharge) * 0.04;
    const matTotal = matBase + surcharge + matCess;
    let h = '';
    h += `<div class="mat-row"><span class="ml">Book Profit u/s 115JB</span><span class="mv">${fmt(bookProfit)}</span></div>`;
    h += `<div class="mat-row"><span class="ml">MAT Rate</span><span class="mv">${(matRate*100).toFixed(1)}%</span></div>`;
    h += `<div class="mat-row"><span class="ml">MAT (before surcharge/cess)</span><span class="mv">${fmt(matBase)}</span></div>`;
    if (surcharge) h += `<div class="mat-row"><span class="ml">Surcharge</span><span class="mv">${fmt(surcharge)}</span></div>`;
    h += `<div class="mat-row"><span class="ml">Cess @ 4%</span><span class="mv">${fmt(matCess)}</span></div>`;
    h += `<div class="mat-row mt"><span class="ml">Total MAT Payable</span><span class="mv">${fmt(matTotal)}</span></div>`;
    h += `<p style="font-size:11px;margin-top:8px;color:#78350F">Tax payable = <strong>MAX(Normal Tax, MAT)</strong>. If MAT &gt; Normal Tax, excess = MAT Credit u/s 115JAA (carry fwd 15 yrs).</p>`;
    document.getElementById('matResultRows').innerHTML = h;
    document.getElementById('matResult').style.display = 'block';

  } else if (ac.canAmt && document.getElementById('amtSection').style.display !== 'none') {
    const ati = parseFloat(document.getElementById('amtAti').value) || 0;
    if (!ati) { document.getElementById('amtResult').style.display='none'; return; }
    const amtBase = ati * 0.185;
    let surcharge = 0;
    if (ac.group === 'firm' && ati > 1e7) surcharge = amtBase * 0.12;
    const amtCess = (amtBase + surcharge) * 0.04;
    const amtTotal = amtBase + surcharge + amtCess;
    let h = '';
    h += `<div class="mat-row"><span class="ml">Adjusted Total Income (ATI)</span><span class="mv">${fmt(ati)}</span></div>`;
    h += `<div class="mat-row"><span class="ml">AMT Rate</span><span class="mv">18.5%</span></div>`;
    h += `<div class="mat-row"><span class="ml">AMT (before surcharge/cess)</span><span class="mv">${fmt(amtBase)}</span></div>`;
    if (surcharge) h += `<div class="mat-row"><span class="ml">Surcharge</span><span class="mv">${fmt(surcharge)}</span></div>`;
    h += `<div class="mat-row"><span class="ml">Cess @ 4%</span><span class="mv">${fmt(amtCess)}</span></div>`;
    h += `<div class="mat-row mt"><span class="ml">Total AMT Payable</span><span class="mv">${fmt(amtTotal)}</span></div>`;
    h += `<p style="font-size:11px;margin-top:8px;color:#78350F">Tax payable = <strong>MAX(Normal Tax, AMT)</strong>. If AMT &gt; Normal Tax, excess = AMT Credit u/s 115JD (carry fwd 15 yrs).</p>`;
    document.getElementById('amtResultRows').innerHTML = h;
    document.getElementById('amtResult').style.display = 'block';
  }
}

/* ── Advance Tax Schedule ─────────────────────────────────────────── */
function renderAdvanceTaxSchedule(totalTax, tdsPaid, tcsPaid) {
  // Only for PY 2026-27
  if (currentYear !== '2026-27') {
    document.getElementById('advanceTaxCard').style.display = 'none';
    return;
  }
  const netLiability = Math.max(0, totalTax - tdsPaid - tcsPaid);
  if (netLiability < 10000) {
    document.getElementById('advanceTaxCard').style.display = 'none';
    return;
  }

  // Instalments for PY 2026-27 (u/s 211)
  const today = new Date();
  const instalments = [
    { due: new Date('2026-06-15'), cumPct: 15,  pct: 15,  label: '1st Instalment' },
    { due: new Date('2026-09-15'), cumPct: 45,  pct: 30,  label: '2nd Instalment' },
    { due: new Date('2026-12-15'), cumPct: 75,  pct: 30,  label: '3rd Instalment' },
    { due: new Date('2027-03-15'), cumPct: 100, pct: 25,  label: '4th Instalment' },
  ];

  // For presumptive income assessees (Sec 44AD/44ADA) — single instalment
  const isPrespumptive = false; // could add toggle later

  let h = `<div style="margin-bottom:12px;padding:10px 14px;background:#F0FDF4;border:1px solid #BBF7D0;border-radius:8px;font-size:11px;color:#065F46">
    <strong>💡 Total Tax Liability (after TDS/TCS):</strong> ${fmt(netLiability)} — Advance Tax required (≥ ₹10,000)
  </div>`;

  h += `<table class="at-table">
    <thead>
      <tr>
        <th>Instalment</th>
        <th style="text-align:center">% of Tax</th>
        <th>Due Date</th>
        <th style="text-align:right">Amount Due</th>
        <th style="text-align:right">Cumulative</th>
      </tr>
    </thead>
    <tbody>`;

  let cumAmount = 0;
  instalments.forEach((inst, idx) => {
    const amt = Math.round(netLiability * inst.pct / 100);
    const cumAmt = Math.round(netLiability * inst.cumPct / 100);
    cumAmount = cumAmt;

    const isPast = today > inst.due;
    const isNext = !isPast && (idx === 0 || today > instalments[idx-1].due);
    const rowClass = isPast ? 'overdue' : (isNext ? 'upcoming' : '');
    const statusBadge = isPast
      ? '<span style="font-size:10px;background:#FEE2E2;color:#B91C1C;padding:2px 7px;border-radius:10px;margin-left:6px;font-weight:700">Due</span>'
      : (isNext ? '<span style="font-size:10px;background:#FEF3C7;color:#92400E;padding:2px 7px;border-radius:10px;margin-left:6px;font-weight:700">Next</span>' : '');

    const dueDateStr = inst.due.toLocaleDateString('en-IN', { day:'numeric', month:'short', year:'numeric' });

    h += `<tr class="${rowClass}">
      <td><strong>${inst.label}</strong>${statusBadge}</td>
      <td class="pct">${inst.pct}%</td>
      <td class="due">${dueDateStr}</td>
      <td class="amt-cell">${fmt(amt)}</td>
      <td class="amt-cell"><span class="cumul">Cumul: </span>${fmt(cumAmt)}</td>
    </tr>`;
  });

  h += `</tbody>
    <tfoot>
      <tr style="background:#EFF6FF;font-weight:800">
        <td colspan="3" style="padding:10px;font-size:12px">Total Advance Tax</td>
        <td></td>
        <td class="amt-cell" style="color:var(--brand);font-size:13px">${fmt(netLiability)}</td>
      </tr>
    </tfoot>
  </table>`;

  h += `<div style="margin-top:12px;font-size:11px;color:var(--muted);line-height:1.7">
    <strong>📌 Notes:</strong>
    <ul style="margin:6px 0 0 16px;padding:0">
      <li>Instalments calculated on <em>estimated</em> total income for PY 2026-27.</li>
      <li>Assessees under Sec 44AD / 44ADA (presumptive) may pay <strong>entire advance tax by 15 March 2027</strong> (single instalment).</li>
      <li>Senior citizens (60+) with <em>no</em> business income are <strong>exempt</strong> from advance tax u/s 207.</li>
      <li>If advance tax paid &lt; 90% of assessed tax → interest u/s 234B @ 1%/month on shortfall.</li>
      <li>Shortfall in each instalment → interest u/s 234C @ 1%/month for 3 months (1 month for last instalment).</li>
    </ul>
  </div>`;

  document.getElementById('advanceTaxTable').innerHTML = h;
  document.getElementById('advanceTaxCard').style.display = 'block';
}

/* ── Main Computation ─────────────────────────────────────────────── */
function computeForRegime(isNew) {
  const c = cfg();
  const ac = getAssesseeCfg();
  const t = document.getElementById('assesseeType').value;
  const name = document.getElementById('assesseeName').value.trim();
  const isResident = (ac.isNRI || (document.getElementById('residentialStatus') && document.getElementById('residentialStatus').value === 'nri')) ? false : true;

  const isCo   = ac.group === 'company';
  const isFirm = ac.group === 'firm';
  const isCoop = ac.group === 'coop';
  const isLocal= ac.group === 'local';
  const isInd  = ac.group === 'individual';

  /* ── Read income from the right panel ── */
  let grossSalary=0, exemptAllow=0, stdDed=0, salaryIncome=0;
  let houseRaw=0, loanInt=0, houseIncome=0, houseLossCapped=0;
  let businessIncome=0;
  let stcg111a=0, stcgOther=0, ltcg112a=0, ltcgOther=0;
  let stcg111aPreJuly=0, ltcg112aPreJuly=0, ltcgOtherPreJuly=0;
  let ltcg112aPreJulyExemptAmt=0;
  let otherIncome=0, winnings=0;
  // B/F Capital Losses — Issue 1 fix: new inputs read here
  let bf_stcl=0, bf_ltcl=0;

  if (isInd) {
    grossSalary  = v('grossSalary');
    exemptAllow  = isNew ? 0 : v('exemptAllow');
    stdDed       = v('stdDeduction');
    salaryIncome = Math.max(0, grossSalary - exemptAllow - stdDed);
    houseRaw     = v('houseIncome');
    loanInt      = v('homeLoanInterest');
    houseIncome  = houseRaw - loanInt;
    houseLossCapped = Math.max(Math.min(0, houseIncome), -200000);
    businessIncome= v('businessIncome');
    stcg111a     = v('stcg111a');
    stcgOther    = v('stcgOther');
    ltcg112a     = v('ltcg112a');
    ltcgOther    = v('ltcgOther');
    const isTrans= c.hasTransitional || false;
    if (isTrans) {
      stcg111aPreJuly = v('stcg111aPreJuly');
      ltcg112aPreJuly = v('ltcg112aPreJuly');
      ltcgOtherPreJuly= v('ltcgOtherPreJuly');
      ltcg112aPreJulyExemptAmt = Math.min(ltcg112aPreJuly, c.ltcg112aExemptOld||100000);
    }
    otherIncome  = v('otherIncome');
    winnings     = v('winningsIncome');
    // Issue 1 fix: read B/F carry-forward losses (always positive from user)
    bf_stcl      = Math.max(0, v('bf_stcl'));
    bf_ltcl      = Math.max(0, v('bf_ltcl'));

  } else if (isCo) {
    businessIncome= v('co_businessIncome');
    stcg111a     = v('co_stcg111a');
    stcgOther    = v('co_stcgOther');
    ltcg112a     = v('co_ltcg112a');
    ltcgOther    = v('co_ltcgOther');
    otherIncome  = v('co_otherIncome');
    // Sync book profit to MAT field
    const bp = v('co_bookProfit');
    const matEl = document.getElementById('matBookProfit');
    if (matEl && bp) matEl.value = bp;

  } else if (isFirm) {
    businessIncome= v('firm_businessIncome');
    stcg111a     = v('firm_stcg111a');
    stcgOther    = v('firm_stcgOther');
    ltcg112a     = v('firm_ltcg112a');
    ltcgOther    = v('firm_ltcgOther');
    otherIncome  = v('firm_otherIncome');
    const ati = v('firm_amtAti');
    const amtEl = document.getElementById('amtAti');
    if (amtEl && ati) amtEl.value = ati;

  } else if (isCoop) {
    businessIncome= v('coop_businessIncome');
    otherIncome  = v('coop_otherIncome');

  } else if (isLocal) {
    businessIncome= v('local_income');
  }

  const isTrans = c.hasTransitional || false;

  const normalIncome = salaryIncome + Math.max(0, houseIncome) + businessIncome + stcgOther + otherIncome;
  const normalAfterLoss = Math.max(0, normalIncome + houseLossCapped);

  // Deductions (only individuals, not companies/firms/coop/local)
  let totalDeductions = 0;
  if (isInd) {
    const ded80ccd2 = v('ded80ccd2');
    if (isNew) {
      totalDeductions = ded80ccd2;
    } else {
      totalDeductions = Math.min(v('ded80c'), 150000) + Math.min(v('ded80ccd1b'), 50000) +
        ded80ccd2 + v('ded80d') + v('ded80e') + v('ded80g') + v('ded80tta') + v('dedOther');
    }
  }

  const normalTaxable = Math.max(0, normalAfterLoss - totalDeductions);

  let normalTax, surchargeNormal, surchargeSpecial, totalSurcharge, slabResult;
  let rebate87a = 0;

  if (isCo || isFirm || isLocal) {
    const flatRes = computeForCompanyFirm(ac, normalTaxable);
    normalTax = flatRes.baseTax;
    surchargeNormal = flatRes.surcharge;
    surchargeSpecial = 0;
    totalSurcharge = flatRes.surcharge;
    slabResult = { tax: flatRes.baseTax, breakup: [{ from:0, to:normalTaxable, rate:ac.flatRate, amount:normalTaxable, tax:flatRes.baseTax }] };

    // Special rate taxes for company/firm
    const taxSTCG111A = stcg111a * c.stcg111aRate;
    const taxLTCG112A = Math.max(0, ltcg112a - c.ltcg112aExempt) * c.ltcg112aRate;
    const taxLTCGOther= ltcgOther * c.ltcgOtherRate;
    const totalSpecialTax = taxSTCG111A + taxLTCG112A + taxLTCGOther;
    const ltcg112aExemptAmt = Math.min(ltcg112a, c.ltcg112aExempt);
    const totalIncome = normalTaxable + stcg111a + ltcg112a + ltcgOther;
    const flatRes2 = computeForCompanyFirm(ac, normalTaxable);
    const cess = flatRes2.cess;
    const totalTax = flatRes2.totalTax + totalSpecialTax;
    const tdsPaid = v('tds'); const tcsPaid = v('tcs'); const advTax = v('advanceTax');
    const totalPrepaid = tdsPaid + tcsPaid + advTax;
    const netPayable = totalTax - totalPrepaid;

    return {
      yearLabel:c.label, ayLabel:c.ayLabel, isFuture:c.isFuture, isTrans:false,
      name, isNew, isResident, c, ac, assesseeType:t,
      grossSalary:0, exemptAllow:0, stdDed:0, salaryIncome:0,
      houseRaw:0, loanInt:0, houseIncome:0, houseLossCapped:0,
      businessIncome, stcg111a, stcgOther, ltcg112a, ltcg112aExemptAmt, ltcgOther,
      stcg111aPreJuly:0, ltcg112aPreJuly:0, ltcg112aPreJulyExemptAmt:0, ltcgOtherPreJuly:0,
      otherIncome, winnings:0,
      normalIncome:normalAfterLoss, totalDeductions:0, normalTaxable,
      slabResult:{ tax:flatRes2.baseTax, breakup:[{ from:0, to:normalTaxable, rate:ac.flatRate, amount:normalTaxable, tax:flatRes2.baseTax }] },
      normalTax:flatRes2.baseTax, rebate87a:0, normalTaxAfterRebate:flatRes2.baseTax,
      taxSTCG111A, taxLTCG112A, taxLTCGOther, taxWinnings:0,
      taxSTCG111APreJuly:0, taxLTCG112APreJuly:0, taxLTCGOtherPreJuly:0,
      totalSpecialTax, totalIncome, surchargeNormal:flatRes2.surcharge, surchargeSpecial:0, totalSurcharge:flatRes2.surcharge,
      cess, totalTax, tdsPaid, tcsPaid, advTax, totalPrepaid, netPayable,
      isFlatRate:true, flatRate:ac.flatRate,
    };
  }

  if (isCoop) {
    // Co-operative society slabs
    slabResult = calcSlabTax(normalTaxable, COOP_SLABS);
    normalTax  = slabResult.tax;
    const surcharge = normalTaxable > 1e7 ? normalTax * 0.12 : 0;
    const cess = (normalTax + surcharge) * 0.04;
    const totalTax = normalTax + surcharge + cess;
    const tdsPaid = v('tds'); const tcsPaid = v('tcs'); const advTax = v('advanceTax');
    const totalPrepaid = tdsPaid + tcsPaid + advTax;
    const netPayable = totalTax - totalPrepaid;
    return {
      yearLabel:c.label, ayLabel:c.ayLabel, isFuture:c.isFuture, isTrans:false,
      name, isNew:false, isResident, c, ac, assesseeType:t,
      grossSalary:0, exemptAllow:0, stdDed:0, salaryIncome:0,
      houseRaw:0, loanInt:0, houseIncome:0, houseLossCapped:0,
      businessIncome, stcg111a:0, stcgOther:0, ltcg112a:0, ltcg112aExemptAmt:0, ltcgOther:0,
      stcg111aPreJuly:0, ltcg112aPreJuly:0, ltcg112aPreJulyExemptAmt:0, ltcgOtherPreJuly:0,
      otherIncome, winnings:0,
      normalIncome:normalAfterLoss, totalDeductions:0, normalTaxable,
      slabResult, normalTax, rebate87a:0, normalTaxAfterRebate:normalTax,
      taxSTCG111A:0, taxLTCG112A:0, taxLTCGOther:0, taxWinnings:0,
      taxSTCG111APreJuly:0, taxLTCG112APreJuly:0, taxLTCGOtherPreJuly:0,
      totalSpecialTax:0, totalIncome:normalTaxable,
      surchargeNormal:surcharge, surchargeSpecial:0, totalSurcharge:surcharge,
      cess, totalTax, tdsPaid, tcsPaid, advTax, totalPrepaid, netPayable,
    };
  }

  // Individual / HUF / AOP / Trust / Artificial Person — slab-based
  const slabs = isNew ? c.newSlabs : getOldSlabs();
  slabResult = calcSlabTax(normalTaxable, slabs);
  normalTax  = slabResult.tax;

  if (isResident) {
    const rebateCfg = isNew ? c.rebateNew : c.rebateOld;
    if (normalTaxable <= rebateCfg.limit) {
      rebate87a = Math.min(normalTax, rebateCfg.max);
    }
  }
  normalTax = Math.max(0, normalTax - rebate87a);

  // ── Issue 2 fix: Guard negative CG inputs — user shouldn't enter negative here ──
  // Capture raw (possibly negative) values first so we can report current-year losses as C/F notes.
  const raw_stcg111a = stcg111a;
  const raw_stcgOther = stcgOther;
  const raw_ltcg112a  = ltcg112a;
  const raw_ltcgOther = ltcgOther;
  // Current-year STCL = sum of all negative CG fields (absolute value = loss to c/f next year)
  const cy_stcl = Math.max(0, -raw_stcg111a) + Math.max(0, -raw_stcgOther);
  const cy_ltcl = Math.max(0, -raw_ltcg112a) + Math.max(0, -raw_ltcgOther);
  stcg111a  = Math.max(0, stcg111a);
  stcgOther = Math.max(0, stcgOther);
  ltcg112a  = Math.max(0, ltcg112a);
  ltcgOther = Math.max(0, ltcgOther);
  if (isTrans) {
    stcg111aPreJuly  = Math.max(0, stcg111aPreJuly);
    ltcg112aPreJuly  = Math.max(0, ltcg112aPreJuly);
    ltcgOtherPreJuly = Math.max(0, ltcgOtherPreJuly);
  }

  // ── Issue 1 fix: B/F Loss Set-off u/s 74 ──
  // Rule: STCL c/f can be set off against ANY capital gain (STCG or LTCG).
  //       LTCL c/f can ONLY be set off against LTCG.
  // We apply set-off in the most tax-efficient order:
  //   Step 1: Use B/F STCL against LTCG 112A first (highest-rate savings), then LTCG Other, then STCG 111A.
  //   Step 2: Use B/F LTCL against remaining LTCG 112A, then LTCG Other.
  // Gains are floored at 0 — losses cannot create negative taxable income under this head.
  let stcl_remaining = bf_stcl;
  let ltcl_remaining = bf_ltcl;

  // STCL set-off: absorb into LTCG 112A first (reduces special-rate tax)
  const stcl_vs_ltcg112a = Math.min(stcl_remaining, ltcg112a);
  ltcg112a      -= stcl_vs_ltcg112a;
  stcl_remaining -= stcl_vs_ltcg112a;
  // Then into LTCG Other
  const stcl_vs_ltcgOther = Math.min(stcl_remaining, ltcgOther);
  ltcgOther      -= stcl_vs_ltcgOther;
  stcl_remaining  -= stcl_vs_ltcgOther;
  // Then into LTCG Other (Pre-July if transitional)
  if (isTrans) {
    const stcl_vs_ltcgOtherPJ = Math.min(stcl_remaining, ltcgOtherPreJuly);
    ltcgOtherPreJuly -= stcl_vs_ltcgOtherPJ;
    stcl_remaining   -= stcl_vs_ltcgOtherPJ;
    const stcl_vs_ltcg112aPJ = Math.min(stcl_remaining, ltcg112aPreJuly);
    ltcg112aPreJuly -= stcl_vs_ltcg112aPJ;
    stcl_remaining  -= stcl_vs_ltcg112aPJ;
  }
  // Finally into STCG 111A
  const stcl_vs_stcg111a = Math.min(stcl_remaining, stcg111a);
  stcg111a       -= stcl_vs_stcg111a;
  stcl_remaining  -= stcl_vs_stcg111a;

  // LTCL set-off: only against LTCG
  const ltcl_vs_ltcg112a = Math.min(ltcl_remaining, ltcg112a);
  ltcg112a       -= ltcl_vs_ltcg112a;
  ltcl_remaining  -= ltcl_vs_ltcg112a;
  const ltcl_vs_ltcgOther = Math.min(ltcl_remaining, ltcgOther);
  ltcgOther      -= ltcl_vs_ltcgOther;
  ltcl_remaining  -= ltcl_vs_ltcgOther;
  if (isTrans) {
    const ltcl_vs_ltcgOtherPJ = Math.min(ltcl_remaining, ltcgOtherPreJuly);
    ltcgOtherPreJuly -= ltcl_vs_ltcgOtherPJ;
    ltcl_remaining   -= ltcl_vs_ltcgOtherPJ;
    const ltcl_vs_ltcg112aPJ = Math.min(ltcl_remaining, ltcg112aPreJuly);
    ltcg112aPreJuly -= ltcl_vs_ltcg112aPJ;
    ltcl_remaining  -= ltcl_vs_ltcg112aPJ;
  }
  // Store how much B/F loss was actually utilised and remaining c/f
  const bf_stcl_utilised = bf_stcl - stcl_remaining;
  const bf_ltcl_utilised = bf_ltcl - ltcl_remaining;
  const bf_stcl_cf = stcl_remaining;   // still to be carried forward
  const bf_ltcl_cf = ltcl_remaining;

  const taxSTCG111A = Math.max(0, stcg111a) * c.stcg111aRate;
  const taxLTCG112A = Math.max(0, ltcg112a - c.ltcg112aExempt) * c.ltcg112aRate;
  const taxLTCGOther= Math.max(0, ltcgOther) * c.ltcgOtherRate;
  const taxWinnings = winnings * 0.30;

  let taxSTCG111APreJuly=0, taxLTCG112APreJuly=0, taxLTCGOtherPreJuly=0;
  if (isTrans) {
    taxSTCG111APreJuly = Math.max(0, stcg111aPreJuly) * (c.stcg111aRateOld||0.15);
    taxLTCG112APreJuly = Math.max(0, ltcg112aPreJuly - ltcg112aPreJulyExemptAmt) * (c.ltcg112aRateOld||0.10);
    taxLTCGOtherPreJuly= Math.max(0, ltcgOtherPreJuly) * (c.ltcgOtherRateOld||0.20);
  }

  const totalSpecialTax = taxSTCG111A + taxLTCG112A + taxLTCGOther + taxWinnings +
                          taxSTCG111APreJuly + taxLTCG112APreJuly + taxLTCGOtherPreJuly;

  const totalIncome = normalTaxable + stcg111a + stcg111aPreJuly + ltcg112a + ltcg112aPreJuly +
                      ltcgOther + ltcgOtherPreJuly + winnings;

  surchargeNormal  = calcSurcharge(normalTax, totalIncome, isNew);
  surchargeSpecial = calcSurchargeCapped(totalSpecialTax, totalIncome);
  totalSurcharge   = surchargeNormal + surchargeSpecial;

  // Marginal relief: proper formula — if tax+surcharge exceeds tax-at-threshold + excess income
  // Note: 'slabs' is already declared above via const slabs = isNew ? c.newSlabs : getOldSlabs()
  const mrResult = calcMarginalRelief(normalTax, totalSpecialTax, totalIncome, totalSurcharge, slabs, isNew, normalTaxable);
  const marginalRelief = mrResult.relief;
  const marginalReliefDetail = mrResult.detail;
  totalSurcharge = Math.max(0, totalSurcharge - marginalRelief);

  const totalBeforeCess = normalTax + totalSpecialTax + totalSurcharge;
  const cess     = totalBeforeCess * 0.04;
  const totalTax = totalBeforeCess + cess;
  const tdsPaid  = v('tds'); const tcsPaid = v('tcs'); const advTax = v('advanceTax');
  const totalPrepaid = tdsPaid + tcsPaid + advTax;
  const netPayable   = totalTax - totalPrepaid;
  const ltcg112aExemptAmt = Math.min(ltcg112a, c.ltcg112aExempt);

  return {
    yearLabel:c.label, ayLabel:c.ayLabel, isFuture:c.isFuture, isTrans,
    name, isNew, isResident, c, ac, assesseeType:t,
    grossSalary, exemptAllow, stdDed, salaryIncome,
    houseRaw, loanInt, houseIncome, houseLossCapped,
    businessIncome,
    stcg111a, stcgOther, ltcg112a, ltcg112aExemptAmt, ltcgOther,
    stcg111aPreJuly, ltcg112aPreJuly, ltcg112aPreJulyExemptAmt, ltcgOtherPreJuly,
    otherIncome, winnings,
    bf_stcl, bf_ltcl, bf_stcl_utilised, bf_ltcl_utilised, bf_stcl_cf, bf_ltcl_cf,
    cy_stcl, cy_ltcl,
    normalIncome:normalAfterLoss, totalDeductions, normalTaxable,
    slabResult,
    normalTax: normalTax + rebate87a, rebate87a,
    normalTaxAfterRebate: normalTax,
    taxSTCG111A, taxLTCG112A, taxLTCGOther, taxWinnings,
    taxSTCG111APreJuly, taxLTCG112APreJuly, taxLTCGOtherPreJuly,
    totalSpecialTax, totalIncome,
    surchargeNormal, surchargeSpecial, totalSurcharge,
    marginalRelief: marginalRelief||0, marginalReliefDetail: marginalReliefDetail||{},
    cess, totalTax, tdsPaid, tcsPaid, advTax, totalPrepaid, netPayable,
  };
}

/* ── Render result rows ────────────────────────────────────────────── */
function row(lbl, val, cls) {
  return '<div class="result-row '+(cls||'')+'"><span class="lbl">'+lbl+'</span><span class="val">'+val+'</span></div>';
}

function renderResult(r) {
  const c = r.c;
  const ac = r.ac;
  let h = '';

  // Assessee badge
  h += `<div class="assessee-badge">👤 ${ac.label || ''}</div>`;

  h += row('Gross Salary', fmt(r.grossSalary));
  if (!r.isNew && !r.isFlatRate) h += row('Less: Exempt Allowances', fmt(-r.exemptAllow), 'sub');
  if (!r.isFlatRate) h += row('Less: Standard Deduction (₹'+c.stdDeduction.toLocaleString('en-IN')+')', fmt(-r.stdDed), 'sub');
  h += row('Net Salary Income (Head 1)', fmt(r.salaryIncome));
  h += '<div style="height:6px"></div>';
  h += row('House Property Income (Head 2)', fmt(r.houseIncome));
  if (r.houseLossCapped < 0) h += row('Loss set-off (max ₹2L)', fmt(r.houseLossCapped), 'sub');
  h += row('Business Income (Head 3)', fmt(r.businessIncome));

  // Determine whether any CG or BF loss exists
  const _hasCG = r.stcg111a || r.stcgOther || r.ltcg112a || r.ltcgOther ||
                 r.stcg111aPreJuly || r.ltcg112aPreJuly || r.ltcgOtherPreJuly ||
                 r.bf_stcl || r.bf_ltcl;
  if (_hasCG) {
    const totalCG = r.stcg111a + r.stcgOther + r.ltcg112a + r.ltcgOther +
                    r.stcg111aPreJuly + r.ltcg112aPreJuly + r.ltcgOtherPreJuly;
    h += row('Capital Gains (Head 4) — Net after B/F set-off', fmt(totalCG));
  }
  if (r.isTrans && (r.stcg111aPreJuly || r.ltcg112aPreJuly || r.ltcgOtherPreJuly)) {
    if (r.stcg111aPreJuly) h += row('STCG 111A pre-July @ '+(c.stcg111aRateOld*100)+'%', fmt(r.stcg111aPreJuly), 'sub');
    if (r.ltcg112aPreJuly) h += row('LTCG 112A pre-July (exempt ₹'+(c.ltcg112aExemptOld/100000)+'L) @ '+(c.ltcg112aRateOld*100)+'%', fmt(r.ltcg112aPreJuly), 'sub');
    if (r.ltcgOtherPreJuly) h += row('LTCG Other pre-July @ '+c.ltcgOtherLabelOld, fmt(r.ltcgOtherPreJuly), 'sub');
  }
  if (r.stcg111a) h += row((r.isTrans?'STCG 111A post-July':'STCG u/s 111A')+' @ '+(c.stcg111aRate*100)+'%', fmt(r.stcg111a), 'sub');
  if (r.stcgOther) h += row('STCG — Other (slab rate)', fmt(r.stcgOther), 'sub');
  if (r.ltcg112a) h += row((r.isTrans?'LTCG 112A post-July':'LTCG u/s 112A')+' (exempt ₹'+(c.ltcg112aExempt/100000)+'L)', fmt(r.ltcg112a), 'sub');
  if (r.ltcgOther) h += row((r.isTrans?'LTCG Other post-July':'LTCG — Other')+' @ '+c.ltcgOtherLabel, fmt(r.ltcgOther), 'sub');
  // B/F loss set-off display rows
  if (r.bf_stcl_utilised > 0) h += row('Less: B/F STCL set off u/s 74', fmt(-r.bf_stcl_utilised), 'sub');
  if (r.bf_ltcl_utilised > 0) h += row('Less: B/F LTCL set off u/s 74', fmt(-r.bf_ltcl_utilised), 'sub');

  // ── C/F note box: remaining B/F loss not absorbed + current-year fresh loss ──
  {
    const cfNotes = [];
    if (r.bf_stcl_cf > 0)
      cfNotes.push('Unadjusted B/F Short-Term Capital Loss of <strong>₹' + fmt(r.bf_stcl_cf) + '</strong> is still available to carry forward to the next year (can be set off against any capital gain).');
    if (r.bf_ltcl_cf > 0)
      cfNotes.push('Unadjusted B/F Long-Term Capital Loss of <strong>₹' + fmt(r.bf_ltcl_cf) + '</strong> is still available to carry forward to the next year (can only be set off against LTCG).');
    if (r.cy_stcl > 0)
      cfNotes.push('Current-year Short-Term Capital Loss of <strong>₹' + fmt(r.cy_stcl) + '</strong> will be carried forward to the next assessment year and can be set off against any capital gain (u/s 74). It can be carried forward for up to <strong>8 assessment years</strong>.');
    if (r.cy_ltcl > 0)
      cfNotes.push('Current-year Long-Term Capital Loss of <strong>₹' + fmt(r.cy_ltcl) + '</strong> will be carried forward to the next assessment year and can only be set off against Long-Term Capital Gains (u/s 74). It can be carried forward for up to <strong>8 assessment years</strong>.');
    if (cfNotes.length > 0) {
      h += '<div style="margin:10px 0 4px;padding:11px 14px;background:#FFF7ED;border:1.5px solid #FED7AA;border-radius:9px;font-size:12px;line-height:1.75;color:#92400E">'
         + '<div style="font-weight:700;margin-bottom:6px;font-size:12px">📌 Capital Loss Carry-Forward Note</div>'
         + cfNotes.map(n => '• ' + n).join('<br>')
         + '</div>';
    }
  }

  h += row('Other Sources (Head 5)', fmt(r.otherIncome + r.winnings));
  if (r.winnings) h += row('Winnings @ 30%', fmt(r.winnings), 'sub');

  h += '<div style="height:4px;border-top:2px solid var(--border);margin:10px 0"></div>';
  const gtiTotal = r.normalIncome + r.stcg111a + r.ltcg112a + r.ltcgOther + r.winnings + r.stcg111aPreJuly + r.ltcg112aPreJuly + r.ltcgOtherPreJuly;
  h += row('Gross Total Income', fmt(gtiTotal), 'total');
  if (r.totalDeductions > 0) h += row('Less: Deductions Ch VI-A', fmt(-r.totalDeductions));
  h += row('Total Taxable Income (Normal)', fmt(r.normalTaxable), 'total');

  h += '<div style="height:4px;border-top:2px solid var(--border);margin:10px 0"></div>';
  if (r.isFlatRate) {
    h += row('Tax @ '+(r.flatRate*100)+'% (flat)', fmt(r.normalTax));
  } else {
    h += row('Tax on Normal Income (slab)', fmt(r.normalTax));
    if (r.rebate87a > 0) h += row('Less: Rebate u/s 87A', fmt(-r.rebate87a), 'sub');
    h += row('Tax after Rebate', fmt(r.normalTaxAfterRebate));
  }

  if (r.totalSpecialTax > 0) {
    h += '<div style="height:6px"></div>';
    if (r.taxSTCG111APreJuly) h += row('Tax STCG 111A pre-July @ '+(c.stcg111aRateOld*100)+'%', fmt(r.taxSTCG111APreJuly), 'sub');
    if (r.taxLTCG112APreJuly) h += row('Tax LTCG 112A pre-July @ '+(c.ltcg112aRateOld*100)+'%', fmt(r.taxLTCG112APreJuly), 'sub');
    if (r.taxLTCGOtherPreJuly) h += row('Tax LTCG Other pre-July @ '+(c.ltcgOtherRateOld*100)+'%', fmt(r.taxLTCGOtherPreJuly), 'sub');
    if (r.taxSTCG111A) h += row('Tax '+(r.isTrans?'STCG 111A post-July':'STCG 111A')+' @ '+(c.stcg111aRate*100)+'%', fmt(r.taxSTCG111A), 'sub');
    if (r.taxLTCG112A) h += row('Tax '+(r.isTrans?'LTCG 112A post-July':'LTCG 112A')+' @ '+(c.ltcg112aRate*100)+'%', fmt(r.taxLTCG112A), 'sub');
    if (r.taxLTCGOther) h += row('Tax '+(r.isTrans?'LTCG Other post-July':'LTCG Other')+' @ '+c.ltcgOtherLabel, fmt(r.taxLTCGOther), 'sub');
    if (r.taxWinnings) h += row('Tax on Winnings @ 30%', fmt(r.taxWinnings), 'sub');
    h += row('Total Special Rate Tax', fmt(r.totalSpecialTax));
  }

  if (r.totalSurcharge > 0) h += row('Surcharge', fmt(r.totalSurcharge));
  if (r.marginalRelief > 0) h += row('Less: Marginal Relief', fmt(-r.marginalRelief), 'sub');
  h += row('Health & Education Cess @ 4%', fmt(r.cess));
  h += row('Total Tax Liability', fmt(r.totalTax), 'total');

  // Detailed explanation section
  const _baseTax = (r.normalTaxAfterRebate||0) + (r.totalSpecialTax||0);
  const _surchargeGross = (r.surchargeNormal||0) + (r.surchargeSpecial||0);
  const _mrAmt = r.marginalRelief||0;
  const _mrd = r.marginalReliefDetail||{};
  h += `<div style="margin-top:12px">
    <button onclick="var d=this.nextElementSibling;d.style.display=d.style.display==='none'?'block':'none';this.textContent=d.style.display==='block'?'▼ Hide Details':'▶ Show Detailed Explanation'"
      style="background:none;border:1px solid var(--brand);color:var(--brand);border-radius:6px;padding:6px 14px;font-size:12px;cursor:pointer;width:100%">▶ Show Detailed Explanation</button>
    <div style="display:none;background:#F0F9FF;border-radius:8px;padding:12px;margin-top:6px;font-size:12px;line-height:1.8">
      <b>📊 Step-by-step Tax Computation</b><br><br>
      <b>1. Income &amp; Base Tax</b><br>
      &nbsp;• Normal Taxable Income: ₹${fmt(r.normalTaxable||0)}<br>
      &nbsp;• Gross Total Income (GTI): ₹${fmt(r.totalIncome||0)}<br>
      &nbsp;• Normal slab tax: ₹${fmt(r.normalTaxAfterRebate||0)}${r.rebate87a ? ' (after 87A rebate)' : ''}<br>
      ${(r.totalSpecialTax||0) > 0 ? `&nbsp;• Capital gains tax (special rate): ₹${fmt(r.totalSpecialTax)}<br>` : ''}
      &nbsp;• <b>Total base tax: ₹${fmt(_baseTax)}</b><br><br>
      <b>2. Surcharge</b><br>
      &nbsp;• GTI ₹${fmt(r.totalIncome||0)} → surcharge bracket: ${(r.totalIncome||0)>50000000?'> ₹5Cr':(r.totalIncome||0)>20000000?'> ₹2Cr':(r.totalIncome||0)>10000000?'> ₹1Cr':(r.totalIncome||0)>5000000?'> ₹50L':'≤ ₹50L (Nil)'}<br>
      &nbsp;• Gross surcharge: ₹${fmt(_surchargeGross)}<br>
      ${_mrAmt > 0 ?
        `&nbsp;• <span style="color:#059669"><b>Marginal Relief applies! ₹${fmt(_mrAmt)}</b></span><br>
      &nbsp;&nbsp;&nbsp;Actual tax+surcharge (₹${fmt(_mrd.actualWithSurcharge||0)}) &gt; Max allowed (₹${fmt(_mrd.maxAllowed||0)}) by ₹${fmt(_mrAmt)}<br>
      &nbsp;&nbsp;&nbsp;Max allowed = Tax at ₹${fmt(_mrd.threshold||0)} threshold (₹${fmt((_mrd.taxAtTh||0)+(_mrd.surchargeAtTh||0))}) + Excess ₹${fmt(_mrd.excess||0)}`
        :
        `&nbsp;• Marginal Relief: <b>₹0 — not applicable</b><br>
      &nbsp;&nbsp;&nbsp;(Excess income above threshold >> surcharge amount)`}<br>
      &nbsp;• <b>Net surcharge: ₹${fmt(r.totalSurcharge||0)}</b><br><br>
      <b>3. Health &amp; Education Cess @ 4%</b><br>
      &nbsp;• 4% × ₹${fmt(_baseTax + (r.totalSurcharge||0))} = <b>₹${fmt(r.cess||0)}</b><br><br>
      <b>4. Total Tax Liability = ₹${fmt(r.totalTax||0)}</b>
    </div>
  </div>`;

  h += '<div style="height:4px;border-top:2px solid var(--border);margin:10px 0"></div>';
  if (r.tdsPaid) h += row('Less: TDS', fmt(-r.tdsPaid), 'sub');
  if (r.tcsPaid) h += row('Less: TCS', fmt(-r.tcsPaid), 'sub');
  if (r.advTax) h += row('Less: Advance Tax', fmt(-r.advTax), 'sub');

  const cls = r.netPayable > 0 ? 'payable' : 'refund';
  const lbl = r.netPayable > 0 ? 'Net Tax Payable' : 'Refund Due';
  h += row(lbl, fmt(Math.abs(r.netPayable)), 'total ' + cls);
  return h;
}

function renderSlabs(result) {
  const slabs = result.slabResult.breakup;
  let h = '<table class="slab-table"><thead><tr><th>Slab</th><th style="text-align:right">Income</th><th style="text-align:right">Rate</th><th style="text-align:right">Tax</th></tr></thead><tbody>';
  for (const s of slabs) {
    h += '<tr><td>₹'+Math.round(s.from).toLocaleString('en-IN')+' – ₹'+(s.to===Infinity?'∞':Math.round(s.to).toLocaleString('en-IN'))+'</td>';
    h += '<td class="amt">'+fmt(s.amount)+'</td><td class="amt">'+(s.rate*100).toFixed(0)+'%</td><td class="amt">'+fmt(s.tax)+'</td></tr>';
  }
  h += '<tr style="font-weight:800;border-top:2px solid var(--border)"><td>Total</td><td></td><td></td><td class="amt">'+fmt(result.slabResult.tax)+'</td></tr>';
  h += '</tbody></table>';
  return h;
}

/* ── MAT/AMT in results panel ─────────────────────────────────────── */
function renderMatAmtResult(result) {
  const t = document.getElementById('assesseeType').value;
  const ac = ASSESSEE_CFG[t];
  const matCard = document.getElementById('matAmtResultCard');

  if (ac.canMat) {
    const bookProfit = parseFloat(document.getElementById('matBookProfit').value) || 0;
    if (!bookProfit) { matCard.style.display = 'none'; return; }
    const matRate = parseFloat(document.getElementById('matRate').value) || 0.15;
    const matBase = bookProfit * matRate;
    let surcharge = 0;
    if (ac.surchargeRateLow && bookProfit > (ac.surchargeThreshold||1e7)) surcharge = matBase * (ac.surchargeRateHigh||0);
    const matCess = (matBase + surcharge) * 0.04;
    const matTotal = matBase + surcharge + matCess;
    const normalTax = result.totalTax;
    const isMatApplicable = matTotal > normalTax;

    document.getElementById('matAmtResTitle').textContent = 'MAT Summary u/s 115JB';
    let h = '';
    h += `<div style="display:flex;gap:10px;margin-bottom:12px;flex-wrap:wrap">
      <div style="flex:1;min-width:120px;padding:12px;background:#EFF6FF;border-radius:10px;text-align:center">
        <div style="font-size:11px;color:var(--muted);margin-bottom:4px">Normal Tax</div>
        <div style="font-size:16px;font-weight:800;color:var(--brand)">${fmt(normalTax)}</div>
      </div>
      <div style="flex:1;min-width:120px;padding:12px;background:#FEF3C7;border-radius:10px;text-align:center">
        <div style="font-size:11px;color:var(--muted);margin-bottom:4px">MAT</div>
        <div style="font-size:16px;font-weight:800;color:#B45309">${fmt(matTotal)}</div>
      </div>
      <div style="flex:1;min-width:120px;padding:12px;background:${isMatApplicable?'#FEF3C7':'#F0FDF4'};border-radius:10px;text-align:center">
        <div style="font-size:11px;color:var(--muted);margin-bottom:4px">Tax Payable</div>
        <div style="font-size:16px;font-weight:800;color:${isMatApplicable?'#B45309':'#065F46'}">${fmt(Math.max(normalTax, matTotal))}</div>
      </div>
    </div>`;
    if (isMatApplicable) {
      h += `<div style="padding:10px 14px;background:#FEF3C7;border:1px solid #FDE68A;border-radius:8px;font-size:12px;color:#92400E;font-weight:600">
        ⚠️ MAT applies — MAT (${fmt(matTotal)}) &gt; Normal Tax (${fmt(normalTax)}). MAT Credit u/s 115JAA = ${fmt(matTotal - normalTax)} (carry fwd up to 15 years).
      </div>`;
    } else {
      h += `<div style="padding:10px 14px;background:#F0FDF4;border:1px solid #BBF7D0;border-radius:8px;font-size:12px;color:#065F46;font-weight:600">
        ✅ Normal Tax (${fmt(normalTax)}) &gt; MAT (${fmt(matTotal)}). Normal tax provisions apply.
      </div>`;
    }
    document.getElementById('matAmtResBody').innerHTML = h;
    matCard.style.display = 'block';
  } else {
    matCard.style.display = 'none';
  }
}

/* ── CALCULATE ────────────────────────────────────────────────────── */
function calculateTax() {
  const c = cfg();
  const panel = document.getElementById('resultPanel');
  const preInfo = document.getElementById('preCalcInfo');
  const ac = getAssesseeCfg();
  const isCo  = ac.group === 'company';
  const isFirm = ac.group === 'firm';

  if (currentRegime === 'both' && !isCo && !isFirm) {
    const rNew = computeForRegime(true);
    const rOld = computeForRegime(false);

    document.getElementById('singleResult').style.display = 'none';
    document.getElementById('compareResult').style.display = 'block';
    document.getElementById('compareYearLabel').textContent = c.label + (c.isFuture ? ' (Estimated)' : '');

    const diff = Math.abs(rNew.totalTax - rOld.totalTax);
    const winner = rNew.totalTax <= rOld.totalTax ? 'New Regime' : 'Old Regime';
    document.getElementById('regimeWinner').innerHTML = '<strong>' + winner + '</strong> saves you more tax';
    document.getElementById('savingsAmt').textContent = 'Save ' + fmt(diff);

    let ct = '';
    ct += cmpRow('Taxable Income', rNew.normalTaxable, rOld.normalTaxable, true);
    ct += cmpRow('Tax on Normal Income', rNew.normalTaxAfterRebate, rOld.normalTaxAfterRebate, true);
    ct += cmpRow('Tax on Special Income', rNew.totalSpecialTax, rOld.totalSpecialTax, true);
    ct += cmpRow('Surcharge', rNew.totalSurcharge, rOld.totalSurcharge, true);
    ct += cmpRow('Cess', rNew.cess, rOld.cess, true);
    ct += cmpRow('Total Tax', rNew.totalTax, rOld.totalTax, true);
    ct += cmpRow('Net Payable/Refund', rNew.netPayable, rOld.netPayable, true);
    document.getElementById('compareBody').innerHTML = ct;

    document.getElementById('newRegimeDetail').innerHTML = renderResult(rNew);
    document.getElementById('oldRegimeDetail').innerHTML = renderResult(rOld);

    document.getElementById('slabRegimeLabel').textContent = 'Tax Year ' + c.ayLabel.replace('AY ','');
    document.getElementById('slabBody').innerHTML =
      '<h3 style="font-size:13px;font-weight:700;margin-bottom:8px">🆕 New Regime Slabs</h3>' +
      renderSlabs(rNew) +
      '<div style="height:16px"></div>' +
      '<h3 style="font-size:13px;font-weight:700;margin-bottom:8px">📜 Old Regime Slabs</h3>' +
      renderSlabs(rOld);

    // Regime bar chart
    setTimeout(() => renderRegimeBarChart(rNew, rOld), 400);

    // Advance tax for 2026-27
    renderAdvanceTaxSchedule(rNew.totalTax, rNew.tdsPaid, rNew.tcsPaid);
    renderMatAmtResult(rNew);
  } else {
    const isNew = currentRegime === 'new' || isCo || isFirm;
    const result = computeForRegime(isNew);

    document.getElementById('singleResult').style.display = 'block';
    document.getElementById('compareResult').style.display = 'none';

    let label = '';
    if (isCo) label = '🏢 Company';
    else if (isFirm) label = '🤝 Firm / LLP';
    else label = isNew ? '🆕 New Regime' : '📜 Old Regime';

    document.getElementById('resultTitle').textContent = 'Tax Computation — ' + label;
    document.getElementById('resultSubtitle').textContent =
      (result.name ? result.name + ' · ' : '') + c.label + (c.isFuture ? ' (Estimated)' : '');

    document.getElementById('resultBody').innerHTML = renderResult(result);
    document.getElementById('slabRegimeLabel').textContent = (isCo ? 'Company' : isFirm ? 'Firm/LLP' : (isNew ? 'New Regime' : 'Old Regime')) + ' · Tax Year ' + c.ayLabel.replace('AY ','');
    document.getElementById('slabBody').innerHTML = renderSlabs(result);

    // Advance tax for 2026-27
    renderAdvanceTaxSchedule(result.totalTax, result.tdsPaid, result.tcsPaid);
    renderMatAmtResult(result);
  }

  document.getElementById('futureDisclaimer').style.display = c.isFuture ? 'block' : 'none';

  panel.classList.add('show');
  preInfo.style.display = 'none';
  document.getElementById('slabCard').style.display = 'block';
  document.getElementById('result-section').scrollIntoView({ behavior: 'smooth', block: 'start' });
  toast('Tax calculated for ' + c.label + '!');
}

function cmpRow(label, valNew, valOld, lowerBetter) {
  const nw = fmt(valNew), ol = fmt(valOld);
  let nCls = '', oCls = '';
  if (lowerBetter) {
    if (valNew < valOld) nCls = 'winner'; else if (valOld < valNew) oCls = 'winner';
  }
  return '<tr><td style="text-align:left;font-weight:500">'+label+'</td><td class="'+nCls+'">'+nw+'</td><td class="'+oCls+'">'+ol+'</td></tr>';
}

function resetForm() {
  document.querySelectorAll('input[type=number]').forEach(i => {
    if (i.id === 'stdDeduction') i.value = cfg().stdDeduction;
    else i.value = '';
  });
  document.getElementById('assesseeName').value = '';
  // Reset assessee buttons
  document.querySelectorAll('.at-btn').forEach(b => b.classList.remove('active'));
  const firstBtn = document.querySelector('.at-btn[data-val="individual_below60"]');
  if (firstBtn) firstBtn.classList.add('active');
  document.getElementById('assesseeType').value = 'individual_below60';
  document.getElementById('resultPanel').classList.remove('show');
  document.getElementById('preCalcInfo').style.display = 'block';
  document.getElementById('singleResult').style.display = 'none';
  document.getElementById('compareResult').style.display = 'none';
  document.getElementById('advanceTaxCard').style.display = 'none';
  document.getElementById('matAmtResultCard').style.display = 'none';
  onAssesseeTypeChange();
  setRegime('new');
  toast('Form reset');
}

function toast(msg) {
  const t = document.getElementById('toast');
  t.textContent = msg;
  t.classList.add('show');
  setTimeout(() => t.classList.remove('show'), 3000);
}

// Initialize
setRegime('new');
updateRefSlabs();
onAssesseeTypeChange();

/* ═══════════════════════════════════════════
   ANIMATION ENGINE
   ═══════════════════════════════════════════ */

/* ── 1. Scroll reveal ── */
function initReveal() {
  const obs = new IntersectionObserver((entries) => {
    entries.forEach(e => { if (e.isIntersecting) { e.target.classList.add('visible'); obs.unobserve(e.target); } });
  }, { threshold: 0.08 });
  document.querySelectorAll('.reveal').forEach(el => obs.observe(el));
}
initReveal();

/* ── 2. Nav scroll shadow ── */
window.addEventListener('scroll', () => {
  document.querySelector('nav').classList.toggle('scrolled', window.scrollY > 10);
});

/* ── 3. Progress bar ── */
function showProgress() {
  const bar = document.getElementById('calcProgress');
  bar.style.width = '0';
  bar.style.transition = 'none';
  requestAnimationFrame(() => {
    bar.style.transition = 'width .35s ease';
    bar.style.width = '70%';
    setTimeout(() => { bar.style.width = '95%'; }, 350);
  });
}
function finishProgress() {
  const bar = document.getElementById('calcProgress');
  bar.style.width = '100%';
  setTimeout(() => { bar.style.width = '0'; bar.style.transition = 'none'; }, 400);
}

/* ── 4. Button ripple ── */
function addRipple(btn, e) {
  const rect = btn.getBoundingClientRect();
  const size = Math.max(rect.width, rect.height) * 2;
  const r = document.createElement('span');
  r.className = 'ripple';
  r.style.cssText = `width:${size}px;height:${size}px;left:${e.clientX - rect.left - size/2}px;top:${e.clientY - rect.top - size/2}px`;
  btn.appendChild(r);
  r.addEventListener('animationend', () => r.remove());
}

/* ── 5. Number counter ── */
function animateCounter(el, target, prefix, suffix, duration) {
  const start = performance.now();
  const startVal = 0;
  function update(now) {
    const progress = Math.min((now - start) / duration, 1);
    const ease = 1 - Math.pow(1 - progress, 3);
    const current = Math.round(startVal + (target - startVal) * ease);
    el.textContent = prefix + current.toLocaleString('en-IN') + suffix;
    if (progress < 1) requestAnimationFrame(update);
  }
  requestAnimationFrame(update);
}

function animateAllCounters() {
  document.querySelectorAll('.val').forEach(el => {
    const text = el.textContent.trim();
    const isNeg = text.startsWith('-₹');
    const clean = text.replace(/[₹,\-]/g, '');
    const num = parseFloat(clean);
    if (!isNaN(num) && num > 0) {
      animateCounter(el, num, isNeg ? '-₹' : '₹', '', 900);
    }
  });
}

/* ── 6. Donut chart ── */
const CIRC = 2 * Math.PI * 54; // 339.3
const DONUT_COLORS = ['#2563EB', '#F59E0B', '#10B981', '#EF4444'];
const DONUT_LABELS = ['Base Tax', 'Surcharge', 'Cess', 'Special Rate Tax'];
const DONUT_IDS    = ['donut-base', 'donut-surcharge', 'donut-cess', 'donut-special'];

function renderDonutChart(baseTax, surcharge, cess, specialTax) {
  const wrap = document.getElementById('taxChartWrap');
  const total = baseTax + surcharge + cess + specialTax;
  if (total <= 0) { wrap.style.display = 'none'; return; }
  wrap.style.display = 'block';

  const values = [baseTax, surcharge, cess, specialTax];
  document.getElementById('donut-center-val').textContent = '₹' + Math.round(total).toLocaleString('en-IN');

  // Build SVG segments — need 4 circles layered with correct dashoffset
  // Remove old special segment if exists
  const oldSpecial = document.getElementById('donut-special');
  if (oldSpecial) oldSpecial.remove();

  let offsetDeg = 0; // starts at top (adjusted by -85° in CSS)
  const ids = ['donut-base', 'donut-surcharge', 'donut-cess'];
  const colors = ['#2563EB', '#F59E0B', '#10B981'];
  const mainVals = [baseTax, surcharge, cess];

  // Also add special if needed
  if (specialTax > 0) {
    const svg = document.querySelector('.donut-svg');
    const circle = document.createElementNS('http://www.w3.org/2000/svg', 'circle');
    circle.setAttribute('id', 'donut-special');
    circle.setAttribute('class', 'donut-segment');
    circle.setAttribute('cx', '70'); circle.setAttribute('cy', '70'); circle.setAttribute('r', '54');
    circle.setAttribute('fill', 'none'); circle.setAttribute('stroke', '#EF4444');
    circle.setAttribute('stroke-width', '22');
    circle.setAttribute('stroke-dasharray', '0 339.3');
    circle.setAttribute('stroke-dashoffset', '84.8');
    circle.setAttribute('stroke-linecap', 'round');
    svg.appendChild(circle);
    ids.push('donut-special'); colors.push('#EF4444'); mainVals.push(specialTax);
  }

  let cumOffset = CIRC * 0.25; // start at top
  ids.forEach((id, i) => {
    const el = document.getElementById(id);
    if (!el) return;
    const pct = mainVals[i] / total;
    const dash = pct * CIRC;
    const gap  = CIRC - dash;
    el.setAttribute('stroke', colors[i]);
    // Animate after paint
    setTimeout(() => {
      el.style.strokeDasharray = `${dash} ${gap}`;
      el.style.strokeDashoffset = cumOffset;
    }, 80 + i * 60);
    cumOffset -= dash;
  });

  // Legend
  const legend = document.getElementById('donutLegend');
  const allLabels = ['Base Tax', 'Surcharge', 'Cess', 'Special Rate'];
  legend.innerHTML = ids.map((id, i) => {
    if (mainVals[i] <= 0) return '';
    const pct = ((mainVals[i] / total) * 100).toFixed(1);
    return `<div class="donut-legend-item">
      <span class="donut-dot" style="background:${colors[i]}"></span>
      <span class="donut-label">${allLabels[i]}</span>
      <span class="donut-val">${pct}%</span>
    </div>`;
  }).join('');
}

/* ── 7. Regime bar chart ── */
function renderRegimeBarChart(rNew, rOld) {
  const wrap = document.getElementById('regimeChartWrap');
  const chart = document.getElementById('regimeBarChart');
  const maxVal = Math.max(rNew.totalTax, rOld.totalTax, 1);

  const rows = [
    { label: 'Taxable Income', nv: rNew.normalTaxable, ov: rOld.normalTaxable },
    { label: 'Base Tax', nv: rNew.normalTaxAfterRebate, ov: rOld.normalTaxAfterRebate },
    { label: 'Total Tax', nv: rNew.totalTax, ov: rOld.totalTax },
    { label: 'Net Payable', nv: Math.max(0,rNew.netPayable), ov: Math.max(0,rOld.netPayable) },
  ];

  const maxAll = Math.max(...rows.map(r => Math.max(r.nv, r.ov)), 1);

  chart.innerHTML = rows.map(r => {
    const nPct = (r.nv / maxAll * 100).toFixed(1);
    const oPct = (r.ov / maxAll * 100).toFixed(1);
    const nWinner = r.nv <= r.ov;
    return `<div style="margin-bottom:14px">
      <div style="font-size:11px;font-weight:700;color:var(--muted);margin-bottom:6px;text-transform:uppercase;letter-spacing:.04em">${r.label}</div>
      <div class="bar-row">
        <div class="bar-label" style="color:#2563EB;font-size:10px">🆕 New</div>
        <div class="bar-track">
          <div class="bar-fill" style="background:${nWinner?'#2563EB':'#93C5FD'}" data-pct="${nPct}">
            <span>₹${Math.round(r.nv/1000)}K</span>
          </div>
        </div>
        <div class="bar-val" style="color:${nWinner?'#2563EB':'var(--muted)'}">₹${Math.round(r.nv).toLocaleString('en-IN')}</div>
      </div>
      <div class="bar-row" style="margin-top:4px">
        <div class="bar-label" style="color:#F59E0B;font-size:10px">📜 Old</div>
        <div class="bar-track">
          <div class="bar-fill" style="background:${!nWinner?'#F59E0B':'#FCD34D'}" data-pct="${oPct}">
            <span>₹${Math.round(r.ov/1000)}K</span>
          </div>
        </div>
        <div class="bar-val" style="color:${!nWinner?'#F59E0B':'var(--muted)'}">₹${Math.round(r.ov).toLocaleString('en-IN')}</div>
      </div>
    </div>`;
  }).join('');

  // Animate bars after DOM paint
  requestAnimationFrame(() => {
    setTimeout(() => {
      document.querySelectorAll('.bar-fill').forEach(bar => {
        bar.style.width = bar.dataset.pct + '%';
      });
    }, 100);
  });
}

/* ── 8. Confetti burst ── */
function fireConfetti() {
  const colors = ['#2563EB','#10B981','#F59E0B','#EF4444','#8B5CF6','#EC4899'];
  for (let i = 0; i < 60; i++) {
    const piece = document.createElement('div');
    piece.className = 'confetti-piece';
    piece.style.cssText = `
      left: ${Math.random()*100}vw;
      background: ${colors[Math.floor(Math.random()*colors.length)]};
      width: ${4+Math.random()*6}px;
      height: ${4+Math.random()*6}px;
      border-radius: ${Math.random()>.5?'50%':'2px'};
      animation-duration: ${1.5+Math.random()*2}s;
      animation-delay: ${Math.random()*.5}s;
      opacity: 1;
    `;
    document.body.appendChild(piece);
    piece.addEventListener('animationend', () => piece.remove());
  }
}

/* ── Override calculateTax to wire animations ── */
const _origCalc = calculateTax;
calculateTax = function(e) {
  const btn = document.getElementById('calcBtn');

  // Ripple
  if (e && btn) addRipple(btn, e);

  // Spinner
  if (btn) btn.classList.add('loading');

  // Progress bar
  showProgress();

  // Small delay to show spinner, then compute
  setTimeout(() => {
    _origCalc();
    if (btn) btn.classList.remove('loading');
    finishProgress();

    // Counter animation
    setTimeout(animateAllCounters, 200);

    // Add reveal to result cards
    document.querySelectorAll('#resultPanel .card, #resultPanel > div').forEach((el, i) => {
      el.classList.add('reveal');
      el.style.transitionDelay = (i * 0.07) + 's';
      setTimeout(() => el.classList.add('visible'), 50 + i * 70);
    });

    // Wire up donut chart from result data
    // (called from calculateTax internals via hook below)

  }, 380);
};

/* Hook into renderResult to trigger donut */
const _origRenderResult = renderResult;
renderResult = function(r) {
  const html = _origRenderResult(r);
  // Schedule donut render
  setTimeout(() => {
    const baseTax = r.normalTaxAfterRebate || r.normalTax || 0;
    const surcharge = r.totalSurcharge || 0;
    const cess = r.cess || 0;
    const special = r.totalSpecialTax || 0;
    renderDonutChart(baseTax, surcharge, cess, special);

    // Confetti if zero tax
    if ((r.totalTax || 0) === 0 || (r.netPayable || 0) <= 0) {
      fireConfetti();
    }
  }, 500);
  return html;
};

/* Hook into calculateTax for regime bar chart */
const _origCmpRow = cmpRow;
let _lastRNew = null, _lastROld = null;
const _origCalcTax2 = calculateTax;
// Patch the compare path via renderResult hook on compare
const _origCalcTaxFinal = calculateTax;
calculateTax = (function(prev) {
  return function(e) {
    prev(e);
    // Regime bar chart rendered after compute
    setTimeout(() => {
      if (document.getElementById('compareResult').style.display !== 'none') {
        // bar chart data is set by calculateTax — read from compare table
        const rows = document.querySelectorAll('#compareBody tr');
        if (rows.length >= 3) {
          // parse from DOM (simpler than re-running compute)
          document.getElementById('regimeChartWrap').style.display = 'block';
        }
      }
    }, 500);
  };
})(calculateTax);

/* Store last comparison data for bar chart */
const _origCalculateTaxInner = window.calculateTax;

</script>
<a href="https://wa.me/918427651580" target="_blank" class="wa-float" title="WhatsApp Support"><svg viewBox="0 0 24 24"><path d="M17.472 14.382c-.297-.149-1.758-.867-2.03-.967-.273-.099-.471-.148-.67.15-.197.297-.767.966-.94 1.164-.173.199-.347.223-.644.075-.297-.15-1.255-.463-2.39-1.475-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.298-.347.446-.52.149-.174.198-.298.298-.497.099-.198.05-.371-.025-.52-.075-.149-.669-1.612-.916-2.207-.242-.579-.487-.5-.669-.51-.173-.008-.371-.01-.57-.01-.198 0-.52.074-.792.372-.272.297-1.04 1.016-1.04 2.479 0 1.462 1.065 2.875 1.213 3.074.149.198 2.096 3.2 5.077 4.487.709.306 1.262.489 1.694.625.712.227 1.36.195 1.871.118.571-.085 1.758-.719 2.006-1.413.248-.694.248-1.289.173-1.413-.074-.124-.272-.198-.57-.347m-5.421 7.403h-.004a9.87 9.87 0 01-5.031-1.378l-.361-.214-3.741.982.998-3.648-.235-.374a9.86 9.86 0 01-1.51-5.26c.001-5.45 4.436-9.884 9.888-9.884 2.64 0 5.122 1.03 6.988 2.898a9.825 9.825 0 012.893 6.994c-.003 5.45-4.437 9.884-9.885 9.884m8.413-18.297A11.815 11.815 0 0012.05 0C5.495 0 .16 5.335.157 11.892c0 2.096.547 4.142 1.588 5.945L.057 24l6.305-1.654a11.882 11.882 0 005.683 1.448h.005c6.554 0 11.89-5.335 11.893-11.893a11.821 11.821 0 00-3.48-8.413z"/></svg></a>

<button class="help-btn" onclick="openHelp()" title="How to use this tool">?</button>
<div class="help-overlay" id="helpOverlay">
  <div class="help-modal">
    <div class="help-modal-head"><h3>How to Use — Income Tax Calculator</h3><button class="help-close" onclick="closeHelp()">&#10005;</button></div>
    <div class="help-modal-body"><div class="help-step"><div class="help-step-num">1</div><div class="help-step-body"><h4>Select Year</h4><p>Choose the Assessment Year and assessee type (Individual, HUF, Firm, Company, etc.).</p></div></div><div class="help-step"><div class="help-step-num">2</div><div class="help-step-body"><h4>Enter Income</h4><p>Fill income under Salary, House Property, Business/Profession, Capital Gains, and Other Sources.</p></div></div><div class="help-step"><div class="help-step-num">3</div><div class="help-step-body"><h4>Add Deductions</h4><p>Enter 80C, 80D, HRA, and other deductions (applicable under old regime).</p></div></div><div class="help-step"><div class="help-step-num">4</div><div class="help-step-body"><h4>View Result</h4><p>Tax under old and new regime is compared automatically side by side.</p></div></div><div class="help-step"><div class="help-step-num">5</div><div class="help-step-body"><h4>Advance Tax</h4><p>Scroll down to see the quarterly advance tax schedule.</p></div></div><div class="help-tip">⚠️ For estimation only. Verify with the latest CBDT notifications and consult a CA for complex cases.</div></div>
  </div>
</div>
<script>function openHelp(){document.getElementById('helpOverlay').classList.add('open')}function closeHelp(){document.getElementById('helpOverlay').classList.remove('open')}document.getElementById('helpOverlay').addEventListener('click',function(e){if(e.target===this)closeHelp()})</script>
</body></html>"""


# ══════════════════════════════════════════════════════════════════════════════
#  TDS CALCULATOR TEMPLATE
# ══════════════════════════════════════════════════════════════════════════════

TDS_CALC_T = r"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>TDS / TCS Calculator (IT Act 2025) – CA Toolkit</title>

<style>
""" + BASE_CSS + """
.nav-links{display:flex;gap:20px;list-style:none}
.nav-links a{text-decoration:none;color:var(--muted);font-size:13px;font-weight:500}
.nav-links a:hover{color:var(--brand)}
.hero{text-align:center;padding:32px 24px 16px;max-width:760px;margin:0 auto}
.hero-badge{display:inline-flex;align-items:center;gap:6px;background:#ECFDF5;color:#065F46;
            border:1px solid #A7F3D0;border-radius:99px;padding:5px 14px;font-size:12px;font-weight:600;margin-bottom:12px}
h1{font-size:clamp(20px,4vw,32px);font-weight:800;line-height:1.15;letter-spacing:-.5px;margin-bottom:8px}
h1 em{font-style:normal;color:var(--accent)}
.hero p{font-size:13px;color:var(--muted);line-height:1.7;max-width:520px;margin:0 auto}
.act-note{max-width:1100px;margin:0 auto;padding:0 24px 10px}
.act-box{background:#EFF6FF;border:1px solid #BFDBFE;border-radius:8px;padding:9px 14px;
         font-size:12px;color:#1e40af;display:flex;align-items:flex-start;gap:6px}
/* Toggle */
.toggle-wrap{max-width:1100px;margin:0 auto;padding:0 24px 16px;display:flex;gap:10px}
.toggle-btn{flex:1;padding:12px;border-radius:10px;border:2px solid var(--border);
            font-family:inherit;font-size:14px;font-weight:700;cursor:pointer;
            background:var(--white);color:var(--muted);transition:all .2s}
.toggle-btn.active{background:var(--brand);color:#fff;border-color:var(--brand)}
.toggle-btn:hover:not(.active){border-color:var(--brand);color:var(--brand)}
/* Layout */
.main{max-width:1100px;margin:0 auto;padding:0 24px 48px;
      display:grid;grid-template-columns:1.1fr 1fr;gap:20px;align-items:start}
@media(max-width:800px){.main{grid-template-columns:1fr}}
.card{background:var(--white);border-radius:var(--radius);border:1px solid var(--border);
      box-shadow:var(--shadow);overflow:hidden;margin-bottom:16px}
.card:last-child{margin-bottom:0}
.card-head{padding:14px 18px;border-bottom:1px solid var(--border);display:flex;align-items:center;gap:10px}
.card-head .icon{width:30px;height:30px;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:15px}
.card-head h2{font-size:14px;font-weight:700}
.card-head p{font-size:12px;color:var(--muted);margin-top:1px}
.card-body{padding:16px}
.field{margin-bottom:13px}
.row2{display:grid;grid-template-columns:1fr 1fr;gap:12px}
label{display:block;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.04em;color:var(--muted);margin-bottom:4px}
.hint{font-size:11px;color:var(--muted);margin-top:3px}
select,input[type=number],input[type=date]{width:100%;border:1.5px solid var(--border);border-radius:8px;
  padding:8px 11px;font-family:inherit;font-size:13px;color:var(--ink);background:var(--white);
  transition:border-color .2s;outline:none}
select:focus,input:focus{border-color:var(--brand)}
.btn{width:100%;background:var(--brand);color:#fff;border:none;border-radius:8px;
     padding:11px;font-family:inherit;font-size:14px;font-weight:700;cursor:pointer;transition:background .2s}
.btn:hover{background:var(--brand-d)}
/* Results */
.result-section{display:none;margin-top:14px}
.rboxes{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:12px}
.rbox{border-radius:10px;padding:14px 16px}
.rbox-main {background:#EFF6FF;border:1.5px solid #BFDBFE}
.rbox-int  {background:#FFFBEB;border:1.5px solid #FDE68A}
.rbox-total{background:#1D4ED8;border:1.5px solid #1D4ED8;grid-column:1/-1}
.rbox .val {font-size:22px;font-weight:800;margin-bottom:2px}
.rbox .lbl {font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;opacity:.75}
.rbox .sub {font-size:11px;margin-top:5px;opacity:.8}
.rbox-main  .val{color:#1D4ED8}.rbox-main  .lbl{color:#1D4ED8}
.rbox-int   .val{color:#92400E}.rbox-int   .lbl{color:#92400E}
.rbox-total .val{color:#fff;font-size:26px}
.rbox-total .lbl{color:rgba(255,255,255,.75)}
.rbox-total .sub{color:rgba(255,255,255,.8);font-size:12px}
.detail-table{width:100%;border-collapse:collapse;font-size:12px;margin-top:10px}
.detail-table td{padding:7px 2px;border-bottom:1px solid var(--border)}
.detail-table tr:last-child td{border:none;font-weight:700;font-size:13px}
.detail-table td:last-child{text-align:right;font-weight:600}
.ontime-box{background:#ECFDF5;border:1.5px solid #A7F3D0;border-radius:8px;
            padding:12px 14px;font-size:13px;color:#065F46;font-weight:600;margin-top:14px;text-align:center}
.note-box{background:#FFFBEB;border:1px solid #FDE68A;border-radius:8px;
          padding:10px 12px;font-size:11px;color:#92400E;margin-top:10px;line-height:1.6}
.info-box{background:#EFF6FF;border:1px solid #BFDBFE;border-radius:8px;
          padding:10px 12px;font-size:11px;color:#1e40af;margin-top:10px;line-height:1.6}
/* Rate tables */
.rate-table{width:100%;border-collapse:collapse;font-size:11px}
.rate-table th{text-align:left;font-size:10px;letter-spacing:.06em;text-transform:uppercase;
               color:var(--muted);border-bottom:1.5px solid var(--border);padding:5px 6px}
.rate-table td{padding:6px;border-bottom:1px solid var(--border);vertical-align:top;line-height:1.5}
.rate-table tr:last-child td{border:none}
.rate-table tr:hover td{background:#F9FAFB}
.code{background:#EFF6FF;color:var(--brand);font-size:10px;font-weight:700;
      padding:1px 5px;border-radius:4px;font-family:monospace;white-space:nowrap}
.tcs-code{background:#F5F3FF;color:#5B21B6;font-size:10px;font-weight:700;
          padding:1px 5px;border-radius:4px;font-family:monospace;white-space:nowrap}
footer{background:#0f1b2d;color:#9CA3AF;font-size:12px;padding:0}
.ft-main{display:grid;grid-template-columns:2fr 1fr 1.4fr;gap:40px;padding:40px 48px;max-width:1200px;margin:0 auto}
.ft-brand-name{color:#fff;font-size:18px;font-weight:800;margin-bottom:12px}
.ft-brand-desc{font-size:12.5px;line-height:1.75;color:#9CA3AF;max-width:340px;text-align:justify}
.ft-col-title{color:#fff;font-size:14px;font-weight:700;margin-bottom:14px}
.ft-links{list-style:none;padding:0;margin:0}
.ft-links li{margin-bottom:8px}
.ft-links a{color:#9CA3AF;text-decoration:none;font-size:13px;transition:color .2s}
.ft-links a:hover{color:#fff}
.ft-contact-name{color:#fff;font-weight:700;font-size:13px;margin-bottom:6px}
.ft-contact-addr{color:#9CA3AF;font-size:12px;line-height:1.7;margin-bottom:10px}
.ft-contact-line{color:#9CA3AF;font-size:12px;margin-bottom:4px}
.ft-socials{display:flex;gap:14px;margin-top:12px}
.ft-socials a{color:#9CA3AF;transition:color .2s}
.ft-socials a:hover{color:#fff}
.ft-socials svg{width:20px;height:20px;fill:currentColor}
.ft-bottom{background:#0a1422;border-top:1px solid #1e2d42;padding:12px 48px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px}
.ft-bottom-left{font-size:11px;color:#6B7280}
.ft-bottom-right{font-size:11px;color:#6B7280}
@media(max-width:768px){.ft-main{grid-template-columns:1fr;padding:28px 20px;gap:24px}.ft-bottom{padding:12px 20px;flex-direction:column;text-align:center}}
</style></head><body>

<nav>
  <a href="/" class="logo">CA<span class="logo-dot"></span><span>Toolkit</span></a>
  
  <div class="nav-right">
    {% if username %}<div class="nav-user"><span class="nav-avatar">{{ username[0].upper() }}</span><strong>{{ username }}</strong></div><div class="nav-sep"></div>{% if is_admin %}<a href="/admin" class="nav-btn ghost">⚙ Admin</a>{% endif %}{% endif %}
    <a href="/" class="nav-btn dash">⬅ Dashboard</a>
    {% if username %}<a href="/logout" class="nav-link">Sign out</a>{% else %}<a href="/login" class="nav-btn">Sign In →</a>{% endif %}
  </div>
</nav>

<section class="hero">
  <div class="hero-badge">🆓 Free · No Login Required</div>
  <h1>TDS / TCS Calculator — <em>IT Act 2025</em></h1>
  <p>Calculate TDS or TCS liability, late deposit interest and total payable amount as per new Section 393 / Section 394 of IT Act 2025.</p>
</section>

<div class="act-note">
  <div class="act-box">ℹ️ <span><strong>IT Act 2025 (w.e.f. 1 Apr 2026):</strong> TDS consolidated under Section 393 (non-salary) &amp; Section 392 (salary). TCS under Section 394. Numeric payment codes replace old section numbers in returns. Rates &amp; thresholds unchanged.</span></div>
</div>

<!-- TDS / TCS TOGGLE -->
<div class="toggle-wrap">
  <button class="toggle-btn active" id="btnTDS" onclick="switchMode('tds')">📑 TDS — Tax Deducted at Source</button>
  <button class="toggle-btn" id="btnTCS" onclick="switchMode('tcs')">🧾 TCS — Tax Collected at Source</button>
</div>

<div class="main">
  <!-- LEFT: INPUT -->
  <div>
    <div class="card">
      <div class="card-head">
        <div class="icon" style="background:#EFF6FF" id="formIcon">📑</div>
        <div>
          <h2 id="formTitle">TDS Calculator</h2>
          <p id="formSub">IT Act 2025 · Tax Year 2026-27</p>
        </div>
      </div>
      <div class="card-body">

        <div class="field">
          <label id="sectionLabel">Nature of Payment (TDS)</label>
          <select id="mainSection" onchange="updateHint()">
            <option value="">— Select Payment Type —</option>
          </select>
          <p class="hint" id="sectionHint">Select a payment type to see rate and threshold</p>
        </div>

        <div class="field">
          <label id="amtLabel">Payment Amount (₹)</label>
          <input type="number" id="paymentAmt" placeholder="e.g. 100000" min="0"/>
          <p class="hint" id="amtHint">Gross payment amount before TDS deduction</p>
        </div>

        <hr style="border:none;border-top:1.5px dashed var(--border);margin:14px 0"/>
        <p style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;color:var(--muted);margin-bottom:12px">Late Deposit Interest (Optional)</p>

        <div class="row2">
          <div class="field">
            <label id="d1label">Date of Deduction</label>
            <input type="date" id="deductionDate"/>
            <p class="hint" id="d1hint">When TDS was deducted</p>
          </div>
          <div class="field">
            <label>Date of Actual Deposit</label>
            <input type="date" id="depositDate"/>
            <p class="hint">When you paid the challan</p>
          </div>
        </div>

        <button class="btn" id="calcBtn" onclick="calculate()">Calculate TDS &amp; Interest →</button>

        <!-- RESULTS -->
        <div class="result-section" id="resultSection">
          <div class="ontime-box" id="ontimeBox" style="display:none"></div>
          <div id="lateBoxes" style="display:none">
            <div class="rboxes">
              <div class="rbox rbox-main">
                <div class="lbl" id="r-main-lbl">TDS Amount</div>
                <div class="val" id="r-main"></div>
                <div class="sub" id="r-main-sub"></div>
              </div>
              <div class="rbox rbox-int">
                <div class="lbl">Interest u/s 201(1A)</div>
                <div class="val" id="r-int"></div>
                <div class="sub" id="r-int-sub"></div>
              </div>
              <div class="rbox rbox-total">
                <div class="lbl">Total Amount Payable</div>
                <div class="val" id="r-total"></div>
                <div class="sub" id="r-total-sub"></div>
              </div>
            </div>
            <table class="detail-table">
              <tr><td id="d-amt-lbl">Payment Amount</td><td id="d-payment"></td></tr>
              <tr><td>New Section (IT Act 2025)</td><td id="d-newsec"></td></tr>
              <tr><td>Old Section (for reference)</td><td id="d-oldsec"></td></tr>
              <tr><td>Payment Code</td><td id="d-code"></td></tr>
              <tr><td id="d-rate-lbl">TDS Rate</td><td id="d-rate"></td></tr>
              <tr><td id="d-tax-lbl">TDS Amount</td><td id="d-tds"></td></tr>
              <tr><td id="d-date1-lbl">Date of Deduction</td><td id="d-ddate"></td></tr>
              <tr><td>Due Date for Deposit</td><td id="d-due"></td></tr>
              <tr><td>Actual Deposit Date</td><td id="d-adate"></td></tr>
              <tr><td>Delay (months)</td><td id="d-months"></td></tr>
              <tr><td>Interest Rate</td><td>1.5% per month</td></tr>
              <tr><td>Interest Amount</td><td id="d-intamt"></td></tr>
              <tr><td id="d-total-lbl" style="color:var(--brand)">Total Payable</td><td id="d-total" style="color:var(--brand)"></td></tr>
            </table>
            <div class="note-box">⚠ As per IT Act 2025, a fractional month is counted as a full month for interest calculation. Interest runs from date of deduction/collection to actual date of deposit.</div>
          </div>
          <div id="basicBox" style="display:none">
            <div class="rboxes">
              <div class="rbox rbox-main" style="grid-column:1/-1">
                <div class="lbl" id="b-main-lbl">TDS Amount</div>
                <div class="val" id="b-main"></div>
                <div class="sub" id="b-sub"></div>
              </div>
            </div>
            <table class="detail-table">
              <tr><td id="b-amt-lbl">Payment Amount</td><td id="b-payment"></td></tr>
              <tr><td>New Section (IT Act 2025)</td><td id="b-newsec"></td></tr>
              <tr><td>Old Section (for reference)</td><td id="b-oldsec"></td></tr>
              <tr><td>Payment Code</td><td id="b-code"></td></tr>
              <tr><td id="b-rate-lbl">TDS Rate</td><td id="b-rate"></td></tr>
              <tr><td id="b-tax-lbl">TDS Amount</td><td id="b-tds2"></td></tr>
              <tr><td id="b-net-lbl">Net Payment to Payee</td><td id="b-net"></td></tr>
            </table>
            <div class="info-box">ℹ Enter deduction and deposit dates above to also calculate late deposit interest.</div>
          </div>
        </div>

      </div>
    </div>
  </div>

  <!-- RIGHT: RATE CHARTS + DUE DATES -->
  <div>
    <!-- TDS rate chart -->
    <div id="tdsRateCard" class="card">
      <div class="card-head">
        <div class="icon" style="background:#FFFBEB">📋</div>
        <div><h2>TDS Quick Rate Chart</h2><p>Section 393 — IT Act 2025</p></div>
      </div>
      <div class="card-body" style="padding:0;overflow-x:auto">
        <table class="rate-table">
          <thead><tr><th>Code</th><th>Old Sec</th><th>Nature</th><th>Rate</th><th>Threshold</th></tr></thead>
          <tbody>
            <tr><td><span class="code">1001</span></td><td>192</td><td>Salary</td><td>Slab</td><td>Basic exemption</td></tr>
            <tr><td><span class="code">1021</span></td><td>194A</td><td>Interest (Bank/PO)</td><td>10%</td><td>₹50,000</td></tr>
            <tr><td><span class="code">1022</span></td><td>194A</td><td>Interest (Others)</td><td>10%</td><td>₹10,000</td></tr>
            <tr><td><span class="code">1023</span></td><td>194C</td><td>Contractor (Ind)</td><td>1%</td><td>₹30K/₹1L pa</td></tr>
            <tr><td><span class="code">1024</span></td><td>194C</td><td>Contractor (Others)</td><td>2%</td><td>₹30K/₹1L pa</td></tr>
            <tr><td><span class="code">1006</span></td><td>194H</td><td>Commission/Brokerage</td><td>2%</td><td>₹20,000</td></tr>
            <tr><td><span class="code">1008</span></td><td>194I(a)</td><td>Rent (P&amp;M)</td><td>2%</td><td>₹50K/mo</td></tr>
            <tr><td><span class="code">1009</span></td><td>194I(b)</td><td>Rent (Land/Bldg)</td><td>10%</td><td>₹50K/mo</td></tr>
            <tr><td><span class="code">1036</span></td><td>194IA</td><td>Immovable Property</td><td>1%</td><td>₹50L</td></tr>
            <tr><td><span class="code">1027</span></td><td>194J(b)</td><td>Professional Fees</td><td>10%</td><td>₹50,000</td></tr>
            <tr><td><span class="code">1026</span></td><td>194J(a)</td><td>Technical Services</td><td>2%</td><td>₹50,000</td></tr>
            <tr><td><span class="code">1031</span></td><td>194Q</td><td>Purchase of Goods</td><td>0.1%</td><td>₹50L pa</td></tr>
            <tr><td><span class="code">1039</span></td><td>194S</td><td>VDA/Crypto</td><td>1%</td><td>₹10,000</td></tr>
            <tr><td><span class="code">1041</span></td><td>194T</td><td>Partner Salary</td><td>10%</td><td>₹20,000</td></tr>
          </tbody>
        </table>
      </div>
    </div>

    <!-- TCS rate chart -->
    <div id="tcsRateCard" class="card" style="display:none">
      <div class="card-head">
        <div class="icon" style="background:#F5F3FF">🧾</div>
        <div><h2>TCS Quick Rate Chart</h2><p>Section 394 — IT Act 2025</p></div>
      </div>
      <div class="card-body" style="padding:0;overflow-x:auto">
        <table class="rate-table">
          <thead><tr><th>Code</th><th>Old Sec</th><th>Nature of Goods/Transaction</th><th>Rate</th><th>Threshold</th></tr></thead>
          <tbody>
            <tr><td><span class="tcs-code">2001</span></td><td>206C(1)</td><td>Alcoholic Liquor for Human Consumption</td><td>2%</td><td>Nil</td></tr>
            <tr><td><span class="tcs-code">2002</span></td><td>206C(1)</td><td>Tendu Leaves</td><td>2%</td><td>Nil</td></tr>
            <tr><td><span class="tcs-code">2003</span></td><td>206C(1)</td><td>Timber (forest lease)</td><td>2%</td><td>Nil</td></tr>
            <tr><td><span class="tcs-code">2004</span></td><td>206C(1)</td><td>Timber (other than forest lease)</td><td>2%</td><td>Nil</td></tr>
            <tr><td><span class="tcs-code">2005</span></td><td>206C(1)</td><td>Any other forest produce</td><td>2%</td><td>Nil</td></tr>
            <tr><td><span class="tcs-code">2006</span></td><td>206C(1)</td><td>Scrap</td><td>2%</td><td>Nil</td></tr>
            <tr><td><span class="tcs-code">2007</span></td><td>206C(1)</td><td>Minerals (coal/lignite/iron ore)</td><td>2%</td><td>Nil</td></tr>
            <tr><td><span class="tcs-code">2009</span></td><td>206C(1F)</td><td>Motor Vehicle &gt; ₹10L</td><td>1%</td><td>₹10L</td></tr>
            <tr><td><span class="tcs-code">2010</span></td><td>206C(1G)</td><td>Foreign Remittance (LRS) — Education/Medical</td><td>2%</td><td>₹10L</td></tr>
            <tr><td><span class="tcs-code">2011</span></td><td>206C(1G)</td><td>Foreign Remittance (LRS) — Other purposes</td><td>20%</td><td>₹10L</td></tr>
            <tr><td><span class="tcs-code">2012</span></td><td>206C(1G)</td><td>Overseas Tour Package</td><td>2%</td><td>Nil</td></tr>
            <tr><td><span class="tcs-code">2013</span></td><td>206C(1H)</td><td>Sale of Goods &gt; ₹50L</td><td>0.1%</td><td>₹50L pa</td></tr>
            <tr><td><span class="tcs-code">2014</span></td><td>206C(1)</td><td>Parking lot / Toll Plaza / Mining &amp; Quarrying</td><td>2%</td><td>Nil</td></tr>
          </tbody>
        </table>
      </div>
    </div>

    <!-- Due dates -->
    <div class="card">
      <div class="card-head">
        <div class="icon" style="background:#F0FDF4">📅</div>
        <div><h2 id="dueDateTitle">TDS Deposit Due Dates</h2><p>Rule 218 — IT Rules 2026</p></div>
      </div>
      <div class="card-body">
        <div id="tdsduedates" style="font-size:12px;line-height:2;color:var(--muted)">
          <p><strong style="color:var(--ink)">April – February:</strong> 7th of the following month</p>
          <p><strong style="color:var(--ink)">March deductions:</strong> 30th April</p>
          <p><strong style="color:var(--ink)">Sec 194IA/194IB/194M/194S:</strong> 30 days from end of deduction month</p>
          <p style="margin-top:8px;color:var(--red)"><strong>Late interest:</strong> 1.5% per month · Fractional month = full month</p>
        </div>
        <div id="tcsduedates" style="display:none;font-size:12px;line-height:1.9;color:var(--muted)">
          <div style="margin-bottom:10px;padding:8px 12px;background:#F0FDF4;border:1px solid #BBF7D0;border-radius:8px;font-size:11px;color:#065F46">
            <strong>IT Act 2025 Reference:</strong> Section 394 (TCS) · Rule 219 (IT Rules 2026) · Challan 281
          </div>
          <p><strong style="color:var(--ink)">📅 Deposit Due Date:</strong></p>
          <p>Collections <strong>April – February</strong> → <strong>7th of following month</strong></p>
          <p>Collections in <strong>March</strong> → <strong>7th April</strong> of next FY</p>
          <div style="margin:10px 0;border-top:1px solid var(--border)"></div>
          <p><strong style="color:var(--ink)">🗓️ Quarterly Return — Form 27EQ:</strong></p>
          <table style="width:100%;border-collapse:collapse;font-size:11px;margin:6px 0">
            <thead><tr style="background:#F9FAFB">
              <th style="padding:5px 8px;border:1px solid var(--border);text-align:left">Quarter</th>
              <th style="padding:5px 8px;border:1px solid var(--border);text-align:left">Period</th>
              <th style="padding:5px 8px;border:1px solid var(--border);text-align:left">Due Date</th>
            </tr></thead>
            <tbody>
              <tr><td style="padding:5px 8px;border:1px solid var(--border)">Q1</td><td style="padding:5px 8px;border:1px solid var(--border)">Apr – Jun</td><td style="padding:5px 8px;border:1px solid var(--border);font-weight:600;color:var(--ink)">15th July</td></tr>
              <tr><td style="padding:5px 8px;border:1px solid var(--border)">Q2</td><td style="padding:5px 8px;border:1px solid var(--border)">Jul – Sep</td><td style="padding:5px 8px;border:1px solid var(--border);font-weight:600;color:var(--ink)">15th October</td></tr>
              <tr><td style="padding:5px 8px;border:1px solid var(--border)">Q3</td><td style="padding:5px 8px;border:1px solid var(--border)">Oct – Dec</td><td style="padding:5px 8px;border:1px solid var(--border);font-weight:600;color:var(--ink)">15th January</td></tr>
              <tr><td style="padding:5px 8px;border:1px solid var(--border)">Q4</td><td style="padding:5px 8px;border:1px solid var(--border)">Jan – Mar</td><td style="padding:5px 8px;border:1px solid var(--border);font-weight:600;color:var(--ink)">15th May</td></tr>
            </tbody>
          </table>
          <div style="margin-top:10px;padding:8px 12px;background:#FEF2F2;border:1px solid #FECACA;border-radius:8px;font-size:11px">
            <strong style="color:#991B1B">⚠ Default Consequences:</strong><br>
            <span style="color:#991B1B">Non-collection/deposit: Interest <strong>1%/month</strong> u/s 394(6) · Fractional month = full month<br>
            Late 27EQ filing: <strong>₹200/day</strong> u/s 267 (max = TCS amount)</span>
          </div>
        </div>
      </div>
    </div>
  </div>
</div>

<footer>
  <div class="ft-main">
    <div>
      <div class="ft-brand-name">CA<span>Toolkit</span></div>
      <p class="ft-brand-desc">CA Toolkit is a comprehensive utility platform built by a CA Article from Ludhiana, Punjab, providing automation tools for Indian Chartered Accountants. The platform saves hours of manual work every year — from Balance Sheet year-shift to GST reconciliation, tax calculations, and more.</p>
    </div>
    <div>
      <div class="ft-col-title">Know More</div>
      <ul class="ft-links">
        <li><a href="/">Home</a></li>
        <li><a href="/">BS Year Shift</a></li>
        <li><a href="/tool/tb-to-bs">TB → Balance Sheet</a></li>
        <li><a href="/tool/tax-calculator">Tax Calculator</a></li>
        <li><a href="/privacy">Privacy Policy</a></li>
      </ul>
    </div>
    <div>
      <div class="ft-col-title">Contact Us</div>
      <div class="ft-contact-name">CA Toolkit</div>
      <div class="ft-contact-addr">Built for Indian Chartered Accountants<br/>Created by CA Article · Ludhiana, Punjab</div>
    </div>
  </div>
  <div class="ft-bottom">
    <span class="ft-bottom-left">©2026 CA Toolkit · All Rights Reserved</span>
    <span class="ft-bottom-right">Built for Indian CAs · Ludhiana, Punjab</span>
  </div>
</footer>

<script>
// ── DATA ─────────────────────────────────────────────────────────────────────

const TDS_DATA = {
  "1001":{rate:0,    thresh:0,       label:"Salary",                       newSec:"Sec 392",                    oldSec:"Sec 192",     note:"Slab rate"},
  "1004":{rate:10,   thresh:50000,   label:"PF Accumulated Balance",        newSec:"Sec 392(7)",                 oldSec:"Sec 192A",    note:"No PAN: 20%"},
  "1005":{rate:2,    thresh:20000,   label:"Insurance Commission",          newSec:"Sec 393(1) Sl.1(i)",         oldSec:"Sec 194D",    note:"Ind: 2%, Others: 10%"},
  "1006":{rate:2,    thresh:20000,   label:"Commission/Brokerage",          newSec:"Sec 393(1) Sl.1(ii)",        oldSec:"Sec 194H",    note:""},
  "1008":{rate:2,    thresh:50000,   label:"Rent – Machinery/Plant",        newSec:"Sec 393(1) Sl.2(ii).D(a)",  oldSec:"Sec 194I(a)", note:"Monthly threshold"},
  "1009":{rate:10,   thresh:50000,   label:"Rent – Land/Building",          newSec:"Sec 393(1) Sl.2(ii).D(b)",  oldSec:"Sec 194I(b)", note:"Monthly threshold"},
  "1010":{rate:2,    thresh:50000,   label:"Rent by Ind/HUF",               newSec:"Sec 393(1) Sl.2(i)",         oldSec:"Sec 194IB",   note:"Per month. Reduced from 5% to 2%"},
  "1011":{rate:10,   thresh:0,       label:"JDA Consideration",             newSec:"Sec 393(1) Sl.3(ii)",        oldSec:"Sec 194IC",   note:""},
  "1012":{rate:10,   thresh:500000,  label:"Land Acquisition Comp.",        newSec:"Sec 393(1) Sl.3(iii)",       oldSec:"Sec 194LA",   note:"Threshold ₹5L"},
  "1013":{rate:10,   thresh:10000,   label:"Mutual Fund Units",             newSec:"Sec 393(1) Sl.4(i)",         oldSec:"Sec 194K",    note:""},
  "1014":{rate:10,   thresh:0,       label:"Business Trust – Interest",     newSec:"Sec 393(1) Sl.4(ii)",        oldSec:"Sec 194LBA",  note:""},
  "1017":{rate:10,   thresh:0,       label:"Investment Fund Income",        newSec:"Sec 393(1) Sl.4(iii)",       oldSec:"Sec 194LBB",  note:""},
  "1018":{rate:10,   thresh:0,       label:"Securitisation Trust",          newSec:"Sec 393(1) Sl.4(iv)",        oldSec:"Sec 194LBC",  note:""},
  "1019":{rate:10,   thresh:10000,   label:"Interest on Securities",        newSec:"Sec 393(1) Sl.5(i)",         oldSec:"Sec 193",     note:""},
  "1020":{rate:10,   thresh:100000,  label:"Interest – Senior Citizen",     newSec:"Sec 393(1) Sl.5(ii).D(a)",  oldSec:"Sec 194A",    note:"Threshold ₹1L"},
  "1021":{rate:10,   thresh:50000,   label:"Interest – Bank/Post Office",   newSec:"Sec 393(1) Sl.5(ii).D(b)",  oldSec:"Sec 194A",    note:"Threshold ₹50K"},
  "1022":{rate:10,   thresh:10000,   label:"Interest – Others",             newSec:"Sec 393(1) Sl.5(iii)",       oldSec:"Sec 194A",    note:"Threshold ₹10K"},
  "1023":{rate:1,    thresh:30000,   label:"Contractor – Ind/HUF",          newSec:"Sec 393(1) Sl.6(i).D(a)",   oldSec:"Sec 194C",    note:"Single ₹30K / Annual ₹1L"},
  "1024":{rate:2,    thresh:30000,   label:"Contractor – Others",           newSec:"Sec 393(1) Sl.6(i).D(b)",   oldSec:"Sec 194C",    note:"Single ₹30K / Annual ₹1L"},
  "1026":{rate:2,    thresh:50000,   label:"Technical Services/Royalty",    newSec:"Sec 393(1) Sl.6(iii).D(a)", oldSec:"Sec 194J(a)", note:""},
  "1027":{rate:10,   thresh:50000,   label:"Professional Fees",             newSec:"Sec 393(1) Sl.6(iii).D(b)", oldSec:"Sec 194J(b)", note:""},
  "1028":{rate:10,   thresh:0,       label:"Director Remuneration",         newSec:"Sec 393(1) Sl.6(iii).D(b)", oldSec:"Sec 194J(b)", note:"No threshold"},
  "1029":{rate:10,   thresh:10000,   label:"Dividends",                     newSec:"Sec 393(1) Sl.7",            oldSec:"Sec 194",     note:""},
  "1030":{rate:2,    thresh:100000,  label:"Life Insurance Proceeds",       newSec:"Sec 393(1) Sl.8(i)",         oldSec:"Sec 194DA",   note:"On taxable portion"},
  "1031":{rate:0.1,  thresh:5000000, label:"Purchase of Goods",             newSec:"Sec 393(1) Sl.8(ii)",        oldSec:"Sec 194Q",    note:"Annual > ₹50L"},
  "1033":{rate:10,   thresh:20000,   label:"Benefit/Perquisite",            newSec:"Sec 393(1) Sl.8(iv)",        oldSec:"Sec 194R",    note:""},
  "1035":{rate:0.1,  thresh:500000,  label:"E-Commerce Operator",           newSec:"Sec 393(1) Sl.8(vi)",        oldSec:"Sec 194O",    note:"Annual > ₹5L"},
  "1036":{rate:1,    thresh:5000000, label:"Purchase of Immovable Property",newSec:"Sec 393(1) Sl.3(i)",         oldSec:"Sec 194IA",   note:"Threshold ₹50L"},
  "1037":{rate:5,    thresh:5000000, label:"Contractor/Prof by Ind/HUF",    newSec:"Sec 393(1) Sl.6(iv)",        oldSec:"Sec 194M",    note:"Annual > ₹50L"},
  "1038":{rate:2,    thresh:2000000, label:"Cash Withdrawal",               newSec:"Sec 393(1) Sl.8(vii)",       oldSec:"Sec 194N",    note:"3% if no ITR filed"},
  "1039":{rate:1,    thresh:10000,   label:"VDA/Crypto",                    newSec:"Sec 393(1) Sl.8(viii)",      oldSec:"Sec 194S",    note:"₹50K for specified persons"},
  "1040":{rate:30,   thresh:10000,   label:"Lottery/Puzzle Winnings",       newSec:"Sec 393(1) Sl.8(ix)",        oldSec:"Sec 194B",    note:""},
  "1041":{rate:10,   thresh:20000,   label:"Partner Salary/Remuneration",   newSec:"Sec 393(1) Sl.6(v)",         oldSec:"Sec 194T",    note:"Threshold ₹20K pa"},
};

const TCS_DATA = {
  "2001":{rate:2,    thresh:0,        label:"Alcoholic Liquor for Human Consumption", newSec:"Sec 394(1)(i)",   oldSec:"Sec 206C(1)(a)",  note:"Increased from 1% to 2% w.e.f. 01.04.2026"},
  "2002":{rate:2,    thresh:0,        label:"Tendu Leaves",                           newSec:"Sec 394(1)(ii)",  oldSec:"Sec 206C(1)(b)",  note:"Reduced from 5% to 2% w.e.f. 01.04.2026"},
  "2003":{rate:2,    thresh:0,        label:"Timber – Forest Lease",                  newSec:"Sec 394(1)(iii)", oldSec:"Sec 206C(1)(c)",  note:"Reduced from 2.5% to 2% w.e.f. 01.04.2026"},
  "2004":{rate:2,    thresh:0,        label:"Timber – Other than Forest Lease",       newSec:"Sec 394(1)(iv)",  oldSec:"Sec 206C(1)(d)",  note:"Reduced from 2.5% to 2% w.e.f. 01.04.2026"},
  "2005":{rate:2,    thresh:0,        label:"Any Other Forest Produce",               newSec:"Sec 394(1)(v)",   oldSec:"Sec 206C(1)(e)",  note:"Reduced from 2.5% to 2% w.e.f. 01.04.2026"},
  "2006":{rate:2,    thresh:0,        label:"Scrap",                                  newSec:"Sec 394(1)(vi)",  oldSec:"Sec 206C(1)(f)",  note:"Increased from 1% to 2% w.e.f. 01.04.2026"},
  "2007":{rate:2,    thresh:0,        label:"Minerals (Coal/Lignite/Iron Ore)",       newSec:"Sec 394(1)(vii)", oldSec:"Sec 206C(1)(g)",  note:"Increased from 1% to 2% w.e.f. 01.04.2026"},
  "2009":{rate:1,    thresh:1000000,  label:"Motor Vehicle > ₹10L",                   newSec:"Sec 394(1F)",     oldSec:"Sec 206C(1F)",    note:"On sale consideration"},
  "2010":{rate:2,    thresh:1000000,  label:"Foreign Remittance (LRS) – Education/Medical > ₹10L",newSec:"Sec 394(1G)(i)",oldSec:"Sec 206C(1G)",   note:"Reduced from 5% to 2%. Nil if loan from bank. Threshold now ₹10L"},
  "2011":{rate:20,   thresh:1000000,  label:"Foreign Remittance (LRS) – Other > ₹10L",newSec:"Sec 394(1G)(ii)", oldSec:"Sec 206C(1G)",    note:"20% above ₹10L. Threshold changed from ₹7L to ₹10L"},
  "2012":{rate:2,    thresh:0,        label:"Overseas Tour Package",                  newSec:"Sec 394(1G)(iii)",oldSec:"Sec 206C(1G)",    note:"Flat 2% (was 5%/20%). Threshold removed w.e.f. 01.04.2026"},
  "2013":{rate:0.1,  thresh:5000000,  label:"Sale of Goods > ₹50L",                   newSec:"Sec 394(1H)",     oldSec:"Sec 206C(1H)",    note:"Annual turnover > ₹10Cr"},
  "2014":{rate:2,    thresh:0,        label:"Parking Lot / Toll Plaza / Mining",       newSec:"Sec 394(1)(viii)",oldSec:"Sec 206C(1)(h)",  note:""},
};

const TDS_SPECIAL_30 = ["1036","1010","1037","1039"];

let currentMode = "tds";

// ── Build dropdowns ───────────────────────────────────────────────────────────

function buildTDSOptions(){
  return `<option value="">— Select Payment Type —</option>
    <optgroup label="── Salary ──">
      <option value="1001">Salary (Sec 392) — Slab rate</option>
      <option value="1004">PF Accumulated Balance — 10%</option>
    </optgroup>
    <optgroup label="── Commission &amp; Brokerage ──">
      <option value="1005">Insurance Commission (Old: 194D) — 2%</option>
      <option value="1006">Commission / Brokerage (Old: 194H) — 2%</option>
    </optgroup>
    <optgroup label="── Rent ──">
      <option value="1008">Rent – Machinery/Plant (Old: 194I(a)) — 2%</option>
      <option value="1009">Rent – Land/Building (Old: 194I(b)) — 10%</option>
      <option value="1010">Rent by Individual/HUF (Old: 194IB) — 2%</option>
    </optgroup>
    <optgroup label="── Property ──">
      <option value="1011">JDA Consideration (Old: 194IC) — 10%</option>
      <option value="1012">Compensation – Land Acquisition (Old: 194LA) — 10%</option>
      <option value="1036">Purchase of Immovable Property (Old: 194IA) — 1%</option>
    </optgroup>
    <optgroup label="── Interest ──">
      <option value="1019">Interest on Securities (Old: 193) — 10%</option>
      <option value="1020">Interest – Senior Citizen (Old: 194A) — 10%</option>
      <option value="1021">Interest – Bank/Post Office (Old: 194A) — 10%</option>
      <option value="1022">Interest – Others (Old: 194A) — 10%</option>
    </optgroup>
    <optgroup label="── Investment Income ──">
      <option value="1013">Mutual Fund Units (Old: 194K) — 10%</option>
      <option value="1029">Dividends (Old: 194) — 10%</option>
      <option value="1014">Business Trust – Interest (Old: 194LBA) — 10%</option>
    </optgroup>
    <optgroup label="── Contractor &amp; Professional ──">
      <option value="1023">Contractor – Individual/HUF (Old: 194C) — 1%</option>
      <option value="1024">Contractor – Others/Company (Old: 194C) — 2%</option>
      <option value="1026">Technical Services/Royalty (Old: 194J(a)) — 2%</option>
      <option value="1027">Professional Fees (Old: 194J(b)) — 10%</option>
      <option value="1028">Director Remuneration (Old: 194J(b)) — 10%</option>
      <option value="1037">Contractor/Prof by Ind/HUF (Old: 194M) — 5%</option>
      <option value="1041">Partner Salary/Remuneration (Old: 194T) — 10%</option>
    </optgroup>
    <optgroup label="── Other Payments ──">
      <option value="1030">Life Insurance Proceeds (Old: 194DA) — 2%</option>
      <option value="1031">Purchase of Goods (Old: 194Q) — 0.1%</option>
      <option value="1033">Benefit/Perquisite (Old: 194R) — 10%</option>
      <option value="1035">E-Commerce Operator (Old: 194O) — 0.1%</option>
      <option value="1038">Cash Withdrawal (Old: 194N) — 2%</option>
      <option value="1039">VDA / Crypto (Old: 194S) — 1%</option>
      <option value="1040">Lottery/Puzzle Winnings (Old: 194B) — 30%</option>
    </optgroup>`;
}

function buildTCSOptions(){
  return `<option value="">— Select Nature of Goods/Transaction —</option>
    <optgroup label="── Goods (All rationalised to 2%) ──">
      <option value="2001">Alcoholic Liquor for Human Consumption — 2%</option>
      <option value="2002">Tendu Leaves — 2%</option>
      <option value="2003">Timber – Forest Lease — 2%</option>
      <option value="2004">Timber – Other than Forest Lease — 2%</option>
      <option value="2005">Any Other Forest Produce — 2%</option>
      <option value="2006">Scrap — 2%</option>
      <option value="2007">Minerals (Coal/Lignite/Iron Ore) — 2%</option>
      <option value="2014">Parking Lot / Toll Plaza / Mining &amp; Quarrying — 2%</option>
    </optgroup>
    <optgroup label="── High Value Transactions ──">
      <option value="2009">Motor Vehicle Sale &gt; ₹10 Lakh — 1%</option>
      <option value="2013">Sale of Goods &gt; ₹50L (Annual) — 0.1%</option>
    </optgroup>
    <optgroup label="── Foreign Remittance (LRS) ──">
      <option value="2010">Foreign Remittance – Education/Medical &gt; ₹10L — 2%</option>
      <option value="2011">Foreign Remittance – Other Purposes &gt; ₹10L — 20%</option>
      <option value="2012">Overseas Tour Package — 2% (flat)</option>
    </optgroup>`;
}

// ── Toggle mode ───────────────────────────────────────────────────────────────

function switchMode(mode){
  currentMode = mode;
  const sel = document.getElementById("mainSection");
  sel.innerHTML = mode==="tds" ? buildTDSOptions() : buildTCSOptions();
  document.getElementById("sectionHint").textContent = "Select a payment type to see rate and threshold";

  const isTDS = mode==="tds";
  document.getElementById("btnTDS").className = "toggle-btn"+(isTDS?" active":"");
  document.getElementById("btnTCS").className = "toggle-btn"+(!isTDS?" active":"");
  document.getElementById("formIcon").textContent     = isTDS?"📑":"🧾";
  document.getElementById("formTitle").textContent    = isTDS?"TDS Calculator":"TCS Calculator";
  document.getElementById("formSub").textContent      = isTDS?"IT Act 2025 · Tax Year 2026-27":"IT Act 2025 · Tax Year 2026-27";
  document.getElementById("sectionLabel").textContent = isTDS?"Nature of Payment (TDS)":"Nature of Goods / Transaction (TCS)";
  document.getElementById("amtLabel").textContent     = isTDS?"Payment Amount (₹)":"Sale / Collection Amount (₹)";
  document.getElementById("amtHint").textContent      = isTDS?"Gross payment amount before TDS deduction":"Gross sale/receipt amount before TCS collection";
  document.getElementById("d1label").textContent      = isTDS?"Date of Deduction":"Date of Collection";
  document.getElementById("d1hint").textContent       = isTDS?"When TDS was deducted":"When TCS was collected";
  document.getElementById("calcBtn").textContent      = isTDS?"Calculate TDS & Interest →":"Calculate TCS & Interest →";
  document.getElementById("r-main-lbl").textContent   = isTDS?"TDS Amount":"TCS Amount";
  document.getElementById("b-main-lbl").textContent   = isTDS?"TDS Amount":"TCS Amount";
  document.getElementById("d-rate-lbl").textContent   = isTDS?"TDS Rate":"TCS Rate";
  document.getElementById("d-tax-lbl").textContent    = isTDS?"TDS Amount":"TCS Amount";
  document.getElementById("d-date1-lbl").textContent  = isTDS?"Date of Deduction":"Date of Collection";
  document.getElementById("d-amt-lbl").textContent    = isTDS?"Payment Amount":"Sale Amount";
  document.getElementById("b-amt-lbl").textContent    = isTDS?"Payment Amount":"Sale Amount";
  document.getElementById("b-rate-lbl").textContent   = isTDS?"TDS Rate":"TCS Rate";
  document.getElementById("b-tax-lbl").textContent    = isTDS?"TDS Amount":"TCS Amount";
  document.getElementById("b-net-lbl").textContent    = isTDS?"Net Payment to Payee":"Amount Receivable from Buyer";
  document.getElementById("d-total-lbl").textContent  = isTDS?"Total Payable (TDS + Interest)":"Total Payable (TCS + Interest)";
  document.getElementById("tdsRateCard").style.display = isTDS?"block":"none";
  document.getElementById("tcsRateCard").style.display = !isTDS?"block":"none";
  document.getElementById("tdsduedates").style.display = isTDS?"block":"none";
  document.getElementById("tcsduedates").style.display = !isTDS?"block":"none";
  document.getElementById("dueDateTitle").textContent = isTDS?"TDS Deposit Due Dates":"TCS Deposit Due Dates";

  // Update interest section label for TCS
  const intLbl = document.getElementById("r-int");
  if(intLbl){
    const lbl = intLbl.closest(".rbox")?.querySelector(".lbl");
    if(lbl) lbl.textContent = isTDS?"Interest u/s 201(1A)":"Interest u/s 206C(7)";
  }

  document.getElementById("resultSection").style.display = "none";
}

// ── Hint update ───────────────────────────────────────────────────────────────

function updateHint(){
  const code = document.getElementById("mainSection").value;
  const el   = document.getElementById("sectionHint");
  if(!code){ el.textContent="Select a payment type to see rate and threshold"; return; }
  const data = currentMode==="tds" ? TDS_DATA : TCS_DATA;
  const d    = data[code];
  if(!d) return;
  el.textContent = (d.rate===0?"Rate: Slab rate":"Rate: "+d.rate+"%")
    + (d.thresh?" · Threshold: ₹"+Math.round(d.thresh).toLocaleString("en-IN"):" · No threshold")
    + (d.note?" · "+d.note:"");
}

// ── Due date calc ─────────────────────────────────────────────────────────────

function getDueDate(deductDate, code, mode){
  const d     = new Date(deductDate);
  const month = d.getMonth();
  const year  = d.getFullYear();
  if(mode==="tds" && TDS_SPECIAL_30.includes(code)){
    const endOfMonth = new Date(year, month+1, 0);
    return new Date(endOfMonth.getTime() + 30*24*60*60*1000);
  }
  if(month===2) return new Date(year, 3, 30); // March → 30 April
  return new Date(year, month+1, 7);          // Others → 7th next month
}

function calcMonthsLate(dueDate, depositDate){
  if(new Date(depositDate) <= new Date(dueDate)) return 0;
  let months=0, cur=new Date(dueDate);
  while(cur < new Date(depositDate)){ cur.setMonth(cur.getMonth()+1); months++; }
  return months;
}

// ── Main calculate ─────────────────────────────────────────────────────────────

function calculate(){
  const code   = document.getElementById("mainSection").value;
  const amt    = parseFloat(document.getElementById("paymentAmt").value);
  const dDate  = document.getElementById("deductionDate").value;
  const aDate  = document.getElementById("depositDate").value;

  if(!code){ alert("Please select a payment type."); return; }
  if(!amt||amt<=0){ alert("Please enter a valid amount."); return; }

  const data = currentMode==="tds" ? TDS_DATA : TCS_DATA;
  const d    = data[code];
  if(!d){ alert("Data not found."); return; }

  const isTCS      = currentMode==="tcs";
  const intRate    = isTCS ? 0.01 : 0.015; // TCS: 1%/mo, TDS: 1.5%/mo
  const intSecLbl  = isTCS ? "u/s 206C(7)" : "u/s 201(1A)";

  const belowThresh = d.thresh && amt < d.thresh;
  const tax         = belowThresh ? 0 : (d.rate===0 ? 0 : Math.round(amt * d.rate / 100));
  const net         = isTCS ? amt + tax : amt - tax;

  const fmt = n => "₹"+Math.round(n).toLocaleString("en-IN");
  const fmtDate = dt => new Date(dt).toLocaleDateString("en-IN",{day:"2-digit",month:"short",year:"numeric"});

  document.getElementById("resultSection").style.display = "block";

  if(!dDate||!aDate){
    // Basic result only
    document.getElementById("ontimeBox").style.display = "none";
    document.getElementById("lateBoxes").style.display = "none";
    document.getElementById("basicBox").style.display  = "block";
    document.getElementById("b-main").textContent = belowThresh?"No "+(isTCS?"TCS":"TDS"):d.rate===0?"Slab Rate":fmt(tax);
    document.getElementById("b-sub").textContent  = belowThresh?"Below threshold of "+fmt(d.thresh):d.rate===0?"Compute at slab rate":d.rate+"% on "+fmt(amt);
    document.getElementById("b-payment").textContent = fmt(amt);
    document.getElementById("b-newsec").textContent  = d.newSec;
    document.getElementById("b-oldsec").textContent  = d.oldSec+" (ref only)";
    document.getElementById("b-code").textContent    = code;
    document.getElementById("b-rate").textContent    = d.rate===0?"Slab rate":d.rate+"%";
    document.getElementById("b-tds2").textContent    = belowThresh?"Nil (below threshold)":d.rate===0?"As per slab":fmt(tax);
    document.getElementById("b-net").textContent     = isTCS?fmt(net)+" (incl. TCS)":fmt(net);
    return;
  }

  const dueDate    = getDueDate(dDate, code, currentMode);
  const monthsLate = calcMonthsLate(dueDate, aDate);
  const interest   = Math.round(tax * intRate * monthsLate);
  const total      = tax + interest;
  const isOnTime   = monthsLate===0;

  document.getElementById("basicBox").style.display = "none";

  if(isOnTime||belowThresh||d.rate===0){
    document.getElementById("ontimeBox").style.display = "block";
    document.getElementById("lateBoxes").style.display = "none";
    if(belowThresh) document.getElementById("ontimeBox").textContent = "No "+(isTCS?"TCS":"TDS")+" — Below threshold of "+fmt(d.thresh);
    else if(d.rate===0) document.getElementById("ontimeBox").textContent = "Salary TDS — compute at applicable slab rate";
    else document.getElementById("ontimeBox").textContent = "✓ Deposit is on time — No interest. "+(isTCS?"TCS":"TDS")+": "+fmt(tax);
    return;
  }

  document.getElementById("ontimeBox").style.display = "none";
  document.getElementById("lateBoxes").style.display = "block";

  document.getElementById("r-main").textContent     = fmt(tax);
  document.getElementById("r-main-sub").textContent = d.rate+"% on "+fmt(amt);
  document.getElementById("r-int").textContent      = fmt(interest);
  document.getElementById("r-int-sub").textContent  = (intRate*100)+"% × "+monthsLate+" month"+(monthsLate>1?"s":"");
  document.getElementById("r-total").textContent    = fmt(total);
  document.getElementById("r-total-sub").textContent= (isTCS?"TCS":"TDS")+" "+fmt(tax)+" + Interest "+fmt(interest);
  document.getElementById("d-payment").textContent  = fmt(amt);
  document.getElementById("d-newsec").textContent   = d.newSec;
  document.getElementById("d-oldsec").textContent   = d.oldSec+" (ref only)";
  document.getElementById("d-code").textContent     = code;
  document.getElementById("d-rate").textContent     = d.rate+"%";
  document.getElementById("d-tds").textContent      = fmt(tax);
  document.getElementById("d-ddate").textContent    = fmtDate(dDate);
  document.getElementById("d-due").textContent      = dueDate.toLocaleDateString("en-IN",{day:"2-digit",month:"short",year:"numeric"});
  document.getElementById("d-adate").textContent    = fmtDate(aDate);
  document.getElementById("d-months").textContent   = monthsLate+" month"+(monthsLate>1?"s":"")+" (fractional = full month)";
  document.getElementById("d-intamt").textContent   = fmt(interest)+" "+intSecLbl;
  document.getElementById("d-total").textContent    = fmt(total);
}

// Init
document.getElementById("mainSection").innerHTML = buildTDSOptions();
</script>
<a href="https://wa.me/918427651580" target="_blank" class="wa-float" title="WhatsApp Support"><svg viewBox="0 0 24 24"><path d="M17.472 14.382c-.297-.149-1.758-.867-2.03-.967-.273-.099-.471-.148-.67.15-.197.297-.767.966-.94 1.164-.173.199-.347.223-.644.075-.297-.15-1.255-.463-2.39-1.475-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.298-.347.446-.52.149-.174.198-.298.298-.497.099-.198.05-.371-.025-.52-.075-.149-.669-1.612-.916-2.207-.242-.579-.487-.5-.669-.51-.173-.008-.371-.01-.57-.01-.198 0-.52.074-.792.372-.272.297-1.04 1.016-1.04 2.479 0 1.462 1.065 2.875 1.213 3.074.149.198 2.096 3.2 5.077 4.487.709.306 1.262.489 1.694.625.712.227 1.36.195 1.871.118.571-.085 1.758-.719 2.006-1.413.248-.694.248-1.289.173-1.413-.074-.124-.272-.198-.57-.347m-5.421 7.403h-.004a9.87 9.87 0 01-5.031-1.378l-.361-.214-3.741.982.998-3.648-.235-.374a9.86 9.86 0 01-1.51-5.26c.001-5.45 4.436-9.884 9.888-9.884 2.64 0 5.122 1.03 6.988 2.898a9.825 9.825 0 012.893 6.994c-.003 5.45-4.437 9.884-9.885 9.884m8.413-18.297A11.815 11.815 0 0012.05 0C5.495 0 .16 5.335.157 11.892c0 2.096.547 4.142 1.588 5.945L.057 24l6.305-1.654a11.882 11.882 0 005.683 1.448h.005c6.554 0 11.89-5.335 11.893-11.893a11.821 11.821 0 00-3.48-8.413z"/></svg></a>

<button class="help-btn" onclick="openHelp()" title="How to use this tool">?</button>
<div class="help-overlay" id="helpOverlay">
  <div class="help-modal">
    <div class="help-modal-head"><h3>How to Use — TDS/TCS Calculator</h3><button class="help-close" onclick="closeHelp()">&#10005;</button></div>
    <div class="help-modal-body"><div class="help-step"><div class="help-step-num">1</div><div class="help-step-body"><h4>Select Section</h4><p>Choose the TDS/TCS section (e.g. 194C, 194J, 206C etc.).</p></div></div><div class="help-step"><div class="help-step-num">2</div><div class="help-step-body"><h4>Enter Amount</h4><p>Enter the payment/receipt amount.</p></div></div><div class="help-step"><div class="help-step-num">3</div><div class="help-step-body"><h4>Check Threshold</h4><p>The tool shows whether TDS is applicable based on annual threshold.</p></div></div><div class="help-step"><div class="help-step-num">4</div><div class="help-step-body"><h4>View Rate</h4><p>See applicable TDS/TCS rate, deductible amount, and net payable.</p></div></div><div class="help-tip">💡 Updated for IT Act 2025 new payment codes (Sections 393/394).</div></div>
  </div>
</div>
<script>function openHelp(){document.getElementById('helpOverlay').classList.add('open')}function closeHelp(){document.getElementById('helpOverlay').classList.remove('open')}document.getElementById('helpOverlay').addEventListener('click',function(e){if(e.target===this)closeHelp()})</script>
</body></html>"""


DEP_CALC_T = r"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Depreciation Calculator – CA Toolkit</title>

<style>
""" + BASE_CSS + """
.hero{text-align:center;padding:40px 24px 28px;max-width:700px;margin:0 auto}
.hero-badge{display:inline-flex;align-items:center;gap:6px;background:#ECFDF5;
            color:#065F46;border:1px solid #A7F3D0;border-radius:99px;
            padding:5px 14px;font-size:12px;font-weight:600;margin-bottom:14px}
h1{font-size:clamp(22px,4vw,34px);font-weight:800;line-height:1.15;letter-spacing:-.5px;margin-bottom:10px}
h1 em{font-style:normal;color:var(--accent)}
.hero p{font-size:14px;color:var(--muted);line-height:1.7;max-width:500px;margin:0 auto}
.main{max-width:1000px;margin:0 auto;padding:28px 24px 48px}
.card{background:var(--white);border-radius:var(--radius);border:1px solid var(--border);
      box-shadow:var(--shadow);overflow:hidden;margin-bottom:20px}
.card-head{padding:14px 20px;border-bottom:1px solid var(--border);display:flex;align-items:center;gap:10px}
.card-head .icon{width:30px;height:30px;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:15px}
.card-head h2{font-size:14px;font-weight:700}
.card-head p{font-size:12px;color:var(--muted);margin-top:1px}
.card-body{padding:20px}
.form-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:14px}
@media(max-width:700px){.form-grid{grid-template-columns:1fr}}
.field{margin-bottom:0}
label{display:block;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.04em;color:var(--muted);margin-bottom:5px}
.hint{font-size:11px;color:var(--muted);margin-top:4px}
input[type=number],input[type=text],select{width:100%;border:1.5px solid var(--border);border-radius:8px;
  padding:9px 12px;font-family:inherit;font-size:13px;color:var(--ink);background:var(--white);
  transition:border-color .2s;outline:none}
input:focus,select:focus{border-color:var(--brand)}
.row2{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-top:14px}
.btn{background:var(--brand);color:#fff;border:none;border-radius:10px;
     padding:11px 24px;font-family:inherit;font-size:14px;font-weight:700;
     cursor:pointer;transition:background .2s;margin-top:16px}
.btn:hover{background:var(--brand-d)}
.result-section{display:none}
.summary-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:20px}
@media(max-width:700px){.summary-grid{grid-template-columns:1fr 1fr}}
.summary-box{background:var(--white);border:1px solid var(--border);border-radius:10px;padding:16px}
.summary-box .val{font-size:20px;font-weight:800;color:var(--brand);margin-bottom:4px}
.summary-box .lbl{font-size:11px;color:var(--muted)}
table{width:100%;border-collapse:collapse;font-size:12px}
th{text-align:left;font-size:10px;letter-spacing:.06em;text-transform:uppercase;
   color:var(--muted);border-bottom:1.5px solid var(--border);padding:7px 10px}
td{padding:9px 10px;border-bottom:1px solid var(--border)}
tr:last-child td{border:none;font-weight:700;background:#F9FAFB}
td:not(:first-child){text-align:right}
.tag-it{background:#EFF6FF;color:var(--brand);font-size:10px;font-weight:700;
        padding:2px 7px;border-radius:99px}
.tag-ca{background:#F5F3FF;color:#5B21B6;font-size:10px;font-weight:700;
        padding:2px 7px;border-radius:99px}
footer{background:#0f1b2d;color:#9CA3AF;font-size:12px;padding:0}
.ft-main{display:grid;grid-template-columns:2fr 1fr 1.4fr;gap:40px;padding:40px 48px;max-width:1200px;margin:0 auto}
.ft-brand-name{color:#fff;font-size:18px;font-weight:800;margin-bottom:12px}
.ft-brand-desc{font-size:12.5px;line-height:1.75;color:#9CA3AF;max-width:340px;text-align:justify}
.ft-col-title{color:#fff;font-size:14px;font-weight:700;margin-bottom:14px}
.ft-links{list-style:none;padding:0;margin:0}
.ft-links li{margin-bottom:8px}
.ft-links a{color:#9CA3AF;text-decoration:none;font-size:13px;transition:color .2s}
.ft-links a:hover{color:#fff}
.ft-contact-name{color:#fff;font-weight:700;font-size:13px;margin-bottom:6px}
.ft-contact-addr{color:#9CA3AF;font-size:12px;line-height:1.7;margin-bottom:10px}
.ft-contact-line{color:#9CA3AF;font-size:12px;margin-bottom:4px}
.ft-socials{display:flex;gap:14px;margin-top:12px}
.ft-socials a{color:#9CA3AF;transition:color .2s}
.ft-socials a:hover{color:#fff}
.ft-socials svg{width:20px;height:20px;fill:currentColor}
.ft-bottom{background:#0a1422;border-top:1px solid #1e2d42;padding:12px 48px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px}
.ft-bottom-left{font-size:11px;color:#6B7280}
.ft-bottom-right{font-size:11px;color:#6B7280}
@media(max-width:768px){.ft-main{grid-template-columns:1fr;padding:28px 20px;gap:24px}.ft-bottom{padding:12px 20px;flex-direction:column;text-align:center}}
</style></head><body>

<nav>
  <a href="/" class="logo">CA<span class="logo-dot"></span><span>Toolkit</span></a>
  <div class="nav-right">
    {% if username %}
    <span class="nav-user">👤 <strong>{{ username }}</strong></span>
    {% if is_admin %}<a href="/admin" class="nav-btn">Admin</a>{% endif %}
    <a href="/logout" class="nav-link">Sign out</a>
    {% else %}
    <a href="/login" class="nav-btn">Sign In</a>
    {% endif %}
    <a href="/" class="nav-btn dash">⬅ Dashboard</a>
  </div>
</nav>

<section class="hero">
  <div class="hero-badge">🆓 Free Tool · No Login Required</div>
  <h1>Depreciation Calculator</h1>
  <p>Calculate depreciation under <strong>Companies Act 2013</strong> (WDV/SLM) and <strong>Income Tax Act</strong>. Get full year-wise schedule instantly.</p>
</section>

<div class="main">
  <div class="card">
    <div class="card-head">
      <div class="icon" style="background:#F5F3FF">🏭</div>
      <div><h2>Asset Details</h2><p>Enter asset information to generate depreciation schedule</p></div>
    </div>
    <div class="card-body">
      <div class="form-grid">
        <div class="field">
          <label>Asset Name</label>
          <input type="text" id="assetName" placeholder="e.g. Machinery, Vehicle"/>
        </div>
        <div class="field">
          <label>Cost of Asset (₹)</label>
          <input type="number" id="assetCost" placeholder="e.g. 500000" min="0"/>
        </div>
        <div class="field">
          <label>Date of Purchase</label>
          <input type="date" id="purchaseDate"/>
        </div>
        <div class="field">
          <label>Asset Block (IT Act)</label>
          <select id="itBlock">
            <option value="15">15% — Furniture, Fittings</option>
            <option value="15b">15% — Ships</option>
            <option value="30">30% — Motor Cars (not used for hire)</option>
            <option value="40">40% — Motor Taxis, Buses (hire)</option>
            <option value="40b">40% — Machinery (general)</option>
            <option value="60">60% — Computers &amp; Software</option>
            <option value="80">80% — Energy saving devices</option>
            <option value="100">100% — Books, Scientific research</option>
            <option value="10">10% — Buildings (residential)</option>
            <option value="5">5% — Buildings (other)</option>
          </select>
        </div>
        <div class="field">
          <label>Asset Class (Companies Act)</label>
          <select id="caClass">
            <option value="15_wdv">Buildings — Factory (5% SLM / 15 yr WDV)</option>
            <option value="10_wdv">Buildings — Other (10% SLM / 10 yr WDV)</option>
            <option value="15_plant">Plant &amp; Machinery General (15% SLM)</option>
            <option value="30_plant">Plant &amp; Machinery (30% SLM — certain)</option>
            <option value="20_furn">Furniture &amp; Fixtures (10% SLM)</option>
            <option value="25_comp">Computers &amp; Peripherals (40% SLM)</option>
            <option value="20_veh">Vehicles — Motor Car (20% SLM)</option>
            <option value="30_veh">Vehicles — Motor Cycle (30% SLM)</option>
            <option value="10_off">Office Equipment (20% SLM)</option>
          </select>
        </div>
        <div class="field">
          <label>Method (Companies Act)</label>
          <select id="caMethod">
            <option value="slm">SLM — Straight Line Method</option>
            <option value="wdv">WDV — Written Down Value</option>
          </select>
        </div>
      </div>
      <div class="row2">
        <div class="field">
          <label>Number of Years to Project</label>
          <input type="number" id="numYears" value="5" min="1" max="20"/>
        </div>
        <div class="field">
          <label>Salvage / Residual Value (₹)</label>
          <input type="number" id="salvageVal" value="0" min="0"/>
          <p class="hint">Under Companies Act, minimum 5% of cost</p>
        </div>
      </div>
      <button class="btn" onclick="calcDep()">Generate Depreciation Schedule →</button>
    </div>
  </div>

  <div class="result-section" id="resultSection">
    <div class="summary-grid" id="summaryGrid"></div>

    <div class="card">
      <div class="card-head">
        <div class="icon" style="background:#EFF6FF">📊</div>
        <div><h2>Income Tax Act Schedule <span class="tag-it">IT Act</span></h2>
             <p>WDV method — Block of assets basis</p></div>
      </div>
      <div class="card-body" style="padding:0;overflow-x:auto">
        <table id="itTable">
          <thead><tr><th>FY</th><th>Opening WDV</th><th>Additions</th><th>Depreciation</th><th>Closing WDV</th></tr></thead>
          <tbody id="itBody"></tbody>
        </table>
      </div>
    </div>

    <div class="card" style="margin-top:20px">
      <div class="card-head">
        <div class="icon" style="background:#F5F3FF">📋</div>
        <div><h2>Companies Act 2013 Schedule <span class="tag-ca">Companies Act</span></h2>
             <p id="caMethodLabel">SLM method</p></div>
      </div>
      <div class="card-body" style="padding:0;overflow-x:auto">
        <table id="caTable">
          <thead><tr><th>FY</th><th>Opening WDV</th><th>Depreciation</th><th>Closing WDV</th><th>Acc. Dep.</th></tr></thead>
          <tbody id="caBody"></tbody>
        </table>
      </div>
    </div>
  </div>
</div>

<footer>
  <div class="ft-main">
    <div>
      <div class="ft-brand-name">CA<span>Toolkit</span></div>
      <p class="ft-brand-desc">CA Toolkit is a comprehensive utility platform built by a CA Article from Ludhiana, Punjab, providing automation tools for Indian Chartered Accountants. The platform saves hours of manual work every year — from Balance Sheet year-shift to GST reconciliation, tax calculations, and more.</p>
    </div>
    <div>
      <div class="ft-col-title">Know More</div>
      <ul class="ft-links">
        <li><a href="/">Home</a></li>
        <li><a href="/">BS Year Shift</a></li>
        <li><a href="/tool/tb-to-bs">TB → Balance Sheet</a></li>
        <li><a href="/tool/tax-calculator">Tax Calculator</a></li>
        <li><a href="/privacy">Privacy Policy</a></li>
      </ul>
    </div>
    <div>
      <div class="ft-col-title">Contact Us</div>
      <div class="ft-contact-name">CA Toolkit</div>
      <div class="ft-contact-addr">Built for Indian Chartered Accountants<br/>Created by CA Article · Ludhiana, Punjab</div>
    </div>
  </div>
  <div class="ft-bottom">
    <span class="ft-bottom-left">©2026 CA Toolkit · All Rights Reserved</span>
    <span class="ft-bottom-right">Built for Indian CAs · Ludhiana, Punjab</span>
  </div>
</footer>

<script>
const IT_RATES = {"15":15,"15b":15,"30":30,"40":40,"40b":40,"60":60,"80":80,"100":100,"10":10,"5":5};
const CA_RATES = {
  "15_wdv":  {slm:6.67, wdv:15,  life:15},
  "10_wdv":  {slm:10,   wdv:10,  life:10},
  "15_plant":{slm:6.67, wdv:15,  life:15},
  "30_plant":{slm:10,   wdv:30,  life:10},
  "20_furn": {slm:10,   wdv:10,  life:10},
  "25_comp": {slm:40,   wdv:40,  life:3},
  "20_veh":  {slm:20,   wdv:20,  life:5},
  "30_veh":  {slm:30,   wdv:30,  life:4},
  "10_off":  {slm:20,   wdv:20,  life:5},
};

function fmt(n){ return "₹"+Math.round(n).toLocaleString("en-IN"); }

function calcDep(){
  const cost = parseFloat(document.getElementById("assetCost").value);
  const name = document.getElementById("assetName").value || "Asset";
  const pd   = document.getElementById("purchaseDate").value;
  const itBl = document.getElementById("itBlock").value;
  const caCl = document.getElementById("caClass").value;
  const meth = document.getElementById("caMethod").value;
  const yrs  = Math.min(parseInt(document.getElementById("numYears").value)||5, 20);
  const salv = Math.max(parseFloat(document.getElementById("salvageVal").value)||0, cost*0.05);

  if(!cost||cost<=0){alert("Enter a valid asset cost.");return;}
  if(!pd){alert("Enter purchase date.");return;}

  const purchaseYear = parseInt(pd.split("-")[0]);
  const purchaseMon  = parseInt(pd.split("-")[1]);
  // IT Act: if purchased after 3 Oct (i.e. used < 180 days), half rate in first year
  const halfRate = purchaseMon >= 10 || (purchaseMon === 9 && parseInt(pd.split("-")[2]) > 3);

  const itRate = IT_RATES[itBl] / 100;
  const caInfo = CA_RATES[caCl];
  const caRate = meth === "slm" ? caInfo.slm/100 : caInfo.wdv/100;

  // IT Schedule (WDV)
  let itWDV = cost, itRows = "";
  for(let i=0;i<yrs;i++){
    const fy = `FY ${purchaseYear + i}-${String(purchaseYear+i+1).slice(-2)}`;
    const additions = i===0 ? cost : 0;
    const rate = (i===0 && halfRate) ? itRate/2 : itRate;
    const dep = Math.round(itWDV * rate);
    const closing = itWDV - dep;
    itRows += `<tr><td>${fy}</td><td>${fmt(itWDV)}</td><td>${i===0?fmt(additions):"—"}</td><td>${fmt(dep)}</td><td>${fmt(closing)}</td></tr>`;
    itWDV = closing;
    if(itWDV <= 0) break;
  }
  document.getElementById("itBody").innerHTML = itRows;

  // CA Schedule
  let caWDV = cost, caAcc = 0, caRows = "";
  document.getElementById("caMethodLabel").textContent = meth.toUpperCase() + " method";
  for(let i=0;i<yrs;i++){
    const fy = `FY ${purchaseYear + i}-${String(purchaseYear+i+1).slice(-2)}`;
    let dep;
    if(meth === "slm"){
      dep = Math.round((cost - salv) * caRate);
      if(caWDV - dep < salv) dep = Math.max(0, caWDV - salv);
    } else {
      dep = Math.round(caWDV * caRate);
      if(caWDV - dep < salv) dep = Math.max(0, caWDV - salv);
    }
    caAcc += dep;
    const closing = caWDV - dep;
    caRows += `<tr><td>${fy}</td><td>${fmt(caWDV)}</td><td>${fmt(dep)}</td><td>${fmt(closing)}</td><td>${fmt(caAcc)}</td></tr>`;
    caWDV = closing;
    if(caWDV <= salv) break;
  }
  document.getElementById("caBody").innerHTML = caRows;

  // Summary
  const itDep1 = cost * ((halfRate?itRate/2:itRate));
  const caDep1 = meth==="slm" ? (cost-salv)*caRate : cost*caRate;
  document.getElementById("summaryGrid").innerHTML =
    `<div class="summary-box"><div class="val">${fmt(cost)}</div><div class="lbl">Asset Cost</div></div>
     <div class="summary-box"><div class="val">${fmt(itDep1)}</div><div class="lbl">Year 1 Dep (IT Act)</div></div>
     <div class="summary-box"><div class="val">${fmt(caDep1)}</div><div class="lbl">Year 1 Dep (Co. Act)</div></div>
     <div class="summary-box"><div class="val">${fmt(salv)}</div><div class="lbl">Residual Value</div></div>`;

  document.getElementById("resultSection").style.display = "block";
  document.getElementById("resultSection").scrollIntoView({behavior:"smooth"});
}
</script>
<a href="https://wa.me/918427651580" target="_blank" class="wa-float" title="WhatsApp Support"><svg viewBox="0 0 24 24"><path d="M17.472 14.382c-.297-.149-1.758-.867-2.03-.967-.273-.099-.471-.148-.67.15-.197.297-.767.966-.94 1.164-.173.199-.347.223-.644.075-.297-.15-1.255-.463-2.39-1.475-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.298-.347.446-.52.149-.174.198-.298.298-.497.099-.198.05-.371-.025-.52-.075-.149-.669-1.612-.916-2.207-.242-.579-.487-.5-.669-.51-.173-.008-.371-.01-.57-.01-.198 0-.52.074-.792.372-.272.297-1.04 1.016-1.04 2.479 0 1.462 1.065 2.875 1.213 3.074.149.198 2.096 3.2 5.077 4.487.709.306 1.262.489 1.694.625.712.227 1.36.195 1.871.118.571-.085 1.758-.719 2.006-1.413.248-.694.248-1.289.173-1.413-.074-.124-.272-.198-.57-.347m-5.421 7.403h-.004a9.87 9.87 0 01-5.031-1.378l-.361-.214-3.741.982.998-3.648-.235-.374a9.86 9.86 0 01-1.51-5.26c.001-5.45 4.436-9.884 9.888-9.884 2.64 0 5.122 1.03 6.988 2.898a9.825 9.825 0 012.893 6.994c-.003 5.45-4.437 9.884-9.885 9.884m8.413-18.297A11.815 11.815 0 0012.05 0C5.495 0 .16 5.335.157 11.892c0 2.096.547 4.142 1.588 5.945L.057 24l6.305-1.654a11.882 11.882 0 005.683 1.448h.005c6.554 0 11.89-5.335 11.893-11.893a11.821 11.821 0 00-3.48-8.413z"/></svg></a>
</body></html>"""


# ══════════════════════════════════════════════════════════════════════════════
#  ADMIN PANEL
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
#  MSME DISALLOWANCE CALCULATOR
# ══════════════════════════════════════════════════════════════════════════════

MSME_T = r"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>MSME Disallowance Calculator – CA Toolkit</title>

<style>
""" + BASE_CSS + """
.hero{text-align:center;padding:32px 24px 16px;max-width:760px;margin:0 auto}
.hero-badge{display:inline-flex;align-items:center;gap:6px;background:#ECFDF5;color:#065F46;
            border:1px solid #A7F3D0;border-radius:99px;padding:5px 14px;font-size:12px;font-weight:600;margin-bottom:12px}
h1{font-size:clamp(20px,4vw,32px);font-weight:800;line-height:1.15;letter-spacing:-.5px;margin-bottom:8px}
h1 em{font-style:normal;color:var(--accent)}
.hero p{font-size:13px;color:var(--muted);line-height:1.7;max-width:560px;margin:0 auto}
.wrap{max-width:1100px;margin:0 auto;padding:16px 24px 48px}
.card{background:var(--white);border-radius:var(--radius);border:1px solid var(--border);
      box-shadow:var(--shadow);overflow:hidden;margin-bottom:20px}
.card-head{padding:14px 18px;border-bottom:1px solid var(--border);display:flex;align-items:center;gap:10px}
.card-head .icon{width:30px;height:30px;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:15px}
.card-head h2{font-size:14px;font-weight:700}
.card-head p{font-size:12px;color:var(--muted);margin-top:1px}
.card-body{padding:18px}
.field{margin-bottom:14px}
label{display:block;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.04em;color:var(--muted);margin-bottom:4px}
.hint{font-size:11px;color:var(--muted);margin-top:3px}
input[type=file],input[type=number],input[type=date]{width:100%;border:1.5px solid var(--border);border-radius:8px;
  padding:8px 11px;font-family:inherit;font-size:13px;color:var(--ink);background:var(--white);
  transition:border-color .2s;outline:none}
input:focus{border-color:var(--brand)}
.row2{display:grid;grid-template-columns:1fr 1fr;gap:14px}
.btn{background:var(--brand);color:#fff;border:none;border-radius:8px;padding:10px 20px;
     font-family:inherit;font-size:13px;font-weight:700;cursor:pointer;transition:background .2s}
.btn:hover{background:var(--brand-d)}
.btn-full{width:100%;padding:12px;font-size:14px}
/* Format table */
.fmt-table{width:100%;border-collapse:collapse;font-size:12px}
.fmt-table th{text-align:left;font-size:10px;letter-spacing:.06em;text-transform:uppercase;
              color:var(--muted);border-bottom:1.5px solid var(--border);padding:6px 10px;background:#F9FAFB}
.fmt-table td{padding:8px 10px;border-bottom:1px solid var(--border);vertical-align:top}
.fmt-table tr:last-child td{border:none}
.col-req{background:#EFF6FF;color:var(--brand);font-size:10px;font-weight:700;
         padding:1px 6px;border-radius:4px;font-family:monospace}
/* Results */
.summary-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:20px}
@media(max-width:700px){.summary-grid{grid-template-columns:1fr 1fr}}
.sbox{background:var(--white);border:1px solid var(--border);border-radius:10px;padding:14px 16px}
.sbox.red{background:#FEF2F2;border-color:#FECACA}
.sbox.green{background:#ECFDF5;border-color:#A7F3D0}
.sbox.yellow{background:#FFFBEB;border-color:#FDE68A}
.sbox .val{font-size:20px;font-weight:800;margin-bottom:3px}
.sbox .lbl{font-size:11px;color:var(--muted);font-weight:500}
.sbox.red .val{color:#991B1B}
.sbox.green .val{color:#065F46}
.sbox.yellow .val{color:#92400E}
/* Result table */
.res-table{width:100%;border-collapse:collapse;font-size:12px}
.res-table th{text-align:left;font-size:10px;letter-spacing:.06em;text-transform:uppercase;
              color:var(--muted);border-bottom:1.5px solid var(--border);padding:7px 8px;
              background:#F9FAFB;position:sticky;top:0}
.res-table td{padding:8px;border-bottom:1px solid var(--border);vertical-align:middle}
.res-table tr:hover td{background:#F9FAFB}
.row-ok{background:#F0FDF4}
.row-warn{background:#FFFBEB}
.row-over{background:#FEF2F2}
.badge-ok{background:#ECFDF5;color:#065F46;font-size:10px;font-weight:700;padding:2px 8px;border-radius:99px}
.badge-warn{background:#FFFBEB;color:#92400E;font-size:10px;font-weight:700;padding:2px 8px;border-radius:99px}
.badge-over{background:#FEF2F2;color:#991B1B;font-size:10px;font-weight:700;padding:2px 8px;border-radius:99px}
.badge-na{background:#F3F4F6;color:var(--muted);font-size:10px;font-weight:700;padding:2px 8px;border-radius:99px}
.note-box{background:#FEF2F2;border:1px solid #FECACA;border-radius:8px;padding:10px 14px;
          font-size:12px;color:#991B1B;margin-top:12px;line-height:1.7}
.info-box{background:#EFF6FF;border:1px solid #BFDBFE;border-radius:8px;padding:10px 14px;
          font-size:12px;color:#1e40af;margin-bottom:16px;line-height:1.7}
.dl-btn{background:var(--green);color:#fff;border:none;border-radius:8px;padding:8px 16px;
        font-family:inherit;font-size:12px;font-weight:600;cursor:pointer;margin-left:8px}
footer{background:#0f1b2d;color:#9CA3AF;font-size:12px;padding:0}
.ft-main{display:grid;grid-template-columns:2fr 1fr 1.4fr;gap:40px;padding:40px 48px;max-width:1200px;margin:0 auto}
.ft-brand-name{color:#fff;font-size:18px;font-weight:800;margin-bottom:12px}
.ft-brand-desc{font-size:12.5px;line-height:1.75;color:#9CA3AF;max-width:340px;text-align:justify}
.ft-col-title{color:#fff;font-size:14px;font-weight:700;margin-bottom:14px}
.ft-links{list-style:none;padding:0;margin:0}
.ft-links li{margin-bottom:8px}
.ft-links a{color:#9CA3AF;text-decoration:none;font-size:13px;transition:color .2s}
.ft-links a:hover{color:#fff}
.ft-contact-name{color:#fff;font-weight:700;font-size:13px;margin-bottom:6px}
.ft-contact-addr{color:#9CA3AF;font-size:12px;line-height:1.7;margin-bottom:10px}
.ft-contact-line{color:#9CA3AF;font-size:12px;margin-bottom:4px}
.ft-socials{display:flex;gap:14px;margin-top:12px}
.ft-socials a{color:#9CA3AF;transition:color .2s}
.ft-socials a:hover{color:#fff}
.ft-socials svg{width:20px;height:20px;fill:currentColor}
.ft-bottom{background:#0a1422;border-top:1px solid #1e2d42;padding:12px 48px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px}
.ft-bottom-left{font-size:11px;color:#6B7280}
.ft-bottom-right{font-size:11px;color:#6B7280}
@media(max-width:768px){.ft-main{grid-template-columns:1fr;padding:28px 20px;gap:24px}.ft-bottom{padding:12px 20px;flex-direction:column;text-align:center}}
</style></head><body>

<nav>
  <a href="/" class="logo">CA<span class="logo-dot"></span><span>Toolkit</span></a>
  <div class="nav-right">
    {% if username %}<span class="nav-user">👤 <strong>{{ username }}</strong></span>
    {% if is_admin %}<a href="/admin" class="nav-btn">Admin</a>{% endif %}
    <a href="/logout" class="nav-link">Sign out</a>
    {% else %}<a href="/login" class="nav-btn">Sign In</a>{% endif %}
    <a href="/" class="nav-btn dash">⬅ Dashboard</a>
  </div>
</nav>

<section class="hero">
  <div class="hero-badge">🆓 Free · No Login Required</div>
  <h1>MSME Disallowance Calculator</h1>
  <p>Check creditor payments against MSME time limits under <strong>Section 43B(h)</strong>. Upload your creditors list and instantly see which payments are overdue and the total disallowance amount.</p>
</section>

<div class="wrap">

  <div class="info-box">
    ℹ️ <strong>Section 43B(h) — IT Act:</strong> Payments to MSME suppliers must be made within 15 days (no agreement) or 45 days (written agreement) from date of invoice. Any unpaid amount beyond the limit is <strong>disallowed</strong> as a deduction in the year of computation and allowed only in the year of actual payment.
  </div>

  <!-- FORMAT GUIDE -->
  <div class="card">
    <div class="card-head">
      <div class="icon" style="background:#FFFBEB">📋</div>
      <div><h2>Required Excel Format</h2><p>Your upload must follow this column structure exactly</p></div>
    </div>
    <div class="card-body" style="padding:0;overflow-x:auto">
      <table class="fmt-table">
        <thead><tr><th>#</th><th>Column Name</th><th>Format</th><th>Example</th><th>Notes</th></tr></thead>
        <tbody>
          <tr><td>A</td><td><span class="col-req">Creditor Name</span></td><td>Text</td><td>ABC Enterprises</td><td>Name of the MSME supplier</td></tr>
          <tr><td>B</td><td><span class="col-req">Invoice Date</span></td><td>DD/MM/YYYY</td><td>01/04/2025</td><td>Date of invoice / bill received</td></tr>
          <tr><td>C</td><td><span class="col-req">Invoice Amount</span></td><td>Number</td><td>50000</td><td>Total invoice amount (₹)</td></tr>
          <tr><td>D</td><td><span class="col-req">Payment Date</span></td><td>DD/MM/YYYY or blank</td><td>20/05/2025</td><td>Leave blank if payment not yet made</td></tr>
          <tr><td>E</td><td><span class="col-req">Amount Paid</span></td><td>Number or 0</td><td>50000</td><td>Amount actually paid (0 if unpaid)</td></tr>
          <tr><td>F</td><td><span class="col-req">Written Agreement</span></td><td>Yes / No</td><td>No</td><td>Yes = 45 day limit, No = 15 day limit</td></tr>
          <tr><td>G</td><td><span class="col-req">MSME Category</span></td><td>Micro / Small / Medium</td><td>Small</td><td>MSME registration category of supplier</td></tr>
        </tbody>
      </table>
      <div style="padding:12px 16px;border-top:1px solid var(--border);display:flex;align-items:center;gap:12px;flex-wrap:wrap">
        <span style="font-size:12px;color:var(--muted)">Download blank template to fill in your data:</span>
        <button class="btn" onclick="downloadTemplate()" style="padding:7px 14px;font-size:12px">⬇ Download Template (.xlsx)</button>
      </div>
    </div>
  </div>

  <!-- UPLOAD -->
  <div class="card">
    <div class="card-head">
      <div class="icon" style="background:#EFF6FF">📁</div>
      <div><h2>Upload Creditors List</h2><p>Excel file (.xlsx) following the format above</p></div>
    </div>
    <div class="card-body">
      <div class="row2">
        <div class="field">
          <label>Upload Excel File (.xlsx)</label>
          <input type="file" id="msmeFile" accept=".xlsx"/>
          <p class="hint">Must follow the format shown above</p>
        </div>
        <div class="field">
          <label>Assessment Year</label>
          <input type="number" id="assessYear" value="2026" min="2020" max="2030"/>
          <p class="hint">Year for which disallowance is computed</p>
        </div>
      </div>
      <button class="btn btn-full" onclick="processMSME()">Analyse &amp; Calculate Disallowance →</button>
      <div id="msmeError" style="display:none;margin-top:10px;background:#FEF2F2;border:1px solid #FECACA;
           border-radius:8px;padding:10px 14px;font-size:13px;color:#991B1B"></div>
    </div>
  </div>

  <!-- RESULTS -->
  <div id="msmeResults" style="display:none">
    <div class="summary-grid" id="summaryGrid"></div>

    <div class="card">
      <div class="card-head">
        <div class="icon" style="background:#FEF2F2">📊</div>
        <div>
          <h2>Creditor-wise Analysis</h2>
          <p>Overdue payments highlighted in red · Near-due in yellow · On time in green</p>
        </div>
        <div style="margin-left:auto">
          <button class="dl-btn" onclick="exportResults()">⬇ Export Results</button>
        </div>
      </div>
      <div style="overflow-x:auto">
        <table class="res-table" id="resultsTable">
          <thead><tr>
            <th>Creditor</th><th>Category</th><th>Invoice Date</th><th>Invoice Amt</th>
            <th>Paid Amt</th><th>Payment Date</th><th>Limit</th><th>Due Date</th>
            <th>Days Overdue</th><th>Unpaid Amt</th><th>Status</th>
          </tr></thead>
          <tbody id="resultsBody"></tbody>
        </table>
      </div>
      <div class="note-box" id="disallowNote"></div>
    </div>
  </div>

</div>

<footer>
  <div class="ft-main">
    <div>
      <div class="ft-brand-name">CA<span>Toolkit</span></div>
      <p class="ft-brand-desc">CA Toolkit is a comprehensive utility platform built by a CA Article from Ludhiana, Punjab, providing automation tools for Indian Chartered Accountants. The platform saves hours of manual work every year — from Balance Sheet year-shift to GST reconciliation, tax calculations, and more.</p>
    </div>
    <div>
      <div class="ft-col-title">Know More</div>
      <ul class="ft-links">
        <li><a href="/">Home</a></li>
        <li><a href="/">BS Year Shift</a></li>
        <li><a href="/tool/tb-to-bs">TB → Balance Sheet</a></li>
        <li><a href="/tool/tax-calculator">Tax Calculator</a></li>
        <li><a href="/privacy">Privacy Policy</a></li>
      </ul>
    </div>
    <div>
      <div class="ft-col-title">Contact Us</div>
      <div class="ft-contact-name">CA Toolkit</div>
      <div class="ft-contact-addr">Built for Indian Chartered Accountants<br/>Created by CA Article · Ludhiana, Punjab</div>
    </div>
  </div>
  <div class="ft-bottom">
    <span class="ft-bottom-left">©2026 CA Toolkit · All Rights Reserved</span>
    <span class="ft-bottom-right">Built for Indian CAs · Ludhiana, Punjab</span>
  </div>
</footer>

<script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
<script>
const fmt = n => "₹" + Math.round(n).toLocaleString("en-IN");

function parseDate(val){
  if(!val) return null;
  if(val instanceof Date) return val;
  // Try DD/MM/YYYY
  const s = String(val).trim();
  const parts = s.split("/");
  if(parts.length===3) return new Date(parts[2], parts[1]-1, parts[0]);
  // Try Excel serial
  if(!isNaN(val)){
    const d = new Date((val - 25569) * 86400 * 1000);
    return new Date(d.getFullYear(), d.getMonth(), d.getDate());
  }
  return new Date(val);
}

function daysBetween(d1, d2){
  return Math.round((d2 - d1) / (1000*60*60*24));
}

function fmtDate(d){
  if(!d) return "—";
  return d.toLocaleDateString("en-IN",{day:"2-digit",month:"short",year:"numeric"});
}

let processedRows = [];

function processMSME(){
  const file = document.getElementById("msmeFile").files[0];
  const errEl = document.getElementById("msmeError");
  errEl.style.display = "none";

  if(!file){ showErr("Please select an Excel file."); return; }

  const reader = new FileReader();
  reader.onload = function(e){
    try{
      const wb   = XLSX.read(e.target.result, {type:"binary", cellDates:true});
      const ws   = wb.Sheets[wb.SheetNames[0]];
      const rows = XLSX.utils.sheet_to_json(ws, {header:1, raw:false});

      if(rows.length < 2){ showErr("File appears empty. Please check the format."); return; }

      // Skip header row
      const data = rows.slice(1).filter(r => r && r[0]);
      if(data.length === 0){ showErr("No data rows found. Please check the format."); return; }

      processedRows = [];
      let totalInvoice=0, totalPaid=0, totalDisallow=0, totalOverdue=0;
      const today = new Date();

      const tbody = document.getElementById("resultsBody");
      tbody.innerHTML = "";

      data.forEach((row, i) => {
        const name        = String(row[0]||"").trim();
        const invoiceDate = parseDate(row[1]);
        const invoiceAmt  = parseFloat(String(row[2]||"0").replace(/,/g,""))||0;
        const paymentDate = parseDate(row[3]);
        const paidAmt     = parseFloat(String(row[4]||"0").replace(/,/g,""))||0;
        const hasAgreement= String(row[5]||"No").trim().toLowerCase()==="yes";
        const category    = String(row[6]||"MSME").trim();

        if(!name||!invoiceDate) return;

        const limitDays = hasAgreement ? 45 : 15;
        const dueDate   = new Date(invoiceDate);
        dueDate.setDate(dueDate.getDate() + limitDays);

        const refDate  = paymentDate || today;
        const daysDiff = daysBetween(dueDate, refDate);
        const unpaid   = Math.max(0, invoiceAmt - paidAmt);
        const isOverdue = daysDiff > 0;
        const isNearDue = !isOverdue && daysDiff > -7;

        let status, rowClass, badge;
        if(!isOverdue && paidAmt >= invoiceAmt){
          status="✓ Paid on time"; rowClass="row-ok"; badge=`<span class="badge-ok">Paid On Time</span>`;
        } else if(isOverdue && unpaid > 0){
          status="⚠ Overdue"; rowClass="row-over"; badge=`<span class="badge-over">Overdue</span>`;
          totalDisallow += unpaid;
          totalOverdue++;
        } else if(!isOverdue && unpaid > 0){
          status="Near due"; rowClass="row-warn"; badge=`<span class="badge-warn">Pending</span>`;
        } else if(isOverdue && paidAmt >= invoiceAmt){
          status="Paid late"; rowClass="row-warn"; badge=`<span class="badge-warn">Paid Late</span>`;
        } else {
          status="—"; rowClass=""; badge=`<span class="badge-na">N/A</span>`;
        }

        totalInvoice += invoiceAmt;
        totalPaid    += paidAmt;

        processedRows.push({name,category,invoiceDate,invoiceAmt,paidAmt,paymentDate,
          limitDays,dueDate,daysDiff,unpaid,status,isOverdue});

        const tr = document.createElement("tr");
        tr.className = rowClass;
        tr.innerHTML = `
          <td><strong>${name}</strong></td>
          <td>${category}</td>
          <td>${fmtDate(invoiceDate)}</td>
          <td>${fmt(invoiceAmt)}</td>
          <td>${fmt(paidAmt)}</td>
          <td>${fmtDate(paymentDate)}</td>
          <td>${limitDays} days</td>
          <td>${fmtDate(dueDate)}</td>
          <td style="font-weight:700;color:${isOverdue&&unpaid>0?"#991B1B":daysDiff>-7?"#92400E":"#065F46"}">${isOverdue?"+"+daysDiff+" days":daysDiff===0?"Today":Math.abs(daysDiff)+" days left"}</td>
          <td style="font-weight:700;color:${unpaid>0?"#991B1B":"#065F46"}">${fmt(unpaid)}</td>
          <td>${badge}</td>`;
        tbody.appendChild(tr);
      });

      // Summary
      const ayear = document.getElementById("assessYear").value;
      document.getElementById("summaryGrid").innerHTML = `
        <div class="sbox"><div class="val">${processedRows.length}</div><div class="lbl">Total Creditors</div></div>
        <div class="sbox yellow"><div class="val">${fmt(totalInvoice)}</div><div class="lbl">Total Invoice Value</div></div>
        <div class="sbox ${totalOverdue>0?"red":"green"}"><div class="val">${totalOverdue}</div><div class="lbl">Overdue Creditors</div></div>
        <div class="sbox ${totalDisallow>0?"red":"green"}"><div class="val">${fmt(totalDisallow)}</div><div class="lbl">Disallowance u/s 43B(h)</div></div>`;

      document.getElementById("disallowNote").innerHTML = totalDisallow > 0
        ? `⚠ <strong>Total disallowance u/s 43B(h) for AY ${ayear}-${parseInt(ayear)+1}: ${fmt(totalDisallow)}</strong><br>
           This amount will be added back to income and disallowed as a deduction. It will be allowed only in the year when actual payment is made to the MSME supplier.`
        : `✓ No disallowance applicable. All MSME payments are within the prescribed time limits.`;

      document.getElementById("msmeResults").style.display = "block";
      document.getElementById("msmeResults").scrollIntoView({behavior:"smooth"});

    } catch(err){
      showErr("Error reading file: "+err.message+". Please ensure file follows the required format.");
    }
  };
  reader.readAsBinaryString(file);
}

function showErr(msg){
  const el = document.getElementById("msmeError");
  el.textContent = msg; el.style.display = "block";
}

function downloadTemplate(){
  const wb = XLSX.utils.book_new();
  const data = [
    ["Creditor Name","Invoice Date","Invoice Amount","Payment Date","Amount Paid","Written Agreement","MSME Category"],
    ["ABC Enterprises","01/04/2025","50000","20/04/2025","50000","No","Small"],
    ["XYZ Traders","15/04/2025","120000","","0","Yes","Micro"],
    ["PQR Industries","01/05/2025","80000","20/06/2025","80000","No","Medium"],
  ];
  const ws = XLSX.utils.aoa_to_sheet(data);
  ws["!cols"] = [{wch:20},{wch:14},{wch:16},{wch:14},{wch:12},{wch:18},{wch:16}];
  XLSX.utils.book_append_sheet(wb, ws, "Creditors");
  XLSX.writeFile(wb, "MSME_Creditors_Template.xlsx");
}

function exportResults(){
  if(!processedRows.length) return;
  const data = [["Creditor","Category","Invoice Date","Invoice Amt","Paid Amt","Payment Date","Limit","Due Date","Days Overdue","Unpaid Amt","Status"]];
  processedRows.forEach(r => {
    data.push([r.name,r.category,fmtDate(r.invoiceDate),r.invoiceAmt,r.paidAmt,
      fmtDate(r.paymentDate),r.limitDays+" days",fmtDate(r.dueDate),
      r.isOverdue?"+"+r.daysDiff+" days":Math.abs(r.daysDiff)+" days left",r.unpaid,r.status]);
  });
  const wb = XLSX.utils.book_new();
  const ws = XLSX.utils.aoa_to_sheet(data);
  XLSX.utils.book_append_sheet(wb, ws, "MSME Analysis");
  XLSX.writeFile(wb, "MSME_Disallowance_Analysis.xlsx");
}
</script>
<a href="https://wa.me/918427651580" target="_blank" class="wa-float" title="WhatsApp Support"><svg viewBox="0 0 24 24"><path d="M17.472 14.382c-.297-.149-1.758-.867-2.03-.967-.273-.099-.471-.148-.67.15-.197.297-.767.966-.94 1.164-.173.199-.347.223-.644.075-.297-.15-1.255-.463-2.39-1.475-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.298-.347.446-.52.149-.174.198-.298.298-.497.099-.198.05-.371-.025-.52-.075-.149-.669-1.612-.916-2.207-.242-.579-.487-.5-.669-.51-.173-.008-.371-.01-.57-.01-.198 0-.52.074-.792.372-.272.297-1.04 1.016-1.04 2.479 0 1.462 1.065 2.875 1.213 3.074.149.198 2.096 3.2 5.077 4.487.709.306 1.262.489 1.694.625.712.227 1.36.195 1.871.118.571-.085 1.758-.719 2.006-1.413.248-.694.248-1.289.173-1.413-.074-.124-.272-.198-.57-.347m-5.421 7.403h-.004a9.87 9.87 0 01-5.031-1.378l-.361-.214-3.741.982.998-3.648-.235-.374a9.86 9.86 0 01-1.51-5.26c.001-5.45 4.436-9.884 9.888-9.884 2.64 0 5.122 1.03 6.988 2.898a9.825 9.825 0 012.893 6.994c-.003 5.45-4.437 9.884-9.885 9.884m8.413-18.297A11.815 11.815 0 0012.05 0C5.495 0 .16 5.335.157 11.892c0 2.096.547 4.142 1.588 5.945L.057 24l6.305-1.654a11.882 11.882 0 005.683 1.448h.005c6.554 0 11.89-5.335 11.893-11.893a11.821 11.821 0 00-3.48-8.413z"/></svg></a>
</body></html>"""


# ══════════════════════════════════════════════════════════════════════════════
#  CAPITAL GAINS CALCULATOR
# ══════════════════════════════════════════════════════════════════════════════

CG_CALC_T = r"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Capital Gains Calculator – CA Toolkit</title>

<style>
""" + BASE_CSS + """
.hero{text-align:center;padding:32px 24px 16px;max-width:760px;margin:0 auto}
.hero-badge{display:inline-flex;align-items:center;gap:6px;background:#ECFDF5;color:#065F46;
            border:1px solid #A7F3D0;border-radius:99px;padding:5px 14px;font-size:12px;font-weight:600;margin-bottom:12px}
h1{font-size:clamp(20px,4vw,32px);font-weight:800;line-height:1.15;letter-spacing:-.5px;margin-bottom:8px}
h1 em{font-style:normal;color:var(--accent)}
.hero p{font-size:13px;color:var(--muted);line-height:1.7;max-width:520px;margin:0 auto}
.wrap{max-width:1100px;margin:0 auto;padding:16px 24px 48px;display:grid;grid-template-columns:1fr 1fr;gap:20px;align-items:start}
@media(max-width:800px){.wrap{grid-template-columns:1fr}}
.card{background:var(--white);border-radius:var(--radius);border:1px solid var(--border);
      box-shadow:var(--shadow);overflow:hidden;margin-bottom:16px}
.card:last-child{margin-bottom:0}
.card-head{padding:14px 18px;border-bottom:1px solid var(--border);display:flex;align-items:center;gap:10px}
.card-head .icon{width:30px;height:30px;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:15px}
.card-head h2{font-size:14px;font-weight:700}
.card-head p{font-size:12px;color:var(--muted);margin-top:1px}
.card-body{padding:16px}
.field{margin-bottom:13px}
.row2{display:grid;grid-template-columns:1fr 1fr;gap:12px}
label{display:block;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.04em;color:var(--muted);margin-bottom:4px}
.hint{font-size:11px;color:var(--muted);margin-top:3px}
select,input[type=number],input[type=text]{width:100%;border:1.5px solid var(--border);border-radius:8px;
  padding:8px 11px;font-family:inherit;font-size:13px;color:var(--ink);background:var(--white);
  transition:border-color .2s;outline:none}
select:focus,input:focus{border-color:var(--brand)}
.btn{width:100%;background:var(--brand);color:#fff;border:none;border-radius:8px;
     padding:11px;font-family:inherit;font-size:14px;font-weight:700;cursor:pointer;transition:background .2s}
.btn:hover{background:var(--brand-d)}
/* Tabs */
.tabs{display:flex;gap:0;margin-bottom:16px;border-radius:8px;overflow:hidden;border:1px solid var(--border)}
.tab{flex:1;padding:9px;font-family:inherit;font-size:12px;font-weight:600;cursor:pointer;
     background:var(--white);color:var(--muted);border:none;transition:all .2s}
.tab.active{background:var(--brand);color:#fff}
/* Result boxes */
.rboxes{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:14px}
.rbox{border-radius:10px;padding:13px 15px}
.rbox-blue{background:#EFF6FF;border:1.5px solid #BFDBFE}
.rbox-green{background:#ECFDF5;border:1.5px solid #A7F3D0}
.rbox-red{background:#FEF2F2;border:1.5px solid #FECACA}
.rbox-total{background:#1D4ED8;border:1.5px solid #1D4ED8;grid-column:1/-1}
.rbox .val{font-size:20px;font-weight:800;margin-bottom:2px}
.rbox .lbl{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.05em;opacity:.75}
.rbox .sub{font-size:11px;margin-top:4px;opacity:.8}
.rbox-blue  .val{color:#1D4ED8}.rbox-blue  .lbl{color:#1D4ED8}
.rbox-green .val{color:#065F46}.rbox-green .lbl{color:#065F46}
.rbox-red   .val{color:#991B1B}.rbox-red   .lbl{color:#991B1B}
.rbox-total .val{color:#fff;font-size:22px}.rbox-total .lbl{color:rgba(255,255,255,.75)}
.rbox-total .sub{color:rgba(255,255,255,.8);font-size:11px}
.dtable{width:100%;border-collapse:collapse;font-size:12px;margin-top:10px}
.dtable td{padding:6px 2px;border-bottom:1px solid var(--border)}
.dtable tr:last-child td{border:none;font-weight:700}
.dtable td:last-child{text-align:right;font-weight:600}
.compare-grid{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-top:12px}
.cbox{border-radius:10px;padding:13px 15px;border:1.5px solid var(--border);background:var(--white)}
.cbox.winner{border-color:var(--green);background:#ECFDF5}
.cbox h3{font-size:12px;font-weight:700;margin-bottom:8px;color:var(--ink)}
.cbox .ctax{font-size:20px;font-weight:800;color:var(--brand);margin-bottom:3px}
.cbox.winner .ctax{color:#065F46}
.cbox .csub{font-size:11px;color:var(--muted)}
.winner-badge{background:var(--green);color:#fff;font-size:10px;font-weight:700;
              padding:2px 8px;border-radius:99px;display:inline-block;margin-bottom:6px}
.reverse-box{background:#F5F3FF;border:1.5px solid #DDD6FE;border-radius:10px;padding:14px 16px;margin-top:12px}
.reverse-box h3{font-size:12px;font-weight:700;color:#5B21B6;margin-bottom:8px}
.reverse-box .rval{font-size:22px;font-weight:800;color:#5B21B6;margin-bottom:3px}
.reverse-box .rsub{font-size:11px;color:var(--muted)}
.info-box{background:#EFF6FF;border:1px solid #BFDBFE;border-radius:8px;padding:10px 14px;
          font-size:12px;color:#1e40af;margin-bottom:14px;line-height:1.7}
.note-box{background:#FFFBEB;border:1px solid #FDE68A;border-radius:8px;
          padding:10px 12px;font-size:11px;color:#92400E;margin-top:10px;line-height:1.6}
/* CII table */
.cii-table{width:100%;border-collapse:collapse;font-size:11px}
.cii-table th{text-align:left;font-size:10px;letter-spacing:.06em;text-transform:uppercase;
              color:var(--muted);border-bottom:1.5px solid var(--border);padding:5px 6px}
.cii-table td{padding:5px 6px;border-bottom:1px solid var(--border)}
.cii-table tr:last-child td{border:none}
.cii-table tr:hover td{background:#F9FAFB}
.cii-table .highlight{background:#EFF6FF;font-weight:700}
footer{background:#0f1b2d;color:#9CA3AF;font-size:12px;padding:0}
.ft-main{display:grid;grid-template-columns:2fr 1fr 1.4fr;gap:40px;padding:40px 48px;max-width:1200px;margin:0 auto}
.ft-brand-name{color:#fff;font-size:18px;font-weight:800;margin-bottom:12px}
.ft-brand-desc{font-size:12.5px;line-height:1.75;color:#9CA3AF;max-width:340px;text-align:justify}
.ft-col-title{color:#fff;font-size:14px;font-weight:700;margin-bottom:14px}
.ft-links{list-style:none;padding:0;margin:0}
.ft-links li{margin-bottom:8px}
.ft-links a{color:#9CA3AF;text-decoration:none;font-size:13px;transition:color .2s}
.ft-links a:hover{color:#fff}
.ft-contact-name{color:#fff;font-weight:700;font-size:13px;margin-bottom:6px}
.ft-contact-addr{color:#9CA3AF;font-size:12px;line-height:1.7;margin-bottom:10px}
.ft-contact-line{color:#9CA3AF;font-size:12px;margin-bottom:4px}
.ft-socials{display:flex;gap:14px;margin-top:12px}
.ft-socials a{color:#9CA3AF;transition:color .2s}
.ft-socials a:hover{color:#fff}
.ft-socials svg{width:20px;height:20px;fill:currentColor}
.ft-bottom{background:#0a1422;border-top:1px solid #1e2d42;padding:12px 48px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px}
.ft-bottom-left{font-size:11px;color:#6B7280}
.ft-bottom-right{font-size:11px;color:#6B7280}
@media(max-width:768px){.ft-main{grid-template-columns:1fr;padding:28px 20px;gap:24px}.ft-bottom{padding:12px 20px;flex-direction:column;text-align:center}}
</style></head><body>

<nav>
  <a href="/" class="logo">CA<span class="logo-dot"></span><span>Toolkit</span></a>
  <div class="nav-right">
    {% if username %}<span class="nav-user">👤 <strong>{{ username }}</strong></span>
    {% if is_admin %}<a href="/admin" class="nav-btn">Admin</a>{% endif %}
    <a href="/logout" class="nav-link">Sign out</a>
    {% else %}<a href="/login" class="nav-btn">Sign In</a>{% endif %}
    <a href="/" class="nav-btn dash">⬅ Dashboard</a>
  </div>
</nav>

<section class="hero">
  <div class="hero-badge">🆓 Free · No Login Required</div>
  <h1>Capital Gains Tax Calculator</h1>
  <p>Calculate LTCG / STCG on property, shares, mutual funds and more. Compare old regime (with indexation) vs new regime. Includes reverse calculator — find the <strong>sale price for zero tax</strong>.</p>
</section>

<div class="wrap">
  <!-- LEFT: INPUT -->
  <div>
    <div class="card">
      <div class="card-head">
        <div class="icon" style="background:#EFF6FF">💰</div>
        <div><h2>Capital Gains Calculator</h2><p>IT Act 2025 · Tax Year 2026-27</p></div>
      </div>
      <div class="card-body">

        <div class="info-box">ℹ️ Under IT Act 2025: LTCG on property is <strong>12.5% without indexation</strong>. Old regime (20% with indexation) applicable for assets purchased before 23 Jul 2023. Choose as applicable.</div>

        <div class="field">
          <label>Asset Type</label>
          <select id="assetType" onchange="updateAssetUI()">
            <option value="property">Immovable Property (Land/Building)</option>
            <option value="equity">Listed Equity Shares / Equity MF</option>
            <option value="debt">Debt Mutual Funds / Bonds</option>
            <option value="gold">Gold / Gold ETF / Sovereign Gold Bond</option>
            <option value="unlisted">Unlisted Shares</option>
            <option value="other">Other Capital Assets</option>
          </select>
        </div>

        <div class="row2">
          <div class="field">
            <label>Date of Purchase</label>
            <input type="text" id="purchaseDate" placeholder="DD/MM/YYYY" maxlength="10" oninput="autoDateSlash(this)" autocomplete="off" inputmode="numeric"/>
            <p class="hint">Original acquisition date</p>
          </div>
          <div class="field">
            <label>Date of Sale</label>
            <input type="text" id="saleDate" placeholder="DD/MM/YYYY" maxlength="10" oninput="autoDateSlash(this)" autocomplete="off" inputmode="numeric"/>
            <p class="hint">Date of transfer/sale</p>
          </div>
        </div>

        <div class="row2">
          <div class="field">
            <label>Purchase Price (₹)</label>
            <input type="number" id="purchasePrice" placeholder="e.g. 2000000" min="0"/>
            <p class="hint">Cost of acquisition</p>
          </div>
          <div class="field">
            <label>Sale Price (₹)</label>
            <input type="number" id="salePrice" placeholder="e.g. 5000000" min="0"/>
            <p class="hint">Full value of consideration</p>
          </div>
        </div>

        <div class="row2">
          <div class="field">
            <label>Improvement Cost (₹)</label>
            <input type="number" id="improveCost" placeholder="0" min="0" value="0"/>
            <p class="hint">Cost of any improvements made</p>
          </div>
          <div class="field">
            <label>Transfer Expenses (₹)</label>
            <input type="number" id="transferCost" placeholder="0" min="0" value="0"/>
            <p class="hint">Brokerage, registration, legal fees</p>
          </div>
        </div>

        <div class="field" id="exemptionField">
          <label>Exemption Claimed</label>
          <select id="exemptionType">
            <option value="0">None</option>
            <option value="54">Sec 54 — Residential House Property</option>
            <option value="54B">Sec 54B — Agricultural Land</option>
            <option value="54EC">Sec 54EC — NHAI/REC Bonds (Max ₹50L)</option>
            <option value="54F">Sec 54F — Any LTCG → Residential House</option>
          </select>
        </div>
        <div class="field" id="exemptionAmtField">
          <label>Exemption Amount (₹)</label>
          <input type="number" id="exemptionAmt" placeholder="0" min="0" value="0"/>
          <p class="hint">Amount claimed under selected exemption</p>
        </div>

        <div class="row2">
          <div class="field">
            <label>Tax Year (Previous Year)</label>
            <select id="taxYear" onchange="updateCIITable()">
              <option value="2025-26" selected>PY 2025-26 (AY 2026-27)</option>
              <option value="2026-27">PY 2026-27 (AY 2027-28) · Future</option>
              <option value="2024-25">PY 2024-25 (AY 2025-26)</option>
              <option value="2023-24">PY 2023-24 (AY 2024-25)</option>
            </select>
            <p class="hint">FY in which asset is sold / will be sold</p>
          </div>
          <div class="field">
            <label>Assessee Type</label>
            <select id="assesseeType">
              <option value="individual">Individual / HUF</option>
              <option value="firm">Firm / LLP</option>
              <option value="company">Company</option>
            </select>
          </div>
        </div>

        <button class="btn" onclick="calcCG()">Calculate Capital Gains →</button>

        <!-- RESULTS -->
        <div id="cgResults" style="display:none;margin-top:16px">
          <div id="cgTypeLabel" style="font-size:13px;font-weight:700;margin-bottom:10px;color:var(--ink)"></div>

          <!-- Regime comparison -->
          <div class="compare-grid" id="compareGrid"></div>

          <!-- Detail breakdown -->
          <div id="detailBreakdown"></div>

          <!-- Reverse calculator result -->
          <div class="reverse-box" id="reverseBox" style="display:none">
            <h3>🔄 Reverse Calculator — Zero Tax Sale Price</h3>
            <div class="rval" id="revSalePrice"></div>
            <div class="rsub" id="revSub"></div>
          </div>

          <div class="note-box" id="cgNote"></div>
        </div>

      </div>
    </div>
  </div>

  <!-- RIGHT: INFO PANELS -->
  <div>
    <div class="card">
      <div class="card-head">
        <div class="icon" style="background:#FFFBEB">📋</div>
        <div><h2>Tax Rates — IT Act 2025</h2><p>LTCG &amp; STCG at a glance</p></div>
      </div>
      <div class="card-body" style="padding:0;overflow-x:auto">
        <table class="cii-table">
          <thead><tr><th>Asset</th><th>Holding</th><th>Rate</th><th>Threshold</th></tr></thead>
          <tbody>
            <tr><td>Immovable Property</td><td>&gt;24 months</td><td>12.5% (no idx) / 20% (with idx pre Jul-23)</td><td>₹1.25L (old ₹1L)</td></tr>
            <tr><td>Listed Equity / Eq MF</td><td>&gt;12 months</td><td>12.5%</td><td>₹1.25L exempt</td></tr>
            <tr><td>Listed Equity / Eq MF</td><td>≤12 months</td><td>20% (STCG)</td><td>Nil</td></tr>
            <tr><td>Debt MF / Bonds</td><td>Any</td><td>Slab rate</td><td>Nil</td></tr>
            <tr><td>Gold / Gold ETF</td><td>&gt;24 months</td><td>12.5%</td><td>Nil</td></tr>
            <tr><td>Unlisted Shares</td><td>&gt;24 months</td><td>12.5%</td><td>Nil</td></tr>
            <tr><td>Unlisted Shares</td><td>≤24 months</td><td>Slab rate</td><td>Nil</td></tr>
            <tr><td>Any STCG (others)</td><td>≤24/36 months</td><td>Slab rate</td><td>Nil</td></tr>
          </tbody>
        </table>
      </div>
    </div>

    <div class="card">
      <div class="card-head">
        <div class="icon" style="background:#F5F3FF">📈</div>
        <div><h2>Cost Inflation Index (CII)</h2><p>As notified by CBDT</p></div>
      </div>
      <div class="card-body" style="padding:0;max-height:260px;overflow-y:auto">
        <table class="cii-table">
          <thead><tr><th>FY</th><th>CII</th><th>FY</th><th>CII</th></tr></thead>
          <tbody>
            <tr><td>2001-02</td><td>100</td><td>2014-15</td><td>240</td></tr>
            <tr><td>2002-03</td><td>105</td><td>2015-16</td><td>254</td></tr>
            <tr><td>2003-04</td><td>109</td><td>2016-17</td><td>264</td></tr>
            <tr><td>2004-05</td><td>113</td><td>2017-18</td><td>272</td></tr>
            <tr><td>2005-06</td><td>117</td><td>2018-19</td><td>280</td></tr>
            <tr><td>2006-07</td><td>122</td><td>2019-20</td><td>289</td></tr>
            <tr><td>2007-08</td><td>129</td><td>2020-21</td><td>301</td></tr>
            <tr><td>2008-09</td><td>137</td><td>2021-22</td><td>317</td></tr>
            <tr><td>2009-10</td><td>148</td><td>2022-23</td><td>331</td></tr>
            <tr><td>2010-11</td><td>167</td><td>2023-24</td><td>348</td></tr>
            <tr><td>2011-12</td><td>184</td><td>2024-25</td><td>363</td></tr>
            <tr><td>2012-13</td><td>200</td><td>2025-26</td><td>380</td></tr>
            <tr><td>2013-14</td><td>220</td><td>2026-27</td><td>TBA</td></tr>
          </tbody>
        </table>
      </div>
    </div>

    <div class="card">
      <div class="card-head">
        <div class="icon" style="background:#F0FDF4">🏠</div>
        <div><h2>Key Exemptions</h2><p>Reduce your capital gains tax</p></div>
      </div>
      <div class="card-body">
        <div style="font-size:12px;line-height:2;color:var(--muted)">
          <p><strong style="color:var(--ink)">Sec 54</strong> — LTCG on residential property → invest in new house (within 2yr purchase / 3yr construct)</p>
          <p><strong style="color:var(--ink)">Sec 54B</strong> — LTCG on agricultural land → invest in new agricultural land</p>
          <p><strong style="color:var(--ink)">Sec 54EC</strong> — Any LTCG → NHAI/REC bonds (max ₹50L, lock-in 5yr)</p>
          <p><strong style="color:var(--ink)">Sec 54F</strong> — LTCG on any asset → invest in residential house (net consideration)</p>
          <p style="margin-top:8px;color:var(--red)"><strong>Note:</strong> Exemptions available only on LTCG. Claim only if actually investing.</p>
        </div>
      </div>
    </div>
  </div>
</div>

<footer>
  <div class="ft-main">
    <div>
      <div class="ft-brand-name">CA<span>Toolkit</span></div>
      <p class="ft-brand-desc">CA Toolkit is a comprehensive utility platform built by a CA Article from Ludhiana, Punjab, providing automation tools for Indian Chartered Accountants. The platform saves hours of manual work every year — from Balance Sheet year-shift to GST reconciliation, tax calculations, and more.</p>
    </div>
    <div>
      <div class="ft-col-title">Know More</div>
      <ul class="ft-links">
        <li><a href="/">Home</a></li>
        <li><a href="/">BS Year Shift</a></li>
        <li><a href="/tool/tb-to-bs">TB → Balance Sheet</a></li>
        <li><a href="/tool/tax-calculator">Tax Calculator</a></li>
        <li><a href="/privacy">Privacy Policy</a></li>
      </ul>
    </div>
    <div>
      <div class="ft-col-title">Contact Us</div>
      <div class="ft-contact-name">CA Toolkit</div>
      <div class="ft-contact-addr">Built for Indian Chartered Accountants<br/>Created by CA Article · Ludhiana, Punjab</div>
    </div>
  </div>
  <div class="ft-bottom">
    <span class="ft-bottom-left">©2026 CA Toolkit · All Rights Reserved</span>
    <span class="ft-bottom-right">Built for Indian CAs · Ludhiana, Punjab</span>
  </div>
</footer>

<script>
// CII table
const CII = {2001:100,2002:105,2003:109,2004:113,2005:117,2006:122,2007:129,2008:137,
             2009:148,2010:167,2011:184,2012:200,2013:220,2014:240,2015:254,2016:264,
             2017:272,2018:280,2019:289,2020:301,2021:317,2022:331,2023:348,2024:363,
             2025:380, 2026:400}; // CII 2026 = estimated, not yet notified by CBDT

const TAX_YEAR_FY = {"2023-24":2023,"2024-25":2024,"2025-26":2025,"2026-27":2026};
const TAX_YEAR_AY = {"2023-24":"AY 2024-25","2024-25":"AY 2025-26","2025-26":"AY 2026-27","2026-27":"AY 2027-28"};

function getSelectedSaleFY(){
  const el = document.getElementById("taxYear");
  return TAX_YEAR_FY[(el||{value:"2025-26"}).value] || 2025;
}

function updateCIITable(){
  const saleFY = getSelectedSaleFY();
  document.querySelectorAll(".cii-table tr").forEach(tr=>{
    const c = tr.cells[0]; if(!c) return;
    const yr = parseInt((c.textContent||"").replace("FY ","").split("-")[0]);
    tr.classList.toggle("highlight", yr === saleFY);
  });
  if(document.getElementById("cgResults").style.display !== "none") calcCG();
}

const fmt   = n => "₹"+Math.round(n).toLocaleString("en-IN");
const fmtPct= n => n.toFixed(2)+"%";

function parseMyDate(s){
  if(!s) return null;
  const p=String(s).trim().split("/");
  if(p.length===3) return new Date(parseInt(p[2]),parseInt(p[1])-1,parseInt(p[0]));
  return new Date(s);
}

function getFY(d){ return d.getMonth()>=3 ? d.getFullYear() : d.getFullYear()-1; }

function holdingMonths(d1,d2){
  return (d2.getFullYear()-d1.getFullYear())*12 + (d2.getMonth()-d1.getMonth());
}

function getCII(fy){ return CII[fy] || 380; }

function updateAssetUI(){
  const t = document.getElementById("assetType").value;
  const showExemption = ["property","equity","other","gold","unlisted"].includes(t);
  document.getElementById("exemptionField").style.display    = showExemption?"block":"none";
  document.getElementById("exemptionAmtField").style.display = showExemption?"block":"none";
}

function calcCG(){
  const asset     = document.getElementById("assetType").value;
  const pd        = parseMyDate(document.getElementById("purchaseDate").value);
  const sd        = parseMyDate(document.getElementById("saleDate").value);
  const pp        = parseFloat(document.getElementById("purchasePrice").value)||0;
  const sp        = parseFloat(document.getElementById("salePrice").value)||0;
  const ic        = parseFloat(document.getElementById("improveCost").value)||0;
  const tc        = parseFloat(document.getElementById("transferCost").value)||0;
  const exemAmt   = parseFloat(document.getElementById("exemptionAmt").value)||0;
  const exemType  = document.getElementById("exemptionType").value;
  const assessee  = document.getElementById("assesseeType").value;

  if(!pd||!sd){ alert("Please enter both purchase and sale dates."); return; }
  if(!pp||!sp){ alert("Please enter purchase price and sale price."); return; }
  if(sd<=pd){ alert("Sale date must be after purchase date."); return; }

  const months = holdingMonths(pd,sd);
  const pyFY   = getFY(pd);
  const selectedTY  = (document.getElementById("taxYear")||{value:"2025-26"}).value||"2025-26";
  const syFY        = getSelectedSaleFY();
  const ayLabel     = TAX_YEAR_AY[selectedTY]||"AY 2026-27";
  const isFutureTY  = selectedTY === "2026-27";

  // Determine LTCG/STCG threshold
  let ltcgMonths = 24;
  if(asset==="equity") ltcgMonths = 12;
  const isLTCG = months >= ltcgMonths;
  const cgType = isLTCG ? "Long-Term Capital Gain (LTCG)" : "Short-Term Capital Gain (STCG)";

  // Net sale consideration
  const netSale = sp - tc;

  // ── New Regime (no indexation) ────────────────────────────────────────────
  let newCOA = pp + ic;
  let newCG  = Math.max(0, netSale - newCOA - exemAmt);
  let newRate, newExempt=0, newTax=0;

  if(!isLTCG){
    // STCG
    if(asset==="equity") newRate=20;
    else newRate=0; // slab
  } else {
    if(asset==="equity"){ newRate=12.5; newExempt=125000; }
    else if(asset==="debt") newRate=0; // slab
    else newRate=12.5;
  }
  if(newRate>0){
    const taxableNew = Math.max(0, newCG - newExempt);
    newTax = Math.round(taxableNew * newRate / 100);
  }

  // ── Old Regime (with indexation) — only for LTCG on property purchased pre Jul 2023 ──
  const preJul23 = pd < new Date(2023,6,23);
  const showOldRegime = isLTCG && (asset==="property"||asset==="other"||asset==="gold") && preJul23;
  let oldCOA=0, oldCG=0, oldRate=0, oldExempt=0, oldTax=0;

  if(showOldRegime){
    const ciiPurchase = getCII(pyFY);
    const ciiSale     = getCII(syFY);
    oldCOA  = Math.round((pp + ic) * ciiSale / ciiPurchase);
    oldCG   = Math.max(0, netSale - oldCOA - exemAmt);
    oldRate = 20;
    const taxableOld = Math.max(0, oldCG - oldExempt);
    oldTax  = Math.round(taxableOld * oldRate / 100);
  }

  // ── Reverse calculator ─────────────────────────────────────────────────────
  // Sale price at which capital gains tax = 0 (new regime)
  // Formula: netSale = COA + exemAmt + exempt_threshold
  //  → sp - tc = COA + exemAmt + exempt_threshold
  //  → sp = COA + tc + exemAmt + exempt_threshold
  let zeroTaxSale = null;
  let zeroTaxNote = "";
  if(newRate > 0){
    // Fixed-rate: solve for sp where taxable gain = 0
    zeroTaxSale = newCOA + tc + exemAmt + newExempt;
    zeroTaxNote = "Sell at or below this price → zero tax (New Regime, " + newRate + "% rate). Capital gain will be ≤ ₹0 / within exempt threshold.";
  } else if(newRate === 0 && isLTCG){
    // Debt/Slab LTCG — zero tax when gain is zero (COA = sale)
    zeroTaxSale = newCOA + tc + exemAmt;
    zeroTaxNote = "Slab-rate LTCG. Zero capital gain (no addition to income) when sold at or below this price.";
  } else if(!isLTCG && newRate === 0){
    // STCG slab rate — zero capital gain
    zeroTaxSale = newCOA + tc + exemAmt;
    zeroTaxNote = "Slab-rate STCG. Zero capital gain (no addition to income) when sold at or below this price.";
  }

  // ── Render results ─────────────────────────────────────────────────────────
  document.getElementById("cgTypeLabel").innerHTML =
    `<span style="background:${isLTCG?"#EFF6FF":"#FFFBEB"};color:${isLTCG?"var(--brand)":"#92400E"};
     padding:4px 12px;border-radius:99px;font-size:12px">${cgType} · ${months} months holding</span>
     <span style="margin-left:8px;background:#F0FDF4;color:#065F46;padding:4px 12px;border-radius:99px;font-size:12px;font-weight:600">
       ${selectedTY} (${ayLabel})${isFutureTY?" · Projected":""}
     </span>`;

  // Compare grid
  let compareHTML = "";
  const winner = (showOldRegime && oldTax < newTax) ? "old" : "new";

  compareHTML += `<div class="cbox ${winner==="new"?"winner":""}">
    ${winner==="new"?'<span class="winner-badge">✓ Lower Tax</span><br>':""}
    <h3>New Regime (No Indexation)</h3>
    <div class="ctax">${newRate===0?"Slab Rate":fmt(newTax)}</div>
    <div class="csub">${newRate===0?"Tax at applicable slab rate":newRate+"% on "+fmt(Math.max(0,newCG-newExempt))}</div>
  </div>`;

  if(showOldRegime){
    compareHTML += `<div class="cbox ${winner==="old"?"winner":""}">
      ${winner==="old"?'<span class="winner-badge">✓ Lower Tax</span><br>':""}
      <h3>Old Regime (With Indexation)</h3>
      <div class="ctax">${fmt(oldTax)}</div>
      <div class="csub">20% on ${fmt(Math.max(0,oldCG))} (Indexed COA: ${fmt(oldCOA)})</div>
    </div>`;
  } else {
    compareHTML += `<div class="cbox" style="background:#F9FAFB;border-style:dashed">
      <h3 style="color:var(--muted)">Old Regime (Indexation)</h3>
      <div class="ctax" style="color:var(--muted);font-size:14px">Not Applicable</div>
      <div class="csub">${!isLTCG?"STCG — no indexation benefit":!preJul23?"Asset purchased after 23 Jul 2023":"Not applicable for this asset type"}</div>
    </div>`;
  }
  document.getElementById("compareGrid").innerHTML = compareHTML;

  // Detail breakdown (new regime)
  let detailHTML = `<div class="card" style="margin-top:12px">
    <div class="card-head"><div class="icon" style="background:#EFF6FF">🧮</div>
    <div><h2>Computation (New Regime)</h2><p>Step-by-step breakdown</p></div></div>
    <div class="card-body" style="padding:12px 16px">
    <table class="dtable">
      <tr><td>Full Value of Consideration (Sale Price)</td><td>${fmt(sp)}</td></tr>
      <tr><td>Less: Transfer Expenses</td><td>(${fmt(tc)})</td></tr>
      <tr><td>Net Sale Consideration</td><td>${fmt(netSale)}</td></tr>
      <tr><td>Less: Cost of Acquisition</td><td>(${fmt(pp)})</td></tr>
      <tr><td>Less: Cost of Improvement</td><td>(${fmt(ic)})</td></tr>
      <tr><td>Capital Gain (Before Exemption)</td><td>${fmt(Math.max(0,netSale-pp-ic))}</td></tr>
      ${exemAmt>0?`<tr><td>Less: Exemption u/s ${exemType}</td><td>(${fmt(exemAmt)})</td></tr>`:""}
      <tr><td>Taxable Capital Gain</td><td>${fmt(newCG)}</td></tr>
      ${newExempt>0?`<tr><td>Less: Basic Exemption (₹1.25L)</td><td>(${fmt(Math.min(newExempt,newCG))})</td></tr>`:""}
      <tr><td>Tax @ ${newRate===0?"Slab":newRate+"%"}</td><td><strong>${newRate===0?"As per slab":fmt(newTax)}</strong></td></tr>
    </table></div></div>`;

  if(showOldRegime){
    const ciiP = getCII(pyFY), ciiS = getCII(syFY);
    detailHTML += `<div class="card" style="margin-top:12px">
      <div class="card-head"><div class="icon" style="background:#F5F3FF">📊</div>
      <div><h2>Computation (Old Regime with Indexation)</h2><p>CII ${pyFY}-${pyFY+1}: ${ciiP} → ${syFY}-${syFY+1}: ${ciiS}</p></div></div>
      <div class="card-body" style="padding:12px 16px">
      <table class="dtable">
        <tr><td>Net Sale Consideration</td><td>${fmt(netSale)}</td></tr>
        <tr><td>Indexed Cost of Acquisition (${pp} × ${ciiS}/${ciiP})</td><td>(${fmt(oldCOA)})</td></tr>
        <tr><td>Capital Gain (After Indexation)</td><td>${fmt(oldCG)}</td></tr>
        ${exemAmt>0?`<tr><td>Less: Exemption u/s ${exemType}</td><td>(${fmt(exemAmt)})</td></tr>`:""}
        <tr><td>Tax @ 20%</td><td><strong>${fmt(oldTax)}</strong></td></tr>
      </table></div></div>`;
  }

  document.getElementById("detailBreakdown").innerHTML = detailHTML;

  // Reverse calculator — always show if zeroTaxSale computed
  if(zeroTaxSale !== null && zeroTaxSale > 0){
    document.getElementById("reverseBox").style.display = "block";
    document.getElementById("revSalePrice").innerHTML =
      `<span style="font-size:22px;font-weight:800;color:var(--brand)">${fmt(zeroTaxSale)}</span>`;
    document.getElementById("revSub").innerHTML =
      `<span style="font-size:12px;color:var(--muted)">${zeroTaxNote}</span>` +
      (showOldRegime && isLTCG ? `<br><span style="font-size:12px;color:var(--muted);margin-top:4px;display:block">` +
        `Old Regime (indexed): Sell at or below <strong>${fmt(oldCOA + tc + exemAmt)}</strong> for zero tax.</span>` : "");
  } else {
    document.getElementById("reverseBox").style.display = "none";
  }

  // Note
  let note = "";
  if(newRate===0) note = "Tax at applicable slab rate. Add capital gain to total income and apply applicable tax slab.";
  else if(!isLTCG) note = "Short-term capital gain — taxed at "+newRate+"% (equity) or slab rate (others).";
  else note = `LTCG taxed at ${newRate}%. ${showOldRegime?"Both regimes shown — choose the one with lower tax.":""}`;
  if(exemAmt>0) note += ` Exemption of ${fmt(exemAmt)} claimed u/s ${exemType}.`;
  document.getElementById("cgNote").textContent = "⚠ " + note + " This is an estimate — verify with actual CII and consult your CA.";

  document.getElementById("cgResults").style.display = "block";
  document.getElementById("cgResults").scrollIntoView({behavior:"smooth"});
}

updateAssetUI();

function autoDateSlash(el) {
  // Auto-insert "/" after DD and MM as user types digits
  let v = el.value.replace(/[^0-9]/g, ""); // strip non-digits
  let out = "";
  if (v.length > 2) out = v.slice(0,2) + "/" + v.slice(2);
  else out = v;
  if (v.length > 4) out = v.slice(0,2) + "/" + v.slice(2,4) + "/" + v.slice(4,8);
  el.value = out;
}
</script>
<a href="https://wa.me/918427651580" target="_blank" class="wa-float" title="WhatsApp Support"><svg viewBox="0 0 24 24"><path d="M17.472 14.382c-.297-.149-1.758-.867-2.03-.967-.273-.099-.471-.148-.67.15-.197.297-.767.966-.94 1.164-.173.199-.347.223-.644.075-.297-.15-1.255-.463-2.39-1.475-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.298-.347.446-.52.149-.174.198-.298.298-.497.099-.198.05-.371-.025-.52-.075-.149-.669-1.612-.916-2.207-.242-.579-.487-.5-.669-.51-.173-.008-.371-.01-.57-.01-.198 0-.52.074-.792.372-.272.297-1.04 1.016-1.04 2.479 0 1.462 1.065 2.875 1.213 3.074.149.198 2.096 3.2 5.077 4.487.709.306 1.262.489 1.694.625.712.227 1.36.195 1.871.118.571-.085 1.758-.719 2.006-1.413.248-.694.248-1.289.173-1.413-.074-.124-.272-.198-.57-.347m-5.421 7.403h-.004a9.87 9.87 0 01-5.031-1.378l-.361-.214-3.741.982.998-3.648-.235-.374a9.86 9.86 0 01-1.51-5.26c.001-5.45 4.436-9.884 9.888-9.884 2.64 0 5.122 1.03 6.988 2.898a9.825 9.825 0 012.893 6.994c-.003 5.45-4.437 9.884-9.885 9.884m8.413-18.297A11.815 11.815 0 0012.05 0C5.495 0 .16 5.335.157 11.892c0 2.096.547 4.142 1.588 5.945L.057 24l6.305-1.654a11.882 11.882 0 005.683 1.448h.005c6.554 0 11.89-5.335 11.893-11.893a11.821 11.821 0 00-3.48-8.413z"/></svg></a>
</body></html>"""


ADMIN_T = """<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Admin – CA Toolkit</title>
<style>
""" + BASE_CSS + """
.wrap{max-width:1100px;margin:0 auto;padding:28px 24px}
h1{font-size:20px;font-weight:800;margin-bottom:4px}
.sub{font-size:13px;color:var(--muted);margin-bottom:24px}
.alert{padding:10px 14px;border-radius:8px;font-size:13px;margin-bottom:16px}
.as{background:#ECFDF5;border:1px solid #A7F3D0;color:#065F46}
.ae{background:#FEF2F2;border:1px solid #FECACA;color:#991B1B}
.section{background:var(--white);border:1px solid var(--border);border-radius:var(--radius);margin-bottom:20px;overflow:hidden}
.sec-head{padding:14px 18px;border-bottom:1px solid var(--border);display:flex;align-items:center;justify-content:space-between}
.sec-head h2{font-size:14px;font-weight:700}
.sec-body{padding:18px}
label{display:block;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.04em;color:var(--muted);margin-bottom:4px}
.field{margin-bottom:12px}
input[type=text],input[type=password],select{width:100%;border:1.5px solid var(--border);border-radius:8px;
  padding:8px 11px;font-family:inherit;font-size:13px;color:var(--ink);background:var(--white);outline:none;transition:border-color .2s}
input:focus,select:focus{border-color:var(--brand)}
.form-row{display:grid;grid-template-columns:1fr 1fr 1fr auto;gap:10px;align-items:end}
@media(max-width:640px){.form-row{grid-template-columns:1fr}}
.btn{background:var(--brand);color:#fff;border:none;border-radius:8px;padding:9px 16px;
     font-family:inherit;font-size:13px;font-weight:600;cursor:pointer;transition:background .2s}
.btn:hover{background:var(--brand-d)}
.sm{padding:5px 10px;font-size:11px;border-radius:6px;border:none;cursor:pointer;font-family:inherit;font-weight:600}
.sg{background:#ECFDF5;color:#065F46}.sg:hover{background:#A7F3D0}
.rr{background:#FEF2F2;color:#991B1B}.rr:hover{background:#FECACA}
.am{background:#EFF6FF;color:var(--brand)}.am:hover{background:#BFDBFE}
table{width:100%;border-collapse:collapse;font-size:12px}
th{text-align:left;font-size:10px;letter-spacing:.08em;text-transform:uppercase;color:var(--muted);
   border-bottom:1.5px solid var(--border);padding:7px 10px}
td{padding:10px;border-bottom:1px solid var(--border);vertical-align:middle}
tr:last-child td{border-bottom:none}
tr:hover td{background:#F9FAFB}
.bdg{display:inline-block;font-size:10px;font-weight:700;padding:2px 7px;border-radius:99px;text-transform:uppercase}
.b-free{background:#F3F4F6;color:var(--muted)}
.b-starter{background:#ECFDF5;color:#065F46}
.b-standard{background:#EFF6FF;color:var(--brand)}
.b-pro{background:#FFFBEB;color:#92400E}
.b-firm{background:#F5F3FF;color:#5B21B6}
.warn{color:var(--red);font-weight:700}
.ok{color:var(--green);font-weight:700}
.progress-wrap{display:flex;align-items:center;gap:8px;min-width:120px}
.progress-bg{flex:1;background:#F3F4F6;border-radius:99px;height:5px;overflow:hidden}
.progress-fill{height:100%;border-radius:99px}
</style></head><body>
<nav>
  <a href="/" class="logo">CA<span class="logo-dot"></span><span>Toolkit</span></a>
  <div class="nav-right">
    <a href="/" class="nav-btn dash">⬅ Dashboard</a>
    <a href="/logout" class="nav-link">Sign out</a>
  </div>
</nav>
<div class="wrap">
  <h1>⚙ Admin Panel</h1>
  <p class="sub">Create accounts, manage plans, track usage.</p>
  {% if msg %}<div class="alert {{ 'ae' if 'error' in msg.lower() or 'cannot' in msg.lower() else 'as' }}">{{ msg }}</div>{% endif %}

  <!-- CREATE USER -->
  <div class="section">
    <div class="sec-head"><h2>➕ Create New User</h2></div>
    <div class="sec-body">
      <form method="POST" action="/admin/create">
        <div class="form-row">
          <div class="field"><label>Username</label>
            <input type="text" name="username" placeholder="e.g. rahul_ca" required/></div>
          <div class="field"><label>Password</label>
            <input type="password" name="password" placeholder="Min 6 chars" required/></div>
          <div class="field"><label>Plan</label>
            <select name="plan">
              <option value="free">Free (2 uploads)</option>
              <option value="starter">Starter (10 uploads · ₹60)</option>
              <option value="standard" selected>Standard (25 uploads · ₹130)</option>
              <option value="pro">Professional (60 uploads · ₹270)</option>
              <option value="firm">Firm (150 uploads · ₹600)</option>
              <option value="ca">CA Firm (500 uploads · ₹1000)</option>
            </select></div>
          <div class="field"><label>&nbsp;</label>
            <button class="btn" type="submit">Create User</button></div>
        </div>
      </form>
    </div>
  </div>

  <!-- USERS TABLE -->
  <div class="section">
    <div class="sec-head">
      <h2>👥 All Users ({{ users|length }})</h2>
      <span style="font-size:12px;color:var(--muted)">Uploads remaining shown in green/red</span>
    </div>
    <div class="sec-body" style="padding:0;overflow-x:auto">
      <table>
        <thead><tr>
          <th>#</th><th>Username</th><th>Plan</th>
          <th>Uploads Used</th><th>Remaining</th>
          <th>Valid Till</th><th>Joined</th><th>Actions</th>
        </tr></thead>
        <tbody>
        {% for u in users %}
        <tr>
          <td style="color:var(--muted)">{{ u.id }}</td>
          <td><strong>{{ u.username }}</strong>
            {% if u.is_admin %}<span class="bdg" style="background:#EFF6FF;color:var(--brand);margin-left:4px">Admin</span>{% endif %}
          </td>
          <td><span class="bdg b-{{ u.plan }}">{{ u.plan }}</span></td>
          <td>
            <div class="progress-wrap">
              <div class="progress-bg">
                <div class="progress-fill"
                     style="width:{{ [u.uploads_used*100//u.uploads_total if u.uploads_total else 0, 100]|min }}%;
                            background:{{ '#EF4444' if u.remaining==0 else '#F59E0B' if u.remaining<=3 else '#10B981' }}">
                </div>
              </div>
              <span style="font-size:11px;white-space:nowrap">{{ u.uploads_used }} / {{ u.uploads_total }}</span>
            </div>
          </td>
          <td class="{{ 'warn' if u.remaining==0 else 'ok' if u.remaining > 5 else '' }}">
            {{ u.remaining }}
          </td>
          <td style="font-size:11px;color:var(--muted)">
            {{ u.validity_end[:10] if u.validity_end else '—' }}
          </td>
          <td style="font-size:11px;color:var(--muted)">{{ u.created_at[:10] }}</td>
          <td>
            {% if not u.is_admin %}
            <div style="display:flex;gap:6px;flex-wrap:wrap">
              <!-- Add uploads dropdown -->
              <form method="POST" action="/admin/addplan" style="display:flex;gap:4px;align-items:center">
                <input type="hidden" name="uid" value="{{ u.id }}"/>
                <select name="plan" style="padding:4px 6px;font-size:11px;border-radius:6px;border:1px solid var(--border);width:auto">
                  <option value="starter">+10</option>
                  <option value="standard" selected>+25</option>
                  <option value="pro">+60</option>
                  <option value="firm">+150</option>
                  <option value="ca">+500</option>
                </select>
                <button class="sm am" type="submit">Add</button>
              </form>
              <form method="POST" action="/admin/delete" style="display:inline"
                    onsubmit="return confirm('Delete {{ u.username }}? This cannot be undone.')">
                <input type="hidden" name="uid" value="{{ u.id }}"/>
                <button class="sm rr" type="submit">Delete</button>
              </form>
            </div>
            {% else %}
            <span style="font-size:11px;color:var(--muted)">—</span>
            {% endif %}
          </td>
        </tr>
        {% endfor %}
        </tbody>
      </table>
    </div>
  </div>
</div>
<footer>
  <div class="ft-bottom" style="justify-content:center">
    <span class="ft-bottom-left">©2026 CA Toolkit · Admin Panel · All Rights Reserved</span>
  </div>
</footer>
<a href="https://wa.me/918427651580" target="_blank" class="wa-float" title="WhatsApp Support"><svg viewBox="0 0 24 24"><path d="M17.472 14.382c-.297-.149-1.758-.867-2.03-.967-.273-.099-.471-.148-.67.15-.197.297-.767.966-.94 1.164-.173.199-.347.223-.644.075-.297-.15-1.255-.463-2.39-1.475-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.298-.347.446-.52.149-.174.198-.298.298-.497.099-.198.05-.371-.025-.52-.075-.149-.669-1.612-.916-2.207-.242-.579-.487-.5-.669-.51-.173-.008-.371-.01-.57-.01-.198 0-.52.074-.792.372-.272.297-1.04 1.016-1.04 2.479 0 1.462 1.065 2.875 1.213 3.074.149.198 2.096 3.2 5.077 4.487.709.306 1.262.489 1.694.625.712.227 1.36.195 1.871.118.571-.085 1.758-.719 2.006-1.413.248-.694.248-1.289.173-1.413-.074-.124-.272-.198-.57-.347m-5.421 7.403h-.004a9.87 9.87 0 01-5.031-1.378l-.361-.214-3.741.982.998-3.648-.235-.374a9.86 9.86 0 01-1.51-5.26c.001-5.45 4.436-9.884 9.888-9.884 2.64 0 5.122 1.03 6.988 2.898a9.825 9.825 0 012.893 6.994c-.003 5.45-4.437 9.884-9.885 9.884m8.413-18.297A11.815 11.815 0 0012.05 0C5.495 0 .16 5.335.157 11.892c0 2.096.547 4.142 1.588 5.945L.057 24l6.305-1.654a11.882 11.882 0 005.683 1.448h.005c6.554 0 11.89-5.335 11.893-11.893a11.821 11.821 0 00-3.48-8.413z"/></svg></a>
</body></html>"""

# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — AUTH
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/login", methods=["GET"])
def login_page():
    if "uid" in session: return redirect(url_for("dashboard"))
    return render_template_string(LOGIN_T, error=None, email=CONTACT_EMAIL)

@app.route("/login", methods=["POST"])
def login_post():
    u = request.form.get("username", "").strip()
    p = request.form.get("password", "")
    user = get_user_by_name(u)
    if not user or user["password"] != _hash(p):
        return render_template_string(LOGIN_T, error="Invalid username or password.", email=CONTACT_EMAIL)
    session.clear()
    session["uid"] = user["id"]
    return redirect(url_for("dashboard"))

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login_page"))

# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — DASHBOARD & TOOLS
# ══════════════════════════════════════════════════════════════════════════════

def user_ctx(user):
    """Build common template context for a user."""
    used  = user["uploads_used"]
    total = user["uploads_total"]
    left  = uploads_remaining(user)
    pct   = min(int(used * 100 / total) if total else 0, 100)
    return dict(
        username=user["username"],
        plan=user["plan"],
        plan_label=PLANS.get(user["plan"], {}).get("label", user["plan"].title()),
        is_admin=bool(user["is_admin"]),
        uploads_used=used,
        uploads_total=total,
        uploads_left=left,
        uploads_remaining=left,
        bar_pct=pct,
        validity_end=user["validity_end"],
        contact_email=CONTACT_EMAIL,
        contact_upi=CONTACT_UPI,
    )


# ══════════════════════════════════════════════════════════════════════════════
#  CA TOOLS HUB PAGE  — /ca-tools-hub
# ══════════════════════════════════════════════════════════════════════════════

CA_TOOLS_HUB_TEMPLATE = """<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>CA Tools Hub – CA Toolkit</title>
<style>""" + BASE_CSS + """
.hub-grid{display:grid;grid-template-columns:1fr;gap:24px;margin-top:8px}
.hub-card{background:var(--card);border:1.5px solid var(--border);border-radius:var(--radius);
  padding:28px 26px;position:relative;overflow:hidden;
  transition:border-color .22s,box-shadow .25s,transform .22s}
.hub-card:hover{border-color:var(--brand);
  box-shadow:0 0 0 3px rgba(20,184,166,.09),0 18px 52px rgba(15,118,110,.18);
  transform:translateY(-4px)}
.hub-card::before{content:'';position:absolute;top:0;left:0;right:0;height:4px;
  background:var(--hc-grad,linear-gradient(90deg,var(--brand),var(--accent)))}
.hub-card-head{display:flex;align-items:flex-start;gap:16px;margin-bottom:14px}
.hub-icon{width:54px;height:54px;border-radius:14px;display:flex;align-items:center;
  justify-content:center;font-size:26px;flex-shrink:0;box-shadow:0 2px 10px rgba(0,0,0,.07)}
.hub-tag{display:inline-flex;align-items:center;font-size:10px;font-weight:700;
  padding:3px 10px;border-radius:99px;margin-bottom:7px;letter-spacing:.04em;
  background:var(--brand-l);color:var(--brand-d);border:1px solid var(--brand-m)}
.hub-name{font-family:var(--font-head);font-size:19px;font-weight:800;
  color:var(--ink);letter-spacing:-.3px;margin-bottom:4px;line-height:1.2}
.hub-tagline{font-size:13px;color:var(--muted);line-height:1.6}
.hub-desc{font-size:13.5px;color:#374151;line-height:1.85;margin-bottom:18px}
.hub-features{display:flex;flex-wrap:wrap;gap:7px;margin-bottom:20px}
.hub-feat{display:inline-flex;align-items:center;gap:5px;font-size:12px;font-weight:500;
  color:var(--brand-d);background:var(--brand-l);border:1px solid var(--brand-m);
  padding:4px 11px;border-radius:99px}
.hub-cta{display:inline-flex;align-items:center;gap:8px;
  background:linear-gradient(135deg,var(--brand),var(--brand-d));color:#fff;
  padding:10px 24px;border-radius:99px;font-size:13px;font-weight:700;
  text-decoration:none;transition:transform .18s,box-shadow .18s,opacity .18s;
  box-shadow:0 4px 16px rgba(15,118,110,.30)}
.hub-cta:hover{transform:translateY(-2px);box-shadow:0 7px 24px rgba(15,118,110,.40);opacity:.95}
.hub-cta:active{transform:scale(0.97)}
.hub-cta-row{display:flex;align-items:center;gap:12px;flex-wrap:wrap}
.hub-note{font-size:11.5px;color:var(--muted);font-style:italic}
</style></head><body>
""" + _PAGE_NAV + """
<div class="page-wrap anim-up">
  <div class="page-hero">
    <div class="page-eyebrow">🧰 CA Tools Hub</div>
    <h1 class="page-title">Companion tools for <em>CA students &amp; firms</em></h1>
    <p class="page-sub">Standalone products built alongside CA Toolkit — each solving a specific problem faced by CA students and practitioners.</p>
  </div>
  <div class="page-divider"></div>

  <div class="hub-grid">

    <!-- Habit Tracker -->
    <div class="hub-card anim-up anim-d1" style="--hc-grad:linear-gradient(90deg,#6D28D9,#8B5CF6,#F59E0B)">
      <div class="hub-card-head">
        <div class="hub-icon" style="background:linear-gradient(135deg,#EDE9FE,#DDD6FE)">📅</div>
        <div>
          <div class="hub-tag" style="background:#EDE9FE;color:#5B21B6;border-color:#C4B5FD">Free · For CA Students</div>
          <div class="hub-name">Habit Tracker for CA Students</div>
          <div class="hub-tagline">Build consistent study habits and track your daily CA exam preparation.</div>
        </div>
      </div>
      <p class="hub-desc">
        CA preparation demands months of consistent, disciplined study. This Habit Tracker is built specifically for CA students — log daily study hours, track subject-wise progress, and build streaks that keep you accountable through Foundation, Intermediate, and Final.
        <br/><br/>
        Log habits like revision, mock tests, ICAI material reading, and practice questions. See weekly completion rates at a glance and stay on track toward your exam date. No distractions, no social feed — just you and your goals.
      </p>
      <div class="hub-features">
        <span class="hub-feat">✓ Daily habit logging</span>
        <span class="hub-feat" style="background:#EDE9FE;color:#5B21B6;border-color:#C4B5FD">✓ Streak tracking</span>
        <span class="hub-feat">✓ Subject-wise progress</span>
        <span class="hub-feat" style="background:#EDE9FE;color:#5B21B6;border-color:#C4B5FD">✓ Free to use</span>
        <span class="hub-feat">✓ Works on mobile</span>
      </div>
      <div class="hub-cta-row">
        <a href="https://habit-tracker-9hcr.onrender.com" target="_blank" rel="noopener"
           class="hub-cta" style="background:linear-gradient(135deg,#6D28D9,#7C3AED)">
          Open Habit Tracker ↗
        </a>
        <span class="hub-note">Free · Opens in new tab · No login needed</span>
      </div>
    </div>

  </div>

  <div class="page-divider"></div>
  <div class="page-section">
    <h2>Request a tool</h2>
    <p>Every tool here comes from a real problem faced by CA students or practitioners. If you have a repetitive task that could be automated, or a tool that would genuinely help your CA preparation, reach out — the best ideas come from people in the field.</p>
    <ul>
      <li>WhatsApp: <a href="https://wa.me/918427651580">+91 84276 51580</a></li>
      <li>Email: <a href="mailto:sumitverma2880@gmail.com">sumitverma2880@gmail.com</a></li>
    </ul>
  </div>
</div>
""" + _PAGE_FOOTER + """
</body></html>"""


# ══════════════════════════════════════════════════════════════════════════════
#  PRICING PAGE  — /pricing
# ══════════════════════════════════════════════════════════════════════════════

PRICING_TEMPLATE = """<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Pricing – CA Toolkit</title>
<style>""" + BASE_CSS + """
.plans-wrap{max-width:960px;margin:0 auto}
.plans-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:18px;margin-bottom:28px}
@media(max-width:700px){.plans-grid{grid-template-columns:repeat(2,1fr)}}
@media(max-width:440px){.plans-grid{grid-template-columns:1fr}}
.plan-card{background:var(--card);border:1.5px solid var(--border);border-radius:var(--radius);
  padding:22px 18px;position:relative;transition:border-color .22s,box-shadow .22s,transform .2s}
.plan-card:hover{border-color:var(--brand);box-shadow:var(--shadow-md);transform:translateY(-4px)}
.plan-card.popular{border-color:var(--brand);box-shadow:var(--shadow-md)}
.plan-card.popular::before{content:'Most Popular';position:absolute;top:-11px;left:50%;
  transform:translateX(-50%);background:linear-gradient(135deg,var(--brand),var(--brand-d));
  color:#fff;font-size:10px;font-weight:700;padding:3px 12px;border-radius:99px;white-space:nowrap;
  letter-spacing:.04em}
.plan-name{font-size:10.5px;font-weight:700;color:var(--muted);text-transform:uppercase;
  letter-spacing:.06em;margin-bottom:8px}
.plan-price{font-family:var(--font-mono,monospace);font-size:28px;font-weight:700;
  color:var(--ink);margin-bottom:2px;letter-spacing:-.02em}
.plan-uploads{font-family:var(--font-mono,monospace);font-size:13px;font-weight:600;
  color:var(--brand-d);margin-bottom:2px}
.plan-validity{font-size:11px;color:var(--muted);margin-bottom:14px}
.plan-features{list-style:none;padding:0;margin-bottom:18px}
.plan-features li{font-size:12px;color:var(--ink2);padding:4px 0;display:flex;gap:6px;
  border-bottom:1px solid var(--border)}
.plan-features li:last-child{border:none}
.plan-features li::before{content:"✓";color:var(--green);font-weight:700;flex-shrink:0}
.plan-cta{display:block;text-align:center;padding:9px;border-radius:9px;
  font-size:12px;font-weight:700;text-decoration:none;transition:all .2s;
  border:1.5px solid var(--brand);color:var(--brand);background:transparent}
.plan-cta:hover{background:var(--brand);color:#fff}
.plan-card.popular .plan-cta{background:var(--brand);color:#fff}
.plan-card.popular .plan-cta:hover{background:var(--brand-d)}
/* Payment box */
.pay-box{background:var(--white);border:1.5px solid var(--border);border-radius:var(--radius);
  padding:28px;max-width:560px;margin:0 auto}
.pay-grid{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-top:20px}
@media(max-width:480px){.pay-grid{grid-template-columns:1fr}}
.pay-item{background:var(--bg);border:1.5px solid var(--border);border-radius:12px;
  padding:16px 18px}
.pay-item .pi-label{font-size:10px;font-weight:700;text-transform:uppercase;
  letter-spacing:.06em;color:var(--muted);margin-bottom:6px}
.pay-item .pi-value{font-size:14px;font-weight:600;color:var(--ink)}
.pay-item .pi-value a{color:var(--brand-d);text-decoration:none}
.pay-item.upi-item{border-color:var(--brand-m);background:var(--brand-l)}
.pay-item.upi-item .pi-label{color:var(--brand-d)}
.pay-item.upi-item .pi-value{font-family:var(--font-mono,monospace);font-size:13px;
  letter-spacing:-.01em;color:var(--brand-dk)}
.refund-note{font-size:11.5px;color:var(--muted);text-align:center;margin-top:16px;line-height:1.7}
.refund-note strong{color:var(--red)}
</style></head><body>
""" + _PAGE_NAV + """
<div class="page-wrap anim-up">
  <div class="page-hero">
    <div class="page-eyebrow">💳 Pricing</div>
    <h1 class="page-title">Simple, <em>upload-based</em> pricing</h1>
    <p class="page-sub">No subscriptions. Buy uploads when you need them — they stack if you recharge before expiry.</p>
  </div>
  <div class="page-divider"></div>

  <div class="plans-wrap">
    <div class="plans-grid">
      <div class="plan-card anim-up anim-d1">
        <div class="plan-name">Free</div>
        <div class="plan-price">₹0</div>
        <div class="plan-uploads">2 uploads</div>
        <div class="plan-validity">Try it out</div>
        <ul class="plan-features"><li>All features</li><li>All sheet types</li><li>Up to 20 MB</li></ul>
        <a href="/" class="plan-cta">Get Started</a>
      </div>
      <div class="plan-card anim-up anim-d2">
        <div class="plan-name">Starter</div>
        <div class="plan-price">₹60</div>
        <div class="plan-uploads">10 uploads</div>
        <div class="plan-validity">3 month validity</div>
        <ul class="plan-features"><li>All features</li><li>All sheet types</li><li>Up to 20 MB</li></ul>
        <a href="#pay" class="plan-cta">Buy via UPI ↓</a>
      </div>
      <div class="plan-card popular anim-up anim-d3">
        <div class="plan-name">Standard</div>
        <div class="plan-price">₹130</div>
        <div class="plan-uploads">25 uploads</div>
        <div class="plan-validity">3 month validity</div>
        <ul class="plan-features"><li>All features</li><li>Priority support</li><li>Up to 20 MB</li></ul>
        <a href="#pay" class="plan-cta">Buy via UPI ↓</a>
      </div>
      <div class="plan-card anim-up anim-d4">
        <div class="plan-name">Professional</div>
        <div class="plan-price">₹270</div>
        <div class="plan-uploads">60 uploads</div>
        <div class="plan-validity">3 month validity</div>
        <ul class="plan-features"><li>All features</li><li>Priority support</li><li>Up to 20 MB</li></ul>
        <a href="#pay" class="plan-cta">Buy via UPI ↓</a>
      </div>
      <div class="plan-card anim-up anim-d5">
        <div class="plan-name">Firm</div>
        <div class="plan-price">₹600</div>
        <div class="plan-uploads">150 uploads</div>
        <div class="plan-validity">3 month validity</div>
        <ul class="plan-features"><li>All features</li><li>WhatsApp support</li><li>Up to 20 MB</li></ul>
        <a href="#pay" class="plan-cta">Buy via UPI ↓</a>
      </div>
      <div class="plan-card anim-up anim-d6">
        <div class="plan-name">CA Firm</div>
        <div class="plan-price">₹1,000</div>
        <div class="plan-uploads">500 uploads</div>
        <div class="plan-validity">3 month validity</div>
        <ul class="plan-features"><li>All features + GST Recon</li><li>WhatsApp support</li><li>Best for CA firms</li></ul>
        <a href="#pay" class="plan-cta">Buy via UPI ↓</a>
      </div>
    </div>

    <div class="page-divider"></div>

    <!-- Payment section — UPI only, email only in contact -->
    <div class="pay-box" id="pay">
      <div style="text-align:center;margin-bottom:6px">
        <div class="page-eyebrow" style="justify-content:center">How to Purchase</div>
        <p style="font-size:13.5px;color:var(--muted);line-height:1.8;margin-top:8px">
          Pay via UPI below and send your payment screenshot on WhatsApp or email.
          Your account will be upgraded within a few hours.
        </p>
      </div>
      <div class="pay-grid">
        <div class="pay-item upi-item">
          <div class="pi-label">💳 UPI Payment</div>
          <div class="pi-value">{{ contact_upi }}</div>
        </div>
        <div class="pay-item">
          <div class="pi-label">💬 WhatsApp</div>
          <div class="pi-value"><a href="https://wa.me/918427651580">+91 84276 51580</a></div>
        </div>
      </div>
      <p class="refund-note">
        <strong>No refund</strong> after first upload of a paid plan is used ·
        Unused uploads stack when you recharge before expiry
      </p>
    </div>
  </div>
</div>
""" + _PAGE_FOOTER + """
</body></html>"""

@app.route("/")
def dashboard():
    if "uid" in session:
        user = get_user_by_id(session["uid"])
        ctx = user_ctx(user)
    else:
        ctx = dict(
            username=None, plan="free", plan_label="Free",
            is_admin=False, uploads_used=0, uploads_total=2,
            uploads_left=2, uploads_remaining=2, bar_pct=0,
            validity_end=None, contact_email=CONTACT_EMAIL,
            contact_upi=CONTACT_UPI,
        )
    return render_template_string(DASHBOARD_T, **ctx)


@app.route("/bs-shift")
def bs_shift_redirect():
    return redirect("/")

@app.route("/privacy")
def privacy_page():
    if "uid" in session:
        user = get_user_by_id(session["uid"])
        ctx = user_ctx(user)
    else:
        ctx = dict(username=None, plan="free", plan_label="Free",
            is_admin=False, uploads_used=0, uploads_total=2,
            uploads_left=2, uploads_remaining=2, bar_pct=0,
            validity_end=None, contact_email=CONTACT_EMAIL, contact_upi=CONTACT_UPI)
    return render_template_string(PRIVACY_TEMPLATE, **ctx)

@app.route("/story")
def story_page():
    return render_template_string(STORY_TEMPLATE)

@app.route("/how-to-use")
def how_to_use_page():
    return render_template_string(HOW_TO_USE_TEMPLATE)

@app.route("/ca-tools-hub")
def ca_tools_hub_page():
    return render_template_string(CA_TOOLS_HUB_TEMPLATE)

@app.route("/pricing")
def pricing_page():
    if "uid" in session:
        user = get_user_by_id(session["uid"])
        ctx = user_ctx(user)
    else:
        ctx = dict(username=None, plan="free", plan_label="Free",
            is_admin=False, uploads_used=0, uploads_total=2,
            uploads_left=2, uploads_remaining=2, bar_pct=0,
            validity_end=None, contact_email=CONTACT_EMAIL, contact_upi=CONTACT_UPI)
    return render_template_string(PRICING_TEMPLATE, **ctx)

@app.route("/tool/converter")
@login_required
def tool_converter():
    user = get_user_by_id(session["uid"])
    return render_template_string(CONVERTER_T, **user_ctx(user))

# ── T-Shape BS Converter tool page ────────────────────────────────────────────
TSHAPE_T = """<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>T-Shape BS Converter &ndash; CA Toolkit</title>
<style>
""" + BASE_CSS + """
.nav-links{display:flex;gap:20px;list-style:none}
.nav-links a{text-decoration:none;color:var(--muted);font-size:13px;font-weight:500;transition:color .2s}
.nav-links a:hover{color:var(--brand)}
.hero{text-align:center;padding:56px 24px 40px;max-width:700px;margin:0 auto}
.hero-badge{display:inline-flex;align-items:center;gap:6px;background:var(--brand-l);
            color:var(--brand-d);border:1px solid var(--brand-m);border-radius:99px;
            padding:5px 14px;font-size:12px;font-weight:600;margin-bottom:18px}
h1{font-size:clamp(24px,4vw,40px);font-weight:800;line-height:1.15;letter-spacing:-.5px;margin-bottom:14px}
h1 em{font-style:normal;color:var(--accent)}
.hero p{font-size:15px;color:var(--muted);line-height:1.7;max-width:520px;margin:0 auto 28px}
.stats{display:flex;justify-content:center;gap:36px;flex-wrap:wrap;padding:16px 24px;
       background:var(--white);border-top:1px solid var(--border);border-bottom:1px solid var(--border)}
.stat-n{font-size:20px;font-weight:800;color:var(--brand-d)}
.stat-l{font-size:11px;color:var(--muted);margin-top:2px}
.main{max-width:1160px;margin:0 auto;padding:40px 24px;
      display:grid;grid-template-columns:1fr 1fr;gap:24px;align-items:start}
@media(max-width:768px){.main{grid-template-columns:1fr}}
.card{background:var(--white);border-radius:var(--radius);border:1px solid var(--border);
      box-shadow:var(--shadow);overflow:hidden}
.card-head{padding:16px 20px;border-bottom:1px solid var(--border);
           display:flex;align-items:center;gap:10px}
.card-head .icon{width:32px;height:32px;border-radius:8px;display:flex;
                 align-items:center;justify-content:center;font-size:16px}
.card-head h2{font-size:14px;font-weight:700}
.card-head p{font-size:12px;color:var(--muted);margin-top:1px}
.card-body{padding:20px}
.usage-row{display:flex;justify-content:space-between;align-items:center;
           font-size:12px;font-weight:600;margin-bottom:5px}
.usage-bar-bg{background:#F3F4F6;border-radius:99px;height:6px;overflow:hidden;margin-bottom:14px}
.usage-bar-fill{height:100%;border-radius:99px;background:linear-gradient(90deg,var(--brand),var(--brand-d));transition:width .4s}
.field{margin-bottom:16px}
label{display:block;font-size:11px;font-weight:600;text-transform:uppercase;
      letter-spacing:.04em;color:var(--muted);margin-bottom:5px}
.hint{font-size:11px;color:var(--muted);margin-top:4px}
.dropzone{border:2px dashed var(--border);border-radius:10px;padding:24px 14px;
          text-align:center;cursor:pointer;transition:all .2s;position:relative;background:var(--bg)}
.dropzone:hover,.dropzone.drag{border-color:var(--brand);background:var(--brand-l)}
.dropzone input{position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%;pointer-events:none}
.dz-icon{font-size:26px;margin-bottom:6px}
.dz-text{font-size:12px;color:var(--muted)}
.dz-text strong{color:var(--brand-d)}
.dz-file{font-size:12px;font-weight:600;color:var(--green);margin-top:5px;display:none}
.row2{display:grid;grid-template-columns:1fr 1fr;gap:12px}
input[type=number],input[type=text]{width:100%;border:1.5px solid var(--border);
  border-radius:8px;padding:9px 12px;font-family:inherit;font-size:13px;
  color:var(--ink);background:var(--white);transition:border-color .2s;outline:none}
input:focus{border-color:var(--brand)}
.btn{width:100%;background:linear-gradient(135deg,var(--brand),var(--brand-d));color:#fff;border:none;
     border-radius:10px;padding:12px;font-family:inherit;font-size:14px;font-weight:700;cursor:pointer;
     transition:opacity .2s;display:flex;align-items:center;justify-content:center;gap:8px}
.btn:hover{opacity:.88}
.btn:disabled{opacity:.5;cursor:not-allowed}
.spinner{width:16px;height:16px;border:2px solid rgba(255,255,255,.3);
         border-top-color:#fff;border-radius:50%;animation:spin .7s linear infinite;display:none}
@keyframes spin{to{transform:rotate(360deg)}}
#status{margin-top:12px;border-radius:8px;padding:12px 14px;font-size:13px;display:none;line-height:1.6}
#status.success{background:#ECFDF5;border:1px solid #A7F3D0;color:#065F46}
#status.error{background:#FEF2F2;border:1px solid #FECACA;color:#991B1B}
.log-list{margin-top:6px;padding-left:14px;font-size:11px;color:#374151;line-height:2}
.dl-btn{display:none;margin-top:10px;width:100%;background:var(--green);color:#fff;
        border:none;border-radius:10px;padding:11px;font-family:inherit;font-size:13px;
        font-weight:600;cursor:pointer;text-decoration:none;text-align:center;transition:background .2s}
.dl-btn:hover{background:#059669}
.steps{padding:0;list-style:none;counter-reset:step}
.steps li{display:flex;gap:10px;align-items:flex-start;padding:12px 0;border-bottom:1px solid var(--border)}
.steps li:last-child{border:none}
.steps li::before{counter-increment:step;content:counter(step);min-width:24px;height:24px;
                  background:linear-gradient(135deg,var(--brand),var(--brand-d));color:#fff;border-radius:50%;
                  display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:700;margin-top:1px}
.steps li strong{display:block;font-size:12px;font-weight:600;margin-bottom:2px}
.steps li span{font-size:11px;color:var(--muted)}
.limit-banner{max-width:640px;margin:0 auto;padding:0 24px}
.limit-box{background:#FEF2F2;border:1px solid #FECACA;border-radius:var(--radius);
           padding:20px 24px;text-align:center;margin-top:16px}
.limit-box h3{font-size:15px;font-weight:700;color:#991B1B;margin-bottom:8px}
.limit-box p{font-size:13px;color:#7F1D1D;line-height:1.7;margin-bottom:10px}
.limit-box a{color:var(--brand);font-weight:600;text-decoration:none}
.info-box{background:var(--brand-l);border:1px solid var(--brand-m);border-radius:10px;
          padding:12px 16px;font-size:12px;color:var(--brand-dk);line-height:1.7;margin-bottom:16px}
.info-box strong{color:var(--brand-dk)}
@media(max-width:480px){.row2{grid-template-columns:1fr}}
</style></head><body>

<nav>
  <a href="/" class="logo">CA<span class="logo-dot"></span><span>Toolkit</span></a>
  <ul class="nav-links">
    <li><a href="/#tool">Tools</a></li>
    <li><a href="/how-to-use">How to Use</a></li>
  </ul>
  <div class="nav-right">
    {% if username %}
    <div class="nav-user">
      <span class="nav-avatar">{{ username[0].upper() }}</span>
      <strong>{{ username }}</strong>
      <span class="badge b-{{ plan }}">{{ plan_label }}</span>
    </div>
    <div class="nav-sep"></div>
    {% if is_admin %}<a href="/admin" class="nav-btn ghost">&#9881; Admin</a>{% endif %}
    {% endif %}
    <a href="/" class="nav-btn dash">&#8678; Dashboard</a>
    {% if username %}<a href="/logout" class="nav-link">Sign out</a>
    {% else %}<a href="/login" class="nav-btn">Sign In &rarr;</a>{% endif %}
  </div>
</nav>

<section class="hero">
  <div class="hero-badge">&#128209; T-Shape &rarr; Comparative BS</div>
  <h1>T-Shaped Balance Sheet &rarr;<br/><em>Comparative Format</em> Automatically</h1>
  <p>Upload a T-shaped balance sheet (.xls) &mdash; the tool extracts every annexure and fills the PY column of the output template. CY column left blank (yellow) for you to fill.</p>
</section>

{% if uploads_left == 0 %}
<div class="limit-banner">
  <div class="limit-box">
    <h3>&#128274; No uploads remaining</h3>
    <p>You've used all your uploads. Contact us to recharge your account.</p>
    <p>&#128231; <a href="mailto:{{ contact_email }}">{{ contact_email }}</a> &nbsp;|&nbsp;
       &#128179; UPI: <strong>{{ contact_upi }}</strong></p>
  </div>
</div>
{% endif %}

<div class="stats">
  <div class="stat"><div class="stat-n">PY Auto</div><div class="stat-l">Filled from XLS</div></div>
  <div class="stat"><div class="stat-n">All annexures</div><div class="stat-l">Capital, debtors, creditors &amp; more</div></div>
  <div class="stat"><div class="stat-n">&lt;15 sec</div><div class="stat-l">Processing time</div></div>
  <div class="stat"><div class="stat-n">Year auto-read</div><div class="stat-l">No manual input needed</div></div>
</div>

<div class="main" id="tool">
  <div class="card">
    <div class="card-head">
      <div class="icon" style="background:var(--brand-l)">&#128209;</div>
      <div>
        <h2>Convert T-Shaped Balance Sheet</h2>
        <p>{{ plan_label }} &middot; {{ uploads_left }} upload{{ 's' if uploads_left != 1 else '' }} remaining</p>
      </div>
    </div>
    <div class="card-body">
      <div class="usage-row">
        <span style="color:var(--muted)">Uploads used</span>
        <span><strong>{{ uploads_used }}</strong> / {{ uploads_total }}
          {% if validity_end %}<span style="color:#9CA3AF;font-weight:400"> &middot; expires {{ validity_end[:10] }}</span>{% endif %}
        </span>
      </div>
      <div class="usage-bar-bg"><div class="usage-bar-fill" style="width:{{ bar_pct }}%"></div></div>
      <div class="info-box">
        <strong>&#9432; .xls format only.</strong> Upload the T-shaped XLS file. The tool auto-detects the format, reads the financial year from the file, and fills the PY column automatically.
      </div>
      <div class="field">
        <label>Upload T-Shaped XLS</label>
        <div class="dropzone" id="dropzone">
          <div class="dz-icon">&#128196;</div>
          <div class="dz-text">Drag &amp; drop or <strong>click to browse</strong></div>
          <div class="dz-text" style="margin-top:3px">.xls files only</div>
          <div class="dz-file" id="dzFile"></div>
          <input type="file" id="xlFile" accept=".xls">
        </div>
      </div>
      <div class="field">
        <label>Client Name (optional)</label>
        <input type="text" id="outputName" placeholder="e.g. M/S Ashok Kumar Gupta &amp; Co.">
        <div class="hint">Used as client name in output. Leave blank to auto-detect from file.</div>
      </div>
      <button class="btn" id="processBtn" onclick="processFile()">
        <div class="spinner" id="spinner"></div>
        <span id="btnText">&#9889; Convert &amp; Download</span>
      </button>
      <div id="status"></div>
      <a class="dl-btn" id="dlBtn" download>&#11015; Download Output</a>
    </div>
  </div>

  <div style="display:flex;flex-direction:column;gap:24px">

    <div class="card">
      <div class="card-head">
        <div class="icon" style="background:var(--brand-l)">&#128218;</div>
        <div><h2>How to Use</h2><p>Step-by-step guide</p></div>
      </div>
      <div class="card-body">
        <ol class="steps">
          <li><strong>Upload the .xls file</strong><span>Click or drag the T-shaped XLS. Must be .xls (not .xlsx).</span></li>
          <li><strong>Client Name (optional)</strong><span>Overrides the entity name auto-detected from the XLS. Leave blank if unsure.</span></li>
          <li><strong>Click Convert</strong><span>The tool reads the financial year from the file and fills the PY column automatically. Takes under 15 seconds.</span></li>
          <li><strong>Download &amp; verify</strong><span>Open in Excel. PY column is filled. Yellow cells = CY inputs for you to enter.</span></li>
        </ol>
      </div>
    </div>

    <div class="card">
      <div class="card-head">
        <div class="icon" style="background:#EDE9FE">&#128196;</div>
        <div><h2>Output Format</h2><p>Download the blank output template</p></div>
      </div>
      <div class="card-body">
        <p style="font-size:13px;color:var(--muted);line-height:1.7;margin-bottom:16px">
          The tool fills the <strong>PY column</strong> of this template automatically from the T-shaped XLS.
          Download it to see the exact sheet structure and what each section looks like before you upload.
        </p>
        <a href="/tshape-sample" download="Output_sample_format.xlsx"
           style="display:flex;align-items:center;justify-content:center;gap:10px;
                  background:linear-gradient(135deg,var(--brand),var(--brand-d));color:#fff;
                  border-radius:10px;padding:12px 20px;text-decoration:none;
                  font-size:13px;font-weight:700;transition:opacity .2s"
           onmouseover="this.style.opacity='.88'" onmouseout="this.style.opacity='1'">
          &#11015; Download Output Template (.xlsx)
        </a>
        <div style="margin-top:14px;background:var(--brand-l);border-radius:8px;padding:12px;
                    font-size:11px;color:var(--brand-dk);line-height:1.8;border:1px solid var(--brand-m)">
          <strong>Sheets included:</strong> bs &middot; p&amp;l &middot; notes to accounts &middot; capital &middot;
          notes to bs &middot; notes to p&amp;l &middot; Fixed Assets C.Yr. &middot; Fixed Assets P.Yr. &middot;
          Details &middot; GROSS PROFIT<br/>
          <strong>PY column</strong> = auto-filled &nbsp;&middot;&nbsp;
          <strong>CY column</strong> = yellow (you fill these)
        </div>
      </div>
    </div>

  </div>
</div>

<script>
const dz=document.getElementById('dropzone'),fi=document.getElementById('xlFile'),dzFile=document.getElementById('dzFile');
if(dz&&fi){
  dz.addEventListener('dragover',e=>{e.preventDefault();dz.classList.add('drag');});
  dz.addEventListener('dragleave',()=>dz.classList.remove('drag'));
  dz.addEventListener('drop',e=>{e.preventDefault();dz.classList.remove('drag');
    if(e.dataTransfer.files.length){fi.files=e.dataTransfer.files;showFile(fi.files[0]);}});
  fi.addEventListener('change',()=>{if(fi.files.length)showFile(fi.files[0]);});
}
function showFile(f){
  if(!f.name.toLowerCase().endsWith('.xls')||f.name.toLowerCase().endsWith('.xlsx')){
    showStatus('error','Please upload a .xls file (not .xlsx).');
    return;
  }
  dzFile.textContent='\\u2713 '+f.name;dzFile.style.display='block';
}
async function processFile(){
  const f=fi?fi.files[0]:null,
        oNm=document.getElementById('outputName').value.trim(),
        btn=document.getElementById('processBtn'),
        sp=document.getElementById('spinner'),
        bt=document.getElementById('btnText'),
        dl=document.getElementById('dlBtn');
  if(!f){showStatus('error','Please select a .xls file first.');return;}
  if(!f.name.toLowerCase().endsWith('.xls')||f.name.toLowerCase().endsWith('.xlsx')){
    showStatus('error','Only .xls files are supported.');return;}
  btn.disabled=true;sp.style.display='block';bt.textContent='Converting\\u2026';
  dl.style.display='none';showStatus('','');
  const fd=new FormData();
  fd.append('file',f);
  fd.append('output_name',oNm);
  /* closing_year/new_year not needed — server auto-derives from XLS */
  try{
    const res=await fetch('/process',{method:'POST',body:fd});
    const ct=res.headers.get('content-type')||'';
    if(!ct.includes('application/json')){
      showStatus('error','Server error. Please try again.');return;}
    const data=await res.json();
    if(data.status==='success'){
      const logHtml='<ul class="log-list">'+data.log.map(l=>`<li>${l}</li>`).join('')+'</ul>';
      showStatus('success','\\u2713 Done! Your comparative BS is ready.'+logHtml);
      dl.href='/download/'+data.file_id+'?fn='+encodeURIComponent(data.filename);
      dl.download=data.filename;
      dl.textContent='\\u2b07  Download \\u2014 '+data.filename;
      dl.style.display='block';
    }else{showStatus('error','\\u2717 '+data.message);}
  }catch(e){showStatus('error','Network error: '+e.message);}
  finally{btn.disabled=false;sp.style.display='none';bt.textContent='\\u26a1 Convert & Download';}
}
function showStatus(t,m){const e=document.getElementById('status');e.className=t;e.innerHTML=m;e.style.display=m?'block':'none';}
</script>
</body></html>"""

@app.route("/tool/tshape")
@login_required
def tool_tshape():
    user = get_user_by_id(session["uid"])
    return render_template_string(TSHAPE_T, **user_ctx(user))

@app.route("/tshape-sample")
def tshape_sample_download():
    """Serve the Output_sample_format.xlsx as a downloadable sample."""
    import base64 as _b64, io
    data = _b64.b64decode(_TSHAPE_SAMPLE_B64)
    return send_file(
        io.BytesIO(data),
        download_name="Output_sample_format.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/tool/tax-calculator")
def tool_tax_calculator():
    if "uid" in session:
        user = get_user_by_id(session["uid"])
        ctx = user_ctx(user)
    else:
        ctx = dict(
            username=None, plan="free", plan_label="Free",
            is_admin=False, uploads_used=0, uploads_total=2,
            uploads_left=2, uploads_remaining=2, bar_pct=0,
            validity_end=None, contact_email=CONTACT_EMAIL,
            contact_upi=CONTACT_UPI,
        )
    return render_template_string(TAX_CALC_T, **ctx)

@app.route("/tool/tds-calculator")
def tool_tds_calculator():
    if "uid" in session:
        user = get_user_by_id(session["uid"])
        ctx = user_ctx(user)
    else:
        ctx = dict(
            username=None, plan="free", plan_label="Free",
            is_admin=False, uploads_used=0, uploads_total=2,
            uploads_left=2, uploads_remaining=2, bar_pct=0,
            validity_end=None, contact_email=CONTACT_EMAIL,
            contact_upi=CONTACT_UPI,
        )
    return render_template_string(TDS_CALC_T, **ctx)

@app.route("/tool/depreciation-calculator")
def tool_depreciation_calculator():
    if "uid" in session:
        user = get_user_by_id(session["uid"])
        ctx = user_ctx(user)
    else:
        ctx = dict(
            username=None, plan="free", plan_label="Free",
            is_admin=False, uploads_used=0, uploads_total=2,
            uploads_left=2, uploads_remaining=2, bar_pct=0,
            validity_end=None, contact_email=CONTACT_EMAIL,
            contact_upi=CONTACT_UPI,
        )
    return render_template_string(DEP_CALC_T, **ctx)

# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — PROCESS & DOWNLOAD
# ══════════════════════════════════════════════════════════════════════════════

def _convert_xls_to_xlsx(xls_path, xlsx_path):
    """Convert legacy .xls to .xlsx using xlrd + openpyxl with formatting."""
    import xlrd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb_xls = xlrd.open_workbook(xls_path, formatting_info=True)
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    colour_map = wb_xls.colour_map
    def xlrd_color_to_hex(idx):
        if idx is None or idx < 8 or idx > 63: return None
        rgb = colour_map.get(idx)
        if rgb and rgb != (0, 0, 0): return f'{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}'
        return None

    border_styles = {0: None, 1: 'thin', 2: 'medium', 3: 'dashed',
                     4: 'dotted', 5: 'thick', 6: 'double', 7: 'hair'}

    for sheet_name in wb_xls.sheet_names():
        xls_ws = wb_xls.sheet_by_name(sheet_name)
        xlsx_ws = wb_out.create_sheet(title=sheet_name)

        for row_idx in range(xls_ws.nrows):
            for col_idx in range(xls_ws.ncols):
                ctype = xls_ws.cell_type(row_idx, col_idx)
                if ctype == xlrd.XL_CELL_EMPTY:
                    continue
                val = xls_ws.cell_value(row_idx, col_idx)
                cell = xlsx_ws.cell(row=row_idx + 1, column=col_idx + 1)

                if ctype == xlrd.XL_CELL_NUMBER and val == int(val):
                    val = int(val)
                if ctype == xlrd.XL_CELL_DATE:
                    try:
                        from datetime import datetime
                        dt = xlrd.xldate_as_tuple(val, wb_xls.datemode)
                        cell.value = datetime(*dt)
                    except:
                        cell.value = val
                else:
                    cell.value = val

                # Copy formatting
                try:
                    xf = wb_xls.xf_list[xls_ws.cell_xf_index(row_idx, col_idx)]
                    font_xls = wb_xls.font_list[xf.font_index]

                    cell.font = Font(
                        name=font_xls.name or 'Calibri',
                        size=font_xls.height / 20 if font_xls.height else 11,
                        bold=font_xls.bold, italic=font_xls.italic,
                        underline='single' if font_xls.underline_type else None,
                    )

                    ha = {0: None, 1: 'left', 2: 'center', 3: 'right',
                          5: 'justify'}.get(xf.alignment.hor_align)
                    va = {0: 'top', 1: 'center', 2: 'bottom'}.get(
                        xf.alignment.vert_align, 'bottom')
                    cell.alignment = Alignment(
                        horizontal=ha, vertical=va,
                        wrap_text=bool(xf.alignment.text_wrapped),
                        indent=xf.alignment.indent_level,
                    )

                    fmt_str = wb_xls.format_map.get(xf.format_key)
                    if fmt_str:
                        cell.number_format = fmt_str.format_str

                    def _side(style_idx):
                        s = border_styles.get(style_idx)
                        return Side(style=s) if s else Side()
                    brd = xf.border
                    cell.border = Border(
                        left=_side(brd.left_line_style),
                        right=_side(brd.right_line_style),
                        top=_side(brd.top_line_style),
                        bottom=_side(brd.bottom_line_style),
                    )

                    bg_idx = xf.background.pattern_colour_index
                    bg_hex = xlrd_color_to_hex(bg_idx)
                    if bg_hex and xf.background.fill_pattern:
                        cell.fill = PatternFill('solid', fgColor=bg_hex)
                except Exception:
                    pass

        # Merged cells (after data so we don't write to merged cells)
        for crange in xls_ws.merged_cells:
            r1, r2, c1, c2 = crange
            xlsx_ws.merge_cells(
                start_row=r1 + 1, start_column=c1 + 1,
                end_row=r2, end_column=c2)

        # Column widths
        for c, ci in xls_ws.colinfo_map.items():
            if ci.width:
                xlsx_ws.column_dimensions[get_column_letter(c + 1)].width = ci.width / 256

        # Row heights
        for r, rh in xls_ws.rowinfo_map.items():
            if rh.height:
                xlsx_ws.row_dimensions[r + 1].height = rh.height / 20

    wb_out.save(xlsx_path)

def _convert_xlsb_to_xlsx(xlsb_path: str, xlsx_path: str) -> None:
    """
    Convert a .xlsb (Excel Binary Workbook) to .xlsx using pyxlsb + openpyxl.

    pyxlsb reads the BIFF12 binary format reliably without requiring any
    external system tools (LibreOffice is not available on Render).

    What is preserved: all cell values (numbers, text, dates, booleans).
    What is NOT preserved: cell formatting, colors, borders, merged cells,
    print settings, formulas. For the year-shift tool this is acceptable
    because only cell values are shifted; formatting is irrelevant to the
    algorithm. The output .xlsx will have plain formatting.

    After conversion the year-shift runs on the clean .xlsx exactly like any
    natively-uploaded .xlsx file.

    REQUIRES: pyxlsb  (pip install pyxlsb)
    Add 'pyxlsb' to requirements.txt so Render installs it on deploy.
    """
    try:
        import pyxlsb
    except ImportError:
        raise ImportError(
            "pyxlsb is not installed. Add 'pyxlsb' to requirements.txt and redeploy. "
            "Run: pip install pyxlsb"
        )

    from openpyxl import Workbook
    from datetime import datetime

    # pyxlsb serial-date epoch: days since 1899-12-30
    _EPOCH = datetime(1899, 12, 30)

    # pyxlsb formula placeholder values — when pyxlsb cannot decode a
    # formula's cached result it returns the raw BIFF12 record type as a
    # hex string (e.g. '0x17' = BrtFmlaError, '0x7' = BrtFmlaNum).
    # These are formula artifacts from cross-sheet references, NOT real values.
    # Treat them as empty so they don't pollute the converted xlsx.
    _XLSB_FORMULA_ARTIFACTS = frozenset({"0x17", "0x7", "0x6", "0x9"})

    def _xlsb_val(cell_val):
        """Convert pyxlsb cell value to a Python type openpyxl can write."""
        if cell_val is None:
            return None
        if isinstance(cell_val, bool):
            return cell_val
        if isinstance(cell_val, (int, float)):
            # Dates in xlsb are stored as floats. We cannot reliably distinguish
            # date-formatted numbers without the style table (pyxlsb does not
            # expose it). Return as float — the year-shift only reads numeric
            # values and date strings, so this is safe.
            return cell_val
        if isinstance(cell_val, str):
            # Filter out pyxlsb BIFF12 formula-record-type artifacts
            if cell_val in _XLSB_FORMULA_ARTIFACTS:
                return None
            return cell_val
        # Fallback: coerce to string
        return str(cell_val)

    wb_out = Workbook()
    wb_out.remove(wb_out.active)  # remove default empty sheet

    with pyxlsb.open_workbook(xlsb_path) as wb_in:
        for sheet_name in wb_in.sheets:
            ws_out = wb_out.create_sheet(title=sheet_name)
            with wb_in.get_sheet(sheet_name) as ws_in:
                for row in ws_in.rows():
                    for cell in row:
                        val = _xlsb_val(cell.v)
                        if val is None:
                            continue
                        # pyxlsb cell.r and cell.c are 0-based
                        r = cell.r
                        c = cell.c
                        # Guard: pyxlsb occasionally returns r=-1 or c=-1
                        # for header/metadata rows — skip those.
                        if r < 0 or c < 0:
                            continue
                        ws_out.cell(row=r + 1, column=c + 1, value=val)

    wb_out.save(xlsx_path)


# ── T-shaped balance sheet detection & routing ────────────────────────────────
import sys as _sys
import traceback as _traceback

# Build a list of candidate directories to search for tshape_processor.py
_search_dirs = []
try:
    _search_dirs.append(os.path.dirname(os.path.abspath(__file__)))
except Exception:
    pass
_search_dirs += [
    os.path.abspath('.'),
    os.path.abspath(os.path.join('.', '..')),
    '/opt/render/project/src',
    '/opt/render/project',
]
for _d in _search_dirs:
    if _d not in _sys.path and os.path.isfile(os.path.join(_d, 'tshape_processor.py')):
        _sys.path.insert(0, _d)
        print(f"[tshape] added to sys.path: {_d}", flush=True)
        break

try:
    from tshape_processor import process_tshape as _process_tshape
    _TSHAPE_AVAILABLE = True
    print("[tshape] import OK", flush=True)
except Exception as _e:
    _TSHAPE_AVAILABLE = False
    print(f"[tshape] import FAILED: {type(_e).__name__}: {_e}", flush=True)
    _traceback.print_exc()


def _is_tshape_xls(xls_path: str) -> bool:
    """
    Return True if the .xls file is a T-shaped balance sheet.
    Detection: LIABILITIES and ASSETS appear in the same row within the
    first 15 rows — the hallmark of a T-shaped BS layout.
    """
    try:
        import xlrd as _xlrd
        _wb = _xlrd.open_workbook(xls_path)
        _ws = _wb.sheet_by_index(0)
        for _r in range(min(15, _ws.nrows)):
            _row_str = ' '.join(
                str(_ws.cell_value(_r, _c)).upper()
                for _c in range(min(_ws.ncols, 25))
            )
            if 'LIABILIT' in _row_str and 'ASSET' in _row_str and 'AMOUNT' in _row_str:
                return True
    except Exception:
        pass
    return False


@app.route("/process", methods=["POST"])
@login_required
def process_file():
    try:
        user = get_user_by_id(session["uid"])
        if not user["is_admin"] and uploads_remaining(user) <= 0:
            return jsonify({"status": "error",
                "message": f"No uploads remaining. Contact {CONTACT_EMAIL} to recharge."})
        if "file" not in request.files:
            return jsonify({"status": "error", "message": "No file uploaded."})
        f = request.files["file"]
        orig_name = f.filename.lower()
        is_xls  = orig_name.endswith(".xls") and not orig_name.endswith(".xlsx")
        is_xlsb = orig_name.endswith(".xlsb")
        if not (orig_name.endswith(".xlsx") or is_xls or is_xlsb):
            return jsonify({"status": "error",
                "message": "Only .xlsx, .xls, and .xlsb files are supported."})
        on = request.form.get("output_name", "").strip()
        h  = uuid.uuid4().hex
        ip = os.path.join(UPLOAD_DIR, f"{h}_in.xlsx")
        op = os.path.join(OUTPUT_DIR, f"{h}_out.xlsx")

        # ── Save file & detect T-shaped XLS before any conversion ─────────────
        raw_tmp   = None
        is_tshape = False
        try:
            if is_xls:
                raw_tmp = os.path.join(UPLOAD_DIR, f"{h}_in.xls")
                f.save(raw_tmp)
                # Detect BEFORE conversion — always, regardless of processor availability
                is_tshape = _is_tshape_xls(raw_tmp)
                if not is_tshape:
                    _convert_xls_to_xlsx(raw_tmp, ip)
            elif is_xlsb:
                raw_tmp = os.path.join(UPLOAD_DIR, f"{h}_in.xlsb")
                f.save(raw_tmp)
                _convert_xlsb_to_xlsx(raw_tmp, ip)
            else:
                f.save(ip)
        except Exception as e:
            for p in (raw_tmp, ip):
                if p and not is_tshape:
                    try: os.remove(p)
                    except: pass
            return jsonify({"status": "error", "message": f"File conversion error: {e}"})
        finally:
            if raw_tmp and not is_tshape:
                try: os.remove(raw_tmp)
                except: pass

        if is_tshape:
            # ── T-shaped path ─────────────────────────────────────────────────
            if not _TSHAPE_AVAILABLE:
                try: os.remove(raw_tmp)
                except: pass
                return jsonify({"status": "error",
                                "message": "tshape_processor.py not found on server. "
                                           "Please ensure it is uploaded to the repository."})
            client_name = on or None
            base_name   = on or os.path.splitext(f.filename)[0]
            fname       = f"{base_name}.xlsx"
            try:
                # Find template on disk; fall back to embedded base64 copy
                import base64 as _b64, io as _io
                base_dir = os.path.dirname(os.path.abspath(__file__))
                template_path = None
                for _tn in ('Output_sample_format.xlsx',
                            'Output sample format.xlsx',
                            'output_sample_format.xlsx'):
                    _tp = os.path.join(base_dir, _tn)
                    if os.path.exists(_tp):
                        template_path = _tp
                        break
                if not template_path:
                    # Write embedded template to a temp file
                    _tmp_tpl = os.path.join(UPLOAD_DIR, f"{h}_template.xlsx")
                    with open(_tmp_tpl, 'wb') as _tf:
                        _tf.write(_b64.b64decode(_TSHAPE_SAMPLE_B64))
                    template_path = _tmp_tpl
                result = _process_tshape(
                    input_path=raw_tmp,
                    output_path=op,
                    template_path=template_path,
                    client_name=client_name,
                    cy_year=None,   # auto-derived: py_year_end + 1
                )
                if result.get("status") != "success":
                    return jsonify({"status": "error",
                                    "message": result.get("message", "T-shape processing failed")})
            except Exception as e:
                return jsonify({"status": "error", "message": f"T-shape error: {e}"})
            finally:
                try: os.remove(raw_tmp)
                except: pass
                # Clean up tmp template if we created one
                _tmp_tpl_path = os.path.join(UPLOAD_DIR, f"{h}_template.xlsx")
                if os.path.exists(_tmp_tpl_path):
                    try: os.remove(_tmp_tpl_path)
                    except: pass

        else:
            # ── Standard year-shift path ──────────────────────────────────────
            try:
                cy = int(request.form.get("closing_year", 0))
                ny = int(request.form.get("new_year", cy + 1))
            except ValueError:
                return jsonify({"status": "error", "message": "Invalid year values."})
            if ny != cy + 1:
                return jsonify({"status": "error",
                                "message": "New year must be closing year + 1."})
            import re as _re
            base_name = on or os.path.splitext(f.filename)[0]
            for _ in range(3):
                base_name = _re.sub(r'[_\-]+\d{4}[-_]\d{2,4}$', '', base_name).strip('_- ')
                base_name = _re.sub(r'[_\-]+\d{4}$', '', base_name).strip('_- ')
            fname = f"{base_name}_{ny}.xlsx"
            try:
                result = process(ip, op, cy, ny)
                _rollover_fixed_assets(op, str(ny), result.get("log", []), source_path=ip)
            except Exception as e:
                return jsonify({"status": "error", "message": str(e)})
            finally:
                try: os.remove(ip)
                except: pass

        log_usage(user["id"], fname)
        return jsonify({"status": "success", "log": result["log"], "file_id": h, "filename": fname})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Unexpected error: {e}"}), 500

@app.route("/download/<fid>")
@login_required
def download(fid):
    if not re.fullmatch(r"[a-f0-9]{32}", fid): return "Invalid ID", 400
    path = os.path.join(OUTPUT_DIR, f"{fid}_out.xlsx")
    if not os.path.exists(path): return "File not found or expired.", 404
    fn = request.args.get("fn", f"bs_shift_{fid[:8]}.xlsx")
    if not fn.endswith(".xlsx"): fn += ".xlsx"
    return send_file(path, as_attachment=True,
        download_name=fn,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — ADMIN
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/admin")
@admin_required
def admin_panel():
    raw   = all_users()
    users = []
    for u in raw:
        d = dict(u)
        d["remaining"] = uploads_remaining(u)
        users.append(d)
    return render_template_string(ADMIN_T, users=users, msg=request.args.get("msg", ""))

@app.route("/admin/create", methods=["POST"])
@admin_required
def admin_create():
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")
    plan_key = request.form.get("plan", "free")
    if not username or len(username) < 3:
        return redirect(url_for("admin_panel", msg="Username must be at least 3 characters."))
    if not re.match(r"^[a-zA-Z0-9_]+$", username):
        return redirect(url_for("admin_panel", msg="Username: only letters, numbers, underscores."))
    if len(password) < 6:
        return redirect(url_for("admin_panel", msg="Password must be at least 6 characters."))
    if get_user_by_name(username):
        return redirect(url_for("admin_panel", msg=f"Username '{username}' already exists."))
    if plan_key not in PLANS:
        plan_key = "free"
    create_user(username, password, plan_key)
    plan_label = PLANS[plan_key]["label"]
    return redirect(url_for("admin_panel",
        msg=f"✓ User '{username}' created on {plan_label} plan ({PLANS[plan_key]['uploads']} uploads)."))

@app.route("/admin/addplan", methods=["POST"])
@admin_required
def admin_addplan():
    uid      = int(request.form.get("uid"))
    plan_key = request.form.get("plan", "standard")
    if plan_key not in PLANS: plan_key = "standard"
    user = get_user_by_id(uid)
    if not user: return redirect(url_for("admin_panel", msg="User not found."))
    old_rem = uploads_remaining(user)
    add_uploads(uid, plan_key)
    extra = PLANS[plan_key]["uploads"]
    return redirect(url_for("admin_panel",
        msg=f"✓ Added {extra} uploads to '{user['username']}'. Total remaining: {old_rem + extra}."))

@app.route("/admin/delete", methods=["POST"])
@admin_required
def admin_delete():
    uid = int(request.form.get("uid"))
    if uid == session["uid"]:
        return redirect(url_for("admin_panel", msg="Cannot delete your own account."))
    user = get_user_by_id(uid)
    if not user: return redirect(url_for("admin_panel", msg="User not found."))
    name = user["username"]
    del_user(uid)
    return redirect(url_for("admin_panel", msg=f"✓ User '{name}' deleted."))

@app.route("/tool/msme-calculator")
def tool_msme_calculator():
    if "uid" in session:
        user = get_user_by_id(session["uid"]); ctx = user_ctx(user)
    else:
        ctx = dict(
            username=None, plan="free", plan_label="Free",
            is_admin=False, uploads_used=0, uploads_total=2,
            uploads_left=2, uploads_remaining=2, bar_pct=0,
            validity_end=None, contact_email=CONTACT_EMAIL,
            contact_upi=CONTACT_UPI,
        )
    return render_template_string(MSME_T, **ctx)

@app.route("/tool/capital-gains-calculator")
def tool_cg_calculator():
    if "uid" in session:
        user = get_user_by_id(session["uid"]); ctx = user_ctx(user)
    else:
        ctx = dict(
            username=None, plan="free", plan_label="Free",
            is_admin=False, uploads_used=0, uploads_total=2,
            uploads_left=2, uploads_remaining=2, bar_pct=0,
            validity_end=None, contact_email=CONTACT_EMAIL,
            contact_upi=CONTACT_UPI,
        )
    return render_template_string(CG_CALC_T, **ctx)

# ══════════════════════════════════════════════════════════════════════════════
#  TEMPLATE — GST RECONCILIATION
# ══════════════════════════════════════════════════════════════════════════════

GST_RECON_T = r"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
<meta http-equiv="Pragma" content="no-cache">
<meta http-equiv="Expires" content="0">
<title>GST Reconciliation – CA Toolkit</title>

<style>
""" + BASE_CSS + """
.nav-links{display:flex;gap:20px;list-style:none}
.nav-links a{text-decoration:none;color:var(--muted);font-size:13px;font-weight:500;transition:color .2s}
.nav-links a:hover{color:var(--brand)}
.hero{text-align:center;padding:40px 24px 30px;max-width:700px;margin:0 auto}
.hero-badge{display:inline-flex;align-items:center;gap:6px;background:#FEF3C7;
            color:#92400E;border:1px solid #FDE68A;border-radius:99px;
            padding:5px 14px;font-size:12px;font-weight:600;margin-bottom:18px}
h1{font-size:clamp(22px,3.5vw,34px);font-weight:800;line-height:1.15;letter-spacing:-.5px;margin-bottom:10px}
h1 em{font-style:normal;color:var(--accent)}
.hero p{font-size:14px;color:var(--muted);line-height:1.7;max-width:520px;margin:0 auto}
.main{max-width:900px;margin:0 auto;padding:30px 24px}
.card{background:var(--white);border-radius:var(--radius);border:1px solid var(--border);
      box-shadow:var(--shadow);overflow:hidden;margin-bottom:24px}
.card-head{padding:14px 20px;border-bottom:1px solid var(--border);display:flex;align-items:center;gap:10px}
.card-head .icon{width:32px;height:32px;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:16px}
.card-head h2{font-size:14px;font-weight:700}
.card-body{padding:20px}
.field{margin-bottom:16px}
label{display:block;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.04em;color:var(--muted);margin-bottom:5px}
.hint{font-size:11px;color:var(--muted);margin-top:4px}
.dropzone{border:2px dashed var(--border);border-radius:10px;padding:24px 14px;text-align:center;cursor:pointer;transition:all .2s;position:relative;background:var(--bg)}
.dropzone:hover,.dropzone.drag{border-color:var(--brand);background:#EFF6FF}
.dropzone input{position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%;pointer-events:none}
.dz-icon{font-size:26px;margin-bottom:6px}
.dz-text{font-size:12px;color:var(--muted)}
.dz-text strong{color:var(--brand)}
.dz-file{font-size:12px;font-weight:600;color:var(--green);margin-top:5px;display:none}
.row2{display:grid;grid-template-columns:1fr 1fr;gap:12px}
input[type=text]{width:100%;border:1.5px solid var(--border);border-radius:8px;padding:9px 12px;font-size:13px;font-family:inherit;outline:none;transition:border .2s;box-sizing:border-box}
input[type=text]:focus{border-color:var(--brand)}
.btn{width:100%;padding:12px;background:var(--brand);color:#fff;border:none;border-radius:10px;font-size:14px;font-weight:700;cursor:pointer;display:flex;align-items:center;justify-content:center;gap:8px;transition:background .2s;font-family:inherit}
.btn:hover{background:#1E40AF}
.btn:disabled{background:#93C5FD;cursor:not-allowed}
.spinner{display:none;width:16px;height:16px;border:2.5px solid #fff;border-top-color:transparent;border-radius:50%;animation:spin .7s linear infinite}
@keyframes spin{to{transform:rotate(360deg)}}
.status{padding:14px 16px;border-radius:10px;font-size:13px;margin-top:16px;display:none;line-height:1.6}
.status.success{background:#ECFDF5;border:1px solid #A7F3D0;color:#065F46}
.status.error{background:#FEF2F2;border:1px solid #FECACA;color:#991B1B}
.dl-link{display:none;margin-top:14px;padding:12px 18px;background:#ECFDF5;border:1px solid #A7F3D0;border-radius:10px;color:#065F46;font-weight:700;font-size:13px;text-decoration:none;text-align:center}
.dl-link:hover{background:#D1FAE5}
.info-box{background:#FEF3C7;border:1px solid #FDE68A;border-radius:10px;padding:14px 16px;font-size:12px;color:#92400E;line-height:1.7;margin-bottom:16px}
.log-list{margin:10px 0 0;padding:0;list-style:none;font-size:12px;line-height:1.8}
.log-list li{padding:2px 0}
.mapping-row{display:grid;grid-template-columns:80px 1fr 30px;gap:8px;align-items:center;margin-bottom:8px}
.mapping-row input{font-size:13px}
.mapping-row .remove-btn{background:none;border:none;color:#EF4444;cursor:pointer;font-size:18px;padding:0}
#add-mapping{background:none;border:1px dashed var(--border);border-radius:8px;padding:8px;font-size:12px;color:var(--muted);cursor:pointer;width:100%;margin-top:4px}
#add-mapping:hover{border-color:var(--brand);color:var(--brand)}
</style><script>
function gstDrop(e,dzId,inputId,sfId){
  e.preventDefault();e.stopPropagation();
  var dz=document.getElementById(dzId);if(dz)dz.classList.remove('drag');
  var inp=document.getElementById(inputId);if(!inp)return;
  var files=e.dataTransfer&&e.dataTransfer.files;if(!files||!files.length)return;
  try{var dt=new DataTransfer();dt.items.add(files[0]);inp.files=dt.files;}catch(_){}
  if(inp.files&&inp.files.length){
    var sf=document.getElementById(sfId);
    if(sf){sf.textContent='✓ '+inp.files[0].name;sf.style.display='block';}
    if(sfId==='sf-gst')setTimeout(function(){if(typeof detectStateCodes==='function')detectStateCodes(inp.files[0]);},0);
  }
}
function gstDragOver(e,dzId){
  e.preventDefault();e.stopPropagation();
  if(e.dataTransfer)e.dataTransfer.dropEffect='copy';
  var dz=document.getElementById(dzId);if(dz)dz.classList.add('drag');
}
function gstDragLeave(e,dzId){
  var dz=document.getElementById(dzId);
  if(dz&&!dz.contains(e.relatedTarget))dz.classList.remove('drag');
}
</script>
</head><body>
<nav>
  <a href="/" class="logo">CA<span class="logo-dot"></span><span>Toolkit</span></a>
  <div class="nav-right">
    {% if username %}<div class="nav-user"><span class="nav-avatar">{{ username[0].upper() }}</span><strong>{{ username }}</strong></div><div class="nav-sep"></div>{% if is_admin %}<a href="/admin" class="nav-btn ghost">⚙ Admin</a>{% endif %}{% endif %}
    <a href="/" class="nav-btn dash">⬅ Dashboard</a>
    {% if username %}<a href="/logout" class="nav-link">Sign out</a>{% else %}<a href="/login" class="nav-btn">Sign In →</a>{% endif %}
  </div>
</nav>

<div class="hero">
  <div class="hero-badge">📊 GST Reconciliation</div>
  <h1>Sales <em>Books vs GSTR 3B</em> <span style="font-size:11px;background:#16a34a;color:#fff;padding:2px 8px;border-radius:99px;vertical-align:middle;font-weight:600">v3.0</span></h1>
  <p>Upload your month-wise sales summary and GSTR 3B PDFs to get an instant reconciliation report showing differences by state and month.</p>
</div>

<div class="main">
  <div class="card">
    <div class="card-head">
      <div class="icon" style="background:#FEF3C7">📄</div>
      <div><h2>Upload Files</h2><p style="font-size:12px;color:var(--muted)">Sales summary (.xlsx) + GSTR 3B PDFs (.zip)</p></div>
    </div>
    <div class="card-body">

      <div class="info-box">
        <strong>📊 Sales Summary:</strong> Month in col A, sales value in col B (or separate column per branch/state).<br>
        <strong>📁 GSTR 3B ZIP:</strong> ZIP with sub-folders named by 2-digit state code (e.g. 05/, 09/). GSTR 3B PDFs inside. Reads Table 3.1 A+B+C+E only (excludes D — reverse charge).
      </div>

      <div style="padding:12px 14px;background:#F0FDF4;border:1px solid #BBF7D0;border-radius:8px;margin-bottom:16px">
        <div style="font-size:12px;font-weight:700;color:#065F46;margin-bottom:8px">⬇ Download Sales Format Template</div>
        <div style="display:flex;gap:10px;flex-wrap:wrap">
          <a href="/gst-template/consolidated"
             style="display:inline-flex;align-items:center;gap:6px;padding:7px 14px;background:#059669;color:#fff;border-radius:6px;font-size:12px;font-weight:600;text-decoration:none">
            📥 Consolidated Template
          </a>
          <a href="/gst-template/branchwise"
             style="display:inline-flex;align-items:center;gap:6px;padding:7px 14px;background:#0284C7;color:#fff;border-radius:6px;font-size:12px;font-weight:600;text-decoration:none">
            📥 Branch / State-wise Template
          </a>
        </div>
        <div style="font-size:11px;color:#065F46;margin-top:6px">
          Fill in your figures and upload. Do not change column headers or row order.
        </div>
      </div>

      <!-- CONSOLIDATED CHECKBOX -->
      <div class="field" style="margin-bottom:14px">
        <label style="display:flex;align-items:center;gap:10px;cursor:pointer;font-size:13px;font-weight:600;text-transform:none;letter-spacing:0">
          <input type="checkbox" id="consolidated-chk" onchange="onConsolidatedChange()"
            style="width:18px;height:18px;accent-color:var(--brand);cursor:pointer;flex-shrink:0"/>
          <span>
            <strong>Consolidated Sales Data</strong>
            <span style="font-weight:400;color:var(--muted);margin-left:6px">— tick if your Excel has ONE total column for all branches combined</span>
          </span>
        </label>
        <div id="consolidated-hint" style="display:none;margin-top:8px;padding:8px 12px;background:#EFF6FF;border:1px solid #BFDBFE;border-radius:8px;font-size:12px;color:#1E40AF">
          ✅ <strong>Consolidated mode:</strong> Tool will compare your single total sales figure against the <em>sum of all GSTR 3B states</em> combined. No column mapping needed.
        </div>
        <div id="split-hint" style="margin-top:8px;padding:8px 12px;background:#F0FDF4;border:1px solid #BBF7D0;border-radius:8px;font-size:12px;color:#065F46">
          📍 <strong>Location-wise mode:</strong> Map each state code to its column header in your Excel below.
        </div>
      </div>

      <div class="field">
        <label>Sales Summary (Excel)</label>
        <div class="dropzone" id="dz-sales" ondragover="gstDragOver(event,'dz-sales')" ondragleave="gstDragLeave(event,'dz-sales')" ondrop="gstDrop(event,'dz-sales','file-sales','sf-sales')">
          <div class="dz-icon">📊</div>
          <div class="dz-text"><strong>Click or drag</strong> your sales summary .xlsx</div>
          <div class="dz-file" id="sf-sales"></div>
          <input type="file" id="file-sales" accept=".xlsx,.xls" onchange="pickFile(this,'sf-sales')">
        </div>
      </div>

      <div class="field">
        <label>GSTR 3B PDFs (ZIP file)</label>
        <div class="dropzone" id="dz-gst" ondragover="gstDragOver(event,'dz-gst')" ondragleave="gstDragLeave(event,'dz-gst')" ondrop="gstDrop(event,'dz-gst','file-gst','sf-gst')">
          <div class="dz-icon">📁</div>
          <div class="dz-text"><strong>Click or drag</strong> your GSTR 3B ZIP file</div>
          <div class="dz-file" id="sf-gst"></div>
          <input type="file" id="file-gst" accept=".zip" onchange="pickFile(this,'sf-gst')">
        </div>
      </div>

      <div class="field" id="mapping-field">
        <label>State Code → Sales Column Mapping</label>
        <p class="hint" style="margin-bottom:8px">
          Enter the exact column header from your Excel for each state code (e.g. DRH/LDH, HOSUR, RUDRAPUR).
        </p>
        <div id="mapping-container"></div>
        <button id="add-mapping" onclick="addMapping()">+ Add Mapping</button>
      </div>

      <div class="field" id="consolidated-col-field" style="display:none">
        <label>Column Name in your Excel (Sales column header)</label>
        <input type="text" id="consolidated-col-input" placeholder="e.g. Total Sales, Sales, ALL PLANTS — leave blank to auto-detect first numeric column"/>
        <p class="hint">Enter the exact header of the column containing total sales. Leave blank to auto-detect.</p>
      </div>

      <div class="field">
        <label>Output File Name (optional)</label>
        <input type="text" id="output-name" placeholder="GST_Reconciliation">
      </div>

      <button class="btn" id="proc-btn" onclick="doProcess()">
        <span id="bt">⚡ Process & Download</span>
        <div class="spinner" id="sp"></div>
      </button>

      <div class="status" id="status"></div>
      <a class="dl-link" id="dl-link" href="#">⬇ Download</a>
    </div>
  </div>
</div>

<section style="background:var(--white);border-top:1px solid var(--border);border-bottom:1px solid var(--border);padding:48px 24px">
  <h2 style="text-align:center;font-size:24px;font-weight:800;margin-bottom:6px">Simple Pricing</h2>
  <p style="text-align:center;color:var(--muted);font-size:13px;margin-bottom:32px">Upload-based · 3-month validity · Shared across all premium tools</p>
  <div style="max-width:1080px;margin:0 auto;display:grid;grid-template-columns:repeat(6,1fr);gap:14px">
    <div style="border:1.5px solid var(--border);border-radius:var(--radius);padding:20px 16px">
      <div style="font-size:10px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.06em;margin-bottom:6px">Free</div>
      <div style="font-size:24px;font-weight:800;margin-bottom:2px">₹0</div>
      <div style="font-size:12px;font-weight:700;color:var(--brand);margin-bottom:2px">2 uploads</div>
      <div style="font-size:10px;color:var(--muted);margin-bottom:14px">Try it out</div>
      <ul style="list-style:none;margin-bottom:16px;font-size:11px"><li style="padding:3px 0">✓ All premium tools</li><li style="padding:3px 0">✓ BS + GST Recon</li></ul>
      <a href="#" style="display:block;text-align:center;padding:8px;border-radius:7px;font-size:12px;font-weight:700;background:var(--bg);color:var(--ink);text-decoration:none;border:1px solid var(--border)">Get Started</a>
    </div>
    <div style="border:1.5px solid var(--border);border-radius:var(--radius);padding:20px 16px">
      <div style="font-size:10px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.06em;margin-bottom:6px">Starter</div>
      <div style="font-size:24px;font-weight:800;margin-bottom:2px">₹60</div>
      <div style="font-size:12px;font-weight:700;color:var(--brand);margin-bottom:2px">10 uploads</div>
      <div style="font-size:10px;color:var(--muted);margin-bottom:14px">3 month validity</div>
      <ul style="list-style:none;margin-bottom:16px;font-size:11px"><li style="padding:3px 0">✓ All premium tools</li><li style="padding:3px 0">✓ BS + GST Recon</li></ul>
      <a href="#gst-contact" style="display:block;text-align:center;padding:8px;border-radius:7px;font-size:12px;font-weight:700;background:var(--bg);color:var(--ink);text-decoration:none;border:1px solid var(--border)">Contact to Buy</a>
    </div>
    <div style="border:1.5px solid var(--brand);border-radius:var(--radius);padding:20px 16px;position:relative">
      <div style="position:absolute;top:-10px;left:50%;transform:translateX(-50%);background:var(--brand);color:#fff;font-size:10px;font-weight:700;padding:2px 10px;border-radius:99px;white-space:nowrap">Most Popular</div>
      <div style="font-size:10px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.06em;margin-bottom:6px">Standard</div>
      <div style="font-size:24px;font-weight:800;margin-bottom:2px">₹130</div>
      <div style="font-size:12px;font-weight:700;color:var(--brand);margin-bottom:2px">25 uploads</div>
      <div style="font-size:10px;color:var(--muted);margin-bottom:14px">3 month validity</div>
      <ul style="list-style:none;margin-bottom:16px;font-size:11px"><li style="padding:3px 0">✓ All premium tools</li><li style="padding:3px 0">✓ Priority support</li></ul>
      <a href="#gst-contact" style="display:block;text-align:center;padding:8px;border-radius:7px;font-size:12px;font-weight:700;background:var(--brand);color:#fff;text-decoration:none">Contact to Buy</a>
    </div>
    <div style="border:1.5px solid var(--border);border-radius:var(--radius);padding:20px 16px">
      <div style="font-size:10px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.06em;margin-bottom:6px">Professional</div>
      <div style="font-size:24px;font-weight:800;margin-bottom:2px">₹270</div>
      <div style="font-size:12px;font-weight:700;color:var(--brand);margin-bottom:2px">60 uploads</div>
      <div style="font-size:10px;color:var(--muted);margin-bottom:14px">3 month validity</div>
      <ul style="list-style:none;margin-bottom:16px;font-size:11px"><li style="padding:3px 0">✓ All premium tools</li><li style="padding:3px 0">✓ Priority support</li></ul>
      <a href="#gst-contact" style="display:block;text-align:center;padding:8px;border-radius:7px;font-size:12px;font-weight:700;background:var(--bg);color:var(--ink);text-decoration:none;border:1px solid var(--border)">Contact to Buy</a>
    </div>
    <div style="border:1.5px solid var(--border);border-radius:var(--radius);padding:20px 16px">
      <div style="font-size:10px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.06em;margin-bottom:6px">Firm</div>
      <div style="font-size:24px;font-weight:800;margin-bottom:2px">₹600</div>
      <div style="font-size:12px;font-weight:700;color:var(--brand);margin-bottom:2px">150 uploads</div>
      <div style="font-size:10px;color:var(--muted);margin-bottom:14px">3 month validity</div>
      <ul style="list-style:none;margin-bottom:16px;font-size:11px"><li style="padding:3px 0">✓ All premium tools</li><li style="padding:3px 0">✓ WhatsApp support</li></ul>
      <a href="#gst-contact" style="display:block;text-align:center;padding:8px;border-radius:7px;font-size:12px;font-weight:700;background:var(--bg);color:var(--ink);text-decoration:none;border:1px solid var(--border)">Contact to Buy</a>
    </div>
    <div style="border:1.5px solid var(--border);border-radius:var(--radius);padding:20px 16px">
      <div style="font-size:10px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.06em;margin-bottom:6px">CA Firm</div>
      <div style="font-size:24px;font-weight:800;margin-bottom:2px">₹1,000</div>
      <div style="font-size:12px;font-weight:700;color:var(--brand);margin-bottom:2px">500 uploads</div>
      <div style="font-size:10px;color:var(--muted);margin-bottom:14px">3 month validity</div>
      <ul style="list-style:none;margin-bottom:16px;font-size:11px"><li style="padding:3px 0">✓ All premium tools</li><li style="padding:3px 0">✓ Best for CA firms</li></ul>
      <a href="#gst-contact" style="display:block;text-align:center;padding:8px;border-radius:7px;font-size:12px;font-weight:700;background:var(--bg);color:var(--ink);text-decoration:none;border:1px solid var(--border)">Contact to Buy</a>
    </div>
  </div>
  <p style="text-align:center;font-size:11px;color:var(--muted);margin-top:16px">⚠ No refund after first upload is used · Unused uploads stack when you recharge</p>
</section>

<section style="max-width:700px;margin:0 auto;padding:48px 24px" id="gst-contact">
  <h2 style="text-align:center;font-size:20px;font-weight:800;margin-bottom:16px">Purchase a Plan</h2>
  <p style="text-align:center;color:var(--muted);font-size:13px;margin-bottom:24px">Pay via UPI and send your payment screenshot to our email. Account upgraded within a few hours.</p>
  <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;max-width:400px;margin:0 auto">
    <div style="text-align:center">
      <strong style="font-size:11px;color:var(--muted);text-transform:uppercase">Email</strong><br>
      <a href="mailto:{{ contact_email }}" style="font-size:13px;color:var(--brand)">{{ contact_email }}</a>
    </div>
    <div style="text-align:center">
      <strong style="font-size:11px;color:var(--muted);text-transform:uppercase">UPI Payment</strong><br>
      <span style="font-size:13px;font-weight:600">{{ contact_upi }}</span>
    </div>
  </div>
</section>

<div id="toast" style="position:fixed;bottom:24px;right:24px;background:#065F46;color:#fff;padding:12px 20px;border-radius:10px;font-size:13px;font-weight:600;opacity:0;transition:opacity .3s;pointer-events:none;z-index:999"></div>

<script>
function pickFile(inp, sfId){
  const sf=document.getElementById(sfId);
  if(inp.files.length){sf.textContent='✓ '+inp.files[0].name;sf.style.display='block';}
  if(sfId==='sf-gst' && inp.files[0]){detectStateCodes(inp.files[0]);}
}

// Drag-and-drop for GST upload zones
// pointer-events:none on the input passes drags to the .dropzone div.
// dragleave fires on child-element crossings too, so we use relatedTarget
// to ignore those false leaves and only clear 'drag' when truly leaving the zone.
['file-sales', 'file-gst'].forEach(function(inputId) {
  var inp  = document.getElementById(inputId);
  var dz   = inp ? inp.closest('.dropzone') : null;
  var sfId = inputId.replace('file-', 'sf-');
  if (!inp || !dz) return;

  dz.addEventListener('dragenter', function(e) {
    e.preventDefault(); e.stopPropagation();
    dz.classList.add('drag');
  });
  dz.addEventListener('dragover', function(e) {
    e.preventDefault(); e.stopPropagation();
    dz.classList.add('drag');
    e.dataTransfer.dropEffect = 'copy';
  });
  dz.addEventListener('dragleave', function(e) {
    e.preventDefault();
    // Only remove highlight when the drag truly leaves the zone (not a child)
    if (!dz.contains(e.relatedTarget)) {
      dz.classList.remove('drag');
    }
  });
  dz.addEventListener('drop', function(e) {
    e.preventDefault(); e.stopPropagation();
    dz.classList.remove('drag');
    var files = e.dataTransfer && e.dataTransfer.files;
    if (!files || !files.length) return;
    var f = files[0];
    try {
      var dt = new DataTransfer();
      dt.items.add(f);
      inp.files = dt.files;
    } catch(_) {}
    pickFile(inp, sfId);
  });
  // Clicking anywhere on the zone opens the file picker
  dz.addEventListener('click', function() { inp.click(); });
});

// India state name -> GST state code (for ZIP folder detection)
const _INDIA_SC = {'jammu and kashmir':'01','jammu & kashmir':'01','j&k':'01','himachal pradesh':'02','punjab':'03','chandigarh':'04','uttarakhand':'05','haryana':'06','delhi':'07','new delhi':'07','rajasthan':'08','uttar pradesh':'09','u.p':'09','u.p.':'09','bihar':'10','sikkim':'11','arunachal pradesh':'12','nagaland':'13','manipur':'14','mizoram':'15','tripura':'16','meghalaya':'17','assam':'18','west bengal':'19','jharkhand':'20','odisha':'21','orissa':'21','chhattisgarh':'22','chattisgarh':'22','madhya pradesh':'23','gujarat':'24','maharashtra':'27','andhra pradesh':'28','karnataka':'29','goa':'30','lakshadweep':'31','kerala':'32','tamil nadu':'33','tamilnadu':'33','puducherry':'34','pondicherry':'34','andaman and nicobar':'35','telangana':'36','ladakh':'38'};
function _folderToSC(seg) {
  var nm = /(?:^|[\/_\s-])(0[1-9]|[1-3][0-9])(?:$|[\/_\s-])/.exec(seg);
  if (nm) return nm[1];
  if (/^(0[1-9]|[1-3][0-9])$/.test(seg.trim())) return seg.trim();
  var sl = seg.toLowerCase();
  var keys = Object.keys(_INDIA_SC).sort(function(a,b){return b.length-a.length;});
  for (var i=0; i<keys.length; i++) { if (sl.indexOf(keys[i]) !== -1) return _INDIA_SC[keys[i]]; }
  return null;
}
async function detectStateCodes(file) {
  var ab   = await file.arrayBuffer();
  var text = new TextDecoder('utf-8', {fatal: false}).decode(new Uint8Array(ab));
  var codes = new Map();
  // Split on forward-slash to get ZIP path segments. No regex = no escape issues.
  var parts = text.split('/');
  for (var i = 0; i < parts.length; i++) {
    var seg = parts[i].trim();
    if (!seg || seg.length < 2 || seg.length > 80) continue;
    var sc = _folderToSC(seg);
    if (sc && !codes.has(sc)) codes.set(sc, seg);
  }
  if (codes.size > 0) {
    document.getElementById('mapping-container').innerHTML = '';
    var sorted = Array.from(codes.entries()).sort(function(a,b){return a[0].localeCompare(b[0]);});
    sorted.forEach(function(pair) { addMapping(pair[0], ''); });

    // FIX Bug 1: Show a visible banner prompting the user to fill in column headers.
    // Without this, the user has no idea the mapping rows need input — the button
    // silently rejects because code is set but col is blank.
    var mf = document.getElementById('mapping-field');
    var existingBanner = document.getElementById('autodetect-banner');
    if (!existingBanner) {
      var banner = document.createElement('div');
      banner.id = 'autodetect-banner';
      banner.style.cssText = 'margin-top:10px;padding:10px 14px;background:#FEF9C3;border:1.5px solid #FDE047;border-radius:8px;font-size:12px;color:#713F12;font-weight:600';
      banner.innerHTML = '⚠️ ' + codes.size + ' state codes detected from your ZIP. <strong>Please fill in the matching column header</strong> from your Excel sales file for each row below, then click Process & Download.';
      mf.appendChild(banner);
    }
    // Scroll the mapping section into view
    mf.scrollIntoView({behavior: 'smooth', block: 'center'});
  }
}

function onConsolidatedChange(){
  const chk = document.getElementById('consolidated-chk').checked;
  document.getElementById('consolidated-hint').style.display      = chk ? 'block' : 'none';
  document.getElementById('split-hint').style.display             = chk ? 'none'  : 'block';
  document.getElementById('mapping-field').style.display          = chk ? 'none'  : 'block';
  document.getElementById('consolidated-col-field').style.display = chk ? 'block' : 'none';
}

function addMapping(code,col){
  const container=document.getElementById('mapping-container');
  const div=document.createElement('div');div.className='mapping-row';
  // FIX Bug 3: When auto-detected (code set, col blank), highlight the column
  // input in amber so the user immediately sees what needs to be filled.
  const colStyle = (code && !col) ? 'border:2px solid #F59E0B;background:#FFFBEB;' : '';
  const colPlaceholder = (code && !col) ? '⚠ Enter your Excel column header here' : 'Column header (e.g. DRH/LDH)';
  div.innerHTML=`<input type="text" class="map-code" value="${code||''}" placeholder="03">
    <input type="text" class="map-col" value="${col||''}" placeholder="${colPlaceholder}" style="${colStyle}"
      oninput="this.style.border='';this.style.background='';this.placeholder='Column header (e.g. DRH/LDH)'">
    <button class="remove-btn" onclick="this.parentElement.remove()">✕</button>`;
  container.appendChild(div);
}

function showStatus(t,m){const e=document.getElementById('status');e.className=t;e.innerHTML=m;e.style.display=m?'block':'none';}
function toast(msg){const t=document.getElementById('toast');t.textContent=msg;t.style.opacity='1';setTimeout(()=>t.style.opacity='0',3000);}

async function doProcess(){
  const btn=document.getElementById('proc-btn'),sp=document.getElementById('sp'),bt=document.getElementById('bt');
  const dl=document.getElementById('dl-link');
  const salesFile=document.getElementById('file-sales').files[0];
  const gstFile=document.getElementById('file-gst').files[0];
  if(!salesFile){showStatus('error','✗ Please upload your Sales Summary Excel file.');return;}
  if(!gstFile){showStatus('error','✗ Please upload your GSTR 3B ZIP file.');return;}

  // Check consolidated mode
  const chkEl=document.getElementById('consolidated-chk');
  const isConsolidated = chkEl && chkEl.checked;

  // Collect state-column mappings (location-wise mode only)
  const mappings={};
  if(!isConsolidated){
    const rows=document.querySelectorAll('.mapping-row');
    for(const r of rows){
      const code=r.querySelector('.map-code').value.trim();
      const col=r.querySelector('.map-col').value.trim();
      if(code&&col)mappings[code]=col;
    }
    if(Object.keys(mappings).length===0){
      // FIX Bug 2: Check if rows exist but have empty column values (auto-detect case).
      const allRows = document.querySelectorAll('.mapping-row');
      const hasIncomplete = Array.from(allRows).some(r => r.querySelector('.map-code').value.trim() && !r.querySelector('.map-col').value.trim());
      const msg = hasIncomplete
        ? '✗ Please fill in the <strong>Column Header</strong> field for each state row (the second input box, e.g. "DRH/LDH", "Total Sales").'
        : '✗ Please add at least one State Code → Column mapping, or tick Consolidated Sales.';
      showStatus('error', msg);
      // Scroll the error message into view so user can see it
      const statusEl = document.getElementById('status');
      if(statusEl) statusEl.scrollIntoView({behavior:'smooth', block:'center'});
      return;
    }
  }

  // Build FormData FIRST, then append all fields
  const fd=new FormData();
  fd.append('sales_file',salesFile);
  fd.append('gst_file',gstFile);
  fd.append('mappings',JSON.stringify(mappings));
  fd.append('output_name',document.getElementById('output-name').value.trim());

  // Consolidated params
  if(isConsolidated){
    const colInputEl=document.getElementById('consolidated-col-input');
    const colName=(colInputEl?colInputEl.value:'').trim();
    fd.append('consolidated_mode','true');
    fd.append('consolidated_col',colName);
  }

  btn.disabled=true;sp.style.display='inline-block';bt.textContent='Processing…';
  showStatus('info','⏳ Processing — this may take 30–60 seconds for large ZIP files…');
  dl.style.display='none';

  // Live counter so user knows it's working
  let _secs=0;
  const _timer=setInterval(()=>{
    _secs++;
    bt.textContent=`Processing… ${_secs}s`;
  },1000);

  // 5-minute timeout — large ZIPs (many states × months) can exceed 3 min
  // on free-tier Render; deduplication (server-side) halves parse count but
  // we keep a generous client timeout to avoid false "Timed out" errors.
  const _ctrl=new AbortController();
  const _tout=setTimeout(()=>_ctrl.abort(),300000);

  try{
    const res=await fetch('/gst-process',{method:'POST',body:fd,
      credentials:'include',signal:_ctrl.signal});
    clearTimeout(_tout);
    const ct=res.headers.get('content-type')||'';
    if(!ct.includes('application/json')){
      showStatus('error','✗ Server error (not JSON). Please try again.');return;
    }
    const data=await res.json();
    if(data.status==='success'){
      const logHtml='<ul class="log-list">'+data.log.map(l=>`<li>${l}</li>`).join('')+'</ul>';
      showStatus('success','✓ Reconciliation complete! ('+_secs+'s)'+logHtml);
      dl.href='/download/'+data.file_id+'?fn='+encodeURIComponent(data.filename);dl.download=data.filename;
      dl.textContent='⬇  Download — '+data.filename;dl.style.display='block';
      toast('Reconciliation done!');
    }else{showStatus('error','✗ '+data.message);}
  }catch(e){
    clearTimeout(_tout);
    if(e.name==='AbortError'){
      showStatus('error','✗ Timed out after 5 minutes. Try splitting the ZIP into smaller batches or contact support.');
    }else{
      showStatus('error','✗ Network error: '+e.message);
    }
  }
  finally{
    clearInterval(_timer);
    btn.disabled=false;sp.style.display='none';bt.textContent='⚡ Process & Download';
  }
}
</script>
<a href="https://wa.me/918427651580" target="_blank" class="wa-float" title="WhatsApp Support"><svg viewBox="0 0 24 24"><path d="M17.472 14.382c-.297-.149-1.758-.867-2.03-.967-.273-.099-.471-.148-.67.15-.197.297-.767.966-.94 1.164-.173.199-.347.223-.644.075-.297-.15-1.255-.463-2.39-1.475-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.298-.347.446-.52.149-.174.198-.298.298-.497.099-.198.05-.371-.025-.52-.075-.149-.669-1.612-.916-2.207-.242-.579-.487-.5-.669-.51-.173-.008-.371-.01-.57-.01-.198 0-.52.074-.792.372-.272.297-1.04 1.016-1.04 2.479 0 1.462 1.065 2.875 1.213 3.074.149.198 2.096 3.2 5.077 4.487.709.306 1.262.489 1.694.625.712.227 1.36.195 1.871.118.571-.085 1.758-.719 2.006-1.413.248-.694.248-1.289.173-1.413-.074-.124-.272-.198-.57-.347m-5.421 7.403h-.004a9.87 9.87 0 01-5.031-1.378l-.361-.214-3.741.982.998-3.648-.235-.374a9.86 9.86 0 01-1.51-5.26c.001-5.45 4.436-9.884 9.888-9.884 2.64 0 5.122 1.03 6.988 2.898a9.825 9.825 0 012.893 6.994c-.003 5.45-4.437 9.884-9.885 9.884m8.413-18.297A11.815 11.815 0 0012.05 0C5.495 0 .16 5.335.157 11.892c0 2.096.547 4.142 1.588 5.945L.057 24l6.305-1.654a11.882 11.882 0 005.683 1.448h.005c6.554 0 11.89-5.335 11.893-11.893a11.821 11.821 0 00-3.48-8.413z"/></svg></a>

<button class="help-btn" onclick="openHelp()" title="How to use this tool">?</button>
<div class="help-overlay" id="helpOverlay">
  <div class="help-modal">
    <div class="help-modal-head"><h3>How to Use — GST Reconciliation</h3><button class="help-close" onclick="closeHelp()">&#10005;</button></div>
    <div class="help-modal-body"><div class="help-step"><div class="help-step-num">1</div><div class="help-step-body"><h4>Upload Sales Excel</h4><p>Upload your books sales summary Excel with month-wise, state-wise data.</p></div></div><div class="help-step"><div class="help-step-num">2</div><div class="help-step-body"><h4>Upload GSTR-3B ZIP</h4><p>Zip all your GSTR-3B PDFs (one per month) and upload the ZIP file.</p></div></div><div class="help-step"><div class="help-step-num">3</div><div class="help-step-body"><h4>Review Mappings</h4><p>Map your Excel column headers to the required fields on screen.</p></div></div><div class="help-step"><div class="help-step-num">4</div><div class="help-step-body"><h4>Process</h4><p>Click Process to generate the reconciliation report.</p></div></div><div class="help-step"><div class="help-step-num">5</div><div class="help-step-body"><h4>Download Report</h4><p>Download the Excel report with month-wise and state-wise differences highlighted.</p></div></div><div class="help-tip">📌 Export GSTR-3B PDFs from the GST portal and ZIP them before uploading.</div></div>
  </div>
</div>
<script>function openHelp(){document.getElementById('helpOverlay').classList.add('open')}function closeHelp(){document.getElementById('helpOverlay').classList.remove('open')}document.getElementById('helpOverlay').addEventListener('click',function(e){if(e.target===this)closeHelp()})</script>
</body></html>"""

# ══════════════════════════════════════════════════════════════════════════════
#  GST RECONCILIATION — PROCESSING LOGIC
# ══════════════════════════════════════════════════════════════════════════════

_INDIA_STATE_CODE = {
    "jammu and kashmir":"01","jammu & kashmir":"01","j&k":"01",
    "himachal pradesh":"02","himachal":"02","punjab":"03","chandigarh":"04",
    "uttarakhand":"05","uttaranchal":"05","haryana":"06","delhi":"07","new delhi":"07",
    "rajasthan":"08","uttar pradesh":"09","u.p":"09","u.p.":"09","up":"09",
    "bihar":"10","sikkim":"11","arunachal pradesh":"12","nagaland":"13",
    "manipur":"14","mizoram":"15","tripura":"16","meghalaya":"17","assam":"18",
    "west bengal":"19","wb":"19","jharkhand":"20","odisha":"21","orissa":"21",
    "chhattisgarh":"22","chattisgarh":"22","madhya pradesh":"23","m.p":"23","mp":"23",
    "gujarat":"24","daman and diu":"25","dadra and nagar haveli":"26",
    "maharashtra":"27","andhra pradesh":"28","ap":"28","karnataka":"29","goa":"30",
    "lakshadweep":"31","kerala":"32","tamil nadu":"33","tamilnadu":"33","tn":"33",
    "puducherry":"34","pondicherry":"34","andaman and nicobar":"35",
    "telangana":"36","andhra pradesh (new)":"37","ladakh":"38",
}

def _folder_to_state_code(folder_name):
    import re as _r
    s = folder_name.strip()
    m = _r.search(r'(?:^|[/_\s-])(0[1-9]|[1-3][0-9])(?:$|[/_\s-])', s)
    if m: return m.group(1)
    if _r.fullmatch(r'(0[1-9]|[1-3][0-9])', s): return s
    sl = s.lower()
    for name in sorted(_INDIA_STATE_CODE, key=len, reverse=True):
        if name in sl: return _INDIA_STATE_CODE[name]
    return None


def _parse_gstr3b_pdf(pdf_path):
    """Extract Table 3.1(a) data from a GSTR-3B PDF using pymupdf (fitz).

    pymupdf is ~40× faster than pdfplumber (7 ms vs 300 ms per PDF) because
    it reads the raw PDF text stream directly without rendering.  It outputs
    each field on its own line (label then value), making the parse simpler
    and more reliable than pdfplumber's inline text.

    Falls back to a bytes-based call so the function also accepts raw bytes
    (for in-memory ZIP extraction without writing to disk).
    """
    import re

    result = {
        'taxable_value': 0.0, 'igst': 0.0, 'cgst': 0.0,
        'sgst': 0.0, 'cess': 0.0, 'total_tax': 0.0,
        'period': None, 'year': None, 'gstin': None,
        'trade_name': None, 'state_code': None,
    }

    def _num(s):
        try:
            return float(str(s).strip().replace(',', ''))
        except Exception:
            return 0.0

    try:
        import fitz  # pymupdf
        if isinstance(pdf_path, (bytes, bytearray)):
            doc = fitz.open(stream=pdf_path, filetype="pdf")
        else:
            doc = fitz.open(pdf_path)
        text = doc[0].get_text()
        doc.close()
    except Exception:
        return result

    lines = [ln.strip() for ln in text.split('\n')]

    # ── Metadata ─────────────────────────────────────────────────────────────
    # pymupdf puts each label on one line and the value on the next line.
    for i, line in enumerate(lines):
        nxt = lines[i + 1] if i + 1 < len(lines) else ''
        if line == 'Period':
            result['period'] = nxt or result['period']
        elif line == 'Year':
            result['year'] = nxt or result['year']
        elif line == 'GSTIN of the supplier':
            result['gstin'] = nxt
            if len(nxt) >= 2 and nxt[:2].isdigit():
                result['state_code'] = nxt[:2]
        elif 'Trade name, if any' in line:
            result['trade_name'] = nxt or result['trade_name']

    # ── Table 3.1 rows A, B, C, E — sum all four, exclude D (reverse charge) ─
    # Structure per row: label line(s), then 5 values: taxable, igst, cgst, sgst, cess
    # Row markers (as they appear in PDF text):
    #   (a) Outward taxable supplies (other than zero rated, nil rated and exempted)
    #   (b) Outward taxable supplies (zero rated)
    #   (c ) Other outward supplies (nil rated, exempted)
    #   (d) Inward supplies (liable to reverse charge)   ← EXCLUDED
    #   (e) Non-GST outward supplies

    def _extract_row_values(lines, start_idx, end_label_substr):
        """From start_idx, skip past the label (until end_label_substr found),
        then collect the next 5 numeric/dash tokens as [taxable, igst, cgst, sgst, cess]."""
        j = start_idx
        # advance until we hit the end of the label text
        while j < len(lines) and end_label_substr not in lines[j].lower():
            j += 1
        nums = []
        k = j + 1
        while k < len(lines) and len(nums) < 5:
            v = lines[k].replace(',', '').strip()
            if re.match(r'^-?\d+\.?\d*$', v):
                nums.append(float(v))
            elif v in ('-', '\u2013', '\u2014', ''):
                nums.append(0.0)
            elif v.startswith('('):
                break   # hit the next row label — stop
            k += 1
        return nums if len(nums) >= 5 else None

    row_a = row_b = row_c = row_e = None
    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped.startswith('(a)') and 'Outward taxable' in stripped and row_a is None:
            row_a = _extract_row_values(lines, i, 'exempted')
        elif stripped.startswith('(b)') and 'zero rated' in stripped.lower() and row_b is None:
            row_b = _extract_row_values(lines, i, 'zero rated')
        elif stripped.startswith('(c') and 'nil rated' in stripped.lower() and row_c is None:
            row_c = _extract_row_values(lines, i, 'exempted')
        elif stripped.startswith('(e)') and 'Non-GST' in stripped and row_e is None:
            row_e = _extract_row_values(lines, i, 'supplies')

    # Sum A + B + C + E for each column
    def _col(rows, idx):
        return sum(r[idx] for r in rows if r is not None)

    rows_abce = [r for r in [row_a, row_b, row_c, row_e] if r is not None]
    if rows_abce:
        result['taxable_value'] = _col(rows_abce, 0)
        result['igst']          = _col(rows_abce, 1)
        result['cgst']          = _col(rows_abce, 2)
        result['sgst']          = _col(rows_abce, 3)
        result['cess']          = _col(rows_abce, 4)
        result['total_tax']     = result['igst'] + result['cgst'] + result['sgst'] + result['cess']

    return result


def _month_key(period_name, fy_str):
    """Convert GSTR3B period 'December' + year '2025-26' to 'Dec-25' format.
    Handles both full names ('January') and abbreviations ('Jan')."""
    month_abbr = {
        'january': 'Jan', 'february': 'Feb', 'march': 'Mar', 'april': 'Apr',
        'may': 'May', 'june': 'Jun', 'july': 'Jul', 'august': 'Aug',
        'september': 'Sep', 'october': 'Oct', 'november': 'Nov', 'december': 'Dec',
        # Also accept 3-letter abbreviations directly
        'jan': 'Jan', 'feb': 'Feb', 'mar': 'Mar', 'apr': 'Apr',
        'jun': 'Jun', 'jul': 'Jul', 'aug': 'Aug', 'sep': 'Sep',
        'oct': 'Oct', 'nov': 'Nov', 'dec': 'Dec',
    }
    abbr = month_abbr.get(period_name.lower().strip())
    if not abbr:
        return None
    # FY "2025-26" means Apr 2025 - Mar 2026
    # Apr-Mar months: Apr,May,...,Dec use first year; Jan,Feb,Mar use second
    try:
        fy_start = int(fy_str.split('-')[0])
    except:
        return None
    if abbr in ('Jan', 'Feb', 'Mar'):
        yr = fy_start + 1
    else:
        yr = fy_start
    return f"{abbr}-{str(yr)[2:]}"


def _process_gst_reconciliation(sales_path, gst_zip_path, mappings, output_path):
    """Process GST reconciliation and generate output Excel."""
    from openpyxl import load_workbook as lb
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    import tempfile, shutil

    log = []

    # --- 1. Read Sales Summary ---
    wb_sales = lb(sales_path, read_only=True, data_only=True)
    ws = wb_sales[wb_sales.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=1, max_row=500, values_only=False))
    wb_sales.close()

    # Find header row (first row with "Month" in col A or similar)
    header_idx = None
    for i, row in enumerate(rows):
        a_val = str(row[0].value or '').strip().lower()
        if a_val in ('month', 'months', 'period'):
            header_idx = i
            break
    if header_idx is None:
        # Try: any row with 1+ non-empty cells where at least one is a text header (non-numeric)
        for i, row in enumerate(rows):
            non_empty = [c for c in row if c.value is not None]
            has_text = any(
                isinstance(c.value, str) and c.value.strip()
                and c.value.strip().lower() not in ('total', 'grand total', 'subtotal')
                for c in non_empty
            )
            has_numeric = any(isinstance(c.value, (int, float)) for c in non_empty)
            if len(non_empty) >= 1 and has_text and not has_numeric:
                header_idx = i
                break
    if header_idx is None:
        # Final fallback: first row with any non-empty cell
        for i, row in enumerate(rows):
            if any(c.value is not None for c in row):
                header_idx = i
                break
    if header_idx is None:
        return {'status': 'error', 'message': 'Could not find header row in sales file.'}

    header_row = rows[header_idx]
    col_headers = {str(c.value).strip(): c.column - 1 for c in header_row if c.value}
    log.append(f"Sales columns found: {', '.join(col_headers.keys())}")

    # FIX: Find Month column dynamically (not hardcoded to col 0)
    month_col_idx = 0
    for hdr, idx in col_headers.items():
        if hdr.lower() in ('month', 'months', 'period'):
            month_col_idx = idx
            break

    def _parse_sales_val(val):
        if val is None: return 0.0
        if isinstance(val, (int, float)): return float(val)
        s = str(val).strip().replace(',', '')
        try: return float(s)
        except: return 0.0

    # Read month-wise sales data
    _DT_MNAMES = {1:'January',2:'February',3:'March',4:'April',5:'May',6:'June',
                  7:'July',8:'August',9:'September',10:'October',11:'November',12:'December'}
    sales_data = {}
    for row in rows[header_idx + 1:]:
        month_val = row[month_col_idx].value if len(row) > month_col_idx else None
        if month_val is None:
            continue
        if hasattr(month_val, 'month') and hasattr(month_val, 'year'):
            month_val = f"{_DT_MNAMES[month_val.month]}-{str(month_val.year)[2:]}"
        ms = str(month_val).strip()
        if 'total' in ms.lower() or 'grand' in ms.lower():
            continue
        row_data = {}
        for hdr, idx in col_headers.items():
            if hdr.lower() in ('month', 'months', 'period'):
                continue
            try:
                val = row[idx].value if len(row) > idx else None
                row_data[hdr] = _parse_sales_val(val)
            except:
                row_data[hdr] = 0.0
        if ms not in sales_data:
            sales_data[ms] = row_data
        else:
            if sum(abs(v) for v in row_data.values()) > sum(abs(v) for v in sales_data[ms].values()):
                sales_data[ms] = row_data

    log.append(f"Sales months found: {', '.join(sales_data.keys())}")

    # --- 2. Extract GSTR 3B data from ZIP ---
    gst_data = {}  # {state_code: {month_key: {taxable, igst, cgst, sgst, cess, total_tax}}}
    trade_names = {}

    tmpdir = tempfile.mkdtemp()
    try:
        import zipfile as zf

        # Read PDFs directly from ZIP bytes — no extractall, no temp-file I/O.
        # Also deduplicate: some client ZIPs contain each PDF twice (once in a
        # named-city folder like "Ludhiana/" and again under the state-code
        # folder "03/").  Skip any filename already seen.
        pdf_tasks = []
        _seen_paths: set = set()
        _folder_sc_hint: dict = {}
        _sc_display_name: dict = {}
        with zf.ZipFile(gst_zip_path, 'r') as _z:
            for entry in _z.infolist():
                fname = entry.filename.split('/')[-1]
                if not fname.lower().endswith('.pdf'):
                    continue
                if 'certificate' in entry.filename.lower():
                    continue
                if entry.filename.lower() in _seen_paths:
                    continue
                _seen_paths.add(entry.filename.lower())
                import re as _rd
                for seg in reversed(entry.filename.rstrip('/').split('/')[:-1]):
                    sc = _folder_to_state_code(seg)
                    if sc:
                        _folder_sc_hint[entry.filename] = sc
                        if sc not in _sc_display_name:
                            disp = _rd.sub(r'^(gst\s*3b\s*|gst\s+)', '', seg.strip(), flags=_rd.IGNORECASE).strip() or seg.strip()
                            _sc_display_name[sc] = disp
                        break
                pdf_bytes = _z.read(entry.filename)
                pdf_tasks.append((pdf_bytes, fname, entry.filename))

        log.append(f"Found {len(pdf_tasks)} GSTR 3B PDFs — parsing...")
        if _sc_display_name:
            log.append(f"  State hints: {_sc_display_name}")

        from concurrent.futures import ThreadPoolExecutor, as_completed
        def _parse_task(args):
            pdf_bytes, fname, entry_path = args
            return fname, entry_path, _parse_gstr3b_pdf(pdf_bytes)

        with ThreadPoolExecutor(max_workers=4) as pool:
            futures = {pool.submit(_parse_task, t): t for t in pdf_tasks}
            for future in as_completed(futures):
                try:
                    fname, entry_path, result = future.result(timeout=30)
                except Exception as _e:
                    log.append(f"  ⚠ PDF parse error: {_e}")
                    continue

                if result and not result.get('state_code'):
                    result['state_code'] = _folder_sc_hint.get(entry_path)

                if result and result['state_code'] and result['period']:
                    sc = result['state_code']

                    # ── CRITICAL FIX: derive month from FILENAME not PDF period ──
                    mk = None
                    fn_base = os.path.splitext(fname)[0]
                    fn_parts = fn_base.split('_')
                    fn_mmyyyy = fn_parts[-1] if fn_parts else ''
                    if len(fn_mmyyyy) == 6 and fn_mmyyyy.isdigit():
                        mm   = int(fn_mmyyyy[:2])
                        yyyy = int(fn_mmyyyy[2:])
                        _MNAMES = {1:'January',2:'February',3:'March',4:'April',
                                   5:'May',6:'June',7:'July',8:'August',
                                   9:'September',10:'October',11:'November',12:'December'}
                        fn_period = _MNAMES.get(mm, result['period'])
                        fn_fy_start = yyyy - 1 if mm <= 3 else yyyy
                        fn_fy = f"{fn_fy_start}-{str(fn_fy_start+1)[2:]}"
                        mk = _month_key(fn_period, fn_fy)
                        if mk and mk != _month_key(result['period'], result['year'] or ''):
                            log.append(f"  ℹ Quarterly filer: PDF says '{result['period']}' "
                                       f"but filename says {fn_period} → using '{mk}'")

                    if not mk:
                        mk = _month_key(result['period'], result['year'] or '')

                    if mk:
                        if sc not in gst_data:
                            gst_data[sc] = {}
                        if mk in gst_data[sc]:
                            existing = gst_data[sc][mk]
                            existing['taxable_value'] += result['taxable_value']
                            existing['igst']          += result['igst']
                            existing['cgst']          += result['cgst']
                            existing['sgst']          += result['sgst']
                            existing['cess']          += result['cess']
                            existing['total_tax']     += result['total_tax']
                        else:
                            gst_data[sc][mk] = result
                        if result.get('trade_name'):
                            trade_names[sc] = result['trade_name']
                        log.append(f"Parsed: State {sc} / {mk} — taxable ₹{result['taxable_value']:,.2f}")
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

    if not gst_data:
        return {'status': 'error', 'message': 'No valid GSTR 3B PDFs found in the ZIP.'}

    for sc, disp in _sc_display_name.items():
        if sc not in trade_names:
            trade_names[sc] = disp

    # ── Normalize month keys so sales & GSTR3B match ─────────────────
    # Sales might have "April","May" or "Dec-25","Jan-26" etc.
    # GSTR3B uses "Dec-25","Mar-26" from _month_key().
    import re as _re
    _MF = {'january':'jan','february':'feb','march':'mar','april':'apr',
           'may':'may','june':'jun','july':'jul','august':'aug',
           'september':'sep','october':'oct','november':'nov','december':'dec'}

    def _norm_month(raw):
        s = str(raw).strip().lower().rstrip('.')
        # "Dec-25" / "Jan-26"
        m = _re.match(r'^([a-z]+)-?(\d{2,4})$', s)
        if m:
            mon = _MF.get(m.group(1), m.group(1)[:3])
            return f"{mon}-{m.group(2)[-2:]}"
        # "April" / "December" (full, no year)
        if s in _MF:
            return _MF[s]
        if s[:3] in _MF.values():
            return s[:3]
        return s

    # Normalize sales keys
    sales_data = {_norm_month(k): v for k, v in sales_data.items()}
    # Normalize GSTR3B month keys
    for sc in list(gst_data.keys()):
        gst_data[sc] = {_norm_month(mk): d for mk, d in gst_data[sc].items()}

    log.append(f"Normalized: Sales={sorted(sales_data.keys())}")
    if gst_data:
        sample_sc = next(iter(gst_data))
        log.append(f"  GSTR3B[{sample_sc}]={sorted(gst_data[sample_sc].keys())}")

    # ── KEY FIX: Align month keys when sales has no year suffix ──────────────
    # Sales: {"apr": {...}, "may": {...}, ...}   (bare 3-letter, no year)
    # GSTR3B: {"apr-25": {...}, "may-25": {...}, ...}  (with 2-digit year)
    # If ALL sales keys are bare (no "-") but GSTR3B keys all have "-",
    # remap sales keys by prepending the matching year from GSTR3B.
    sales_bare   = all('-' not in k for k in sales_data.keys())
    gstr_with_yr = any('-' in mk for sc_d in gst_data.values() for mk in sc_d.keys())

    if sales_bare and gstr_with_yr:
        # Build a bare→with-year map from GSTR3B keys
        # e.g. "apr" → "apr-25", "jan" → "jan-26"
        bare_to_full = {}
        for sc_d in gst_data.values():
            for mk in sc_d.keys():
                if '-' in mk:
                    bare = mk.split('-')[0]  # "apr-25" → "apr"
                    bare_to_full[bare] = mk   # last one wins (same across states)
        if bare_to_full:
            remapped = {}
            for bare_k, v in sales_data.items():
                full_k = bare_to_full.get(bare_k, bare_k)  # fallback: keep original
                remapped[full_k] = v
            sales_data = remapped
            log.append(f"Month key alignment: Sales remapped to year-suffixed keys: {sorted(sales_data.keys())}")

    # Also handle reverse: GSTR3B bare, Sales with year
    gstr_bare   = all('-' not in mk for sc_d in gst_data.values() for mk in sc_d.keys())
    sales_with_yr = any('-' in k for k in sales_data.keys())
    if gstr_bare and sales_with_yr:
        sales_bare_to_full = {}
        for k in sales_data.keys():
            if '-' in k:
                bare = k.split('-')[0]
                sales_bare_to_full[bare] = k
        for sc in list(gst_data.keys()):
            remapped = {}
            for mk, d in gst_data[sc].items():
                full_k = sales_bare_to_full.get(mk, mk)
                remapped[full_k] = d
            gst_data[sc] = remapped
        log.append(f"Month key alignment: GSTR3B remapped to year-suffixed keys")

    # --- 3. Build output Excel ---
    wb = Workbook()
    ws_out = wb.active
    ws_out.title = "Reconciliation"

    # Styles
    hdr_font = Font(name='Arial', bold=True, size=11)
    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    hdr_font_w = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    sub_fill = PatternFill('solid', fgColor='D6E4F0')
    num_fmt = '#,##0.00'
    thin = Side(style='thin', color='B4B4B4')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    diff_neg_font = Font(name='Arial', color='CC0000', bold=True, size=10)
    diff_pos_font = Font(name='Arial', color='006600', bold=True, size=10)
    data_font = Font(name='Arial', size=10)

    # Title
    entity = trade_names.get(list(gst_data.keys())[0], 'Entity')
    ws_out['A1'] = f"{entity} — GST Reconciliation (Sales Books vs GSTR 3B)"
    ws_out['A1'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')
    ws_out.merge_cells('A1:I1')
    ws_out['A2'] = "Table 3.1: Points A + B + C + E (Excludes D — Reverse Charge)"
    ws_out['A2'].font = Font(name='Arial', italic=True, size=10, color='666666')
    ws_out.merge_cells('A2:I2')

    # Headers (row 4)
    headers = ['Month', 'GST State Code', 'Sale in Books',
               'Sale in GSTR 3B\n(Excl. Tax)', 'Tax Amount\n(IGST+CGST+SGST+Cess)',
               'Difference 1\n(Books − GSTR3B)',
               'Difference 2\n(Books − GSTR3B incl. Tax)',
               'IGST', 'CGST', 'SGST']
    for c, h in enumerate(headers, 1):
        cell = ws_out.cell(row=4, column=c, value=h)
        cell.font = hdr_font_w
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border

    # ── Detect consolidated vs split-by-location mode ──────────────────────
    forced_consolidated = "__consolidated__" in mappings
    consolidated_hint   = mappings.pop("__consolidated__", "") if forced_consolidated else ""
    non_empty_cols = [v.strip() for v in mappings.values() if v and v.strip()]
    is_consolidated = forced_consolidated or (not non_empty_cols or len(set(c.lower() for c in non_empty_cols)) == 1)

    consolidated_col = None
    if is_consolidated:
        if forced_consolidated and consolidated_hint and consolidated_hint != "__auto__":
            matched = next((h for h in col_headers if h.lower()==consolidated_hint.lower()), consolidated_hint)
            non_empty_cols = [matched]
        if non_empty_cols:
            consolidated_col = next((h for h in col_headers if h.lower()==non_empty_cols[0].lower()), non_empty_cols[0])
        else:
            for hdr in col_headers:
                if hdr.lower() in ("month","months","period"): continue
                if any(mdata.get(hdr,0) for mdata in sales_data.values()):
                    consolidated_col = hdr; break
        if consolidated_col:
            log.append(f"✅ Consolidated mode — books column: '{consolidated_col}'")
        else:
            return {"status":"error","message":"Could not find sales data column."}

    all_state_codes = sorted(gst_data.keys())
    valid_mappings = {}
    if is_consolidated:
        for sc in all_state_codes: valid_mappings[sc] = consolidated_col
    else:
        for sc, col_name in mappings.items():
            if not col_name or not col_name.strip(): continue
            matched = next((h for h in col_headers if h.lower()==col_name.lower().strip()), None)
            if matched: valid_mappings[sc] = matched
            else: log.append(f"⚠ Column '{col_name}' not found for state {sc} — skipped")
        if not valid_mappings:
            return {"status":"error","message":"No valid mappings found."}

    log.append(f"Mode: {'Consolidated' if is_consolidated else 'Split'} · States: {', '.join(sorted(valid_mappings.keys()))}")

    # GSTR3B months ONLY — never include sales-only months in output
    gstr_months_set = {mk for sc_d in gst_data.values() for mk in sc_d.keys()}
    all_months = sorted(gstr_months_set, key=lambda x: _month_sort_key(x))

    def _pretty(mk):
        p = mk.split("-"); return f"{p[0].capitalize()}-{p[1]}" if len(p)==2 else mk.capitalize()

    def _sales_val(mk, col):
        v = sales_data.get(mk,{}).get(col) or sales_data.get(mk.split("-")[0],{}).get(col) or 0.0
        return float(v)

    def _sc_lbl(sc):
        d = trade_names.get(sc,""); return f"{sc} — {d}" if d else sc

    row_num = 5
    month_totals = {}

    for month in all_months:
        pretty = _pretty(month)
        first = True

        if is_consolidated:
            gt=tt=it=ct=st=0.0
            for sc in all_state_codes:
                ge=gst_data.get(sc,{}).get(month)
                gv=ge["taxable_value"] if ge else 0.0; tv=ge["total_tax"] if ge else 0.0
                iv=ge["igst"] if ge else 0.0; cv=ge["cgst"] if ge else 0.0; sv=ge["sgst"] if ge else 0.0
                gt+=gv; tt+=tv; it+=iv; ct+=cv; st+=sv
                nil="  ← Nil Return" if gv==0 else ""
                vals=[pretty if first else "",_sc_lbl(sc)+nil,"",gv,tv,"","",iv,cv,sv]
                for c,v in enumerate(vals,1):
                    cell=ws_out.cell(row=row_num,column=c,value=v)
                    cell.font=data_font; cell.border=border
                    if c>=4 and isinstance(v,(int,float)):
                        cell.number_format=num_fmt; cell.alignment=Alignment(horizontal="right")
                    if c<=2: cell.alignment=center
                    if gv==0 and c>=4: cell.font=Font(name="Arial",size=10,color="999999")
                first=False; row_num+=1
            bv=_sales_val(month,consolidated_col)
            d1=bv-gt; d2=bv-(gt+tt)
            sv2=[f"{pretty} — Books vs GSTR3B Total","",bv,gt,tt,d1,d2,it,ct,st]
            for c,v in enumerate(sv2,1):
                cell=ws_out.cell(row=row_num,column=c,value=v)
                cell.font=Font(name="Arial",bold=True,size=10); cell.fill=sub_fill; cell.border=border
                if c>=3 and isinstance(v,(int,float)):
                    cell.number_format=num_fmt; cell.alignment=Alignment(horizontal="right")
                if c in(6,7) and isinstance(v,(int,float)):
                    cc="CC0000" if v<-0.5 else("006600" if v>0.5 else"000000")
                    cell.font=Font(name="Arial",bold=True,size=10,color=cc); cell.fill=sub_fill
            month_totals[month]={"books":bv,"gstr":gt,"tax":tt,"igst":it,"cgst":ct,"sgst":st}
            row_num+=2; continue

        for sc in all_state_codes:
            if sc not in valid_mappings: continue
            bv=_sales_val(month,valid_mappings[sc])
            ge=gst_data.get(sc,{}).get(month)
            gv=ge["taxable_value"] if ge else 0.0; tv=ge["total_tax"] if ge else 0.0
            iv=ge["igst"] if ge else 0.0; cv=ge["cgst"] if ge else 0.0; sv=ge["sgst"] if ge else 0.0
            if month not in month_totals:
                month_totals[month]={"books":0,"gstr":0,"tax":0,"igst":0,"cgst":0,"sgst":0}
            mt=month_totals[month]
            mt["books"]+=bv;mt["gstr"]+=gv;mt["tax"]+=tv;mt["igst"]+=iv;mt["cgst"]+=cv;mt["sgst"]+=sv
            vals=[pretty if first else "",_sc_lbl(sc),bv,gv,tv,bv-gv,bv-(gv+tv),iv,cv,sv]
            for c,v in enumerate(vals,1):
                cell=ws_out.cell(row=row_num,column=c,value=v)
                cell.font=data_font; cell.border=border
                if c>=3 and isinstance(v,(int,float)):
                    cell.number_format=num_fmt; cell.alignment=Alignment(horizontal="right")
                if c in(6,7) and isinstance(v,(int,float)):
                    cell.font=diff_neg_font if v<-0.5 else(diff_pos_font if v>0.5 else data_font)
                if c<=2: cell.alignment=center
            first=False; row_num+=1

        if month in month_totals:
            mt=month_totals[month]
            sv2=[f"{pretty} Total","",mt["books"],mt["gstr"],mt["tax"],
                 mt["books"]-mt["gstr"],mt["books"]-mt["gstr"]-mt["tax"],mt["igst"],mt["cgst"],mt["sgst"]]
            for c,v in enumerate(sv2,1):
                cell=ws_out.cell(row=row_num,column=c,value=v)
                cell.font=Font(name="Arial",bold=True,size=10); cell.fill=sub_fill; cell.border=border
                if c>=3 and isinstance(v,(int,float)):
                    cell.number_format=num_fmt; cell.alignment=Alignment(horizontal="right")
                if c in(6,7) and isinstance(v,(int,float)):
                    cc="CC0000" if v<-0.5 else("006600" if v>0.5 else"000000")
                    cell.font=Font(name="Arial",bold=True,size=10,color=cc); cell.fill=sub_fill
            row_num+=1
        row_num+=1

    grand={"books":0,"gstr":0,"tax":0,"igst":0,"cgst":0,"sgst":0}
    for mt in month_totals.values():
        for k in grand: grand[k]+=mt[k]
    gv2=["GRAND TOTAL","",grand["books"],grand["gstr"],grand["tax"],
         grand["books"]-grand["gstr"],grand["books"]-grand["gstr"]-grand["tax"],
         grand["igst"],grand["cgst"],grand["sgst"]]
    for c,v in enumerate(gv2,1):
        cell=ws_out.cell(row=row_num,column=c,value=v)
        cell.font=Font(name="Arial",bold=True,size=11,color="FFFFFF")
        cell.fill=PatternFill("solid",fgColor="1F4E79"); cell.border=border
        if c>=3 and isinstance(v,(int,float)):
            cell.number_format=num_fmt; cell.alignment=Alignment(horizontal="right")

    widths=[22,26,18,20,20,20,22,15,13,13]
    for i,w in enumerate(widths,1): ws_out.column_dimensions[get_column_letter(i)].width=w
    ws_out.freeze_panes="A5"
    wb.save(output_path)
    log.append(f"Output: {len(month_totals)} months (GSTR3B only) x {len(all_state_codes)} states")
    return {"status":"success","log":log}



def _month_sort_key(ms):
    """Sort key for month strings like 'Apr-25', 'Dec-25', 'Jan-26', or bare 'apr','may'."""
    month_order = {'apr':1,'may':2,'jun':3,'jul':4,'aug':5,'sep':6,
                   'oct':7,'nov':8,'dec':9,'jan':10,'feb':11,'mar':12}
    parts = ms.lower().split('-')
    if len(parts) == 2:
        m = month_order.get(parts[0][:3], 0)
        try: y = int(parts[1])
        except: y = 0
        return (y, m)
    # Bare month name (no year)
    m = month_order.get(ms.lower()[:3], 0)
    if m: return (0, m)
    return (99, 99)


# ── GST routes ────────────────────────────────────────────────────────────────

# ── GST Sales Template Data (base64-encoded) ──────────────────────────────────
_GST_CONS_B64   = "UEsDBBQAAAAIACZBxVxGx01IlQAAAM0AAAAQAAAAZG9jUHJvcHMvYXBwLnhtbE3PTQvCMAwG4L9SdreZih6kDkQ9ip68zy51hbYpbYT67+0EP255ecgboi6JIia2mEXxLuRtMzLHDUDWI/o+y8qhiqHke64x3YGMsRoPpB8eA8OibdeAhTEMOMzit7Dp1C5GZ3XPlkJ3sjpRJsPiWDQ6sScfq9wcChDneiU+ixNLOZcrBf+LU8sVU57mym/8ZAW/B7oXUEsDBBQAAAAIACZBxVx4wQ9+7gAAACsCAAARAAAAZG9jUHJvcHMvY29yZS54bWzNksFKxDAQhl9Fcm8n7eKioZuL4klBcEHxFpLZ3WDThGSk3bc3jbtdRB9AyCUzf775BtLpILSP+Bx9wEgW09Xk+iEJHTbsQBQEQNIHdCrVOTHk5s5Hpyhf4x6C0h9qj9ByvgaHpIwiBTOwCguRyc5ooSMq8vGEN3rBh8/YF5jRgD06HChBUzfA5DwxHKe+gwtghhFGl74LaBZiqf6JLR1gp+SU7JIax7EeVyWXd2jg7enxpaxb2SGRGjTmV8kKOgbcsPPk19Xd/faByZa364rnc73lN4LfiqZ9n11/+F2EnTd2Z/+x8VlQdvDrX8gvUEsDBBQAAAAIACZBxVyZXJwjEAYAAJwnAAATAAAAeGwvdGhlbWUvdGhlbWUxLnhtbO1aW3PaOBR+76/QeGf2bQvGNoG2tBNzaXbbtJmE7U4fhRFYjWx5ZJGEf79HNhDLlg3tkk26mzwELOn7zkVH5+g4efPuLmLohoiU8nhg2S/b1ru3L97gVzIkEUEwGaev8MAKpUxetVppAMM4fckTEsPcgosIS3gUy9Zc4FsaLyPW6rTb3VaEaWyhGEdkYH1eLGhA0FRRWm9fILTlHzP4FctUjWWjARNXQSa5iLTy+WzF/NrePmXP6TodMoFuMBtYIH/Ob6fkTlqI4VTCxMBqZz9Wa8fR0kiAgsl9lAW6Sfaj0xUIMg07Op1YznZ89sTtn4zK2nQ0bRrg4/F4OLbL0otwHATgUbuewp30bL+kQQm0o2nQZNj22q6RpqqNU0/T933f65tonAqNW0/Ta3fd046Jxq3QeA2+8U+Hw66JxqvQdOtpJif9rmuk6RZoQkbj63oSFbXlQNMgAFhwdtbM0gOWXin6dZQa2R273UFc8FjuOYkR/sbFBNZp0hmWNEZynZAFDgA3xNFMUHyvQbaK4MKS0lyQ1s8ptVAaCJrIgfVHgiHF3K/99Ze7yaQzep19Os5rlH9pqwGn7bubz5P8c+jkn6eT101CznC8LAnx+yNbYYcnbjsTcjocZ0J8z/b2kaUlMs/v+QrrTjxnH1aWsF3Pz+SejHIju932WH32T0duI9epwLMi15RGJEWfyC265BE4tUkNMhM/CJ2GmGpQHAKkCTGWoYb4tMasEeATfbe+CMjfjYj3q2+aPVehWEnahPgQRhrinHPmc9Fs+welRtH2Vbzco5dYFQGXGN80qjUsxdZ4lcDxrZw8HRMSzZQLBkGGlyQmEqk5fk1IE/4rpdr+nNNA8JQvJPpKkY9psyOndCbN6DMawUavG3WHaNI8ev4F+Zw1ChyRGx0CZxuzRiGEabvwHq8kjpqtwhErQj5iGTYacrUWgbZxqYRgWhLG0XhO0rQR/FmsNZM+YMjszZF1ztaRDhGSXjdCPmLOi5ARvx6GOEqa7aJxWAT9nl7DScHogstm/bh+htUzbCyO90fUF0rkDyanP+kyNAejmlkJvYRWap+qhzQ+qB4yCgXxuR4+5Xp4CjeWxrxQroJ7Af/R2jfCq/iCwDl/Ln3Ppe+59D2h0rc3I31nwdOLW95GblvE+64x2tc0LihjV3LNyMdUr5Mp2DmfwOz9aD6e8e362SSEr5pZLSMWkEuBs0EkuPyLyvAqxAnoZFslCctU02U3ihKeQhtu6VP1SpXX5a+5KLg8W+Tpr6F0PizP+Txf57TNCzNDt3JL6raUvrUmOEr0scxwTh7LDDtnPJIdtnegHTX79l125COlMFOXQ7gaQr4Dbbqd3Do4npiRuQrTUpBvw/npxXga4jnZBLl9mFdt59jR0fvnwVGwo+88lh3HiPKiIe6hhpjPw0OHeXtfmGeVxlA0FG1srCQsRrdguNfxLBTgZGAtoAeDr1EC8lJVYDFbxgMrkKJ8TIxF6HDnl1xf49GS49umZbVuryl3GW0iUjnCaZgTZ6vK3mWxwVUdz1Vb8rC+aj20FU7P/lmtyJ8MEU4WCxJIY5QXpkqi8xlTvucrScRVOL9FM7YSlxi84+bHcU5TuBJ2tg8CMrm7Oal6ZTFnpvLfLQwJLFuIWRLiTV3t1eebnK56Inb6l3fBYPL9cMlHD+U751/0XUOufvbd4/pukztITJx5xREBdEUCI5UcBhYXMuRQ7pKQBhMBzZTJRPACgmSmHICY+gu98gy5KRXOrT45f0Usg4ZOXtIlEhSKsAwFIRdy4+/vk2p3jNf6LIFthFQyZNUXykOJwT0zckPYVCXzrtomC4Xb4lTNuxq+JmBLw3punS0n/9te1D20Fz1G86OZ4B6zh3OberjCRaz/WNYe+TLfOXDbOt4DXuYTLEOkfsF9ioqAEativrqvT/klnDu0e/GBIJv81tuk9t3gDHzUq1qlZCsRP0sHfB+SBmOMW/Q0X48UYq2msa3G2jEMeYBY8wyhZjjfh0WaGjPVi6w5jQpvQdVA5T/b1A1o9g00HJEFXjGZtjaj5E4KPNz+7w2wwsSO4e2LvwFQSwMEFAAAAAgAJkHFXGTPnIzcAgAARQkAABgAAAB4bC93b3Jrc2hlZXRzL3NoZWV0MS54bWyFlmtvmzAUhv+KxaRq+xIuubYlSE2r7iL1oqbdPjvkAFaNzWyTtP9+tqGQbVw+JNjG7/scG44P4ZGLV5kBKPSWUybXTqZUceG6Ms4gx3LCC2D6TsJFjpXuitSVhQC8t6KcuoHnLdwcE+ZEoR17FFHIS0UJg0eBZJnnWLxvgPLj2vGdj4EnkmbKDLhRWOAUtqBeikehe27jsic5MEk4QwKStXPlX2z8hRHYGT8JHOVJG5ml7Dh/NZ3v+7XjmYiAQqyMBdaXA1wDpcZJx/G7NnUaphGetj/cb+3i9WJ2WMI1p7/IXmVrZ+WgPSS4pOqJH79BvaC58Ys5lfYfHau5Om4Ul1LxvBbrCHLCqit+qzfiRBB4PYKgFgQ27gpko7zBCkeh4EckzGztZhp2qVatgyPMPJWtEvou0ToVfd0+oyeIOYsJJdhu1NmnVeAHl+iaM8kp2WMFe7TFFCSqNiJ0lSYbvRvrnyY22KDBBhYb9GDPPvnBKvBml+iWUIqkdU9IWgp9JUyzaZkztJmgG47uH55RnGGWAoqr8Uy/fyDkZCCSaRPJ1EYy7YnkjjOV/e1jZZth2TNXmFa7MhDErAliZt1mPW5XhSC0K4hKNrcyk2D/EeYNYT5IuMPvXf7zMf9F478Y9P9RMugCLMYAywawHAHQzhUsxwCrBrAafgZlqtOtC7EaQ5w3iPNBxBYKBfkORBflfIzie21Oe4Och1jxHkqtHMKcHB3+IOaeH3pXU0uHOO1Z4QeDnBuI+znBKKc9Cfzp8CuGWakrUydmOoppc90fTvZb2IlezmjC+23G+2MpL+LOk80fzXq/TXu/yuDF0DnYCal0Jr2TaPty93kz09V7/iV0kyg86JmHU657UsFyEKmt01If9yVTVVVpRutvgeBiYyvgv+PmG8FWxtam+sDQ25ESJhGFRFt6k6Vev6hqdtVRvLB1cseVrrq2WdUZM0HfTzhXHx0DaL6coj9QSwMEFAAAAAgAJkHFXHNazGUqAwAAFBAAAA0AAAB4bC9zdHlsZXMueG1s3Vhhb9owEP0rUX7AQghkZCJIJQVp0jZVaj/sqyEOWHLiLDEV9NfPZ4ckgI/RtVOlBaHYd37vns9nGzGt5YHTxy2l0tnnvKhjdytl+cXz6vWW5qT+JEpaKE8mqpxI1a02Xl1WlKQ1gHLuDQeD0MsJK9zZtNjly1zWzlrsChm7A9ebTTNRdJbQNQY1lOTUeSY8dhPC2apieizJGT8Y8xAMa8FF5UglhcauD5b6xbh90wOVDU/OClGB0TMRrsRZNRxdiGqzUnoH/iK4Gy9P4gxvo2QYZTj6PJrM+5TRG0Uu9XOZjD9TtuMHr5dgxepXrTgY5+0ij1xjmE1LIiWtiqXqaIw2Xricpv10KNUqbypy8Idj92ZALThLIeQm6Wfpfr64Wy40TQ/6RtKuPqyk+qXSsRJVSqs2IUP3aJpNOc2kgldss4W3FCXkWUgpctVIGdmIguhsHRF9pKP3a+zKrd5vJ2WRzO/HC1MIMLSJcSNCj9VybgSokUfdNyLM4N7EmobK15py/ggkP7M2ab6i2meOOVK+pnCaOFBtx6bKdNM0NKYDgfpshrtHO/krWqdkz0LOd2oGhe7/2glJHyqasb3u77M2Psbud+zDM3ZSlvxwx9mmyKmZ+80BZ1NyxDlbUbEXFQ226VoZqDkO9xkuaohP+eNEBZ2ooC/K/0hRIyRTr1BwXhyjf8re0z5GSu99tL8Pu9ds1N5pcHIWtFYHrqrY/QG/SHhH4ax2jEtWNL0tS1NaXBwJil6SlfrJc8Kvxqc0Izsun1pn7Hbt7zRluzxqRz3AtJpRXfsbnKF+2N7XKhYrUrqnadJ01aF4cp2YBwDnnu6Ov/RgGOOze8CHxcEUYBiDwuL8T/OZoPMxPkzbxOqZoJgJijEomyfRHyyOHROpxz7TKAqCMMQymiRWBQmWtzCEr50N0wYILA5Eel2u8dXGK+R6HWBreq1CsJnilYjNFM81eOx5A0QU2VcbiwMIbBWw2oH49jhQU3ZMEMCqYtqwHYx7ogjzQC3aazQMkeyE8LGvD7ZLgiCK7B7w2RUEAeaB3Yh7MAWgAfMEgb4Hz+4j73hPed3/ALPfUEsDBBQAAAAIACZBxVyXirscwAAAABMCAAALAAAAX3JlbHMvLnJlbHOdkrluwzAMQH/F0J4wB9AhiDNl8RYE+QFWog/YEgWKRZ2/r9qlcZALGXk9PBLcHmlA7TiktoupGP0QUmla1bgBSLYlj2nOkUKu1CweNYfSQETbY0OwWiw+QC4ZZre9ZBanc6RXiFzXnaU92y9PQW+ArzpMcUJpSEszDvDN0n8y9/MMNUXlSiOVWxp40+X+duBJ0aEiWBaaRcnToh2lfx3H9pDT6a9jIrR6W+j5cWhUCo7cYyWMcWK0/jWCyQ/sfgBQSwMEFAAAAAgAJkHFXEs11FQ5AQAAKgIAAA8AAAB4bC93b3JrYm9vay54bWyNUdFuwjAM/JUqH7AWtCENUV6GtiFNGxoT76F1qUUSV44Lg6+f26oa0l72lNzZutxdFmfi457omHx7F2JuapFmnqaxqMHbeEcNBJ1UxN6KQj6ksWGwZawBxLt0mmWz1FsMZrkYtTac3gISKAQpKNkRO4Rz/J13MDlhxD06lEtu+rsDk3gM6PEKZW4yk8Sazq/EeKUg1m0LJudyMxkGO2DB4g+97Ux+2X3sGbH7T6tGcjPLVLBCjtJv9PpWPZ5AlwfUCj2jE+CVFXhhahsMh05GU6Q3MfoexnMocc7/qZGqCgtYUdF6CDL0yOA6gyHW2ESTBOshN1vrICbb1nvLly6XPrQuh4yi5m4a4znqgNflYHP0VkKFAcp3lYvKa0/FhpPu6HWm9w+TR+2jde5JuY/wRrYco47ftPwBUEsDBBQAAAAIACZBxVwkHpuirQAAAPgBAAAaAAAAeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJlbHO1kT0OgzAMha8S5QA1UKlDBUxdWCsuEAXzIxISxa4Kty+FAZA6dGGyni1/78lOn2gUd26gtvMkRmsGymTL7O8ApFu0ii7O4zBPahes4lmGBrzSvWoQkii6QdgzZJ7umaKcPP5DdHXdaXw4/bI48A8wvF3oqUVkKUoVGuRMwmi2NsFS4stMlqKoMhmKKpZwWiDiySBtaVZ9sE9OtOd5Fzf3Ra7N4wmu3wxweHT+AVBLAwQUAAAACAAmQcVcZZB5khkBAADPAwAAEwAAAFtDb250ZW50X1R5cGVzXS54bWytk01OwzAQha8SZVslLixYoKYbYAtdcAFjTxqr/pNnWtLbM07aSqASFYVNrHjevM+el6zejxGw6J312JQdUXwUAlUHTmIdIniutCE5SfyatiJKtZNbEPfL5YNQwRN4qih7lOvVM7Ryb6l46XkbTfBNmcBiWTyNwsxqShmjNUoS18XB6x+U6kSouXPQYGciLlhQiquEXPkdcOp7O0BKRkOxkYlepWOV6K1AOlrAetriyhlD2xoFOqi945YaYwKpsQMgZ+vRdDFNJp4wjM+72fzBZgrIyk0KETmxBH/HnSPJ3VVkI0hkpq94IbL17PtBTluDvpHN4/0MaTfkgWJY5s/4e8YX/xvO8RHC7r8/sbzWThp/5ovhP15/AVBLAQIUAxQAAAAIACZBxVxGx01IlQAAAM0AAAAQAAAAAAAAAAAAAACAAQAAAABkb2NQcm9wcy9hcHAueG1sUEsBAhQDFAAAAAgAJkHFXHjBD37uAAAAKwIAABEAAAAAAAAAAAAAAIABwwAAAGRvY1Byb3BzL2NvcmUueG1sUEsBAhQDFAAAAAgAJkHFXJlcnCMQBgAAnCcAABMAAAAAAAAAAAAAAIAB4AEAAHhsL3RoZW1lL3RoZW1lMS54bWxQSwECFAMUAAAACAAmQcVcZM+cjNwCAABFCQAAGAAAAAAAAAAAAAAAgIEhCAAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1sUEsBAhQDFAAAAAgAJkHFXHNazGUqAwAAFBAAAA0AAAAAAAAAAAAAAIABMwsAAHhsL3N0eWxlcy54bWxQSwECFAMUAAAACAAmQcVcl4q7HMAAAAATAgAACwAAAAAAAAAAAAAAgAGIDgAAX3JlbHMvLnJlbHNQSwECFAMUAAAACAAmQcVcSzXUVDkBAAAqAgAADwAAAAAAAAAAAAAAgAFxDwAAeGwvd29ya2Jvb2sueG1sUEsBAhQDFAAAAAgAJkHFXCQem6KtAAAA+AEAABoAAAAAAAAAAAAAAIAB1xAAAHhsL19yZWxzL3dvcmtib29rLnhtbC5yZWxzUEsBAhQDFAAAAAgAJkHFXGWQeZIZAQAAzwMAABMAAAAAAAAAAAAAAIABvBEAAFtDb250ZW50X1R5cGVzXS54bWxQSwUGAAAAAAkACQA+AgAABhMAAAAA"
_GST_BRANCH_B64 = "UEsDBBQAAAAIACZBxVxGx01IlQAAAM0AAAAQAAAAZG9jUHJvcHMvYXBwLnhtbE3PTQvCMAwG4L9SdreZih6kDkQ9ip68zy51hbYpbYT67+0EP255ecgboi6JIia2mEXxLuRtMzLHDUDWI/o+y8qhiqHke64x3YGMsRoPpB8eA8OibdeAhTEMOMzit7Dp1C5GZ3XPlkJ3sjpRJsPiWDQ6sScfq9wcChDneiU+ixNLOZcrBf+LU8sVU57mym/8ZAW/B7oXUEsDBBQAAAAIACZBxVx4wQ9+7gAAACsCAAARAAAAZG9jUHJvcHMvY29yZS54bWzNksFKxDAQhl9Fcm8n7eKioZuL4klBcEHxFpLZ3WDThGSk3bc3jbtdRB9AyCUzf775BtLpILSP+Bx9wEgW09Xk+iEJHTbsQBQEQNIHdCrVOTHk5s5Hpyhf4x6C0h9qj9ByvgaHpIwiBTOwCguRyc5ooSMq8vGEN3rBh8/YF5jRgD06HChBUzfA5DwxHKe+gwtghhFGl74LaBZiqf6JLR1gp+SU7JIax7EeVyWXd2jg7enxpaxb2SGRGjTmV8kKOgbcsPPk19Xd/faByZa364rnc73lN4LfiqZ9n11/+F2EnTd2Z/+x8VlQdvDrX8gvUEsDBBQAAAAIACZBxVyZXJwjEAYAAJwnAAATAAAAeGwvdGhlbWUvdGhlbWUxLnhtbO1aW3PaOBR+76/QeGf2bQvGNoG2tBNzaXbbtJmE7U4fhRFYjWx5ZJGEf79HNhDLlg3tkk26mzwELOn7zkVH5+g4efPuLmLohoiU8nhg2S/b1ru3L97gVzIkEUEwGaev8MAKpUxetVppAMM4fckTEsPcgosIS3gUy9Zc4FsaLyPW6rTb3VaEaWyhGEdkYH1eLGhA0FRRWm9fILTlHzP4FctUjWWjARNXQSa5iLTy+WzF/NrePmXP6TodMoFuMBtYIH/Ob6fkTlqI4VTCxMBqZz9Wa8fR0kiAgsl9lAW6Sfaj0xUIMg07Op1YznZ89sTtn4zK2nQ0bRrg4/F4OLbL0otwHATgUbuewp30bL+kQQm0o2nQZNj22q6RpqqNU0/T933f65tonAqNW0/Ta3fd046Jxq3QeA2+8U+Hw66JxqvQdOtpJif9rmuk6RZoQkbj63oSFbXlQNMgAFhwdtbM0gOWXin6dZQa2R273UFc8FjuOYkR/sbFBNZp0hmWNEZynZAFDgA3xNFMUHyvQbaK4MKS0lyQ1s8ptVAaCJrIgfVHgiHF3K/99Ze7yaQzep19Os5rlH9pqwGn7bubz5P8c+jkn6eT101CznC8LAnx+yNbYYcnbjsTcjocZ0J8z/b2kaUlMs/v+QrrTjxnH1aWsF3Pz+SejHIju932WH32T0duI9epwLMi15RGJEWfyC265BE4tUkNMhM/CJ2GmGpQHAKkCTGWoYb4tMasEeATfbe+CMjfjYj3q2+aPVehWEnahPgQRhrinHPmc9Fs+welRtH2Vbzco5dYFQGXGN80qjUsxdZ4lcDxrZw8HRMSzZQLBkGGlyQmEqk5fk1IE/4rpdr+nNNA8JQvJPpKkY9psyOndCbN6DMawUavG3WHaNI8ev4F+Zw1ChyRGx0CZxuzRiGEabvwHq8kjpqtwhErQj5iGTYacrUWgbZxqYRgWhLG0XhO0rQR/FmsNZM+YMjszZF1ztaRDhGSXjdCPmLOi5ARvx6GOEqa7aJxWAT9nl7DScHogstm/bh+htUzbCyO90fUF0rkDyanP+kyNAejmlkJvYRWap+qhzQ+qB4yCgXxuR4+5Xp4CjeWxrxQroJ7Af/R2jfCq/iCwDl/Ln3Ppe+59D2h0rc3I31nwdOLW95GblvE+64x2tc0LihjV3LNyMdUr5Mp2DmfwOz9aD6e8e362SSEr5pZLSMWkEuBs0EkuPyLyvAqxAnoZFslCctU02U3ihKeQhtu6VP1SpXX5a+5KLg8W+Tpr6F0PizP+Txf57TNCzNDt3JL6raUvrUmOEr0scxwTh7LDDtnPJIdtnegHTX79l125COlMFOXQ7gaQr4Dbbqd3Do4npiRuQrTUpBvw/npxXga4jnZBLl9mFdt59jR0fvnwVGwo+88lh3HiPKiIe6hhpjPw0OHeXtfmGeVxlA0FG1srCQsRrdguNfxLBTgZGAtoAeDr1EC8lJVYDFbxgMrkKJ8TIxF6HDnl1xf49GS49umZbVuryl3GW0iUjnCaZgTZ6vK3mWxwVUdz1Vb8rC+aj20FU7P/lmtyJ8MEU4WCxJIY5QXpkqi8xlTvucrScRVOL9FM7YSlxi84+bHcU5TuBJ2tg8CMrm7Oal6ZTFnpvLfLQwJLFuIWRLiTV3t1eebnK56Inb6l3fBYPL9cMlHD+U751/0XUOufvbd4/pukztITJx5xREBdEUCI5UcBhYXMuRQ7pKQBhMBzZTJRPACgmSmHICY+gu98gy5KRXOrT45f0Usg4ZOXtIlEhSKsAwFIRdy4+/vk2p3jNf6LIFthFQyZNUXykOJwT0zckPYVCXzrtomC4Xb4lTNuxq+JmBLw3punS0n/9te1D20Fz1G86OZ4B6zh3OberjCRaz/WNYe+TLfOXDbOt4DXuYTLEOkfsF9ioqAEativrqvT/klnDu0e/GBIJv81tuk9t3gDHzUq1qlZCsRP0sHfB+SBmOMW/Q0X48UYq2msa3G2jEMeYBY8wyhZjjfh0WaGjPVi6w5jQpvQdVA5T/b1A1o9g00HJEFXjGZtjaj5E4KPNz+7w2wwsSO4e2LvwFQSwMEFAAAAAgAJkHFXPiPksM/BAAA2REAABgAAAB4bC93b3Jrc2hlZXRzL3NoZWV0MS54bWyNmG9zmzgQxr+Kxp3p9F5csABjTG3P1EFc72bSy8Rt77WMZcMVECdE3Hz7W/4EU4fFfpEYePZZ6bfCeMXyJNWPIhJCk59pkhWrSaR17hlGEUYi5cWdzEUGykGqlGs4VUejyJXg+9qUJoY5nTpGyuNssl7W1x7VeilLncSZeFSkKNOUq5eNSORpNaGT1wtP8THS1QVjvcz5UWyF/pY/Kjgzuiz7OBVZEcuMKHFYTT5RL6BOZagjvsfiVPSOSYWyk/JHdfLnfjWZVjMSiQh1lYLDx7O4F0lSZYJ5/NcmnXRjVsb+8Wv2oIYHmB0vxL1M/on3OlpN3AnZiwMvE/0kT59FCzSr8oUyKer/5NTEwrxJWBZapq0ZZpDGWfPJf7aF6BnMKWIwW4N5q8FqDdatBrs12LcaZq1hdqvBaQ31YhpNsepK+1zz9VLJE1FVNGSrDurlqt1Q4Dir7qytVqDG4NPrP7ZfyZMIZRbGSczrxX7/zjWp+ZFsFM/CyNhqrsXvp7gQZMsTUZBmRZeGhuGrJEYIfzBsN7bZjW3WY5vI2O/fUdM1p/ZHmEGe8FAQoCnTjETwFRGqgFLoiLzIUlU3YMkTsmumVFRTIhlPRXFHHnhOBA8joiWJdUGA6AtpIkK5FyTOiI4EqDK5G5k0rHR1C1pd2bubsqOxahoLoXmQmY5+HaC2bcZtj2X2L9+RD1PrtwHz/bj5MzwMeMbB7Qy5/XG3L5IoBu98yMuu0PKIK15EWsHo5mCGYDzDV6l5MrIgdld3u85jI3k+5SpOhure2Ga1rXrAnouKKj6qMFQJGsWBa4f19tvDh43tMRsKclgvnyHueYBt1rHNRtke+MsQ2QwlQxUfVRiqBLNLspnHZuNkTkfmjJL9VWZiCM1B0VDFRxWGKoFzieZ4zBlHm3do8ytoyeCqzVE0VPFRhaFKML9Em3tsPo7mdmju+HetPMKzcQjOReFQxUcVhiqBewnneswdh1t0cItRuK3ItUh3Qg3xLVA+VPFRhaFKsLjkW3hsMc5Hp+ff+uko4d+hlghf6xwCxCUflxguBa3Ug6RTDwxXMHstDR3F/CKf0XVsrYOcqOTjEsOloJX6nBQ46RXOc/tEzVFOX4Q4p4lzopKPSwyXglbqc5rAaV7htM6c1vgDlWclNDqDmBaOiUo+LjFcClqpj2kBpnUF89zH0PFGJhA7hXLizQwu+bjEcCmgbxoaCh0NvdLS0HNPQ681NSocbJQp3tfgko9LDJcC+qa3odDc0CvdDT23N7TpIOa3d7ctZONz++3iZmDclvsy+t727rFo/020b3s+Fs3eRDN7sAJtwS6jA9sL0HoZvR1pKtSxfndQwIaszHSzQeyutu8nTC+od+WX16v3FvVO95ymeekBt9ExzgqSiAOknN7NYUVVs2VrTrTM633vTmrYztWHzcayCgD9IKV+PakG6N7mrP8HUEsDBBQAAAAIACZBxVyOgk7SRAMAANQQAAANAAAAeGwvc3R5bGVzLnhtbN1YbW+bMBD+K4gfMAIkLExJpIQGadI2VWo/7KsTTGLJvAycLumvn88mQBJflq7tJo2owr7z89zj83GgTmpx4PRhS6mw9hnP66m9FaL85Dj1ekszUn8oSppLT1pUGRFyWm2cuqwoSWoAZdzxBoPAyQjL7dkk32VxJmprXexyMbUHtjObpEXeWQJbG+RSklHrifCpHRHOVhVTa0nG+EGbPTCsC15UlpBS6NR2wVI/a7erZ6Cy4clYXlRgdHSEK3FWDUcXotqspN6Bu/Tno/gkjncbJcMog+HH4XjRpwxfKTJW12Uyfk/Zrh+8XIIRq2615GCct4c8srVhNimJELTKYzlRGGW8cFnN+PFQylPeVOTgeiP7ZkBdcJZAyE3Uz9LdYjmPl4qmB30laVcfb0gaj+N5HKGk6iZzvCqqhFZtlj37aJpNOE2FhFdss4W7KEo4vEKIIpODhJFNkRN1BEdEH2mpJjC1xVY9xCe1Fi3uRktdXbC0iXEjQq1Vcm4EyJVH3Tci9OLexpqBzNeacv4AJN/TNmmupNqnlu5TnxNoURaU8HEoM90MNY2eQKA+m+bu0YZ/RGuV7KkQi53cQa7mP3aFoPcVTdlezfdpGx9jdzt274ydlCU/zDnb5BnVe7854GxCjjhrW1TsWUaDZ38tDVT32H2Ki/LwLf87UX4nyu+Lct9TlPWzIuUj3YumiV5VOETS9gI555Uy/Fvso459+EbsAyO7967aX8HuND2h13hO2k5rteBVO7W/wRcV7yis1Y5xwfJmtmVJQvOL7iPpBVnJT7YTfrk+oSnZcfHYOqd2N/5KE7bLwnbVPWyrWdWNv0C7doP2e0PGYnlC9zSJmqnsvydvLn0B4NzTfaNcejCM9pk94MPiYAowjEZhcf6n/YzR/Wgfpm1s9IxRzBjFaJTJE6kfFseMCeVl3mkY+n4QYBmNIqOCCMtbEMCfmQ3TBggsDkR6Wa7x08Yr5HodYGd6rUKwneKViO0UzzV4zHkDRBiaTxuLAwjsFLDagfjmOFBTZozvw6li2rAnGPeEIeaBWjTXaBAg2QngZz4f7Cnx/TA0e8BnVuD7mAeeRtyDKQANmMf31Xvw7H3kHN9TTvd/jNkvUEsDBBQAAAAIACZBxVyXirscwAAAABMCAAALAAAAX3JlbHMvLnJlbHOdkrluwzAMQH/F0J4wB9AhiDNl8RYE+QFWog/YEgWKRZ2/r9qlcZALGXk9PBLcHmlA7TiktoupGP0QUmla1bgBSLYlj2nOkUKu1CweNYfSQETbY0OwWiw+QC4ZZre9ZBanc6RXiFzXnaU92y9PQW+ArzpMcUJpSEszDvDN0n8y9/MMNUXlSiOVWxp40+X+duBJ0aEiWBaaRcnToh2lfx3H9pDT6a9jIrR6W+j5cWhUCo7cYyWMcWK0/jWCyQ/sfgBQSwMEFAAAAAgAJkHFXEs11FQ5AQAAKgIAAA8AAAB4bC93b3JrYm9vay54bWyNUdFuwjAM/JUqH7AWtCENUV6GtiFNGxoT76F1qUUSV44Lg6+f26oa0l72lNzZutxdFmfi457omHx7F2JuapFmnqaxqMHbeEcNBJ1UxN6KQj6ksWGwZawBxLt0mmWz1FsMZrkYtTac3gISKAQpKNkRO4Rz/J13MDlhxD06lEtu+rsDk3gM6PEKZW4yk8Sazq/EeKUg1m0LJudyMxkGO2DB4g+97Ux+2X3sGbH7T6tGcjPLVLBCjtJv9PpWPZ5AlwfUCj2jE+CVFXhhahsMh05GU6Q3MfoexnMocc7/qZGqCgtYUdF6CDL0yOA6gyHW2ESTBOshN1vrICbb1nvLly6XPrQuh4yi5m4a4znqgNflYHP0VkKFAcp3lYvKa0/FhpPu6HWm9w+TR+2jde5JuY/wRrYco47ftPwBUEsDBBQAAAAIACZBxVwkHpuirQAAAPgBAAAaAAAAeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJlbHO1kT0OgzAMha8S5QA1UKlDBUxdWCsuEAXzIxISxa4Kty+FAZA6dGGyni1/78lOn2gUd26gtvMkRmsGymTL7O8ApFu0ii7O4zBPahes4lmGBrzSvWoQkii6QdgzZJ7umaKcPP5DdHXdaXw4/bI48A8wvF3oqUVkKUoVGuRMwmi2NsFS4stMlqKoMhmKKpZwWiDiySBtaVZ9sE9OtOd5Fzf3Ra7N4wmu3wxweHT+AVBLAwQUAAAACAAmQcVcZZB5khkBAADPAwAAEwAAAFtDb250ZW50X1R5cGVzXS54bWytk01OwzAQha8SZVslLixYoKYbYAtdcAFjTxqr/pNnWtLbM07aSqASFYVNrHjevM+el6zejxGw6J312JQdUXwUAlUHTmIdIniutCE5SfyatiJKtZNbEPfL5YNQwRN4qih7lOvVM7Ryb6l46XkbTfBNmcBiWTyNwsxqShmjNUoS18XB6x+U6kSouXPQYGciLlhQiquEXPkdcOp7O0BKRkOxkYlepWOV6K1AOlrAetriyhlD2xoFOqi945YaYwKpsQMgZ+vRdDFNJp4wjM+72fzBZgrIyk0KETmxBH/HnSPJ3VVkI0hkpq94IbL17PtBTluDvpHN4/0MaTfkgWJY5s/4e8YX/xvO8RHC7r8/sbzWThp/5ovhP15/AVBLAQIUAxQAAAAIACZBxVxGx01IlQAAAM0AAAAQAAAAAAAAAAAAAACAAQAAAABkb2NQcm9wcy9hcHAueG1sUEsBAhQDFAAAAAgAJkHFXHjBD37uAAAAKwIAABEAAAAAAAAAAAAAAIABwwAAAGRvY1Byb3BzL2NvcmUueG1sUEsBAhQDFAAAAAgAJkHFXJlcnCMQBgAAnCcAABMAAAAAAAAAAAAAAIAB4AEAAHhsL3RoZW1lL3RoZW1lMS54bWxQSwECFAMUAAAACAAmQcVc+I+Swz8EAADZEQAAGAAAAAAAAAAAAAAAgIEhCAAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1sUEsBAhQDFAAAAAgAJkHFXI6CTtJEAwAA1BAAAA0AAAAAAAAAAAAAAIABlgwAAHhsL3N0eWxlcy54bWxQSwECFAMUAAAACAAmQcVcl4q7HMAAAAATAgAACwAAAAAAAAAAAAAAgAEFEAAAX3JlbHMvLnJlbHNQSwECFAMUAAAACAAmQcVcSzXUVDkBAAAqAgAADwAAAAAAAAAAAAAAgAHuEAAAeGwvd29ya2Jvb2sueG1sUEsBAhQDFAAAAAgAJkHFXCQem6KtAAAA+AEAABoAAAAAAAAAAAAAAIABVBIAAHhsL19yZWxzL3dvcmtib29rLnhtbC5yZWxzUEsBAhQDFAAAAAgAJkHFXGWQeZIZAQAAzwMAABMAAAAAAAAAAAAAAIABORMAAFtDb250ZW50X1R5cGVzXS54bWxQSwUGAAAAAAkACQA+AgAAgxQAAAAA"

_TSHAPE_SAMPLE_B64 = "UEsDBBQABgAIAAAAIQBg4Ct7wQEAAPQLAAATAAgCW0NvbnRlbnRfVHlwZXNdLnhtbCCiBAIooAACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADMltFO2zAUhu+R9g6Rb6fGLdsYm5pywdjVNJBgD2Ds08aqY1s+B2jfficGqmoqrSoizTeJEtv//52TyP6nF6vOVY+Q0AbfiEk9FhV4HYz1i0b8ufs5OhcVkvJGueChEWtAcTH7cDK9W0fAild7bERLFL9LibqFTmEdIngemYfUKeLHtJBR6aVagDwdj8+kDp7A04h6DTGb/oC5enBUXa349TPJvfWiunye11s1QsXorFbEoPLRm39MRmE+txpM0A8dS9cYEyiDLQB1ro7JsmO6BSIuDIXc6ZnA4XGmL1XVvDKDYWsjfuTS33DoR96u6mXdNX+OZA1UNyrRb9Vx7XLl5FNIy/sQlvV+kWNbk1tUd8r6V+49/nkyynybDAzS15eFj+Q4LYTjUyEcnwvh+FIIx1khHF8L4TgvhONbIRyTcSkgpeyok/+1pQKf/8kr98v6Jcrtp6Fbs6194LwhzjQg8/X9FFnmgCHS2gEOfcxn0UPOrUpgbonT0mJwgG3tAxxaOX3ZciQZuAkb3X3+HCFvUojIKTXB8QCvkbBfPYosBIksbELhrnC1ceSI++6K+z/bGzA7vGXO7LO/AAAA//8DAFBLAwQUAAYACAAAACEAtVUwI/QAAABMAgAACwAIAl9yZWxzLy5yZWxzIKIEAiigAAIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAKySTU/DMAyG70j8h8j31d2QEEJLd0FIuyFUfoBJ3A+1jaMkG92/JxwQVBqDA0d/vX78ytvdPI3qyCH24jSsixIUOyO2d62Gl/pxdQcqJnKWRnGs4cQRdtX11faZR0p5KHa9jyqruKihS8nfI0bT8USxEM8uVxoJE6UchhY9mYFaxk1Z3mL4rgHVQlPtrYawtzeg6pPPm3/XlqbpDT+IOUzs0pkVyHNiZ9mufMhsIfX5GlVTaDlpsGKecjoieV9kbMDzRJu/E/18LU6cyFIiNBL4Ms9HxyWg9X9atDTxy515xDcJw6vI8MmCix+o3gEAAP//AwBQSwMEFAAGAAgAAAAhAHbF34sqBAAA0wwAAA8AAAB4bC93b3JrYm9vay54bWysV2GP4jYQ/V6p/yG1UPcTJE5CgHTDiV3gitT2ELu9U6WTViYxi7WJnTpmYXW6/95xSMCErbRdiiDBsf3mzcybwVx/2GWp9UxlwQSPEO44yKI8FgnjjxH6837a7iOrUIQnJBWcRuiFFujD8McfrrdCPi2FeLIAgBcRWiuVh7ZdxGuakaIjcsphZiVkRhQM5aNd5JKSpFhTqrLUdh0nsDPCONojhPItGGK1YjEdi3iTUa72IJKmRAH9Ys3yokbL4rfAZUQ+bfJ2LLIcIJYsZeqlBEVWFoezRy4kWabg9g53rZ2EdwAf7MDFrS3B1JmpjMVSFGKlOgBt70mf+Y8dG+OTEOzOY/A2JN+W9JnpHB5YyeCdrIIDVnAEw87FaBikVWolhOC9E6174Oai4fWKpfTzXroWyfM/SKYzlSIrJYWaJEzRJEI9GIotPT4Ar+Qmv9mwFGY9x3X7yB4e5DyXVkJXZJOqexByDQ+VEQQDt6tXgjBGqaKSE0VvBVegw8qvSzVXYt+uBSjcWtC/N0xSKCzQF/gKVxKHZFnMiVpbG5lGaBx+HYstTwUU1VdDjeRc+v9BjyTWTtrg5Z7J/nvTYyAkw1pzcyUt+D4b/wZxvyPPkAXIdVIV6QzCjL0HHssQP3yb+o478buj9q3T89t+d+y2b7yboN2f4JtgMr6dTgP8HZyRQRgLslHrKsEaOkI+ZPNs6neyq2ewE25YcqTxzalebX1vXOq579ph3co+M7otjlLQQ2v3hfFEbCPUxg60wpfT4bac/MIStY6Q67ldENf+2a+UPa6BMXZ9/VCR5UI3qQj1PfCBxIo903uyhAW6LKSreUfohO94z3cKr7a+nPC1DcJlSwXi5d3iZRksoROW4zL8IPlQ48tZgrW35sr8Z5Llv0DVHJZDazssd5vLuVC0sJQAH2Kx4erEjmfs9Jo7Y5IzRUxD5nL/Xw2duKKrt3alLEfTlQO3c58gd4d9QdPUfD4x3B+Uv3gKVLxmSUKhpR529po7p2xHE2tUFFQV1m3H+kt2zEBiY2/ZZky2J3vnzb3Y9HTQtDumirD0JPa+mWSnueHj4tPdnTVffJrO7k2GoL6jNM60MR250CBNJYF44TjwanBwqRW7ViPdlU0yXdAVlXCmoKDQs2eG8VIv9mu7oCEzThPd3wHDGFVaf9ilPOvMJePqYQRnDN3xY5Le6ZOGlr+Dhsvip9aohcPWtOXja9vAeA+gh4aVmCvUjy2vdzHqAA2rtFaokxb2BhfD9tDw6hWZXlVGZq2ef7GNftPGXs5HG17/YhvQKIdXpoxr9GnLvTz4cJ64OmttR/5d72L+vmlhWdTYkOTAvRgcjghH+lX3O1roXx4fDAYauP9L3AM0hO570OJZKMxSheKHso7hiKZv5Q/bADtu2Rzt+v/I8B8AAAD//wMAUEsDBBQABgAIAAAAIQBJc4gtZwEAAPMJAAAaAAgBeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJlbHMgogQBKKAAAQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAC8ls1qwzAMx++DvUPwfXXS744mPWwMCjtt3QMYR01CEjtY3ta+/UwHSQtF7cH4YpCMpR9/yZbXm0PbRD9gsNIqZckoZhEoqfNKFSn72r09LVmEVqhcNFpByo6AbJM9Pqw/oBHWHcKy6jByURSmrLS2e+YcZQmtwJHuQLmdvTatsM40Be+ErEUBfBzHc27OY7DsIma0zVNmtrnLvzt2LvPt2Hq/ryS8avndgrJXUvBfbWosAawLKkwBNmW9C/lpZzlyxIxfh0kmPmngYMEo0bxXqh6Azr3Iz62EQvNKdodOEwpmEbhoC7Jo48A0yZjE8SqOFI18KUWlhgbqXRRFaE1oSeY+K4SlMJB/WuPeMBxkuXCTBQrdLhSMV2XuuNZzUpkktDTkkzcLTDMjtfFKg/bYuEnbT6l/m8zvJnfQQZnEFM40MM2UglkFhlmRhfIqjXU/LRj65GTy09pfHX7xVcv+AAAA//8DAFBLAwQUAAYACAAAACEAQnPRqF0KAADVLwAAGAAAAHhsL3dvcmtzaGVldHMvc2hlZXQxLnhtbKxabW/byBH+XqD/gdGHpsUhEpcUSUm1fYgkshegVxyauxb9SEuUTUQSdST9dof+987sC/eVsmozSEzn2Znhzuzs7rPDvfr++bD3Hou6Kavj9YiM/ZFXHDfVtjzeXY9++Tn7NBt5TZsft/m+OhbXo5eiGX1/88c/XD1V9bfmvihaDywcm+vRfdueFpNJs7kvDnkzrk7FEVp2VX3IW/hvfTdpTnWRb6nSYT8JfD+eHPLyOGIWFvUlNqrdrtwU62rzcCiOLTNSF/u8hf439+WpEdYOm0vMHfL628Pp06Y6nMDEbbkv2xdqdOQdNosvd8eqzm/34PczmeYb77mGvwH8C8VrKG696VBu6qqpdu0YLE9Yn23355P5JN90lmz/LzJDppO6eCxxAKWp4G1dIlFnK5DGwjcaiztjGK568VBur0e/+/zPJ3gS/OHLH6Ltv6Obq20JI4xeeXWxux59JotsSkaTmyuaQP8qi6dG+d37raoOXzf5vvgHZtwectmHXO7Qr5iqf89fqocWVa9H83jkYQ7fVtU3RL5A33x4bVPsiw1mk5fD47FYFXsw9gNOg19pR+BX6MSk64X6u+hRRrP+p9rbFrv8Yd+uqv2/y217D68dE38eJtFINP2zevqhKO/uW+jxdDyFQGGmLbYv66LZQIpDr8ZhhK/cVHuwDz+9Q4lzFVI0f6bPJ2Y7GkfRNJ6h8aZ9wawFmc1D01YH/nYav84CDDG1AE9hgXTdu8gCjCu1AE9ugczGs1mckMt7AS5TG/AUNnzbBvTxjCfgMbUBzzP9OG8D8oHawLxg8QwgfDIMt0XTZiUO09muJGJY4nAmXZLD3hnEJGJDSlNpnbf5zVVdPXmwDuCLTzmuqmQB/cHkwM44UmMKKbtBlc+ocz2KCYwFiDeAP96EQXI1ecQXcamlkJrSjELFlQ2tbSi1oUyDJtD1rv8QabP/JBmHc+1PX6oLh9AIOgSD2jkUTKeGQ0Iqlg7Z0NqGUhvKNEhzCIKqOoT9ElnA5jodM9fU7cYHTVyPwkhGfsmgqZ/IvguILjJ0YNcCmndSKYOCBDrchcbXA5MxmdiHjjpkNO8gT9/rHZowvGNQTNROEr2TKymDCxym45pD/lzpdxzoeikTChI12YNpZITAKRVKKS0IuGoqc+4tQ4wmIAgxbDwyY4185TI+TB4p4xv+rZjUlPjdmK8FxNZwDFXKbUVse0AoYxCZh3STUuck5MF7/UMTzD9cXIyxXPJWdbrGxmxdcRklu4WW9CG1oYxBLrdgvX2vW2gCc1fteWgMG5MJEnXYYiPfVkxoKl1ZS9MYMTOLWWsEROXmanezyU9lm+8//I3AVN6hvDmlhXxkyMNIGPJaZiN3eWdmowkWIjFPlwxSnF1ZyNpWSxkUQf6jz19/+fHPabJIk7/0uGyLZ8kic4hrHsPa8V6P0YThMYNUjy1kzRFY9ESgUgZFvoQyAdmzFGb8u7tObch5amTdUjRDaOUKFJm5zKXUZBaQ6hzHNO86jG512sAQg9m8ZZWlNl6br1wIuq86GRvbj5Tqth8BgZOUQRkbD2+OfDYHPx6rtmi8tvJum48f1mTek8Zn1VKHmh41g0+9KWqM3OAq101h3JrBlprRNrQWkDbsTFMfdoE5hn0A/oTEVtl9zAVaNKtZbZDflZSRw83ZleYcw3TnBEbDp4/OAPSJCP50bg/iQvomlFgTl5nSJi6H+nKaNffkdAh7uHM34t1xq6UONT1qA/AtPBpYO7d5QOBCRtRgoqrnohWX0qImGBdbCQy7KdfQs4SpYEjweKfSL3IZ/6JHPThDuw9H8rBnc6QlfQMcFDWOklhLHtPUHOWQNgcYFsHxAfdpY50LZ305cU4tdajpMbqMzF0aI0m+5Irn4nGJtU7YRI5wSIuR4GPuGPXuBefU0tBW02M0AJsjkpfJ3dFYR5ZcSNtBYdj1WWNTPqHG1xrDasqbe9YaZNTutUbwQNe2mzrU9JgNwAeJJIQyZsbkWnIhnXUgndejZtNGocijZtgFgkAXuggqBY7ZmNj8nx6ls7NqqUNNL+IMQEUDtGHQaI6prMOG1g7NlGMieejhgZAFhKfv+ODQyEADImNr6M5fRlRp2Rb37Pty821Zseqgu4LblbVY2UwjYrTSCGFSyYu5Q0mZjrwISOF0Kccif0ZTJZ19B2HrmVWGbDb7LnPI6nG5jIpqcfm5Or1eHgsc/JRjWqZYlHXt0Ew5FkEVSUQr4xjx7VNXMAA/pTbM6o8595dCSh1pYlbFpJAcalE+VOokXEx3kskR3ybhwQA8ldroLwGJZs07s7YlhaR3kv52x2YupnvH5JzeDcAnAxefNA8aXEhb44khtJJC0ke7WpdyMd1HXsJzjeAANbzAwRw5prsUmcxRSkmfpDE8LBs8KuUaHYnMyudi631umqJtvNXY+089/vjhS9xHJjt1vulr6j8J9RDeeq76FQxQIaQ2zG2sI4lKhSExGRLXVOm2YgyDZmikvLmHec/69vpOzcmQHGr6sj4Aqwzsat+SY9oKbnNGh2bKMa1QGMQLCE/vXu+oFYJG5tLQnR+AHgacHkLt31HpXYpmbWE0T+5SSE4wuwyZcjF90WByroUxHIDGURuvldy4kE5+rfOnlOqcVKx3qz/H2CywK+G9zdrAhgNUG6mNV11nnE4/K5m0X1hSP+1JLki/qJinJa5irwWnP+WH01/3UHLEL7zOI1Oni5+I1RN8pwtz6fzSGQ5Qd6Q2zPiZNRoupMfPZEXCkho/SRoxfgZlTrlGz1JK/L61tNMzQ4dF3tSlp6fdEF+LJd/rP6KHvHipFg2JuW1LITnhpHGadkamplylL26954lOzx23184W4QAEldow0808pnMhne+YRSApJOMmaSqNm5GhUPPkx3RnQYhEfeewTs+5d7v09HwbgPrCFSO7lEqsecpLotpXlcQk+NyWynkU8zRy5pdQ3t6XcXFv5ETB1RXxlDj09MgNQKhDB6HmmEp7bGjt0Ew5ptGeEEocYdRHexwaGWhkLg3d+currfQOzuUljtBReRXYuRKHlJFTzjaVcrGuxBFAjSOEOe7eBkW1ldVDMhDOHMJ6ZC5jw/9/kQMLPGY5TGByW1vZ0NqhCeV09vlOqUlnHIvZnSLdqwFobuj4LM4xfTU1z0FSSA6ttOW4TgKFcId3gubSSzCad9MBaC61YRzzOKZOZRtaOzRTIaYOD8eAQlgfiOBy6yVf/i/8+EGtwaddVkWg5eAlx8JIvcSC35616rQi1Q2Vw1oqsFgmLl7Qxfwm7N4SGx92U5ZdbzwU9R29Udt4m+oBL7eGQE47lN30XU4XcC8M3m3geAOYhs3EAzjj0XBK8zdXp/yu+DGv78pj4+2LHb1Hm/izOCZTkoRBMAuneH2uZldv/bGjrcW6JbRMZ35IpnO4ShfM5wF+RL6tWrhU29N4D5fcC7iLBfd2Fa0gmuN90F0FFwZ6GsFl7PXXon04eQ3eZYZrhtDHqi7hIjC94n49OlV1W+dlC11f4H3q+suW5VJ3If/mfwAAAP//AwBQSwMEFAAGAAgAAAAhAKpQz9yICAAAdyIAABgAAAB4bC93b3Jrc2hlZXRzL3NoZWV0Mi54bWysWm1z47YR/p6Z/gceP+TSD5YIvuklkjIniZzcTO/SqZM2zTeagiyOSUEhafvcTv97F28EAVK0xtbNneVb7j7YXTwAllgtfvpW5NYTLquMHJc2Gjm2hY8p2WXH+6X926/xzdS2qjo57pKcHPHSfsGV/dPqL98tnkn5UB0wri1AOFZL+1DXp/l4XKUHXCTViJzwEZ7sSVkkNfy3vB9XpxInO2ZU5GPXccJxkWRHmyPMy0swyH6fpXhL0scCH2sOUuI8qcH/6pCdKolWpJfAFUn58Hi6SUlxAoi7LM/qFwZqW0U6/3x/JGVyl0Pc35CfpNa3Ev668M+TwzB5Z6QiS0tSkX09AuQx97kb/mw8Gydpg9SN/yIY5I9L/JTRCVRQ7ttcQkGD5Sow741gYQNG01XOH7Pd0v6vI/7cwCeiP5wbB9EfrT//s1eLXQYzTKOySrxf2p/QPHYn9ni1YAT6Z4afq9bv1n8IKW7TJMdfKeNy4LIDXG6kt5Sqf0teyGNNTcVjSuI7Qh6o6DM458C4Fc5xSulkVX/KkX+Pt4gOPW7Gbv8u/YgZ1/9eWju8Tx7zekPyf2W7+rC0ZyPkzLxJYMtH/yDPP+Ps/lCDI/7Ih/RQfs13L1tcpUBscGXkBXTIlOSADz+tIqMrFIiZfGOfzyZ2Vb9QqjJPGwOYR2YAn8LAn46m03CCptSfxsZKH6uaFMJhHQOmj2HAp8BA3sj3ke+E7qUYECPDgE+J4XQxwMsBP2AshgGfEmM68jzX8VDLj2GMUGDA5wAGpHnAj4mchNCbqnjUJLcnYswnkBFnm9TJalGSZwvWOgxRnRK6c6I5OMOoEI7osuulAnCAWn2iZks7ROADsKSqS6DsfnVXffiEFuP9avFEd9M5IKdABdhwK1w+YXtlfRnfWr//+w9r88toMX4C+qYCcS0Rp4xudJBNVwT85+MqragrijXRGCJtwoXA3hIuNaPhzni4EOzTyvVDIwSh5ToqhK5o2xVFXVGsibQQgP7tEKhfkiZ86bNJ7Zs+cJvPHoUALaQcXUsRX3Is/VLkNuFshcgJG1EkRWxTZPAxF4UOHNqUHCxbTpMrLRjg7QXBUFa6sF/0s7IJi4ItbS/wG+/WXBQiSGfjCjC0zbyN0qEbHQ19K0QOLHGd3lt/kN5fSY2tr8TgdsTh3AkkpvHC9Se6H3GvlqdIpiWObnitdfsWFlAImi4RJCe1QWmpAxRTnqu5ZNnaSCVFga0UTRkq8nTYiD8OIL902/h4hLRVVk2s0/dJcfox//hhO+V5Xhljxa9aRl1LLXEQyXsTRyHMxCHD0TVXQlNtyzCpx5V8fsQy5knsGU8c0K3N1Yg/HkoczTXdgDuZe9U06jHVUgeb/XtTRyEgdSEURYpPrkE6qQTbuFIyKLQRSuog2EoztV9FXBTARkd5dvvblx+iYB6Ffz2Toq56HMzjHnUtLfTd5J1LkULwYlBuQWspUtzYcNFspvZjoYTUjhdxUYDURh5LEUuM5jqQ872uU4jOjBq0XQslcKo1o4GxD3Ml31GebyV4a0q5SItPirrxwcS/O0CGYS53g41roWQsd2dmhCjUtBglfujyJW+kJRLPh9a8C+P0rvnXbaMeW40idM7eyxGGYabQ5IhQMlJo1FgbodXeMht0sWcaJpF4PpRAH3bv/gTyUnLANuqx1RNo1JxvOapp/dI5cgyirIWSnsCpucqElpZAiS4SaBQmkTAZTCDY9CeQYw8msGurJ/AKFS/iNale7Jj1u1AyGGhkYyO0tAQKdLmEoQTRTm1hMpTAAJzpTyDHHkpgj62ewMuq7MFXBjhieoqezhrmWkYGjXRsBJaWQQkvKGhsnJEwGawYz1KQYw9lcPoaBa9QbiNVbzeHvJSF2sHYOTWEZavWEYa+0zoZhUyrdpAzh9ydq3d6LGKwiPssdEZdoYpGqoxW+VBVs5RthJ7GFlE3a9GL+hbWmjSNhWnA6yM9gisUs/T2o1P7eOaLQKPVnmMUGEtnI9Xakyzw4Q1NhhQJNTnJ0eQmQrDF928dej0bT27iHl09LVcoZhGvN+leqyZWVKrwOqQmlsu0ie2aRgIucJRp3MjYEHoEV6hp6T1P57j1jJe3tdDSNzvPeJnZCC0tSgWv5lUUsVqUUtaNEu6Y3l2VMYzXCluhZARpvowJrXaQLfQmSCHjO3H3DbV57A9dGblXKEgZxmsFqVAyQjfXrdDSQudFY3sBREJNY3Ej65nfKxSNrioam3UoZDSmZh0KmRZB1zQSatrx4sLx4qJzx0uPRQwWcZ+Ffkt7hYrPVRWfil5cVmrRc5kWfdc0EnD6/InSDI4Z2pXR7pkvv9ikh3B9yNKHNaF9mNdu3l1RKWlXKCgwXwYatfap45kvA1KrdehIUfvQEbIAbnTpPQocQjdAhzOnjqEcg3Lco6yn67L6inUnZLp+JafXb7pZQ4YvdMUCXk7pa4DLNBZ067WoBy6WModdf+phXVYmXXi97fYUTFI2aV8ztu6M+QWt0PIdVRtthQzNlCzqkcVyhL7wrlBDQTOVHrW+07q0FzK3ddO1ETIP2qnNLX2j1+oG9chow5Yd57w05DPEe6i8FVbg8h5vcJ5XVkoeadvTA543Ut75Xftz6BTA2IacdoTZjZUpd2GjY/uCgl8tDi8nXObZ8QG6qc3vYgCY3pyk7DsES/vjL/UBl9Zn+BZEgT9CWw2atll1ypOXpc0fZezRmZa2f7alTXeqthen5B5/Scr77FhZOd6znu/EmYYh8tHEc92p50/gNCh5m9gZ9Tyr6UqEJ/4UWqD+LIR+7Gzm0pfeO1JDN/fMwwN8DQNDtwV6zC0rN5jRbuaeQE/gzEMIgXp9i+vHk1XRbvvSpl1kUmbQtBYJPJGyLpOsBtfntONfft7x5nnzlZHV/wEAAP//AwBQSwMEFAAGAAgAAAAhANDnrFurCgAAYT4AABgAAAB4bC93b3Jrc2hlZXRzL3NoZWV0My54bWysm1tz27oVhd870/+gsDM9T5FEipJj1fKZIxG3mabJSXp/Y2Ta1kQSVVK2k3b63wuQhIiNxTiuwsw5EfNpb5BYuHARFK5+/rLbDh6zotzk+0UQDsfBINuv85vN/m4R/OXP/PWbYFAe0/1Nus332SL4mpXBz9e//c3VU158Lu+z7DjQJezLRXB/PB7mo1G5vs92aTnMD9lef3ObF7v0qP9Z3I3KQ5GlN1XSbjuKxuPZaJdu9kFdwrx4SRn57e1mnSX5+mGX7Y91IUW2TY/6+sv7zaG0pe3WLylulxafHw6v1/nuoIv4tNlujl+rQoPBbj1Xd/u8SD9tdb2/hHG6Hnwp9H+R/n9iT1NxONNusy7yMr89DnXJo/qasfqXo8tRuj6VhPV/UTFhPCqyx41pwLao6LxLCqensqK2sMmZhc1OhRm5ivnD5mYR/Gfc/HmtP0Pz1/j1WPeF6sh+99/g+upmo1vY1GpQZLeL4Jdw/mEaB6Prq6oD/XWTPZXO8eBRg0VwSO+ype5ln98bTbKnYPDvPN99XKfb7KPprH9Mv+YPR5OsO/tYd3bTjT/l+WeDlL68sT5zmW2ztelQg1R/PGarbLtdBMuZHgn/qq5FH+rrGJ0uxD22F8Wrjv++GNxkt+nD9vghf5LZ5u7+qE8cD2MtiOlR85uvSVaudVfWpx5OTLHrfKvL0H8PdhszJHVPTL9Un0+bm+P9IpgO4ziMx7NoGgzWD+Ux3/2t/iJs0utE3YBVov5sEsNwGI4vJxfP5+m2qvL0p827fEneRZOnP22ePnzmAvXMUp1If9oEfc5nEi6bBP1pE7SOzySEuoFr8UxL1yJFw9nsm+KNavGrhk3SY3p9VeRPAz0wdSuUh9RMc+E81BdcNeNsaAZJVzPq9jNZv5i0RTCLdL10A5fHQveu2+tP5atfwqvR7fXVo5n65rrktZ5i9OxYZsVjFlwP3o4+Dv7+j38OVu+GV6NH3b3WTYnLtkTTVcxJVogSRAwRRyQQSUSKoJGW6KSTFuQcnUyao5NW6fE6it94dW+jTnVHlCBiiDgigUgiUgSRuutGJnV/tmeYYN2ZLqJqzJp2XCJaIUoQMUQckUAkESmCSPX0UHt59UzwIpi8qVrSTDvkj+78pFPX0fFY9yszTOrmD72oVR3lSpYgYog4IoFIIlIEETH0FAr9fDIetlPraa7/zgxhCqqEsl16WZNZ5IoxiWZUspWNavtPgogh4ogEIolIEUTEMDdGd3J8tuObYK/jI1ohShAxRByRQCQRKYJI9fRdjFTPzFT2tlPf5av7w7O1NmU44yHyBkD9NR0A4ST22ryJmp6mjAQIA8KBCCASiHIJUcM8D7iNfY4apoxFEI/bmiwt0sqepoBwMvEUaKIcBYAwIByIACKBKCBv35yE/1N7+K49fN8e/toefqgOiYbawPywhqYMraEjYUPGVMKpJyHkJUAYEA5EAJFAFJC3l62E7eG79vB9e/hre/ihOiQSGmv3wxpWhVARLdIqtgbLnIuEJZjJEHFEApFEpAii9fad6DnjL6wNnBmA7WibXvhW00bp6jtj0r8T2cKcQYmIIeKIBCKJSBFE1fH951nq1BbPHVphg7yx5Sm2smGuFFAYwyiOSCCSiBRBVArfjn5DivaBtH6QeqlxCWunSFRqkKeSZ+NXmJkgYog4IoFIIlIEUZV8V9u3So21dabp0Lpdd4aBsMSGtZkMEUckEElEiiAqie9tzxpDtVv0Zhj/gS60UXSGufRuWTbMHVZNpqsOII6JApFEpAii6vhm9yx1arNJxk6DvLHjPwOFkJkgYog4IoFIIlIEUSlcY1zf6v9vXxyC81xaRJWYjf1OgW4YC2OIOCKBSCJSBFElXFN8rhLgN5dmoanyydTReQ8PKxvmDg+0xRjFEQlEEpEiiCrhWttzlUBrG7betnVl6GRtmDsroJfFKIFIIlIEkWpHrhs9s9pVGfVTkWO3Yq/fL09hZM6M/YnChjmdAhFDxBEJRBKRIoiq43rWc9VpzKhzIzUrJDg8Yn942DBXCSiMYRRHJBBJRIogqoTrT89VAu1p1NrTdmUUwhIb5gwPRByRQCQRKYJotV0vem61rbl0n1nC2FshWJql9qpTuMPjjX/zsFFunwA/yzCKIxKIJCJFEBXHtaDfFsf49Mnp/vpSlx6h/2wQXWDUq6re/fUU5qwwdjDWwXgHEx1MdjBFGVXKdabndiNwiUvzHg3nkQt/9c0Js4Ms6WCsg/EOJjqY7GCKMiqIa0bPFQS9aNTpRS/8xTgnrBWkTbWMdcTxDiY6mOxgijIqSA+WNEJLapG7KIRhCSKGiCMSiCQiRRCtdg/+M2oXZV374Y2C5SmM2I8ZDBZcqbWZ7t0H12oxSiCSiBRBVJ0ePGmEntQi7znFX2/FzAQRQ8QRCUQSkSKIKDHpwaZWZdA1U4uoErGvBGYmiBgijkggkogUQVSJHizpBC2pRZ4S/pIpZiaIGCKOSCCSiBRBVIkeLOkELalF7pSJYQkihogjEogkIkUQrXYPlnRSW8bTW27fizbf61+buAvosb9qbMMcM4qIIeKIBCKJSBFEZXmZGX3+pw3oOScda56WudWGTIZRHJFAJBEpgmi1e3CWk9pGnnqDf+dsvvd6wwzmBjCoic107pyIOCKBSCJSBFFZevCX5pnFe1FpER0X/vunFWYmiBgijkggkogUQVSJHozlBI2lRWSWxIVNzGSIOCKBSCJSBNFq92As9YAgv33wbMGy+d4fFzBLoqO0me64QEeJUQKRRKQIorL04Cgn6Cgt8hyl//4MMxNEDBFHJBBJRIogokTcg6OsyqC/gWqQ6ROnBS3LnPsFIoaIIxKIJCJFEK12D/Yxru3j6X7h3QiWzfe+e/B7gw1zZcHlTYziiAQiiUgRRGXpwUvG6CUtIuMigoVezEwQMUQckUAkESmCqBI92MsY37Bb5M0Q/gtTzEwQMUQckUAkESmCqBI9OMoYHaVFXp/w16gwM0HEEHFEApFEpAiiSvRgMvUrY99NWUTmSjSRmMkQcUQCkUSkCKLV7sFExrWJPM2V/k+Vmu+9ufLCX/a3Ye5cCfaUYRRHJBBJRIogKksPjjJGR2mR98wJMwSaTCyMIeKIBCKJSBFElejBZMb49hzRClGCiCHiiAQiiUgRRGvdg4eM0UMiWiFKEDFEHJFAJBEpgkitpz34xaoMugKJaIUoQcQQcUQCkUSkCKK17sEuTnG1EdEKUYKIIeKIBCKJSBFEa92DG5yiG2xQOKtX1fR2qXpL1e8+MP6q3lV13Ry7m0xWTV40a39XnHQw1sF4BxNnXofsKEtRRlXswUlO0UlaZPZA6jfLtYrr9LA5pttX35WyLo9KiYw1J3HjeAcTDaNN+tPh9+nu8IftT9+7HNlRpKKMKvoDjvS07bPeHbjLirtqe2g5WOcPZhNnrLvjidY7V5dRPDevtXWu/810rrfVINd7XfUeuw4e6YKqraftia+vzJ7Xt2lxt9mXg212W+0k1bf3ot5sqjcl6RbOD2Z/qdme9Ck/6h2j9l/3ekN2pjefmb2ng9s8P9p/6JObcj9mx4fDoDT7ZxfBhS4pLzZ6s2q13VrvWMyLY5Fujvpkc7O3t1A31WWPTpvDr/8HAAD//wMAUEsDBBQABgAIAAAAIQDrBg5rUAoAAIQsAAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDQueG1srJpZc+PGEYDfU5X/QPMhTmpLJAYXQUSkS8SVrcomW5E31xtEghJrSYIGIGllV/57eg7MTQkracte2h8ajemenp5GYy5/+nbYjx6qpt3Vx8UYTZzxqDqu683ueLsYf/k5v4jGo7Yrj5tyXx+rxfipasc/LX//u8vHuvna3lVVNwINx3Yxvuu6Uzydtuu76lC2k/pUHeHKtm4OZQf/29xO21NTlRty02E/dR0nnB7K3XFMNcTNEB31drtbV2m9vj9Ux44qaap92cH427vdqe21HdZD1B3K5uv96WJdH06g4ma333VPROl4dFjHH2+PdVPe7MHub8gv16NvDfzjwr9e/xjCjScdduumbuttNwHNUzpm0/z5dD4t11yTaf8gNcifNtXDDk+gUOW+bkgo4Lpcocx7pbKQK8PuauL73WYx/s1hfy7gF+G/nAvHw39Jf/43Xl6SOPncLC9P5W11XXVfTp+b0XbX/Vx/BgCxOp4uL6dcarODgMBOGDXVdjG+QnHhzbAIkfjnrnpspf8e/VrXh+t1ua/+hgN0D+ocCH1Or3Fk/7V8qu87fCtcRnAZx/xNXX/F6CPY4uBhVvtqjaNvVMLPQ5VUe9B2BRa3v/QjuUKOHwWzUIwYa+hHL48tJ8sFDN1U2/J+3yX1/l+7TXe3GM8nyJl7s2DcX/pH/fiXand718Hg/IkPHsYhGm+e0qpdw9qA4U28AD9yXe/Bdvh7dNjhRQ6xXX4jv49U92wShr4TuqC87Z5wuIPM+r7t6gN7OnE21wCxQTTAL9OA5t+pAvxDVMBvr2ImP/2mart8h217diRgNlETCjUBd9QgW+CpRAP89gOJvk8DZEjqz9CLYDxMi5gvPgwcr3QuSEymZVcuL5v6cQQrH8xsTyXOoyiGoZBZDSd4EVpnFaYT33WFb1uMQxfsB1+1XQMRuV3etD9cocvpdnn5gHNrDJrXMKmQftuqeajGy9Gn6fXo3//57yj5++Ry+oDHxTSuhEYcOfghiYlSE2Umyk1UKGgKxnMPgK2v8QC+DXsAnIY9APY/LN0ArFesElLcKhOlJspMlJuoUJBiFc4D0rw+P5tYGFY6SVvE9StGIrKMyWQYJDVIZpDcIIVMlBFDCMsjxs7tlyDNTiRYbYaA72lUYhWLsTeXpgTN9CmhQj7NUMQ0g6QGyQySG6SQiWIazm/SZLzGNKxiMZZGvTJIYpDUIJlBcoMUMlHsgAl5qx1YBV424CuxbDxXWza9lDRJFCGfbGZkutMeibjNTJSbqOgRZFA+CIcPQTEZbw4vTx1OmwHsZva0yeMTK4NJhB2dP9abaaZbZBDShBKLkOvPVU2pRSgMVZmMyUB2FCOKVJncNmrtWYVtQMEZl+LK/o2rAavA5cecpF19hdOrPiRHYZSrWZVQoQDNyM5VILZvLcWgaZAJuT6DZybKexQRbT+e/lAeTn/e//hD5vp0P9T1Furzk+hDGl1k0Yccxon3T0leiUfIbW91HlYBzgtcntxXFPmeSPeUBGjOUcqQ63CU9YiWaTiZ5j0S2oseEfWKNbi4fas5RAe2R2xWK8Z8XxjEUACZB5cq118+/TGJ4mT+J/v8pFw+4PJpFKfn5DMuH3L5LIqzc/I5l6cBiMeTR3F+Tr7QxpMg5wOM8QKe+wF0PR80SKvyXrMBER165va11bLiUlLmZgx7vl9CKWdCLuNMpPicM+olsbBynBO1hUIWLKxkEt8BvFXgiU4QAkchcBQCR5mrXI1HrRjE2RoWKn0lGVqFIFqVoVAsp5WFJYx50o5mEcssLLewgjPiZtUsrRo8Zxap/gO8jd3tNpuKvrNR4/kr3wtbHKL1nWq8yRImpxhvimUWdbmFFZyReFKNH1ZYShY/W2Mio+5bmSgxUcqQ7JnMFMtNVChItW1YZTnYNrPKRGaZaaKUIdU2s9Q07ywUpNo2rNocbBst+uQKGhkoMVHKkGqbcWdu3lkoSLVtYFkpVuHzMcnKL5FMV1DXkHch6Q3ORClDqm3GnZBrackQilRVKNpU44YVeIMnjpVzsnEUeXL9jiKtJk0QlxLbDiscpdScmWI5QwjsFSUkmmubXaHcqbpgWJk22AWsPpNdwJEYYeiplXsC1Zv21pgypM64IZaLOyUHIO3toVD0q00VrbDDO0s4wV3E791Z8C5O3+6h8CSFayj1KSwssbDUwjILyy2sUJlqpt4+e37BSrbQIgUptpgscU2WWlhmYbmFFSpTbdEbYYNt6esdeV5MlrgmSy0ss7DcwgqVqbZoBc8LC02al74GkW0xWeKaLLWwzMJyCytUptqi1S+iRTvcKlqzvGe7FjdwWAdYdDZNllrk4J3YuDe3sEJlqlO0wuc1TqH1idrBRbNA7+AKMWGoyVLyEYM2hHmjwMJyCytUphqqVUHD55yWKNI7CG7W4xd/KeckFpZaWGZhuYUVKlMted+aBzfeX272Mim522ui1ESZiXITFQpSrX3fIsg1iyATJSZKTZSZKDdRoSDVtPctblxafbzUEuZiUmeBMaUpzJnUFbaw3MIKzuTC8kwX0xvWuhKd4aHFHlH8UpfYJmS2ia1SM62KS21SRqO4F3q2U2wdut4qtg9KdLiVWPOGda4G+5aVU2cax+Rp4Hq5c4wivXPMpFCAP8Nvl4XXt3hnrhNFfjRxwygghwq0zxqpdqfrIoQm4QXynDlsO6SbNXdQOAnm/E+k7UdZryPErWb45Bg5oR/4E71pzx+FtI407LRnGmfa6BIPGmceNM7gN/f6xpnvI+QjZ+L0NormuzpxZicNth+jkTZ44ljtKDetPcrkrjVDKBDbXsqZaD9mnIlkknMmveRyRpKJaqDZU3uTgayglLvYHmVyF5shFOAvH6yL7aE48dy+j/1yEPalKw1f3HiGaY7BT72Kl2OwV0GDC6uAKInBrb2Ks2HZ3+ny8UNsxeD8M334orc4pMNNPA+i0oOo9CAq4YWXrJrBUfm+vUCvL2Tlr5lGT5xLSTsXY4hY9bAMQj/05/5EW+wpF6Mfulx3Hs08/RteLzTDHnpYIj+az13IQlrizbVnzgIv8MOJE/TJSkshkNdYvdh73gfP++B5Hzw/IOepy8XsVFo764MTAi3BlT4znDqiBa70HYsxuc9sEcssLLewgjOzyQ6P0L9l2S3Eh4ve1mQnj8KVvFTdW1jCmGJ8/xIgbs0st+YWVnBmNtnhyNubP+QRHeoZBxMlJkoZUnpaplhuInxSTzTIaLzSk330jNSham7Jwbp2tK7v8dE2vOlyyk78ufN4BZUjvPXpVzw/XsEqslxx/Ri/3tquBHCFOFjXBicLyXdW4/lwg01TGK/IVmboQTH+TodPhQkD6XHHT2Vzuzu2o321Jef4Zk4UhrDfzzzXjTx/BjtuQ4/+ORPLta4+4dN/Mz9yPPg4H8LRPkhF+NvMTd3Bob4zF+/gdG4F347h3KB0lxvM8bG2bV135y6CCeyQ5v1pVDc7OH1IDuQuxns4Odyuy1MFA47x8c/m44aazM8PL/8PAAD//wMAUEsDBBQABgAIAAAAIQCnWWf6hRoAAJp+AAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDUueG1svF3rcyM5bv+eqvwPWn/IJXW3tvqpRzy+Gqnbla26vWxlbvP6ppHlsWpty5E0r6TyvwckARIA2ZbopLK1O96BATb4wg8A2ejrP357ehx92ewP293zu4vicnwx2jyvd3fb50/vLn79y+2P04vR4bh6vls97p437y6+bw4Xf7z567+6/rrb/3Z42GyOI2jh+fDu4uF4fJlfXR3WD5un1eFy97J5ht/c7/ZPqyP8df/p6vCy36zurNDT41U5HrdXT6vt84VrYb4/p43d/f12vel2689Pm+eja2S/eVwdQf/Dw/blQK09rc9p7mm1/+3zy4/r3dMLNPFx+7g9freNXoye1vOfPj3v9quPj9Dvb0W9Wo++7eHfEv6r6DGWHj3pabve7w67++MltHzldI67P7uaXa3WvqW4/2c1U9RX+82XrZnA0FT5NpWKxrdVhsaqNzbW+sbMcO3nn7d37y7+a4z//Ag/C/PH+Mdxbf5g//z3xc21XSe/7G+uX1afNh82x19fftmP7rfHv+x+AQKs1Yurm+srz3W3hQVhBmG039y/u3hfzG+LtjQ8luWft5uvB/b/o+Pu5U+b++Ny8/gI3FV9MfrP3e7pw3plZnvasL/+2SxhYBLED2bp/2n1fff5aJoGdQqQMZvi4273myH9BJ0dm35sHjdrszxHK/jxZcOeePgP0vV9Ma6nzaQNfTJNUP+48rd2Q8FQ3G3uV58fj8vd479s744PoN9l29bjtgQ98Hf/tPv6D5vtp4cjqFdfQhftap3ffe82hzVsH1DwsjKPXO8eYXDgz9HT1pgBWP2rb/bnV9f25LKqynFVmMYPx+9mjNqL0cfN4Xi7Na1fjNafD8fdE6pi58Y3B0vJNgc/sbm2vWyaup1ORHOvNAELyDYBP7GJAv43aPKKKHTbisJPFC2LzKeDlrYJM8NuPKCJ6bSdFOf3AIbLjWlbTYMqbM58Z8yydhNiV263Oq5urve7ryMwEDDQh5eVMbfFHFq0Mwtt8L0z9pPh535g4mHGTZvvTaMwny2YepjLw3EPq/b+5uPhh/fF9dX9zfUXY6Dn8Nw1TDvY8MNm/2VzcTP6+erD6F//7d9Hy3+8vL76YrTGFhehRbO4zEOWMamLSX1MuiXSxO4NGAc/GLCe+GAkFzf10fDGfVyvXrbH1eMP70vX0Zs/746bw8iAFgDg6GW1P45296Pjw2Z0u31ePa+3q8fRh+PquDHocxgZTvvb75vVHoDzbnP3h1FVHI6jn1f79cMfRuW4bNTYBE382MSkLib1MemWSPHYmN3BFsqrY2N4YWwmY2sKzGwtiOQ2sp1AIlmbaue0i0k9ktqZW0uwkr7cjOUI3BJPrDVsDL28y+zlbBqB/TFtaRl3Ba7jm0rNBbFOmLZlM5FcS+Ka+vHpkDSBn2bL2G6WtZrqPslVBa4rvpqNKdRbG8wkWdZzN7NpxvR+FiYTSTPw7YKyzVT10nHVTegkyYVF0CMJ7LaxUlx9sEZnLzjDq3RE0gxWNNNxpnR0XFxHknMwZhZqj6SEjjDNZ+toeJWOgcR0bNXqXjouriPKzZiOSEroaPztczeu4VU6BhLXEdY/t89Lx8V1RDmuI5ISOsLmPltHw/vuorJ73VqNhaMUY7bSkMSfjqTE0wtYx2c/3jLL5yOpbBDv7OZt9XZArqIImndEm0ysafnw689/203m3fTvED3UUugT/P1k3if4xVYqFMq/arwts+qgg8yiCDZgiWyzQAKraKembgKtR1oxsbZOapUDt4VDJj7tSKqN4xa2eAvIy5dmh2wFXwtESy2GHKCzPqMaKwdGxeuGp0NJqRWKprRSQPb6DDqYEGNFkMNhtNSmpiuQTYwVgU5sok2Ecv7GcaZeaBWwhc9grWfQsTXl1O6SbnNcbR8PP8AYDm2TtECfEJArMgdzCgcLoj8Bh14xlh1Kyrkn0cQoc5SBnr0+9w4IhFYx8nRghKylFxM9jCEFB5GTKjg7L1QgNOBoPNFAZ58C8DNmHiHSmnIWLGRRz6EDgzbSPYtL9CDRpyTk7HMUOtlJhBPwJsnrXoB5tNBktoU3SBMNlchVjJn3h7SmMvH9/c3vu2IK/8H4mGBJu7u95u6BG6Ah4ha9KznIgYI20msuQWUX6p7pDi5sO+8umN3vkQQgpn0442t7A3FiSHvLbMIG69ZI5TlUnJqa0kGF8NfBpNih1P665xUO+1ixLYmNYzfRyGXHKLejR70/jFZHiN0ux9VlHLX1A+L9gHiwiXJgOFqdHBiHLzVYeb9mS4Ir7su3OmAhLtiNJNkRbczQnmhTl0sTsTVHsJOaOsypYScFTQmbhEdfKrhflsjGjYinscCDaFNrbOSgclg7qarDGqmqo5Vmo3pDUJU6+LC5L2XviAZ7inreEw2yedIcSK05eOEGL6eXYA1fz2NgCCSG2tGk/mUb7QrELDHURONDTdiW2Ng52FZiJCRUdTSjash+IJAJvYjG9SJgTOiVA3jgl0RhE9JKZieXxMctP9HAsw/zzaMpOcc5EAWQadSqwd0Im8jRToQqKClCFaLxUKVs5l3ZDgFxQgKW8rxPSYheAgqejRrvLTO44GPWS6RJH7zV24+4uFFDGgtke2IDg63hrVLw9kZMfW/bMTPFzB3SVBxR6j1IbHytU3OhtZ7YUp1Q2PrmTjjYlZ1wNGlIila7RBUitugENgeb3m8M4gOYiaZChW1v7gVhJJ+KVEgXea8VcolOUETHvNeqnHdVObRpqJVJkOhBAv6LJeSmSaQ6i3z/zh5caathDrNMzuV1xxa5hGNLtCk5tlX1+84AYdKx1dw9cPcJbtlxjtaANG+eeQfarYu+XWapcrSCmdAl0njiIyb1JAluBi3fW6S1bSK7riLPN3fCQW3rDv6wE4TwYSctq5jWJWg90lqRl47S7/jQVL84vv9vJoeg36UA4CipKyDtbVeRShcsIPB0uMe9r6LR0SaxMU+r8zQw7OzkqjNepn3WCad+QB6c/bT8gFdf5eRsLTOAhzPtOOXoQRh/3DugReQre9GwSDuklbNCJltMPjO9ad2ztECfEJD7ViV9zZI/w1+t0KsBp9h7NZ4GK4x1N8JKdH7gIMdHMV7UZWB9bsmkEdPddY00gKgmVCeBPiEgulvnJJkts5nTEG8tPE0ua42lyFZO+eGSebbJ6zsrbNPM1RRgaDYEQ9gKJSSMRA8SfUpCdjMn1Vxjwlh0E2nG7Qpz2WjnDUULF76500NqDqbGuwyST6qak3+u0R0BF8QvO08Tu0yf9aEL5BLgqCg1hqd9lDqoIO4+x8z4B0v5fkh+wMzUCZ/pjB1oxfTiRLfJQBzbgco0L0kUYMXvQE9jjp6nxacGdU4m3DKDqrDLw6w5Z0b5pPpUfYmiJTcW1BzsDrv366EMHTI2sFoMY59glEtR+TCvRuwQnFh8E51KZRyKstWLEdlEp5A24bsGHwG7RzvatfJUXlcV3QyhKmYDxAFJtL0xAyGsmKPBEgu7G9tP6ZnwPGrIxJ1Ih9ToPAiNQ46BL+7oVB/ZxOAiTQwuPiKltAJ+g4nN5BJ6+fowI9wLpRGWRQ6qKPWpZI1sQmmkCaXxESml3wbkNQK5UBoRWikdrY4EkGNzpVAaH5FQulFwbEZ6cmpxWCEwJ2B/vTnxNAEClfY2PRtztTyNrWhPczf3xH2OHGht8MiW4eMCabCDfaqMSPySBp3ihm1GXLA5fDCDtFQw0+QAq2UGzw8OisOYIj5yRSNSh5KsOz2SKnO85WEoTIWwuU3OQa9lBi0x+2XjD3KI9X0s4hXXAaLLMQ4wS4jzPRKiYG32JI8+6jOjjwH5fkh+wC1ochDWMpstwa8HBYQNCw1p3KCjbMNSd72kyQnLAcmGEIzrFUAy6EWIyNxllJV6OT6kSb1yELEhxOJ6IdjxHDbyCTceaVIvJ5vUSyHgqxDSEChxvRK5deSTejk+qZegyfHKiW7NJjLpKLG+ArSFeSQc4/PoaFIvQZN6KRx7fbwIWvh4BfQKehFUcb0whBTrXtCEXm1O5GiZ1XghTZyREI3vR6SJ8ZI0qVcOHrWIKnwekSb1cnxifSGf1MvxpdY9XNLPuDmIuCL0CllrP4+2UZOgYPOINKmXk03qlQM4LQZUQi/EDG4nkE/q5fikXoIm5zHH3rd4Miv0Sth75JN6OT6pl6BJvXLsfYv2Hvwn70YQzeTUvTfQ6JgIuUoeoCOtAUkTv9mESQO3Ttp6KGGSkAAsm/cpCdnLHPRoET3A+Qq9jMKiJbKxSKmLSb1vLHizt0hLeXWQ3c3YVggnQlFHY1otbZtwsMDcz5jUE+kcr868c3D+tWEEF3Cc8AWFrqCzCbVMFrZhMADCqyujM31i434d0mrj2wi/7rxsz4D46ZsiNtl0i+LuOpZcejmAZy7gGyAG5ygsPaLxvGsZxT4oWs7YzVffnMu7/u7lb1ZPL3//+Lsf4DbRQOZ1UOQ2ISI6Ci8iZNzTdsnSGibad9Q2YDrPIhKklTMW0SGNpyi8qHUUpFo5AApxrT2DY4H6Ammn1iSx8TWJtIIf57ezeTcZD1m4hEQPEn1KQnYzB49NDG6XGbNwSOOGIyZ1Man3jcVH93DJLGNJIBhza2YbEKZrGZO6mNQjqXXX5dweJVriBM2cvJ3/jgGis1AUD2/Dvl3aNqXZjUk9klBROaU50GzfyTObh+0nh9Zw1uQdrJiri0k9klqAW5CUOuUAKZyPumXGdIpIy5iri0k9klrAzkinHMyEBWF1GuPJx//Ba3PYZAmmMww0RnWM1iX4eqS1iZv7QMpYkwiyqlv/n2/IWX0Bu8UwYMAohiGm9SibHIYc/ATj6+w3W3EEn2xyIlKHgvxKEpJaZ9DFLoAjvozXfQjqgv9T0hGUSlgtbMOR/zPRqVli41iDtNrEccz/gSTakA+j33sbbuE2774sXEzNGB7KhDLLZRtwngC74j/RZwHIJlweL2pcnvhW9+Cv5fzmgCkkf52VY66MpzFXBmkiWENaywKqnvjgboC2dNMcPLXMZhC5XoSxXC8MeHnQjbJSL+RL6ZUDn+Y9aut8cL2IxvXCgJe7fijLXT/fXOz6mVf/z38rj9CSXXG0DZwOR4hNbEfXnHD9pgW8dTYY3GIrXKIHiT4lIVdrDiaDZ6stJJKY7VvGpC4m9dQWu4x3i7RUbDvNwWnLDN4TXyQU7gZLjlzilciIq0cutOTOGSRawhmEjZCxbBB4IYVH0W1FJwnKbC9sw3o5FSaukq95uiYLsSHpMerO1JnvQeCTNThkvgcBbyeLgTnzMpEVA0NkMnkhQ6SM+QK51HX4cWTzHXYX7AJhh6KFs0x2fnv/0PgWq3mj+vwXYRG9+esctgFzmga6+A5N9TsSyCU0RVoNiTtyFXtJEzsbOpShKQKp0BQTvVJTfWXLPgZuLvExRVoDt0X5DSx40XPoXumARJ+SkL3MQVtzocHenGWv1yBNzYe+FoNcspeYSOaAgnyNo0lNc/AX3vqLNXW01twU9ysneokZJVt2DapDWgOXenyydDaed7PBS84JCVhr8z4lIXuZg+aAznEvE6lq5JOj7/jEvqX23A6Reik4P9P8zBwSizoKuozCApm09dEHysgme4FIL16Dx0R5qhc5aD3DcNlFvfgifjjIfM36OC6pKbVm9nTsG/uH6V/LWchBcUhG2tUh9A8Hnkz/yCY5Lqk/taZtUgV55OQ9Uv98KQHvKMcSspc5PsCMwJm5jkg7ZZMQ6oXlpdYGZmno11L/nKgZbtnHs+Ro0lKVRbQhiI29Teyba5ipasFUDVZcSEj0sxZM1cmaCybfkVFVAsGcL8fCNuHQ3KdviCiWHxErOGLw6E1EhHQxCYUp6pShHSafnUvqNrtrAt7XFYgBq0r5jIGPv9Ztnw+X2+BYwoL4DN4LmQ3hNzVB3D1w9wlu1cUc8DYvnevog2g8/Ah8ofgKvKnthOUb/Uhk0rfEmIpBinEOhDtudW+KiPyGl2fkGzkQg12AN8jRLQBblrg85YITYkr3IAeeoUCdHXF+p6qiig7mbXZZAMxzg5Hx2pmbTSJACY3yuiQka9JIbziA840q+cwYxRSoydlzdDOJO2S6lNTCNWouJvAQptaYFdjEsqVHsEJCgdOaErWllIeArykBNauIABQnRE8BMgv+kM0TzUmTn+Jypg/sgzC7RU7EhoXlsKLdY5Co+qK8hbf3Bb0G2Rcksuuoy2JMnEJxR2zZmQ0oLohKceUAvF1xTHdLxZEoFSdOobgjKsUFUSmeCNLhFcU3rB4H7aVUHIlSceIUiqNnIEdcEKXieQWiqBwU5HHC2qayUbDjhqOrwovy2iBIbOCitI+vYGVDDZTx4LvX1FKDryrYF3iMkNkTJ94kLXS5KCz3Cbsyc5dj8adSjgT6Euw80j1R3a4iYuNub7vcSSDGR7qFreqkK97Vl2/QGy+AmZR8SAvpExH3PKM1T7bUOrj3bBDm0XqAUlkY3sObLsFh88SE6bUFo86t41ZgeSm5Q4go7GsVXdcIwnzToHDDz6GIEyqvRgcCha0ldb7C6AyILY3lqPidfqg7RpxCO0dU2hExpV0WFttCV2amOVwRURgczym0czirtCNiSrucSNy8nWxcV6UdEqV2xCm0Q5CUM0vElHY5cbYpIpHQLgWPnlNo5zjV2BExpV1OfFxgfSk1dikM9JxCO8eptCNi/J63KaGX4QNipSdlQB1SiVftyihV754U2VQn28ArBgFLigqwpBguOkj1puDOJwlBXb8KsCQlJKFTlJ2CNfV68TKsJFVCojNApyfya5xlGdXV8nzsQhYcYdjV18Axpo0zARihspa/5RcVWFT80MUx1NZK8KtuZoXTWOmq5td0CiTyW05E44dICRrksTHBzuLzWyImIzVbfeps44y1qsQ7OvA2udvX7MWdZYoIefmYE1TGeHvgUidGm8iV7kNWvIylqUrIk/rTMDqiMuV8ZbTpuWHJhlBkHGG75+NFPZHYmto+bwk3B+Rzw01b4er8OaYzbXasWWCRLFHFiIgVZKeCP4OcLRT9Cf6MJKodkwXAVAKLZbIXrna8Kf0j4sUiygmgcCXqCvoW2WuvocX4aA5eps2x3Jbb6AbDOnyy6Bo1bDwuKE2dSpnWwObE6+5BmCdx/IMT4GMLaJ2/IhzImi4wY0xE7r5jYS6lHXEK7YiY0i4LuLFwltLOIa8ZzpAw9eW5eP7RiwvtSDylHQdumK0TQIYYLceOiEI7R1RjR5xCOyImtMsqmVX4mll8Zj2Ra0fFr3iZhSDOtfPiKe04Op4cO6xVJWfWE9mV0YKqWkntML7kb1cRZwV5FX1zqbBVpWhXnNYOcUvMLBamqqCy22vhPdWvkvoiOsJ1W+ehVFRcPnJOUL5EVjgrilmllbXlpM7vm0sAK6MVxb1Yo0obrQgcqZaV7C49QyweIqamhyeVT0+PgzG1eByxhdeDg1nwxaa4WaBKV3LxUJsp7TiMndbOxXxKO0dU2jmiNAtYQIrXuoM1QG2mtOOYdVo7F/Mp7RxRaeeISjsSFzNLxJR2PI48rV0KjrDgktLOcSrtUnCE4mmzwOHotHYEHcKkIhHyOmzdOaLSjsTF2GEc6UyN2thZcIRVjtTMIp5I7VJw5MWFdhhKprQTxY5Ojh0WG5LaEVFoh0Q5dlSrSKRziZjULguOsIJQxercLgpP5Aa/LKKjilSZIhJuwG8NQbipx1TUg2/fpIRg70NJpqSQXCy2XtDZKIDVhSqTd/auKzxFx0ieT7rf6u72sqBqRewlP6LJA1NkRKLqAg/2Tq8oxBNWVRXmjIh8NyJRrSjH2TCMh2FGogMRpV0WRmF1IDPAwbH2RKEdXkMSCErFhaR2mOtMascx6mQOBssLKe0QZORuTGEUiquxc5xNUrusuMqWHzKlKMXYIchI7VIYheJKO8x1JrXLynVi9SCer4NNjclKeMs0hMaSqFZTViDkK/jw4yYkSoe0LPSV1QL51PJH3IHLJME+1SZJWA9X3ceW4CtsLEkIQmbjnDpwsrV7zo5LsdKPMsiEW+xNp4JKDAnLI2ALT5WQsXEvt8rZ0JWCXg/8sH5Pxa+6Qu1Jd7dE1E+topJjgY+nmFG4gXdPXXxQm28D1AX84a9zR5GCEoJJMJ8IACHAjLiyoepwVi4T6wDJ3GCCuCwSRCji6aI1Lt6niLdETKYCsyoOFamSQykiqBxnLkHlVDozQQSVHWda5azsJZX2ASDw2Uuqf6oTXovCc4uM+STKKnk+njH3RLDZb3tbKzw/aiLzda0iqxaR44br/PwdaSIaUxiccCpbxE6mYWIJQ4MlhbUoiGqnZAErlSSS2lHwF8ABlp0j1lI74uRX/KhUEZTlifIatlzQ2VaVChNJ7Sj4E9o5otKOOIV2BKwp7bKAlcoTSe0csLYMWGHsHFFpR5xCO8LllHZZEExFiqR2DkWVdo6otCNOoR2BcEq7rENE+ByXvUMrtUsFf54zbADYFQiY7GU92BVETGiXVa+ooIJF7JWiRSDCY3wIEl0s9aL8rh8VLAINg/NiareAzkOlDeh5DRMyOx+cl5SQNAK2OtHZwRXVQeIXCWLaEmrku0njrktM6yWfUozffz0ZMmExI/6WvvncsXpLHhSLaPA9q4gGinGaU8x9W9l9fvaw2x/td0/d15UXE6hxb944t9/+nn973N+99YvTk6v9dv1wBx+5NVhpnrPcPd9t7QeT7deRF3U9N7UQjcH0atxcP232n+y3lA+j9e6z+ZaxuZrsqU5POLc0Lq2r6q5+BwedxkdO/g6+Ht3bJyqZ99BtG/Ap+mJWzZfusCv6zQx+Y1e8/g1ck57DJUPXs9Ad98Fr+H7sp+3zYfQI36g2n2mGsd67Lznb/4evV1sqrJOPuyN8eJn+9gBfWN/AVzDNh53h67TwOVv8Cwwffkj788vo4D5zbYqd2C9q09ek4O+7/RY+bWu/pw4f+4Xp2K+2R3j43Hy8e//TnVUX6lwu4Fvuv/mxh+X/tHr+vHq05KWbEJMv+bj/bWQkTXUK//1jUA24zdeMEQVdYzDB/qvyN/8DAAD//wMAUEsDBBQABgAIAAAAIQA+hB+CYhAAAFRLAAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDYueG1svFxfk9u4DX/vTL+D4oemmbnYlmTJf7q7N7Gl7WXau8vc3rW9vim2dtcT23Jl7SZpp9+9IAlSAEl7bd1MM7fZ3M8ABEIgAII0r779st0Ez2V9WFe7617YH/aCcresVuvdw3Xvl59v3056waEpdqtiU+3K697X8tD79ub3v7v6XNWfDo9l2QQgYXe47j02zX42GByWj+W2OPSrfbmDT+6rels08L/1w+Cwr8tiJZm2m0E0HKaDbbHe9ZSEWX2OjOr+fr0ss2r5tC13jRJSl5uiAf0Pj+v9QUvbLs8Rty3qT0/7t8tquwcRH9ebdfNVCu0F2+Xs/cOuqouPGxj3l3BULIMvNfwXwU+sHyNx50nb9bKuDtV90wfJA6WzO/zpYDoolkaSO/6zxISjQV0+r8ULbEVF3VQKEyMraoXFHYWlRpgwVz17Wq+ue/8Z4p+38DsUfw3fDhPxF/nz397NlfSTD/XN1b54KO/K5pf9hzq4Xzc/Vx8AAF/tDW6uBoZqtQaHEEYI6vL+uvcunN1OxoJEUvxtXX4+kH8H/66q7d2y2JQ/CAfdgLghuL5B74Rn/7X4Wj01gvW6N4UhCJf/WFWfBPIehjIUWpabcimcLyjg13O5KDcg7C9hCNPmX1IT8W+jqWDVWlOdbuU0gQGuyvviadMsqs3f16vm8bo36afpaJhGSU9/9lP1+bty/fDYgNKj/ghMK3xztvqalYclTApQrB+LRy6rDYwZ/g62azG5waeLL/L3Z0f2ofkq3DxFPsUBPiA54DdyJMP+aBSiPoYnWD4dmmqLGsvhmqeC3aQM0B9lRGF/MknH4WQM2FkyUq15Gk9gvKe1H6hhy9eeFU1xc1VXnwOYXOKV7AsRqsIZSJT2S/vCzz32S+DdLgXXO8EGhomABUx7aGr45P7m4+HVu/BqcH9z9SzC1wwkL8F+EOEOZf1c9m6C7wd3wT9+/Wew+LF/NXiGl75EifNWonhJ4iELF8pcKGfQAIZlxgaj6DI2weaObVns102xefUuUgO8+aFqykMgQjmkhWBf1E1Q3QfNYxncrnfFbrkuNsFdUzSliMmHQFDKT7+WRQ3pZFWuvgni8NAE3xf18vGbIBpGiWWTVhNjExfKXChnELMJeF4Xmwi26148GsqpIN7OXEGjxCALB8kcJKcIUwxc+HzFjB8KruseUWLuIAsHyRSSxui9IO35Zshtn7c0Ik5RtxIT1Joy8bAPstSEN2HoSAzSU0iIESYN5dR5vf9Dsd3/afP6VQZeIGbQTRhb7qA54O2KSSfVjiy9F5pIBTvxpjINwZDIZG2fmI/wib+iawqHjqHsifuuU+YvSbs9KQ0+xGnPbArW6/D2BZcwYeuDcw3RVxtF7VNVZNFUMhdKKFNQkrYenjOIqTvupK7gstTVkEoxMuppiOimoCRVSURQ5Qxiuona9OyIbiaS4BK6TdoJrqEpcbcUwjsN2wskSlqzZQpK0kg6dpaic9kTjJPlLhkbFSjRYVSCCzJb2E6GOUJRa96FC2UulGuotdCtgtJUymLqhlA3ddBXssFrSNo3PdfYiGjswTLE0hhCrAkPttEJkU4ptxrzDcNTHsRJH9DLYp2s/LzBTr/20Jqic8NCo11oky0MGYl3BuMBzwTWs8KcFpLg87HEyXV09go5Et3CiyqRtshSaTxOiPtKUcKSNMDF4cSaloaMRBHEWBgxmHwEd+GLSoVWaV0qkEgC809FFxpKwtAqdxaajAYTxHQ0ufvl+z9mYTSDnzdH4oqHIwcO+HE5+IC7lSChqhRYkNEYjTIeLPNgucFIoEHMG2k85Uh4tIJvX5NK42E0bQN+qDGSjTxYhtgLkUYJE0RtpEHMF2m6VQAhZvI0tquoEEKfKqPsqtrwYHjAOmpqTyEtuq0uMmRN498UWZRgW0i3yEIrEbBuu/69MEBj1ZG2r34eaowmHo21zpkhXRq3hso5xqcZLVC0ymc4rK40aHqzK+RQ1yw0OobCFVjRYshodNT1iKrH82h8LL5owpHsYNCVQUirlAuGpuoIHucRgzjfroc1RtVWWJJS42tM9VjYkpjWJedrGAk2UZdQ49sZWxOx1BQmdmoyZGQUiCXE/XKNQbSxV2ARLUsuGIVqFTA7S1EwsjG81XZRNXWURlZShGXIypVWdJhjeTeCFgEXKO0pAiLEqHMYjJpV0XENNeZxDprxL9AQsztzDqeVoUsAPjOhmcJmZqTJ6CgUxkehMc8oaBq/YBQqjXPnQAyco23AIMY8QWGJymtyPZlHGvNoSFP2BRriypuWg6IDKiYm0xAxpqHCuIYa82hIs/EFGmLKZBoixjT0rL9FKxFGwt+yxjwa+hLfy1kECjI3kMEahLUiNREPZKG99DVk1Ff1unys6pE///Tj3V3w4acfb9///PpVLlo6siix12Uoyxs3uuXLCFPhmBR4GmNxQ6dMOgqd4lq/z5E3SaU8Htm6pT0oPZ2eg8bGPExYeX5hyKjSOuupvoNaKUxmWZS8eZsdzeYoia4v8nAyg+n75q2vBuDNVCubevccdL8vVjk0hN86mMw1pnKc6k15sMyD5YilEymP62Xlx9N6qZQFyzGil4stYhfLPFiOmFevbikwxnSnXE9aaY4Y7T5jYjOjyJDm9DqFENllhuikOJ3e6OjmknnRmMPGYivs/oZ0dkUIUWsSO+TIR4llPet2RHbI0WTjtj+UacxqVMBa5ZKu7hEp+WkpR/odYnnUofsl2UQyo0awS3xDxEpRy54LQyWLdNXhRSyB96InYM4xPololoZwdHprzCysYRWkMjIdhF0qGyI6iMje/1kYMjoKzNpsFAzjo+i2rhZbI7KuoKOwazpNxHtRQ3u1ZchIsEaMvwvM9mpkfBTdGu4xZntwqDbkelrumo5WTIhxDTG7+zTs1naPdZamGraLWFN1ajqmIWZp5gkM4zakWfoCf8YszWzoWZzGnsUpYtyGmKY9NoTdRTdsTPqTKftzae9ZShVRheQ3g9EyI7E7QIaKOC5iCURgEdxllRGPZlmcHutHejgg7MxyHwd7YbAzyMxxMotLYtjqoNWFxmh14cEyD5YbeW2muUXM14UUeavL9q3K2awLKUXBSGBBYNwfsWlby2YInU7uhMg0ITXmaUKOfAk/6YtF1lnttPZ8hJQETgfFmV0D6PMDITSY2LLD8MB7N62Iid0kR6pwQjpwmtMuAWK93Dhrx+OIlPy0lCMlwOiiEoAYTi2bR9NTNYAULqwLL6Zt2kR2EYBkIZn5mWFtU2rOMT4H/UVAB4c4pyyAjWyZdfm4JnZCRTI+LsU6IlMkR7pEhW4+Ln9Z0GFcWChMSAIT9YAoHmAY7QxWGNfZxXLkTcbuhhTkvk4xBhP+hLS2pSihIXzWTjXHf5ATDsHpcWSGc9LG/1Eyy0bH478uGkaGAzxuBm/nhf0o2IvvNF5VAoQQtUzRI0XZMVXRsZjqQDlyphO3BTPqtuyXbEIXqp8qCnjMVxjTz4FyLQ2Wcc45nUtW57DnqI4KEK082MKDZR4sR8y3CobN/fa9wow7mdclsTjCQPXyrM49dJkHyxHz6mWtzk/rpTJ3KroQ5FBR5+N/InWJU3ekfll4sMyD5RxjcU70pE1d8qKt1cLdGdP/8cSf1Nc2Q6uWCUMeupxj3AzdeuKJSsaROqyrGi+IkdDuIBkip2szQuTMXJp1sf3c6YidzqdTpwrTlZF9cmKeIM8UHVvtDlu12sJQkUaMxn5bFXZESrcqLKFpHg15xuEAyQbpEYowk0IMxvZfI7szi2RkxYRIAlWYLIVv11/KVfDucCjhVOyiH/xa91+/+i6FHTdvZ/wk+wfNHru7tXwCeLYL4nH/3HrHdHcSLAmmbfk4NxitRkd23W6o2jIkQyyBbrdZTCbpLEuOFhMejhw44OeFYiLptosg2SD90PIOMRoBsOvQlkkOTY5IqiTxl9OtkEhUNcCjE3YX2sITqcghDkReiE5KEB5fYerCEVGWU0RjcDQ+f5VonEkKEp0Jd5Goz9CGVk9irnlSengltVcIhooMW2N2eEpOnmiz3Dg/IiU/LeXYqd+L9vHbRSIchJULjGkbZuaIpWKGmJI+HtoddENGls+I4TpJbd9yjL//i3YxiNa4jcG0Rkws1chK1k43cJZWVkfguqYEQIxrjXsgnhUfnJByvPZ4KiBa424G01rv0lOt46G9mS8fCcUM0xo37kk7MUc63zpVuDmt3879forkExmMeojeueda2xkMWbnWuJnPtGYY9xBP/XKWrXXtQbVGjHlIPHROtCsyrrXCuIcwjGt9UbFAPASX+8zWiFla20kRzvu4fq23AdriI0c6r4dcdBCAaK0zObW13htgs3HixBBFlpLvCmTQTZTHF5iHMIzb+qJ0TLTGvQFma71fwP3a7makioxrrbcLqK0ZxrW+KFETrXFvgGmt9wu41k7kwzzMYojeQqBaM4xpDem1UwyRfFYM0Zjl1/ZxMSRjsxExNhs5xrXumBthY8LJjRqztLb3O5AsJd8zyRBLxu00yTnGte6YG8WWr2gZUg/RGNfaOcyJZCnkQ5MbEeO2Pp4bodnYzUM8uVHKEjv5zK+dw0NIxm3tyY1I54t84465UfIJW5MT+gaDadRWT6Gd0ZGM+7UnNyKdV+uOuVF+L9bWWudLqnU0cbT25EYUxz1E50ZpGO7XHXOjOMck/ZraWmNtMbdAOm5YTyJEukSdpuMqdkyEMLNdFTVGVcSsRyMx8nIr6qznsWLHrCcO5zpW1BhVEVMcU9GT4lCe14odUxwcs3NV1BhV0ZPPkJdbUecz14qwp9gpWkk+awYZjKiIGPNFxJiKBvOo2DF5wS6DY0WDURUVHVcRj17TCgx5fS8a+hDdrOjJVFKWivlmwwsxFuARY8mUY2xGTzqmJclnJVON0e9rI8ZV1DmI5HukQ7W5ih1zkLixwM73GmMqKjquok44VEWGcRU7JhyxCeuo2C7G2heN2YVWTcjLX7TOLlJtrmLH7AJbmq6KuKSCRXWrosbIOTLk5SrqjONR8aLsYlpcYoNVGJGcjJ9rjJ11tGtRJEqh/WuqOsQSiJemXZqGs2xytF3q4ciBI/dxqDei7k5RV2QcqrqRNzio21NAcfgW4RBGJC/3mX3Z1KuuV8qMB/V6+biCmzjEaWXxoEW1W63lnSnyghRhJbEtYnS4udqW9YO8SuUQLKsncaUJfNWRwEpL2HyBdrB80YZBfzKFT2QUtT6Be2HgGg14nI3D1yZlVW3jCXwAfUQPRxLDJ9L17KeH0NiGr4W5PJnYDPfh4VDY2/tJAp/I3qb9lBg0U3vB9icj+ARczn0+nHEFacrcrY3VFTtwN8fDencINuW9vEJmPJxALBqF4ziKJvFIrDNqdevMsO/5rKn24uKZ8WgyjMPRNIUra6bTSHxD9WPVwOUw4rqc4BHufyrhO/1wQQ2hiZKpCIj3FVwzcuRDGApeA/S0Dw7i3p7rnqi85X1A+hoKeEBVr+EKEnkbFFzGAr5WF+sGFJ+Jq4fq9yv56uFbYnO4iepT61zifpndU7GR8AI9DhzuY/0pEJxiK1neYgOPBEp9+ZARBN5r7sO6+R8AAAD//wMAUEsDBBQABgAIAAAAIQAzlBZazBUAAChqAAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDcueG1spF1rcxs7cv2eqvwHXqYqTkorivMkh7G0ZZHzcHZvcmudTbL5RlMji3UlUSFp2d6t/PccYAAMHo156Lqy184ZdAMNoBv9wAzf//770+PktT6e9ofn62kwm08n9fPucLd//nI9/fN/FJfL6eR03j7fbR8Pz/X19Ed9mv7+5u//7v23w/HX00Ndnyfg8Hy6nj6czy+rq6vT7qF+2p5mh5f6GU/uD8en7Rn/7/HL1enlWG/vONHT41U4n6dXT9v987ThsDoO4XG4v9/v6s1h9/Wpfj43TI714/aM8Z8e9i8nye1pN4Td0/b469eXy93h6QUsPu8f9+cfnOl08rRbffzyfDhuPz9C7u9BvN1Nvh/xfyH+F8luOO709LTfHQ+nw/15Bs5XzZhd8bOr7Gq7U5xc+QexCeKrY/26ZwvYsgrfNqQgUbzClln0RmapYsam67j6ur+7nv5tLv5c4u+A/Wd+OU/Zf7Q//ze9eX+3xwozqSbH+v56+iFY/SEI5tOrm/d8B/3nvv520v49eQVwPX3Zfqlvsc1+/YVNSv1tOjkfXv5Y35/X9eMj44Kh/PVwePq02z7Wn9j+/eP2x+HrmbHD/p9j/7Od/flw+JVBHzHiOQZzqh/rHdtjky3+eq0bbnmQQjv+l4+P/Rtju1KD0/8tB1pwbfjlOLmr77dfH8/rw+N/7e/OD9fTbBbMs2iRTOWjPx2+VfX+y8MZo4pnMSaQ7cDV3Y9Nfdph62Ncs4j1uDs8gj3+O3naMxXGzt1+539/s1mfzj/YXo4TQddQYKU5Bf4WFBGYtG0nu6+n8+FJDDQwaDGbnHahaDHYJInTJRNlGA9YGM4Df4v+wUPOxjAWmRSaLd8wqQM1UfiH7Dek+v1cn87Fnq3DtGsqAjmPQRotsV7d47hqlo3vls32vL15fzx8m8AEsKl/2TKDGqwgFl//dMa0kVj/BFtzx6g+MLLraRpiQ2Kkp/MRT+5vPp9++hC8v7q/ef/KjOwKnHfYALDDp/r4Wk9vJj9ffZr891/+Z7L+99n7q1fs2Z3geNtyZJuMdbJ2oY0L5S5UuFDpQpULfTSgK0yRmifMyFvmiZG58/Tu+XCuTzAVk5d/3D69/Mvju58+hM283fwbf8TOMZyJk5ft8Tw53E/OD/Wk2D9vn3f77ePk03l7rtmBdJpM0JI//VFvjzhL7+q7302i4HSe/Lw97h5+NwnnYWJNdTsoNdUutHGh3IUKFypdqHKhjwZkTDWUXJ9q0hCJLVKyttfTENolRakMyGAMLRm+hmqrMyosYYwBs60O/PUmjLDNje3btlJz6kIbF8pdqHChsoEMOQ3IkJOZwsE6reRkVGwq+ZnHFfDWhdYutHGh3IUKFyolpC2eARlCsaNvvFCMytwftxKC7dbWM1LrafSKg2boXvzA2lp9NVAa4W/VVwJN17fOWjaC+VWNgmBpttqIVnG7PLkLFS5USkibZTHShEOGvMz9tmY5jnzHgdo6jMqSvIFwNCnFXDdQvDQmfplZcgpeia5u8cJslMsOsV3aeY3NRgXJyZrWkuRkGcyKatSO25hALOLgDcPaWtMmoARGTNuc1hytZauYn7rrCM/ZmXszt+bSaMcZPmyP9d20cXE32apiLfbc59z4uOReLg0l7J3Tc/EGmvINNHz8mMKEz4QuITk2Y6Vg6IYvFW9srZXCdMWN7bWSrcKIL9a7Yv+9vpt8OJ1qnN7r2eQvx9m7n9bp/IJ+skmxqOTiSsY4mJjv9enPP/+Th3mUrjZR+s80n9waYAefIFhtwsDDp9D4sIWAt4dR0SP6hz/lxU8egfkzMVTxb91alqqbZvN3DDeMIXbkGW6l+DSGphku47YO5iv0IunMMZg7yHKeuzyVDywAsLVdYfoOSiwtXstWcLYpJe9+nGuP22XxzxoXl5ad+wRFd3elfIygr3tX5lievGN5xHSF7vIEWB61C7uWx/LZu5en8Ud1D+uWhVhsyULY//ZkTlpXQYQpslnGRb5YZxfYQ5dYN5/mWgSb7GIDgo2PIFcDafZJs1cv8uwCynuJFfarS0GTFtkF9PUSi9mlaWKYCM/ZUpbZBdTiEnNPi1XRXVXZBXTtsvJ0ZSrUCNefp1YchRLOf6K5HSwFw9olmXJFNgSWE1hBYCWBVRJLuXdmymRFHd270PXzbwOBRbqzEwTOOSOapa2ft5GkGpYTWEFgJYFVCgt57kmPjwMr6OgW0/X8bzkDdsLqtjBMLX9tLZtFTNlcj0c+hptMPM7l45Qdxq7X0k1ddlNXBnVzIGO4MFaJewaZW8SKbbrnjghpeF7QnjvXExGUcaPP9NFc+f0NSa5b5PHnexWlfpvTZDghSdx4S/5zCvZkVXU4I2KwxvHRMdguMyjl7h0STrSq40STQ2p8e93hSLFJlH/WdaLp4Sg7abt3ChGQBhLr9jhEq5j5WISSice0Q5LLPrgLQChZJ3XZTV3Jx8Kdbhy2BeZv0adkemjbP3dESBsIrM8dkM2kOxAk8AdS+AOIYmlPvqVoXTS4BAn+l8ItkHSEO5yrQYneYOLgFaTwCjy9FRqF1lsBugJ0RVdvpdUbTBt8ghQ+gaeziu4MphJOQQqngBbNtI56SN2/cERQzY4U7gHoXoHA0qY4wfLNG9lOw3ICKwisJLBKYU2pxkgn68Fnr1Aha21FnxJDEkdLFaS2W6CasQMReWRkMPUcwDqcw4wy9sgC4FxYByE2Kg5dcqNqzNTWsYNu7n683lB79beRF7+NvPTPRCM7lcdopoY5j03SihvtQRKbxQM9UGSr3V1mUXm1UMaMWk6WwNYEhkBd7BktYUhgBYGVCtNyhiZmiqcHWmPEkzGXLp6LrdkhbqWmNwSWE1hBYKXCdPGM+M8UT49KxognKxG6eC62xlHmiudiOdGuILBSYbp4/qpIqAcoY8STsUor3przMosIGwJD7O+IXBBYqTBdFH/hI9SDkDGiEPEI59WIoqdirSTzWjXTwi4CywmsILCSwCoTM/emHjr0nyNE7MBKuKy+1V2jUK10v9EtUshmepWCwAoCKxWmr7UcsBtwhrYvzGwq8sCiZq4uM/Q4yZyNfbTKiozmL4h2fdULxU5P4DvlC9UKw/XXL2hedgGD5mVXMFQrPYlg+Qp/kI0iftHD3GejvGfmITv+isTElQEeWYSRHd8r0gWRP/5FpMtvWfmV9lBkLw25Gz8qFuFidRsufUlxaxgdfLLVLRV7NMnSYdKEMCqUNOXwUcCNu/VGw5Xis1RJ2XW4XIF/T9AUjvK9eWtbl6Q/3lkkkZTw0dyqgVyydRRYNQP5ZOPfD6J7eC8djMOFhzFmydell2YTetznXEnpq9goSbMVoj5foUWxIZNYpXrsK5CoXrBp/PWcyuhGpLPCDJsm69k00ajYhre2No3COrMUshVZWZNC5t6tochZpcJJU+TqMau6UOU11QPqRcjD+1bLGKXTTake9xmsHKuV+1XcmgyeF4lQyIqIgRmWnV++kyX/Xg+Ct7ZXS5Zq9DKJHYdKQuwqMg6NkM7jbRCHQj2Y4q3D7HId+cqfkiEEtxiycI77C8SiyqE6o/ATFf6h+4lKNTxueOyQkR5eMwGs4DRseOY6UkFYNoMGNfcphzpEEVEIkxj+VrebFNbcCuUJFALLCawgsJLAKoXFTr0BPkpbvcfG68yE8sb2rhURUU9ZRZJGPIfDT3UUlRv3Bi6znA5UMh2sILCSwJAzFrTc3zRXVQ/OesUkqkeRxIyySmxdX1vLZp6ySvfjvPtx0f247H5cyce8pCPOoSiGZYv7ziE9GuydOyIEZFuDu7GdvotsRZZV5CEB1fZ4rIq8cQrMGoUi93ob8O089zoUY1/dQjHHPSC/tyHZIGni96AqrwuryPu8EaTaOgo6leLTSrOOUFyL+oprUNwRhoIIkDkDu7jmXNKQrTwlE/nYUzJR1KQvUnRTl93UKLg125j3LRUIJSfqSpBpfPQQu1eBiGoTyykwBYr1mM/xDGQj7gG5GepoAc+AtcGxjsWHqbrAysMzkHVE+9ab6tVh2LCgEr35W4iKtxCVbyFqJoDNJI/gbHeCH8G6t2Muox674988+QzTNtIvIGJ63O7k62v4BRLT/QIXyxVt264gsJLAKoURfoEesfbuWaJYhBOw2bPGgZlaN1Bx57JphsIA7c/CqPI22LUwfYhdsWN9BUA/M7ZjuZCEL+sdgZ+oeEtP5VuIGuGbygnhnzsyGTsWR/1wm80bK+fO5KPXXfo2A85vLWll8tF9a6lB/C2Zh/3dXd28BtXpgbK3B+yMmMLMm+hthtkcg+XwDu7ZyPw39/pxu2JAzle16s75ymZ6zpfACgIrFablfAUWErfTcbga+4JZsuZ6+uDZIFxkzpVlwLV0r8DMdG+wtKzARjQLu2+ry1bsgpA/3UvzstO9NC873Uu2ag9ec2NZbvLgqSQ8Zlz5aCrtevo7cN54UM0ap/Bd8QFvKYW4Bu29Mqko9GuliiparQMVDQSLRZzO7D5zxUHP5ygO6Qq1KxlPYNXjLJ4tl5n4A09Dv4NcWLza8YfSJ1ksF3hf0nrTo1R0Mrkt+1dOfJzE2XIxszdaRYx+HcP9jZX7m6V4hzGbtUM1F9nyggcvMuEQM2+uPyySrbSLOmqecNMVKTN1hzcM0iXeybVeXpAMjPvtYsJwgRMMiACQ27e8u+8UeTRFmsKlSmcLS4GKTg4h4+C7T18qUj3ikeNG1j33Vh0qRdomiNcxPPVYdRaG0Ryvu+oDNtfZctkHrzPhvceD7orJViK6UGvc3i12Fla/ACYzq2KCcON2heuovrKM1VlTllGk6QpFZF8e1jtOpbPWOEtJIa6yt5L5Qu5K9aGvHy6nxX2X0+DWOwcb3kpmhV9tBcc57JynlfSSWM8VNtVMXmGLcYUtxhU2VsnklSqPygo/v8lf3lxsQLcB3UbR+YyzRQhrfQEtvswVYRKH2TKYpZG0ybZJtjjARl9Aiy8LxcFjkmVoIi65gwzqdoklE5J6LbLVIUz0BRT4EptAUEZBFqTQ1ag1LqauGsU1OIE8MItmv2nRiYgmloGC7t9ITLsUr9q1WE5gBYGVBFZJjLgUz65Tqdfk2Js4Q+0UJ7Q2tcR6MrmymX5BnsByAisIrCSwSmHufQV2T+9tIhMvDnFmTk7KTurKVp6krnzsuSsvH3vuyndTl93UlUEtclIJXuxJiNfLDKVhnvbbppGIwDgzexqde/OylUrwigPno9dbVRR6oVdRRauP7QkX4s8sXcSJ/ByJtYi54qWncRWvdPWxPfKW8yiKZ1mylMys4KGQvFQuV/JRZ2AUxMtZEC6klc2siKe0hqMOxY/qUEyXcKlmwXIuR2F5wZUahZbKTUKsfSjP7iCJsiSaZXEgeLQszM1g3MwbY0aIoJi9c+u4tU6eV7by5HnlY0+eV1HTed5u6rKbupKP09bxXCcR5pV4B9KcRuMG4JhpJKJpdgVrwEtzqpl0MZLgYp3g8jF7vY682mNTbECxAcVGUfTrkhybvCwPFlCwy1yx6Fchi0UBFtCry0Kx6FchkwVs3gX2/iXWSUjer0HWIGBOL6BWl9gCgsVgDTJuTEofJEF88Rs8z4TIDEhMfzNPYvodfALLCawgsJLAKoW5d/AT4+7kmG1PBMWcGUt36tfxg4V9vU0186SNk3RV8TbI+8JGrOMlNEJeD6Odbj9Llgfmrze83gRJEkazJXNmxR/riwqDuMT44A8c2nQZNGzsjzcUw7gscOQFmZ0cGUQbLdNFhCzHMhOCWBPcTJ437SxmI5pHizQLZ/F8KdjQX/9IjDum2CGdOV7emk5EY9ebXks3I7fU8q+cAwpgc8KztOOIbuZGhGCcA3DULd+KfYXKkzdvvwXFySyfXGDtyzX4wkFzLqTtpQl828DBcgIrFNZGMCWB4d1Fzo/3akpmO9+dU4QIwUnWSyxtkuHi6wbud5x6Pl+w1vjIuyMbAssJrCCwUmIL7QUV9h2unnFUis79BkzquNjYBuEc3yrT/2COxuUjOFt7mwin/DfPasunnVUXy+UY0J9sVxBYKTFjVomvdvXOshjDgphl23ft3pKuw1oiMGPbNM7aq0mViZk6YHt53R26rl0Jx9Lt0MDMDu3CArMm3m/LqZee4A3Y2lcKzBS1aUco+5gbH6l7lpcCM3sT7aiVHHNNAubPla7B4qbixVPYlWgnxmDOq5Us7F5H9ywp0wYz+xNYs5nM/qyS+sCX15hOWyXPUmDtsVDpiNErUz0V6PfVbnlj07aUAjNW0cTM/qxa8UApsR8cKQVm9ty0c3frwqotd64mb2zL2dgYszcDM+Ucc3FyQZgdgZn9GabI7G/MDcYFYXUEZvYn2hHayD8EKu8z9+4bwtZwBpZZFVhI9TfG2iwIayMwQxtlO6q/MdZmQVgbgZn9iXZUf1Rpot+KLwi7IzBN+3XE3DVjrvEsCEsjMFNK0Y6Qkn2UbrC14Y0tLRSY0Z9sR/U35mbKkrAxEmt4N2eGiRnzuRxjZXhjWz7DkxH9+b2b5Rgrwxvb/QnLY8hnYKZ8Y6wM+1yufTJJzOjPb2WWlm/TabV5Y1s+YXmM/gzMlG+MlVkSVkZiRn9+n2Y5xsrwxrZ8hkUR+8VvZZZjfBreGLGw5sWXAhMWWvQnbBClfyPsS7ls7EY8b68qViZmrFU2wpaUvDFOG523iZm8R9iNMmvshsnbwEzeI2xEmQkPwxi3gZm8R9iDMhPehMHbwEzeI3S/zES8YvA2MJP3CD0vs0Z/zfk2MJP3CJ0u4Zfz2MsYt4GZvFv9DVc9XlCZifjD4G1gJu9WVwfwFnGFwdvATN6tXg7gTehlZmAGb/61fXHI9zNvWluaaYEW+1Y3h7AnlDOY+7UzQOZPuihD2BP62bBQIlmjbzV0CHtCRQOssbZHLfatkg5hT2hpMPeraTBv9XQIe0JRGxa+yWlVdQh7QleDuV9Zg/kIbb1tWpsbE9+bdHUY35R0Qbyp6oL4FqQL4mOPhF2wQGuRR1gGyOGaBsjhgpDDBSGHC0IOF4QchA2yQEuOEVYIcrhmCHK4IORwQcjhgpDDBSEHYe8s0JRD+8Rx77a9Zb92Yh8xuJDrgri96YK4zumCuAPogviAqAvic34G2MjR/LpJ83sVT/XxC/8dlNNkd/jKfpCEFXEVqn61Bb/hgNSujYe4J0Hgt3G0ukUp36VY4wm7nk7wilf4Qr+L3+Ja6C0+oUHwwhP2eROCBh+h4DUNa7xrXBrknK5asW/esx+bwa86fNk/nyaP+I0Z9pssi/kyTYM4WET4TG0UswTRsfkZF1YQOLzwNrjQjKp0ls7TMMtC9rHSz4czfmPF8/ABP5tU44NouMGsUeEDqixIuj/gByo8DyEgG+On+vz1ZXJiP3mDa/UYx+G4x49V8B9Nup4+4ted8OylxkhX7Cd6jh/vmvS4+o2nm/8HAAD//wMAUEsDBBQABgAIAAAAIQDVB0WuixQAAPtwAAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDgueG1spF1tc+M4cv6eqvwHnT7k5jJlSXyVrNi6WlmkxNvd5Ooml7dvGlkeq8ayfJLGs5NU/nsaQJNEE+iWyGytx1YD3Q08jbcHgMi7P/62f+m9b4+n3eH1vh8MRv3e9nVzeNy9frnv//Vf85tJv3c6r18f1y+H1+19/8f21P/j7O//7u774fj19Lzdnntg4fV0338+n9+mw+Fp87zdr0+Dw9v2FVKeDsf9+gwfj1+Gp7fjdv2olfYvw3A0Sof79e61byxMj9fYODw97TbbxWHzbb99PRsjx+3L+gzlPz3v3k6ltf3mGnP79fHrt7ebzWH/BiY+71525x/aaL+330yLL6+H4/rzC9T7tyBeb3q/HeH/EH6i0o2WO572u83xcDo8nQdgeWjK7Fb/dng7XG8qS279rzITxMPj9n2nAlibCrsVKUgqW2FtLOpoLK2MKbiO02+7x/v+/4zwvxv4Hah/RjejsfrH+u9/+7O7xx1EWNWqd9w+3fd/Cqa/jNP+cHanG9C/7bbfT9bfvXcQ3Pff1l+2c2hlX/+sMNl+7/f++3DYf9qsVQwnt9bHf1YN8+W+fxtYwk+qQf+y/nH4dlYOjIpq6Z8Ph69KUkANRlC40/Zlu1FtrreGX+/bh+0L2Moj6D6nv+nyqr+hsMOqtPbfZclz3Tv+fOw9bp/W317Ofzl8X213X57P0BXjQQyoqWY3ffyx2J420N7B+SBSZjeHF7AB//b2O9Vvobmuf9O/v+8ez8/3/SgZBKPbaJz0e5tvp/Nh/+9GHqC20YMgaz34jXrBeBBF4SgKQtA8nX8o3JSRz9vTOd+pcokGIdDaIPwuDU4GSRKnkwslgbpqRfhdl6RSvNI7lFMbgd+lkWQwmaTj4JL7FDXhd6kJf7av/xjtwG+0A23uytLDSKtLD79RNxxdhR240Irw+/9V+ABar2lG8EdpKR6kaTxKVXMQGlJQtUD4o1QN69JfCUFQNkj1R2lmZIVQtjM03UJ3ucX6vJ7dHQ/fezCuQqFOb2s1SwVTZdrbraA/qbw/qcz3/TSGpgjt/XQ+Qn9/mn0+/e6n4G74NLt7V/PVFOxtoG/AlHbaHt+3/Vnv1+Gn3n/853/1Hv5lcDd8h+6+QYvz2qLqusrJgytauKLMFeWuaOmKVq6ocEV/QtHIDCmqXD+7ol+IaAiIVrACktfDqjJbsAKo77MwAUgJVHWuCipXtHBFmSvKXdHSFa1cUUFEpMYwrpEa65E6GcBYITYppQZ1H2GT0nWPwriqO/EBmVwf6QAKdX7ebb7OD2YU9jlMwLJpxMqIQhs8q0asPU7GDbDrTBXYrmjhijJXlLuipStauaKCiAgQagaye+2VYCu1+34Mo2FV8yRq1NzkGUPTrtGZ0DwPJk8CY2Gd55bmWVyRJ6vylCDnRjJJ9Fysut0S89jluR1RXyvU0ssKHeQCJdD6qhLe1h2KYKlmsw5YKrX7fhJWJZ0bydhuV7dhAznUInkaEVj48tQdQtcv8+VJqK/c5LlVvaNs53FjUFminbiqxcpIJgTvlFouMA9M5TW6jZr+nOqOlQaWSwI7KHeBXalR2I2EVDNMmrijGsG90aoXV+TJqjxVi63cl5Il5rEwNRLAtMxToAQWNBWCodUVCVKKZzoNtO2op4zAHF93kbmRjO3mAeviRns1mRLoTtaEPwfk1Gw/+2l/+PZ67u1ee385NWb2hVfz4QrNzKu5uEIzp1XC5Ul2heYSfZLxooHFyuSZIGBoPb/CeoGaOO6i5krWJE0ANLt0FqUGhMcao4wExqhqWjOSZFxJFo4kcyS5Y2eJeaxWbyST2ldhJAFMLor/2esltcDuUkOtp8eDuiMFo+ZsjrnGdeEeUBTVpVu4oswV5a6tZZnLqjuKJrWoIIq08s1F+JXTuaIWjfiiiNQUV6ngX/XZxty5QI3Un5zJyXntD5iADsHz+rh97JudCUie5opVnHZ6c2AeBB8f4Aec3oBlHEUaJVqW1Yp1eYPGtLbCZIDW5xKSp6vSZdCffUCfyt8f/hFsD4PR6OMHKIH+AFttf/AXo5CrBsnTovQT9mdQ1Rvw7diigW7SgmsDbZbddkfWrmEDwm7SuDjnAi0mZ2iQaQcmitpfA3UTWujPTtNaosmIC6QpkBtIZVIPEY5JgzhXDD3O2ToU/I4MJTAUhYBvRAR8JDIc+GJyhj5Y8Ct/im/Pgwh6UQS9KIJWDatHPRc7vQiLzYFvkg34qpdom8qe6iVR2Usi/UHqJaRoORRpxRXpZ6ilaibB+DZOBvUKjUbJy/Eu88jAkCUSJSMiUUJKxUVJTM7QBxulyp9/LFTJ0LIBhHkQQwRjiGAMiMPa3h9BrBIXQZPMjoPoDrQhutqf8qWiG5fRjfUHKbpilQoNCFQJyHUOVVl5qkKD25G3BobOkeAiT7XHPyNiwrNAI+xEJ2rnqA2NyR/cZJqHaqIFNID5zIMEApxAgBMAHWYwf4CxWlyAkcRyE10yXZUuxzrI2qfyp4KclEFO9AcpyBWUvqoVQTItSj8TFegEAu1WiQa6I6kODLMkgUZabQfaiNhAi8kZ+mB7ceXPM9HpUxbPRIfF5gKJfNkJpGoteuHtmeikYkzU+pmf6Ai1hkZ57SrDkFICvhGRIdSIWPDF5CwQk3NMdnuZ6Vf+VQYWmwMfqXYr8Ktqe9rABfAJW28BvuGIBHyks3bLNyIWfDE5C8TkHJPbgY/F5sBH5tsK/Kra7cEnPLkF+C5RDpDh2uAbEQu+mJyhQXbYqfy1GHaw2Bz4SMBbgS8VQ275IaHw14Ov9eguBYrsYQdFHPhyciYn57W/68Evi82Aj8l+csOM+WIxLoBPthBagO9uIei5njJLFLHgizsMmaxtlk0cpWMmXDTJMUtMbge+qYR/9LsAPqH1LcB3aX1oREFc79k9oEyhX+7aLTyyzCPLPfaWlaze+1pVMmvDiurS403CpVvU2LBFODWxt+saG/Hz0OSiINQEugbBlWWoa4OVe+wtK5kNQunXBoGUhYJAqGoLECqqWoMQKdJMj3td9hrK7FVOzuTkHJM5gqOTYbl6C+QmBDYJP4sQ2Ks6qvWSG7THdlGZvZbuFFH/gA6VM2A2YUlfQ6Cv8EFgNnKdisoJ7BNC/W9WnsrQiBP+2iLiFX+1I94405vre0WNgVemtKjBUVo5Oa/9eSltqCitYiFqNw52OKF8EHbgtCFw2pDjtGiUDbvNaVVktU1lT0W25KwhcFb4IEXW5qwQPOCkXJF+hiIpthQFaToYh5PqMh1zHqzrXJ23tQiyy121qUZEZe6KGmxERW0TLm4qhRh6+RO6ZCPGkVd9l28VptMVHKKYVgJbWxDSFJpJCiFNVUhT3IYIU/1BCinHeLWjAhwVlSNYdEFlIeZwMt0YfWiH7UiFQ5cKo4isSWUqjBpsLGUqXPvz987xNK/ggAY+D8cA+xh65xighwMw/6Asc2X0yezI65nS3algC6oHDj2VslsVYUe2rPUanMFly5iLXbbKbFnWzjHZv16E3UJ/X5PpMtpsib/El2GHVtoqCjsSZq3XwN8lzJiLxV8mzLJ2jskt8ZcZM9psib9EmS/gr29Fd5hrtB7FH0X2+IQiDn85OZOTzYVudq5h2n9ZboY0Y3I7/OuKe8j7Jfw70ubIpc0oIviLvHiBGtz8ICfntT9ftTn8sdwc/ia5Jf4Scb6Ef0fmDNukzZsPKCL4ywfiqMHiL2rntb82+GO5OfzFE3Fm/pULIo//egVRjj8wjl17adc9EteW6FIXRezwIx+Jy9pm7dN6+JHPxNFly+ZvH4DbN2CuWP7Amru+9dQCfvesW1tSX42x9o1QRvaNPLLMI8s99paVzNoyqWTWlgnVJevxyCbQLSps2B7ZNgonjYtBc228iUFNoKttI8xn45J5ZLnH3hJlUMXS3KrKZkNg3GI4KAT20WgLCCp2ad0YbV4vnkfuYSmK2F4oH5bK2nntz0tSdLLaPlBH4hHwQvhZRMANI5e36XvNSzRYktHmRW9M5u48VP70ebhxqJwBEY1KIhoBEYUPAhGVK1VUTtRhOFRm5akMjbhNQltE3OWgkXsciyI2vPJxrKyd1/784QUOqr4LpEc6tTEYAQeFn0UEHDTiOCgaZUPMndea/YZoPF2VPuG8BPYbjFPlUIV5jPsN0Vh/kMLMHfKa/QZwVFSO1AYhVGrlqRQNtc1nW4TaPfyNXDqLIjbUMp2VtfPan285c+uns6jEhlI6/YXg+ShyIRYEDnYkOhvZdLYF/MgKrXvS2lJjOSPS1QVqsKtJUTuv/bWBn7JZZ7A0yf7lDAu/xGYvwB/bJ8DXw6/VKJlFkb2YRxHX+uXkTE7Oa38t4C/LjYv5JvyY3A5+sSCX4LfJbAv4XS4bV5Su+poAilj45SNgWTuv/bWBn3JZB36Jy3KtXyzIJfhtLtsCfpfKwiJesVvS+mUqixrc4CMn57W/NvBTKuvAL1FZFv6q4p6CXILfPpJuAb9LZdU3Q5vwi1x1gRos/KJ2XvtrAz+lsg78Jrnl4CNR2Uvwd6OyeoXVGPvdg2/MxQ4+4rl4JmubFSy7k8CsfMpyc2M/d7qtlsps6+cuW2sleeUTdyPWWq0Bv3uxGnORjQSPLPPIcpRZY9mydGpxaBRBc62+IUkUySI77sagtVrjq2JR1PyWPeYiY29Nkas9BMxG9hA8stw1t0SRvYWAIlJ9m8nT6nejk7FLJ1FE6iofaaIGO9DJR5q1Py+djIFOqu+U6+auLhzEQCfhZxEDnYw5OlnWy/TE2BkIbToJbNHYVPaALcYlW4yBLcIHgS3KZS+g7EVVdjgGz6HMK0+ZaSS7scXYZYsoIpEU6eACNdhIitp57c83UTMXDcpyc5Hi2KIh/vFkuqoQBhMQygk0jwmEcqJCOUHiH0/0BymU0oEp3AuTGKZ6dkL1PdUWqwyXYWpLjUWezDBRgw2ZzDBrf21CRhim27lEhukcPZhHM8gFkeHXDaD9eYVWo9MciuwegyJulSEnZ3KyGdjYLy4yPaYsN9NjMJlZ5DHw1xX3tYML8HdjmImhYvazMlBE4K8pZDXNYjYyzXpkuWtuiSJ7mkWRPc0SRTI4q8OGDj1dqzW/kD5uPEpgjrlI9WuGV1fflWWoakOSu+aWKCLVryhZtcgiirT63ehU4tIpFJG6ynQKNbiBTk7Oa3/eVUYSwSpDLQrUSlxdnErgm6/ws0jg27TqERzei1NlvUxPdJ4/w7EtM3WBy1XlEo5CPqBP5Q+mrqT8tm0C37aFD8LUJdetAEdF5QjOQACLm5WnTjTS3Zhb4h5CoohEWr6yjBpspEXtvPbnG8qYKxBluZlIisTN/22EQi6H82UPin434qaeyNJ49AOKCPry9WHUYNEXtfPaXxv0sdwM+vbt4OZZOtxT9R4YyOW4gH43Kpm4V31RRNCXr/qiBou+fNW39tcGfSw3gz5301cPlRz63K1drXQB/W5MVo9wjdWcezCKudjVnHwwKmubCYRdzXEjD7l868wh3LmniD53hnkN+t3YZ+KyTxSRyycoI3tGHlnmkeUee8vSrbVpVGWzdo2oKh1pu1G3xHAcevek+QS6OeaiENSErF7RubIMdemKzuSz7S0xH1nSldlsCIgqgSDtdkCn1XSHs+6eTJprWsxlj4Ao4vqgnJzJyXntz7vO08lqjQfXNObp6OMD/CzS0Q2YZdZ4ZTWZPV1M5u6eVP7gIsQHdKicwQIvHeHeRDrSH4QFnlyponQCd15nAMDNylMZGvFujC11zwRRRMIr329FDW6Ck5Pz2p8/vPBkqRSX8XAVFkIMj5aCn0UKj5ZKuUdLlfXiQsydGZp1fAqPlqp8wgYlhFk7VQ5VmMtnS6XwbCn4IIWZuzNr7p6Ao6JypHYToVIrT6VoqLsR1tQ9f0QRCbV8/ogabKhF7bz251nLQHC9X6Uoy82FUjp/NE9ZcL/KIhdE3htJuxFmrUYXMygi8MuEGTVY+OXzx9pfG/jl80e06d+aYuGXzh/Nk/rYbxKl3VisVmvA754/Yi52HpPPH2XtHJMh3G3gR/bNtX6JxrLwS+ePl+C3aax60uGVN8lTl8eiiDR/mceiBtv8ZR5b+2uDP+GxQXMxjzZbNv/q5NVXkAujj01k2+DvMtnUvUOMIrb9y3eIZe289tcGf8JkXfwlKsu2f4nKXmr/NpWt8G/7kF495zceKu0yW8zFRkNmtrK2WVNxzJadiwmzdaMhUVs2GhK1vRQNm9p2j4ZhurcwK9WPuW48pSHFPHrwnn3666/qeaPTeTpmnir6QBScB6ZCMixrlU31oEAwqizCE1OnD6zFBWsRbfhWT1kXJVMweASEf6bynzct0ROQ1+obEzIGK8CgqDCAHXqFgX6sK4uByV6WTCvo57OyCn/CIsBrc2DQmYHLm9+//cN6//ZPL7//HTzEUv5WeWrvJKjmJb7hQOeG70BYvH1eyuCSS3XD0pNv4ZFlHlnusbf05Ft5ZAXVJZRiTB7pdKmiOnejoqXMrqgn38Ijyzyy3GNv6cm38sgKqksratPkixFVL0dQD1O3I1rKSEXdfAuPbuaR5R57S0++lUdWUF1aUZskXq6oIVG0oigjFXXzLdSD5RsgZR5ZXsose0tPvpVHVlBdWlGbjl2uqGEetKIoIxV18y3UCx+cirqyvMxHKurmW3nsFVSXVrS58hYHI/WuqWoyi8Lm6yBUsrr6qp6Y8jT7COPiRxhb/ftlK53rfTZJbuNgwDxMBQZY67nuF0cQyF2XLmg+xl0b06Wz6+C86sG8eSKIb8fwkqjmY4bNW8fMK5D22+MX/X6yU2+jXmSgn817V0nx7WoxfA9IjdJNuXr+t08ewuNSlXxYm5/dqbev/bo+ftm9nnov2yf1nprB6NZ+sdsIBoujecHZaABrn/PhTb3TTL9h7HCG15SVn57hTYFbeDOKet9Z7+lwOJcfwKdy82l7/vbWO5kXu6kx73DcwQvS9HsA7/sv8MJCSHvbgrepeuvcsXg0j+OvXls4+z8AAAD//wMAUEsDBBQABgAIAAAAIQBXPso9mQ0AANo7AAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDkueG1spFttk+K4Ef6eqvwHjqrkdrMZwDbvGbhasA1c5V7qNpe3byyYGWoBE9szu5tU/nseSS1bsmQP42ztDMyj7larWy2p9XL/3ZfzqfUcJekxvszaTqfXbkWXXbw/Xh5m7V//Et6N2600217221N8iWbtr1Ha/m7+29/cf46TT+ljFGUtSLiks/Zjll2n3W66e4zO27QTX6MLSg5xct5m+DN56KbXJNruOdP51HV7vWH3vD1e2kLCNLlFRnw4HHeRH++eztElE0KS6LTNoH/6eLymUtp5d4u48zb59HS928XnK0R8PJ6O2VcutN0676abh0ucbD+e0O4vTn+7a31J8N/Fjyer4bhR0/m4S+I0PmQdSO4Knc3mT7qT7naXSzLbf5MYp99Noucjc2Ahym2mkjPIZbmFMK+hsGEujJkrmT4d97P2f3r07w6fDvvVu+u9Z7+Uf/9tz+/3R3iYtaqVRIdZ+70z/d6btLvze96B/nqMPqfK99YzgFn7un2IFuhln35mNok+t1v/juPzh92W+XA8Uf78kXXM06w9cRTwA+vQf95+jZ8yVoFgYT39Yxx/YsgGLehBuTQ6RTvW51pbfDxHy+gEWUsHwtJ/cX3ZdyjbzbVVv0vNQx4dPyetfXTYPp2yX+LP6+j48JghFPudPqzGut10/9WP0h36OyrveEzsLj5BBn63zkcWt+iu2y/88/Nxnz3O2t6g4/Qm3mjQbu2e0iw+/03gXKmcD07mfPgkPmfU8Ty35zkuONPsK7MbE/IxSrPwyPSqFQhHc4H4lALHncGgPxy/oAnayhnxWWiSM95YO/TkQvAphQw64/Fw5LxU/ZA48Sk58fX17R+RHHySHPS5G7XHSMu1xyfxur2bbIcqOCM+/y/lHQz+ohvhi5TU7wyH/d6QdYeajsQ6vmDFF8nqFtrfaAJHdkj2RYrpKS6sl9MVYcFDzt9m2/l9En9uYVxlYXndslnKmcJK1qhCODHS94x21h720RPR3dMsQbgf5h/Tb947993D/P6ZTVdTiNshNDCjpVHyHLXnrR+6H1p//8c/W8ufOvfdZ0T7jiQuCoksclklSxPyTSgwodCEVia0NqGNBnVhl9w4MLVqHD74DDtAs8fj7tMiFkFvs9kAphFGYzKY0RD3zGjAn+fjUckOBVFuBxPyTSgwodCEVia0NqGNBml2gO6GHQYdDAO13YVxzdp99Kq84QOv1HBBM4ILCuOMdZqloBkg8gqaiU7j30AT5DTSxqFAxgM+b7DetyIaVZ9JT69rTVy9nGtDCAySazhBSFBX10yJ2GlgSsY1aw/cvMqFQEZqr5q4JcMRl0ZTcoBvo+nrcgIbzUCnCQXNhMWG7OX9wgQ8FFYkp5+3Yi2QsWbuoS55QzSYNwrjFi3VjMsmZmUw4/H6cj9lXLpxBaI1xu2XusGS2DTrlrqufwNNkNPk3TKvXiIrolEsJxBYTtJsCMEcmdvJ9YqBRjMUm8HLhnrtwMZkYNYowmAhkJHaB7DSKnVKQTSAtsocsoDT2QQyf3+Ony5Z63hp/ZKWJgvfyrm8gTOwcvo3cIZ6k2jGC27gXFGd2phQssVa0IzJYCQ9vEH6hjhpbCXOdT2n1gMQTg1ChXFhBa2MQwLBOJRPXAIZjHLEN5DAQEJDzopolD4vkHFR10YgDuYPllCoUzdLzl8/FDAuPhQUMeT0ypO1IBoVii0F4hWK+QYSGEhoyFkRjdJkgYz7fKX1ZjF+txy//cNq3HV6vXdv/DH7it2CtxQ9pSFqo9agWQf9poF1GBf3vjLCjEtD3kIQcY2f5+Uxs67QrysMRGF/Utg4LCAsQ7lKj9sk2rdFVhxOpiFbK6dHnpcuJu+Wk3f+5C5AEPKhpqTbilrHTP08d0rz21pRzqxrPZmuHdYXUZfTnv/y068/+m/evOF1wl2T36Fi9tF13/6x0llUv8dcbbRmM5lueAqNGtz2PJzcrc12aC5mKUsDH3O2kpO9sjUWRFXh5dpSv7Y0oFLN0QrGfFN2nFS5wnNqfZphhbcwcBgiN1Kk4QzGw0c/lUe3eymjunERwp1r2L20FloQVZXdRSZjL/VreQMq1e0u5DGs1CdFUNlMt5INqfKGouPN3hA8GF8tarzgDVsK9/KSkA8chjfKEwFRVXlD5FNV3qgrDUiy7g3B8UpvCCavyhuKFjd7g0RWDFRyzEUxxt+7tYO1emm81eOlWXLpiJwLSxFlMlKSLp5xLIiqykOUtwnjlIYVv5Y3oFLdQzLjtZsmdDxMSSydwSCCOqHduyV+UNMd5FVMS7KhVR5U2nCzB8l2VR5kxVARexhQGR40VdM92Cyn5ZYwlhNGjFH6Z/XRkmRUxVgdb0C8ugcFR0WM8XWZZf4RTJUxpmhxs4dIZJWHqBMNmYf68BAy9toYa5YYOyKlLMVYeQuHqKpijNLSihirKw1Isu4hwWHxED9bCDkPOu+IxdcA8TVAfA0QX1jQWZd9spFV8aVoeLP3yG5V3mPFUHHMvDeA90zV9PiyZes3zGEiRyx5r7S3s3Aok6yIr7pSv5Y3oFLde0JetfdYOUwzYd4bwntDeG8I78nM1lj7USOrvKfof7P3SGSV90hFLBHhviHcZ+qmu69Zqs3zifLwOCmlJQuiqgo+ypgrgq+uNCDJuvsER7X7ZA6EJAiqwX8j+G8E/2FYt0cf7ShU+U9R8Wb/kcgq/0kdWRoF3ZC4vTB4NttKcCiZV/ZKCFJNurRgvgULLFhowVayWmUPgSD0kXyvUFLxbFrvrc22BhyRvWKHWt05Ke8NEJVuADOz9y10gQULLdiKMMzGsrVrgvJdFAd7Ac4EibkzoY0UB+k5/qjZSpFiTYu5zRJtzmakGOXJlahExJn7KVRcsf6pLQ1qS0Ot4sN84fbeLfHju707cFZEs2xTRTSrNR6KXRIhmomFS9weNkvcHv9Su10i6+JhPoe+d2uLXvpxXrPU3KX0U+3bHltyaUeZRFXpqdrcnJjtfgyK0vyIQKuNuceBexxYzoEd6TTW2CiRDalyj5qaK+7hoplY5h6HucfhX+rdo+bscI8D95h66e5plqu7lI8qAy1B2jhjwXwLFliw0IKtZLXKOEOQOtBKKsuw0SzvdemEUst7x8awUWSixYGyifkkTTVUYMFCC7YiTB1oCdIMkOea5R16t1nayNmMtNGIxtq0kWRUDZu1aSPxausiBWM3IhZuH9HYR6D0ETZmbiZOQ2VDRDSWjxfXqo7qYMlFs0jss0js8y/1kajmkohE5IsWnfRIbJYv8otR+vEQQXokFimc7Jy+hS6wYKEFW8lq1UjMs7Z8ySOpLJHYLL9y6YRPj8TyAp2odAMUWVBhABMLLLyhBVsRpkVinvgUBsjzGiMSm2UoLi20dQOUt6yJqiJDqS31a0sDKtUjUctQEIlIQvADSYjEqiRENoTmxfJ9D1ULNRK5aBaJIxaJI/6lPhLVzASRiOTDopMeic2SDzdPPtTd0dL9kAVRVblGOXs0Vp5+LW9ApbprhDxKHuEaHGbiB5LgGmQK1vxQNqTKNdr5aH7uJkQz14yZa3BMii/1rsmPZ9kAHkKntUUn3TXNMiU3P0WVkbkgSB8jLGmRhS6wYKEFW8lq1UEyP9Msxoj8ELI8RnjNshzOVjpRd8s3qhZEpRnAgvkWLLBgoQVbEaYOkgSpyxVJZc4SXrPkgbMZy5XSRZAFUVUlD1RcsV6pLQ0U0Xn2oFWH9YqHJT5+fA/Zg1eVPciWVISiqoUySgrRCEWPZQ4eMgd8qQ1FWQ8ldtBpbdFJC0W2x9nkgqKZOXBJ7NKicqHFgvkWLLBgoQVbEab1xPwcLg9FSWXpiWrmwE6qbzxixvlp+fYOQXpzLXmChS6wYKEFW8lqlZGHIC3wKvME2Klw7muaS+teJS/kosreLY588sWYhS6wYKEFWxGmeTdPKArv5otyY6BV196vaW5+WJPPK5650F5aMN+CBRYstGArwrTmmotvSWXpzOriO2/ua+8iemJtq15GJKg/Ue5pWTDfggUWLLRgK8IG42KvkaB8r9HDKYY3ZOPfkPYavSH/o2avUbbGvNXmqSv15tYS61CMccqV2vLFbLmQZjfZD/MPv/7wZjGaLtAU+0ptyVVjscUZjAtMKJ+GnIbdkcLYziQuR1NmHLtEv1oiv7nD+5KxMA0acQnNrNrX1LWiugb88RXP6tcFZLvGheLpJrcCZlJmhfVouq60gqDGCgpHYvq9MK6Y/Uj6e9JihANiOA+13n17/f32fP3T6dtvcKZff+DhqTkH62T1t/zFytnBtd9i4JGYOouadD6vCd1B4Q0sWJhjhbyVhW5twTY6r75sUFfwLzV0gyd2bPrEay1mU/jlDi9fkI9XGFM8bxNvbdI4yT5k2yySD/b60w3bK+PvQ6dfTsm+6avEUTc57h73eNGD46x7Vs8yvuyP/P0dPQ7kXaSbqzC/P0fJA3+Wl7Z27LY1v1FYoPmjQjyMgUdzasJdqM5v2xZi5vfsceEP2+TheElbp+jA3sV0ehP13WIPy+dEvN/rdTC2ZPGVPdnjD+jiDK/w5F+PeAgb4SkGe87XOsRxJv+ALqyaD1H2dG2l4t3iCL0zTo54/8efuc7aJ7zHRdk1Qm1T9qgy2ezFMJq/yp3/DwAA//8DAFBLAwQUAAYACAAAACEAfF1xaroSAABqcAAAGQAAAHhsL3dvcmtzaGVldHMvc2hlZXQxMC54bWysXW1z28YR/t6Z/geGX9JOxiIBAuDLSMrYJDz1TNN66qRv32gKsjgWBZWkZDud/vfu3e3d7R4W5AmBJ7GSR7uH3b2XBwsc9i5//Lq7HzxX+8O2frgaJhfj4aB62NQ324dPV8Nffn77ajYcHI7rh5v1ff1QXQ2/VYfhj9e//93ll3r/+XBXVccBtPBwuBreHY+Pi9HosLmrduvDRf1YPcBvbuv9bn2E/91/Gh0e99X6Rivt7kfpeFyMduvtw9C0sNjHtFHf3m431arePO2qh6NpZF/dr49g/+Fu+3iwre02Mc3t1vvPT4+vNvXuEZr4uL3fHr/pRoeD3Wbx7tNDvV9/vAe/vybZejP4uod/Uvh3Yi+j8caVdtvNvj7Ut8cLaHlkbG66Px/NR+uNa6npf1QzSTbaV89b1YG+qbSbSUnu2kp9Y5OOjRWuMRWu/eJpe3M1/O8Y/7yCn4n6a/xqPFN/kT//G15f6nHyfn99+bj+VH2ojr88vt8PbrfHn+v3AMBYHY6uL0dO6mYLA0IFYbCvbq+Gr5NFmWRaRov8fVt9OWCj6r8Hz/DX1VC1/QZG5ef3KobVl+Hg17refdis76u/qJF7D9cZw5xw6Ac15P+8/lY/HVUz+Gs1GT7W9WcFvQMnx8r+6r7aqGE5WMOP52pZ3UNrb5NxATPqP9rIt8kC/j+b5dPCO6PasI5Rq9/qmQQxuKlu10/3x7/VX/5UbT/dHcGG7CKDCKshurj5tqoOG5gbYMXFRDW7qe/Bc/h7sNuqOQ5De/1V//yyvTneXQ3zi8kkHU+SNB8ONk+HY737h/mFDp9ThBGhFeEnKk6Li6LIxsUZRbBNK8JPVEzmUYpgj1aEn6iYgk8nbB0ZX3WPr9bH9fXlvv4ygIkFTh8e12qZShbQmg5acaHGuBQ0iJZSeq20roZFBhoQzsNxD916e/3x8N3r5HJ0e335rFauBTS8gfEIi9uh2j9Xw+vBT6MPg3/+69+D5V8vLkfP0JkbbPGNb1F1jLrIsgmtmlDJoBF45VwDJzq4prTAtXRqXAPHnq/TogjM9VLO3Ca0akIlg5i5sBR0MFdpXQ2zYq4HtIrbmya0NFAyTZ3UqgmVDGK2wQANbUsh8MIYySBgZowoHRhVE2hVjREdyGkaxNELuTgaKJuOva1NqGQQsxXGpGSrmb9uYTgzwFUjaoADyTvj80lgvBXysV8aKB3PvPEWmrO58v1DfawOg2M9+Hj4/rtVZqbN9evDYH0cTOB2Y3KRjtOcX7GMaatsaQtgnHEsXGrRJQvAyWmvZKFL55kfbAbKpjAaXKCSIujmpZUi489CZiVWA7e0kG6eGQmTMdpIJRsYiRAZ/ksDZdPcd5SB8pmHSgsZFqKLi7r7iw2bkg0sQohaZCBmkYHy2dQZWVqoaRGMr2iLlGxgEULUIgMxiwzELLJQ06IE7g+iTdLCgU0Wo0YhxqxCjHWdwwS7AtI7TXWGYNigTxBjdhmM22UwbpfFBLsCxjptl2ESbhdizC6DcbsMxu2ymGBXQE2n7UIWoYtEgpTEVok5X9uWTohMScTy2UzfYHz45ac/rKaLVZL+EdfLcbBAOoW5UyincMMrKLAFJhEYLskvYEU67SzyF3F2qZsCSja3rpoLV4jlc3PjqFc7h+lFkVsTcNhpGwwp8IFgsHBZDghsmVgxGnKD5fOJZuxGgPHXsyL4NXfgJaySCLRiMTaSLT9Qaw2Wz/3dQom6uekUbtdLiCQRmAQxHtjGSLaK1FBLJX5grpIchnLeOpSbGiVowL9NDe7kS7gpMYQC9xqe0xGjt4tLAVshls8JYzpMWEYChtK5RsQcMwQDeRi5xcggz+D5g5Xy9ybLxGK0H5Cu5lO9QKySyQ+rZNq2mnDhEoRLQZhnHS8hvVQJq5tkckclYEsBWwlYyTFul5TpnY++us3Xd8Iw0fxt/CxMh7yUz4ea2Iq0ZuVKjnGbpRQuwmZDZemY5nCNu/fUSvl79aWArRzGb+BXcG+uMt1zt+wt6mWLestdeiplhxGR8GTsey8tgrTijW4dKAsmhe8+o+pzm5WVmpLbUcRyoAP10Iml390oNfWUetJkTACZyQajJts0kZpsMNFkIXeMuAtQT3nwdvqkyci2zGSDUZMtJ1OTPek2ohwwbeSimnrOPWkyUi4z2WDUZEvM1GQkZmlgBCQca7Kn45MmY2bHTDYYNdnmf9Tk9mwvDSg11mSf+Z00GVM/ZrLBqMk2QaQmt6eDaTe21WomDTtpMuaGzGSDUZNtBklNbs8XJwF1RkZZq503GcXYIocYMdlK0UUOMWnFmHRjVa0WYTLmlDTKqEpNtpkniTJKiSZ3I9WJzzdPDQwU41E2qtRkm5RSk9sTUPUMsfFs9Dz7abWIKGN2yqLcYD9sLGMDw0iJUe7GfnAjGEMlKMaj3GA/K8VMbme/STf202oRURbYD1XpwBDYD6XEKHdjv0kc+6EYj3KD/awUi3I7+026sZ9Wi4iywH6oSqMssB9KiVHuxn6TOPZDMR7lBvtZKRbldvabdGM/rRYRZYH9UJVGWWA/lJKinHVjP6123mQUY1FGjJhspWiUERNN7sZ+8PY5ZpFDMW6yUaUmC+yHmqLJ3dgvi2M/FOMmN9jPSrEot7Nf1o39tFrEwBDYD1VplO0TXELYKCVGuRv7ZXHsh2I8yg32s1Isyu3sp95ld7jH0GoRURbYD1VplAX2Qykxyt3YL4tjPxTjUW6wn5ViUW5nv6wb+2m1iCgL7IeqNMoC+6GUGOVu7KdeHUc8FEAxHuUG+1kpFuV29su6sZ9Wi4iywH6oSqMssB9KSVHOu7GfVjtvMoqxKCNGTLZSNMqIiSZ3Y788jv1QjJvcYD8rxUxuf8OYd2M/rRYRZczqaCKFqjTKQu6HUmKUu7FfHvfkE8V4lBu5n5ViUW7P/fJu7KfVIqIsPPlEVRpl4cknSolR7sZ+edyTTxTz5i0byAoRll8jls/1I3q+y0fgPrUHbs7+yK9vc7dHKfdUaB98v0Esg1eK5EHH1D+YN7vTUIxu9LGN6WfhWqpEKTHkAhV28CDuqWhuxfyrsCVi1AMrRT1ofyqaC8zYwYM4otQXU7ODemBUqQe2MeqBwQoYTeHT81wgyg4eGKpTb+BPPRrTFws9MKrUA9sY9aD9mWkh8ObLPdCtnF96nBjpA8SIB06KeICYNA8KgUY7eBDHqvpiQR8gRj2wjVEP2lm1EFi1gwc+xXRrkW45NNfIUXN97qi24epdMulksSqStr0F2K7qDqtRgkYpafDdggIbd3DVkjPMTPdSOR2H7/QLK0aIAzH4RsJv3EPMR6RERBxuAjl38MBnqr6zLEZnh8FoZ1kpOrYwLZ1rR3nABWLuYK7naW+uxai5BqPmWimS+BftryOLXmhZt2KWI2+uZVdqrsGouQIHY2viYOiFgwvPwd5cgXBRjporEC5Kieb2QriFJ1xvrsVodBvs6jTN5hm90BQZLDRZ60Ljk1S30IBGKWnwcd8thYV98CrrPrNZyEmRzUIOI5uFHGaWyVWR/rAqYO+L3vcR7oxD4QJWJuVqCcKlIMz3Nb9ksxDsiDebhfxUfOMw4omArQSs5Bi3S+Zk/zlL7D56teMz3OAkYEsBWwlYyTFuc7fcFnYYNjc4peNwiwwRc1tkBGwlYKXDdMdxo7tlt1PDjGd2ODkpssMJsQReqlk/Vk4OJg/5nAd2PsVtcWrRL9v0W/Y4TWVm7jDo4h4j6+sFN1eIZSbjNJt3ESvGZPMuYlJmCp8UND9AGV90cCMuudbXC92wj5sJ16AcdwOZXEiwYftyT27EPW7W1wvdMKq8NwzG3cAHzpIbvTA8PPiJeZjrxAiLWgxeYPoJZ5rjPjCMLxK90P40Ls92YtQHVGU+YFbNZgXDuA+9ZNrTuEzbiVEfUJX5YDDeDwxjPsx6ybV1K+dzbSdGfLAY9QEx5gPHuA+9ZNuzuGzbiVEf7B0BmQ8ox30wcohxH3rJt2dxr3SdGPUBVVk/4H0EnQ+oK/rQSyI9i3vK7cSoD6jKfDAY7weG8X7oJZWexbG1E6M+oCrzwWDcB4ZxH3rJr+GbwRh+cGLUB1RlPhiM+8Aw7kMvSTc8C4ryQcjDrSrzQeBplBPnQy88DV9KRvkgJOdWlfkg8DTKiT70wtPwrVuUD0LGblWZDwJPo5zoQy88PYvjaSdG54PA0yjH50M7T6s3N41PwF/8bki3cp6nnRjxwWK0HxBjPnCMzWn4UrAPH+J4Wl8suPe2GPOBcbJ5w4Vy0lia98LTupWIfrB0TvtB4GlsjvcD427eD73wNHxMGTOnnRj1QeBplOM+tPO0ei/Vw3yI42l9sXAsCTyNctyHdp6G1359+BDH0/pioQ8CT6Mc96Gdp+e98LRuJWI+CDxtVdmcFnga5cQ53QtPww6RqPkg8LRVZT4IPI1yog+98DQ8bIjyQeBpq8p8EHga5UQfeuFpeOEU5YMVo+uSwNPYHJ8P7TytC0j99oXJNHN+Rng54oYDaV9YkDkSgIwlEsj6elieTDMxnlheZ54ImbVtMfCkPbdO4EPwXjyJy67N5YKl1oG8T4QE20pKMyRRH4z3MbriuNtcruGJwN5WMuiTdv6GSmj9eBLH4OZyDU8EDreSgSftLJ6Me6Fx00zMPLF8z+aJwOS2xcCTdi431eqCymkvfzFvmonxRKBzp8zniUDoVlKeJ71QeqK+to/Y8OzlWJ+gMvdEoHWrLnvSC7EnUL0rzhOB2p0y90Qgdyspe9ILvSfjOH73cqxPBIa3ksE8OcHxYREo9VFvh3lCykOd2mOXODnqiQVZnyDIPeEg53hd9em3z3hbPOrMbsHEyTFPJI5HycCTExyf9MPxupmItcvJMU+E3DxBycCT9uw80fWeeuiTSI7H8lJsE6qxAcLAR5fwKN1KijNeqkDVZZ5EcjxWnwo8kTgeJYM+OcHxuqpUD30Sl6snWMQq8ETieJQMPDnB8bqSVA+exD1ZT1whLDZPUJmPLonjUV0eXf1wvC1/pR76+Y2b03CzTOJKZzFPJI5HyYLUnyqtuuxJPxxv61mdXYUljrfKvE8sx/tKWuBJ+xvxRJen6mF0RXI8VsMK5onE8SgZ9MkJjtdlp367J7Z61bk+cXJ0dLnSV+S9coIg98SBplA2rVmU6GJUPXgS99zdXC7MtLAgFucTWyWLzRMHSp70w/FYvurctw2Jk2N9InE8SgZ9coLjpdpXHZgRS1Wd98TeCzBPpDweWww8OZHH69pWPYyuSI53pbSYJxLHo2TgyQmO1yWvevAkkuNdhS3micTxKBl4coLjdSWsHjyJ5HhXeIt5InE8SgaeMOLn+YkqWd7D8y7dTMRdvZOD4Pp7gTnUUqSFGZeJlWNkiWDgXPsWuCSswHWmNi7QrrepUe9Ot9ZuM9SXlfJ1Aqq6f37/d9ANQuoeU0oWK2+dq3NpC3TBfhi3ETmxINSicZsMLWg3pJviubBbH7xrrTmKDVGdcqoL6Ao63O+O1bsSW74LSqO4zxEkEAoE4wZ4qEjivRRAqNNpJQVi7FiyK7E1u7idNh8nnyg5SW4nShIQ7LSgZKdA4DHjyFbgUpvqyMdV4QkGCZHzA0kAoUSqIecMWrSSYDoDg6EgPHmPMt1w55lt7fpIAf2RB9nXLoFguW0Q5iTd2T6FSsgxxTvbGijbGmjZ2q4KvHapLmL0zGJMJoflf8ohWOyLfAJuleHDOrmUM6o0fx/0pvDMXuU1UScq+G+t1ScHwmPVNKwl6+WYe0aZ3txje8p8MiyNHIKBI8L796hhaUg3U1uj3YxqViB2hcDYWmB1WccYMIUTb9wHm9BbsDRP2stBY/NUCWahKggtKQWeC7cGUZ7blJ2tzU0Q1mYE+ZrXBMFaC0prnpDcR9mJWTpfm5sg2Ikgt7MJgp0WlOyUOH7a4RuPxFYc45bbrJyNJAS55U0QLLegYHnHKmOJLTPG7BRAKH8vsbQAwilUJ1hal/CKPWBDnWfV+MxMAsG6piRU52+C+ows0qaZT+ZULXOA0qHeHz8c18cKz9WawffTesnVx74tvt7vb7oeNjYd7bebuxs4pylV52XBhZb1w81Wn5mlz8fSF4JFb+SMuL7cVftP+iytw2BTP6ljrtTa61By+pdaLkM8XcDRRBIOx4XBYwbhN1NQMeX5w7byxRu9dI+8Seawsp/W+0/bh8PgvroF88YX0/GsKCD600maziaZ+k5ubw7uEn93rB+1VjaDM7myeQHna83nqXom/7E+wuFcLb+8g7P1KjhfBnaAEq0Uvt0CHr2t4Ryell+C03jE2tPj4HH9WO0/bH+FA63U9FIHocEBQbBG6JPX7PeQMKLr/RYOGdPn7sHZV9B3+/X2CI4t1CFv+3c3ZlK6UwKv/w8AAP//AwBQSwMEFAAGAAgAAAAhAM5G2LKmBwAAGx8AABkAAAB4bC93b3Jrc2hlZXRzL3NoZWV0MTEueG1srFlbk9o2FH7vTP+D44deHhZjGwOmQCZge5qZpM2UpLc3rxGLZzGittlLO/3vPUe2ZMkygdDdSZbl87nrHOnIZ/r6KdsZDyQvUrqfmXavbxpkn9B1ur+bmZ8+Rjdj0yjKeL+Od3RPZuYzKczX86+/mj7S/L7YElIaIGFfzMxtWR4mllUkW5LFRY8eyB6ebGiexSV8ze+s4pCTeM2Ysp3l9PtDK4vTvVlJmOSXyKCbTZqQgCbHjOzLSkhOdnEJ9hfb9FBwaVlyibgszu+Ph5uEZgcQcZvu0vKZCTWNLJm8vdvTPL7dgd9P9iBOjKcc/jnw3+VqGK5pytIkpwXdlD2QbFU26+77lm/FiZCk+3+RGHtg5eQhxQVsRDnXmWR7QpbTCHOvFDYUwjBc+eSYrmfmP/365wY+bfzVv+n7+Ev6+decT1mefMjn00N8R1ak/HT4kBubtPxIPwAAuWpa86klqNYpJAQGwcjJZma+sSeRM0ISRvFrSh4L6W+jjG9XZEeSkoBNtmmU9PCObMol2e2AGZz/m9JslcS49p4nff0JExpoFHCFhfAufqbHEjVVT7FCbim9R+QtaOmjU0wnWhknZfpAKn0rGzQUfzHD8W/hGLJyJ2UXIlZVEI812cTHXfkLffyRpHfbEnwZ9AYQbUzXyfo5IEUCdQLKey6KTegOwgC/jSzFeoc0j5/Y52O6LrdguNPzvMFwPAKLkmNR0uy36gELt2CEADFG+KwZXVjjzzDAU8YAn5xh2HNdp+/azudVgTeMEz4bGwcDe9AfnuEEuYwTPrnOfu8izmHNCZ+c0+sNhydVWlVcWaoFcRnPpzl9NKCgIcDFIcbt0Z6AMFwgx+2h3dt0vSbVElRhEwvYuXoDSJ4EJb5BkTNzaGPSzsyizOHJZn5bvHpjT63NfPqA2+kEtCaQurDjFiR/IObceG+tjN//+NNY/tybWg+QVUktcdFIxAxBJUsdCnQo1KFIgSyIgggFpMqJUHyZ/yjnZf1vJAr/dSjQoVCHIgVS/IfMfxH/Uc7M9IesnHGxFhqy1JBAQ0INiWREMR1KTzPd6UFn0LHNNImKXNJCQZo+zB1v0Eq+hkoEX4cCHQp1KFIgxQPcX/U6BPpymyb3C4obZ7c7nig8lDEzYctqIq9DSx0KdCjUoUiBFONh47gi/MiF5tb7BAu/57ai30Vk+55KteRUkNy45TBRtj9UqYJLFIYXKYzOKVTCM3qBtUUZGKzqkGRVxSFMEu600w7gspPKHbVCo4sPLxIfnROvBAJb9f+b5CgDA8GaEHYWLDgEtkir33JxeRFVwKmaOIcXMUbnqJRA+FcVDHKh66BK8hMOVeWwrKlchx273+5pSQroHo3DN3F2+GH37avAHlfn8LzfKqOzrOEp1qBi9cZwPZNsA0WybWFDxffSSIGUINkg64pthbFBmGy/2QY5Bj1g00GgeKCDQHEskOikmhq1QhwKstMxhr0HW512iKPzrKHOqoalq2U7f9hhO4buKmGpMSUsHJPDotOFXJ7EGwmM8apWd3VXF1hdtSuurea8024QazJ3IK0vx5qdIrA5JpV3BxYJrMORrjbpAkeq3gXDL2dWqwIXdk32mep16i5aq97zvOEp3kDwyoeJ7fut+pVUiAK+RC3sxa1qUJPjuv7NrtopNaVrTKrqJadTKp3TQcCb07PfWo9QsDYJE6niVEe62rgLkqNurZTarDGlNjkm16ZOF7LbOm5tstUKr2r1df2bzZsgOaVtv93BcTK3uRAsBSY7UouTjA4FnewIp+uoza5O64Lw8yZGPjFqTLGaY7LVHJN3FB2LbIVXDX9XW3SB1bzjkK2uMcVqjslWc0y2Wsci3HabM1K1+roeBvc/1sTIVvO+Q86QphdpTmeOyVbrWMR1VPWu3vGv6yngSqVZzTE51gKTYi0wyeoOLFJ5VauvO/LxZsVWT27KnPamvhBkp7sZBy7FXe3MUuIVyyQw2WG9g4hUXtXh67oFhx/vssNwPLS6BUHWrNJSwjratkA8ln3Sm4lIFa36dF3jADc+fRHhjt/2qWkcmndEcjOhtaKBkCz7xFmkDVfQ6Ruuc93BzdhavSjH5PNOYHI5NQc3dzTs4I1UXnUlrjul8aVovXXJd5z2+yJOVrdwoTO6WX16/93Cnywc9/tTddSczviyFCYDjGnpT6DETjAFtSZvXL36xpcCYQcWCaxjAbuOfhdf0H/BeyeHn8iNHQuBSTcugcnLWfF6YzkJOdbIi2peb9zhw0u8X4FRTLW4oFv0g4P23ZpTuSN2uRbrOjy5rvzYHwsGtqanGIJagzeWG/G2HZBT9abgC7GhPwmdU2Jx0lTtIg1D5MMAqoOhKpVqblWNCgqal6syLkk9uXI8yEoQyOaqk6ddvr52mjey8jTZrmEggWcP6lnS/TrFWSnXxW50ljBhPs1IfsfGUoWR0COOjlzgFWgzXWPZ08YHE3gJCxtHG3cgGGys1YivBnvv4/wu3RfGDoZvOKUCv/NqkMX+hrEcQ6F+b2kJ0yj+bQuTZAIViXMtY0PhLUz9BZTXA8PjwSjqAR4cjzRPYRDG5sQzcwcDbXh2IKBtglPJ/O26mieKsfb8PwAAAP//AwBQSwMEFAAGAAgAAAAhAJfEk+eoDAAAGjQAABkAAAB4bC93b3Jrc2hlZXRzL3NoZWV0MTIueG1snFttj9u4Ef5eoP/B1adcg5VFSqIkY9eHxLYk++6Cw6XXFij6wfFqd43Y1p7t3SQt+t87fBNFjuSXLJL1+vHMkJw3coby7Y9ft5vBa7U/rOvdnUf8wBtUu1V9v9493nm//y2/Sb3B4bjc3S839a66875VB+/H8Z//dPul3n8+PFXVcQASdoc77+l4fB4Nh4fVU7VdHvz6udrBJw/1frs8wtv94/DwvK+W94JpuxnSIGDD7XK986SE0f4SGfXDw3pVTevVy7baHaWQfbVZHmH+h6f180FL264uEbdd7j+/PN+s6u0ziPi03qyP34RQb7BdjeaPu3q//LSBdX8l0XI1+LqHfxT+h3oYgaORtuvVvj7UD0cfJA/lnPHys2E2XK4aSXj9F4kh0XBfva65AY0o+n1TInEjixph4XcKY40wrq796GV9f+f9N1A/N/BK+K/gJnjPf7V+/ueNb+/XYGG+qsG+erjz3pHRT2HoDce3woH+vq6+HFp/D14BuPOel4/Ve/Cyz79ynVRfvMGxfv65ejhOqs2GC4GZ/Keutx9Xy031gfsmoBlpgR+5T/+8/Fa/HPkYEBMBxAT39k91/ZlDc1hFABM8VJtqxf1usISX10oOkZMMIuYPMWf+N0x42My4/beefS4i5Nf94L56WL5sjpN684/1/fHpzqOhH0UkChiNPf3pb/WXslo/Ph1hYrHPQK/cMUf336bVYQURAVPzhZZW9QZGgN+D7ZpHNjj08qt4/aKkt4Qfjt+4k0feYPVyONZbNQHCZ9+IAI8QIuBViQhhrJAGIeETvEgGqF/IgFclgzA/jiOWJpeKgEkKEfCqRSQXjg5DCFZ4bViv1QKoXMiAVy0D7MCUlS5SAkxXiIBXI+I6JUBWFiLgVYnIrl0IuKn0CO7e0iVIeq0Q0vgVC1NjEctzG53wUJBuKQJiujwux7f7+ssAMh/IOTwv+T5CRlxop1uDM3Lad5wYCKlwT468R8gEIVOEzBCSI6RASKkQGWV89DmiWbSRISyxWScEz+Xr5MT2OhEyQcgUITOE5AgpEFIqpLVORLNoI9Y6IcAvXycnttcpERaBS0FKO0C2fR0n0e3wlbuQcoOJIeJpiltiiqEZhnIMFRgqMTTH0MKCLBXA3C0ViJzNfFDZ8Wm9+vy+5jm829NjWLB0dS7EVo1EbNWkiaMaQ9SoBkMzDOUYKjBUYmiOoYUFWarhuf7iaOfEtgoUAvmzcY44dDSgaCBBNjRpatNMFQ1kQkOT2TSzC2jyhkYru5AIDZoMVSoEDN6MlQX2WHNFY/LaomP0jDRclk5hQ7pcp5zY1qlEqIn2iUT44cjMmDoaVFztMM0cS8y6aJxQzrtoYnusQtKE7flERhciXEolJ270PlcIvJhVMFvyQtIwHpk62WRmpZaW+VHjYs/lxLaWJdLWskSsVWVONE8Vl6Vlx5dnF9DkDU3jpxJhqfFThbSyPkIWDdJKzsbqlr54+djW13clQS7E1qNEaNLMe6IQFe/HPWTQh/G/yL/fffgw+6f/l4LE4E8PfB95t61fdsfBejf47eA7/qykqIygpExA1+c4Z52c0ws4807O2QWcheRkyi/UbPMLOEvFCenFeLyT+eY2jZJeXCB90ck5P81peQ0voS6OMk5se4dEQpNHJwox+XiKkBlCcoQUEmHMxItCjCfOEc2ijVjrJHD2vuLYy6mdc6+E4nbqIoljyYkY5s4LW8vH0AxDOYYKBbVVoKGWDjDVwoJsLVx3+O84/UsIDG7cuUMLiqqthQbiiYESlgbM2ZRnvBoBrbeUl2OoUBCoheedCSFvp/AfmG+AWqWPTvml5mxrTw6phL1R0n74K9AOoR3x9g3I5W9+4K0qJdyZ9sKeEMzvZt47EdsaV5UoUIQhn5SQ7ZOBs+dOFKPlk5KxBc0wVY6hQkGWT0pZrK1VBZngXViMthauKmB4W8mNTFUYQIA2KZYEbgmjGGHJ3AHTJGV+4O7+iobJbgwvc2aGTe/kOYYKzah9MgSfDMEnQ/BJ2K7FltY5ZKk5EzEt4p5TLcFieU/LfXXvyU7dnISjOeU5by0aZeC/YmTuv6H235C/afsvTMRnAcuaH0cNC3s1sDhwaL2KJAsSP2K0g9m2qluTnW4z4OKLSMjONIFbe2gqZdWMBX4Um/YmMrASahLTzEgwBkZUhaKSeQLZAT4eFZSfbcEOxIOkFIEDROAAETgAeKJ0gDOzK/UgyhecU/m8dw7S+jD91zEJo8gPWNLowDm2L04vBD4eLURbFRZCvTEsDGyvF5AwaCGGKW2Em0xj2/6qopPgqlNBTl5zK29FZeU1KcvKawjKMSOcWUVasfKagtp5DVEtMONPCkpkR7rdkiJXVY6C2jmEqMrM2n75xmR1axSjSnUkolkCnV/n/G2ItN/PMJRjqFAQDgXp/NINO8csNWufg6vqUGTRdpC1HJwScHAWGh906uTFyemJcypECSUx9TOSGKXYDmzVnrCvnE5euPgkqta07JQiOykqmbxoQIPYj0KTl12LWeRRwmJocTlmnemRTYbLMVQo6KQNGWyRxGfOCbfUrH02VBXrKRuSAI6MMPVQJxKUpE4JUTaME5ZSUBfTQnrKYmLVxWdtiQtgIQBOpbYtnSlPNJW0JWGMQMy5BpTCreyEoNxIanoHCjodcl1Dlpqzz1qqfjwZcVFGfbgYczo5J+ek4ywiNPLhwqm7jyY2mqb4PGsaXH0KAcg0TkaYaCplmjAmgd+akmxnGyKTDnGZiqkKBZ22TeeYpWbtM05TzEKh05cOA8p8BreCrnVO8WrrkCiNfBI2h4Wedif0Vlul8zkzCWrnKkWWzk4EuUc5xah2rSiIwjD13Wbz1FA1dsJQjqFCQSft1DNoqXl7DHVStEjEr2M4cMfgdkmm05VTnywsIWOY7s2c6oozhDSXJuC0geLuaZlSq7w/aylc3gsBKKDcUkpTtcp7DM0MJA6mUQp32I6b5pitUNCZ40WHsFJz9lmpXer3hBMhhPo0TaJmE3biypodWAlqfaqbDhBJcK1L28FoX0ha5f5Z6+Byn6qq3d6JHJ1ONJVMdzFjIVyaOzuRoTFRhFoCOaYqFHTSOp1DlpqzzzpNy6A/2aVR5sdRag5ITmd+Yc0OrEPBOhAqovqKUgLJjgQdxyvbSlY74qyVcDsCrhxEF8u2ktuU0VTSSlGQEJ86R62poTFWUsJbxztMVSjoTK7rGLLUnH1WUq2WU+cFBtsry0ITQm57wZodWAnaC1S3F8IIKvg4aw51Qd+eZLUXzloJtxdoZ3sBxVLTCRA5DJofPnF7ClpSq6eAoRxDhYJOZ7quIUvN2WcldR188lTHYHOlsemVuHWUNTuwEjQCqG4EQPUElUFkOi19VrIaAWethBsB4mmnO89pBLgHPEXVbgRgaIahHEOFgtqNAA21GgGYamFBdkaxqv6zWsAXxqKz5O7KAdKC6g3IjJKkCTwy5l4ia0ltX23Ymv4XppLNrTvvpK92DlkqYdAe5iEUuQ1O/fEpXw3hoOoTmvQ2LC0h4KsMfBUSrsz7CaU+6btkhnvNa862uNIXApCHOhvTRFHB3XPzkIiCIMxMcke3xjmmKhTUvknWUPsBoqaE1uIXFpXtoVfVyBTXyApymrVujayp1J4XRxkcat0HaqaGyigGV8mYqlCQvhii6VuQ9XZG05uc6hvRqHvQUvP2ZdQL6mRoR3A3TfUZ3Q3QhT0/mC54qZ5XmLIkzHyWZmrfNB5kW8q6r4Vc8l23/OLKwKnRVHVtMsNEU7WdVlG1LkgwVY6hQkFM9ETlAyQKSsyI8ioDskzGr/Z6juju7Zsl2dIV7Ad2bH+XroQU55E5Wc9m1p2+04NVbDFfClw3pTSjmZ8EmbqbcIrfaZvcWrl4uhk+Hi0ECbTlQ2/88fdf3kxJMJrS7Ad94wktzQTa86gh2CuZN1VFwoB+aGdxdgmnY41C8WRRlwXVaCHJCIvgJsE5bZWXMMMynUe5LuEicZjFoZ9FBBXQwhmlcuH59JPzDsKEQS8sCqAQ4T+mLrb97qoqnD+P5D6SiaAJpppiaIahHEMFhkoNmc1oriHzgNdCQ+Ja116yKZrYCA4Up/rlOXy3gK84EY9CSTHyuX35mPK22j+KJ/wPgxV/fEhcRN02qLz5fB9CQPCd08Xh/kosYmjEjG/5dxV+We4f17vDYANfUeCP7sP+uJdP94u/4csLAoWY/lQf4cF8/e4JvkhTwU0wf9h/8FDXR/0GBudyP1bHl+fBgX/LAdYE7PV+Dd8OEF+TufM28H0e+Oy5gtFG/EsZ+/m9SJ7D5ls94/8DAAD//wMAUEsDBBQABgAIAAAAIQBE1Ib/wgAAAEMBAAAkAAAAeGwvd29ya3NoZWV0cy9fcmVscy9zaGVldDEwLnhtbC5yZWxzhI/BisIwFEX3A/5DeHuT1oUMQ1M3IrhV5wNi+toG25eQ9xT9e7McZcDl5XDP5Tab+zypG2YOkSzUugKF5GMXaLDwe9otv0GxOOrcFAktPJBh0y6+mgNOTkqJx5BYFQuxhVEk/RjDfsTZsY4JqZA+5tlJiXkwyfmLG9Csqmpt8l8HtC9Ote8s5H1Xgzo9Uln+7I59Hzxuo7/OSPLPhEk5kGA+okg5yEXt8oBiQet39p7rSp8DgWkb83K9fQIAAP//AwBQSwMEFAAGAAgAAAAhAAPTAOs7BwAAiiAAABMAAAB4bC90aGVtZS90aGVtZTEueG1s7FlLbxs3EL4X6H9Y7F3Ra1cPI3KgZ9zEL8RKihwpidLS4i6FJWVbCAIUyamXAgXSopcCvfVQFA3QAA166Y8xkKBNf0SHXElLWlRtJzaQFrYvWu43w48zw+Hs8Padk5A6RzjmhEU1N38r5zo46rMBiUY192G3k6m4DhcoGiDKIlxzZ5i7dzY//eQ22hABDrED8hHfQDU3EGKykc3yPgwjfotNcATvhiwOkYDHeJQdxOgY9IY0W8jlStkQkch1IhSC2r3hkPSx05Uq3c2F8jaFx0hwOdCn8YFUjQ0JhR2M8xLBZ7xJY+cI0ZoL8wzYcRefCNehiAt4UXNz6s/Nbt7Ooo25EBVrZDW5jvqby80FBuOCmjMe9ZaTep7vlepL/QpAxSquXW6X2qWlPgVA/T6sNOFi6iwXmt4cq4GSnxbdrXKrmDfwmv7iCue6L/8NvAIl+r0VfKfTBCsaeAVK8P4K3m9UGy1TvwIl+NIKvpyrt7yyoV+BAkqi8Qo655eKzcVql5Aho1tWeNX3OuXCXHmKgmhYRpecYsgisS7WQnTI4g4AJJAiQSJHzCZ4iPoQxU1ESS8mzjYZBUJOgzYw0t4nQ32+MiRndHg/JhNRc+9NEOyLVOub169Pn706ffbb6fPnp89+0bUbclsoGuly7378+u/vv3D++vWHdy++SaY+i+c6/u3PX779/Y9/Uw+bSaP17cu3r16++e6rP396YdFej1FPh3dJiLmzi4+dByyEBSrrmHxwL76cRDdAxJBAAei2qG6LwADuzhC14RrYNOGjGPKIDXh3emhwPQjiqSCWme8HoQHcYYw2WGw1wH05l2bh7jQa2SePpzruAUJHtrmbKDIc3J5OIIESm8pmgA2a+xRFAo1whIUj37ExxpbVPSbEsOsO6ceMs6FwHhOngYjVJF3SMwIpFdoiIfhlZiMIrjZss/PIaTBqW3ULH5lI2BaIWsh3MTXMeBdNBQptKrsopLrBt5EIbCQPZnFfx7W5AE+PMGVOe4A5t8nsxbBezen3IYfY3b5DZ6GJjAUZ23RuI8Z0ZIuNmwEKJ1bOJAp07Gd8DCGKnH0mbPAdZu4Q+Qx+QNFadz8i2HD3+YngIaRPnVIaIPLNNLb48i5m5n6c0SHCtixTj0Mju9ZjYo2OxnRkhPY2xhQdowHGzsPPLAwabGLYPCV9L4CssoVtgXUPmbEqnyPMoRCSlctqitwm3AjZAzxia/jszM4knhmKQhSv07wLXjdCF84xayrdo/2xDtwlUOBBvFiNssdBhxbc7XVa9wNknF3ymdvjdRYb/rvIHoN9eXjZfQky+NIykNgvbJsuosYEacB0EZQQtnQLIob7UxF5riqxqVVuaG7a1A1Q+hgVTUiic8ub6yps7CXKFZQ0dsUfUsysSxpbZ0qYdbj/YOHSQtNoH8NZsZqVbuqWm7rF/d/XLev28k21clOt3FQrtu+ra6lW0gIFape0U6P6NuHats2QUHogZhRvc9W54fDNMujAoGopqb7iso03CeCnXA9MYOBGMVIyTszE50QEBwGaQNcnr5qQIz5XPeLOhHHoNqph1Q7FZ3Sr3tE03GGDpFuZz8vOZGJCjkQ6nvOX49BpEgm6VE47cEv1iu1IdUoXBKTsZUhok5kkihYS5cWgNNLCfhYSamVXwqJqYVGR6heuWmEB1JZegY9qBz7Fa67vJV1gaLdBAT6QfkpcvfCudM6VenqdMakeATloic8jIPV0VXJduzy5uiTULuBpg4QWbiYJLQwD+NSdR6feNr9KX1dTlxr0pCkWuyGlUa5ch69lEjmTG2ikZwoaOcc1t1T04WakjyY1dwhdX/gZTiB2uPyuQnQEVyd9EScb/n0yyyTmooV4kBhcJZ0kG4RE4NihJKy5cvnLaKCRyiGKW74ACeGjJVeFtPKxkQOnm07GwyHuC93t2ohq4ysAZPgkV1jfKvH3B0tJNgV3HwSDY6dHp/EDBCHml/PSgAPC4Wogn1hzQOA2a5nI0vg7czDN065+naRiKBlHdBKg+YmiJ/MErlL5ko56ShatTjkwoGEC83l+EPZG8oD94FP3/KNaWk5LmumZaWQVeWrak+n1HfIaq/QQNVglqVt9U/M011UXuQ4C1XpKnHPqXuBA0KilkxnUJOPVNCxz9nzUpHaFBYFmidIauy3PCKsl3vfkB7mzUSsPiEVdqbaBuvbWb6ZZ7xCSRwvuAKdU8OTW7wQa/1D0JbeISdqALXMi5lsDfjnTmNTcJzm/7jULfjOTq/jtjFf0cpmKXy9m6r5fzLf9fK7VKDyFg0UEYd5Prtw7cElBZ/OLdzW+cvkeLu5hbvVZmGXqcj2riKvL93xh/eW7QyDpPCkVOtVitVHKVIv1TsZrNSqZarPUyLRKzXKr02r6lWrnqescKbBXLza9UruSKeWbzYxXykn6lWqm7BUKda9cr7S9+tN5GQMrT5LJ3BZgXsVr8x8AAAD//wMAUEsDBBQABgAIAAAAIQDOyISu7B0AAKzfAQANAAAAeGwvc3R5bGVzLnhtbOw9a28bOZLfD7j/IGgOh9nFOlLr4cnD9szkYdwAu4sAyQELXA6BLLdjIXp4pPasvYf970t2S91sdRdZJKvYdOb8IbFkiawu1vvFsx8fVsveb+l2t9isz/vJs2G/l67nm+vF+st5/78/Xp487/d22Wx9PVtu1ul5/zHd9X+8+Pd/O9tlj8v0w22aZj2xxHp33r/NsruXg8FufpuuZrtnm7t0Lf5ys9muZpl4uf0y2N1t09n1Tn5ptRyMhsPTwWq2WPeLFV6u5phFVrPt1/u7k/lmdTfLFleL5SJ7zNfq91bzl798WW+2s6ulAPUhmczmvYfkdDvqPWwPm+TvNvZZLebbzW5zkz0T6w42NzeLedoE98XgxWA2r1YSK7utlEwHw1Ht2R+2jitNBtv0t4U8vv7F2fp+dbnKdr355n6dnfdPy7d6xV9+uT7vT8b9XnEobzbXAk2fv/9j77s/fffd8Nlw+PkPr+TLT98f3vhUvPGfv95vslcnxX8//ph/7KfPf+gPDlsq6yenk6MNesoGvVefxctPJ+WOxRvHG8h3f/rcAzaYAht8rq2+f2W39Gl96QIpABg/1D9b4qz3qtiz933xv4LM/R/2mPzuU+/VT5+gpxScp57T//zHyeXz4fB/r8XPpz996q3Ez6de8fuj+JFADvYEcHF2s1lXdDARCMvp/uXX9ebv60v5N8Hsgjrkxy7Odv/o/TZbincSuch8s9xse5ngYkEd+Tvr2SotPvFmtlxcbRfyYzez1WL5WLw9km/kjL//3Goh2DCHqNiho32Gpuf5uE2v7oXAynp/+dD6UAMT/Iv1dfqQCr56nlMJClO1Ra8k6g4HkANcodsRvKNVft4uZkvz092rgLRQwtGqOlKAHzCnlOoBkaCV+Mm/zkegtbM4AtWfVNpZym7dGoRHzOkIIeFzelOu8XzpsGVLeqd1/naiXGvCrx23JwRH1IJ8gBoERzLOFgeeX1dk0vbL1Xn/UvwMxQ9e8JpVVFMCFNJGs6EvSRJv0KbMeaBXtCvZBnVNyIUaRczwoAa3gZMIMuLEQ6dqkGG7Ko0lQ2p+St5uGjneOmtSlz8e4sB7pdrTuUnb2hJepkFNlI7JkEROEhqBKbXLsX6xP18dzbloYWYjGG9sG4HP/YCd8BUXy2UZlRhJv1O8cXEmAjhZul1fihe9/e8fH++E17kWsSZJMIPic4ZPf9nOHpPRVPlC/j2x79Vmey1iW4d4yHQiti7euzhbpjeZ2GK7+HIr/882d+Lfq02WiQDQxdn1YvZls54tJRCHb6jfFEExEf8672e3In518JlLX1CEQOQX5Rb7HVCfz2HJQUF9XIB8gBj1+eLhzM92wMpTgtn2PIoT7xTNzCRkQ82BTvzp8RUCYkfccQqbb0acXW/uRSgflK9YgRZIWBugVcR194xQkh+L5jgiwBi0E88Du6mdEMiPwoJw07Ku3LH3Xiuj3WiANb9hMMGaXyDgai3cGhR2BQyZZREc/VpdrYOGU1u3curTA4ZUyJNobGvyOvbfOtDYqHNHUkxQ8ygOUarxwCPSRZzmEZI2lNBEWMTYxkgomZhfRLmqco3k6VSuulofTBJqlV4v7lc28TbDNypzz0XDM4JzQKBhCzsNdTB/APlj2Ms2fFinHdIHQWKnJf5qTw/IvZ6wsv22LbcQxlIosd/chySmz65j3eG2N2dQ1BxJpsVejh2JbotQSVTKilioHvv1kWj9qDRgGKWP5NenkOl0d8WPrfmnkHhlUi3HqGCOrzoH+JB0y5VyRymths9e+iwENs6+ykEUTczT5fKDrG74201ZOTGWzQIPN0qbh2gUkqXxsqNE/ipKOPa/FlUSxQsBYO1L+96Q4msJ+LXe7O5u+SjbFPLFi1dih+rV67ywo3r983LxZb1K1S+8326ydJ7lfU15MRoMieyIaX+AAJAoiJR9LU8NkUU7UXGiMPyhT7RT2kKd6O+Kxmt8Hw+VdAqJQiWxyB/Z/KaTg3+9X12l28u8o5NMMl+ciZ6yQnjLRtRsMZetaELRyhK6u1KI95ab+VfZZ5XL8sHDDayP8GLUSfYqX4LFzLHme1Gpy05PO1JOjMcWiEVzwFTSNKryRlSkCqY1o1BC7BhgJwb6/y+Jzu6Dtb03vPE0ooif0EpvoPoThXehOBanz8dOnkXv4abdxVDoZASrs8O3VSssbyyAfBzrtYqVFb9EelCVprvdbBf/EMcodd1c+C1p0S2tU2uyrR7QzsXjCHJQHCchyPJXIAgNZatVqrIbXbu7QHxtdw3XS2q0WsvqSXp/387uPqYPwnnMj1T7WAnsvJYkhn6uhJpK6kdk92AaY66d+i0oVkYi8g4HHb1KrUR5yi0soxiMSCZKIKjG4g86vBwiDk0pkZzm8yj2QQ38+nVr+sCuRdTj8AreVdWH1pjWaeQAxya2bzKlNLz2eg08JPHQ9TPC4NBbBufzTIrjhZCTCG6zhww+XWVLKmTozhyyCcYCDjemUB8AXJ0RafgnsmNDIrUeAj1EoKp8GTlWpfo9iJDIQVXFNwiqMOgYua+xeiecgCKvOEBFkVccoKrkBVnb7hYPSnrJmVN7H62QRPEK2obp1w0n1OxI8NQ8CezYDMCcZTdmF8bCHTuLSKTVToztMLvSxx7EZDg+Zy73Jo9i/3a+LjV0dqERMIYgYiZu+lslfetn0xn6xsVYY0ggFQmwdCGsyk+yp2wWukpK5daOr84AhiLCDU8LCjOQxgwTCBxJCDjOCARPw1hxxI8+RrVPauoDkpr0oHXUmlxs+wQkOVVIXlbW9+J160AwKWmCYUOKCAx+8a5dwMghVfQyCCmF045fPNoDi8MPRoRum26wX2BeCWhw5QLog5BcBNesVQiRK6GKfKoOOYShSUNA8+d1ZM0lRa4KijHk88I9YjcuMYamJ++HR2lGGWPBcQgpFKhxhNNQoMYRTlNBDRVMC7EnkXDDnGQ3cT8VslDZ9BB70h8ciB3ixKqCHXBLT/l0rDIQW1pSp2YHuCjEU0Fo9kS7m8dqUGtSka2qRiChRZtFD9jilVoYHF1ZpjcINLhm2kHJB4LWoScnEskMDKRPxmGbhLMvirBSa7kSwrNsmuls3roGUFzKyRdWB/k9joM3MMrN08cLqPn5FFbkMSnMMXoKDpfEGjoFGCCzxZDq5zfb6HPZoK5ukAdLEkhXqIwgYsZ4Ez7EFRwxijk6ok7UOLRijIzF9eIk+XoxRtBJNVyiw0FpRJe4OIGuswNcDJ3Goud3f5hcsIcWJk1/TpHTIOyy5oSy4LzWF4rHGFmsG9ySooYGxiJVHYKLDIH4LomRV2Qph3ctE3gO6PIUF07EF5voVpdM2dbNxHJaqP6qEQgS+qwshC0OJJCmqQqQbHr7ulBG9h1p9vrX9tSaYxW0FUggXYHWhVUljFVlJ4gcEdHiE0foajAXgcHhEbpV0NYiQ2ZmOfQk8ut8MpnqUa4GcgGHdo7u/OSVbLJ9uuBsp3JWx5K8uiEKqTi+zKR/JWEAgW6lBWVQIZJu+5H1FAFPXWdWNnUhiMYVWeMxI73JcQQkVdaqd9MmqloTJGrIht0sdIIQp//4RA0alZY06qaoLTjNn2ZJA1ikc16gUqGq5UMQREtoj6jlAznGYtQFmEY2hw7CvVBAESJjzrgvzrulnlvTkFpGDHNCoCO9ugRCQ+FcsSHmzbYbLdRlorUHY9uVoyAZBJYxlT9Ci1q7k1erjSATjCqoZdnbCYFDpg8s4YEYgyOSTdBVBaGPMcaNCXfVi2IgIMugJYtFhTh4HJgyHVjY+V3BieFfsp5Pxx5LHC4rKG1w6RguIaOAkpueEtClY+UOtM9sQxw9lNHab54c0A4MlQIN1wJeP2pQg5YqNMazNur9JwV0mWF5SlCXFmgcQGOULk3KrtUtRTIVb5GvtQaF2Ig2yocASx0mFgYo62jD2DuwhtOxToYAneoPIlm1XTvwXBxSOakp1pVhldYkUXMCKFn6F9wTXcWolMviF7OLTGDYtLQbDpFh5jYPrzEn0EF71CuqLf1Gf9/NRXVkn2AjCJ2T3SruIEEUut/Jmr5wpoBTNKXLQgTM2YT0GVQFThCaA/UeVa2eOVBjN+Wv60IGS2vGRhugZpQTlAoETRvCSTUcJj3Dmi7tWg1VzjYHjmbAJU9ShIYtKUo0wQmCFG0B0rZtNXqrgLpnrt9nXJ4xKWGCLZZ2Rf/xrt5JJB0qIJHIVqdrSvIn3ZQZ1GUyOBoqsEmNCD1BoIZrL9d036tBOqgKrjtKs279o24YkZVBrSKYoxcXa48w0b5GCIHRCwo1Bzbuc1WHmgQcBBBLiZ3XVVWcRX9myQbiibb1UxFRwUnFvl0qKPHsS+KQUCansrx9P3SQ6fCsq3RNzMhubLBEK13zWr6V7OYJ9p4eA8EAcSiuE6bk6sAyxhaV5FTq/sOITjTQpLH1yj8jBrtZ7NQ13KohCiGbye/1SB83IWLCoxzNb9ksBMqCkClvmlBNpD2pYPSHu95Nn25Cp9AiolV0CKV7+WqVTgyUckWIMHR2wqt1EKFWMfgLnVL0wB+zKm0m1aI5ayRoGM3OkdcgaFrHm9OkYsmSiUBziTUWSgRkM4QRFpUqdYJFRjTXGRBJ4E6Uv0uyhif55yH+GgD51YohOzaIJ48je26ox69htDazN0SqCqPIvlj2eLAYk7Y5YBzMvBOJeWBuBO28LGGn2A2Gy5rTh5jrOX2M40aRRRewYizQcawxBtC6awAcow064SUAm0FDsDFPFXkm6/bwrNh2MdaoPDCqdvEwlW7osDpGhAS8xBanhD0tT+KxiZY97U/I2AkVULdIAqEMCV64SWQzlQ9pSqxS5Vaao4Ri1Mq8goEoBuyLSrc4SSfeOjPVkDrOzUnb3bsfkI1M69ERRTx5cxs2UndsXV+pqxE03jnRXnqCbJSl6UegqMvRzSGLoGvavmQNTGNzVTHF1VtC40MyFKuC1X22xgqUeOjCzYyhIwjsLORxghHHhag08zws30HFu9vtYv314+ZykcnB26ayTcTIAMJKRP8WHnQOlaDhzjUxrVTvgugdi3pZn3t3vKfsLtbX6dqPRDgKQc29qT085GBjHn6QLPvccCVgBYLb8HU6JG5UfjFc0q7gUSvTwLsLkHTCtQU5g/YfW5MTS819iIsvSOPFUaIXFBZROAMgdPHgMob8OPEkDtB/CFG2R8pywa6coHFx8bfa8d9DgnBRQqvnozIQKheFrM3EP+mASTvGY8VhailJbjsELRb8zVy6HDko79DBC5cMPJ7snGBHZ6ebeROM4xm6nu7YhKRi/lDXp5kCKKClo5b9/Hq/ydL32/Rm8VDdhuboxHHMguEprKEKf3rWKLpw4e/x8Nh6mtHUpctaQTOWKCZboBIrCB4mS5HZ3jHgVoPjpJ5CdQC5B667EmYYG5DFUeCc4chx9Z8/vGNwwAiF5TeCkgzNS6MgHa5hrZGkE6t7NF0Wa85bpptA2MqaSkYdfELfIdBE118pjAqD6jzHGIWIcH2zdqNnMKhp3gtGyQUTZ2sBg/jm6mxFUfSIx/O0k8SIRHSSjIZTaQGU5p5M2PS/pY26H0SD0CGJ2L/lAlBoUrXvdjQ2Yu1abQs17NvBqaAWsSspaiF9TX+FpEo/0K6+EkxPRuZdg6C2ekjL7egtBOu5Zg0EV0IBnJrNWgviF32HYI6n4F+xWcKPJXe440CJ0nZDEPrZcE55POiGH6pmTZ+gJ3ibdKMEq7OyGjV0YH1ZUsgSJYKEjgFcq1IeCFejeJqEMUcbabYbTK9RBYXsGjW+FWEZclB80wnFDShEZYhdp+1gFmccMmqMLEN0T1ZvQFKuSFbDauZCpaQEwk24tLLlLGUE7LwzUFDWlnO2o0s+gQIrrCk5xgfGjuc2O3RBJ7daVlRA6ifOaioI2m+1mIrSOgbtN77aGaqWcs8JcNYuEY1G5kiWtjUhtWoLKY33wefuLeX2FK9t7RdoiHHdOOLV7wA6o93k0G0aSzoNkZAGnIRvUW+rq8d0UPcjHrOcE4Ag7fIFTm2EGGgFRBk5AQValFFoGc5prfCQN7KV1ElZE2pVwF4Khom5OxiiYioqsfE9u9FqRNZ1CEIlFaQcNiJpcy8Yika7Bs1UqNrcCjFxiRhhcOXpe7/UNiY+zFWNTNv3ArobfCWjFBoP34rCOWcPTLOWBo0buZH2KpuiVGgJTZBo8y8i5c5se+QsQeuMyPFhAO13kenl6L0ZUZkGHPQWwamy2/7WQ8rNgRbekc9GtW0bZ+Gu3qC1M7yh1fW9oCcUusYVTDrUfq6D7nEgmyyIda8bwIfJAbvGcmiCuRxWKxYjnFSIvGQcohyyAJtzyhOtL134HB2rdeG6GEnK2VvGcDDVpPWWHLpzFQdHgBPJ16he8ZaWgNeb7XW6rTrHWUaxcof6zDcSx3UHKYet31msQWVW62HC1HEPpSQHgiXgeMTWKY9kUyoM6hLN2YFG1MQDrtLBBQ/R7HaGptGUB3OAjHqRZWRfEmc9NYTfaaghBVYaS/aLtaYFR6HqRmgUbC2L6WOtUN/1YXR2bULoUBtm0IA6P8n4WDmQlxoRgaDq2Kn6gCyv1ALzZVS1jR5nCxNeBMDB9kA8ch9jvsje1qcyAxzUXXE20YGWQZSVg6Ak4LAT/WfwQ+B2Aq3NOHjuEId1JukUM0294ZN5XXcbpuaDr8Jcg2MKUogiQgPKg07Q6hxpjWgqMsoOfGJuWCfGjonHavNhQEONT6Q5lgUlp5jZuQ2MM9+op4vl9DAhSqUnBDyLcN2PJJFWSDaO+UKXJPlkusYEousNo+NBVUZDoQU+ExOBVYyI4L2x1e6+CJSI6CqNS3hxD0+RWnuTHcpi73rGerTXISEPvT7+GJIF4RSXq1GBuIqjUczCZlI4a1+cbTTp2jgiuNCOsTSEaha+HFgbKGboxKxgnDAm29lDtE/ClfYSjPUOWdNjjCuaXEdMyDxqozVo2AZZcgZVz7u6I5Y5LmkyU806n0hTlWoxmaNsXYxtsIgxjAZB5DyJvG5GQct7CjRdHaz3E5HMglY8HOjaAI5bCS3nVKkBKAjM0BaviWTBWxgIJsNBayckBcuQTxGizqmtTbvUjSNjTz6IGLbwdbu5VDk30vRvn3cQ2FZlumcQT3LElbJqbAZSflSnbp+7xITDmlHZsCFzo9KF0Boi00Y6V4ZKhfkUhEHIxN8lyTu4C5RTXTGRwuATqK2RLHFg4gUQAjalaOTOkI2p/sNJ1Gwb/oIqz5sxlBgjuCf1RSfVnqCl6mmqunTX0WXYKhnoJKIhpHAAKMc5m9hoRB0EcB9pW3dJSUfsON2bAw4RKmXeoWa83u9Nu5lnbPSYWeo1GOgYIPdNWEDUVbn0BjwNz4SaG4KIRRjuVBrTpTjpLop4Ag4vlQlJyZCQQVgWNYfg/nCnoIsZI4/h4NF/k4jBDgA/VWTWCGph/aY52QVTZukegqiIw924esegtFAqy3Z0Og4T8NDe1WT5WCBy1SZW8T0PjDnSSyjwSkcGw7DIYRUe2HK01SM5TD5+BMNdpb9hJ3CpEWZMG0QKXpmBihO+sqSsW/DQ3djY26Nxqtag/bpr68aBz8OadnXD0Smu8OALDecw29k6e+HRzGZJTuwP5Nx4g4n/BG+URnQDYOCWt4B32uFN9BzxVz3i2GHaSPR12MHlIWabXarhngOdEsFFlGT8pCsWoX6UDqmL+FEC1tkSQ95h743bk+AH/nbc92+ZwAhkQhFpuOjL+mPDPqu91ywYepp6evztaIRpuE61Bmk5qggovjTukNnRj6J2wxhDKSyy1r4iFOMShct2Ol/jVxe1+KJGz0ouNasI5srCOZRE6AObojyfxK2mo2P/D+HygXXe4UBnZhw+34II8Bj1hqzzw/QQqyIYKq8bGdKLbOPm64IVBK+kdM90MXXYGLwwzJCAdBryTWJfs1MAeWQehJiqqoEa4hHU4lpNdY6MjGXw0DCH+luBOJZsH4TxaAUHOMaBQzajB01glFusKDXejRkbz6GL893y99DyptKsSplqPQAI3RNDaW+EupqK5bz8+U7J1zH+40jBbr0d0GaGWjlPcjaUizBRs9UQ0DHYluYZCtLyv8y8txo5crJ9yAFKZpcfAaqcFN9y0waBeZPrXh8Yw5VmYEGFIiiSEAOdPL7RwjgixWQW6ONgoBoMnHRHBL4cMeEmraGmDc8gm/4woE1ZI6Ho3k999Fr7ZOiGV7IMtu1wHcVzgIANlxQilA6ByvoQ7AsSNyd1g9fUdLMp567oWUvMmXIzKYAzAZ6Ozglef+CO1U6UB6flDk7mC5c51gloaV1u1tkv1+I6XuubVI+1qLLYBAq5gVMjmipZvU0G8n9KehGmDrrEncDRADW23aWExsc6aEQoKmWo0bcbrg82/yZOSU44GKBSHYiChijgwgEKGoabtmytPoUfJlAI3jceoGFpOavZbmCpRQMJVkZxM4z7LBiVisAkGtVYM5cb5CGgkg6vtZcXY7fPWmyoR0fud8AUDFTDnw4HFISoqhvdU1WQUD54nGGuiEZKkRiJDoSpQ5oDy/ZAI4514CHoHncDTjxjyDAGTCNPEE50gXd1RShOKUjJUwqVIIgja8mSgAMo9AIcnKyHuw+l6zuAUNfD4W698wwxeN+YjnkUDEcTkurRzXyt4ZuuhIz/Bb+Qz8Ahftygrc2RxFsC8XURKS4qaLAGD3w1p3tZXl8EPkq4Zl+aeFNQI5Pgwh+Qczn8aEy5vcXFBqDV8wQuZQBhZ7QidbfDOisksrsOzMYVTeCX0ST2QTCnr+Ot4DkHVLPlTuhjo2H9UcykbuhgEjbLyyQoQIg4BBtc4aXY0uhEOJ2/jpx2igAykHjVV5ioaUJ01vFpNuA/hU5vhWxQs0B+vd9k6ftterN4EDno1pAHWwef6tNiSKdLWHFt3I2gQLALeB2r2okK+gp16lEgbtIcoA/IkBj2D3mAfYdUjpNt94UEaF9nIudV4O6q9JtAoGwpDY7WLT1rqDXZcvyW2IYjVVihn4fAakCZMf6PSzlt0ugvQvhrTqJzRKDPvWMgcJ0ZqWA8wBCDxlUAWZeKHQt7HRtKrrG6ZddmcSg+hUaLvRdIFaqnAtHuBmWwwowDY21q1CQZ5DAJP3qxGituFTTyR54Lo+Cddd3qYD09RyjAP/IMV0JSFbJ4KAi4TSBURguo78YU7IZzUBjzD4zy3Wi6gBOo0IevrbTGC5lgnqWmdEGhuBGImMaslSeZN+2QcXSu/WCeLpd/u9ldnMlfPmSPy3TXm2/u19l5f5z0lbd769kqPe+/2axWM+Xerav7xTJbrGUHwlje/Vguo36+lwx7o15p/8oG4ofty/uF+NL/Dfc/J+L/qfxnWP0zTIo//lOzcrkqetGRedEyUiez3ShIx4hFeyWPJ3ID1LoT47rlovIX1JpT85oVqDJJjlr1FLNqta6IyODW/QGxbkkEaLw+x6xajW9BQ/sCsW418UxKBRRyfzYuW3GXrJBCLfrauGgJqRTCqDXfGNcskZrfR4Ba9K1m0XcPQuT0XksxdLJY9/662a5my4OAkpfN4uB+p9nivx7v0u1ysf6qMDCa2S41C9eBFcSgSNNhm8wrvtBLkgqSfLIXBouJTpgeFp4qC2OJM9EJ1P3CleSTdIoCVydRD6sKxXI4avS6OolarlsJVfS6OqlarltxKVazJDq5ul+3WhUrrROdVN2vWjJ/PqIRdWY6qbpftSy6yu9aQK2qk6n7VUsjcYSVVIlOpO5XLePLeZsGCladTN2vWglArLZKdFL1fbqVFp4im9Ao0MnVw7IlbY3R8kCVpbmBWdiVF2fXDzeldZnLt2x2tUzrVqfY5Tq9md0vs4/lH8/71e9/Sa8X9yvBn/tPvV/8tsnyJc771e9/Xny5zURrlLBI04fsz7ss/793v10Ik/Pd6x9evH13OTp5Pnz9/GQyTqcnL6av355MJ29ev317+WI4Gr75p5Arq+V69/IhmZz3b7Ps7uVgsJvfpqvZ7tlqMd9udpub7Nl8sxpsbm4W83Swu9ums+vdbZpmq+VAyMQXgxeD1WyxFla0WOTlbik+td0/7B74D9V7533lRQF+fpe6AF+F/cXodPjzNBmeXI6HycnkdPb85PnpeHpyOU1Gb08nr99NL6cK7FM32JPhIEkq4Kcvs8UqFRrwcFaHE1LfFYckXmoeYnA4icFO+hkfJKYu/gUAAP//AwBQSwMEFAAGAAgAAAAhAG2eD8ekEgAAYjwAABQAAAB4bC9zaGFyZWRTdHJpbmdzLnhtbLRb23Lb2LF9d5X/YZeqzjl0WSYtaeJJHFspCAQtJLwVScn2eQPJLREZEmBwkcw85TfyNk/5kPmUfMlZ3RsAKfSGzExy5mFKBjb2pXdfVq9ufvjDt81aPegkDePo48lZ++2J0tEiXobR/ceTm1nvzW9PVJoF0TJYx5H+eLLT6ckfLl+++JCmmcK3UfrxZJVl2/edTrpY6U2QtuOtjvDmLk42QYZ/JveddJvoYJmutM42687527fvOpsgjE7UIs6j7OPJu/MfTlQehX/JtWueXFycn1x+SMPLD9lly9nQQxVGapK2X33oZJcfOvTKvB4HSRYu8nWQpPVXfv1BKxCft+by0UI+WopHbp4kGttah8E8XIdZqMX6o2ylE7X4/sBZnAVrsXuxfWc69WbT+rhhHL0p1wjSVGdiH+MEl5Jku1M1XgfYMu5TeX/Jw+2GDkD/8iNc8n04X2tln6KlGwVg/8AN0hVPPQ+in9Q8wMILISFSoffpNlhAtaAjqU4e9MnldBUnmZrpZKP6cRClPI2zfOAZVP3403yzCZKdiu9UGt5H4V244DMuWLmgyGobr8OF5XpmK5wWwzbbINrRwCjONFZL8DiCumX6PgnWagsFo9lxmeoujLCLEE9hFZkm+QlpT/SDjnKMTeKNIsEHGayrQTl8WNtGy6sXd88qgnui4arlv/Z9cSH+bX0e7xuMMdXpe/Fis13HO63VXEf6LsxSpc3Q+sCuxr3gwHQEvodgg8sJ/2oeNHxk9L54KQ5uTtL0VpzhVooi+FZuVxystMos+CaFAQVMO8qo1zaJH0Jye3y1mDHRaxwKWpDFSgfJOoTp7vCHzZogsU6rH6fpKwU/x5pBQ1Xri//6i38rbuZXaNpRqjZN2moYt+sn/Rxmq2USPAbrVC3zhA8F5aUtCutZkbpDu+EjcCrVgcml6ZNTCYeb6IcwzlP1lY88/iqPO5o5/fpnlUO4IodwZXcIlzTql5/hkn75ecF/aviph2BtMzRWs19+JgdD/7d6mEujbUep8csXDfo8xdyJxYH08iQKsxwCLP/772Cz/f3LF73wGz0WquPC1+QZwm1dOM5yiQhi8RJdvcwX/KbjLP+cp5nV5TiZOlPONgnX6vzt+VtxY/sQqTqOPUQ4k5nv3vSdiYgvJmY4iBKDYLEKI53s6gvc6lW4WMvzGuEfHQad+rxX9QdmQu9Y72VWLr2gfTJos8Vd18baAxVsL9M2A3RhVbhmvVSOCUOISPXVR4+QZPrPv/1d9fJoaX+r3GAbAhqU00ibMq8RvpMYioIFv2fvGlBBi+WMXFtO3ZTt575BTOFp6sHYPryY/OqViN3Vsuq1al0JP3IznHruzcTrqv7IGdYPfwG4etGGvp8Lp3Y9cQbOcDrxlT+ceZOhM/NHQ+mSOBK8wUVt1DxOkvgRnlKYZit8pRAT6BZUnGeMhsmjQtIAC/CcahMukvhUqXQTrDGI8NRGL8N8Q14BjksnsExEYXrz+/peW+F+/sPpy9kXuK0wixMsxWgyWwGbDMyKU16Q1huY9bz9YkImN8Pu5KtyIU1/NpJWPkuCpVbjYBcABEokHcFIsAsbiuIPJ3qh4adt3wJ0q2mhLy5cXLhky7iPY6GGAONqr1zfG7x4pbpxPs/ucoGde4zTNBKMVBieUbrnRnS9meP31ainev4XaJ8Vdze4BI4En9vd9q3QdqcL0UMRhYedODPvKIdjnfeZjThTNRJ282niYb2Jml1Lm+p706n1xdTp17d42fXGE8/12bjq6vYWxvkDGedZ/c3Zb9+qrvNVCOG/jpKA40+UOxoaQeIQR33kOscOHA3GNySb+qZ7QlR2qfduJkN/BpelnGGXtIf/rs/22fM/Xc/U1LUIdT8DY4lyjvoU474znCkzZOC41/7Qm3ytD7r1rn1X3tuTfLr+jb+3dBVkDG3n+j6MIvJ5RRZkA5KW7zQc0zNfdENkFpl6Pjor1Qpez18v3sgE3AR3kxE1JBnGZYacNFmzYjg0eLW6DIYBciyDieG5bSNmRbYsJj0Ems8hNuNuIYAGr2lXr0bgA0amRouQ1zWezhZVipBlItZBwDqIVmJC+ObmCa1h6tesgm0rVWioQlJGRBGHWxBE2yBc4i0SNIqyO5Xm2y3naaSprKoItxhBQVkDq2LUngUgpa2ni3Ypj5ExLcItBXxpHQTrZEyhwKUoy1PIj4ms4j0QhUCjFTgE2vl8Z3Y5z3dILsFnEfRgDAE8xRn22bvSYgbTAUWeRXaqiHu7V49I6viMxQqFZW2DHbM4GwrfkAyxFJVc5noXQyD0LNhuY9oPEGKwK1EiC6kmI5FNthBnOYE1J8O61bkAUhjuYBMU+qu0EaRHGLPxL5FV7+iom+An5mKK7bYekTHgglbBA4Q21zqqhJRDXsXGm3dOR6IbfaXmGE+ygZqrAGkUQTN6u9+lSregMO5CnB1KBPRhBlQSlidePn+bkFiS02xkOQcqWmpogzZ+V9CadJ+1aK9HrEjqLk8Y/R2c6mBdugbeC/TMXISipIYGm6Om+WKhAREhG6Y0TiGHDMkinq+gD0hiHlcYT2I7vFsg1nmM+2FGbJHlgLc7o8qlohnEewByAYNZC4o7UNs82cYpJcjQm2WYYor4kYg8YyEBYLTJcEl9DKsDuIsU5eULc1WlZZxfCMsQ94a0m1IuTnBpz8zRLck5PCZhluGEy/gxUqATSG9heHiAw4t5CpZtBlaIDfDsd+/O6oMaUh1IsEj5DJEI9gz8IXaFq6D7T5Z8+L0xH9CUxu0Afat7sHKgHiFtfKO3ZLX43o+WYSD5zwPSs1ia7nlsJz3t2yY9JoFhkTUS/L2vASmmwbuC5Q4e03ZFcxz+UXeRn8zej0Nz+9OX+yUStjTYBYsg0XcaVDvcWGxyoEckU7zdBWUTKe7VkNjYPT2JiaItzlDfnEWwB1bZRBSX/HlBnxvUVRHokuWpr1oywkiSYjDU5OmbhoQpHAqP+itOTJwr+RqEonmAowpN7dbnOS6PyFCPIaWCDZTsNmVKZGGLgnEwrH5bGZRVpcecURnVPpikJHPrAaC+O9yQ8fvwNNBt1jguXSzFyZD8ERt7WGFB9Hz63/thWM/8ms2yACPP5MyqC6+5jk1VhA2fSlUIUSlODUQI1xjTCUvl0nd3BGDrhyRj4rDGUW2ZkxXnODDOs2CpoygETClObCBh4cEbAO3zYz53b+lK4e8O2ECRgM1iWWsq2N+5xvEAI+D4Wr7vv7mVRYZbX1bVwGA8W2KqSwjaty5LCISC4DThGJGyo4yDqk0C84YmzvMUvghkNM6z1gHUE1wwLKGIPQUMIufOUKJSKeP1hXjHebLAVUpuoxumi3WcEo17WAMoQRQZ5D15GXJBJiZVyIGs0pipCZ2m6MmlkocgXDMmqlw9KUC2U8QBA7y+V1J5L5/Q1DD+w4DGesgqxaULhE1yD3TlDaEFdSeEEiqrLZJwXm6+fhc9qlLxuTGCnIGN556MBmri9UEbdNWYmGJPpO89GjOaXSN5bhhxVVJsEEEqymZ2w91/Q/4phargvslvLDl5JNPKMpwuz1jWUI1ggYJByj4WSsUBIUsMhU5KRRoU3nG1z3g4nu2oYOUXmrpFQlKUjaC4uIwiRySgDu3Es6fV2FO1rqqYQVnFZAcRFFAEsx21g66eo0RDx1Im3wJQLsoewF0E9h4hHzgxDCjCGp6TuID05n8mZ0XWgyB5F1J1no5hypTlzRsYKTxaapAizhYmqnB79G3hNgoewpSOCgyCQBJRKJ+TVd2tsbYWTt5+57fI4cB3dqaoBofwtVjkgThDUweGf9xjVOMKUGVFFQ+n5nTEZCkHmc8e/RdpSf14Qy77Yg5bdZcc/qaxHlyUsVBTRG9DIYZCKBQGuJZW3IRtVfVGnb9XXAn4n/R7ZP8ThgV3yjCW0a2J3YmKsIfCtxKchl5SqCLvyDoBbAHwrZNT4GxClQqOIdiSe60+TNmMeOY2ZyI8NYkADRYqJ2dHBU8aYNSHs8GKGyZde/+mftLzNhxZFKL/Y6mI9TXwodg+xI5oCGdZbHejkcpJOHDWBvzFSLpMxikobyKnhFemYjzJGl0EnGGWD08PgnBdVPCsfm/UgKmMOtCZTcK6r/4bzpgWnFbVf5PulOko8Q96eaouzhCuBgEizilV5C7qInFMlMagqmonqhhmzH4m60TWkC+mAlkokKLrDLyJqLh5fc+dTXwwkxDRbOYPPwlP/8kD1+ighnCUz6rIyeNc3GB05fc9gakmI0SXa+at66L8zGS2Oxr1vfqOLvtQjOM8a4igvFZ/BNkk1m6N83T1E116HIWyM6jrD8aQ1h8dfyhejsbeECJU09nI/VN94+ObiXvtTGUo7fqg1WfK+4Kv5Wu70/w0GYG5H09GuLXjJA0ojdraxAON/NVzjrtM8P23/uhmyl/IS6oyxHH9rFQ+qCuS/SBufzStRCbvAkbhopXJsDng08jwCHsamlZcQPfGU7ORIsDWXPoafYZCT7kSiyLZlKn7vu9AEVGuOXLbRVcJNy1VoKUO4v+FdqcKKNQlOZs4XQ9gzPX8W+cKptIaHRQlyTkHqohzYBIYMUXqndognVqlL19wUxIJjVmfCjNT5KRvmZDeZrKj4/hVKbYYqunfXvX/4WR1cTYu8euPUV+C0C5BzypaEUmRo0IsjIT63t6wCjHmoe4kkQdelVmRH1XZRn2iqoR6CqhdFVw/WQqupr8OMGpC3U+Iz9xNKXOkahzTAWJTRXMegcuyncJByot6sS2dqPoBhG5zlb3lSvWjF5YGgbJ/4LXsHShKFcyaPmlBtDUamMH2lkXVcuT0DTVXIyai/AwNSwQjKhiyvePogWg94mSYLk/qQhzl8iG1V6EhxFqn4N4r6k65F3fcedJVe9Aa0aiLdiEMAUKrT1QPToV41PpV/2iYnH+pM9XeucQdf1nRjObKZrTDIKdalgFPg1rL0s9WIrGi3+SH+mG4EQ7Ij/Bi1ZUiRjW/sYI4gRmto8QqVeFdYs4ntFD9tf0yOVyMFlk8R055Tlj27bs2GHYKMaY4QwUqiifIpsGtpGgpZQoTSXhB2SJ1A/FxZ0I1N64wXDc8XF2SIYhehvghyg5FhcsUyYL7BGFF5ds4MpkJpe+8HvFzETJNrrFRWbAqYcAdcYUDmwUrLNnTp9mbYc4uzg5w+2/suL26SfG+ygsOWioNXSzaKk2qAH15bj2rZolVh2UCa0lQnjaoUnn0mBNS+lJlJkLHbJlJw6b+o2mUWKNS9mZjE98QgCKA6bju6AbgtzeiVhjPQFoPnSV4d3E2namBA3yuTpG+Wc62b7l/pv+/j7rtd9rLGgMJXMmvApNFo5VsCODZmImqKzUbOcVIEUtMCzvH0AZqsIIa9UlbZcsXQPr1bkudGGC/C+IMRCwiENE2T6my0uSLnnx05aHmTQxmWeLp5VQbFPiA9/kdEnPQkQ5gSg0E3MVP1HOXeQhufv7Ppf5Cd/Zu/OKMGQLK66WJPXHW4vXhJMZ/0Cz1tRraGwoiXH2CkteurSFDAQhJbcMvTXWobLq3B+in7b/2X4uUAJIv0kWbM+otli7p7q0zdDmjm96Mx30faZtA28WvbYgta+iWHkCJkM5T263QeG7HpF/UlGd6Zh4MU0cMK7EYqfu+uZEQYH3vPvKChIqXAhijDLtd4RdYjd1SbtG5JuVRtJzXp7R2qsmvV0Eo2qMGMX7wJPboomkqCQoOtHsruJirPFwTQqpv5OrG75PTrT8vms2PUtB9I37RNGf67wXq9IiHBiuKH0bUl/uMXxsQfrMve9n3BIFGHVacMoAzQlPucUTTcQb3pixWHW+jqtcetV1R9QIgubsrLqWPbo48UY96fUc1gaZf4Dg5OjBUT1tuOCH9ExYXbzYh0J+lqt0N8RuUkD4SGyuvgkpilkmLvdpTlcu+vicDlluEVTEKv4kQ6fivhhnGGh3za+vOJkjfUOYoeivBCIKd5paC+sGnixjAi2Ui5b6Kt9bZXfkbnEGMHkTl7vALimLVYub6gt3wnn8NMIhRj6R65UBjfbEt1E1iSvz5hLfxOgsQa6EIcxRdsV2ZWBCNjg/itZxsoufUHmSEQZVS6m4TviTR4f0qQ9b5GCSC1uiBm9u/VS2EHRHAy9+/1Gc+Qxvxb6iNWETAEapxT3+yqFoT6hThX/GpH8USP7b3Rept9TNI+nkRF/yFQ9hzC68lIQEwob58/d9//u0f7TYlPGUVEEFrFW6ZddEJ/YawKE7zQB65JytNGkOVOJpBR/e4Ju7zoXSqKn4jCaKiP1zTcR1IatCZ0tbQJS2UEo+F9yW2wHWV01kc5WnRjvsKP96KfhJeoAhA++cd/Eb48v8AAAD//wMAUEsDBBQABgAIAAAAIQBjsaN+wwAAAEMBAAAkAAAAeGwvd29ya3NoZWV0cy9fcmVscy9zaGVldDExLnhtbC5yZWxzhI/NasMwEITvgb6D2Hslu4cQgmVfSiHX/DyAKq9tEXsltJuSvH10TEKgx+FjvmGa7rrM6g8zh0gWal2BQvKxDzRaOB1/PjegWBz1bo6EFm7I0LUfq2aPs5NS4ikkVsVCbGESSVtj2E+4ONYxIRUyxLw4KTGPJjl/diOar6pam/zogPbJqXa9hbzra1DHWyrL/7vjMASP39FfFiR5M2FSDiSYDyhSDnJRuzyiWND6lb3muta/gcC0jXm63t4BAAD//wMAUEsDBBQABgAIAAAAIQBLGL0mwwAAAEMBAAAkAAAAeGwvd29ya3NoZWV0cy9fcmVscy9zaGVldDEyLnhtbC5yZWxzhI/BisIwFEX3gv8Q3t6kdSGDNO1GBLcz+gGZ9LUNti8h7zmMf2+WKgOzvBzuudym+11m9YOZQyQLta5AIfnYBxotXM7HzQcoFke9myOhhTsydO161Xzi7KSUeAqJVbEQW5hE0t4Y9hMujnVMSIUMMS9OSsyjSc5f3YhmW1U7k58d0L441am3kE99Dep8T2X5f3cchuDxEP1tQZI/JkzKgQTzF4qUg1zULo8oFrR+Z++53urvQGDaxrxcbx8AAAD//wMAUEsDBBQABgAIAAAAIQCiNBoYwgAAAEIBAAAjAAAAeGwvd29ya3NoZWV0cy9fcmVscy9zaGVldDgueG1sLnJlbHOEj8GKwjAURfeC/xDefpLqYhBp6kYGunX0A2L62gbbl5D3Zpj+/WSpIri8HO653PrwN0/qFzOHSBY2ugKF5GMXaLBwOX997ECxOOrcFAktLMhwaNar+oSTk1LiMSRWxUJsYRRJe2PYjzg71jEhFdLHPDspMQ8mOX9zA5ptVX2afO+A5sGp2s5CbrsNqPOSyvJ7d+z74PEY/c+MJC8mTMqBBPM3ipSDXNQuDygWtH5mz3mnr4HANLV5eN78AwAA//8DAFBLAwQUAAYACAAAACEAE8QsE8IAAABCAQAAIwAAAHhsL3dvcmtzaGVldHMvX3JlbHMvc2hlZXQyLnhtbC5yZWxzhI/BasMwEETvhfyD2HskO4dQiiVfSiHXJv0ARV7bovZKaLcl+fvo2IRCjsNj3jBdf1kX9YuFYyILrW5AIYU0RJosfJ0+tq+gWDwNfkmEFq7I0LvNS/eJi5da4jlmVtVCbGEWyW/GcJhx9axTRqpkTGX1UmOZTPbh209odk2zN+WvA9ydUx0GC+UwtKBO11yXn7vTOMaA7yn8rEjyz4TJJZJgOaJIPchV7cuEYkHrR/aYd/ocCYzrzN1zdwMAAP//AwBQSwMEFAAGAAgAAAAhADttMkvBAAAAQgEAACMAAAB4bC93b3Jrc2hlZXRzL19yZWxzL3NoZWV0MS54bWwucmVsc4SPwYrCMBRF9wP+Q3h7k9aFDENTNyK4VecDYvraBtuXkPcU/XuzHGXA5eVwz+U2m/s8qRtmDpEs1LoCheRjF2iw8HvaLb9BsTjq3BQJLTyQYdMuvpoDTk5KiceQWBULsYVRJP0Yw37E2bGOCamQPubZSYl5MMn5ixvQrKpqbfJfB7QvTrXvLOR9V4M6PVJZ/uyOfR88bqO/zkjyz4RJOZBgPqJIOchF7fKAYkHrd/aea30OBKZtzMvz9gkAAP//AwBQSwMEFAAGAAgAAAAhADShCZLCAAAAQgEAACMAAAB4bC93b3Jrc2hlZXRzL19yZWxzL3NoZWV0My54bWwucmVsc4SPwWrDMBBE74H8g9h7JCeFUoLlXEIg1zb9AFVeyyL2Smi3pfn76libQo/DY94w7el7ntQXFo6JLOx1AwrJpz5SsPB+u+xeQLE46t2UCC08kOHUbTftK05OaonHmFlVC7GFUSQfjWE/4uxYp4xUyZDK7KTGEkx2/u4CmkPTPJvy2wHdwqmuvYVy7fegbo9cl/93p2GIHs/Jf85I8seEySWSYHlDkXqQq9qVgGJB6zVb5yf9EQlM15rF8+4HAAD//wMAUEsDBBQABgAIAAAAIQBDlhGjwgAAAEIBAAAjAAAAeGwvd29ya3NoZWV0cy9fcmVscy9zaGVldDQueG1sLnJlbHOEj8FqwzAQRO+B/IPYeyQnlFKC5VxCINc2/QBVXssi9kpot6X5++pYm0KPw2PeMO3pe57UFxaOiSzsdQMKyac+UrDwfrvsXkCxOOrdlAgtPJDh1G037StOTmqJx5hZVQuxhVEkH41hP+LsWKeMVMmQyuykxhJMdv7uAppD0zyb8tsB3cKprr2Fcu33oG6PXJf/d6dhiB7PyX/OSPLHhMklkmB5Q5F6kKvalYBiQes1W+cn/REJTNeaxfPuBwAA//8DAFBLAwQUAAYACAAAACEAZPM0IsIAAABCAQAAIwAAAHhsL3dvcmtzaGVldHMvX3JlbHMvc2hlZXQ1LnhtbC5yZWxzhI/BasMwEETvgfyD2HskJ9BSguVcQiDXNv0AVV7LIvZKaLel+fvqWJtCj8Nj3jDt6Xue1BcWjoks7HUDCsmnPlKw8H677F5AsTjq3ZQILTyQ4dRtN+0rTk5qiceYWVULsYVRJB+NYT/i7FinjFTJkMrspMYSTHb+7gKaQ9M8m/LbAd3Cqa69hXLt96Buj1yX/3enYYgez8l/zkjyx4TJJZJgeUORepCr2pWAYkHrNVvnJ/0RCUzXmsXz7gcAAP//AwBQSwMEFAAGAAgAAAAhAExaKnrCAAAAQgEAACMAAAB4bC93b3Jrc2hlZXRzL19yZWxzL3NoZWV0Ni54bWwucmVsc4SPwWrDMBBE74X8g9h7JKcHU4rlXEog1yb9AFVe26L2Smg3If776FibQo/DY94wzfExT+qOmUMkCwddgULysQs0WPi6nvZvoFgcdW6KhBYWZDi2u5fmEycnpcRjSKyKhdjCKJLejWE/4uxYx4RUSB/z7KTEPJjk/I8b0LxWVW3ybwe0K6c6dxbyuTuAui6pLP/vjn0fPH5Ef5uR5I8Jk3IgwXxBkXKQi9rlAcWC1lu2zbX+DgSmbczqefsEAAD//wMAUEsDBBQABgAIAAAAIQBrPw/7wwAAAEIBAAAjAAAAeGwvd29ya3NoZWV0cy9fcmVscy9zaGVldDcueG1sLnJlbHOEj8FqwzAQRO+B/IPYeyQnh7YEy7mEQK5t+gGqvJZF7JXQbkvz99WxNoUeh8e8YdrT9zypLywcE1nY6wYUkk99pGDh/XbZvYBicdS7KRFaeCDDqdtu2lecnNQSjzGzqhZiC6NIPhrDfsTZsU4ZqZIhldlJjSWY7PzdBTSHpnky5bcDuoVTXXsL5drvQd0euS7/707DED2ek/+ckeSPCZNLJMHyhiL1IFe1KwHFgtZrts7P+iMSmK41i+fdDwAAAP//AwBQSwMEFAAGAAgAAAAhAIVRP5nDAAAAQgEAACMAAAB4bC93b3Jrc2hlZXRzL19yZWxzL3NoZWV0OS54bWwucmVsc4SPwWrDMBBE74H8g9h7JCeH0gbLuYRArm36Aaq8lkXsldBuS/P31bE2hR6Hx7xh2tP3PKkvLBwTWdjrBhSST32kYOH9dtk9g2Jx1LspEVp4IMOp227aV5yc1BKPMbOqFmILo0g+GsN+xNmxThmpkiGV2UmNJZjs/N0FNIemeTLltwO6hVNdewvl2u9B3R65Lv/vTsMQPZ6T/5yR5I8Jk0skwfKGIvUgV7UrAcWC1mu2zi/6IxKYrjWL590PAAAA//8DAFBLAwQUAAYACAAAACEAPOGWndEBAADwAwAAIgAAAHhsL2V4dGVybmFsTGlua3MvZXh0ZXJuYWxMaW5rMS54bWycU11vozAQfD+p/8HyezHQprqiQOUCbZEiEhEq3dvJBSex6g9kOyn99zWk12uV5FTdC2Zn17OzY3t60wsOdlQbpmQMA8+HgMpGtUyuY/hY353/hMBYIlvClaQxfKUG3iRnP6a0t1RLwmdMPgNHIk0MN9Z2EUKm2VBBjKc6Kl1mpbQg1oV6jUynKWnNhlIrOAp9/woJwiTcM0Si+Q6JIPp52503SnTEsifGmX0duSAQTVSspdLkiTuxfXAJ+p6HwR9+Bxw0EKzRyqiV9RwhUqsVa+ihzmt0/UXpSPt/XGGAnHnc+WZC32lLPry8Verdy0h/x4i92Ew1W0Gl3dupKXemKGk2rDMQ6Ii1MdRFO/QZfS+JoObTP9gRHsNbPMNlmoPlQ57XEB3k6wpnRXkPcJrOH8tjFYtqflfUAJcZmM2Xy39XLqoir/Pqd4oXRY1nRxrissx/eUcSWb6o8rTAdTEvQfqAq1EMOpgtI5YsqX2fZIjAWFM4P/wP3q94cAIPT+AXJ/BLZ7ZWL8CdYjCZuKChnA/R/RACG0NjtYN3CRZqKy1gElTGm6JdMkVDrVvcfvcdJQ8ij84xGfr/rRnHHS7X+DKH2+Synx9q8gYAAP//AwBQSwMEFAAGAAgAAAAhAEA/ZG1AAwAA7A4AACcAAAB4bC9wcmludGVyU2V0dGluZ3MvcHJpbnRlclNldHRpbmdzMS5iaW7sVstu00AUPWkKgvLqJyBWbPJqUkqjLkjitnLJizgt3YEbTxurqW3ZDkmK2PUr4AdYsEACiW75B5b8C5xxHJqkFRRIVaQwo/HMXI/PnLm+92jycGHDRxOCo7tQUEAVMRSxgEVkkKTlKe1e8N7k0+OsyrUmLACR2ejVr5i/Hf2EmQgieH0jc81gfwfbnM/wGeUsR6TJlUgIJfsZyYHtG8u6qo1so6jlzXvoRr5EV+Yern948zMGt8KXEm+8DPab4BH+Q/2DHvid/9zlYq1U35DHmEcz8hJ5LCGFB6wprDHeY0HUy3yKcZZmHq1xlOdcrorRskxbmnWJfYrZlqNlEa+IqFpO28+bFtYqtZJW2awVVlFb1ZRiEZuW6QpPjqq6I1zNPBTIZVAShqnXe46AVs+VlVxNQaXtD1D0hlDsjoUNe6eq74mKawgXmq9bhu4aqLimsHzdN20L1UqtXsupdaJzXYDwpK23TL+Hsu0e6C0U7FZL9wUq5WCN1uBba49vLQGl7bREF1urtbpayBVRE57dagfASlVNJ5NdNolguyXbEFh39Z5HAHHuaFiZA7YySmnwr7aqVfXto8ef7xPh3RXgmE2WI/bX2d9kO2azqFvJsA5v1qT2+XCQRYLVQyNQwgPoHMdxQJVrBArpUSV3uTLOuU17Ap1AAQ3OOlyboF4m+RcfcpTk/04Qs6+RfrBub8QyvM8+NbVHDJvrjQBJ5bcOSX5kc2jZPcM7k+WdZBRK3pL9eXnvkq9OT4iQ+37A8oTzwinWk/a15Jyi3/+Gs2T5PvRz/wSjtC/fz+Px0Y9nGRf7PH1qzMsX4eMkVenP4/nFD5YyDwe8x6Pj8mNj3M+Sd59lk7y7jPTWL/Oww1zoUCM6zKd4kNEy77dR4n1Kxuoyq8wxK8wcj/7QqSh9/Xse7ONRBU6XUf+cvY9UIKlBgx21ITWTiE6Ib14AfixQOY8KqfN0/RNJ7ZVF9kez/XuaHHfDcT8Wstigp3ag8jsHbSLkw7vlGTSnwjTwSy7whg2N8dFia9A301xWeDkf3Pfl+ERLsswwwbwxGX0an4ecFRlVOm3SPl1l+K6TxTMqTpwK3qUypKc6s74DAAD//wMAUEsDBBQABgAIAAAAIQBAP2RtQAMAAOwOAAAnAAAAeGwvcHJpbnRlclNldHRpbmdzL3ByaW50ZXJTZXR0aW5nczIuYmlu7FbLbtNAFD1pCoLy6icgVmzyalJKoy5I4rZyyYs4Ld2BG08bq6lt2Q5Jitj1K+AHWLBAAolu+QeW/AuccRyapBUUSFWkMKPxzFyPz5y5vvdo8nBhw0cTgqO7UFBAFTEUsYBFZJCk5SntXvDe5NPjrMq1JiwAkdno1a+Yvx39hJkIInh9I3PNYH8H25zP8BnlLEekyZVICCX7GcmB7RvLuqqNbKOo5c176Ea+RFfmHq5/ePMzBrfClxJvvAz2m+AR/kP9gx74nf/c5WKtVN+Qx5hHM/ISeSwhhQesKawx3mNB1Mt8inGWZh6tcZTnXK6K0bJMW5p1iX2K2ZajZRGviKhaTtvPmxbWKrWSVtmsFVZRW9WUYhGblukKT46quiNczTwUyGVQEoap13uOgFbPlZVcTUGl7Q9Q9IZQ7I6FDXunqu+JimsIF5qvW4buGqi4prB83TdtC9VKrV7LqXWic12A8KStt0y/h7LtHugtFOxWS/cFKuVgjdbgW2uPby0Bpe20RBdbq7W6WsgVUROe3WoHwEpVTSeTXTaJYLsl2xBYd/WeRwBx7mhYmQO2Mkpp8K+2qlX17aPHn+8T4d0V4JhNliP219nfZDtms6hbybAOb9ak9vlwkEWC1UMjUMID6BzHcUCVawQK6VEld7kyzrlNewKdQAENzjpcm6BeJvkXH3KU5P9OELOvkX6wbm/EMrzPPjW1Rwyb640ASeW3Dkl+ZHNo2T3DO5PlnWQUSt6S/Xl575KvTk+IkPt+wPKE88Ip1pP2teScot//hrNk+T70c/8Eo7Qv38/j8dGPZxkX+zx9aszLF+HjJFXpz+P5xQ+WMg8HvMej4/JjY9zPknefZZO8u4z01i/zsMNc6FAjOsyneJDRMu+3UeJ9SsbqMqvMMSvMHI/+0Kkoff17HuzjUQVOl1H/nL2PVCCpQYMdtSE1k4hOiG9eAH4sUDmPCqnzdP0TSe2VRfZHs/17mhx3w3E/FrLYoKd2oPI7B20i5MO75Rk0p8I08Esu8IYNjfHRYmvQN9NcVng5H9z35fhES7LMMMG8MRl9Gp+HnBUZVTpt0j5dZfiuk8UzKk6cCt6lMqSnOrO+AwAA//8DAFBLAwQUAAYACAAAACEAQD9kbUADAADsDgAAJwAAAHhsL3ByaW50ZXJTZXR0aW5ncy9wcmludGVyU2V0dGluZ3MzLmJpbuxWy27TQBQ9aQqC8uonIFZs8mpSSqMuSOK2csmLOC3dgRtPG6upbdkOSYrY9SvgB1iwQAKJbvkHlvwLnHEcmqQVFEhVpDCj8cxcj8+cub73aPJwYcNHE4Kju1BQQBUxFLGARWSQpOUp7V7w3uTT46zKtSYsAJHZ6NWvmL8d/YSZCCJ4fSNzzWB/B9ucz/AZ5SxHpMmVSAgl+xnJge0by7qqjWyjqOXNe+hGvkRX5h6uf3jzMwa3wpcSb7wM9pvgEf5D/YMe+J3/3OVirVTfkMeYRzPyEnksIYUHrCmsMd5jQdTLfIpxlmYerXGU51yuitGyTFuadYl9itmWo2URr4ioWk7bz5sW1iq1klbZrBVWUVvVlGIRm5bpCk+OqrojXM08FMhlUBKGqdd7joBWz5WVXE1Bpe0PUPSGUOyOhQ17p6rviYprCBear1uG7hqouKawfN03bQvVSq1ey6l1onNdgPCkrbdMv4ey7R7oLRTsVkv3BSrlYI3W4Ftrj28tAaXttEQXW6u1ulrIFVETnt1qB8BKVU0nk102iWC7JdsQWHf1nkcAce5oWJkDtjJKafCvtqpV9e2jx5/vE+HdFeCYTZYj9tfZ32Q7ZrOoW8mwDm/WpPb5cJBFgtVDI1DCA+gcx3FAlWsECulRJXe5Ms65TXsCnUABDc46XJugXib5Fx9ylOT/ThCzr5F+sG5vxDK8zz41tUcMm+uNAEnltw5JfmRzaNk9wzuT5Z1kFErekv15ee+Sr05PiJD7fsDyhPPCKdaT9rXknKLf/4azZPk+9HP/BKO0L9/P4/HRj2cZF/s8fWrMyxfh4yRV6c/j+cUPljIPB7zHo+PyY2Pcz5J3n2WTvLuM9NYv87DDXOhQIzrMp3iQ0TLvt1HifUrG6jKrzDErzByP/tCpKH39ex7s41EFTpdR/5y9j1QgqUGDHbUhNZOITohvXgB+LFA5jwqp83T9E0ntlUX2R7P9e5ocd8NxPxay2KCndqDyOwdtIuTDu+UZNKfCNPBLLvCGDY3x0WJr0DfTXFZ4OR/c9+X4REuyzDDBvDEZfRqfh5wVGVU6bdI+XWX4rpPFMypOnArepTKkpzqzvgMAAP//AwBQSwMEFAAGAAgAAAAhAMuK9ddBAwAA7A4AACcAAAB4bC9wcmludGVyU2V0dGluZ3MvcHJpbnRlclNldHRpbmdzNC5iaW7sVstu00AUPUkKgvLqJyBWLMiraaGNuiCJ28olL+K0sAM3mTZWXduyHZIUsetXwA+wYIEEEt3yDyz5FzjjODRJKyiQqkhhRuOZuR6fOXN979Hk4cKGjxYER7ehoIAq4ihiHotYQIqWJ7R7wXuDT4+zKtcasABEZmKXv2LuZuwTolFE8PrawpUm+1u4F4kgymeMsxyRJlciIZTso5ID2zeWdVUb2UZRy5t30I18ia3MLq1/ePMzBjfClxJvvAz2m+AR/kP9gx74nf/c5WKtVN+Qx5hDK/ISeTxAGvdZ01hjvMeDqJf5FOcswzxa4yjPuVwVp2WZtgzrA/ZpZluOlkW8IqJqOW0/b1hYq9RKWmWzVlhFbVVTikVsWoYrPDmq6o5wNeNAILeAkmgaer3nCGj1XFnJ1RRU2v4ARW8Ixe5Y2LC3q/quqLhN4ULzdaupu01UXENYvu4btoVqpVav5dQ60bkuQHjc1k3D76Fsu/u6iYJtmrovUCkHa7QG31q7fGsJKG3HFF1srdbqaiFXRE14ttkOgJWqmkmlumwSwXZLdlNg3dV7HgHEmaNhZRbYWlBKg3+1Va2qbx8++nyXCO8uAUdsshyyv8r+OtsRm0XdSoV1eLMWtc+HgyySrB4agRLuQ+c4gX2qXCNQSI8qucOVCc5t2pPoBArY5KzDtUnqZYp/cYmjFP93kph9jfSDdbsjluF99qipPWLYXN8MkFR+65DkRzaHlp1TvDNZ3ilGoeQt2Z+V9w756vSECLnvBSyPOc+fYD1pX0vOafr9bzhLlu9DP/dPMEr74v08Hh/9eJZxscfTp8e8fB4+TlGV/jyeX/xgKfNwwHs8Oi4+Nsb9LHn3WbbIu8tIN3+Zhx3mQoca0WE+JYKMlnn/FCXep2SsLrPKHLPCzPHoD52K0te/58E+HlXgZBn1z+n7SAWSGjTYURtSM4nohPjGOeDHA5XzqJA6T9c/kdReWWR/ONO/p8lxNxz3YyGLDXpqGyq/c9AmQj68W55CcypMA7/kAm/Y0BgfJluDvpnmssLL+eC+L8fHWpJlhgnmjcHo0/g84KzIqNJpk/bpKsN3nSyeUXESVPAulSEz1Zn1HQAA//8DAFBLAwQUAAYACAAAACEA/zcAIUEDAADsDgAAJwAAAHhsL3ByaW50ZXJTZXR0aW5ncy9wcmludGVyU2V0dGluZ3M5LmJpbuxWy27TQBQ9SQqC8uonIFZs8mpSSqMuSOK2csmLOC3dgRtPG6upbdkOSYrY9SvgB1iwQAKJbvkHlvwLnHEcmqQVFEhVpDKj8cxcj8+cub73aApwYcNHC4Kju1BQRA1xlDCPBWSRouUp7V7w3uTT46zGtSYsAJGZ2NWvmLsd+4RoFBG8vpG9ZrC/g61oBFE+Y5zliTS9EgmhZB+VHNi+sayp2tg2ilrZuIde5Etsefbh2oc3P2NwK3wp8SbLcL8pHuE/1D/ogd/5zz0u1sqNdXmMObQiL1HAItJ4wJrGKuM9HkS9zKc4Zxnm0SpHBc7lqjgtS7RlWBfZp5lteVoW8IqIquV0/IJpYbVaL2vVjXpxBfUVTSmVsGGZrvDkqKY7wtXMA4F8FmVhmHqj7whojXxFydcVVDv+EEVvCsXuWli3t2v6rqi6hnCh+bpl6K6BqmsKy9d907ZQq9Yb9bzaIDrXBQhPOnrb9Puo2O6+3kbRbrd1X6BaCdZoTb61dvnWElA6Tlv0sLlSb6jFfAl14dntTgCs1NRMKtVjkwi2W7YNgTVX73sEEGeOhuVZYDOrlIf/arNWU98+evz5PhHeXQGO2GQ5ZH+d/U22IzaLupUK6+hmLWqfDwc5JFk9NAMl3IfOcQL7VLlmoJAeVXKHKxOc27Qn0Q0U0OCsy7VJ6mWKf/EhRyn+7yQxBxrpB+t2xyyj++xRU/vEsLneCJBUfuuQ5Ec2h5adU7wzXd4pRqHkLdmflfcO+er0hAi57wUsjznPn2A9bV9Lzmn6/W84S5bvQz8PTjBO++L9PBkfg3iWcbHH06cnvHwePk5Rlf48nl/8YCnzcMh7MjouPjYm/Sx5D1i2yLvHSG//Mg+7zIUuNaLLfEoEGS3zfgtl3qdkrC6xyhyzwszx6A+dijLQv+fBPh5V4GQZ98/p+0gFkho03FEbUTOJ6IT45jngxwOV86iQOk83OJHUXllkfzgzuKfJcS8cD2Ihh3V6ahsqv3PQIUIhvFueQvNSmIZ+yQfesKExPtpsTfrmMpdlXs6H9305PtaSHDNMMG9MRp/G5wFnJUaVTpu0X64yetfJ4RkVJ0EF71EZMpc6s74DAAD//wMAUEsDBBQABgAIAAAAIQAOb/JkQgMAAOwOAAAoAAAAeGwvcHJpbnRlclNldHRpbmdzL3ByaW50ZXJTZXR0aW5nczEwLmJpbuxW3VLTQBg9pegI/vEIjlfetLS0CHS4sG2ACfbPpiB3GpqFZihJJklti+MdT6Ev4IUXzuiM3PoOXvouejZNpS2MopbBmbo7m939sjl79sv3ndkcXNjw0YDg6B4U5FFBDAUsYBFpJGh5SrsXvDf59DircK0JC0BkOnr9K+buRD9hKoIZvL6ZvmEggrvY4XyKzyhnWSKNr0RCKNlPSQ5s31g2VG1oG0Utbd1HJ/Ilujq7vPHhzc8Y3A5fSrzR0t9vjEf4D/UPeuB3/nOHi7VibVMeYw6NyEvksIQkHrImsc54jwVRL/MpxlmKebTOUY5zuSpGywptKdYl9klmW5aWRbwiomo5LT9nWlgvV4taeauaX0N1TVMKBWxZpis8OarojnA180ggm0ZRGKZe6zoCWi1bUrJVBeWW30fR60Kx2xY27d2Kvi/KriFcaL5uGbproOyawvJ137QtVMrVWjWr1ojOdQHCk5beNP0uSrZ7qDeRt5tN3Rcol4I1Wp1vrX2+tQSUltMUHWyvVWtqPltAVXh2sxUAKxU1lUh02CSC7RZtQ2DD1bseAcSFo2F1FthOK8X+v9quVNS3jx5/fkCEd9eAEzZZjtnPsL/FdsJmUbcSYR3crEHt8+Egg3lWD/VACQ+hcxzHIVWuHiikR5Xc48o45zbt82gHCmhw1ubaeeplgn9xmaME//c8MXsa6Qfr9ocsg/scUFO7xLC53giQVH7rkORHNoeWvXO8M17eCUah5C3ZX5T3Hvnq9IQIuR8ELE85L5xhPW5fS85J+v1vOEuW70M/904wTPvq/TwaH714lnFxwNMnR7x8GT5OUJX+PJ5f/GAp87DPezQ6rj42Rv0sefdYNsi7w0hv/jIP28yFNjWizXyKBxkt834HRd6nZKyusMocs8LM8egPnYrS07/nwT4eVeBsGfbP+ftIBZIa1N9RG1AzieiE+OYl4McClfOokDpP1zuR1F5ZZH883bunyXEnHPdiIYNNemoXKr9z0CJCLrxbnkNzIkx9v2QDb9jQGB9Ntjp9M8lllZfz/n1fjk+1JMMME8wbk9Gn8XnEWYFRpdMm7ZNVBu86GTyj4sSp4B0qQ2qiM+s7AAAA//8DAFBLAwQUAAYACAAAACEA/zcAIUEDAADsDgAAKAAAAHhsL3ByaW50ZXJTZXR0aW5ncy9wcmludGVyU2V0dGluZ3MxMS5iaW7sVstu00AUPUkKgvLqJyBWbPJqUkqjLkjitnLJizgt3YEbTxurqW3ZDkmK2PUr4AdYsEACiW75B5b8C5xxHJqkFRRIVaQyo/HMXI/PnLm+92gKcGHDRwuCo7tQUEQNcZQwjwVkkaLlKe1e8N7k0+OsxrUmLACRmdjVr5i7HfuEaBQRvL6RvWawv4OtaARRPmOc5Yk0vRIJoWQflRzYvrGsqdrYNopa2biHXuRLbHn24dqHNz9jcCt8KfEmy3C/KR7hP9Q/6IHf+c89LtbKjXV5jDm0Ii9RwCLSeMCaxirjPR5EvcynOGcZ5tEqRwXO5ao4LUu0ZVgX2aeZbXlaFvCKiKrldPyCaWG1Wi9r1Y16cQX1FU0plbBhma7w5KimO8LVzAOBfBZlYZh6o+8IaI18RcnXFVQ7/hBFbwrF7lpYt7dr+q6ouoZwofm6ZeiugaprCsvXfdO2UKvWG/W82iA61wUITzp62/T7qNjuvt5G0W63dV+gWgnWaE2+tXb51hJQOk5b9LC5Um+oxXwJdeHZ7U4ArNTUTCrVY5MItlu2DYE1V+97BBBnjoblWWAzq5SH/2qzVlPfPnr8+T4R3l0BjthkOWR/nf1NtiM2i7qVCuvoZi1qnw8HOSRZPTQDJdyHznEC+1S5ZqCQHlVyhysTnNu0J9ENFNDgrMu1Seplin/xIUcp/u8kMQca6Qfrdscso/vsUVP7xLC53giQVH7rkORHNoeWnVO8M13eKUah5C3Zn5X3Dvnq9IQIue8FLI85z59gPW1fS85p+v1vOEuW70M/D04wTvvi/TwZH4N4lnGxx9OnJ7x8Hj5OUZX+PJ5f/GAp83DIezI6Lj42Jv0seQ9Ytsi7x0hv/zIPu8yFLjWiy3xKBBkt834LZd6nZKwuscocs8LM8egPnYoy0L/nwT4eVeBkGffP6ftIBZIaNNxRG1EzieiE+OY54McDlfOokDpPNziR1F5ZZH84M7inyXEvHA9iIYd1emobKr9z0CFCIbxbnkLzUpiGfskH3rChMT7abE365jKXZV7Oh/d9OT7WkhwzTDBvTEafxucBZyVGlU6btF+uMnrXyeEZFSdBBe9RGTKXOrO+AwAA//8DAFBLAwQUAAYACAAAACEAy4r110EDAADsDgAAKAAAAHhsL3ByaW50ZXJTZXR0aW5ncy9wcmludGVyU2V0dGluZ3MxMi5iaW7sVstu00AUPUkKgvLqJyBWLMiraaGNuiCJ28olL+K0sAM3mTZWXduyHZIUsetXwA+wYIEEEt3yDyz5FzjjODRJKyiQqkhhRuOZuR6fOXN979Hk4cKGjxYER7ehoIAq4ihiHotYQIqWJ7R7wXuDT4+zKtcasABEZmKXv2LuZuwTolFE8PrawpUm+1u4F4kgymeMsxyRJlciIZTso5ID2zeWdVUb2UZRy5t30I18ia3MLq1/ePMzBjfClxJvvAz2m+AR/kP9gx74nf/c5WKtVN+Qx5hDK/ISeTxAGvdZ01hjvMeDqJf5FOcswzxa4yjPuVwVp2WZtgzrA/ZpZluOlkW8IqJqOW0/b1hYq9RKWmWzVlhFbVVTikVsWoYrPDmq6o5wNeNAILeAkmgaer3nCGj1XFnJ1RRU2v4ARW8Ixe5Y2LC3q/quqLhN4ULzdaupu01UXENYvu4btoVqpVav5dQ60bkuQHjc1k3D76Fsu/u6iYJtmrovUCkHa7QG31q7fGsJKG3HFF1srdbqaiFXRE14ttkOgJWqmkmlumwSwXZLdlNg3dV7HgHEmaNhZRbYWlBKg3+1Va2qbx8++nyXCO8uAUdsshyyv8r+OtsRm0XdSoV1eLMWtc+HgyySrB4agRLuQ+c4gX2qXCNQSI8qucOVCc5t2pPoBArY5KzDtUnqZYp/cYmjFP93kph9jfSDdbsjluF99qipPWLYXN8MkFR+65DkRzaHlp1TvDNZ3ilGoeQt2Z+V9w756vSECLnvBSyPOc+fYD1pX0vOafr9bzhLlu9DP/dPMEr74v08Hh/9eJZxscfTp8e8fB4+TlGV/jyeX/xgKfNwwHs8Oi4+Nsb9LHn3WbbIu8tIN3+Zhx3mQoca0WE+JYKMlnn/FCXep2SsLrPKHLPCzPHoD52K0te/58E+HlXgZBn1z+n7SAWSGjTYURtSM4nohPjGOeDHA5XzqJA6T9c/kdReWWR/ONO/p8lxNxz3YyGLDXpqGyq/c9AmQj68W55CcypMA7/kAm/Y0BgfJluDvpnmssLL+eC+L8fHWpJlhgnmjcHo0/g84KzIqNJpk/bpKsN3nSyeUXESVPAulSEz1Zn1HQAA//8DAFBLAwQUAAYACAAAACEACaFOoZEKAAD5NgAAEAAAAHhsL2NhbGNDaGFpbi54bWx0W9tuHDcMfS/QfzD2vbVnb7aLOMV6RjPSPrcfYDjbJoAvgW0U7d93NjoUKZ7RU5CzlIaieKf86fd/n58u/jm9vX97fblbdb9erS5OL4+vX769/H23+vOP8Zeb1cX7x8PLl4en15fT3eq/0/vq988///Tp8eHpsf/68O3lYt7h5f1u9fXj4/tvl5fvj19Pzw/vv75+P73Mv/z1+vb88DH/9+3vy/fvb6eHL+9fT6eP56fL9dXV/vJ53mD1+dPjxdvdKu3nb327W6271cXTzMvqEj+Em+sfP9wUZGCk2+x/EHVbXdetbz02LGH4sF26X9NKhkKXGTMLB4ZCt2EuFqDC68UsTT19VLGIQEaGAkMDQz1D9wzFXT7WfBHliwzFXRZ4RUVQ3O3kUnUvguJuy1QEJaVyItplcVasEBS3dKy4m/U9q5xyR1Dc5qux2zMUt0V9y14MJYXqM4wL3G1ZKgzFLR+eoaSQk902a3p1OIKSUvnls7U6CW4JiluWM0NJIfeRDV8AQ3HDF8BQ3LByMxQ3LFSGkkKeYRbqhqC4YdkxFDcsO4YiXJ29RobimkXEUFLIHWvNsmMoKeSXs8dYs8c4+3+nUQzFNUuFodix4jAUO5YKQxGe3UqYoaSQO3zHh2coduwuGYpzwPEiYigp5FwNApI5yYDzmhC713hUReO4ZyVgKCnkwxnLYU9QUsgvZwHtCUoK+eUsuT0HGoX8cnYLe4KSQn45uwVkFeYukkJ+OVvGnqCkUL08cYRJ7B8Tx9LE7iaxN0tsfoltLbHOJ3a/ieNyYs+Y2HkkNpbEppw4dUkctBKnLonDfuKInDg2JfbFiZ1a4iwrdPBz8z+VGQ4dMonZdupsucNR/A/ziqwse5Mh591vCzJ1mWY2/mrXCXwYSmiFQXBIi2TPqkjEqWc1KFkSQyKbs7JUbIzExkBIT8iIgKZsBGI1EM1ANAPR9CSEnlb1/HVovBHCwFDP0D1DoxQcKlCJVFbGMDsDRYZGWE9FlZ1rdV/5Ui0VjlgtzG61YmIBQiFkuWcItm/26jeZr1lPq2ppZNLYoZYzKsdQRLgzRIREGEd1pmwvlXxQFZozISWqRJZNz0BT40xxzXJrHZ9Je6RAV2pwEv8VEnvQ+D8Ky2XZ4SZzbI9VIq8z0zWHZOZMYoERgaRPFlq40X1J/urb76HTXiuGIi5xOpIcVd/mq4SenD1UpWaSIleakLMYdTKJtSxJL0CFKCHLckKaJzm1IRroOsICku9Bv3boEN8NA3pZ7hZVzEVsLcmLkdk8piW8iFaNzaQZigtZ6o0/TiAtDUj99HwSITjsWRq5vFoEifmSKzTcH66zxdsvkJgHlnzoIE7jdKCpXuEGJg0N0gOSTJ8xHEi9gdiTgEabaKBRrQagngKA1bEf1clGkwqkpCYx2MEnK1FAempKEISe2fvVGU7BRS3vJZfyZ75nV9czFJExVNZc8q7ShtSMS6BhQdnWOXxZXSBoVsns0iud9NBwkwOtISJEmnyGBg09o9zrLGxDw4gGlErWB8/Uvd6cuxXtovofGp3NAdUHJ6qoG+gH5FZG1fzJBtYidJWMXkEDLOIdixTEnF1TSzkga7J3KRfnw5NPh0NJt4tKNc4ulN4xhEJf8mnJUlzMGvLHTYhHZaRmKaf22X9PDRJsRoExaA+/zssaeGzgqUWPdNdUFBC/8VBZKfSgkrf5Y4ndk3ObS9lyH7fi2139dZsVYVaQ6qDDrS+yhlvxCDXl7P3zXEC/hcKRI5W6SYlmxnHCmNe6T4fy3AY4lNSe31mv5CCqgvCHdjmqezou2kSG9FoE4I8Ly/ECQ01jdkCzmb5VYqrfGR0cO/nJDTKDoFGiCJpOCiBnVTFmwLgJSNrOe67LJMfpQYfm8LkLV2lI2CCdp4o+c2S+hy1mW6h1DJ0JwpFOEA6PoDsH9PCMfFDUGQRljMoDntTelIQHf/grajVAJcl/iZjMxZBCYngxs1j3QLDWMATZspYXmdd3AbWmdAtFx3y5dTjsSiz3P1BEP1CwPpTQ7CJkMRn14cvGMs8S4TR8UnQoZuQY41BPAVEmlEY/2IPMjsLXf6FreAUpJ60jU//jhruY7TqRNNxuaLjdQG43NN1u0U1X7jQ1lNIw8tOlC+NTVTKZgH4haWjpJNVcRTg9E/B20qITbZmQ9Jj8euuL0WkrDROdTmYD11UjrRpplaTpppdGqwKtkrTMpBq0aqBVPZ2rp1U9rYqFxklygVIkUFP2aK1aefrMbSSaQMhAiJS+urN0+cz9wodZxLdpZdxnaMq3XAUrHYfr4mBL78NCcCsWgv+3EIKEheCoDCQ+1UKoKy2ECtBCiKUWgt+2ELE6oiahZiAeDDAu1rHcPPT0pdujXEyNT06NT04YOfitJwQa/iTLVfJ65aJvcNErF5Wz7RtchLJPKXYb5witHRrnGGjnobHz0Nh5aOwc9ez1fLKxv7xX8ZLuy/5y9h7u3VPK9NwaZpmW1NUOpeSjNMOKFaIvS6UWCIlNWBXpMoo482AIRqrBd0RuYDx964goBqqutA8TU2ux5Cb6ZXltcS51XA9Hkl1ngtKd8X39bA7mBA05jQ05TX6DqbFBIGGFxnllZmOCmlLWiap83KfJ2Qf6Ox0aZwj+DHMC9iN78hsceU4UqQUud0Nq3rre5UNI55cUs1HmyAhv5ryS0X3e3pTeZZZTbGYsJXO9dsK3tCgWq54jSp3wS03n1JGK0NAoQkOjCD1gvuErhlBqFP/FUjq5CacUqi6P19qurkhKLSju61CqpYKUOrIgEKTt4ubvGsSXvwfqaU9AzkZbj4vxhsF7t5GKbqS957ZOLQf9QZiWl1Azbcl8ZfBoIcoPJpk9KdVRoVpBJhR45iNHhTwt5TBHaRHM56kUb5Ipt+WhZEWeFrNBS1sgRyse29Aq5GkpfztK+CB+Za5s9y0Zod+XRH4US+V9KaM74qHS+Yu1zDjjOSpU06JFaa4NLUraFS1KwsciHMdFA0eLkvY5NuglMpnY2kNupoFREOe3GjgM6ixt92o49xY9nlr7aMSo7XBgO5T3QJWJsGbhdY61Vq5FjgrVB5DnRNWAlkbmMho/G5d7r7Kc4k+l571schTF4CfsXJOD65GFNC3ITfdyvg5KbT4iLWs7I5X3fCYcSsvSyU73c0FH7sRHRXGjHi/z9OKDGxN2eUPJObOUrjWLY+GktCMaPEhd4Jt/0zW9MQg00AiNTOrIr97l1VY1BJSuoTIJFSRB4RWg71DiRZdG1XsNmaXfLu1rf1mILjT/bPWc5z5eoxstswYvRMlOfMdYpk2G7zLDdVxStnHfiP2BMxXEPzPLaGU519KD81kCDkzZA/wc4dLUrtURM3MzCMI10Zjbp62h3LDjzE8TRmm5u7wOZawdyy1Tyl2ZWRCuW5ERlk/5mOJ1obosjlE8lboabYjXIbqB98QaArTnbCTCqSDOZciEwnsSaJEzSXkGSAm5PE+ntwLyDmb+oR63YIUxhtYe8mbQdtLxkIA8xrLvlr8pmvXMTY3kmKVFI+U2US5faiiX6jQ1y9VXStMyLOUba1g+JxVcDWaWLzNIluzL/5bA+9al9QsrLssfrX3+HwAA//8DAFBLAwQUAAYACAAAACEA/zcAIUEDAADsDgAAJwAAAHhsL3ByaW50ZXJTZXR0aW5ncy9wcmludGVyU2V0dGluZ3M3LmJpbuxWy27TQBQ9SQqC8uonIFZs8mpSSqMuSOK2csmLOC3dgRtPG6upbdkOSYrY9SvgB1iwQAKJbvkHlvwLnHEcmqQVFEhVpDKj8cxcj8+cub73aApwYcNHC4Kju1BQRA1xlDCPBWSRouUp7V7w3uTT46zGtSYsAJGZ2NWvmLsd+4RoFBG8vpG9ZrC/g61oBFE+Y5zliTS9EgmhZB+VHNi+sayp2tg2ilrZuIde5Etsefbh2oc3P2NwK3wp8SbLcL8pHuE/1D/ogd/5zz0u1sqNdXmMObQiL1HAItJ4wJrGKuM9HkS9zKc4Zxnm0SpHBc7lqjgtS7RlWBfZp5lteVoW8IqIquV0/IJpYbVaL2vVjXpxBfUVTSmVsGGZrvDkqKY7wtXMA4F8FmVhmHqj7whojXxFydcVVDv+EEVvCsXuWli3t2v6rqi6hnCh+bpl6K6BqmsKy9d907ZQq9Yb9bzaIDrXBQhPOnrb9Puo2O6+3kbRbrd1X6BaCdZoTb61dvnWElA6Tlv0sLlSb6jFfAl14dntTgCs1NRMKtVjkwi2W7YNgTVX73sEEGeOhuVZYDOrlIf/arNWU98+evz5PhHeXQGO2GQ5ZH+d/U22IzaLupUK6+hmLWqfDwc5JFk9NAMl3IfOcQL7VLlmoJAeVXKHKxOc27Qn0Q0U0OCsy7VJ6mWKf/EhRyn+7yQxBxrpB+t2xyyj++xRU/vEsLneCJBUfuuQ5Ec2h5adU7wzXd4pRqHkLdmflfcO+er0hAi57wUsjznPn2A9bV9Lzmn6/W84S5bvQz8PTjBO++L9PBkfg3iWcbHH06cnvHwePk5Rlf48nl/8YCnzcMh7MjouPjYm/Sx5D1i2yLvHSG//Mg+7zIUuNaLLfEoEGS3zfgtl3qdkrC6xyhyzwszx6A+dijLQv+fBPh5V4GQZ98/p+0gFkho03FEbUTOJ6IT45jngxwOV86iQOk83OJHUXllkfzgzuKfJcS8cD2Ihh3V6ahsqv3PQIUIhvFueQvNSmIZ+yQfesKExPtpsTfrmMpdlXs6H9305PtaSHDNMMG9MRp/G5wFnJUaVTpu0X64yetfJ4RkVJ0EF71EZMpc6s74DAAD//wMAUEsDBBQABgAIAAAAIQAUZF4VawEAAJoCAAARAAgBZG9jUHJvcHMvY29yZS54bWwgogQBKKAAAQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAB8kklvwjAUhO+V+h8i340XttYKQV0El1aq1FSterPsB0QkjmWbAv++ToCULuoxmZlPM09Op7uqTD7A+aI2E8R6FCVgVK0Ls5ygl3yGr1DigzRalrWBCdqDR9Ps8iJVVqjawZOrLbhQgE8iyXih7AStQrCCEK9WUEnfiw4TxUXtKhnip1sSK9VaLoFwSkekgiC1DJI0QGw7IjoiteqQduPKFqAVgRIqMMET1mPkyxvAVf7PQKucOasi7G3cdKx7ztbqIHbunS8643a77W37bY3Yn5G3x4fndiouTHMrBShLtRLKgQy1y+YbZx1ASMnZz+aApfThMd56UYC+3Wc3a7+SIZlvbJAp+a2fIk+uMAF0xikfYsYwG+d0JNhQDPl7lzuZYpF296EN6CQuEYfdJ+W1f3efz1DD45iOMKc5Y2LABR1E3o98s+wArI7N/yeOMB1jfp0zKgZjwYdnxBMga0t/f03ZJwAAAP//AwBQSwMEFAAGAAgAAAAhAAv6xFU4AgAAvAYAABAACAFkb2NQcm9wcy9hcHAueG1sIKIEASigAAEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAnFVRb9owEH6ftP/g5WF9Kgm0qiZmUiEo66RuREA77WkyzgWsJrZlHwj26+cQGsIamrG38/m7u8+fz2d6u8lSsgZjhZI9r90KPAKSq1jIRc97nI0uP3nEIpMxS5WEnrcF692G79/RyCgNBgVY4lJI2/OWiLrr+5YvIWO25bal20mUyRi6pVn4KkkEh6Hiqwwk+p0guPFhgyBjiC91mdArMnbX+L9JY8VzfvZpttWOcEj7WqeCM3SnDL8JbpRVCZK7DYeU+tVN6thNga+MwG0YUL+6pFPOUhi4xGHCUgvUPzjoPbBctIgJY0O6xu4aOCpDrPjtZLv2yJxZyOn0vDUzgkl0tHJYsdjZqbZowh/KPNslAFrqO0Dh3JlVbNUW12G7s0M4601kkew7yyAmEyYXcFaNdn2NnGVxWlf8WIeZwBTsOImYwRpZOldVXXbkClUKnvMjdqUW+iPL9Gd3cQd1SksqdB2JijDO1Uoea1iiONMCWUOGE9XLCm/RiKK7WnojsXHK9611t0sGLfLTtJpx0WncEJCJtF6mL5PxdEqiyXj0dVZfpN8JOkXf7JusFGhuP0RGSPzVN8Bqg/cSNsH2BJtgFzXCXJwXVKjUHFSVpRn9qp/OCJnbM8D7bmqO+Fega8ETAh69179e6EBlmsmte8il9SDks33UMzVkCC+T79hJp0tmIHbDspyMpYPeu6Fn0jzJYJmPnPgF83ojn9NPxWcUtm9awVXgRnDFR/3DtxP+AQAA//8DAFBLAwQUAAYACAAAACEAQD9kbUADAADsDgAAJwAAAHhsL3ByaW50ZXJTZXR0aW5ncy9wcmludGVyU2V0dGluZ3M2LmJpbuxWy27TQBQ9aQqC8uonIFZs8mpSSqMuSOK2csmLOC3dgRtPG6upbdkOSYrY9SvgB1iwQAKJbvkHlvwLnHEcmqQVFEhVpDCj8cxcj8+cub73aPJwYcNHE4Kju1BQQBUxFLGARWSQpOUp7V7w3uTT46zKtSYsAJHZ6NWvmL8d/YSZCCJ4fSNzzWB/B9ucz/AZ5SxHpMmVSAgl+xnJge0by7qqjWyjqOXNe+hGvkRX5h6uf3jzMwa3wpcSb7wM9pvgEf5D/YMe+J3/3OVirVTfkMeYRzPyEnksIYUHrCmsMd5jQdTLfIpxlmYerXGU51yuitGyTFuadYl9itmWo2URr4ioWk7bz5sW1iq1klbZrBVWUVvVlGIRm5bpCk+OqrojXM08FMhlUBKGqdd7joBWz5WVXE1Bpe0PUPSGUOyOhQ17p6rviYprCBear1uG7hqouKawfN03bQvVSq1ey6l1onNdgPCkrbdMv4ey7R7oLRTsVkv3BSrlYI3W4Ftrj28tAaXttEQXW6u1ulrIFVETnt1qB8BKVU0nk102iWC7JdsQWHf1nkcAce5oWJkDtjJKafCvtqpV9e2jx5/vE+HdFeCYTZYj9tfZ32Q7ZrOoW8mwDm/WpPb5cJBFgtVDI1DCA+gcx3FAlWsECulRJXe5Ms65TXsCnUABDc46XJugXib5Fx9ylOT/ThCzr5F+sG5vxDK8zz41tUcMm+uNAEnltw5JfmRzaNk9wzuT5Z1kFErekv15ee+Sr05PiJD7fsDyhPPCKdaT9rXknKLf/4azZPk+9HP/BKO0L9/P4/HRj2cZF/s8fWrMyxfh4yRV6c/j+cUPljIPB7zHo+PyY2Pcz5J3n2WTvLuM9NYv87DDXOhQIzrMp3iQ0TLvt1HifUrG6jKrzDErzByP/tCpKH39ex7s41EFTpdR/5y9j1QgqUGDHbUhNZOITohvXgB+LFA5jwqp83T9E0ntlUX2R7P9e5ocd8NxPxay2KCndqDyOwdtIuTDu+UZNKfCNPBLLvCGDY3x0WJr0DfTXFZ4OR/c9+X4REuyzDDBvDEZfRqfh5wVGVU6bdI+XWX4rpPFMypOnArepTKkpzqzvgMAAP//AwBQSwMEFAAGAAgAAAAhAEA/ZG1AAwAA7A4AACcAAAB4bC9wcmludGVyU2V0dGluZ3MvcHJpbnRlclNldHRpbmdzNS5iaW7sVstu00AUPWkKgvLqJyBWbPJqUkqjLkjitnLJizgt3YEbTxurqW3ZDkmK2PUr4AdYsEACiW75B5b8C5xxHJqkFRRIVaQwo/HMXI/PnLm+92jycGHDRxOCo7tQUEAVMRSxgEVkkKTlKe1e8N7k0+OsyrUmLACR2ejVr5i/Hf2EmQgieH0jc81gfwfbnM/wGeUsR6TJlUgIJfsZyYHtG8u6qo1so6jlzXvoRr5EV+Yern948zMGt8KXEm+8DPab4BH+Q/2DHvid/9zlYq1U35DHmEcz8hJ5LCGFB6wprDHeY0HUy3yKcZZmHq1xlOdcrorRskxbmnWJfYrZlqNlEa+IqFpO28+bFtYqtZJW2awVVlFb1ZRiEZuW6QpPjqq6I1zNPBTIZVAShqnXe46AVs+VlVxNQaXtD1D0hlDsjoUNe6eq74mKawgXmq9bhu4aqLimsHzdN20L1UqtXsupdaJzXYDwpK23TL+Hsu0e6C0U7FZL9wUq5WCN1uBba49vLQGl7bREF1urtbpayBVRE57dagfASlVNJ5NdNolguyXbEFh39Z5HAHHuaFiZA7YySmnwr7aqVfXto8ef7xPh3RXgmE2WI/bX2d9kO2azqFvJsA5v1qT2+XCQRYLVQyNQwgPoHMdxQJVrBArpUSV3uTLOuU17Ap1AAQ3OOlyboF4m+RcfcpTk/04Qs6+RfrBub8QyvM8+NbVHDJvrjQBJ5bcOSX5kc2jZPcM7k+WdZBRK3pL9eXnvkq9OT4iQ+37A8oTzwinWk/a15Jyi3/+Gs2T5PvRz/wSjtC/fz+Px0Y9nGRf7PH1qzMsX4eMkVenP4/nFD5YyDwe8x6Pj8mNj3M+Sd59lk7y7jPTWL/Oww1zoUCM6zKd4kNEy77dR4n1Kxuoyq8wxK8wcj/7QqSh9/Xse7ONRBU6XUf+cvY9UIKlBgx21ITWTiE6Ib14AfixQOY8KqfN0/RNJ7ZVF9kez/XuaHHfDcT8Wstigp3ag8jsHbSLkw7vlGTSnwjTwSy7whg2N8dFia9A301xWeDkf3Pfl+ERLsswwwbwxGX0an4ecFRlVOm3SPl1l+K6TxTMqTpwK3qUypKc6s74DAAD//wMAUEsDBBQABgAIAAAAIQAh6NCQHQEAALQBAAAtAAAAeGwvZXh0ZXJuYWxMaW5rcy9fcmVscy9leHRlcm5hbExpbmsxLnhtbC5yZWxzhJBfS8MwFMXfBb9DCfi4pu1ARNaO/dVC54abb3m5tLdtWJqUJEr77b2CQweCb7nncH/n5szmQ6eCD7ROGp2yOIxYgLo0ldRNyt5O28kDC5wHXYEyGlM2omPz7PZm9ooKPC25VvYuIIp2KWu97x85d2WLHbjQ9KjJqY3twNNoG95DeYYGeRJF99z+ZrDsihnkVcpsXsUsOI09Jf/PNnUtS1yb8r1D7f+I4Dh4tBpUIfX5AL4lNtgGfcpqqZAu50Ic0VIfYjfeJdGF5cQBrJfoRP6y3pPxVOyXi4Ieq+fNLl8tiqNIoiSaJLFYggJdInmuRfRiKymQph/9+KWTMqW6pyHtxeGg3OWWnanot5vvSxnPZvyq6+wTAAD//wMAUEsDBBQABgAIAAAAIQD/NwAhQQMAAOwOAAAnAAAAeGwvcHJpbnRlclNldHRpbmdzL3ByaW50ZXJTZXR0aW5nczguYmlu7FbLbtNAFD1JCoLy6icgVmzyalJKoy5I4rZyyYs4Ld2BG08bq6lt2Q5Jitj1K+AHWLBAAolu+QeW/AuccRyapBUUSFWkMqPxzFyPz5y5vvdoCnBhw0cLgqO7UFBEDXGUMI8FZJGi5SntXvDe5NPjrMa1JiwAkZnY1a+Yux37hGgUEby+kb1msL+DrWgEUT5jnOWJNL0SCaFkH5Uc2L6xrKna2DaKWtm4h17kS2x59uHahzc/Y3ArfCnxJstwvyke4T/UP+iB3/nPPS7Wyo11eYw5tCIvUcAi0njAmsYq4z0eRL3MpzhnGebRKkcFzuWqOC1LtGVYF9mnmW15Whbwioiq5XT8gmlhtVova9WNenEF9RVNKZWwYZmu8OSopjvC1cwDgXwWZWGYeqPvCGiNfEXJ1xVUO/4QRW8Kxe5aWLe3a/quqLqGcKH5umXoroGqawrL133TtlCr1hv1vNogOtcFCE86etv0+6jY7r7eRtFut3VfoFoJ1mhNvrV2+dYSUDpOW/SwuVJvqMV8CXXh2e1OAKzU1Ewq1WOTCLZbtg2BNVfvewQQZ46G5VlgM6uUh/9qs1ZT3z56/Pk+Ed5dAY7YZDlkf539TbYjNou6lQrr6GYtap8PBzkkWT00AyXch85xAvtUuWagkB5VcocrE5zbtCfRDRTQ4KzLtUnqZYp/8SFHKf7vJDEHGukH63bHLKP77FFT+8Swud4IkFR+65DkRzaHlp1TvDNd3ilGoeQt2Z+V9w756vSECLnvBSyPOc+fYD1tX0vOafr9bzhLlu9DPw9OME774v08GR+DeJZxscfTpye8fB4+TlGV/jyeX/xgKfNwyHsyOi4+Nib9LHkPWLbIu8dIb/8yD7vMhS41ost8SgQZLfN+C2Xep2SsLrHKHLPCzPHoD52KMtC/58E+HlXgZBn3z+n7SAWSGjTcURtRM4nohPjmOeDHA5XzqJA6Tzc4kdReWWR/ODO4p8lxLxwPYiGHdXpqGyq/c9AhQiG8W55C81KYhn7JB96woTE+2mxN+uYyl2Vezof3fTk+1pIcM0wwb0xGn8bnAWclRpVOm7RfrjJ618nhGRUnQQXvURkylzqzvgMAAP//AwBQSwECLQAUAAYACAAAACEAYOAre8EBAAD0CwAAEwAAAAAAAAAAAAAAAAAAAAAAW0NvbnRlbnRfVHlwZXNdLnhtbFBLAQItABQABgAIAAAAIQC1VTAj9AAAAEwCAAALAAAAAAAAAAAAAAAAAPoDAABfcmVscy8ucmVsc1BLAQItABQABgAIAAAAIQB2xd+LKgQAANMMAAAPAAAAAAAAAAAAAAAAAB8HAAB4bC93b3JrYm9vay54bWxQSwECLQAUAAYACAAAACEASXOILWcBAADzCQAAGgAAAAAAAAAAAAAAAAB2CwAAeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJlbHNQSwECLQAUAAYACAAAACEAQnPRqF0KAADVLwAAGAAAAAAAAAAAAAAAAAAdDgAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1sUEsBAi0AFAAGAAgAAAAhAKpQz9yICAAAdyIAABgAAAAAAAAAAAAAAAAAsBgAAHhsL3dvcmtzaGVldHMvc2hlZXQyLnhtbFBLAQItABQABgAIAAAAIQDQ56xbqwoAAGE+AAAYAAAAAAAAAAAAAAAAAG4hAAB4bC93b3Jrc2hlZXRzL3NoZWV0My54bWxQSwECLQAUAAYACAAAACEA6wYOa1AKAACELAAAGAAAAAAAAAAAAAAAAABPLAAAeGwvd29ya3NoZWV0cy9zaGVldDQueG1sUEsBAi0AFAAGAAgAAAAhAKdZZ/qFGgAAmn4AABgAAAAAAAAAAAAAAAAA1TYAAHhsL3dvcmtzaGVldHMvc2hlZXQ1LnhtbFBLAQItABQABgAIAAAAIQA+hB+CYhAAAFRLAAAYAAAAAAAAAAAAAAAAAJBRAAB4bC93b3Jrc2hlZXRzL3NoZWV0Ni54bWxQSwECLQAUAAYACAAAACEAM5QWWswVAAAoagAAGAAAAAAAAAAAAAAAAAAoYgAAeGwvd29ya3NoZWV0cy9zaGVldDcueG1sUEsBAi0AFAAGAAgAAAAhANUHRa6LFAAA+3AAABgAAAAAAAAAAAAAAAAAKngAAHhsL3dvcmtzaGVldHMvc2hlZXQ4LnhtbFBLAQItABQABgAIAAAAIQBXPso9mQ0AANo7AAAYAAAAAAAAAAAAAAAAAOuMAAB4bC93b3Jrc2hlZXRzL3NoZWV0OS54bWxQSwECLQAUAAYACAAAACEAfF1xaroSAABqcAAAGQAAAAAAAAAAAAAAAAC6mgAAeGwvd29ya3NoZWV0cy9zaGVldDEwLnhtbFBLAQItABQABgAIAAAAIQDORtiypgcAABsfAAAZAAAAAAAAAAAAAAAAAKutAAB4bC93b3Jrc2hlZXRzL3NoZWV0MTEueG1sUEsBAi0AFAAGAAgAAAAhAJfEk+eoDAAAGjQAABkAAAAAAAAAAAAAAAAAiLUAAHhsL3dvcmtzaGVldHMvc2hlZXQxMi54bWxQSwECLQAUAAYACAAAACEARNSG/8IAAABDAQAAJAAAAAAAAAAAAAAAAABnwgAAeGwvd29ya3NoZWV0cy9fcmVscy9zaGVldDEwLnhtbC5yZWxzUEsBAi0AFAAGAAgAAAAhAAPTAOs7BwAAiiAAABMAAAAAAAAAAAAAAAAAa8MAAHhsL3RoZW1lL3RoZW1lMS54bWxQSwECLQAUAAYACAAAACEAzsiEruwdAACs3wEADQAAAAAAAAAAAAAAAADXygAAeGwvc3R5bGVzLnhtbFBLAQItABQABgAIAAAAIQBtng/HpBIAAGI8AAAUAAAAAAAAAAAAAAAAAO7oAAB4bC9zaGFyZWRTdHJpbmdzLnhtbFBLAQItABQABgAIAAAAIQBjsaN+wwAAAEMBAAAkAAAAAAAAAAAAAAAAAMT7AAB4bC93b3Jrc2hlZXRzL19yZWxzL3NoZWV0MTEueG1sLnJlbHNQSwECLQAUAAYACAAAACEASxi9JsMAAABDAQAAJAAAAAAAAAAAAAAAAADJ/AAAeGwvd29ya3NoZWV0cy9fcmVscy9zaGVldDEyLnhtbC5yZWxzUEsBAi0AFAAGAAgAAAAhAKI0GhjCAAAAQgEAACMAAAAAAAAAAAAAAAAAzv0AAHhsL3dvcmtzaGVldHMvX3JlbHMvc2hlZXQ4LnhtbC5yZWxzUEsBAi0AFAAGAAgAAAAhABPELBPCAAAAQgEAACMAAAAAAAAAAAAAAAAA0f4AAHhsL3dvcmtzaGVldHMvX3JlbHMvc2hlZXQyLnhtbC5yZWxzUEsBAi0AFAAGAAgAAAAhADttMkvBAAAAQgEAACMAAAAAAAAAAAAAAAAA1P8AAHhsL3dvcmtzaGVldHMvX3JlbHMvc2hlZXQxLnhtbC5yZWxzUEsBAi0AFAAGAAgAAAAhADShCZLCAAAAQgEAACMAAAAAAAAAAAAAAAAA1gABAHhsL3dvcmtzaGVldHMvX3JlbHMvc2hlZXQzLnhtbC5yZWxzUEsBAi0AFAAGAAgAAAAhAEOWEaPCAAAAQgEAACMAAAAAAAAAAAAAAAAA2QEBAHhsL3dvcmtzaGVldHMvX3JlbHMvc2hlZXQ0LnhtbC5yZWxzUEsBAi0AFAAGAAgAAAAhAGTzNCLCAAAAQgEAACMAAAAAAAAAAAAAAAAA3AIBAHhsL3dvcmtzaGVldHMvX3JlbHMvc2hlZXQ1LnhtbC5yZWxzUEsBAi0AFAAGAAgAAAAhAExaKnrCAAAAQgEAACMAAAAAAAAAAAAAAAAA3wMBAHhsL3dvcmtzaGVldHMvX3JlbHMvc2hlZXQ2LnhtbC5yZWxzUEsBAi0AFAAGAAgAAAAhAGs/D/vDAAAAQgEAACMAAAAAAAAAAAAAAAAA4gQBAHhsL3dvcmtzaGVldHMvX3JlbHMvc2hlZXQ3LnhtbC5yZWxzUEsBAi0AFAAGAAgAAAAhAIVRP5nDAAAAQgEAACMAAAAAAAAAAAAAAAAA5gUBAHhsL3dvcmtzaGVldHMvX3JlbHMvc2hlZXQ5LnhtbC5yZWxzUEsBAi0AFAAGAAgAAAAhADzhlp3RAQAA8AMAACIAAAAAAAAAAAAAAAAA6gYBAHhsL2V4dGVybmFsTGlua3MvZXh0ZXJuYWxMaW5rMS54bWxQSwECLQAUAAYACAAAACEAQD9kbUADAADsDgAAJwAAAAAAAAAAAAAAAAD7CAEAeGwvcHJpbnRlclNldHRpbmdzL3ByaW50ZXJTZXR0aW5nczEuYmluUEsBAi0AFAAGAAgAAAAhAEA/ZG1AAwAA7A4AACcAAAAAAAAAAAAAAAAAgAwBAHhsL3ByaW50ZXJTZXR0aW5ncy9wcmludGVyU2V0dGluZ3MyLmJpblBLAQItABQABgAIAAAAIQBAP2RtQAMAAOwOAAAnAAAAAAAAAAAAAAAAAAUQAQB4bC9wcmludGVyU2V0dGluZ3MvcHJpbnRlclNldHRpbmdzMy5iaW5QSwECLQAUAAYACAAAACEAy4r110EDAADsDgAAJwAAAAAAAAAAAAAAAACKEwEAeGwvcHJpbnRlclNldHRpbmdzL3ByaW50ZXJTZXR0aW5nczQuYmluUEsBAi0AFAAGAAgAAAAhAP83ACFBAwAA7A4AACcAAAAAAAAAAAAAAAAAEBcBAHhsL3ByaW50ZXJTZXR0aW5ncy9wcmludGVyU2V0dGluZ3M5LmJpblBLAQItABQABgAIAAAAIQAOb/JkQgMAAOwOAAAoAAAAAAAAAAAAAAAAAJYaAQB4bC9wcmludGVyU2V0dGluZ3MvcHJpbnRlclNldHRpbmdzMTAuYmluUEsBAi0AFAAGAAgAAAAhAP83ACFBAwAA7A4AACgAAAAAAAAAAAAAAAAAHh4BAHhsL3ByaW50ZXJTZXR0aW5ncy9wcmludGVyU2V0dGluZ3MxMS5iaW5QSwECLQAUAAYACAAAACEAy4r110EDAADsDgAAKAAAAAAAAAAAAAAAAAClIQEAeGwvcHJpbnRlclNldHRpbmdzL3ByaW50ZXJTZXR0aW5nczEyLmJpblBLAQItABQABgAIAAAAIQAJoU6hkQoAAPk2AAAQAAAAAAAAAAAAAAAAACwlAQB4bC9jYWxjQ2hhaW4ueG1sUEsBAi0AFAAGAAgAAAAhAP83ACFBAwAA7A4AACcAAAAAAAAAAAAAAAAA6y8BAHhsL3ByaW50ZXJTZXR0aW5ncy9wcmludGVyU2V0dGluZ3M3LmJpblBLAQItABQABgAIAAAAIQAUZF4VawEAAJoCAAARAAAAAAAAAAAAAAAAAHEzAQBkb2NQcm9wcy9jb3JlLnhtbFBLAQItABQABgAIAAAAIQAL+sRVOAIAALwGAAAQAAAAAAAAAAAAAAAAABM2AQBkb2NQcm9wcy9hcHAueG1sUEsBAi0AFAAGAAgAAAAhAEA/ZG1AAwAA7A4AACcAAAAAAAAAAAAAAAAAgTkBAHhsL3ByaW50ZXJTZXR0aW5ncy9wcmludGVyU2V0dGluZ3M2LmJpblBLAQItABQABgAIAAAAIQBAP2RtQAMAAOwOAAAnAAAAAAAAAAAAAAAAAAY9AQB4bC9wcmludGVyU2V0dGluZ3MvcHJpbnRlclNldHRpbmdzNS5iaW5QSwECLQAUAAYACAAAACEAIejQkB0BAAC0AQAALQAAAAAAAAAAAAAAAACLQAEAeGwvZXh0ZXJuYWxMaW5rcy9fcmVscy9leHRlcm5hbExpbmsxLnhtbC5yZWxzUEsBAi0AFAAGAAgAAAAhAP83ACFBAwAA7A4AACcAAAAAAAAAAAAAAAAA80EBAHhsL3ByaW50ZXJTZXR0aW5ncy9wcmludGVyU2V0dGluZ3M4LmJpblBLBQYAAAAAMAAwADwOAAB5RQEAAAA="

@app.route("/gst-template/<ttype>")
def gst_template_download(ttype):
    """Serve pre-built sales summary Excel templates."""
    import base64 as _b64, io
    if ttype == "consolidated":
        data  = _b64.b64decode(_GST_CONS_B64)
        fname = "GST_Sales_Consolidated_Template.xlsx"
    else:
        data  = _b64.b64decode(_GST_BRANCH_B64)
        fname = "GST_Sales_BranchWise_Template.xlsx"
    from flask import send_file
    return send_file(
        io.BytesIO(data),
        download_name=fname,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/tool/gst-reconciliation")
@login_required
def tool_gst_reconciliation():
    user = get_user_by_id(session["uid"])
    from flask import make_response
    resp = make_response(render_template_string(GST_RECON_T, **user_ctx(user)))
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


# NOTE: gunicorn timeout should be >= 180s for large ZIPs.
# In gunicorn.conf.py set: timeout = 300
@app.route("/gst-process", methods=["POST"])
@login_required
def gst_process():
    try:
        user = get_user_by_id(session["uid"])
        if not user["is_admin"] and uploads_remaining(user) <= 0:
            return jsonify({"status": "error",
                "message": f"No uploads remaining. Contact {CONTACT_EMAIL} to recharge."})

        if "sales_file" not in request.files or "gst_file" not in request.files:
            return jsonify({"status": "error", "message": "Please upload both files."})

        sales_f = request.files["sales_file"]
        gst_f = request.files["gst_file"]

        if not sales_f.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({"status": "error", "message": "Sales file must be .xlsx or .xls"})
        if not gst_f.filename.lower().endswith('.zip'):
            return jsonify({"status": "error", "message": "GSTR 3B file must be a .zip"})

        try:
            mappings = json.loads(request.form.get("mappings", "{}"))
        except:
            return jsonify({"status": "error", "message": "Invalid mapping data."})

        # Handle consolidated checkbox
        consolidated_mode = request.form.get("consolidated_mode", "").lower() == "true"
        consolidated_col  = request.form.get("consolidated_col", "").strip()
        if consolidated_mode:
            mappings["__consolidated__"] = consolidated_col or "__auto__"

        if not mappings:
            return jsonify({"status": "error", "message": "Please provide at least one state-column mapping or tick Consolidated Sales."})

        on = request.form.get("output_name", "").strip()
        h = uuid.uuid4().hex
        sales_path = os.path.join(UPLOAD_DIR, f"{h}_sales.xlsx")
        gst_path = os.path.join(UPLOAD_DIR, f"{h}_gst.zip")
        op = os.path.join(OUTPUT_DIR, f"{h}_out.xlsx")

        try:
            orig = sales_f.filename.lower()
            if orig.endswith('.xls') and not orig.endswith('.xlsx'):
                xls_tmp = os.path.join(UPLOAD_DIR, f"{h}_sales.xls")
                sales_f.save(xls_tmp)
                _convert_xls_to_xlsx(xls_tmp, sales_path)
                try: os.remove(xls_tmp)
                except: pass
            else:
                sales_f.save(sales_path)
            gst_f.save(gst_path)

            result = _process_gst_reconciliation(sales_path, gst_path, mappings, op)
        except Exception as e:
            return jsonify({"status": "error", "message": f"Processing error: {e}"})
        finally:
            for p in (sales_path, gst_path):
                try: os.remove(p)
                except: pass

        if result['status'] != 'success':
            return jsonify(result)

        fname = f"{on or 'GST_Reconciliation'}.xlsx"
        log_usage(user["id"], fname)
        return jsonify({"status": "success", "log": result["log"],
                        "file_id": h, "filename": fname})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Unexpected error: {e}"}), 500


# ═══════════════════════════════════════════════════════════════════════════════
#  TRIAL BALANCE → BALANCE SHEET ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

try:
    from tb_processor import analyze_trial_balance, process_tb_to_bs
    TB_PROCESSOR_AVAILABLE = True
except ImportError:
    TB_PROCESSOR_AVAILABLE = False


@app.route("/tool/tb-to-bs")
def tb_to_bs_page():
    if "uid" not in session:
        return redirect("/login")
    user = get_user_by_id(session["uid"])
    if not user:
        return redirect("/login")
    ctx = user_ctx(user)
    return render_template_string(TB_BS_TEMPLATE, **ctx)


@app.route("/tb-analyse", methods=["POST"])
def tb_analyse():
    if "uid" not in session:
        return jsonify({"status": "error", "message": "Session expired — please refresh the page and log in again"}), 401
    user = get_user_by_id(session["uid"])
    if not user:
        return jsonify({"status": "error", "message": "Session expired — please refresh the page and log in again"}), 401
    if not TB_PROCESSOR_AVAILABLE:
        return jsonify({"status": "error", "message": "TB processor not available on this server"}), 500

    try:
        tb_file = request.files.get("tb_file")
        if not tb_file:
            return jsonify({"status": "error", "message": "No trial balance file uploaded"})

        import tempfile, os
        tmp = tempfile.mkdtemp()
        # Preserve original extension for PDF detection
        orig_name = tb_file.filename or "tb.xlsx"
        ext = ".pdf" if orig_name.lower().endswith(".pdf") else ".xlsx"
        tb_path = os.path.join(tmp, "tb" + ext)
        tb_file.save(tb_path)

        result = analyze_trial_balance(tb_path)

        try:
            os.remove(tb_path)
            os.rmdir(tmp)
        except Exception:
            pass

        if "error" in result:
            return jsonify({"status": "error", "message": result["error"]})

        return jsonify({"status": "success", **result})

    except Exception as e:
        import traceback
        return jsonify({"status": "error", "message": f"Analysis failed: {e}\n{traceback.format_exc()}"}), 500


@app.route("/tb-read-bs", methods=["POST"])
def tb_read_bs():
    """Read Capital Account and Fixed Assets sheets from uploaded BS template."""
    if "uid" not in session:
        return jsonify({"status": "error", "message": "Not logged in"}), 401
    try:
        import tempfile, os, re
        from openpyxl import load_workbook
        from openpyxl.cell import MergedCell
        from openpyxl.utils import column_index_from_string

        bs_file = request.files.get("bs_file")
        if not bs_file:
            return jsonify({"status": "error", "message": "No BS template uploaded"})

        tmp = tempfile.mkdtemp()
        bs_path = os.path.join(tmp, "bs.xlsx")
        bs_file.save(bs_path)

        wb = load_workbook(bs_path, read_only=True, data_only=True)
        # Parallel formula-view workbook — used to resolve cells whose
        # data_only value is None because they hold a formula with no
        # cached <v> (common after tb_processor.py's openpyxl round-trip,
        # which drops cached values for formulas it didn't itself compute).
        # E.g. capital!B8 = "=F40" (the proprietor's name lives in F40, and
        # B8 just references it) — data_only reads B8 as None, but we can
        # resolve it ourselves for simple same-sheet cell references.
        wb_f = load_workbook(bs_path, read_only=True, data_only=False)
        result = {"capital": None, "fixed_assets": None}

        def _resolve_cell(ws_do, ws_f, row, col):
            """Return ws_do.cell(row,col).value, or — if that's None and the
            formula-view cell is a simple same-sheet reference like '=F40' —
            the resolved value of the referenced cell instead."""
            val = ws_do.cell(row, col).value
            if val is not None:
                return val
            try:
                fval = ws_f.cell(row, col).value
            except Exception:
                return val
            if isinstance(fval, str) and fval.startswith('='):
                m = re.match(r'^=\$?([A-Z]+)\$?(\d+)$', fval.strip())
                if m:
                    ref_col_letters, ref_row = m.group(1), int(m.group(2))
                    ref_col = column_index_from_string(ref_col_letters)
                    try:
                        return ws_do.cell(ref_row, ref_col).value
                    except Exception:
                        return val
            return val

        # ── Read Capital Account sheet ──────────────────────────────────
        cap_sheet = None
        for sn in wb.sheetnames:
            if "capital" in sn.lower():
                cap_sheet = sn
                break

        if cap_sheet:
            ws = wb[cap_sheet]
            ws_f = wb_f[cap_sheet]
            rows_data = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=50, max_col=15, values_only=False), start=1):
                r = []
                for col_idx, c in enumerate(row, start=1):
                    if isinstance(c, MergedCell):
                        r.append(None)
                    else:
                        r.append(_resolve_cell(ws, ws_f, row_idx, col_idx))
                rows_data.append(r)

            # Find the header row with "Sr. No." or "Name of Proprietor/Partner"
            partners = []
            header_row_idx = None
            name_col = None  # which column has names
            opening_col = None  # which column has opening balance

            for i, row in enumerate(rows_data):
                row_str = " ".join(str(v or "").lower() for v in row)
                if ("sr" in row_str and "name" in row_str) or \
                   ("name of" in row_str and ("proprietor" in row_str or "partner" in row_str)):
                    header_row_idx = i
                    # Find which col is "Name" and which is "As at 1st April"
                    for ci, val in enumerate(row):
                        vl = str(val or "").lower().strip()
                        if "name" in vl and ("proprietor" in vl or "partner" in vl or "of" in vl):
                            name_col = ci
                        if "as at" in vl and "april" in vl:
                            opening_col = ci
                    break

            if header_row_idx is None:
                # Fallback: look for "As at 1st April" row
                for i, row in enumerate(rows_data):
                    row_str = " ".join(str(v or "").lower() for v in row)
                    if "as at" in row_str and "april" in row_str:
                        header_row_idx = i
                        for ci, val in enumerate(row):
                            if isinstance(val, str) and "as at" in val.lower() and "april" in val.lower():
                                opening_col = ci
                                break
                        name_col = 1  # default
                        break

            # Stop words — rows containing these are NOT partners
            _cap_stop = {"total", "previous year", "previous year (py)", "py",
                         "chartered accountant", "ca.", "auditor", "partner ",
                         "proprietor", "director", "secretary", "for ",
                         "sd/-", "authorised", "firm", "registration",
                         "arun gupta", "arun kumar",  # signatory names from image
                         }

            if header_row_idx is not None:
                nc = name_col if name_col is not None else 1
                oc = opening_col if opening_col is not None else 2
                for i in range(header_row_idx + 1, min(header_row_idx + 12, len(rows_data))):
                    row = rows_data[i]
                    if not row or all(v is None for v in row):
                        break  # blank row = end of data

                    name_val = row[nc] if len(row) > nc else None
                    if not isinstance(name_val, str) or len(name_val.strip()) < 2:
                        continue

                    nm = name_val.strip()
                    nm_low = nm.lower()

                    # Remove parentheses wrapper if present
                    if nm.startswith("("):
                        nm = nm.lstrip("(").rstrip(")")
                        nm_low = nm.lower()

                    # Skip stop words
                    if any(sw in nm_low for sw in _cap_stop):
                        continue  # skip this row, keep scanning (don't break)

                    # Skip if too short or looks like a label
                    if len(nm) < 3:
                        continue
                    if "account" in nm_low and "capital" not in nm_low:
                        continue

                    # Skip rows that have no number at all (pure text rows / signatures)
                    has_any_number = any(
                        isinstance(row[ci], (int, float)) and row[ci] != 0
                        for ci in range(2, min(len(row), 10))
                    )
                    # Also check if opening is there (row might have 0 opening for new partner)
                    has_sr_no = isinstance(row[0], (int, float))

                    if not has_any_number and not has_sr_no:
                        continue  # no Sr. No. and no numbers = not a data row

                    # Get opening balance
                    opening = 0
                    if len(row) > oc and isinstance(row[oc], (int, float)):
                        opening = float(row[oc])
                    else:
                        for ci in range(oc, min(oc + 3, len(row))):
                            if isinstance(row[ci], (int, float)) and row[ci] != 0:
                                opening = float(row[ci])
                                break

                    partners.append({"name": nm, "opening": opening, "row": i + 1})

            if partners:
                # Detect column layout from header row
                cap_columns = []
                col_map = {}  # {field_key: column_index (1-based)}
                if header_row_idx is not None:
                    hrow = rows_data[header_row_idx]
                    for ci, hv in enumerate(hrow):
                        hs = str(hv or "").lower().strip()
                        if "introduced" in hs or "capital intro" in hs:
                            cap_columns.append({"key": "introduced", "label": "Capital Introduced", "col": ci+1})
                            col_map["introduced"] = ci + 1
                        elif "interest" in hs and "capital" in hs:
                            cap_columns.append({"key": "interest_on_capital", "label": "Interest on Capital", "col": ci+1})
                            col_map["interest_on_capital"] = ci + 1
                        elif "salary" in hs:
                            cap_columns.append({"key": "salary", "label": "Salary", "col": ci+1})
                            col_map["salary"] = ci + 1
                        elif "withdraw" in hs:
                            cap_columns.append({"key": "withdrawals", "label": "Withdrawals", "col": ci+1})
                            col_map["withdrawals"] = ci + 1

                result["capital"] = {
                    "sheet": cap_sheet,
                    "partners": partners,
                    "columns": cap_columns,
                    "col_map": col_map,
                }

        # ── Read Fixed Assets sheet ─────────────────────────────────────
        fa_sheet = None
        for sn in wb.sheetnames:
            sl = sn.lower()
            if "fixed asset" in sl or "fa " in sl or sl.startswith("fa") or "ppe" in sl:
                fa_sheet = sn
                break

        if fa_sheet:
            # ── Build opening WDV lookup from Fixed Assets P. Yr. sheet ──
            # Col I (index 8) of P.Yr. sheet = closing WDV = opening for current year
            # This is the authoritative source — C.Yr. col B is just =P.Yr.!I9 formula
            py_opening = {}  # {asset_name_lower: opening_wdv}
            py_sheet_name = None
            for sn in wb.sheetnames:
                if "p. yr" in sn.lower() or "p.yr" in sn.lower() or \
                   ("fixed" in sn.lower() and "p" in sn.lower()):
                    py_sheet_name = sn
                    break

            if py_sheet_name:
                ws_py = wb[py_sheet_name]
                py_rows = []
                for row in ws_py.iter_rows(min_row=1, max_row=45, max_col=12, values_only=False):
                    r = []
                    for c in row:
                        r.append(None if isinstance(c, MergedCell) else c.value)
                    py_rows.append(r)

                # Find header to locate closing WDV col (W.D.V AS ON 31.03.xxxx)
                py_closing_col = 8  # col I default (0-indexed)
                py_rate_col = None
                for i, row in enumerate(py_rows[:8]):
                    row_str = " ".join(str(v or "").lower() for v in row)
                    if "w.d.v" in row_str and ("31.03" in row_str or "closing" in row_str):
                        for ci, val in enumerate(row):
                            vl = str(val or "").lower()
                            if ("w.d.v" in vl or "as on" in vl) and ci > 4:
                                py_closing_col = ci
                        break

                # Read asset names and closing WDV
                import re as _re2
                _date_re2 = _re2.compile(r'^\d{1,2}[./]\d{1,2}[./]\d{2,4}$')

                # Use col I (index 8) for closing WDV if it has values,
                # otherwise fall back to col B (index 1, opening WDV of P.Yr.)
                # Col I is a formula (=F-H) and loses its cached <v> after an
                # openpyxl round-trip, while col B holds plain numeric constants
                # that survive.  Either way, the goal is to give the user a
                # non-zero reference figure to start from.
                ws_py_f = wb_f[py_sheet_name]
                for row in py_rows[5:]:
                    nm = str(row[0] or "").strip()
                    if not nm or len(nm) < 2:
                        continue
                    if nm.isupper():  # skip category headers
                        continue
                    if _date_re2.match(nm):
                        continue
                    if any(sw in nm.lower() for sw in
                           {"total", "particular", "addition", "amount", "rate",
                            "w.d.v", "building", "property", "chair"}):
                        continue
                    closing = row[py_closing_col] if len(row) > py_closing_col else None
                    if not isinstance(closing, (int, float)):
                        # col I is None → fall back to col B (opening WDV of P.Yr.)
                        closing = row[1] if len(row) > 1 else None
                    if isinstance(closing, (int, float)):
                        py_opening[nm.lower().strip()] = float(closing)

            ws = wb[fa_sheet]
            ws_f_fa = wb_f[fa_sheet]
            # Build a parallel formula-view row list for resolving cross-sheet
            # refs in col A (asset names like ='Fixed Assets P. Yr.'!A10) and
            # col B (opening WDV like ='Fixed Assets P. Yr.'!I10) that have no
            # cached <v> after an openpyxl round-trip.
            rows_data_f = []
            for row in ws_f_fa.iter_rows(min_row=1, max_row=60, max_col=15, values_only=False):
                r = []
                for c in row:
                    if isinstance(c, MergedCell):
                        r.append(None)
                    else:
                        r.append(c.value if hasattr(c, 'value') else None)
                rows_data_f.append(r)

            rows_data = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=60, max_col=15, values_only=False), start=1):
                r = []
                for col_idx, c in enumerate(row, start=1):
                    if isinstance(c, MergedCell):
                        r.append(None)
                    else:
                        r.append(c.value if hasattr(c, 'value') else None)
                rows_data.append(r)

            # Find header row (PARTICULARS / W.D.V / ADDITIONS / SALE / RATE)
            fa_header_row = None
            wdv_col = None  # opening WDV column
            rate_col = None
            for i, row in enumerate(rows_data):
                row_str = " ".join(str(v or "").lower() for v in row)
                if ("particular" in row_str or "w.d.v" in row_str) and \
                   ("addition" in row_str or "rate" in row_str or "sale" in row_str):
                    fa_header_row = i
                    for ci, val in enumerate(row):
                        vl = str(val or "").lower().strip()
                        if "w.d.v" in vl or "opening" in vl or "01.04" in vl or "as on" in vl:
                            wdv_col = ci
                        if vl in ("rate", "%", "rate %"):
                            rate_col = ci
                    break

            # Skip words for FA — not actual assets
            import re as _re
            _fa_skip = {"particular", "w.d.v", "amount", "total", "grand total",
                        "rate", "addition", "sale", "depreciation",
                        "as on", "note", "ca.", "chartered", "auditor",
                        "sd/-", "partner", "proprietor", "director", "for ",
                        "property, plant", "intangible asset",
                        "amount in rs", "amount in"}

            # Date pattern: 01.04.2024, 31.03.2025, 1.4.2024 etc
            _date_re = _re.compile(r'^\d{1,2}[./]\d{1,2}[./]\d{2,4}$')
            # Note/reference number pattern: "7 Property, Plant..." or starts with digit+space
            _note_re = _re.compile(r'^\d+\s+\w')

            assets = []
            start_row = (fa_header_row + 1) if fa_header_row is not None else 5
            wc = wdv_col if wdv_col is not None else 1  # default col B for opening WDV
            rc = rate_col  # rate column

            for i in range(start_row, min(start_row + 40, len(rows_data))):
                row = rows_data[i]
                row_f = rows_data_f[i] if i < len(rows_data_f) else row
                if not row or all(v is None for v in row):
                    continue

                # Get asset name from col A — if data_only is None (formula cell
                # with no cached <v>), check the formula-view cell: if it's a
                # cross-sheet ref like ='Fixed Assets P. Yr.'!A10, look that
                # name up in py_opening keys as a fallback.
                name_val = row[0] if len(row) > 0 else None
                if name_val is None and len(row_f) > 0:
                    fval = row_f[0]
                    if isinstance(fval, str) and fval.startswith('=') and '!' in fval:
                        # Cross-sheet ref: find matching name from py_opening
                        for known_name in py_opening:
                            name_val = known_name
                            break
                        # Better: use formula-view of py sheet directly
                        m = __import__('re').search(r"!([A-Z]+)(\d+)$", fval)
                        if m:
                            ref_col_s = m.group(1)
                            ref_row_i = int(m.group(2))
                            from openpyxl.utils import column_index_from_string as _c2i
                            try:
                                ws_py_do_tmp = wb[py_sheet_name]
                                name_val = ws_py_do_tmp.cell(ref_row_i, _c2i(ref_col_s)).value
                            except Exception:
                                pass
                if name_val is None and len(row) > 1:
                    name_val = row[1]

                if not isinstance(name_val, str) or len(name_val.strip()) < 2:
                    continue

                nm = name_val.strip()
                nm_low = nm.lower()

                # Skip "Total" row — stop scanning
                if nm_low.strip() in ("total", "grand total"):
                    break

                # Skip if it matches any stop word
                if any(sw in nm_low for sw in _fa_skip):
                    continue

                # Skip dates like "01.04.2024"
                if _date_re.match(nm):
                    continue

                # Skip note numbers like "7 Property, Plant and Equipment"
                if _note_re.match(nm):
                    continue

                # Skip pure numbers
                if nm.replace(",", "").replace(".", "").replace("-", "").isdigit():
                    continue

                # Skip ALL-CAPS category headers (PLANT & MACHINERY, VEHICLE, etc.)
                has_own_number = False
                if len(row) > wc and isinstance(row[wc], (int, float)) and row[wc] != 0:
                    has_own_number = True
                if nm.isupper() and not has_own_number:
                    continue

                # ── Opening WDV: prefer P.Yr. closing col I (authoritative) ──
                # The C.Yr. col B is a formula (=P.Yr.!I9) — data_only may be stale.
                # Match by name to P.Yr. lookup table first.
                opening_wdv = py_opening.get(nm_low.strip(), None)
                if opening_wdv is None:
                    # Fallback: read C.Yr. col B (data_only cached value)
                    if len(row) > wc and isinstance(row[wc], (int, float)):
                        opening_wdv = float(row[wc])
                    else:
                        opening_wdv = 0

                rate = 0
                if rc is not None and len(row) > rc and isinstance(row[rc], (int, float)):
                    rate = float(row[rc])
                else:
                    # Try to find rate in later columns (look for value 5-100)
                    for ci in range(max(wc + 3, 5), min(len(row), 12)):
                        v = row[ci]
                        if isinstance(v, (int, float)) and 5 <= v <= 100:
                            rate = float(v)
                            break

                assets.append({
                    "name": nm,
                    "opening_wdv": opening_wdv,
                    "rate": rate,
                    "row": i + 1,
                })

            if assets:
                result["fixed_assets"] = {
                    "sheet": fa_sheet,
                    "assets": assets,
                }

        wb.close()
        wb_f.close()
        try:
            os.remove(bs_path)
            os.rmdir(tmp)
        except:
            pass

        return jsonify({"status": "success", **result})

    except Exception as e:
        import traceback
        return jsonify({"status": "error", "message": f"Failed to read BS: {e}\n{traceback.format_exc()}"}), 500



def _rollover_fixed_assets(output_path, cy_year, log, source_path=None):
    """
    Fixed-assets rollover for the year-shift tool.

    Behavior:
    1. Output Fixed Assets P. Yr. becomes a value snapshot of the uploaded
       Fixed Assets C. Yr. sheet.
    2. Output Fixed Assets C. Yr. keeps formulas, but additions/sale inputs are
       cleared so the new year opens from the mirrored PY closing balances.

    PERFORMANCE FIX (2026-07-10): moved workbook opens AFTER the FA sheet
    detection check so files without FA sheets (e.g. Deluxe format) return
    immediately without paying the openpyxl load cost. Also uses
    _fast_load_workbook to strip bloated styles.xml (38k+ named styles) that
    cause 7s+ load times, preventing Render request timeouts.
    """
    from openpyxl.cell import MergedCell as _MC
    import re as _re
    import zipfile as _zf

    # ── Fast sheet-name check BEFORE opening any workbook ────────────────────
    # Avoids paying 4 × ~7s openpyxl opens when no FA sheet exists.
    try:
        with _zf.ZipFile(output_path, "r") as _zi:
            _wb_xml = _zi.read("xl/workbook.xml").decode("utf-8", errors="replace")
        import re as _re2
        _snames_quick = _re2.findall(r'<sheet\b[^>]*name="([^"]+)"', _wb_xml)
    except Exception:
        _snames_quick = []
    _cy_quick, _py_quick = detect_fixed_asset_sheet_names(_snames_quick)
    if not _cy_quick:
        log.append("⚠ FA C.Yr. sheet not found — skipping FA rollover")
        return

    # ── Use fast loader to strip bloated styles before openpyxl opens ────────
    try:
        from processor import _fast_load_workbook as _flwb
    except ImportError:
        from openpyxl import load_workbook as _flwb

    def _lwb(path, **kwargs):
        return _flwb(path, **kwargs)

    wb = _lwb(output_path)
    wb_do = _lwb(output_path, data_only=True)
    src_wb = src_wb_do = None
    if source_path:
        src_wb = _lwb(source_path)
        src_wb_do = _lwb(source_path, data_only=True)

    try:
        cy_sn, py_sn = detect_fixed_asset_sheet_names(wb.sheetnames)
        if not cy_sn:
            log.append("⚠ FA C.Yr. sheet not found")
            return

        ws_cy = wb[cy_sn]
        ws_cy_do = wb_do[cy_sn]

        src_sheetnames = src_wb.sheetnames if src_wb else wb.sheetnames
        src_cy_sn, _ = detect_fixed_asset_sheet_names(src_sheetnames)
        src_cy_ws = src_wb[src_cy_sn] if src_wb and src_cy_sn else wb[cy_sn]
        src_cy_ws_do = src_wb_do[src_cy_sn] if src_wb_do and src_cy_sn else wb_do[cy_sn]

        # Detect FA layout from current CY sheet
        op_col = 2; ag_col = 3; al_col = 4; sl_col = 5; rt_col = 7; cl_col = 9
        data_start = 9
        date_row = 7
        for r in range(1, 15):
            vals = []
            for c in range(1, 12):
                cell = ws_cy.cell(r, c)
                vals.append("" if isinstance(cell, _MC) else str(cell.value or "").lower().strip())
            row_str = " ".join(vals)
            if "01.04" in row_str or "31.03" in row_str:
                date_row = r
            if "greater" in row_str or ("addition" in row_str and "sale" in row_str):
                data_start = r + 1
                for ci, v in enumerate(vals, 1):
                    if "w.d.v" in v and ci < 3: op_col = ci
                    elif "greater" in v: ag_col = ci
                    elif "less" in v and v != "less": al_col = ci
                    elif v in ("sale", "sales"): sl_col = ci
                    elif v in ("%", "rate", "rate %"): rt_col = ci
                    elif "w.d.v" in v and ci > 5: cl_col = ci

        # 1) Mirror source CY sheet → output PY sheet as a direct row-by-row copy.
        #
        # The PY sheet should be a COMPLETE HISTORICAL RECORD of the uploaded
        # CY year — including all additions (Camera DVR, Steel Angle, Car
        # additions etc.) and any sales (Property sale). This is the factual
        # data for the prior year and must not be filtered or omitted.
        #
        # Approach: copy every cell from src_cy_ws directly into ws_py,
        # row-for-row, column-for-column. For formula cells, write the
        # RESOLVED VALUE (from src_cy_ws_do) rather than the formula text,
        # since the PY sheet is a value-snapshot, not a live calculation sheet.
        # Exception: same-sheet formulas that have NO cross-sheet refs and
        # whose resolved value would be None (e.g. =B8+C8... on a section
        # header row) are left as None (blank), matching the source display.
        copied = 0
        if py_sn:
            ws_py = wb[py_sn]

            # First, blank the entire PY sheet so stale data from the old
            # prior-prior-year template doesn't bleed through.
            merged_children_py = set()
            for merged in ws_py.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged.bounds
                for rr in range(min_row, max_row + 1):
                    for cc in range(min_col, max_col + 1):
                        if (rr, cc) != (min_row, min_col):
                            merged_children_py.add((rr, cc))

            for r in range(1, ws_py.max_row + 1):
                for c in range(1, min(ws_py.max_column, src_cy_ws.max_column) + 1):
                    if (r, c) in merged_children_py:
                        continue
                    from openpyxl.cell import MergedCell as _MC2
                    tgt = ws_py.cell(r, c)
                    if isinstance(tgt, _MC2):
                        continue
                    tgt.value = None

            # Copy source CY → output PY row-by-row, cell-by-cell.
            # Find the last row that belongs to the FA table: the last row
            # that has EITHER a non-empty col A (asset/section name or Total)
            # OR a numeric rate value in col G (the Rate % column).
            # This excludes stray formula-overflow values that appear below
            # the table boundary in columns B-I with no corresponding label.
            src_last_row = 1
            for _r in range(src_cy_ws.max_row, 0, -1):
                _a = src_cy_ws_do.cell(_r, 1).value
                _g = src_cy_ws_do.cell(_r, rt_col).value
                if _a is not None or isinstance(_g, (int, float)):
                    src_last_row = _r
                    break

            for r in range(1, src_last_row + 1):
                for c in range(1, 10):  # strictly cols 1-9 only
                    src_cell_f = src_cy_ws.cell(r, c)
                    from openpyxl.cell import MergedCell as _MC3
                    if isinstance(src_cell_f, _MC3):
                        continue
                    val = src_cy_ws_do.cell(r, c).value
                    if val is None:
                        raw = src_cell_f.value
                        if not (isinstance(raw, str) and raw.startswith('=')):
                            val = raw
                    from openpyxl.cell import MergedCell as _MC4
                    tgt = ws_py.cell(r, c)
                    if isinstance(tgt, _MC4):
                        continue
                    tgt.value = val
                    # Copy cell formatting (border, font, fill, alignment,
                    # number_format) so the table borders render correctly
                    # in Excel — without this, the Total row's medium border
                    # is missing and it appears outside the table visually.
                    import copy
                    if src_cell_f.has_style:
                        tgt.border    = copy.copy(src_cell_f.border)
                        tgt.font      = copy.copy(src_cell_f.font)
                        tgt.fill      = copy.copy(src_cell_f.fill)
                        tgt.alignment = copy.copy(src_cell_f.alignment)
                        tgt.number_format = src_cell_f.number_format
                    if val is not None:
                        copied += 1

            # Copy column widths from source CY so PY columns match
            import copy
            for col_letter, col_dim in src_cy_ws.column_dimensions.items():
                ws_py.column_dimensions[col_letter].width = col_dim.width
                ws_py.column_dimensions[col_letter].hidden = col_dim.hidden
            # Copy row heights for rows we've written
            for row_num in range(1, src_last_row + 1):
                if row_num in src_cy_ws.row_dimensions:
                    src_rd = src_cy_ws.row_dimensions[row_num]
                    ws_py.row_dimensions[row_num].height = src_rd.height
                    ws_py.row_dimensions[row_num].hidden = src_rd.hidden

            log.append(f"✓ FA PY: copied source CY sheet into '{py_sn}' ({copied} cells)")

            # 1b) Fix CY opening-WDV formulas after PY restructure.
            #
            # The CY sheet's B column contains cross-sheet formulas like
            # ='Fixed Assets P. Yr.'!I9 that pull the opening WDV from the
            # original PY sheet. Those row numbers were hardcoded to the
            # OLD PY layout (e.g. Battery was at PY row 9). Now that we've
            # replaced the PY sheet with a copy of the CY sheet, the same
            # assets appear at the SAME row numbers as in CY (Battery is
            # now at PY row 10, matching CY row 10). The old formula
            # ='Fixed Assets P. Yr.'!I9 now picks up a section-header
            # (PLANT & MACHINERY, I=blank) instead of Battery's closing WDV.
            #
            # Fix: for each CY B-column cross-sheet formula referencing PY,
            # check whether the referenced PY row in the NEW PY sheet still
            # holds the correct asset name. If not, find the correct row in
            # the new PY (same row as the CY asset row) and update the
            # formula to reference that row instead.
            import re as _re_cyfix
            for r in range(1, ws_cy.max_row + 1):
                from openpyxl.cell import MergedCell as _MC5

                # B column: opening WDV formula e.g. ='Fixed Assets P. Yr.'!I9
                b_cell = ws_cy.cell(r, 2)
                if not isinstance(b_cell, _MC5):
                    bval = b_cell.value
                    if (isinstance(bval, str) and bval.startswith('=')
                            and py_sn.lower() in bval.lower()):
                        m = _re_cyfix.search(r'!I(\d+)', bval)
                        if m:
                            old_py_row = int(m.group(1))
                            new_py_row = r  # new PY = copy of CY, same row
                            if new_py_row != old_py_row:
                                b_cell.value = bval.replace(
                                    f'!I{old_py_row}', f'!I{new_py_row}')

                # A column: asset name formula e.g. ='Fixed Assets P. Yr.'!A9
                # FIX: remap A-col refs the same way as B-col so asset names
                # resolve correctly in the restructured PY sheet.
                a_cell = ws_cy.cell(r, 1)
                if not isinstance(a_cell, _MC5):
                    aval = a_cell.value
                    if (isinstance(aval, str) and aval.startswith('=')
                            and py_sn.lower() in aval.lower()):
                        m_a = _re_cyfix.search(r'!A(\d+)', aval)
                        if m_a:
                            old_a_row = int(m_a.group(1))
                            new_a_row = r  # new PY row = CY row
                            if new_a_row != old_a_row:
                                a_cell.value = aval.replace(
                                    f'!A{old_a_row}', f'!A{new_a_row}')

            # 1c) Fix bs sheet cross-references after PY restructure.
            #
            # The bs sheet contains formulas that reference specific row
            # numbers in the FA P.Yr sheet — e.g. ='Fixed Assets P. Yr.'!I38
            # which pointed to the TOTAL row in the ORIGINAL PY sheet (R38).
            # After our rollover, the new PY sheet is a copy of the CY sheet
            # where the Total row is at a DIFFERENT row (e.g. R42 for Chetan
            # Textiles). The old formula now references the wrong row (Property
            # data row = 0 instead of the Total = 183,827).
            #
            # Also: bs!F8 = '=capital!G11' (PY capital) and similar cross-
            # sheet references can be replaced by plain values by
            # processor.process — restore them from the source file.
            _bs_sn     = next((s for s in wb.sheetnames      if s.lower() == 'bs'), None)
            _bs_src_sn = next((s for s in src_wb.sheetnames  if s.lower() == 'bs'), None) if src_wb else None
            if _bs_sn and _bs_src_sn:
                ws_bs = wb[_bs_sn]
                ws_bs_src = src_wb[_bs_src_sn]

                # Find the new Total row in the new PY sheet.
                # Search by "total" label first; if not found (new PY is a copy
                # of CY whose total row label is the firm name, not "Total"),
                # fall back to finding the last row with a SUM formula in col I.
                new_py_total_row = None
                for _r in range(1, ws_py.max_row + 1):
                    _a = str(ws_py.cell(_r, 1).value or "").strip().lower()
                    if _a == "total":
                        new_py_total_row = _r
                        break
                if new_py_total_row is None:
                    from openpyxl.cell import MergedCell as _MCt
                    for _r in range(ws_py.max_row, 0, -1):
                        _i_cell = ws_py.cell(_r, 9)
                        if isinstance(_i_cell, _MCt):
                            continue
                        _iv = _i_cell.value
                        if isinstance(_iv, str) and _iv.startswith('=') and 'SUM' in _iv.upper():
                            new_py_total_row = _r
                            break
                        if isinstance(_iv, (int, float)) and _iv > 0:
                            new_py_total_row = _r
                            break

                import re as _re_bs
                for r in range(1, ws_bs.max_row + 1):
                    for c in range(1, ws_bs.max_column + 1):
                        src_val = ws_bs_src.cell(r, c).value
                        cur_val = ws_bs.cell(r, c).value
                        if not (isinstance(src_val, str) and src_val.startswith('=')):
                            continue
                        # If the source had a formula but the generated output
                        # has a plain number, restore the formula (fixes capital
                        # reference bs!F8='=capital!G11' being replaced by value)
                        if not (isinstance(cur_val, str) and cur_val.startswith('=')):
                            ws_bs.cell(r, c).value = src_val
                            cur_val = src_val
                        # Fix FA P.Yr row references in bs formulas
                        if py_sn and py_sn in cur_val and new_py_total_row:
                            # Find the old PY total row from the SOURCE bs formula
                            m = _re_bs.search(
                                r"'Fixed Assets P\. Yr\.'!I(\d+)",
                                cur_val
                            )
                            if m:
                                old_ref_row = int(m.group(1))
                                if old_ref_row != new_py_total_row:
                                    new_formula = cur_val.replace(
                                        f'!I{old_ref_row}',
                                        f'!I{new_py_total_row}'
                                    )
                                    ws_bs.cell(r, c).value = new_formula


        cy_data_rows = set()
        rate_values_found = sum(
            1 for r in range(data_start, min(ws_cy.max_row + 1, data_start + 40))
            if isinstance(ws_cy_do.cell(r, rt_col).value, (int, float))
            and ws_cy_do.cell(r, rt_col).value in (5, 10, 15, 20, 25, 30, 40, 60, 100)
        )
        if rate_values_found == 0:
            for try_col in range(6, 11):
                cnt = sum(
                    1 for r in range(data_start, min(ws_cy.max_row + 1, data_start + 40))
                    if isinstance(ws_cy_do.cell(r, try_col).value, (int, float))
                    and ws_cy_do.cell(r, try_col).value in (5, 10, 15, 20, 25, 30, 40, 60, 100)
                )
                if cnt >= 2:
                    rt_col = try_col
                    break
        for r in range(data_start, ws_cy.max_row + 1):
            if isinstance(ws_cy.cell(r, rt_col), _MC):
                continue
            rate_v = ws_cy_do.cell(r, rt_col).value
            if not isinstance(rate_v, (int, float)):
                continue
            nm = str(ws_cy.cell(r, 1).value or "").strip().lower()
            if nm not in ("total", "grand total"):
                cy_data_rows.add(r)

        header_skip = {"additions", "greater than", "less than", "sale", "180 days", "amount in rs.", "particulars", "w.d.v", "amount in rs", "180days"}
        cleared = 0
        for r in range(1, ws_cy.max_row + 1):
            is_data_row = r in cy_data_rows
            for col in (ag_col, al_col, sl_col):
                cell = ws_cy.cell(r, col)
                if isinstance(cell, _MC):
                    continue
                v = cell.value
                if v is None:
                    continue
                vs = str(v).strip().lower()
                if vs in header_skip:
                    continue
                if isinstance(v, str) and 'sum' in vs:
                    continue
                if is_data_row:
                    if isinstance(v, (int, float)):
                        cell.value = None
                        cleared += 1
                    elif isinstance(v, str) and v.startswith('='):
                        body = v[1:].strip()
                        import re as _re2
                        is_arith = _re2.sub(r'[\d\s\.\+\-\*\/\(\)]+', '', body) == ''
                        if is_arith:
                            cell.value = None
                            cleared += 1
                else:
                    if isinstance(v, (int, float)):
                        cell.value = None
                        cleared += 1
                    elif isinstance(v, str) and not v.startswith('='):
                        if _re.fullmatch(r'\d{1,2}[./]\d{1,2}[./]\d{2,4}', v.strip()):
                            cell.value = None
                            cleared += 1

        # Keep CY dates aligned to the new year; PY dates come from mirrored source CY
        try:
            new_oy = int(cy_year) - 1
            new_cy = int(cy_year)
            for r in range(1, max(date_row, data_start) + 2):
                for c in range(1, 12):
                    cell = ws_cy.cell(r, c)
                    if isinstance(cell, _MC):
                        continue
                    v = str(cell.value or "").strip()
                    if _re.fullmatch(r"\d{1,2}[./]04[./]\d{4}", v):
                        cell.value = f"01.04.{new_oy}"
                    elif _re.fullmatch(r"\d{1,2}[./]03[./]\d{4}", v):
                        cell.value = f"31.03.{new_cy}"
        except Exception:
            pass

        try:
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
            wb.calculation.calcMode = 'auto'
        except Exception:
            pass

        wb.save(output_path)
        log.append(f"✓ FA CY: additions/sale inputs cleared ({cleared} cells)")
        log.append("✓ FA rollover complete")
    finally:
        wb.close()
        wb_do.close()
        if src_wb:
            src_wb.close()
        if src_wb_do:
            src_wb_do.close()


def _inject_cap_fa(output_path, cap_entries, fa_entries, log):
    """Inject user-entered Capital A/c and Fixed Assets values into the output BS."""
    from openpyxl import load_workbook
    from openpyxl.cell import MergedCell
    wb = load_workbook(output_path)

    def _is_formula(v):
        return isinstance(v, str) and v.startswith("=")

    def _safe_write(ws, row, col, value):
        """Write to cell only if it's not merged and not a formula."""
        try:
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                return False
            if _is_formula(str(cell.value or "")):
                return False
            cell.value = round(float(value), 2)
            return True
        except Exception:
            return False

    # ── Capital Account injection ──────────────────────────────────────
    if cap_entries:
        cap_sheet = None
        for sn in wb.sheetnames:
            if "capital" in sn.lower():
                cap_sheet = sn
                break
        if cap_sheet:
            ws = wb[cap_sheet]
            for entry in cap_entries:
                row = entry.get("row")
                if not row:
                    continue
                # Write each field to its column using column index from entry
                fields = [
                    ("introduced", "Capital Introduced"),
                    ("interest_on_capital", "Interest on Capital"),
                    ("salary", "Salary"),
                    ("withdrawals", "Withdrawals"),
                ]
                for field_key, field_label in fields:
                    val = entry.get(field_key, 0)
                    col_idx = entry.get(f"{field_key}_col")
                    if val and col_idx:
                        if _safe_write(ws, row, col_idx, val):
                            log.append(f"✓ {cap_sheet}!{chr(64+col_idx)}{row} ({field_label}) = {float(val):,.2f}")

    # ── Fixed Assets injection ─────────────────────────────────────────
    if fa_entries:
        fa_sheet = None
        for sn in wb.sheetnames:
            sl = sn.lower()
            if "fixed asset" in sl or "fa " in sl or sl.startswith("fa") or "ppe" in sl:
                fa_sheet = sn
                break
        if fa_sheet:
            ws = wb[fa_sheet]

            # ── FIX Bug 1: Handle row=0 entries (new assets added by user) ──
            # When the user adds a completely new asset (e.g. Generator, Mobile)
            # via "Add Row", row=0 because it has no existing row in the template.
            # We INSERT a new row in the correct FA section (determined by rate:
            # 15%→Plant & Machinery, 40%→Computers, 10%→Furniture/Building, etc.).
            new_asset_entries = [e for e in fa_entries if not e.get('row')]
            existing_entries  = [e for e in fa_entries if e.get('row')]

            if new_asset_entries:
                # Rate → section keyword mapping (standard IT Act rates)
                RATE_TO_SECTION_KW = {
                    0:  ['land'],
                    5:  ['building'],
                    10: ['furniture', 'building'],
                    15: ['plant', 'machiner'],
                    20: ['vehicle'],
                    30: ['vehicle'],
                    40: ['computer'],
                }

                # Scan the sheet for section header rows (ALL-CAPS, no number in col B)
                section_info = []
                prev_header_row = None
                prev_header_name = None
                for r_idx in range(1, 60):
                    a_val = ws.cell(r_idx, 1).value
                    b_val = ws.cell(r_idx, 2).value
                    if a_val and isinstance(a_val, str):
                        nm = a_val.strip()
                        # Section header: ALL-CAPS, no numeric value in col B
                        if nm.isupper() and len(nm) > 3 and not isinstance(b_val, (int, float)):
                            if prev_header_name is not None:
                                section_info.append({
                                    'header_row': prev_header_row,
                                    'name': prev_header_name,
                                    'end_row': r_idx - 1,
                                })
                            prev_header_row = r_idx
                            prev_header_name = nm
                        # Total row marks end of all sections
                        elif nm.lower() in ('total', 'grand total') and prev_header_name:
                            section_info.append({
                                'header_row': prev_header_row,
                                'name': prev_header_name,
                                'end_row': r_idx - 1,
                            })
                            prev_header_name = None
                            break

                def _section_for_rate(rate, section_label=""):
                    """Find the FA sheet section that matches by label first, then by rate."""
                    rate_int = int(rate) if rate else 15
                    kws = RATE_TO_SECTION_KW.get(rate_int, ['plant', 'machiner'])
                    # If a section label was provided by the user (from the dropdown), try that first
                    if section_label:
                        sec_low = section_label.lower()
                        for sec in section_info:
                            if any(kw in sec['name'].lower() for kw in sec_low.split()):
                                return sec
                    # Fall back to rate-based matching
                    for sec in section_info:
                        sec_low = sec['name'].lower()
                        if any(kw in sec_low for kw in kws):
                            return sec
                    # Default: first Plant & Machinery section
                    for sec in section_info:
                        if 'plant' in sec['name'].lower() or 'machin' in sec['name'].lower():
                            return sec
                    return section_info[0] if section_info else None

                for entry in new_asset_entries:
                    name  = (entry.get('name') or '').strip()
                    rate  = float(entry.get('rate') or 15)
                    gt    = float(entry.get('additions_gt180') or 0)
                    lt    = float(entry.get('additions_lt180') or 0)
                    sale  = float(entry.get('sale') or 0)
                    if not name or (gt == 0 and lt == 0 and sale == 0):
                        continue

                    sec = _section_for_rate(rate, entry.get('section', ''))
                    if not sec:
                        log.append(f"⚠ {fa_sheet}: no section found for new asset '{name}' (rate {rate}%) — skipped")
                        continue

                    # Find the last data row in this section (last non-empty named asset row)
                    insert_at = sec['header_row'] + 1  # default: right after header
                    for scan_r in range(sec['header_row'] + 1, sec['end_row'] + 1):
                        a = ws.cell(scan_r, 1).value
                        if a and isinstance(a, str) and a.strip() and not a.strip().isupper():
                            insert_at = scan_r + 1  # insert after this row

                    # Insert a blank row, then fill it
                    ws.insert_rows(insert_at)

                    ws.cell(insert_at, 1).value = name
                    ws.cell(insert_at, 2).value = 0  # opening WDV (new asset)
                    if gt:
                        ws.cell(insert_at, 3).value = round(gt, 2)
                    if lt:
                        ws.cell(insert_at, 4).value = round(lt, 2)
                    if sale:
                        ws.cell(insert_at, 5).value = round(sale, 2)
                    total_cost = round(gt + lt - sale, 2)
                    # Use formulas (not hardcoded values) so Excel recalculates correctly
                    ws.cell(insert_at, 6).value = f"=B{insert_at}+C{insert_at}+D{insert_at}-E{insert_at}"
                    ws.cell(insert_at, 7).value = rate
                    ws.cell(insert_at, 8).value = (
                        f"=(B{insert_at}+C{insert_at}-E{insert_at})*G{insert_at}/100"
                        f"+(D{insert_at}*G{insert_at}/200)"
                    )
                    ws.cell(insert_at, 9).value = f"=F{insert_at}-H{insert_at}"

                    # FIX (2026-07-16 v2): After insert_rows openpyxl does NOT update
                    # formula text. We must expand SUM ranges and shift same-sheet refs.
                    #
                    # CRITICAL RULE: ONLY touch same-sheet formulas.
                    # Cross-sheet refs like ='Fixed Assets P. Yr.'!I35 point to the PY
                    # sheet's OWN fixed layout. Shifting them (I35→I36) is WRONG — the
                    # PY sheet was not modified by insert_rows on the CY sheet.
                    # Only bare same-sheet refs (=B9+C9, =F9-H9, =SUM(B8:B41)) need updating.
                    #
                    # Three formula types in the FA sheet, and what we do with each:
                    #
                    # Type 1 — =SUM(B8:B41)   → expand END by 1 if end >= insert_at
                    #           (total row: covers all assets; must include new row)
                    #           START is never moved (always row 8, before any insert)
                    #
                    # Type 2 — =B32+C32+D32-E32  (col F, same-row arithmetic)
                    #           =F32-H32            (col I)
                    #           These reference only the CURRENT row, which openpyxl
                    #           already shifts correctly when rows move. We leave them alone.
                    #           (Actually _repair_fa_formula handles these in Pass 2.)
                    #
                    # Type 3 — ='Fixed Assets P. Yr.'!I35  (col B cross-sheet)
                    #           NEVER touch — PY sheet layout is unchanged.
                    #
                    # Summary: only expand =SUM(...) end-of-range. Skip everything else.
                    import re as _fa_re
                    _sum_pat = _fa_re.compile(
                        r'=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', _fa_re.IGNORECASE
                    )
                    for _fr in range(1, 120):
                        for _fc in range(1, 15):
                            _fv = ws.cell(_fr, _fc).value
                            if not (isinstance(_fv, str) and _fv.startswith('=')):
                                continue
                            # Skip cross-sheet formulas completely — never shift them.
                            if '!' in _fv:
                                continue
                            _m = _sum_pat.match(_fv)
                            if not _m:
                                continue  # not a SUM — leave all other same-sheet formulas
                                          # to _repair_fa_formula (Pass 2) which handles
                                          # per-row arithmetic refs correctly
                            # Expand SUM range end if it reaches at or past insert_at.
                            # Start is left alone (always row 8 = before any insertion).
                            c1, r1, c2, r2 = (
                                _m.group(1), int(_m.group(2)),
                                _m.group(3), int(_m.group(4))
                            )
                            if r2 >= insert_at:
                                _new_fv = f'=SUM({c1}{r1}:{c2}{r2 + 1})'
                                ws.cell(_fr, _fc).value = _new_fv
                                log.append(
                                    f"✓ {fa_sheet}!{chr(64+_fc)}{_fr}: SUM expanded "
                                    f"{_fv} → {_new_fv} (new asset row at {insert_at})"
                                )

                    log.append(
                        f"✓ {fa_sheet}: inserted new asset '{name}' at row {insert_at} "
                        f"(section: {sec['name']}, rate: {rate}%) — "
                        f">180d={gt:,.2f}, <180d={lt:,.2f}"
                    )

            # ── Existing rows: write additions/sale ────────────────────────
            # FIX (FA Formula Shift Bug):
            # openpyxl reads shared formulas (Excel's 'si' attribute) and
            # stores them as expanded per-cell formulas.  But when a PREVIOUS
            # tb-process run wrote plain numeric values into F/H/I of a row
            # (e.g. Machine row 12: F12=5200000, H12=780000, I12=4420000),
            # those cells lost their formula text.  Because Excel stored the
            # NEXT row's formulas as shared-formula references to the base cell
            # (row 12), openpyxl now reads them as =B12+C12+D12-E12 instead of
            # the correct =B13+C13+D13-E13 — one row off for every row below.
            #
            # Fix strategy (two passes):
            #   Pass 1: For each entry row, ALWAYS re-write F, H, I with
            #           correct per-row formulas (even if they look like plain
            #           values — the plain value IS the corruption).
            #   Pass 2: After all entries are written, scan all FA data rows
            #           and repair any F/H/I formula that still references a
            #           row number other than its own (the shifted-formula
            #           symptom: row 13 having =B12+... instead of =B13+...).
            #
            # This is safe because:
            #   - The formulas are always =B{r}+C{r}+D{r}-E{r} etc. — we know
            #     the canonical form.
            #   - We only touch F, H, I — never B, C, D, E, G (user data cols).
            #   - If a cell already has the correct formula we leave it alone.

            def _fa_formula_for_row(r):
                """Return the canonical (F, H, I) formulas for FA row r."""
                f_col = f"=B{r}+C{r}+D{r}-E{r}"
                # Rate column is G (7); depreciation formula uses G{r}
                h_col = f"=(B{r}+C{r}-E{r})*G{r}/100+(D{r}*G{r}/200)"
                i_col = f"=F{r}-H{r}"
                return f_col, h_col, i_col

            def _repair_fa_formula(ws_fa, r):
                """
                Repair F, H, I for row r if they are:
                  (a) a plain numeric value (formula was previously over-written), OR
                  (b) a formula referencing a row number other than r (shared-formula shift).
                Returns True if a repair was made.
                """
                import re as _re
                repaired = False
                for col_idx, get_formula in (
                    (6,  lambda r: f"=B{r}+C{r}+D{r}-E{r}"),
                    (8,  lambda r: f"=(B{r}+C{r}-E{r})*G{r}/100+(D{r}*G{r}/200)"),
                    (9,  lambda r: f"=F{r}-H{r}"),
                ):
                    cell = ws_fa.cell(r, col_idx)
                    v = cell.value
                    correct = get_formula(r)
                    needs_repair = False
                    if v is None:
                        pass  # blank — leave it (Land has no H/I formula)
                    elif isinstance(v, (int, float)):
                        # Plain value where a formula should be → repair
                        needs_repair = True
                    elif isinstance(v, str) and v.startswith("="):
                        # Formula that references wrong row → repair
                        # Extract first row number referenced in the formula
                        nums = _re.findall(r'\d+', v)
                        if nums and int(nums[0]) != r:
                            needs_repair = True
                    if needs_repair:
                        cell.value = correct
                        repaired = True
                return repaired

            for entry in existing_entries:
                row = entry.get("row")
                if not row:
                    continue
                fields = [
                    ("additions_gt180", 3, "Addition >180d"),
                    ("additions_lt180", 4, "Addition <180d"),
                    ("sale", 5, "Sale"),
                ]
                wrote_any = False
                for field_key, default_col, label in fields:
                    val = entry.get(field_key, 0)
                    col_idx = entry.get(f"{field_key}_col", default_col)
                    if val:
                        if _safe_write(ws, row, col_idx, val):
                            log.append(f"✓ {fa_sheet}!{chr(64+col_idx)}{row} ({label}) = {float(val):,.2f}")
                            wrote_any = True

                # Pass 1: Always repair F/H/I for the row we just wrote into
                if wrote_any:
                    if _repair_fa_formula(ws, row):
                        log.append(f"✓ {fa_sheet} row {row}: repaired F/H/I formulas (shared-formula fix)")

            # Pass 2: Scan ALL data rows in the FA sheet and repair any
            # shifted shared-formula references — catches rows BELOW the
            # injection target that reference the wrong row number.
            # FIX: previously used a_s.isupper() to skip section headers, but
            # asset names like "E.P.B.X", "COMPUTER", "LAND" are also all-caps
            # and their isupper() returns True — causing them to be skipped.
            # Better heuristic: a SECTION HEADER row has no formula or numeric
            # value in col B (opening WDV), while an asset data row always has
            # either a number or a cross-sheet formula there.
            max_fa_row = 60
            for r_scan in range(8, max_fa_row + 1):
                a_val = ws.cell(r_scan, 1).value
                b_val = ws.cell(r_scan, 2).value
                # Skip completely blank rows
                if not a_val:
                    continue
                if not isinstance(a_val, str):
                    continue
                a_s = a_val.strip()
                if not a_s:
                    continue
                # Skip total/subtotal rows
                if a_s.lower() in ("total", "grand total"):
                    continue
                # FIX: distinguish section headers from asset rows using col B:
                # - Section header: B is None/empty and row has no F formula
                # - Asset row: B has a number, a formula, or 0
                b_is_empty = (b_val is None or str(b_val).strip() == "")
                f_val = ws.cell(r_scan, 6).value
                f_is_empty = (f_val is None)
                if b_is_empty and f_is_empty:
                    # This is a section header (PLANT & MACHINERY, VEHICLE etc.) — skip
                    continue
                if _repair_fa_formula(ws, r_scan):
                    log.append(f"✓ {fa_sheet} row {r_scan} ('{a_s}'): repaired shifted FA formulas")

    wb.save(output_path)
    wb.close()

@app.route("/tb-process", methods=["POST"])
def tb_process():
    if "uid" not in session:
        return jsonify({"status": "error", "message": "Not logged in"}), 401
    user = get_user_by_id(session["uid"])
    if not user:
        return jsonify({"status": "error", "message": "Not logged in"}), 401
    
    if not TB_PROCESSOR_AVAILABLE:
        return jsonify({"status": "error", "message": "TB processor not available"}), 500
    
    try:
        import tempfile, os, json, shutil

        tb_file = request.files.get("tb_file")
        bs_file = request.files.get("bs_file")

        if not tb_file or not bs_file:
            return jsonify({"status": "error", "message": "Both Trial Balance and BS template files are required"})

        # 🔍 Safely parse user mappings from frontend
        raw_mappings = request.form.get("user_mappings", "{}")
        try:
            user_mapping = json.loads(raw_mappings)
            # Clean keys & values: strip spaces, ignore empty/auto
            user_mapping = {
                str(k).strip(): str(v).strip() 
                for k, v in user_mapping.items() 
                if v and str(v).strip().lower() not in ("", "auto", "none", "ignore")
            }
        except Exception:
            user_mapping = {}

        # Read Capital & FA user entries (if any)
        raw_cap = request.form.get("capital_entries", "[]")
        raw_fa  = request.form.get("fa_entries", "[]")
        try:
            cap_entries = json.loads(raw_cap)
        except Exception:
            cap_entries = []
        try:
            fa_entries = json.loads(raw_fa)
        except Exception:
            fa_entries = []

        client_name = request.form.get("client_name", "Balance_Sheet").strip()
        cy_year = request.form.get("cy_year", "2025").strip()

        tmp = tempfile.mkdtemp()
        tb_orig = tb_file.filename or "tb.xlsx"
        tb_ext = ".pdf" if tb_orig.lower().endswith(".pdf") else ".xlsx"
        tb_path  = os.path.join(tmp, "tb" + tb_ext)
        bs_path  = os.path.join(tmp, "bs_template.xlsx")
        out_path = os.path.join(tmp, "bs_output.xlsx")

        tb_file.save(tb_path)
        bs_file.save(bs_path)

        # Process using the updated tb_processor
        result = process_tb_to_bs(
            tb_path, bs_path, out_path,
            user_mapping=user_mapping,
        )

        if result.get("status") == "error":
            try: shutil.rmtree(tmp, ignore_errors=True)
            except: pass
            return jsonify(result)

        # ── NOTE: FA rollover intentionally NOT run here ────────────────
        # _rollover_fixed_assets() is designed for the YEAR-SHIFT tool,
        # where CY data becomes PY data for a new fiscal year. The TB→BS
        # tool is a different workflow: it fills CY figures into an
        # EXISTING template whose "Fixed Assets P. Yr." sheet already
        # holds last year's correct closing data (e.g. Equipment/Car/
        # Motor Cycle WDV figures). Running the rollover here treated
        # that sheet as if it needed to be "shifted", overwriting its
        # rows/headers with the CY sheet's layout and wiping the real
        # data — causing the Fixed Assets note (and PPE on the BS) to
        # show 0 in the generated output.
        #
        # The BS template's FA C.Yr / FA P.Yr sheets are preserved
        # as-is by process_tb_to_bs() above; no further FA processing
        # is needed for this tool.

        # ── Inject Capital & FA user entries ──────────────────────────────
        if cap_entries or fa_entries:
            _inject_cap_fa(out_path, cap_entries, fa_entries, result.get("log", []))

        h = os.urandom(16).hex()
        dest = os.path.join(OUTPUT_DIR, h + "_out.xlsx")
        shutil.move(out_path, dest)

        try: shutil.rmtree(tmp, ignore_errors=True)
        except: pass

        safe_name = "".join(c if c.isalnum() or c in "-_ " else "_" for c in client_name)
        fname = f"{safe_name}_BS_{cy_year}.xlsx"
        log_usage(user["id"], fname)

        # Extract aggregated P&L figures for UI success page
        aggregated_vals = result.get("aggregated", {}) or {}
        revenue_val        = result.get("revenue", aggregated_vals.get("revenue", 0))
        other_income_val   = result.get("other_income", aggregated_vals.get("other_income", 0))
        direct_expenses_val= aggregated_vals.get("direct_expenses", 0)
        opening_stock_val  = result.get("opening_stock", aggregated_vals.get("opening_stock", 0))
        closing_stock_val  = result.get("closing_stock", aggregated_vals.get("inventories", 0))
        purchases_val      = aggregated_vals.get("purchases", 0)
        employee_exp_val   = aggregated_vals.get("employee_expenses", 0)
        other_exp_val      = aggregated_vals.get("other_expenses", 0)
        depreciation_val   = aggregated_vals.get("depreciation", 0)
        finance_cost_val   = aggregated_vals.get("finance_cost", 0)
        tax_expense_val    = aggregated_vals.get("tax_expense", 0)

        total_assets_val   = result.get("total_assets", 0)
        total_liab_val     = result.get("total_liabilities", 0)
        net_profit_val     = result.get("net_profit", 0)
        diff_val           = abs(float(total_assets_val) - float(total_liab_val))

        return jsonify({
            "status":    "success",
            "log":      result.get("log", []),
            "file_id":  h,
            "filename": fname,
            "aggregated":      aggregated_vals,
            "revenue":         revenue_val,
            "other_income":    other_income_val,
            "direct_expenses": direct_expenses_val,
            "opening_stock":   opening_stock_val,
            "closing_stock":   closing_stock_val,
            "purchases":       purchases_val,
            "employee_expenses": employee_exp_val,
            "other_expenses":  other_exp_val,
            "depreciation":    depreciation_val,
            "finance_cost":    finance_cost_val,
            "tax_expense":     tax_expense_val,
            "tally": {
                "balanced":       bool(result.get("tally_ok", False)),
                "total_assets":   total_assets_val,
                "total_liabilities": total_liab_val,
                "difference":     diff_val,
                "profit":         net_profit_val,
                "user_mappings_applied": len(user_mapping or {}),
            },
        })

    except Exception as e:
        import traceback
        return jsonify({
            "status":   "error",
            "message": f"Processing failed: {e}\n{traceback.format_exc()}"
        }), 500


# ── TB→BS Page Template ───────────────────────────────────────────────────────
TB_BS_TEMPLATE = """<!DOCTYPE html><html lang="en"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Balance Sheet from Trial Balance – CA Toolkit</title>
<link rel=\"stylesheet\" href=\"https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap\"/>
<style>
""" + BASE_CSS + """
nav{background:#fff;border-bottom:1px solid var(--border);padding:0 24px;height:56px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:100;box-shadow:0 1px 4px rgba(0,0,0,.06)}
.hero{background:linear-gradient(135deg,#0F172A,#1E3A5F);color:#fff;padding:40px 24px 32px;text-align:center}
.hero h1{font-size:clamp(22px,4vw,32px);font-weight:800;margin-bottom:8px}
.hero p{color:#94A3B8;font-size:14px;max-width:600px;margin:0 auto}
.hero-badge{display:inline-flex;align-items:center;gap:6px;background:rgba(255,255,255,.1);border:1px solid rgba(255,255,255,.15);border-radius:20px;padding:4px 14px;font-size:11px;font-weight:600;color:#CBD5E1;margin-bottom:14px}
.page{max-width:1000px;margin:0 auto;padding:24px 16px 60px}
.card{background:#fff;border:1px solid var(--border);border-radius:14px;box-shadow:0 2px 8px rgba(0,0,0,.05);margin-bottom:20px;overflow:hidden}
.card-head{display:flex;align-items:center;gap:14px;padding:16px 20px;border-bottom:1px solid var(--border);background:#FAFAFA}
.card-head .icon{width:36px;height:36px;border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:18px;flex-shrink:0}
.card-head h2{font-size:15px;font-weight:700;margin:0}
.card-head p{font-size:12px;color:var(--muted);margin:2px 0 0}
.card-body{padding:20px}
.row2{display:grid;grid-template-columns:1fr 1fr;gap:16px}
.field label{display:block;font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;color:var(--muted);margin-bottom:6px}
.upload-zone{border:2px dashed var(--border);border-radius:10px;padding:24px 20px;text-align:center;cursor:pointer;transition:all .2s;background:#FAFAFA;position:relative;min-height:90px}
.upload-zone:hover,.upload-zone.drag{border-color:var(--brand);background:#EFF6FF}
.upload-zone input[type=file]{position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%}
.upload-zone .uzicon{font-size:26px;margin-bottom:6px}
.upload-zone .uztitle{font-size:13px;font-weight:600;color:var(--ink)}
.upload-zone .uzsub{font-size:11px;color:var(--muted);margin-top:3px}
.uz-done{display:none;margin-top:8px;padding:6px 12px;background:#ECFDF5;border-radius:6px;font-size:11px;font-weight:700;color:#065F46}
select,input[type=text]{width:100%;border:1.5px solid var(--border);border-radius:8px;padding:8px 10px;font-family:inherit;font-size:13px;box-sizing:border-box}
select:focus,input:focus{outline:none;border-color:var(--brand)}
.btn-main{width:100%;padding:14px;background:var(--brand);color:#fff;border:none;border-radius:10px;font-size:14px;font-weight:700;cursor:pointer;font-family:inherit;transition:all .2s}
.btn-main:hover{background:#1D4ED8}
.btn-main:disabled{background:#93C5FD;cursor:not-allowed}
.btn-sec{padding:10px 20px;background:#F3F4F6;color:var(--ink);border:1.5px solid var(--border);border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;font-family:inherit}
.btn-sec:hover{background:#E5E7EB}

/* Steps */
.steps{display:flex;margin-bottom:20px;border-radius:10px;overflow:hidden;border:1px solid var(--border)}
.step-item{flex:1;padding:10px 8px;text-align:center;font-size:11px;font-weight:600;color:var(--muted);background:#F9FAFB;border-right:1px solid var(--border);transition:all .2s}
.step-item:last-child{border-right:none}
.step-item.active{background:var(--brand);color:#fff}
.step-item.done{background:#ECFDF5;color:#065F46}
.step-num{display:block;font-size:15px;margin-bottom:2px}

/* Mapping table */
.map-table{width:100%;border-collapse:collapse;font-size:12px}
.map-table th{padding:8px 10px;border-bottom:2px solid var(--border);font-size:10px;text-transform:uppercase;letter-spacing:.05em;color:var(--muted);background:#F9FAFB;text-align:left}
.map-table td{padding:7px 10px;border-bottom:1px solid var(--border);vertical-align:middle}
.map-table tr:hover td{background:#F9FAFB}
.acc-name{font-weight:600;color:var(--ink);font-size:12px}
.acc-grp{font-size:10px;color:var(--muted)}
.amt{font-weight:700;text-align:right;white-space:nowrap;font-size:12px}
.amt.cr{color:var(--green)}
.amt.dr{color:#2563EB}
.conf-pill{display:inline-flex;align-items:center;padding:2px 8px;border-radius:10px;font-size:10px;font-weight:700;white-space:nowrap}
.conf-high{background:#ECFDF5;color:#065F46}
.conf-med{background:#FFFBEB;color:#92400E}
.conf-low{background:#FEF2F2;color:#B91C1C}
.conf-user{background:#EFF6FF;color:#1E40AF}
.map-sel{width:100%;border:1.5px solid var(--border);border-radius:6px;padding:5px 7px;font-size:11px;font-family:inherit;background:#fff;cursor:pointer}
.map-sel:focus{border-color:var(--brand);outline:none}
.map-sel.changed{border-color:#F59E0B;background:#FFFBEB;font-weight:700}
.map-sel.user{border-color:var(--brand);background:#EFF6FF;font-weight:700}

/* Summary */
.sum-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:10px;margin-bottom:16px}
.sum-card{padding:10px;background:#F9FAFB;border:1px solid var(--border);border-radius:10px;text-align:center}
.sum-val{font-size:18px;font-weight:800;color:var(--brand)}
.sum-lbl{font-size:10px;color:var(--muted);font-weight:600;text-transform:uppercase;margin-top:2px}

/* Spinner */
.spinner{width:44px;height:44px;border:4px solid #E5E7EB;border-top-color:var(--brand);border-radius:50%;animation:spin .8s linear infinite;margin:0 auto 14px}
@keyframes spin{to{transform:rotate(360deg)}}

/* Result */
.result-ok{padding:16px;background:#F0FDF4;border:1.5px solid #BBF7D0;border-radius:10px;text-align:center}
.result-err{padding:16px;background:#FEF2F2;border:1.5px solid #FECACA;border-radius:10px;text-align:center}
.trow{display:flex;justify-content:space-between;padding:7px 0;border-bottom:1px solid rgba(0,0,0,.05);font-size:13px}
.trow:last-child{border:none}
.tlbl{color:var(--muted);font-weight:500}
.tval{font-weight:700}
.note-box{font-size:11px;color:var(--muted);line-height:1.7;padding:10px 14px;background:#FFFBEB;border:1px solid #FDE68A;border-radius:8px;margin-top:12px}
footer{background:#0f1b2d;color:#9CA3AF;font-size:12px;padding:0}
.ft-main{display:grid;grid-template-columns:2fr 1fr 1.4fr;gap:40px;padding:40px 48px;max-width:1200px;margin:0 auto}
.ft-brand-name{color:#fff;font-size:18px;font-weight:800;margin-bottom:12px}
.ft-brand-desc{font-size:12.5px;line-height:1.75;color:#9CA3AF;max-width:340px;text-align:justify}
.ft-col-title{color:#fff;font-size:14px;font-weight:700;margin-bottom:14px}
.ft-links{list-style:none;padding:0;margin:0}
.ft-links li{margin-bottom:8px}
.ft-links a{color:#9CA3AF;text-decoration:none;font-size:13px;transition:color .2s}
.ft-links a:hover{color:#fff}
.ft-contact-name{color:#fff;font-weight:700;font-size:13px;margin-bottom:6px}
.ft-contact-addr{color:#9CA3AF;font-size:12px;line-height:1.7;margin-bottom:10px}
.ft-contact-line{color:#9CA3AF;font-size:12px;margin-bottom:4px}
.ft-socials{display:flex;gap:14px;margin-top:12px}
.ft-socials a{color:#9CA3AF;transition:color .2s}
.ft-socials a:hover{color:#fff}
.ft-socials svg{width:20px;height:20px;fill:currentColor}
.ft-bottom{background:#0a1422;border-top:1px solid #1e2d42;padding:12px 48px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px}
.ft-bottom-left{font-size:11px;color:#6B7280}
.ft-bottom-right{font-size:11px;color:#6B7280}
@media(max-width:768px){.ft-main{grid-template-columns:1fr;padding:28px 20px;gap:24px}.ft-bottom{padding:12px 20px;flex-direction:column;text-align:center}}
@media(max-width:600px){.row2{grid-template-columns:1fr}.steps{flex-direction:column}}
</style></head><body>
<nav>
  <a href="/" class="logo">CA<span class="logo-dot"></span><span>Toolkit</span></a>
  <div class="nav-right">
    <span class="nav-user">👤 <strong>{{ username }}</strong>
      <span class="badge b-{{ plan }}">{{ plan_label }}</span></span>
    <a href="/" class="nav-btn dash">⬅ Dashboard</a>
    <a href="/logout" class="nav-link">Sign out</a>
  </div>
</nav>
<section class="hero">
  <div class="hero-badge">📋 Premium · Zero Formatting Change</div>
  <h1>Balance Sheet from Trial Balance</h1>
  <p>Upload trial balance + BS template. Auto-maps accounts, lets you correct, then injects CY figures.</p>
</section>

<div class="page">
  <div class="steps">
    <div class="step-item active" id="s1"><span class="step-num">1</span>Upload</div>
    <div class="step-item" id="s2b"><span class="step-num">2</span>Capital &amp; FA</div>
    <div class="step-item" id="s2"><span class="step-num">3</span>Review Mapping</div>
    <div class="step-item" id="s3"><span class="step-num">4</span>Download</div>
  </div>

  <!-- STEP 1 -->
  <div id="step1">
    <div class="card">
      <div class="card-head"><div class="icon" style="background:#EFF6FF">📤</div>
        <div><h2>Upload Files</h2><p>Trial Balance (.xlsx or .pdf) and Balance Sheet template (.xlsx)</p></div></div>
      <div class="card-body">
        <div class="row2">
          <div class="field">
            <label>Trial Balance</label>
            <div class="upload-zone" id="tbZone">
              <input type="file" id="tbFile" accept=".xlsx,.xls,.pdf" onchange="onFile(this,'tb')"/>
              <div class="uzicon">📊</div>
              <div class="uztitle">Click or drag Trial Balance</div>
              <div class="uzsub">Tally / Busy / Manual — .xlsx or .pdf</div>
            </div>
            <div class="uz-done" id="tbDone"></div>
          </div>
          <div class="field">
            <label>Balance Sheet Template</label>
            <div class="upload-zone" id="bsZone">
              <input type="file" id="bsFile" accept=".xlsx" onchange="onFile(this,'bs')"/>
              <div class="uzicon">📋</div>
              <div class="uztitle">Click or drag BS Template</div>
              <div class="uzsub">PY filled · CY blank · formatting intact</div>
            </div>
            <div class="uz-done" id="bsDone"></div>
          </div>
        </div>
        <div class="row2" style="margin-top:14px">
          <div class="field">
            <label>Financial Year (CY)</label>
            <select id="cyYear">
              <option value="2025">2024-25 (31 March 2025)</option>
              <option value="2026">2025-26 (31 March 2026)</option>
            </select>
          </div>
          <div class="field">
            <label>Client / Firm Name</label>
            <input type="text" id="clientName" placeholder="XYZ Enterprises..."/>
          </div>
        </div>
        <div style="margin-top:18px">
          <button class="btn-main" id="analyseBtn" onclick="doAnalyse()" disabled>
            🔍 Analyse Trial Balance
          </button>
        </div>
      </div>
    </div>
  </div>

  <!-- STEP 2 (now Capital & FA — shown first after upload) -->
  <div id="step2b" style="display:none">
    <div class="card">
      <div class="card-head"><div class="icon" style="background:#FEF3C7">📋</div>
        <div><h2>Capital Account &amp; Fixed Assets</h2>
        <p>Enter additions, withdrawals (capital) and additions, sales (fixed assets) from ledger.</p></div></div>
      <div class="card-body">

        <div style="background:#FEF3C7;border:1px solid #FDE68A;border-radius:10px;padding:12px 16px;font-size:12px;color:#92400E;margin-bottom:16px;line-height:1.7">
          <strong>Why this step?</strong> The Trial Balance only has closing balances. Capital A/c needs opening + additions + withdrawals from the ledger. Same for Fixed Assets — additions &amp; sales come from ledger, not TB.
        </div>

        <div style="margin-bottom:24px">
          <h3 style="font-size:14px;font-weight:700;margin-bottom:10px">👤 Owner's Capital Account</h3>
          <div id="capTableWrap" style="overflow-x:auto">
            <p style="color:var(--muted);font-size:12px">Loading from BS template...</p>
          </div>
        </div>

        <div style="margin-bottom:20px">
          <h3 style="font-size:14px;font-weight:700;margin-bottom:10px;display:flex;align-items:center;gap:8px">
            🏭 Fixed Assets Chart
            <button onclick="openAddAssetModal()" style="margin-left:auto;background:var(--brand);color:#fff;border:none;border-radius:6px;padding:5px 14px;font-size:11px;cursor:pointer;font-weight:600">+ Add Asset</button>
          </h3>
          <div id="faTableWrap" style="overflow-x:auto">
            <p style="color:var(--muted);font-size:12px">Loading from BS template...</p>
          </div>
        </div>

        <!-- Add Asset Modal -->
        <div id="addAssetModal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.45);z-index:2000;align-items:center;justify-content:center;padding:16px">
          <div style="background:#fff;border-radius:12px;max-width:420px;width:100%;box-shadow:0 24px 60px rgba(0,0,0,.2);padding:24px">
            <h3 style="font-size:15px;font-weight:800;margin-bottom:16px">➕ Add New Asset</h3>
            <div style="margin-bottom:12px">
              <label style="display:block;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.04em;color:var(--muted);margin-bottom:4px">Asset Name</label>
              <input type="text" id="newAssetName" placeholder="e.g. Generator, Laptop, Shed" style="width:100%;border:1.5px solid var(--border);border-radius:8px;padding:9px 12px;font-family:inherit;font-size:13px;outline:none"/>
            </div>
            <div style="margin-bottom:12px">
              <label style="display:block;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.04em;color:var(--muted);margin-bottom:4px">Asset Section</label>
              <select id="newAssetSection" onchange="onAssetSectionChange()" style="width:100%;border:1.5px solid var(--border);border-radius:8px;padding:9px 12px;font-family:inherit;font-size:13px;outline:none;background:#fff">
                <option value="plant_15">Plant &amp; Machinery (15%)</option>
                <option value="vehicle_15">Vehicle (15%)</option>
                <option value="computer_40">Computers &amp; Peripherals (40%)</option>
                <option value="furniture_10">Furniture &amp; Fixtures (10%)</option>
                <option value="building_10">Building (10%)</option>
                <option value="land_0">Land (0% — No depreciation)</option>
                <option value="other_custom">Other (specify rate)</option>
              </select>
            </div>
            <div id="customRateWrap" style="display:none;margin-bottom:12px">
              <label style="display:block;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.04em;color:var(--muted);margin-bottom:4px">Depreciation Rate %</label>
              <input type="number" id="newAssetRate" placeholder="e.g. 15" min="0" max="100" style="width:100%;border:1.5px solid var(--border);border-radius:8px;padding:9px 12px;font-family:inherit;font-size:13px;outline:none"/>
            </div>
            <div style="background:#EFF6FF;border:1px solid #BFDBFE;border-radius:8px;padding:10px 12px;font-size:11px;color:#1E40AF;margin-bottom:16px">
              ℹ️ <strong>Section determines where in the Fixed Assets chart this asset goes.</strong> Rate is pre-filled based on IT Act. You can customize for special cases.
            </div>
            <div style="display:flex;gap:10px">
              <button onclick="closeAddAssetModal()" style="flex:1;background:#F3F4F6;color:var(--ink);border:none;border-radius:8px;padding:10px;font-family:inherit;font-size:13px;font-weight:600;cursor:pointer">Cancel</button>
              <button onclick="confirmAddAsset()" style="flex:2;background:var(--brand);color:#fff;border:none;border-radius:8px;padding:10px;font-family:inherit;font-size:13px;font-weight:700;cursor:pointer">Add to Chart →</button>
            </div>
          </div>
        </div>

        <div style="display:flex;gap:12px">
          <button class="btn-sec" onclick="goStep(1)">← Back</button>
          <button class="btn-main" style="flex:1" onclick="goStep(2)">
            Next → Review Mapping
          </button>
        </div>
      </div>
    </div>
  </div>

  <!-- STEP 3 (Review Mapping — two-panel BS | P&L) -->
  <div id="step2" style="display:none">
    <div class="card">
      <div class="card-head"><div class="icon" style="background:#EFF6FF">🗂️</div>
        <div><h2>Review &amp; Confirm Account Mapping</h2>
          <p id="mapSub">Verify auto-detected heads — change any using the dropdown</p></div></div>
      <div class="card-body">
        <div id="tbFormatBox" style="padding:10px 14px;background:#EFF6FF;border:1px solid #BFDBFE;border-radius:8px;font-size:12px;color:#1E40AF;margin-bottom:14px"></div>
        <div class="sum-grid" id="sumGrid"></div>
        <div id="preChecks"></div>

        <div style="margin-bottom:10px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px">
          <div style="font-size:12px;color:var(--muted)">
            🟢 Auto &nbsp;|&nbsp; 🟡 Review &nbsp;|&nbsp; 🔴 Manual &nbsp;|&nbsp; 🔵 Changed
          </div>
          <button class="btn-sec" onclick="expandAll()">Expand All Groups</button>
        </div>

        <!-- Tab-based BS | P&L layout -->
        <div style="display:flex;gap:0;margin-bottom:16px">
          <button id="tabBS" onclick="switchTab('bs')" style="flex:1;padding:12px;font-size:14px;font-weight:700;border:2px solid var(--brand);border-radius:10px 0 0 10px;cursor:pointer;background:var(--brand);color:#fff;transition:all .2s">📊 Balance Sheet</button>
          <button id="tabPL" onclick="switchTab('pl')" style="flex:1;padding:12px;font-size:14px;font-weight:700;border:2px solid var(--brand);border-left:none;border-radius:0 10px 10px 0;cursor:pointer;background:#fff;color:var(--brand);transition:all .2s">📈 Profit &amp; Loss</button>
        </div>
        <div id="bsPanel" style="max-height:65vh;overflow-y:auto"></div>
        <div id="plPanel" style="max-height:65vh;overflow-y:auto;display:none"></div>

        <div style="margin-top:20px;display:flex;gap:12px">
          <button class="btn-sec" onclick="goStep('2b')">← Back</button>
          <button class="btn-main" id="generateBtn" onclick="doGenerate()" style="flex:1">
            ✅ Generate Balance Sheet
          </button>
        </div>
      </div>
    </div>
  </div>

  <!-- PROCESSING -->
  <div id="loadWrap" style="display:none">
    <div class="card"><div class="card-body" style="text-align:center;padding:48px">
      <div class="spinner"></div>
      <div style="font-size:15px;font-weight:700">Generating Balance Sheet...</div>
      <div style="font-size:12px;color:var(--muted);margin-top:6px" id="loadMsg">Applying your mapping and injecting figures...</div>
    </div></div>
  </div>

  <!-- STEP 3 -->
  <div id="step3" style="display:none">
    <div class="card">
      <div class="card-head"><div class="icon" style="background:#ECFDF5">✅</div>
        <div><h2>Balance Sheet Ready</h2><p id="resSub"></p></div></div>
      <div class="card-body">
        <div id="resBox"></div>
        <div style="margin-top:18px;display:flex;gap:12px">
          <button class="btn-sec" onclick="goStep(2)">← Back to Mapping</button>
          <a id="dlBtn" class="btn-main" style="flex:1;text-decoration:none;display:flex;align-items:center;justify-content:center;gap:8px" href="#">
            📥 Download Balance Sheet
          </a>
        </div>
        <button class="btn-sec" style="width:100%;margin-top:10px" onclick="location.reload()">🔄 New Client</button>
      </div>
    </div>
    <div class="note-box">⚠️ <strong>Always verify:</strong> Total Assets = Total Liabilities · All figures match TB · Profit matches capital account · Notes sheets populated correctly.</div>
  </div>
</div>

<footer>
  <div class="ft-main">
    <div>
      <div class="ft-brand-name">CA<span>Toolkit</span></div>
      <p class="ft-brand-desc">CA Toolkit is a comprehensive utility platform built by a CA Article from Ludhiana, Punjab, providing automation tools for Indian Chartered Accountants. The platform saves hours of manual work every year — from Balance Sheet year-shift to GST reconciliation, tax calculations, and more.</p>
    </div>
    <div>
      <div class="ft-col-title">Know More</div>
      <ul class="ft-links">
        <li><a href="/">Home</a></li>
        <li><a href="/">BS Year Shift</a></li>
        <li><a href="/tool/tb-to-bs">TB → Balance Sheet</a></li>
        <li><a href="/tool/tax-calculator">Tax Calculator</a></li>
        <li><a href="/privacy">Privacy Policy</a></li>
      </ul>
    </div>
    <div>
      <div class="ft-col-title">Contact Us</div>
      <div class="ft-contact-name">CA Toolkit</div>
      <div class="ft-contact-addr">Built for Indian Chartered Accountants<br/>Created by CA Article · Ludhiana, Punjab</div>
      <div class="ft-contact-line">Support · <a href="https://wa.me/918427651580" style="color:#9CA3AF">WhatsApp Chat</a></div>
      <div class="ft-socials">
        <a href="https://wa.me/918427651580" target="_blank" title="WhatsApp"><svg viewBox="0 0 24 24"><path d="M17.472 14.382c-.297-.149-1.758-.867-2.03-.967-.273-.099-.471-.148-.67.15-.197.297-.767.966-.94 1.164-.173.199-.347.223-.644.075-.297-.15-1.255-.463-2.39-1.475-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.298-.347.446-.52.149-.174.198-.298.298-.497.099-.198.05-.371-.025-.52-.075-.149-.669-1.612-.916-2.207-.242-.579-.487-.5-.669-.51-.173-.008-.371-.01-.57-.01-.198 0-.52.074-.792.372-.272.297-1.04 1.016-1.04 2.479 0 1.462 1.065 2.875 1.213 3.074.149.198 2.096 3.2 5.077 4.487.709.306 1.262.489 1.694.625.712.227 1.36.195 1.871.118.571-.085 1.758-.719 2.006-1.413.248-.694.248-1.289.173-1.413-.074-.124-.272-.198-.57-.347m-5.421 7.403h-.004a9.87 9.87 0 01-5.031-1.378l-.361-.214-3.741.982.998-3.648-.235-.374a9.86 9.86 0 01-1.51-5.26c.001-5.45 4.436-9.884 9.888-9.884 2.64 0 5.122 1.03 6.988 2.898a9.825 9.825 0 012.893 6.994c-.003 5.45-4.437 9.884-9.885 9.884m8.413-18.297A11.815 11.815 0 0012.05 0C5.495 0 .16 5.335.157 11.892c0 2.096.547 4.142 1.588 5.945L.057 24l6.305-1.654a11.882 11.882 0 005.683 1.448h.005c6.554 0 11.89-5.335 11.893-11.893a11.821 11.821 0 00-3.48-8.413z"/></svg></a>
      </div>
    </div>
  </div>
  <div class="ft-bottom">
    <span class="ft-bottom-left">©2026 CA Toolkit · All Rights Reserved · <a href="/privacy" style="color:#6B7280;text-decoration:none">Privacy Policy</a> · <span style="color:#EF4444">No refund after first upload is used</span></span>
    <span class="ft-bottom-right">Built for Indian CAs · Ludhiana, Punjab</span>
  </div>
</footer>

<script>
// ═══════════════════════════════════════
//  STATE
// ═══════════════════════════════════════
let tbFile = null, bsFile = null;
let analysisData = null;
// KEY STORAGE: maps account unique key → currently selected bs_head
// This is what gets sent to the server on Generate
let userMappings = {};

const BS_HEADS = [
  {v:"capital",       l:"Owner's Capital / Partners Capital"},
  {v:"lt_borrowings", l:"Long Term Borrowings"},
  {v:"st_borrowings", l:"Short Term Borrowings"},
  {v:"trade_payables",l:"Trade Payables (Creditors)"},
  {v:"advance_from_customer", l:"Advance from Customer (under Sundry Creditors)"},
  {v:"other_cl",      l:"Other Current Liabilities"},
  {v:"st_provisions", l:"Short Term Provisions"},
  {v:"fixed_assets",  l:"Fixed Assets / PPE"},
  {v:"investments",   l:"Non-Current Investments"},
  {v:"inventories",   l:"Closing Stock / Inventories"},
  {v:"trade_rec",     l:"Trade Receivables (Debtors)"},
  {v:"advance_to_supplier", l:"Advance to Supplier / Customer (under Sundry Debtors)"},
  {v:"cash_bank",     l:"Cash and Bank Balances"},
  {v:"stla",          l:"Short Term Loans & Advances"},
  {v:"other_ca",      l:"Other Current Assets"},
  {v:"revenue",       l:"Revenue from Operations"},
  {v:"other_income",  l:"Other Income"},
  {v:"opening_stock", l:"Opening Stock"},
  {v:"purchases",     l:"Purchases"},
  {v:"direct_expenses",l:"Direct Expenses"},
  {v:"employee_expenses",l:"Employee / Salary Expenses"},
  {v:"finance_cost",  l:"Finance Cost / Bank Interest"},
  {v:"depreciation",  l:"Depreciation"},
  {v:"other_expenses",l:"Other Expenses"},
  {v:"tax_expense",   l:"Tax Expense"},
  {v:"ignore",        l:"⊘ Ignore / Skip"},
];

const HEAD_LABEL = Object.fromEntries(BS_HEADS.map(h=>[h.v, h.l]));

// ═══════════════════════════════════════
//  STEPS
// ═══════════════════════════════════════
function goStep(n) {
  document.getElementById('step1').style.display = n===1?'block':'none';
  document.getElementById('step2b').style.display = n==='2b'?'block':'none';
  document.getElementById('step2').style.display = n===2?'block':'none';
  document.getElementById('step3').style.display = n===3?'block':'none';
  document.getElementById('loadWrap').style.display = 'none';
  // Step order: s1=1, s2b=2, s2=3, s3=4
  const order = {s1:1, s2b:2, s2:3, s3:4};
  const curVal = n===1?1 : n==='2b'?2 : n===2?3 : n===3?4 : 0;
  ['s1','s2b','s2','s3'].forEach(id=>{
    const s = document.getElementById(id);
    s.className = 'step-item' + (order[id]===curVal?' active':(order[id]<curVal?' done':''));
  });
  window.scrollTo({top:0,behavior:'smooth'});
}

// ═══════════════════════════════════════
//  FILE UPLOAD + DRAG & DROP
// ═══════════════════════════════════════
function onFile(inp, type) {
  const f = inp.files[0]; if(!f) return;
  _setFile(f, type);
}

function _setFile(f, type) {
  if (type==='tb') {
    tbFile = f;
    // Clear stale session mappings whenever a new TB file is loaded
    try { sessionStorage.removeItem('tb_mappings'); sessionStorage.removeItem('tb_mappings_fp'); } catch(e) {}
    document.getElementById('tbDone').style.display='block';
    document.getElementById('tbDone').textContent='✓ '+f.name;
    document.getElementById('tbZone').style.borderColor='var(--green)';
  } else {
    bsFile = f;
    document.getElementById('bsDone').style.display='block';
    document.getElementById('bsDone').textContent='✓ '+f.name;
    document.getElementById('bsZone').style.borderColor='var(--green)';
  }
  document.getElementById('analyseBtn').disabled = !(tbFile && bsFile);
}

// Drag-and-drop for both upload zones
['tbZone','bsZone'].forEach(id => {
  const zone = document.getElementById(id);
  if (!zone) return;
  const type = id === 'tbZone' ? 'tb' : 'bs';
  zone.addEventListener('dragover', e => { e.preventDefault(); zone.classList.add('drag'); });
  zone.addEventListener('dragleave', () => zone.classList.remove('drag'));
  zone.addEventListener('drop', e => {
    e.preventDefault(); zone.classList.remove('drag');
    const f = e.dataTransfer.files[0];
    if (f && (f.name.endsWith('.xlsx') || f.name.endsWith('.xls') || (type==='tb' && f.name.endsWith('.pdf')))) {
      _setFile(f, type);
      // Update the hidden input too
      const dt = new DataTransfer(); dt.items.add(f);
      zone.querySelector('input[type=file]').files = dt.files;
    }
  });
});

// ═══════════════════════════════════════
//  ANALYSE
// ═══════════════════════════════════════
async function doAnalyse() {
  const btn = document.getElementById('analyseBtn');
  btn.disabled = true; btn.textContent = '⏳ Analysing...';
  document.getElementById('step1').style.display='none';
  document.getElementById('loadWrap').style.display='block';
  document.getElementById('loadMsg').textContent = 'Reading trial balance and auto-classifying accounts...';

  const fd = new FormData();
  fd.append('tb_file', tbFile);

  try {
    const res = await fetch('/tb-analyse', {method:'POST', body:fd, credentials:'include'});
    const data = await res.json();

    if (data.status !== 'success') {
      document.getElementById('loadWrap').style.display='none';
      document.getElementById('step1').style.display='block';
      btn.disabled=false; btn.textContent='🔍 Analyse Trial Balance';
      if (res.status === 401) {
        // Session expired — redirect to login
        if (confirm('Session expired. Click OK to go to login page.')) {
          window.location.href = '/login?next=' + encodeURIComponent(window.location.pathname);
        }
      } else {
        alert('Error: ' + data.message);
      }
      return;
    }

    analysisData = data;
    userMappings = {};
    (data.accounts || []).forEach(a => {
      userMappings[a.key] = a.bs_head || 'ignore';
    });

    // Restore saved mappings from sessionStorage ONLY if they belong to
    // the same TB file (matched by a fingerprint of account keys).
    // Without this guard, stale mappings from a previous session (e.g.
    // all accounts mapped to 'other_cl' by an old buggy version) get
    // restored over the correct new auto-classifications.
    try {
      const saved = JSON.parse(sessionStorage.getItem('tb_mappings') || '{}');
      const savedFp = sessionStorage.getItem('tb_mappings_fp') || '';
      const currentFp = Object.keys(userMappings).sort().slice(0,5).join('|');
      if (savedFp === currentFp) {
        Object.keys(saved).forEach(k => {
          if (saved[k] && userMappings.hasOwnProperty(k)) userMappings[k] = saved[k];
        });
      } else {
        // Different TB file — discard stale mappings
        sessionStorage.removeItem('tb_mappings');
        sessionStorage.removeItem('tb_mappings_fp');
      }
    } catch(e) {}

    buildMappingUI(data);
    // Go to Capital & FA step first, then user proceeds to mapping
    goToCapFA();

  } catch(e) {
    alert('Network error: '+e);
    document.getElementById('step1').style.display='block';
    document.getElementById('loadWrap').style.display='none';
    btn.disabled=false; btn.textContent='🔍 Analyse Trial Balance';
  }
}

// ═══════════════════════════════════════
//  BUILD MAPPING UI
// ═══════════════════════════════════════
function buildMappingUI(data) {
  const accts = data.accounts || [];
  const fi = data.detection || {};
  const colLetter = (idx) => idx != null && idx >= 0 ? String.fromCharCode(65 + idx) : null;
  const drL = colLetter(fi.debit_col);
  const crL = colLetter(fi.credit_col);
  const netL = colLetter(fi.net_col);
  let colInfo = '';
  if (drL && crL) { colInfo = `Dr: <strong>${drL}</strong> | Cr: <strong>${crL}</strong>`; }
  else if (netL) { colInfo = `Amount: <strong>${netL}</strong>`; }
  else { colInfo = `Amounts: <strong>auto</strong>`; }

  document.getElementById('tbFormatBox').innerHTML =
    `<strong>📊 Detected:</strong> ${fi.format_type ? 'Format '+fi.format_type : 'Auto'} &nbsp;|&nbsp; ` +
    `Name col: <strong>${colLetter(fi.account_col)||'A'}</strong> &nbsp;|&nbsp; ` +
    colInfo + ` &nbsp;|&nbsp; ` +
    `<strong>${accts.length}</strong> accounts`;

  const hi = accts.filter(a=>a.confidence==='high').length;
  const me = accts.filter(a=>a.confidence==='med').length;
  const lo = accts.filter(a=>a.confidence==='low').length;
  document.getElementById('sumGrid').innerHTML = `
    <div class="sum-card"><div class="sum-val">${accts.length}</div><div class="sum-lbl">Total</div></div>
    <div class="sum-card"><div class="sum-val" style="color:var(--green)">${hi}</div><div class="sum-lbl">Auto ✅</div></div>
    <div class="sum-card"><div class="sum-val" style="color:#F59E0B">${me}</div><div class="sum-lbl">Review ⚠️</div></div>
    <div class="sum-card"><div class="sum-val" style="color:#EF4444">${lo}</div><div class="sum-lbl">Manual ❌</div></div>`;

  const checks = data.pre_checks || [];
  document.getElementById('preChecks').innerHTML = checks.map(c=>
    `<div style="padding:6px 12px;border-radius:6px;font-size:12px;margin-bottom:6px;
      background:${c.ok?'#F0FDF4':'#FFFBEB'};color:${c.ok?'#065F46':'#92400E'};
      border:1px solid ${c.ok?'#BBF7D0':'#FDE68A'}">${c.ok?'✅':'⚠️'} ${c.message}</div>`
  ).join('');

  document.getElementById('mapSub').textContent =
    `${accts.length} accounts · ${hi} auto-mapped · ${me+lo} need review`;

  rebuildPanels();
}

// BS heads go in left panel, P&L heads in right panel
const BS_HEAD_KEYS = ['capital','lt_borrowings','st_borrowings',
  'trade_payables','advance_from_customer','other_cl',
  'st_provisions','fixed_assets','investments','inventories',
  'trade_rec','advance_to_supplier','cash_bank','stla','other_ca'];
const PL_HEAD_KEYS = ['revenue','other_income','opening_stock','purchases','direct_expenses',
  'employee_expenses','finance_cost','depreciation','other_expenses','tax_expense'];

function rebuildPanels() {
  const accts = analysisData?.accounts || [];
  // Group by CURRENT userMappings value
  const groups = {};
  accts.forEach(a => {
    const h = userMappings[a.key] || a.bs_head || 'ignore';
    if (!groups[h]) groups[h] = [];
    groups[h].push(a);
  });

  // Low confidence accounts with no user override
  const lowConf = accts.filter(a => a.confidence === 'low' && !(userMappings[a.key] && userMappings[a.key] !== 'ignore'));

  // ── FIX (Issue 1): Use the server-provided manual_review list (full
  // account objects from tb_processor) and split them by suggested_side
  // so the user sees them at the BOTTOM of the BS / P&L panels with a
  // dropdown populated from bs_head_options. Submission still happens
  // through userMappings (keyed by `name_row`).
  const manualReview   = analysisData?.manual_review || [];
  const bsHeadOptions  = analysisData?.bs_head_options || [];
  const manualBS = manualReview.filter(m => (m.suggested_side || 'asset') === 'asset');
  const manualPL = manualReview.filter(m => (m.suggested_side || 'asset') === 'liability');

  let bsHtml = '';
  if (lowConf.length) bsHtml += buildGroup('❌ Needs Manual Mapping', lowConf, true, true);
  BS_HEAD_KEYS.forEach(h => {
    const g = groups[h] || [];
    if (g.length) bsHtml += buildGroup(HEAD_LABEL[h]||h, g, false, false);
  });
  if (groups['ignore']?.length) bsHtml += buildGroup('Ignored', groups['ignore'], false, false);
  // Append server-provided Manual rows at the BOTTOM of BS panel
  if (manualBS.length) bsHtml += buildManualGroup('🔍 Manual Review — Dr Balances (Asset side)', manualBS, bsHeadOptions);
  document.getElementById('bsPanel').innerHTML = bsHtml || '<p style="padding:16px;color:var(--muted);font-size:12px">No BS accounts</p>';

  let plHtml = '';
  PL_HEAD_KEYS.forEach(h => {
    const g = groups[h] || [];
    if (g.length) plHtml += buildGroup(HEAD_LABEL[h]||h, g, false, false);
  });
  // Append server-provided Manual rows at the BOTTOM of P&L panel
  if (manualPL.length) plHtml += buildManualGroup('🔍 Manual Review — Cr Balances (Income / Liability side)', manualPL, bsHeadOptions);
  document.getElementById('plPanel').innerHTML = plHtml || '<p style="padding:16px;color:var(--muted);font-size:12px">No P&L accounts</p>';
}

// ── NEW: Build a group panel for server-provided manual_review items.
// These items only carry {name,row,group,net,dr_cr,bs_head,suggested_side}
// (no `key`), so we look up the matching account in analysisData.accounts
// (which DOES have `key`) to wire the dropdown back into userMappings.
function _findAcctKeyForManual(m) {
  const accts = analysisData?.accounts || [];
  // Primary match: by row number
  if (m.row != null) {
    const byRow = accts.find(a => a.row === m.row && a.name === m.name);
    if (byRow) return byRow.key;
    const byRowOnly = accts.find(a => a.row === m.row);
    if (byRowOnly) return byRowOnly.key;
  }
  // Fallback: by name only
  const byName = accts.find(a => (a.name || '').toUpperCase() === (m.name || '').toUpperCase());
  if (byName) return byName.key;
  // Last resort: synthesize a key in the same format tb_processor uses
  return `${m.name}_${m.row || 0}`;
}

function buildManualGroup(title, manualItems, headOptions) {
  if (!manualItems || !manualItems.length) return '';
  const total = manualItems.reduce((s,m) => s + Math.abs(m.net || 0), 0);
  const opts = (headOptions && headOptions.length)
    ? headOptions
    : BS_HEADS.map(h => ({key: h.v, label: h.l}));

  const rows = manualItems.map(m => {
    const net = m.net || 0;
    const drcr = m.dr_cr || (net < 0 ? 'Cr' : 'Dr');
    const amtCls = drcr === 'Cr' ? 'cr' : 'dr';
    const amtStr = drcr + ' ₹' + Math.abs(net).toLocaleString('en-IN', {maximumFractionDigits:2});
    const key = _findAcctKeyForManual(m);
    const currentHead = userMappings[key] || m.bs_head || 'ignore';
    const selOpts = opts.map(o =>
      `<option value="${escHtml(o.key)}"${currentHead === o.key ? ' selected' : ''}>${escHtml(o.label)}</option>`
    ).join('') + `<option value="ignore"${currentHead === 'ignore' ? ' selected' : ''}>⊘ Ignore / Skip</option>`;

    return `<tr>
      <td><div class="acc-name">${escHtml(m.name)}</div><div class="acc-grp">${escHtml(m.group||'')}</div></td>
      <td class="amt ${amtCls}">${amtStr}</td>
      <td><span class="conf-pill conf-low">❌ Manual</span></td>
      <td><select class="map-sel" data-key="${escHtml(key)}" onchange="onMapChange(this)">
        ${selOpts}
      </select></td>
    </tr>`;
  }).join('');

  const id = 'grp_manual_' + title.replace(/[^a-z0-9]/gi,'_');
  return `<div style="margin-bottom:8px;border-radius:8px;overflow:hidden;border:1px solid #FDE68A">
    <div style="padding:10px 14px;background:#FFFBEB;display:flex;justify-content:space-between;align-items:center;cursor:pointer;user-select:none" onclick="toggleGroup('${id}',this)">
      <span style="font-size:13px;font-weight:700;color:#92400E"><span class="grp-arrow">▼</span> ${escHtml(title)}</span>
      <span style="font-size:12px;font-weight:600;color:var(--ink)">₹${Math.round(total).toLocaleString('en-IN')} <span style="color:var(--muted);font-weight:400">(${manualItems.length})</span></span>
    </div>
    <div id="${id}">
      <table class="map-table">
        <thead><tr><th>Account Name</th><th style="text-align:right">Balance</th><th>Status</th><th>Map To BS Head</th></tr></thead>
        <tbody>${rows}</tbody>
      </table>
    </div>
  </div>`;
}

function buildGroup(title, accounts, highlight, startExpanded) {
  const total = accounts.reduce((s,a)=>s+Math.abs(a.net||0),0);
  const rows = accounts.map(a => {
    const net = a.net || 0;
    const amtCls = net < 0 ? 'cr' : 'dr';
    const amtStr = (net<0?'Cr ':'Dr ') + '₹' + Math.abs(net).toLocaleString('en-IN',{maximumFractionDigits:2});
    const conf = a.confidence || 'low';
    const pill = conf==='high' ? '<span class="conf-pill conf-high">✅ Auto</span>'
               : conf==='med'  ? '<span class="conf-pill conf-med">⚠️ Review</span>'
               : conf==='user' ? '<span class="conf-pill conf-user">🔵 User</span>'
               :                 '<span class="conf-pill conf-low">❌ Manual</span>';

    // Use CURRENT mapping from userMappings for selected value
    const currentHead = userMappings[a.key] || a.bs_head || 'ignore';
    const selOpts = BS_HEADS.map(h =>
      `<option value="${h.v}"${currentHead===h.v?' selected':''}>${h.l}</option>`
    ).join('');

    return `<tr>
      <td><div class="acc-name">${escHtml(a.name)}</div><div class="acc-grp">${escHtml(a.group||'')}</div></td>
      <td class="amt ${amtCls}">${amtStr}</td>
      <td>${pill}</td>
      <td><select class="map-sel" data-key="${escHtml(a.key)}" onchange="onMapChange(this)">
        ${selOpts}
      </select></td>
    </tr>`;
  }).join('');

  const bg = highlight ? '#FFFBEB' : '#FAFAFA';
  const border = highlight ? '1px solid #FDE68A' : '1px solid var(--border)';
  const id = 'grp_' + title.replace(/[^a-z0-9]/gi,'_');
  const collapsed = startExpanded ? '' : 'display:none';
  const arrow = startExpanded ? '▼' : '▶';

  return `<div style="margin-bottom:8px;border-radius:8px;overflow:hidden;border:${border}">
    <div style="padding:10px 14px;background:${bg};display:flex;justify-content:space-between;align-items:center;cursor:pointer;user-select:none" onclick="toggleGroup('${id}',this)">
      <span style="font-size:13px;font-weight:700"><span class="grp-arrow">${arrow}</span> ${escHtml(title)}</span>
      <span style="font-size:12px;font-weight:600;color:var(--ink)">₹${Math.round(total).toLocaleString('en-IN')} <span style="color:var(--muted);font-weight:400">(${accounts.length})</span></span>
    </div>
    <div id="${id}" style="${collapsed}">
      <table class="map-table">
        <thead><tr><th>Account Name</th><th style="text-align:right">Balance</th><th>Status</th><th>Map To BS Head</th></tr></thead>
        <tbody>${rows}</tbody>
      </table>
    </div>
  </div>`;
}

function toggleGroup(id, headerEl) {
  const el = document.getElementById(id);
  if (!el) return;
  const showing = el.style.display === 'none';
  el.style.display = showing ? '' : 'none';
  // Update arrow
  if (headerEl) {
    const arrow = headerEl.querySelector('.grp-arrow');
    if (arrow) arrow.textContent = showing ? '▼' : '▶';
  }
}
function expandAll() {
  document.querySelectorAll('[id^="grp_"]').forEach(el => { el.style.display = ''; });
  document.querySelectorAll('.grp-arrow').forEach(el => { el.textContent = '▼'; });
}

function switchTab(tab) {
  const bsP = document.getElementById('bsPanel');
  const plP = document.getElementById('plPanel');
  const bsB = document.getElementById('tabBS');
  const plB = document.getElementById('tabPL');
  if (tab === 'bs') {
    bsP.style.display = ''; plP.style.display = 'none';
    bsB.style.background = 'var(--brand)'; bsB.style.color = '#fff';
    plB.style.background = '#fff'; plB.style.color = 'var(--brand)';
  } else {
    bsP.style.display = 'none'; plP.style.display = '';
    plB.style.background = 'var(--brand)'; plB.style.color = '#fff';
    bsB.style.background = '#fff'; bsB.style.color = 'var(--brand)';
  }
}

// ═══════════════════════════════════════
//  KEY FIX: onMapChange stores to userMappings immediately
// ═══════════════════════════════════════
function onMapChange(sel) {
  const key = sel.dataset.key;
  const val = sel.value;
  userMappings[key] = val;
  sel.classList.add('changed');
  try {
    sessionStorage.setItem('tb_mappings', JSON.stringify(userMappings));
    const fp = Object.keys(userMappings).sort().slice(0,5).join('|');
    sessionStorage.setItem('tb_mappings_fp', fp);
  } catch(e) {}
  // Track expanded groups before rebuild
  const expanded = new Set();
  document.querySelectorAll('[id^="grp_"]').forEach(el => {
    if (el.style.display !== 'none') expanded.add(el.id);
  });
  rebuildPanels();
  // Restore expanded state
  expanded.forEach(id => {
    const el = document.getElementById(id);
    if (el) {
      el.style.display = '';
      const hdr = el.previousElementSibling;
      if (hdr) { const a = hdr.querySelector('.grp-arrow'); if (a) a.textContent = '▼'; }
    }
  });
}

// ═══════════════════════════════════════
//  STEP 2.5: CAPITAL & FIXED ASSETS
// ═══════════════════════════════════════
let capData = null, faData = null;

async function goToCapFA() {
  document.getElementById('step1').style.display = 'none';
  document.getElementById('loadWrap').style.display = 'block';
  document.getElementById('loadMsg').textContent = 'Reading Capital Account & Fixed Assets from BS template...';

  try {
    const fd = new FormData();
    fd.append('bs_file', bsFile);
    const res = await fetch('/tb-read-bs', {method:'POST', body:fd, credentials:'include'});
    const data = await res.json();
    document.getElementById('loadWrap').style.display = 'none';

    if (data.status !== 'success') {
      // If reading fails, still proceed — user can fill manually later
      capData = null; faData = null;
      buildCapTable(null);
      buildFATable(null);
      goStep('2b');
      return;
    }

    capData = data.capital;
    faData  = data.fixed_assets;
    buildCapTable(capData);
    buildFATable(faData);
    goStep('2b');

  } catch(e) {
    document.getElementById('loadWrap').style.display = 'none';
    capData = null; faData = null;
    buildCapTable(null);
    buildFATable(null);
    goStep('2b');
  }
}

function buildCapTable(cap) {
  const wrap = document.getElementById('capTableWrap');
  if (!cap || !cap.partners || !cap.partners.length) {
    wrap.innerHTML = '<p style="color:var(--muted);font-size:12px">No capital account sheet found in BS template. You can fill it manually in Excel after download.</p>';
    return;
  }
  // Detect columns from the template data
  const cols = cap.columns || [];
  const hasInterest = cols.some(c => c.key === 'interest_on_capital');
  const hasSalary = cols.some(c => c.key === 'salary');

  let html = '<table style="width:100%;border-collapse:collapse;font-size:12px">';
  html += '<tr style="background:#F1F5F9"><th style="padding:8px;text-align:left;border:1px solid var(--border)">Name</th>';
  html += '<th style="padding:8px;text-align:right;border:1px solid var(--border)">Opening</th>';
  html += '<th style="padding:8px;text-align:center;border:1px solid var(--border)">Capital Introduced</th>';
  if (hasInterest) html += '<th style="padding:8px;text-align:center;border:1px solid var(--border)">Interest on Capital</th>';
  if (hasSalary) html += '<th style="padding:8px;text-align:center;border:1px solid var(--border)">Salary</th>';
  html += '<th style="padding:8px;text-align:center;border:1px solid var(--border)">Withdrawals</th></tr>';

  cap.partners.forEach((p, i) => {
    html += `<tr>
      <td style="padding:8px;border:1px solid var(--border);font-weight:600">${escHtml(p.name)}
        <input type="hidden" class="cap-row" value="${p.row}">
        <input type="hidden" class="cap-cols" value='${JSON.stringify(cap.col_map||{})}'>
      </td>
      <td style="padding:8px;border:1px solid var(--border);text-align:right;color:var(--muted)">₹${(p.opening||0).toLocaleString('en-IN',{maximumFractionDigits:2})}</td>
      <td style="padding:4px;border:1px solid var(--border)"><input type="number" class="cap-intro" data-idx="${i}" step="0.01" value="0" style="width:100%;padding:6px;border:1.5px solid var(--border);border-radius:6px;text-align:right;font-size:12px"></td>`;
    if (hasInterest) html += `<td style="padding:4px;border:1px solid var(--border)"><input type="number" class="cap-interest" data-idx="${i}" step="0.01" value="0" style="width:100%;padding:6px;border:1.5px solid var(--border);border-radius:6px;text-align:right;font-size:12px"></td>`;
    if (hasSalary) html += `<td style="padding:4px;border:1px solid var(--border)"><input type="number" class="cap-salary" data-idx="${i}" step="0.01" value="0" style="width:100%;padding:6px;border:1.5px solid var(--border);border-radius:6px;text-align:right;font-size:12px"></td>`;
    html += `<td style="padding:4px;border:1px solid var(--border)"><input type="number" class="cap-wd" data-idx="${i}" step="0.01" value="0" style="width:100%;padding:6px;border:1.5px solid var(--border);border-radius:6px;text-align:right;font-size:12px"></td>
    </tr>`;
  });
  html += '</table>';
  wrap.innerHTML = html;
}

function buildFATable(fa) {
  const wrap = document.getElementById('faTableWrap');
  if (!fa || !fa.assets || !fa.assets.length) {
    wrap.innerHTML = '<p style="color:var(--muted);font-size:12px">No fixed assets sheet found in BS template. You can fill it manually in Excel after download.</p>';
    return;
  }
  let html = '<table id="faTable" style="width:100%;border-collapse:collapse;font-size:12px">';
  html += '<tr style="background:#F1F5F9"><th style="padding:8px;text-align:left;border:1px solid var(--border)">Asset</th>';
  html += '<th style="padding:8px;text-align:right;border:1px solid var(--border)">Opening WDV</th>';
  html += '<th style="padding:8px;text-align:center;border:1px solid var(--border)">Additions &gt;180d</th>';
  html += '<th style="padding:8px;text-align:center;border:1px solid var(--border)">Additions &lt;180d</th>';
  html += '<th style="padding:8px;text-align:center;border:1px solid var(--border)">Sale</th>';
  html += '<th style="padding:8px;text-align:right;border:1px solid var(--border)">Rate %</th></tr>';
  fa.assets.forEach((a, i) => {
    html += buildFARow(a, i);
  });
  html += '</table>';
  wrap.innerHTML = html;
}

function buildFARow(a, i) {
  return `<tr class="fa-row" data-idx="${i}">
    <td style="padding:8px;border:1px solid var(--border);font-weight:600">${escHtml(a.name)}<input type="hidden" class="fa-rownum" value="${a.row}"></td>
    <td style="padding:8px;border:1px solid var(--border);text-align:right;color:var(--muted)">₹${(a.opening_wdv||0).toLocaleString('en-IN',{maximumFractionDigits:2})}</td>
    <td style="padding:4px;border:1px solid var(--border)"><input type="number" class="fa-add-gt" step="0.01" value="0" style="width:100%;padding:6px;border:1.5px solid var(--border);border-radius:6px;text-align:right;font-size:12px"></td>
    <td style="padding:4px;border:1px solid var(--border)"><input type="number" class="fa-add-lt" step="0.01" value="0" style="width:100%;padding:6px;border:1.5px solid var(--border);border-radius:6px;text-align:right;font-size:12px"></td>
    <td style="padding:4px;border:1px solid var(--border)"><input type="number" class="fa-sale" step="0.01" value="0" style="width:100%;padding:6px;border:1.5px solid var(--border);border-radius:6px;text-align:right;font-size:12px"></td>
    <td style="padding:8px;border:1px solid var(--border);text-align:right;color:var(--muted)">${a.rate||0}%</td>
  </tr>`;
}

// ── Add Asset modal functions (Bug 1 fix: proper section dropdown) ──
const FA_SECTION_RATES = {
  'plant_15':     { label: 'PLANT & MACHINERY', rate: 15 },
  'vehicle_15':   { label: 'VEHICLE',            rate: 15 },
  'computer_40':  { label: 'COMPUTERS',          rate: 40 },
  'furniture_10': { label: 'FURNITURE AND FIXTURES', rate: 10 },
  'building_10':  { label: 'BUILDING',           rate: 10 },
  'land_0':       { label: 'LAND',               rate: 0 },
  'other_custom': { label: null,                 rate: 15 },
};

function openAddAssetModal() {
  document.getElementById('newAssetName').value = '';
  document.getElementById('newAssetSection').value = 'plant_15';
  document.getElementById('newAssetRate').value = '15';
  document.getElementById('customRateWrap').style.display = 'none';
  const modal = document.getElementById('addAssetModal');
  modal.style.display = 'flex';
  setTimeout(() => document.getElementById('newAssetName').focus(), 100);
}

function closeAddAssetModal() {
  document.getElementById('addAssetModal').style.display = 'none';
}

function onAssetSectionChange() {
  const sec = document.getElementById('newAssetSection').value;
  const info = FA_SECTION_RATES[sec] || { rate: 15 };
  const customWrap = document.getElementById('customRateWrap');
  if (sec === 'other_custom') {
    customWrap.style.display = 'block';
    document.getElementById('newAssetRate').value = '15';
  } else {
    customWrap.style.display = 'none';
    document.getElementById('newAssetRate').value = info.rate;
  }
}

function confirmAddAsset() {
  const name = document.getElementById('newAssetName').value.trim();
  if (!name) { alert('Please enter an asset name.'); return; }
  const sec = document.getElementById('newAssetSection').value;
  const info = FA_SECTION_RATES[sec] || { rate: 15 };
  const rate = sec === 'other_custom'
    ? (parseFloat(document.getElementById('newAssetRate').value) || 15)
    : info.rate;
  const sectionLabel = info.label || 'PLANT & MACHINERY';

  addFARow(name, rate, sectionLabel);
  closeAddAssetModal();
}

function addFARow(name, rate, sectionLabel) {
  let tbl = document.getElementById('faTable');
  if (!tbl) {
    // Create table if it doesn't exist yet
    const wrap = document.getElementById('faTableWrap');
    wrap.innerHTML = '<table id="faTable" style="width:100%;border-collapse:collapse;font-size:12px"><tr style="background:#F1F5F9"><th style="padding:8px;text-align:left;border:1px solid var(--border)">Asset</th><th style="padding:8px;text-align:right;border:1px solid var(--border)">Opening WDV</th><th style="padding:8px;text-align:center;border:1px solid var(--border)">Additions &gt;180d</th><th style="padding:8px;text-align:center;border:1px solid var(--border)">Additions &lt;180d</th><th style="padding:8px;text-align:center;border:1px solid var(--border)">Sale</th><th style="padding:8px;text-align:right;border:1px solid var(--border)">Rate %</th></tr></table>';
    tbl = document.getElementById('faTable');
  }
  const idx = tbl.querySelectorAll('.fa-row').length;
  const tr = document.createElement('tr');
  tr.className = 'fa-row';
  tr.dataset.idx = idx;
  // FIX (2026-07-16): New assets (row=0) show an editable text input for the name
  // so the user can correct typos or rename before submitting. Static text was used
  // before, making the name uneditable once the row was added.
  tr.innerHTML = `
    <td style="padding:6px;border:1px solid var(--border);font-weight:600">
      <input type="text" class="fa-name-input" value="${escHtml(name)}"
        style="width:100%;border:1.5px solid var(--brand);border-radius:5px;
               padding:4px 7px;font-size:12px;font-weight:600;font-family:inherit;
               background:#EFF6FF;outline:none"
        placeholder="Asset name…">
      <input type="hidden" class="fa-rownum" value="0">
      <input type="hidden" class="fa-rate-val" value="${rate}">
      <input type="hidden" class="fa-section-val" value="${escHtml(sectionLabel||'PLANT & MACHINERY')}">
      <span style="display:block;font-size:10px;color:var(--muted);font-weight:400;margin-top:2px">${escHtml(sectionLabel||'')} · editable</span>
    </td>
    <td style="padding:8px;border:1px solid var(--border);text-align:right;color:var(--muted)">₹0</td>
    <td style="padding:4px;border:1px solid var(--border)"><input type="number" class="fa-add-gt" step="0.01" value="0" style="width:100%;padding:6px;border:1.5px solid var(--border);border-radius:6px;text-align:right;font-size:12px"></td>
    <td style="padding:4px;border:1px solid var(--border)"><input type="number" class="fa-add-lt" step="0.01" value="0" style="width:100%;padding:6px;border:1.5px solid var(--border);border-radius:6px;text-align:right;font-size:12px"></td>
    <td style="padding:4px;border:1px solid var(--border)"><input type="number" class="fa-sale" step="0.01" value="0" style="width:100%;padding:6px;border:1.5px solid var(--border);border-radius:6px;text-align:right;font-size:12px"></td>
    <td style="padding:8px;border:1px solid var(--border);text-align:right;color:var(--muted)">${rate}%</td>`;
  tbl.appendChild(tr);
}

// Close modal on backdrop click
document.addEventListener('click', function(e) {
  const modal = document.getElementById('addAssetModal');
  if (modal && e.target === modal) closeAddAssetModal();
});

function collectCapEntries() {
  const entries = [];
  document.querySelectorAll('.cap-row').forEach((el, i) => {
    const row = parseInt(el.value);
    let colMap = {};
    try { colMap = JSON.parse(document.querySelectorAll('.cap-cols')[i]?.value || '{}'); } catch(e) {}
    const intro = parseFloat(document.querySelectorAll('.cap-intro')[i]?.value) || 0;
    const interest = parseFloat(document.querySelectorAll('.cap-interest')[i]?.value) || 0;
    const salary = parseFloat(document.querySelectorAll('.cap-salary')[i]?.value) || 0;
    const wd = parseFloat(document.querySelectorAll('.cap-wd')[i]?.value) || 0;
    if (row && (intro || interest || salary || wd)) {
      entries.push({
        row,
        introduced: intro, introduced_col: colMap.introduced || 4,
        interest_on_capital: interest, interest_on_capital_col: colMap.interest_on_capital || 5,
        salary: salary, salary_col: colMap.salary || 6,
        withdrawals: wd, withdrawals_col: colMap.withdrawals || 7,
      });
    }
  });
  return entries;
}

function collectFAEntries() {
  const entries = [];
  document.querySelectorAll('.fa-row').forEach(tr => {
    const row = parseInt(tr.querySelector('.fa-rownum')?.value) || 0;
    const gt = parseFloat(tr.querySelector('.fa-add-gt')?.value) || 0;
    const lt = parseFloat(tr.querySelector('.fa-add-lt')?.value) || 0;
    const sale = parseFloat(tr.querySelector('.fa-sale')?.value) || 0;
    // FIX (2026-07-16): read name from the editable input (.fa-name-input) for
    // new assets (row=0). Old assets use static text extracted from the first <td>.
    let name = '';
    const nameInput = tr.querySelector('.fa-name-input');
    if (nameInput) {
      // New asset row — name is in the editable text input
      name = nameInput.value.trim();
    } else {
      // Existing asset row — name is static text in first <td>
      const nameEl = tr.querySelector('td:first-child');
      const nameNode = nameEl ? Array.from(nameEl.childNodes).find(n => n.nodeType === 3) : null;
      name = nameNode ? nameNode.textContent.trim() : (nameEl ? nameEl.textContent.trim() : '');
    }
    const rateEl = tr.querySelector('.fa-rate-val');
    const rate = rateEl ? (parseFloat(rateEl.value) || 0) : 0;
    const sectionEl = tr.querySelector('.fa-section-val');
    const section = sectionEl ? (sectionEl.value || '') : '';
    // For existing rows (row>0): only send if there's actual data
    if (row && (gt || lt || sale)) {
      entries.push({row, name, rate, section, additions_gt180: gt, additions_lt180: lt, sale});
    } else if (!row && name && (gt || lt || sale)) {
      // New asset: row=0, must send name, rate, section for server-side insertion
      entries.push({row: 0, name, rate, section, additions_gt180: gt, additions_lt180: lt, sale});
    }
  });
  return entries;
}

// ═══════════════════════════════════════
//  GENERATE — sends ALL data to server
// ═══════════════════════════════════════
async function doGenerate() {
  // Collect ALL current dropdown values
  document.querySelectorAll('.map-sel').forEach(sel => {
    const key = sel.dataset.key;
    if (key) userMappings[key] = sel.value;
  });

  document.getElementById('step2').style.display = 'none';
  document.getElementById('loadWrap').style.display = 'block';
  document.getElementById('loadMsg').textContent = 'Applying mapping and injecting figures into Balance Sheet...';

  const fd = new FormData();
  fd.append('tb_file', tbFile);
  fd.append('bs_file', bsFile);
  fd.append('cy_year', document.getElementById('cyYear').value);
  fd.append('client_name', document.getElementById('clientName').value || 'Client');
  fd.append('user_mappings', JSON.stringify(userMappings));
  fd.append('capital_entries', JSON.stringify(collectCapEntries()));
  fd.append('fa_entries', JSON.stringify(collectFAEntries()));

  try {
    const res = await fetch('/tb-process', {method:'POST', body:fd, credentials:'include'});
    const data = await res.json();
    document.getElementById('loadWrap').style.display = 'none';

    if (data.status !== 'success') {
      alert('Error: ' + data.message);
      document.getElementById('step2').style.display = 'block';
      return;
    }

    // Skip tally page — download directly
    const fn = data.filename || 'Balance_Sheet.xlsx';
    const dlUrl = '/download/' + data.file_id + '?fn=' + encodeURIComponent(fn);
    
    // Trigger download
    const a = document.createElement('a');
    a.href = dlUrl; a.download = fn; a.click();

    // ── FIX (Issue 2 & 3): Show P&L summary cards (Revenue, Other Income,
    // Direct Expenses, etc.) on the success page. Server now returns these
    // values from result["aggregated"] via /tb-process JSON.
    const t = data.tally || {};
    const ok = !!t.balanced;
    const fmtMoney = (n) => '₹' + (Math.round(n||0)).toLocaleString('en-IN');
    const plRows = [
      ['Revenue from Operations', data.revenue],
      ['Other Income',            data.other_income],          // NEW Issue 2
      ['Opening Stock',           data.opening_stock],
      ['Purchases',               data.purchases],
      ['Direct Expenses',         data.direct_expenses],       // NEW Issue 3
      ['Employee / Salary Exp.',  data.employee_expenses],
      ['Finance Cost',            data.finance_cost],
      ['Depreciation',            data.depreciation],
      ['Other Expenses',          data.other_expenses],
      ['Tax Expense',             data.tax_expense],
      ['Closing Stock',           data.closing_stock],
    ].filter(r => r[1] != null && Math.abs(r[1]) > 0.005);

    const plCardsHtml = plRows.map(r => `
      <div class="sum-card" style="padding:10px 12px;text-align:left">
        <div class="sum-lbl" style="margin:0 0 4px">${escHtml(r[0])}</div>
        <div class="sum-val" style="font-size:15px">${fmtMoney(r[1])}</div>
      </div>`).join('');

    document.getElementById('resBox').innerHTML = `
      <div style="text-align:center;padding:24px 20px 16px">
        <div style="font-size:48px;margin-bottom:8px">✅</div>
        <h2 style="font-size:20px;font-weight:800;margin-bottom:6px">Balance Sheet Downloaded!</h2>
        <p style="color:var(--muted);font-size:13px;margin-bottom:18px">${escHtml(fn)}</p>
        <a href="${dlUrl}" class="btn-main" style="display:inline-flex;padding:12px 32px;text-decoration:none">
          ⬇ Download Again
        </a>
      </div>

      <div class="${ok?'result-ok':'result-ok'}" style="margin-top:6px">
        <div style="font-size:16px;font-weight:800;color:#065F46">
          Balance Sheet Generated ✅
        </div>
      </div>

      <div style="margin-top:14px">
        <div class="trow"><span class="tlbl">Profit / (Loss)</span><span class="tval">${fmtMoney(t.profit)}</span></div>
        <div class="trow"><span class="tlbl">User Mapping Overrides Applied</span><span class="tval" style="color:var(--brand)">${t.user_mappings_applied||0}</span></div>
      </div>

      ${plCardsHtml ? `
      <h3 style="margin:20px 0 8px;font-size:13px;font-weight:700;color:var(--ink);text-transform:uppercase;letter-spacing:.04em">P&amp;L Summary</h3>
      <div class="sum-grid">${plCardsHtml}</div>` : ''}

      ${data.log ? '<div style="margin-top:10px;padding:10px;background:#F9FAFB;border-radius:8px;font-size:10px;color:var(--muted);max-height:120px;overflow-y:auto">'+data.log.slice(-12).map(l=>'<div>'+escHtml(l)+'</div>').join('')+'</div>' : ''}
    `;
    document.getElementById('resSub').textContent = fn;
    goStep(3);

  } catch(e) {
    alert('Error: '+e);
    document.getElementById('step2').style.display = 'block';
    document.getElementById('loadWrap').style.display = 'none';
  }
}

// ═══════════════════════════════════════
//  RESULT
// ═══════════════════════════════════════
function buildResult(data) {
  const t = data.tally || {};
  const ok = t.balanced;
  document.getElementById('resSub').textContent =
    (document.getElementById('clientName').value||'Balance Sheet') + ' · CY figures populated';

  document.getElementById('resBox').innerHTML = `
    <div class="result-ok">
      <div style="font-size:22px;margin-bottom:6px">🎉</div>
      <div style="font-size:16px;font-weight:800;color:#065F46">Balance Sheet Generated ✅</div>
    </div>
    <div style="margin-top:14px">
      <div class="trow"><span class="tlbl">Profit / (Loss)</span><span class="tval">₹${fmt(t.profit)}</span></div>
      <div class="trow"><span class="tlbl">User Mapping Overrides Applied</span><span class="tval" style="color:var(--brand)">${t.user_mappings_applied||0}</span></div>
    </div>
    ${data.log ? '<div style="margin-top:10px;padding:10px;background:#F9FAFB;border-radius:8px;font-size:10px;color:var(--muted);max-height:100px;overflow-y:auto">'+data.log.slice(-10).map(l=>'<div>'+escHtml(l)+'</div>').join('')+'</div>' : ''}`;

  const dlBtn = document.getElementById('dlBtn');
  dlBtn.href = '/download/' + data.file_id + '?fn=' + encodeURIComponent(data.filename);
  dlBtn.setAttribute('download', data.filename);
}

function fmt(n) { return (Math.round(n||0)).toLocaleString('en-IN'); }
function escHtml(s) { return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;'); }

// Drag & drop
['tbZone','bsZone'].forEach(id=>{
  const el=document.getElementById(id);
  el.addEventListener('dragover',e=>{e.preventDefault();el.classList.add('drag')});
  el.addEventListener('dragleave',()=>el.classList.remove('drag'));
  el.addEventListener('drop',e=>{
    e.preventDefault();el.classList.remove('drag');
    const f=e.dataTransfer.files[0]; if(!f) return;
    const type=id==='tbZone'?'tb':'bs';
    onFile({files:[f]},type);
    if(type==='tb') {
      tbFile=f;
      // Clear stale mappings from any previous session whenever a new
      // TB file is picked — prevents old classifications (e.g. all accounts
      // mapped to 'other_cl' by a previous buggy version) from overwriting
      // the fresh auto-classifications returned by the server.
      try {
        sessionStorage.removeItem('tb_mappings');
        sessionStorage.removeItem('tb_mappings_fp');
      } catch(e) {}
    } else { bsFile=f; }
    document.getElementById('analyseBtn').disabled=!(tbFile&&bsFile);
  });
});
</script><a href="https://wa.me/918427651580" target="_blank" class="wa-float" title="WhatsApp Support"><svg viewBox="0 0 24 24"><path d="M17.472 14.382c-.297-.149-1.758-.867-2.03-.967-.273-.099-.471-.148-.67.15-.197.297-.767.966-.94 1.164-.173.199-.347.223-.644.075-.297-.15-1.255-.463-2.39-1.475-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.298-.347.446-.52.149-.174.198-.298.298-.497.099-.198.05-.371-.025-.52-.075-.149-.669-1.612-.916-2.207-.242-.579-.487-.5-.669-.51-.173-.008-.371-.01-.57-.01-.198 0-.52.074-.792.372-.272.297-1.04 1.016-1.04 2.479 0 1.462 1.065 2.875 1.213 3.074.149.198 2.096 3.2 5.077 4.487.709.306 1.262.489 1.694.625.712.227 1.36.195 1.871.118.571-.085 1.758-.719 2.006-1.413.248-.694.248-1.289.173-1.413-.074-.124-.272-.198-.57-.347m-5.421 7.403h-.004a9.87 9.87 0 01-5.031-1.378l-.361-.214-3.741.982.998-3.648-.235-.374a9.86 9.86 0 01-1.51-5.26c.001-5.45 4.436-9.884 9.888-9.884 2.64 0 5.122 1.03 6.988 2.898a9.825 9.825 0 012.893 6.994c-.003 5.45-4.437 9.884-9.885 9.884m8.413-18.297A11.815 11.815 0 0012.05 0C5.495 0 .16 5.335.157 11.892c0 2.096.547 4.142 1.588 5.945L.057 24l6.305-1.654a11.882 11.882 0 005.683 1.448h.005c6.554 0 11.89-5.335 11.893-11.893a11.821 11.821 0 00-3.48-8.413z"/></svg></a>

<button class="help-btn" onclick="openHelp()" title="How to use this tool">?</button>
<div class="help-overlay" id="helpOverlay">
  <div class="help-modal">
    <div class="help-modal-head"><h3>How to Use — Trial Balance → Balance Sheet</h3><button class="help-close" onclick="closeHelp()">&#10005;</button></div>
    <div class="help-modal-body"><div class="help-step"><div class="help-step-num">1</div><div class="help-step-body"><h4>Upload Trial Balance</h4><p>Upload your Excel trial balance with account names and debit/credit balances.</p></div></div><div class="help-step"><div class="help-step-num">2</div><div class="help-step-body"><h4>Upload BS Template</h4><p>Upload your existing Balance Sheet template with CY column cells ready to fill.</p></div></div><div class="help-step"><div class="help-step-num">3</div><div class="help-step-body"><h4>Enter Details</h4><p>Set client name, financial year, and review auto-mapped accounts.</p></div></div><div class="help-step"><div class="help-step-num">4</div><div class="help-step-body"><h4>Fixed Assets & Capital</h4><p>Enter additions, sales, and capital account movements if prompted.</p></div></div><div class="help-step"><div class="help-step-num">5</div><div class="help-step-body"><h4>Generate & Download</h4><p>Click Generate — CY figures are injected into your BS template. Download instantly.</p></div></div><div class="help-tip">💡 Your BS template's formatting and formulas are never changed — only the CY figures are filled in.</div></div>
  </div>
</div>
<script>function openHelp(){document.getElementById('helpOverlay').classList.add('open')}function closeHelp(){document.getElementById('helpOverlay').classList.remove('open')}document.getElementById('helpOverlay').addEventListener('click',function(e){if(e.target===this)closeHelp()})</script>
</body></html>"""


init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
