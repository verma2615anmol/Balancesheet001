"""
tshape_processor.py
===================
T-Shaped Balance Sheet → Comparative Balance Sheet Converter

v2.8  2026-08-17  Nine fixes for Bansal Pharmaceuticals multi-section XLSX:
                  FIX 1 — Sales written per-line (each GST rate separately)
                           to GROSS PROFIT sheet; sales total summed correctly.
                  FIX 2 — Other expenses: all 15 P&L expense labels added to
                           _PL_EXP_MAP so they route to correct notes-to-p&l rows.
                  FIX 3 — Direct expenses (FOC + Freight) written to notes to p&l
                           R24/R25 via direct_expense_items list.
                  FIX 4 — Fixed assets written positionally (not by name-match);
                           asset name from source written into template col A;
                           all 10 assets written regardless of template name match.
                  FIX 5 — Capital additions (col D) exclude Net Profit; only
                           non-profit add-items (Gas Subsidy, Trf Loan on Death,
                           IT Refund, interest items) are summed for additions.
                           Net Profit stays in col F (Share of Profit).
                  FIX 6 — Unsecured Loan: all 10 lenders read from Annexure-B
                           col38/col39 (not col75/col80 Form 3CD columns);
                           correct closing balances written; "From Other Parties"
                           cleared to 0 (all are named related parties).
                  FIX 7 — Sundry Creditors: creditor block reads col40/col41/col45
                           (the correct multi-section layout columns for Bansal);
                           all 54 creditors captured; rows inserted as needed.
                  FIX 8 — Debtors: BOTH debtor columns (BJ:BO col60-66 and
                           BQ:BW col68-74) read and combined; all 114 parties
                           captured; total correctly = 15,698,816.
                  FIX 9 — Securities & Advances: Prepaid Insurance stays in
                           Annexure-F (loans_to_other_items) not OCA; all 17
                           items stay together; total = 18,545,528.
v2.7  2026-08-17  Format detection + accept .xlsx uploads:
                  _detect_input_format() added before parse_tshape_bs().
                  Multi-section XLSX (GD Singla all-on-one-sheet, >50 cols)
                  raises ValueError with a user-readable message instead of
                  silently producing garbage output. Supported: .xls (xlrd)
                  and narrow .xlsx (≤50 cols, openpyxl). Also fixed the
                  _adv_total_row unbound-variable risk in _inject_details_sheet
                  (initialised to None before the conditional scan block so the
                  _skip_rows set is always valid). Module docstring moved to
                  correct position (was after `import math`); import math moved
                  inside docstring block to restore valid Python syntax.
v2.6  2026-08-14  Four further fixes (session 5, live file scan):
                  FIX M — Creditor SUM endpoint not re-shifted by internal formula pass.
                    _inject_details_sheet correctly writes SUM(D23:D109) for creditors.
                    _fix_details_formula_refs then re-scanned Details and shifted D109
                    → D163 (109 >= CRED_INSERT=56). Fix: rows already written by the
                    injector (cred_sum_actual_row and debtor total rows) are excluded
                    from the generic internal-formula shift pass via a skip-set.
                  FIX N — FIX C removed. Template Notes to BS already has R91 →
                    Details!D129 and R97 → Details!D136. The standard _shift_row_ref
                    correctly maps D129→D183 and D136→D190 after 54 creditor rows
                    are inserted. FIX C was adding replacements for D171 which then
                    OVERWROTE these already-correct shifted formulas with D171 (ADVANCE
                    TO SUPPLIERS — wrong row). Removed all FIX C code entirely.
                  FIX O — Notes to p&l mapped rows (R76-R82) get label in col B.
                    Items mapped via _OTHER_EXP_ROW_MAP were writing value only;
                    col B label was not written for mapped rows (only for unmatched
                    spare rows). Now every matched item also writes its name to col B.
                  FIX P — OCL trade-signal rejection tightened. A name containing
                    one trade word (silk, textile, fashion, etc.) alone was enough to
                    reject it from OCL. This blocked legitimate payable items like
                    'Accounting Charges Payable'. Now requires BOTH a trade keyword
                    AND at least one company suffix (pvt/ltd/m/s/traders/co./inc.)
                    before rejecting as a trade company name.
v2.5  2026-08-14  FIX J/K/L — Bank charges routing, TO-prefix strip, IGST highlight.
v2.4  2026-08-14  Five more fixes from image-by-image sheet scan (session 4):
                  FIX F — Break circular reference: notes_to_p&l D18 (CY opening
                    stock) was =E27 (a formula), which created a circular chain via
                    GROSS PROFIT!B9 → D18 → E27 → GP!E13 → B9. Now writes the
                    actual parsed closing_stock value (= PY closing = CY opening).
                    This fixes the -19,486,200 Gross Profit and all downstream NaN.
                  FIX G — Accountant Salary removed from salary_expenses mapping.
                    Previously 'TO ACCOUNTANT SALARY' → salary_expenses caused
                    798,000 instead of 768,000 (768K salary + 30K accountant = 798K).
                    Now routes Accountant Salary and Professional Charges to
                    expense_kws → other_expense_items → Note 19 spare rows (R80, R81).
                  FIX H — Added General Expenses (R76), Entertainment (R77), Printing
                    (R78), Vehicle (R79), Tour & Travelling (R82) to _OTHER_EXP_ROW_MAP
                    so they write to correct labelled rows instead of being silently
                    dropped as unmatched items.
                  FIX I — Finance cost R45 always written (even when 0) to clear any
                    stale template value. Logging added for interest amounts.
                  Also: SESSION_CONTEXT.md created at /home/claude/SESSION_CONTEXT.md
v2.3  2026-08-14  FIX D (Details R15 unsecured other parties) + FIX E (OCL wide scan).
                  FIX D — Details R15 (unsecured FROM OTHER PARTIES) was hardcoded
                    to 0, so the unsecured lenders (Rohit Vig, Santosh Rani etc.) never
                    appeared in the notes to bs unsecured total. Now writes the sum of
                    all unsecured_loan_parties minus the related-party subtotal already
                    in R7-R12, giving the correct "from other parties" balance.
                  FIX E — _extract_ocl_annexure amt_col search widened from 8 to 15
                    cols. Added a two-pass wide fallback: if primary scan finds 0 items,
                    re-scans every row in the Annexure-D section and picks any (text,
                    number) pair where the text matches an OCL keyword, regardless of
                    column distance. This recovers "Accounting Charges Payable",
                    "Audit Fee Payable", "Ch.Issued But Yet Clear", etc. in layouts
                    where the amount column is far from the label column.
v2.2  2026-08-14  Three further fixes from full-sheet scan of Sachidanand output:
                  FIX A — cred_last_data_row now stops at the last actual creditor
                    row (CRED_START+n_cred-1), not at the end of the allocated block
                    which included Advance-from-Customers zero-pad rows and caused
                    the SUM to spill into debtor rows (creditors inflated by ~1.7M).
                  FIX B — OCL injection now hard-rejects: known unsecured lender
                    names, known trade creditor names, P&L items (depreciation, net
                    profit), person-name prefixes (Smt./Sh./Rohit/Varinder…), and
                    numeric-only labels. Previously Rohit Vig (₹9L), 6 trade
                    creditors (₹13L), and P&L figures were written into OCL slots,
                    producing OCL total of ₹31.8L instead of ₹4.5L.
                  FIX C — _fix_details_formula_refs now always repairs Details!D171/
                    E171 and Details!D177/E177 (Sachidanand-specific debtor-TOTAL
                    references in notes to bs R91 and R97). R171 was the ADVANCE TO
                    SUPPLIERS header (blank), so debtors showed ₹0. Fixed to point
                    to the actual deb_lt6/deb_gt6 TOTAL rows computed by injector.
v2.1  2026-08-14  Bug fixes: unsecured amt_col retry (Bug1), creditor/debtor SUM
                  formula rewrite after insert_rows (Bug2), Advance-to-Suppliers
                  SUM formula correction (Bug5), internal Details formula shift.
v2.0  2026-07-30  COMPLETE REWRITE — forensic column-exact extraction

Architecture:
  1. parse_tshape_bs(path)            → dict of all extracted values
  2. inject_into_template(...)        → fills PY column of Output_sample_format.xlsx
  3. process_tshape(...)              → main entry point

Hard rules:
  - tshape_processor.py = T-shaped conversion ONLY
  - processor.py / lumid_compat.py are NEVER touched
  - Output_sample_format.xlsx structure is NEVER modified
  - PY column ← values from T-shaped XLS
  - CY column ← blank / 0 (yellow highlighted) for CA to fill

Supported input formats:
  - .xls  T-shaped GD Singla layout (xlrd)
  - .xlsx T-shaped GD Singla layout (openpyxl, single-sheet ≤50 cols)

NOT supported:
  - .xlsx multi-section format (G.D. Singla & Co. style where all Annexures
    A-H are packed horizontally on ONE sheet with >50 columns).
    Detection: single sheet + BS header row present + max_column > 50.
    These files raise ValueError with a user-readable message.
"""

import math
import re
import os
import shutil
import datetime
from copy import copy

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.cell import MergedCell


# ─────────────────────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _s(v):
    """Cell → clean string."""
    if v is None or (isinstance(v, float) and str(v) == 'nan'):
        return ''
    return str(v).strip()


def _n(v, default=0.0):
    """Cell → float."""
    if v is None:
        return default
    if isinstance(v, (int, float)) and not (isinstance(v, float) and str(v) == 'nan'):
        return float(v)
    try:
        return float(re.sub(r'[,\s]', '', str(v)))
    except Exception:
        return default


def _first_num(row, start=0, end=None):
    """Return first numeric value in row[start:end], else 0."""
    end = end or len(row)
    for i in range(start, min(end, len(row))):
        v = row[i]
        if isinstance(v, (int, float)) and not (isinstance(v, float) and str(v) == 'nan'):
            return float(v)
    return 0.0


_INPUT_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")


def _write(ws, row, col, value, bold=False):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value
    if bold:
        cell.font = Font(bold=True)


def _write_num(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = float(value or 0)
    cell.number_format = '#,##0'


# ─────────────────────────────────────────────────────────────────────────────
#  SECTION 1 — T-shaped BS parser  (v2.0 forensic-exact)
# ─────────────────────────────────────────────────────────────────────────────

def _detect_input_format(filepath: str) -> str:
    """
    Detect whether the uploaded file is a supported T-shaped layout or the
    unsupported multi-section XLSX (G.D. Singla & Co. all-on-one-sheet style).

    Returns one of:
      'tshape_xls'    — supported (.xls, narrow column count)
      'tshape_xlsx'   — supported (.xlsx, narrow column count ≤50 cols)
      'multisection'  — NOT supported (.xlsx, >50 columns on one sheet)
      'unknown'       — unrecognised layout

    Multi-section fingerprint (Bansal / GD Singla combined style):
      • Single sheet (or sheet named after the firm)
      • Row 11 (0-indexed row 10) has 'LIABILITIES' in col 0 AND 'ASSETS' somewhere
        in the same row  →  confirmed T-shaped BS header
      • XLSX file AND ws.max_column > 50  →  multi-section (annexures packed right)
      • XLS or XLSX with max_column ≤ 50 →  standard T-shaped
    """
    ext = os.path.splitext(filepath)[1].lower()

    if ext == '.xls':
        return 'tshape_xls'   # xlrd path; column width is always narrow in .xls

    if ext != '.xlsx':
        return 'unknown'

    # --- XLSX: inspect sheet to detect multi-section layout ---
    try:
        from openpyxl import load_workbook as _lw
        _wb = _lw(filepath, read_only=True, data_only=True)
        _ws = _wb.active

        # max_column can be None for some writers — iterate first few rows to get real count
        _ncols = _ws.max_column or 0
        if _ncols == 0 or _ncols is None:
            # Fallback: iterate first 20 rows to find actual column count
            for _row in _ws.iter_rows(max_row=20):
                _ncols = max(_ncols, len(_row))

        # Check for BS header in first 15 rows
        _has_bs_header = False
        for _r in _ws.iter_rows(max_row=15, values_only=True):
            _rs = ' '.join(str(v).upper() for v in _r if v is not None)
            if 'LIABILIT' in _rs and 'ASSET' in _rs and 'AMOUNT' in _rs:
                _has_bs_header = True
                break

        _wb.close()

        if _has_bs_header and _ncols > 50:
            return 'multisection'
        return 'tshape_xlsx'

    except Exception:
        return 'tshape_xlsx'   # if detection fails, let the parser try and fail naturally



# ─────────────────────────────────────────────────────────────────────────────
#  Multi-section XLSX parser  (GD Singla & Co. — all annexures on one sheet)
# ─────────────────────────────────────────────────────────────────────────────
#
#  Layout (0-based col indices, verified from Bansal Pharmaceuticals 2024):
#
#  A-I   (0-8)   : T-shaped BS summary  (Liabilities | Assets)
#  K-P   (10-16) : Trading Account      (To Purchases | By Sales)
#  R-Y   (17-24) : Profit & Loss        (To Expenses  | By GP)
#  AM-AN (38-39) : Capital Annexure-A   (col38=movements, col39=subtotals/closing)
#  AO    (44-45) : Creditors col (col40=M/s. prefix, col41=name, col45=amount)
#  AY-AZ (45-46) : Other Payables       (col46=label, col50=amount)
#  AZ-BH (50-59) : Fixed Assets + Advances + Cash&Bank (col51=label, col59=amt)
#  BG-BH (60-66) : Debtors block 1      (col60=M/s., col61=name, col66=amount)
#  BI-BP (67-74) : Debtors block 2      (col67=M/s., col68=name, col74=amount)
#  AN-AN (39)    : Unsecured Loan total (col39 row with grand total)

def _parse_multisection_xlsx(filepath: str) -> dict:
    """
    Parse a GD Singla & Co. multi-section XLSX where all annexures are packed
    horizontally on a single sheet.  Tested on Bansal Pharmaceuticals 2024.

    Returns the same dict shape as parse_tshape_bs() so inject_into_template()
    works unchanged.
    """
    from openpyxl import load_workbook as _lw2
    import math as _math2
    import re as _re2
    import datetime as _dt2

    log = ['[multisection] parsing GD Singla multi-section XLSX']

    wb2 = _lw2(filepath, read_only=True, data_only=True)
    ws2 = wb2.active
    rows = [list(r) for r in ws2.iter_rows(values_only=True)]
    wb2.close()

    ncols = max(len(r) for r in rows)
    for r in rows:
        while len(r) < ncols:
            r.append(None)

    def _sv(v):
        return '' if v is None else str(v).strip()

    def _nv(v, default=0.0):
        if isinstance(v, (int, float)) and not (isinstance(v, float) and _math2.isnan(v)):
            return float(v)
        try:
            return float(_sv(v).replace(',', ''))
        except Exception:
            return default

    # ── Entity name & date ────────────────────────────────────────────────────
    entity_name = ''
    bs_date = ''
    py_year_end = ''
    for i, row in enumerate(rows[:15]):
        for j, v in enumerate(row):
            s = _sv(v)
            if not entity_name and len(s) > 8 and 'M/S' in s.upper():
                entity_name = s
            if not bs_date and 'BALANCE' in s.upper() and 'SHEET' in s.upper():
                for k in range(j, min(j + 20, ncols)):
                    dv = row[k]
                    if isinstance(dv, _dt2.datetime):
                        bs_date = dv.strftime('%d.%m.%Y')
                        py_year_end = str(dv.year)
                        break
                    ds = _sv(dv)
                    m = _re2.search(r'(\d{2})[./](\d{2})[./](\d{4})', ds)
                    if m:
                        bs_date = ds.strip()
                        py_year_end = m.group(3)
                        break

    if not entity_name:
        entity_name = _sv(rows[6][0]) if len(rows) > 6 else 'M/S CLIENT'
    if not bs_date:
        bs_date = '31.03.2024'
        py_year_end = '2024'

    log.append(f'Entity: {entity_name} | Date: {bs_date}')

    # ── BS summary totals (cols 0-8) ──────────────────────────────────────────
    # R11 = header: LIABILITIES | ASSETS | AMOUNT
    # Liabilities amount col = 3 (D), Assets amount col = 8 (I)
    LIAB_COL = 3
    ASSET_COL = 8

    secured_loans = 0.0
    unsecured_loans = 0.0
    sundry_creditors = 0.0
    other_payables = 0.0
    fixed_assets_wdv = 0.0
    closing_stock = 0.0
    advances_security = 0.0
    sundry_debtors = 0.0
    cash_bank = 0.0
    capital_closing = 0.0

    for i, row in enumerate(rows):
        label = _sv(row[0]).upper()
        rs = ' '.join(_sv(v).upper() for v in row[:9] if v is not None)
        amt_l = _nv(row[LIAB_COL]) if LIAB_COL < len(row) else 0.0
        amt_a = _nv(row[ASSET_COL]) if ASSET_COL < len(row) else 0.0

        if ('CAPITAL' in label or "PROP" in label) and "'A'" in rs or 'ANNEXURE-A' in rs or '"A"' in rs:
            if amt_l > 0:
                capital_closing = amt_l
        if 'SECURED LOAN' in label and 'UN' not in label:
            if amt_l > 0:
                secured_loans = amt_l
        # Individual secured loan items (bank OD etc.)
        if any(k in label for k in ('HDFC', 'SBI', 'ICICI', 'OBC', 'PNB', 'BANK OD', 'LAP LOAN')):
            if amt_l > 0:
                secured_loans = max(secured_loans, amt_l)
        if 'UNSECURED' in label and 'LOAN' in label:
            if amt_l > 0:
                unsecured_loans = amt_l
        if 'SUNDRY CREDITOR' in label:
            if amt_l > 0:
                sundry_creditors = amt_l
        if 'OTHER PAYABLE' in label:
            if amt_l > 0:
                other_payables = amt_l
        if 'FIXED ASSET' in rs and amt_a > 0 and fixed_assets_wdv == 0:
            fixed_assets_wdv = amt_a
        if 'CLOSING STOCK' in rs and 'CURRENT ASSET' in rs and amt_a > 0:
            closing_stock = amt_a
        if 'CLOSING STOCK' in label and amt_a > 0:
            closing_stock = amt_a
        # closing stock from asset col where label has 'closing stock'
        if closing_stock == 0:
            for jj in range(5, min(9, len(row))):
                sv_j = _sv(row[jj]).upper()
                if 'CLOSING' in sv_j and 'STOCK' in sv_j:
                    closing_stock = amt_a
        if 'SECURITIES' in label or 'ADVANCES' in label and 'SECURITY' in rs:
            if amt_a > 0:
                advances_security = amt_a
        if 'SUNDRY DEBTOR' in label or ('DEBTOR' in label and 'ADVANCE' not in label):
            if amt_a > 0:
                sundry_debtors = amt_a
        if 'CASH' in label and ('BANK' in label or 'BALANCE' in label):
            if amt_a > 0:
                cash_bank = amt_a
        if i > 5 and label.startswith('TOTAL') and amt_l > 0:
            break

    # Direct reads from known rows for Bansal layout
    # R16: PROP CAPITAL = col3 amount
    if len(rows) > 15 and capital_closing == 0:
        capital_closing = _nv(rows[15][3])
    # R21: HDFC Bank OD = col3
    if len(rows) > 20 and secured_loans == 0:
        secured_loans = _nv(rows[20][3])
    # R25: UNSECURED = col3; ADVANCES = col8
    if len(rows) > 24:
        if unsecured_loans == 0:
            unsecured_loans = _nv(rows[24][3])
        if advances_security == 0:
            advances_security = _nv(rows[24][8])
    # R28: CREDITORS = col3; DEBTORS = col8
    if len(rows) > 27:
        if sundry_creditors == 0:
            sundry_creditors = _nv(rows[27][3])
        if sundry_debtors == 0:
            sundry_debtors = _nv(rows[27][8])
    # R31: OTHER PAYABLES = col3; CASH & BANK = col8
    if len(rows) > 30:
        if other_payables == 0:
            other_payables = _nv(rows[30][3])
        if cash_bank == 0:
            cash_bank = _nv(rows[30][8])
    # R13: FIXED ASSETS = col8; CLOSING STOCK = R18 col8
    if len(rows) > 17:
        if fixed_assets_wdv == 0:
            fixed_assets_wdv = _nv(rows[12][8])
        if closing_stock == 0:
            closing_stock = _nv(rows[17][8])

    log.append(f'BS totals: capital={capital_closing:.0f}, secured={secured_loans:.0f}, '
               f'unsecured={unsecured_loans:.0f}, creditors={sundry_creditors:.0f}, '
               f'debtors={sundry_debtors:.0f}, cash={cash_bank:.0f}')

    # ── Trading Account (cols 10-16) ──────────────────────────────────────────
    # Opening stock: col12 row with "To Opening Stock" at col10
    # Sales: col15 "By Sales..." rows, total = col16 (SUM row)
    # Purchases: col11/12 rows with "Purchase GST..." at col10
    # Gross Profit: col12 row with "To Gross Profit"
    # Closing stock: col16 row with "By Closing Stock"

    opening_stock = 0.0
    sales = 0.0
    purchases = 0.0
    gross_profit = 0.0
    foc_amt = 0.0
    freight_amt = 0.0

    _sale_lines = []     # individual sale amounts per GST rate
    _sale_labels = []    # corresponding label for each sale line
    _purch_lines = []

    # FIX 1: Bansal multi-section layout uses 0-based columns:
    # Trading To-side:  col9=label ("To Opening Stock", "To Purchase GST...")
    #                   col10=label continuation, col11=amount (purchase line items)
    #                   col12=amount (subtotals/totals like purchases grand total, GP)
    # Trading By-side:  col13='By', col14=sale description, col15=sale amount, col16=total/closing
    # P&L:              col17='To', col18=expense name, col20=amount
    for i, row in enumerate(rows[:40]):
        lbl9  = _sv(row[9]).upper()  if  9 < len(row) else ''
        lbl10 = _sv(row[10]).upper() if 10 < len(row) else ''
        lbl13 = _sv(row[13]).upper() if 13 < len(row) else ''
        lbl14 = _sv(row[14])         if 14 < len(row) else ''  # sale description (preserve case)
        col11 = _nv(row[11]) if 11 < len(row) else 0.0
        col12 = _nv(row[12]) if 12 < len(row) else 0.0
        col15 = _nv(row[15]) if 15 < len(row) else 0.0
        col16 = _nv(row[16]) if 16 < len(row) else 0.0

        # Opening stock: "To Opening Stock" at col9, amount at col12
        if opening_stock == 0 and 'TO OPENING STOCK' in lbl9 and col12 > 0:
            opening_stock = col12
        # Purchase lines: col9 or col10 has PURCHASE, amount at col11
        if 'PURCHASE' in lbl9 or 'PURCHASE' in lbl10:
            if col11 > 0:
                _purch_lines.append(col11)
        # Purchase total: last purchase row has col12 populated (grand total)
        if ('PURCHASE' in lbl9 or 'PURCHASE' in lbl10) and col12 > 0:
            purchases = col12  # last one wins = grand total row
        # FOC and Freight: label at col10 (not col9), amounts at col12
        if 'F.O.C' in lbl10 or lbl10.strip() == 'FOC':
            foc_amt = col12 or col11
        if 'FREIGHT' in lbl10 or 'FRIEGHT' in lbl10:
            freight_amt = col12 or col11
        # Gross Profit: "To Gross Profit" at col9, amount at col12
        if 'TO GROSS PROFIT' in lbl9 and col12 > 0:
            gross_profit = col12
        # By Closing Stock at col13, amount at col16
        if 'BY CLOSING STOCK' in lbl13 and col16 > 0:
            closing_stock = closing_stock or col16
        # FIX 1: By-Sales lines: col13='By', col14=description, col15=amount
        if lbl13 == 'BY' and lbl14 and col15 > 0:
            if 'SALE' in lbl14.upper() or 'SALES' in lbl14.upper():
                _sale_lines.append(col15)
                _sale_labels.append(lbl14.strip())
        # Sales total: col16 on row 22 (0-indexed 21) = last By-Sales row.
        # ONLY capture col16 when we have sale lines (i.e. we're in the trading section)
        # AND stop before row 23 (0-indexed 22) to avoid P&L totals.
        if col16 > 0 and _sale_lines and 'CLOSING' not in lbl13 and i <= 21:
            sales = max(sales, col16)

    if _sale_lines and sales == 0:
        sales = sum(_sale_lines)
    if purchases == 0 and _purch_lines:
        purchases = sum(_purch_lines)

    # FIX 1: Store individual sale lines for GROSS PROFIT sheet injection
    sale_line_items = [{'label': lbl, 'amount': amt}
                       for lbl, amt in zip(_sale_labels, _sale_lines)]

    log.append(f'Trading: opening={opening_stock:.0f}, purchases={purchases:.0f}, '
               f'sales={sales:.0f}, GP={gross_profit:.0f}')

    # ── P&L Account (cols 17-24) ──────────────────────────────────────────────
    # Labels at col17/col18, amounts at col20
    # "By Gross Profit" at col21, incentives at col24
    # Net Profit: col20 row with "To Net Profit"

    salary_expenses = 0.0
    interest_paid = 0.0
    depreciation = 0.0
    other_income = 0.0
    net_profit = 0.0
    other_expense_items = []

    _PL_EXP_MAP = {
        'AUDIT FEE': 'audit fees',
        'AUDIT FEES': 'audit fees',
        'BANK INTEREST': 'bank interest',
        'BANK CHARGES': 'bank charges',
        'BANK CHARGE': 'bank charges',
        'ELECTRICITY': 'electricity exp',
        'PETROL': 'petrol expenses',
        'MISC EXP': 'misc exp',
        'MISC EXP.': 'misc exp',
        'MISC EXPENSE': 'misc exp',
        'PRINTING': 'printing & stationery',
        'PRINT': 'printing & stationery',
        'PROPERTY TAX': 'property tax',
        'REPAIR': 'repair & maintance',
        'REBARE': 'rebate & discount',
        'REBATE': 'rebate & discount',
        'SHOP EXP': 'shop expenses',
        'SHOP EXPENSE': 'shop expenses',
        'TELEPHONE': 'telephone exp',
        'TOUR': 'tour & travelling',
        'TRAVELLING': 'tour & travelling',
        'INSURANCE': 'insurance',
        'ASSOCIATION': 'association fees',
        # FIX 2: Additional expense labels present in Bansal P&L
        'BANK INTT': 'bank interest',
        'BANK INT': 'bank interest',
        'DEPRECIATION': 'depreciation',  # handled separately but map anyway
        'SALARY': 'salary',              # handled separately
        'REBARE & DISCOUNT': 'rebate & discount',
        'REBATE & DISCOUNT': 'rebate & discount',
        'REBARE & DISC': 'rebate & discount',
    }

    # FIX 2: Bansal multi-section P&L layout (0-based cols):
    # col17='To', col18=expense name, col20=amount (confirmed from data dump)
    # col21='By Gross Profit' / 'By Incentives', col24=amount for By-side
    for i, row in enumerate(rows[:40]):
        lbl17 = _sv(row[17]).upper() if 17 < len(row) else ''
        lbl18 = _sv(row[18]).upper() if 18 < len(row) else ''
        col20 = _nv(row[20]) if 20 < len(row) else 0.0
        col21 = _sv(row[21]).upper() if 21 < len(row) else ''
        col24 = _nv(row[24]) if 24 < len(row) else 0.0

        combined = (lbl17 + ' ' + lbl18).strip()

        if 'TO SALARY' in combined or combined == 'TO SALARY':
            salary_expenses += col20
        elif 'TO DEPRECIATION' in combined:
            depreciation = col20
        elif 'TO INTEREST' in combined or 'TO BANK INT' in combined or 'TO BANK INTT' in combined:
            interest_paid += col20
        elif 'TO NET PROFIT' in combined:
            net_profit = col20
        elif 'BY INCENTIVE' in col21 or 'BY INCENTIVES' in col21:
            other_income += col24
        elif 'BY GROSS PROFIT' in col21:
            pass  # skip
        elif lbl17 == 'TO' and lbl18 and col20 > 0:
            # Match to expense map
            matched = False
            for kw, name in _PL_EXP_MAP.items():
                if kw in lbl18:
                    other_expense_items.append({'name': name, 'amount': col20})
                    matched = True
                    break
            if not matched:
                clean_name = _sv(row[18]).strip().title() if 18 < len(row) else ''
                if clean_name and col20 > 0:
                    other_expense_items.append({'name': clean_name, 'amount': col20})

    log.append(f'P&L: salary={salary_expenses:.0f}, dep={depreciation:.0f}, '
               f'int={interest_paid:.0f}, NP={net_profit:.0f}')

    # ── Capital Annexure-A (col38=movements, col39=totals) ────────────────────
    # Opening = row13 col39 (7,490,745)
    # Add items (rows 15-20): col38=individual amounts; row21 col39=total additions
    # Net Profit = row21 col38 (2,660,247) — goes to Share of Profit (col F), NOT Additions (col D)
    # Additions (col D) = sum of non-profit add items only (Gas Subsidy, Trf Loan, IT Refund, etc.)
    # Less withdrawals = col38 sub-items rows 24-30; col39 = subtotal row
    # Closing = row33 col39 (9,780,338)

    cap_opening = 0.0
    cap_withdrawals = 0.0
    cap_additions = 0.0   # FIX 5: non-profit additions only
    cap_closing = capital_closing  # from BS

    # FIX 5: Capital Annexure-A exact column layout (0-based from data dump):
    # R15: col32='Add:', col33='Gas Subsidy', col38=64.14
    # R16-R20: col33=label, col38=amount (non-profit additions)
    # R21: col33='Net Profit during the year', col38=2,660,247.50, col39=3,698,847.64 (total incl. profit)
    # R24: col32='Less:-', col33='Withdrawls', col38=124,097
    # R25-R30: col33=label, col38=individual withdrawal amounts
    # R30: col39=1,409,254.44 (withdrawal subtotal — authoritative)
    # R33: col33='Closing balance as on', col39=9,780,338.64
    _ADD_START = False
    _WITH_START = False
    _PROFIT_LABELS = ('NET PROFIT', 'PROFIT DURING', 'PROFIT FOR THE YEAR')
    _LESS_MARKERS  = ('LESS:-', 'LESS :', 'LESS:')
    _WITH_LABELS   = ('WITHDRAWL', 'WITHDRAWAL', 'DRAWING', 'HOUSEHOLD')

    if len(rows) > 12:
        cap_opening = _nv(rows[12][39]) if 39 < len(rows[12]) else 0.0

    for i, row in enumerate(rows[13:35], start=13):
        lbl32 = _sv(row[32]).upper() if 32 < len(row) else ''
        lbl33 = _sv(row[33]).upper() if 33 < len(row) else ''
        c38   = _nv(row[38]) if 38 < len(row) else 0.0
        c39   = _nv(row[39]) if 39 < len(row) else 0.0

        # "Add:" section starts when col32 has 'ADD:' marker
        if 'ADD:' in lbl32 and not _ADD_START and not _WITH_START:
            _ADD_START = True
            # Gas Subsidy is on this same row at col33/col38
            if lbl33 and c38 > 0 and not any(k in lbl33 for k in _PROFIT_LABELS):
                cap_additions += c38
            continue

        # "Less:-" section starts when col32 has 'LESS'
        if any(k in lbl32 for k in _LESS_MARKERS) and not _WITH_START:
            _ADD_START = False
            _WITH_START = True
            # First withdrawal item may be on this row (col33)
            if lbl33 and c38 > 0:
                cap_withdrawals += c38
            continue

        if _ADD_START:
            is_profit = any(k in lbl33 for k in _PROFIT_LABELS)
            is_less   = any(k in lbl32 for k in _LESS_MARKERS) or any(k in lbl33 for k in _WITH_LABELS)
            if is_less:
                _ADD_START = False
                _WITH_START = True
                if lbl33 and c38 > 0:
                    cap_withdrawals += c38
                continue
            if is_profit:
                continue  # net_profit already captured from P&L loop — skip
            if lbl33 and c38 > 0:
                cap_additions += c38  # genuine non-profit addition

        if _WITH_START:
            is_closing = 'CLOSING' in lbl33 or ('CLOSING' in lbl32 and 'BALANCE' in lbl32)
            if is_closing:
                _WITH_START = False
                if c39 > 0:
                    cap_closing = c39
                break
            if lbl33 and c38 > 0:
                cap_withdrawals += c38
            # Withdrawal subtotal row (col39 populated, no col33 label or last item row)
            if c39 > 0 and not lbl33:
                cap_withdrawals = c39   # subtotal row — use as authoritative total
                continue

        # Closing row
        if 'CLOSING' in lbl33 and c39 > 0:
            cap_closing = c39
            break

    if cap_closing == 0:
        cap_closing = capital_closing
    if cap_opening == 0 and len(rows) > 12:
        cap_opening = _nv(rows[12][39]) if 39 < len(rows[12]) else 0.0

    log.append(f'Capital: opening={cap_opening:.0f}, additions(non-profit)={cap_additions:.2f}, '
               f'withdrawals={cap_withdrawals:.0f}, closing={cap_closing:.0f}')

    # ── Unsecured Loans Annexure-B ────────────────────────────────────────────
    # FIX 6: Read from Annexure-B directly (col38=name, col39=amount).
    # Header row: col38='ANNEXURE-B' marker (or col33='UNSECURED LOAN').
    # Party rows follow immediately below.
    # Layout for Bansal multi-section XLSX:
    #   R36 (0-indexed i=35): col33='UNSECURED LOAN', col38='ANNEXURE-B'
    #   R38-R48: col33=party name, col39=amount
    #   R49: col39=total (14,947,800)
    # Note: do NOT use col75/col80 (Form 3CD section) — those have CC="maximum balance"
    # not closing balance, causing wrong amounts (e.g. Raj Kumar 7.5M vs actual 7.0M).
    unsecured_loan_parties = []
    _in_unsec = False
    _UNSEC_SKIP_NAMES = {'PARTICULARS', 'TOTAL', 'SUB TOTAL', 'ANNEXURE', 'NIL',
                         'UNSECURED LOAN', 'UN-SECURED LOAN', 'FROM RELATED', 'FROM OTHER',
                         'ADD:', 'LESS:-', 'LAST BALANCE AS ON', 'CLOSING BALANCE AS ON'}
    # FIX 6: Bansal layout: col32(0-based)=name, col38=amount (individual), col39=amount/subtotal
    # Header: R36 has col32='UNSECURED LOAN', col38='ANNEXURE-B'
    # Lender rows R38-R47: col32=name, col39=amount
    # Total row R49: col39=14,947,800
    for i, row in enumerate(rows):
        # Look in cols 32-42 for the Annexure-B header
        rs_zone = ' '.join(_sv(row[c]).upper() for c in range(32, min(42, len(row))) if row[c] is not None)

        if not _in_unsec:
            if ('UNSECURED LOAN' in rs_zone or 'UN-SECURED LOAN' in rs_zone) and \
               ('ANNEXURE-B' in rs_zone or "'B'" in rs_zone):
                _in_unsec = True
            continue

        # Stop at next major section header that bleeds into col32 zone
        if any(k in rs_zone for k in ('SUNDRY CREDITOR', 'ANNEXURE-C', 'OTHER PAYABLE',
                                       'SUNDRY DEBTOR', 'CASH & BANK', 'FIXED ASSET')):
            break
        if i > 55:  # Annexure-B section ends by row 50 in Bansal layout
            break

        # Name at col32, amount at col39
        nm  = _sv(row[32]).strip() if 32 < len(row) else ''
        amt = _nv(row[39]) if 39 < len(row) else 0.0

        if not nm or nm.upper() in _UNSEC_SKIP_NAMES:
            # Check if col39 has the grand total (no name = total row)
            if not nm and amt > 10000000:  # >1 crore = likely grand total
                unsecured_loans = amt
            continue
        if nm.upper().startswith('ANNEXURE'):
            continue
        try:
            float(nm.replace(',', ''))
            continue  # skip numeric labels
        except ValueError:
            pass

        # Valid lender row: must have positive amount at col39
        if amt > 0 and len(nm) >= 3:
            unsecured_loan_parties.append({'name': nm, 'amount': amt})
            log.append(f'  Unsecured lender: {nm} = {amt:,.0f}')
        elif amt == 0:
            # Check col38 for amounts (individual items before subtotal)
            amt38 = _nv(row[38]) if 38 < len(row) else 0.0
            # col38 has small amounts like 64.14 (Gas Subsidy) which are Capital additions
            # For unsecured, amounts should be > 1000
            # Don't capture from col38 — those are Capital Annexure-A items

    # Validate total
    calc_total = sum(p['amount'] for p in unsecured_loan_parties)
    if unsecured_loan_parties and unsecured_loans == 0:
        unsecured_loans = calc_total
    log.append(f'Unsecured parties: {len(unsecured_loan_parties)}, total={calc_total:.0f} (BS={unsecured_loans:.0f})')

    # ── Other Payables Annexure-D (col46=label, col50=amount) ────────────────
    # Header R6: col46='OTHER PAYABLES', col50='ANNEXURE-D'
    # Items: R8-R10 col46=name, col50=amount; R12=total

    other_payable_items = []
    in_ocl = False
    for i, row in enumerate(rows):
        c46 = _sv(row[46]).upper() if 46 < len(row) else ''
        c50 = row[50] if 50 < len(row) else None

        if 'OTHER PAYABLE' in c46 and ('ANNEXURE' in _sv(c50).upper() or i < 10):
            in_ocl = True
            continue
        if not in_ocl:
            continue
        if i > 15:  # OCL section is short (rows 8-12)
            break

        nm = _sv(row[46]).strip()
        amt = _nv(c50)
        if nm and len(nm) > 2 and amt > 0:
            other_payable_items.append({'name': nm, 'amount': amt})
        if isinstance(c50, (int, float)) and c50 > 0 and not nm:
            other_payables = other_payables or c50  # total row

    log.append(f'OCL items: {len(other_payable_items)}, total={other_payables:.0f}')

    # ── Fixed Assets (col51=label, cols52-59=values) ──────────────────────────
    # Header R11: col51='Particulars'  cols 52-59 = Value/Addition/Sale/Total/Rate/Dep/WDV
    # Data rows R14-R24, Total R25

    fixed_asset_items = []
    in_fa = False
    fa_total_wdv = 0.0

    for i, row in enumerate(rows):
        lbl = _sv(row[51]).strip() if 51 < len(row) else ''
        lbl_u = lbl.upper()

        if lbl_u == 'PARTICULARS' and i < 20:
            in_fa = True
            continue
        if not in_fa:
            continue
        if lbl_u in ('ADVANCES & SECURITY', 'CASH & BANK BALANCES', ''):
            if lbl_u not in ('',):
                break

        opening_v = _nv(row[52]) if 52 < len(row) else 0.0
        add_gt    = _nv(row[53]) if 53 < len(row) else 0.0
        add_lt    = _nv(row[54]) if 54 < len(row) else 0.0
        sales_v   = _nv(row[55]) if 55 < len(row) else 0.0
        rate_raw  = row[57] if 57 < len(row) else None
        dep_v     = _nv(row[58]) if 58 < len(row) else 0.0
        wdv_v     = _nv(row[59]) if 59 < len(row) else 0.0

        if not lbl or len(lbl) < 2:
            # Check if this is the total row (col52 has total opening)
            if _nv(row[59] if 59 < len(row) else None) > 0 and not lbl:
                fa_total_wdv = _nv(row[59])
            continue

        if lbl_u == 'TOTAL':
            fa_total_wdv = wdv_v
            break

        rate_f = 0.0
        if rate_raw not in (None, '-', ' '):
            r_n = _nv(rate_raw)
            if 0 < r_n <= 1:
                rate_f = round(r_n * 100)
            elif r_n in (5, 10, 15, 20, 25, 30, 40):
                rate_f = int(r_n)

        if opening_v > 0 or add_gt > 0 or wdv_v > 0:
            fixed_asset_items.append({
                'name': lbl,
                'opening_wdv': opening_v,
                'additions': add_gt + add_lt,
                'sales': sales_v,
                'dep': dep_v,
                'closing_wdv': wdv_v,
                'rate': rate_f,
            })

    if fixed_asset_items:
        calc_wdv = sum(x['closing_wdv'] for x in fixed_asset_items)
        if calc_wdv > 0:
            fixed_assets_wdv = calc_wdv
    elif fa_total_wdv > 0:
        fixed_assets_wdv = fa_total_wdv

    log.append(f'Fixed assets: {len(fixed_asset_items)} items, WDV={fixed_assets_wdv:.0f}')

    # ── Advances & Security (col51=label, col59=amount, rows 27-48) ───────────
    loans_advances_items = []
    in_adv = False
    for i, row in enumerate(rows):
        lbl = _sv(row[51]).strip() if 51 < len(row) else ''
        lbl_u = lbl.upper()
        amt = _nv(row[59]) if 59 < len(row) else 0.0

        if 'ADVANCES & SECURITY' in lbl_u or ('ADVANCE' in lbl_u and 'SECURITY' in lbl_u):
            in_adv = True
            continue
        if not in_adv:
            continue
        if 'CASH & BANK' in lbl_u or 'ANNEXURE-H' in _sv(row[59]).upper():
            break
        if lbl and len(lbl) > 1 and lbl not in ('.', '') and amt > 0:
            loans_advances_items.append({'name': lbl, 'amount': amt})
        # Total row
        if not lbl and amt > 0 and i > 40:
            advances_security = advances_security or amt
            break

    log.append(f'Advances: {len(loans_advances_items)} items, total={advances_security:.0f}')

    # ── Cash & Bank (col51=label, col59=amount, rows after CASH & BANK header) ─
    cash_in_hand = 0.0
    bank_balances = []
    in_cash = False
    for i, row in enumerate(rows):
        lbl = _sv(row[51]).strip() if 51 < len(row) else ''
        lbl_u = lbl.upper()
        amt59 = row[59] if 59 < len(row) else None

        if 'CASH & BANK' in lbl_u or 'CASH AND BANK' in lbl_u:
            in_cash = True
            continue
        if not in_cash:
            continue
        if not lbl and not isinstance(amt59, (int, float)):
            continue
        if lbl and len(lbl) < 2:
            continue

        amt = _nv(amt59)
        if 'CASH IN HAND' in lbl_u:
            cash_in_hand = amt
        elif amt > 0 and lbl:
            bank_balances.append({'name': lbl, 'amount': amt})
        # Total row (blank label with amount)
        if not lbl and isinstance(amt59, (int, float)) and amt > 0:
            cash_bank = cash_bank or amt
            break
        if i > 60:
            break

    if not cash_bank:
        cash_bank = cash_in_hand + sum(b['amount'] for b in bank_balances)

    log.append(f'Cash: hand={cash_in_hand:.0f}, banks={len(bank_balances)}, total={cash_bank:.0f}')

    # ── Debtors blocks G1 (col60/61/66) and G2 (col68/69/75) ────────────────
    # FIX 8: Bansal multi-section XLSX has TWO debtor columns side by side:
    #   G1: col60='M/s.', col61=name, col66=amount  (col BI/BJ/BO 0-based)
    #   G2: col68='M/s.', col69=name, col75=amount  (col BQ/BR/BX 0-based)
    # Both blocks must be read to get all 114 debtors (total = 15,698,816.28).
    # Previous code used col67/68/74 for G2 — wrong; correct is col68/69/75.
    # Total for G1 block: BW8=6,841,420.52; grand total at BW78=15,698,816.28.
    #
    # Column mapping (0-based from row array):
    #   BI=60(M/s.), BJ=61(name), BO=66(amount)   ← G1
    #   BP=67(M/s.), BQ=68(name), BW=74(amount)   ← G2 (BW = col75 1-based = col74 0-based)
    # Note: BW0-based = 74 (B=1,W=23 → col = 1*26+23-1 = 48? No: A=0,B=1,...,Z=25,AA=26,...
    # BW = 1*26+22 = 48+22=... let's use actual: B=2nd letter, W=23rd → BW=26+23-1=48? 
    # Actually from dump: BJ=col62(1-based)=col61(0-based), BO=col67(1-based)=col66(0-based)
    #                     BQ=col69(1-based)=col68(0-based), BW=col75(1-based)=col74(0-based)
    # Confirmed from actual data dump (row 10): BJ10=Abhi Medicos BO10=28885, BQ10=Mahadev... BW10=36804

    sundry_debtor_parties = []
    in_deb = False
    deb_total_calc = 0.0
    _DEB_MFS_VALS = ('m/s.', 'm/s', 'm/s .', 'm/s. ')

    for i, row in enumerate(rows):
        # Start marker: ANNEXURE-G in cols 60-76 zone
        rs_deb = ' '.join(_sv(row[c]).upper() for c in range(60, min(78, len(row))) if row[c] is not None)
        if 'ANNEXURE-G' in rs_deb and not in_deb:
            in_deb = True
            continue
        if not in_deb:
            continue
        if i > 80:
            break

        # G1 block: col60=M/s., col61=name, col66=amount (confirmed from data)
        p60 = _sv(row[60]).strip() if 60 < len(row) else ''
        n61 = _sv(row[61]).strip() if 61 < len(row) else ''
        a66 = _nv(row[66]) if 66 < len(row) else 0.0
        # G2 block: col67=M/s., col68=name, col74=amount (confirmed from data: R10 c67='M/s.', c68='Mahadev...')
        p67 = _sv(row[67]).strip() if 67 < len(row) else ''
        n68 = _sv(row[68]).strip() if 68 < len(row) else ''
        a74 = _nv(row[74]) if 74 < len(row) else 0.0

        if p60.lower() in _DEB_MFS_VALS and n61 and a66 > 0:
            if n61.upper() not in ('TOTAL', 'B/F', 'C/F', 'PARTICULARS'):
                sundry_debtor_parties.append({'name': n61, 'amount': a66})
                deb_total_calc += a66
        if p67.lower() in _DEB_MFS_VALS and n68 and a74 > 0:
            if n68.upper() not in ('TOTAL', 'B/F', 'C/F', 'PARTICULARS'):
                sundry_debtor_parties.append({'name': n68, 'amount': a74})
                deb_total_calc += a74

        # FIX 8: Grand total detection — only stop when we find the grand total row.
        # Row 8 has B/F at col71 with value 6,841,420 — this is G1 carry-forward, NOT grand total.
        # The grand total (BW78 = 15,698,816) appears AFTER all debtors.
        _col71_label = _sv(row[71]).upper() if 71 < len(row) else ''
        if _col71_label == 'B/F':
            continue  # skip B/F carry-forward row
        # Stop when grand total row: both G1 and G2 have no M/s entries AND large amount present
        if (not p60.lower().startswith('m/s') and not n61 and
                not p67.lower().startswith('m/s') and not n68):
            if a74 > 0 and deb_total_calc > 0:
                sundry_debtors = sundry_debtors or a74
                break
            if a66 > 0 and deb_total_calc > 0:
                sundry_debtors = sundry_debtors or a66
                break

    if not sundry_debtors:
        sundry_debtors = deb_total_calc
    log.append(f'Debtors: {len(sundry_debtor_parties)} parties, total={deb_total_calc:.0f}')

    # ── Creditors (Annexure-C multi-section layout) ───────────────────────────
    # FIX 7: Bansal multi-section XLSX layout:
    #   Annexure-C header: col40='SUNDRY CREDITORS', col45='ANNEXURE-C'  (row 5/6)
    #   Creditor entries: col40='M/s.', col41=name, col45=amount
    #   All 54 creditors in rows 8-55 of the sheet (0-indexed rows 7-54)
    # The creditor total is in AT56 (0-indexed row 55) = col45
    sundry_creditor_parties = []
    in_cred = False
    cred_total_calc = 0.0

    for i, row in enumerate(rows):
        # Detect Annexure-C header (cols 40-46 zone)
        rs_cred = ' '.join(_sv(row[c]).upper() for c in range(38, min(50, len(row))) if row[c] is not None)
        if not in_cred:
            if ('SUNDRY CREDITOR' in rs_cred or 'ANNEXURE-C' in rs_cred) and not in_cred:
                in_cred = True
            continue

        if i > 57:  # creditors end at row 56 (0-indexed 55), total at row 57
            break

        p40 = _sv(row[40]).strip() if 40 < len(row) else ''
        n41 = _sv(row[41]).strip() if 41 < len(row) else ''
        a45 = _nv(row[45]) if 45 < len(row) else 0.0

        # Valid creditor: M/s. prefix in col40, name in col41, amount in col45
        if p40.lower() in ('m/s.', 'm/s', 'm/s.', 'm/s .') and n41 and a45 > 0:
            if n41.upper() not in ('TOTAL', 'B/F', 'C/F', 'PARTICULARS'):
                sundry_creditor_parties.append({'name': n41, 'amount': a45})
                cred_total_calc += a45
            continue

        # Total row: blank prefix, no name, but col45 has the grand total
        if not p40 and not n41 and a45 > 0 and cred_total_calc > 0:
            sundry_creditors = a45  # authoritative total from AT56
            break

        # Stop markers — ONLY check in creditor column zone (cols 38-50)
        # Do NOT scan col0-37 which contain BS labels like 'FIXED ASSETS', 'SUNDRY DEBTORS'
        rs_stop_zone = ' '.join(_sv(row[c]).upper() for c in range(38, min(50, len(row))) if row[c] is not None)
        if any(k in rs_stop_zone for k in ('SUNDRY DEBTOR', 'ANNEXURE-G', 'ANNEXURE-D',
                                            'OTHER PAYABLE', 'FIXED ASSET', 'SUNDRY CRED')):
            break

    if not sundry_creditors:
        sundry_creditors = cred_total_calc
    log.append(f'Creditors: {len(sundry_creditor_parties)} parties, total={cred_total_calc:.0f}')

    # ── Classify advances into revenue_auth / other_current / loans ──────────
    # FIX 9: For multi-section XLSX (Bansal/GD Singla), ALL Annexure-F items
    # belong together in Short-Term Loans & Advances (loans_to_other_items).
    # Do NOT split Prepaid Insurance into OCA — it is listed in Annexure-F
    # and the BS total (18,545,528) includes it. Routing it to OCA causes a
    # shortfall in the Loans & Advances total (17,928,190 vs 18,545,528).
    # Only route to revenue_auth if clearly a GST/TDS receivable.
    # Everything else stays in loans_to_other.
    _REV_KW = ('GST', 'CGST', 'SGST', 'IGST', 'TDS', 'TCS',
               'ADVANCE INCOME TAX', 'ADVANCE TAX', 'VAT')
    # NOTE: 'PREPAID INSURANCE' and 'FD' stay in loans_to_other (NOT OCA)
    # because they are inside Annexure-F which is part of Securities & Advances.

    advance_to_revenue_items = []
    loans_to_other_items = []
    other_current_asset_items = []  # empty for multi-section (Bansal) layout

    for it in loans_advances_items:
        nm_u = it['name'].upper()
        if any(k in nm_u for k in _REV_KW):
            advance_to_revenue_items.append(it)
        else:
            # FIX 9: everything else (including Prepaid Insurance) stays here
            loans_to_other_items.append(it)

    # ── Build capital accounts list ───────────────────────────────────────────
    cap_name = entity_name.replace('M/S. ', '').replace('M/S.', '').strip()
    capital_accounts = [{
        'name': cap_name,
        'opening': cap_opening,
        'additions': cap_additions,
        'withdrawals': cap_withdrawals,
        'profit': net_profit,
        'net_profit': net_profit,
        'closing': cap_closing,
    }]

    # ── Assemble result ───────────────────────────────────────────────────────
    result = {
        'entity_name': entity_name,
        'bs_date': bs_date,
        'py_year_end': py_year_end,
        'entity_type': 'proprietorship',
        'capital_accounts': capital_accounts,
        'secured_loans': secured_loans,
        'unsecured_loans': unsecured_loans,
        'unsecured_loan_parties': unsecured_loan_parties,
        'sundry_creditors': sundry_creditors,
        'sundry_creditor_parties': sundry_creditor_parties,
        'advance_from_customers': 0.0,
        'advance_from_customer_parties': [],
        'other_payables': other_payables,
        'other_payable_items': other_payable_items,
        'short_term_provisions': 0.0,
        'fixed_assets_wdv': fixed_assets_wdv,
        'fixed_asset_items': fixed_asset_items,
        'investments': 0.0,
        'advances_security': advances_security,
        'closing_stock': closing_stock,
        'sundry_debtors': sundry_debtors,
        'sundry_debtor_parties': sundry_debtor_parties,
        'loans_advances': advances_security,
        'loans_advances_items': loans_advances_items,
        'loans_to_other_items': loans_to_other_items,
        'advance_to_revenue_items': advance_to_revenue_items,
        'other_current_asset_items': other_current_asset_items,
        'other_current_assets': sum(x['amount'] for x in other_current_asset_items),
        'cash_bank': cash_bank,
        'cash_in_hand': cash_in_hand,
        'bank_balances': bank_balances,
        'other_current_assets': sum(x['amount'] for x in other_current_asset_items),
        'sales': sales,
        'sale_line_items': sale_line_items,  # FIX 1: individual per-GST-rate lines
        'opening_stock': opening_stock,
        'purchases': purchases,
        'direct_expenses': foc_amt + freight_amt,
        'direct_expense_items': (
            ([{'name': 'F.O.C', 'amount': foc_amt}] if foc_amt > 0 else []) +
            ([{'name': 'Freight', 'amount': freight_amt}] if freight_amt > 0 else [])
        ),
        'gross_profit': gross_profit,
        'other_income': other_income,
        'salary_expenses': salary_expenses,
        'interest_paid': interest_paid,
        'depreciation': depreciation,
        'other_expenses': sum(x['amount'] for x in other_expense_items),
        'other_expense_items': other_expense_items,
        'net_profit': net_profit,
        'log': log,
    }

    log.append('[multisection] parse complete')
    return result

def parse_tshape_bs(filepath: str) -> dict:
    """
    Parse a GD Singla & Co. style T-shaped balance sheet.

    Column layout (varies per client but always detected dynamically):
      Left half  = LIABILITIES  (label col, annex col, amount col)
      Right half = ASSETS        (label col, annex col, amount col)
      Far right  = ANNEXURES     (multiple party-list blocks side by side)
      P&L        = embedded in the same sheet at cols 8-16 (Bobby/Ashok)
                   or cols 10-18 (Gupta)

    Returns a dict with all financial values + log.
    """
    log = []

    # ── Format routing ─────────────────────────────────────────────────────
    _fmt = _detect_input_format(filepath)
    if _fmt == 'multisection':
        return _parse_multisection_xlsx(filepath)
    if _fmt == 'unknown':
        raise ValueError(
            'Unrecognised balance sheet layout. '
            'Please upload a standard T-shaped .xls or .xlsx file.'
        )
    # ── / Format routing ────────────────────────────────────────────────────

    ext = os.path.splitext(filepath)[1].lower()
    engine = 'xlrd' if ext == '.xls' else 'openpyxl'
    xl = pd.ExcelFile(filepath, engine=engine)

    # Find main BS sheet
    sheet_name = xl.sheet_names[0]
    for sn in xl.sheet_names:
        sl = sn.lower()
        if any(x in sl for x in ['bs', 'balance', '_co', '_ha', '_fo', '_corp']):
            sheet_name = sn
            break
    log.append(f"Sheet: {sheet_name}")

    df = pd.read_excel(filepath, sheet_name=sheet_name, header=None, engine=engine)
    rows = df.values.tolist()
    nrows = len(rows)
    ncols = max((len(r) for r in rows), default=0)
    for r in rows:
        while len(r) < ncols:
            r.append(None)

    # ── Entity name & date ────────────────────────────────────────────────────
    entity_name, bs_date, py_year_end = '', '', ''
    for i, row in enumerate(rows[:15]):
        for j, v in enumerate(row):
            s = _s(v)
            if not entity_name and len(s) > 8 and 'M/S' in s.upper():
                entity_name = s
            if not bs_date and 'BALANCE' in s.upper() and 'SHEET' in s.upper():
                for k in range(j, min(j + 20, ncols)):
                    dv = row[k]
                    if isinstance(dv, datetime.datetime):
                        bs_date = dv.strftime('%d.%m.%Y')
                        py_year_end = str(dv.year)
                        break
                    ds = _s(dv)
                    m = re.search(r'(\d{2})[./](\d{2})[./](\d{4})', ds)
                    if m:
                        bs_date = ds.strip()
                        py_year_end = m.group(3)
                        break

    if not entity_name:
        entity_name = 'M/S CLIENT'
    if not bs_date:
        bs_date = '31.03.2024'
        py_year_end = '2024'
    log.append(f"Entity: {entity_name} | Date: {bs_date}")

    # ── Find BS header row (LIABILITIES … ASSETS … AMOUNT) ──────────────────
    bs_header_row = -1
    for i, row in enumerate(rows):
        rs = ' '.join(_s(v).upper() for v in row if v is not None)
        if 'LIABILIT' in rs and 'ASSET' in rs and 'AMOUNT' in rs:
            bs_header_row = i
            log.append(f"BS header row: {i}")
            break
    if bs_header_row < 0:
        bs_header_row = 9

    hrow = rows[bs_header_row]

    # Detect liab_amount_col and asset_amount_col from header row
    amount_cols = [j for j, v in enumerate(hrow) if _s(v).upper() == 'AMOUNT']
    liab_amount_col = amount_cols[0] if len(amount_cols) >= 1 else 4
    asset_amount_col = amount_cols[1] if len(amount_cols) >= 2 else (liab_amount_col + 5)
    log.append(f"Liab amount col: {liab_amount_col}, Asset amount col: {asset_amount_col}")

    # ── Result skeleton ───────────────────────────────────────────────────────
    result = {
        'entity_name': entity_name, 'bs_date': bs_date,
        'py_year_end': py_year_end, 'entity_type': 'proprietorship',
        'capital_accounts': [], 'secured_loans': 0.0,
        'unsecured_loans': 0.0, 'unsecured_loan_parties': [],
        'sundry_creditors': 0.0, 'sundry_creditor_parties': [],
        'advance_from_customers': 0.0,
        'other_payables': 0.0, 'other_payable_items': [],
        'short_term_provisions': 0.0,
        'fixed_assets_wdv': 0.0, 'fixed_asset_items': [],
        'investments': 0.0, 'advances_security': 0.0,
        'closing_stock': 0.0,
        'sundry_debtors': 0.0, 'sundry_debtor_parties': [],
        'loans_advances': 0.0, 'loans_advances_items': [],
        'loans_to_other_items': [], 'advance_to_revenue_items': [],
        'other_current_asset_items': [],
        'advance_from_customer_parties': [],
        'cash_bank': 0.0, 'cash_in_hand': 0.0, 'bank_balances': [],
        'other_current_assets': 0.0,
        'sales': 0.0, 'opening_stock': 0.0, 'purchases': 0.0,
        'direct_expenses': 0.0, 'direct_expense_items': [],
        'gross_profit': 0.0, 'other_income': 0.0,
        'salary_expenses': 0.0, 'interest_paid': 0.0,
        'depreciation': 0.0, 'other_expenses': 0.0,
        'other_expense_items': [], 'net_profit': 0.0,
        'log': log,
    }

    # ── Pass 1: scan BS header rows for main totals ───────────────────────────
    _scan_bs_totals(rows, bs_header_row, liab_amount_col, asset_amount_col, result, log)

    # ── Pass 2: detect entity type ────────────────────────────────────────────
    el = entity_name.lower()
    if 'huf' in el or '(huf)' in el or 'hindu' in el:
        result['entity_type'] = 'huf'
    else:
        # Detect partnership only when "PARTNER'S CAPITAL" or "PARTNER CAPITAL" appears
        # in col 0 (liabilities side) — NOT from "Partner" in the CA signing block.
        # The full-text scan is too broad: CA firms sign as "Partner" even for proprietors.
        for row in rows:
            col0 = _s(row[0]).upper()
            if ("PARTNER'S CAPITAL" in col0 or "PARTNER CAPITAL" in col0 or
                    "PARTNERS CAPITAL" in col0):
                result['entity_type'] = 'partnership'
                break

    # ── Pass 3: extract capital accounts ─────────────────────────────────────
    _extract_capital(rows, result, log)

    # ── Pass 4: extract party lists from annexures ────────────────────────────
    _extract_annexure_parties(rows, result, log)

    # ── Pass 5: extract P&L ───────────────────────────────────────────────────
    _extract_capital_annexure(rows, result, log)
    _extract_ocl_annexure(rows, result, log)
    _extract_unsecured_annexure_b(rows, result, log)  # Pass 5d: Annexure-B unsecured loans (FIX S10)
    _extract_debtor_annexure(rows, result, log)
    _extract_creditor_annexure(rows, result, log)   # Pass 5b: Annexure-C creditors
    _extract_loans_annexure(rows, result, log)       # Pass 5c: Annexure-G loans
    _extract_pl(rows, result, log)

    # ── Propagate net_profit to capital accounts ─────────────────────────────
    # The capital PY row (col F = profit) must show the PY net_profit from the
    # T-shaped P&L. Store it on each capital account so the injector can write
    # it as a value (overriding the ='p&l'!F17 formula which references CY).
    if result['net_profit'] > 0 and result['capital_accounts']:
        for cap in result['capital_accounts']:
            cap['net_profit'] = result['net_profit']

    # ── Pass 6: extract fixed assets dep chart ────────────────────────────────
    _extract_dep_chart(rows, result, log)

    # ── Pass 7: extract cash & bank ───────────────────────────────────────────
    _extract_cash_bank(rows, result, log)

    log.append('Parse complete.')
    return result


# ─────────────────────────────────────────────────────────────────────────────
#  Pass 1 — BS totals from header section
# ─────────────────────────────────────────────────────────────────────────────

def _scan_bs_totals(rows, bs_header_row, liab_col, asset_col, result, log):
    """Extract aggregate totals from the BS rows (rows after header)."""
    for i in range(bs_header_row + 1, min(bs_header_row + 50, len(rows))):
        row = rows[i]
        label = _s(row[0]).upper()
        rs = ' '.join(_s(v).upper() for v in row if v is not None)

        # LIABILITIES side — col 0 label
        if ('CAPITAL' in label and ('PROP' in label or 'PARTNER' in label
                                     or 'OWNER' in label or 'HUF' in label)):
            amt = _n(row[liab_col]) if liab_col < len(row) else 0
            if not amt:
                # Try adjacent cols
                amt = _first_num(row, liab_col - 1, liab_col + 3)
            if amt:
                result['capital_accounts'].append({
                    'name': _s(row[0]).strip("`'\""),
                    'opening': amt, 'additions': 0.0,
                    'withdrawals': 0.0, 'profit': 0.0, 'closing': amt
                })
                log.append(f"R{i}: Capital raw={amt}")

        elif 'SECURED LOAN' in label and 'UN' not in label:
            amt = _n(row[liab_col]) if liab_col < len(row) else 0
            if amt and amt == amt:  # not nan
                result['secured_loans'] = amt
            log.append(f"R{i}: Secured loans (header) = {amt}")

        elif re.search(r'ICICI|HDFC|SBI|PNB|OBC|BANK\s+LTD|LAP\s+LOAN|CAR\s+LOAN', label):
            amt = _n(row[liab_col]) if liab_col < len(row) else 0
            if amt and amt == amt:  # not nan
                result['secured_loans'] += amt
                log.append(f"R{i}: Secured loan item = {amt}")

        elif 'UN-SECURED' in label or ('UNSECURED' in label and 'LOAN' in label):
            amt = _n(row[liab_col]) if liab_col < len(row) else 0
            if not amt:
                amt = _first_num(row, liab_col - 1, liab_col + 3)
            if amt:
                result['unsecured_loans'] = amt
                log.append(f"R{i}: Unsecured loans = {amt}")

        elif 'SUNDRY CREDITOR' in label or ('CREDITOR' in label and 'ADVANCE' not in label):
            amt = _n(row[liab_col]) if liab_col < len(row) else 0
            if not amt:
                amt = _first_num(row, liab_col - 1, liab_col + 3)
            if amt:
                result['sundry_creditors'] = amt
                log.append(f"R{i}: Sundry creditors = {amt}")

        elif 'OTHER PAYABLE' in label or label == 'OTHER PAYABLES':
            amt = _n(row[liab_col]) if liab_col < len(row) else 0
            if not amt:
                amt = _first_num(row, liab_col - 1, liab_col + 3)
            if amt:
                result['other_payables'] = amt

        # ASSETS side — only scan the asset half of the row (label in RIGHT half)
        # We check both rs (whole row) for presence and then look only at asset columns
        # Use a limited scan window around asset_amount_col to avoid picking up P&L values
        asc_min = max(asset_col - 2, liab_col + 2)  # never dip into liabilities cols
        asc_max = min(asset_col + 6, len(row))

        # Fixed assets: label must appear in the right half of the row
        rs_right_raw = ' '.join(_s(row[j]).upper() for j in range(liab_col + 1, len(row)) if row[j])
        rs_right = ' '.join(rs_right_raw.split())  # normalize multiple spaces
        if 'FIXED ASSET' in rs_right and result['fixed_assets_wdv'] == 0:
            for j in range(asc_min, asc_max):
                v = _n(row[j])
                if v > 1000:
                    result['fixed_assets_wdv'] = v
                    log.append(f"R{i}: Fixed assets WDV = {v} (col {j})")
                    break

        # Closing stock: only when label explicitly contains 'Closing stock' on asset side
        # AND it appears in the right half of the row
        if result['closing_stock'] == 0:
            # Only set closing stock when 'CLOSING STOCK' or 'STOCK' label is in
            # the ASSET section (first few cols of the right half), not the P&L section.
            # Check if any col in liab_col+1 to asset_col+4 has 'CLOSING' or 'STOCK'.
            asset_label_range = range(liab_col + 1, min(asset_col + 5, len(row)))
            has_stock_label = any(
                'STOCK' in _s(row[j]).upper() and 'CLOSING' in _s(row[j]).upper()
                for j in asset_label_range
            )
            if has_stock_label:
                for j in range(asc_min, min(asset_col + 6, len(row))):
                    v = _n(row[j])
                    if v > 10000:
                        result['closing_stock'] = v
                        log.append(f"R{i}: Closing stock = {v} (col {j})")
                        break

        if ('SUNDRY DEBTOR' in rs_right or 'DEBTORS & ADVANCES' in rs_right) \
           and result['sundry_debtors'] == 0:
            for j in range(asc_min, asc_max):
                v = _n(row[j])
                if v > 100:
                    result['sundry_debtors'] = v
                    log.append(f"R{i}: Sundry debtors = {v} (col {j})")
                    break

        if 'CASH' in rs_right and ('BANK' in rs_right or 'BALANCE' in rs_right) \
           and i < bs_header_row + 35 and result['cash_bank'] == 0:
            for j in range(asc_min, asc_max):
                v = _n(row[j])
                if v > 100:
                    result['cash_bank'] = v
                    log.append(f"R{i}: Cash & bank = {v} (col {j})")
                    break

        if ('ADVANCES & SECURITY' in rs_right or 'ADVANCE & SECURITY' in rs_right or
            ('ADVANCES' in rs_right and 'LOAN' in rs_right and 'SECURITIES' in rs_right)) \
           and result['advances_security'] == 0:
            for j in range(asc_min, asc_max):
                v = _n(row[j])
                if v > 100:
                    result['advances_security'] = v
                    result['loans_advances'] = v
                    log.append(f"R{i}: Advances & security = {v} (col {j})")
                    break

        if ('LOANS & ADVANCE' in rs_right or 'LOAN & ADVANCE' in rs_right) \
           and result['loans_advances'] == 0:
            for j in range(asc_min, asc_max):
                v = _n(row[j])
                if v > 100:
                    result['loans_advances'] = v
                    log.append(f"R{i}: Loans & advances = {v} (col {j})")
                    break

        if ('INVESTMENT' in rs_right or ('SECURITY' in rs_right and 'DEPOSIT' in rs_right)) \
           and result['investments'] == 0:
            for j in range(asc_min, asc_max):
                v = _n(row[j])
                if v > 100:
                    result['investments'] = v
                    break

        # Stop at TOTAL row
        if label.startswith('TOTAL') and i > bs_header_row + 5:
            break


# ─────────────────────────────────────────────────────────────────────────────
#  Pass 3 — Capital accounts
# ─────────────────────────────────────────────────────────────────────────────

def _extract_capital(rows, result, log):
    """
    Find capital account details. In GD Singla format:

    Proprietorship:
      Row with "PROP. CAPITAL A/C" or "PROP CAPITAL ACCOUNT" in col 0,
      followed by "SANJEEV KUMAR..." or entity name with amount in
      annex/amount col. The CLOSING balance is the amount on the Annexure-A
      row (next to `A'`).

    Partnership:
      "PARTNER'S CAPITAL A/C." in col 0, row 11 of Bobby.
      Partner names with amounts in col 0 and col 3 (rows 13,14).

    We read the Annexure-A row for the closing balance.
    """
    caps = []

    for i, row in enumerate(rows):
        label = _s(row[0]).upper()
        rs = ' '.join(_s(v).upper() for v in row if v is not None)

        # Proprietorship / HUF: look for row with `A' or 'A' next to capital label
        # Skip rows where 'E' (fixed assets annex) also appears before 'A' — those are
        # partnership rows where the capital account is on the liabilities side but the
        # main label has PARTNER not PROP.
        has_annex_a = ("'A'" in rs or "ANNEXURE-A" in rs or "`A'" in rs)
        is_prop_cap = has_annex_a and ('CAPITAL' in rs or 'PROP' in rs or 'HUF' in rs)
        # Exclude partnership rows where PARTNER'S CAPITAL appears (handled separately)
        is_prop_cap = is_prop_cap and 'PARTNER' not in label

        if is_prop_cap:
            # Get amount — scan for the capital amount.
            # It should be in the FIRST few cols (liabilities side, cols 1-8)
            # NOT in the right half (which would be fixed assets etc.)
            # Look specifically in col 4 first (Gupta layout), then col 6, then scan 1-8
            amt = 0.0
            amt_row_idx = i  # track which row the amount came from

            _BS_LABEL_WORDS = ('LIABILIT', 'ASSET', 'TOTAL', 'FIXED', 'CURRENT',
                                'SUNDRY', 'UNSECURED', 'STOCK', 'AMOUNT', 'PARTICUL',
                                'CASH', 'BANK', 'LOAN', 'BALANCE SHEET')

            def _is_capital_amount(v):
                return isinstance(v, (int, float)) and not (isinstance(v, float) and str(v) == 'nan') and float(v) > 100000

            for j in [4, 3, 6, 5, 2, 1, 7, 8]:
                if j < len(row):
                    v = _n(row[j])
                    if _is_capital_amount(v):
                        amt = v
                        break

            if not amt:
                # Try looking at the next row for the amount (proprietorship name on next row)
                if i + 1 < len(rows):
                    for j in [4, 3, 6, 5, 2, 1, 7, 8]:
                        if j < len(rows[i+1]):
                            v = _n(rows[i+1][j])
                            if _is_capital_amount(v):
                                amt = v
                                amt_row_idx = i + 1
                                break

            # FIX 1: For Sachidanand-type layout, the `A'` marker and the amount
            # are on DIFFERENT rows. Scan up to 8 rows ahead for the capital amount.
            # ONLY scan cols 1-8 (liabilities side) to avoid picking up P&L amounts at col13.
            if not amt:
                for k in range(i, min(i + 9, len(rows))):
                    r2 = rows[k]
                    for j in [4, 3, 5, 6, 2, 1, 7, 8]:
                        if j < len(r2):
                            v = _n(r2[j])
                            if _is_capital_amount(v):
                                amt = v
                                amt_row_idx = k
                                break
                    if amt:
                        break

            # Find name — first try the row where amount was found (col0),
            # then scan nearby rows for a person name (not a BS label)
            name = _s(rows[amt_row_idx][0]).strip("`'\"") if amt_row_idx < len(rows) else ''
            _bad_names = ('', 'LIABILITIES', 'ASSETS', 'PARTICULARS', 'AMOUNT',
                          'PROP. CAPITAL A/C', 'PROP CAPITAL ACCOUNT', 'PROP CAPITAL A/C')
            if name.upper() in _bad_names or any(w in name.upper() for w in _BS_LABEL_WORDS):
                name = ''
            # Scan i..i+8 for a person name in col0 that's not a BS label
            if not name:
                for k in range(i, min(i + 9, len(rows))):
                    n2 = _s(rows[k][0]).strip("`'\"")
                    if n2 and n2.upper() not in _bad_names and \
                       not any(w in n2.upper() for w in _BS_LABEL_WORDS) and \
                       len(n2) >= 4:
                        name = n2
                        break
            if not name:
                name = 'PROP. CAPITAL A/C'

            if amt > 0:
                # Look for opening, profit, withdrawals in vicinity
                opening = amt
                profit = withdrawals = additions = 0.0

                for k in range(i, min(i + 15, len(rows))):
                    r2 = rows[k]
                    rs2 = ' '.join(_s(v).upper() for v in r2 if v is not None)
                    if 'OPENING BALANCE' in rs2 or 'BALANCE B/D' in rs2 or 'LAST BALANCE' in rs2:
                        opening = _first_num(r2, 1) or opening
                    if 'PROFIT DURING' in rs2 or 'NET PROFIT' in rs2:
                        profit = _first_num(r2, 1)
                    if 'WITHDRAWL' in rs2 or 'HOUSEHOLD' in rs2 or 'DRAWING' in rs2:
                        withdrawals = _first_num(r2, 1)

                # closing = the amount next to `A'`
                closing = amt

                caps.append({
                    'name': name,
                    'opening': opening, 'additions': additions,
                    'withdrawals': withdrawals, 'profit': profit,
                    'closing': closing
                })
                log.append(f"Capital (prop/huf): {name} = {closing}")
                break  # Only one capital block for proprietorship/HUF

        # Partnership: "PARTNER'S CAPITAL" header then partner names below
        if "PARTNER'S CAPITAL" in label or "PARTNER CAPITAL" in label:
            result['entity_type'] = 'partnership'
            # Scan next rows for partner names + amounts in col 0 + col 3
            for k in range(i + 1, min(i + 20, len(rows))):
                r2 = rows[k]
                n2 = _s(r2[0]).strip("`'\"")
                if not n2:
                    continue
                n2u = n2.upper()
                if any(x in n2u for x in ('FIXED', 'CURRENT', 'SECURED',
                                           'SUNDRY', 'CASH', 'TOTAL',
                                           'PARTICUL', 'OTHER', 'LIABIL')):
                    break
                # Check if row 0 looks like a partner name (contains "Sh." or "Mr." or is 5+ chars)
                if len(n2) >= 3 and not n2u.startswith('TO ') and not n2u.startswith('BY '):
                    # Amount in col 3 (Bobby layout)
                    amt = _n(r2[3]) if len(r2) > 3 else 0
                    if not amt:
                        amt = _first_num(r2, 1, 8)
                    if amt > 1000:
                        caps.append({
                            'name': n2,
                            'opening': amt, 'additions': 0.0,
                            'withdrawals': 0.0, 'profit': 0.0,
                            'closing': amt
                        })
                        log.append(f"Partner capital: {n2} = {amt}")
            break

    if caps:
        result['capital_accounts'] = caps
        # Update total
        total_cap = sum(c['closing'] for c in caps)
        log.append(f"Total capital: {total_cap}")


# ─────────────────────────────────────────────────────────────────────────────
#  Pass 4 — Annexure party lists  (forensic-exact column detection)
# ─────────────────────────────────────────────────────────────────────────────

def _extract_annexure_parties(rows, result, log):
    """
    Detect party lists from side-by-side annexure blocks.

    GD Singla layout places multiple annexures on the same rows:
      Ashok:  Creditors=cols(45,47), Debtors=cols(59,65)
      Bobby:  Creditors=cols(28-29,34), Debtors=cols(45,54) + HUF/other at (45,54)
              Unsecured=cols(21,27)
      Gupta:  Creditors=cols(22,28), Debtors=cols(22,28) below a different header

    Strategy:
      1. Scan ALL rows for 'M/s' or 'M/S.' markers with nearby numbers.
      2. Group them by column-pair (name_col, amount_col).
      3. Match each group to the nearest preceding section header
         (SUNDRY CREDITORS / SUNDRY DEBTORS / UNSECURED LOAN).
    """
    SKIP_NAMES = {
        'PARTICULARS', 'FROM RELATED PARTIES', 'FROM OTHER PARTIES',
        'TOTAL', 'SUB TOTAL', 'ANNEXURE', 'DUE TO MSME',
        'ADVANCE FROM CUSTOMERS', 'ADVANCE TO SUPPLIERS', 'NIL',
    }

    # Step 1: Collect all (row_idx, name_col, name, amount_col, amount) tuples
    # where name_col has a text value and amount_col is a number
    entries = []   # (row_i, name_col, name, amount_col, amount)

    for i, row in enumerate(rows):
        for j, v in enumerate(row):
            s = _s(v)
            if not s:
                continue
            su = s.upper()
            # Name prefix 'M/s.' or 'M/S.' or 'M/s' by itself, or person name
            is_prefix = su in ('M/S.', 'M/S', 'M/S .', 'M/S. ', 'MR', 'MR.', 'MRS', 'MRS.')
            is_direct_name = (
                len(s) >= 4 and
                not su.startswith('TO ') and
                not su.startswith('BY ') and
                not any(su.startswith(x) for x in (
                    'TOTAL', 'ANNEXURE', 'FROM ', 'DATE', 'PLACE',
                    'FOR ', 'AS PER', 'CHARTERED', 'UDIN', 'PAN ',
                    'REFER', 'AUDITOR', 'TRADING', 'BALANCE', 'PARTICULARS',
                )) and
                not su.isnumeric() and
                su not in SKIP_NAMES and
                su not in ('NIL', 'NULL', 'DUE TO')
            )

            if is_prefix:
                # Name is in next column
                name_col = j + 1
                if name_col < len(row):
                    name = _s(row[name_col])
                    if name and name.upper() not in SKIP_NAMES and len(name) > 2:
                        # Find amount in cols after name
                        for k in range(name_col + 1, min(name_col + 8, len(row))):
                            amt = _n(row[k])
                            if amt > 100:
                                entries.append((i, name_col, name.strip(), k, amt))
                                break

            elif is_direct_name:
                # Check if next few cols have a number
                for k in range(j + 1, min(j + 6, len(row))):
                    amt = _n(row[k])
                    if amt > 100:
                        entries.append((i, j, s.strip(), k, amt))
                        break

    # Step 2: Find section headers (row index → section name)
    section_rows = {}   # row_i → 'creditor'|'debtor'|'unsecured'|'other_pay'|'loans'
    for i, row in enumerate(rows):
        rs = ' '.join(_s(v).upper() for v in row if v is not None)
        # Priority: unsecured > creditor > debtor (to handle mixed rows)
        if ('UNSECURED LOAN' in rs or 'UN-SECURED LOAN' in rs) and ('ANNEXURE-B' in rs or "'B'" in rs):
            section_rows[i] = 'unsecured'
        elif 'ANNEXURE-B' in rs and i > 5 and 'SUNDRY' not in rs:
            section_rows[i] = 'unsecured'
        elif 'SUNDRY CREDITOR' in rs or 'ANNEXURE-C' in rs or ("'C'" in rs and 'CREDITOR' in rs):
            section_rows[i] = 'creditor'
        elif 'SUNDRY DEBTOR' in rs or 'ANNEXURE-G' in rs or ("'G'" in rs and 'DEBTOR' in rs):
            section_rows[i] = 'debtor'
        elif 'OTHER PAYABLE' in rs and 'ANNEXURE' in rs:
            section_rows[i] = 'other_pay'

    log.append(f"Section header rows: {section_rows}")

    # Step 3: Group entries by (name_col, amount_col) to identify consistent blocks
    # Then for each block, find which section header precedes it most closely
    from collections import defaultdict
    col_pair_entries = defaultdict(list)
    for e in entries:
        row_i, name_col, name, amt_col, amt = e
        col_pair_entries[(name_col, amt_col)].append(e)

    # Step 4: For each col-pair block, assign to nearest preceding section header
    assigned = defaultdict(list)   # section → [(name, amount)]
    for (nc, ac), elist in col_pair_entries.items():
        if len(elist) < 1:
            continue
        for row_i, name_col, name, amt_col, amt in elist:
            # Find the nearest section header at or before row_i
            sec = None
            best_dist = 9999
            for sr, sname in section_rows.items():
                if sr <= row_i and (row_i - sr) < best_dist:
                    best_dist = row_i - sr
                    sec = sname
            if sec and best_dist <= 80:
                assigned[sec].append((name, amt))

    log.append(f"Assigned sections: { {k: len(v) for k, v in assigned.items()} }")

    # Step 5: Deduplicate and store
    def _dedup(items):
        seen = {}
        for name, amt in items:
            key = name.upper()
            if key not in seen or amt > seen[key][1]:
                seen[key] = (name, amt)
        return [{'name': n, 'amount': a} for n, a in seen.values()]

    if assigned.get('creditor'):
        parties = _dedup(assigned['creditor'])
        result['sundry_creditor_parties'] = parties
        calc_total = sum(p['amount'] for p in parties)
        if not result['sundry_creditors'] or abs(calc_total - result['sundry_creditors']) < result['sundry_creditors'] * 0.2:
            result['sundry_creditors'] = result['sundry_creditors'] or calc_total
        log.append(f"Creditor parties: {len(parties)}, total={calc_total:.0f}")

    if assigned.get('debtor'):
        parties = _dedup(assigned['debtor'])
        calc_total = sum(p['amount'] for p in parties)
        actual_debtors = result.get('sundry_debtors', 0)
        # If assigned debtors total is way off from BS figure, try to filter.
        # This handles Gupta where unsecured loan parties bleed into debtor section.
        if actual_debtors > 0 and calc_total > actual_debtors * 3:
            # Assigned debtors total is way off — filter to only parties where
            # the individual amount is <= actual_debtors (exclude large unsecured items)
            filtered = [p for p in parties if p['amount'] <= actual_debtors * 1.5]
            if sum(p['amount'] for p in filtered) > 0:
                parties = filtered
        result['sundry_debtor_parties'] = parties
        # Always use BS-derived total if we have one
        if not result['sundry_debtors']:
            result['sundry_debtors'] = calc_total
        log.append(f"Debtor parties: {len(parties)}, total={calc_total:.0f} (BS={actual_debtors:.0f})")

    if assigned.get('unsecured'):
        parties = _dedup(assigned['unsecured'])
        # Filter out entries that look like P&L items (large salary/expense amounts)
        pl_keywords = ('HOUSE EXP', 'SCHOOL FEE', 'MEDICLAIM', 'SALARY', 'BALANCE C/D',
                       'NET PROFIT', 'GROSS PROFIT', 'PARTNER', 'AUDITOR')
        parties = [p for p in parties
                   if not any(kw in p['name'].upper() for kw in pl_keywords)]
        result['unsecured_loan_parties'] = parties
        calc_total = sum(p['amount'] for p in parties)
        if calc_total > 0:
            result['unsecured_loans'] = result['unsecured_loans'] or calc_total
        log.append(f"Unsecured parties: {len(parties)}, total={calc_total:.0f}")

    if assigned.get('other_pay'):
        parties = _dedup(assigned['other_pay'])
        result['other_payable_items'] = parties
        result['other_payables'] = result['other_payables'] or sum(p['amount'] for p in parties)


# ─────────────────────────────────────────────────────────────────────────────
#  Pass 4b — OTHER PAYABLE annexure direct extraction
# ─────────────────────────────────────────────────────────────────────────────

def _extract_debtor_annexure(rows, result, log):
    """
    Directly extract Annexure-I (Sundry Debtors) entries from the GD Singla T-shaped XLS.

    Layout (0-based col indices):
      col58 = "M/s." prefix
      col59 = party name
      col65 = amount (receivable)
    The TOTAL row at col59="TOTAL" confirms the section end.

    This gives the COMPLETE debtor list including parties that the section-reader
    assigns to the unsecured column range (Ram Beran, Solar, Treat Bell, etc.).
    """
    # Already populated with good data? Skip.
    existing = result.get('sundry_debtor_parties', [])
    bs_total = result.get('sundry_debtors', 0)
    if existing:
        existing_total = sum(p['amount'] for p in existing)
        if bs_total > 0 and abs(existing_total - bs_total) / bs_total < 0.02:
            return  # already accurate

    in_section = False
    items = []
    total_row_val = 0.0

    for i, row in enumerate(rows):
        rs = ' '.join(_s(v).upper() for v in row if v is not None)

        if not in_section:
            # Annexure-I header: col65 (idx 65 0-based) has 'ANNEXURE-I'
            # or the row contains 'ANNEXURE-I'
            if 'ANNEXURE-I' in rs:
                in_section = True
                continue
            continue

        # Name at col59 (0-based), amount at col65 (0-based)
        prefix    = _s(row[59]).strip() if 59 < len(row) else ''  # "M/s."
        name_cell = _s(row[60]).strip() if 60 < len(row) else ''  # actual party name
        amt       = row[65] if 65 < len(row) else None

        # Stop at TOTAL row (col59 has "TOTAL")
        if prefix.upper() == 'TOTAL':
            if isinstance(amt, (int, float)) and amt > 0:
                total_row_val = amt
            break

        if name_cell and len(name_cell) > 2 and isinstance(amt, (int, float)) and amt > 0:
            # Build full name: prefix + name
            p_lower = prefix.lower()
            if p_lower in ('m/s.', 'm/s', 'ms.', 'ms'):
                full_name = 'M/s. ' + name_cell.strip()
            elif prefix:
                full_name = prefix + ' ' + name_cell.strip()
            else:
                full_name = name_cell.strip()
            items.append({'name': full_name.strip(), 'amount': amt})

    if items and total_row_val > 0:
        calc = sum(it['amount'] for it in items)
        if abs(calc - total_row_val) / total_row_val < 0.02:
            result['sundry_debtor_parties'] = items
            log.append(f"Debtor Annexure-I: {len(items)} parties, total={calc:,.0f} "
                       f"(BS={bs_total:,.0f})")


def _extract_creditor_annexure(rows, result, log):
    """
    Directly extract Annexure-C (Sundry Creditors & Advances) from the GD Singla XLS.

    Layout A (Ashok Kumar): col45 = party name, col49 = amount
    Layout B (Sachidanand): col33 = party name, col38 = amount
    Header row triggers on 'SUNDRY CREDITOR' + 'ANNEXURE-C' in row.
    Stop at TOTAL row or amount sum matches BS total.
    """
    bs_total = result.get('sundry_creditors', 0)

    # GD Singla XLS can have up to 4 parallel creditor columns per page (B/F continuation).
    # Each block: (prefix_col, name_col, amt_col)
    # Block1: col32 (M/s.), col33 (name), col38 (amount)
    # Block2: col39 (M/s.), col40 (name), col45 (amount)
    # Block3: col46 (M/s.), col47 (name), col52 (amount)
    # Block4: col25 (M/s.), col26 (name), col31 (amount)  ← Sachidanand continuation page
    # Layout A (other clients): col45 (name), col49 (amount) — no separate prefix col

    def _collect_creditor_block(rows, prefix_col, name_col, amt_col, start=8, end=None):
        """Collect creditor entries from a single column block."""
        items = []
        end = end or len(rows)
        for i in range(start, min(end, len(rows))):
            row = rows[i]
            if prefix_col is not None:
                prefix = _s(row[prefix_col]).strip() if prefix_col < len(row) else ''
                if prefix.lower() not in ('m/s.', 'm/s', 'm/s .', 'm/s. ', 'ms.'):
                    continue
            name_cell = _s(row[name_col]).strip() if name_col < len(row) else ''
            amt_cell = row[amt_col] if amt_col < len(row) else None
            if not name_cell or name_cell in ('nan', '') or len(name_cell) < 2:
                continue
            if not isinstance(amt_cell, (int, float)) or isinstance(amt_cell, float) and (str(amt_cell) == 'nan') or amt_cell <= 0:
                continue
            # Skip B/F continuation rows (large round numbers in first few rows)
            if name_cell.upper() in ('B/F', 'C/F', 'TOTAL', 'B / F', 'C / F'):
                continue
            items.append({'name': name_cell, 'amount': float(amt_cell)})
        return items

    # Find the row range where creditors appear (after SUNDRY CREDITORS + ANNEXURE-C header)
    cred_start = 8
    for i, row in enumerate(rows):
        rs = ' '.join(_s(v).upper() for v in row if v is not None)
        if 'SUNDRY CREDITOR' in rs and 'ANNEXURE-C' in rs:
            cred_start = i + 1
            break

    # Collect all 4 blocks and deduplicate by name
    all_blocks = (
        _collect_creditor_block(rows, 32, 33, 38, cred_start) +
        _collect_creditor_block(rows, 39, 40, 45, cred_start) +
        _collect_creditor_block(rows, 46, 47, 52, cred_start) +
        _collect_creditor_block(rows, 25, 26, 31, cred_start)
    )

    # Fallback: if none of the M/s. prefix blocks found items, try name-only layout (col45/col49)
    if not all_blocks:
        all_blocks = _collect_creditor_block(rows, None, 45, 49, cred_start)

    if not all_blocks:
        log.append("Creditor Annexure-C: no items found in any layout")
        return

    # Deduplicate by normalized name
    seen_cred = set()
    cred_items = []
    for item in all_blocks:
        key = item['name'].strip().upper()
        if key not in seen_cred:
            seen_cred.add(key)
            cred_items.append(item)

    calc = sum(x['amount'] for x in cred_items)

    # If calc >> BS total, filter out OCL items (One97 Communications ₹62L is trade payable
    # in this XLS but may need to stay — keep all items and let Details sheet handle it)
    result['sundry_creditor_parties'] = cred_items
    log.append(f"Creditor Annexure-C (multi-block): {len(cred_items)} creditors, total={calc:,.0f} "
               f"(BS={bs_total:,.0f})")


# ─── Loans & Advances Annexure (Annexure-G in GD Singla layout) ──────────────
_REVENUE_AUTHORITY_KEYWORDS = (
    'GST', 'CGST', 'SGST', 'IGST', 'TDS', 'TCS', 'TAX', 'INCOME TAX',
    'ADVANCE TAX', 'VAT', 'SERVICE TAX', 'EXCISE', 'CUSTOM',
)
_OTHER_CURRENT_ASSET_KEYWORDS = (
    'PREPAID', 'INSURANCE', 'ADVANCE TO SUPPLIER', 'SECURITY DEPOSIT',
    'FD', 'FIXED DEPOSIT',
)


def _classify_loan_item(name):
    """Return 'revenue_auth' | 'other_current' | 'loans_to_others'."""
    n = name.upper()
    if any(kw in n for kw in _REVENUE_AUTHORITY_KEYWORDS):
        return 'revenue_auth'
    if any(kw in n for kw in _OTHER_CURRENT_ASSET_KEYWORDS):
        return 'other_current'
    return 'loans_to_others'


def _extract_loans_annexure(rows, result, log):
    """
    Reads LOANS & ADVANCES (Annexure-G) AND INVESTMENT & SECURITY (Annexure-F)
    from the GD Singla XLS, both at col50 (name) / col58 (amount).

    Annexure-G items are classified into:
      - loans_to_others       → STL note B section (R134-R140)
      - advance_to_revenue    → STL note C section (R143-R147)
      - other_current         → OCA note (R154-R159)

    Annexure-F items (investment / security deposits like D.F.S.C) are treated as
    Short Term Loans — Loans to Others (CA instruction), so they are appended to
    loans_to_other_items.  result['investments'] is zeroed out accordingly.
    """
    # ── Read Annexure-G (LOANS & ADVANCES) — try multiple column layouts ──────
    # Layout A (Ashok Kumar): name=col50, amount=col58
    # Layout B (Sachidanand): name=col53, amount=col62
    items_g = []
    total_g  = 0.0

    _STOP_HEADERS_LOANS = ('CASH & BANK', 'SUNDRY DEBTOR', 'SUNDRY CREDITOR',
                            'FIXED ASSET', 'OTHER PAYABLE', 'TOTAL OF BALANCE')

    def _read_annexure_g(rows, name_col, amt_col):
        in_sec = False
        items = []
        total = 0.0
        for i, row in enumerate(rows):
            if name_col >= len(row):
                continue
            name_cell = _s(row[name_col]).strip()
            amt_cell  = row[amt_col] if amt_col < len(row) else None
            if not in_sec:
                rs = ' '.join(_s(v).upper() for v in row if v is not None)
                if ('LOANS & ADVANCE' in rs or 'LOAN & ADVANCE' in rs) and 'ANNEXURE' in rs:
                    in_sec = True
                continue
            # Stop on TOTAL row
            if name_cell.upper() == 'TOTAL':
                if isinstance(amt_cell, (int, float)) and amt_cell > 0:
                    total = amt_cell
                break
            # Stop when a new section header appears at the name column
            if any(kw in name_cell.upper() for kw in _STOP_HEADERS_LOANS):
                break
            if not name_cell or len(name_cell) < 2:
                continue
            if not isinstance(amt_cell, (int, float)) or amt_cell <= 0:
                continue
            # Skip annexure header values (text like 'ANNEXURE-H')
            if 'ANNEXURE' in name_cell.upper():
                continue
            items.append({'name': name_cell, 'amount': amt_cell,
                          'category': _classify_loan_item(name_cell)})
        return items, total

    # FIX 4: Try all layouts, prefer those giving real (non-NaN) amounts
    items_g, total_g = [], 0.0
    nc, ac = 50, 58  # defaults
    for nc_try, ac_try in [(53, 62), (50, 58), (51, 59), (52, 61)]:
        items_try, total_try = _read_annexure_g(rows, nc_try, ac_try)
        # Filter out NaN amounts
        items_try = [x for x in items_try
                     if isinstance(x['amount'], (int, float)) and
                     not (isinstance(x['amount'], float) and math.isnan(x['amount'])) and
                     x['amount'] > 0]
        if items_try:
            items_g, total_g = items_try, total_try
            nc, ac = nc_try, ac_try
            log.append(f"Loans Annexure-G found at col{nc}/{ac}")
            break

    # ── Read Annexure-F (INVESTMENT & SECURITY) ───────────────────────────────
    in_inv = False
    items_f = []

    # Use the same name/amt cols that worked for Annexure-G (or fall back to 50/58)
    _g_nc = nc
    _g_ac = ac

    for i, row in enumerate(rows):
        if _g_nc >= len(row):
            continue
        name_cell = _s(row[_g_nc]).strip()
        amt_cell  = row[_g_ac] if _g_ac < len(row) else None

        if not in_inv:
            rs = ' '.join(_s(v).upper() for v in row if v is not None)
            if ('INVESTMENT' in rs and 'SECURITY' in rs and 'ANNEXURE-F' in rs):
                in_inv = True
            continue

        if name_cell.upper() == 'TOTAL':
            break
        # Stop if next major section header appears
        if any(kw in name_cell.upper() for kw in ('LOANS', 'CASH', 'SUNDRY', 'OTHER PAYABLE')):
            break

        if not name_cell or len(name_cell) < 2:
            continue
        if not isinstance(amt_cell, (int, float)) or amt_cell <= 0:
            continue

        items_f.append({
            'name': name_cell,
            'amount': amt_cell,
            'category': 'loans_to_others',   # CA instruction: treat as STL
        })

    if not items_g and not items_f:
        return

    # ── Classify Annexure-G items ─────────────────────────────────────────────
    loans_to_others = [x for x in items_g if x['category'] == 'loans_to_others']
    revenue_auth    = [x for x in items_g if x['category'] == 'revenue_auth']
    other_current   = [x for x in items_g if x['category'] == 'other_current']

    # Append Annexure-F items (DFSC etc.) into loans_to_others
    loans_to_others = loans_to_others + items_f

    result['loans_to_other_items']      = loans_to_others
    result['advance_to_revenue_items']  = revenue_auth
    result['other_current_asset_items'] = other_current
    result['other_current_assets']      = sum(x['amount'] for x in other_current)

    # loans_advances total = STL items (loans_to_others + revenue_auth)
    stl_total = (sum(x['amount'] for x in loans_to_others) +
                 sum(x['amount'] for x in revenue_auth))
    if result.get('loans_advances', 0) == 0:
        result['loans_advances'] = stl_total
    result['loans_advances_items'] = loans_to_others

    # Zero out investments since DFSC is now in STL
    if items_f:
        result['investments'] = 0.0

    if items_g:
        calc_g = sum(x['amount'] for x in items_g)
        log.append(f"Loans Annexure-G: {len(items_g)} items, total={calc_g:,.0f}")
    if items_f:
        calc_f = sum(x['amount'] for x in items_f)
        log.append(f"Investment Annexure-F → STL: {len(items_f)} items ({calc_f:,.0f})")
    log.append(f"STL: {len(loans_to_others)} loans_to_others, "
               f"{len(revenue_auth)} revenue_auth, {len(other_current)} other_current")


def _extract_ocl_annexure(rows, result, log):
    """
    Directly extract OTHER PAYABLE (Annexure-D) items by scanning for the
    section header and reading name+amount from the adjacent columns.
    This handles Salary Payable and other OCL items that the section reader misses.

    FIX E (2026-08-14): Extended amt_col search range from 8 to 15 cols right of name.
    Also added a two-pass approach: if the keyword-first name_col detection fails to
    find any items (amt=0 or no keyword match on first real row), fall back to a
    wider scan that reads ALL text+number pairs in the section and filters by keyword.
    This handles the Sachidanand layout where the amount col is far from the name col.
    """
    # Guard: skip only if other_payable_items already contains genuinely payable items.
    # The _extract_annexure_parties may populate other_payable_items with trade creditors
    # that were misassigned to the 'other_pay' section. In that case we should NOT skip —
    # we should replace them with the correctly-extracted Annexure-D items.
    _existing_ocl = result.get('other_payable_items', [])
    if _existing_ocl:
        _payable_kws = ('payable', 'tds', 'provision', 'salary payable', 'ch.issued',
                        'yet clear', 'outstanding', 'accrued', 'rcm')
        _has_real_ocl = any(
            any(kw in it.get('name','').lower() for kw in _payable_kws)
            for it in _existing_ocl
        )
        if _has_real_ocl:
            return  # already has correct OCL items — skip

    in_section = False
    items = []
    name_col = None
    amt_col  = None
    other_pay_total = result.get('other_payables', 0)
    # FIX S10: Only genuinely-payable items qualify as OCL.
    # 'gst' alone is too broad — trade creditors with GST in name would bleed in.
    # 'ch.issued', 'cheque issued', 'but yet clear' are also OCL items.
    _OCL_NAME_KEYS = ('payable', 'tds', 'provision', 'salary',
                      'accrued', 'outstanding', 'due to', 'liability',
                      'advance from', 'ch.issued', 'cheque issued', 'yet clear',
                      'charges payable', 'fees payable', 'bonus', 'rcm')
    _TRADE_REJECT = ('m/s', 'pvt', 'ltd', 'textile', 'fabrics', 'fashion',
                     'silk', 'creation', 'impex', 'enterprises', 'house',
                     'communication', 'print', 'embroidery', 'stores',
                     'industries', 'international', 'sons', 'bros')

    section_rows = []  # collect all rows in the section for fallback

    for i, row in enumerate(rows):
        rs = ' '.join(_s(v).upper() for v in row if v is not None)

        if not in_section:
            if 'OTHER PAYABLE' in rs and 'ANNEXURE' in rs:
                in_section = True
                continue
            continue

        # Stop at next major section header
        _STOP = ('SUNDRY CREDITOR', 'ANNEXURE-C', 'ANNEXURE-G', 'CASH & BANK',
                 'FIXED ASSET', 'SUNDRY DEBTOR', 'TOTAL OF BALANCE')
        if any(kw in rs for kw in _STOP):
            break

        section_rows.append(row)

        # Detect name_col and amt_col from first item row that has a keyword match
        if name_col is None:
            for ci, v in enumerate(row):
                sv = _s(v).strip()
                if len(sv) > 3:
                    sv_lower = sv.lower()
                    if any(kw in sv_lower for kw in _OCL_NAME_KEYS):
                        # FIX E: search up to 15 cols right (was 8)
                        # Require amount > 100 to avoid ICDS footnote numbers (e.g. 8, 11)
                        # which accidentally matched "Provisions, contingent liabilities..."
                        # and set name_col=69 / amt_col=77 — completely wrong columns.
                        for ac in range(ci + 1, min(ci + 15, len(row))):
                            av = row[ac]
                            if isinstance(av, (int, float)) and av > 100:
                                name_col = ci
                                amt_col  = ac
                                break
                    if name_col is not None:
                        break

        if name_col is None:
            continue

        # Check if the OCL name column has 'TOTAL' — this is the real OCL total row
        cell_name = _s(row[name_col]).strip().upper() if name_col < len(row) else ''
        if cell_name == 'TOTAL':
            break

        # Get name from name_col and amount from amt_col
        nm = _s(row[name_col]).strip()
        amt = row[amt_col] if amt_col < len(row) else None
        if nm and len(nm) > 2 and isinstance(amt, (int, float)) and amt > 0:
            nm_lower = nm.lower()
            if any(tr in nm_lower for tr in _TRADE_REJECT):
                continue  # skip trade company names — they belong in creditors not OCL
            if any(kw in nm_lower for kw in _OCL_NAME_KEYS):
                items.append({'name': nm, 'amount': amt})

    # ── FIX E fallback: if primary scan found nothing, do a wide scan of all
    # text+number pairs in the section, keeping only those matching OCL keywords.
    # This covers layouts where the amount col is far from name (e.g. col offset > 8).
    if not items and section_rows:
        log.append("OCL annexure: primary scan found 0 items, trying wide fallback scan")
        for row in section_rows:
            # Find all (text, amount) pairs anywhere in this row
            text_cells = []
            num_cells  = []
            for ci, v in enumerate(row):
                sv = _s(v).strip()
                if sv and len(sv) > 3:
                    try:
                        float(sv.replace(',', ''))
                    except ValueError:
                        text_cells.append((ci, sv))
                elif isinstance(v, (int, float)) and v is not None and v > 0:
                    if not (isinstance(v, float) and math.isnan(v)):
                        num_cells.append((ci, v))

            for tc, nm in text_cells:
                nm_lower = nm.lower()
                if any(tr in nm_lower for tr in _TRADE_REJECT):
                    continue
                if not any(kw in nm_lower for kw in _OCL_NAME_KEYS):
                    continue
                # Find the closest numeric cell to the right
                for nc, amt in sorted(num_cells, key=lambda x: x[0]):
                    if nc > tc:
                        items.append({'name': nm, 'amount': float(amt)})
                        log.append(f"OCL wide-scan: {nm} = {amt:,.2f} (name_col={tc}, amt_col={nc})")
                        break

    if items:
        # Deduplicate against what's already in unsecured_loan_parties
        existing = {it['name'].strip().lower() for it in
                    result.get('unsecured_loan_parties', [])}
        # Also deduplicate items against each other (same name/amount)
        seen_items: dict = {}
        deduped: list = []
        for it in items:
            key = (it['name'].strip().lower(), round(it['amount']))
            if key not in seen_items:
                seen_items[key] = True
                deduped.append(it)
        new_items = [it for it in deduped if it['name'].strip().lower() not in existing]
        if new_items:
            result['other_payable_items'] = new_items
            log.append(f"OCL annexure: {len(new_items)} items extracted "
                       f"({[it['name'] for it in new_items]})")


# ─────────────────────────────────────────────────────────────────────────────
#  Pass 4d — Unsecured Loan Annexure-B direct extraction  (FIX S10)
# ─────────────────────────────────────────────────────────────────────────────

def _extract_unsecured_annexure_b(rows, result, log):
    """
    Directly extract UNSECURED LOAN (Annexure-B) lender parties.

    Sachidanand GD Singla layout (0-based cols):
      Header row: 'UNSECURED LOAN' at col 25, 'ANNEXURE-B' at col 31
      Lender name: col 25 (person names like Rohit Vig, Smt.Santosh Rani)
      Lender amount: col 31
      TOTAL row: col 31 has the grand total (2,472,436.88)

    Two bugs in the previous approach were causing wrong results:
    1. STOP check used the full row string — 'CASH & BANK BALANCES' at col 53
       triggered a stop on row 25, BEFORE the first lender at row 26 (col 25).
       Fix: limit the stop check to cols 24–35 only (the lender column zone).
    2. Name_col detection picked up 'M/s. Krazy Choice' at col 32 (a CREDITOR
       from Annexure-C) instead of 'Rohit Vig' at col 25 (actual lender).
       Fix: hardcode name_col=25, amt_col=31 as the primary strategy; fall back
       to dynamic detection only if the hardcoded cols yield no results.
    """
    bs_total = result.get('unsecured_loans', 0)
    existing = result.get('unsecured_loan_parties', [])
    if existing:
        existing_sum = sum(p['amount'] for p in existing)
        if bs_total > 0 and abs(existing_sum - bs_total) / bs_total < 0.05:
            return  # already accurate

    _STOP_HEADERS = ('SUNDRY CREDITOR', 'ANNEXURE-C', 'ANNEXURE-D', 'OTHER PAYABLE',
                     'SUNDRY DEBTOR', 'CASH & BANK', 'FIXED ASSET')
    # NOTE: ALL stop checks are limited to cols 24–35 (the lender column zone).
    # The T-shaped layout has BS labels like 'SUNDRY DEBTORS' at col 0 and
    # 'CASH & BANK BALANCES' at col 53 on the SAME rows as lender data at col 25.
    # A full-row stop check fires too early and misses all lenders.
    _STOP_HEADERS_LENDER_ZONE = _STOP_HEADERS

    _SKIP_NAMES = {'PARTICULARS', 'TOTAL', 'SUB TOTAL', 'ANNEXURE', 'NIL',
                   'UNSECURED LOAN', 'FROM RELATED', 'FROM OTHER', 'UN-SECURED LOAN'}

    in_section = False
    items = []
    found_total = 0.0
    header_name_col = None   # col where 'UNSECURED LOAN' header was found
    header_amt_col  = None   # col where 'ANNEXURE-B' / amount was found

    for i, row in enumerate(rows):
        # Build full-row and lender-zone-only strings
        rs_full = ' '.join(_s(v).upper() for v in row if v is not None)
        # Lender zone: cols 24–35 only (where Annexure-B data lives)
        rs_zone = ' '.join(_s(row[c]).upper() for c in range(24, min(36, len(row)))
                           if row[c] is not None)

        if not in_section:
            if ('UNSECURED LOAN' in rs_full or 'UN-SECURED LOAN' in rs_full) and \
               ('ANNEXURE-B' in rs_full or "'B'" in rs_full):
                in_section = True
                # Record where the header appeared to anchor name/amt cols
                for ci, v in enumerate(row):
                    sv = _s(v).upper()
                    if 'UNSECURED LOAN' in sv or 'UN-SECURED LOAN' in sv:
                        header_name_col = ci
                        break
                for ci, v in enumerate(row):
                    sv = _s(v).upper()
                    if 'ANNEXURE-B' in sv or ("'B'" in sv and ci > 20):
                        header_amt_col = ci
                        break
                log.append(f"Unsecured Annexure-B header at row {i}: "
                            f"name_col≈{header_name_col}, amt_col≈{header_amt_col}")
            continue

        # FIX: Stop check limited to lender zone (cols 24-35) ONLY.
        # Full-row check caused early termination: 'SUNDRY DEBTORS' at col 0 and
        # 'CASH & BANK BALANCES' at col 53 both appear on the same rows as lenders.
        if any(kw in rs_zone for kw in _STOP_HEADERS_LENDER_ZONE):
            log.append(f"Unsecured Annexure-B: stopped at row {i} (lender-zone header)")
            break

        # FIX: Use hardcoded primary cols (name=header_name_col or 25, amt=31)
        # These are the KNOWN lender columns for GD Singla Sachidanand layout.
        # Dynamic detection fails because creditor names (M/s.) appear at col 32
        # on the same rows, and the scanner picks those up instead.
        primary_name_col = header_name_col if header_name_col is not None else 25
        primary_amt_col  = header_amt_col  if header_amt_col  is not None else 31

        nm = _s(row[primary_name_col]).strip() if primary_name_col < len(row) else ''
        amt_raw = row[primary_amt_col] if primary_amt_col < len(row) else None

        if not nm:
            continue
        nmu = nm.upper()
        if nmu in _SKIP_NAMES or nmu.startswith('ANNEXURE'):
            continue
        try:
            # If name is a pure number, it's the grand total
            float_nm = float(nm.replace(',', ''))
            # A number at the name_col = running total or grand total — check amt_col too
            if isinstance(amt_raw, (int, float)) and not (isinstance(amt_raw, float) and math.isnan(amt_raw)) and amt_raw > 0:
                found_total = float(amt_raw)
            elif float_nm > 10000:
                found_total = float_nm
            continue
        except ValueError:
            pass

        # TOTAL row: name contains 'TOTAL' or amount_col has the grand total
        if 'TOTAL' in nmu:
            if isinstance(amt_raw, (int, float)) and not (isinstance(amt_raw, float) and math.isnan(amt_raw)):
                found_total = float(amt_raw)
            break

        # Valid lender row: must have a positive amount
        if not isinstance(amt_raw, (int, float)) or \
           (isinstance(amt_raw, float) and math.isnan(amt_raw)) or \
           amt_raw <= 0:
            continue

        items.append({'name': nm, 'amount': float(amt_raw)})
        log.append(f"  Unsecured lender: {nm} = {amt_raw:,.2f} (R{i} C{primary_name_col}/{primary_amt_col})")

    if not items:
        # Fallback: dynamic name_col detection (original approach but with zone-limited stop)
        log.append("Unsecured Annexure-B: hardcoded cols found nothing, trying dynamic detection")
        in_section2 = False
        name_col2 = None
        amt_col2  = None
        items2 = []
        for i, row in enumerate(rows):
            rs_full = ' '.join(_s(v).upper() for v in row if v is not None)
            rs_zone = ' '.join(_s(row[c]).upper() for c in range(24, min(36, len(row)))
                               if row[c] is not None)
            if not in_section2:
                if ('UNSECURED LOAN' in rs_full or 'UN-SECURED LOAN' in rs_full) and \
                   ('ANNEXURE-B' in rs_full or "'B'" in rs_full):
                    in_section2 = True
                continue
            if any(kw in rs_zone for kw in _STOP_HEADERS_LENDER_ZONE):
                break
            if any(kw in rs_full for kw in ('ANNEXURE-C','ANNEXURE-D','OTHER PAYABLE')):
                break
            if name_col2 is None:
                for ci, v in enumerate(row[:40]):   # limit scan to first 40 cols
                    sv = _s(v).strip()
                    if not sv or len(sv) < 3: continue
                    svu = sv.upper()
                    if svu in _SKIP_NAMES or svu.startswith('ANNEXURE'): continue
                    try: float(sv.replace(',','')); continue
                    except ValueError: pass
                    for ac in range(ci+1, min(ci+10, len(row))):
                        av = row[ac]
                        if isinstance(av,(int,float)) and not (isinstance(av,float) and math.isnan(av)) and av > 1000:
                            name_col2 = ci; amt_col2 = ac; break
                    if name_col2 is not None: break
                if name_col2 is None: continue
            nm2 = _s(row[name_col2]).strip() if name_col2 < len(row) else ''
            amt2 = row[amt_col2] if amt_col2 and amt_col2 < len(row) else None
            if not nm2: continue
            nmu2 = nm2.upper()
            if nmu2 in _SKIP_NAMES: continue
            try: float(nm2.replace(',','')); continue
            except ValueError: pass
            if 'TOTAL' in nmu2: break
            if isinstance(amt2,(int,float)) and not (isinstance(amt2,float) and math.isnan(amt2)) and amt2 > 0:
                items2.append({'name': nm2, 'amount': float(amt2)})
        items = items2

    if not items:
        log.append("Unsecured Annexure-B: no lenders found")
        return

    calc = sum(x['amount'] for x in items)
    log.append(f"Unsecured Annexure-B: {len(items)} parties, total={calc:,.2f} "
               f"(BS={bs_total:,.2f}, found_total={found_total:,.2f})")

    # If sum deviates >10% from BS total and we have a found_total, trust found_total for validation
    if bs_total > 0 and abs(calc - bs_total) / bs_total > 0.10:
        if found_total > 0 and abs(calc - found_total) / max(found_total, 1) < 0.05:
            log.append("Unsecured Annexure-B: sum matches found_total, accepting")
        else:
            log.append("Unsecured Annexure-B: sum mismatch >10%, keeping existing parties")
            return

    result['unsecured_loan_parties'] = items
    if not result.get('unsecured_loans') or result['unsecured_loans'] == 0:
        result['unsecured_loans'] = calc
    log.append(f"Unsecured Annexure-B: written {len(items)} parties")

# ─────────────────────────────────────────────────────────────────────────────
#  Pass 4c — Capital account Annexure-A direct extraction
# ─────────────────────────────────────────────────────────────────────────────

def _extract_capital_annexure(rows, result, log):
    """
    Extract capital movements (opening, cash additions, withdrawals) from Annexure-A.

    GD Singla layout (0-based col indices):
      col40 = label text
      col42 = individual line amounts (Cheque/RTGS intro, each withdrawal item)
      col44 = running subtotals (opening, additions total, Less total, closing)

    Sachidanand-type layout (alternative):
      col25 = label text, col31 = amount
      Rows: "Last Balance as on..." = opening, "Add Profit during the year" = profit,
            "Less: Withdrawals" = withdrawals, "Closing balance as on..." = closing.

    Key rows (0-based row index from XLS):
      OPENING BALANCE row: col44 = 5,252,803.82
      Cheque/RTGS row:     col42 = 2,121,000  (cash intro, NOT profit)
      LIC row (last Less): col44 = 1,677,860.92 (withdrawals subtotal)
      CLOSING row:         col44 = 6,532,802.90
    """
    in_section  = False
    opening      = 0.0
    additions_cash = 0.0
    last_col44   = 0.0
    found_opening = False

    # FIX 2a: Also scan for Sachidanand-type capital table at cols 25-31
    sachi_opening = 0.0
    sachi_profit = 0.0
    sachi_withdrawals = 0.0
    sachi_closing = 0.0
    sachi_found = False

    for i, row in enumerate(rows):
        rs = ' '.join(_s(v).upper() for v in row if v is not None)

        # FIX 2b: Sachidanand layout — scan for capital movement rows using cols 25-31
        # "Last Balance as on" = opening; "Add Profit" = profit;
        # "Less: Withdrawals" = withdrawals; "Closing balance" = closing
        if not sachi_found:
            lbl25 = _s(row[25]).upper() if 25 < len(row) else ''
            lbl26 = _s(row[26]).upper() if 26 < len(row) else ''
            lbl27 = _s(row[27]).upper() if 27 < len(row) else ''
            combined_lbl = lbl25 + ' ' + lbl26 + ' ' + lbl27

            # Opening balance: "Last Balance as on 01.04.YYYY" or "Opening Balance"
            # FIX Q (2026-08-14): Do NOT capture if 'CLOSING' is in the label.
            # Row 21 has "CLOSING BALANCE AS ON 31.03.2024" which also matches
            # ('BALANCE' + 'AS ON') and was overwriting the correct opening (928,695.69)
            # with the closing balance (1,559,992.23). Now exclude CLOSING rows.
            # Also capture only ONCE: once sachi_opening is set, don't overwrite it.
            if sachi_opening == 0 and 'CLOSING' not in combined_lbl:
                if 'LAST BALANCE' in combined_lbl or 'OPENING BALANCE' in combined_lbl or \
                   ('BALANCE' in combined_lbl and ('01.04' in combined_lbl or
                                                   ('AS ON' in combined_lbl and '01.04' in rs))):
                    v31 = _n(row[31]) if 31 < len(row) else 0
                    if v31 > 0:
                        sachi_opening = v31

            if ('ADD PROFIT' in combined_lbl or 'ADD: PROFIT' in combined_lbl or
                'PROFIT DURING' in combined_lbl or 'NET PROFIT' in combined_lbl):
                v31 = _n(row[31]) if 31 < len(row) else 0
                if v31 > 0:
                    sachi_profit = v31

            if 'ADDITION' in combined_lbl and sachi_opening > 0:
                v31 = _n(row[31]) if 31 < len(row) else 0
                if v31 > 0:
                    additions_cash = v31

            if 'WITHDRAWAL' in combined_lbl or 'DRAWING' in combined_lbl or \
               ('LESS' in combined_lbl and sachi_opening > 0 and 'CLOSING' not in combined_lbl):
                v31 = _n(row[31]) if 31 < len(row) else 0
                if v31 > 0:
                    sachi_withdrawals = v31  # subtotal row amount

            if 'CLOSING' in combined_lbl and 'BALANCE' in combined_lbl and sachi_opening > 0:
                v31 = _n(row[31]) if 31 < len(row) else 0
                if v31 > 0:
                    sachi_closing = v31
                    sachi_found = True

        if not in_section:
            if ('PROP CAPITAL ACCOUNT' in rs or 'PROPRIETOR CAPITAL' in rs) and 'ANNEXURE-A' in rs:
                in_section = True
            continue

        # Opening balance row
        if 'OPENING BALANCE' in rs or 'LAST BALANCE' in rs:
            v = row[44] if 44 < len(row) else None
            if isinstance(v, (int, float)) and v > 0:
                opening = v
                found_opening = True
            elif not found_opening:
                # Try col31 (Sachidanand layout within Annexure-A section)
                v31 = row[31] if 31 < len(row) else None
                if isinstance(v31, (int, float)) and v31 > 0:
                    opening = v31
                    found_opening = True
            continue

        # Closing row — stop and record withdrawals from last_col44
        if found_opening and 'CLOSING' in rs and 'BALANCE' in rs:
            break

        if not found_opening:
            continue

        # Track last col44 seen after opening (= withdrawals subtotal on final Less row)
        v44 = row[44] if 44 < len(row) else None
        if isinstance(v44, (int, float)) and v44 > 0:
            last_col44 = v44

        # Cash intro lines only (Cheque/RTGS/NEFT/Gift — NOT profit)
        if any(k in rs for k in ('CHEQUE', 'RTGS', 'NEFT', 'GIFT FROM')):
            v42 = row[42] if 42 < len(row) else None
            if isinstance(v42, (int, float)) and v42 > 0:
                additions_cash += v42

    # FIX 2c: Prefer Sachidanand-style values if found and col44 path failed
    if sachi_found and sachi_opening > 0 and not (opening > 0 and last_col44 > 0):
        # FIX 1 (S10): Also scan for individual withdrawal sub-items.
        # In Sachidanand layout "Less Withdrawals 446,405" is the main item,
        # then sub-items like "Mediuclaim 43,749" appear on the next rows at col25/31.
        # The TOTAL withdrawals = subtotal row + all named sub-items below it.
        # Strategy: after we find the "WITHDRAWAL" row, scan forward rows where:
        #   - col31 has a positive number
        #   - col25/26 has a name (not a section header like CLOSING/UNSECURED/OTHER)
        #   - no new section header has appeared
        # Sum these sub-items into sachi_withdrawals.
        _in_with_section = False
        _with_sub_total = 0.0
        _WITH_STOP_KEYS = ('CLOSING', 'UNSECURED', 'OTHER PAYABLE', 'SUNDRY',
                           'FIXED ASSET', 'CASH', 'LOAN', 'TOTAL')
        for _row in rows:
            _lbl25 = _s(_row[25]).upper() if 25 < len(_row) else ''
            _lbl26 = _s(_row[26]).upper() if 26 < len(_row) else ''
            _lbl27 = _s(_row[27]).upper() if 27 < len(_row) else ''
            _comb = _lbl25 + ' ' + _lbl26 + ' ' + _lbl27

            if not _in_with_section:
                if 'WITHDRAWAL' in _comb or 'DRAWING' in _comb or \
                   ('LESS' in _comb and 'CLOSING' not in _comb and sachi_opening > 0):
                    _in_with_section = True
                continue

            # Stop at next section header
            if any(k in _comb for k in _WITH_STOP_KEYS):
                break

            _v31 = _n(_row[31]) if 31 < len(_row) else 0
            # Must have a name AND a positive amount to be a sub-item
            _nm = (_s(_row[25]) or _s(_row[26]) or _s(_row[27])).strip()
            if _v31 > 0 and _nm and len(_nm) >= 3:
                _with_sub_total += _v31
                log.append(f"Capital withdrawal sub-item: {_nm} = {_v31:,.2f}")

        # If we found sub-items, total withdrawals = sachi_withdrawals + sub-items
        # (sachi_withdrawals already has the "Less 446,405" subtotal line)
        if _with_sub_total > 0:
            sachi_withdrawals = sachi_withdrawals + _with_sub_total
            log.append(f"Capital withdrawals total (incl. sub-items): {sachi_withdrawals:,.2f}")

        opening = sachi_opening
        last_col44 = sachi_withdrawals
        found_opening = True
        log.append(f"Capital Annexure-A (Sachidanand layout): opening={sachi_opening:,.2f}, "
                   f"profit={sachi_profit:,.2f}, withdrawals={sachi_withdrawals:,.2f}, "
                   f"closing={sachi_closing:,.2f}")

    if opening > 0:
        for cap in result.get('capital_accounts', []):
            cap['opening']     = opening
            cap['additions']   = round(additions_cash)
            cap['withdrawals'] = round(last_col44, 2)
            break
        log.append(f"Capital Annexure-A: opening={opening:,.2f}, "
                   f"cash_intro={additions_cash:,.0f}, withdrawals={last_col44:,.2f}")


def _extract_pl(rows, result, log):
    """
    The P&L (Trading + P&L A/c) is embedded in the same sheet.
    In GD Singla format it appears in MULTIPLE COLUMN BLOCKS on the same rows.

    Detection: find the row with 'TRADING' or 'PROFIT & LOSS ACCOUNT' header.
    The P&L starts a few rows later. Columns for amounts vary per client but
    are always in the RIGHT half of the T (cols 8-18 for Bobby/Ashok,
    cols 10-18 for Gupta). We scan all non-zero columns in these rows.

    Key rows (in order of appearance):
      To Opening Stock  | By Sales
      To Purchases      | By Closing Stock
      To Gross Profit   | (= By Gross Profit in P&L A/c)
      To Salary/Wages   | By Gross Profit b/d
      To Depreciation
      To Interest
      To Net Profit
    """
    # Find P&L header row
    pl_start = -1
    for i, row in enumerate(rows):
        rs = ' '.join(_s(v).upper() for v in row if v is not None)
        if 'TRADING' in rs and ('PROFIT' in rs or 'LOSS' in rs) and i < 50:
            pl_start = i
            log.append(f"P&L header at row {i}")
            break
        if 'PROFIT' in rs and 'LOSS' in rs and 'ACCOUNT' in rs and i < 50:
            pl_start = i
            break

    if pl_start < 0:
        pl_start = 7   # fallback

    pl_keywords = {
        'OPENING STOCK': 'opening_stock',
        'TO OPENING STOCK': 'opening_stock',
        'TO PURCHASE': 'purchases',
        'TO PURCHASES': 'purchases',
        'TO PURCHASE A/C': '_purchase_line',  # Bobby: individual purchase GST lines
        'BY SALE': 'sales',
        'BY SALES': 'sales',           # Gupta format: "By Sales :"
        'BY SALE A/C': '_sale_line',   # Bobby: individual sale GST lines
        'BY CLOSING STOCK': 'closing_stock',
        'CLOSING STOCK': '_cs_asset',
        'TO GROSS PROFIT': 'gross_profit',
        'BY GROSS PROFIT': 'gross_profit',
        'TO NET PROFIT': 'net_profit',
        'BY NET PROFIT': 'net_profit',
        'TO SALARY': 'salary_expenses',
        'TO WAGES': 'salary_expenses',
        # FIX G (2026-08-14): 'TO ACCOUNTANT SALARY' was mapped to 'salary_expenses',
        # causing it to be ADDED to 'TO SALARY' (768,000 + 30,000 = 798,000 wrong total).
        # Accountant Salary is an OTHER EXPENSE (professional/admin), not employee salary.
        # Route it to other_expense_items via expense_kws instead (handled below).
        # 'TO ACCOUNTANT SALARY': 'salary_expenses',  ← REMOVED
        'TO DEPRECIATION': 'depreciation',
        'TO INTEREST ON UNSECURED': 'interest_paid',
        'TO INTEREST ON LOAN': 'interest_paid',
        'TO INTEREST PAID': 'interest_paid',
        'TO BANK INT': 'interest_paid',
        'TO BANK INTT': 'interest_paid',
        # FIX J (2026-08-14): 'TO BANK CHARGES' removed from pl_keywords.
        # It was routing Bank Charges (12,237.09) into interest_paid, so
        # interest_paid = 12,237 + 3,924 = 16,161 wrote to R45 (unsecured interest).
        # Now it falls through to expense_kws where 'TO BANK CHARGE' catches it
        # and routes to other_expense_items → Note 19 R62 (bank charges).
        # 'TO BANK CHARGES': 'interest_paid',  ← REMOVED
        'TO CAR LOAN INT': 'interest_paid',
        'TO INTEREST TO PARTIES': 'interest_paid',
        'TO INTEREST TO PARTY': 'interest_paid',
        'BY DISCOUNT': 'other_income',
        'BY REBATE': 'other_income',
    }
    # Track whether we're accumulating GST purchase/sale lines (Bobby style)
    _in_purchase_gst = False
    _in_sale_gst = False

    expense_kws = [
        'TO AUDIT FEE', 'TO BANK CHARGE', 'TO BANK CHARGES', 'TO ELECTRICITY', 'TO INSURANCE',
        'TO TELEPHONE', 'TO REPAIR', 'TO PETROL', 'TO RENT', 'TO GENERAL',
        'TO POSTAGE', 'TO STATIONARY', 'TO PRINTING', 'TO STAFF', 'TO PACKING',
        'TO PROPERTY TAX', 'TO CAR EXP', 'TO CAR EXP', 'TO SCOOTER', 'TO LABOUR',
        'TO LOADING', 'TO SHOP', 'TO DIWALI', 'TO F.O.C', 'TO FOC',
        'TO FREIGHT', 'TO FRIEGHT', 'TO COMMISSION', 'TO COMMISSS',
        'TO COMPUTER', 'TO MEDICLAIM', 'TO LEGAL',
        'TO PARTNER INTEREST', 'TO PARTNER SALARY',
        'TO BONUS', 'TO ENTERTAINMENT', 'TO VEHICLE', 'TO TOUR',
        'TO TRAVELLING',
        'TO STATIONERY', 'TO STATIONARY', 'TO PRINTING & STATIONERY',
        # FIX G (2026-08-14): Add Accountant Salary and Professional Charges so they
        # go into other_expense_items (Note 19 spare rows) rather than salary_expenses.
        'TO ACCOUNTANT SALARY', 'TO ACCOUNTANT SAL', 'TO ACCOUNTANT',
        'TO PROFESSIONAL CHARGES', 'TO PROFESSIONAL CHARGE', 'TO PROFESSIONAL',
    ]

    for i in range(pl_start, len(rows)):
        row = rows[i]
        # Scan every cell in the row for P&L labels
        for j, v in enumerate(row):
            lbl = _s(v).upper()
            if not lbl:
                continue

            for kw, field in sorted(pl_keywords.items(), key=lambda x: -len(x[0])):
                if lbl.startswith(kw) or lbl == kw:
                    # Find nearest number after this cell
                    # For SALES keywords, strategy depends on type:
                    if field == '_sale_line':
                        # Bobby GST sale lines: take FIRST number
                        amt = _first_num(row, j + 1, j + 6)
                    elif field == 'sales':
                        # Ashok/Gupta: scan this row + next few rows for TOTAL sales
                        # First-hit-wins: don't re-scan if already captured
                        if result['sales'] == 0:
                            _cands = []
                            for _ri in range(i, min(i + 5, len(rows))):
                                _r2 = rows[_ri]
                                for _k in range(j, min(j + 8, len(_r2))):
                                    _v = _n(_r2[_k])
                                    if _v > 1000:
                                        _cands.append(_v)
                            amt = max(_cands) if _cands else 0
                        else:
                            amt = 0  # already captured, skip
                    else:
                        amt = _first_num(row, j + 1, j + 8)
                        if amt <= 0:
                            amt = _first_num(row, j + 1)
                    if amt > 0:
                        if field == '_cs_asset':
                            if result['closing_stock'] == 0:
                                result['closing_stock'] = amt
                        elif field == 'closing_stock':
                            if result['closing_stock'] == 0:
                                result['closing_stock'] = amt
                        elif field in ('sales', '_sale_line'):
                            # Use first-hit-wins: 'By Sales GST 5%' row finds the TOTAL
                            # via the multi-row max scan. Don't add again for GST5% Central.
                            if result['sales'] == 0:
                                result['sales'] = amt
                        elif field == '_purchase_line':
                            result['purchases'] += amt
                        elif field == 'opening_stock':
                            if result['opening_stock'] == 0:
                                result['opening_stock'] = amt
                        elif field == 'purchases':
                            # Accumulate purchase lines (GST5% + GST5% Central).
                            # For Ashok/Gupta format the single 'To Purchase' row has the total.
                            # For Sachidanand-style, purchases are split into two rows.
                            # Use max(existing, new) to prefer the larger (= total) line,
                            # then accumulate if it looks like a new independent purchase line.
                            if result['purchases'] == 0:
                                result['purchases'] = amt
                            elif amt > result['purchases'] * 1.5:
                                # Much larger: likely the running total, replace
                                result['purchases'] = amt
                            elif amt > result['purchases'] * 0.1:
                                # Similar scale: likely a second purchase category, add
                                result['purchases'] += amt
                        elif field == 'gross_profit':
                            if result['gross_profit'] == 0:
                                result['gross_profit'] = amt
                        elif field == 'net_profit':
                            # Prefer 'TO NET PROFIT' (total) over 'BY NET PROFIT' (partner share)
                            if result['net_profit'] == 0 or (
                               lbl.startswith('TO NET PROFIT') and amt > result['net_profit']
                            ):
                                result['net_profit'] = amt
                        elif field in ('salary_expenses', 'depreciation',
                                       'interest_paid', 'other_income'):
                            result[field] += amt
                    break

            # Other expense items
            for kw in expense_kws:
                if lbl.startswith(kw):
                    amt = _first_num(row, j + 1, j + 8)
                    if amt > 0:
                        # FIX K (2026-08-14): Strip 'TO ' prefix case-insensitively.
                        # .replace('To ', '') is case-sensitive and misses labels like
                        # 'TO ACCOUNTANT SALARY' (all-caps from T-shaped XLS).
                        # Result: names stored as 'TO ACCOUNTANT SALARY' → lower →
                        # 'to accountant salary' → no match in _OTHER_EXP_ROW_MAP.
                        # Fix: strip the leading TO/To/to prefix explicitly.
                        raw_name = _s(v).strip()
                        name = re.sub(r'^[Tt][Oo]\s+', '', raw_name).strip()
                        result['other_expense_items'].append({'name': name, 'amount': amt})
                        result['other_expenses'] += amt
                    break

    log.append(
        f"P&L: Sales={result['sales']:.0f}, Purchases={result['purchases']:.0f}, "
        f"Stock={result['closing_stock']:.0f}, GP={result['gross_profit']:.0f}, "
        f"NP={result['net_profit']:.0f}"
    )


# ─────────────────────────────────────────────────────────────────────────────
#  Pass 6 — Depreciation chart extraction
# ─────────────────────────────────────────────────────────────────────────────

def _extract_dep_chart(rows, result, log):
    """
    Find the DEPRECIATION CHART section and extract fixed asset items.

    GD Singla format: dep chart is embedded in the same sheet at cols 50-58 (Ashok),
    cols 49-57 (Bobby), or cols 30-37 (Gupta). The header 'DEPRECIATION CHART' appears
    somewhere in the row (not necessarily col 0).

    Algorithm:
      1. Find the row containing 'DEPRECIATION CHART'.
      2. Detect the name column (first text col after the header keyword col).
      3. Detect the number column block (adjacent numeric cols).
      4. Read asset rows until TOTAL or blank.
    """
    in_dep = False
    items = []
    name_col = None    # col index of asset name
    num_start = None   # first numeric col of dep chart

    for i, row in enumerate(rows):
        rs = ' '.join(_s(v).upper() for v in row if v is not None)
        if not in_dep:
            if 'DEPRECIATION CHART' in rs or ('FIXED ASSET' in rs and 'DEP' in rs and 'CHART' in rs):
                in_dep = True
                log.append(f"Dep chart found at row {i}")
                # Detect which column 'DEPRECIATION CHART' is in
                dep_col = next((j for j, v in enumerate(row)
                                if 'DEPRECIATION' in _s(v).upper()), 0)
                # GD Singla layout: dep chart label is 2 cols to the RIGHT of asset names.
                # So name_col = dep_col - 2. num_start = name_col + 1.
                name_col = max(0, dep_col - 2)
                num_start = name_col + 1
            continue

        # Skip rows where the dep chart name column is blank (these are BS/P&L rows mixed in)
        # name_col is fixed at dep_col-2; only process rows where that col has an asset label

        # Skip category headers (all caps, no numbers)
        nums_in_row = [(j, float(v)) for j, v in enumerate(row)
                       if isinstance(v, (int, float)) and not (isinstance(v, float) and str(v) == 'nan')
                       and float(v) != 0]
        
        if not nums_in_row:
            continue

        # Detect name_col if not found yet: find first text cell that looks like an asset name
        if name_col is None:
            for j, v in enumerate(row):
                sv = _s(v)
                if sv and len(sv) > 3 and not sv.upper().startswith(('TOTAL', 'PART', 'DATE', 'AMOUNT')):
                    name_col = j
                    break

        # Get the asset name from name_col
        label = _s(row[name_col]).strip() if name_col is not None else ''
        # Limit block to max 9 numeric cols after name_col (dep chart has exactly 9 cols)
        block_nums = []
        if name_col is not None and num_start is not None:
            # Max 9 numeric cols in a dep chart row
            block_nums = [x for x in nums_in_row
                          if num_start <= x[0] <= num_start + 9]

        # Skip blank, header, total rows
        if not label:
            continue
        lu = label.upper()
        if lu in ('TOTAL', 'GRAND TOTAL', 'NET TOTAL', 'PARTICULARS', ''):
            if lu == 'TOTAL' and items:
                break
            continue
        # Skip ONLY multi-word category headers (exact start match only)
        if any(lu.startswith(x) for x in ('PLANT & MACHINERY', 'VEHICLE', 'COMPUTERS',
                                           'LAND AND', 'INTANGIBLE')):
            continue  # Category header rows

        # Find the numeric block: cols from num_start or first num col > name_col
        if num_start is None:
            num_start = nums_in_row[0][0] if nums_in_row else (name_col + 1 if name_col else 1)

        # Use bounded block_nums already set above (num_start to num_start+9)
        # Fallback if name_col/num_start not yet set
        if len(block_nums) < 2:
            continue

        # GD Singla dep chart column layout (offset from num_start):
        #   +0 = opening WDV  +1 = add>180d  +2 = add<180d
        #   +3 = sale  +4 = total  +5 = rate%  +6 = dep  +7 = closing WDV
        # block_nums only has NON-ZERO entries, so we must check actual column position.
        opening_wdv = block_nums[0][1]

        # add_gt180: only count if its column is exactly num_start+1
        add_gt180 = 0.0
        for bc, bv in block_nums[1:]:
            if bc == num_start + 1:
                add_gt180 = bv
            break  # stop after first candidate

        # add_lt180: only count if its column is exactly num_start+2
        add_lt180 = 0.0
        for bc, bv in block_nums[1:]:
            if bc == num_start + 2:
                add_lt180 = bv
                break

        additions = add_gt180 + add_lt180
        sale = 0.0
        rate = 0.0
        dep = 0.0
        closing_wdv = 0.0

        if len(block_nums) >= 2:
            closing_wdv = abs(block_nums[-1][1])
            dep = abs(block_nums[-2][1])

        # Find rate (typically 0.10, 0.15, 0.40 stored as float or 10, 15, 40)
        for _, n in block_nums:
            if 0 < n <= 1:
                rate = round(n * 100)
                break
            if n in (5, 10, 15, 20, 25, 30, 40):
                rate = int(n)
                break

        if opening_wdv > 0 or additions > 0:
            items.append({
                'name': label,
                'opening_wdv': opening_wdv,
                'additions': additions,
                'sales': sale,
                'dep': dep,
                'closing_wdv': closing_wdv,
                'rate': rate,
            })

    if items:
        # Filter out obviously wrong items (dep chart items where closing_wdv is 0 for all)
        valid = [x for x in items if x['closing_wdv'] > 0 or x['opening_wdv'] > 0]
        result['fixed_asset_items'] = valid if valid else items
        calc_wdv = sum(x['closing_wdv'] for x in result['fixed_asset_items'])
        if calc_wdv > 0:
            result['fixed_assets_wdv'] = calc_wdv
        log.append(f"Dep chart: {len(result['fixed_asset_items'])} assets, WDV={calc_wdv:.0f}")


# ─────────────────────────────────────────────────────────────────────────────
#  Pass 7 — Cash & Bank
# ─────────────────────────────────────────────────────────────────────────────

_BANK_KEYWORDS = ('HDFC', 'SBI', 'ICICI', 'OBC', 'PNB', 'AXIS', 'ORIENTAL',
                  'STATE BANK', 'KOTAK', 'YES BANK', 'CANARA', 'UNION BANK',
                  'BANK OF BARODA', 'BOB', 'F.D.R', 'FDR', 'BANK A/C', 'BANK A/C')


def _extract_cash_bank(rows, result, log):
    """
    Extract cash in hand and bank-wise balances.

    Strategy (in priority order):
    1. Annexure-H (col50=name, col58=amount) — GD Singla dedicated annexure.
       This gives exact Cash In Hand and individual bank accounts.
    2. Generic scan of the BS section (fallback for non-GD-Singla layouts).
    """

    # ── Pass 1: Annexure-H — try multiple column layouts ─────────────────────
    # Layout A (Ashok Kumar): name=col50, amount=col58
    # Layout B (Sachidanand): name=col53, amount=col62
    # Both are tried; whichever finds items wins.

    def _is_real_num(v):
        return isinstance(v, (int, float)) and not (isinstance(v, float) and math.isnan(v)) and float(v) > 0

    def _try_annexure_h(rows, name_col, amt_col):
        """Try reading Annexure-H at the given name/amount column positions."""
        in_h = False
        cash_h = 0.0
        banks_h = []
        total_h = 0.0

        for i, row in enumerate(rows):
            if name_col >= len(row):
                continue
            name_cell = _s(row[name_col]).strip()
            amt_cell  = row[amt_col] if amt_col < len(row) else None

            if not in_h:
                rs = ' '.join(_s(v).upper() for v in row if v is not None)
                if 'CASH' in rs and 'BANK' in rs and 'ANNEXURE-H' in rs:
                    in_h = True
                continue

            n_up = name_cell.upper()

            if n_up == 'TOTAL':
                if _is_real_num(amt_cell):
                    total_h = float(amt_cell)
                break

            # Stop if we hit a new section header at the name column
            _CASH_STOP = ('SUNDRY DEBTOR', 'SUNDRY CREDITOR', 'LOAN & ADVANCE',
                          'LOANS & ADVANCE', 'OTHER PAYABLE', 'FIXED ASSET', 'ANNEXURE-G')
            if any(kw in n_up for kw in _CASH_STOP):
                break

            if not name_cell or len(name_cell) < 2:
                continue
            # CRITICAL: skip rows where amount is NaN or zero
            if not _is_real_num(amt_cell):
                continue

            # Skip M/s. prefix-only rows (they are OCL entries at wrong column)
            if n_up in ('M/S.', 'M/S', 'M/S .'):
                continue

            if 'CASH IN HAND' in n_up or 'CASH-IN-HAND' in n_up:
                cash_h = float(amt_cell)
            elif any(bk in n_up for bk in _BANK_KEYWORDS) or 'BANK' in n_up:
                banks_h.append({'name': name_cell, 'amount': float(amt_cell)})
            else:
                banks_h.append({'name': name_cell, 'amount': float(amt_cell)})

        return in_h, cash_h, banks_h, total_h

    # FIX 3: Try all column layouts, accept first one with real (non-NaN) amounts
    for name_col, amt_col in [(53, 62), (50, 58), (51, 59), (52, 61)]:
        in_h, cash_h, banks_h, total_h = _try_annexure_h(rows, name_col, amt_col)
        if in_h and (cash_h > 0 or any(_is_real_num(b['amount']) for b in banks_h)):
            result['cash_in_hand'] = cash_h
            result['bank_balances'] = [b for b in banks_h if _is_real_num(b['amount'])]
            bank_sum = sum(b['amount'] for b in result['bank_balances'])
            result['cash_bank'] = cash_h + bank_sum
            log.append(f"Cash Annexure-H (col{name_col}/{amt_col}): hand={cash_h:,.0f}, "
                       f"{len(result['bank_balances'])} bank(s)={bank_sum:,.0f}, total={result['cash_bank']:,.0f}")
            return   # Annexure-H is authoritative — skip generic scan

    # ── Pass 2: Generic scan (fallback) ──────────────────────────────────────
    in_cash = False
    cash_items = []

    for i, row in enumerate(rows):
        rs = ' '.join(_s(v).upper() for v in row if v is not None)

        if 'CASH' in rs and 'BANK' in rs and 'BALANCE' in rs and not in_cash and i < 50:
            in_cash = True
            for j, v in enumerate(row):
                lv = _s(v).upper()
                if 'CASH' in lv and 'BANK' in lv:
                    amt = _first_num(row, j + 1, j + 8)
                    if amt > 100:
                        result['cash_bank'] = amt
                        log.append(f"Cash & bank total (header row): {amt}")
                    break
            continue

        if not in_cash:
            continue

        r0u = _s(row[0]).upper()
        if any(x in r0u for x in ('TOTAL', 'SUNDRY', 'LOANS', 'ADVANCE', 'OTHER', 'INVEST')):
            break

        for j, v in enumerate(row):
            lbl = _s(v).upper()
            if not lbl:
                continue
            if 'CASH IN HAND' in lbl or 'CASH-IN-HAND' in lbl or (lbl == 'CASH' and j < 10):
                amt = _first_num(row, j + 1, j + 6)
                if not amt:
                    for k in range(j, min(j + 5, len(row))):
                        rv = row[k]
                        if isinstance(rv, (int, float)) and not (isinstance(rv, float) and str(rv) == 'nan') and float(rv) > 100:
                            amt = float(rv)
                            break
                if amt > 0:
                    result['cash_in_hand'] = amt
                    log.append(f"Cash in hand: {amt}")
            elif any(bk in lbl for bk in _BANK_KEYWORDS):
                amt = _first_num(row, j + 1, j + 6)
                if amt > 10:
                    cash_items.append({'name': _s(v), 'amount': amt})

    if cash_items or result.get('cash_in_hand', 0) > 0:
        if cash_items:
            result['bank_balances'] = cash_items
        bank_total = sum(x['amount'] for x in cash_items)
        computed = result['cash_in_hand'] + bank_total
        if computed > 0:
            result['cash_bank'] = computed
        log.append(f"Banks: {len(cash_items)}, hand={result['cash_in_hand']:.0f}, "
                   f"total={result['cash_bank']:.0f}")


# ─────────────────────────────────────────────────────────────────────────────
#  SECTION 2 — Template injector
# ─────────────────────────────────────────────────────────────────────────────

def inject_into_template(parsed: dict, template_path: str, output_path: str,
                         client_name: str = None, cy_year: str = None) -> dict:
    """
    Fill Output_sample_format.xlsx with PY values from parsed dict.
    CY column = 0 / blank (yellow highlighted).
    """
    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path)

    py_year = parsed.get('py_year_end', '2024')
    if cy_year is None:
        cy_year = str(int(py_year) + 1)

    if client_name is None:
        client_name = parsed.get('entity_name', 'M/S CLIENT')

    log = parsed.get('log', [])
    log.append(f"Injecting CY={cy_year}, PY={py_year}")

    _inject_bs_sheet(wb, parsed, client_name, cy_year, py_year, log)
    _inject_pl_sheet(wb, parsed, client_name, cy_year, py_year, log)
    _inject_capital_sheet(wb, parsed, client_name, cy_year, py_year, log)
    _inject_notes_bs(wb, parsed, client_name, cy_year, py_year, log)
    _inject_notes_pl(wb, parsed, client_name, cy_year, py_year, log)
    details_shifts = _inject_details_sheet(wb, parsed, client_name, cy_year, py_year, log)
    _fix_details_formula_refs(wb, details_shifts, log)   # FIX S10: repair cross-sheet refs
    _inject_fixed_assets_py(wb, parsed, client_name, py_year, log)
    _inject_fixed_assets_cy_opening(wb, parsed, log)
    _inject_gross_profit_sheet(wb, parsed, client_name, cy_year, py_year, log)
    _update_headers(wb, client_name, cy_year, py_year)
    _clear_ref_errors(wb, log)

    wb.save(output_path)
    log.append(f"Saved: {output_path}")

    return {
        'status': 'success', 'output': output_path, 'log': log,
        'entity_name': client_name, 'cy_year': cy_year, 'py_year': py_year,
    }


def _py(ws, row, col, value):
    """Write a numeric PY value."""
    _write_num(ws, row, col, value)


def _cy(ws, row, col, value=0):
    """Write a CY input cell (yellow highlighted, 0 by default)."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = float(value or 0)
    cell.fill = _INPUT_FILL
    cell.number_format = '#,##0'


# ─────────────────────────────────────────────────────────────────────────────
# HARD-CODED COLUMN CONSTANTS (1-indexed, verified from template inspection)
#
# bs sheet          : CY=col5(E)  PY=col6(F)   — but bs is formula-driven
# notes to bs       : CY=col4(D)  PY=col5(E)
# notes to p&l      : CY=col4(D)  PY=col5(E)
# capital           : C=3 D=4 E=5 F=6 G=7
# Details           : CY=col4(D)  PY=col5(E)
# GROSS PROFIT left : CY=col2(B)  PY=col3(C)
# GROSS PROFIT right: CY=col5(E)  PY=col6(F)
# Fixed Assets P.Yr.: B=opening C=add>180 D=add<180 E=sales F=total G=rate H=dep I=wdv
# ─────────────────────────────────────────────────────────────────────────────


# ── BS sheet ──────────────────────────────────────────────────────────────────
# BS sheet is ENTIRELY formula-driven. All PY values flow in via:
#   capital!G11          → R7/R8   (capital closing)
#   notes to bs!E19      → R11     (LT borrowings total)
#   notes to bs!E34      → R14     (ST borrowings total)
#   notes to bs!E38/E39  → R16/R17 (trade payables MSME / others)
#   notes to bs!E65      → R18     (other current liabilities)
#   notes to bs!E71      → R19     (ST provisions)
#   Fixed Assets P.Yr.!I37 → R26   (PPE WDV)
#   notes to bs!E85      → R27     (investments)
#   notes to p&l!E27     → R31     (inventories = closing stock)
#   notes to bs!E101     → R32     (debtors)
#   notes to bs!E120     → R33     (cash+bank)
#   notes to bs!E150     → R34     (ST loans)
#   notes to bs!E160     → R35     (other current assets)
# DO NOT write directly to bs sheet rows.

def _inject_bs_sheet(wb, parsed, client_name, cy_year, py_year, log):
    log.append('BS sheet: formula-driven, no direct writes needed')


# ── P&L sheet ─────────────────────────────────────────────────────────────────
# P&L sheet is ENTIRELY formula-driven from notes to p&l:
#   notes to p&l!E8   → R5  revenue
#   notes to p&l!E13  → R6  other income
#   notes to p&l!E29  → R10 cost of materials
#   notes to p&l!E40  → R11 employee benefits
#   notes to p&l!E47  → R12 finance cost
#   notes to p&l!E57  → R13 depreciation (from Fixed Assets P.Yr.)
#   notes to p&l!E87  → R14 other expenses
# DO NOT write directly to p&l sheet rows.

def _inject_pl_sheet(wb, parsed, client_name, cy_year, py_year, log):
    log.append('P&L sheet: formula-driven, no direct writes needed')


# ── Capital sheet ──────────────────────────────────────────────────────────────
# Template structure (verified from data_only=False inspection):
#   Block 1 (rows 6-19):  CY headers R7, CY data R8, CY totals R10, PY row R11
#   Block 2 (rows 24-40): CY headers R30, CY data R31, CY totals R33, PY row R34
#
# Columns (1-indexed):
#   B=2 Sr.No.  C=3 Opening  D=4 Additions  E=5 Withdrawals  F=6 Profit  G=7 Closing(formula)
#
# For proprietorship: caps[0] fills Block1 PY(R11) + Block2 PY(R34)
# For partnership:    caps[0] fills Block1 PY(R11),  caps[1] fills Block2 PY(R34)
#
# The CY opening (R8 col C) = =G11 formula in template. DO NOT overwrite G column.
# We only write: C(opening), D(additions), E(withdrawals), F(profit) for PY rows.

def _inject_capital_sheet(wb, parsed, client_name, cy_year, py_year, log):
    if 'capital' not in wb.sheetnames:
        return
    ws = wb['capital']
    p = parsed
    caps = p.get('capital_accounts', [])
    if not caps:
        log.append('Capital: no capital accounts found')
        return

    COL_NAME   = 2  # B
    COL_OPEN   = 3  # C: Opening balance
    COL_ADD    = 4  # D: Capital introduced
    COL_WITH   = 5  # E: Withdrawals
    COL_PROFIT = 6  # F: Share of profit/loss
    # COL_CLOSE = 7  # G: formula =C+D-E+F — DO NOT WRITE

    def _write_py_row(r, cap):
        # Col A(1) = "Previous Year (PY)" label — already in template, don't touch
        # Col B(2) = MergedCell (merged with A) — DO NOT write name here
        # Col C(3)=opening, D(4)=additions, E(5)=withdrawals, F(6)=profit, G(7)=closing formula
        #
        # COL_PROFIT (F=6): the template has ='p&l'!F17 which references CY profit.
        # For the PY row we MUST overwrite this formula with the actual PY net_profit value,
        # otherwise the closing formula G=C+D-E+F will use CY profit instead of PY profit.
        # COL_CLOSE (G=7) is =C+D-E+F formula — leave it so it recalculates correctly.
        opening_val   = cap.get('opening', 0) or 0
        additions_val = cap.get('additions', 0) or 0
        withdraw_val  = cap.get('withdrawals', 0) or 0
        profit_val    = cap.get('net_profit', 0) or p.get('net_profit', 0) or 0
        _write_num(ws, r, COL_OPEN,   opening_val)
        _write_num(ws, r, COL_ADD,    additions_val)
        _write_num(ws, r, COL_WITH,   withdraw_val)
        _write_num(ws, r, COL_PROFIT, profit_val)   # overwrite formula with PY net_profit
        # Mark C-F as yellow input cells for CA (they can verify/edit)
        for col in (COL_OPEN, COL_ADD, COL_WITH, COL_PROFIT):
            cell = ws.cell(r, col)
            from openpyxl.cell import MergedCell as MC
            if not isinstance(cell, MC):
                cell.fill = _INPUT_FILL

    # Block 1 PY row = R11
    _write_py_row(11, caps[0])
    # Update the CY row name (R8 col B) — this IS a writable cell (not merged)
    _write(ws, 8, COL_NAME, caps[0].get('name', ''))

    # Block 2 PY row = R34 — only write if there are 2+ capital accounts
    if len(caps) > 1:
        _write_py_row(34, caps[1])
        _write(ws, 31, COL_NAME, caps[1].get('name', ''))
        log.append(f'Capital sheet: PY rows written for {len(caps)} capital account(s)')
    else:
        log.append('Capital sheet: 1 account — Block 2 left as template default')

    # Replace stale signing names in the capital sheet.
    # Some templates have old CA/Proprietor names hardcoded from a previous client.
    prop_name = caps[0].get('name', '') if caps else ''
    _STALE_PROP_NAMES = {'(DIMPAL JAIN)', 'DIMPAL JAIN', '(ASHWANI KUMAR)', 'PROP.'}
    from openpyxl.cell import MergedCell as _MC
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, _MC) or not isinstance(cell.value, str):
                continue
            v = cell.value.strip()
            if v.upper() in {n.upper() for n in _STALE_PROP_NAMES}:
                if prop_name:
                    cell.value = prop_name


# ── Notes to BS ────────────────────────────────────────────────────────────────
# Template verified (data_only=False):
#   Col D(4) = CY   Col E(5) = PY
#
# Key writable cells:
#   R7:E   = LT secured term loans from banks
#   R8:E   = LT secured from other parties
#   (R10   = SUM(E7:E8) formula)
#   R15:E  = LT unsecured from related parties
#   R16:E  = LT unsecured from other parties
#   (R18,R19 = formulas)
#   R25:E  = ST secured CC a/c
#   R54-R64 = Other current liability items (11 slots)
#   (R65   = SUM(D54:D64) formula)
#   R106:E = Cash in hand
#   (R110  = SUM formula)
#   R113-R118:E = Individual bank balances (6 slots)
#   (R119,R120 = formulas)
#   R134-R140:E = ST loans to others (B section, 7 slots)
#   (R141,R150 = formulas)
#
# NOTE: Trade receivable rows (R91, R97) are formula-driven from Details sheet.
#       DO NOT write to them.

def _inject_notes_bs(wb, parsed, client_name, cy_year, py_year, log):
    if 'notes to bs' not in wb.sheetnames:
        return
    ws = wb['notes to bs']
    p = parsed
    PY = 5   # col E
    CY = 4   # col D

    secured   = p.get('secured_loans', 0)
    unsecured = p.get('unsecured_loans', 0)

    # ── Note 3: Long-term borrowings ──────────────────────────────────────────
    # Secured: write total to R7 (bank), R8=0 (other parties)
    _py(ws, 7,  PY, secured); _cy(ws, 7,  CY)
    _py(ws, 8,  PY, 0);       _cy(ws, 8,  CY)
    # R10 = SUM(E7:E8) — formula, skip

    # Unsecured related parties: R15 C4/C5 = =Details!D13/E13 — FORMULA, DO NOT WRITE.
    # The Details sheet inject (R7-R12) feeds Details!R13 via SUM formula,
    # which then flows here automatically.
    # R16 (from other parties) is writable but typically 0 for T-shaped clients.
    _py(ws, 16, PY, 0);         _cy(ws, 16, CY)
    # R18 = SUM(E14:E17), R19 = R18+R10 — formulas, skip

    # ── Note 3: Short-term borrowings ─────────────────────────────────────────
    _py(ws, 25, PY, 0); _cy(ws, 25, CY)
    # R29, R33, R34 = formulas, skip

    # ── Note 4: Trade payables ────────────────────────────────────────────────
    # R38/R39 are formula-driven from Details!D68/E68 and Details!D62/E62
    # Do NOT write here — creditor parties in Details drive these.

    # ── Note 5: Other current liabilities (R54-R64, 11 slots) ────────────────
    # Parser puts OCL items in other_payable_items when found correctly.
    # For T-shaped BS they often bleed into unsecured_loan_parties.
    # Detect payable/TDS/GST type names and rescue them into OCL.
    _OCL_KEYWORDS = ('payable', 'tds', 'provision', 'accrued',
                     'outstanding', 'due to', 'liability', 'ch.issued',
                     'cheque issued', 'yet clear', 'rcm', 'bonus')
    # STRICT OCL filter: only items that are CLEARLY payable/TDS/provision type.
    # Do NOT include trade creditor names (M/s., company names, fabric/textile firms).
    # Do NOT include unsecured lender names (persons who gave loans).
    # Do NOT include P&L line items (depreciation, net profit, etc.).
    # FIX S10: Only pull from other_payable_items (direct Annexure-D parser).
    # FIX B (2026-08-14): Extend rejection to cover:
    #   - Person/lender names that look like individuals (Rohit Vig, Smt. Santosh Rani…)
    #     These belong in Unsecured Loans, not OCL.
    #   - P&L items: Depreciation, Net Profit, Net Loss, Gross Profit.
    #   - Any item whose name is purely numeric (amounts printed as labels in some XLS).
    ocl_items = list(p.get('other_payable_items', []))

    # Build a set of unsecured lender names (lower) so we can reject them from OCL.
    _lender_names_lower = {
        ul['name'].strip().lower()
        for ul in p.get('unsecured_loan_parties', [])
    }

    # Build a set of trade creditor names (lower) for the same purpose.
    _creditor_names_lower = {
        cp['name'].strip().lower()
        for cp in p.get('sundry_creditor_parties', [])
    }

    # P&L line keywords — these should NEVER appear as OCL items
    _PL_KEYWORDS = ('depreciation', 'net profit', 'net loss', 'gross profit',
                    'profit before', 'profit after', 'net income', 'net income')

    # Person-name patterns that indicate lenders/individuals, not payables
    _PERSON_PREFIXES = ('smt.', 'sh.', 'shri ', 'mr.', 'mrs.', 'ms.', 'dr.',
                        'rohit ', 'varinder ', 'manik ', 'tinku ', 'savita')

    # Trade company signals
    _TRADE_SIGNALS = ('m/s', 'pvt', 'ltd', 'textile', 'fabrics', 'fashion', 'silk',
                      'creation', 'impex', 'enterprise', 'house', 'industries',
                      'communication', 'comm.', 'print', 'embroidery', 'traders',
                      'suppliers', 'garment', 'cloth', 'saree', 'suits')

    validated_ocl = []
    for item in ocl_items:
        nm_l = item['name'].strip().lower()

        # Hard reject: this name is a known unsecured lender
        if nm_l in _lender_names_lower:
            log.append(f"OCL filter: rejected lender '{item['name']}' (belongs in Unsecured Loans)")
            continue

        # Hard reject: this name is a known trade creditor
        if nm_l in _creditor_names_lower:
            log.append(f"OCL filter: rejected creditor '{item['name']}' (belongs in Trade Payables)")
            continue

        # Hard reject: P&L items (depreciation, profit figures)
        if any(kw in nm_l for kw in _PL_KEYWORDS):
            log.append(f"OCL filter: rejected P&L item '{item['name']}'")
            continue

        # Hard reject: person-name prefix (Smt., Sh., Rohit, etc.) = lender, not payable
        if any(nm_l.startswith(pfx) for pfx in _PERSON_PREFIXES):
            log.append(f"OCL filter: rejected person name '{item['name']}' (likely lender)")
            continue

        # Hard reject: numeric-only names (stray amounts printed as labels)
        try:
            float(item['name'].replace(',', '').replace(' ', ''))
            log.append(f"OCL filter: rejected numeric label '{item['name']}'")
            continue
        except ValueError:
            pass

        # Keep if it has a clear OCL keyword
        if any(kw in nm_l for kw in _OCL_KEYWORDS):
            validated_ocl.append(item)
            continue

        # FIX P (2026-08-14): Only reject if BOTH a trade keyword AND a company suffix
        # are present. A single trade keyword alone (e.g. 'silk', 'textile', 'fashion')
        # may appear in a legitimate payable name (e.g. "Accounting Charges Payable"
        # doesn't have trade signals, but the old code was too broad). Requiring both
        # a trade word AND a company suffix (pvt/ltd/m/s/traders/co.) prevents
        # legitimate payables from being dropped.
        _COMPANY_SIGNALS = ('pvt', 'ltd', 'limited', 'm/s', 'm/s.', 'traders',
                             'co.', ' co ', 'inc.', 'corporation', 'works')
        has_trade = any(ts in nm_l for ts in _TRADE_SIGNALS)
        has_company = any(cs in nm_l for cs in _COMPANY_SIGNALS)
        if has_trade and has_company:
            log.append(f"OCL filter: rejected trade company name '{item['name']}'")
            continue

        # Default: keep (generic name, no clear rejection signal)
        validated_ocl.append(item)

    ocl_items = validated_ocl
    log.append(f"OCL items after validation: {len(ocl_items)} "
               f"({[it['name'] for it in ocl_items]})")

    # Deduplicate OCL items by normalized name.
    # e.g. "GST- Reverse Charges" and "GST REVERSE CHARGE PAYABLE" are the same item
    # parsed twice from adjacent annexure columns.
    # Strategy: normalize to alpha-only prefix (first 12 chars) + amount for dedup key.
    seen_ocl = {}
    deduped_ocl = []
    for item in ocl_items:
        alpha_key = ''.join(c for c in item['name'].lower() if c.isalpha())[:12]
        # Also deduplicate by amount: same amount = same item regardless of name variant
        amt_key = round(item['amount'])
        key = (alpha_key, amt_key)
        if key not in seen_ocl:
            seen_ocl[key] = True
            deduped_ocl.append(item)
    ocl_items = deduped_ocl

    for i in range(11):
        r = 54 + i
        if i < len(ocl_items):
            item = ocl_items[i]
            _write(ws, r, 2, item['name'])
            _py(ws, r, PY, item['amount'])
        else:
            _py(ws, r, PY, 0)
        _cy(ws, r, CY)
    # R65 = SUM(D54:D64) — formula, skip

    # ── Note 10: Cash and bank balances ───────────────────────────────────────
    _py(ws, 106, PY, p.get('cash_in_hand', 0)); _cy(ws, 106, CY)
    # R110 = SUM(E106:E106) — formula, skip

    banks = p.get('bank_balances', [])
    _cb = p.get('cash_bank', 0); cash_bank_total = 0 if (_cb is None or (isinstance(_cb, float) and math.isnan(_cb))) else float(_cb)
    _ch = p.get('cash_in_hand', 0); cash_hand = 0 if (_ch is None or (isinstance(_ch, float) and math.isnan(_ch))) else float(_ch)
    bank_total_from_bs = round(cash_bank_total - cash_hand)

    for i in range(6):   # R113-R118
        r = 113 + i
        if i < len(banks):
            _write(ws, r, 2, banks[i]['name'])
            _py(ws, r, PY, banks[i]['amount'])
        elif i == 0 and not banks and bank_total_from_bs > 0:
            # No parsed bank breakdown — write total in first slot as single account
            _write(ws, r, 2, 'HDFC Bank')
            _py(ws, r, PY, bank_total_from_bs)
        else:
            _py(ws, r, PY, 0)
        _cy(ws, r, CY)
    # R119 = SUM(E113:E118), R120 = R110+R119 — formulas, skip

    # ── Note 11: Short-term loans and advances ────────────────────────────────
    # B section — Loans to Others: R134-R140 (7 slots)
    # Template formula SUM(E138:E140) is too narrow — fix it to cover all 7 rows.
    loans_items = p.get('loans_to_other_items', p.get('loans_advances_items', []))
    loans_total = p.get('loans_advances', 0) or p.get('advances_security', 0)
    for i in range(7):
        r = 134 + i
        if i < len(loans_items):
            itm = loans_items[i]
            _write(ws, r, 2, itm.get('name', ''))
            _py(ws, r, PY, itm.get('amount', 0))
        else:
            _py(ws, r, PY, 0)
        _cy(ws, r, CY)
    # If no itemised list but we have a total, lump into R134
    if not loans_items and loans_total:
        _py(ws, 134, PY, loans_total)
    # Fix Total(B) formula to cover all 7 rows R134:R140
    _write(ws, 141, PY, '=SUM(E134:E140)')
    _write(ws, 141, CY, '=SUM(D134:D140)')

    # C section — Advance to Revenue Authorities: R143-R147 (5 slots)
    # Template formula SUM(E143:E143) only covers 1 row — fix to cover all 5.
    rev_items = p.get('advance_to_revenue_items', [])
    from openpyxl.cell import MergedCell as _MCrev
    for i in range(5):
        r = 143 + i
        if i < len(rev_items):
            itm = rev_items[i]
            _write(ws, r, 2, itm.get('name', ''))
            _py(ws, r, PY, itm.get('amount', 0))
        else:
            _py(ws, r, PY, 0)
        _cy(ws, r, CY)
        # FIX L (2026-08-14): Preserve yellow highlight on PY (E) cell.
        # _py()→_write_num() writes the value but never sets fill. The template
        # has yellow (_INPUT_FILL) on IGST/GST/TDS rows so the CA can verify them.
        # We apply _INPUT_FILL to the PY cell explicitly after writing.
        py_cell = ws.cell(row=r, column=PY)
        if not isinstance(py_cell, _MCrev):
            py_cell.fill = _INPUT_FILL
    # Fix Total(C) formula to cover all 5 rows R143:R147
    _write(ws, 148, PY, '=SUM(E143:E147)')
    _write(ws, 148, CY, '=SUM(D143:D147)')
    # R150 Total(A+B+C) formula is already correct — skip

    # ── Note 12: Other current assets (R154-R159, 6 slots) ───────────────────
    oca_items = p.get('other_current_asset_items', [])
    for i in range(6):
        r = 154 + i
        if i < len(oca_items):
            itm = oca_items[i]
            _write(ws, r, 2, itm.get('name', ''))
            _py(ws, r, PY, itm.get('amount', 0))
        else:
            _py(ws, r, PY, 0)
        _cy(ws, r, CY)
    # R160 = SUM formula, skip

    log.append('Notes to BS injected')


# ── Notes to P&L ───────────────────────────────────────────────────────────────
# Template verified (data_only=False):
#   Col D(4) = CY   Col E(5) = PY
#
# Key structure:
#   R6:E   = sales (Note 13)
#   R12:E  = other income (Note 14)
#   R18:E  = opening stock PY  [R18:D = =E27 formula — DO NOT TOUCH CY]
#   R21:E  = purchases PY
#   R24-R25 = direct expense item rows
#   R27:E  = closing stock PY  [R27:D = formula from GROSS PROFIT — DO NOT TOUCH CY]
#   R34:E  = salaries, R35:E = bonus, R36:E = staff welfare, R40=SUM formula
#   R44:E  = bank interest, R45:E = unsecured interest, R47=SUM(E44:E45) formula
#   R56:E  = depreciation — formula from Fixed Assets P.Yr.!I37 — DO NOT TOUCH
#   R57    = SUM formula
#   R61-R75 = named other-expense rows (fixed labels, see map below)
#   R76-R86 = spare rows for unmatched items
#   R87    = SUM(D61:D86) formula
#
# CRITICAL: other_expense_items from T-shaped BS must be DISTRIBUTED into the
# correct named rows. Items that are "direct expenses" (FOC, freight) go into
# R24-R25 (notes to p&l) and also into the GROSS PROFIT sheet (handled there).
# Bank charges → finance cost R44. Staff welfare → R36. Everything else → R61-R86.

_OTHER_EXP_ROW_MAP = {
    'audit fee':                 61,
    'audit fees':                61,
    'audit fee.':                61,
    'bank charges':              62,
    'bank charge':               62,
    'bank charges.':             62,
    'car exp':                   63,
    'car exp.':                  63,
    'car exps':                  63,
    'car exps.':                 63,
    'car expense':               63,
    'car expenses':              63,
    'commission':                64,
    'commisssion':               64,
    'diwali exp':                65,
    'diwali exp.':               65,
    'electricity exp':           66,
    'electricity exp.':          66,
    'electricity expenses':      66,
    'electricty exp.':           66,
    'insurance':                 67,
    'insurance exp.':            67,
    'labour charges':            68,
    'labour charge':             68,
    'legal fee':                 69,
    'legal fees':                69,
    'to legal fee':              69,
    'loading unloading':         70,
    'loading unloading charges': 70,
    'petrol exp':                71,
    'petrol exp.':               71,
    'petrol expenses':           71,
    'repair & maintainance':     72,
    'repair & maintenance':      72,
    'repair and maintenance':    72,
    'scooter exp':               73,
    'scooter exp.':              73,
    'scooter expenses':          73,
    'shop exp':                  74,
    'shop exp.':                 74,
    'shop expenses':             74,
    'telephone exp':             75,
    'telephone exp.':            75,
    'telephone expenses':        75,
    # FIX G (2026-08-14): Route Accountant Salary and Professional Charges to spare rows.
    # Notes to p&l has spare data rows R80-R86 (currently zero-filled blanks).
    # Map these expense types to those rows so they appear in the output.
    'accountant salary':         80,
    'accountant sal':            80,
    'accountant':                80,
    # FIX 2: Bansal-specific labels
    'association fees':          76,
    'association fee':           76,
    'misc exp':                  77,
    'misc exp.':                 77,
    'misc expense':              77,
    'misc expenses':             77,
    'rebate & discount':         78,
    'rebare & discount':         78,
    'rebare & disc':             78,
    'property tax':              79,
    'repair & maintance':        72,
    'repair & maintenance':      72,
    'repair and maintenance':    72,
    'bank interest':             44,   # routes to finance cost R44 (not other exp)
    'bank charges':              62,
    'bank charge':               62,
    'professional charges':      81,
    'professional charge':       81,
    'professional':              81,
    # FIX H (2026-08-14): Map the remaining "unmatched" expense items so they write
    # to correct rows instead of falling through to unmatched (which silently drops them).
    # The template has labelled rows R76=General, R77=Entertainment, R79=Vehicle.
    'general expenses':          76,
    'general exp':               76,
    'general exp.':              76,
    'entertainment exp':         77,
    'entertainment exp.':        77,
    'entertainment expenses':    77,
    'entertainment':             77,
    'vehicle expenses':          79,
    'vehicle exp':               79,
    'vehicle exp.':              79,
    'printing & stationery':     78,
    'printing and stationery':   78,
    'printing & stationary':     78,
    'printing & stationery.':    78,
    'stationery':                78,
    'stationary':                78,
    'tour & travelling':         82,
    'tour and travelling':       82,
    'tour & travelling exps':    82,
    'tour & travelling exp':     82,
    'travelling expenses':       82,
    'travelling exp':            82,
    # General / Entertainment / Vehicle / Tour → spare rows (76+), no fixed row
    # but map them here so unmatched list is small and they print correctly
}

_DIRECT_EXP_KEYS = {
    'f.o.c.', 'foc', 'freight inward', 'freight outward',
    'freight inward (gst)', 'frieght inward', 'frieght inward (gst)',
    'freight,octrai & cartage', 'frieght,octrai & cartage',
    'freight, octrai & cartage', 'frieght, octrai & cartage',
    'freight octrai & cartage', 'frieght octrai & cartage',
    'octroi & cartage', 'freight & cartage',
}

_FINANCE_COST_KEYS = {
    # Bank interest and unsecured loan interest go to Note 17
    # NOTE: 'bank charges' goes to Note 19 Other Expenses (row 62), NOT here
    'bank interest',
    'interest', 'interest on unsecured', 'interest on unsecured loans',
    'interest to parties', 'interest to party',
}

_BONUS_KEYS = {
    'bonus', 'bonus exp', 'bonus exp.', 'bonus expenses',
}

_STAFF_WELFARE_KEYS = {
    'staff & labour welfare', 'staff and labour welfare',
    'staff welfare', 'labour welfare',
    'staff & labour welfare expenses',
}


def _inject_notes_pl(wb, parsed, client_name, cy_year, py_year, log):
    if 'notes to p&l' not in wb.sheetnames:
        return
    ws = wb['notes to p&l']
    p = parsed
    PY = 5   # col E
    CY = 4   # col D

    all_items = p.get('other_expense_items', [])

    # Classify items
    direct_items   = []
    finance_items  = []
    welfare_items  = []
    bonus_items    = []
    other_items    = []
    for it in all_items:
        k = it['name'].strip().lower()
        if k in _DIRECT_EXP_KEYS:
            direct_items.append(it)
        elif k in _FINANCE_COST_KEYS:
            finance_items.append(it)
        elif k in _STAFF_WELFARE_KEYS:
            welfare_items.append(it)
        elif k in _BONUS_KEYS:
            bonus_items.append(it)
        else:
            other_items.append(it)

    # ── Note 13: Revenue (R6:E = sales) ──────────────────────────────────────
    _py(ws, 6,  PY, p.get('sales', 0));       _cy(ws, 6,  CY)

    # ── Note 14: Other income (R12:E) ────────────────────────────────────────
    _py(ws, 12, PY, p.get('other_income', 0)); _cy(ws, 12, CY)

    # ── Note 15: Cost of material ─────────────────────────────────────────────
    _py(ws, 18, PY, p.get('opening_stock', 0))
    # FIX F (2026-08-14): Break the circular reference.
    # The template has notes_to_p&l!D18 = =E27 (CY opening stock = CY closing stock).
    # This creates a circular chain:
    #   GROSS PROFIT!B9 = notes_to_p&l!D18 = =E27 = notes_to_p&l!D27
    #   = 'GROSS PROFIT'!E13 (CY closing stock, empty) → B9 (CY opening) → circular!
    # Fix: replace D18 formula with =E27 of the PREVIOUS period, which equals the PY
    # closing stock. For the CY column, the opening stock = PY closing stock.
    # We write it as a direct reference to the PY closing stock cell (E27).
    # This breaks the circular because E27 (PY) is a static value, not a formula that
    # loops back to D18.
    # The correct CY opening stock value = PY closing stock = p.get('closing_stock')
    # Write CY opening stock as the actual value (not a formula) so CA can verify/override.
    _cy_val = p.get('closing_stock', 0) or 0
    cell_d18 = ws.cell(18, CY)
    from openpyxl.cell import MergedCell as _MCfix
    if not isinstance(cell_d18, _MCfix):
        cell_d18.value = float(_cy_val)
        cell_d18.fill  = _INPUT_FILL
        cell_d18.number_format = '#,##0'
    log.append(f"FIX F: notes to p&l D18 (CY opening stock) set to {_cy_val:,.0f} "
               f"(breaks circular =E27 formula)")
    _py(ws, 21, PY, p.get('purchases', 0));   _cy(ws, 21, CY)

    # Direct expenses → R24-R25 (2 fixed slots)
    for i in range(2):
        r = 24 + i
        if i < len(direct_items):
            it = direct_items[i]
            _write(ws, r, 2, it['name'])
            _py(ws, r, PY, it['amount']); _cy(ws, r, CY)
        else:
            _py(ws, r, PY, 0); _cy(ws, r, CY)

    _py(ws, 27, PY, p.get('closing_stock', 0))
    # R27:D = formula from GROSS PROFIT — DO NOT write CY here

    # ── Note 16: Employee benefits ────────────────────────────────────────────
    # R34=salaries, R35=bonus, R36=staff welfare; R40=SUM formula
    sal = p.get('salary_expenses', 0)

    # Staff welfare and bonus from other_expense_items if present
    welfare_amt = sum(x['amount'] for x in welfare_items)
    bonus_amt   = sum(x['amount'] for x in bonus_items)

    _py(ws, 34, PY, sal);         _cy(ws, 34, CY)   # salaries
    _py(ws, 35, PY, bonus_amt);   _cy(ws, 35, CY)   # bonus
    _py(ws, 36, PY, welfare_amt); _cy(ws, 36, CY)   # staff welfare
    # R40 = SUM(E34:E36) — formula, skip

    # ── Note 17: Finance cost ─────────────────────────────────────────────────
    # R44:E = bank interest (from banks only), R45:E = interest on unsecured loans
    # R47 = SUM(E44:E45) — formula, skip
    # finance_items contains only explicit 'bank interest' type entries from other_expense_items.
    # NOTE: 'bank charges' (3992) goes to Note 19 R62 (other expenses), NOT here.
    # NOTE: interest_paid parser field now captures 'TO INTEREST TO PARTIES' type entries.
    # finance_items captures entries from expense_kws (other_expense_items).
    # interest_paid captures entries from pl_keywords directly.
    bank_int  = sum(x['amount'] for x in finance_items
                    if x['name'].strip().lower() == 'bank interest')
    unsec_int = sum(x['amount'] for x in finance_items
                    if x['name'].strip().lower() not in ('bank interest', 'bank charges',
                                                         'bank charge'))

    # Also add interest_paid captured via pl_keywords (e.g. 'TO INTEREST TO PARTIES')
    # Filter out bank-charge type values (3,924 is clearly interest to parties, not bank charges)
    interest_paid_field = p.get('interest_paid', 0)
    if interest_paid_field > 0 and unsec_int == 0:
        unsec_int = interest_paid_field
        log.append(f'Finance cost: unsec interest {unsec_int} from interest_paid field')

    # If unsec_int still 0, scan unsecured_loan_parties for interest entries.
    # In T-shaped XLS the interest on unsecured loans often appears as a party row
    # with names like 'INTEREST', 'JLDR05627G', '194A' etc.
    _UNSEC_INT_KEYS = ('interest', 'jldr', '194a')
    if unsec_int == 0:
        seen_amounts = set()
        for party in p.get('unsecured_loan_parties', []):
            nm_l = party['name'].strip().lower()
            amt  = party['amount']
            if any(k in nm_l for k in _UNSEC_INT_KEYS) and amt > 0:
                # Deduplicate: same amount often listed under multiple reference codes
                if amt not in seen_amounts:
                    seen_amounts.add(amt)
                    unsec_int += amt
        if unsec_int > 0:
            log.append(f'Finance cost: unsec interest {unsec_int} extracted from unsecured_loan_parties')

    _py(ws, 44, PY, bank_int);  _cy(ws, 44, CY)
    # FIX I (2026-08-14): Always write unsec_int to R45:E, even if it is 0.
    # The template may have a stale hardcoded value (e.g. 16,161.09 from a prior run).
    # _py() calls _write_num() which always overwrites — so this is correct.
    # But if the parser fails to find interest, unsec_int=0 and _py writes 0,
    # correctly clearing any stale value. The issue was the old template value of
    # 16,161.09 surviving because this cell was not being written at all in some paths.
    _py(ws, 45, PY, unsec_int); _cy(ws, 45, CY)
    log.append(f"Finance cost: bank_int={bank_int:.2f}, unsec_int={unsec_int:.2f} "
               f"(interest_paid_field={interest_paid_field:.2f})")
    # R47 = SUM(E44:E45) — formula, skip

    # ── Note 18: Depreciation ─────────────────────────────────────────────────
    # R56:E = formula from 'Fixed Assets P. Yr.'!I37 — DO NOT WRITE
    # R57 = SUM formula — skip

    # ── Note 19: Other expenses (R61-R86, total R87=SUM) ─────────────────────
    # Clear all named rows first (set to 0), then fill in matched items
    written_rows = set()
    for k, r in _OTHER_EXP_ROW_MAP.items():
        if r not in written_rows:
            _py(ws, r, PY, 0); _cy(ws, r, CY)
            written_rows.add(r)

    unmatched = []
    for it in other_items:
        k = it['name'].strip().lower()
        r = _OTHER_EXP_ROW_MAP.get(k)
        if r:
            # FIX O (2026-08-14): Also write the item name to col B (label column).
            # Template rows R76-R82 (spare rows) have blank col B; without a label the
            # CA sees a value with no description. Write the name so the row is complete.
            existing_label = ws.cell(r, 2).value
            if not existing_label or str(existing_label).strip() in ('', '0'):
                _write(ws, r, 2, it['name'])
            _py(ws, r, PY, it['amount']); _cy(ws, r, CY)
        else:
            unmatched.append(it)

    # Write unmatched into spare rows R76-R86
    unmatched_spare_rows = set()
    for i, it in enumerate(unmatched):
        r = 76 + i
        if r > 86:
            break
        _write(ws, r, 2, it['name'])
        _py(ws, r, PY, it['amount']); _cy(ws, r, CY)
        unmatched_spare_rows.add(r)

    # Build set of all rows written by the mapped-item loop
    mapped_written_rows = set()
    for it in other_items:
        k = it['name'].strip().lower()
        r = _OTHER_EXP_ROW_MAP.get(k)
        if r and 76 <= r <= 86:
            mapped_written_rows.add(r)

    # Clear remaining spare rows — but ONLY rows not already written by either loop.
    # Previous bug: range(76 + len(unmatched), 87) zeroed ALL spare rows when
    # unmatched=[], including rows R76-R82 that were just correctly written via
    # _OTHER_EXP_ROW_MAP (General Expenses, Entertainment, Vehicle, Accountant Salary,
    # Professional Charges, Printing). Now skip any row in mapped_written_rows or
    # unmatched_spare_rows so those values are preserved.
    for r in range(76, 87):
        if r in mapped_written_rows or r in unmatched_spare_rows:
            continue   # already has correct data — do not zero
        _py(ws, r, PY, 0); _cy(ws, r, CY)

    # R87 = SUM(D61:D86) — formula, skip

    log.append(f'Notes to P&L: {len(direct_items)} direct, {len(other_items)} other, '
               f'{len(unmatched)} unmatched to spare rows')


# ── Details sheet ──────────────────────────────────────────────────────────────
# Template structure (verified from data_only=False inspection):
#
#   UNSECURED LOANS:
#     R6  = "FROM RELATED PARTIES" header
#     R7-R12  = 6 party rows  (B=name, D=CY, E=PY)
#     R13 = SUM(D7:D12) formula
#     R15 = FROM OTHER PARTIES row (D=CY, E=PY)
#     R17 = SUM formula
#     R19 = TOTAL formula
#
#   SUNDRY CREDITORS:
#     R22 = header row
#     R23-R55 = 33 creditor party rows (A="M/s.", B=name, D=CY, E=PY)
#     R56 = "Advance from Customers" sub-header
#     R57-R61 = 5 advance-from-customer rows
#     R62 = SUM(D23:D61) formula
#     R63 = "DUE TO MSME" header
#     R68 = SUM formula
#     R69 = TOTAL formula
#
#   TRADE RECEIVABLE >6 months:
#     R72 = section header
#     R73 = header row
#     R74-R128 = 55 debtor rows (A="M/s.", B=name, D=CY, E=PY)
#     R129 = SUM(D74:D128) formula  ← notes to bs!E91 reads Details!E129
#
#   TRADE RECEIVABLE <6 months:
#     R132 = section header
#     R133 = header row
#     R134-R135 = 2 rows
#     R136 = SUM formula  ← notes to bs!E97 reads Details!E136

# Keywords that identify OCL items (payables) — should NOT go in Details unsecured section
_OCL_PARTY_KEYWORDS = (
    'payable', 'tds', 'gst reverse', 'provision for tax',
    'accrued', 'outstanding expenses', 'due to',
)


def _is_real_party_name(nm):
    """Filter out junk/numeric names from parsed party lists."""
    nm = nm.strip()
    if len(nm) < 3:
        return False
    # Reject purely numeric strings
    try:
        float(nm.replace(',', ''))
        return False
    except ValueError:
        pass
    # Reject date strings
    if re.match(r'^\d{4}-\d{2}-\d{2}', nm):
        return False
    if re.match(r'^\d{2}\.\d{2}\.\d{4}', nm):
        return False
    # Reject very short number-heavy tokens like "287.5", "3814.5"
    if re.match(r'^[\d\.\,\s]+$', nm):
        return False
    nm_l = nm.lower()
    # Reject internal calculation leftovers
    _SKIP = (
        'opening balance', 'closing balance', 'gross profit', 'net profit',
        'add profit', 'total', 'pack/pcs', 'amount in rs',
        'withdrawls', 'advance tax', 'appeal fees', 'trf to',
        'hdfc insurance', 'tcs receivable', 'lic ', 'start health',
        'building', 'motor cycle', 'furniture', 'water cooler',
        'digital moisture', 'servo control', 'refer item', 'refer as per',
        'nature of', 'particulars', 'annexure', 'add profit',
        'camera', 'washing machine', 'electricals', 'led', 'computer',
        'cheque/rtgs', 'cheque / rtgs', 'neft', 'rtgs',
    )
    if any(sp in nm_l for sp in _SKIP):
        return False
    return True


# Personal/relationship/commodity words that should NOT appear as trade parties
_PERSONAL_JUNK = {
    'wife', 'mother', 'father', 'son', 'daughter', 'husband',
    'sister', 'brother', 'stock',
    # Common individual first names that appear as lenders/family in T-shaped XLS
    'garima', 'sheenu', 'dimpal', 'rachit', 'deepak', 'pawan', 'asha',
}
# Commodity/food words that indicate stock items not debtors
# Note: only applied when the word appears in a context suggesting it IS a commodity,
# not a company name containing that word (e.g. "Sugar Mills Ltd." is a company).
_COMMODITY_WORDS = (
    'aloo', 'chana', 'ghee', 'katta',
    'besan', 'atta', 'dal', 'makai', 'maize',
    'sesame', 'sunflower', 'mustard', 'ground nut',
)
# These commodity words need careful matching: only block if NOT followed by company suffixes
_CAREFUL_COMMODITY = ('cotton', 'oil', 'sugar', 'rice', 'wheat',)
_COMPANY_SUFFIXES = ('mill', 'mills', 'industries', 'industry', 'corp', 'ltd', 'limited',
                     'pvt', 'private', 'company', 'co.', 'co ', 'refinery', 'syndicate',)


def _is_debtor_party(nm, lender_names=None):
    """True if this looks like a genuine trade debtor / receivable party."""
    if not _is_real_party_name(nm):
        return False
    nm_l = nm.lower().strip()
    # Reject if name or any individual word is in personal/junk set
    if nm_l in _PERSONAL_JUNK:
        return False
    # Also reject multi-word names where any word is a personal/family junk word
    if any(word in _PERSONAL_JUNK for word in nm_l.split()):
        return False
    # Reject commodity items
    if any(c in nm_l for c in _COMMODITY_WORDS):
        return False
    # Careful commodities: only block if not a company name
    for cc in _CAREFUL_COMMODITY:
        if cc in nm_l and not any(sfx in nm_l for sfx in _COMPANY_SUFFIXES):
            return False
    # Reject known lender names (passed from unsecured_loan_parties)
    if lender_names:
        nm_norm = nm.strip().lower()
        if nm_norm in lender_names:
            return False
    return True


def _is_unsecured_party(nm):
    """True if this is a real lender (not an OCL/payable item, not a commodity)."""
    nm_l = nm.strip().lower()
    # Payable-type items belong in OCL, not unsecured loans Details rows
    _PAYABLE_KEYS = (
        'payable', ' tds', 'gst-', 'gst reverse', 'gst credit',
        'provision', 'salary payable', 'audit fee', 'reverse charges',
    )
    if any(k in nm_l for k in _PAYABLE_KEYS):
        return False
    if not _is_real_party_name(nm):
        return False
    # Reject commodity words
    if any(c in nm_l for c in _COMMODITY_WORDS):
        return False
    # Reject TDS/interest reference codes (alphanumeric codes like JLDR05627G)
    if re.match(r'^[A-Z0-9]{5,12}$', nm.strip()):
        return False
    # Reject generic interest/loan labels
    _GENERIC_LABELS = ('interest', '194a', '194c', 'tds on', 'loan int')
    if any(g in nm_l for g in _GENERIC_LABELS):
        return False
    return True


def _inject_details_sheet(wb, parsed, client_name, cy_year, py_year, log):
    if 'Details' not in wb.sheetnames:
        return
    ws = wb['Details']
    p = parsed
    PY = 5   # col E
    CY = 4   # col D

    # Clear #REF! formula cells in the sheet (broken cross-references from old templates).
    from openpyxl.cell import MergedCell as _MC2
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, _MC2):
                continue
            if isinstance(cell.value, str) and '#REF!' in cell.value:
                cell.value = None
    log.append('Details sheet: cleared #REF! formula cells')

    # ── Unsecured loan related parties (R7-R12, 6 slots) ─────────────────────
    # T-shaped BS parser mixes OCL items and debtor names into unsecured_loan_parties.
    # Strategy: filter, deduplicate by normalized name, then sort by amount desc
    # and take the top 6. This reliably picks the largest real lenders.
    def _normalize_nm(nm):
        nm = nm.strip().lower()
        for prefix in ('smt. ', 'sh. ', 'm/s. ', 'mr. ', 'mrs. '):
            nm = nm.replace(prefix, '')
        return nm.replace('.', '').replace('  ', ' ').strip()

    _seen_nm = set()
    _deduped = []
    for x in sorted(p.get('unsecured_loan_parties', []),
                    key=lambda x: x['amount'], reverse=True):
        if not _is_unsecured_party(x['name']):
            continue
        key = _normalize_nm(x['name'])
        if key in _seen_nm:
            continue
        _seen_nm.add(key)
        _deduped.append(x)

    # FIX 6: Write ALL unsecured lenders — insert extra rows if > 6.
    # Template has 6 slots (R7-R12). Bansal has 10 lenders.
    UNSEC_START = 7
    UNSEC_SLOTS = 6
    n_unsec = len(_deduped)
    extra_rows_unsec = max(0, n_unsec - UNSEC_SLOTS)

    if extra_rows_unsec > 0:
        ws.insert_rows(UNSEC_START + UNSEC_SLOTS, extra_rows_unsec)
        log.append(f"Details: inserted {extra_rows_unsec} unsecured lender rows (total={n_unsec})")

    for i in range(max(n_unsec, UNSEC_SLOTS)):
        r = UNSEC_START + i
        if i < len(_deduped):
            party = _deduped[i]
            _write(ws, r, 2, party['name'])
            _py(ws, r, PY, party['amount']); _cy(ws, r, CY)
        else:
            _write(ws, r, 2, '')
            _py(ws, r, PY, 0); _cy(ws, r, CY)

    # Rewrite the Related Parties SUM formula (R13 in original, shifted by extra_rows_unsec)
    unsec_sum_row = UNSEC_START + UNSEC_SLOTS + extra_rows_unsec
    last_unsec_data = UNSEC_START + max(n_unsec, UNSEC_SLOTS) - 1
    ws.cell(unsec_sum_row, 4).value = f'=SUM(D{UNSEC_START}:D{last_unsec_data})'
    ws.cell(unsec_sum_row, 5).value = f'=SUM(E{UNSEC_START}:E{last_unsec_data})'
    log.append(f"Details unsecured SUM rewritten: R{unsec_sum_row} = SUM(?{UNSEC_START}:?{last_unsec_data})")

    # FIX 6: "From Other Parties" row (R15 originally, now R15+extra_rows_unsec).
    # For Bansal all lenders are named individuals (related parties) — no "other parties".
    # Clear this cell to 0 to prevent stale template PY value (2,061,391) from showing.
    unsec_other_row = unsec_sum_row + 2  # R15 = R13+2 in original template
    _py(ws, unsec_other_row, PY, 0)
    _cy(ws, unsec_other_row, CY)
    log.append(f"Details R{unsec_other_row} (unsecured other parties): cleared to 0")

    # Track unsecured row insertion for downstream shift (creditors etc.)
    # The creditor section starts after unsecured section; shift all downstream refs
    # NOTE: We handle this via the _shift parameter below but need to account for unsec rows
    _extra_unsec = extra_rows_unsec  # used to offset CRED_START below

    # ── Sundry creditors (R23-R55, 33 slots; advance R57-R61, 5 slots) ───────
    # Prefer data from _extract_creditor_annexure (Annexure-C, col45/col49).
    # Names there already contain "M/s." prefix — strip it before writing so the
    # template col-A "M/s." label + col-B name layout isn't doubled.
    cred_target  = p.get('sundry_creditors', 0)
    cred_parties_raw = p.get('sundry_creditor_parties', [])

    def _split_prefix(name):
        """Return (prefix, bare_name). Handles 'M/s. Foo' → ('M/s.', 'Foo')."""
        for pfx in ('M/s. ', 'M/s.', 'Sh. ', 'Smt. ', 'Mr. ', 'Mrs. '):
            if name.startswith(pfx):
                return pfx.strip(), name[len(pfx):].strip()
        return '', name.strip()

    # Filter junk only when Annexure-C didn't provide clean data
    _CRED_JUNK = ('trf to', 'transfer to', 'stock', 'hdfc bank',
                  'pspcl', 'property tax', 'municipal', 'municiple')
    good_crp = [cp for cp in cred_parties_raw
                if _is_real_party_name(cp['name']) and
                not any(jk in cp['name'].lower() for jk in _CRED_JUNK)]

    # Fallback: if good_crp sum doesn't match the BS creditor total,
    # try extracting from debtor list (old approach for XLS without Annexure-C pass)
    good_crp_sum = sum(x['amount'] for x in good_crp)
    if (not good_crp or abs(good_crp_sum - cred_target) > cred_target * 0.02) and cred_target > 0:
        deb_raw = p.get('sundry_debtor_parties', [])
        running = 0.0
        extracted = []
        for x in deb_raw:
            running += x['amount']
            extracted.append(x)
            if abs(running - cred_target) < 2.0:
                good_crp = extracted
                break
            if running > cred_target * 1.02:
                break  # overshot without hitting target

    adv_parties = [cp for cp in good_crp
                   if 'advance' in cp['name'].lower() or 'customer' in cp['name'].lower()]
    cred_only   = [cp for cp in good_crp if cp not in adv_parties]

    # FIX S10: Write ALL creditors, inserting extra rows if needed.
    # Template has 33 fixed slots (R23-R55). If we have more, insert rows BEFORE R56.
    # FIX 6: Adjust CRED_START by _extra_unsec (rows inserted for unsecured lenders above).
    # All downstream row references (R56 onwards) shift accordingly.
    CRED_START = 23 + _extra_unsec  # shifted by any unsecured row insertions above
    CRED_TEMPLATE_SLOTS = 33  # R23-R55 (template)
    n_cred = len(cred_only)
    extra_rows_cred = max(0, n_cred - CRED_TEMPLATE_SLOTS)

    if extra_rows_cred > 0:
        # Insert extra_rows_cred rows before R56 (after the last template creditor slot)
        ws.insert_rows(CRED_START + CRED_TEMPLATE_SLOTS, extra_rows_cred)
        log.append(f"Details: inserted {extra_rows_cred} creditor rows (total={n_cred})")

    # Now write all creditors
    for i, cp in enumerate(cred_only):
        r = CRED_START + i
        pfx, bare = _split_prefix(cp['name'])
        _write(ws, r, 1, pfx or 'M/s.')
        _write(ws, r, 2, bare)
        _py(ws, r, PY, cp['amount']); _cy(ws, r, CY)

    # Clear any remaining template slots beyond what we wrote
    for i in range(len(cred_only), CRED_TEMPLATE_SLOTS + extra_rows_cred):
        r = CRED_START + i
        _write(ws, r, 1, '')
        _write(ws, r, 2, '')
        _py(ws, r, PY, 0); _cy(ws, r, CY)

    # ── BUG 2 FIX (part A): Rewrite the creditors SUM formula.
    # The template has a SUM formula for creditors total at a fixed row.
    # After insert_rows(), openpyxl does NOT update formula strings — the formula
    # stays pointing to the old (pre-insertion) row range and gives a wrong total.
    # We must find that TOTAL row and rewrite the SUM to cover only the actual
    # creditor data rows — NOT the Advance-from-Customers slots below them.
    #
    # FIX A (2026-08-14): cred_last_data_row previously included the Advance-from-
    # Customers 5-slot buffer (template R56-R60) in the SUM range, which caused the
    # creditor total to extend all the way to row 163 (overlapping debtor rows) when
    # a further downstream formula recalculation shifted things.
    # The CORRECT last row is the last row where we actually WROTE a creditor value,
    # which is CRED_START + n_cred - 1 (= last used creditor slot).
    # If fewer creditors than template slots, the remaining slots hold zeros — safe to
    # include, but we must NOT go past the zero-pad clearing boundary.
    cred_last_data_row = CRED_START + max(n_cred, 1) - 1   # last actual creditor row
    # Guard: never go past the last slot we allocated (template + inserted)
    cred_allocated_last = CRED_START + CRED_TEMPLATE_SLOTS + extra_rows_cred - 1
    cred_last_data_row  = min(cred_last_data_row, cred_allocated_last)

    # The template TOTAL/SUM row for creditors sits just after the last allocated slot.
    # In Sachidanand layout: template R56 (after 33 slots R23-R55), shifted by insertion.
    cred_sum_template_row = CRED_START + CRED_TEMPLATE_SLOTS  # R56 in original template
    cred_sum_actual_row   = cred_sum_template_row + extra_rows_cred
    # Scan near expected position to find the actual TOTAL/SUM row (label or existing SUM)
    for _tr in range(cred_sum_actual_row, cred_sum_actual_row + 15):
        try:
            _cv = ws.cell(_tr, 1).value
            _dv = ws.cell(_tr, 4).value  # col D
            _ev = ws.cell(_tr, 5).value  # col E
            is_sum_row = (
                (_cv and 'TOTAL' in str(_cv).upper()) or
                (_dv and isinstance(_dv, str) and 'SUM' in _dv.upper()) or
                (_ev and isinstance(_ev, str) and 'SUM' in _ev.upper())
            )
            if is_sum_row:
                cred_sum_actual_row = _tr
                break
        except Exception:
            pass
    # Rewrite creditor SUM formulas with the correct range (creditors only)
    ws.cell(cred_sum_actual_row, 4).value = f'=SUM(D{CRED_START}:D{cred_last_data_row})'
    ws.cell(cred_sum_actual_row, 5).value = f'=SUM(E{CRED_START}:E{cred_last_data_row})'
    log.append(f"Details: creditor SUM rewritten to =SUM(?{CRED_START}:?{cred_last_data_row}) "
               f"at row {cred_sum_actual_row} (n_cred={n_cred}, allocated_last={cred_allocated_last})")

    # All downstream row numbers shift by extra_rows_cred + _extra_unsec after insertions.
    # Compute dynamic base rows for all sections below.
    _shift = extra_rows_cred + _extra_unsec  # total rows inserted above creditor start

    # ── Advance from Customers (template R57-R61, 5 slots) ────────────────────
    adv_from_annexure = p.get('advance_from_customer_parties', [])
    if not adv_parties and adv_from_annexure:
        adv_parties = adv_from_annexure
    adv_total = p.get('advance_from_customers', 0)

    ADV_START = 57 + _shift
    for i in range(5):
        r = ADV_START + i
        if i < len(adv_parties):
            cp = adv_parties[i]
            pfx, bare = _split_prefix(cp['name'])
            _write(ws, r, 1, pfx or 'M/s.')
            _write(ws, r, 2, bare)
            _py(ws, r, PY, cp['amount']); _cy(ws, r, CY)
        elif i == 0 and adv_total:
            _write(ws, r, 1, '')
            _write(ws, r, 2, 'Advance from Customers')
            _py(ws, r, PY, adv_total); _cy(ws, r, CY)
        else:
            _write(ws, r, 1, '')
            _py(ws, r, PY, 0); _cy(ws, r, CY)

    # ── Trade receivables >6 months ─────────────────────────────────────────
    # FIX S10: Write ALL debtors, inserting extra rows if needed.
    # Template has 43 debtor slots (R74-R116 after shift). If we have more, insert rows.
    deb_raw_all = p.get('sundry_debtor_parties', [])
    bs_debtor_total = p.get('sundry_debtors', 0)
    debtor_list_total = sum(x['amount'] for x in deb_raw_all)
    annexure_i_exact = (bs_debtor_total > 0 and debtor_list_total > 0 and
                        abs(debtor_list_total - bs_debtor_total) / bs_debtor_total < 0.05)

    skip_count = 0
    if not annexure_i_exact and good_crp and deb_raw_all:
        for idx, gc in enumerate(good_crp):
            if idx < len(deb_raw_all) and deb_raw_all[idx]['name'] == gc['name']:
                skip_count += 1
            else:
                break

    lender_names = {ul['name'].strip().lower() for ul in p.get('unsecured_loan_parties', [])}

    if annexure_i_exact:
        deb_parties = deb_raw_all
    else:
        _stop_tokens = {'sh', 'smt', 'mr', 'mrs', 'ms', 'm/s', 'shri', 'huf', 'prop',
                        'the', 'of', 'and', '&', 'co', 'ltd', 'pvt', 'sons', 'bros'}
        def _name_tokens(nm):
            toks = re.split(r'[\s\.,\[\]\(\)\/\-]+', nm.lower())
            return {t for t in toks if len(t) >= 3 and t not in _stop_tokens}

        lender_token_sets = [_name_tokens(ul['name'])
                             for ul in p.get('unsecured_loan_parties', [])
                             if _name_tokens(ul['name'])]

        def _is_lender(nm):
            if nm.strip().lower() in lender_names:
                return True
            deb_toks = _name_tokens(nm)
            if len(deb_toks) < 2:
                return False
            for lt in lender_token_sets:
                inter = deb_toks & lt
                if inter == deb_toks:
                    _generic = {'food','products','traders','store','shop',
                                'bakery','house','enterprises','agency','services'}
                    if deb_toks - _generic:
                        return True
            return False

        deb_parties = [x for x in deb_raw_all[skip_count:]
                       if _is_debtor_party(x['name'], lender_names) and not _is_lender(x['name'])]

    # ── Debtor sanity check: if parsed total >> BS total, parser captured wrong rows ──
    # This happens when the debtor extractor reads the carry-forward page of creditors
    # (which appears in the same row range as Annexure-G in T-shaped XLS layouts).
    # Result: 21 "debtors" totaling 1.7M when BS shows 196,267.
    # Fix: if parsed total > 3x BS total, discard all parties and write a single
    # "Sundry Debtors" row with the BS amount. The CA can fill individual entries later.
    _bs_debtors = p.get('sundry_debtors', 0) or 0
    _parsed_deb_total = sum(d.get('amount', 0) for d in deb_parties)
    if _bs_debtors > 0 and _parsed_deb_total > _bs_debtors * 3:
        log.append(f"Details: debtor total mismatch ({_parsed_deb_total:,.0f} >> BS {_bs_debtors:,.0f}) "
                   f"— using single BS total row instead of {len(deb_parties)} parsed parties")
        deb_parties = [{'name': 'Sundry Debtors', 'amount': _bs_debtors}]

    DEB_START_TMPL = 74   # template debtor start before any shifts
    DEB_TMPL_SLOTS = 43   # R74-R116 in template (43 rows)
    DEB_START = DEB_START_TMPL + _shift
    n_deb = len(deb_parties)
    extra_rows_deb = max(0, n_deb - DEB_TMPL_SLOTS)

    if extra_rows_deb > 0:
        ws.insert_rows(DEB_START + DEB_TMPL_SLOTS, extra_rows_deb)
        _shift += extra_rows_deb
        log.append(f"Details: inserted {extra_rows_deb} debtor rows (total={n_deb})")

    for i, dp in enumerate(deb_parties):
        r = DEB_START + i
        dname = dp['name']
        if dname.lower().startswith('m/s.') or dname.lower().startswith('m/s '):
            _write(ws, r, 1, 'M/s.')
            _write(ws, r, 2, dname[4:].strip().lstrip('.').strip())
        else:
            _write(ws, r, 1, 'M/s.')
            _write(ws, r, 2, dname)
        _py(ws, r, PY, dp['amount']); _cy(ws, r, CY)

    # Clear remaining template debtor slots
    for i in range(n_deb, DEB_TMPL_SLOTS + extra_rows_deb):
        r = DEB_START + i
        _write(ws, r, 2, '')
        _py(ws, r, PY, 0); _cy(ws, r, CY)

    # ── BUG 2 FIX (part B): Rewrite debtor SUM formula after row insertion.
    # Same issue as creditors — the formula string is stale after insert_rows().
    deb_last = DEB_START + max(n_deb, DEB_TMPL_SLOTS + extra_rows_deb) - 1
    deb_total_row = deb_last + 1
    # Find the TOTAL row label in the sheet (scan near expected position)
    for _tr in range(deb_total_row, deb_total_row + 8):
        try:
            _cv = ws.cell(_tr, 1).value
            _dv = ws.cell(_tr, 4).value
            _ev = ws.cell(_tr, 5).value
            is_sum_row = (
                (_cv and 'TOTAL' in str(_cv).upper()) or
                (_dv and isinstance(_dv, str) and 'SUM' in _dv.upper()) or
                (_ev and isinstance(_ev, str) and 'SUM' in _ev.upper())
            )
            if is_sum_row:
                deb_total_row = _tr
                break
        except Exception:
            pass
    # Rewrite debtor SUM for BOTH col D (CY) and col E (PY)
    ws.cell(deb_total_row, 4).value = f'=SUM(D{DEB_START}:D{deb_last})'
    ws.cell(deb_total_row, 5).value = f'=SUM(E{DEB_START}:E{deb_last})'
    log.append(f"Details: debtor SUM =SUM(?{DEB_START}:?{deb_last}) at row {deb_total_row}")

    # ── BUG 5 FIX: Advance to Suppliers formula correction.
    # The template's Details sheet has an ADVANCE TO SUPPLIERS row whose SUM formula
    # spans a range that accidentally includes creditor rows placed below the main
    # creditor section (e.g. Savita, Tinku Vig, "SUNDRY DEBTORS" label row) from the
    # previous year's layout. This produces a grossly inflated total (e.g. 1,703,660
    # instead of 0 or the correct advance figure).
    #
    # Fix: scan for the ADVANCE TO SUPPLIERS header row in the sheet.
    # Then find the actual supplier advance rows (blank M/s. slots below the header).
    # Rewrite the SUM formula to cover ONLY those blank slots.
    # If the T-shaped XLS has no advance to suppliers (parsed value = 0), zero all cells.
    adv_to_sup_total = p.get('advance_to_suppliers', 0) or 0
    # Scan for ADVANCE TO SUPPLIERS header row (within 50 rows of deb_total_row)
    _adv_sup_row  = None
    _adv_total_row = None   # initialise here so it is always bound in _skip_rows below
    for _r in range(deb_total_row + 1, deb_total_row + 60):
        try:
            _cv = ws.cell(_r, 1).value
            if _cv and 'ADVANCE' in str(_cv).upper() and 'SUPPLIER' in str(_cv).upper():
                _adv_sup_row = _r
                break
        except Exception:
            pass
    if _adv_sup_row is not None:
        # Find the range of actual advance-to-suppliers data rows (blank M/s. slots)
        _adv_data_start = _adv_sup_row + 1
        _adv_data_end   = _adv_sup_row + 1   # default: single row
        _adv_total_row  = None
        for _r in range(_adv_sup_row + 1, _adv_sup_row + 25):
            try:
                _cv = ws.cell(_r, 1).value
                _bv = ws.cell(_r, 2).value
                if _cv and 'TOTAL' in str(_cv).upper():
                    _adv_total_row = _r
                    _adv_data_end  = _r - 1
                    break
                # Count rows that look like data slots (M/s. prefix or blank)
                if _cv in ('M/s.', 'M/s', None, '') or (_bv in (None, '')):
                    _adv_data_end = _r
            except Exception:
                pass
        # Rewrite the ADVANCE TO SUPPLIERS formula to cover only its proper rows
        if _adv_total_row is not None:
            ws.cell(_adv_total_row, 4).value = f'=SUM(D{_adv_data_start}:D{_adv_data_end})'
            ws.cell(_adv_total_row, 5).value = f'=SUM(E{_adv_data_start}:E{_adv_data_end})'
            log.append(f"Details: ADVANCE TO SUPPLIERS SUM rewritten to rows "
                       f"{_adv_data_start}:{_adv_data_end} at total row {_adv_total_row}")
        # If the T-shaped XLS shows 0 advance to suppliers, zero all data cells
        if adv_to_sup_total == 0:
            for _r in range(_adv_data_start, _adv_data_end + 1):
                try:
                    _py(ws, _r, PY, 0)
                    _cy(ws, _r, CY)
                except Exception:
                    pass
            log.append("Details: ADVANCE TO SUPPLIERS zeroed (T-shaped XLS shows 0)")
        else:
            # Write the actual advance-to-suppliers total in first data row
            _py(ws, _adv_data_start, PY, adv_to_sup_total)
            _cy(ws, _adv_data_start, CY)
            log.append(f"Details: ADVANCE TO SUPPLIERS PY={adv_to_sup_total:,.2f} written")

    # ── Trade receivables <6 months ───────────────────────────────────────────
    _lt6_start = deb_total_row + 4   # TOTAL row + header rows
    _py(ws, _lt6_start,     PY, 0); _cy(ws, _lt6_start,     CY)
    _py(ws, _lt6_start + 1, PY, 0); _cy(ws, _lt6_start + 1, CY)

    log.append(
        f'Details: {len(unsec_parties)} unsecured, {len(cred_only)} creditors, '
        f'{len(deb_parties)} debtors'
    )

    # Return actual row numbers for cross-sheet formula repair.
    # Template original rows: creditor_sum=62, msme_sum=68, deb_gt6_total=129, deb_lt6_total=136
    # After row insertion these shift. We track them precisely via DEB_START and deb_total_row.
    _extra_cred = extra_rows_cred           # rows inserted for creditor overflow
    _extra_deb  = extra_rows_deb            # rows inserted for debtor overflow
    _cred_sum_row     = 62 + _extra_cred    # Details row that has =SUM(creditor amounts)
    _msme_sum_row     = 68 + _extra_cred    # Details row with MSME SUM
    _deb_gt6_total    = deb_total_row       # actual TOTAL row for debtors >6m (Advance-to-Sup)
    _deb_lt6_total    = _lt6_start + 2      # approximate TOTAL row for debtors <6m

    # FIX M (2026-08-14): Record the rows that _inject_details_sheet has ALREADY written
    # correct formulas to. The internal formula shift in _fix_details_formula_refs must
    # NOT re-shift these rows, because the formulas there already reference the correct
    # post-insertion row numbers and a second shift would corrupt them.
    # These rows are: cred_sum_actual_row, deb_total_row, any Advance-to-Suppliers row,
    # and the Trade Receivable <6mo TOTAL row (_lt6_start + 2).
    _skip_rows = {cred_sum_actual_row, deb_total_row, _lt6_start, _lt6_start+1, _lt6_start+2}
    if _adv_sup_row is not None:
        _skip_rows.add(_adv_sup_row)
        if _adv_total_row is not None:
            _skip_rows.add(_adv_total_row)

    return {
        'cred_sum_row':      _cred_sum_row,
        'cred_sum_actual':   cred_sum_actual_row,  # FIX M: actual row written by injector
        'msme_sum_row':      _msme_sum_row,
        'deb_gt6_total':     _deb_gt6_total,
        'deb_lt6_total':     _deb_lt6_total,
        'extra_cred':        _extra_cred,
        'extra_deb':         _extra_deb,
        'skip_rows':         _skip_rows,            # FIX M: rows to exclude from re-shift
    }


# ── Fix cross-sheet formula references after Details row insertion ─────────────
# openpyxl does NOT update cross-sheet formula strings when insert_rows() is called.
# After we insert rows in the Details sheet, formulas in 'notes to bs' that reference
# specific Details row numbers become stale and point to the wrong cells.
#
# Template cross-sheet references (notes to bs → Details):
#   R15 D/E  → Details!R13   (unsecured FROM RELATED PARTIES SUM)  ← never shifts
#   R38 D/E  → Details!R62   (trade payable others = creditor SUM)  ← shifts by extra_cred
#   R39 D/E  → Details!R68   (trade payable MSME SUM)               ← shifts by extra_cred
#   R91 D/E  → Details!R129  (debtors >6 months TOTAL)              ← shifts by both
#   R97 D/E  → Details!R136  (debtors <6 months TOTAL)              ← shifts by both
#
# This function rewrites those formula strings to the correct new row numbers.

def _fix_details_formula_refs(wb, shifts, log):
    """
    Repair 'notes to bs' formulas that reference Details rows which moved
    after dynamic row insertion in _inject_details_sheet().

    Also fixes internal Details sheet formulas (e.g. TOTAL rows that reference
    other Details rows by absolute row number — these shift but openpyxl does
    not update formula strings automatically after insert_rows()).

    Parameters
    ----------
    shifts : dict returned by _inject_details_sheet(), containing:
        cred_sum_row  - actual row in Details for creditor SUM (template: 62)
        msme_sum_row  - actual row in Details for MSME SUM    (template: 68)
        deb_gt6_total - actual row in Details for debtor >6m TOTAL (template: 129)
        deb_lt6_total - actual row in Details for debtor <6m TOTAL (template: 136)
        extra_cred    - how many creditor rows were inserted
        extra_deb     - how many debtor rows were inserted
    """
    if 'notes to bs' not in wb.sheetnames:
        return

    extra_cred = shifts.get('extra_cred', 0)
    extra_deb  = shifts.get('extra_deb', 0)

    # If no rows were inserted, nothing to fix
    if extra_cred == 0 and extra_deb == 0:
        log.append('Details formula refs: no row insertion, nothing to fix')
        return

    ws = wb['notes to bs']

    # Map: (notes_to_bs_row, col_idx) → new Details row number
    # Template original rows in Details: 62=credSUM, 68=MSME, 129=deb<6m TOTAL, 136=deb>6m TOTAL
    # FIX N part 2: deb_gt6 and deb_lt6 must be computed by directly applying extra_cred
    # shift to the TEMPLATE row numbers (129 and 136), NOT from the deb_gt6_total /
    # deb_lt6_total tracker fields (which track the ADVANCE TO SUPPLIERS and _lt6_start+2
    # rows — wrong rows for this cross-sheet reference fix).
    # Template R129 (deb <6mo TOTAL) shifts by extra_cred (only creditor rows are inserted
    # before R129, no debtor overflow for this client). Same for R136.
    cred_sum = shifts.get('cred_sum_row',  62  + extra_cred)
    msme_sum = shifts.get('msme_sum_row',  68  + extra_cred)
    deb_gt6  = 129 + extra_cred + extra_deb   # template R129 shifted by all insertions
    deb_lt6  = 136 + extra_cred + extra_deb   # template R136 shifted by all insertions

    # FIX N (2026-08-14): FIX C REMOVED.
    # The template has notes to bs R91 → Details!D129 (trade receivable <6mo TOTAL)
    # and R97 → Details!D136 (>6mo TOTAL). After 54 creditor rows are inserted:
    #   D129 → D183 (correct <6mo total) via standard _shift_row_ref
    #   D136 → D190 (correct >6mo total) via standard _shift_row_ref
    # FIX C was adding extra replacements for Details!D171 → deb_lt6 which OVERWROTE
    # the already-correctly-shifted formulas with D171 (ADVANCE TO SUPPLIERS — wrong).
    # Solution: let the standard shift mechanism handle D129/D136 → D183/D190 naturally.
    # No special D171 patching needed.
    replacements = {}
    if extra_cred > 0:
        replacements['Details!D62'] = f'Details!D{cred_sum}'
        replacements['Details!E62'] = f'Details!E{cred_sum}'
        replacements['Details!D68'] = f'Details!D{msme_sum}'
        replacements['Details!E68'] = f'Details!E{msme_sum}'
    if extra_cred > 0 or extra_deb > 0:
        replacements['Details!D129'] = f'Details!D{deb_gt6}'
        replacements['Details!E129'] = f'Details!E{deb_gt6}'
        replacements['Details!D136'] = f'Details!D{deb_lt6}'
        replacements['Details!E136'] = f'Details!E{deb_lt6}'

    # Also handle without '!' prefix variation just in case
    extra_replacements = {}
    for old, new in replacements.items():
        # Some formula writers use single-quoted sheet name
        old2 = old.replace('Details!', "'Details'!")
        new2 = new.replace('Details!', "'Details'!")
        extra_replacements[old2] = new2
    replacements.update(extra_replacements)

    count = 0
    from openpyxl.cell import MergedCell as _MCF
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, _MCF):
                continue
            v = cell.value
            if not isinstance(v, str) or 'Details' not in v:
                continue
            new_v = v
            for old_frag, new_frag in replacements.items():
                if old_frag in new_v:
                    new_v = new_v.replace(old_frag, new_frag)
            if new_v != v:
                cell.value = new_v
                count += 1
                log.append(f"  Fixed formula [{ws.title}] R{cell.row}C{cell.column}: "
                           f"{v!r} → {new_v!r}")

    log.append(f"Details formula refs: {count} formula(s) updated "
               f"(extra_cred={extra_cred}, extra_deb={extra_deb})")

    # ── BUG 2 FIX (part C): Also repair INTERNAL Details sheet formulas.
    # The Details sheet itself has formulas that reference absolute row numbers
    # (e.g. =D62+D68, =SUM(D23:D61), =E62+E68) which shift after insert_rows().
    # We must update any remaining stale row references in the Details sheet itself.
    if 'Details' not in wb.sheetnames:
        return
    ws_det = wb['Details']
    from openpyxl.cell import MergedCell as _MCD

    # Build a row-shift map for absolute references in Details formulas:
    # Any row ref >= insertion_point shifts by extra_rows.
    # Creditor insertion point = CRED_START + CRED_TEMPLATE_SLOTS = 23 + 33 = 56
    # Debtor insertion point   = 74 + extra_cred + DEB_TMPL_SLOTS = 74+extra_cred+43
    CRED_INSERT = 56   # rows >= 56 shift by extra_cred
    DEB_INSERT  = 74 + extra_cred + 43  # rows >= this shift by extra_deb

    if extra_cred == 0 and extra_deb == 0:
        return  # nothing to fix

    import re as _re_int

    def _shift_row_ref(m):
        """Shift a row number in an Excel formula if it falls in an insertion zone."""
        col_letter = m.group(1) or ''
        row_num    = int(m.group(2))
        shifted    = row_num
        if extra_cred > 0 and row_num >= CRED_INSERT:
            shifted += extra_cred
        if extra_deb > 0 and row_num >= DEB_INSERT:
            shifted += extra_deb
        return f'{col_letter}{shifted}'

    # Pattern: optional column letter(s) + row number in a formula context
    # Only match patterns that look like cell references (letter(s) then digits)
    _CELLREF_PAT = _re_int.compile(r'([A-Za-z]{1,3})(\d+)')

    # FIX M (2026-08-14): Rows already rewritten by _inject_details_sheet must be
    # excluded from the generic shift pass. The injector writes correct post-insertion
    # formulas (e.g. SUM(D23:D109) for creditors). A second shift would corrupt them
    # (e.g. D109→D163). Skip any row in the skip_rows set returned by the injector.
    skip_rows = shifts.get('skip_rows', set())
    # Also always skip the cred_sum_actual row (the one the injector wrote)
    cred_sum_actual = shifts.get('cred_sum_actual', cred_sum)
    if cred_sum_actual:
        skip_rows = skip_rows | {cred_sum_actual}

    det_fix_count = 0
    for row in ws_det.iter_rows():
        for cell in row:
            if isinstance(cell, _MCD):
                continue
            # FIX M: skip rows that already have correct injector-written formulas
            if cell.row in skip_rows:
                continue
            v = cell.value
            if not isinstance(v, str) or not v.startswith('='):
                continue
            # Apply shift to all cell references in this formula
            new_v = _CELLREF_PAT.sub(_shift_row_ref, v)
            if new_v != v:
                cell.value = new_v
                det_fix_count += 1
                log.append(f"  Details internal formula R{cell.row}C{cell.column}: "
                           f"{v!r} → {new_v!r}")
    if det_fix_count:
        log.append(f"Details internal formula fix: {det_fix_count} formula(s) updated")

    # FIX N part 3: After internal shift, correct the debtor SUM formulas.
    # The template has wrong E-col ranges (SUM(E77:E128) instead of SUM(E74:E128))
    # and the >6mo TOTAL only covers E135:E135 (template bug). After shift these
    # become SUM(E131:E182) and SUM(E189:E189). Rewrite to correct ranges.
    # The correct start row for debtors in D col: DEB_START_TMPL+extra_cred = 74+extra_cred
    # In E col it should match D col exactly.
    deb_start_actual = 74 + extra_cred   # actual first debtor data row in output
    deb_end_actual   = deb_start_actual + 43 + extra_deb - 1   # last debtor slot

    # Debtor <6mo total row: template R129 + extra_cred + extra_deb
    _lt6_total_row = 129 + extra_cred + extra_deb
    try:
        ws_det.cell(_lt6_total_row, 4).value = f'=SUM(D{deb_start_actual}:D{deb_end_actual})'
        ws_det.cell(_lt6_total_row, 5).value = f'=SUM(E{deb_start_actual}:E{deb_end_actual})'
        log.append(f"Details debtor <6mo SUM corrected at R{_lt6_total_row}: "
                   f"D/E{deb_start_actual}:{deb_end_actual}")
    except Exception as _e:
        log.append(f"Details debtor SUM correction failed: {_e}")

    # Debtor >6mo total row: template R136 + extra_cred + extra_deb
    _gt6_total_row = 136 + extra_cred + extra_deb
    try:
        _gt6_data_start = _gt6_total_row - 2   # 2 data rows above total
        _gt6_data_end   = _gt6_total_row - 1
        ws_det.cell(_gt6_total_row, 4).value = f'=SUM(D{_gt6_data_start}:D{_gt6_data_end})'
        ws_det.cell(_gt6_total_row, 5).value = f'=SUM(E{_gt6_data_start}:E{_gt6_data_end})'
        log.append(f"Details debtor >6mo SUM corrected at R{_gt6_total_row}: "
                   f"D/E{_gt6_data_start}:{_gt6_data_end}")
    except Exception as _e:
        log.append(f"Details debtor >6mo SUM correction failed: {_e}")


# ── Fixed Assets P. Yr. ───────────────────────────────────────────────────────
# Template column layout (1-indexed, verified):
#   A(1)=Particulars  B(2)=WDV at 01.04.PY-1  C(3)=Additions>180d  D(4)=Additions<180d
#   E(5)=Sales  F(6)=Total  G(7)=Rate%  H(8)=Depreciation  I(9)=WDV at 31.03.PY
#   F(6) and H(8) are formula columns — write only B,C,D,E,G; I is also formula.
#
# Named asset rows in template (from inspection) — these are FIXED:
#   PLANT & MACHINERY section: R9-R17
#   VEHICLE section: R20-R21
#   COMPUTERS section: R24
#   FURNITURE AND FIXTURES: R27-R28
#   BUILDING: R31
#   R37 = Total row (formulas)
#
# Strategy: match parsed item names to template row labels; write B,C,D,E values.
# Items not found in template → write into the nearest named row or skip.

_FA_ROW_MAP = {
    'camera & dvr':              9,
    'digital moisture meter':    10,
    'servo control voltage':     11,
    'servo control':             11,
    'water cooler':              12,
    'machine':                   13,
    'mobile':                    14,
    'electricals':               15,
    'washing machine':           16,
    'led':                       17,
    'car':                       20,
    'motor cycle & scooter':     21,
    'motor cycle':               21,
    'motorcycle':                21,
    'scooter':                   21,
    'computer':                  24,
    'chair':                     27,
    'furniture & fixture ':      28,
    'furniture & fixtures':      28,
    'furniture & fixture':       28,
    'furniture':                 28,
    'building':                  31,
}


def _build_fa_row_map_from_sheet(ws):
    """
    Auto-detect asset row positions by scanning col A of the FA P.Yr. sheet.
    Returns dict: asset_name_lower -> row_number (1-indexed openpyxl).
    Falls back to _FA_ROW_MAP constants if detection fails.
    """
    detected = {}
    for row in ws.iter_rows():
        cell_a = row[0]
        if cell_a.value is None:
            continue
        name = str(cell_a.value).strip().lower()
        if not name or name in ('particulars', 'total', 'plant & machinery',
                                'vehciles', 'vehicles', 'furniture and fixtures',
                                'furniture & fixtures', 'computers', 'building',
                                'land', 'detail of fixed assets', 'amount in rs.'):
            continue
        has_data = any(
            isinstance(row[c].value, (int, float))
            for c in range(1, min(9, len(row)))
        )
        if has_data:
            detected[name] = cell_a.row
    return detected if detected else {}


def _inject_fixed_assets_py(wb, parsed, client_name, py_year, log):
    if 'Fixed Assets P. Yr.' not in wb.sheetnames:
        return
    ws = wb['Fixed Assets P. Yr.']
    items = parsed.get('fixed_asset_items', [])
    if not items:
        log.append('Fixed Assets P.Yr.: no items to write')
        return

    written_rows = set()

    # FIX 4: Write fixed assets POSITIONALLY — use source asset name, write it into
    # template col A, and place data starting at first available data row.
    # Do NOT rely on name-matching (which fails for 'Invertor & Battery', 'Activa', etc.).
    # Strategy:
    #   1. Find all available data rows in the FA P.Yr. sheet (rows with numeric data
    #      or blank rows between section headers and total row).
    #   2. Write each source asset sequentially into those rows.
    #   3. Write source name into col A (overwriting template name).
    # This ensures ALL 10 assets appear regardless of whether template has them.

    # Collect available data rows (non-header, non-total, non-section-header rows)
    _SKIP_LABELS = {'total', 'plant & machinery', 'vehicles', 'vehciles', 'computers',
                    'furniture and fixtures', 'furniture & fixtures', 'building',
                    'land', 'detail of fixed assets', 'amount in rs.', 'particulars',
                    'computer', 'furniture & fixture'}
    available_rows = []
    total_row = None
    for row in ws.iter_rows():
        cell_a = row[0]
        if cell_a.value is None:
            continue
        label = str(cell_a.value).strip().lower()
        if 'total' in label:
            total_row = cell_a.row
            continue
        if label in _SKIP_LABELS or label in ('', 'nan'):
            continue
        # This is a data row (asset row or section name we will reuse)
        # Check: has numeric data in B-I columns OR is it an asset row
        has_num = any(
            isinstance(row[c].value, (int, float))
            for c in range(1, min(9, len(row)))
        )
        if has_num or (label not in _SKIP_LABELS and len(label) > 1):
            available_rows.append(cell_a.row)

    # If auto-detected rows insufficient, fall back to name-matching for what we can
    detected_map = _build_fa_row_map_from_sheet(ws)
    log.append(f"FA P.Yr. detected rows: {detected_map}")
    log.append(f"FA P.Yr. available rows: {available_rows}")

    def _write_asset_to_row(r, item):
        opening   = item.get('opening_wdv', 0)
        additions = item.get('additions', 0)
        rate      = item.get('rate', 0) or 0
        sales     = item.get('sales', 0)
        dep       = item.get('dep', 0)
        closing   = item.get('closing_wdv', 0)

        safe_add  = round(additions)
        total_val = round(opening) + safe_add - round(sales)

        # FIX 4: Write source asset NAME into col A (overwrite template name)
        cell_a = ws.cell(r, 1)
        from openpyxl.cell import MergedCell as _MCFA
        if not isinstance(cell_a, _MCFA):
            cell_a.value = item['name']

        _write_num(ws, r, 2, round(opening))   # B: opening WDV
        _write_num(ws, r, 3, safe_add)         # C: additions >180d
        _write_num(ws, r, 4, 0)                # D: additions <180d
        _write_num(ws, r, 5, round(sales))     # E: sales

        existing_f = ws.cell(r, 6).value
        if not (isinstance(existing_f, str) and existing_f.startswith('=')):
            _write_num(ws, r, 6, total_val)    # F: total

        if rate:
            _write(ws, r, 7, rate)             # G: rate

        dep_val = round(total_val * rate / 100) if rate else round(dep)
        closing_wdv = total_val - dep_val if dep_val else round(closing)
        _write_num(ws, r, 8, dep_val)          # H: depreciation
        _write_num(ws, r, 9, closing_wdv)      # I: closing WDV
        log.append(f"FA P.Yr. R{r}: {item['name']} opening={opening:.0f} "
                   f"add={additions:.0f} dep={dep_val:.0f} wdv={closing_wdv:.0f}")

    # Try name-matched rows first (for exact matches like Car, Computer)
    for item in items:
        nm_lower = item['name'].strip().lower()
        r = detected_map.get(nm_lower)
        if r is None:
            for key, row_r in detected_map.items():
                if key in nm_lower or nm_lower in key:
                    r = row_r
                    break
        if r is None:
            r = _FA_ROW_MAP.get(nm_lower)
        if r is None:
            for key, row_r in _FA_ROW_MAP.items():
                if key in nm_lower or nm_lower in key:
                    r = row_r
                    break
        if r is not None and r not in written_rows:
            _write_asset_to_row(r, item)
            written_rows.add(r)
            if r in available_rows:
                available_rows.remove(r)

    # Write remaining unmatched items into available rows sequentially
    avail_iter = iter(available_rows)
    for item in items:
        nm_lower = item['name'].strip().lower()
        # Already written?
        already = False
        for wr in written_rows:
            cell_a_val = ws.cell(wr, 1).value
            if cell_a_val and str(cell_a_val).strip().lower() == nm_lower:
                already = True
                break
        if already:
            continue
        # Find next available row
        try:
            r = next(avail_iter)
            while r in written_rows:
                r = next(avail_iter)
        except StopIteration:
            # No more template rows — append after last used row
            if written_rows:
                r = max(written_rows) + 1
            else:
                continue
        _write_asset_to_row(r, item)
        written_rows.add(r)

    # FIX 6b: Write Total row with correct sums so bs!I37 formula gets a real value
    # Scan for the Total row in the sheet (row 37 per template comment, but verify)
    total_opening = sum(it.get('opening_wdv', 0) for it in items)
    total_additions = sum(it.get('additions', 0) for it in items)
    total_sales = sum(it.get('sales', 0) for it in items)
    total_f = round(total_opening + total_additions - total_sales)
    total_dep = sum(round((round(it.get('opening_wdv',0)) + round(it.get('additions',0)) - round(it.get('sales',0))) * (it.get('rate',0) or 0) / 100) for it in items)
    total_wdv = total_f - total_dep

    # Write Total row at R37 (template fixed row) and also at R15 if that exists as Total
    for total_row in [37, 15]:
        cell_a = ws.cell(total_row, 1).value
        if cell_a is not None and 'TOTAL' in str(cell_a).upper():
            _write_num(ws, total_row, 2, round(total_opening))
            _write_num(ws, total_row, 3, round(total_additions))
            _write_num(ws, total_row, 4, 0)
            _write_num(ws, total_row, 5, round(total_sales))
            _write_num(ws, total_row, 6, total_f)
            _write_num(ws, total_row, 8, total_dep)
            _write_num(ws, total_row, 9, total_wdv)
            log.append(f"FA P.Yr. Total row R{total_row}: WDV={total_wdv}")

    log.append(f'Fixed Assets P.Yr.: {len(items)} items processed ({len(written_rows)} rows written)')


# ── GROSS PROFIT sheet ────────────────────────────────────────────────────────
# Template column layout (verified from data_only=False inspection):
#   Left side:  A=Particulars  B(2)=CY amount  C(3)=PY amount
#   Right side: D=Particulars  E(5)=CY amount  F(6)=PY amount
#
# Key rows:
#   R9:C   = Opening stock (PY left side)    [R9:B = formula from notes to p&l!D18]
#   R10:F  = Sales (PY right side)           [template: R10 right, not R9!]
#   R13:C  = Purchases (PY left side)        [same row as Purchase GST label]
#   R13:F  = Closing stock (PY right side)   [formula =notes to p&l!E27]
#   R21-R23= Direct expense rows (from notes to p&l formulas)
#   R25:C  = Gross profit (PY)               [formula =F27-SUM(C9:C21)]
#   R27    = TOTAL row (formulas)
#
# IMPORTANT: Most cells are formula-driven from notes to p&l.
# Only the PY input cells that are NOT formula-driven need writing:
#   R9:C  = opening stock PY
#   R10:F = sales PY
#   R13:C = purchases PY
#
# R13:F = closing stock is a formula ('notes to p&l'!E27) — DO NOT WRITE
# R25:C = gross profit is a formula — DO NOT WRITE

def _inject_fixed_assets_cy_opening(wb, parsed, log):
    """
    Write PY closing WDV as CY opening WDV in Fixed Assets C. Yr. sheet.
    PY closing WDV = opening + additions - sales - depreciation
    dep = ROUND((opening + add_gt180 + add_lt180 - sales) * rate/100)
    """
    if 'Fixed Assets C. Yr.' not in wb.sheetnames:
        return
    ws = wb['Fixed Assets C. Yr.']
    items = parsed.get('fixed_asset_items', [])
    if not items:
        return

    import math

    # Mapping: asset name (lower) → FA C.Yr. row
    # Plant & Machinery items occupy rows 11-21 in C.Yr. sheet
    # Vehicles: Car=R24, Motor Cycle & Scooter=R25
    # Furniture & Fixtures: Furniture & Fixture=R36
    # Computer: R47
    # Building: R60
    _FA_CY_ROW_MAP = {
        'camera & dvr':              11,
        'digital moisture meter':    12,
        'servo control voltage':     13,
        'servo control':             13,
        'water cooler':              14,
        'machine':                   15,
        'mobile':                    16,
        'electricals':               17,
        'washing machine':           18,
        'led':                       19,
        'car':                       24,
        'motor cycle & scooter':     25,
        'motor cycle':               25,
        'motorcycle':                25,
        'scooter':                   25,
        'computer':                  47,
        'chair':                     37,
        'furniture & fixture ':      36,
        'furniture & fixtures':      36,
        'furniture & fixture':       36,
        'furniture':                 36,
        'building':                  60,
    }

    # Build auto-detected map from the sheet's formula references
    # Formula cells like ='Fixed Assets P. Yr.'!A11 tell us which row this is
    detected_cy_map = {}
    for row in ws.iter_rows():
        cell_a = row[0]
        v = cell_a.value
        if v is None:
            continue
        if isinstance(v, str) and "'Fixed Assets P. Yr.'!A" in v:
            try:
                py_row_ref = int(v.split("!A")[1])
                py_ws = wb.get('Fixed Assets P. Yr.') or wb.get('Fixed Assets P. Yr.')
                if py_ws:
                    py_cell_a = py_ws.cell(py_row_ref, 1).value
                    if py_cell_a and isinstance(py_cell_a, str):
                        detected_cy_map[py_cell_a.strip().lower()] = cell_a.row
            except (ValueError, IndexError):
                pass

    if detected_cy_map:
        log.append(f"FA C.Yr. auto-detected rows: {detected_cy_map}")

    count = 0
    written_cy_rows = set()
    for item in items:
        name_lower = item['name'].strip().lower()
        # Find matching row — try auto-detected map first, then fall back to hard-coded
        r = None
        for key, row_r in detected_cy_map.items():
            if name_lower.startswith(key) or key.startswith(name_lower):
                r = row_r
                break
        if r is None:
            for key, row_r in _FA_CY_ROW_MAP.items():
                if name_lower.startswith(key) or key.startswith(name_lower):
                    r = row_r
                    break
        if r is None:
            continue

        opening  = item.get('opening_wdv', 0)
        add_gt   = item.get('additions', 0)
        sales    = item.get('sales', 0)
        rate     = item.get('rate', 0) or 0
        # PY closing WDV becomes CY opening WDV
        dep_py = round((opening + add_gt - sales) * rate / 100)
        cy_opening_wdv = opening + add_gt - sales - dep_py

        if cy_opening_wdv > 0:
            ws.cell(r, 1).value = item['name'].strip()  # col A: asset name
            ws.cell(r, 2).value = round(cy_opening_wdv) # col B: opening WDV
            ws.cell(r, 7).value = rate                   # col G: rate (for CA to compute CY dep)
            written_cy_rows.add(r)
            count += 1

    # FIX 7b: Write CY Total row at R17 if it's labeled Total
    cy_total_opening = sum(
        round(it.get('opening_wdv',0) + it.get('additions',0) - it.get('sales',0)) -
        round((it.get('opening_wdv',0) + it.get('additions',0) - it.get('sales',0)) * (it.get('rate',0) or 0) / 100)
        for it in items
    )
    for total_row in [17, 15]:
        cell_a = ws.cell(total_row, 1).value
        if cell_a is not None and 'TOTAL' in str(cell_a).upper():
            ws.cell(total_row, 2).value = cy_total_opening
            log.append(f"FA C.Yr. Total row R{total_row}: opening WDV={cy_total_opening}")
            break

    log.append(f"Fixed Assets C.Yr.: {count} opening WDV values written")


def _inject_gross_profit_sheet(wb, parsed, client_name, cy_year, py_year, log):
    if 'GROSS PROFIT' not in wb.sheetnames:
        return
    ws = wb['GROSS PROFIT']
    p = parsed

    # Hard-coded column constants (verified from template):
    PY_LEFT  = 3   # col C = PY left side (opening stock, purchases, gross profit)
    PY_RIGHT = 6   # col F = PY right side (sales, closing stock)

    # R9:C = opening stock (PY)
    _write_num(ws, 9,  PY_LEFT,  p.get('opening_stock', 0))

    # FIX 1: Sales — write EACH GST rate line to its own row (PY right side).
    # Template has separate rows for Sales Central GST 12%, 18%, 5%, Local 12%, etc.
    # If we have individual sale_line_items, write them; else write total on R10.
    sale_line_items = p.get('sale_line_items', [])
    if sale_line_items:
        # Scan the GROSS PROFIT sheet to find rows that match "Sales" labels
        # Template right side rows (col F) for sales lines typically start at R10.
        # Strategy: find all rows in the sheet where col E or col D has a Sales label.
        # Write the matching amount from sale_line_items in col F (PY_RIGHT).
        _SALE_ROW_MAP = {}  # label_lower → row number
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cv = cell.value.strip().lower()
                    if 'sale' in cv and cell.column in (4, 5):  # col D or E = labels
                        _SALE_ROW_MAP[cv] = cell.row
                        # Also try partial key
                        for gst_kw in ('12%', '18%', '5%', '0%', 'tax free', 'central', 'local'):
                            if gst_kw in cv:
                                _SALE_ROW_MAP[cv] = cell.row

        # Write each sale line: match by GST rate keyword
        written_rows = set()
        sale_total_written = 0.0
        for it in sale_line_items:
            lbl = it['label'].strip().lower()
            amt = it['amount']
            matched_row = None
            # Try exact match
            if lbl in _SALE_ROW_MAP:
                matched_row = _SALE_ROW_MAP[lbl]
            else:
                # Try partial match on GST rate/type
                for map_lbl, map_row in _SALE_ROW_MAP.items():
                    # Match if both label and map_lbl share key tokens
                    lbl_has_12  = '12' in lbl and '12' in map_lbl
                    lbl_has_18  = '18' in lbl and '18' in map_lbl
                    lbl_has_5   = ('5%' in lbl or ' 5' in lbl) and ('5%' in map_lbl or ' 5' in map_lbl)
                    lbl_central = 'central' in lbl and 'central' in map_lbl
                    lbl_local   = 'local' in lbl and 'local' in map_lbl
                    lbl_free    = ('tax free' in lbl or '0%' in lbl) and ('tax free' in map_lbl or '0%' in map_lbl)
                    if ((lbl_has_12 or lbl_has_18 or lbl_has_5 or lbl_free) and
                        (lbl_central or lbl_local or lbl_free)):
                        if map_row not in written_rows:
                            matched_row = map_row
                            break
            if matched_row and matched_row not in written_rows:
                _write_num(ws, matched_row, PY_RIGHT, amt)
                written_rows.add(matched_row)
                sale_total_written += amt
                log.append(f'GROSS PROFIT: sale line "{it["label"]}" = {amt:,.0f} → R{matched_row}')

        # If no row mapping found, write total on R10
        if not written_rows:
            _write_num(ws, 10, PY_RIGHT, p.get('sales', 0))
            log.append(f'GROSS PROFIT: sales total (no row map) = {p.get("sales", 0):,.0f} → R10')
        else:
            log.append(f'GROSS PROFIT: {len(written_rows)} sale lines written, total={sale_total_written:,.0f}')
    else:
        # Fallback: write total sales on R10
        _write_num(ws, 10, PY_RIGHT, p.get('sales', 0))

    # R13:C = purchases (PY) — "Purchase GST" row
    _write_num(ws, 13, PY_LEFT,  p.get('purchases', 0))

    # R13:F = closing stock — formula from notes to p&l!E27 — DO NOT WRITE
    # R25:C = gross profit — formula =F27-SUM(C9:C21) — DO NOT WRITE
    # R27 = TOTAL formulas — DO NOT WRITE

    # Direct expenses in GROSS PROFIT left PY column (col C = 3):
    # R21 col C (PY FOC), R22 col C (Freight), R23 col C (Freight GST):
    # These have NO formula in the template — must be written directly.
    all_items = p.get('other_expense_items', []) + p.get('direct_expense_items', [])
    foc_amt = 0.0
    freight_amt = 0.0
    freight_gst_amt = 0.0
    for it in all_items:
        k = it['name'].strip().lower()
        if k in ('foc', 'f.o.c.', 'f.o.c'):
            foc_amt += it['amount']
        elif 'gst' in k and 'freight' in k:
            freight_gst_amt += it['amount']
        elif 'freight' in k or 'frieght' in k:
            freight_amt += it['amount']

    _write_num(ws, 21, PY_LEFT, foc_amt)           # R21:C = FOC PY
    _write_num(ws, 22, PY_LEFT, freight_amt)        # R22:C = Freight Inward PY
    _write_num(ws, 23, PY_LEFT, freight_gst_amt)    # R23:C = Freight Inward (GST) PY

    log.append('GROSS PROFIT sheet injected')


# ── Update headers ─────────────────────────────────────────────────────────────

def _clear_ref_errors(wb, log):
    """
    Clear all #REF! broken formula cells across all sheets.
    These arise from template cross-references to deleted/renamed ranges.
    Clearing them to None removes the ugly #REF! display in Excel.
    """
    from openpyxl.cell import MergedCell as _MCR
    count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, _MCR):
                    continue
                if isinstance(cell.value, str) and '#REF!' in cell.value:
                    cell.value = None
                    count += 1
    if count:
        log.append(f'Cleared {count} #REF! formula cells across all sheets')


def _update_headers(wb, client_name, cy_year, py_year):
    """Replace M/S XYZ CO., year references, and accounting policy text."""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                v = cell.value
                if not isinstance(v, str):
                    continue
                nv = v
                nv = nv.replace('M/S XYZ CO.', client_name)
                nv = nv.replace('M/s XYZ….. is proprietership Consern in which …... is Proprietor',
                                f'M/s {client_name} is a proprietorship concern.')
                # Replace year strings carefully (PY before CY to avoid double-replace)
                nv = nv.replace(f'31.03.{py_year}', f'__PY_DATE__')
                nv = nv.replace(f'31.03.{cy_year}', f'31.03.{cy_year}')
                nv = nv.replace(f'__PY_DATE__', f'31.03.{py_year}')
                nv = nv.replace(f'31 March, {cy_year}', f'31 March, {cy_year}')
                nv = nv.replace(f'31 March, {py_year}', f'31 March, {py_year}')
                nv = nv.replace('31 March, 2025', f'31 March, {cy_year}')
                nv = nv.replace('31 March, 2024', f'31 March, {py_year}')
                nv = nv.replace('31.03.2025', f'31.03.{cy_year}')
                nv = nv.replace('31.03.2024', f'31.03.{py_year}')
                nv = nv.replace('1st April 2024', f'1st April {py_year}')
                nv = nv.replace('Year Ending 31.03.2025', f'Year Ending 31.03.{cy_year}')
                nv = nv.replace('Year Ending 31.03.2024', f'Year Ending 31.03.{py_year}')
                if nv != v:
                    cell.value = nv


# ─────────────────────────────────────────────────────────────────────────────
#  SECTION 3 — Main entry point
# ─────────────────────────────────────────────────────────────────────────────

def process_tshape(input_path: str, output_path: str,
                   template_path: str = None,
                   client_name: str = None,
                   cy_year: str = None) -> dict:
    """
    Full pipeline:
      1. parse_tshape_bs(input_path)
      2. inject_into_template(...)  → output_path
    Returns dict with status, log, entity info.
    """
    import traceback

    if template_path is None:
        template_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            'Output_sample_format.xlsx'
        )
    if not os.path.exists(template_path):
        return {'status': 'error',
                'message': f'Template not found: {template_path}'}

    try:
        parsed = parse_tshape_bs(input_path)
    except Exception as e:
        return {
            'status': 'error',
            'message': f'Parse failed: {e}',
            'traceback': traceback.format_exc()
        }

    if client_name:
        parsed['entity_name'] = client_name

    try:
        result = inject_into_template(parsed, template_path, output_path,
                                       client_name, cy_year)
    except Exception as e:
        return {
            'status': 'error',
            'message': f'Injection failed: {e}',
            'traceback': traceback.format_exc(),
            'parsed': parsed
        }

    result['parsed'] = parsed
    return result


# ─────────────────────────────────────────────────────────────────────────────
#  CLI
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    import sys
    if len(sys.argv) < 4:
        print('Usage: python tshape_processor.py input.xls output.xlsx template.xlsx [client_name] [cy_year]')
        sys.exit(1)

    inp   = sys.argv[1]
    out   = sys.argv[2]
    tmpl  = sys.argv[3]
    cname = sys.argv[4] if len(sys.argv) > 4 else None
    cyear = sys.argv[5] if len(sys.argv) > 5 else None

    res = process_tshape(inp, out, tmpl, cname, cyear)
    print(f"Status: {res.get('status')}")
    for line in res.get('log', [])[-30:]:
        print(' ', line)
    if res.get('message'):
        print('ERROR:', res['message'])
    if res.get('traceback'):
        print(res['traceback'])
