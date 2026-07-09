"""
lumid_compat.py
===============
Compatibility shim for **Lumid-format** balance sheets
(e.g. "Final_Lumid_Biotech_Pvt__Ltd_BALANCE_SHEET_31_03_2025_in_Rupees.xlsx").

This file is a STANDALONE wrapper around the existing processor.py.
It does NOT modify processor.py in any way.  It:
  1. Detects whether the uploaded file is Lumid-format.
  2. If yes — patches the column-pair map that the processor would otherwise
     compute automatically, fixing five specific issues found in this template.
  3. Calls the standard processor.process() for everything else.
  4. If the file is NOT Lumid-format — falls through to the standard
     processor.process() unchanged.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
COMPATIBILITY ISSUES FOUND IN LUMID FORMAT  (analysis date: 2025-06-29)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Issue 1 – NOA 1-2  (two annexures on one sheet, column layout differs)
  Annexure 1 (rows ~8-55):
    E=CY Count, F=CY Amount (Rs Lakhs), G=PY Count, H=PY Amount
    Auto-detection finds E→G (count→count) but MISSES F→H (amount→amount).
    The amount column (F/H) has no year-bearing header of its own —
    it is labelled "Amount" in row 11, not "As At 31st March".
  Annexure 2 (rows ~57-75):
    F=CY, H=PY. Date header is split over two rows (row 59="As At",
    row 60="31st March 2025"), and row 60 > HEADER_SCAN_ROWS=20, so
    it is never seen by _detect_columns → Annexure 2 is entirely missed.
  Fix: hard-map NOA 1-2  → [(E,G), (F,H)]

Issue 2 – NOA 3-6  (four annexures on one sheet, two different column layouts)
  Annexure 3 (rows ~8-22):
    E=CY_NonCurrent, H=PY_NonCurrent  ← auto-detected correctly as E→H
    F=CY_CurrentMaturities, I=PY_CurrentMaturities  ← not detected (F/I unlabelled)
    G=CY_Total,             J=PY_Total              ← not detected
    In practice F and G are empty in this client's data, so E→H is sufficient.
  Annexures 4/5/6 (rows ~27-78):
    H=CY, J=PY — date headers are on rows 30/42/52, all beyond HEADER_SCAN_ROWS=20.
    NONE of these are detected automatically.
  Conflict: H is PY in Ann3 rows 8-22 AND CY in Ann4-6 rows 27-78.
    A single static E→H and H→J pair would corrupt Ann3 rows (H would be
    both cleared as CY-new and retained as PY-old simultaneously).
  Fix: hard-map NOA 3-6 → [(E,H)] for Ann3 section (rows 8-22 only have
    E and H populated; tool handles row-level: only rows with CY values are
    shifted) + add (H,J) for Ann4-6. Because E is empty in rows 27-78 and H
    is empty in rows 8-22 (as PY col H is never populated with CY data in
    Ann3 — it only receives the shifted value), this does NOT create a
    conflict at the cell level.  The tool writes E→H for rows where E has
    data; it writes H→J for rows where H has data AND J already exists.
    Since Ann3 rows have H=PY_data (copied FROM old E), and Ann4-6 rows
    have H=CY_data (to be copied to J), the pair (H,J) is safe once
    (E,H) has already shifted Ann3 correctly.
  Fix: hard-map NOA 3-6 → [(E,H), (H,J)]
    Caveat: This requires two passes in the right ORDER so that (E,H) runs
    before (H,J). processor._scan_workbook reads both pairs in the same pass
    so we supply them as ordered pairs and rely on the tool's row-by-row
    processing.  The tool processes pairs independently per-column so
    both can coexist.

Issue 3 – CASH FLOW  (stale date headers + wrong hardcoded SHEET_COL_MAP entry)
  The sheet contains date headers "For the Year ended 31st March, 2018" and
  "31st March, 2019" (never updated since FY2018-19), so auto-detection finds
  nothing.  The fallback SHEET_COL_MAP entry in processor.py is:
    "cash flow": [("D", "F")]
  but the actual layout is D=CY, E=PY (column F is empty).  The hardcoded
  D→F entry shifts data to the wrong column.
  Fix: override the CASH FLOW entry to ("D", "E").

Issue 4 – DEP COMPANIES ACT (2)  (Fixed-Asset schedule wrongly shifted)
  Auto-detection picks up B→D, H→I, L→N based on date-like strings in
  row 11 ("As at 1st April 2024", "As at 31 March 2025").  These are
  opening/closing WDV columns in a depreciation schedule, NOT CY/PY
  financial data columns.  Shifting them overwrites WDV and net-block
  cells with wrong values.
  Fix: add "DEP COMPANIES ACT (2)" and "DEP COMPANIES ACT" to TEXT_ONLY
  (date-text update only, no column shift).

Issue 5 – HEADER_SCAN_ROWS  (too small for this multi-annexure format)
  The standard value of 20 misses date headers beyond row 20.
  For this template the issues are addressed by Issues 1-4 above (hard-map
  overrides and TEXT_ONLY additions), so no global change is needed.
  No change to processor.py is made.
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

from __future__ import annotations

import re
import zipfile
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.cell import MergedCell

# ── Import the real processor ────────────────────────────────────────────────
import processor as _proc

# ── Lumid-format fingerprint ─────────────────────────────────────────────────
# We identify a Lumid-format file by the presence of the specific annexure
# sheet names used in this template family.  A simple heuristic that will
# not trigger on any other client's workbook.
_LUMID_SHEET_SIGNATURES = {
    "noa 1-2",
    "noa 3-6",
    "noa 7-12",
    "noa 13-17",
    "noa 18-22",
    "noa 23-27",
}

# Minimum number of the above sheets that must be present to call it Lumid-format
_LUMID_MIN_MATCH = 4


def _is_lumid_format(filepath: str) -> bool:
    """Return True if the workbook looks like a Lumid-format balance sheet."""
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        names_lower = {s.strip().lower() for s in wb.sheetnames}
        wb.close()
        matches = _LUMID_SHEET_SIGNATURES & names_lower
        return len(matches) >= _LUMID_MIN_MATCH
    except Exception:
        return False


# ── Regner-format fingerprint ────────────────────────────────────────────────
# Regner Impex (and similar CA templates with the same NOA numbering scheme)
# uses a unique sheet naming pattern: NOA 7-9, NOA 7-11, NOA 19-22, DEP IT ACT.
# These do NOT appear in Lumid-format workbooks.
# Issues fixed for this template family:
#   Issue R1 – NOA 1-2: auto-detects D→F (count col correct) but misses E→G
#              (amount col, no year header). Fix: add (E,G) pair.
#   Issue R2 – DEP COMPANIES ACT / DEP IT ACT: "As at 1st April YYYY" headers
#              trigger wrong CY/PY shift on WDV schedule columns. Fix: TEXT_ONLY.
#   Issue R3 – CASH FLOW: stale "2018/2019" headers → auto-detects D→F.
#              Correct: D→E. Fixed globally in SHEET_COL_MAP, confirmed here.
_REGNER_SHEET_SIGNATURES = {"noa 7-9", "noa 19-22", "dep it act"}
_REGNER_MIN_MATCH = 3

def _is_regner_format(filepath: str) -> bool:
    """Return True if the workbook looks like a Regner/similar CA format."""
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        names_lower = {s.strip().lower() for s in wb.sheetnames}
        wb.close()
        return len(_REGNER_SHEET_SIGNATURES & names_lower) >= _REGNER_MIN_MATCH
    except Exception:
        return False

_REGNER_TEXT_ONLY_EXTRA = {
    # Depreciation schedules — WDV opening/closing columns must not be shifted
    "dep companies act", "dep it act",
}

_REGNER_COL_OVERRIDES: dict[str, list[tuple[str, str]]] = {
    # NOA 1-2: auto-detects (D,F) for count col but misses (E,G) for amount col
    "NOA 1-2": [("D", "F"), ("E", "G")],
}

# ── Pooja/GD-Singla horizontal-format fingerprint ───────────────────────────
# This template family (e.g. POOJA_INDUSTRIES, similar GD Singla CA templates)
# packs ALL schedules horizontally into a SINGLE sheet called "POOJA-I" or
# similar.  Each schedule occupies a 4-column block:
#   col N   = labels, col N+1 = note ref, col N+2 = CY figures, col N+3 = PY figures
#
# The auto-detector finds wrong column pairs because the year-header strings
# (e.g. "31st March, 2025") appear in the PY column (N+3) of each block, not
# the CY column.  The detector pairs them with the wrong adjacent column.
#
# Detection: the sheet name starts with the company name (no standard pattern),
# but the layout is identified by finding "Figures as at the end of the" and
# "current reporting period" / "of the prev reporting period" split over two
# rows in the same column area — this is the Revised Schedule VI BS header.

_POOJA_HEADER_MARKER = "figures as at the end of the"  # lower-cased


def _is_pooja_format(filepath: str) -> bool:
    """
    Return True if the workbook uses the Pooja/GD-Singla horizontal layout:
    all BS, P&L and schedule blocks packed side-by-side in one sheet, with
    the 'Figures as at the end of the / current reporting period' split header.
    """
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        try:
            for sname in wb.sheetnames:
                ws = wb[sname]
                for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
                    for cell in row:
                        if (isinstance(cell, str) and
                                _POOJA_HEADER_MARKER in cell.lower()):
                            return True
        finally:
            wb.close()
    except Exception:
        pass
    return False


def _pooja_sheet_col_map(filepath: str) -> dict[str, list[tuple[str, str]]]:
    """
    Build the corrected shift_map for a Pooja/GD-Singla format workbook.

    Returns: { sheet_name: [(cy_letter, py_letter), ...] }

    Two header patterns are detected within the first 20 rows of each sheet:

    Pattern 1 — Revised Schedule VI "Figures" split header (BS / P&L sections):
      Col X  = "Figures as at the end of the" (CY label)
      Col X+1 = "Figures as at the end" (PY label, split across same row)
      → pair: (X, X+1)
      Example: BS at cols C/D → pair C→D; P&L at cols G/H → pair G→H.

    Pattern 2 — Notes with count + amount sub-columns (Share Capital etc.):
      Col X   = "As at 31 March YYYY_CY"  (CY header, e.g. 2025)
      Col X+2 = "As at 31 March YYYY_PY"  (PY header, e.g. 2024)
      Data layout: X=CY_count, X+1=CY_amount, X+2=PY_count, X+3=PY_amount
      → pairs: (X, X+2) for count column AND (X+1, X+3) for amount column
      Example: cols 11/13 → pairs K→M and L→N.

    Pattern 1 and Pattern 2 are distinguished by whether two year-labelled
    headers appear 1 apart (Pattern 1) or 2 apart (Pattern 2).
    """
    from openpyxl.utils import get_column_letter
    import re as _re

    cy_year_re = _re.compile(r'as at\s+\d+\s+\w+\s*,?\s*(\d{4})', _re.IGNORECASE)
    result: dict = {}

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        try:
            for sname in wb.sheetnames:
                ws = wb[sname]
                pairs: list[tuple[str, str]] = []
                seen_cy_cols: set[int] = set()

                # Pass 1 — "Figures as at the end of the" / "Figures as at the end"
                # split header (always gap=1, CY then PY on same row)
                for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
                    for c_idx, cell in enumerate(row, 1):
                        if not isinstance(cell, str):
                            continue
                        if _POOJA_HEADER_MARKER in cell.lower():
                            if c_idx not in seen_cy_cols:
                                seen_cy_cols.add(c_idx)
                                pairs.append((
                                    get_column_letter(c_idx),
                                    get_column_letter(c_idx + 1),
                                ))

                # Pass 2 — "As at 31 March YYYY" date headers.
                # Collect all (col, year) tuples from header rows, then resolve pairs.
                date_headers: dict[int, int] = {}  # col_idx → year
                for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
                    for c_idx, cell in enumerate(row, 1):
                        if not isinstance(cell, str):
                            continue
                        m = cy_year_re.search(cell)
                        if m:
                            yr = int(m.group(1))
                            if c_idx not in date_headers:
                                date_headers[c_idx] = yr

                # Determine CY year (most recent) from the found headers
                if date_headers:
                    max_yr = max(date_headers.values())
                    for c_idx, yr in sorted(date_headers.items()):
                        if yr != max_yr:
                            continue   # this is a PY header — skip
                        if c_idx in seen_cy_cols:
                            continue   # already handled by Pass 1
                        # Check if PY header is at gap=1 or gap=2
                        py_gap1 = date_headers.get(c_idx + 1)
                        py_gap2 = date_headers.get(c_idx + 2)
                        if py_gap1 is not None and py_gap1 < max_yr:
                            # Gap=1: simple CY→PY pair
                            seen_cy_cols.add(c_idx)
                            pairs.append((
                                get_column_letter(c_idx),
                                get_column_letter(c_idx + 1),
                            ))
                        elif py_gap2 is not None and py_gap2 < max_yr:
                            # Gap=2: count+amount sub-columns
                            # Add (X→X+2) for count AND (X+1→X+3) for amount
                            seen_cy_cols.update([c_idx, c_idx + 1])
                            pairs.append((
                                get_column_letter(c_idx),
                                get_column_letter(c_idx + 2),
                            ))
                            pairs.append((
                                get_column_letter(c_idx + 1),
                                get_column_letter(c_idx + 3),
                            ))

                if pairs:
                    result[sname] = pairs
        finally:
            wb.close()
    except Exception:
        pass

    return result


# ── Lumid-specific column-pair overrides ─────────────────────────────────────
# Keys are the exact sheet names as they appear in the workbook (case-sensitive
# because _scan_workbook uses them as-is).  The values are the corrected
# (cy_letter, py_letter) pairs that replace whatever auto-detection would find.
#
# These are discovered from analysis of the Lumid template layout.

def _lumid_sheet_col_map(wb_sheetnames: list[str]) -> dict:
    """
    Build the corrected shift_map for all Lumid-format sheets.

    Returns a dict  { exact_sheet_name: [(cy_letter, py_letter), ...] }
    for sheets that need an override.  Sheets not listed here are handled
    by the standard processor logic.
    """
    overrides: dict = {}
    for sn in wb_sheetnames:
        sl = sn.strip().lower()

        # ── NOA 1-2 ──────────────────────────────────────────────────────────
        # Annexure 1 (Share Capital): E=CY_count, F=CY_amount, G=PY_count, H=PY_amount
        # Annexure 2 (Reserves):       F=CY_amount, H=PY_amount
        # The overlap of F and H across both annexures means (F,H) covers both.
        # E→G (count) is kept for completeness but share-count rarely changes.
        if sl == "noa 1-2" or sl == "noa 1-2 ":
            overrides[sn] = [("E", "G"), ("F", "H")]

        # ── NOA 3-6 ──────────────────────────────────────────────────────────
        # Annexure 3  (rows  ~8-22): E=CY_NonCurrent,  H=PY_NonCurrent → E→H
        # Annexures 4-6 (rows 27-78): H=CY,            J=PY            → H→J
        # NOTE: H plays two roles across the two section groups.
        # This is safe because in practice F and G are empty in Ann3,
        # and E is empty in Ann4-6, so there is no cell-level overlap.
        elif sl == "noa 3-6":
            overrides[sn] = [("E", "H"), ("H", "J")]

        # ── NOA 7-12 ─────────────────────────────────────────────────────────
        # Auto-detection gets F→H which is correct (verified in analysis).
        # No override needed — let the standard detector handle it.

        # ── NOA 13-17 ────────────────────────────────────────────────────────
        # Auto-detection finds G→I (correct) AND N→P (WRONG — must be dropped).
        #
        # The N→P problem:
        #   Col N row 9 = "As At 31st March 2025" (CY header) → auto-detect
        #   picks N as a CY column with P as the paired PY column.
        #   BUT col N rows 11-35 are EMPTY (no CY data there — the debtor aging
        #   sub-table is a current-year-only schedule with no comparative column
        #   filled in).  Col P rows 11-35 contain the PY aging detail values
        #   (individual debtor amounts) and P36 = PY total (127.31 Lakhs).
        #
        #   The tool's BUG-6 logic correctly avoids clearing PY cells whose
        #   corresponding CY cell has NO value — but the tool still calls
        #   _process_sheet_xml with both pairs active.  The problem is that
        #   N36=0 (an explicit zero) DOES get treated as a CY value, so P36
        #   gets overwritten with 0, destroying the PY aging total (127.31→0).
        #   Additionally the date header in N9 triggers a string-type clear
        #   on P9, wiping the PY period label.
        #
        # Fix: hard-override NOA 13-17 → [(G,I)] only. Drop N→P entirely.
        # The aging sub-table has no comparative data to shift; its PY figures
        # in col P must be preserved as-is for the new year's PY reference.
        elif sl == "noa 13-17":
            overrides[sn] = [("G", "I")]

        # ── NOA 18-22 ────────────────────────────────────────────────────────
        # Auto-detection gets E→G — correct.  No override needed.

        # ── NOA 23-27 ────────────────────────────────────────────────────────
        # Auto-detection finds G→I — correct for Annexures 23-26 (rows 10-77),
        # BUT WRONG for Annexure 27 EPS table (rows 80-95).
        #
        # The Ann27 EPS problem:
        #   Ann27 "Earning Per Share" (rows 80-95) has a 4-column layout:
        #     F83="31st March 2025" (CY), G83="31st March 2024" (PY),
        #     H83="31st March 2024" (PY), I83="31st March 2024" (PY).
        #   Col G in rows 83-95 holds PY data (EPS figures, share count etc.),
        #   NOT CY data.  The G→I shift incorrectly:
        #     1. Moves G83-G95 values (PY data) into I column cells.
        #     2. Clears G83-G95, destroying the PY EPS column entirely.
        #   Result: G86, G87, G88, G91, G92, G93, G95 all wiped to blank.
        #
        # Fix: hard-override NOA 23-27 → [(G,I)] with a row-range restriction
        # so only rows 1-79 (Ann23-26 data) are read as CY for col G.
        # Rows 80+ (Ann27 EPS) are excluded from the CY scan → G values there
        # are never treated as CY data → they won't be shifted or cleared.
        # This is implemented via row_restrictions in _rescan_changed_sheets.
        elif sl == "noa 23-27":
            overrides[sn] = [("G", "I")]

        # ── CASH FLOW ────────────────────────────────────────────────────────
        # Standard SHEET_COL_MAP fallback uses D→F (wrong for this template).
        # Actual layout: D=CY, E=PY.
        elif sl == "cash flow":
            overrides[sn] = [("D", "E")]

    return overrides


# Sheets that are Fixed-Asset / depreciation SCHEDULES — they must NOT have
# CY/PY column-shifting applied.  We add the Lumid-specific sheets to the
# set that the processor treats as text-only.
_LUMID_TEXT_ONLY_EXTRA = {
    "dep companies act (2)",
    "dep companies act",
    "dep income tax (3)",
    "dep income tax",
    "deferred tax (2)",
    "deferred tax",
    "sheet1",
    "ratios",
}


# ── Public entry point ───────────────────────────────────────────────────────

def process(input_path: str, output_path: str,
            closing_year: int, new_year: int) -> dict:
    """
    Year-shift a balance sheet, with compatibility shims for known template families.

    Template families handled:
      • Lumid-format (NOA 1-2 … NOA 23-27 multi-sheet layout):
          Column-pair overrides, DEP/deferred-tax TEXT_ONLY, CASH FLOW remap,
          PY-column formula freeze, XML repair.
      • Pooja/GD-Singla horizontal format (all schedules in one wide sheet):
          Auto-detects correct CY→PY column pairs from the Revised Schedule VI
          "Figures as at the end of the / current reporting period" header.
          Overrides the wrong pairs that the standard auto-detector finds.

    For all other files:
      • Falls through directly to processor.process() unchanged.
    """
    # ── Pooja/GD-Singla horizontal format ───────────────────────────────────
    if _is_pooja_format(input_path):
        col_overrides = _pooja_sheet_col_map(input_path)
        if col_overrides:
            # Run standard scan first
            sizes, shift_map, cy_values, cy_formulas, cap_data = \
                _proc._scan_workbook(input_path, closing_year)

            # Replace auto-detected pairs with correct pairs for every
            # sheet that has a Pooja-style header
            for sn, corrected_pairs in col_overrides.items():
                shift_map[sn] = corrected_pairs

            # Re-scan CY values using the corrected column pairs.
            # _rescan_changed_sheets expects all workbook sheet names as arg 3.
            _wb_tmp2 = load_workbook(input_path, read_only=True, data_only=True)
            _all_sn_pooja = _wb_tmp2.sheetnames
            _wb_tmp2.close()
            _rescan_changed_sheets(
                input_path, closing_year,
                _all_sn_pooja,
                col_overrides,
                shift_map, cy_values, cy_formulas,
            )

            pairs = _proc._date_replacements(str(closing_year), str(new_year))
            log: list[str] = []
            log.append("ℹ Pooja/GD-Singla horizontal format detected — correcting column pairs")
            for sn, cp in col_overrides.items():
                desc = ", ".join(f"{c}→{p}" for c, p in cp)
                log.append(f"  {sn}: {desc}")

            result = _run_with_patched_maps(
                input_path, output_path,
                closing_year, new_year,
                sizes, shift_map, cy_values, cy_formulas, cap_data,
                pairs, log,
                set(),  # no Lumid-specific TEXT_ONLY extras
            )

            # ── Pooja: freeze PY formula columns ─────────────────────────────
            # In the Pooja-I sheet, BOTH the CY columns (C, G, K, L) AND the PY
            # columns (D, H, M, N) are built entirely from intra-sheet formula
            # references (e.g. D18=+M85, D17=N24, H17=+H15+H16, M22=N22/10).
            #
            # After the year-shift:
            #   • CY cols (C/G/K/L) — _pooja_blank_cy_formulas() blanks them
            #     below (formulas + values stripped → truly blank for CA to fill).
            #   • PY cols (D/H/M/N) — the shift correctly copies old-CY cached
            #     values into them, but the FORMULA in each PY cell still points
            #     to the schedule columns (which now hold shifted/cleared data).
            #     When the CA clicks "Enable Editing", Excel recalculates the PY
            #     formula against the now-shifted schedule: D18=+M85 → M85 is the
            #     NEW CY col (blank) → D18 recalculates to 0 instead of 45.8L.
            #     This makes the BS total drop from ₹25.26 Cr to ₹20.68 Cr.
            #
            # Fix: freeze every formula cell in the PY columns to its cached <v>
            # value (remove the <f> tag) so Enable Editing cannot recalculate it.
            # Same mechanism as _freeze_py_columns used for Lumid BAL SHEET/P L.
            #
            # Determine PY col letters for each shifted sheet
            pooja_py_cols: dict[str, list[str]] = {}
            for sn, cpairs in col_overrides.items():
                pooja_py_cols[sn] = [py for _cy, py in cpairs]

            _freeze_py_columns(output_path, pooja_py_cols)
            log.append(
                "🔒 Pooja: PY columns frozen (formulas → plain values) "
                "so Enable Editing preserves correct prior-year figures"
            )

            # ── Pooja: blank CY formula columns ──────────────────────────────
            # After freezing PY cols, blank the CY formula cells.
            # See _pooja_blank_cy_formulas docstring for full explanation.
            _pooja_blank_cy_formulas(output_path, col_overrides)
            log.append(
                "🔒 Pooja: CY formula cells blanked — Enable Editing will "
                "not recalculate stale refs in the new-year column"
            )

            return result

    # ── Regner / similar CA format ───────────────────────────────────────────
    if _is_regner_format(input_path):
        sizes, shift_map, cy_values, cy_formulas, cap_data = \
            _proc._scan_workbook(input_path, closing_year)

        wb_tmp_r = load_workbook(input_path, read_only=True, data_only=True)
        sheetnames_r = wb_tmp_r.sheetnames
        wb_tmp_r.close()

        # Apply NOA 1-2 column override (adds missing amount pair E→G)
        regner_overrides = {}
        for sn in sheetnames_r:
            if sn in _REGNER_COL_OVERRIDES:
                shift_map[sn] = _REGNER_COL_OVERRIDES[sn]
                regner_overrides[sn] = _REGNER_COL_OVERRIDES[sn]

        # Remove DEP sheets from shift_map (TEXT_ONLY treatment)
        regner_text_only_names = set()
        for sn in sheetnames_r:
            if sn.strip().lower() in _REGNER_TEXT_ONLY_EXTRA:
                shift_map.pop(sn, None)
                regner_text_only_names.add(sn)

        # ── BAL SHEET and P L: allow normal D→E column shift ────────────────
        # These sheets are formula-driven summaries in the original xlsb, but
        # after pyxlsb→xlsx conversion all formulas become plain cached values.
        # The D→E shift is therefore correct and necessary:
        #   • D col (CY Rupees) → cleared (blank for fresh FY26 data entry)
        #   • E col (PY) ← gets old D values (FY25 in Rupees as new PY reference)
        # Old E Lakhs values are overwritten — acceptable since they were stale
        # cached formula results that would be wrong after the NOA sheets shift.
        # The new E (Rupees) gives the CA the correct PY rupee reference.
        # NOTE: the processor.py stale-reference fix (clear_v for formulas whose
        # column refs are all outside the shift range) ensures that any formula
        # cells in D that reference columns like DG or Y (outside D/E) are fully
        # cleared rather than kept, so they cannot recalculate back to old values.

        # Re-scan CY values for sheets whose column pairs changed
        if regner_overrides:
            _rescan_changed_sheets(
                input_path, closing_year, sheetnames_r, regner_overrides,
                shift_map, cy_values, cy_formulas,
            )

        pairs = _proc._date_replacements(str(closing_year), str(new_year))
        log: list[str] = []
        log.append("ℹ Regner-format detected — applying compatibility overrides")
        if regner_overrides:
            log.append(f"  Column-pair overrides: {list(regner_overrides.keys())}")
        if regner_text_only_names:
            log.append(f"  TEXT_ONLY (dep/summary sheets): {sorted(regner_text_only_names)}")

        original_text_only_r = _proc.TEXT_ONLY_SHEETS
        _proc.TEXT_ONLY_SHEETS = original_text_only_r | _REGNER_TEXT_ONLY_EXTRA

        result = _run_with_patched_maps(
            input_path, output_path,
            closing_year, new_year,
            sizes, shift_map, cy_values, cy_formulas, cap_data,
            pairs, log,
            regner_text_only_names,
        )

        _proc.TEXT_ONLY_SHEETS = original_text_only_r

        # Run XML repair — strips empty <definedNames/>, <workbookProtection/>,
        # stale calcChain, and cell-type malformations that arise from the
        # xlsb→xlsx conversion (openpyxl writes these empty tags when building
        # a workbook from scratch, and some Excel versions flag them as corrupt).
        _repair_worksheet_xml(output_path)
        log.append(
            "🔧 Regner: XML repair pass — definedNames/calcChain cleaned"
        )
        return result

    if not _is_lumid_format(input_path):
        # Not Lumid-format, not Pooja-format, not Regner-format → standard processor
        return _proc.process(input_path, output_path, closing_year, new_year)

    # ── Lumid-format path ────────────────────────────────────────────────────
    # Step 1: Run the standard scan to get sizes, cy_values, cy_formulas etc.
    sizes, shift_map, cy_values, cy_formulas, cap_data = \
        _proc._scan_workbook(input_path, closing_year)

    # Step 2: Get the corrected column-pair overrides for this template.
    wb_tmp = load_workbook(input_path, read_only=True, data_only=True)
    sheetnames = wb_tmp.sheetnames
    wb_tmp.close()

    col_overrides = _lumid_sheet_col_map(sheetnames)

    # Step 3: Apply overrides to shift_map.
    # For sheets listed in col_overrides, REPLACE the auto-detected pairs.
    # For sheets not listed, keep whatever the standard detector found.
    for sn, corrected_pairs in col_overrides.items():
        shift_map[sn] = corrected_pairs

    # Step 4: For Lumid TEXT_ONLY extras, REMOVE them from shift_map
    # so they only receive date-string updates (no column shifting).
    # We must also ensure they don't get mistakenly added back by SHEET_COL_MAP.
    lumid_text_only_names = set()
    for sn in sheetnames:
        sl = sn.strip().lower()
        if sl in _LUMID_TEXT_ONLY_EXTRA:
            shift_map.pop(sn, None)     # remove from column-shift list
            lumid_text_only_names.add(sn)

    # Step 5: Re-scan cy_values and cy_formulas using the CORRECTED pairs.
    # For sheets whose pairs changed, the existing cy_values may be from the
    # wrong columns.  Re-read those sheets.
    _rescan_changed_sheets(
        input_path, closing_year, sheetnames, col_overrides,
        shift_map, cy_values, cy_formulas
    )

    # Step 6: Build a patched TEXT_ONLY set that includes Lumid extras.
    # We monkey-patch the module-level constant temporarily so that the
    # process() function's inner logic respects our additions.
    original_text_only = _proc.TEXT_ONLY_SHEETS
    _proc.TEXT_ONLY_SHEETS = original_text_only | _LUMID_TEXT_ONLY_EXTRA

    # Step 7: Override SHEET_COL_MAP fallback for cash flow.
    # The standard SHEET_COL_MAP has "cash flow": [("D","F")] which is wrong
    # for this template.  We temporarily override it.
    original_col_map = dict(_proc.SHEET_COL_MAP)
    _proc.SHEET_COL_MAP["cash flow"] = [("D", "E")]

    # Step 8: Build the date-replacement pairs (same as standard)
    pairs = _proc._date_replacements(str(closing_year), str(new_year))
    log: list[str] = []
    log.append("ℹ Lumid-format detected — applying compatibility overrides")
    log.append(f"  Column-pair overrides: {list(col_overrides.keys())}")
    log.append(f"  TEXT_ONLY extras: {sorted(lumid_text_only_names)}")
    log.append(f"  CASH FLOW remapped: D→E (was D→F)")

    # Step 9: Execute the ZIP-level edit loop (copied from processor.process()
    # but using our patched shift_map).
    result = _run_with_patched_maps(
        input_path, output_path,
        closing_year, new_year,
        sizes, shift_map, cy_values, cy_formulas, cap_data,
        pairs, log,
        lumid_text_only_names,
    )

    # Step 10: Restore the temporarily patched module globals.
    _proc.TEXT_ONLY_SHEETS = original_text_only
    _proc.SHEET_COL_MAP.clear()
    _proc.SHEET_COL_MAP.update(original_col_map)

    return result


# ── Internal helpers ─────────────────────────────────────────────────────────

def _noa36_section_boundary(ws) -> int:
    """
    Find the row number where NOA 3-6's second section (Annexure 4 onwards)
    begins.  This is the first row > 22 where column E (col 5) contains an
    'ANNEXURE' label or a deferred-tax section header.

    Returns the boundary row (rows < boundary belong to Annexure 3,
    rows >= boundary belong to Annexures 4-6).  Defaults to 27 if not found.
    """
    for r in range(23, 90):
        for c in range(3, 7):
            v = ws.cell(r, c).value
            if isinstance(v, str) and (
                "ANNEXURE" in v.upper()
                or "DEFERRED" in v.upper()
            ):
                return r
    return 27   # safe fallback


def _rescan_changed_sheets(
    filepath: str,
    closing_year: int,
    sheetnames: list[str],
    col_overrides: dict,
    shift_map: dict,
    cy_values: dict,
    cy_formulas: dict,
) -> None:
    """
    For sheets whose column pairs were overridden, re-read the CY column
    values from the corrected columns (the original scan may have read the
    wrong columns).

    Special handling for NOA 3-6 (Issue 2 / H-column conflict):
    ─────────────────────────────────────────────────────────────
    NOA 3-6 is mapped as [(E,H), (H,J)].  The sheet has two section groups:
      Ann 3  (rows  ~8-26):  E = CY_NonCurrent, H = PY_NonCurrent
      Ann 4-6 (rows 27-78):  H = CY,            J = PY

    Within a single processor pass, both pairs are applied to every row.
    The conflict arises in Ann 3 rows 14/16/20/22 where H has numeric data
    (PY values = 175):
      • (E→H) correctly shifts E → H (H becomes the new PY value)
      • (H→J) then sees the NEWLY written H value as "CY", clears H,
        and writes it to J — destroying the correctly-shifted H value.

    Fix: restrict the cy_values for column H to ONLY rows >= section_boundary
    (i.e. Ann4-6 rows where H genuinely is the CY column).  For Ann3 rows,
    H has no cy_value entry → (H→J) makes no change there → no conflict.
    """
    changed_sheets = set(col_overrides.keys())
    if not changed_sheets:
        return

    # Get the sheet-file map from the ZIP
    with zipfile.ZipFile(filepath) as z:
        smap = _proc._sheet_file_map(z)

        _f_cell_re = re.compile(rb'<c\b[^>]*\br="([A-Z]+\d+)"[^>]*>[^<]*<f[ />]')
        formula_refs_new: dict = {}
        formula_texts_new: dict = {}
        for sn in changed_sheets:
            sf = smap.get(sn, "")
            if not sf:
                continue
            try:
                xml_data = z.read(sf)
                refs = set(m.group(1).decode() for m in _f_cell_re.finditer(xml_data))
                formula_refs_new[sn] = refs

                ftexts: dict = {}
                for cm in re.finditer(
                    rb'<c\b[^>]*\br="([A-Z]+\d+)"[^>]*>(.*?)</c>',
                    xml_data, re.DOTALL
                ):
                    ref_b, body_b = cm.group(1), cm.group(2)
                    has_real_v = bool(re.search(rb'<v[^>]*>[^<]+</v>', body_b))
                    if has_real_v:
                        continue
                    fm = re.search(rb'<f[^>]*>([^<]*)</f>', body_b)
                    if fm:
                        try:
                            ftexts[ref_b.decode()] = fm.group(1).decode("utf-8", "replace")
                        except Exception:
                            pass
                formula_texts_new[sn] = ftexts
            except Exception:
                formula_refs_new[sn] = set()
                formula_texts_new[sn] = {}

    # Re-read the corrected CY columns with openpyxl
    wb = load_workbook(filepath, read_only=True, data_only=True)
    try:
        for sn in changed_sheets:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            det = shift_map.get(sn, [])
            if not det:
                continue

            sheet_frefs = formula_refs_new.get(sn, set())
            sheet_ftexts = formula_texts_new.get(sn, {})

            # Per-column row-range restrictions (for multi-section sheets).
            # {col_letter: (min_row, max_row)} — None means no restriction.
            row_restrictions: dict[str, tuple[int, int] | None] = {}

            sl = sn.strip().lower()
            if sl == "noa 3-6":
                # Detect boundary between Ann3 and Ann4 dynamically.
                boundary = _noa36_section_boundary(ws)
                # E column: only Ann3 rows (rows < boundary) are the CY section
                row_restrictions["E"] = (1, boundary - 1)
                # H column: only Ann4-6 rows (rows >= boundary) are CY
                row_restrictions["H"] = (boundary, 10_000)
            elif sl == "noa 13-17":
                # N→P pair has been dropped (override = [(G,I)] only).
                # No row restriction needed for G — all rows are correct CY.
                pass
            elif sl == "noa 23-27":
                # G→I is correct for Ann23-26 (rows 10-77).
                # Ann27 EPS (rows 80-95) has G=PY data, NOT CY — exclude it.
                # Row restriction: only scan G for CY values up to row 79.
                row_restrictions["G"] = (1, 79)

            cy_vals_sheet: dict = {}
            cy_frows_sheet: dict = {}

            for cy_l, _ in det:
                ci = _proc._col_idx(cy_l)
                vals: dict = {}
                frows: set = set()

                # Determine the allowed row range for this column
                rng = row_restrictions.get(cy_l)  # (min_row, max_row) or None
                min_r = rng[0] if rng else 1
                max_r = rng[1] if rng else 10_000

                for row in ws.iter_rows(min_col=ci, max_col=ci):
                    for cell in row:
                        if isinstance(cell, MergedCell):
                            continue
                        if not hasattr(cell, 'row') or cell.row is None:
                            continue
                        rn = cell.row

                        # Apply row-range restriction
                        if rn < min_r or rn > max_r:
                            continue

                        ref = f"{cy_l}{rn}"

                        if ref in sheet_frefs:
                            frows.add(rn)
                            if isinstance(cell.value, (int, float)) and cell.value is not None:
                                vals[rn] = float(cell.value)
                            elif cell.value is None:
                                ftext = sheet_ftexts.get(ref)
                                evaluated = _proc._try_eval_arithmetic_formula(ftext)
                                vals[rn] = evaluated if evaluated is not None else 0.0
                        elif isinstance(cell.value, (int, float)) and cell.value is not None:
                            vals[rn] = float(cell.value)
                        elif isinstance(cell.value, str) and cell.value.strip() in ("-", "—", "–"):
                            vals[rn] = 0.0

                cy_vals_sheet[cy_l] = vals
                cy_frows_sheet[cy_l] = frows

            cy_values[sn] = cy_vals_sheet
            cy_formulas[sn] = cy_frows_sheet
    finally:
        wb.close()


def _sanitize_cash_flow_errors(input_path: str) -> str:
    """
    Pre-processing pass: strip t="e" (error-type) attributes and stale #REF!
    formulas from the CASH FLOW sheet (sheet13) in the Lumid template.

    WHY THIS IS NEEDED
    ──────────────────
    The CASH FLOW sheet's E column (PY data) was built using cross-sheet
    formulas that reference rows which no longer exist in the P L / BAL SHEET
    source sheets.  Those cells have been broken since the template was first
    created — they carry  t="e"  with formulas like  'P L'!#REF!  and cached
    value "#REF!".

    When the D→E year-shift writes the correct numeric CY→PY values into those
    E cells, processor._process_sheet_xml correctly replaces the <v> content
    with the new number but leaves the t="e" attribute and the stale <f> tag
    untouched (it only rewrites <v>…</v>).  The resulting cell XML:

        <c r="E10" s="31" t="e"><f>'P L'!#REF!</f><v>1.1679…</v></c>

    violates OOXML §3.18.11 (a t="e" cell must have an error string in <v>,
    not a number) and triggers Excel's "Repaired Records: Cell information from
    /xl/worksheets/sheet13.xml" dialog every time the output is opened.

    FIX
    ───
    Before the main ZIP-edit loop runs, rewrite the CASH FLOW sheet so that
    every t="e" cell has its  t="e"  attribute and  <f>…</f>  tag removed,
    leaving a plain empty cell.  The year-shift then fills the <v> cleanly.

    Implementation: rewrites to a temp file and returns the new path so the
    caller can use it as the input to the main ZIP loop.  If any error occurs
    the original path is returned unchanged (safe fallback).
    """
    import zipfile as _zipfile, os as _os, tempfile as _tmp

    CASH_FLOW_NAME = "CASH FLOW"   # exact sheet name in the Lumid template

    try:
        with _zipfile.ZipFile(input_path, "r") as zi:
            smap = _proc._sheet_file_map(zi)
            cf_file = smap.get(CASH_FLOW_NAME, "")
            if not cf_file:
                return input_path          # sheet not found — no-op

            all_names = zi.namelist()
            all_data  = {n: zi.read(n) for n in all_names}
            all_infos = {item.filename: item for item in zi.infolist()}

        # Patch the CASH FLOW sheet XML
        raw  = all_data[cf_file]
        text = raw.decode("utf-8", "replace")

        def _clear_error_cell(cm):
            full = cm.group(0)
            # Only touch cells in column E (ref like E10, E27, etc.)
            ref_m = re.search(r'\br="([A-Z]+)(\d+)"', full)
            if not ref_m or ref_m.group(1) != "E":
                return full
            # Remove t="e" attribute
            fixed = re.sub(r'\s*t="e"', '', full, count=1)
            # Remove the stale <f>…</f> tag (contains #REF!)
            fixed = re.sub(r'<f[^>]*>.*?</f>', '', fixed, flags=re.DOTALL)
            fixed = re.sub(r'<f[^>]*/>', '', fixed)
            # Remove the stale <v>#REF!</v> cached value
            fixed = re.sub(r'<v>\s*#REF!\s*</v>', '', fixed)
            return fixed

        patched = re.sub(
            r'<c\b[^>]*\bt="e"[^>]*>.*?</c>',
            _clear_error_cell, text, flags=re.DOTALL
        )

        if patched == text:
            return input_path              # nothing changed — no-op

        all_data[cf_file] = patched.encode("utf-8")

        # Write to a temp file
        fd, tmp_path = _tmp.mkstemp(suffix=".xlsx")
        _os.close(fd)
        with _zipfile.ZipFile(tmp_path, "w", _zipfile.ZIP_DEFLATED) as zo:
            for name in all_names:
                info = all_infos[name]
                zo.writestr(info, all_data[name])

        return tmp_path

    except Exception:
        # Any failure → return original path unchanged (safe fallback)
        return input_path


def _run_with_patched_maps(
    input_path: str,
    output_path: str,
    closing_year: int,
    new_year: int,
    sizes: dict,
    shift_map: dict,
    cy_values: dict,
    cy_formulas: dict,
    cap_data: dict,
    pairs: list,
    log: list,
    lumid_text_only_names: set,
) -> dict:
    """
    Re-implement the ZIP-level loop from processor.process() using the
    patched shift_map and cy_values.  This avoids calling _scan_workbook
    a second time while still running the full edit pipeline.
    """
    import zipfile as _zipfile

    # Compute effective-big-sheet set (same logic as processor.process)
    BIG_ROWS = _proc.BIG_ROWS
    BIG_COLS = _proc.BIG_COLS

    def _effective_big(n, r, c):
        if r > BIG_ROWS:
            return True
        if c > BIG_COLS:
            sm_pairs = shift_map.get(n, [])
            if sm_pairs:
                from openpyxl.utils import column_index_from_string
                max_sm_col = max(
                    column_index_from_string(col)
                    for pair in sm_pairs for col in pair
                )
                return max_sm_col > BIG_COLS
            return True
        return False

    big_names = {n for n, (r, c) in sizes.items() if _effective_big(n, r, c)}
    date_style_indices = _proc._build_date_style_set(input_path)

    RAW_DATA_SHEETS  = _proc.RAW_DATA_SHEETS
    CAPITAL_SHEET_NAMES = _proc.CAPITAL_SHEET_NAMES
    # TEXT_ONLY_SHEETS already patched at module level by caller

    ext_count = 0

    # ── Pre-processing: sanitize CASH FLOW sheet ─────────────────────────────
    # sheet13 (CASH FLOW) was authored with its E column (PY data) populated
    # via cross-sheet formulas that reference rows which no longer exist in the
    # source sheets (deleted or renumbered over successive years).  Those cells
    # carry t="e" (error type) with formulas like  'P L'!#REF!  and cached
    # value "#REF!".
    #
    # When the D→E year-shift runs, it writes the correct numeric CY→PY values
    # into those E cells.  The processor's regex correctly replaces the <v>
    # content with the new number — but it does NOT touch the cell's t= attribute
    # or its <f> tag (those are outside the <v>…</v> it rewrites).  The result is
    # a cell like:
    #
    #   <c r="E10" s="31" t="e"><f>'P L'!#REF!</f><v>1.1679...</v></c>
    #
    # which is invalid per OOXML §3.18.11: a cell with t="e" MUST have an error
    # value (#REF!, #VALUE!, etc.) in <v> — a number there is undefined behaviour.
    # Excel detects this as corruption and shows the "Repaired Records: Cell
    # information from /xl/worksheets/sheet13.xml" dialog.
    #
    # Fix: before the shift loop opens the ZIP for writing, rewrite the input ZIP
    # in-place (via a temp file) to strip t="e" and the stale #REF! formulas from
    # every E-column cell in CASH FLOW.  This leaves them as plain empty numeric
    # cells, which the shift then fills correctly.
    input_path = _sanitize_cash_flow_errors(input_path)

    with _zipfile.ZipFile(input_path, "r") as zi:
        smap = _proc._sheet_file_map(zi)

        with _zipfile.ZipFile(output_path, "w", _zipfile.ZIP_DEFLATED) as zo:
            for item in zi.infolist():
                fn   = item.filename
                data = zi.read(fn)

                # Skip external links
                if "externalLinks" in fn:
                    continue

                # Shared strings — date-text replacement only
                if fn == "xl/sharedStrings.xml":
                    data = _proc._update_shared_strings(data, pairs)
                    zo.writestr(item, data)
                    continue

                # Worksheet XMLs
                if fn.startswith("xl/worksheets/") and fn.endswith(".xml"):
                    sheet_name = next(
                        (sn for sn, sf in smap.items() if sf == fn), None
                    )
                    sl = (sheet_name or "").strip().lower()

                    # Strip external formulas from all sheets
                    before_len = len(data)
                    data = _proc._strip_external_formulas(data)
                    if len(data) != before_len:
                        ext_count += 1

                    if sheet_name and sheet_name in shift_map:
                        # Full CY→PY shift + date update
                        data = _proc._process_sheet_xml(
                            data,
                            shift_map[sheet_name],
                            cy_values.get(sheet_name, {}),
                            cy_formulas.get(sheet_name, {}),
                            date_style_indices,
                            shift_map,
                        )
                        data = _proc._update_inline_strings(data, pairs)
                        desc = ", ".join(
                            f"{c}→{p}" for c, p in shift_map[sheet_name]
                        )
                        log.append(
                            f"✓ {sheet_name}: CY→PY copied ({desc}), "
                            f"CY constants cleared, dates updated"
                        )
                    elif sheet_name and sl in RAW_DATA_SHEETS:
                        log.append(f"— {sheet_name}: skipped (raw data sheet)")
                    elif sheet_name and sheet_name.strip().lower() in CAPITAL_SHEET_NAMES:
                        cap_info = cap_data.get(sheet_name, {})
                        if cap_info:
                            data = _proc._process_capital_sheet(data, cap_info)
                            log.append(
                                f"✓ {sheet_name}: capital CY→PY shifted, "
                                f"CY opening updated, additions/withdrawals cleared"
                            )
                        data = _proc._update_inline_strings(data, pairs)
                    elif sheet_name and sheet_name in lumid_text_only_names:
                        # Lumid-specific TEXT_ONLY extra: just update date strings
                        data = _proc._update_inline_strings(data, pairs)
                        log.append(f"· {sheet_name}: dates updated (Lumid TEXT_ONLY)")
                    elif sheet_name and sheet_name not in big_names:
                        data = _proc._update_inline_strings(data, pairs)
                        log.append(f"· {sheet_name}: dates updated")
                    elif sheet_name in big_names:
                        log.append(
                            f"* {sheet_name}: preserved unchanged "
                            f"(large sheet {sizes[sheet_name][0]} rows)"
                        )

                    zo.writestr(item, data)
                    continue

                # workbook.xml.rels — remove external-link references
                if fn == "xl/workbook.xml.rels" or fn.endswith(".rels"):
                    text_rels = data.decode("utf-8", errors="replace")
                    if "externalLinks" in text_rels:
                        text_rels = re.sub(
                            r'<Relationship[^>]*Target="externalLinks[^"]*"[^>]*/>\s*',
                            '', text_rels
                        )
                        data = text_rels.encode("utf-8")

                # workbook.xml — remove <externalReferences> section
                if fn == "xl/workbook.xml":
                    text_wb = data.decode("utf-8", errors="replace")
                    if "<externalReferences" in text_wb:
                        text_wb = re.sub(
                            r'<externalReferences\b[^>]*>.*?</externalReferences>',
                            '', text_wb, flags=re.DOTALL
                        )
                        data = text_wb.encode("utf-8")

                # [Content_Types].xml — remove external link content-type
                # overrides (bug fixed 2026-06-30: this step existed in
                # processor.py but was missing from this duplicated loop,
                # leaving Content_Types claiming externalLinkN.xml parts
                # existed after they'd already been stripped from the zip
                # above. That package inconsistency is what triggered
                # Excel's "repaired/removed unreadable content" dialog on
                # Lumid-format files specifically, since only this code
                # path — not processor.py's — handles them.)
                if fn == "[Content_Types].xml":
                    text_ct = data.decode("utf-8", errors="replace")
                    if "externalLink" in text_ct:
                        text_ct = re.sub(
                            r'<Override[^>]*externalLink[^>]*/>\s*',
                            '', text_ct
                        )
                        data = text_ct.encode("utf-8")

                zo.writestr(item, data)

    if ext_count:
        log.append(
            f"🔗 External references converted to values in {ext_count} sheet(s)"
        )

    # ── Repair XML malformations introduced by the cell-edit regex passes ────
    # The surgical XML regex editing in processor._process_sheet_xml can leave
    # three classes of malformed cell XML when it clears CY cells or rewrites
    # formula cells in this complex multi-column template:
    #
    #   (a) t="str" with no <f>: a cell typed as "string formula result" must
    #       always have a <f> child.  If the formula was removed (CY cell clear),
    #       the t="str" attribute must be removed too — otherwise Excel sees a
    #       string-type cell with no formula and reports "XML error".
    #
    #   (b) t="s" with no <v>: a shared-string cell must have a <v> child that
    #       contains the shared-string index.  If the <v> was removed (CY clear),
    #       the t="s" attribute must be removed — or the cell will be reported
    #       as unreadable.
    #
    #   (c) <v>…</v> appearing BEFORE <f>…</f> in the same cell: the OOXML spec
    #       requires the order <f> then <v>.  Some regex insertions produce the
    #       reverse order, which triggers Excel's "repairs required" dialog.
    #
    # We also drop calcChain.xml from the package entirely.  After a year-shift
    # many formula cells are cleared, re-typed, or moved, leaving the calcChain
    # stale.  Excel always rebuilds it on open; keeping a stale one causes the
    # "Removed Records: Formula from /xl/calcChain.xml" line in the repair log.
    _repair_worksheet_xml(output_path)

    # ── Freeze PY columns in summary sheets ──────────────────────────────────
    # BAL SHEET and P L have cross-sheet formulas in their PY (E) column.
    # After the year-shift, the cached values are correct but Excel's live
    # recalculation would change them (NOA sheets now have shifted/cleared cols).
    # Convert those PY formula cells to plain hardcoded values so the PY column
    # is stable regardless of Excel's recalculation order.
    _freeze_py_columns(output_path, {"BAL SHEET": ["E"], "P L": ["E"]})
    log.append("🔒 BAL SHEET & P L: PY column formulas frozen to preserve correct values")

    # Clean up the sanitized temp file created by _sanitize_cash_flow_errors
    # (it has a different path than the original input_path argument).
    try:
        import os as _os2
        original_arg = _run_with_patched_maps.__dict__.get("_orig_input")
        if original_arg and input_path != original_arg and _os2.path.exists(input_path):
            _os2.unlink(input_path)
    except Exception:
        pass   # non-critical — /tmp is cleaned up by the OS anyway

    return {"status": "success", "log": log}


def _repair_worksheet_xml(output_path: str) -> None:
    """
    Fix four classes of XML malformations left by the cell-edit regex passes,
    and remove the stale calcChain.xml that causes "Removed Records" repair noise.

    (a) t="str" cells with no <f>: remove the t="str" attribute.
    (b) t="s" cells with no <v>: remove the t="s" attribute.
    (c) <v>…</v> before <f>…</f>: swap them to <f>…</f><v>…</v>.
    (d) Drop calcChain.xml and its Content_Types entry entirely.
    (e) t="e" cells whose <v> is now a number (not an error string): remove t="e"
        and strip the stale <f> tag.  This catches any t="e" / #REF! cells that
        survived _sanitize_cash_flow_errors (e.g. in other sheets) and then had a
        numeric value written into them by the year-shift.  An OOXML t="e" cell
        with a numeric <v> is undefined behaviour and triggers Excel repair.
    (f) Remove empty <workbookProtection/> tag from workbook.xml.
        openpyxl writes this tag when building a workbook from scratch (e.g.
        during xlsb→xlsx conversion).  An empty <workbookProtection/> with no
        attributes is invalid — Excel shows "We found a problem with some
        content" popup on open and requires recovery.  Strip it entirely;
        workbooks without this tag open cleanly with no protection applied.
    """
    import zipfile as _zipfile, os as _os

    tmp = output_path + ".repair_tmp"

    with _zipfile.ZipFile(output_path, "r") as zi:
        with _zipfile.ZipFile(tmp, "w", _zipfile.ZIP_DEFLATED) as zo:
            for item in zi.infolist():
                fn = item.filename

                # Drop calcChain.xml — always stale after a year-shift
                if fn == "xl/calcChain.xml":
                    continue

                data = zi.read(fn)

                # Remove calcChain reference from Content_Types
                if fn == "[Content_Types].xml":
                    text = data.decode("utf-8", errors="replace")
                    text = re.sub(
                        r'<Override[^>]*calcChain[^>]*/>\s*', '', text
                    )
                    data = text.encode("utf-8")

                # (f) Strip empty tags from workbook.xml that openpyxl writes
                #     when building a workbook from scratch (xlsb→xlsx).
                #
                # workbookProtection: empty tag is invalid OOXML — triggers
                #   Excel's "We found a problem" recovery popup on open.
                #
                # definedNames: empty <definedNames /> triggers "Removed
                #   Records: Cell information from /xl/worksheets/sheetN.xml"
                #   repair dialog in Excel 2016+ (bug fixed 2026-07-09).
                #
                # workbookPr: empty <workbookPr /> is benign in most versions
                #   but can cause issues in some; strip it too.
                if fn == "xl/workbook.xml":
                    text = data.decode("utf-8", errors="replace")
                    # Empty workbookProtection
                    text = re.sub(r'\s*<workbookProtection\s*/>', '', text)
                    text = re.sub(r'\s*<workbookProtection\b[^>]*/>', '', text)
                    text = re.sub(
                        r'\s*<workbookProtection\b[^>]*>.*?</workbookProtection>',
                        '', text, flags=re.DOTALL
                    )
                    # Empty definedNames (no children)
                    text = re.sub(r'\s*<definedNames\s*/>', '', text)
                    text = re.sub(
                        r'\s*<definedNames\b[^>]*>\s*</definedNames>',
                        '', text, flags=re.DOTALL
                    )
                    # Empty workbookPr
                    text = re.sub(r'\s*<workbookPr\s*/>', '', text)
                    data = text.encode("utf-8")

                # Repair worksheet XMLs only
                if fn.startswith("xl/worksheets/") and fn.endswith(".xml"):
                    text = data.decode("utf-8", errors="replace")

                    # (a) Cells with text <v> values but no type attribute cause
                    #     openpyxl to try casting the text as a number → crash.
                    #     These arise when the year-shift tool copies a string value
                    #     from a t="str" formula cell into the PY column as a plain
                    #     value but loses the t="str" attribute in the process.
                    #
                    #     Fix: for any paired <c> cell that has NO t= attribute,
                    #     no <f> tag, and a <v> whose content is NOT numeric,
                    #     add t="str" so openpyxl reads it as a string.
                    #
                    #     We do NOT remove t="str" from cells — that was wrong and
                    #     caused exactly this crash.
                    def _add_missing_t_str(cm):
                        full = cm.group(0)
                        # Skip if already has a type attribute
                        if re.search(r'\bt="', full):
                            return full
                        # Skip if has a formula (type inferred from formula result)
                        if '<f' in full:
                            return full
                        v_m = re.search(r'<v>([^<]*)</v>', full)
                        if not v_m:
                            return full
                        val = v_m.group(1).strip()
                        # If value is numeric — no fix needed
                        try:
                            float(val)
                            return full
                        except (ValueError, OverflowError):
                            pass
                        # Text value with no type — add t="str"
                        return re.sub(r'(<c\b)', r'\1 t="str"', full, count=1)

                    text = re.sub(
                        r'<c\b(?![^>]*\bt=)[^>]*>(?:(?!</c>).)*?</c>',
                        _add_missing_t_str, text, flags=re.DOTALL
                    )

                    # (b) t="s" cells with no <v>: strip t="s"
                    def _fix_s_no_v(cm):
                        full = cm.group(0)
                        if '<v>' in full or '<v ' in full:
                            return full  # has shared-string ref — fine
                        return re.sub(r'\s*t="s"', '', full, count=1)

                    text = re.sub(
                        r'<c\b[^>]*\bt="s"[^>]*>(?:(?!</c>).)*?</c>',
                        _fix_s_no_v, text, flags=re.DOTALL
                    )

                    # (c) <v> before <f>: swap order to <f>…<v>
                    def _fix_v_before_f(cm):
                        full = cm.group(0)
                        # Extract <v>…</v> and <f>…</f> parts
                        v_m = re.search(r'<v>([^<]*)</v>', full)
                        f_m = re.search(r'<f[^>]*>.*?</f>', full, re.DOTALL)
                        if not (v_m and f_m):
                            return full
                        v_pos = full.index('<v>')
                        f_pos = full.index('<f')
                        if v_pos < f_pos:
                            # <v> comes before <f> — swap them
                            v_tag = v_m.group(0)
                            f_tag = f_m.group(0)
                            # Remove both tags then re-insert in correct order
                            inner = re.sub(r'<v>[^<]*</v>', '', full, flags=re.DOTALL)
                            inner = re.sub(r'<f[^>]*>.*?</f>', '', inner, flags=re.DOTALL)
                            inner = inner.replace('</c>', f'{f_tag}<v>{v_m.group(1)}</v></c>', 1)
                            return inner
                        return full

                    text = re.sub(
                        r'<c\b(?=[^>]*>(?!/))[^>]*>.*?</c>',
                        _fix_v_before_f, text, flags=re.DOTALL
                    )

                    # (d2) t="n" cells with no <v>: strip t="n".
                    # A numeric-typed cell with no value child is invalid OOXML
                    # and causes Excel's "Repairs required" / "Removed Records"
                    # dialog on open.  This arises specifically when the xlsb→xlsx
                    # conversion writes formula-cached values as plain t="n" cells,
                    # and the year-shift then clears the <v> tag without removing
                    # the t="n" attribute.  The primary fix is in processor.py's
                    # clear_v action, but we add a safety-net pass here too so any
                    # t="n" empty cells that slipped through are caught.
                    def _fix_n_no_v(cm):
                        full = cm.group(0)
                        if '<v>' in full or '<v ' in full:
                            return full  # has value — fine
                        if '<f' in full:
                            return full  # has formula — value will be recalculated
                        return re.sub(r'\s*t="n"', '', full, count=1)

                    text = re.sub(
                        r'<c\b[^>]*\bt="n"[^>]*>(?:(?!</c>).)*?</c>',
                        _fix_n_no_v, text, flags=re.DOTALL
                    )

                    # (e) t="e" cells with a numeric <v>: invalid per OOXML.
                    # After the year-shift, a t="e" (error-type) cell in the PY
                    # destination column may have had a real numeric value written
                    # into its <v> tag while the t="e" attribute and the stale
                    # #REF! <f> tag were left untouched.  Excel reports that as
                    # corrupt.  Fix: remove t="e" and strip the stale <f> tag
                    # so the cell becomes a plain numeric constant.
                    def _fix_e_type_numeric(cm):
                        full = cm.group(0)
                        v_m = re.search(r'<v>([^<]*)</v>', full)
                        if not v_m:
                            return full
                        try:
                            float(v_m.group(1))
                        except (ValueError, OverflowError):
                            return full   # not numeric — genuine error cell, leave it
                        # Numeric value inside t="e" cell — sanitize
                        fixed = re.sub(r'\s*t="e"', '', full, count=1)
                        fixed = re.sub(r'<f[^>]*>.*?</f>', '', fixed, flags=re.DOTALL)
                        fixed = re.sub(r'<f[^>]*/>', '', fixed)
                        return fixed

                    text = re.sub(
                        r'<c\b[^>]*\bt="e"[^>]*>.*?</c>',
                        _fix_e_type_numeric, text, flags=re.DOTALL
                    )

                    data = text.encode("utf-8")

                zo.writestr(item, data)

    _os.replace(tmp, output_path)


def _pooja_blank_cy_formulas(
    output_path: str,
    col_overrides: dict,
) -> None:
    """
    For every Pooja-format sheet, strip the <f> formula tag AND the <v> value
    from all cells in the CY columns (the first element of each shift pair),
    leaving those cells truly blank.

    WHY: The Pooja-I sheet stores BS/P&L totals as live intra-sheet formulas
    in the CY columns (e.g. C17 = =+L18, C18 = =+L85, G71 = =+C71).
    The year-shift correctly copies the cached <v> to the PY column but
    cannot clear the live <f> formula — only constant cells get blanked.
    When the CA enables editing Excel recalculates the still-live CY formulas
    against the now-shifted schedule columns → wrong (prior-year) figures.

    Fix: strip <f> AND <v> from every formula cell in each CY column so those
    cells become identical to the blank constant cells the shift already cleared.

    NOTE: <v xml:space="preserve"> variants must also be matched — the plain
    <v>[^<]*</v> pattern misses them; use <v\\b[^>]*>[^<]*</v> instead.
    """
    import zipfile as _zipfile
    import re as _re

    if not col_overrides:
        return

    cy_cols_by_sheet: dict[str, list[str]] = {
        sn: [cy for cy, _py in pairs]
        for sn, pairs in col_overrides.items()
    }

    with _zipfile.ZipFile(output_path, "r") as zi:
        smap = _proc._sheet_file_map(zi)
        all_items = {item.filename: (item, zi.read(item.filename))
                     for item in zi.infolist()}

    modified: dict[str, tuple] = {}

    for sheet_name, cy_cols in cy_cols_by_sheet.items():
        sf = smap.get(sheet_name, "")
        if not sf or sf not in all_items:
            continue

        item, xml_bytes = all_items[sf]
        text = xml_bytes.decode("utf-8", errors="replace")

        for cy_col in cy_cols:
            col_re = _re.compile(
                rf'<c\b[^>]*\br="{_re.escape(cy_col)}(\d+)"[^>]*/>\s*'
                rf'|<c\b[^>]*\br="{_re.escape(cy_col)}(\d+)"[^>]*>.*?</c>',
                _re.DOTALL
            )

            def _blank_cy_formula_cell(cm, _cy_col=cy_col):
                full = cm.group(0)
                has_f = bool(
                    _re.search(r'<f[^>]*>.*?</f>', full, _re.DOTALL) or
                    _re.search(r'<f[^>]*/>', full)
                )
                if not has_f:
                    return full  # constant cell — already handled by main shift
                # Strip <f> tag
                full = _re.sub(r'<f[^>]*>.*?</f>', '', full, flags=_re.DOTALL)
                full = _re.sub(r'<f[^>]*/>', '', full)
                # Strip <v> tag — including <v xml:space="preserve">…</v> variants
                full = _re.sub(r'<v\b[^>]*>[^<]*</v>', '', full)
                # Remove t= attribute (t="n"/"str" with no <v> is invalid OOXML)
                full = _re.sub(r'\s*\bt="[^"]*"', '', full, count=1)
                # If now just an empty paired tag, collapse to self-closing
                inner = _re.sub(r'<c\b[^>]*>', '', full, count=1)
                inner = _re.sub(r'</c>\s*$', '', inner).strip()
                if not inner:
                    r_m = _re.search(
                        rf'\br="{_re.escape(_cy_col)}\d+"', full
                    )
                    s_m = _re.search(r'\bs="(\d+)"', full)
                    r_attr = r_m.group(0) if r_m else ''
                    s_attr = f' s="{s_m.group(1)}"' if s_m else ''
                    return f'<c {r_attr}{s_attr}/>'
                return full

            text = col_re.sub(_blank_cy_formula_cell, text)

        modified[sf] = (item, text.encode("utf-8"))

    if not modified:
        return

    import os as _os
    tmp = output_path + ".pooja_blank_tmp"
    with _zipfile.ZipFile(output_path, "r") as zi:
        with _zipfile.ZipFile(tmp, "w", _zipfile.ZIP_DEFLATED) as zo:
            for item in zi.infolist():
                if item.filename in modified:
                    _, new_bytes = modified[item.filename]
                    zo.writestr(item, new_bytes)
                else:
                    zo.writestr(item, zi.read(item.filename))
    _os.replace(tmp, output_path)


def _freeze_py_columns(
    output_path: str,
    sheets_to_freeze: dict,
) -> None:
    """
    Convert formula cells in designated PY columns to plain hardcoded values.

    WHY THIS IS NEEDED FOR THE LUMID TEMPLATE:
    ─────────────────────────────────────────────────────────────────────────
    The Lumid BAL SHEET and P L sheets have their PY (E) column built entirely
    from cross-sheet formulas (e.g. E19 = 'NOA 3-6'!H22). After the year-shift:

    1. The tool correctly moves old CY constants/formula-cached values → E column.
    2. The CACHED <v> values in E-column cells are correct (verified: both BAL SHEET
       and P L totals balance to the correct PY figures in data_only mode).
    3. BUT when Excel opens the file and recalculates live, the cross-sheet formulas
       now point to NOA sheet columns that themselves contain a mix of old-shifted
       data and newly-cleared CY blanks. The recalculated values differ from the
       correct cached values, making the PY column show wrong figures.

    THE FIX:
    ─────────────────────────────────────────────────────────────────────────
    For each designated sheet and its PY column letter(s):
      - Remove <f>…</f> from every PY cell that has a formula AND a cached <v>.
      - Keep the <v> tag as a plain constant (the correct shifted value).
      - For formula cells with NO <v> (e.g. E22 = SUM(D19:D23) with blank CY ref):
        set the cell value to 0 and remove the formula.

    This is Lumid-specific because other templates do NOT have formula-driven
    PY columns in their summary sheets; they use plain value PY columns that
    are safe to keep as-is after a shift.
    ─────────────────────────────────────────────────────────────────────────

    sheets_to_freeze: {sheet_name: [py_col_letter, ...]}
      e.g. {"BAL SHEET": ["E"], "P L": ["E"]}
    """
    import zipfile as _zipfile

    with _zipfile.ZipFile(output_path, "r") as zi:
        smap = _proc._sheet_file_map(zi)
        all_items = {item.filename: (item, zi.read(item.filename))
                     for item in zi.infolist()}

    modified = {}
    for sheet_name, py_cols in sheets_to_freeze.items():
        sf = smap.get(sheet_name, "")
        if not sf or sf not in all_items:
            continue
        item, xml_bytes = all_items[sf]
        text = xml_bytes.decode("utf-8", errors="replace")

        for py_col in py_cols:
            # Match all cells in this column.
            # IMPORTANT: use TWO separate alternatives — self-closing (<c ... />) and
            # paired (<c ...>...</c>) — to prevent a self-closing E-col cell from
            # consuming the following non-E-col cell via the .*?</c> greedy path.
            col_re = re.compile(
                rf'<c\b[^>]*\br="{re.escape(py_col)}(\d+)"[^>]*/>\s*'   # self-closing
                rf'|<c\b[^>]*\br="{re.escape(py_col)}(\d+)"[^>]*>.*?</c>',  # paired
                re.DOTALL
            )

            def _freeze_cell(cm, _py_col=py_col):
                full = cm.group(0)
                # Only process cells that have a formula
                f_match = re.search(r'<f[^>]*>.*?</f>', full, re.DOTALL)
                f_sc_match = re.search(r'<f[^>]*/>', full)
                has_formula = bool(f_match or f_sc_match)
                if not has_formula:
                    return full

                # Check for cached value and cell type
                v_match = re.search(r'<v>([^<]*)</v>', full)
                t_attr = re.search(r'\bt="([^"]+)"', full)
                cell_type = t_attr.group(1) if t_attr else ''

                # Skip error cells — keep them as-is (#REF! etc. are pre-existing)
                if cell_type == 'e':
                    return full

                # String-typed formula cells (t="str"): the formula evaluates to a
                # text string (e.g. ='BAL SHEET'!A67 → "FOR AKSHIT MAHESHWARY...").
                # Convert these to shared-string references isn't straightforward here;
                # instead keep the formula so Excel can display the text correctly.
                # These cells are in P L E column (footer rows) and don't affect totals.
                if cell_type == 'str':
                    return full

                if v_match and v_match.group(1).strip():
                    # Has a valid numeric cached value — freeze: remove formula, keep value
                    full = re.sub(r'<f[^>]*>.*?</f>', '', full, flags=re.DOTALL)
                    full = re.sub(r'<f[^>]*/>', '', full)
                    # Convert self-closing to paired if needed (value is now present)
                    if full.rstrip().endswith('/>'):
                        full = re.sub(r'/>\s*$', '></c>', full.rstrip())
                else:
                    # No cached value (blank result) — replace formula with explicit 0
                    full = re.sub(r'<f[^>]*>.*?</f>', '', full, flags=re.DOTALL)
                    full = re.sub(r'<f[^>]*/>', '', full)
                    # Ensure the cell has a <v>0</v>
                    if '<v>' not in full:
                        if full.rstrip().endswith('/>'):
                            full = re.sub(r'/>\s*$', '><v>0</v></c>', full.rstrip())
                        else:
                            full = full.replace('</c>', '<v>0</v></c>')
                    else:
                        full = re.sub(r'<v>[^<]*</v>', '<v>0</v>', full)
                return full

            text = col_re.sub(_freeze_cell, text)

        modified[sf] = (item, text.encode("utf-8"))

    if not modified:
        return

    # Re-write the ZIP with the frozen sheets
    import shutil, tempfile, os
    tmp = output_path + ".freeze_tmp"
    with _zipfile.ZipFile(output_path, "r") as zi:
        with _zipfile.ZipFile(tmp, "w", _zipfile.ZIP_DEFLATED) as zo:
            for item in zi.infolist():
                if item.filename in modified:
                    _, new_bytes = modified[item.filename]
                    zo.writestr(item, new_bytes)
                else:
                    zo.writestr(item, zi.read(item.filename))
    os.replace(tmp, output_path)



__all__ = ["process", "_is_lumid_format"]


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 5:
        print("Usage: lumid_compat.py INPUT OUTPUT CLOSING_YEAR NEW_YEAR")
        sys.exit(1)
    res = process(sys.argv[1], sys.argv[2], int(sys.argv[3]), int(sys.argv[4]))
    print(f"Status: {res['status']}")
    for line in res.get("log", []):
        print(" ", line)
